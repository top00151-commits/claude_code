"""
╔══════════════════════════════════════════════════════════════╗
║  KNK PMS V4 — shared/knk_standard.py                         ║
║  스펙 §20 체크리스트 20항 통합 표준화                          ║
║  (스펙 §19 원칙 1: apply_knk_standard.py 1개만 실행)           ║
╚══════════════════════════════════════════════════════════════╝

이 모듈 = 엑셀 파일을 스펙 v2026.04에 정확히 맞추는 치유 스크립트.
- build.py/sync.py 실행 직후 호출
- 기존 데이터(value)는 건드리지 않고 서식·레이아웃만 재적용
- VML 메모 박스 크기 강제(380×110)는 저장 후 zip 재작성으로 처리
"""
import os
import re
import zipfile
import shutil
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation

from shared.styles import (
    KNK_RED, KNK_DARK_RED, KNK_TEXT,
    FT_DATA, FT_BOLD, FT_HEAD,
    AL_C, AL_L, AL_R,
    FILL_KNK_RED, FILL_KNK_DARK, FILL_WHITE, FILL_ALT,
    THIN, SHEET_PASSWORD,
    R3_GUIDE, R4_AREA,
    ACCT, ACCT_USD, PCT,
    apply_auto_widths, header_std_width,
)

# ═══════════════════════════════════════════════════════════════
# 스펙 상수
# ═══════════════════════════════════════════════════════════════
ROW_HEIGHTS = {1: 32, 2: 20, 3: 22, 4: 36}   # R1~R4
DATA_ROW_HEIGHT = 22                          # R5+
DDAY_FMT = '+0"일";-0"일";"0일"'
QTY_FMT  = '#,##0'
TEXT_FMT = '@'                                # 날짜/코드는 텍스트 유지
R2_SLOGAN_TMPL = "㈜케이엔케이  |  {purpose}  |  HAIST Innovation  |  Human & AI create the Best"
R2_TS_TMPL = "업데이트: {ts}"

# 긴 텍스트(좌측 indent=1) 헤더 키워드
LONG_TEXT_KEYS = ("품명", "모델", "비고", "메모", "이슈사항", "규격", "메이커", "업체")
# 우측정렬(금액·단가) 헤더 키워드
RIGHT_ALIGN_KEYS = ("금액", "총액", "단가", "인보이스")
# 중앙·숫자포맷(수량 등) 헤더 키워드
QTY_KEYS = ("수량", "BOX수")

R1_LEFT_ALIGN = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=False)
LEFT_INDENT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center", wrap_text=True, indent=1)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ═══════════════════════════════════════════════════════════════
# 1. R1 — 좌측정렬 indent=1 (스펙 §2, 체크 [1])
# ═══════════════════════════════════════════════════════════════
def apply_r1(ws, title, max_col):
    # 기존 R1 merge 해제 후 재적용
    to_unmerge = [str(mg) for mg in ws.merged_cells.ranges if mg.min_row == 1]
    for rng in to_unmerge:
        ws.unmerge_cells(rng)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    cell = ws.cell(1, 1)
    cell.value = title
    cell.fill = FILL_KNK_RED
    cell.font = Font(name="맑은 고딕", size=12, color="FFFFFF", bold=True)
    cell.alignment = R1_LEFT_ALIGN
    cell.border = THIN
    ws.row_dimensions[1].height = ROW_HEIGHTS[1]


# ═══════════════════════════════════════════════════════════════
# 2. R2 — 좌측 슬로건 병합 + 우측 끝 업데이트 일시 (스펙 §3, 체크 [2][3])
# ═══════════════════════════════════════════════════════════════
def apply_r2(ws, purpose, max_col, ts=None, slogan_override=None):
    """
    slogan_override: 지정 시 R2_SLOGAN_TMPL을 우회하고 이 텍스트를 그대로 사용.
                     (예: 시트별 안내문구 삽입)
    """
    if ts is None:
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

    # 기존 R2 merge 해제
    to_unmerge = [str(mg) for mg in ws.merged_cells.ranges if mg.min_row == 2]
    for rng in to_unmerge:
        ws.unmerge_cells(rng)

    # 좌측(A2 ~ C{max_col-1}) 병합 — 슬로건
    if max_col >= 2:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, max_col - 1))
    c1 = ws.cell(2, 1)
    c1.value = slogan_override if slogan_override else R2_SLOGAN_TMPL.format(purpose=purpose)
    c1.fill = FILL_KNK_DARK
    c1.font = Font(name="맑은 고딕", size=9, color="FFFFFF")
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=False)
    c1.border = THIN

    # 우측 끝 단일 셀 — 업데이트 일시
    if max_col >= 2:
        c2 = ws.cell(2, max_col)
        c2.value = R2_TS_TMPL.format(ts=ts)
        c2.fill = FILL_KNK_DARK
        c2.font = Font(name="맑은 고딕", size=9, color="FFFFFF", italic=True)
        c2.alignment = Alignment(horizontal="right", vertical="center", indent=1, wrap_text=False)
        c2.border = THIN

    ws.row_dimensions[2].height = ROW_HEIGHTS[2]


# ═══════════════════════════════════════════════════════════════
# 3. R3 — 가이드행 (스펙 §4, 체크 [4])
# ═══════════════════════════════════════════════════════════════
def apply_r3(ws, max_col, r3_map):
    to_unmerge = [str(mg) for mg in ws.merged_cells.ranges if mg.min_row == 3]
    for rng in to_unmerge:
        ws.unmerge_cells(rng)
    for c in range(1, max_col + 1):
        gtype = r3_map.get(c, "auto")
        g = R3_GUIDE.get(gtype, R3_GUIDE["auto"])
        cell = ws.cell(3, c)
        cell.value = g["text"]
        cell.fill = PatternFill("solid", fgColor=g["fill"])
        cell.font = Font(name="맑은 고딕", size=8, bold=True, color=g["fc"])
        cell.alignment = AL_C
        cell.border = THIN
    ws.row_dimensions[3].height = ROW_HEIGHTS[3]


# ═══════════════════════════════════════════════════════════════
# 4. R4 — 헤더 + 메모 (스펙 §5, §14, 체크 [5][6])
# ═══════════════════════════════════════════════════════════════
def apply_r4(ws, max_col, labels, r4_map, comments=None):
    for c in range(1, max_col + 1):
        area = r4_map.get(c, "manage")
        color = R4_AREA.get(area, "4A4A4A")
        cell = ws.cell(4, c)
        if c <= len(labels):
            cell.value = labels[c - 1]
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = FT_HEAD
        cell.alignment = AL_C
        cell.border = THIN
        # 메모
        if comments:
            text = comments.get(c) or _default_comment_for(cell.value)
            if text:
                cm = Comment(text, "KNK PMS")
                cm.width = 500    # v2026.04d: 380→500 (상세 메모 수용)
                cm.height = 220   # v2026.04d: 110→220
                cell.comment = cm
    ws.row_dimensions[4].height = ROW_HEIGHTS[4]


def _default_comment_for(header):
    """헤더명 → 상세 메모 텍스트 (스펙 §14 v2026.04d 확장)

    매칭 규칙: header 문자열에 key가 **포함**되면 그 메모를 반환.
    구체적 키가 먼저 평가되도록 순서 중요 (예: 고객사담당자 > 고객사).
    """
    if not header:
        return None
    h = str(header).replace("\n", "").replace(" ", "")

    rules = [
        # ── 식별정보 ──
        ("NO", (
            "▣ NO (일련번호)\n\n"
            "• sync 실행 시 1부터 자동 채번\n"
            "• 정렬 후 번호 재부여 (수동 입력 불가)\n"
            "• 최상단이 가장 납기 임박한 진행중 건"
        )),
        ("관리코드", (
            "▣ 관리코드 (프로젝트 고유 번호)\n\n"
            "• 포맷: [SEQ 3자리][사업부][YYMM]  예: 001T2604\n"
            "  - 사업부 T=검사기, M=자동화\n"
            "  - YYMM = 채번 시점 연월\n\n"
            "🟢 [신규 프로젝트]\n"
            "  비워두고 영업단계='수주확정' 입력 후 .bat 실행\n"
            "  → 자동 채번\n\n"
            "🟠 [추가 PO — 기존 검사기 추가 발주]\n"
            "  사용자가 기존 코드 직접 입력 (예: 005T2604)\n"
            "  → .bat 실행 시 그대로 유지 (재채번 X)\n\n"
            "⚠️ 같은 관리코드를 여러 행에 둘 때:\n"
            "  진척률·진행상태는 행마다 별도 관리되나\n"
            "  부서 입력 파일에서는 1개로 통합됨 (마지막 행 기준)\n"
            "  → 가급적 한 관리코드는 한 행으로 운영 권장"
        )),
        ("수주번호", (
            "▣ 수주번호\n\n"
            "• 포맷: [사업부]-[수주일 YYMMDD]  예: T-260501\n"
            "• 동일 수주일 복수 건: 뒤에 -N  예: T-260501-2\n\n"
            "• 관리코드가 채번된 후 자동 생성\n"
            "  - C17 수주일이 비어있으면 오늘 날짜로 부여\n"
            "  - 정확한 수주일 입력 후 .bat 실행 권장\n"
            "• R3=🔒자동 (사용자 수정 비권장 — sync 시 일관성 유지)"
        )),

        # ── 영업·고객 ──
        ("고객사담당자", (
            "▣ 고객사담당자 (고객사 PIC)\n\n"
            "• 고객사 측 담당자 이름 + 직책 직접 입력\n"
            "• 예: '장영남 수석', '김용우 책임'\n"
            "• 출하·납품·A/S 시 연락처로 활용"
        )),
        ("고객사", (
            "▣ 고객사 (거래처명)\n\n"
            "• 직접 입력 — 고객사 회사명\n"
            "• 예: 삼성전자, 드림텍, 파인텍, KNK VINA(삼성전자)\n"
            "• 대시보드 고객사별 매출 집계 기준"
        )),
        ("제품구분", (
            "▣ 제품구분 (검사기 분류)\n\n"
            "• 드롭다운 선택:\n"
            "  PBA    — PCB Board Assembly 검사기\n"
            "  TSP    — Touch Screen Panel 검사기\n"
            "  SENSOR — 센서 검사기\n"
            "  기타   — 위에 해당되지 않는 검사기\n"
            "• 대시보드 제품군별 집계에 사용"
        )),
        ("모델", (
            "▣ 모델 (장비 모델명)\n\n"
            "• 직접 입력 (좌측 정렬)\n"
            "• 예: SM-F776U PROXIMITY, IHVN-JG1\n"
            "• 고객사 제품 코드 또는 내부 모델명"
        )),
        ("품명", (
            "▣ 품명 (장비 명칭 / 작업 내용)\n\n"
            "• 직접 입력 (좌측 정렬)\n"
            "• 예: 기능검사기 MODIFY, ASSY 검사기\n"
            "• 서술적 이름 — 경영진이 한눈에 이해 가능하게"
        )),

        # ── PO / 영업단계 ──
        ("PO유형", (
            "▣ PO유형 (발주 성격)\n\n"
            "• 드롭다운:\n"
            "  신규 — 최초 수주 (신규 관리코드 채번)\n"
            "  추가 — 기존 프로젝트 확장 (기존 코드 직접 입력)\n"
            "  수정 — 기존 PO 조건 변경\n"
            "  개조 — 기 납품 장비 개조\n"
            "  기타 — 위에 해당되지 않는 경우\n\n"
            "💡 추가 PO 운영 팁:\n"
            "  - 같은 검사기 단순 추가 발주: 기존 행에서 수량·금액 누적\n"
            "  - 별도 추적 필요: 새 행 + 기존 관리코드 직접 입력\n"
            "    (단 부서 진척률은 1개로 통합됨에 주의)"
        )),
        ("영업단계", (
            "▣ 영업단계 (Sales Pipeline)\n\n"
            "• 드롭다운:\n"
            "  제안작성 — 제안서 작성 중 (코드 미채번)\n"
            "  제안제출 — 제안서 제출 완료 (코드 미채번)\n"
            "  수주확정 ★ — 수주 확정 → 관리코드 자동 채번\n"
            "  납품 — 제작·납품 진행\n"
            "  개조 — 기 납품 건 개조\n"
            "  A/S  — 사후 서비스\n\n"
            "📋 정렬 규칙 (sync 시 자동):\n"
            "  ① 진행중·수주예정 — 납기일 가까운 순\n"
            "  ② 제안작성·제안제출 — 수주일 가까운 순\n"
            "  ③ 보류 — 납기일 순\n"
            "  ④ 취소 — 납기일 순\n"
            "  ※ 납품완료/완료는 4_완료이력으로 자동 이동\n\n"
            "💡 제안 단계는 관리코드 없이도 보존됨 (수주확정 전까지 추적)"
        )),
        ("담당영업", (
            "▣ 담당영업 (영업 담당자)\n\n"
            "• 영업 인원 이름 직접 입력\n"
            "• 예: 이현, 오경환, 이해림, 안지연\n"
            "• 담당영업별 실적 집계 기준\n"
            "• 관리코드 미발행 건의 사업부 분류에도 사용"
        )),
        ("PM", (
            "▣ PM (Project Manager)\n\n"
            "• 프로젝트 매니저 이름 직접 입력\n"
            "• 수주확정 후 PM이 지정되면 기입\n"
            "• 부서 간 조율·납기 관리 책임자"
        )),

        # ── 수주 정보 ──
        ("수량", (
            "▣ 수량\n\n"
            "• 숫자 입력 (중앙 정렬, 천단위 콤마)\n"
            "• 정수 또는 소수 허용\n"
            "• 금액 = 수량 × 단가 (sync 시 자동 계산)"
        )),
        ("통화", (
            "▣ 통화 (KRW / USD)\n\n"
            "• 드롭다운 — 수주 계약 통화\n"
            "  KRW — 국내 거래 (원화)\n"
            "  USD — 해외 거래 (베트남·수출)\n"
            "• 단가·금액 포맷이 통화에 따라 자동 변경\n"
            "  KRW → ₩#,##0       USD → $#,##0.00"
        )),
        ("단가", (
            "▣ 단가 (계약 단가)\n\n"
            "• 숫자 입력 (우측 정렬)\n"
            "• 통화별 포맷 자동 적용 (₩ 또는 $)\n"
            "• sync 시 금액(= 수량 × 단가) 자동 재계산"
        )),
        ("금액", (
            "▣ 금액 (계약 총액)\n\n"
            "• 자동 계산 = 수량 × 단가\n"
            "• 수동 수정 잠김 — 수량·단가 변경 후 sync 재실행\n"
            "• 통화별 포맷 (₩ / $) 자동 적용\n"
            "• 대시보드 매출 집계 기준"
        )),
        ("수주일", (
            "▣ 수주일 (공식 수주 일자)\n\n"
            "• 입력 YYYY-MM-DD  예: 2026-04-15\n"
            "• 허용 형식(모두 자동 정규화):\n"
            "  2026/4/15, 2026.4.15, 260415, 20260415\n"
            "• 수주번호 생성·월별 집계 기준"
        )),
        ("납기일", (
            "▣ 납기일 (납품 약속일)\n\n"
            "• 입력 YYYY-MM-DD  예: 2026-08-31\n"
            "• 수주일과 동일한 허용 형식\n"
            "• D-day 자동 계산 기준\n"
            "• 정렬 기준 — 납기 임박한 건이 상단"
        )),

        # ── 현황 ──
        ("출하경로", (
            "▣ 출하경로 (물류 루트)\n\n"
            "• 드롭다운:\n"
            "  K→고객사       — 한국 내 직접 납품\n"
            "  K→V→K→고객사 — 한국→베트남→한국→고객사\n"
            "  K→V→고객사    — 한국→베트남→고객사 (해외)"
        )),
        ("전체", (
            "▣ 전체 진척률(%)\n\n"
            "• 자동 계산 — 참여 부서 소계의 평균\n"
            "  예: 3개 부서 진척률 80/50/30 → 53%\n"
            "• sync 시 갱신, 수동 수정 불가\n"
            "• 100% = 전 부서 완료 상태"
        )),
        ("D-day", (
            "▣ D-day (납기 카운터)\n\n"
            "• 자동 계산 = (오늘 - 납기일)\n"
            "  음수 (-)  → 남은 일수 (예: -5일 = 5일 남음)\n"
            "  0        → 오늘이 납기일\n"
            "  양수 (+)  → 지연 일수 (예: +3일 = 3일 지남)\n"
            "• 납품완료/취소/보류 상태는 공란 처리"
        )),
        ("진행상태", (
            "▣ 진행상태\n\n"
            "• 드롭다운:\n"
            "  수주예정 — 채번 전 대기 상태\n"
            "  진행중 ★ — 작업 진행 중 (정렬 최상단)\n"
            "  납품완료 — 완료 → 4_완료이력 자동 이관\n"
            "  취소    — 수주 취소\n"
            "  보류    — 일시 중단"
        )),
        ("비고", (
            "▣ 비고 (자유 메모)\n\n"
            "• 자유 텍스트 입력 (좌측 정렬)\n"
            "• 특이사항·요청사항·이슈 메모\n"
            "• 대시보드 집계에는 포함되지 않음"
        )),

        # ── 부서담당자 (부서 이름 매칭) ──
        ("부서담당자", (
            "▣ 부서담당자\n\n"
            "• 3가지 입력 방식:\n"
            "  1) '제외' 선택 → 이 프로젝트 해당 부서 미참여\n"
            "     (부서 파일에 등록 안 됨)\n"
            "  2) 빈 칸 → 부서 참여 확정, 담당자 미정\n"
            "  3) 담당자 이름 직접 입력 → 부서 파일 자동 반영\n"
            "• 부서 파일에서 담당자 입력 시 PMS에 역반영"
        )),

        # ── 소모품 전용 ──
        ("품목코드", (
            "▣ 품목코드\n\n"
            "• 직접 입력 — 내부 자재 관리 코드\n"
            "• 예: BP-001, PIN-T450\n"
            "• 출하관리와 연계 검색 키"
        )),
        ("단위", (
            "▣ 단위\n\n"
            "• 드롭다운: EA / SET / M / KG / BOX / ROLL / PAIR / 기타\n"
            "• 수량의 단위"
        )),
        ("규격", (
            "▣ 규격 (Specification)\n\n"
            "• 직접 입력 (좌측 정렬)\n"
            "• 제품 치수·전압·재질 등\n"
            "• 예: 10×10×5mm, 24V, SUS304"
        )),
        ("메이커", (
            "▣ 메이커 (제조사)\n\n"
            "• 제조사 이름 입력\n"
            "• 예: KEYENCE, OMRON, SMC, 자체제작"
        )),
        ("기본단가", (
            "▣ 기본단가 (KRW)\n\n"
            "• 숫자 입력 (우측 정렬, ₩ 표시)\n"
            "• 내부 표준 단가 (견적 기준)\n"
            "• 실제 출하 단가는 출하관리에서 별도 입력"
        )),
        ("카테고리", (
            "▣ 카테고리\n\n"
            "• 드롭다운: 지그부품 / 검사핀 / 커넥터 / 케이블 / 기타\n"
            "• 소모품 분류 기준"
        )),
        ("등록일", (
            "▣ 등록일\n\n"
            "• sync 실행 시 자동 입력 (수동 수정 불가)"
        )),

        # ── 부서명 직접 매칭 (설계팀, 검사기팀, 개발혁신팀, 품질팀, 제조기술1/2팀,
        #    구매팀, 가공팀, 전장설계팀, 소프트웨어팀, 베트남 등) ──
        ("팀", (
            "▣ 부서담당자\n\n"
            "• 3가지 입력 방식:\n"
            "  1) '제외' 선택 → 이 부서 미참여\n"
            "     (해당 부서 파일에 프로젝트 안 올라감)\n"
            "  2) 빈 칸 → 부서 참여 확정, 담당자 미정\n"
            "  3) 담당자 이름 직접 입력 → 부서 파일 자동 반영\n"
            "• 부서 파일에서 담당자 입력 시 PMS에 역반영"
        )),
        ("베트남", (
            "▣ 베트남 담당자\n\n"
            "• 3가지 입력 방식:\n"
            "  1) '제외' 선택 → 베트남 미참여\n"
            "  2) 빈 칸 → 베트남 참여 확정, 담당자 미정\n"
            "  3) 담당자 이름 직접 입력 → 베트남 파일 자동 반영\n"
            "• 주로 해외 출하 / 현지 조립 프로젝트에 활용"
        )),

        # ═══════════════════════════════════════════════════════════
        # 부서 입력 파일 전용 (v3.0)
        # ═══════════════════════════════════════════════════════════
        ("전체진척률", (
            "▣ 전체 진척률 (%)\n\n"
            "• 자동 계산 — 세부항목 평균\n"
            "• 수정 불가 (sync 시 자동 갱신)\n"
            "• 0% = 미착수 / 100% = 완료"
        )),
        ("담당자", (
            "▣ 부서 담당자\n\n"
            "• 부서 내 실제 담당자 이름 입력\n"
            "  예: 정민규, 한재운, 이찬\n\n"
            "• 입력 시: PMS에도 자동 반영\n"
            "• 빈 칸: 부서 참여 확정, 담당자 미정"
        )),
        ("제안서", (
            "▣ 제안서 완료일\n\n"
            "★ 날짜 입력 (% 아님 ⚠️)\n\n"
            "• 형식: YYYY-MM-DD  예: 2026-05-15\n"
            "• 허용: 26-5-15, 260515, 2026.5.15 (자동 정규화)\n"
            "• 실제 제안서 작성·제출 완료된 날짜 기입\n"
            "• 비워두면 미완료 상태\n"
            "• 예정일이 아닌 ★실제 완료일★ 만 입력"
        )),
        # 키는 \n 제거 형태 (헤더가 \n 제거되어 매칭됨)
        ("가공의뢰입고일", (
            "▣ 가공의뢰 입고일\n\n"
            "★ 날짜 입력 (% 아님 ⚠️)\n\n"
            "• 형식: YYYY-MM-DD  예: 2026-07-10\n"
            "• 가공팀 의뢰분 또는 외주 가공품의\n"
            "  ★실제 입고된 날짜★ 기입 (내부·외주 무관)\n\n"
            "• 비워두면: 대시보드 외주현황에 '진행중' 표시\n"
            "• 입고일 입력 시: '입고완료'로 자동 분류"
        )),
        ("외주입고일", (
            "▣ 외주 입고일 (양산 외주발주)\n\n"
            "★ 날짜 입력 (% 아님 ⚠️)\n\n"
            "• 형식: YYYY-MM-DD\n"
            "• 외부업체에서 양산 가공품이 실제 입고된 날짜\n"
            "• 비워두면: 대시보드 외주현황에 '진행중'\n"
            "• 입고일 입력 시: '입고완료' 자동 분류"
        )),
        ("가공입고일", (
            "▣ 가공 입고일 (베트남)\n\n"
            "★ 날짜 입력 (% 아님 ⚠️)\n\n"
            "• 형식: YYYY-MM-DD\n"
            "• 베트남 외주 가공품 실제 입고일\n"
            "• 비워두면 진행중 / 입력 시 입고완료"
        )),
        ("설계완료일", (
            "▣ 설계 완료일 (3D/2D 설계 완료)\n\n"
            "★ 날짜 입력 (% 아님 ⚠️)\n\n"
            "• 형식: YYYY-MM-DD  예: 2026-06-30\n"
            "• 3D/2D 설계가 실제로 완료된 날짜\n"
            "• 비워두면 미완료\n"
            "• 예정일 아닌 실제 완료일"
        )),
        # ── v3.1: 01 검사기 일정 중심 운영 (부서별 시작일·완료일) ──
        ("부서시작일", (
            "▣ 부서 시작일\n\n"
            "★ 날짜 입력 (% 아님 ⚠️)\n\n"
            "• 형식: YYYY-MM-DD  예: 2026-05-10\n"
            "• 허용: 26-5-10, 260510, 2026.5.10 (자동 정규화)\n"
            "• 부서가 ★실제로 작업을 시작한 날짜★ 기입\n"
            "• PO 후 작업 착수일 (지시 받은 날 아님)\n\n"
            "• 비워두면: 미착수 상태\n"
            "• v3.1 검사기 운영: 진척률 % 대신 시작·완료일로 추적\n"
            "  (2주 단위 빠른 작업 사이클이라 % 입력이 비효율)"
        )),
        ("부서완료일", (
            "▣ 부서 완료일\n\n"
            "★ 날짜 입력 (% 아님 ⚠️)\n\n"
            "• 형식: YYYY-MM-DD  예: 2026-05-24\n"
            "• 허용: 26-5-24, 260524, 2026.5.24 (자동 정규화)\n"
            "• 부서 작업이 ★실제로 완료된 날짜★ 기입\n"
            "• 예정일 아닌 ★실제 완료일★ 만 입력\n\n"
            "• 비워두면: PMS 진행현황에서 '미완료'로 집계\n"
            "• 입력 시: PMS 2_진행현황에 자동 표시 +\n"
            "  전체 진척률 = (완료 부서 수) / (참여 부서 수) 자동 계산\n"
            "• '제외' 부서는 분모에서 제외됨"
        )),
        ("외주발주일", (
            "▣ 외주 발주일 (양산 외주발주)\n\n"
            "★ 날짜 입력 (% 아님 ⚠️)\n\n"
            "• 형식: YYYY-MM-DD  예: 2026-05-15\n"
            "• 외부업체에 양산 가공품을 ★실제 발주한 날짜★ 기입\n"
            "• PO·발주서가 외부업체에 송부된 날짜\n\n"
            "• 외주입고일과 짝 — 발주 후 입고까지 추적\n"
            "• 비워두면: 외주 미발주 상태\n"
            "• 구매팀 전용 컬럼 (양산 외주 추적)"
        )),
        ("상태", (
            "▣ 부서 작업 상태\n\n"
            "• 드롭다운 선택:\n"
            "  미착수 — 작업 시작 안 함\n"
            "  진행중 — 작업 진행 중\n"
            "  완료   — 부서 작업 완료\n"
            "  N/A    — 해당 부서 무관 프로젝트\n\n"
            "• 부서장이 매주 상태 갱신 권장"
        )),
        ("메모", (
            "▣ 메모 (부서 자유 입력)\n\n"
            "• 자유 텍스트 입력\n"
            "• 활용:\n"
            "  - 이슈사항 (예: 자재 결품, 외주 지연)\n"
            "  - 변경 요청 (예: 사양 변경, 납기 조정)\n"
            "  - 다른 부서·PM에게 공유할 정보\n"
            "• 대시보드 집계에는 미포함"
        )),
    ]
    for key, txt in rules:
        if key.replace(" ", "") in h:
            return txt
    # Fallback: 부서 파일의 세부항목 (설계검토·3D2D·BOM·가공발주·PCB·H/W 등)
    # — 일반 형식상 "% 진척률 입력"으로 해석
    return (
        f"▣ {header} (세부항목 진척률)\n\n"
        f"★ 진척률 % 입력 (날짜 아님 ⚠️)\n\n"
        f"• 0~1 사이 소수로 입력\n"
        f"  예: 0.5 = 50%, 0.8 = 80%, 1.0 = 100%\n"
        f"• 0 또는 빈 칸 = 미착수\n"
        f"• 부서장·담당자가 매주 진척률 갱신\n\n"
        f"• 전체진척률(%)은 세부항목 평균으로 자동 계산"
    )


# ═══════════════════════════════════════════════════════════════
# 5. 데이터행 서식 (스펙 §6, §7, 체크 [7][8][15][16])
# ═══════════════════════════════════════════════════════════════
def apply_data_rows(ws, max_col, qty_cols=None, money_cols=None,
                    currency_col=None, long_text_cols=None, right_cols=None,
                    row_start=5, row_end=2000):
    """
    컬럼별 정렬 + 포맷 일괄 적용.
    - qty_cols: 수량류 (중앙, #,##0)
    - money_cols: 금액류 (우측, 통화별 포맷). currency_col 지정 시 행별로 KRW/USD 결정
    - long_text_cols: 긴 텍스트 (좌측 indent=1)
    - right_cols: 기타 우측정렬
    """
    qty_cols = set(qty_cols or [])
    money_cols = set(money_cols or [])
    long_text_cols = set(long_text_cols or [])
    right_cols = set(right_cols or [])

    for r in range(row_start, row_end + 1):
        bg = FILL_ALT if (r - row_start) % 2 == 1 else FILL_WHITE
        # 행별 통화 결정
        row_fmt = ACCT
        if currency_col:
            cur = ws.cell(r, currency_col).value
            if cur and str(cur).strip().upper() == "USD":
                row_fmt = ACCT_USD

        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.fill = bg
            cell.font = FT_DATA
            cell.border = THIN
            if c in qty_cols:
                cell.alignment = CENTER_ALIGN
                cell.number_format = QTY_FMT
            elif c in money_cols:
                cell.alignment = RIGHT_ALIGN
                cell.number_format = row_fmt
            elif c in long_text_cols:
                cell.alignment = LEFT_INDENT_ALIGN
            elif c in right_cols:
                cell.alignment = RIGHT_ALIGN
            else:
                cell.alignment = CENTER_ALIGN
        # R5+ 행 높이
        if ws.row_dimensions[r].height is None:
            ws.row_dimensions[r].height = DATA_ROW_HEIGHT


# ═══════════════════════════════════════════════════════════════
# 6. 헤더 기반 컬럼 자동 탐지 (config를 모르는 파일에도 적용 가능)
# ═══════════════════════════════════════════════════════════════
def detect_column_roles(ws, max_col):
    """R4 헤더를 읽어 qty/money/long_text/currency 컬럼을 자동 분류."""
    qty_cols, money_cols, long_text_cols = [], [], []
    currency_col = None
    dday_col = None
    for c in range(1, max_col + 1):
        h = ws.cell(4, c).value
        if not h:
            continue
        hs = str(h).replace("\n", "")
        if "통화" in hs:
            currency_col = c
        if "D-day" in hs:
            dday_col = c
        if any(k in hs for k in QTY_KEYS):
            qty_cols.append(c)
        elif any(k in hs for k in RIGHT_ALIGN_KEYS):
            money_cols.append(c)
        elif any(k in hs for k in LONG_TEXT_KEYS):
            long_text_cols.append(c)
    return {
        "qty": qty_cols,
        "money": money_cols,
        "long_text": long_text_cols,
        "currency_col": currency_col,
        "dday_col": dday_col,
    }


# ═══════════════════════════════════════════════════════════════
# 7. D-day 포맷 (스펙 §15, 체크 [12][13])
# ═══════════════════════════════════════════════════════════════
def apply_dday_format(ws, dday_col, status_col, row_start=5, row_end=2000):
    """D-day number_format 부여 + 납품완료/취소/보류 → 빈 칸"""
    if not dday_col:
        return
    inactive = {"납품완료", "완료", "취소", "보류"}
    for r in range(row_start, row_end + 1):
        cell = ws.cell(r, dday_col)
        if status_col:
            st = str(ws.cell(r, status_col).value or "").strip()
            if st in inactive:
                cell.value = None
        cell.number_format = DDAY_FMT


# ═══════════════════════════════════════════════════════════════
# 8. 시트 보호 (스펙 §9, 체크 [20])
# ═══════════════════════════════════════════════════════════════
def apply_sheet_protection(ws, max_col, r3_map, row_end=2000, password=None):
    password = password or SHEET_PASSWORD
    locked_types = {"auto", "auto_calc", "auto_sum"}
    unlocked = Protection(locked=False)
    locked = Protection(locked=True)
    for c in range(1, max_col + 1):
        gtype = r3_map.get(c, "auto")
        p = locked if gtype in locked_types else unlocked
        for r in range(5, row_end + 1):
            ws.cell(r, c).protection = p
    for r in range(1, 5):
        for c in range(1, max_col + 1):
            ws.cell(r, c).protection = locked
    ws.protection.sheet = True
    ws.protection.password = password
    ws.protection.formatColumns = False
    ws.protection.formatRows = False


# ═══════════════════════════════════════════════════════════════
# 8a. 데이터 이후 빈 행 서식 제거 — Ctrl+End가 마지막 데이터 행으로 가도록
# ═══════════════════════════════════════════════════════════════
def trim_empty_row_formatting(ws, max_col, row_start=5, row_end=2000):
    """
    데이터 있는 마지막 행 이후의 빈 행들을 **물리적으로 삭제**.
    결과: Excel에서 Ctrl+End → 마지막 데이터 행으로 정확히 이동.
    """
    # 데이터 있는 마지막 행 탐지 (row_end까지만 스캔)
    data_last = row_start - 1
    scan_end = min(row_end, ws.max_row)
    for r in range(row_start, scan_end + 1):
        for c in range(1, max_col + 1):
            if ws.cell(r, c).value is not None:
                data_last = r
                break

    # 최소 1행은 유지 (전혀 데이터 없는 경우 row_start에 빈 행 1개)
    if data_last < row_start:
        data_last = row_start

    # 데이터 이후 row_end까지의 빈 행들 일괄 삭제
    rows_to_delete = scan_end - data_last
    if rows_to_delete > 0:
        ws.delete_rows(data_last + 1, rows_to_delete)


# ═══════════════════════════════════════════════════════════════
# 8b. 드롭다운 재정비 (스펙 §10, 체크 [20] 관련)
# ═══════════════════════════════════════════════════════════════
def apply_dropdowns(ws, dropdown_map, row_start=5, row_end=2000):
    """
    기존 DataValidation 전부 제거 후 dropdown_map 기준으로 재생성.

    dropdown_map: {col: (options_list, allow_blank_bool)}

    마이그레이션·컬럼 재구성 후 드롭다운이 이전 위치에 남는 문제를 해결.
    """
    # 기존 DV 전부 제거
    try:
        ws.data_validations.dataValidation = []
    except Exception:
        pass

    # 재생성
    for col, entry in dropdown_map.items():
        options, allow_blank = entry if isinstance(entry, tuple) else (entry, True)
        formula = '"' + ",".join(options) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=bool(allow_blank))
        dv.error = "목록에서 선택하세요"
        dv.errorTitle = "입력 오류"
        col_letter = get_column_letter(col)
        dv.add(f"{col_letter}{row_start}:{col_letter}{row_end}")
        ws.add_data_validation(dv)


# ═══════════════════════════════════════════════════════════════
# 9. VML 메모 박스 크기 강제 380×110 (스펙 §14.4, 체크 [6])
# ═══════════════════════════════════════════════════════════════
def fix_vml_comment_size(xlsx_path, width_px=500, height_px=220):
    """
    xlsx zip 내부 xl/drawings/commentsDrawingN.vml 파일의
    style="... width:Wpx; height:Hpx" (또는 pt) 값을 강제 재작성.

    openpyxl 3.x는 Comment.width/height를 VML에 px 단위로 저장하므로
    Comment 객체에 380/110을 넣었다면 이 함수는 no-op에 가깝다.
    구버전 호환 + 안전장치.
    """
    tmp_path = xlsx_path + ".tmp"
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        names = zin.namelist()
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for name in names:
                data = zin.read(name)
                if name.startswith("xl/drawings/") and name.endswith(".vml"):
                    text = data.decode("utf-8", errors="ignore")
                    text = re.sub(r"width\s*:\s*[\d.]+\s*px", f"width:{width_px}px", text)
                    text = re.sub(r"height\s*:\s*[\d.]+\s*px", f"height:{height_px}px", text)
                    text = re.sub(r"width\s*:\s*[\d.]+\s*pt", f"width:{width_px}px", text)
                    text = re.sub(r"height\s*:\s*[\d.]+\s*pt", f"height:{height_px}px", text)
                    data = text.encode("utf-8")
                zout.writestr(name, data)
    shutil.move(tmp_path, xlsx_path)


# ═══════════════════════════════════════════════════════════════
# 10. 전체 시트 정규화 (메인 엔트리)
# ═══════════════════════════════════════════════════════════════
def normalize_sheet(ws, spec, log=None):
    """
    단일 시트를 스펙에 맞춰 정규화.
    spec (dict):
      - title:      R1 타이틀
      - purpose:    R2 슬로건의 목적 텍스트
      - max_col:    최대 컬럼
      - labels:     R4 헤더 리스트
      - r3_map:     {col: gtype}
      - r4_map:     {col: area}
      - comments:   {col: text} (선택, 없으면 기본 규칙으로 자동)
      - freeze:     "F5" 또는 "A5"
      - qty_cols, money_cols, long_text_cols, right_cols, currency_col, dday_col, status_col
        (선택, 없으면 R4 헤더로 자동 탐지)
      - row_end:    보호/서식 적용 범위 (기본 2000)
    """
    mc = spec["max_col"]

    # [1] R1
    apply_r1(ws, spec["title"], mc)
    # [2][3] R2 (slogan 지정 시 R2_SLOGAN_TMPL 우회)
    apply_r2(ws, spec.get("purpose", "업무용"), mc, slogan_override=spec.get("slogan"))
    # [4] R3
    apply_r3(ws, mc, spec.get("r3_map", {c: "auto" for c in range(1, mc + 1)}))
    # [5][6] R4 + 메모
    apply_r4(ws, mc, spec["labels"], spec.get("r4_map", {}), spec.get("comments") or {
        c: None for c in range(1, mc + 1)
    })

    # 컬럼 역할 자동 탐지 (헤더 기반)
    roles = detect_column_roles(ws, mc)
    qty_cols = spec.get("qty_cols") or roles["qty"]
    money_cols = spec.get("money_cols") or roles["money"]
    long_text_cols = spec.get("long_text_cols") or roles["long_text"]
    currency_col = spec.get("currency_col") or roles["currency_col"]
    dday_col = spec.get("dday_col") or roles["dday_col"]
    status_col = spec.get("status_col")

    row_end = spec.get("row_end", 2000)
    # [7][8][15][16] 데이터행 서식
    apply_data_rows(ws, mc,
                    qty_cols=qty_cols,
                    money_cols=money_cols,
                    currency_col=currency_col,
                    long_text_cols=long_text_cols,
                    right_cols=spec.get("right_cols"),
                    row_end=row_end)

    # [12][13] D-day
    if dday_col:
        apply_dday_format(ws, dday_col, status_col, row_end=row_end)

    # [9] 글틀고정 — "auto"면 labels에서 "품명" 위치 찾아 그 다음 열 기준 고정
    fr = spec.get("freeze", "auto")
    if fr == "auto":
        target_col = None
        labels = spec.get("labels", [])
        for i, lab in enumerate(labels, start=1):
            if "품명" in str(lab or "").replace("\n", ""):
                target_col = i
                break
        fr = f"{get_column_letter(target_col + 1)}5" if target_col else "A5"
    ws.freeze_panes = fr

    # [10] 컬럼 너비
    apply_auto_widths(ws, mc)

    # 드롭다운 재정비 (구 위치 잔재 전부 제거 + 새로 생성)
    dropdown_map = spec.get("dropdown_map")
    if dropdown_map is not None:
        apply_dropdowns(ws, dropdown_map, row_end=row_end)

    # [20] 시트 보호
    apply_sheet_protection(ws, mc, spec.get("r3_map", {}), row_end=row_end)

    # ★ v2026.04g 맨 마지막: used range 축소 + 빈 행 삭제 + 입력 셀 unlock 보장
    # buffer=30: 신규 입력 영역 30행 사전 마련 (위 데이터와 동일 서식 적용)
    # min_end=30: 데이터 0건이어도 R30까지 입력 영역 유지
    _shrink_used_range(ws, mc,
                       row_start=5,
                       buffer=spec.get("trim_buffer", 30),
                       min_end=spec.get("trim_min_end", 30),
                       r3_map=spec.get("r3_map", {}),
                       log=log)

    if log:
        log(f"  ✓ {ws.title}: mc={mc}, freeze={fr}, qty={qty_cols}, money={money_cols}, "
            f"long={long_text_cols}, dday={dday_col}")


def _shrink_used_range(ws, mc, row_start=5, buffer=1, min_end=14, r3_map=None, log=None):
    """
    used range 최소화 + 빈 행 정리 + 입력 셀 unlock 재적용.

    동작:
      1) 데이터 마지막 행(data_last) 측정
      2) active_end = max(data_last + buffer, min_end)
      3) active_end 너머의 모든 행 delete (sheetData·row_dims·cell 메타 모두 정리)
      4) DataValidation 범위를 active_end로 축소
      5) ★ R5~R{active_end} 입력 컬럼 unlock 재적용 (delete_rows 후 default-lock 방지)

    효과:
      - Ctrl+End → 데이터 마지막 행 + buffer로 정확히 이동
      - 새로 입력하려는 빈 행도 unlock 상태로 유지 (잠겨있던 R109+ 문제 해결)
    """
    # 1) 데이터 마지막 행 측정
    data_last = row_start - 1
    for r in range(row_start, ws.max_row + 1):
        for c in range(1, mc + 1):
            if ws.cell(r, c).value is not None:
                data_last = r
                break
    active_end = max(data_last + buffer, min_end)

    # 2) active_end 너머의 모든 행 삭제 (cell·row 노드 모두 제거)
    rows_to_delete = max(0, ws.max_row - active_end)
    if rows_to_delete > 0:
        ws.delete_rows(active_end + 1, rows_to_delete)

    # 3) row_dimensions에서 active_end 너머 제거 (delete_rows로도 못 정리되는 메타)
    rd_removed = 0
    for r in list(ws.row_dimensions.keys()):
        if r > active_end:
            del ws.row_dimensions[r]
            rd_removed += 1

    # 4) DataValidation 범위 축소
    dv_shrunk = 0
    for dv in ws.data_validations.dataValidation:
        new_refs = []
        for ref in str(dv.sqref).split():
            m = re.match(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", ref)
            if m:
                col1, r1, col2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
                if r2 > active_end:
                    new_refs.append(f"{col1}{r1}:{col2}{active_end}")
                    dv_shrunk += 1
                else:
                    new_refs.append(ref)
            else:
                new_refs.append(ref)
        dv.sqref = " ".join(new_refs)

    # 5) ★ R5~R{active_end} 입력 셀 unlock 재적용 (가장 중요)
    # apply_sheet_protection 이후 delete_rows를 거치면 일부 셀의 protection 메타가 누락되거나
    # default(locked=True)로 돌아갈 수 있음. R{data_last+1}~R{active_end}의 빈 입력 행도
    # 명시적으로 unlock 처리해 사용자가 신규 행에 입력 가능하도록 보장.
    unlock_cells = 0
    unlock_cols_set = set()
    if r3_map:
        unlocked_p = Protection(locked=False)
        locked_p   = Protection(locked=True)
        locked_types = {"auto", "auto_calc", "auto_sum"}
        for c, gtype in r3_map.items():
            if gtype not in locked_types:
                unlock_cols_set.add(c)
        for r in range(row_start, active_end + 1):
            for c in range(1, mc + 1):
                if c in unlock_cols_set:
                    ws.cell(r, c).protection = unlocked_p
                    unlock_cells += 1
                else:
                    ws.cell(r, c).protection = locked_p

    # 6) ★★ 컬럼 단위 default protection (cell 메타 없는 빈 행에도 적용)
    # Ctrl+End를 active_end로 좁혀도, 사용자가 active_end 너머 임의 행으로 이동해
    # 입력 시도하면 cell이 없어 Excel은 default(=locked) 적용 → 입력 불가.
    # column.protection.locked=False를 설정하면 빈 셀에도 column default가 적용되어
    # 입력 컬럼은 어느 행에서나 입력 가능.
    col_unlocked = 0
    if r3_map:
        for c in range(1, mc + 1):
            col_letter = get_column_letter(c)
            cd = ws.column_dimensions[col_letter]
            if c in unlock_cols_set:
                cd.protection = Protection(locked=False)
                col_unlocked += 1
            else:
                cd.protection = Protection(locked=True)

    if log:
        log(f"    ↳ used range: data_last={data_last} active_end={active_end} "
            f"deleted-rows={rows_to_delete} row_dims-{rd_removed} DV축소={dv_shrunk} "
            f"unlock_cells={unlock_cells} col_unlocked={col_unlocked}")


# ═══════════════════════════════════════════════════════════════
# 11. 파일 단위 정규화 (여러 시트 한 번에)
# ═══════════════════════════════════════════════════════════════
def normalize_file(xlsx_path, sheet_specs, fix_vml=True, log=None):
    """
    xlsx_path의 각 시트에 sheet_specs[시트명]을 적용.
    시트명이 specs에 없으면 건드리지 않음.
    fix_vml=True 이면 저장 후 VML 메모 크기 380×110 강제 수정.
    """
    if log:
        log(f"▶ {os.path.basename(xlsx_path)}")

    wb = load_workbook(xlsx_path)
    for sheet_name, spec in sheet_specs.items():
        if sheet_name not in wb.sheetnames:
            if log:
                log(f"  ⚠ 시트 없음: {sheet_name}")
            continue
        ws = wb[sheet_name]
        normalize_sheet(ws, spec, log=log)

    wb.save(xlsx_path)

    if fix_vml:
        fix_vml_comment_size(xlsx_path)
        if log:
            log(f"  ✓ VML 메모 박스 500×220 px 강제 적용")

    if log:
        log(f"  ✓ 저장 완료")
