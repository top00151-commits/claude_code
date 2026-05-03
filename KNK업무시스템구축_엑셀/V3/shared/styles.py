"""
╔══════════════════════════════════════════════════════════════╗
║  KNK PMS V4 — shared/styles.py                               ║
║  공통 스타일·R3 가이드·시트보호·Comment·서식 유틸리티          ║
║  ※ 각 유형별 config.py에서 import하여 사용                    ║
║  ※ 이 파일은 스타일/유틸만 담당 — 컬럼·시트 정의는 각 config  ║
╚══════════════════════════════════════════════════════════════╝
"""

import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════
# A. 날짜·연도 (전역)
# ═══════════════════════════════════════════════════════════════
TODAY = datetime.date.today()
YEAR  = TODAY.year
YYMM  = TODAY.strftime("%y%m")

# ═══════════════════════════════════════════════════════════════
# B. KNK 브랜드 컬러
# ═══════════════════════════════════════════════════════════════
KNK_RED      = "A5282C"
KNK_DARK_RED = "8B1E22"
KNK_GRAY     = "4A4A4A"
KNK_BG_WHITE = "FFFFFF"
KNK_BG_ALT   = "F5F5F5"
KNK_TEXT      = "1A1A1A"

# ═══════════════════════════════════════════════════════════════
# C. openpyxl 스타일 상수
# ═══════════════════════════════════════════════════════════════
FT_DATA  = Font(name="맑은 고딕", size=9, color=KNK_TEXT)
FT_BOLD  = Font(name="맑은 고딕", size=9, color=KNK_TEXT, bold=True)
FT_HEAD  = Font(name="맑은 고딕", size=9, color="FFFFFF", bold=True)
FT_R3    = Font(name="맑은 고딕", size=8, bold=True)  # R3 가이드행 폰트

AL_C  = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_L  = Alignment(horizontal="left",   vertical="center", wrap_text=True, indent=1)
AL_R  = Alignment(horizontal="right",  vertical="center", wrap_text=True)
AL_L0 = Alignment(horizontal="left",   vertical="center", wrap_text=True)     # indent 없음
AL_R0 = Alignment(horizontal="right",  vertical="center", wrap_text=True, indent=1)  # 금액 (약간 여백)

THIN = Border(
    left=Side("thin", color="D9D9D9"), right=Side("thin", color="D9D9D9"),
    top=Side("thin", color="D9D9D9"), bottom=Side("thin", color="D9D9D9"))

FILL_WHITE  = PatternFill("solid", fgColor=KNK_BG_WHITE)
FILL_ALT    = PatternFill("solid", fgColor=KNK_BG_ALT)
FILL_INPUT  = PatternFill("solid", fgColor="E8F5E9")

# KNK 브랜드 Fill
FILL_KNK_RED   = PatternFill("solid", fgColor=KNK_RED)
FILL_KNK_DARK  = PatternFill("solid", fgColor=KNK_DARK_RED)
FILL_KNK_GRAY  = PatternFill("solid", fgColor=KNK_GRAY)

# 상태 컬러
FILL_PROGRESS = PatternFill("solid", fgColor="FFF8E1")   # 진행중 (노란)
FILL_DONE     = PatternFill("solid", fgColor="E8F5E9")   # 완료 (녹색)
FILL_DELAY    = PatternFill("solid", fgColor="FFEBEE")   # 지연 (연한 레드)
FILL_WAIT     = PatternFill("solid", fgColor=KNK_BG_ALT) # 대기 (그레이)

# 부서별 구분 색상 (2_진행현황/4_완료이력 헤더용 — 최대 8부서)
DEPT_HEADER_COLORS = [
    "2E5090",  # 남색 (설계팀)
    "00695C",  # 진녹 (검사기팀/전장설계팀)
    "6A1B9A",  # 보라 (개발혁신팀/소프트웨어팀)
    "BF360C",  # 갈색 (품질팀/구매팀)
    "37474F",  # 진회 (제조기술팀)
    "1B5E20",  # 진초록 (구매팀)
    "4E342E",  # 다크브라운 (가공팀)
    "01579B",  # 딥블루 (베트남)
]

# 시트 탭 색상
TAB_INPUT = "00C853"   # 초록 — 입력용
TAB_VIEW  = "A5282C"   # KNK RED — 확인용 핵심
TAB_VIEW2 = "9E9E9E"   # 회색 — 확인용 보조

# ═══════════════════════════════════════════════════════════════
# D. 숫자 포맷
# ═══════════════════════════════════════════════════════════════
ACCT     = '₩#,##0;(₩#,##0);"-"'        # KRW
ACCT_USD = '$#,##0.00;($#,##0.00);"-"'  # USD
PCT      = '0%'

def money_fmt(currency="KRW"):
    """통화에 따른 숫자 포맷 반환"""
    return ACCT_USD if str(currency).strip().upper() == "USD" else ACCT

# ═══════════════════════════════════════════════════════════════
# E. R3 가이드행 디자인 시스템 (12가지 유형)
# ═══════════════════════════════════════════════════════════════
R3_GUIDE = {
    "auto":       {"fill": "E0E0E0", "fc": KNK_TEXT, "text": "🔒자동"},
    "auto_calc":  {"fill": "E0E0E0", "fc": KNK_TEXT, "text": "🔒자동계산"},
    "auto_sum":   {"fill": "E0E0E0", "fc": KNK_TEXT, "text": "🔒자동합산"},
    "auto_input": {"fill": "FFE0B2", "fc": KNK_TEXT, "text": "자동·입력"},
    "input":      {"fill": "C8E6C9", "fc": KNK_TEXT, "text": "✏입력"},
    "select":     {"fill": "C8E6C9", "fc": KNK_TEXT, "text": "✏선택▼"},
    "sel_input":  {"fill": "C8E6C9", "fc": KNK_TEXT, "text": "✏선택/입력"},
    "po_input":   {"fill": "FFE0B2", "fc": KNK_TEXT, "text": "PO후입력"},
    "po_select":  {"fill": "FFE0B2", "fc": KNK_TEXT, "text": "PO후선택▼"},
    "dept_sel":   {"fill": "CE93D8", "fc": KNK_TEXT, "text": "📋 선택▼"},
    "memo":       {"fill": "FFE0B2", "fc": KNK_TEXT, "text": "✏메모"},
    "ship":       {"fill": "E0E0E0", "fc": KNK_TEXT, "text": "🚚"},
}

# R4 헤더행 영역별 색상 (배경색 HEX, 글자색=흰색)
R4_AREA = {
    "id":      KNK_RED,    # 식별정보
    "sales":   "1B3A5C",   # 영업정보
    "po":      "00695C",   # 수주정보
    "status":  "37474F",   # 진행현황
    "ship":    "BF360C",   # 출하정보
    "dept":    "6A1B9A",   # 부서선택
    "manage":  "37474F",   # 관리항목
    "procure": "1B5E20",   # 자재조달
    "payment": "B08F36",   # 매출·수금 (v2026.04e — 황갈색)
    # v3.1.2: 부서 세부항목 [%·예정일] 짝 그룹 (옵션 B 시각 강화)
    "pair_a":  "0277BD",   # 짝 A — 깊은 청 (항목 1, 3, 5번)
    "pair_b":  "EF6C00",   # 짝 B — 깊은 주황 (항목 2, 4, 6번)
}

# ═══════════════════════════════════════════════════════════════
# F. 시트 보호
# ═══════════════════════════════════════════════════════════════
SHEET_PASSWORD = "knk2026"


# ═══════════════════════════════════════════════════════════════
# G. 유틸리티 함수
# ═══════════════════════════════════════════════════════════════

def apply_r3_guide(ws, max_col, r3_map, row=3):
    """
    R3 가이드행 적용.
    ws: worksheet, max_col: 최대 컬럼, r3_map: {컬럼번호: 가이드유형}
    기존 merge 해제 후, 각 셀에 유형별 색상+텍스트 표시.
    """
    # 기존 R3 merge 해제
    merges_to_remove = []
    for mg in ws.merged_cells.ranges:
        if mg.min_row <= row <= mg.max_row:
            merges_to_remove.append(mg)
    for mg in merges_to_remove:
        ws.unmerge_cells(str(mg))

    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        guide_type = r3_map.get(c, "auto")
        g = R3_GUIDE.get(guide_type, R3_GUIDE["auto"])
        cell.value = g["text"]
        cell.fill = PatternFill("solid", fgColor=g["fill"])
        cell.font = Font(name="맑은 고딕", size=8, bold=True, color=g["fc"])
        cell.alignment = AL_C
        cell.border = THIN


def apply_r4_header(ws, max_col, labels, r4_area_map, area_colors=None, row=4):
    """
    R4 헤더행 적용.
    labels: 컬럼명 리스트, r4_area_map: {컬럼번호: 영역키}
    area_colors: 영역별 색상 딕셔너리 (기본=R4_AREA)
    """
    if area_colors is None:
        area_colors = R4_AREA
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.value = labels[c - 1] if c <= len(labels) else ""
        area_key = r4_area_map.get(c, "manage")
        color = area_colors.get(area_key, KNK_GRAY)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = FT_HEAD
        cell.alignment = AL_C
        cell.border = THIN


def apply_protection(ws, max_col, r3_map, extra_unlock=None, row_start=5, row_end=2000, password=None):
    """
    시트 보호 적용 (R3 기반).
    auto/auto_calc/auto_sum → 잠금, 나머지 → 편집 허용.
    extra_unlock: 추가 해제할 컬럼 리스트
    """
    if password is None:
        password = SHEET_PASSWORD

    locked_types = {"auto", "auto_calc", "auto_sum"}
    unlock_cols = set()
    for c, gtype in r3_map.items():
        if gtype not in locked_types:
            unlock_cols.add(c)
    if extra_unlock:
        unlock_cols.update(extra_unlock)

    unlocked = Protection(locked=False)
    locked   = Protection(locked=True)

    for r in range(row_start, row_end + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).protection = unlocked if c in unlock_cols else locked

    # R1~R4 행 잠금
    for r in range(1, row_start):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).protection = locked

    ws.protection.sheet = True
    ws.protection.password = password
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False


def make_comment(text, author="KNK PMS"):
    """R4 Comment 생성 (크기 자동 조정)"""
    c = Comment(text, author)
    lines = text.count("\n") + 1
    max_line = max(len(l) for l in text.split("\n")) if text else 10
    c.width = min(max(max_line * 7, 150), 400)
    c.height = min(max(lines * 18, 50), 300)
    return c


def setup_r1(ws, title, max_col, row=1):
    """R1 타이틀행 (전체 병합, KNK RED 배경, 좌측정렬 indent=1 — 스펙 §2)"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1)
    cell.value = title
    cell.fill = FILL_KNK_RED
    cell.font = Font(name="맑은 고딕", size=12, color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=False)
    cell.border = THIN
    ws.row_dimensions[row].height = 32


def format_data_rows(ws, max_col, money_cols=None, money_usd_cols=None,
                     pct_cols=None, row_start=5, row_end=2000):
    """
    데이터행 서식 적용.
    교대 배경 + 금액 우측정렬 + 통화포맷 + 퍼센트포맷
    """
    if money_cols is None:
        money_cols = []
    if money_usd_cols is None:
        money_usd_cols = []
    if pct_cols is None:
        pct_cols = []

    for r in range(row_start, row_end + 1):
        bg = FILL_ALT if r % 2 == 0 else FILL_WHITE
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = bg
            cell.font = FT_DATA
            cell.border = THIN
            if c in money_cols:
                cell.alignment = AL_R
                cell.number_format = ACCT
            elif c in money_usd_cols:
                cell.alignment = AL_R
                cell.number_format = ACCT_USD
            elif c in pct_cols:
                cell.alignment = AL_C
                cell.number_format = PCT
            else:
                cell.alignment = AL_C


def auto_fit_columns(ws, min_width=5, max_width=40):
    """컬럼 너비 자동 조정 — 스펙 §8 (min=5, max=40, 한글 2폭 환산)

    R1(타이틀 병합) · R2(슬로건 병합)는 측정 제외 —
    A열만 값 보유해서 A열 폭이 타이틀 길이로 부풀어오르는 문제 회피.
    """
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_width
        for cell in col_cells[:50]:
            if cell.row <= 2:   # R1·R2 병합 타이틀 스킵
                continue
            if cell.value:
                for line in str(cell.value).split("\n"):
                    cell_len = sum(2 if ord(ch) > 127 else 1 for ch in line.strip()) + 2
                    if cell_len > max_len:
                        max_len = cell_len
        ws.column_dimensions[col_letter].width = min(max_len, max_width)


# ═══════════════════════════════════════════════════════════════
# H. 스펙 §8.1 표준 너비 룩업 (헤더 텍스트 매칭)
# ═══════════════════════════════════════════════════════════════
_STD_WIDTH_RULES = [
    (("NO",), 5),
    (("관리코드", "수주번호"), 14),
    (("단위",), 6),
    (("수량", "BOX수량", "BOX수"), 8),
    (("통화",), 6),
    (("환율",), 9),
    (("단가",), 15),                          # v2026.04c: 13→15 (₩ 렌더 여유)
    (("금액", "총액", "인보이스"), 18),        # v2026.04c: 16→18 (큰 금액 #### 방지)
    (("일",), 12),
    (("상태", "PO유형"), 10),
    (("고객사",), 18),
    (("담당자", "PM", "영업", "담당"), 14),
    (("품명",), 24),
    (("모델",), 18),
    (("규격", "메이커", "업체"), 14),
    (("비고", "메모", "이슈사항"), 28),
    (("진행률", "D-day"), 9),
    (("부서담당자",), 14),
]


def header_std_width(header):
    """헤더 텍스트 → 표준 너비 (스펙 §8.1). 매칭 없으면 None."""
    if not header:
        return None
    h = str(header).replace("\n", "").strip()
    for keys, w in _STD_WIDTH_RULES:
        for k in keys:
            if k in h:
                return w
    return None


def _render_value(v, fmt):
    """셀 값을 number_format에 맞춰 렌더링된 문자열로 변환 (폭 측정용)."""
    if v is None:
        return ""
    if fmt in (None, "", "General", "@"):
        return str(v)
    if isinstance(v, (int, float)):
        # 통화 prefix (₩ 또는 $)
        prefix = ""
        if "₩" in fmt:
            prefix = "₩"
        elif "$" in fmt:
            prefix = "$"
        # 소수 자리수 추정
        decimals = 0
        if ".00" in fmt or "0.00" in fmt:
            decimals = 2
        try:
            body = f"{v:,.{decimals}f}" if "#,##0" in fmt or "," in fmt else f"{v:.{decimals}f}"
        except (ValueError, TypeError):
            body = str(v)
        return prefix + body
    return str(v)


def auto_width_smart(ws, col, row_start=4, row_end=300, min_w=5, max_w=40):
    """
    스펙 §8.2 데이터 기반 자동조정.
    - 헤더(R4) + 데이터(R5+) 모두 측정, 한글 2폭 환산
    - 셀 number_format 고려 (₩·$·,·소수점 포함 렌더링 기준)
    - 표준 너비가 있으면 max(표준, 데이터+1) 적용
    """
    header = ws.cell(4, col).value
    def w(s):
        return sum(2 if ord(c) > 127 else 1 for c in s)
    h_max = 0
    if header:
        for line in str(header).split("\n"):
            h_max = max(h_max, w(line.strip()))
    data_max = 0
    for r in range(max(5, row_start), min(ws.max_row + 1, row_end + 1)):
        cell = ws.cell(r, col)
        if cell.value is not None:
            rendered = _render_value(cell.value, cell.number_format)
            for line in rendered.split("\n"):
                data_max = max(data_max, w(line.strip()))
    std = header_std_width(header)
    base = max(h_max, data_max)
    if std:
        return max(min_w, min(max_w, max(std, base + 1)))
    return max(min_w, min(max_w, base + 2))


def apply_auto_widths(ws, max_col, row_start=4, row_end=300, min_w=5, max_w=40):
    """전체 컬럼 auto_width_smart 일괄 적용"""
    for c in range(1, max_col + 1):
        w = auto_width_smart(ws, c, row_start, row_end, min_w, max_w)
        ws.column_dimensions[get_column_letter(c)].width = w


def fmt_cell(ws, row, col, value, is_money=False, is_usd=False):
    """sync에서 셀 쓰기 유틸리티 (서식 유지)"""
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if is_money:
        cell.alignment = AL_R
        cell.number_format = ACCT_USD if is_usd else ACCT
