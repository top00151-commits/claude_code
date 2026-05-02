# -*- coding: utf-8 -*-
"""
Phase 2-A: 부서 입력 파일에 마일스톤 날짜·입고일 컬럼 추가
- 위치: 세부항목 다음, 상태·메모 직전
- 추가 컬럼:
  * 설계팀:        제안서 완료일·설계 완료일·가공의뢰 입고일 (3)
  * 소프트웨어팀:  제안서 완료일·설계 완료일 (2)
  * 구매팀:        외주 입고일 (1)
  * 베트남:        가공 입고일 (1)
- R3 = ✏입력 (po_input 또는 input), R4 = po 영역색
- 그 외 부서: 변경 없음
"""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

V3 = Path(__file__).parent

THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)

# R3 가이드: ✏입력 (녹색)
R3_INPUT_FILL = "C8E6C9"
R3_INPUT_TEXT = "✏입력"
# R4 영역색: 매출과 같은 황금 (날짜성 데이터)
R4_DATE_COLOR = "B08F36"

# 부서별 추가 컬럼 정의
NEW_COLS = {
    "설계팀": ["제안서\n완료일", "설계\n완료일", "가공의뢰\n입고일"],
    "소프트웨어팀": ["제안서\n완료일", "설계\n완료일"],
    "구매팀": ["외주\n입고일"],
    "베트남": ["가공\n입고일"],
}


def find_dept(filename):
    for d in NEW_COLS.keys():
        if d in filename:
            return d
    return None


def add_cols(file_path, dept):
    wb = load_workbook(file_path)
    ws = wb.active

    # 현재 끝 = 상태(끝-1), 메모(끝)
    cur_max = ws.max_column
    state_col = cur_max - 1   # 상태
    memo_col = cur_max         # 메모

    new_labels = NEW_COLS[dept]
    n_new = len(new_labels)

    # 상태 컬럼 직전(state_col 위치)에 신규 컬럼 n_new개 삽입
    ws.insert_cols(state_col, n_new)

    # 신규 컬럼 R3·R4 채우기
    for i, label in enumerate(new_labels):
        c = state_col + i
        # R3
        cell3 = ws.cell(3, c)
        cell3.value = R3_INPUT_TEXT
        cell3.fill = PatternFill("solid", fgColor=R3_INPUT_FILL)
        cell3.font = Font(name="맑은 고딕", size=8, bold=True, color="2E7D32")
        cell3.alignment = AL_C
        cell3.border = THIN
        # R4
        cell4 = ws.cell(4, c)
        cell4.value = label
        cell4.fill = PatternFill("solid", fgColor=R4_DATE_COLOR)
        cell4.font = Font(name="맑은 고딕", size=9, bold=True, color="FFFFFF")
        cell4.alignment = AL_C
        cell4.border = THIN

    # 컬럼 너비
    for i in range(n_new):
        c_letter = get_column_letter(state_col + i)
        ws.column_dimensions[c_letter].width = 12

    new_max = ws.max_column
    wb.save(file_path)
    return cur_max, new_max, new_labels


print("=" * 70)
print("Phase 2-A: 부서 입력 파일 마일스톤·입고일 컬럼 추가")
print("=" * 70)

import glob
files = sorted(
    glob.glob(str(V3 / "01_검사기_완제품/KNK_*_*_입력_*.xlsx"))
    + glob.glob(str(V3 / "02_자동화_완제품/KNK_*_*_입력_*.xlsx"))
)

for fp in files:
    fname = Path(fp).name
    dept = find_dept(fname)
    if dept is None:
        print(f"  [SKIP] {fname} (변경 대상 아님)")
        continue
    before, after, labels = add_cols(fp, dept)
    print(f"  ✓ {fname}: {before}열 → {after}열 (+{len(labels)}: {labels})")

print("=" * 70)
print("완료. 다음 단계: apply_standard.py 실행으로 스펙 정규화")
