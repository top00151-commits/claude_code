"""
╔══════════════════════════════════════════════════════════════╗
║  KNK PMS V4 — 10_대시보드 gen_dashboard.py                    ║
║  5개 소스 통합 집계 → 경영현황 대시보드 생성                   ║
║  실행: python gen_dashboard.py                                ║
╚══════════════════════════════════════════════════════════════╝
"""

import os, sys, glob, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import *


# ═══════════════════════════════════════════════════════════════
# 데이터 수집
# ═══════════════════════════════════════════════════════════════

def _val(ws, r, c):
    v = ws.cell(row=r, column=c).value
    return v if v is not None else ""

def _num(ws, r, c):
    v = ws.cell(row=r, column=c).value
    try: return float(v)
    except (TypeError, ValueError): return 0


# R4 헤더 텍스트 기반 컬럼 인덱스 룩업 (캐시 포함)
_header_cache = {}
def _header_index(ws, header_text):
    """ws의 R4에서 header_text와 부분 일치하는 첫 컬럼 번호 반환 (없으면 None)."""
    key = (id(ws), header_text)
    if key in _header_cache:
        return _header_cache[key]
    norm_target = str(header_text).replace("\n", "").replace(" ", "").strip()
    found = None
    for c in range(1, ws.max_column + 1):
        h = ws.cell(4, c).value
        if not h:
            continue
        norm = str(h).replace("\n", "").replace(" ", "").strip()
        if norm_target in norm:
            found = c
            break
    _header_cache[key] = found
    return found

def _val_by_header(ws, r, header_text):
    c = _header_index(ws, header_text)
    return _val(ws, r, c) if c else ""

def _num_by_header(ws, r, header_text):
    c = _header_index(ws, header_text)
    return _num(ws, r, c) if c else 0


def _collect_product_data(source_dir, type_name):
    """완제품(01, 02) PMS에서 프로젝트 데이터 수집"""
    projects = []
    pms_files = glob.glob(os.path.join(source_dir, "KNK_*_PMS_*.xlsx"))
    if not pms_files:
        return projects

    fp = pms_files[0]
    try:
        wb = load_workbook(fp, data_only=True)
        ws = wb["1_프로젝트등록"]
        for r in range(5, ws.max_row + 1):
            code = _val(ws, r, 2)
            if not code:
                continue
            # v2026.04c: 01 검사기(23열, C6=제품구분)와 02 자동화(22열)의 컬럼 위치가 다름.
            # R4 헤더 텍스트 기반 룩업으로 구조 차이 흡수.
            projects.append({
                "type": type_name,
                "code": str(code),
                "sj":       _val_by_header(ws, r, "수주번호"),
                "cust":     _val_by_header(ws, r, "고객사"),
                "cust_pic": _val_by_header(ws, r, "고객사담당자"),
                "prodtype": _val_by_header(ws, r, "제품구분"),    # 01만 존재 (02는 None)
                "model":    _val_by_header(ws, r, "모델"),
                "prod":     _val_by_header(ws, r, "품명"),
                "amount":   _num_by_header(ws, r, "금액"),
                "currency": _val_by_header(ws, r, "통화"),
                "status":   _val_by_header(ws, r, "진행상태"),
                "prog":     _num_by_header(ws, r, "진척률"),
                "due":      _val_by_header(ws, r, "납기일"),
                "dday":     _num_by_header(ws, r, "D-day"),
                "update": "",
            })
        wb.close()
    except Exception as e:
        print(f"  [WARN] {type_name} 읽기 실패: {e}")

    return projects


def _collect_parts_data(source_dir, type_name):
    """부품출하(03, 04) 마스터에서 데이터 수집"""
    projects = []
    master_files = glob.glob(os.path.join(source_dir, "KNK_*_마스터_*.xlsx"))
    if not master_files:
        return projects

    fp = master_files[0]
    try:
        wb = load_workbook(fp, data_only=True)
        ws = wb.active
        for r in range(5, ws.max_row + 1):
            code = _val(ws, r, 2)
            if not code:
                continue
            projects.append({
                "type": type_name,
                "code": str(code),
                "sj": _val(ws, r, 3),
                "cust": _val(ws, r, 4),
                "prod": _val(ws, r, 5),
                "amount": _num(ws, r, 16),
                "currency": "KRW",
                "status": _val(ws, r, 19) if ws.max_column >= 19 else "",
                "prog": _num(ws, r, 11),
                "due": "",
                "dday": 0,
                "update": "",
                "inv_usd": _num(ws, r, 17),
            })
        wb.close()
    except Exception as e:
        print(f"  [WARN] {type_name} 읽기 실패: {e}")

    return projects


# ═══════════════════════════════════════════════════════════════
# 매출 사이클 데이터 수집 (v2026.04e 신규)
# ═══════════════════════════════════════════════════════════════
def _parse_date(v):
    """문자열/datetime → date. 실패 시 None."""
    import datetime as _dt
    if v is None or v == "":
        return None
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%y-%m-%d", "%y/%m/%d"):
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _collect_invoice_data(source_dir, type_name):
    """4_완료이력(발행 예정·미수금) + 8_매출마감(완납) 데이터 수집"""
    result = {"pending_invoice": [], "pending_payment": [], "settled": []}
    pms_files = glob.glob(os.path.join(source_dir, "KNK_*_PMS_*.xlsx"))
    if not pms_files:
        return result
    fp = pms_files[0]
    try:
        wb = load_workbook(fp, data_only=True)

        def read_sheet(ws, bucket_key):
            for r in range(5, ws.max_row + 1):
                code = _val(ws, r, 2)
                if not code:
                    continue
                inv = _val_by_header(ws, r, "계산서")
                pay = _val_by_header(ws, r, "입금일")
                row = {
                    "type": type_name,
                    "code": str(code),
                    "sj":       _val_by_header(ws, r, "수주번호"),
                    "cust":     _val_by_header(ws, r, "고객사"),
                    "prod":     _val_by_header(ws, r, "품명"),
                    "amount":   _num_by_header(ws, r, "금액"),
                    "currency": _val_by_header(ws, r, "통화") or "KRW",
                    "due":      _val_by_header(ws, r, "납기일"),
                    "status":   _val_by_header(ws, r, "진행상태"),
                    "inv_date": _parse_date(inv),
                    "pay_date": _parse_date(pay),
                }
                if bucket_key == "archive":
                    if not row["inv_date"]:
                        result["pending_invoice"].append(row)
                    elif not row["pay_date"]:
                        result["pending_payment"].append(row)
                elif bucket_key == "settled":
                    result["settled"].append(row)

        if "4_완료이력" in wb.sheetnames:
            read_sheet(wb["4_완료이력"], "archive")
        if "8_매출마감" in wb.sheetnames:
            read_sheet(wb["8_매출마감"], "settled")
        wb.close()
    except Exception as e:
        print(f"  [WARN] {type_name} 매출 데이터 읽기 실패: {e}")
    return result


def _collect_active_for_forecast(source_dir, type_name):
    """진행중 프로젝트 — 납기 기반 월별 매출 예측용"""
    active = []
    pms_files = glob.glob(os.path.join(source_dir, "KNK_*_PMS_*.xlsx"))
    if not pms_files:
        return active
    fp = pms_files[0]
    try:
        wb = load_workbook(fp, data_only=True)
        if "1_프로젝트등록" in wb.sheetnames:
            ws = wb["1_프로젝트등록"]
            for r in range(5, ws.max_row + 1):
                code = _val(ws, r, 2)
                if not code:
                    continue
                status = _val_by_header(ws, r, "진행상태")
                if status in ("납품완료", "완료", "취소"):
                    continue
                due = _parse_date(_val_by_header(ws, r, "납기일"))
                active.append({
                    "type": type_name,
                    "code": str(code),
                    "cust": _val_by_header(ws, r, "고객사"),
                    "prod": _val_by_header(ws, r, "품명"),
                    "amount": _num_by_header(ws, r, "금액"),
                    "currency": _val_by_header(ws, r, "통화") or "KRW",
                    "due": due,
                    "status": status,
                })
        wb.close()
    except Exception as e:
        print(f"  [WARN] {type_name} 진행중 데이터 읽기 실패: {e}")
    return active


def _collect_outsource_data():
    """v3.0: 부서 파일에서 가공의뢰·외주 입고일 수집
    - 01 설계팀: '가공의뢰\\n입고일' (의뢰부서='설계')
    - 01 구매팀: '외주\\n입고일' (의뢰부서='구매')
    - 01 베트남: '가공\\n입고일' (의뢰부서='베트남')
    - 02 동일 + 02 설계팀
    """
    rows = []
    targets = [
        ("검사기", "01_검사기_완제품", "설계팀",  "설계", "가공의뢰\n입고일"),
        ("검사기", "01_검사기_완제품", "구매팀",  "구매", "외주\n입고일"),
        ("검사기", "01_검사기_완제품", "베트남",  "베트남", "가공\n입고일"),
        ("자동화", "02_자동화_완제품", "설계팀",  "설계", "가공의뢰\n입고일"),
        ("자동화", "02_자동화_완제품", "구매팀",  "구매", "외주\n입고일"),
        ("자동화", "02_자동화_완제품", "베트남",  "베트남", "가공\n입고일"),
    ]
    for biz, biz_dir, dept_name, owner, target_header in targets:
        source_dir = SOURCES.get(biz_dir)
        if not source_dir:
            continue
        pattern = os.path.join(source_dir, f"KNK_*{dept_name}_입력_*.xlsx")
        files = glob.glob(pattern)
        if not files:
            continue
        fp = files[0]
        try:
            wb = load_workbook(fp, data_only=True)
            ws = wb.active
            # target_header 컬럼 인덱스 찾기
            target_col = None
            for c in range(1, ws.max_column + 1):
                h = ws.cell(4, c).value
                if h == target_header:
                    target_col = c
                    break
            if not target_col:
                wb.close()
                continue
            # 데이터 행 수집
            for r in range(5, ws.max_row + 1):
                code = _val(ws, r, 2)
                if not code:
                    continue
                rows.append({
                    "biz": biz,
                    "code": str(code),
                    "sj":   _val(ws, r, 3),
                    "cust": _val(ws, r, 4),
                    "prod": _val(ws, r, 6),
                    "owner": owner,
                    "in_date": _val(ws, r, target_col),
                })
            wb.close()
        except Exception as e:
            print(f"  [WARN] 외주 수집 실패 ({fp}): {e}")
    return rows


def _build_outsource_sheet(wb, rows):
    """외주현황 시트 — 가공의뢰·외주 입고 추적 (v3.0)"""
    sheet_name = "외주현황"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = TAB_VIEW

    labels = ["NO", "사업부", "관리코드", "수주번호", "고객사", "품명",
              "의뢰부서", "입고일", "상태"]
    mc = len(labels)
    setup_r1(ws, f"㈜케이엔케이 │ 외주현황 (가공의뢰·외주 입고 추적) │ {YEAR}", mc)
    r3 = {c: "auto" for c in range(1, mc + 1)}
    apply_r3_guide(ws, mc, r3)
    r4 = {c: "id" for c in range(1, mc + 1)}
    r4[7] = "dept"     # 의뢰부서
    r4[8] = "payment"  # 입고일
    r4[9] = "status"
    apply_r4_header(ws, mc, labels, r4)

    # 정렬: 미입고 먼저, 그 안에서 의뢰부서·관리코드 순
    rows_sorted = sorted(rows, key=lambda x: (
        0 if not x["in_date"] else 1,    # 미입고 먼저
        x["owner"],
        x["code"],
    ))

    row = 5
    for idx, r in enumerate(rows_sorted, 1):
        ws.cell(row=row, column=1).value = idx
        ws.cell(row=row, column=2).value = r["biz"]
        ws.cell(row=row, column=3).value = r["code"]
        ws.cell(row=row, column=4).value = r["sj"]
        ws.cell(row=row, column=5).value = r["cust"]
        ws.cell(row=row, column=6).value = r["prod"]
        ws.cell(row=row, column=7).value = r["owner"]
        ws.cell(row=row, column=8).value = r["in_date"]
        ws.cell(row=row, column=9).value = "입고완료" if r["in_date"] else "진행중"
        row += 1

    format_data_rows(ws, mc, row_start=5, row_end=max(row, 200))
    apply_protection(ws, mc, r3)
    auto_fit_columns(ws)


def _collect_consumables_data(source_dir):
    """소모품(05) 출하관리에서 데이터 수집"""
    projects = []
    files = glob.glob(os.path.join(source_dir, "KNK_*_출하대장_*.xlsx"))
    if not files:
        return projects

    fp = files[0]
    try:
        wb = load_workbook(fp, data_only=True)
        ws = wb["출하관리"]
        for r in range(5, ws.max_row + 1):
            if not _val(ws, r, 3):  # 품명 없으면 스킵
                continue
            projects.append({
                "type": "소모품",
                "code": str(_val(ws, r, 1)),
                "sj": _val(ws, r, 2),
                "cust": "",
                "prod": _val(ws, r, 3),
                "amount": _num(ws, r, 12),
                "currency": _val(ws, r, 10),
                "status": _val(ws, r, 25),
                "prog": 0,
                "due": "",
                "dday": 0,
                "update": "",
            })
        wb.close()
    except Exception as e:
        print(f"  [WARN] 소모품 읽기 실패: {e}")

    return projects


# ═══════════════════════════════════════════════════════════════
# 대시보드 생성
# ═══════════════════════════════════════════════════════════════

def generate():
    """경영현황 대시보드 생성"""
    print("=" * 60)
    print(f"  KNK PMS V4 — 경영현황 대시보드 생성")
    print(f"  시각: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # 데이터 수집
    all_projects = []
    type_summaries = {}

    # 2026-04-15: 03/04 부품출하 보류 — 01/02/05 3-모듈로 축소
    source_configs = [
        ("01_검사기_완제품", "검사기_완제품", "product"),
        ("02_자동화_완제품", "자동화_완제품", "product"),
        ("05_소모품", "소모품", "consumables"),
    ]

    for key, type_name, category in source_configs:
        src = SOURCES.get(key, "")
        if not os.path.exists(src):
            print(f"  [SKIP] {key} 폴더 없음")
            continue

        if category == "product":
            data = _collect_product_data(src, type_name)
        elif category == "parts":
            data = _collect_parts_data(src, type_name)
        else:
            data = _collect_consumables_data(src)

        all_projects.extend(data)

        # 유형별 집계
        active = [p for p in data if p.get("status") not in ("납품완료", "완료", "취소")]
        done = [p for p in data if p.get("status") in ("납품완료", "완료")]
        total_krw = sum(p["amount"] for p in data if p.get("currency", "KRW") != "USD")
        total_usd = sum(p["amount"] for p in data if p.get("currency") == "USD")
        inv_usd = sum(p.get("inv_usd", 0) for p in data)
        progs = [p["prog"] for p in active if p["prog"] and p["prog"] > 0]
        avg_prog = sum(progs) / len(progs) if progs else 0

        type_summaries[type_name] = {
            "active": len(active), "done": len(done), "total": len(data),
            "krw": total_krw, "usd": total_usd, "inv_usd": inv_usd,
            "avg_prog": avg_prog,
            "delayed": sum(1 for p in active if p.get("dday", 0) and p["dday"] < 0),
        }
        print(f"  {type_name}: {len(data)}건 수집")

    # 대시보드 생성
    fp = dashboard_path()
    wb = Workbook()

    # ── 시트1: 매출총괄 ──
    ws1 = wb.active
    ws1.title = "매출총괄"
    ws1.sheet_properties.tabColor = TAB_VIEW

    mc1 = SUMMARY_MAX_COL
    setup_r1(ws1, f"㈜케이엔케이 │ 경영현황 │ 매출총괄 │ {YEAR}", mc1)
    r3 = {c: "auto" for c in range(1, mc1 + 1)}
    apply_r3_guide(ws1, mc1, r3)
    r4 = {c: "id" for c in range(1, mc1 + 1)}
    apply_r4_header(ws1, mc1, SUMMARY_LABELS, r4)

    row = 5
    for type_name, s in type_summaries.items():
        ws1.cell(row=row, column=1).value = type_name
        ws1.cell(row=row, column=2).value = s["active"]
        ws1.cell(row=row, column=3).value = s["done"]
        ws1.cell(row=row, column=4).value = s["total"]
        ws1.cell(row=row, column=5).value = s["krw"]
        ws1.cell(row=row, column=5).number_format = ACCT
        ws1.cell(row=row, column=5).alignment = AL_R
        ws1.cell(row=row, column=6).value = s["usd"]
        ws1.cell(row=row, column=6).number_format = ACCT_USD
        ws1.cell(row=row, column=6).alignment = AL_R
        ws1.cell(row=row, column=7).value = s["inv_usd"]
        ws1.cell(row=row, column=7).number_format = ACCT_USD
        ws1.cell(row=row, column=7).alignment = AL_R
        ws1.cell(row=row, column=8).value = s["avg_prog"] / 100 if s["avg_prog"] else None
        ws1.cell(row=row, column=8).number_format = PCT
        ws1.cell(row=row, column=9).value = s["delayed"]
        row += 1

    # 합계행
    ws1.cell(row=row, column=1).value = "합계"
    ws1.cell(row=row, column=1).font = FT_BOLD
    for c in range(2, 10):
        total = sum(ws1.cell(row=r, column=c).value or 0 for r in range(5, row))
        if c == 8:  # 평균 진척률은 평균
            vals = [ws1.cell(row=r, column=c).value or 0 for r in range(5, row) if ws1.cell(row=r, column=c).value]
            ws1.cell(row=row, column=c).value = sum(vals) / len(vals) if vals else None
            ws1.cell(row=row, column=c).number_format = PCT
        else:
            ws1.cell(row=row, column=c).value = total
        if c in (5,):
            ws1.cell(row=row, column=c).number_format = ACCT
            ws1.cell(row=row, column=c).alignment = AL_R
        if c in (6, 7):
            ws1.cell(row=row, column=c).number_format = ACCT_USD
            ws1.cell(row=row, column=c).alignment = AL_R

    format_data_rows(ws1, mc1, money_cols=[5], money_usd_cols=[6, 7], pct_cols=[8], row_start=5, row_end=row)
    apply_protection(ws1, mc1, r3, row_end=row + 10)
    auto_fit_columns(ws1)

    # ── 시트2: 전체 프로젝트 목록 ──
    ws2 = wb.create_sheet("프로젝트목록")
    ws2.sheet_properties.tabColor = TAB_VIEW

    mc2 = PROJECT_MAX_COL
    setup_r1(ws2, f"㈜케이엔케이 │ 경영현황 │ 전체 프로젝트 │ {YEAR}", mc2)
    r3_2 = {c: "auto" for c in range(1, mc2 + 1)}
    apply_r3_guide(ws2, mc2, r3_2)
    r4_2 = {c: "id" for c in range(1, mc2 + 1)}
    apply_r4_header(ws2, mc2, PROJECT_LABELS, r4_2)

    row = 5
    for idx, p in enumerate(all_projects, 1):
        ws2.cell(row=row, column=1).value = idx
        ws2.cell(row=row, column=2).value = p["type"]
        ws2.cell(row=row, column=3).value = p["code"]
        ws2.cell(row=row, column=4).value = p["sj"]
        ws2.cell(row=row, column=5).value = p["cust"]
        ws2.cell(row=row, column=6).value = p["prod"]
        ws2.cell(row=row, column=7).value = p["amount"]
        ws2.cell(row=row, column=7).number_format = ACCT
        ws2.cell(row=row, column=7).alignment = AL_R
        ws2.cell(row=row, column=8).value = p["status"]
        ws2.cell(row=row, column=9).value = p["prog"] / 100 if p["prog"] else None
        ws2.cell(row=row, column=9).number_format = PCT
        ws2.cell(row=row, column=10).value = p["due"]
        ws2.cell(row=row, column=11).value = p["dday"] if p["dday"] else None
        ws2.cell(row=row, column=12).value = p["update"]
        row += 1

    format_data_rows(ws2, mc2, money_cols=[7], pct_cols=[9], row_start=5, row_end=max(row, 2000))
    apply_protection(ws2, mc2, r3_2)
    auto_fit_columns(ws2)

    # ══════════════════════════════════════════════════════════
    # v2026.04e 매출 사이클 시트 4개 추가
    # ══════════════════════════════════════════════════════════
    invoice_01 = _collect_invoice_data(SOURCES["01_검사기_완제품"], "검사기")
    invoice_02 = _collect_invoice_data(SOURCES["02_자동화_완제품"], "자동화")
    active_01 = _collect_active_for_forecast(SOURCES["01_검사기_완제품"], "검사기")
    active_02 = _collect_active_for_forecast(SOURCES["02_자동화_완제품"], "자동화")

    all_pending_inv = invoice_01["pending_invoice"] + invoice_02["pending_invoice"]
    all_pending_pay = invoice_01["pending_payment"] + invoice_02["pending_payment"]
    all_settled     = invoice_01["settled"] + invoice_02["settled"]
    all_active      = active_01 + active_02

    # 시트: 매출 관련 4개
    _build_pending_invoice_sheet(wb, all_pending_inv)
    _build_pending_payment_sheet(wb, all_pending_pay)
    _build_settled_ytd_sheet(wb, all_settled)
    _build_forecast_sheet(wb, all_active)

    # 시트: 이번주 출하 Gantt (검사기 D-day [-7~+3])
    _build_week_shipment_sheet(wb, all_active)

    # 시트: 진행률 매트릭스 (전체 프로젝트 × 부서)
    _build_progress_matrix_sheet(wb, all_active)

    # 시트: 외주현황 (v3.0 — 설계·구매·베트남 부서 파일에서 입고일 수집)
    outsource_rows = _collect_outsource_data()
    _build_outsource_sheet(wb, outsource_rows)

    # 시트0: 오늘의 현황 (맨 앞에 삽입)
    _build_today_sheet(wb, type_summaries, all_pending_inv, all_pending_pay, all_active)

    wb.save(fp)
    print(f"\n  대시보드 저장: {fp}")
    print(f"  총 {len(all_projects)}건 프로젝트 + 매출사이클 4시트 + Gantt + 매트릭스")
    print("=" * 60)
    return fp


# ═══════════════════════════════════════════════════════════════
# 매출 사이클 시트 빌더 (v2026.04e)
# ═══════════════════════════════════════════════════════════════
def _build_pending_invoice_sheet(wb, rows):
    """매출_발행예정 — 납품완료 + 계산서 미발행"""
    ws = wb.create_sheet("매출_발행예정")
    ws.sheet_properties.tabColor = TAB_VIEW
    labels = ["NO", "매출유형", "관리코드", "수주번호", "고객사", "품명",
              "금액", "통화", "납기일", "납기\n경과일", "비고"]
    mc = len(labels)
    setup_r1(ws, f"㈜케이엔케이 │ 경영현황 │ 계산서 발행 예정 │ {YEAR}", mc)
    r3 = {c: "auto" for c in range(1, mc + 1)}
    apply_r3_guide(ws, mc, r3)
    r4 = {c: ("id" if c <= 1 else "sales") for c in range(1, mc + 1)}
    apply_r4_header(ws, mc, labels, r4)

    import datetime as _dt
    today = _dt.date.today()
    # 경과일 기준 정렬 (오래된 것 위)
    rows = sorted(rows, key=lambda x: _parse_date(x.get("due")) or _dt.date.max)

    total_krw = total_usd = 0
    row = 5
    for i, r_ in enumerate(rows, 1):
        due = _parse_date(r_.get("due"))
        elapsed = (today - due).days if due else None
        ws.cell(row, 1).value = i
        ws.cell(row, 2).value = r_["type"]
        ws.cell(row, 3).value = r_["code"]
        ws.cell(row, 4).value = r_["sj"]
        ws.cell(row, 5).value = r_["cust"]
        ws.cell(row, 6).value = r_["prod"]
        ws.cell(row, 7).value = r_["amount"]
        ws.cell(row, 7).number_format = ACCT_USD if str(r_["currency"]).upper() == "USD" else ACCT
        ws.cell(row, 8).value = r_["currency"]
        ws.cell(row, 9).value = r_["due"]
        ws.cell(row, 10).value = elapsed
        ws.cell(row, 11).value = ""
        if str(r_["currency"]).upper() == "USD":
            total_usd += r_["amount"]
        else:
            total_krw += r_["amount"]
        row += 1

    # 합계 행
    ws.cell(row, 1).value = "합계"
    ws.cell(row, 1).font = FT_BOLD
    ws.cell(row, 6).value = f"총 {len(rows)}건"
    ws.cell(row, 6).font = FT_BOLD
    ws.cell(row, 7).value = total_krw if total_krw else total_usd
    ws.cell(row, 7).number_format = ACCT_USD if total_usd else ACCT
    ws.cell(row, 7).font = FT_BOLD
    ws.cell(row, 7).alignment = AL_R

    format_data_rows(ws, mc, money_cols=[7], row_start=5, row_end=row)
    apply_protection(ws, mc, r3)
    auto_fit_columns(ws)


def _build_pending_payment_sheet(wb, rows):
    """매출_미수금 — 계산서 발행 + 입금 대기 (경과일 분류)"""
    ws = wb.create_sheet("매출_미수금")
    ws.sheet_properties.tabColor = TAB_VIEW
    labels = ["NO", "매출유형", "관리코드", "수주번호", "고객사", "품명",
              "금액", "통화", "계산서\n발행일", "경과일", "구간", "비고"]
    mc = len(labels)
    setup_r1(ws, f"㈜케이엔케이 │ 경영현황 │ 미수금 │ {YEAR}", mc)
    r3 = {c: "auto" for c in range(1, mc + 1)}
    apply_r3_guide(ws, mc, r3)
    r4 = {c: ("id" if c <= 1 else "sales") for c in range(1, mc + 1)}
    apply_r4_header(ws, mc, labels, r4)

    import datetime as _dt
    today = _dt.date.today()
    rows = sorted(rows, key=lambda x: x.get("inv_date") or _dt.date.max)

    total_krw = total_usd = 0
    row = 5
    for i, r_ in enumerate(rows, 1):
        inv_date = r_.get("inv_date")
        elapsed = (today - inv_date).days if inv_date else None
        if elapsed is None:
            bucket = ""
        elif elapsed <= 30:
            bucket = "30일 이내"
        elif elapsed <= 60:
            bucket = "30~60일"
        elif elapsed <= 90:
            bucket = "60~90일"
        else:
            bucket = "90일 초과 ⚠"
        ws.cell(row, 1).value = i
        ws.cell(row, 2).value = r_["type"]
        ws.cell(row, 3).value = r_["code"]
        ws.cell(row, 4).value = r_["sj"]
        ws.cell(row, 5).value = r_["cust"]
        ws.cell(row, 6).value = r_["prod"]
        ws.cell(row, 7).value = r_["amount"]
        ws.cell(row, 7).number_format = ACCT_USD if str(r_["currency"]).upper() == "USD" else ACCT
        ws.cell(row, 8).value = r_["currency"]
        ws.cell(row, 9).value = inv_date.strftime("%Y-%m-%d") if inv_date else ""
        ws.cell(row, 10).value = elapsed
        ws.cell(row, 11).value = bucket
        if str(r_["currency"]).upper() == "USD":
            total_usd += r_["amount"]
        else:
            total_krw += r_["amount"]
        row += 1

    ws.cell(row, 1).value = "합계"
    ws.cell(row, 1).font = FT_BOLD
    ws.cell(row, 6).value = f"총 {len(rows)}건"
    ws.cell(row, 6).font = FT_BOLD
    ws.cell(row, 7).value = total_krw if total_krw else total_usd
    ws.cell(row, 7).number_format = ACCT_USD if total_usd else ACCT
    ws.cell(row, 7).font = FT_BOLD
    ws.cell(row, 7).alignment = AL_R

    format_data_rows(ws, mc, money_cols=[7], row_start=5, row_end=row)
    apply_protection(ws, mc, r3)
    auto_fit_columns(ws)


def _build_settled_ytd_sheet(wb, rows):
    """매출_완납YTD — 올해 완납된 매출 월별 집계"""
    ws = wb.create_sheet("매출_완납YTD")
    ws.sheet_properties.tabColor = TAB_VIEW
    labels = ["월", "검사기\n건수", "검사기\n금액", "자동화\n건수", "자동화\n금액",
              "합계\n건수", "합계 금액(KRW)", "합계 금액(USD)"]
    mc = len(labels)
    setup_r1(ws, f"㈜케이엔케이 │ 경영현황 │ 완납 YTD │ {YEAR}", mc)
    r3 = {c: "auto" for c in range(1, mc + 1)}
    apply_r3_guide(ws, mc, r3)
    r4 = {c: "id" for c in range(1, mc + 1)}
    apply_r4_header(ws, mc, labels, r4)

    import datetime as _dt
    # 월별 집계 (올해만)
    monthly = {m: {"t_count": 0, "t_amt": 0, "m_count": 0, "m_amt": 0,
                   "krw": 0, "usd": 0} for m in range(1, 13)}
    for r_ in rows:
        pay = r_.get("pay_date")
        if not pay or pay.year != YEAR:
            continue
        m = pay.month
        if r_["type"] == "검사기":
            monthly[m]["t_count"] += 1
            monthly[m]["t_amt"] += r_["amount"]
        elif r_["type"] == "자동화":
            monthly[m]["m_count"] += 1
            monthly[m]["m_amt"] += r_["amount"]
        if str(r_["currency"]).upper() == "USD":
            monthly[m]["usd"] += r_["amount"]
        else:
            monthly[m]["krw"] += r_["amount"]

    row = 5
    tot_t_c = tot_t_a = tot_m_c = tot_m_a = tot_krw = tot_usd = 0
    for m in range(1, 13):
        data = monthly[m]
        ws.cell(row, 1).value = f"{m:02d}월"
        ws.cell(row, 2).value = data["t_count"]
        ws.cell(row, 3).value = data["t_amt"]
        ws.cell(row, 3).number_format = ACCT
        ws.cell(row, 4).value = data["m_count"]
        ws.cell(row, 5).value = data["m_amt"]
        ws.cell(row, 5).number_format = ACCT
        ws.cell(row, 6).value = data["t_count"] + data["m_count"]
        ws.cell(row, 7).value = data["krw"]
        ws.cell(row, 7).number_format = ACCT
        ws.cell(row, 8).value = data["usd"]
        ws.cell(row, 8).number_format = ACCT_USD
        tot_t_c += data["t_count"]; tot_t_a += data["t_amt"]
        tot_m_c += data["m_count"]; tot_m_a += data["m_amt"]
        tot_krw += data["krw"];     tot_usd += data["usd"]
        row += 1

    ws.cell(row, 1).value = "합계"
    ws.cell(row, 1).font = FT_BOLD
    ws.cell(row, 2).value = tot_t_c
    ws.cell(row, 3).value = tot_t_a; ws.cell(row, 3).number_format = ACCT
    ws.cell(row, 4).value = tot_m_c
    ws.cell(row, 5).value = tot_m_a; ws.cell(row, 5).number_format = ACCT
    ws.cell(row, 6).value = tot_t_c + tot_m_c
    ws.cell(row, 7).value = tot_krw; ws.cell(row, 7).number_format = ACCT
    ws.cell(row, 8).value = tot_usd; ws.cell(row, 8).number_format = ACCT_USD
    for c in (1, 2, 3, 4, 5, 6, 7, 8):
        ws.cell(row, c).font = FT_BOLD

    format_data_rows(ws, mc, money_cols=[3, 5, 7], money_usd_cols=[8], row_start=5, row_end=row)
    apply_protection(ws, mc, r3)
    auto_fit_columns(ws)


def _build_forecast_sheet(wb, active_rows):
    """매출_예측캘린더 — 진행중 프로젝트의 월별 예상 매출 (6개월)

    패턴:
      - 검사기: 납기월 = 예상 계산서월, 계산서월 + 1.5개월 = 예상 입금월
      - 자동화: 납기월 = 예상 계산서월, 계산서월 + 1개월 = 예상 입금월
    """
    ws = wb.create_sheet("매출_예측캘린더")
    ws.sheet_properties.tabColor = TAB_VIEW
    import datetime as _dt
    today = _dt.date.today()

    # 6개월 월 컬럼 (이번달 ~ +5)
    months = []
    cur = _dt.date(today.year, today.month, 1)
    for i in range(6):
        y, m = cur.year, cur.month + i
        while m > 12:
            y += 1; m -= 12
        months.append(f"{y}-{m:02d}")

    labels = ["구분", "지표"] + months + ["합계"]
    mc = len(labels)
    setup_r1(ws, f"㈜케이엔케이 │ 경영현황 │ 매출 예측 캘린더 (6개월) │ {YEAR}", mc)
    r3 = {c: "auto" for c in range(1, mc + 1)}
    apply_r3_guide(ws, mc, r3)
    r4 = {c: "id" for c in range(1, mc + 1)}
    apply_r4_header(ws, mc, labels, r4)

    # 검사기·자동화 / 예상 계산서·예상 입금 분리
    grids = {
        ("검사기", "계산서"): {m: 0 for m in months},
        ("검사기", "입금"):   {m: 0 for m in months},
        ("자동화", "계산서"): {m: 0 for m in months},
        ("자동화", "입금"):   {m: 0 for m in months},
    }

    def _add_months(base, n):
        y = base.year; m = base.month + n
        while m > 12: y += 1; m -= 12
        while m < 1:  y -= 1; m += 12
        return _dt.date(y, m, 1)

    for p in active_rows:
        if not p["due"]:
            continue
        inv_date = _dt.date(p["due"].year, p["due"].month, 1)
        if p["type"] == "검사기":
            pay_date = _add_months(inv_date, 1)  # 검사기 약 45일, 월 단위로 +1
        else:
            pay_date = _add_months(inv_date, 1)
        inv_key = f"{inv_date.year}-{inv_date.month:02d}"
        pay_key = f"{pay_date.year}-{pay_date.month:02d}"
        if inv_key in grids[(p["type"], "계산서")]:
            grids[(p["type"], "계산서")][inv_key] += p["amount"]
        if pay_key in grids[(p["type"], "입금")]:
            grids[(p["type"], "입금")][pay_key] += p["amount"]

    row = 5
    for (tp, metric) in [("검사기", "계산서"), ("검사기", "입금"),
                         ("자동화", "계산서"), ("자동화", "입금")]:
        ws.cell(row, 1).value = tp
        ws.cell(row, 2).value = f"예상 {metric}"
        total = 0
        for i, m in enumerate(months, 3):
            v = grids[(tp, metric)][m]
            ws.cell(row, i).value = v if v else None
            ws.cell(row, i).number_format = ACCT
            total += v
        ws.cell(row, mc).value = total if total else None
        ws.cell(row, mc).number_format = ACCT
        ws.cell(row, mc).font = FT_BOLD
        row += 1

    format_data_rows(ws, mc, money_cols=list(range(3, mc + 1)), row_start=5, row_end=row)
    apply_protection(ws, mc, r3)
    auto_fit_columns(ws)


def _build_today_sheet(wb, summary, pending_inv, pending_pay, active):
    """오늘의 현황 — 경영자 매일 아침 1페이지 요약 (시트 최상단)"""
    ws = wb.create_sheet("오늘의현황", 0)
    ws.sheet_properties.tabColor = TAB_VIEW

    import datetime as _dt
    today = _dt.date.today()

    labels = ["구분", "지표", "값", "비고"]
    mc = len(labels)
    setup_r1(ws, f"㈜케이엔케이 │ 오늘의 현황 │ {today.strftime('%Y-%m-%d')}", mc)
    r3 = {c: "auto" for c in range(1, mc + 1)}
    apply_r3_guide(ws, mc, r3)
    r4 = {c: "id" for c in range(1, mc + 1)}
    apply_r4_header(ws, mc, labels, r4)

    inv_krw = sum(r["amount"] for r in pending_inv if str(r["currency"]).upper() != "USD")
    inv_usd = sum(r["amount"] for r in pending_inv if str(r["currency"]).upper() == "USD")
    pay_krw = sum(r["amount"] for r in pending_pay if str(r["currency"]).upper() != "USD")
    pay_usd = sum(r["amount"] for r in pending_pay if str(r["currency"]).upper() == "USD")

    # 90일+ 미수금 (경보)
    overdue = 0
    overdue_amt = 0
    for r in pending_pay:
        inv_d = r.get("inv_date")
        if inv_d and (today - inv_d).days > 90:
            overdue += 1
            overdue_amt += r["amount"]

    # 진행중 요약
    n_active = len(active)
    delayed = sum(1 for p in active if p.get("due") and (today - p["due"]).days > 0)
    amt_active_krw = sum(p["amount"] for p in active if str(p["currency"]).upper() != "USD")

    rows_data = [
        ("영업 현황", "진행중 프로젝트", f"{n_active}건", f"지연 {delayed}건"),
        ("영업 현황", "검사기", f"{summary.get('검사기_완제품', {}).get('active', 0)}건", ""),
        ("영업 현황", "자동화", f"{summary.get('자동화_완제품', {}).get('active', 0)}건", ""),
        ("", "", "", ""),
        ("매출 현황", "계산서 발행 예정 (KRW)", inv_krw, f"{len(pending_inv)}건"),
        ("매출 현황", "계산서 발행 예정 (USD)", inv_usd, ""),
        ("매출 현황", "미수금 (KRW)", pay_krw, f"{len(pending_pay)}건"),
        ("매출 현황", "미수금 (USD)", pay_usd, ""),
        ("경보", "⚠ 90일 초과 미수금", overdue_amt, f"{overdue}건" if overdue else "없음"),
        ("", "", "", ""),
        ("진행 예측", "진행중 총 수주 (KRW)", amt_active_krw, "납기 도달 시 매출 예정"),
    ]

    row = 5
    for (kw1, kw2, v, note) in rows_data:
        ws.cell(row, 1).value = kw1
        ws.cell(row, 2).value = kw2
        if isinstance(v, (int, float)) and v:
            ws.cell(row, 3).value = v
            ws.cell(row, 3).number_format = ACCT_USD if "USD" in kw2 else ACCT
        else:
            ws.cell(row, 3).value = v
        ws.cell(row, 4).value = note
        row += 1

    # 마지막 섹션 — 오늘 납기 건
    ws.cell(row, 1).value = "오늘/이번주 납기"
    ws.cell(row, 1).font = FT_BOLD
    row += 1
    today_due = [p for p in active if p.get("due") and 0 <= (p["due"] - today).days <= 7]
    today_due.sort(key=lambda p: p["due"])
    for p in today_due[:10]:
        d_days = (p["due"] - today).days
        ws.cell(row, 1).value = p["type"]
        ws.cell(row, 2).value = f"{p['code']} — {p['prod']}"
        ws.cell(row, 3).value = p["amount"]
        ws.cell(row, 3).number_format = ACCT
        ws.cell(row, 4).value = f"D{-d_days:+d}일 ({p['due'].strftime('%m-%d')})"
        row += 1
    if not today_due:
        ws.cell(row, 1).value = "— 없음 —"
        row += 1

    format_data_rows(ws, mc, money_cols=[3], row_start=5, row_end=row)
    apply_protection(ws, mc, r3)
    auto_fit_columns(ws)


# ═══════════════════════════════════════════════════════════════
# 이번주 출하 Gantt (검사기 D-day [-7,+3]) — v2026.04f
# ═══════════════════════════════════════════════════════════════
def _build_week_shipment_sheet(wb, active_rows):
    """이번주 검사기 출하 Gantt (7일 창, D-day [-7~+3])"""
    import datetime as _dt
    from openpyxl.styles import PatternFill as _PF, Font as _F, Alignment as _A, Border as _B, Side as _S

    ws = wb.create_sheet("이번주_출하", 1)   # 오늘의현황(0) 다음
    ws.sheet_properties.tabColor = TAB_VIEW

    today = _dt.date.today()
    # 7일 창 = 오늘 ~ +6일 (주말 포함)
    days = [today + _dt.timedelta(days=i) for i in range(7)]
    weekdays = ["월","화","수","목","금","토","일"]

    # 필터: 검사기 AND D-day ∈ [-7, +3]
    targets = []
    for p in active_rows:
        if p["type"] != "검사기":
            continue
        if not p["due"]:
            continue
        dday = (today - p["due"]).days   # 양수=지연
        if -7 <= dday <= 3:
            p["_dday"] = dday
            targets.append(p)
    targets.sort(key=lambda x: x["due"])

    # 헤더: 좌측 8 + 7일
    left_labels = ["분류", "관리코드", "고객사", "품명", "수량", "금액", "유형", "수주일"]
    day_labels = [f"{d.month}/{d.day}\n({weekdays[d.weekday()]})" for d in days]
    labels = left_labels + day_labels
    mc = len(labels)

    setup_r1(ws, f"㈜케이엔케이 │ 이번주 검사기 출하 │ {days[0].strftime('%m/%d(%a)')} ~ {days[-1].strftime('%m/%d(%a)')}", mc)
    r3 = {c: "auto" for c in range(1, mc + 1)}
    apply_r3_guide(ws, mc, r3)
    r4 = {c: ("id" if c <= 1 else "sales" if c <= len(left_labels) else "po") for c in range(1, mc + 1)}
    apply_r4_header(ws, mc, labels, r4)

    # 색상
    gray_fill   = _PF("solid", fgColor="BDBDBD")       # 작업기간
    today_fill  = _PF("solid", fgColor="FFF9C4")       # 오늘 컬럼
    weekend_fill = _PF("solid", fgColor="EEEEEE")      # 주말 컬럼
    overdue_fill = _PF("solid", fgColor="FFCDD2")      # 지연 배경
    green_font  = _F(name="맑은 고딕", size=9, bold=True, color="2E7D32")  # 출하
    red_font    = _F(name="맑은 고딕", size=9, bold=True, color="C62828")  # 납품
    black_font  = _F(name="맑은 고딕", size=9, bold=True, color="1A1A1A")  # 시료
    center      = _A(horizontal="center", vertical="center", wrap_text=True)

    # 요일 헤더 배경 (주말)
    for i, d in enumerate(days):
        col = len(left_labels) + 1 + i
        if d == today:
            ws.cell(4, col).fill = today_fill
            ws.cell(4, col).font = _F(name="맑은 고딕", size=9, bold=True, color="1A1A1A")
        elif d.weekday() >= 5:
            ws.cell(4, col).fill = weekend_fill

    row = 5
    for p in targets:
        is_overdue = p["_dday"] > 0
        is_done = p["status"] in ("납품완료", "완료")

        # 좌측 고정 컬럼
        ws.cell(row, 1).value = p["type"]
        ws.cell(row, 2).value = p["code"]
        ws.cell(row, 3).value = p["cust"]
        ws.cell(row, 4).value = p["prod"]
        ws.cell(row, 5).value = 1   # 수량은 원본에 없으므로 기본 1
        ws.cell(row, 6).value = p["amount"]
        ws.cell(row, 6).number_format = ACCT
        ws.cell(row, 6).alignment = AL_R
        # PO유형·수주일은 원본 데이터에 없어 공란
        ws.cell(row, 7).value = p.get("potype", "")
        ws.cell(row, 8).value = ""

        # 달력 영역
        for i, d in enumerate(days):
            col = len(left_labels) + 1 + i
            cell = ws.cell(row, col)
            cell.alignment = center

            # 배경 우선순위: 오늘 > 주말 > 기본
            if d == today:
                cell.fill = today_fill
            elif d.weekday() >= 5:
                cell.fill = weekend_fill

            # 납기일 = 이 날짜면 라벨 표시
            if d == p["due"]:
                if is_done:
                    cell.value = "납품"
                    cell.font = red_font
                else:
                    cell.value = "출하"
                    cell.font = green_font
                if is_overdue and not is_done:
                    cell.fill = overdue_fill
            # 납기일이 지난 경우 (지연건) 오늘 셀에 "지연 Nd" 표기
            elif is_overdue and d == today:
                cell.value = f"지연 +{p['_dday']}일"
                cell.font = red_font
                cell.fill = overdue_fill
            # 수주일~납기일 범위는 회색
            # (수주일 정보 없어서 임시: 오늘 창 시작 ~ 납기일 사이 회색)
            elif p["due"] and days[0] <= d <= p["due"]:
                if cell.fill.fgColor.value in (None, "00000000"):
                    cell.fill = gray_fill
        row += 1

    # 합계 행
    if targets:
        ws.cell(row, 1).value = "합계"
        ws.cell(row, 1).font = FT_BOLD
        total_amt = sum(p["amount"] for p in targets)
        ws.cell(row, 6).value = total_amt
        ws.cell(row, 6).number_format = ACCT
        ws.cell(row, 6).font = FT_BOLD
        ws.cell(row, 6).alignment = AL_R
        ws.cell(row, 4).value = f"총 {len(targets)}건"
        ws.cell(row, 4).font = FT_BOLD
    else:
        ws.cell(row, 1).value = "이번주 출하 예정 건 없음"
        ws.cell(row, 1).font = FT_BOLD
    row += 1

    # format_data_rows는 fill을 덮어쓰므로 건너뛰고 폰트·테두리만 수동 적용
    from shared.styles import THIN as _THIN, FT_DATA as _FT_DATA, apply_auto_widths
    for r_ in range(5, row + 1):
        for c_ in range(1, mc + 1):
            cell = ws.cell(r_, c_)
            cell.border = _THIN
            if not cell.font.bold:
                cell.font = _FT_DATA
    apply_protection(ws, mc, r3)
    # 스펙 §8.1 표준 너비 룩업 적용 (A1 병합 타이틀 영향 배제)
    apply_auto_widths(ws, mc, row_start=4, row_end=row, min_w=5, max_w=40)
    # 날짜 컬럼은 "지연 +3일" 같은 값 수용 위해 12로 상향
    from openpyxl.utils import get_column_letter
    for i in range(len(days)):
        col_letter = get_column_letter(len(left_labels) + 1 + i)
        ws.column_dimensions[col_letter].width = 12


# ═══════════════════════════════════════════════════════════════
# 진행률 매트릭스 (전체 프로젝트 × 부서)
# ═══════════════════════════════════════════════════════════════
def _build_progress_matrix_sheet(wb, active_rows):
    """프로젝트 × 부서 진척률 매트릭스"""
    import datetime as _dt
    from openpyxl.styles import PatternFill as _PF, Font as _F, Alignment as _A
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("진행률_매트릭스", 2)   # 이번주_출하 다음
    ws.sheet_properties.tabColor = TAB_VIEW

    today = _dt.date.today()

    # 각 모듈의 2_진행현황 시트에서 부서별 소계 읽기
    def _read_progress(source_dir, type_name, depts):
        """2_진행현황 시트에서 관리코드별 부서 소계 읽기"""
        data = {}   # {code: {dept: subtotal_pct}}
        pms_files = glob.glob(os.path.join(source_dir, "KNK_*_PMS_*.xlsx"))
        if not pms_files:
            return data
        try:
            wb_p = load_workbook(pms_files[0], data_only=True)
            if "2_진행현황" not in wb_p.sheetnames:
                return data
            ws_p = wb_p["2_진행현황"]
            # 헤더에서 각 부서 소계 컬럼 위치 찾기
            dept_cols = {}
            for c in range(1, ws_p.max_column + 1):
                h = str(ws_p.cell(4, c).value or "").replace("\n", "")
                for d in depts:
                    if d in h and "소계" in h:
                        dept_cols[d] = c
                        break
            # 데이터 읽기
            for r in range(5, ws_p.max_row + 1):
                code = ws_p.cell(r, 2).value
                if not code:
                    continue
                row_data = {}
                for dept, col in dept_cols.items():
                    v = ws_p.cell(r, col).value
                    if isinstance(v, (int, float)):
                        row_data[dept] = v * 100 if v <= 1 else v
                    else:
                        row_data[dept] = None
                data[str(code)] = row_data
            wb_p.close()
        except Exception as e:
            print(f"  [WARN] {type_name} 진행현황 읽기 실패: {e}")
        return data

    # v3.0: 가공팀 폐기 (의뢰 부서가 추적)
    depts_01 = ["설계팀","검사기팀","개발혁신팀","품질팀","제조기술1팀","구매팀","베트남"]
    depts_02 = ["설계팀","전장설계팀","소프트웨어팀","구매팀","제조기술2팀","베트남"]
    prog_01 = _read_progress(SOURCES["01_검사기_완제품"], "검사기", depts_01)
    prog_02 = _read_progress(SOURCES["02_자동화_완제품"], "자동화", depts_02)

    # PMS 부서 컬럼 (제외 여부) — 별도 파일 열어서 확인
    def _read_exclusions(source_dir, depts):
        """1_프로젝트등록에서 관리코드별 '제외' 부서 집합 반환"""
        excluded = {}   # {code: set of excluded depts}
        pms_files = glob.glob(os.path.join(source_dir, "KNK_*_PMS_*.xlsx"))
        if not pms_files:
            return excluded
        try:
            wb_p = load_workbook(pms_files[0], data_only=True)
            ws_p = wb_p["1_프로젝트등록"]
            dept_cols = {}
            for c in range(1, ws_p.max_column + 1):
                h = str(ws_p.cell(4, c).value or "").replace("\n", "").strip()
                for d in depts:
                    if h == d:
                        dept_cols[d] = c
                        break
            for r in range(5, ws_p.max_row + 1):
                code = ws_p.cell(r, 2).value
                if not code:
                    continue
                exc = set()
                for dept, col in dept_cols.items():
                    v = str(ws_p.cell(r, col).value or "").strip()
                    if v == "제외":
                        exc.add(dept)
                excluded[str(code)] = exc
            wb_p.close()
        except Exception as e:
            print(f"  [WARN] 제외 정보 읽기 실패: {e}")
        return excluded

    exc_01 = _read_exclusions(SOURCES["01_검사기_완제품"], depts_01)
    exc_02 = _read_exclusions(SOURCES["02_자동화_완제품"], depts_02)

    # 모든 부서 통합 컬럼 목록 (01+02 union, 순서 유지)
    all_depts_set = set(depts_01 + depts_02)
    all_depts = [d for d in depts_01 if d in all_depts_set]
    for d in depts_02:
        if d not in all_depts:
            all_depts.append(d)
    # v3.0 all_depts: [설계팀, 검사기팀, 개발혁신팀, 품질팀, 제조기술1팀, 구매팀, 베트남, 전장설계팀, 소프트웨어팀, 제조기술2팀]

    # 헤더
    left_labels = ["NO", "분류", "관리코드", "고객사", "품명", "담당영업", "납기", "D-day", "전체%"]
    labels = left_labels + all_depts
    mc = len(labels)

    setup_r1(ws, f"㈜케이엔케이 │ 진행률 매트릭스 (전체) │ {today.strftime('%Y-%m-%d')}", mc)
    r3 = {c: "auto" for c in range(1, mc + 1)}
    apply_r3_guide(ws, mc, r3)
    r4 = {c: ("id" if c <= 3 else "sales" if c <= 6 else "status" if c <= 9 else "dept") for c in range(1, mc + 1)}
    apply_r4_header(ws, mc, labels, r4)

    # 색상 팔레트 — 진척률 5단계
    def _prog_fill(pct):
        if pct is None:
            return None
        if pct >= 100:
            return _PF("solid", fgColor="2E7D32")  # 진녹
        if pct >= 71:
            return _PF("solid", fgColor="A5D6A7")  # 연녹
        if pct >= 31:
            return _PF("solid", fgColor="FFF176")  # 노랑
        if pct >= 1:
            return _PF("solid", fgColor="FFE0B2")  # 연노랑
        return _PF("solid", fgColor="E0E0E0")      # 회색 (0%)

    def _dday_fill(dday):
        if dday is None:
            return None
        if dday > 0:
            return _PF("solid", fgColor="C62828")   # 빨강 (지연)
        if dday >= -3:
            return _PF("solid", fgColor="FB8C00")   # 주황 (임박)
        return _PF("solid", fgColor="43A047")       # 녹 (여유)

    white_font = _F(name="맑은 고딕", size=9, bold=True, color="FFFFFF")

    # 정렬: D-day 양수(지연) → 임박 → 여유 순
    active_rows_sorted = sorted(active_rows, key=lambda p: (
        -9999 if not p["due"] else (today - p["due"]).days * -1   # D-day 음수일수록 위 (임박)
    ))
    # 역순 수정: 지연(양수) 먼저, 그 다음 D-day 작은(음수로 큰) 순
    active_rows_sorted = sorted(active_rows, key=lambda p: (
        -(today - p["due"]).days if p["due"] else 9999   # 오름차순: 양수 지연이 가장 작음? → 역순 필요
    ), reverse=True)
    # 간단히: D-day 크기 (오늘-due)가 큰 것 위 (지연 먼저)
    active_rows_sorted = sorted(active_rows,
        key=lambda p: -((today - p["due"]).days if p["due"] else -9999))

    row = 5
    for i, p in enumerate(active_rows_sorted, 1):
        dday = (today - p["due"]).days if p["due"] else None

        ws.cell(row, 1).value = i
        ws.cell(row, 2).value = p["type"]
        ws.cell(row, 3).value = p["code"]
        ws.cell(row, 4).value = p["cust"]
        ws.cell(row, 5).value = p["prod"]
        ws.cell(row, 6).value = ""   # 담당영업 — 간이 데이터에 없어 공란
        ws.cell(row, 7).value = p["due"].strftime("%Y-%m-%d") if p["due"] else ""
        # D-day
        ws.cell(row, 8).value = dday
        ws.cell(row, 8).number_format = '+0"일";-0"일";"0일"'
        fill = _dday_fill(dday)
        if fill:
            ws.cell(row, 8).fill = fill
            ws.cell(row, 8).font = white_font

        # 부서 진척률 조회
        progs = prog_01.get(p["code"]) if p["type"] == "검사기" else prog_02.get(p["code"])
        excs = exc_01.get(p["code"], set()) if p["type"] == "검사기" else exc_02.get(p["code"], set())
        # 전체 평균
        if progs:
            vals = [v for d, v in progs.items() if v is not None and d not in excs]
            avg = sum(vals) / len(vals) if vals else None
        else:
            avg = None
        ws.cell(row, 9).value = (avg / 100) if avg is not None else None
        ws.cell(row, 9).number_format = PCT

        # 각 부서 셀
        valid_depts = depts_01 if p["type"] == "검사기" else depts_02
        for di, dept in enumerate(all_depts):
            col = len(left_labels) + 1 + di
            if dept not in valid_depts:
                ws.cell(row, col).value = "—"
                continue
            if dept in excs:
                ws.cell(row, col).value = "제외"
                ws.cell(row, col).fill = _PF("solid", fgColor="F5F5F5")
                continue
            v = progs.get(dept) if progs else None
            if v is None:
                ws.cell(row, col).value = None
                ws.cell(row, col).fill = _PF("solid", fgColor="E0E0E0")
            else:
                ws.cell(row, col).value = v / 100
                ws.cell(row, col).number_format = PCT
                fill = _prog_fill(v)
                if fill:
                    ws.cell(row, col).fill = fill
                # 100%일 때 흰 글자
                if v >= 100:
                    ws.cell(row, col).font = white_font
        row += 1

    # format_data_rows 대신 폰트·테두리만 적용 (fill 보존)
    from shared.styles import THIN as _THIN, FT_DATA as _FT_DATA, apply_auto_widths
    for r_ in range(5, row + 1):
        for c_ in range(1, mc + 1):
            cell = ws.cell(r_, c_)
            cell.border = _THIN
            if not cell.font.bold:
                cell.font = _FT_DATA
    apply_protection(ws, mc, r3)
    # 스펙 §8.1 표준 너비 룩업 + 데이터 기반 자동 조정
    apply_auto_widths(ws, mc, row_start=4, row_end=row, min_w=5, max_w=40)
    # 부서 컬럼은 "100%" + 헤더 "제조기술1팀" 수용해 최소 10으로 보정
    for di in range(len(all_depts)):
        col_letter = get_column_letter(len(left_labels) + 1 + di)
        cur = ws.column_dimensions[col_letter].width or 0
        if cur < 10:
            ws.column_dimensions[col_letter].width = 10


if __name__ == "__main__":
    generate()
