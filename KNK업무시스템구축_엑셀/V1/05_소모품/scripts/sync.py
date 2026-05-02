"""
╔══════════════════════════════════════════════════════════════╗
║  KNK PMS V4 — 05_소모품 sync.py                               ║
║  출하관리 INV 자동계산                                         ║
║  실행: python sync.py                                         ║
╚══════════════════════════════════════════════════════════════╝
"""

import os, sys, datetime, logging
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import *


def _setup_logger():
    os.makedirs(log_dir(), exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = os.path.join(log_dir(), f"sync_{TYPE_CODE}_{ts}.log")
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)


def _num(ws, r, c):
    v = ws.cell(row=r, column=c).value
    try: return float(v)
    except (TypeError, ValueError): return 0


def _val(ws, r, c):
    v = ws.cell(row=r, column=c).value
    return v if v is not None else ""


def sync_all():
    log = _setup_logger()
    log.info("=" * 60)
    log.info(f"  KNK PMS V4 — {TYPE_NAME} 동기화 시작")
    log.info("=" * 60)

    fp = consumables_path()
    if not os.path.exists(fp):
        log.error(f"파일 없음: {fp}")
        return

    wb = load_workbook(fp)
    ws = wb["출하관리"]

    changed = False
    for r in range(5, ws.max_row + 1):
        qty = _num(ws, r, LI["수량"])
        uprice = _num(ws, r, LI["단가"])
        rate = _num(ws, r, LI["상승률"])
        exrate = _num(ws, r, LI["환율"])
        duty_pct = _num(ws, r, LI["DUTY"])

        if not qty and not uprice:
            continue

        # 금액 = 수량 × 단가
        amount = qty * uprice
        fmt_cell(ws, r, LI["금액"], amount, is_money=True)

        if rate:
            inv_up = uprice * rate
            inv_amt = amount * rate
            fmt_cell(ws, r, LI["INV단가"], inv_up, is_money=True)
            fmt_cell(ws, r, LI["INV금액"], inv_amt, is_money=True)

            if exrate and exrate > 0:
                inv_up_usd = inv_up / exrate
                inv_amt_usd = inv_amt / exrate
                fmt_cell(ws, r, LI["INV단가USD"], inv_up_usd, is_money=True, is_usd=True)
                fmt_cell(ws, r, LI["INV금액USD"], inv_amt_usd, is_money=True, is_usd=True)

                if duty_pct:
                    fmt_cell(ws, r, LI["관세KRW"], inv_amt * duty_pct, is_money=True)
                    fmt_cell(ws, r, LI["관세USD"], inv_amt_usd * duty_pct, is_money=True, is_usd=True)

            changed = True

    if changed:
        wb.save(fp)
        log.info("  INV 자동계산 완료")
    else:
        wb.save(fp)
        log.info("  계산 대상 없음")

    log.info("=" * 60)
    log.info("  동기화 완료!")
    log.info("=" * 60)


if __name__ == "__main__":
    sync_all()
