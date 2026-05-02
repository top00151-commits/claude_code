"""
26년작업일정표.xlsx → 01 검사기 / 02 자동화 PMS 1회성 임포트 스크립트.

분류 룰:
  - 관리코드가 NNNT{YYMM} → 01 검사기
  - 관리코드가 NNNM{YYMM} → 02 자동화
  - 관리코드 공란 → 담당영업 기준:
      "이해림", "배승진" → 02 자동화
      그 외(이현, 오경환, 안지연 등) → 01 검사기

제품구분(01만 해당) 자동 추론 (품명·모델명 키워드):
  - "SENSOR" / "센서" → SENSOR
  - "TSP" → TSP
  - "PBA" → PBA
  - 그 외 → 공란 (사용자가 직접 선택)

실행: python 98_문서/import_schedule.py
그 후: 각 모듈 폴더에서 python scripts/sync.py + apply_standard.py
"""
import os
import re
import sys
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE = os.path.dirname(SCRIPT_DIR)  # HAIST_WORKS_baby

SCHEDULE_FP = os.path.join(SCRIPT_DIR, "26년작업일정표.xlsx")
PMS_01 = os.path.join(BASE, "01_검사기_완제품", "KNK_검사기_완제품_PMS_2026.xlsx")
PMS_02 = os.path.join(BASE, "02_자동화_완제품", "KNK_자동화_완제품_PMS_2026.xlsx")

AUTO_SALES = {"이해림", "배승진"}   # 02 자동화 담당영업
PRODTYPE_RULES = [
    (("SENSOR", "센서"), "SENSOR"),
    (("TSP",),           "TSP"),
    (("PBA",),           "PBA"),
]


def _norm(v):
    """셀 값 정규화 — datetime → YYYY-MM-DD, 숫자는 그대로."""
    import datetime
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        return v.strip() or None
    return v


def _infer_prodtype(model, prod):
    text = f"{model or ''} {prod or ''}".upper()
    for keywords, code in PRODTYPE_RULES:
        if any(k.upper() in text for k in keywords):
            return code
    return None  # 공란


def _read_schedule():
    wb = load_workbook(SCHEDULE_FP, data_only=True)
    ws = wb["26년4월"]
    rows = []
    for r in range(6, ws.max_row + 1):
        vals = [_norm(ws.cell(r, c).value) for c in range(1, 14)]
        cust, cust_pic, code, model, prod, qty, up, amt, po, sj, due, status, sales = vals
        if not cust and not prod:
            continue
        rows.append({
            "cust": cust, "cust_pic": cust_pic, "code": code,
            "model": model, "prod": prod, "qty": qty, "up": up, "amt": amt,
            "po": po, "sj": sj, "due": due, "status": status, "sales": sales,
        })
    return rows


def _classify(rows):
    t_rows, m_rows = [], []
    pat = re.compile(r"^\d{3}([TM])\d{4}$")
    for row in rows:
        c = str(row["code"] or "").strip()
        m = pat.match(c)
        if m:
            (t_rows if m.group(1) == "T" else m_rows).append(row)
        else:
            row["code"] = None   # 빈 문자열 정규화
            if str(row["sales"] or "").strip() in AUTO_SALES:
                m_rows.append(row)
            else:
                t_rows.append(row)
    return t_rows, m_rows


def _write_pms(fp, rows, is_01):
    """PMS 1_프로젝트등록 시트 R15+에 데이터 기록.
    sync가 정렬하므로 R15부터 순서대로 넣으면 된다.
    """
    wb = load_workbook(fp)
    ws = wb["1_프로젝트등록"]

    # 컬럼 인덱스 — 01은 23열(제품구분 C6), 02는 22열(제품구분 없음)
    if is_01:
        C = dict(CODE=2, SJNUM=3, CUST=4, CUSTPIC=5, PRODTYPE=6,
                 MODEL=7, PROD=8, POTYPE=9, STAGE=10, SALES=11, PM=12,
                 QTY=13, CUR=14, UP=15, TOT=16, SJ=17, DUE=18,
                 LOG=19, PROG=20, DDAY=21, STATUS=22, NOTE=23)
    else:
        C = dict(CODE=2, SJNUM=3, CUST=4, CUSTPIC=5,
                 MODEL=6, PROD=7, POTYPE=8, STAGE=9, SALES=10, PM=11,
                 QTY=12, CUR=13, UP=14, TOT=15, SJ=16, DUE=17,
                 LOG=18, PROG=19, DDAY=20, STATUS=21, NOTE=22)

    start_row = 15
    for i, row in enumerate(rows):
        r = start_row + i
        if row["code"]:
            ws.cell(r, C["CODE"]).value = row["code"]
        # 수주번호는 sync가 채번
        ws.cell(r, C["CUST"]).value    = row["cust"]
        ws.cell(r, C["CUSTPIC"]).value = row["cust_pic"]
        if is_01:
            ws.cell(r, C["PRODTYPE"]).value = _infer_prodtype(row["model"], row["prod"])
        ws.cell(r, C["MODEL"]).value = row["model"]
        ws.cell(r, C["PROD"]).value  = row["prod"]
        ws.cell(r, C["POTYPE"]).value = row["po"]
        ws.cell(r, C["STAGE"]).value  = "수주확정"    # 일정표 전제: 수주확정 이상
        ws.cell(r, C["SALES"]).value  = row["sales"]
        if row["qty"] is not None:
            ws.cell(r, C["QTY"]).value = row["qty"]
        ws.cell(r, C["CUR"]).value = "KRW"   # 일정표 금액은 모두 원화
        if row["up"] is not None:
            ws.cell(r, C["UP"]).value = row["up"]
        if row["amt"] is not None:
            ws.cell(r, C["TOT"]).value = row["amt"]
        ws.cell(r, C["SJ"]).value  = row["sj"]
        ws.cell(r, C["DUE"]).value = row["due"]
        ws.cell(r, C["STATUS"]).value = row["status"]

    wb.save(fp)
    return len(rows)


def main():
    print("=" * 60)
    print("  26년작업일정표 → PMS 임포트")
    print("=" * 60)

    if not os.path.exists(SCHEDULE_FP):
        print(f"  ✗ 원본 없음: {SCHEDULE_FP}")
        sys.exit(1)

    rows = _read_schedule()
    print(f"\n  원본 데이터: {len(rows)}건")

    t_rows, m_rows = _classify(rows)
    print(f"  01 검사기: {len(t_rows)}건")
    print(f"  02 자동화: {len(m_rows)}건")

    print(f"\n▶ 01 검사기 PMS 기록")
    n1 = _write_pms(PMS_01, t_rows, is_01=True)
    print(f"  ✓ {n1}건 기록: {os.path.basename(PMS_01)}")

    print(f"\n▶ 02 자동화 PMS 기록")
    n2 = _write_pms(PMS_02, m_rows, is_01=False)
    print(f"  ✓ {n2}건 기록: {os.path.basename(PMS_02)}")

    print("\n" + "=" * 60)
    print("  다음 단계:")
    print("    1. cd 01_검사기_완제품 && python scripts/sync.py && python scripts/apply_standard.py")
    print("    2. cd 02_자동화_완제품 && python scripts/sync.py && python scripts/apply_standard.py")
    print("=" * 60)


if __name__ == "__main__":
    main()
