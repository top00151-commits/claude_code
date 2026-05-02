"""
v2026.04d → v2026.04e 마이그레이션 (1회성)

변경사항:
- 1_프로젝트등록 및 4_완료이력에 컬럼 2개 추가 (계산서 발행일, 입금일)
- 새 시트 8_매출마감 추가 (빈 시트, apply_standard가 서식 적용)

실행 전제: config.py·build.py·sync.py·apply_standard.py 모두 v2026.04e로 수정됨.
실행: python 98_문서/migrate_v2026_04e.py
그 후: 각 모듈 sync.py + apply_standard.py 실행
"""
import os
import sys
import shutil
import datetime
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE = os.path.dirname(SCRIPT_DIR)

MODULES = [
    ("01_검사기_완제품", "KNK_검사기_완제품_PMS_2026.xlsx", 31, 33),   # old_max, new_max
    ("02_자동화_완제품", "KNK_자동화_완제품_PMS_2026.xlsx", 29, 31),
]

SHEETS_TO_EXTEND = ["1_프로젝트등록", "4_완료이력"]


def _backup(fp):
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = fp.replace(".xlsx", f"_backup_v2026_04d_{ts}.xlsx")
    shutil.copy2(fp, backup)
    print(f"  ✓ 백업: {os.path.basename(backup)}")
    return backup


def _extend_sheet(ws, old_max, new_max):
    """시트에 컬럼 2개 추가 (라벨만) — apply_standard가 서식·메모 적용 예정"""
    if old_max == new_max:
        return
    # R4 헤더에 2개 라벨 추가 (old_max + 1 ~ new_max)
    ws.cell(4, old_max + 1).value = "계산서\n발행일"
    ws.cell(4, old_max + 2).value = "입금일"
    return new_max - old_max


def _add_settled_sheet(wb):
    """8_매출마감 시트 추가 (빈 시트, apply_standard가 채움)"""
    if "8_매출마감" in wb.sheetnames:
        return False
    wb.create_sheet("8_매출마감")
    return True


def migrate_module(mod_name, pms_file, old_max, new_max):
    fp = os.path.join(BASE, mod_name, pms_file)
    if not os.path.exists(fp):
        print(f"  ✗ 파일 없음: {fp}")
        return

    print(f"\n▶ {mod_name} ({pms_file})")
    wb = load_workbook(fp)
    sample_ws = wb[SHEETS_TO_EXTEND[0]]
    actual_max = sample_ws.max_column

    if actual_max == new_max:
        print(f"  ⚠ 이미 {new_max}열 구조 — 마이그레이션 스킵")
        wb.close()
        return
    if actual_max != old_max:
        print(f"  ⚠ 예상외 컬럼수 {actual_max} (old 기대 {old_max}) — 수동 확인 필요")
        wb.close()
        return

    _backup(fp)

    # 컬럼 2개 추가 (1_프로젝트등록, 4_완료이력)
    for sh in SHEETS_TO_EXTEND:
        if sh in wb.sheetnames:
            n = _extend_sheet(wb[sh], old_max, new_max)
            print(f"  ✓ {sh}: {old_max}열 → {new_max}열 (+{n or 0}열)")

    # 8_매출마감 시트 신설
    if _add_settled_sheet(wb):
        print(f"  ✓ 8_매출마감 시트 신설")

    wb.save(fp)
    print(f"  ✓ 저장 완료")


def main():
    print("=" * 60)
    print(f"  v2026.04d → v2026.04e 마이그레이션")
    print(f"  - 1/4 시트에 컬럼 2개 추가 (계산서 발행일, 입금일)")
    print(f"  - 8_매출마감 시트 신설")
    print("=" * 60)

    for mod_name, pms_file, old_max, new_max in MODULES:
        migrate_module(mod_name, pms_file, old_max, new_max)

    print("\n" + "=" * 60)
    print("  다음 단계:")
    print("    1. cd 01_검사기_완제품 && python scripts/sync.py && python scripts/apply_standard.py")
    print("    2. cd 02_자동화_완제품 && python scripts/sync.py && python scripts/apply_standard.py")
    print("    3. cd 10_대시보드 && python scripts/gen_dashboard.py")
    print("=" * 60)


if __name__ == "__main__":
    main()
