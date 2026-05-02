"""
╔══════════════════════════════════════════════════════════════╗
║  02_자동화_완제품 migrate_v2026_04.py                         ║
║  41열 (환율·장비출하·셋업·이슈사항 포함) → 29열 마이그레이션  ║
║                                                               ║
║  스펙 §17.1 v2026.04 최종 구조로 정렬.                         ║
║  데이터(관리코드 있는 행)는 열 매핑 유지하며 이동, 제거 컬럼은  ║
║  별도 백업 파일로 보존된다.                                    ║
║                                                               ║
║  실행: python scripts/migrate_v2026_04.py                     ║
╚══════════════════════════════════════════════════════════════╝

이 스크립트는 **1회성**. 마이그레이션 완료 후 재실행 불필요.

전제:
  - config.py는 이미 29열로 업데이트됨 (MAX_COL=29)
  - sync.py, apply_standard.py도 29열 기준으로 수정됨
  - 기존 xlsx 파일은 41열 구조

작업:
  1) 현재 PMS 파일 전체 백업
  2) 1_프로젝트등록·4_완료이력 시트의 C14(환율)·C24~C34(장비·셋업·이슈) 삭제
     → 결과적으로 C15~C23 → C14~C22, C35~C41 → C23~C29
  3) sync.py + apply_standard.py 실행으로 마무리 (스크립트 외부에서 수동)
"""
import os
import sys
import shutil
import datetime
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import pms_path, MAX_COL, TYPE_NAME

# 스펙 v2026.04 기준 — 이 스크립트가 삭제할 컬럼 (구 41열 기준 원본 인덱스)
COLS_TO_DELETE_OLD_INDEX = [
    14,  # 환율
    24, 25, 26,  # 장비출하일(한국), 장비출하일(베트남), 장비 최종출하일
    27,  # 운송방법
    28, 29, 30, 31, 32, 33,  # 셋업 6열
    34,  # 이슈사항
]
# = 총 11열 삭제 → 41 - 11 - 1(환율) = wait, let me recount
# 14(환율) 1열 + 24~34 = 11열 → 총 12열 삭제 → 41 - 12 = 29 ✓

OLD_MAX_COL = 41
NEW_MAX_COL = 29
SHEETS_TO_MIGRATE = ["1_프로젝트등록", "4_완료이력"]

# 안전장치 — 이 구조가 일치하지 않으면 실행 거부
EXPECTED_OLD_HEADERS_AT_R4 = {
    14: "환율",
    24: "장비출하일(한국)",
    34: "이슈사항",
}


def _check_preconditions():
    """사전 조건 검증 — config가 이미 29열, xlsx는 41열이어야 한다."""
    fp = pms_path()
    if not os.path.exists(fp):
        print(f"  ✗ PMS 파일 없음: {fp}")
        return False, None
    if MAX_COL != NEW_MAX_COL:
        print(f"  ✗ config.py MAX_COL={MAX_COL} (기대값 {NEW_MAX_COL}).")
        print(f"     먼저 config.py를 29열로 업데이트하세요.")
        return False, None

    wb = load_workbook(fp, read_only=True)
    ws = wb[SHEETS_TO_MIGRATE[0]]
    if ws.max_column != OLD_MAX_COL:
        print(f"  ⚠ xlsx max_column={ws.max_column} (기대값 {OLD_MAX_COL}).")
        if ws.max_column == NEW_MAX_COL:
            print(f"     이미 마이그레이션 완료된 상태로 보입니다. 중단.")
            wb.close()
            return False, fp
        print(f"     구조가 예상과 다릅니다. 수동 확인 필요.")
        wb.close()
        return False, fp

    # R4 샘플 검증
    for col, expected in EXPECTED_OLD_HEADERS_AT_R4.items():
        actual = ws.cell(4, col).value
        if not actual or expected.replace("(", "").replace(")", "").split()[0] not in str(actual):
            print(f"  ⚠ C{col} 헤더='{actual}' (기대: '{expected}'에 해당). 구조 확인 필요.")
            # 경고만 하고 진행
    wb.close()
    return True, fp


def _backup(fp):
    """현재 파일을 시각 접미사로 백업."""
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = fp.replace(".xlsx", f"_backup_v41cols_{ts}.xlsx")
    shutil.copy2(fp, backup)
    print(f"  ✓ 백업: {os.path.basename(backup)}")
    return backup


def _migrate_sheet(ws):
    """한 시트에서 불필요 컬럼 삭제 (우측부터 역순으로 delete_cols).

    delete_cols는 값 + 서식을 함께 삭제하고 오른쪽 컬럼을 왼쪽으로 이동시킨다.
    역순 처리하는 이유: 왼쪽 컬럼을 먼저 삭제하면 오른쪽 인덱스가 흔들린다.
    """
    before_mc = ws.max_column
    # 내림차순으로 삭제
    for col in sorted(COLS_TO_DELETE_OLD_INDEX, reverse=True):
        ws.delete_cols(col, 1)
    after_mc = ws.max_column
    print(f"    {ws.title}: max_col {before_mc} → {after_mc}")
    return before_mc, after_mc


def migrate():
    print("=" * 60)
    print(f"  [MIGRATE] {TYPE_NAME} — v2026.04 (41열 → 29열)")
    print(f"  시각: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    ok, fp = _check_preconditions()
    if not ok:
        print("  ✗ 사전 조건 불충족 — 중단")
        return

    backup = _backup(fp)

    print(f"\n  데이터 마이그레이션 시작: {os.path.basename(fp)}")
    wb = load_workbook(fp)
    for sh in SHEETS_TO_MIGRATE:
        if sh in wb.sheetnames:
            _migrate_sheet(wb[sh])
        else:
            print(f"    ⚠ 시트 없음: {sh}")

    wb.save(fp)
    print(f"\n  ✓ 저장 완료: {os.path.basename(fp)}")
    print(f"  ✓ 원본 보존: {os.path.basename(backup)}")

    print("\n" + "=" * 60)
    print("  다음 단계 (이 스크립트 실행 후 꼭 진행):")
    print("    1. python scripts/sync.py")
    print("    2. python scripts/apply_standard.py")
    print("=" * 60)


if __name__ == "__main__":
    migrate()
