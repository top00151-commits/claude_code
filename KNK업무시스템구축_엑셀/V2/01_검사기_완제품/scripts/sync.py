"""
╔══════════════════════════════════════════════════════════════╗
║  KNK PMS V4 — 01_검사기_완제품 sync.py                        ║
║  데이터 동기화 · 자동계산 · 집계                               ║
║  실행: python sync.py                                         ║
╚══════════════════════════════════════════════════════════════╝

처리 흐름:
  ① PMS 1_프로젝트등록 읽기 + 관리코드 채번
  ② 부서입력 파일 → 진척률 수집 → PMS 2_진행현황 write-back
  ③ 물류상세 → INV 자동계산
  ④ 물류상세 → 마스터(조달진행현황) 집계 write-back
  ⑤ PMS 3_출하현황 집계 write-back
  ⑥ PMS 4_완료이력 아카이브
  ⑦ 5_매핑조회 · 6_관리코드발행대장 · 7_수주번호생성대장 갱신
"""

import os, sys, glob, datetime, logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import *

# ═══════════════════════════════════════════════════════════════
# 로깅 설정
# ═══════════════════════════════════════════════════════════════
def _setup_logger():
    os.makedirs(log_dir(), exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = os.path.join(log_dir(), f"sync_{TYPE_CODE}_{ts}.log")
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════
# 유틸리티
# ═══════════════════════════════════════════════════════════════
def _val(ws, r, c):
    """셀 값 읽기 (None → "")"""
    v = ws.cell(row=r, column=c).value
    return v if v is not None else ""

def _num(ws, r, c):
    """셀 숫자 읽기 (None/비숫자 → 0)"""
    v = ws.cell(row=r, column=c).value
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0

def _date_val(ws, r, c):
    """셀 날짜 읽기 (datetime / date / 문자열 모두 지원)"""
    v = ws.cell(row=r, column=c).value
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return parse_date(v)


# ─── 날짜 입력/표기 통일 (스킬 §13) ───────────────────────────────
import re as _re
def parse_date(val):
    """다양한 입력 형식을 date로 파싱 (스킬 13.3)"""
    if val is None or val == "":
        return None
    if isinstance(val, datetime.datetime): return val.date()
    if isinstance(val, datetime.date):     return val
    s = str(val).strip().replace(".", "-").replace("/", "-")
    if _re.fullmatch(r"\d{8}", s):
        try: return datetime.datetime.strptime(s, "%Y%m%d").date()
        except ValueError: pass
    if _re.fullmatch(r"\d{6}", s):
        try: return datetime.datetime.strptime(s, "%y%m%d").date()
        except ValueError: pass
    for fmt in ("%Y-%m-%d", "%y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try: return datetime.datetime.strptime(s, fmt).date()
        except ValueError: continue
    return None

def normalize_date(val):
    """어떤 입력이든 'YYYY-MM-DD' 문자열로 변환 (스킬 13.4)"""
    d = parse_date(val)
    return d.strftime("%Y-%m-%d") if d else val

def _normalize_date_cells(ws, date_cols, log):
    """sync STEP 1: 날짜 컬럼 일괄 정규화 (스킬 13.5)"""
    from openpyxl.styles import Alignment
    cnt = 0
    for r in range(5, ws.max_row + 1):
        for col in date_cols:
            cell = ws.cell(r, col)
            raw = cell.value
            if raw in (None, ""):
                continue
            norm = normalize_date(raw)
            if norm != raw:
                cell.value = norm
                cnt += 1
            cell.number_format = "@"
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if cnt:
        log.info(f"  날짜 정규화: {cnt}건")


# ═══════════════════════════════════════════════════════════════
# ① 관리코드 채번
# ═══════════════════════════════════════════════════════════════
def _assign_codes(ws, log):
    """관리코드·수주번호 자동채번 (KNK 표준 v2026.04)
    - 관리코드: {SEQ:03d}{BIZ}{YYMM} 형식 (예: 001T2604)
    - 수주번호: {BIZ}-{YYMMDD} 형식, 동일일자 중복 시 -N (예: T-260401, T-260401-2)
    """
    # ── 관리코드 채번 ──
    # 동일 YYMM 범위에서 최대 일련번호 조회
    existing_codes = set()
    max_seq = 0
    for r in range(5, ws.max_row + 1):
        code = _val(ws, r, C_CODE)
        if not code:
            continue
        existing_codes.add(str(code))
        s = str(code)
        # 신규 형식: 001T2604 (앞 3자리 SEQ, 4번째 BIZ, 끝 4자리 YYMM)
        if len(s) == 8 and s[3] == BIZ_CODE and s[4:] == YYMM:
            try:
                max_seq = max(max_seq, int(s[:3]))
            except ValueError:
                pass

    # ★ v2026.04g 채번 트리거: 영업단계가 명시적으로 STAGE_TRIGGERS에 있을 때만
    # 이전 버그: stage가 None/빈칸이면 str(None)="None" → STAGE_NO_CODE에 없어서 채번됨
    STAGE_TRIGGERS = {"수주확정", "납품", "개조", "A/S"}
    new_count = 0
    for r in range(5, ws.max_row + 1):
        if not _val(ws, r, C_CUST) and not _val(ws, r, C_PROD):
            continue
        stage = str(_val(ws, r, C_STAGE) or "").strip()
        if not _val(ws, r, C_CODE) and stage in STAGE_TRIGGERS:
            max_seq += 1
            new_code = f"{max_seq:03d}{BIZ_CODE}{YYMM}"
            ws.cell(row=r, column=C_CODE).value = new_code
            existing_codes.add(new_code)
            new_count += 1
            log.info(f"  관리코드 채번: R{r} → {new_code}")

    # ── 수주번호 자동채번 (수주일 YYMMDD 기반) ──
    # 기존 수주번호 수집 (중복 검사용)
    existing_sj = set()
    for r in range(5, ws.max_row + 1):
        sj = _val(ws, r, C_SJNUM)
        if sj:
            existing_sj.add(str(sj))

    for r in range(5, ws.max_row + 1):
        code = _val(ws, r, C_CODE)
        if not code or _val(ws, r, C_SJNUM):
            continue
        # 수주일 추출 (없으면 오늘 기준)
        sj_date = _date_val(ws, r, C_SJDATE)
        if not sj_date:
            sj_date = TODAY
        base = f"{BIZ_CODE}-{sj_date.strftime('%y%m%d')}"
        # 중복 시 -N 부여 (둘째부터 -1, -2, -3, ...)
        if base not in existing_sj:
            sj_num = base
        else:
            n = 1
            while f"{base}-{n}" in existing_sj:
                n += 1
            sj_num = f"{base}-{n}"
        ws.cell(row=r, column=C_SJNUM).value = sj_num
        existing_sj.add(sj_num)
        log.info(f"  수주번호 생성: R{r} → {sj_num}")

    # 금액 자동계산 (수량 × 단가)
    for r in range(5, ws.max_row + 1):
        qty = _num(ws, r, C_QTY)
        uprice = _num(ws, r, C_UPRICE)
        if qty and uprice:
            total = qty * uprice
            fmt_cell(ws, r, C_TOTAL, total, is_money=True,
                     is_usd=(str(_val(ws, r, C_CURRENCY)).upper() == "USD"))

    # D-day 계산
    for r in range(5, ws.max_row + 1):
        due = _date_val(ws, r, C_DUE)
        if due:
            dday = (TODAY - due).days
            ws.cell(row=r, column=C_DDAY).value = dday

    return new_count


# ═══════════════════════════════════════════════════════════════
# ①-b 관리코드 일관성 검증 (v2026.04g)
# ═══════════════════════════════════════════════════════════════
def _validate_code_consistency(ws, wb, log):
    """
    같은 관리코드의 모델·품명 일관성 검증·자동 보정.

    동작:
      - 같은 관리코드의 가장 위(R 작은) 행 = 권위(authoritative)
      - 빈 모델/품명 → 권위 값 자동 채움
      - 다른 모델/품명 → 노란색 fill + hover 메모(권위 행/값 표시) + 로그
      - 권위 행은 옅은 파랑 fill + "★ 권위" 메모 표시
      - 별도 시트 9_관리코드불일치에 일괄 보고서 생성

    반환: (fixed, warned, mismatches)
    """
    WARN_FILL = PatternFill("solid", fgColor="FFEB9C")     # 옅은 노랑 (불일치)
    AUTH_FILL = PatternFill("solid", fgColor="DDEBF7")     # 옅은 파랑 (권위 행)

    # 1) 권위 행 수집
    code_authority = {}  # {code: (row, model, prod)}
    for r in range(5, ws.max_row + 1):
        code = _val(ws, r, C_CODE)
        if not code:
            continue
        cs = str(code).strip()
        model = str(_val(ws, r, C_MODEL) or "").strip()
        prod  = str(_val(ws, r, C_PROD) or "").strip()
        if cs not in code_authority and (model or prod):
            code_authority[cs] = (r, model, prod)

    # 2) 같은 코드 다른 행 검증
    fixed = 0
    warned = 0
    mismatches = []  # 보고서용: (code, auth_r, kind, auth_val, cur_r, cur_val, cust)
    auth_rows_with_mismatch = set()  # 권위 행도 표시할 후보

    for r in range(5, ws.max_row + 1):
        code = _val(ws, r, C_CODE)
        if not code:
            continue
        cs = str(code).strip()
        if cs not in code_authority:
            continue
        auth_r, auth_model, auth_prod = code_authority[cs]
        if r == auth_r:
            continue

        cur_model = str(_val(ws, r, C_MODEL) or "").strip()
        cur_prod  = str(_val(ws, r, C_PROD) or "").strip()
        cust = _val(ws, r, C_CUST) or ""

        # 모델
        cell_m = ws.cell(r, C_MODEL)
        if not cur_model and auth_model:
            cell_m.value = auth_model
            fixed += 1
            log.info(f"  관리코드 {cs} 모델 자동 보정: R{r} ← '{auth_model}' (권위 R{auth_r})")
        elif cur_model and cur_model != auth_model:
            cell_m.fill = WARN_FILL
            cell_m.comment = Comment(
                f"⚠️ 관리코드 {cs} 모델 불일치\n\n"
                f"권위(R{auth_r}): {auth_model}\n"
                f"현재(R{r}): {cur_model}\n\n"
                f"같은 검사기면 → 위 권위 값으로 수정\n"
                f"다른 검사기면 → 관리코드 비우고 sync 재실행 (새 코드 부여)",
                "KNK PMS"
            )
            cell_m.comment.width = 380
            cell_m.comment.height = 150
            warned += 1
            mismatches.append((cs, auth_r, "모델", auth_model, r, cur_model, cust))
            auth_rows_with_mismatch.add((cs, auth_r))
            log.warning(f"  ⚠️ R{r} 관리코드 {cs} 모델 불일치: "
                       f"권위(R{auth_r})='{auth_model}' vs 현재='{cur_model}'")

        # 품명
        cell_p = ws.cell(r, C_PROD)
        if not cur_prod and auth_prod:
            cell_p.value = auth_prod
            fixed += 1
            log.info(f"  관리코드 {cs} 품명 자동 보정: R{r} ← '{auth_prod}' (권위 R{auth_r})")
        elif cur_prod and cur_prod != auth_prod:
            cell_p.fill = WARN_FILL
            cell_p.comment = Comment(
                f"⚠️ 관리코드 {cs} 품명 불일치\n\n"
                f"권위(R{auth_r}): {auth_prod}\n"
                f"현재(R{r}): {cur_prod}\n\n"
                f"같은 검사기면 → 위 권위 값으로 수정\n"
                f"다른 검사기면 → 관리코드 비우고 sync 재실행 (새 코드 부여)",
                "KNK PMS"
            )
            cell_p.comment.width = 380
            cell_p.comment.height = 150
            warned += 1
            mismatches.append((cs, auth_r, "품명", auth_prod, r, cur_prod, cust))
            auth_rows_with_mismatch.add((cs, auth_r))
            log.warning(f"  ⚠️ R{r} 관리코드 {cs} 품명 불일치: "
                       f"권위(R{auth_r})='{auth_prod}' vs 현재='{cur_prod}'")

    # 3) 권위 행 표시 (옅은 파랑 + ★ 권위 메모)
    for cs, auth_r in auth_rows_with_mismatch:
        for col in (C_MODEL, C_PROD):
            cell = ws.cell(auth_r, col)
            cell.fill = AUTH_FILL
            cell.comment = Comment(
                f"★ 권위 행 (R{auth_r})\n\n"
                f"관리코드 {cs}의 기준값.\n"
                f"같은 코드를 가진 다른 행(노랑)에서\n"
                f"이 값과 비교됩니다.",
                "KNK PMS"
            )
            cell.comment.width = 280
            cell.comment.height = 110

    # 4) 별도 시트 '9_관리코드불일치' 보고서 생성
    if mismatches:
        _build_inconsistency_report(wb, mismatches, log)
    else:
        # 불일치 없으면 기존 보고 시트 제거
        if "9_관리코드불일치" in wb.sheetnames:
            del wb["9_관리코드불일치"]

    if fixed or warned:
        log.info(f"  관리코드 일관성: 자동보정 {fixed}건 / 경고(노랑셀) {warned}건")
    return fixed, warned, mismatches


def _build_inconsistency_report(wb, mismatches, log):
    """9_관리코드불일치 시트 자동 생성/갱신"""
    sheet_name = "9_관리코드불일치"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    THIN = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL_L = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

    MAX_COL = 8
    # R1 타이틀
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=MAX_COL)
    c1 = ws.cell(1, 1)
    c1.value = f"㈜케이엔케이 │ 관리코드 일관성 불일치 보고 │ {datetime.date.today().isoformat()}"
    c1.fill = PatternFill("solid", fgColor="A5282C")
    c1.font = Font(name="맑은 고딕", size=12, color="FFFFFF", bold=True)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    # R2 슬로건
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=MAX_COL)
    c2 = ws.cell(2, 1)
    c2.value = ("같은 관리코드인데 모델/품명이 다르게 입력된 케이스입니다. "
                "(a) 같은 검사기면 권위 값으로 통일 / (b) 다른 검사기면 새 관리코드 부여")
    c2.fill = PatternFill("solid", fgColor="8B1E22")
    c2.font = Font(name="맑은 고딕", size=9, color="FFFFFF")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    # R3 가이드 (간단)
    for c in range(1, MAX_COL + 1):
        cell = ws.cell(3, c)
        cell.fill = PatternFill("solid", fgColor="E0E0E0")
        cell.font = Font(name="맑은 고딕", size=8, bold=True, color="4A4A4A")
        cell.alignment = AL_C
        cell.border = THIN
    ws.cell(3, 1).value = "🔒자동"
    ws.row_dimensions[3].height = 22

    # R4 헤더
    headers = ["NO", "관리코드", "고객사", "종류",
               "권위 행", "권위 값", "현재 행", "현재 값"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(4, c)
        cell.value = h
        cell.fill = PatternFill("solid", fgColor="A5282C")
        cell.font = Font(name="맑은 고딕", size=9, bold=True, color="FFFFFF")
        cell.alignment = AL_C
        cell.border = THIN
    ws.row_dimensions[4].height = 36

    # R5+ 데이터
    for i, (cs, auth_r, kind, auth_val, cur_r, cur_val, cust) in enumerate(mismatches, 1):
        row = 4 + i
        ws.cell(row, 1).value = i
        ws.cell(row, 2).value = cs
        ws.cell(row, 3).value = cust
        ws.cell(row, 4).value = kind
        ws.cell(row, 5).value = f"R{auth_r}"
        ws.cell(row, 6).value = auth_val
        ws.cell(row, 7).value = f"R{cur_r}"
        ws.cell(row, 8).value = cur_val
        for c in range(1, MAX_COL + 1):
            cell = ws.cell(row, c)
            cell.font = Font(name="맑은 고딕", size=9)
            cell.border = THIN
            if c in (6, 8):
                cell.alignment = AL_L
            else:
                cell.alignment = AL_C
        # 현재 값 셀에 노랑 fill (권위와 다름 강조)
        ws.cell(row, 8).fill = PatternFill("solid", fgColor="FFEB9C")

    # 컬럼 너비
    widths = [5, 14, 18, 8, 10, 35, 10, 35]
    for c, w in enumerate(widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(c)].width = w

    # freeze
    ws.freeze_panes = "A5"

    log.info(f"  보고서 생성: 9_관리코드불일치 ({len(mismatches)}건)")


# ═══════════════════════════════════════════════════════════════
# ② 부서입력 → 진행현황 수집
# ═══════════════════════════════════════════════════════════════
def _collect_dept_progress(log):
    """부서입력 파일에서 진척률 수집 → {관리코드: {부서: {세부: 진척률}}} (v2026.04b)

    컬럼 구조 변경: auto_cols 10 → 11 (C10=담당자 추가).
    담당자(C10) 비어있으면 해당 부서 미참여로 간주 → 전체 평균에서 제외 플래그.
    """
    progress = {}
    for dept in DEPTS:
        fp = dept_path(dept)
        if not os.path.exists(fp):
            log.warning(f"  부서입력 없음: {dept}")
            continue
        wb = load_workbook(fp, data_only=True)
        ws = wb.active
        subs = DEPT_SUB_ITEMS.get(dept, [])
        n_auto = 11  # v2026.04b: 자동연동 컬럼 수 (담당자 포함)

        for r in range(5, ws.max_row + 1):
            code = _val(ws, r, 2)
            if not code:
                continue
            code = str(code)
            if code not in progress:
                progress[code] = {}
            pic = str(_val(ws, r, 10) or "").strip()    # 담당자 (C10)
            dept_data = {"_pic": pic, "_has_pic": bool(pic)}
            for i, sub in enumerate(subs):
                v = _num(ws, r, n_auto + 1 + i)
                dept_data[sub] = v
            vals = [v for v in dept_data.values() if isinstance(v, (int, float)) and v > 0]
            dept_data["_avg"] = sum(vals) / len(vals) if vals else 0
            progress[code][dept] = dept_data
        wb.close()

    return progress


def _write_progress(ws_pms_prog, ws_pms_proj, progress, log):
    """2_진행현황 시트에 write-back"""
    # 프로젝트 목록 수집 (1_프로젝트등록에서 활성 프로젝트만)
    projects = []
    for r in range(5, ws_pms_proj.max_row + 1):
        code = _val(ws_pms_proj, r, C_CODE)
        status = str(_val(ws_pms_proj, r, C_STATUS))
        if code and status not in ("납품완료", "취소"):
            projects.append({
                "code": str(code),
                "sj": _val(ws_pms_proj, r, C_SJNUM),
                "cust": _val(ws_pms_proj, r, C_CUST),
                "prod": _val(ws_pms_proj, r, C_PROD),
                "model": _val(ws_pms_proj, r, C_MODEL),
                "status": status,
            })

    base_cols = 8  # NO, 관리코드, 수주번호, 고객사, 모델, 품명, 진행상태, 전체진척률 (v2026.04b swap)
    row = 5
    for idx, proj in enumerate(projects, 1):
        ws_pms_prog.cell(row=row, column=1).value = idx
        ws_pms_prog.cell(row=row, column=2).value = proj["code"]
        ws_pms_prog.cell(row=row, column=3).value = proj["sj"]
        ws_pms_prog.cell(row=row, column=4).value = proj["cust"]
        ws_pms_prog.cell(row=row, column=5).value = proj["model"]    # v2026.04b swap
        ws_pms_prog.cell(row=row, column=6).value = proj["prod"]     # v2026.04b swap
        ws_pms_prog.cell(row=row, column=7).value = proj["status"]

        col = base_cols + 1
        dept_avgs = []
        code_prog = progress.get(proj["code"], {})
        for dept in DEPTS:
            subs = DEPT_SUB_ITEMS.get(dept, [])
            dept_data = code_prog.get(dept, {})
            for sub in subs:
                v = dept_data.get(sub, 0)
                ws_pms_prog.cell(row=row, column=col).value = v / 100 if v else None
                ws_pms_prog.cell(row=row, column=col).number_format = PCT
                col += 1
            # 소계
            avg = dept_data.get("_avg", 0)
            ws_pms_prog.cell(row=row, column=col).value = avg / 100 if avg else None
            ws_pms_prog.cell(row=row, column=col).number_format = PCT
            if avg > 0:
                dept_avgs.append(avg)
            col += 1

        # 전체진척률
        total_avg = sum(dept_avgs) / len(dept_avgs) if dept_avgs else 0
        ws_pms_prog.cell(row=row, column=8).value = total_avg / 100 if total_avg else None
        ws_pms_prog.cell(row=row, column=8).number_format = PCT

        # PMS 1_프로젝트등록에도 전체진척률 write-back
        for pr in range(5, ws_pms_proj.max_row + 1):
            if str(_val(ws_pms_proj, pr, C_CODE)) == proj["code"]:
                ws_pms_proj.cell(row=pr, column=C_PROG).value = total_avg / 100 if total_avg else None
                ws_pms_proj.cell(row=pr, column=C_PROG).number_format = PCT
                break

        row += 1

    log.info(f"  진행현황 갱신: {len(projects)}건")



# ═══════════════════════════════════════════════════════════════
# ⑤ 출하현황 집계
# ═══════════════════════════════════════════════════════════════
def _sync_ship_summary(ws_ship, ws_proj, log):
    """3_출하현황 시트 집계"""
    ship_rows = []

    # 장비출하 (PMS에서 수집)
    for r in range(5, ws_proj.max_row + 1):
        ship_final = _val(ws_proj, r, C_SHIP_FINAL)
        if not ship_final:
            continue
        ship_rows.append({
            "type": "장비",
            "inv": "",
            "code": str(_val(ws_proj, r, C_CODE)),
            "sj": _val(ws_proj, r, C_SJNUM),
            "cust": _val(ws_proj, r, C_CUST),
            "prod": _val(ws_proj, r, C_PROD),
            "route": _val(ws_proj, r, C_TRANSPORT),
            "trade": "",
            "customs": "",
            "ship_date": ship_final,
            "currency": _val(ws_proj, r, C_CURRENCY),
            "qty": _num(ws_proj, r, C_QTY),
            "box": 0,
            "amount": _num(ws_proj, r, C_TOTAL),
            "inv_usd": 0,
            "customs_status": "",
            "vn_date": "",
        })

    # 코드→프로젝트 매핑 (고객사, 품명)
    code_info = {}
    for r in range(5, ws_proj.max_row + 1):
        c = str(_val(ws_proj, r, C_CODE))
        if c:
            code_info[c] = {
                "cust": _val(ws_proj, r, C_CUST),
                "prod": _val(ws_proj, r, C_PROD),
            }

    # write
    row = 5
    for idx, sr in enumerate(ship_rows, 1):
        info = code_info.get(sr["code"], {})
        ws_ship.cell(row=row, column=1).value = idx
        ws_ship.cell(row=row, column=2).value = sr["type"]
        ws_ship.cell(row=row, column=3).value = sr["inv"]
        ws_ship.cell(row=row, column=4).value = sr["code"]
        ws_ship.cell(row=row, column=5).value = sr["sj"]
        ws_ship.cell(row=row, column=6).value = sr.get("cust") or info.get("cust", "")
        ws_ship.cell(row=row, column=7).value = sr.get("prod") or info.get("prod", "")
        ws_ship.cell(row=row, column=8).value = sr["route"]
        ws_ship.cell(row=row, column=9).value = sr["trade"]
        ws_ship.cell(row=row, column=10).value = sr["customs"]
        ws_ship.cell(row=row, column=11).value = sr["ship_date"]
        ws_ship.cell(row=row, column=12).value = sr["currency"]
        ws_ship.cell(row=row, column=13).value = sr["exrate"]
        ws_ship.cell(row=row, column=14).value = sr["qty"]
        ws_ship.cell(row=row, column=15).value = sr["box"]
        fmt_cell(ws_ship, row, 16, sr["amount"], is_money=True)
        fmt_cell(ws_ship, row, 17, sr["inv_usd"], is_money=True, is_usd=True)
        ws_ship.cell(row=row, column=18).value = sr["customs_status"]
        ws_ship.cell(row=row, column=19).value = sr["vn_date"]
        row += 1

    log.info(f"  출하현황 갱신: {len(ship_rows)}건")


# ═══════════════════════════════════════════════════════════════
# ⑥ 완료이력 아카이브
# ═══════════════════════════════════════════════════════════════
def _archive_completed(ws_proj, ws_archive, log):
    """납품완료 프로젝트를 4_완료이력으로 이동"""
    # 기존 아카이브 코드 수집
    archived = set()
    for r in range(5, ws_archive.max_row + 1):
        c = _val(ws_archive, r, C_CODE)
        if c:
            archived.add(str(c))

    # 아카이브 다음 행 찾기
    next_row = 5
    for r in range(5, ws_archive.max_row + 1):
        if _val(ws_archive, r, C_CODE):
            next_row = r + 1

    count = 0
    for r in range(5, ws_proj.max_row + 1):
        status = str(_val(ws_proj, r, C_STATUS))
        code = str(_val(ws_proj, r, C_CODE))
        if status == "납품완료" and code and code not in archived:
            for c in range(1, MAX_COL + 1):
                ws_archive.cell(row=next_row, column=c).value = _val(ws_proj, r, c)
            next_row += 1
            count += 1

    if count:
        log.info(f"  완료이력 아카이브: {count}건")


def _archive_settled(ws_archive, ws_settled, log):
    """4_완료이력 → 8_매출마감: 계산서발행일 AND 입금일 모두 있는 건 이관 (v2026.04e)"""
    # 이미 매출마감된 코드+수주번호 조합 수집
    settled_keys = set()
    for r in range(5, ws_settled.max_row + 1):
        code = _val(ws_settled, r, C_CODE)
        sj   = _val(ws_settled, r, C_SJNUM)
        if code:
            settled_keys.add((str(code), str(sj) if sj else ""))

    next_row = 5
    for r in range(5, ws_settled.max_row + 1):
        if _val(ws_settled, r, C_CODE):
            next_row = r + 1

    # 완료이력 → 매출마감 이동 대상 찾기
    to_remove = []
    moved = 0
    for r in range(5, ws_archive.max_row + 1):
        code = _val(ws_archive, r, C_CODE)
        sj   = _val(ws_archive, r, C_SJNUM)
        inv  = _val(ws_archive, r, C_INVOICE_DATE)
        pay  = _val(ws_archive, r, C_PAYMENT_DATE)
        if not code:
            continue
        if inv and pay:
            key = (str(code), str(sj) if sj else "")
            if key in settled_keys:
                to_remove.append(r)   # 이미 이관된 건, 완료이력에서 제거만
                continue
            # 매출마감 시트에 복사
            for c in range(1, MAX_COL + 1):
                ws_settled.cell(row=next_row, column=c).value = _val(ws_archive, r, c)
            next_row += 1
            moved += 1
            to_remove.append(r)

    # 완료이력에서 이관된 행 삭제 (큰 번호부터 역순)
    for r in sorted(to_remove, reverse=True):
        ws_archive.delete_rows(r, 1)

    if moved:
        log.info(f"  매출마감 이관: {moved}건 (계산서+입금 완료)")
    elif to_remove:
        log.info(f"  매출마감 중복 제거: {len(to_remove)}건")


# ═══════════════════════════════════════════════════════════════
# ⑦ 매핑·대장 갱신
# ═══════════════════════════════════════════════════════════════
def _sync_mapping_and_ledgers(wb_pms, ws_proj, log):
    """5_매핑조회, 6_관리코드발행대장, 7_수주번호생성대장 갱신"""
    # 5_매핑조회
    ws5 = wb_pms["5_매핑조회(수주번호)"]
    row = 5
    for r in range(5, ws_proj.max_row + 1):
        code = _val(ws_proj, r, C_CODE)
        sj = _val(ws_proj, r, C_SJNUM)
        if not code and not sj:
            continue
        ws5.cell(row=row, column=1).value = row - 4
        ws5.cell(row=row, column=2).value = code
        ws5.cell(row=row, column=3).value = sj
        ws5.cell(row=row, column=4).value = _val(ws_proj, r, C_CUST)
        ws5.cell(row=row, column=5).value = _val(ws_proj, r, C_MODEL)  # v2026.04b swap
        ws5.cell(row=row, column=6).value = _val(ws_proj, r, C_PROD)   # v2026.04b swap
        ws5.cell(row=row, column=7).value = _val(ws_proj, r, C_STATUS)
        ws5.cell(row=row, column=8).value = _val(ws_proj, r, C_SJDATE)
        row += 1

    # 6_관리코드발행대장 — 최신 발행 순(desc) 정렬 (v2026.04g)
    ws6 = wb_pms["6_관리코드발행대장"]
    records6 = []
    for r in range(5, ws_proj.max_row + 1):
        code = _val(ws_proj, r, C_CODE)
        if not code:
            continue
        records6.append({
            "code": str(code),
            "sj":   _val(ws_proj, r, C_SJNUM),
            "cust": _val(ws_proj, r, C_CUST),
            "prod": _val(ws_proj, r, C_PROD),
        })
    # 관리코드 형식 SSSBYYMM (예: 001T2604) → (YYMM, SEQ) desc
    def _code_key(c):
        s = str(c).strip()
        if len(s) == 8:
            try:
                return (s[4:], int(s[:3]))   # ('2605', 1)
            except ValueError:
                pass
        return ("0000", 0)
    records6.sort(key=lambda x: _code_key(x["code"]), reverse=True)
    # 기존 R5+ clear
    for r in range(5, ws6.max_row + 1):
        for c in range(1, 8):
            ws6.cell(r, c).value = None
    for i, rec in enumerate(records6, 1):
        ws6.cell(row=4 + i, column=1).value = i
        ws6.cell(row=4 + i, column=2).value = rec["code"]
        ws6.cell(row=4 + i, column=3).value = rec["sj"]
        ws6.cell(row=4 + i, column=4).value = rec["cust"]
        ws6.cell(row=4 + i, column=5).value = rec["prod"]
        ws6.cell(row=4 + i, column=6).value = TODAY.isoformat()
        ws6.cell(row=4 + i, column=7).value = "자동채번"

    # 7_수주번호생성대장 — 최신 발행 순(desc) 정렬 (v2026.04g)
    ws7 = wb_pms["7_수주번호생성대장"]
    records7 = []
    for r in range(5, ws_proj.max_row + 1):
        sj = _val(ws_proj, r, C_SJNUM)
        if not sj:
            continue
        records7.append({
            "sj":   str(sj),
            "code": _val(ws_proj, r, C_CODE),
            "cust": _val(ws_proj, r, C_CUST),
            "prod": _val(ws_proj, r, C_PROD),
        })
    # 수주번호 형식 B-YYMMDD[-N] → (YYMMDD, N) desc
    def _sj_key(s):
        parts = str(s).strip().split("-")
        if len(parts) >= 2:
            try:
                yymmdd = parts[1]
                n = int(parts[2]) if len(parts) >= 3 else 1
                return (yymmdd, n)
            except (ValueError, IndexError):
                pass
        return ("000000", 0)
    records7.sort(key=lambda x: _sj_key(x["sj"]), reverse=True)
    # 기존 R5+ clear
    for r in range(5, ws7.max_row + 1):
        for c in range(1, 7):
            ws7.cell(r, c).value = None
    for i, rec in enumerate(records7, 1):
        ws7.cell(row=4 + i, column=1).value = i
        ws7.cell(row=4 + i, column=2).value = rec["sj"]
        ws7.cell(row=4 + i, column=3).value = rec["code"]
        ws7.cell(row=4 + i, column=4).value = rec["cust"]
        ws7.cell(row=4 + i, column=5).value = rec["prod"]
        ws7.cell(row=4 + i, column=6).value = TODAY.isoformat()

    log.info(f"  매핑·대장 갱신: 6대장 {len(records6)}건, 7대장 {len(records7)}건 (최신순 정렬)")


# ═══════════════════════════════════════════════════════════════
# 부서입력 파일 자동연동 (PMS → 부서)
# ═══════════════════════════════════════════════════════════════
def _sync_dept_files(ws_proj, log):
    """PMS ↔ 부서입력 파일 양방향 동기화 (v2026.04d)

    동작 규칙:
      - PMS 부서 컬럼 값 == "제외"  → 해당 부서 파일에 프로젝트 미등록
      - PMS 부서 컬럼 값 == (이름)  → 부서 파일 C10(담당자)에 이름 반영
      - PMS 부서 컬럼 값 == (빈 칸) + 부서 파일 C10 == (이름) → 그 이름을 PMS로 역반영
      - 빈 칸 + 빈 칸 → 그대로 유지 (참여 확정, 담당자 미정)

    컬럼 매핑 (부서진척입력 시트):
      C1 NO / C2 관리코드 / C3 수주번호 / C4 고객사 / C5 모델 / C6 품명 /
      C7 PO유형 / C8 영업단계 / C9 진행상태 / C10 담당자 / C11 전체진척률(%)
    """
    # 1단계: PMS 읽기 + PMS 부서 컬럼 값 수집
    projects = []
    pms_dept_vals = {}   # {code: {dept: 값}}
    pms_rows = {}        # {code: row}
    for r in range(5, ws_proj.max_row + 1):
        code = _val(ws_proj, r, C_CODE)
        status = str(_val(ws_proj, r, C_STATUS))
        if not code:
            continue
        code = str(code)
        pms_rows[code] = r
        dept_vals = {}
        for i, dept in enumerate(DEPTS):
            v = ws_proj.cell(r, C_DEPT_START + i).value
            dept_vals[dept] = str(v).strip() if v is not None else ""
        pms_dept_vals[code] = dept_vals
        projects.append({
            "code": code, "sj": _val(ws_proj, r, C_SJNUM),
            "cust": _val(ws_proj, r, C_CUST),
            "model": _val(ws_proj, r, C_MODEL), "prod": _val(ws_proj, r, C_PROD),
            "potype": _val(ws_proj, r, C_POTYPE),
            "stage": _val(ws_proj, r, C_STAGE), "status": status,
            "prog": _val(ws_proj, r, C_PROG),
        })

    # 2단계: 부서 파일에서 C10 담당자 읽기 → PMS 부서 컬럼에 역반영
    #        ("제외"는 유지. 빈 칸 PMS + 이름 있는 파일 → PMS 업데이트)
    dept_pic_from_file = {}  # {code: {dept: 담당자}}
    for dept in DEPTS:
        fp = dept_path(dept)
        if not os.path.exists(fp):
            continue
        wb_d = load_workbook(fp, data_only=True)
        ws_d = wb_d.active
        for r in range(5, ws_d.max_row + 1):
            c = ws_d.cell(r, 2).value
            if c:
                pic = ws_d.cell(r, 10).value
                pic_str = str(pic).strip() if pic is not None else ""
                dept_pic_from_file.setdefault(str(c), {})[dept] = pic_str
        wb_d.close()

    # 부서 파일 → PMS 역반영
    n_writeback = 0
    for code, row in pms_rows.items():
        for i, dept in enumerate(DEPTS):
            pms_val = pms_dept_vals[code][dept]
            file_val = dept_pic_from_file.get(code, {}).get(dept, "")
            if pms_val == "제외":
                continue  # "제외" 유지
            if file_val and file_val != pms_val:
                # 부서 파일 쪽 최신 이름 → PMS에 반영
                ws_proj.cell(row, C_DEPT_START + i).value = file_val
                pms_dept_vals[code][dept] = file_val
                n_writeback += 1
    if n_writeback:
        log.info(f"  부서 파일 → PMS 역반영: {n_writeback}건")

    # 3단계: 부서 파일 기록 (제외 부서 스킵 + PMS 이름을 C10으로 복사)
    for dept in DEPTS:
        fp = dept_path(dept)
        if not os.path.exists(fp):
            continue
        wb_d = load_workbook(fp)
        ws_d = wb_d.active

        # 기존 담당자 + 진척률 값 보존
        existing_pic = {}
        for r in range(5, ws_d.max_row + 1):
            c = ws_d.cell(r, 2).value
            if c:
                existing_pic[str(c)] = ws_d.cell(r, 10).value

        row = 5
        for proj in projects:
            dept_val = pms_dept_vals[proj["code"]][dept]
            if dept_val == "제외":
                continue   # 제외 부서는 미등록
            if proj["status"] in ("납품완료", "취소"):
                continue

            ws_d.cell(row, 1).value  = row - 4
            ws_d.cell(row, 2).value  = proj["code"]
            ws_d.cell(row, 3).value  = proj["sj"]
            ws_d.cell(row, 4).value  = proj["cust"]
            ws_d.cell(row, 5).value  = proj["model"]
            ws_d.cell(row, 6).value  = proj["prod"]
            ws_d.cell(row, 7).value  = proj["potype"]
            ws_d.cell(row, 8).value  = proj["stage"]
            ws_d.cell(row, 9).value  = proj["status"]
            # C10 담당자: PMS 값이 있으면 그걸, 없으면 기존 파일 값 유지
            pic_to_write = dept_val if dept_val else existing_pic.get(proj["code"])
            ws_d.cell(row, 10).value = pic_to_write
            ws_d.cell(row, 11).value = proj["prog"]
            row += 1

        wb_d.save(fp)
    log.info(f"  부서입력 연동: {len(DEPTS)}개 부서")


# ═══════════════════════════════════════════════════════════════
# 물류상세 자동생성 (관리코드 있고 상세파일 없으면)
# ═══════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════
# 메인 동기화
# ═══════════════════════════════════════════════════════════════

def sync_all():
    """전체 동기화"""
    log = _setup_logger()
    log.info("=" * 60)
    log.info(f"  KNK PMS V4 — {TYPE_NAME} 동기화 시작")
    log.info("=" * 60)

    fp = pms_path()
    if not os.path.exists(fp):
        log.error(f"  PMS 파일 없음: {fp}")
        log.error("  build.py를 먼저 실행하세요.")
        return

    wb = load_workbook(fp)
    ws_proj = wb["1_프로젝트등록"]
    ws_prog = wb["2_진행현황"]
    ws_ship = wb["3_출하현황"]
    ws_archive = wb["4_완료이력"]
    ws_settled = wb["8_매출마감"] if "8_매출마감" in wb.sheetnames else None

    # ① 날짜 정규화 (스킬 §13.5)
    log.info("① 날짜 정규화")
    DATE_COLS_PMS = [C_SJDATE, C_DUE]  # v2026.04: 01 검사기 X~AH 삭제로 날짜 컬럼 2개만
    _normalize_date_cells(ws_proj, DATE_COLS_PMS, log)

    # ② 관리코드·수주번호 채번 + 금액 계산
    log.info("② 관리코드·수주번호 채번 + 자동계산")
    new_codes = _assign_codes(ws_proj, log)

    # ②-b 관리코드 일관성 검증 (모델·품명 자동 보정 + 불일치 셀 노랑 + hover 메모 + 보고 시트)
    log.info("②-b 관리코드 일관성 검증")
    _validate_code_consistency(ws_proj, wb, log)

    # ② 부서입력 연동 (PMS → 부서)
    log.info("② 부서입력 연동")
    _sync_dept_files(ws_proj, log)

    # 부서입력 → 진행현황 수집
    log.info("  부서진척률 수집")
    progress = _collect_dept_progress(log)
    _write_progress(ws_prog, ws_proj, progress, log)

    # ⑤ 출하현황 집계 (v2026.04: 01 검사기 X~AH 삭제로 skip)
    # _sync_ship_summary(ws_ship, ws_proj, log)

    # ⑥ 완료이력 아카이브
    log.info("⑥ 완료이력 아카이브")
    _archive_completed(ws_proj, ws_archive, log)

    # ⑥-b 매출마감 이관 (계산서·입금 완료 건)
    if ws_settled is not None:
        log.info("⑥-b 매출마감 이관")
        _archive_settled(ws_archive, ws_settled, log)

    # ⑦ 매핑·대장 갱신
    log.info("⑦ 매핑·대장 갱신")
    _sync_mapping_and_ledgers(wb, ws_proj, log)

    # ⑧ 1_프로젝트등록 정렬 + 상단 10칸 빈 공간 (v2026.04 신규)
    log.info("⑧ 정렬 + 상단 10칸 빈 공간")
    _sort_and_pad_proj(ws_proj, log)

    wb.save(fp)
    log.info("=" * 60)
    log.info("  동기화 완료!")
    log.info("=" * 60)


# ═══════════════════════════════════════════════════════════════
# 정렬 (v2026.04c — 상단 빈 공간 제거, 데이터 R5부터 바로 시작)
# 신규 수주는 Excel에서 Ctrl+End → 맨 아래 빈 행 입력 → .bat 실행
# ═══════════════════════════════════════════════════════════════
DATA_START = 5
DATA_AFTER_PAD = 5   # v2026.04c: 빈 공간 제거 (기존 15)
INACTIVE_STATUSES = {"납품완료", "완료", "취소", "보류"}



def _sort_and_pad_proj(ws, log):
    """1_프로젝트등록 정렬 + 상단 10칸 빈 행 보장 (v2026.04c)

    정렬 우선순위:
      0: 진행중 / 수주예정 — 납기일 ASC
      1: 제안작성 / 제안제출 — 수주일 ASC
      2: 보류 — 납기일 ASC
      3: 취소 — 납기일 ASC

    납품완료/완료 상태는 1_프로젝트등록에서 제외 (4_완료이력에만 존재).
    제안 단계 행은 관리코드 없이도 보존됨 (수주확정 전까지).
    """
    import datetime as _dt
    STATUS_ORDER = {"진행중": 0, "수주예정": 0, "보류": 2, "취소": 3}
    PROPOSAL_STAGES = {"제안작성", "제안제출"}
    ARCHIVE_STATUSES = {"납품완료", "완료"}   # v2026.04c: 1_프로젝트등록에서 제외
    DDAY_FMT = '+0"일";-0"일";"0일"'
    INACTIVE = ARCHIVE_STATUSES | {"취소", "보류"}

    rows = []
    for r in range(DATA_START, ws.max_row + 1):
        code = ws.cell(r, C_CODE).value
        stage = str(ws.cell(r, C_STAGE).value or "").strip()
        status = str(ws.cell(r, C_STATUS).value or "").strip()
        cust = ws.cell(r, C_CUST).value
        prod = ws.cell(r, C_PROD).value

        has_code     = code is not None and str(code).strip() != ""
        is_proposal  = stage in PROPOSAL_STAGES
        has_content  = bool((cust and str(cust).strip()) or (prod and str(prod).strip()))

        # 완전 빈 행 스킵
        if not has_code and not (is_proposal and has_content):
            continue

        # 납품완료/완료는 1_프로젝트등록에서 제외 (4_완료이력에만 존재)
        if status in ARCHIVE_STATUSES:
            continue

        row_data = []
        for c in range(1, MAX_COL + 1):
            cell = ws.cell(r, c)
            row_data.append((cell.value, cell.number_format))

        due_v  = ws.cell(r, C_DUE).value
        due    = parse_date(due_v) if due_v else None
        due_key = due.toordinal() if due else 99999999
        sj_v   = ws.cell(r, C_SJDATE).value
        sj     = parse_date(sj_v) if sj_v else None
        sj_key = sj.toordinal() if sj else 99999999

        if is_proposal and not has_code:
            status_key = 1                    # 제안 단계 그룹
            sort_inner = sj_key               # 수주일 기준
        else:
            status_key = STATUS_ORDER.get(status, 5)
            sort_inner = due_key              # 납기일 기준

        rows.append((status_key, sort_inner, r, row_data, status, is_proposal))

    # (그룹, 내부 정렬키, 원래 row 번호)로 정렬 — 안정성 보장
    rows.sort(key=lambda x: (x[0], x[1], x[2]))

    # R5~끝 클리어
    for r in range(DATA_START, ws.max_row + 1):
        for c in range(1, MAX_COL + 1):
            ws.cell(r, c).value = None

    # R15부터 재기록 + D-day 처리
    today = _dt.date.today()
    n_proposal = 0
    for i, (_, _, _, row_data, status, is_proposal) in enumerate(rows):
        target = DATA_AFTER_PAD + i
        for c, (v, fmt) in enumerate(row_data, start=1):
            cell = ws.cell(target, c)
            cell.value = v
            if fmt: cell.number_format = fmt
        ws.cell(target, C_NO).value = i + 1
        # D-day 처리
        dday_cell = ws.cell(target, C_DDAY)
        if is_proposal or status in INACTIVE:
            dday_cell.value = None           # 제안 단계·비활성 상태는 D-day 공란
        else:
            due_v = ws.cell(target, C_DUE).value
            due = parse_date(due_v) if due_v else None
            if due:
                dday_cell.value = (today - due).days
        dday_cell.number_format = DDAY_FMT
        if is_proposal:
            n_proposal += 1

    log.info(f"  정렬: {len(rows)}건 "
             f"(진행중→제안단계[{n_proposal}]→보류→납품완료→취소) R{DATA_AFTER_PAD}~")



if __name__ == "__main__":
    sync_all()
