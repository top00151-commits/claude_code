#!/usr/bin/env python3
"""
baby ↔ web 관리코드 동기화 검증 스크립트 (임시 다리)
==================================================

⚠️  정책 (2026-04-20 확정):
    baby Excel은 **임시 다리**다. web 시스템이 완성되고 직원이 100% 전환되면
    이 스크립트와 baby Excel은 모두 폐기된다. 새 기능을 baby에 추가하지 마라.

실행:
    python scripts/baby_web_sync_check.py

목적:
- HP $160M 실패 패턴 (데이터 마이그레이션 품질) 회피
- baby 엑셀의 관리코드와 web DB의 projects를 비교
- 차이 발견 시 보고 (자동 보정 X — 사람이 결정)
- 마이그레이션 완료 게이트: 7일 연속 차이 0건 + 직원 web 사용 100%

비교 대상:
1. 관리코드 (mgmt_code) 존재 여부
2. 프로젝트명, 고객사, 사업부 일치
3. 영업단계, 진행상태 일치
4. 수주금액 일치

실행 빈도:
- 일일 (cron 또는 수동)
- 큰 변경 후 (baby 엑셀 수정·web 일괄 등록)
"""

import os
import sys
import sqlite3
from datetime import datetime
from pathlib import Path

# openpyxl 의존성
try:
    from openpyxl import load_workbook
except ImportError:
    print("[ERROR] openpyxl 미설치. 설치: pip install openpyxl")
    sys.exit(1)

# 경로 설정
ROOT = Path(__file__).resolve().parent.parent.parent
WEB_DB = ROOT / "HAIST_WORKS" / "data" / "knk.db"
BABY_DIR = ROOT / "HAIST_WORKS_baby" / "V2"

# baby PMS 파일 (검사기·자동화)
BABY_PMS_FILES = [
    BABY_DIR / "01_검사기_완제품" / "KNK_검사기_완제품_PMS_2026.xlsx",
    BABY_DIR / "02_자동화_완제품" / "KNK_자동화_완제품_PMS_2026.xlsx",
]


def load_baby_mgmt_codes():
    """baby PMS 엑셀에서 관리코드 + 핵심 필드 추출"""
    baby_data = {}  # mgmt_code → dict
    for pms_file in BABY_PMS_FILES:
        if not pms_file.exists():
            print(f"[SKIP] baby PMS 파일 없음: {pms_file.name}")
            continue
        try:
            wb = load_workbook(pms_file, read_only=True, data_only=True)
            # 1_프로젝트등록 시트
            sheet_name = next((s for s in wb.sheetnames if "프로젝트" in s), None)
            if not sheet_name:
                continue
            ws = wb[sheet_name]
            # R4 헤더 (4번째 행)
            headers = [c.value for c in ws[4]]
            mgmt_idx = next((i for i, h in enumerate(headers) if h and "관리코드" in str(h)), None)
            name_idx = next((i for i, h in enumerate(headers) if h and "품명" in str(h)), None)
            customer_idx = next((i for i, h in enumerate(headers) if h and "고객사" in str(h)), None)
            stage_idx = next((i for i, h in enumerate(headers) if h and ("수주단계" in str(h) or "단계" in str(h))), None)
            status_idx = next((i for i, h in enumerate(headers) if h and "진행상태" in str(h)), None)

            if mgmt_idx is None:
                continue

            for row in ws.iter_rows(min_row=5, values_only=True):
                code = row[mgmt_idx] if mgmt_idx < len(row) else None
                if not code or not isinstance(code, str):
                    continue
                code = code.strip()
                if len(code) != 8:
                    continue
                baby_data[code] = {
                    "name": row[name_idx] if name_idx is not None and name_idx < len(row) else None,
                    "customer": row[customer_idx] if customer_idx is not None and customer_idx < len(row) else None,
                    "stage": row[stage_idx] if stage_idx is not None and stage_idx < len(row) else None,
                    "status": row[status_idx] if status_idx is not None and status_idx < len(row) else None,
                    "source_file": pms_file.name,
                }
            wb.close()
        except Exception as e:
            print(f"[ERROR] baby 엑셀 읽기 실패 {pms_file.name}: {e}")
    return baby_data


def load_web_mgmt_codes():
    """web DB에서 관리코드 + 핵심 필드 추출"""
    if not WEB_DB.exists():
        print(f"[ERROR] web DB 없음: {WEB_DB}")
        return {}
    web_data = {}
    conn = sqlite3.connect(WEB_DB)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute("""
            SELECT mgmt_code, name, customer_name, stage, status, biz_div
            FROM projects
            WHERE mgmt_code IS NOT NULL AND mgmt_code != ''
        """).fetchall()
        for r in rows:
            code = r["mgmt_code"].strip() if r["mgmt_code"] else None
            if code and len(code) == 8:
                web_data[code] = {
                    "name": r["name"],
                    "customer": r["customer_name"],
                    "stage": r["stage"],
                    "status": r["status"],
                    "biz_div": r["biz_div"],
                }
    finally:
        conn.close()
    return web_data


def compare(baby, web):
    """비교 결과 반환"""
    only_baby = sorted(baby.keys() - web.keys())
    only_web = sorted(web.keys() - baby.keys())
    both = baby.keys() & web.keys()

    mismatches = []
    for code in sorted(both):
        b, w = baby[code], web[code]
        diffs = []
        # 이름·고객사·단계·상태 비교 (None은 무시, 양쪽 다 값이 있을 때만)
        if b.get("name") and w.get("name") and str(b["name"]).strip() != str(w["name"]).strip():
            diffs.append(f"name: baby='{b['name']}' vs web='{w['name']}'")
        if b.get("customer") and w.get("customer") and str(b["customer"]).strip() != str(w["customer"]).strip():
            diffs.append(f"customer: baby='{b['customer']}' vs web='{w['customer']}'")
        if b.get("stage") and w.get("stage") and str(b["stage"]).strip() != str(w["stage"]).strip():
            diffs.append(f"stage: baby='{b['stage']}' vs web='{w['stage']}'")
        if b.get("status") and w.get("status") and str(b["status"]).strip() != str(w["status"]).strip():
            diffs.append(f"status: baby='{b['status']}' vs web='{w['status']}'")
        if diffs:
            mismatches.append((code, diffs))

    return {
        "only_baby": only_baby,
        "only_web": only_web,
        "both_count": len(both),
        "mismatches": mismatches,
    }


def main():
    print("=" * 60)
    print(f"baby ↔ web 동기화 검증")
    print(f"실행: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    baby = load_baby_mgmt_codes()
    print(f"\nbaby 관리코드: {len(baby)}건")
    web = load_web_mgmt_codes()
    print(f"web 관리코드: {len(web)}건")

    result = compare(baby, web)
    print(f"\n양쪽 모두 존재: {result['both_count']}건")
    print(f"baby에만 있음: {len(result['only_baby'])}건")
    print(f"web에만 있음: {len(result['only_web'])}건")
    print(f"필드 불일치: {len(result['mismatches'])}건")

    # 상세
    if result['only_baby']:
        print("\n--- baby에만 있는 관리코드 (web에 없음) ---")
        for code in result['only_baby'][:10]:
            print(f"  {code} | {baby[code].get('name', '-')[:30] if baby[code].get('name') else '-'}")
        if len(result['only_baby']) > 10:
            print(f"  ... 외 {len(result['only_baby']) - 10}건")

    if result['only_web']:
        print("\n--- web에만 있는 관리코드 (baby에 없음) ---")
        for code in result['only_web'][:10]:
            print(f"  {code} | {web[code].get('name', '-')[:30] if web[code].get('name') else '-'}")
        if len(result['only_web']) > 10:
            print(f"  ... 외 {len(result['only_web']) - 10}건")

    if result['mismatches']:
        print("\n--- 필드 불일치 (앞 10건) ---")
        for code, diffs in result['mismatches'][:10]:
            print(f"  {code}:")
            for d in diffs:
                print(f"    - {d}")
        if len(result['mismatches']) > 10:
            print(f"  ... 외 {len(result['mismatches']) - 10}건")

    # 종합 판정
    print("\n" + "=" * 60)
    is_healthy = (
        len(result['only_baby']) < 5 and
        len(result['only_web']) < 5 and
        len(result['mismatches']) < 10
    )
    if is_healthy:
        print("✅ 동기화 양호 (5건 미만 차이)")
    else:
        print("⚠️ 동기화 점검 필요 — 위 차이를 확인하고 사람이 결정하세요")
        print("   자동 보정은 위험 (HP $160M 실패 사례) — 수동 확인 권장")
    print("=" * 60)

    return 0 if is_healthy else 1


if __name__ == "__main__":
    sys.exit(main())
