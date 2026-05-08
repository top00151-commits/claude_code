# -*- coding: utf-8 -*-
"""
v5H226z30 (2026-05-08) 대표 직접 지시
전체 초기화 → 엑셀 일괄 업로드 (T_검사기 + M_자동화)
"""
import sys, io, os, sqlite3, shutil
from datetime import datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

DB = 'data/knk.db'
XLSX = r'C:\Users\top00\JR\Claude 코드\KNK업무시스템구축\참고자료\프로젝트_일괄등록_양식.xlsx'

# 1. 추가 백업 (스크립트 실행 직전)
ts = datetime.now().strftime('%Y%m%d_%H%M%S')
backup = f'data/knk.db.before_wipe_{ts}'
shutil.copy2(DB, backup)
print(f'[1/4] 백업 생성: {backup} ({os.path.getsize(backup):,} bytes)')

# 2. 초기화 대상 테이블 (프로젝트 관련 전부)
WIPE_TABLES = [
    # 프로젝트 자체
    'projects', 'project_history', 'project_milestones', 'project_phases',
    'project_retros', 'project_burndown_snapshots', 'project_forecasts',
    # 수주
    'orders', 'order_items', 'order_status_history',
    # 견적/제안
    'quotations', 'quotation_items', 'quotation_history',
    # 업무/일일
    'tasks', 'task_comments', 'task_delegations', 'task_reactions',
    # 티켓/이슈
    'tickets', 'ticket_comments', 'issues', 'issues_out', 'issue_logs',
    # 변경관리
    'changes', 'change_impacts', 'change_reads',
    # 소모품
    'consumable_orders', 'consumable_order_items',
    # 수출/통관
    'export_orders', 'commercial_invoices', 'packing_lists', 'packing_items',
    'bills_of_lading', 'customs_declarations', 'fta_certificates', 'fta_certificate_items',
    # 출하/생산/구매
    'shipments', 'work_orders', 'work_order_items', 'production_orders',
    'purchase_orders', 'po_items', 'po_item_project_links',
    # 매출/수금
    'invoices', 'receipts', 'receipts_payment',
    # 품질
    'qc_inspections', 'qc_inspection_items', 'qc_inspection_reports', 'qc_disposition',
    'corrective_actions', 'preventive_actions',
    # 재고 변동 (마스터 parts/suppliers는 보존)
    'stock_movements', 'stock_adjustments', 'stock_audits', 'stock_audit_items',
    # 기타 프로젝트 종속
    'notifications', 'activities', 'team_summaries', 'price_change_history',
    'stock_movements', 'comment_mentions',
]

conn = sqlite3.connect(DB)
conn.execute('PRAGMA foreign_keys = OFF')
c = conn.cursor()

print(f'[2/4] 초기화 시작 ({len(set(WIPE_TABLES))} 테이블)')
total_deleted = 0
existing = {r[0] for r in c.execute("SELECT name FROM sqlite_master WHERE type='table'")}
for t in dict.fromkeys(WIPE_TABLES):  # 중복 제거
    if t not in existing:
        continue
    c.execute(f'SELECT COUNT(*) FROM "{t}"')
    n = c.fetchone()[0]
    c.execute(f'DELETE FROM "{t}"')
    total_deleted += n
    if n: print(f'    DELETE {t}: {n} 행')
# AUTOINCREMENT 카운터 리셋
c.execute("DELETE FROM sqlite_sequence WHERE name IN ({})".format(
    ','.join(['?'] * len(set(WIPE_TABLES)))), list(dict.fromkeys(WIPE_TABLES)))
conn.commit()
print(f'    총 {total_deleted:,} 행 삭제 완료')

# 3. 엑셀 파싱
from openpyxl import load_workbook
wb = load_workbook(XLSX, data_only=True)

# 고객사 매핑 (기존 + 자동 생성)
existing_custs = {r[1].strip(): r[0] for r in c.execute('SELECT id, name FROM customers')}

def get_or_create_customer(name):
    name = (name or '').strip()
    if not name: return None
    if name in existing_custs: return existing_custs[name]
    c.execute("INSERT INTO customers (name) VALUES (?)", (name,))
    cid = c.lastrowid
    existing_custs[name] = cid
    print(f'    + 고객사 자동 생성: {name} (id={cid})')
    return cid

def norm_status(s):
    s = (s or '').strip()
    s = s.replace(' ', '')  # "납품 완료" → "납품완료"
    if s == '신규': return '진행중'
    valid = {'초기협의','제안서전달','견적발행','수주예정','진행중','납품완료','취소','보류'}
    return s if s in valid else '진행중'

def norm_amt(v):
    if v is None or v == '': return 0
    s = str(v).strip()
    if s == '무상' or s == '-': return 0
    s = s.replace(',', '')
    try: return float(s)
    except: return 0

def norm_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if not s: return None
    return s[:10]

def norm_potype(v):
    s = (v or '').strip()
    if s in ('신규','추가','수정','MODIFY','TEST','PO'):
        return {'MODIFY':'수정','TEST':'신규','PO':'신규'}.get(s, s)
    return '신규'

# 4. 일괄 등록
print(f'[3/4] 엑셀 파싱·INSERT')

# 데이터 수집 (mgmt_code 단위로 그룹 안 함 — 각 행 = 1 SO)
all_rows = []
for sn, biz in [('T_검사기','T'), ('M_자동화','M')]:
    ws = wb[sn]
    # row 4는 샘플 placeholder, row 5부터 실데이터
    for r_i, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
        # 컬럼 순서: 프로젝트명/관리코드/PO유형/고객사명/모델명/거래구분/발주일/납기일/단가/수량/통화/상태/PM/영업담당/납품처/비고
        if not row[0] or not str(row[0]).strip(): continue
        if not row[1] or not str(row[1]).strip(): continue  # 관리코드 없으면 skip
        all_rows.append({
            'sheet': sn, 'biz_div': biz, 'row_no': r_i,
            'name': str(row[0]).strip(),
            'mgmt_code': str(row[1]).strip(),
            'po_type': norm_potype(row[2]),
            'customer_name': str(row[3] or '').strip(),
            'model_name': str(row[4] or '').strip(),
            'trade': str(row[5] or '').strip(),  # 내수/수출
            'order_date': norm_date(row[6]),
            'due_date': norm_date(row[7]),
            'unit_price': norm_amt(row[8]),
            'qty': int(row[9] or 1) if row[9] else 1,
            'currency': str(row[10] or 'KRW').strip().upper(),
            'status': norm_status(row[11]),
            'pm_name': str(row[12] or '').strip(),
            'sales_name': str(row[13] or '').strip(),
            'ship_to': str(row[14] or '').strip(),
            'note': str(row[15] or '').strip(),
        })

print(f'    엑셀 유효 행: {len(all_rows)}')

# mgmt_code 별 그룹화 (같은 코드 = 같은 프로젝트, 여러 SO)
groups = {}
for row in all_rows:
    key = row['mgmt_code']
    groups.setdefault(key, []).append(row)
print(f'    프로젝트 (관리코드 unique): {len(groups)}')

# INSERT
proj_count, so_count, item_count = 0, 0, 0
for mgmt_code, rows in groups.items():
    head = rows[0]
    cust_id = get_or_create_customer(head['customer_name'])

    # 통화 우선순위: 첫 행 통화 (보통 동일)
    ccy = head['currency']
    is_export = 1 if any(r['trade'] == '수출' for r in rows) else 0
    total_amount = sum(r['unit_price'] * r['qty'] for r in rows)
    total_qty = sum(r['qty'] for r in rows)

    # 프로젝트 INSERT
    c.execute("""
        INSERT INTO projects (
            name, mgmt_code, biz_div, project_type, customer_id, customer_name,
            model_name, po_type, status, currency, order_date, due_date,
            order_amount, unit_qty, unit_price, pm_name, sales_name,
            is_export, shipment_form, type, equip_type, year_month
        ) VALUES (?, ?, ?, 'NEW_EQUIP', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'ASSEMBLY', ?, ?, ?)
    """, (
        head['name'], mgmt_code, head['biz_div'], cust_id, head['customer_name'],
        head['model_name'], head['po_type'], head['status'], ccy,
        head['order_date'], head['due_date'],
        total_amount, total_qty, head['unit_price'],
        head['pm_name'], head['sales_name'], is_export,
        '신규 장비' if head['biz_div'] in ('T','M') else None,
        head['biz_div'],
        (head['order_date'] or '')[2:7].replace('-','') if head['order_date'] else None
    ))
    pid = c.lastrowid
    proj_count += 1

    # SO 생성 (각 엑셀 행마다 1 SO)
    for idx, row in enumerate(rows, start=1):
        # 수주번호: T-YYMMDD-N 형식 (간단)
        # 수주번호 형식: {사업부}-{YYMMDD}-{시퀀스} (관리코드 미포함)
        # 시퀀스는 후처리(_assign_order_no_seq)에서 (사업부, 발주일) 그룹별로 재부여
        ymd = (row['order_date'] or datetime.now().strftime('%Y-%m-%d')).replace('-','')[2:]
        order_no = f"TMP-{row['biz_div']}-{ymd}-{idx}-{mgmt_code}"  # 임시 (후처리 대상)
        # SO 상태 매핑
        so_status_map = {'진행중':'IN_PRODUCTION','납품완료':'SHIPPED','취소':'CANCELLED','보류':'CONFIRMED'}
        so_status = so_status_map.get(row['status'], 'CONFIRMED')

        c.execute("""
            INSERT INTO orders (
                order_no, project_id, customer_id, order_date, due_date,
                total_amount, status, currency, unit_qty, ship_to, so_type,
                unit_label, unit_note
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'EQUIPMENT', ?, ?)
        """, (
            order_no, pid, cust_id, row['order_date'], row['due_date'],
            row['unit_price'] * row['qty'], so_status, row['currency'],
            row['qty'], row['ship_to'],
            f"{row['po_type']} {row['qty']}대", row['note']
        ))
        oid = c.lastrowid
        so_count += 1

        # order_items (호기 라인) — qty 만큼
        unit_status_map = {'진행중':'진행중','납품완료':'납품완료','취소':'취소','보류':'보류'}
        ust = unit_status_map.get(row['status'], '진행중')
        for u in range(row['qty']):
            c.execute("""
                INSERT INTO order_items (
                    order_id, qty, unit_price, amount, unit_label,
                    line_note, order_date, due_date, ship_to, currency,
                    unit_status, is_export
                ) VALUES (?, 1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                oid, row['unit_price'], row['unit_price'],
                f"{u+1}호기" if row['qty'] > 1 else "1호기",
                row['note'], row['order_date'], row['due_date'],
                row['ship_to'], row['currency'], ust,
                1 if row['trade'] == '수출' else 0
            ))
            item_count += 1

conn.commit()

# === 수주번호 시퀀스 재부여: (사업부, 발주일) 그룹별로 -1, -2, -3 ... ===
from collections import defaultdict
groups_so = defaultdict(list)
for oid, biz, odate in c.execute('''
    SELECT o.id, p.biz_div, o.order_date
    FROM orders o JOIN projects p ON o.project_id = p.id
    ORDER BY o.id
'''):
    ymd = (odate or '').replace('-', '')[2:] if odate else ''
    groups_so[(biz, ymd)].append(oid)
n_renum = 0
for (biz, ymd), ids in groups_so.items():
    for seq, oid in enumerate(ids, start=1):
        new_no = f'{biz}-{ymd}-{seq}' if ymd else f'{biz}-NODATE-{seq}'
        c.execute('UPDATE orders SET order_no = ? WHERE id = ?', (new_no, oid))
        n_renum += 1
conn.commit()

print(f'[4/4] INSERT 완료')
print(f'    프로젝트: {proj_count}건')
print(f'    수주(SO): {so_count}건')
print(f'    호기 라인: {item_count}건')
print(f'    수주번호 재부여: {n_renum}건 (사업부+발주일 그룹별 시퀀스)')

# 최종 검증
print()
print('=== 검증 ===')
for t in ('projects','orders','order_items','customers'):
    n = c.execute(f'SELECT COUNT(*) FROM "{t}"').fetchone()[0]
    print(f'  {t}: {n}')
conn.close()
print()
print('✅ 모든 작업 완료')
