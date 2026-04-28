"""
2026-04-28 대표 지시 — 조직도 이미지 기반 부서별 명단 엑셀 생성.
컬럼: 부서이름 · 직급 · 이름 · 아이디 · 비밀번호
DB(data/knk.db)에서 login_id 가져오고, 직급은 조직도 기준으로 매핑.
"""
import sys, os, sqlite3
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DB   = ROOT / "data" / "knk.db"
OUT  = ROOT / "_전직원_로그인계정_2026-04-28.xlsx"
PW   = "knk1234"

# 조직도(2026-04-28 이미지) 기반 — (부서, 직급, 이름, 팀장여부)
ORG = [
    # 대표이사
    ("대표이사실", "대표이사", "김동후", False),
    ("대표이사실", "대표이사", "김정락", False),

    # 임원
    ("임원실", "상무", "이한빈", False),
    ("임원실", "상무", "윤경호", False),
    ("임원실", "전무", "최홍광", False),

    # 01 기술영업팀
    ("01 기술영업팀",     "이사",     "이해림", True),
    ("01 기술영업팀",     "매니저",   "이현",   False),
    ("01 기술영업팀",     "프로",     "오경환", False),
    ("01 기술영업팀",     "프로",     "배승진", False),
    ("01 기술영업팀",     "프로",     "안지연", False),
    ("01 기술영업팀",     "프로",     "이새롬", False),

    # 02 검사기팀(연구소)
    ("02 검사기팀(연구소)", "상무",     "이한빈", True),
    ("02 검사기팀(연구소)", "이사",     "이치권", False),
    ("02 검사기팀(연구소)", "매니저",   "이성진", False),
    ("02 검사기팀(연구소)", "매니저",   "길희용", False),
    ("02 검사기팀(연구소)", "프로",     "윤광훈", False),
    ("02 검사기팀(연구소)", "프로",     "김태형", False),
    ("02 검사기팀(연구소)", "사원",     "이서준", False),
    ("02 검사기팀(연구소)", "사원",     "김지훈", False),
    ("02 검사기팀(연구소)", "반장",     "지경숙", False),

    # 03 품질팀
    ("03 품질팀",         "매니저",   "김정록", True),
    ("03 품질팀",         "사원",     "정형진", False),

    # 04 설계팀(연구소)
    ("04 설계팀(연구소)", "상무",     "윤경호", True),
    # 04 검사기 sub
    ("04 설계팀(연구소)", "매니저",   "이영준", False),
    ("04 설계팀(연구소)", "프로",     "김범수", False),
    ("04 설계팀(연구소)", "사원",     "안호재", False),
    # 04 자동화 sub
    ("04 설계팀(연구소)", "매니저",   "정민규", False),
    ("04 설계팀(연구소)", "매니저",   "이상천", False),
    ("04 설계팀(연구소)", "매니저",   "한재운", False),
    ("04 설계팀(연구소)", "매니저",   "신광용", False),
    ("04 설계팀(연구소)", "프로",     "김동현", False),
    ("04 설계팀(연구소)", "프로",     "최현규", False),

    # 05 소프트웨어팀(연구소)
    ("05 소프트웨어팀(연구소)", "매니저", "이한중", True),
    ("05 소프트웨어팀(연구소)", "매니저", "최창호", False),
    ("05 소프트웨어팀(연구소)", "매니저", "황정석", False),
    ("05 소프트웨어팀(연구소)", "매니저", "현종필", False),
    ("05 소프트웨어팀(연구소)", "매니저", "이정우", False),
    ("05 소프트웨어팀(연구소)", "매니저", "주진호", False),
    ("05 소프트웨어팀(연구소)", "매니저", "김동욱", False),
    ("05 소프트웨어팀(연구소)", "매니저", "이영준2", False),
    ("05 소프트웨어팀(연구소)", "프로",   "이충희", False),
    ("05 소프트웨어팀(연구소)", "프로",   "박주창", False),
    ("05 소프트웨어팀(연구소)", "프로",   "김기운", False),

    # 06 전장설계팀(연구소)
    ("06 전장설계팀(연구소)", "매니저", "김형렬", True),

    # 07 제조기술1팀
    ("07 제조기술1팀", "매니저", "노충일", True),
    ("07 제조기술1팀", "사원",   "마준영", False),
    ("07 제조기술1팀", "사원",   "이태우", False),
    ("07 제조기술1팀", "사원",   "금진호", False),

    # 08 제조기술2팀
    ("08 제조기술2팀", "매니저", "임택훈", True),
    ("08 제조기술2팀", "매니저", "서재희", False),
    ("08 제조기술2팀", "매니저", "연태흠", False),
    ("08 제조기술2팀", "프로",   "방성기", False),
    ("08 제조기술2팀", "프로",   "김한성", False),
    ("08 제조기술2팀", "사원",   "박지현", False),
    ("08 제조기술2팀", "사원",   "나영훈", False),
    ("08 제조기술2팀", "사원",   "강대성", False),

    # 09 가공팀
    ("09 가공팀", "매니저", "윤영조", True),
    ("09 가공팀", "매니저", "이수빈", False),

    # 10 구매팀
    ("10 구매팀", "매니저", "정성진", True),
    ("10 구매팀", "매니저", "허동준", False),
    ("10 구매팀", "프로",   "오용균", False),
    ("10 구매팀", "프로",   "김선미", False),
    ("10 구매팀", "프로",   "이홍규", False),
    ("10 구매팀", "사원",   "박성준", False),
    ("10 구매팀", "사원",   "란",     False),

    # 11 관리팀
    ("11 관리팀", "매니저", "박지은", True),
    ("11 관리팀", "매니저", "엄혜린", False),
    ("11 관리팀", "프로",   "엄주영", False),
    ("11 관리팀", "프로",   "최혜연", False),

    # 12 베트남법인
    ("12 베트남법인", "법인장", "이용식", True),
    ("12 베트남법인", "매니저", "박지만", False),
    ("12 베트남법인", "매니저", "이용호", False),
    ("12 베트남법인", "부장",   "땀",     False),
    ("12 베트남법인", "차장",   "탕",     False),
    ("12 베트남법인", "차장",   "쑤아잉", False),

    # 13 개발혁신팀(연구소)
    ("13 개발혁신팀(연구소)", "상무",   "최보현", True),
    ("13 개발혁신팀(연구소)", "매니저", "박승환", False),
    ("13 개발혁신팀(연구소)", "사원",   "김수현", False),

    # 14 라이프밸류팀(연구소)
    ("14 라이프밸류팀(연구소)", "매니저", "나재겸", True),
    ("14 라이프밸류팀(연구소)", "매니저", "김기선", False),
    ("14 라이프밸류팀(연구소)", "사원",   "박성수", False),
]

def main():
    # 1) DB → name → login_id 매핑
    con = sqlite3.connect(str(DB))
    con.row_factory = sqlite3.Row
    db_users = {r['name']: r['login_id']
                for r in con.execute("SELECT name, login_id FROM users WHERE is_active=1").fetchall()}

    # 2) 조직도 항목 → 행 빌드
    rows = []
    not_found = []
    for dept, rank, name, is_leader in ORG:
        login_id = db_users.get(name)
        if not login_id:
            not_found.append(name)
            login_id = '(미등록)'
        rows.append((dept, rank, name, login_id, PW, is_leader))

    # 3) 조직도엔 없지만 DB 에 활성 등록된 인원 → 시트 끝에 "기타 등재"
    chart_names = {n for _, _, n, _ in ORG}
    extras = []
    for r in con.execute(
        "SELECT u.name, u.login_id, u.role, t.name as team_name "
        "FROM users u LEFT JOIN teams t ON u.team_id=t.id "
        "WHERE u.is_active=1 ORDER BY u.id"
    ).fetchall():
        if r['name'] in chart_names: continue
        if r['name'] == '시스템관리자': continue
        extras.append((r['team_name'] or '미배정', '(DB)' + (r['role'] or ''),
                       r['name'], r['login_id'], PW, False))
    con.close()

    # 4) Excel 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "전직원 로그인 계정"

    # 헤더
    headers = ["부서", "직급", "이름", "아이디", "비밀번호"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor="3F5C44")  # sage-700
        c.alignment = Alignment(horizontal="center", vertical="center")

    # 본문
    body = rows + ([(None, None, None, None, None, False)] + extras if extras else [])
    cur_dept = None
    for i, item in enumerate(body, 2):
        dept, rank, name, lid, pw, is_leader = item
        if dept is None:
            # 구분선
            for col in range(1, 6):
                cc = ws.cell(row=i, column=col, value="—")
                cc.fill = PatternFill("solid", fgColor="F0F5EB")
                cc.alignment = Alignment(horizontal="center")
            continue
        ws.cell(row=i, column=1, value=dept)
        ws.cell(row=i, column=2, value=rank)
        ws.cell(row=i, column=3, value=name)
        ws.cell(row=i, column=4, value=lid)
        ws.cell(row=i, column=5, value=pw)
        # 팀장 강조
        if is_leader:
            for col in range(1, 6):
                ws.cell(row=i, column=col).font = Font(bold=True, color="A5282C")
        # 부서 변경 시 옅은 구분 음영
        if dept != cur_dept:
            for col in range(1, 6):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor="F7FAF4")
            cur_dept = dept

    # 컬럼 폭
    widths = [22, 10, 12, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # 테두리
    thin = Side(border_style="thin", color="CAD9B8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=5):
        for c in r:
            c.border = border
            if c.row > 1:
                c.alignment = Alignment(horizontal="left", vertical="center")

    # 두 번째 시트 — 안내
    ws2 = wb.create_sheet("안내")
    ws2['A1'] = "KNK HAIST WORKS · 전직원 로그인 계정 (2026-04-28 일괄 리셋)"
    ws2['A1'].font = Font(bold=True, size=13, color="3F5C44")
    notes = [
        "",
        "▣ 일괄 리셋 사유: 운영테스트팀 검증 + 신규 ID 체계(로마자 첫글자) 통일",
        "▣ 비밀번호: 전 직원 동일하게 'knk1234' (개인별 변경 전까지)",
        "▣ 아이디 규칙: 한글 이름 각 음절의 첫 자음·모음 1자 (예: 김정락→kjr · 안지연→ajy)",
        "▣ 동일 약자 충돌 시 숫자 접미사 (예: 김정락=kjr / 김정록=kjr2)",
        "",
        f"▣ 총 등록 인원: {len(rows)}명 (조직도 기준)" + (f" + DB 추가 {len(extras)}명" if extras else ""),
        f"▣ DB 미등록 인원(이름 일치 안 됨): {len(not_found)}명 → " + (', '.join(not_found) if not_found else '없음'),
        "",
        "▣ 첫 로그인 후 권장 액션:",
        "   1) /profile 에서 본인 비밀번호 변경",
        "   2) 이메일·전화·언어 설정 확인",
        "",
        "▣ 시스템 URL: http://localhost:8081/login",
        "▣ admin 계정: id=admin / pw=knk1234 (시스템관리자)",
    ]
    for i, t in enumerate(notes, 2):
        ws2.cell(row=i, column=1, value=t).alignment = Alignment(wrap_text=True)
    ws2.column_dimensions['A'].width = 100

    wb.save(str(OUT))
    print(f"[OK] {OUT}")
    print(f"     본문 {len(rows)}행 · DB추가 {len(extras)}행 · 미등록 {len(not_found)}명")
    if not_found:
        print(f"     미등록: {not_found}")

if __name__ == '__main__':
    main()
