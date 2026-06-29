"""
v5H142 (2026-05-05) — 소모품 발주 전용 도메인
대표 직접 요청: 신규 검사기와 분리 / 관리번호 발급 X / 엑셀 일괄 import + 이미지 자동 압축

핵심 헬퍼:
  - parse_consumable_xlsx(file_path)          : 엑셀 → 라인 list + 이미지 매칭
  - compress_image_bytes(raw, max_dim, quality): bytes → (압축본 bytes, 썸네일 bytes)
  - generate_co_no()                           : CO-YYMMNN 자동 채번
  - match_part_by_name(name)                   : 자재 마스터 LIKE 매칭
  - match_project_by_model(model_use)          : projects.model_name LIKE 매칭
  - co_create / coi_bulk_insert / co_get / coi_list / coi_update / coi_delete
  - recompute_co_total(co_id)
"""
from __future__ import annotations
from io import BytesIO
from datetime import datetime
import os, re, json, shutil

from .database import db_session

# ────────────────────────────────────────────────────────────────────
# 채번 / 상태
# ────────────────────────────────────────────────────────────────────
CO_STATUSES = ["DRAFT", "QUOTED", "CONFIRMED", "SHIPPED", "PAID", "CANCELLED", "HOLD"]
CO_STATUS_LABELS = {
    "DRAFT": "작성중",
    "QUOTED": "견적완료",
    "CONFIRMED": "발주확정",
    "SHIPPED": "출하",
    "PAID": "수금완료",
    "CANCELLED": "취소",
    "HOLD": "보류",   # v5H226z719 (대표 지시): 소모품 '보류' 상태 신설(일괄등록 양식 상태 드롭다운)
}


def generate_co_no(biz_div: str = "M", today=None) -> str:
    """v5H226z248→z338 (대표 확정): 소모품 발주번호 = 수주번호와 동일 형식 [사업부]-[YYMMDD],
    앞글자는 소모품 'C' 고정 (예: C-260602, C-260602-1). 같은 날 첫 건은 접미 없음, 두 번째부터 -1, -2 순차.
    수주번호(orders.order_no)와 충돌 방지를 위해 orders + consumable_orders 양쪽을 스캔."""
    # v5H226z338 (대표 확정): 소모품 발주번호 앞글자 = 소모품 'C' 고정 (수주번호 사업부글자 규칙: 소모품=C).
    #   z248 의 '진행 사업부 T/M/L 사용'을 'C-' 로 통일. 진행 사업부는 consumable_orders.biz_div 컬럼에 별도 보존
    #   (co_no 앞글자로 사업부 판별하는 로직 없음 — 집계는 biz_div 컬럼 기준, 회귀 없음).
    bd = "C"
    d = today or datetime.now()
    yymmdd = d.strftime("%y%m%d")
    base = f"{bd}-{yymmdd}"
    nums = []
    with db_session() as c:
        for tbl, col in (("orders", "order_no"), ("consumable_orders", "co_no")):
            try:
                for r in c.execute(
                    f"SELECT {col} FROM {tbl} WHERE {col}=? OR {col} LIKE ?",
                    (base, base + "-%")
                ):
                    if r[0]:
                        nums.append(r[0])
            except Exception:
                pass
    if not nums:
        return base  # 같은 날 첫 건 → 접미 없음
    max_n = 0
    for on in nums:
        if on == base:
            continue
        m = re.match(rf"^{re.escape(base)}-(\d+)$", on)
        if m:
            max_n = max(max_n, int(m.group(1)))
    return f"{base}-{max_n + 1}"


# ────────────────────────────────────────────────────────────────────
# 엑셀 파싱 — 헤더 자동 감지 + 라인 + 이미지
# ────────────────────────────────────────────────────────────────────
# 헤더 매칭 — 우선순위 순서로 평가 (먼저 매칭된 컬럼은 다른 키에 양보)
# 더 구체적인 키워드(ORDER DATE)를 더 일반적인 것(QTY) 보다 먼저 둠
HEADER_KEYS = [
    ("order_date", ["ORDERDATE", "발주일", "요청일"]),
    ("part_name",  ["SUPPLIERNAME", "품명", "PARTNAME", "ITEMNAME"]),
    ("model_use",  ["MODELUSE", "모델"]),
    ("equip",      ["장비명", "EQUIP"]),                  # v5H226z288: 장비명(미리보기 엑셀 일치)
    ("qty",        ["Q'TY", "QTY", "수량", "QUANTITY"]),
    ("unit",       ["UNIT", "단위"]),
    # v5H226z287: 연결 관리번호(라인별) — '번호'보다 먼저 둬야 '연결관리번호'가 NO 로 안 잡힘
    ("link_mgmt",  ["연결관리번호", "연결관리"]),
    ("no",         ["NO", "번호", "순번"]),
    ("supplier",   ["업체", "VENDOR"]),
    ("spec",       ["SPEC", "규격", "BOM"]),
    # v5H226z285: 단가·금액 (머리글이 '단가 (KRW)' 처럼 통화 포함 수식이어도 '단가'/'금액'으로 인식)
    ("price",      ["단가", "UNITPRICE"]),
    ("amount",     ["금액", "AMOUNT"]),
    ("note",       ["비고", "REMARK"]),                   # v5H226z288: 비고
    # MODEL 단독은 우선순위 낮춤 (MODELUSE 가 못 잡혔을 때만)
    ("model_use",  ["MODEL"]),
]

# v5H226z285: 통화 화이트리스트 + 정규화 (정보란 '통화' 값/머리글 '(KRW)' 에서 추출)
_CCY_WHITELIST = {"KRW", "USD", "VND", "JPY", "CNY", "EUR"}
_CCY_ALIAS = {"₩": "KRW", "원": "KRW", "$": "USD", "달러": "USD", "USD": "USD",
              "₫": "VND", "동": "VND", "¥": "JPY", "엔": "JPY", "JPY": "JPY",
              "€": "EUR", "유로": "EUR", "위안": "CNY", "CNY": "CNY", "RMB": "CNY"}


def _norm_ccy(v):
    """통화 표기 → 코드(KRW 등). 못 알아보면 None."""
    if v is None:
        return None
    s = str(v).strip().upper()
    if s in _CCY_WHITELIST:
        return s
    if s in _CCY_ALIAS:
        return _CCY_ALIAS[s]
    m = re.search(r"[A-Z]{3}", s)
    if m and m.group(0) in _CCY_WHITELIST:
        return m.group(0)
    for k, code in _CCY_ALIAS.items():
        if k in s:
            return code
    return None


_INFO_LABEL_HINTS = ("고객사", "담당자", "연락처", "통화", "거래구분", "납품위치", "납품처",
                     "발주일", "납품일", "납기", "모델명", "장비명", "관리번호", "수주번호",
                     "프로젝트명", "KNK")


def _right_value(ws, r, c, maxc, span=4):
    """라벨(r,c) 바로 오른쪽(최대 span칸) 첫 비어있지 않은 값 — 라벨 옆 값 읽기.
    v5H226z287: 옆 칸이 비어 있으면 '멀리 있는 다른 블록 라벨'을 값으로 오인하지 않게 거리 제한 +
    값이 또 다른 라벨이면 무시(값 없음으로 처리)."""
    end = min(maxc, c + span)
    for cc in range(c + 1, end + 1):
        v = ws.cell(r, cc).value
        if v is None or str(v).strip() == "":
            continue
        sn = re.sub(r"\s+", "", str(v).strip())
        if any(lb in sn for lb in _INFO_LABEL_HINTS):
            return None
        return v
    return None


def read_order_meta(ws, header_row, col_map=None):
    """헤더 위 정보란(좌·우)에서 주문 단위 정보를 라벨 기반으로 모두 읽기 — 셀 위치 무관.
    관리번호·수주번호는 시스템 자동 부여(읽지 않음). 프로젝트명='소모품'은 무시.
    더 구체적인 라벨(연락처·2차 고객사·KNK 담당자)을 먼저 평가해 오인식 방지."""
    meta = {"currency": None, "is_export": None, "customer": None, "customer2": None,
            "ship_to": None, "order_date": None, "due_date": None,
            "cc_name": None, "cc_phone": None, "model_name": None,
            "equip_name": None, "sales_name": None}
    maxc = min(ws.max_column, 30)

    def _sv(x):
        if x is None:
            return None
        s = str(x).strip()
        return s or None

    for r in range(1, max(1, header_row)):
        for c in range(1, maxc + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            tn = re.sub(r"\s+", "", str(v).strip())   # 공백 제거(한글 라벨 비교)
            tU = tn.upper()

            def rv():
                return _right_value(ws, r, c, maxc)

            if meta["currency"] is None and (tn in ("통화", "통화구분") or tU == "CURRENCY"):
                cc = _norm_ccy(rv())
                if cc:
                    meta["currency"] = cc
            elif meta["is_export"] is None and "거래구분" in tn:
                x = rv()
                if x is not None:
                    meta["is_export"] = 1 if ("수출" in str(x) or "EXPORT" in str(x).upper()) else 0
            elif meta["cc_phone"] is None and "연락처" in tn:                 # 고객사 담당자 연락처
                meta["cc_phone"] = _sv(rv())
            elif meta["sales_name"] is None and "담당자" in tn and ("KNK" in tU or "케이엔케이" in tn):
                meta["sales_name"] = _sv(rv())                                # KNK 담당자(영업)
            elif meta["cc_name"] is None and "담당자" in tn:                  # 고객사 담당자
                meta["cc_name"] = _sv(rv())
            elif meta["customer2"] is None and "2차" in tn and "고객사" in tn:
                meta["customer2"] = _sv(rv())
            elif meta["customer"] is None and "고객사" in tn:                 # 1차 고객사 / 고객사
                meta["customer"] = _sv(rv())
            elif meta["ship_to"] is None and tn in ("납품위치", "납품처"):
                meta["ship_to"] = _sv(rv())
            elif meta["order_date"] is None and tn == "발주일":
                meta["order_date"] = _date_str(rv())
            elif meta["due_date"] is None and tn in ("납품일", "납기", "납기일"):
                meta["due_date"] = _date_str(rv())
            elif meta["model_name"] is None and tn == "모델명":
                meta["model_name"] = _sv(rv())
            elif meta["equip_name"] is None and tn == "장비명":
                meta["equip_name"] = _sv(rv())
    # 보조: 단가/금액 머리글의 '(KRW)' 에서 통화 추출
    if meta["currency"] is None and col_map:
        for key in ("price", "amount"):
            if key in col_map:
                cc = _norm_ccy(ws.cell(header_row, col_map[key]).value)
                if cc:
                    meta["currency"] = cc
                    break
    return meta


def _norm(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s)).upper()


def detect_header(ws, max_scan_rows: int = 16) -> tuple[int, dict]:
    """헤더 row 자동 감지. (header_row, col_map) 반환.
    v5H226z283/z285: KNK 표준 양식은 정보란(통화·거래구분 추가로 확장) 아래에 헤더 → 스캔 16행.
    '가장 키 많은 행'을 헤더로 택하므로 상단 라벨(발주일·모델명 메모 등)에 오탐하지 않음.
    col_map = {'no': col_idx, 'part_name': col_idx, ...}  (1-indexed col).
    한 컬럼은 1개 키에만 매핑 (충돌 방지)."""
    best_row = 0
    best_map: dict = {}
    for r in range(1, min(max_scan_rows, ws.max_row) + 1):
        cur_map: dict = {}
        used_cols: set = set()
        for c in range(1, min(ws.max_column, 30) + 1):
            v = _norm(ws.cell(r, c).value)
            if not v:
                continue
            for key, kws in HEADER_KEYS:
                if key in cur_map:
                    continue
                if c in used_cols:
                    continue
                for kw in kws:
                    if _norm(kw) in v:
                        cur_map[key] = c
                        used_cols.add(c)
                        break
        if "part_name" in cur_map and ("qty" in cur_map or "model_use" in cur_map):
            if len(cur_map) > len(best_map):
                best_row = r
                best_map = cur_map
    return best_row, best_map


def parse_consumable_xlsx(file_path: str, image_out_dir: str | None = None) -> dict:
    """엑셀 파싱 → {'lines': [...], 'images': {row:[paths]}, 'header_row': N, 'col_map': {...}}.
    image_out_dir 가 주어지면 이미지 추출 후 압축본/썸네일 저장."""
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True)
    ws = wb.worksheets[0]
    hdr_row, col_map = detect_header(ws)
    lines = []
    if hdr_row == 0 or "part_name" not in col_map:
        return {"lines": [], "images": {}, "header_row": 0, "col_map": {},
                "error": "헤더를 찾지 못했습니다 (NO/MODEL/품명/수량 컬럼이 1~16행 안에 있어야 합니다)"}
    pn_col = col_map["part_name"]
    line_no_seq = 0
    for r in range(hdr_row + 1, ws.max_row + 1):
        pn = ws.cell(r, pn_col).value
        if pn is None or str(pn).strip() == "":
            continue
        line_no_seq += 1
        rec = {
            "row": r,                                           # 엑셀 row (이미지 매칭 키)
            "line_no": line_no_seq,
            "model_use": str(ws.cell(r, col_map["model_use"]).value or "").strip() if "model_use" in col_map else "",
            "part_name": str(pn).strip(),
            "spec":      str(ws.cell(r, col_map["spec"]).value or "").strip() if "spec" in col_map else "",
            "qty":       _to_num(ws.cell(r, col_map["qty"]).value) if "qty" in col_map else 0,
            "unit":      str(ws.cell(r, col_map["unit"]).value or "EA").strip() if "unit" in col_map else "EA",
            "order_date": _date_str(ws.cell(r, col_map["order_date"]).value) if "order_date" in col_map else "",
            "unit_price": _to_num(ws.cell(r, col_map["price"]).value) if "price" in col_map else 0,
            "amount":     _to_num(ws.cell(r, col_map["amount"]).value) if "amount" in col_map else 0,
            # v5H226z287: 연결 관리번호(사용자가 직접 적은 호기/프로젝트 관리번호)
            "link_mgmt":  str(ws.cell(r, col_map["link_mgmt"]).value or "").strip() if "link_mgmt" in col_map else "",
            # v5H226z288: 장비명·비고 (미리보기 엑셀 일치)
            "equip":      str(ws.cell(r, col_map["equip"]).value or "").strip() if "equip" in col_map else "",
            "note":       str(ws.cell(r, col_map["note"]).value or "").strip() if "note" in col_map else "",
        }
        # v5H226z285: 금액 비고 단가·수량 있으면 금액=단가×수량 (수식 캐시 없을 때 보강)
        if not rec["amount"] and rec.get("unit_price") and rec.get("qty"):
            rec["amount"] = rec["unit_price"] * rec["qty"]
        lines.append(rec)
    # 이미지 추출 + 압축
    img_map: dict = {}
    if image_out_dir:
        os.makedirs(image_out_dir, exist_ok=True)
        # v5H226z284 (대표 지시): 행 높이 누적(EMU)으로 이미지 세로 '중심'이 속한 데이터 행에 매칭.
        #   셀 크기·병합·붙인 위치가 달라도 '보이는 위치' 그대로 연결(기존 '앵커행+2'는 전 사진이 2줄 밀렸음).
        _ymax = ws.max_row
        for _im in (getattr(ws, "_images", []) or []):
            _t = getattr(_im.anchor, "to", None) or getattr(_im.anchor, "_to", None)
            if _t is not None and getattr(_t, "row", None) is not None and (_t.row + 3) > _ymax:
                _ymax = _t.row + 3
        _P = _row_y_prefix(ws, _ymax)
        # v5H226z286 (대표 지시): 엑셀의 두 사진 칸 구분 — 사진(PICTURE)=품목사진 / 사진위치(LOCATION)=참고.
        _photo_col, _loc_col = _find_photo_cols(ws, hdr_row)
        try:
            for idx, img in enumerate(getattr(ws, "_images", []) or []):
                try:
                    a = img.anchor
                    raw = img._data()
                    if not raw:
                        continue
                    # 이미지 세로 중심 → 해당 데이터 행. 못 구하면(절대앵커 등) 앵커 상단 행 최근접 폴백.
                    _cy = _image_center_y(a, _P)
                    matched_line = (_line_for_center(lines, _cy, _P) if _cy is not None
                                    else _find_nearest_line(lines, getattr(getattr(a, "_from", None), "row", 0) + 1))
                    if matched_line is None:
                        continue
                    # 이미지가 어느 열(사진/사진위치)에 있는지로 분류
                    _frm = getattr(a, "_from", None)
                    _col = (_frm.col + 1) if (_frm is not None and getattr(_frm, "col", None) is not None) else None
                    _cat = "photo"
                    if _col is not None and _loc_col is not None:
                        if _photo_col is not None:
                            _cat = "loc" if abs(_col - _loc_col) < abs(_col - _photo_col) else "photo"
                        elif _col == _loc_col:
                            _cat = "loc"
                    fn = f"line_{matched_line['line_no']:03d}_{idx+1}.jpg"
                    fn_thumb = f"line_{matched_line['line_no']:03d}_{idx+1}_thumb.jpg"
                    full = os.path.join(image_out_dir, fn)
                    thumb = os.path.join(image_out_dir, fn_thumb)
                    big_bytes, thumb_bytes, info = compress_image_bytes(raw)
                    with open(full, "wb") as f:
                        f.write(big_bytes)
                    with open(thumb, "wb") as f:
                        f.write(thumb_bytes)
                    img_map.setdefault(matched_line["line_no"], []).append({
                        "full": fn, "thumb": fn_thumb, "category": _cat,
                        "orig_size": len(raw), "compressed": len(big_bytes),
                        "info": info,
                    })
                except Exception as e:
                    img_map.setdefault("_errors", []).append(f"img{idx}: {e}")
        except Exception as e:
            img_map["_errors"] = [str(e)]
        # 사진(품목) 우선 정렬 → 미리보기 대표 사진 = 품목 사진
        for _k, _v in img_map.items():
            if isinstance(_v, list):
                _v.sort(key=lambda x: 0 if x.get("category") == "photo" else 1)
    meta = read_order_meta(ws, hdr_row, col_map)   # v5H226z285: 통화·거래구분
    return {"lines": lines, "images": img_map,
            "header_row": hdr_row, "col_map": col_map, "meta": meta,
            "image_count": sum(len(v) for k, v in img_map.items() if isinstance(v, list) and k != "_errors")}


def _find_nearest_line(lines, anchor_row: int):
    """anchor_row 와 가장 가까운(<=) line.row 찾기. 없으면 가장 가까운 라인. (절대앵커 폴백용)"""
    if not lines:
        return None
    candidates = [ln for ln in lines if ln["row"] <= anchor_row + 2]
    if candidates:
        return max(candidates, key=lambda ln: ln["row"])
    return min(lines, key=lambda ln: abs(ln["row"] - anchor_row))


# ── v5H226z284: 이미지 기하학적 매칭(행 높이 기반) — 셀 크기·병합·오프셋 무관 ──
def _row_y_prefix(ws, max_row: int) -> list:
    """행별 세로 위치 누적(EMU). P[r] = 1..r행 높이 합(기본 15pt).
    이미지 앵커의 row(0-index)+rowOff(EMU)로 절대 Y좌표를 계산하는 데 사용."""
    DEFAULT_PT = 15.0
    EMU_PER_PT = 12700.0
    n = max(1, int(max_row)) + 2
    P = [0.0] * (n + 1)
    for r in range(1, n + 1):
        try:
            h = ws.row_dimensions[r].height
        except Exception:
            h = None
        P[r] = P[r - 1] + (h if h else DEFAULT_PT) * EMU_PER_PT
    return P


def _image_center_y(anchor, P):
    """이미지의 세로 '중심' Y(EMU). TwoCellAnchor=from~to 중점,
    OneCellAnchor=from + 높이/2. 계산 불가 시 None."""
    frm = getattr(anchor, "_from", None)
    if frm is None or getattr(frm, "row", None) is None or frm.row + 1 >= len(P):
        return None
    top = P[frm.row] + (getattr(frm, "rowOff", 0) or 0)
    to = getattr(anchor, "to", None) or getattr(anchor, "_to", None)
    if to is not None and getattr(to, "row", None) is not None and to.row + 1 < len(P):
        bot = P[to.row] + (getattr(to, "rowOff", 0) or 0)
    else:
        ext = getattr(anchor, "ext", None)
        cy = getattr(ext, "height", 0) if ext is not None else 0
        bot = top + (cy or 0)
    return (top + bot) / 2.0


def _line_for_center(lines, center_y, P):
    """center_y(EMU)가 속한 데이터 행의 라인. 구간 밖이면 행 밴드 중심 최근접."""
    if not lines or center_y is None:
        return None
    for ln in lines:
        R = ln["row"]
        if R < len(P) and P[R - 1] <= center_y < P[R]:
            return ln
    def _mid(ln):
        R = ln["row"]
        return (P[R - 1] + P[R]) / 2.0 if R < len(P) else P[-1]
    return min(lines, key=lambda ln: abs(_mid(ln) - center_y))


def _find_photo_cols(ws, header_row):
    """헤더에서 '사진(PICTURE)' 열과 '사진위치(PICTURE LOCATION)' 열 식별.
    (photo_col, loc_col) — 1-indexed, 없으면 None. LOCATION/위치 포함=위치, 아니면 사진."""
    photo_col, loc_col = None, None
    maxc = min(ws.max_column, 30)
    for c in range(1, maxc + 1):
        v = _norm(ws.cell(header_row, c).value)
        if not v:
            continue
        if not (("사진" in v) or ("PICTURE" in v) or ("PHOTO" in v) or ("이미지" in v) or ("IMAGE" in v)):
            continue
        if ("LOCATION" in v) or ("위치" in v):
            if loc_col is None:
                loc_col = c
        elif photo_col is None:
            photo_col = c
    return photo_col, loc_col


def _to_num(v):
    if v is None:
        return 0
    try:
        return float(v)
    except (TypeError, ValueError):
        try:
            return float(str(v).replace(",", "").strip())
        except Exception:
            return 0


def _date_str(v):
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def _split_date_bundle(v):
    """v5H226z491b (대표 지시): 세금계산서 발행일 셀이 'YYYY-MM-DD-N'(예 2026-03-23-1)이면
    날짜와 묶음번호(N)로 분리. 같은 '날짜-N'끼리 = 같은 날짜에 한 장으로 묶어 발행한 묶음.
    반환=(깨끗한 날짜 'YYYY-MM-DD', 묶음번호 '' 또는 'N'). 일반 날짜/날짜객체는 (_date_str, '')."""
    import re as _re
    s = _date_str(v)
    m = _re.match(r"^\s*(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})\s*-\s*(\d+)\s*$", s)
    if m:
        y, mo, d, b = m.group(1), int(m.group(2)), int(m.group(3)), m.group(4)
        return f"{y}-{mo:02d}-{d:02d}", b
    return s, ""


# ────────────────────────────────────────────────────────────────────
# 이미지 압축 (Pillow)
# ────────────────────────────────────────────────────────────────────
def compress_image_bytes(raw: bytes, max_dim: int = 1920, quality: int = 92,
                          thumb_dim: int = 320) -> tuple[bytes, bytes, dict]:
    """원본 bytes → (압축본 jpeg bytes, 썸네일 jpeg bytes, info dict)
    v5H226z295(대표 지시): 화질 향상 — 긴변 1920px JPEG q=92, 썸네일 320×160(2:1 와이드) q=92."""
    from PIL import Image, ImageOps
    # v5H226z310 (적대검토): 압축폭탄(decompression bomb) 방어 — 디코드(exif_transpose/resize) 전에
    #   픽셀 수 상한 검사. 작은 용량이라도 거대 비트맵으로 풀리면 OOM. 120MP=고해상 폰사진까지 허용.
    Image.MAX_IMAGE_PIXELS = 120_000_000
    im = Image.open(BytesIO(raw))
    if (im.width or 0) * (im.height or 0) > 120_000_000:
        raise ValueError("이미지 픽셀 수가 너무 큽니다(최대 120MP)")
    im = ImageOps.exif_transpose(im)
    orig_size = im.size
    if im.mode in ("RGBA", "LA"):
        bg = Image.new("RGB", im.size, (255, 255, 255))
        bg.paste(im, mask=im.split()[-1])
        im = bg
    elif im.mode == "P":
        im = im.convert("RGBA")
        bg = Image.new("RGB", im.size, (255, 255, 255))
        bg.paste(im, mask=im.split()[-1])
        im = bg
    elif im.mode != "RGB":
        im = im.convert("RGB")
    # 풀사이즈
    big = im.copy()
    if max(big.size) > max_dim:
        ratio = max_dim / max(big.size)
        big = big.resize((max(1, int(big.size[0] * ratio)),
                          max(1, int(big.size[1] * ratio))), Image.LANCZOS)
    out_big = BytesIO()
    big.save(out_big, "JPEG", quality=quality, optimize=True, progressive=True)
    big_bytes = out_big.getvalue()
    # 썸네일 — v5H226z293/z295(대표 지시): 동일 크기·비율로 가운데 잘라 균일화 + 와이드.
    #   320×160 (2:1 — 가로 넓고 세로 낮게). 원본은 big 에 비율 유지 보관 → 클릭 시 전체 표시.
    thumb_w, thumb_h = thumb_dim, max(1, round(thumb_dim / 2))       # 320×160 (2:1)
    thumb = ImageOps.fit(im, (thumb_w, thumb_h), Image.LANCZOS, centering=(0.5, 0.5))
    out_thumb = BytesIO()
    thumb.save(out_thumb, "JPEG", quality=92, optimize=True)
    thumb_bytes = out_thumb.getvalue()
    info = {
        "orig_dim": orig_size,
        "compressed_dim": big.size,
        "thumb_dim": thumb.size,
    }
    return big_bytes, thumb_bytes, info


# ────────────────────────────────────────────────────────────────────
# 자동 매칭
# ────────────────────────────────────────────────────────────────────
def match_part_by_name(name: str) -> dict | None:
    """parts.part_name 정확일치(+공백/대소문자 정규화)만 자동연결.
    v5H226z262 (대표 지시·연결성 감사): 근본원인 = '첫 단어 LIKE'가 고객사·규격 스코프 없이
      흔한 토큰('CABLE'/'SCREW'/'M3')으로 '아무 자재 1건'에 part_id 를 조용히 붙이던 것.
      재발방지 원칙: **자동연결은 명확한 단일 후보(정확일치)일 때만**. 애매하면 연결하지 않고
      미매칭으로 둬 사용자가 수동 연결하게 한다. (틀린 연결보다 미연결이 안전)
    """
    if not name or not str(name).strip():
        return None
    nm = str(name).strip()
    # 정규화 키(앞뒤·중복 공백 제거 + 대문자) — '동일 품명'의 표기차만 흡수, 다른 품목은 매칭 안 함
    import re as _re_p
    nm_key = _re_p.sub(r"\s+", " ", nm).strip().upper()
    with db_session() as c:
        rows = c.execute(
            "SELECT id, part_no, part_name, std_price, unit FROM parts "
            "WHERE UPPER(TRIM(part_name))=? AND COALESCE(is_active,1)=1", (nm_key,)
        ).fetchall()
        # 정확일치가 '정확히 1건'일 때만 자동연결 (동명 자재 2건 이상이면 애매 → 미매칭)
        if len(rows) == 1:
            d = dict(rows[0]); d["match_level"] = "exact"; return d
    return None


def match_project_by_model(model_use: str, customer_id: int | None = None) -> dict | None:
    """projects.model_name (또는 name) 토큰 매칭. NEW_EQUIP 만 후보.
    v5H226z252: '같은 고객사'로 스코프 — 다른 회사 모델로 오연결되던 문제 수정.
    v5H226z262 (대표 지시·연결성 감사): 근본원인 = 같은 고객사가 동일 모델을 여러 호기 보유하면
      'ORDER BY id DESC'가 무조건 '가장 최근 등록 호기'에 붙이던 것(실제 어느 호기용인지 무관).
      재발방지 원칙: **후보가 정확히 1건일 때만 자동연결**. 2건 이상(애매)이면 연결하지 않고
      미매칭으로 둬 사용자가 어느 호기인지 직접 고르게 한다.
      · customer_id 없으면 자동연결 안 함 / 토큰 2자 미만 매칭 안 함(과매칭 방지)
    """
    if not model_use or not str(model_use).strip():
        return None
    if not customer_id:
        return None  # 고객사 모르면 자동연결 안 함(오연결 방지) → 수동 연결 버튼 사용
    mu = str(model_use).strip()
    # 첫 모델 토큰 (예: 'SM-A576B CTC AUTO' → 'SM-A576B')
    token = mu.split()[0].strip()
    if len(token) < 2:
        return None
    with db_session() as c:
        # 같은 고객사·같은 토큰 후보를 2건까지 조회 → 정확히 1건일 때만 연결(애매하면 보류)
        rows = c.execute(
            "SELECT id, mgmt_code, name, model_name FROM projects "
            "WHERE COALESCE(project_type,'NEW_EQUIP')='NEW_EQUIP' "
            "  AND customer_id = ? "
            "  AND (model_name LIKE ? OR name LIKE ?) "
            "ORDER BY id DESC LIMIT 2",
            (customer_id, f"%{token}%", f"%{token}%")
        ).fetchall()
        if len(rows) == 1:
            d = dict(rows[0]); d["match_level"] = "model_token"; return d
        # len==0 → 미매칭 / len>=2 → 애매(여러 호기) → 자동연결 보류(수동 연결 유도)
    return None


def match_project_by_mgmt(mgmt_code: str) -> dict | None:
    """v5H226z287 (대표 지시): 사용자가 엑셀 '연결 관리번호'에 적은 관리코드로 프로젝트 직접 연결.
    정확일치 1건일 때만 (명시적 입력이므로 모델 추정보다 우선)."""
    if not mgmt_code or not str(mgmt_code).strip():
        return None
    mc = str(mgmt_code).strip().upper()
    with db_session() as c:
        r = c.execute(
            "SELECT id, mgmt_code, name, model_name FROM projects WHERE UPPER(mgmt_code)=? LIMIT 1",
            (mc,)
        ).fetchone()
        if r:
            d = dict(r); d["match_level"] = "mgmt_code"; return d
    return None


def _norm_company(s: str) -> str:
    """상호 정규화 — 법인격((주)/㈜/주식회사/유한회사 등)·괄호·기호·공백 제거 후 소문자."""
    s = (s or "").lower()
    s = re.sub(r"주식회사|유한회사|유한책임회사|\(주\)|\(유\)|\(有\)|㈜|㈲|株式会社", "", s)
    s = re.sub(r"[\s\(\)\[\]\.,\-_/&'\"·｜|]", "", s)
    return s.strip()


def match_customer_by_name(name: str) -> dict | None:
    """v5H226z294 (대표 지시): 엑셀 고객사명 → 등록 고객사 매칭.
    (주)/㈜/주식회사/공백 차이는 무시하고 '상호'로 비교. 원문 정확일치 우선,
    아니면 정규화 상호가 '정확히 1건' 일치할 때만 연결(애매하면 None — 연결성 안전 원칙).
    반환: {'id', 'name'(등록 정식명칭)} 또는 None."""
    if not name or not str(name).strip():
        return None
    raw = str(name).strip()
    n = _norm_company(raw)
    if len(n) < 2:
        return None
    with db_session() as c:
        rows = [dict(r) for r in c.execute("SELECT id, name FROM customers").fetchall()]
    for r in rows:                                   # 1) 원문 정확 일치 우선
        if (r["name"] or "").strip() == raw:
            return {"id": r["id"], "name": r["name"]}
    matches = [r for r in rows if _norm_company(r["name"]) == n]   # 2) 정규화 상호 일치
    if len(matches) == 1:                            # 정확히 1건일 때만(애매하면 미연결)
        return {"id": matches[0]["id"], "name": matches[0]["name"]}
    return None


# ────────────────────────────────────────────────────────────────────
# CRUD
# ────────────────────────────────────────────────────────────────────
def co_create(customer_name: str = "", biz_div: str = "",
              order_date: str = "", due_date: str = "",
              currency: str = "KRW", note: str = "", source_file: str = "",
              created_by: int | None = None) -> tuple[int, str]:
    """v5H216: 소모품 묶음 생성 시 'S' prefix 관리번호 자동 발급.
    v5H218: biz_div(T/M) 추가 — 진행 사업부 별 매출 집계용."""
    # v5H226z248: 발주번호 = 수주번호 형식 [사업부]-[YYMMDD]. 날짜는 발주일 기준(수주번호 관례).
    _co_ref = None
    try:
        if order_date:
            _co_ref = datetime.strptime(str(order_date)[:10], "%Y-%m-%d")
    except Exception:
        _co_ref = None
    # 고객사·사업부는 1회 계산 (재시도와 무관)
    cust_id = None
    if customer_name:
        with db_session() as c:
            r = c.execute("SELECT id FROM customers WHERE name=? LIMIT 1",
                          (customer_name,)).fetchone()
            if r:
                cust_id = r[0]
    _bd = (biz_div or "").strip().upper()
    if _bd not in ("T", "M", "L"):
        _bd = None  # 미선택 허용 (백워드 호환); UI 에서는 검증 강제
    # v5H226z262 (대표 지시·연결성 감사): co_no/mgmt_code UNIQUE 충돌(동시 업로드 race) 시
    #   재발급 후 재시도. 부분 UNIQUE 인덱스(uq_consumable_co_no/mgmt)가 충돌을 IntegrityError 로
    #   드러내면, generate_* 가 (직전에 커밋된 경쟁 행을 보고) 다음 번호를 발급 → 충돌 해소.
    import sqlite3 as _sqlite
    _last_err = None
    for _attempt in range(6):
        co_no = generate_co_no(biz_div, _co_ref)
        # v5H226z244: 소모품 묶음 생성(업로드) 즉시 C 관리코드 부여 — 프로젝트 C소모품 경로와 동일 체계.
        # v5H226z564 (대표 지시): 관리번호도 수주번호처럼 '발주일' 기준 연월로 발번(과거 데이터 대량 이관 시
        #   현재월 2606 으로 몰리던 문제). _co_ref 없으면(발주일 누락) 현재월 폴백.
        try:
            from .database import generate_mgmt_code
            mgmt_code = generate_mgmt_code("C", _co_ref)
        except Exception:
            mgmt_code = None
        try:
            with db_session() as c:
                cocols = {r2[1] for r2 in c.execute("PRAGMA table_info(consumable_orders)").fetchall()}
                _has_mgmt = "mgmt_code" in cocols
                _has_biz = "biz_div" in cocols
                if _has_mgmt and _has_biz:
                    cur = c.execute(
                        """INSERT INTO consumable_orders
                           (co_no, mgmt_code, biz_div, customer_id, customer_name, order_date, due_date,
                            currency, note, source_file, created_by)
                           VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                        (co_no, mgmt_code, _bd, cust_id, customer_name, order_date or "", due_date or "",
                         (currency or "KRW").upper(), note or "", source_file or "",
                         created_by)
                    )
                elif _has_mgmt:
                    cur = c.execute(
                        """INSERT INTO consumable_orders
                           (co_no, mgmt_code, customer_id, customer_name, order_date, due_date,
                            currency, note, source_file, created_by)
                           VALUES (?,?,?,?,?,?,?,?,?,?)""",
                        (co_no, mgmt_code, cust_id, customer_name, order_date or "", due_date or "",
                         (currency or "KRW").upper(), note or "", source_file or "",
                         created_by)
                    )
                else:
                    cur = c.execute(
                        """INSERT INTO consumable_orders
                           (co_no, customer_id, customer_name, order_date, due_date,
                            currency, note, source_file, created_by)
                           VALUES (?,?,?,?,?,?,?,?,?)""",
                        (co_no, cust_id, customer_name, order_date or "", due_date or "",
                         (currency or "KRW").upper(), note or "", source_file or "",
                         created_by)
                    )
                return int(cur.lastrowid), co_no
        except _sqlite.IntegrityError as _ie:
            _last_err = _ie  # co_no/mgmt_code UNIQUE 충돌 → 재발급 후 재시도
            continue
    raise RuntimeError(f"소모품 수주번호/관리코드 발급 충돌 — 재시도 초과: {_last_err}")


def coi_bulk_insert(co_id: int, items: list[dict]) -> int:
    """라인 일괄 INSERT. items: [{line_no, model_use, part_name, spec, qty, unit,
                                  unit_price, part_id, linked_project_id,
                                  image_path, image_thumb_path, note}, ...]"""
    n = 0
    with db_session() as c:
        # v5H226z288: 존재하는 컬럼만 골라 동적 INSERT (추가형 마이그레이션 안전 — equip_name·image_loc_* 등)
        cols_avail = {r2[1] for r2 in c.execute("PRAGMA table_info(consumable_order_items)").fetchall()}
        for it in items:
            qty = float(it.get("qty") or 0)
            up = float(it.get("unit_price") or 0)
            amt = round(qty * up, 2)
            data = {
                "co_id": int(co_id),
                "line_no": int(it.get("line_no") or 0),
                "model_use": (it.get("model_use") or "").strip(),
                "equip_name": (it.get("equip") or "").strip(),
                "part_id": (int(it["part_id"]) if it.get("part_id") else None),
                "part_name": (it.get("part_name") or "").strip(),
                "spec": (it.get("spec") or "").strip(),
                "qty": qty,
                "unit": (it.get("unit") or "EA").strip(),
                "unit_price": up,
                "amount": amt,
                "linked_project_id": (int(it["linked_project_id"]) if it.get("linked_project_id") else None),
                "note": (it.get("note") or "").strip(),
                "image_path": (it.get("image_path") or None),
                "image_thumb_path": (it.get("image_thumb_path") or None),
                "image_loc_path": (it.get("image_loc_path") or None),
                "image_loc_thumb_path": (it.get("image_loc_thumb_path") or None),
            }
            use = [(k, v) for k, v in data.items() if k in cols_avail]
            cols = ", ".join(k for k, _ in use)
            ph = ", ".join("?" for _ in use)
            c.execute(f"INSERT INTO consumable_order_items ({cols}) VALUES ({ph})",
                      [v for _, v in use])
            n += 1
    recompute_co_total(co_id)
    return n


def recompute_co_total(co_id: int) -> float:
    with db_session() as c:
        row = c.execute(
            "SELECT COALESCE(SUM(amount),0) FROM consumable_order_items WHERE co_id=?",
            (int(co_id),)
        ).fetchone()
        total = float(row[0] or 0)
        c.execute("UPDATE consumable_orders SET total_amount=? WHERE id=?",
                  (total, int(co_id)))
    return total


# ────────────────────────────────────────────────────────────────────
# v5H226z368 (대표 지시): 소모품 '통합 일괄등록' — KNK 표준 소모품 발주 양식(24열) 그대로 업로드.
#   1행=안내문구, 2행=머리글, 3행~=품목. '구분 번호'가 같은 줄들 = 한 발주(여러 품목).
#   품목 사진(PICTURE)·사진위치(PICTURE LOCATION) 임베드 이미지도 함께 가져옴(첨부 표준 양식 기준).
# ────────────────────────────────────────────────────────────────────
CO_BULK_HEADERS = ["연결 관리번호", "구분 번호", "모델명", "장비명", "소모품 품명",
                   "소모품 규격 (SPEC)", "사진(PICTURE LOCATION)", "사진 (PICTURE)",
                   "1차 고객사", "2차 고객사", "발주일", "납품일", "통화", "거래구분", "형태",
                   "수량", "단위", "단가", "금액", "고객사 담당자", "고객사 담당자 연락처",
                   "납품위치", "영업담당자", "비고",
                   # v5H226z490 (대표 지시): 거래명세서·세금계산서(1/2/3차) 발행일·금액 — 발주(구분번호) 단위
                   "거래명세서 발행일", "1세금계산서 발행일", "1세금계산서 금액",
                   "2세금계산서 발행일", "2세금계산서 금액", "3세금계산서 발행일", "3세금계산서 금액",
                   # v5H226z719 (대표 지시): 상태(발주 단위·진행중/출하/취소/보류 드롭다운)
                   "상태"]
# 1행 안내 문구(컬럼 위치 1-indexed별)
CO_BULK_HINTS = {1: "있으면 입력(예:012T2601)", 2: "같은 번호=한 발주(여러 품목)", 5: "필수",
                 7: "ALT+크기조정", 8: "ALT+크기조정", 9: "필수(등록 고객사명)",
                 11: "YYYY-MM-DD", 13: "KRW/USD/…", 14: "내수/수출", 15: "제품/상품/기타", 16: "필수",
                 25: "YYYY-MM-DD(선택)", 26: "YYYY-MM-DD", 27: "숫자", 28: "YYYY-MM-DD",
                 29: "숫자", 30: "YYYY-MM-DD", 31: "숫자", 32: "진행중/출하/취소/보류(비우면 초기협의)"}


def build_co_bulk_template_buf():
    """KNK 표준 소모품 발주 양식(24열) 빈 양식(.xlsx) → BytesIO. (openpyxl 필요)
    1행=안내, 2행=머리글, 3행~=품목. 통화·거래구분·형태 드롭다운. 사진 칸은 이미지 붙여넣기."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    import io as _io
    wb = Workbook(); ws = wb.active; ws.title = "소모품"
    fill = PatternFill("solid", fgColor="2F6AA8"); white = Font(color="FFFFFF", bold=True, size=10)
    hintf = Font(color="9AA3AF", italic=True, size=9)
    thin = Side(style="thin", color="DDDDDD"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = [14, 8, 18, 16, 20, 16, 16, 14, 13, 13, 12, 12, 7, 9, 9, 7, 6, 11, 12, 12, 14, 12, 11, 16,
              13, 13, 12, 13, 12, 13, 12, 11]   # v5H226z490: 거래명세서·세금계산서 7열 / z719: 상태 1열(32=AF)
    # 1행 = 안내 문구
    for ci in range(1, len(CO_BULK_HEADERS) + 1):
        hc = ws.cell(1, ci, CO_BULK_HINTS.get(ci, "")); hc.font = hintf; hc.alignment = center
    # 2행 = 머리글
    for ci, h in enumerate(CO_BULK_HEADERS, 1):
        c = ws.cell(2, ci, h); c.font = white; c.fill = fill; c.alignment = center; c.border = border
        ws.column_dimensions[c.column_letter].width = widths[ci - 1]
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"
    # 드롭다운: 통화(M=13)·거래구분(N=14)·형태(O=15)·상태(AF=32·v5H226z719)
    dv_ccy = DataValidation(type="list", formula1='"KRW,USD,VND,JPY,CNY,EUR"', allow_blank=True)
    dv_exp = DataValidation(type="list", formula1='"내수,수출"', allow_blank=True)
    dv_form = DataValidation(type="list", formula1='"제품,상품,기타"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"진행중,출하,취소,보류"', allow_blank=True)
    for dv in (dv_ccy, dv_exp, dv_form, dv_status):
        ws.add_data_validation(dv)
    dv_ccy.add("M3:M2000"); dv_exp.add("N3:N2000"); dv_form.add("O3:O2000"); dv_status.add("AF3:AF2000")
    # v5H226z588: 예시 3줄(3~5행) — '묶음 작성 방식'을 눈에 보이게 2발주 예(구분 1=2품목 / 구분 2=다른 고객사 1품목).
    #   업로드 시 품명 '예)' 접두는 자동 스킵되므로 안전.
    ex_rows = [
        # 발주 A — 구분 1, 첫 줄(발주 정보는 여기서 읽음)
        ["", "1", "WATCH9 검사기", "WATCH9 소모품", "예) Grip Pad", "KNK-P-001", "", "",
         "삼성전자", "", "2026-01-15", "2026-01-30", "KRW", "내수", "상품",
         "6", "EA", "380000", "2280000", "홍길동", "010-0000-0000", "수원 본사", "김영업", "예시 — 같은 구분번호=한 발주",
         "2026-02-05", "2026-02-10", "2280000", "", "", "", "", "진행중"],   # 거래명세서·1세금계산서·상태(예시)
        # 발주 A — 구분 1, 둘째 품목(같은 발주: 발주 정보는 비워도 첫 줄 값을 따름)
        ["", "1", "", "", "예) O-Ring", "KNK-P-002", "", "",
         "", "", "", "", "", "", "",
         "20", "EA", "1500", "30000", "", "", "", "", "예시 — 발주 정보는 첫 줄만 적으면 됨",
         "", "", "", "", "", "", "", ""],
        # 발주 B — 구분 2(번호가 다르면 다른 발주: 고객사도 다름)
        ["", "2", "AOI 장비", "AOI 소모품", "예) Nozzle", "KNK-P-010", "", "",
         "엘지전자", "", "2026-01-18", "2026-02-02", "KRW", "내수", "상품",
         "4", "EA", "120000", "480000", "이담당", "010-1111-2222", "파주 공장", "박영업", "예시 — 구분번호 다르면 다른 발주",
         "", "", "", "", "", "", "", "진행중"],
    ]
    for ex in ex_rows:
        ws.append(ex)
    for rr in (3, 4, 5):
        for ci in range(1, len(CO_BULK_HEADERS) + 1):
            ws.cell(rr, ci).font = Font(color="9AA3AF", italic=True)
    # 안내 시트
    ws2 = wb.create_sheet("작성안내")
    guide = [
        ["소모품 통합 일괄등록 — 작성 안내", ""],
        ["", ""],
        # v5H226z588: '묶음 작성 방식'을 맨 위에 강조 — 예전 단일 양식과 다른 핵심
        ["★ 묶음 작성 방식 (꼭 읽으세요)", "이 통합 양식은 엑셀 한 장에 '여러 발주'를 함께 적습니다(예전 단일 양식처럼 발주 하나당 파일 하나가 아닙니다)."],
        ["  ① 한 줄 = 한 품목", "[소모품] 시트 3행부터 한 줄에 품목 하나씩 입력합니다."],
        ["  ② 같은 구분 번호 = 한 발주", "'구분 번호'가 같은 줄들이 한 발주(여러 품목)로 묶입니다. 발주 정보(고객사·발주일 등)는 그 발주의 '첫 줄'에만 적으면 됩니다."],
        ["  ③ 번호 다르면 다른 발주", "발주가 다르면 구분 번호를 다르게(1, 2, 3…) 적습니다. 한 파일에 여러 고객사·여러 발주를 한 번에 등록할 수 있습니다."],
        ["  예) 묶음 예시", "구분 1: 삼성전자 Grip Pad·O-Ring (한 발주·2품목)  /  구분 2: 엘지전자 Nozzle (다른 발주). → 3행~5행 예시 참고(업로드 시 '예)' 줄은 자동 제외)."],
        ["", ""],
        ["기본 원리", "[소모품] 시트 3행부터 한 줄에 한 품목씩 입력. '구분 번호'가 같은 줄들이 '한 발주(여러 품목)'로 자동으로 묶입니다."],
        ["구분 번호 *", "한 발주 = 한 번호. 같은 발주의 품목은 모두 같은 번호. 발주가 다르면 번호를 다르게(1,2,3…)."],
        ["1차 고객사 *", "등록 고객사명과 (주)/㈜·공백 차이 무시하고 상호로 자동 연결. 못 찾으면 텍스트로만 저장(미리보기에서 표시)."],
        ["발주일·납품일", "YYYY-MM-DD. 발주일이 발주번호(C-YYMMDD) 기준일."],
        ["품명·수량 *", "필수. 단가/금액은 숫자(콤마 없이). 금액 비우면 수량×단가로 자동."],
        ["통화·거래구분·형태", "드롭다운(통화 KRW… / 내수·수출 / 제품·상품·기타). 통화 비우면 KRW, 거래구분 비우면 내수."],
        ["상태", "드롭다운(진행중/출하/취소/보류). 발주(구분번호) 단위 — 첫 줄에 입력. 비우면 '초기협의'. 업로드 시 그 발주 상태로 저장되어 작업일정표 색에 반영됩니다."],
        ["거래명세서·세금계산서", "발주(구분번호) 단위로 한 줄에 입력 — 거래명세서 발행일 + 1/2/3세금계산서 발행일·금액(계약금/중도금/잔금). 날짜 YYYY-MM-DD·금액 숫자. 비우면 미발행. 업로드 후 작업일정표 세금계산서 칸에 표시됩니다."],
        ["세금계산서 묶음 발행(-N)", "같은 날짜에 여러 건을 한 장으로 묶어 발행했으면 발행일 뒤에 '-번호'를 붙이세요(예: 2026-03-23-1). 같은 '날짜-번호'를 가진 줄(2건 이상)이 자동으로 한 장의 묶음 세금계산서로 발행됩니다. -번호 없으면 개별."],
        ["연결 관리번호", "이 품목이 어느 장비(관리번호)의 소모품인지(선택). 예: 012T2601. 못 찾으면 미연결 등록 + 안내."],
        ["사진 / 사진위치", "'사진(PICTURE)'·'사진위치(PICTURE LOCATION)' 칸에 이미지를 붙여넣으면 품목별로 함께 등록됩니다."],
        ["주의", "3행 예시는 지우고 작성. 업로드 → 미리보기(발주 N건·품목 M개·사진 K개) → 확정."],
    ]
    for r in guide:
        ws2.append(r)
    ws2.column_dimensions["A"].width = 18; ws2.column_dimensions["B"].width = 82
    for ci in (1, 2):
        cc = ws2.cell(1, ci); cc.font = white; cc.fill = fill
    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def _co_bulk_detect_header(ws, max_scan: int = 16) -> int:
    """머리글 행 자동 감지 — 알려진 머리글 키워드가 가장 많은 행(>=4). 못 찾으면 0."""
    KEYS = ["품명", "구분", "수량", "고객사", "발주일", "모델", "장비", "통화", "단가", "비고"]
    best_row, best = 0, 0
    for r in range(1, min(max_scan, ws.max_row) + 1):
        score = 0
        for c in range(1, min(ws.max_column, 40) + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            t = re.sub(r"\s+", "", str(v))
            if any(k in t for k in KEYS):
                score += 1
        if score > best:
            best, best_row = score, r
    return best_row if best >= 4 else 0


def _norm_co_form(v) -> str:
    """v5H226z563 (대표 지시): 엑셀 '형태' 값 → 제품/상품/기타 정규화(끝공백·표기흔들림 흡수).
    소모품은 완제품(호기) 개념 없음 → 완제품도 제품으로 흡수. 못 알아보면 '' (확정 시 기본 상품)."""
    t = ("" if v is None else str(v)).strip()
    if not t:
        return ""
    if "상품" in t:
        return "상품"
    if "제품" in t:   # 완제품·반제품·제품 모두 제품으로
        return "제품"
    if "기타" in t:
        return "기타"
    return ""


def _norm_co_status(v) -> str:
    """v5H226z719 (대표 지시): 엑셀 '상태'(진행중/출하/취소/보류) → 소모품 status enum.
    빈칸/미인식='' (확정 시 미적용 → 기본 DRAFT/초기협의 유지). '납품완료'는 출하로 흡수."""
    t = ("" if v is None else str(v)).strip()
    if not t:
        return ""
    if "취소" in t:
        return "CANCELLED"
    if "보류" in t:
        return "HOLD"
    if "출하" in t or "납품완료" in t:
        return "SHIPPED"
    if "진행" in t:
        return "CONFIRMED"
    return ""


def _co_bulk_colmap(header_cells) -> dict:
    """header_cells: [(ci, text), ...] (1-indexed). → {key: ci}. 한 컬럼은 1키에만.
    더 구체적 라벨 먼저(영업담당자→연락처→담당자 순 등 충돌 방지)."""
    cm: dict = {}
    cells = [(ci, re.sub(r"\s+", "", str(t)).upper()) for ci, t in header_cells if t not in (None, "")]

    def take(key, *needles, exclude=()):
        if key in cm:
            return
        for ci, t in cells:
            if ci in cm.values():
                continue
            if any(n.upper() in t for n in needles) and not any(e.upper() in t for e in exclude):
                cm[key] = ci
                return

    take("link_mgmt", "연결관리번호", "연결관리")
    take("group", "구분번호", "구분")
    take("equip", "장비명", "EQUIP")
    take("model_use", "모델명", "MODEL")
    take("part_name", "소모품품명", "품명", "PARTNAME", "SUPPLIERNAME")
    take("spec", "규격", "SPEC")
    take("loc_photo", "PICTURELOCATION", "사진위치", "위치")
    take("photo", "PICTURE", "사진", exclude=("LOCATION", "위치"))   # 리뷰반영: 사진위치 제외 명시
    take("customer2", "2차고객사")
    take("customer", "1차고객사", exclude=("담당자",))               # 리뷰반영: 담당자 컬럼 오매핑 방지
    take("order_date", "발주일", "ORDERDATE")
    take("due_date", "납품일", "납기", "DUEDATE")
    take("currency", "통화", "CURRENCY")
    take("is_export", "거래구분")
    take("form_type", "형태")
    take("status", "상태")   # v5H226z719 (대표 지시): 상태(진행중/출하/취소/보류)
    take("qty", "수량", "QTY", "Q'TY", "QUANTITY")
    take("unit", "단위", "UNIT")
    take("price", "단가", "UNITPRICE")
    take("amount", "금액", "AMOUNT")
    take("sales_name", "영업담당자")
    take("cc_phone", "연락처")
    take("cc_name", "담당자", exclude=("영업", "연락처"))
    take("ship_to", "납품위치", "납품처")
    take("note", "비고", "REMARK")
    # v5H226z490 (대표 지시): 거래명세서·세금계산서(1/2/3차) 발행일·금액 — 발주 단위. (amount 보다 뒤라 '금액' 단독은 위에서 선점됨)
    take("statement_date", "거래명세서")
    take("ti_date1", "1세금계산서발행")
    take("ti_amt1", "1세금계산서금액")
    take("ti_date2", "2세금계산서발행")
    take("ti_amt2", "2세금계산서금액")
    take("ti_date3", "3세금계산서발행")
    take("ti_amt3", "3세금계산서금액")
    if "customer" not in cm:
        take("customer", "고객사", exclude=("담당자", "2차"))
    return cm


def parse_co_bulk_xlsx(file_path: str, image_out_dir: str | None = None) -> dict:
    """KNK 표준 소모품 발주 양식(24열) 파서 — '구분 번호'가 같은 줄 = 한 발주.
    1행=안내, 2행(자동 감지)=머리글, 3행~=품목. 품목 사진/사진위치 임베드 이미지도 추출(image_out_dir).
    반환: {ok, orders:[{group_label, customer_name, cust_ok, customer2, order_date, due_date,
            currency, is_export, cc_name, cc_phone, ship_to, sales_name, biz_div, note,
            items:[{row, line_no, part_name, spec, model_use, equip, qty, unit, unit_price,
                    amount, link_mgmt, note, _imgs:[{full,thumb,category}], _errors}],
            item_count, total_amount, image_count, _errors, _warn}],
           total_orders, total_items, total_images, header_row}."""
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True)   # read_only=False → 이미지(_images) 접근 가능
    ws = None
    for nm in wb.sheetnames:
        if nm != "작성안내":
            ws = wb[nm]; break
    if ws is None:
        ws = wb.worksheets[0]
    hdr = _co_bulk_detect_header(ws)
    if not hdr:
        return {"ok": False, "error": "머리글(구분 번호·소모품 품명·수량…)을 찾지 못했습니다",
                "orders": [], "total_orders": 0, "total_items": 0, "total_images": 0}
    maxc = min(ws.max_column, 40)   # v5H226z490: 컬럼 31개(거래명세서·세금계산서 추가)까지 커버
    cm = _co_bulk_colmap([(c, ws.cell(hdr, c).value) for c in range(1, maxc + 1)])
    if "part_name" not in cm:
        return {"ok": False, "error": "'소모품 품명' 열을 찾지 못했습니다",
                "orders": [], "total_orders": 0, "total_items": 0, "total_images": 0}

    def gv(r, key):
        ci = cm.get(key)
        return ws.cell(r, ci).value if ci else None

    def s(v):
        return "" if v is None else str(v).strip()

    def num(v):
        try:
            return float(str("" if v is None else v).replace(",", "").strip() or 0)
        except Exception:
            return 0.0

    groups: dict = {}; seq = []; all_lines = []; line_seq = 0
    for r in range(hdr + 1, ws.max_row + 1):
        pn = s(gv(r, "part_name"))
        if not pn or pn.startswith("예)"):
            continue   # 빈 행·예시행 스킵
        line_seq += 1
        grp = s(gv(r, "group")) or f"_auto_{s(gv(r, 'customer'))}_{_date_str(gv(r, 'order_date'))}"
        qty = num(gv(r, "qty")); price = num(gv(r, "price")); amt = num(gv(r, "amount"))
        if not amt and qty > 0 and price > 0:   # 리뷰반영: 음수/0 가드
            amt = round(qty * price, 2)
        it = {
            "row": r, "line_no": line_seq, "part_name": pn,
            "spec": s(gv(r, "spec")), "model_use": s(gv(r, "model_use")),
            "equip": s(gv(r, "equip")), "qty": qty, "unit": s(gv(r, "unit")) or "EA",
            "unit_price": price, "amount": amt, "link_mgmt": s(gv(r, "link_mgmt")).upper(),
            "note": s(gv(r, "note")), "_imgs": [], "_errors": [],
        }
        if qty <= 0:
            it["_errors"].append("수량 누락/0")
        if grp not in groups:
            cust = s(gv(r, "customer"))
            _exp_raw = s(gv(r, "is_export")).upper()
            o = {
                "group_label": s(gv(r, "group")), "customer_name": cust,
                "customer2": s(gv(r, "customer2")),
                "order_date": _date_str(gv(r, "order_date")), "due_date": _date_str(gv(r, "due_date")),
                "currency": _norm_ccy(gv(r, "currency")) or "KRW",
                "is_export": 1 if ("수출" in _exp_raw or "EXPORT" in _exp_raw) else 0,
                "cc_name": s(gv(r, "cc_name")), "cc_phone": s(gv(r, "cc_phone")),
                "ship_to": s(gv(r, "ship_to")), "sales_name": s(gv(r, "sales_name")),
                "cust_ok": bool(match_customer_by_name(cust)) if cust else False,
                "biz_div": "", "note": "", "items": [], "_errors": [], "_warn": "",
                # v5H226z563 (대표 지시): 엑셀 '형태'(제품/상품/기타) 반영 — 비우면 기본 상품(확정 시 적용)
                "form_type": _norm_co_form(gv(r, "form_type")),
                # v5H226z719 (대표 지시): 상태(진행중→CONFIRMED·출하→SHIPPED·취소→CANCELLED·보류→HOLD) — 발주 단위
                "status": _norm_co_status(gv(r, "status")),
                # v5H226z490 (대표 지시): 거래명세서·세금계산서(1/2/3차) — 발주 단위(첫 줄에서 읽음)
                "statement_date": _date_str(gv(r, "statement_date")),
                "tax_invoice_amt1": num(gv(r, "ti_amt1")),
                "tax_invoice_amt2": num(gv(r, "ti_amt2")),
                "tax_invoice_amt3": num(gv(r, "ti_amt3")),
            }
            # v5H226z491b (대표 지시): 세금계산서 발행일 'YYYY-MM-DD-N' → 깨끗한 날짜 + 묶음번호(같은 날짜-N끼리 묶음 발행)
            o["tax_invoice_date"], o["ti1_bundle"] = _split_date_bundle(gv(r, "ti_date1"))
            o["tax_invoice_date2"], o["ti2_bundle"] = _split_date_bundle(gv(r, "ti_date2"))
            o["tax_invoice_date3"], o["ti3_bundle"] = _split_date_bundle(gv(r, "ti_date3"))
            if not cust:
                o["_errors"].append("1차 고객사 누락")
            elif not o["cust_ok"]:
                o["_warn"] = f"미등록 고객사 '{cust}' (텍스트로만 저장 · 미연결)"
            if not o["order_date"]:
                o["_errors"].append("발주일 누락")
            groups[grp] = o; seq.append(grp)
        else:
            # 그룹 발주 정보 보강(첫 행이 비었으면 이후 행 값으로 채움)
            o = groups[grp]
            for k_meta, k_src, is_date in (("customer_name", "customer", False), ("customer2", "customer2", False),
                                           ("order_date", "order_date", True), ("due_date", "due_date", True),
                                           ("cc_name", "cc_name", False), ("cc_phone", "cc_phone", False),
                                           ("ship_to", "ship_to", False), ("sales_name", "sales_name", False)):
                if not o.get(k_meta):
                    v = _date_str(gv(r, k_src)) if is_date else s(gv(r, k_src))
                    if v:
                        o[k_meta] = v
            # v5H226z563: 형태(제품/상품/기타) — 첫 줄이 비었으면 이후 줄 값으로 보강
            if not o.get("form_type"):
                _ff = _norm_co_form(gv(r, "form_type"))
                if _ff:
                    o["form_type"] = _ff
            # v5H226z719: 상태도 첫 줄 비었으면 이후 줄에서 보강
            if not o.get("status"):
                _ss = _norm_co_status(gv(r, "status"))
                if _ss:
                    o["status"] = _ss
            # v5H226z505 (대표 지시): 같은 발주(구분번호) 여러 줄 — 세금계산서 금액은 '합산'(줄마다 적은 부분금액
            #   누락 방지). 016처럼 한 관리번호에 345,000+370,000 두 줄이면 1세금계산서=715,000. (첫 줄에만 적던
            #   기존 방식은 이후 줄이 0이라 합산해도 그대로 → 회귀 없음.)
            for _ak, _src in (("tax_invoice_amt1", "ti_amt1"), ("tax_invoice_amt2", "ti_amt2"), ("tax_invoice_amt3", "ti_amt3")):
                _add = num(gv(r, _src))
                if _add:
                    o[_ak] = round((o.get(_ak) or 0) + _add, 2)
            # 발행일·묶음번호(-N): 첫 줄이 비었으면 이후 줄에서 채움
            for _dk, _bk, _src in (("tax_invoice_date", "ti1_bundle", "ti_date1"),
                                   ("tax_invoice_date2", "ti2_bundle", "ti_date2"),
                                   ("tax_invoice_date3", "ti3_bundle", "ti_date3")):
                if not o.get(_dk):
                    _dd2, _bb2 = _split_date_bundle(gv(r, _src))
                    if _dd2:
                        o[_dk] = _dd2; o[_bk] = _bb2
        groups[grp]["items"].append(it)
        all_lines.append(it)

    # 이미지 추출 → 데이터 행 기하 매칭(기존 단일 업로드 z284~z286 로직 재사용)
    total_images = 0
    if image_out_dir and all_lines:
        try:
            os.makedirs(image_out_dir, exist_ok=True)
            _ymax = ws.max_row
            for _im in (getattr(ws, "_images", []) or []):
                _t = getattr(_im.anchor, "to", None) or getattr(_im.anchor, "_to", None)
                if _t is not None and getattr(_t, "row", None) is not None and (_t.row + 3) > _ymax:
                    _ymax = _t.row + 3
            _P = _row_y_prefix(ws, _ymax)
            _photo_col, _loc_col = _find_photo_cols(ws, hdr)
            for idx, img in enumerate(getattr(ws, "_images", []) or []):
                try:
                    a = img.anchor; raw = img._data()
                    if not raw:
                        continue
                    _cy = _image_center_y(a, _P)
                    ml = (_line_for_center(all_lines, _cy, _P) if _cy is not None
                          else _find_nearest_line(all_lines, getattr(getattr(a, "_from", None), "row", 0) + 1))
                    if ml is None:
                        continue
                    _frm = getattr(a, "_from", None)
                    _col = (_frm.col + 1) if (_frm is not None and getattr(_frm, "col", None) is not None) else None
                    _cat = "photo"
                    if _col is not None and _loc_col is not None:
                        if _photo_col is not None:
                            _cat = "loc" if abs(_col - _loc_col) < abs(_col - _photo_col) else "photo"
                        elif _col == _loc_col:
                            _cat = "loc"
                    fn = f"l{ml['line_no']:03d}_{idx + 1}.jpg"
                    fnt = f"l{ml['line_no']:03d}_{idx + 1}_t.jpg"
                    big, thumb, _info = compress_image_bytes(raw)
                    with open(os.path.join(image_out_dir, fn), "wb") as f:
                        f.write(big)
                    with open(os.path.join(image_out_dir, fnt), "wb") as f:
                        f.write(thumb)
                    ml["_imgs"].append({"full": fn, "thumb": fnt, "category": _cat})
                    total_images += 1
                except Exception:
                    continue
            for ln in all_lines:
                ln["_imgs"].sort(key=lambda x: 0 if x.get("category") == "photo" else 1)
        except Exception:
            pass

    orders = []
    for grp in seq:
        o = groups[grp]
        o["item_count"] = len(o["items"])
        o["total_amount"] = round(sum(it["qty"] * it["unit_price"] for it in o["items"]), 2)
        o["image_count"] = sum(len(it["_imgs"]) for it in o["items"])
        if o["items"] and all(it["_errors"] for it in o["items"]):
            o["_errors"].append("유효 품목 없음(모든 품목 행에 오류)")
        orders.append(o)
    return {"ok": True, "orders": orders, "total_orders": len(orders),
            "total_items": sum(o["item_count"] for o in orders),
            "total_images": total_images, "header_row": hdr}


def co_get(co_id: int) -> dict | None:
    with db_session() as c:
        r = c.execute(
            "SELECT * FROM consumable_orders WHERE id=?", (int(co_id),)
        ).fetchone()
        return dict(r) if r else None


def co_list(status: str = "", q: str = "", limit: int = 200) -> list[dict]:
    sql = "SELECT * FROM consumable_orders WHERE 1=1"
    params: list = []
    if status and status in CO_STATUSES:
        sql += " AND status=?"; params.append(status)
    if q:
        sql += " AND (co_no LIKE ? OR customer_name LIKE ? OR note LIKE ?)"
        params += [f"%{q}%", f"%{q}%", f"%{q}%"]
    sql += " ORDER BY id DESC LIMIT ?"; params.append(int(limit))
    with db_session() as c:
        return [dict(r) for r in c.execute(sql, params).fetchall()]


def coi_list(co_id: int) -> list[dict]:
    with db_session() as c:
        rows = c.execute(
            """SELECT i.*,
                      p.mgmt_code AS linked_mgmt_code,
                      p.name      AS linked_project_name,
                      pa.part_no  AS part_no
                 FROM consumable_order_items i
            LEFT JOIN projects p  ON p.id  = i.linked_project_id
            LEFT JOIN parts    pa ON pa.id = i.part_id
                WHERE i.co_id=?
             ORDER BY i.line_no, i.id""",
            (int(co_id),)
        ).fetchall()
        return [dict(r) for r in rows]


# ── v5H226z298: 변경 이력(히스토리 탭) ──
_ORDER_FIELD_LABELS = {
    "customer_name": "고객사", "secondary_customer": "2차 고객사", "ship_to": "납품위치",
    "order_date": "발주일", "due_date": "납기", "cc_name": "고객 담당자", "cc_phone": "연락처",
    "currency": "통화", "is_export": "거래구분", "model_name": "모델명", "equip_name": "장비명",
    "sales_name": "영업담당자", "note": "비고",
    # v5H226z725 (대표 지시): 상세 화면을 엑셀 양식과 일치 — 상태·형태·거래명세서·세금계산서(발주 단위)도
    #   save-all(헤더 경로)로 편집 저장. 라인 표에 칸으로 보이지만 값은 발주(consumable_orders)에 저장.
    "status": "상태", "form_type": "형태", "statement_date": "거래명세서 발행일",
    "tax_invoice_date": "1세금계산서 발행일", "tax_invoice_amt1": "1세금계산서 금액",
    "tax_invoice_date2": "2세금계산서 발행일", "tax_invoice_amt2": "2세금계산서 금액",
    "tax_invoice_date3": "3세금계산서 발행일", "tax_invoice_amt3": "3세금계산서 금액",
}
_LINE_FIELD_LABELS = {"model_use": "모델", "equip_name": "장비명", "part_name": "품명",
                      "spec": "규격", "qty": "수량", "unit": "단위", "unit_price": "단가",
                      "note": "비고", "linked_project_id": "관리번호 연결", "part_id": "자재 연결"}


def co_log_change(co_id, item_id, scope, field, label, old_value, new_value, by_id=None, by_name=""):
    """소모품수주 변경 1건 기록(이력 탭). 실패해도 본 작업은 진행."""
    try:
        with db_session() as c:
            c.execute(
                "INSERT INTO consumable_history(co_id, item_id, scope, field, label, "
                "old_value, new_value, changed_by, changed_by_name) VALUES(?,?,?,?,?,?,?,?,?)",
                (int(co_id), (int(item_id) if item_id else None), scope, field, label,
                 ("" if old_value is None else str(old_value)),
                 ("" if new_value is None else str(new_value)), by_id, by_name or "")
            )
    except Exception:
        pass


def co_history_list(co_id, limit: int = 300) -> list[dict]:
    with db_session() as c:
        return [dict(r) for r in c.execute(
            "SELECT * FROM consumable_history WHERE co_id=? ORDER BY id DESC LIMIT ?",
            (int(co_id), int(limit))
        ).fetchall()]


def coi_get(item_id) -> dict | None:
    """라인 1건 조회(관리번호 검증 등)."""
    with db_session() as c:
        r = c.execute("SELECT * FROM consumable_order_items WHERE id=?", (int(item_id),)).fetchone()
        return dict(r) if r else None


def mgmt_line_matches(line: dict, proj: dict) -> bool:
    """v5H226z301 (대표 지시): 입력한 관리번호(프로젝트)의 모델/장비가 라인과 '맞는지'.
    의미있는 토큰(2자+) 교집합이 있으면 맞음으로 본다(없으면 알람 후 강제연결 가능)."""
    def _tok(s):
        return {t for t in re.findall(r"[A-Za-z0-9가-힣]+", (s or "").upper()) if len(t) >= 2}
    line_tok = _tok(line.get("model_use")) | _tok(line.get("equip_name"))
    proj_tok = _tok(proj.get("model_name")) | _tok(proj.get("name"))
    return bool(line_tok & proj_tok)


def _sim(a, b) -> float:
    """정규화(대문자·공백/기호 제거) 후 문자열 유사도 0~1 (difflib)."""
    from difflib import SequenceMatcher
    na = re.sub(r"[\s\(\)\[\]\.,\-_/]", "", (a or "").upper())
    nb = re.sub(r"[\s\(\)\[\]\.,\-_/]", "", (b or "").upper())
    if not na or not nb:
        return 0.0
    return SequenceMatcher(None, na, nb).ratio()


def suggest_mgmt_for_line(line_model, line_equip, customer_id, threshold: float = 0.9):
    """v5H226z302 (대표 지시): 예상(추천) 관리번호 — 세 조건 모두 충족할 때만.
    ① 같은 1차 고객사  ② 모델명 90%↑ 일치  ③ 장비명 90%↑ 일치.
    여러 후보면 (모델·장비 평균 유사도) 최고 1건 반환 {.., score} 또는 None."""
    if not customer_id:
        return None
    if not str(line_model or "").strip() or not str(line_equip or "").strip():
        return None  # 모델·장비 둘 다 있어야 비교 가능
    with db_session() as c:
        rows = [dict(r) for r in c.execute(
            "SELECT id, mgmt_code, name, model_name, equip_name FROM projects "
            "WHERE customer_id=? AND COALESCE(project_type,'NEW_EQUIP')='NEW_EQUIP'",
            (int(customer_id),)
        ).fetchall()]
    best, best_score = None, 0.0
    for r in rows:
        sm = _sim(line_model, r.get("model_name"))
        se = _sim(line_equip, r.get("equip_name"))
        if sm >= threshold and se >= threshold:
            score = (sm + se) / 2.0
            if score > best_score:
                best, best_score = r, score
    if best:
        best["score"] = round(best_score, 3)
        best["match_level"] = "model_equip_90"
    return best


def co_update_order_field(co_id: int, field: str, value, by_id=None, by_name: str = "") -> tuple:
    """소모품수주(헤더) 정보란 1필드 수정 + 이력. 반환 (성공, 새값, 옛값, 고객사연결여부).
    고객사 변경 시 (주)/주식회사 무시 상호 매칭으로 정식명칭·customer_id 갱신."""
    # v5H226z299 (대표 지시): 1차 고객사는 '등록된 고객사'만 — id 선택으로 customer_id + 정식명칭 동시 설정
    if field == "customer_id":
        new_id = int(value) if str(value).strip().lstrip("-").isdigit() and int(value) > 0 else None
        with db_session() as c:
            row = c.execute("SELECT customer_name FROM consumable_orders WHERE id=?", (int(co_id),)).fetchone()
            old_name = (row[0] if row else "") or ""
            cust = c.execute("SELECT name FROM customers WHERE id=?", (new_id,)).fetchone() if new_id else None
            new_name = (cust[0] if cust else old_name)
            c.execute("UPDATE consumable_orders SET customer_id=?, customer_name=? WHERE id=?",
                      (new_id, new_name, int(co_id)))
        if str(old_name) != str(new_name):
            co_log_change(co_id, None, "order", "customer_name", "고객사", old_name, new_name, by_id, by_name)
        return (True, new_name, old_name, bool(new_id))
    if field not in _ORDER_FIELD_LABELS:
        return (False, None, None, None)
    if field == "is_export":
        new_v = 1 if str(value).strip() in ("수출", "1", "export", "EXPORT") else 0
    elif field == "currency":
        new_v = (str(value).strip().upper() or "KRW")
    else:
        new_v = (str(value).strip() if value is not None else "")
    cust_id = None
    cust_changed = (field == "customer_name")
    cust_ok = None
    if cust_changed:
        m = match_customer_by_name(new_v)
        if m:
            cust_id = m["id"]; new_v = m["name"]; cust_ok = True
        else:
            cust_ok = False
    with db_session() as c:
        cols = {r[1] for r in c.execute("PRAGMA table_info(consumable_orders)").fetchall()}
        if field not in cols:
            return (False, None, None, None)
        row = c.execute(f"SELECT {field} FROM consumable_orders WHERE id=?", (int(co_id),)).fetchone()
        old_v = row[0] if row else None
        c.execute(f"UPDATE consumable_orders SET {field}=? WHERE id=?", (new_v, int(co_id)))
        if cust_changed:
            c.execute("UPDATE consumable_orders SET customer_id=? WHERE id=?", (cust_id, int(co_id)))
    if str(old_v if old_v is not None else "") != str(new_v if new_v is not None else ""):
        co_log_change(co_id, None, "order", field, _ORDER_FIELD_LABELS[field], old_v, new_v, by_id, by_name)
    return (True, new_v, old_v, cust_ok)


def coi_update(item_id: int, fields: dict, by_id=None, by_name: str = "") -> None:
    """라인 인라인 편집 — fields 의 키만 갱신 + 변경 이력 기록."""
    allowed = {"model_use", "equip_name", "part_name", "spec", "qty", "unit",
               "unit_price", "linked_project_id", "part_id", "note"}
    keys = [k for k in fields.keys() if k in allowed]
    if not keys:
        return
    co_id = None
    old = {}
    with db_session() as c:
        cur = c.execute(
            f"SELECT co_id, {', '.join(keys)} FROM consumable_order_items WHERE id=?",
            (int(item_id),)
        ).fetchone()
        if not cur:
            return
        co_id = cur["co_id"]
        old = {k: cur[k] for k in keys}
        sets = ", ".join(f"{k}=?" for k in keys)
        vals = [fields[k] for k in keys] + [int(item_id)]
        c.execute(f"UPDATE consumable_order_items SET {sets} WHERE id=?", vals)
        c.execute(
            "UPDATE consumable_order_items SET amount=ROUND(COALESCE(qty,0)*COALESCE(unit_price,0),2) WHERE id=?",
            (int(item_id),)
        )
    if co_id is not None:
        recompute_co_total(co_id)
        for k in keys:
            if str(old.get(k) if old.get(k) is not None else "") != str(fields[k] if fields[k] is not None else ""):
                co_log_change(co_id, item_id, "line", k, _LINE_FIELD_LABELS.get(k, k),
                              old.get(k), fields[k], by_id, by_name)


def coi_add_blank(co_id: int, by_id=None, by_name: str = "") -> int:
    """v5H226z309 (대표 지시): 상세 화면에서 라인 1줄 직접 추가(빈 줄) → 바로 인라인 입력.
    line_no = 기존 최대+1. 존재하는 컬럼만 동적 INSERT(추가형 마이그 안전)."""
    with db_session() as c:
        cols_avail = {r2[1] for r2 in c.execute("PRAGMA table_info(consumable_order_items)").fetchall()}
        row = c.execute("SELECT COALESCE(MAX(line_no),0) FROM consumable_order_items WHERE co_id=?",
                        (int(co_id),)).fetchone()
        next_no = int(row[0] or 0) + 1
        data = {"co_id": int(co_id), "line_no": next_no, "model_use": "", "equip_name": "",
                "part_name": "", "spec": "", "qty": 0, "unit": "EA", "unit_price": 0,
                "amount": 0, "note": ""}
        use = [(k, v) for k, v in data.items() if k in cols_avail]
        cols = ", ".join(k for k, _ in use)
        ph = ", ".join("?" for _ in use)
        cur = c.execute(f"INSERT INTO consumable_order_items ({cols}) VALUES ({ph})",
                        [v for _, v in use])
        iid = int(cur.lastrowid)
    recompute_co_total(co_id)
    co_log_change(co_id, iid, "line", "add", "라인 추가", "", f"라인 {next_no}", by_id, by_name)
    return iid


def coi_set_image(item_id: int, category: str, image_path: str,
                  image_thumb_path: str, by_id=None, by_name: str = "") -> int | None:
    """v5H226z309 (대표 지시): 라인 1줄에 사진 첨부(직접 업로드). category: 'photo'=사진(PICTURE) /
    'loc'=사진위치(PICTURE LOCATION). 존재하는 컬럼만 갱신. co_id 반환(없으면 None)."""
    cat = "loc" if category == "loc" else "photo"
    full_col, thumb_col = (("image_loc_path", "image_loc_thumb_path") if cat == "loc"
                           else ("image_path", "image_thumb_path"))
    co_id = None
    with db_session() as c:
        cols_avail = {r2[1] for r2 in c.execute("PRAGMA table_info(consumable_order_items)").fetchall()}
        r = c.execute("SELECT co_id FROM consumable_order_items WHERE id=?", (int(item_id),)).fetchone()
        if not r:
            return None
        co_id = r[0]
        sets, vals = [], []
        if full_col in cols_avail:
            sets.append(f"{full_col}=?"); vals.append(image_path)
        if thumb_col in cols_avail:
            sets.append(f"{thumb_col}=?"); vals.append(image_thumb_path)
        if not sets:
            return None   # v5H226z310 (적대검토): image 컬럼 미존재 등으로 미반영이면 성공으로 보고 안 함
        vals.append(int(item_id))
        c.execute(f"UPDATE consumable_order_items SET {', '.join(sets)} WHERE id=?", vals)
    co_log_change(co_id, item_id, "line", f"image_{cat}",
                  ("사진(PICTURE)" if cat == "photo" else "사진위치(LOCATION)"),
                  "", "첨부됨", by_id, by_name)
    return co_id


def coi_delete(item_id: int) -> None:
    with db_session() as c:
        row = c.execute(
            "SELECT co_id, image_path, image_thumb_path FROM consumable_order_items WHERE id=?",
            (int(item_id),)
        ).fetchone()
        if not row:
            return
        c.execute("DELETE FROM consumable_order_items WHERE id=?", (int(item_id),))
    recompute_co_total(row[0])


def co_delete(co_id: int) -> None:
    """헤더 + 라인 + 이미지 디렉토리 일괄 삭제."""
    with db_session() as c:
        c.execute("DELETE FROM consumable_orders WHERE id=?", (int(co_id),))
    # 이미지 폴더 삭제
    img_dir = co_image_dir(co_id)
    if os.path.isdir(img_dir):
        try:
            shutil.rmtree(img_dir, ignore_errors=True)
        except Exception:
            pass


# ────────────────────────────────────────────────────────────────────
# 경로 헬퍼
# ────────────────────────────────────────────────────────────────────
def _uploads_root() -> str:
    """01_HAIST_WORKS/uploads/consumables/"""
    here = os.path.dirname(os.path.abspath(__file__))
    root = os.path.normpath(os.path.join(here, "..", "uploads", "consumables"))
    os.makedirs(root, exist_ok=True)
    return root


def co_image_dir(co_id: int) -> str:
    p = os.path.join(_uploads_root(), str(int(co_id)))
    os.makedirs(p, exist_ok=True)
    return p


def co_image_url(co_id: int, fname: str) -> str:
    """브라우저에서 접근할 URL — /uploads/consumables/{co_id}/{fname}"""
    return f"/uploads/consumables/{int(co_id)}/{fname}"


# ────────────────────────────────────────────────────────────────────
# 프로젝트별 소모품 이력 (project_detail 통합)
# ────────────────────────────────────────────────────────────────────
def get_project_consumable_orders(project_id: int, limit: int = 200) -> dict:
    """프로젝트(장비) 1건에 연결된 consumable_order_items + 합계."""
    with db_session() as c:
        rows = c.execute(
            """SELECT i.id, i.line_no, i.model_use, i.part_name, i.spec,
                      i.qty, i.unit, i.unit_price, i.amount,
                      i.image_thumb_path, i.image_path,
                      co.id   AS co_id,
                      co.co_no, co.order_date, co.customer_name,
                      co.status AS co_status, co.currency
                 FROM consumable_order_items i
                 JOIN consumable_orders co ON co.id = i.co_id
                WHERE i.linked_project_id=?
             ORDER BY co.order_date DESC, co.id DESC, i.line_no
                LIMIT ?""",
            (int(project_id), int(limit))
        ).fetchall()
    out = [dict(r) for r in rows]
    total_amt = sum(float(r["amount"] or 0) for r in out)
    total_qty = sum(float(r["qty"] or 0) for r in out)
    return {"rows": out, "total_amount": round(total_amt, 2),
            "total_qty": round(total_qty, 4), "count": len(out)}
