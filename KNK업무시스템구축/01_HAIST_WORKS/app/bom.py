# -*- coding: utf-8 -*-
# ============================================================
# v5H226z983 (2026-07-15) — 1단계 BOM 보드 (3개월 스프린트 · 대표 지시 "1단계로 진행")
# ------------------------------------------------------------
# 설계·전장·SW팀이 쓰던 BOM 엑셀을 "양식 그대로" 업로드 →
#   ① 변경 어휘 3종 자동 인식: 셀 안 '기존->변경'(빨간 리치텍스트라 색 판정 금지·텍스트로),
#      노랑 배경=신규 추가, 취소선=삭제  (근거: 작업기록/BOM양식_분석_설계팀3종_2026-07-14.md)
#   ② 직전 보드와 자동 비교(추가/변경/삭제) — 색칠·비고 수기 비교 소멸
#   ③ 발주된 라인이 변경/삭제되면 🔴 재검토 플래그
# 원칙:
#   - 헤더 인식 실패 = 중단 + 안내 (위치 추측 절대 금지 — z978 규칙 [[knk_bulk_import_header_strict]])
#   - 부분 목록 업로드(merge 기본)는 기존 라인을 삭제하지 않음 (실측: '1대 구매품 LIST'=부분 목록)
#   - 실삭제 금지(status='삭제') · 모든 변경은 bom_item_history 기록
#   - 자재 마스터 연결은 명확한 단일 후보만 (0단계 정규화 키·별칭 공유 [[feedback_data_connectivity]])
# ============================================================
import os
import re
import json
from datetime import date as _date

from .database import db_session, _normalize_part_key, _logi_now

# ── 헤더 별칭 (3개국어 · 세대 혼재 대응 · 대소문자/공백/줄바꿈 무시 후 '포함' 매칭) ──
# 필드 우선순위 = 리스트 순서 (먼저 매칭된 필드가 열을 차지)
_HEADER_ALIASES = [
    ("mgmt_code",    ["영업관리코드", "관리코드", "관리번호"]),
    ("part_no",      ["PRODUCTCODE", "코드명", "MASANPHAM", "MÃSẢNPHẨM", "품번", "모델명", "PARTNO"]),
    ("part_name",    ["PRODUCTNAME", "제품명", "TENSANPHAM", "TÊNSẢNPHẨM", "품명", "PARTNAME"]),
    ("category",     ["CATEGORY", "구분", "DANHMUC", "DANHMỤC"]),
    ("maker",        ["MANUFACTURER", "제조사", "NHASANXUAT", "NHÀSẢNXUẤT", "MAKER"]),
    ("vendor",       ["VENDOR", "외주사", "NHATHAUPHU", "NHÀTHẦUPHỤ", "협력사"]),
    ("material",     ["MATERIAL", "재질", "VATLIEU", "VẬTLIỆU"]),
    ("finishing",    ["FINISHING", "후처리", "XULYBEMAT", "XỬLÝBỀMẶT", "표면처리"]),
    ("total_qty",    ["UNITTOTAL", "TOTAL", "총수량"]),
    ("unit_count",   ["UNITCOUNT", "수량", "SOLUONG", "SỐLƯỢNG", "QTY"]),
    ("unit_price",   ["UNITPRICE", "단가", "DONGIA", "ĐƠNGIÁ"]),
    ("amount",       ["AMOINT", "AMOUNT", "합계", "TONGCONG", "TỔNGCỘNG", "금액"]),
    ("unit",         ["UNIT", "단위", "DONVI", "ĐƠNVỊ"]),
    ("delivery_text",["DELIVERY", "납기", "THOIGIANGIAOHANG"]),
    ("buy_at",       ["구매처", "KOR/VINA", "KORVINA"]),
    ("unit_code",    ["CODE"]),
    ("line_no",      ["NO."]),
    ("remarks",      ["REMARKS", "비고", "NOTE", "GHICHU", "GHICHÚ"]),
]
# 이 3개가 모두 잡혀야 BOM 시트로 인정 (위치 추측 폴백 없음)
_REQUIRED_ANY = ("part_name", "unit_count")  # + (part_no 또는 material) — 가공품 시트는 모델명이 없음

_MGMT_RE = re.compile(r"\b(\d{3}[A-Z]\d{4})\b")
_ARROW_SPLIT = re.compile(r"\s*->\s*")


def _norm_header_text(v) -> str:
    """헤더 셀 → 비교 키: 줄바꿈·공백·괄호·기호 제거 + 대문자."""
    s = str(v or "")
    s = re.sub(r"[\s\n\r/()\[\]·._-]+", "", s).upper()
    return s


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _num(v, default=0.0):
    if v is None or v == "":
        return default
    try:
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return default


def _is_yellow(cell) -> bool:
    """신규 추가 표기 = 노랑 배경 (실측 FFFFFF00). 리치텍스트 글자색은 판정에 쓰지 않음."""
    try:
        fl = cell.fill
        if fl and fl.patternType == "solid" and fl.fgColor is not None:
            rgb = getattr(fl.fgColor, "rgb", None)
            if isinstance(rgb, str) and rgb.upper().endswith("FFFF00"):
                return True
    except Exception:
        pass
    return False


def _is_strike(cell) -> bool:
    try:
        return bool(cell.font and cell.font.strike)
    except Exception:
        return False


def _split_arrow(raw: str):
    """'기존->변경' 표기 분해. '→'도 인식. 반환 (old, new, is_multi) — 화살표 없으면 (None, raw, False).
    1→N 분해(예: 'A->B+C+D')는 new에 '+' 포함 → is_multi=True (라인 분할은 화면에서 수동)."""
    s = (raw or "").replace("→", "->")
    if "->" not in s:
        return None, (raw or "").strip(), False
    parts = _ARROW_SPLIT.split(s, maxsplit=1)
    old = parts[0].strip()
    new = parts[1].strip() if len(parts) > 1 else ""
    return old, new, ("+" in new)


def detect_header(ws, scan_max: int = 15):
    """r1~15에서 헤더 행 탐지. 반환 (header_row, {col_idx: field}) 또는 (None, {}).
    z978 규칙: 못 찾으면 그 시트는 파싱하지 않는다 (위치 추측 금지)."""
    max_col = min(ws.max_column or 0, 40)
    best = (None, {})
    for r in range(1, min(scan_max, ws.max_row or 0) + 1):
        col_map = {}
        used = set()
        for c in range(1, max_col + 1):
            key = _norm_header_text(ws.cell(r, c).value)
            if not key:
                continue
            for field, aliases in _HEADER_ALIASES:
                if field in used:
                    continue
                if any(a in key for a in aliases):
                    col_map[c] = field
                    used.add(field)
                    break
        score = len(col_map)
        if score > len(best[1]):
            best = (r, col_map)
    hr, cmap = best
    fields = set(cmap.values())
    if not hr:
        return None, {}
    if not all(f in fields for f in _REQUIRED_ANY):
        return None, {}
    if "part_no" not in fields and "material" not in fields:
        return None, {}
    return hr, cmap


def _sheet_item_type(sheet_name: str) -> str:
    n = sheet_name or ""
    if "가공" in n:
        return "가공품"
    if "공용" in n:
        return "공용부"
    return "구매품"


def parse_bom_file(path: str, filename: str = "") -> dict:
    """BOM 엑셀 파싱 → {mgmt_code, sheets: [{sheet, ok, reason, item_type, header_row, items, stats}]}
    items[i] = {필드..., excel_changes: [{field, old, new, is_multi}], is_new_marked, is_deleted_marked}"""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)   # 값(수식 캐시) + 스타일 함께
    out = {"mgmt_code": "", "sheets": []}

    # 프로젝트 코드: ①시트 상단 제목(r1~6) ②파일명 (열은 행 단위에서 추가 확인)
    def _scan_code_top(ws):
        for r in range(1, min(6, ws.max_row or 0) + 1):
            for c in range(1, min(ws.max_column or 0, 20) + 1):
                m = _MGMT_RE.search(str(ws.cell(r, c).value or ""))
                if m:
                    return m.group(1)
        return ""

    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.sheet_state != "visible":
            continue
        hr, cmap = detect_header(ws)
        if not hr:
            out["sheets"].append({"sheet": sn, "ok": False,
                                  "reason": "헤더(품명·수량 등)를 찾지 못함 — BOM 시트가 아니거나 양식이 다릅니다",
                                  "items": [], "stats": {}})
            continue
        if not out["mgmt_code"]:
            out["mgmt_code"] = _scan_code_top(ws)
        item_type = _sheet_item_type(sn)
        items = []
        stats = {"rows": 0, "arrows": 0, "new_marked": 0, "deleted_marked": 0, "multi_split": 0}
        empty_streak = 0
        for r in range(hr + 1, (ws.max_row or hr) + 1):
            row_vals = {}
            cells = {}
            any_val = False
            for c, field in cmap.items():
                cell = ws.cell(r, c)
                cells[field] = cell
                v = cell.value
                if v is not None and str(v).strip() != "":
                    any_val = True
                row_vals[field] = v
            if not any_val:
                empty_streak += 1
                if empty_streak >= 30:      # 데이터 종료로 판단
                    break
                continue
            empty_streak = 0
            pno_raw = _cell_str(row_vals.get("part_no"))
            name_raw = _cell_str(row_vals.get("part_name"))
            if not pno_raw and not name_raw:
                continue                     # 소계/장식 행
            it = {
                "item_type": item_type,
                "category": _cell_str(row_vals.get("category")),
                "unit_code": _cell_str(row_vals.get("unit_code")),
                "line_no": int(_num(row_vals.get("line_no"), 0)) or None,
                "maker": "", "vendor": _cell_str(row_vals.get("vendor")),
                "material": _cell_str(row_vals.get("material")),
                "finishing": _cell_str(row_vals.get("finishing")),
                "unit_count": _num(row_vals.get("unit_count")),
                "total_qty": _num(row_vals.get("total_qty")),
                "unit": _cell_str(row_vals.get("unit")) or "EA",
                "unit_price": _num(row_vals.get("unit_price")),
                "amount": _num(row_vals.get("amount")),
                "delivery_text": _cell_str(row_vals.get("delivery_text")),
                "buy_at": _cell_str(row_vals.get("buy_at")).upper(),
                "remarks": _cell_str(row_vals.get("remarks")),
                "mgmt_code": _cell_str(row_vals.get("mgmt_code")),
                "excel_changes": [],
                "is_new_marked": False,
                "is_deleted_marked": False,
            }
            # 변경 어휘 ①: 셀 안 '기존->변경' (모델명·제조사·품명)
            for f, raw in (("part_no", pno_raw), ("maker", _cell_str(row_vals.get("maker"))),
                           ("part_name", name_raw)):
                old, new, is_multi = _split_arrow(raw)
                it[f] = new
                if old is not None:
                    it["excel_changes"].append({"field": f, "old": old, "new": new, "is_multi": is_multi})
                    stats["arrows"] += 1
                    if is_multi:
                        stats["multi_split"] += 1
            # 변경 어휘 ②: 노랑 배경 = 신규 추가 (주요 셀 아무거나)
            for f in ("part_no", "part_name", "category", "unit_count"):
                if f in cells and _is_yellow(cells[f]):
                    it["is_new_marked"] = True
                    break
            if it["is_new_marked"]:
                stats["new_marked"] += 1
            # 변경 어휘 ③: 취소선 = 삭제
            for f in ("part_no", "part_name"):
                if f in cells and _is_strike(cells[f]):
                    it["is_deleted_marked"] = True
                    break
            if it["is_deleted_marked"]:
                stats["deleted_marked"] += 1
            if not out["mgmt_code"] and it["mgmt_code"]:
                m = _MGMT_RE.search(it["mgmt_code"])
                if m:
                    out["mgmt_code"] = m.group(1)
            if it["total_qty"] == 0 and it["unit_count"]:
                it["total_qty"] = it["unit_count"]
            if it["amount"] == 0 and it["unit_price"]:
                it["amount"] = round((it["total_qty"] or it["unit_count"] or 0) * it["unit_price"], 2)
            items.append(it)
            stats["rows"] += 1
        out["sheets"].append({"sheet": sn, "ok": True, "reason": "", "item_type": item_type,
                              "header_row": hr, "items": items, "stats": stats})
    wb.close()
    if not out["mgmt_code"] and filename:
        m = _MGMT_RE.search(filename)
        if m:
            out["mgmt_code"] = m.group(1)
    return out


# ── 매칭 키: (구분, 모델명) — 모델명 없으면(가공품) 품명 기준 ──
def _item_key(category: str, part_no: str, part_name: str) -> str:
    cat = _normalize_part_key(str(category or ""))
    pno = _normalize_part_key(str(part_no or ""))
    if pno:
        return f"P:{cat}|{pno}"
    return f"N:{cat}|{_normalize_part_key(str(part_name or ''))}"


_COMPARE_FIELDS = ["part_name", "maker", "vendor", "material", "finishing",
                   "unit_count", "total_qty", "unit", "unit_price", "amount",
                   "delivery_text", "unit_code", "remarks"]
_ORDERED_STATES = ("발주", "부분입고", "입고")


def _pair_items(board_items: list, file_items: list):
    """보드(활성) ↔ 파일 라인 매칭.
    같은 키가 여러 개면 등장 순서대로 짝지음(같은 부품이 여러 줄인 실측 케이스).
    화살표 표기 행은 '기존' 모델명 키로도 매칭 시도 (모델명 변경 추적의 핵심)."""
    b_by_key: dict = {}
    for b in board_items:
        b_by_key.setdefault(_item_key(b.get("category"), b.get("part_no"), b.get("part_name")), []).append(b)
    pairs = []      # (board, file_item)
    adds = []
    for f in file_items:
        keys = [_item_key(f.get("category"), f.get("part_no"), f.get("part_name"))]
        for ch in f.get("excel_changes", []):
            if ch["field"] == "part_no" and ch.get("old"):
                keys.append(_item_key(f.get("category"), ch["old"], f.get("part_name")))
            if ch["field"] == "part_name" and ch.get("old") and not f.get("part_no"):
                keys.append(_item_key(f.get("category"), "", ch["old"]))
        matched = None
        for k in keys:
            lst = b_by_key.get(k)
            if lst:
                matched = lst.pop(0)
                break
        if matched is not None:
            pairs.append((matched, f))
        else:
            adds.append(f)
    leftovers = [b for lst in b_by_key.values() for b in lst]
    return pairs, adds, leftovers


def _diff_fields(board: dict, f: dict) -> dict:
    """보드 라인 vs 파일 라인 필드 차이 {field: {old, new}} (숫자는 값 비교·문자는 trim 비교)."""
    out = {}
    if (board.get("part_no") or "") != (f.get("part_no") or ""):
        out["part_no"] = {"old": board.get("part_no"), "new": f.get("part_no")}
    for fld in _COMPARE_FIELDS:
        ov = board.get(fld)
        nv = f.get(fld)
        if fld in ("unit_count", "total_qty", "unit_price", "amount"):
            if abs(_num(ov) - _num(nv)) > 0.0001:
                out[fld] = {"old": _num(ov), "new": _num(nv)}
        else:
            if (str(ov).strip() if ov is not None else "") != (str(nv).strip() if nv is not None else ""):
                # 파일 쪽이 빈값이면 '지움'으로 보지 않고 기존 유지 (부분 양식 관대) — 단 품명은 비교
                if (nv is None or str(nv).strip() == "") and fld not in ("part_name",):
                    continue
                out[fld] = {"old": ov, "new": nv}
    return out


def plan_diff(project_id: int, file_items: list) -> dict:
    """미리보기용 비교 계획 — DB 변경 없음."""
    with db_session() as c:
        board = [dict(r) for r in c.execute(
            "SELECT * FROM bom_items WHERE project_id=? AND status='활성'", (int(project_id),)
        ).fetchall()]
    live = [f for f in file_items if not f.get("is_deleted_marked")]
    dels_marked = [f for f in file_items if f.get("is_deleted_marked")]
    pairs, adds, leftovers = _pair_items(board, live)
    changes = []
    unchanged = 0
    ordered_warn = 0
    for b, f in pairs:
        d = _diff_fields(b, f)
        if d:
            warn = (b.get("order_status") or "미발주") in _ORDERED_STATES
            if warn:
                ordered_warn += 1
            changes.append({"item_id": b["id"], "board": b, "file": f, "fields": d, "ordered": warn})
        else:
            unchanged += 1
    # 취소선 삭제: 살아있는 행과 짝지어지지 않은 나머지(leftovers)에서만 찾음
    #   → 같은 라인이 '미포함'과 '삭제'에 이중 집계되는 것 방지 + 살아있는 짝을 뺏지 않음
    del_hits = []
    remaining = leftovers
    if dels_marked:
        p2, _a2, remaining = _pair_items(leftovers, dels_marked)
        for b, f in p2:
            warn = (b.get("order_status") or "미발주") in _ORDERED_STATES
            if warn:
                ordered_warn += 1
            del_hits.append({"item_id": b["id"], "board": b, "file": f, "ordered": warn})
    return {"adds": adds, "changes": changes, "deletes": del_hits,
            "missing": remaining, "unchanged": unchanged,
            "ordered_warn": ordered_warn, "board_count": len(board)}


def _match_part_id(c, part_no: str):
    """자재 마스터 연결 — 명확한 단일 후보만 (정규화 키 → 별칭). 애매하면 None."""
    nk = _normalize_part_key(str(part_no or ""))
    if not nk:
        return None
    rows = c.execute("SELECT id FROM parts WHERE part_no_norm=?", (nk,)).fetchall()
    if len(rows) == 1:
        return rows[0][0]
    if len(rows) > 1:
        return None
    rows = c.execute("SELECT DISTINCT part_id FROM part_aliases WHERE alias_part_no_norm=?", (nk,)).fetchall()
    if len(rows) == 1:
        return rows[0][0]
    return None


def _hist(c, item_id, project_id, upload_id, change_type, field, old, new,
          source, user_id, note=""):
    c.execute(
        "INSERT INTO bom_item_history (item_id, project_id, upload_id, change_type, field,"
        " old_value, new_value, source, note, changed_by) VALUES (?,?,?,?,?,?,?,?,?,?)",
        (item_id, project_id, upload_id, change_type, field,
         None if old is None else str(old), None if new is None else str(new),
         source, note, user_id or None),
    )


_ITEM_COLS = ["item_type", "category", "unit_code", "line_no", "part_no", "part_name",
              "maker", "vendor", "material", "finishing", "unit_count", "total_qty",
              "unit", "unit_price", "amount", "delivery_text", "buy_at", "remarks"]


def apply_upload(project_id: int, file_items: list, mode: str, user_id: int,
                 filename: str, sheet_names: str) -> dict:
    """비교 결과를 실제 반영 — 단일 트랜잭션.
    merge(기본): 파일에 없는 기존 라인은 건드리지 않음(부분 목록 안전) / replace: 삭제 처리."""
    plan = plan_diff(project_id, file_items)
    now = _logi_now()
    with db_session() as c:
        ver = int(c.execute(
            "SELECT COALESCE(MAX(version_no),0)+1 FROM bom_uploads WHERE project_id=?",
            (int(project_id),)).fetchone()[0])
        c.execute(
            "INSERT INTO bom_uploads (project_id, version_no, source_filename, sheet_names, mode,"
            " added, changed, deleted, missing, ordered_warn, uploaded_by)"
            " VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (int(project_id), ver, filename, sheet_names, mode,
             len(plan["adds"]), len(plan["changes"]), len(plan["deletes"]),
             len(plan["missing"]) if mode == "merge" else 0,
             plan["ordered_warn"], user_id or None))
        upload_id = c.execute("SELECT last_insert_rowid()").fetchone()[0]

        # ① 추가
        for f in plan["adds"]:
            vals = {k: f.get(k) for k in _ITEM_COLS}
            vals["buy_at"] = vals.get("buy_at") if vals.get("buy_at") in ("KOR", "VINA") else ""
            pid_match = _match_part_id(c, f.get("part_no"))
            c.execute(
                f"INSERT INTO bom_items (project_id, {','.join(_ITEM_COLS)}, part_no_norm, part_id,"
                f" first_upload_id, last_upload_id, created_at, updated_at)"
                f" VALUES (?{',?' * len(_ITEM_COLS)},?,?,?,?,?,?)",
                [int(project_id)] + [vals[k] for k in _ITEM_COLS]
                + [_normalize_part_key(str(f.get("part_no") or "")), pid_match,
                   upload_id, upload_id, now, now])
            iid = c.execute("SELECT last_insert_rowid()").fetchone()[0]
            note = "엑셀 신규(노랑) 표기" if f.get("is_new_marked") else ""
            _hist(c, iid, project_id, upload_id, "추가", "", None, f.get("part_no") or f.get("part_name"),
                  "업로드", user_id, note)
            for ch in f.get("excel_changes", []):
                _hist(c, iid, project_id, upload_id, "변경", ch["field"], ch["old"], ch["new"],
                      "엑셀표기(->)", user_id,
                      "⚠ 1→N 다부품 분해 표기 — 확인 필요" if ch.get("is_multi") else "신규 라인의 과거표기")

        # ② 변경
        for chg in plan["changes"]:
            iid = chg["item_id"]
            f = chg["file"]
            sets = []
            params = []
            for fld, ov in chg["fields"].items():
                sets.append(f"{fld}=?")
                params.append(ov["new"])
                _hist(c, iid, project_id, upload_id, "변경", fld, ov["old"], ov["new"],
                      "업로드", user_id)
            if "part_no" in chg["fields"]:
                new_pno = chg["fields"]["part_no"]["new"]
                sets.append("part_no_norm=?")
                params.append(_normalize_part_key(str(new_pno or "")))
                sets.append("part_id=?")
                params.append(_match_part_id(c, new_pno))
            for ch in f.get("excel_changes", []):
                _hist(c, iid, project_id, upload_id, "변경", ch["field"], ch["old"], ch["new"],
                      "엑셀표기(->)", user_id,
                      "⚠ 1→N 다부품 분해 표기 — 확인 필요" if ch.get("is_multi") else "")
            if chg["ordered"]:
                sets.append("review_flag=1")
                _hist(c, iid, project_id, upload_id, "변경", "review_flag", 0, 1, "업로드",
                      user_id, "🔴 발주 후 변경 — 발주 재검토 필요")
            sets.append("last_upload_id=?")
            params.append(upload_id)
            sets.append("updated_at=?")
            params.append(now)
            c.execute(f"UPDATE bom_items SET {', '.join(sets)} WHERE id=?", params + [iid])

        # ③ 삭제 (엑셀 취소선 표기)
        for d in plan["deletes"]:
            iid = d["item_id"]
            c.execute("UPDATE bom_items SET status='삭제', review_flag=CASE WHEN ? THEN 1 ELSE review_flag END,"
                      " last_upload_id=?, updated_at=? WHERE id=?",
                      (1 if d["ordered"] else 0, upload_id, now, iid))
            _hist(c, iid, project_id, upload_id, "삭제", "", d["board"].get("part_no"), None,
                  "엑셀표기(취소선)", user_id,
                  "🔴 발주 후 삭제 — 발주 취소 검토 필요" if d["ordered"] else "")

        # ④ replace 모드: 파일에 없는 활성 라인 삭제 처리
        if mode == "replace":
            for b in plan["missing"]:
                warn = (b.get("order_status") or "미발주") in _ORDERED_STATES
                c.execute("UPDATE bom_items SET status='삭제', review_flag=CASE WHEN ? THEN 1 ELSE review_flag END,"
                          " last_upload_id=?, updated_at=? WHERE id=?",
                          (1 if warn else 0, upload_id, now, b["id"]))
                _hist(c, b["id"], project_id, upload_id, "삭제", "", b.get("part_no"), None,
                      "업로드", user_id,
                      ("🔴 발주 후 삭제 — 발주 취소 검토 필요. " if warn else "") + "전체 교체 모드 — 파일에 없는 라인")
                if warn:
                    plan["ordered_warn"] += 1
            c.execute("UPDATE bom_uploads SET deleted=deleted+?, ordered_warn=? WHERE id=?",
                      (len(plan["missing"]), plan["ordered_warn"], upload_id))
    return {"upload_id": upload_id, "version_no": ver,
            "added": len(plan["adds"]), "changed": len(plan["changes"]),
            "deleted": len(plan["deletes"]) + (len(plan["missing"]) if mode == "replace" else 0),
            "missing": len(plan["missing"]) if mode == "merge" else 0,
            "unchanged": plan["unchanged"], "ordered_warn": plan["ordered_warn"]}


# ── 조회 ──
def get_board(project_id: int, include_deleted: bool = False) -> list:
    sql = ("SELECT bi.*, p.stock_qty AS master_stock, p.part_no AS master_part_no"
           "  FROM bom_items bi LEFT JOIN parts p ON p.id = bi.part_id"
           " WHERE bi.project_id=?")
    if not include_deleted:
        sql += " AND bi.status='활성'"
    sql += " ORDER BY CASE WHEN bi.category IS NULL OR bi.category='' THEN 1 ELSE 0 END, bi.category, bi.line_no, bi.id"
    with db_session() as c:
        rows = [dict(r) for r in c.execute(sql, (int(project_id),)).fetchall()]
    for r in rows:
        r["category"] = r.get("category") or ""   # groupby 안전 (None 섞이면 정렬 오류)
    return rows


def get_item(item_id: int):
    with db_session() as c:
        r = c.execute("SELECT * FROM bom_items WHERE id=?", (int(item_id),)).fetchone()
        return dict(r) if r else None


def get_item_history(item_id: int, limit: int = 100) -> list:
    with db_session() as c:
        return [dict(r) for r in c.execute(
            "SELECT h.*, u.name AS changed_by_name, up.version_no"
            "  FROM bom_item_history h"
            "  LEFT JOIN users u ON u.id = h.changed_by"
            "  LEFT JOIN bom_uploads up ON up.id = h.upload_id"
            " WHERE h.item_id=? ORDER BY h.id DESC LIMIT ?",
            (int(item_id), int(limit))).fetchall()]


_EDITABLE = {"part_no", "part_name", "maker", "vendor", "material", "finishing",
             "category", "unit_code", "unit_count", "total_qty", "unit",
             "unit_price", "delivery_text", "buy_at", "remarks", "order_status",
             "review_flag", "status"}


def update_item(item_id: int, fields: dict, user_id: int, source: str = "수동", note: str = "") -> dict:
    """화면 수정 — 바뀐 필드만 반영 + 필드별 이력. 반환 {field: {old,new}}."""
    cur = get_item(item_id)
    if not cur:
        raise ValueError("BOM 라인을 찾을 수 없습니다")
    changed = {}
    now = _logi_now()
    with db_session() as c:
        sets = []
        params = []
        for f, nv in fields.items():
            if f not in _EDITABLE:
                continue
            if f in ("unit_count", "total_qty", "unit_price"):
                nv2 = _num(nv)
                if abs(_num(cur.get(f)) - nv2) <= 0.0001:
                    continue
                nv = nv2
            elif f == "review_flag":
                nv = 1 if str(nv) in ("1", "true", "True") else 0
                if int(cur.get(f) or 0) == nv:
                    continue
            else:
                nv = (str(nv).strip() if nv is not None else "")
                if (str(cur.get(f)).strip() if cur.get(f) is not None else "") == nv:
                    continue
            changed[f] = {"old": cur.get(f), "new": nv}
            sets.append(f"{f}=?")
            params.append(nv)
            ct = {"buy_at": "구매처", "order_status": "발주상태", "status": "삭제" if nv == "삭제" else "수동",
                  "part_no": "이원화" if f == "part_no" else "수동"}.get(f, "수동")
            _hist(c, item_id, cur["project_id"], None, ct, f, cur.get(f), nv, source, user_id, note)
        if not changed:
            return {}
        if "part_no" in changed:
            sets.append("part_no_norm=?")
            params.append(_normalize_part_key(str(changed["part_no"]["new"] or "")))
            sets.append("part_id=?")
            params.append(_match_part_id(c, changed["part_no"]["new"]))
        # 수량·단가 수정 시 금액 자동 재계산
        if any(f in changed for f in ("unit_count", "total_qty", "unit_price")):
            q = _num(fields.get("total_qty", cur.get("total_qty"))) or _num(fields.get("unit_count", cur.get("unit_count")))
            pr = _num(fields.get("unit_price", cur.get("unit_price")))
            sets.append("amount=?")
            params.append(round(q * pr, 2))
        sets.append("updated_at=?")
        params.append(now)
        c.execute(f"UPDATE bom_items SET {', '.join(sets)} WHERE id=?", params + [int(item_id)])
    return changed


def set_buy_at_bulk(project_id: int, item_ids: list, val: str, user_id: int) -> int:
    val = (val or "").upper()
    if val not in ("KOR", "VINA", ""):
        raise ValueError("구매처는 KOR / VINA / 해제만 가능합니다")
    n = 0
    now = _logi_now()
    with db_session() as c:
        for iid in item_ids:
            r = c.execute("SELECT id, project_id, buy_at FROM bom_items WHERE id=? AND project_id=?",
                          (int(iid), int(project_id))).fetchone()
            if not r or (r["buy_at"] or "") == val:
                continue
            c.execute("UPDATE bom_items SET buy_at=?, updated_at=? WHERE id=?", (val, now, int(iid)))
            _hist(c, int(iid), int(project_id), None, "구매처", "buy_at", r["buy_at"], val, "수동", user_id)
            n += 1
    return n


def find_project_by_code(code: str):
    code = (code or "").strip()
    if not code:
        return None
    with db_session() as c:
        r = c.execute("SELECT id, mgmt_code, name FROM projects WHERE mgmt_code=?", (code,)).fetchone()
        return dict(r) if r else None


def projects_pick_list(limit: int = 300) -> list:
    with db_session() as c:
        return [dict(r) for r in c.execute(
            "SELECT id, mgmt_code, name FROM projects"
            " WHERE COALESCE(mgmt_code,'')<>'' ORDER BY id DESC LIMIT ?", (int(limit),)).fetchall()]


def bom_projects_summary(limit: int = 100) -> list:
    """BOM이 있는 프로젝트 목록 (업로드 화면 하단 안내용)."""
    with db_session() as c:
        return [dict(r) for r in c.execute(
            "SELECT bi.project_id, pr.mgmt_code, pr.name,"
            "       COUNT(*) AS n_items,"
            "       SUM(CASE WHEN bi.review_flag=1 THEN 1 ELSE 0 END) AS n_review,"
            "       MAX(bi.updated_at) AS last_at"
            "  FROM bom_items bi JOIN projects pr ON pr.id=bi.project_id"
            " WHERE bi.status='활성'"
            " GROUP BY bi.project_id ORDER BY last_at DESC LIMIT ?", (int(limit),)).fetchall()]


def list_uploads(project_id: int, limit: int = 20) -> list:
    with db_session() as c:
        return [dict(r) for r in c.execute(
            "SELECT up.*, u.name AS uploaded_by_name FROM bom_uploads up"
            "  LEFT JOIN users u ON u.id=up.uploaded_by"
            " WHERE up.project_id=? ORDER BY up.id DESC LIMIT ?",
            (int(project_id), int(limit))).fetchall()]


def get_recent_changes(project_id: int):
    """v5H226z987 (대표 요구 '기존 내용도 같이 확인'): 최신 업로드에서 바뀐 라인 하이라이트용.
    반환: (latest_upload dict|None, {item_id: {"types": [변경/추가/삭제...], "fields": {field: {old,new}}}})"""
    with db_session() as c:
        up = c.execute(
            "SELECT * FROM bom_uploads WHERE project_id=? ORDER BY id DESC LIMIT 1",
            (int(project_id),)).fetchone()
        if not up:
            return None, {}
        rows = c.execute(
            "SELECT item_id, change_type, field, old_value, new_value"
            "  FROM bom_item_history WHERE upload_id=?", (up["id"],)).fetchall()
    m: dict = {}
    for r in rows:
        e = m.setdefault(r["item_id"], {"types": [], "fields": {}})
        ct = r["change_type"] or ""
        if ct and ct not in e["types"]:
            e["types"].append(ct)
        if r["field"]:
            e["fields"][r["field"]] = {"old": r["old_value"], "new": r["new_value"]}
    return dict(up), m


def get_upload_report(upload_id: int):
    """v5H226z987: 특정 업로드(버전)의 변경 리포트 — 과거 어느 버전이든 '그때 뭐가 바뀌었나' 전체 확인.
    반환: (upload dict|None, [{item_id, part_no, part_name, category, status, entries:[{type,field,old,new,note}]}])"""
    with db_session() as c:
        up = c.execute(
            "SELECT up.*, u.name AS uploaded_by_name, pr.mgmt_code, pr.name AS project_name, pr.id AS pid"
            "  FROM bom_uploads up"
            "  LEFT JOIN users u ON u.id = up.uploaded_by"
            "  LEFT JOIN projects pr ON pr.id = up.project_id"
            " WHERE up.id=?", (int(upload_id),)).fetchone()
        if not up:
            return None, []
        hist = c.execute(
            "SELECT h.item_id, h.change_type, h.field, h.old_value, h.new_value, h.source, h.note,"
            "       bi.part_no, bi.part_name, bi.category, bi.status"
            "  FROM bom_item_history h"
            "  LEFT JOIN bom_items bi ON bi.id = h.item_id"
            " WHERE h.upload_id=? ORDER BY h.item_id, h.id", (int(upload_id),)).fetchall()
    items: dict = {}
    order: list = []
    for r in hist:
        iid = r["item_id"]
        if iid not in items:
            items[iid] = {"item_id": iid, "part_no": r["part_no"], "part_name": r["part_name"],
                          "category": r["category"] or "", "status": r["status"] or "", "entries": []}
            order.append(iid)
        items[iid]["entries"].append({
            "type": r["change_type"] or "", "field": r["field"] or "",
            "old": r["old_value"], "new": r["new_value"],
            "source": r["source"] or "", "note": r["note"] or ""})
    return dict(up), [items[i] for i in order]


def bom_purge_project(project_id: int) -> dict:
    """v5H226z986 (대표 지시 '테스트 데이터 나중에 삭제'): 이 프로젝트의 BOM 데이터 전체 삭제.
    품목·업로드 기록·이력 3테이블을 비움 — 프로젝트 자체는 유지(폐기와 별개).
    되돌릴 수 없으므로 라우트에서 관리번호 타이핑 확인을 거친 뒤에만 호출.
    WP-01(P0-05·게이트 F-03): 운영 BOM 보호 — 테스트 관리번호(999/A 접두)가 아니면
    함수 차원에서도 차단(라우트의 admin/ceo·접두 검사와 2겹)."""
    with db_session() as c:
        _pr = c.execute("SELECT mgmt_code FROM projects WHERE id=?", (int(project_id),)).fetchone()
        _code = (((_pr["mgmt_code"] if _pr else "") or "")).strip()
        if not (_code.startswith("999") or _code.upper().startswith("A")):
            raise ValueError("운영 BOM은 폐기할 수 없습니다 — 테스트 관리번호(999/A 접두)만 허용됩니다")
        n_i = c.execute("SELECT COUNT(*) FROM bom_items WHERE project_id=?", (int(project_id),)).fetchone()[0]
        n_u = c.execute("SELECT COUNT(*) FROM bom_uploads WHERE project_id=?", (int(project_id),)).fetchone()[0]
        n_h = c.execute("SELECT COUNT(*) FROM bom_item_history WHERE project_id=?", (int(project_id),)).fetchone()[0]
        c.execute("DELETE FROM bom_item_history WHERE project_id=?", (int(project_id),))
        c.execute("DELETE FROM bom_items WHERE project_id=?", (int(project_id),))
        c.execute("DELETE FROM bom_uploads WHERE project_id=?", (int(project_id),))
    return {"items": int(n_i or 0), "uploads": int(n_u or 0), "history": int(n_h or 0)}


def export_xlsx(project_id: int) -> bytes:
    """현재 활성 BOM → 엑셀 (실양식 컬럼 순서·협력사 전달/보관용)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    rows = get_board(project_id)
    with db_session() as c:
        pr = c.execute("SELECT mgmt_code, name FROM projects WHERE id=?", (int(project_id),)).fetchone()
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"
    title = f"{(pr['mgmt_code'] if pr else '')} BOM — HAIST WORKS 현재본 ({_date.today().isoformat()})"
    ws.cell(1, 1, title).font = Font(bold=True, size=13)
    headers = ["NO", "구분(CATEGORY)", "CODE", "PRODUCT NAME\n제품명", "PRODUCT CODE\n모델명",
               "MANUFACTURER\n제조사", "외주사", "재질", "후처리", "대당수량", "총수량", "단위",
               "단가", "금액", "납기", "구매처(KOR/VINA)", "발주상태", "재검토", "비고"]
    hf = PatternFill("solid", fgColor="DBEAFE")
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(3, ci, h)
        cell.font = Font(bold=True, size=10)
        cell.fill = hf
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    r = 4
    for i, it in enumerate(rows, start=1):
        vals = [i, it.get("category"), it.get("unit_code"), it.get("part_name"), it.get("part_no"),
                it.get("maker"), it.get("vendor"), it.get("material"), it.get("finishing"),
                it.get("unit_count"), it.get("total_qty"), it.get("unit"),
                it.get("unit_price"), it.get("amount"), it.get("delivery_text"),
                it.get("buy_at"), it.get("order_status"),
                "🔴 재검토" if it.get("review_flag") else "", it.get("remarks")]
        for ci, v in enumerate(vals, start=1):
            ws.cell(r, ci, v)
        if it.get("review_flag"):
            for ci in range(1, len(headers) + 1):
                ws.cell(r, ci).fill = PatternFill("solid", fgColor="FEE2E2")
        r += 1
    widths = [5, 18, 7, 26, 26, 14, 12, 10, 10, 9, 9, 6, 11, 13, 14, 14, 9, 9, 18]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(3, ci).column_letter].width = w
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
