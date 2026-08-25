#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
후보 ①-b — 간이판(기구) + 전장 BOM → 통일판(26칸) 마스터 뼈대 취합
====================================================================
대표 승인 2026-08-25 "바로 진행해.. 진행 완료 후 항상 직접 실행해서 검증해야해."

무엇을 하나
  설계팀 간이판(구매품 PART LIST 초안)과 전장팀 BOM(같은 양식 — 김형렬 팀장 확인)을
  받아, 구매팀이 손으로 하던 「한 시트로 취합」을 대신한다:
  통일판(26칸) 양식에 품목 신상(구분·CODE·품명·형번·제조사·협력사·수량·단위)을 채우고
  수식(TOTAL·발주수량·재고금액·발주금액·합계·SUBTOTAL 합계줄)을 실물 그대로 심는다.

일부러 안 하는 것 (인터뷰 확정 규칙)
  ⛔ 단가·납기·발주일을 지어 넣지 않는다 — 구매팀 몫. 입력에 값이 있어도 옮기지 않고
     「옮기지 않은 값」으로 보고만 한다 (단가 기준은 수불대장 — 허동준 매니저).
  ⛔ 중복 형번 합산 금지 — 줄대로 (유닛별 포장 구분).
  ⛔ B열(NO.)에 번호를 넣지 않는다 — 실물에서 B열은 비교·삭제·추가 같은 상태 낱말 자리다.
  · 「전장 구매품 별도」 자리표시 줄은 취합 시 걷어내고 보고한다.

양식은 어디서 오나
  통일판 빈 양식이 따로 없어 **001M2606 실물에서 도출**했다(--make-template).
  도출본(`양식_통일판_구매품PARTLIST.xlsx`)은 001M·008M 머리글과 칸 단위로 같음을 시험이 보증.

사용법
  python bom_merge_to_master.py 간이판.xlsx 전장BOM.xlsx --code 001M2606 \
      --name "Clip Attach Machine" --author 허동준 --out 마스터뼈대.xlsx
"""
import argparse
import os
import sys
from copy import copy as _copy

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_TEMPLATE = os.path.join(HERE, "양식", "양식_통일판_구매품PARTLIST.xlsx")

ROW0 = 9                      # 통일판 자료 시작줄 (7~8행 = 머리글)
C_FROM, C_TO = 2, 27          # B~AA
SUBTOTAL_COLS = (10, 11, 13, 14, 15, 16, 18, 20, 21, 22, 23)   # J K M N O P R T U V W
PLACEHOLDER = "전장 구매품 별도"


def _norm(v):
    return str(v or "").strip().lower().replace(" ", "").replace("\n", "")


# 간이판 칸 — 자리(열)가 아니라 머리글 이름으로 찾는다 (담당자별 차이 대비)
DRAFT_KEYS = {
    "구분":  lambda h: "category" in h or h.startswith("구분"),
    "CODE": lambda h: h == "code",
    "품명":  lambda h: "productname" in h or "제품명" in h,
    "형번":  lambda h: "productcode" in h or "코드명" in h,
    "제조사": lambda h: "manufacturer" in h or "제조사" in h,
    "협력사": lambda h: "vendor" in h or "외주사" in h,
    "수량":  lambda h: "unitcount" in h or h.startswith("수량"),
    "단위":  lambda h: h.startswith("unit/") or ("unit" in h and "단위" in h),
    "단가":  lambda h: "unitprice" in h or ("단가" in h and "vat" not in h),
    "납기":  lambda h: "delivery/" in h or ("delivery" in h and "납기" in h),
    "발주일": lambda h: "발주일" in h,
    "입고예정": lambda h: "입고예정" in h,
    "비고":  lambda h: "remark" in h or "비고" in h,
}
REQUIRED = ("구분", "품명", "형번", "수량")


def _find_draft_columns(ws):
    for hr in range(1, 11):
        found = {}
        for ci in range(1, (ws.max_column or 0) + 1):
            h = _norm(ws.cell(row=hr, column=ci).value)
            if not h:
                continue
            for key, match in DRAFT_KEYS.items():
                if key not in found and match(h):
                    found[key] = ci
        if all(k in found for k in REQUIRED):
            return hr, found
    return None, {}


def _qty(v):
    try:
        f = float(str(v).strip())
        return int(f) if f == int(f) else f
    except Exception:
        return None


def read_draft_files(paths):
    """간이판 양식 파일들 → (품목 줄, 보고). 머리글을 못 찾으면 ValueError."""
    rows, report = [], {"자리표시제거": 0, "옮기지않음": [], "파일": []}
    for f in paths:
        wb = load_workbook(f, data_only=True)
        ws = wb.active
        hr, col = _find_draft_columns(ws)
        if hr is None:
            raise ValueError(f"[{os.path.basename(f)}] 간이판 양식 머리글(구분·제품명·코드명·수량)을 찾지 못했습니다.\n"
                             f"  → 설계팀 간이판/전장 BOM 양식 파일이 맞는지 확인해 주십시오.")

        def get(r, key):
            ci = col.get(key)
            return ws.cell(row=r, column=ci).value if ci else None

        n = 0
        for r in range(hr + 1, (ws.max_row or hr) + 1):
            name = str(get(r, "품명") or "").strip()
            spec = str(get(r, "형번") or "").strip()
            gubun = str(get(r, "구분") or "").strip()
            if not name and not spec:
                if gubun == PLACEHOLDER:
                    report["자리표시제거"] += 1
                continue
            for key in ("단가", "납기", "발주일", "입고예정"):
                v = get(r, key)
                if v not in (None, "", 0):
                    report["옮기지않음"].append((os.path.basename(f), r, key, str(v)[:20]))
            rows.append({"구분": gubun, "CODE": str(get(r, "CODE") or "").strip(),
                         "품명": name, "형번": spec,
                         "제조사": str(get(r, "제조사") or "").strip(),
                         "협력사": str(get(r, "협력사") or "").strip(),
                         "수량": _qty(get(r, "수량")),
                         "단위": str(get(r, "단위") or "").strip(),
                         "비고": str(get(r, "비고") or "").strip()})
            n += 1
        report["파일"].append((os.path.basename(f), n))
    return rows, report


def make_master_template(source_master, out_path):
    """통일판 실물(001M 등)에서 빈 양식을 도출한다 — 머리글·수식 서식은 남기고 자료만 비운다."""
    wb = load_workbook(source_master)
    ws = wb.active
    base = {ci: _copy(ws.cell(row=ROW0, column=ci)._style) for ci in range(C_FROM, C_TO + 1)}
    for r in range(ROW0, (ws.max_row or ROW0) + 1):
        for ci in range(C_FROM, C_TO + 1):
            c = ws.cell(row=r, column=ci)
            c.value = None
            c._style = _copy(base[ci])
    ws["C2"] = "000M0000 구매 BOM\n(프로젝트명)"
    ws["C5"] = "AUTHOR : "
    ws["J5"] = 1
    wb.save(out_path)
    return out_path


def write_master(rows, code, name, author, out_path, template=DEFAULT_TEMPLATE, sets=1):
    """통일판 뼈대에 품목을 채우고 실물과 같은 수식을 심는다."""
    wb = load_workbook(template)
    ws = wb.active
    base = {ci: _copy(ws.cell(row=ROW0, column=ci)._style) for ci in range(C_FROM, C_TO + 1)}

    ws["C2"] = f"{code} 구매 BOM\n{name}"
    ws["C5"] = f"AUTHOR : {author}"
    ws["J5"] = sets

    def styled(r, ci, value=None):
        c = ws.cell(row=r, column=ci)
        c._style = _copy(base[ci])
        if value is not None:
            c.value = value
        return c

    def put_formulas(r):
        styled(r, 11, f"=J{r}*$J$5")                 # K TOTAL
        styled(r, 15, f"=$K{r}-($M{r}+$N{r})")       # O 발주수량
        styled(r, 21, f"=T{r}*(M{r}+N{r})")          # U 재고금액
        styled(r, 22, f"=O{r}*T{r}")                 # V 발주금액
        styled(r, 23, f"=U{r}+V{r}")                 # W 합계

    unit_defaulted = 0
    r = ROW0
    for x in rows:
        for ci in range(C_FROM, C_TO + 1):
            styled(r, ci)
        styled(r, 3, x["구분"] or None)
        styled(r, 4, x["CODE"] or None)
        styled(r, 5, x["품명"] or None)
        styled(r, 6, x["형번"] or None)
        styled(r, 7, x["제조사"] or None)
        styled(r, 8, x["협력사"] or None)
        styled(r, 10, x["수량"])                      # J 수량
        if x["단위"]:
            styled(r, 12, x["단위"])
        else:
            styled(r, 12, "EA")                       # 실물 관행(여분 줄도 EA 선입력) 그대로
            unit_defaulted += 1
        if x["비고"]:
            styled(r, 27, x["비고"])
        put_formulas(r)
        r += 1

    for _ in range(2):                                # 실물 관행: 여분 수식 줄 2
        for ci in range(C_FROM, C_TO + 1):
            styled(r, ci)
        styled(r, 12, "EA")
        put_formulas(r)
        r += 1
    gap = r                                           # 빈 줄 1
    total_row = r + 1
    last = gap                                        # SUBTOTAL 범위 = 자료+여분+빈줄
    for ci in range(C_FROM, C_TO + 1):
        styled(total_row, ci)
    from openpyxl.utils import get_column_letter as gl
    for ci in SUBTOTAL_COLS:
        L = gl(ci)
        styled(total_row, ci, f"=SUBTOTAL(9,{L}{ROW0}:{L}{last})")

    wb.save(out_path)
    return {"품목": len(rows), "단위기본값": unit_defaulted, "합계줄": total_row}


def main(argv=None):
    ap = argparse.ArgumentParser(description="간이판+전장 BOM → 통일판 마스터 뼈대 취합")
    ap.add_argument("inputs", nargs="*", help="간이판/전장 BOM .xlsx 파일들 (취합 순서대로)")
    ap.add_argument("--code", help="관리번호 (예: 001M2606)")
    ap.add_argument("--name", help="프로젝트 이름")
    ap.add_argument("--author", default="", help="작성자")
    ap.add_argument("--sets", type=int, default=1, help="제작 대수 (기본 1)")
    ap.add_argument("--out", default=None)
    ap.add_argument("--template", default=DEFAULT_TEMPLATE)
    ap.add_argument("--make-template", metavar="실물마스터", help="통일판 실물에서 빈 양식 도출")
    a = ap.parse_args(argv)

    if a.make_template:
        out = a.out or DEFAULT_TEMPLATE
        make_master_template(a.make_template, out)
        print(f"통일판 빈 양식 도출: {out}")
        return 0

    if not a.inputs or not a.code or not a.name:
        ap.error("취합하려면 입력 파일들과 --code --name 이 필요합니다.")
    rows, rep = read_draft_files(a.inputs)
    if not rows:
        print("옮길 품목이 한 줄도 없습니다.")
        return 2
    out = a.out or f"{a.code} 구매품 PART LIST_취합뼈대.xlsx"
    res = write_master(rows, a.code, a.name, a.author, out, a.template, a.sets)

    print("=" * 64)
    print(f"  {a.code} 통일판 마스터 뼈대 — {res['품목']}줄 (대수 {a.sets})")
    print("=" * 64)
    for fn, n in rep["파일"]:
        print(f"  읽음 {fn:<40} {n:>3}줄")
    if rep["자리표시제거"]:
        print(f"  「{PLACEHOLDER}」 자리표시 줄 {rep['자리표시제거']}개 걷어냄 (전장 실물로 대체)")
    if res["단위기본값"]:
        print(f"  단위 빈칸 {res['단위기본값']}줄 → 'EA' 기본값 (실물 관행)")
    if rep["옮기지않음"]:
        print(f"  ⚠ 옮기지 않은 입력값 {len(rep['옮기지않음'])}건 — 단가·납기류는 구매팀 몫(수불대장 기준):")
        for fn, r, k, v in rep["옮기지않음"][:10]:
            print(f"     {fn} r{r} {k}={v}")
    print(f"  수식: TOTAL·발주수량·재고금액·발주금액·합계 + SUBTOTAL 합계줄(r{res['합계줄']}) — 실물과 동일")
    print(f"\n  저장: {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
