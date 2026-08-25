#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
후보 ② (2/2) — 통일판 마스터 → 협력사별 견적요청서 자동 생성
=============================================================
대표 승인 2026-08-25 "바로 진행". 양식·칸 역할 = 허동준 매니저 실물+설명(08-25 10:54) 그대로.

칸 역할 (허동준 매니저 확인)
  · 기존단가(K) = **실적가 — 당사 구매팀 기입** → 마스터 P(기존단가)를 그대로 옮긴다
  · 변경단가(L)·소요일(R)·입고가능일(S) = **매입처 기입** → 비워 둔다
  · 입고요청일(Q) = 당사가 원하는 날짜 → --due 로 넣는다
  · 발주서와 달리 「옛->새」 변경표기는 **그대로 둔다** (매입처가 변경을 봐야 함 — 실물 메일 실측)

사용법
  python bom_make_rfq.py 마스터.xlsx --vendor 파익스 --due 2026-08-29
  python bom_make_rfq.py 마스터.xlsx --vendor 성화기전 --only-missing   (실적가 없는 신규 품목만)
"""
import argparse
import datetime
import os
import re
import sys

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
try:
    from .make_po import read_master             # 패키지로 쓸 때 (앱)
except ImportError:
    sys.path.insert(0, HERE)
    from make_po import read_master              # 명령줄로 바로 돌릴 때

DEFAULT_TEMPLATE = os.path.join(HERE, "양식", "양식_견적요청서.xlsx")
SHEET = "업체별 견적 요청 (단품)"
ITEM0 = 6                                        # 자료 시작줄 (4~5행 = 머리글)
# B순번 C영업 D관리 E코드 F진행수량 G품명 H규격 I메이커 J협력사 K기존단가 L변경단가
# M요청수량 N단위 O기존금액 P변경금액 Q입고요청일 R소요일 S입고가능일 T비고
CLEAR_COLS = (3, 5, 6, 7, 8, 9, 10, 11, 12, 13, 17, 18, 19, 20)


def make_rfq_template(golden, out_path):
    """실물 견적요청서에서 빈 양식을 도출 — 번호·'M'·'EA'·금액수식·합계줄은 남긴다."""
    wb = load_workbook(golden)
    ws = wb[SHEET]
    total_row = _find_total_row(ws)
    for r in range(ITEM0, total_row):
        for ci in CLEAR_COLS:
            ws.cell(row=r, column=ci).value = None
    wb.save(out_path)
    return out_path


def _find_total_row(ws):
    for r in range(ITEM0, (ws.max_row or ITEM0) + 1):
        if str(ws.cell(row=r, column=10).value or "").replace(" ", "") == "합계":
            return r
    raise ValueError("견적요청서 양식에서 합계 줄을 찾지 못했습니다.")


def _split_code(code):
    """'001M2606' → ('001','M','2606'). 형식이 다르면 최대한 나누고 남는 건 앞칸에."""
    m = re.match(r"^(.+?)([A-Z])(\d{4})$", str(code or "").strip())
    return (m.group(1), m.group(2), m.group(3)) if m else (str(code), "", "")


def write_rfq(items, vendor, code, out_path, template=DEFAULT_TEMPLATE, due=""):
    if not items:
        raise ValueError(f"협력사 「{vendor}」 의 산 줄이 없습니다.")
    wb = load_workbook(template)
    ws = wb[SHEET]
    total_row = _find_total_row(ws)
    slots = total_row - ITEM0
    if len(items) > slots:
        raise ValueError(f"양식 칸({slots}줄)보다 품목({len(items)}줄)이 많습니다. 나눠서 뽑아 주십시오.")

    c1, c2, c3 = _split_code(code)
    no_price = arrows = 0
    for i, x in enumerate(items):
        r = ITEM0 + i
        ws.cell(row=r, column=2).value = i + 1          # 순번 다시 매김
        ws.cell(row=r, column=3).value = c1
        ws.cell(row=r, column=4).value = c2
        ws.cell(row=r, column=5).value = c3
        ws.cell(row=r, column=7).value = x["품명"]       # 「옛->새」 그대로
        ws.cell(row=r, column=8).value = x["형번"]
        ws.cell(row=r, column=9).value = x["제조사"]
        ws.cell(row=r, column=10).value = vendor
        if x["기존단가"] in (None, "", 0):
            no_price += 1                               # 신규 — 매입처가 채울 것
        else:
            ws.cell(row=r, column=11).value = x["기존단가"]
        ws.cell(row=r, column=13).value = x["수량"]
        ws.cell(row=r, column=14).value = x["단위"]
        if due:
            ws.cell(row=r, column=17).value = due
        if x["비고"]:
            ws.cell(row=r, column=20).value = x["비고"]
        arrows += sum(1 for f in ("품명", "형번", "제조사") if "->" in str(x[f]))
    wb.save(out_path)
    return {"품목": len(items), "신규단가요청": no_price, "변경표기유지": arrows}


def main(argv=None):
    ap = argparse.ArgumentParser(description="통일판 마스터 → 협력사별 견적요청서 생성")
    ap.add_argument("master")
    ap.add_argument("--vendor", required=True, help="협력사 이름")
    ap.add_argument("--due", default="", help="입고 요청일 (예: 2026-08-29)")
    ap.add_argument("--only-missing", action="store_true", help="실적가 없는 신규 품목만")
    ap.add_argument("--out", default=None)
    ap.add_argument("--template", default=DEFAULT_TEMPLATE)
    ap.add_argument("--make-template", metavar="실물견적요청서")
    a = ap.parse_args(argv)

    if a.make_template:
        make_rfq_template(a.make_template, DEFAULT_TEMPLATE)
        print(f"견적요청서 빈 양식 도출: {DEFAULT_TEMPLATE}")
        return 0

    code, name, sets, rows, vina = read_master(a.master)
    items = [x for x in rows if x["협력사"] == a.vendor]
    if a.only_missing:
        items = [x for x in items if x["기존단가"] in (None, "", 0)]
    due = a.due
    stamp = datetime.date.today().strftime("%Y%m%d")
    out = a.out or f"{code} {a.vendor} 견적요청서_{stamp}.xlsx"
    res = write_rfq(items, a.vendor, code, out, a.template, due)

    print("=" * 64)
    print(f"  {code} {a.vendor} 견적요청서 — {res['품목']}줄")
    print("=" * 64)
    if res["신규단가요청"]:
        print(f"  실적가 없는 신규 품목 {res['신규단가요청']}줄 — 단가 칸 비움(매입처 기입 요청)")
    if res["변경표기유지"]:
        print(f"  「옛->새」 변경표기 {res['변경표기유지']}건 그대로 둠 (매입처가 변경 확인)")
    print(f"  변경단가·소요일·입고가능일 칸 = 매입처 기입 몫 — 비움")
    print(f"\n  저장: {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
