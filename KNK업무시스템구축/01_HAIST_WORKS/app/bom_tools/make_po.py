#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
후보 ② — 통일판 마스터 → 협력사별 발주서 자동 생성
====================================================
대표 승인 2026-08-25 "우선 중요한 사항 먼저 진행" (거래처 목록 데이터는 나중에 얹기로).

무엇을 하나
  통일판 마스터(26칸)에서 협력사를 고르면(또는 --all 전체) 그 협력사의 산 줄만 골라
  실제 발주서 양식(허동준 매니저 실물에서 도출: 머리+표+합계+도장)으로 만들어 준다.

규칙 (인터뷰·실물 실측 그대로)
  · 품목 = 그 협력사의 산 줄 (⛔비교·삭제·실패 줄 제외 — 단가 0원 처리 줄들)
  · 수량 = 수량(J)×대수(J5) − 사내재고(M) − 베트남재고(N)  ← 실물 수식 그대로
  · 단가 = 확정단가(T). 비어 있으면 지어내지 않고 빈칸 + 보고 (단가는 수불대장 몫)
  · VINA 줄은 발주서 대상이 아님 (베트남 담당자 협의 — 허동준 매니저) → 제외 보고
  · 협력사 전화·메일은 비워 둠 — 거래처 목록 데이터가 오면 자동 채움 (대표 결정: 나중에)
  · 구매팀(보내는 쪽) 연락처는 실물 양식에 있던 그대로 유지

사용법
  python bom_make_po.py 마스터.xlsx --vendor 광원전기 --due 2026-08-25
  python bom_make_po.py 마스터.xlsx --all --outdir 발주서묶음/
  (--code·--name 은 마스터 제목칸에서 자동으로 읽는다)
"""
import argparse
import datetime
import os
import re
import sys
from copy import copy as _copy

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_TEMPLATE = os.path.join(HERE, "양식", "양식_발주서.xlsx")

SHEET = "구매요청서"
ITEM0 = 12                     # 품목 표 시작줄
COLS = {"품명": 2, "규격": 3, "MAKER": 4, "수량": 5, "단위": 6, "단가": 7, "금액": 8, "비고": 9}
DEAD_TAGS = ("비교", "삭제", "실패")


def _label_of(text):
    """'관리 코드 : 001M2606' → '관리 코드 : ' (라벨만, 원래 띄어쓰기 보존)"""
    s = str(text or "")
    i = s.find(":")
    return (s[: i + 1] + " ") if i >= 0 else s


def make_po_template(golden_po, out_path):
    """실물 발주서에서 빈 양식을 도출한다 — 표·합계 수식·도장은 남기고 값만 비운다."""
    wb = load_workbook(golden_po)
    for name in [n for n in wb.sheetnames if n != SHEET]:
        del wb[name]
    ws = wb[SHEET]
    total_row = _find_total_row(ws)
    for r in range(ITEM0, total_row - 1):          # 슬롯 (사이 빈 줄 제외)
        for key in ("품명", "규격", "MAKER", "수량", "비고"):
            ws.cell(row=r, column=COLS[key]).value = None
        ws.cell(row=r, column=COLS["단가"]).value = None
        ws.cell(row=r, column=COLS["단위"]).value = "EA"
    for addr in ("A4", "A5", "A6", "A7", "A8"):    # 관리코드·공급업체·전화·메일·요청납기
        ws[addr] = _label_of(ws[addr].value)
    ws["I4"] = None                                # 발주일자 값
    wb.save(out_path)
    return out_path


def _find_total_row(ws):
    for r in range(ITEM0, (ws.max_row or ITEM0) + 1):
        if str(ws.cell(row=r, column=1).value or "").replace(" ", "") == "합계":
            return r
    raise ValueError("발주서 양식에서 합계 줄을 찾지 못했습니다.")


def read_master(path):
    """통일판 마스터 → (제목의 관리번호·프로젝트명, 대수, 산 줄 목록, VINA 제외 수)."""
    ws = load_workbook(path, data_only=True).active
    title = str(ws["C2"].value or "")
    m = re.match(r"\s*(\S+)", title)
    code = m.group(1) if m else ""
    name = title.splitlines()[1].strip() if len(title.splitlines()) > 1 else ""
    sets = ws["J5"].value if isinstance(ws["J5"].value, (int, float)) and ws["J5"].value else 1
    rows, vina = [], 0
    for r in range(9, (ws.max_row or 9) + 1):
        pn = str(ws.cell(row=r, column=5).value or "").strip()
        spec = str(ws.cell(row=r, column=6).value or "").strip()
        if not pn and not spec:
            continue
        tag = str(ws.cell(row=r, column=2).value or "").strip()
        if tag in DEAD_TAGS:
            continue
        vendor = str(ws.cell(row=r, column=8).value or "").strip()
        if vendor.upper() == "VINA":
            vina += 1
            continue
        j = ws.cell(row=r, column=10).value
        mm = ws.cell(row=r, column=13).value or 0
        nn = ws.cell(row=r, column=14).value or 0
        try:
            qty = (float(j) if j not in (None, "") else 0) * float(sets) - float(mm) - float(nn)
            qty = int(qty) if qty == int(qty) else qty
        except Exception:
            qty = j
        rows.append({"협력사": vendor, "품명": pn, "형번": spec,
                     "제조사": str(ws.cell(row=r, column=7).value or "").strip(),
                     "수량": qty,
                     "단위": str(ws.cell(row=r, column=12).value or "").strip() or "EA",
                     "단가": ws.cell(row=r, column=20).value,        # T 확정단가 (발주서용)
                     "기존단가": ws.cell(row=r, column=16).value,    # P 실적가 (견적요청서용)
                     "비고": str(ws.cell(row=r, column=27).value or "").strip()})
    return code, name, sets, rows, vina


def write_po(items, vendor, code, out_path, template=DEFAULT_TEMPLATE,
             date=None, due="", buyer=None):
    """협력사 한 곳의 발주서를 만든다. (품목 수·합계금액·단가 미확정 수) 를 돌려준다."""
    if not items:
        raise ValueError(f"협력사 「{vendor}」 의 산 줄이 없습니다.")
    wb = load_workbook(template)
    ws = wb[SHEET]
    total_row = _find_total_row(ws)
    spacer = total_row - 1
    slots = spacer - ITEM0                          # 표의 빈 칸 수

    n = len(items)
    if n > slots:                                   # 긴 발주서 — 줄을 늘린다
        k = n - slots
        # ⚠ openpyxl 은 줄을 끼워도 병합 칸을 안 밀어준다 (합계·추가요청 상자가
        #   품목 구역을 덮어 값이 가려짐 — 실측) → 끼우기 **전에** 병합부터 내린다.
        from openpyxl.worksheet.cell_range import CellRange
        for s in [str(m) for m in ws.merged_cells.ranges if m.min_row >= spacer]:
            ws.unmerge_cells(s)
            cr = CellRange(s)
            cr.shift(0, k)
            ws.merge_cells(str(cr))
        ws.insert_rows(spacer, k)
        for r in range(spacer, spacer + k):         # 새 줄 서식 = 첫 품목줄 서식
            for ci in range(1, 10):
                ws.cell(row=r, column=ci)._style = _copy(ws.cell(row=ITEM0, column=ci)._style)
            ws.cell(row=r, column=COLS["단위"]).value = "EA"
        for img in ws._images:                      # 도장 그림도 함께 내린다
            for pt in ("_from", "to"):
                anc = getattr(img.anchor, pt, None)
                if anc is not None and anc.row >= spacer - 1:
                    anc.row += k
        spacer += k
        total_row += k
        for r in range(ITEM0, spacer):              # 번호·금액 수식 재보장
            ws.cell(row=r, column=1).value = r - ITEM0 + 1
            ws.cell(row=r, column=COLS["금액"]).value = f"=E{r}*G{r}"
        ws.cell(row=total_row, column=COLS["금액"]).value = f"=SUM(H{ITEM0}:H{spacer})"

    # 실물 규칙(광원전기 발주서 실측): 「옛->새」 변경표기는 발주서에 **새 값만** 적는다
    def _clean(s):
        s = str(s or "")
        return s.split("->")[-1].strip() if "->" in s else s

    no_price = 0
    cleaned = 0
    total = 0
    for i, x in enumerate(items):
        r = ITEM0 + i
        cleaned += sum(1 for f in ("품명", "형번", "제조사") if "->" in str(x[f]))
        ws.cell(row=r, column=COLS["품명"]).value = _clean(x["품명"])
        ws.cell(row=r, column=COLS["규격"]).value = _clean(x["형번"])
        ws.cell(row=r, column=COLS["MAKER"]).value = _clean(x["제조사"])
        ws.cell(row=r, column=COLS["수량"]).value = x["수량"]
        ws.cell(row=r, column=COLS["단위"]).value = x["단위"]
        if x["단가"] in (None, "", 0):
            no_price += 1
        else:
            ws.cell(row=r, column=COLS["단가"]).value = x["단가"]
            try:
                total += float(x["수량"]) * float(x["단가"])
            except Exception:
                pass
        if x["비고"]:
            ws.cell(row=r, column=COLS["비고"]).value = x["비고"]

    ws["A4"] = _label_of(ws["A4"].value) + code
    ws["A5"] = _label_of(ws["A5"].value) + vendor
    ws["A8"] = _label_of(ws["A8"].value) + (due or "")
    ws["I4"] = date or datetime.date.today()
    if buyer:
        ws["I5"] = buyer
    wb.save(out_path)
    return {"품목": n, "합계": int(total), "단가미확정": no_price, "변경정리": cleaned}


def main(argv=None):
    ap = argparse.ArgumentParser(description="통일판 마스터 → 협력사별 발주서 생성")
    ap.add_argument("master", nargs="?", help="통일판 마스터 .xlsx")
    ap.add_argument("--vendor", help="협력사 이름 (한 곳)")
    ap.add_argument("--all", action="store_true", help="산 줄이 있는 모든 협력사")
    ap.add_argument("--code", default=None, help="관리번호 (기본: 마스터 제목에서)")
    ap.add_argument("--due", default="", help="요청 납기 (예: 2026-08-29)")
    ap.add_argument("--date", default=None, help="발주 일자 (기본: 오늘)")
    ap.add_argument("--buyer", default=None, help="발주 담당자 표기 바꿀 때만")
    ap.add_argument("--outdir", default=".")
    ap.add_argument("--template", default=DEFAULT_TEMPLATE)
    ap.add_argument("--make-template", metavar="실물발주서", help="실물 발주서에서 빈 양식 도출")
    a = ap.parse_args(argv)

    if a.make_template:
        out = DEFAULT_TEMPLATE
        make_po_template(a.make_template, out)
        print(f"발주서 빈 양식 도출: {out}")
        return 0
    if not a.master or (not a.vendor and not a.all):
        ap.error("마스터 파일과 --vendor 이름(또는 --all)이 필요합니다.")

    code, name, sets, rows, vina = read_master(a.master)
    code = a.code or code
    date = datetime.date.fromisoformat(a.date) if a.date else datetime.date.today()
    stamp = date.strftime("%Y%m%d")

    by_vendor = {}
    for x in rows:
        if x["협력사"]:
            by_vendor.setdefault(x["협력사"], []).append(x)
    targets = sorted(by_vendor) if a.all else [a.vendor]

    print("=" * 64)
    print(f"  {code} {name} — 발주서 생성 (대수 {sets})")
    print("=" * 64)
    if vina:
        print(f"  VINA 줄 {vina}개 제외 — 베트남 담당자 협의 대상 (발주서 아님)")
    made = 0
    for v in targets:
        items = by_vendor.get(v, [])
        if not items:
            print(f"  ⚠ {v}: 산 줄 없음 — 건너뜀")
            continue
        out = os.path.join(a.outdir, f"{code} {v} 발주서_{stamp}.xlsx")
        res = write_po(items, v, code, out, a.template, date, a.due, a.buyer)
        made += 1
        warn = f" · ⚠단가 미확정 {res['단가미확정']}줄" if res["단가미확정"] else ""
        cl = f" · 변경표기 정리 {res['변경정리']}건" if res["변경정리"] else ""
        print(f"  ✅ {v:<10} {res['품목']:>3}줄 · 합계 {res['합계']:>12,}원{warn}{cl} → {os.path.basename(out)}")
    print(f"\n  발주서 {made}건 생성 · 협력사 전화·메일 칸은 비워 둠(거래처 목록 연결 전 — 손으로 기입)")
    return 0 if made else 2


if __name__ == "__main__":
    sys.exit(main())
