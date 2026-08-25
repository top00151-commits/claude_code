#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
후보 ③ — 부품 단가 사전: 수불부 실적가를 마스터 기존단가에 자동 기입
====================================================================
대표 승인 "순서대로" · 규칙 = 허동준 매니저 확인(08-25 12:42):
  "기본적으로 가장 최근 발주 시의 값을 기입 · 매입처 이원화 시 같은 품목도 단가 차이"

무엇을 하나
  통일판 마스터의 기존단가(P) 빈칸을, 수불부(발주관리대장)의 실적가로 채운다.
  대조 열쇠: ① 형번+협력사 → ② 형번만(다른 매입처 실적 — 표시) · 항상 최근 발주일 우선.
  골든 실측: 001M 실물에서 이 규칙이 허동준 매니저의 손 기입과 100% 일치(117/117).

일부러 안 하는 것
  ⛔ 이미 적힌 기존단가는 절대 덮지 않는다 — 값이 다르면 「단가 변동 후보」로 보고만.
  ⛔ 못 찾은 품목은 지어내지 않는다 — 「신규(견적 대상)」 목록으로 보고.
  ⛔ VINA·비교·삭제·실패 줄은 대상 아님.

사용법
  python bom_fill_prices.py 마스터.xlsx 수불부.xlsx        # → 마스터이름_단가채움.xlsx
"""
import argparse
import datetime
import os
import sys

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

from openpyxl import load_workbook

DEAD_TAGS = ("비교", "삭제", "실패")


def _norm(s):
    return str(s or "").strip().split("->")[-1].strip().upper().replace(" ", "")


def _h(v):
    return str(v or "").strip().lower().replace(" ", "").replace("\n", "")


def read_ledger(path):
    """수불부 발주관리대장 → 두 사전: (형번,협력사)→(발주일,단가,협력사) · 형번→같은 것.
    칸은 이름으로 찾는다 (규격·협력사·발주일자·단가)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    if "발주관리대장" not in wb.sheetnames:
        raise ValueError("수불부 파일에 「발주관리대장」 시트가 없습니다.")
    ws = wb["발주관리대장"]
    col = {}
    for hr, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
        f = {}
        for ci, v in enumerate(row):
            h = _h(v)
            if not h:
                continue
            if "규격" in h and "형번" in h:
                f.setdefault("규격", ci)
            elif h == "협력사":
                f.setdefault("협력사", ci)
            elif "발주일자" in h:
                f.setdefault("발주일", ci)
            elif h == "단가":
                f.setdefault("단가", ci)
        if len(f) == 4:
            col = f
            head = hr
            break
    if not col:
        raise ValueError("수불부 발주관리대장에서 규격·협력사·발주일자·단가 칸을 찾지 못했습니다.")

    by_vk, by_k, n = {}, {}, 0
    for row in ws.iter_rows(min_row=head + 1, values_only=True):
        spec = _norm(row[col["규격"]] if len(row) > col["규격"] else "")
        if not spec:
            continue
        price = row[col["단가"]] if len(row) > col["단가"] else None
        if price in (None, "", 0):
            continue
        ven = str((row[col["협력사"]] if len(row) > col["협력사"] else "") or "").strip()
        d = row[col["발주일"]] if len(row) > col["발주일"] else None
        if not isinstance(d, datetime.datetime):
            d = datetime.datetime.min
        n += 1
        for dic, key in ((by_vk, (spec, ven)), (by_k, spec)):
            old = dic.get(key)
            if old is None or d >= old[0]:
                dic[key] = (d, price, ven)
    return by_vk, by_k, n


def fill_prices(master_path, ledger_path, out_path):
    by_vk, by_k, n_led = read_ledger(ledger_path)
    wb = load_workbook(master_path)
    ws = wb.active
    filled, other_vendor, missing, changed, refs = [], [], [], [], []
    for r in range(9, (ws.max_row or 9) + 1):
        pn = str(ws.cell(row=r, column=5).value or "").strip()
        spec_raw = str(ws.cell(row=r, column=6).value or "").strip()
        if not pn and not spec_raw:
            continue
        if str(ws.cell(row=r, column=2).value or "").strip() in DEAD_TAGS:
            continue
        ven = str(ws.cell(row=r, column=8).value or "").strip()
        if ven.upper() == "VINA":
            continue
        spec = _norm(spec_raw)
        if not ven:
            # 협력사 미정(뼈대 단계) — 채우지 않고 과거 실적을 참고로만 보여준다.
            #   협력사 확정은 구매팀 비교견적 몫 (인터뷰 확정) — 판단을 대신하지 않는다.
            got = by_k.get(spec)
            if got is not None and ws.cell(row=r, column=16).value in (None, "", 0):
                refs.append((r, spec_raw[:30], got[1], got[2]))
            continue
        got = by_vk.get((spec, ven))
        src = "협력사일치"
        if got is None:
            got = by_k.get(spec)
            src = "타매입처"
        cur = ws.cell(row=r, column=16).value          # P 기존단가
        if cur not in (None, "", 0):
            if got is not None and float(got[1]) != float(cur):
                changed.append((r, spec_raw[:30], cur, got[1], got[2]))
            continue                                    # ⛔ 이미 적힌 값은 안 덮는다
        if got is None:
            missing.append((r, ven, spec_raw[:30]))
            continue
        ws.cell(row=r, column=16).value = got[1]
        (filled if src == "협력사일치" else other_vendor).append((r, spec_raw[:30], got[1], got[2]))
    wb.save(out_path)
    return {"수불부항목": n_led, "채움": len(filled), "타매입처": len(other_vendor),
            "신규": missing, "변동후보": changed, "타매입처목록": other_vendor,
            "협력사미정참고": refs}


def main(argv=None):
    ap = argparse.ArgumentParser(description="수불부 실적가 → 마스터 기존단가 자동 기입")
    ap.add_argument("master", help="통일판 마스터 .xlsx")
    ap.add_argument("ledger", help="수불부(발주관리대장) .xlsx")
    ap.add_argument("--out", default=None)
    a = ap.parse_args(argv)
    out = a.out or os.path.splitext(a.master)[0] + "_단가채움.xlsx"
    res = fill_prices(a.master, a.ledger, out)

    print("=" * 64)
    print(f"  기존단가 자동 기입 — 수불부 실적 {res['수불부항목']:,}건 대조")
    print("=" * 64)
    print(f"  ✅ 채움 {res['채움']}줄 (형번+협력사·최근 발주일 — 허동준 매니저 규칙)")
    if res["타매입처"]:
        print(f"  ⚠ 다른 매입처 실적으로 채움 {res['타매입처']}줄 — 매입처 이원화 확인 필요:")
        for r, s, p, v in res["타매입처목록"][:8]:
            print(f"     r{r} {s} = {p:,}원 ({v} 실적)")
    if res["변동후보"]:
        print(f"  ⚠ 단가 변동 후보 {len(res['변동후보'])}줄 — 적힌 값과 최근 실적이 다름(덮지 않음):")
        for r, s, cur, new, v in res["변동후보"][:8]:
            print(f"     r{r} {s}: 적힌 값 {cur:,} vs 최근 실적 {new:,} ({v})")
    if res["협력사미정참고"]:
        print(f"  📌 협력사 미정 줄의 과거 실적 참고 {len(res['협력사미정참고'])}건 — 비교견적 출발점(채우지는 않음):")
        for r, s, p, v in res["협력사미정참고"][:10]:
            print(f"     r{r} {s} = {p:,}원 ({v} 실적)")
    if res["신규"]:
        print(f"  ✍ 실적 없음(견적 대상) {len(res['신규'])}줄:")
        for r, v, s in res["신규"][:10]:
            print(f"     r{r} [{v}] {s}")
    print(f"\n  저장: {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
