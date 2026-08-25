#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
후보 ①-a — 인벤터 유닛 BOM → 구매품 PART LIST(간이판) 변환기
================================================================
대표 승인 2026-08-24 "순서대로 진행해보자.. 하나 완료 후 꼭 검증을 직접 실행해서 해봐."

무엇을 하나
  설계팀이 인벤터에서 유닛(SUB Ass'y)별로 내린 BOM 엑셀들(AA00.xlsx …)을 받아
  ① 한 표로 합치고 ② 구분='구매품' 줄만 남기고 ③ 간이판(000M0000 양식)에 채워
  「구매품 PART LIST 초안」을 만든다.

일부러 안 하는 것 (인터뷰로 확정된 규칙)
  ⛔ 중복 형번 합산 — 유닛별 포장 구분 때문에 줄대로 둔다 (허동준 매니저).
  ⛔ 가공품·용접·프로파일 줄을 조용히 버리기 — 몇 줄을 왜 뺐는지 전부 보고한다.
  ⛔ 빈 품명 지어내기 — 사람 몫 목록으로 알려만 준다 (한재운 매니저: 케이블류 등은 수기).

사람 몫으로 남기는 것 (보고서에 목록으로)
  · 품명이 빈 줄 (인벤터에 DESCRIPTION 없던 부품)
  · 서보모터 줄 = 파란 표시 (케이블을 전장팀과 협의 — 한재운 매니저 규칙)

사용법
  python bom_inventor_to_partlist.py AA00.xlsx AB00.xlsx … --code 005M2606 \
      --name "BUDS 단차 검사기" --author 한재운 --out 산출.xlsx
  (폴더를 주면 그 안의 .xlsx 전부 — 단, 이름에 '정리'가 든 파일은 겹치니 제외)
"""
import argparse
import os
import sys
from copy import copy as _copy

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
# 간이판 양식 — 설계팀 실물(BUDS 키트)에서 자료를 비워 도출한 동봉본
DEFAULT_TEMPLATE = os.path.join(HERE, "양식", "양식_간이판.xlsx")

BUY = "구매품"                      # 간이판으로 옮기는 유일한 구분 값
BLUE = PatternFill("solid", start_color="FF00B0F0", end_color="FF00B0F0")
SERVO_WORDS = ("SERVO", "서보")     # 파란 표시(전장 협의) 후보 — 한재운 매니저 예시 기준


def _norm_header(v):
    return str(v or "").strip().lower().replace(" ", "").replace("\n", "")


# 인벤터 내보내기 칸은 담당자마다 순서가 다를 수 있어(한재운 매니저 확인)
# 자리(열 위치)가 아니라 **칸 이름**으로 찾는다.
ALIASES = {
    "구분":   {"구분"},
    "주제":   {"주제"},
    "code":  {"code"},
    "품명":   {"description", "품명"},
    "형번":   {"부품번호", "productcode", "코드명"},
    "제조사":  {"maker", "제조사", "manufacturer"},
    "협력사":  {"vendor", "협력사", "외주사"},
    "수량":   {"수량"},
    "수량예비": {"q'ty", "qty"},
    "단위":   {"단위", "unit"},
}
REQUIRED = ("구분", "주제", "code", "품명", "형번")


def _find_columns(ws):
    """머리글 줄을 찾아 {우리이름: 열번호} 를 돌려준다. 못 찾으면 (None, 무엇이 없는지)."""
    for hr in range(1, 11):
        found = {}
        for ci in range(1, (ws.max_column or 0) + 1):
            h = _norm_header(ws.cell(row=hr, column=ci).value)
            if not h:
                continue
            for key, names in ALIASES.items():
                if h in names and key not in found:
                    found[key] = ci
        if all(k in found for k in ("구분", "형번")):
            missing = [k for k in REQUIRED if k not in found]
            if "수량" not in found and "수량예비" not in found:
                missing.append("수량")
            return (hr, found, missing)
    return (None, {}, list(REQUIRED))


def _qty(v):
    if v in (None, ""):
        return None
    try:
        f = float(str(v).strip())
        return int(f) if f == int(f) else f
    except Exception:
        return None


def read_inventor_files(paths):
    """유닛 BOM 파일들 → (구매품 줄 목록, 제외 집계, 파일별 통계). 규칙 밖이면 ValueError."""
    files = []
    for p in paths:
        if os.path.isdir(p):
            for fn in sorted(os.listdir(p)):
                if fn.lower().endswith(".xlsx") and "정리" not in fn and not fn.startswith("~$"):
                    files.append(os.path.join(p, fn))
        else:
            files.append(p)
    if not files:
        raise ValueError("읽을 인벤터 BOM 파일이 없습니다.")

    rows, excluded, per_file = [], {}, []
    for f in files:
        wb = load_workbook(f, data_only=True)
        ws = wb["BOM"] if "BOM" in wb.sheetnames else wb.active
        hr, col, missing = _find_columns(ws)
        if hr is None or missing:
            raise ValueError(
                f"[{os.path.basename(f)}] 인벤터 BOM 머리글에서 찾지 못한 칸: {', '.join(missing)}\n"
                f"  → 내보내기 설정에 이 칸들이 있는지 확인해 주십시오 (칸 순서는 상관없습니다).")

        def get(r, key):
            ci = col.get(key)
            return ws.cell(row=r, column=ci).value if ci else None

        n_buy = n_skip = 0
        for r in range(hr + 1, (ws.max_row or hr) + 1):
            name = str(get(r, "품명") or "").strip()
            spec = str(get(r, "형번") or "").strip()
            if not name and not spec:
                continue
            gubun = str(get(r, "구분") or "").strip()
            if gubun != BUY:
                excluded[gubun or "(구분 빈칸)"] = excluded.get(gubun or "(구분 빈칸)", 0) + 1
                n_skip += 1
                continue
            qty = _qty(get(r, "수량"))
            if qty is None:
                qty = _qty(get(r, "수량예비"))
            rows.append({
                "구분": str(get(r, "주제") or "").strip(),
                "CODE": str(get(r, "code") or "").strip(),
                "품명": name,
                "형번": spec,
                "제조사": str(get(r, "제조사") or "").strip(),
                "협력사": str(get(r, "협력사") or "").strip(),
                "수량": qty,
                "단위": str(get(r, "단위") or "").strip(),
                "원본": f"{os.path.basename(f)} r{r}",
            })
            n_buy += 1
        per_file.append((os.path.basename(f), n_buy, n_skip))
    return rows, excluded, per_file


def write_partlist(rows, code, name, author, out_path, template=DEFAULT_TEMPLATE):
    """간이판 양식 사본에 구매품 줄을 채운다. 산출 줄 수를 돌려준다."""
    wb = load_workbook(template)
    ws = wb.active
    ROW0, C_FROM, C_TO = 8, 2, 16          # 자료 시작줄 · B~P

    base_style = {ci: _copy(ws.cell(row=ROW0, column=ci)._style) for ci in range(C_FROM, C_TO + 1)}
    for r in range(ROW0, (ws.max_row or ROW0) + 1):
        for ci in range(C_FROM, C_TO + 1):
            cell = ws.cell(row=r, column=ci)
            cell.value = None
            cell._style = _copy(base_style[ci])

    ws["C2"] = f"{code} 구매품 BOM\nPROJECT : {name}"
    ws["C5"] = f"AUTHOR : {author}"

    servo_rows, empty_name_rows = [], []
    r = ROW0
    for i, x in enumerate(rows, start=1):
        ws.cell(row=r, column=2, value=i)
        ws.cell(row=r, column=3, value=x["구분"])
        ws.cell(row=r, column=4, value=x["CODE"])
        ws.cell(row=r, column=5, value=x["품명"] if x["품명"] else None)
        ws.cell(row=r, column=6, value=x["형번"])
        ws.cell(row=r, column=7, value=x["제조사"] or None)
        ws.cell(row=r, column=8, value=x["협력사"] or None)
        ws.cell(row=r, column=9, value=x["수량"])
        ws.cell(row=r, column=10, value=x["단위"] or None)
        if any(w in x["품명"].upper() for w in SERVO_WORDS):
            for ci in range(2, 6):
                ws.cell(row=r, column=ci).fill = BLUE
            servo_rows.append((i, x["품명"], x["형번"]))
        if not x["품명"]:
            empty_name_rows.append((i, x["형번"], x["원본"]))
        r += 1

    # 관행 그대로: 마지막에 「전장 구매품 별도」 자리표시 줄 (파란 표시)
    ws.cell(row=r, column=2, value=len(rows) + 1)
    ws.cell(row=r, column=3, value="전장 구매품 별도")
    for ci in range(2, 6):
        ws.cell(row=r, column=ci).fill = BLUE

    wb.save(out_path)
    return {"품목": len(rows), "서보표시": servo_rows, "품명빈칸": empty_name_rows}


def main(argv=None):
    ap = argparse.ArgumentParser(description="인벤터 유닛 BOM → 구매품 PART LIST(간이판) 초안")
    ap.add_argument("inputs", nargs="+", help="인벤터 BOM .xlsx 파일들 또는 폴더")
    ap.add_argument("--code", required=True, help="관리번호 (예: 005M2606)")
    ap.add_argument("--name", required=True, help="프로젝트 이름")
    ap.add_argument("--author", default="설계담당자", help="작성자 이름")
    ap.add_argument("--out", default=None, help="산출 파일 경로 (.xlsx)")
    ap.add_argument("--template", default=DEFAULT_TEMPLATE, help="간이판 양식 파일")
    a = ap.parse_args(argv)

    out = a.out or f"{a.code} 구매품 PART LIST_초안.xlsx"
    rows, excluded, per_file = read_inventor_files(a.inputs)
    if not rows:
        print("구매품 줄이 한 줄도 없습니다. 인벤터 구분 칸 값을 확인해 주십시오.")
        return 2
    res = write_partlist(rows, a.code, a.name, a.author, out, a.template)

    print("=" * 64)
    print(f"  {a.code} 구매품 PART LIST 초안 — {res['품목']}줄")
    print("=" * 64)
    for fn, nb, ns in per_file:
        print(f"  읽음 {fn:<18} 구매품 {nb:>3}줄 · 제외 {ns:>3}줄")
    print(f"  제외 내역(간이판에 안 옮긴 줄): "
          + (", ".join(f"{k} {v}줄" for k, v in sorted(excluded.items())) or "없음"))
    if res["품명빈칸"]:
        print(f"\n  ✍ 사람 몫 — 품명이 빈 줄 {len(res['품명빈칸'])}건 (인벤터에 이름이 없던 부품):")
        for no, spec, src in res["품명빈칸"]:
            print(f"     {no:>3}번  {spec[:40]:<40}  ({src})")
    if res["서보표시"]:
        print(f"\n  🔵 파란 표시 {len(res['서보표시'])}건 — 케이블을 전장팀과 협의 (한재운 매니저 규칙):")
        for no, nm, spec in res["서보표시"]:
            print(f"     {no:>3}번  {nm[:22]:<22} {spec[:30]}")
    print(f"\n  저장: {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
