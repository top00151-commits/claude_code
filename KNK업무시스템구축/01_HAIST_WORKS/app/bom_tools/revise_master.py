#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
①-c — 마스터 개정: 설계 변경(추가·수량변경·형번변경·삭제)을 옛 마스터에 증분 반영
==============================================================================
대표 승인 2026-08-26 "진행해.. 검증도 하고" — 배경 문답: "①-a에서 부품이 추가·변경·삭제되면
다시 ①-a에 업로드하면 되는건가?" → 손 기입(협력사·단가·납기) 이후엔 통째 재생성이 파괴적이라
바뀐 유닛 BOM만 받아 마스터 위에 증분으로 반영한다.

실물 근거 (전부 실측 — 008M·001M B열 전수 확인 2026-08-26)
  · 유닛 구분 = D열 CODE (인벤터 파일당 하나: AA00…, 전장 BOM은 「전장」)
    → **새로 올린 유닛만 대조**하고 나머지 유닛(전장 포함)은 한 칸도 건드리지 않는다.
  · 삭제 = 줄을 지우지 않고 B열 「삭제」 + 확정단가 T=0 (수량·기존단가·이력 보존,
    금액·합계는 자동 0) — 008M 삭제줄 7건 전수 T=0·V=0 실측.
  · 추가 = B열 「추가」 줄 (실물 관행 그대로) + 실물 수식(K·O·U·V·W).
  · 형번 변경 = 같은 유닛·같은 품명·같은 제조사 1:1 쌍이면 F열 「옛->새」 (실물 표기).
    확신이 없으면(쌍이 여럿) 추가+삭제로 남기고 후보로 보고만 한다.

일부러 안 하는 것
  ⛔ B열에 낱말이 이미 있는 줄은 덮지 않는다 (비교·실패·VINA… — 보고만).
  ⛔ 단가·납기를 지어 넣지 않는다 — 추가 줄 단가는 ③(수불부)과 구매팀 몫.
  ⛔ 올리지 않은 유닛 줄은 값·수식·서식 무접촉.

사용법
  python revise_master.py 옛마스터.xlsx AA00.xlsx … --out 개정본.xlsx
"""
import argparse
import os
import sys
from copy import copy as _copy

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as _gl

try:
    from . import inventor_to_partlist as _inv
except ImportError:                                      # CLI 직접 실행용
    import inventor_to_partlist as _inv

ROW0 = 9                       # 통일판 자료 시작줄
C_FROM, C_TO = 2, 27           # B~AA
SUBTOTAL_COLS = (10, 11, 13, 14, 15, 16, 18, 20, 21, 22, 23)
MARK_DEL, MARK_ADD = "삭제", "추가"

# 실물 품목줄 수식 (write_master 와 동일 — 001M 골든으로 검증된 패턴)
FORMULA_COLS = {
    11: lambda r: f"=J{r}*$J$5",             # K TOTAL
    15: lambda r: f"=$K{r}-($M{r}+$N{r})",   # O 발주수량
    21: lambda r: f"=T{r}*(M{r}+N{r})",      # U 재고금액
    22: lambda r: f"=O{r}*T{r}",             # V 발주금액
    23: lambda r: f"=U{r}+V{r}",             # W 합계
}


def _norm_spec(v):
    """형번 대조 정규화 — 「옛->새」는 새 형번 기준 (실물 규칙과 동일)."""
    return str(v or "").strip().split("->")[-1].strip().upper().replace(" ", "")


def _norm_txt(v):
    return str(v or "").strip().upper().replace(" ", "")


def _find_total_row(ws):
    for r in range(ROW0, (ws.max_row or ROW0) + 1):
        if str(ws.cell(row=r, column=10).value or "").startswith("=SUBTOTAL"):
            return r
    return None


def _scan_items(ws, total_row):
    """품목줄(E나 F에 값이 있는 줄)을 위에서부터 모은다."""
    items = []
    for r in range(ROW0, total_row):
        e = str(ws.cell(row=r, column=5).value or "").strip()
        f = str(ws.cell(row=r, column=6).value or "").strip()
        if not e and not f:
            continue
        items.append({
            "r": r,
            "B": str(ws.cell(row=r, column=2).value or "").strip(),
            "D": str(ws.cell(row=r, column=4).value or "").strip(),
            "E": e, "F": f,
            "G": str(ws.cell(row=r, column=7).value or "").strip(),
            "J": ws.cell(row=r, column=10).value,
            "T": ws.cell(row=r, column=20).value,
        })
    return items


def _shift_merges(ws, at_row, n):
    """at_row 이후로 걸린 병합 칸을 n줄 아래로 민다 — insert_rows 전에 (실물 발주서 교훈)."""
    moves = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= at_row:
            moves.append((str(rng), rng.min_col, rng.min_row + n, rng.max_col, rng.max_row + n))
        elif rng.max_row >= at_row:                      # 걸침 — 아래로 늘림
            moves.append((str(rng), rng.min_col, rng.min_row, rng.max_col, rng.max_row + n))
    for old, c1, r1, c2, r2 in moves:
        ws.unmerge_cells(old)
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def revise(master_path, inventor_paths, out_path):
    """옛 마스터 + 새 인벤터 유닛 BOM들 → 개정 마스터 + 보고. 규칙 밖이면 ValueError."""
    new_rows, excluded, per_file = _inv.read_inventor_files(inventor_paths)
    if not new_rows:
        raise ValueError("새 인벤터 BOM에 구매품 줄이 한 줄도 없습니다.")
    units = sorted({_norm_txt(x["CODE"]) for x in new_rows if _norm_txt(x["CODE"])})
    if not units:
        raise ValueError("새 인벤터 BOM의 code(SUB Ass'y) 칸이 비어 있어 유닛을 알 수 없습니다.")

    wb = load_workbook(master_path)                      # 수식·서식 보존 열기
    ws = wb.active
    total_row = _find_total_row(ws)
    if total_row is None:
        raise ValueError("마스터에서 SUBTOTAL 합계줄을 찾지 못했습니다 — 통일판 양식이 맞는지 확인해 주십시오.")
    items = _scan_items(ws, total_row)
    if not items:
        raise ValueError("마스터에 품목줄이 없습니다.")

    # 유닛의 「본 블록」 = 같은 D가 연속으로 이어진 가장 긴 구간만 대조한다.
    # 실물 근거: 001M 비교견적 블록(D=AA01~03이 본 블록 밖 r146+에 산재)·BUDS AB00 파일 속
    # code=AA00 교차 줄 — 블록 밖 같은 유닛 줄은 파일 유래가 달라 무접촉+보고가 맞다.
    runs, cur, prev = {}, [], None
    for x in items:
        u = _norm_txt(x["D"])
        if u == prev:
            cur.append(x)
        else:
            if prev is not None:
                runs.setdefault(prev, []).append(cur)
            prev, cur = u, [x]
    if prev is not None:
        runs.setdefault(prev, []).append(cur)
    in_scope, outside, main_run = [], [], {}
    for u in units:
        rs = runs.get(u, [])
        if not rs:
            continue
        main = max(rs, key=len)
        main_run[u] = main
        in_scope += main
        outside += [x for r_ in rs if r_ is not main for x in r_]
    untouched_units = sorted({x["D"] for x in items if x["D"] and _norm_txt(x["D"]) not in units})

    # ── 대조: (유닛, 형번) 키 — 중복 형번은 줄 순서대로 짝지음 (합산 금지 규칙 그대로) ──
    def key_old(x):
        return (_norm_txt(x["D"]), _norm_spec(x["F"]))

    def key_new(x):
        return (_norm_txt(x["CODE"]), _norm_spec(x["형번"]))

    pool = {}
    for x in in_scope:
        pool.setdefault(key_old(x), []).append(x)
    matched, adds = [], []
    for x in new_rows:
        cand = pool.get(key_new(x))
        if cand:
            matched.append((cand.pop(0), x))
        else:
            adds.append(x)
    deletes = [x for lst in pool.values() for x in lst]

    # ── 형번 변경: 같은 유닛·같은 품명·같은 제조사 1:1 쌍만 「옛->새」 (실물 표기) ──
    renames, spec_suggest = [], []
    for unit in units:
        old_u = [x for x in deletes if _norm_txt(x["D"]) == unit]
        new_u = [x for x in adds if _norm_txt(x["CODE"]) == unit]
        for od in list(old_u):
            same = [nw for nw in new_u
                    if _norm_txt(nw["품명"]) == _norm_txt(od["E"])
                    and _norm_txt(nw["제조사"]) == _norm_txt(od["G"])]
            if len(same) == 1 and od["B"] == "":
                nw = same[0]
                renames.append((od, nw))
                deletes.remove(od); old_u.remove(od)
                adds.remove(nw); new_u.remove(nw)
            elif same:
                spec_suggest.append((od["F"], [nw["형번"] for nw in same]))

    report = {"추가": [], "삭제표시": [], "이미표시": [], "수량변경": [],
              "형번개정": [], "형번후보": spec_suggest, "T내림": [],
              "안건드린유닛": untouched_units, "유닛": units,
              "블록밖": [(x["r"], x["D"], x["F"]) for x in outside],
              "제외": excluded, "파일": per_file, "수식보정": 0}

    # ── 제자리 반영 (줄 번호가 아직 안 변한 지금): 수량변경·형번개정·삭제표시 ──
    def _qty_writable(od):                               # J가 수식인 줄은 건드리지 않음
        return not (isinstance(od["J"], str) and od["J"].startswith("="))

    for od, nw in matched:
        if nw["수량"] is not None and od["J"] != nw["수량"] and _qty_writable(od):
            ws.cell(row=od["r"], column=10, value=nw["수량"])
            report["수량변경"].append((od["r"], od["F"], od["J"], nw["수량"]))
    for od, nw in renames:
        new_f = f"{od['F']}->{nw['형번']}"
        ws.cell(row=od["r"], column=6, value=new_f)
        if nw["수량"] is not None and od["J"] != nw["수량"] and _qty_writable(od):
            ws.cell(row=od["r"], column=10, value=nw["수량"])
        report["형번개정"].append((od["r"], od["F"], nw["형번"]))
    for od in deletes:
        if od["B"]:
            report["이미표시"].append((od["r"], od["B"], od["F"]))
            continue
        ws.cell(row=od["r"], column=2, value=MARK_DEL)
        tv = od["T"]
        if isinstance(tv, (int, float)) and tv != 0:
            ws.cell(row=od["r"], column=20, value=0)
            report["T내림"].append((od["r"], od["F"], tv))
        report["삭제표시"].append((od["r"], od["F"]))

    # ── 추가 줄 삽입 — 유닛 블록 끝에, 아래쪽 위치부터 (줄 번호 보존) ──
    last_item = max(x["r"] for x in items)
    pos_of_unit = {u: run[-1]["r"] for u, run in main_run.items()}   # 본 블록 끝에 삽입
    groups = {}
    for nw in adds:
        u = _norm_txt(nw["CODE"])
        pos = pos_of_unit.get(u, last_item) + 1
        groups.setdefault(pos, []).append(nw)

    unit_defaulted = 0
    for pos in sorted(groups, reverse=True):
        batch = groups[pos]
        n = len(batch)
        _shift_merges(ws, pos, n)
        ws.insert_rows(pos, n)
        style_src = pos - 1                               # 바로 위 품목줄 서식을 물려받음
        for i, nw in enumerate(batch):
            r = pos + i
            for ci in range(C_FROM, C_TO + 1):
                c = ws.cell(row=r, column=ci)
                c._style = _copy(ws.cell(row=style_src, column=ci)._style)
                c.value = None
            ws.cell(row=r, column=2, value=MARK_ADD)
            ws.cell(row=r, column=3, value=nw["구분"] or None)
            ws.cell(row=r, column=4, value=nw["CODE"] or None)
            ws.cell(row=r, column=5, value=nw["품명"] or None)
            ws.cell(row=r, column=6, value=nw["형번"] or None)
            ws.cell(row=r, column=7, value=nw["제조사"] or None)
            ws.cell(row=r, column=8, value=nw["협력사"] or None)
            ws.cell(row=r, column=10, value=nw["수량"])
            if nw["단위"]:
                ws.cell(row=r, column=12, value=nw["단위"])
            else:
                ws.cell(row=r, column=12, value="EA")
                unit_defaulted += 1
            report["추가"].append((r, nw["CODE"], nw["형번"]))
    report["단위기본값"] = unit_defaulted

    # ── 수식 재작성 — 삽입으로 밀린 줄의 자기줄 참조 교정 + 새 줄 수식 심기 ──
    total_row = _find_total_row(ws)
    added_rows = {r for r, _u, _f in report["추가"]}
    for r in range(ROW0, total_row):
        e = str(ws.cell(row=r, column=5).value or "").strip()
        f = str(ws.cell(row=r, column=6).value or "").strip()
        is_item = bool(e or f)
        for ci, mk in FORMULA_COLS.items():
            c = ws.cell(row=r, column=ci)
            want = mk(r)
            if r in added_rows and is_item:
                if c.value != want:
                    c.value = want
                    report["수식보정"] += 1
            elif isinstance(c.value, str) and c.value.startswith("=") and c.value != want:
                c.value = want                            # 밀린 줄 — 값 손기입 칸은 안 건드림
                report["수식보정"] += 1
    for ci in SUBTOTAL_COLS:                              # 합계줄 범위 교정
        c = ws.cell(row=total_row, column=ci)
        if str(c.value or "").startswith("=SUBTOTAL"):
            L = _gl(ci)
            c.value = f"=SUBTOTAL(9,{L}{ROW0}:{L}{total_row - 1})"

    wb.save(out_path)
    return report


def main(argv=None):
    ap = argparse.ArgumentParser(description="마스터 개정 — 새 인벤터 유닛 BOM 증분 반영")
    ap.add_argument("master", help="옛 마스터(통일판) .xlsx")
    ap.add_argument("inventor", nargs="+", help="바뀐 유닛의 인벤터 BOM .xlsx 들")
    ap.add_argument("--out", default=None)
    a = ap.parse_args(argv)
    out = a.out or os.path.splitext(a.master)[0] + "_개정.xlsx"
    rep = revise(a.master, a.inventor, out)

    print("=" * 64)
    print(f"  마스터 개정 — 대조 유닛: {', '.join(rep['유닛'])}")
    print("=" * 64)
    for t, lst in (("추가", rep["추가"]), ("삭제표시", rep["삭제표시"]),
                   ("수량변경", rep["수량변경"]), ("형번개정", rep["형번개정"])):
        print(f"  {t} {len(lst)}건" + (":" if lst else ""))
        for x in lst[:10]:
            print("    ", x)
    if rep["T내림"]:
        print(f"  확정단가 0 처리(삭제줄 금액 제외 — 실물 관행): {rep['T내림']}")
    if rep["이미표시"]:
        print(f"  B열 낱말 있어 안 건드린 삭제 후보: {rep['이미표시']}")
    if rep["형번후보"]:
        print(f"  ⚠ 형번 변경 후보(자동 기입 안 함): {rep['형번후보']}")
    if rep["블록밖"]:
        print(f"  📌 본 블록 밖 같은 유닛 줄(무접촉 — 손 확인): {rep['블록밖']}")
    print(f"  안 건드린 유닛: {', '.join(rep['안건드린유닛']) or '없음'} · 수식보정 {rep['수식보정']}칸")
    print(f"\n  저장: {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
