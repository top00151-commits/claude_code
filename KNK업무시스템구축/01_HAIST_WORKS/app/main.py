"""
KNK 일일업무일지 v2 - Phase 1 MVP
Task Card 기반 일일업무 + 팀장 뷰 + 경영진 대시보드 + 개인 히스토리
"""
from fastapi import FastAPI, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
import os, io, calendar, tempfile
from datetime import datetime, timedelta, date
from .i18n import LANGS, t as i18n_t, get_all_translations
from . import menu_catalog as _menu  # Phase 1 (2026-04-29): M-코드 카탈로그
from .database import (db_session, init_db, seed_all, seed_sample_tasks,
                        seed_business_data, seed_recent_tasks_topup, hash_pw,
                        parse_mgmt_xls, parse_mgmt_csv, import_mgmt_rows,
                        regenerate_user_passwords, build_password_csv,
                        add_comment, get_task_comments, delete_comment,
                        get_notifications, count_unread, mark_notification_read,
                        mark_all_read, notify_user,
                        log_activity, log_activity_standalone, get_activities,
                        add_reaction, get_reactions, get_reactions_bulk, get_meta_bulk,
                        notify_status_change, get_user_search,
                        upsert_retro, get_retro, search_all, detect_bottlenecks,
                        delegate_task, get_delegations, resolve_delegation,
                        get_setting, get_settings_all, set_setting,
                        global_search, GLOBAL_SEARCH_CATEGORIES,
                        search_orders, search_customers, search_parts,
                        search_issues, search_tickets, search_users,
                        search_boards, search_exports, search_audits,
                        ceo_dashboard_kpis)

# ═══════════════════════════════════════════════════════════════════════════════
#  HAIST WORKS — main.py 목차 (TABLE OF CONTENTS)
#  v1.0 (2026-04-29 대표 결재 옵션 A — 코드 정리 Phase 1)
# ═══════════════════════════════════════════════════════════════════════════════
#
#  PART I. 기반 (Core)
#    §1.  IMPORTS                                              L 1
#    §2.  APP 초기화                                            L 32
#    §3.  STARTUP + 일일 미작성자 알림 스케줄러                  L 40
#    §4.  Phase 1 — 빅터 자연어 라우팅 + 메뉴 도움말 API          L 220
#    §5.  HELPERS (ctx, get_user, role_home...)                 L 310
#    §6.  Phase 1 — 메뉴 식별번호 헬퍼                           L 395
#
#  PART II. 인증·진입점
#    §7.  AUTH (login/logout)                                  L 624
#    §8.  ROOT (/)                                             L 658
#
#  PART III. 일상 업무 (전 직원)
#    §9.  HOME — 직관적 단일 페이지                              L 669
#    §10. DAILY — 일일 업무카드                                 L 904
#    §11. SUMMARY — 통합 요약                                   L 1137
#    §12. HISTORY — 개인 히스토리                                L 1851
#    §13. CALENDAR — 월간 뷰                                    L 2899
#    §14. FEED — 부서간 피드                                    L 2974
#
#  PART IV. 통신·협업
#    §15. COMMENTS API                                         L 1303
#    §16. TASK DETAIL API                                      L 1337
#    §17. REACTIONS API                                        L 1365
#    §18. 번역 API + set-lang                                   L 1380
#    §19. 업무 위임 API                                         L 1462
#    §20. 멘션 자동완성 API                                     L 1490
#    §21. SIDEBAR TREE API                                     L 1501
#    §22. ACTIVITIES API                                       L 1550
#    §23. NOTIFICATIONS API                                    L 1784
#    §24. CHANGES INFORM (변경 공지)                            L 3926
#    §25. TICKETS (요청 티켓)                                   L 4237
#    §26. ISSUES (이슈·AS DB)                                   L 4548
#    §27. BOARD (게시판)                                        L 4723
#
#  PART V. 분석·보고
#    §28. SEARCH — 통합 검색                                    L 1583
#    §29. RETRO                                                L 1674
#    §30. COCKPIT (팀장/CEO 라이브)                              L 1694
#    §31. BOTTLENECKS                                          L 1770
#    §32. TEAM DASHBOARD                                       L 1904
#    §33. CEO DASHBOARD                                        L 2176
#    §34. WEEKLY                                               L 2368
#
#  PART VI. 프로젝트
#    §35. PROJECT DETAIL                                       L 2749
#    §36. PROGRESS DASHBOARD                                   L 4150
#    §37. PROGRESS GANTT/BURNDOWN                              L 8713
#
#  PART VII. 매출영업 (M-01)
#    §38. CUSTOMER LIST                                        L 2825
#    §39. CUSTOMER DETAIL                                      L 2853
#    §40. SALES S1 가드 / 라이프사이클                           L 7161
#    §41. COMPANY INFO (견적서 헤더)                             L 7246
#    §42. SALES QUOTATIONS                                     L 7273
#    §43. SALES DASHBOARD / FORECAST                           L 7707
#    §44. OUTSTANDING / RECEIPTS                               L 7914
#
#  PART VIII. 수출입 (M-01-10/11)
#    §45. EXPORT P11                                           L 8122
#    §46. EXPORT PRINT                                         L 8555
#    §47. FTA CERTIFICATES                                     L 9894
#
#  PART IX. 자재구매 (M-02)
#    §48. LOGISTICS HUB                                        L 4952
#    §49. SUPPLIERS                                            L 5269
#    §50. PURCHASE ORDERS                                      L 5368
#    §51. STOCK MOVEMENTS                                      L 5539
#    §52. STOCK S2 가드 / 입출고                                L 6870
#    §53. STOCK AUDIT                                          L 5885
#    §54. STOCK AUDIT v2 (첨부파일·close)                        L 6039
#    §55. SAFETY STOCK / REORDER                               L 9667
#    §56. RATES                                                L 5574
#    §57. RATES STRENGTHENING (단가시뮬·알림)                    L 9455
#    §58. PRICES BY DATE                                       L 5621
#    §59. FX RATES                                             L 9740
#
#  PART X. 품질관리 (QMS)
#    §60. QMS                                                  L 9006
#    §61. QMS CAPA / PARETO                                    L 9218
#    §62. QC INSPECTION REPORTS                                L 10102
#
#  PART XI. 생산
#    §63. WORK ORDERS                                          L 10319
#
#  PART XII. 관리·설정 (M-04)
#    §64. ADMIN                                                L 3097
#    §65. MGMT CODE IMPORT                                     L 3194
#    §66. PASSWORD REGENERATION                                L 3231
#    §67. GUIDE                                                L 3270
#    §68. EXTERNAL ASSETS REVIEW                               L 3283
#    §69. COMPANY INFO ADMIN                                   L 3445
#    §70. HIWORKS LINKS (HR)                                   L 3519
#    §71. PROFILE                                              L 3616
#    §72. REMINDERS                                            L 3799
#    §73. EXPORT WEEKLY CSV                                    L 3840
#    §74. PERMISSIONS S3 (위임 1차)                              L 6208
#    §75. PERMISSIONS S3 v3 (그룹·매트릭스)                       L 6458
#    §76. PERMISSIONS REPORT                                   L 6749
#
#  PART XIII. AI 통합
#    §77. VICTOR AI ASK                                        L 6177
#
# ───────────────────────────────────────────────────────────────────────────────
#  사용법:
#   - 특정 영역 수정: PART X 의 §N 라인으로 점프 (에디터 Ctrl+G)
#   - 전체 검색: 파일 내 "§N." 검색 (예: "§42." → SALES QUOTATIONS)
#   - 라우트 검색: "@app.get(" 또는 "@app.post("
#   - 행 수가 정확하지 않을 수 있음 (수정 시 변동) — _INDEX_main_py.md 참조
#  관련 문서: app/menu_catalog.py (사용자용 M-XX-YY 코드)
# ═══════════════════════════════════════════════════════════════════════════════


BASE = os.path.dirname(os.path.dirname(__file__))
app = FastAPI(title="KNK 일일업무일지 v2")
app.add_middleware(SessionMiddleware, secret_key="knk-haist-2026-phase1")
app.mount("/static", StaticFiles(directory=os.path.join(BASE, "static")), name="static")
# v5H142: 업로드 파일(소모품 발주 이미지 등) 정적 서빙
_uploads_root = os.path.join(BASE, "uploads")
os.makedirs(_uploads_root, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=_uploads_root), name="uploads")
tpl = Jinja2Templates(directory=os.path.join(BASE, "app", "templates"))

# v5H5 (2026-05-02) — t() Jinja 전역: 어느 템플릿에서든 {{ t('키', '기본값') }} 호출 가능
# lang은 request.session.get('lang','ko') 로 자동 결정 (캐시 컨텍스트 없을 시 'ko')
def _t_helper(key: str, default: str = "", lang: str = "ko"):
    try:
        from app.i18n import T
        v = T.get(key)
        if isinstance(v, dict):
            return v.get(lang) or v.get("ko") or default or key
    except Exception:
        pass
    return default or key
tpl.env.globals["t"] = _t_helper


# =====================================================
# v5H210 (2026-05-08) — 공휴일 자동 계산 (holidays 라이브러리)
# 음력 환산·대체공휴일 규칙·nghỉ bù 모두 라이브러리가 자동 처리.
# 매년 수동 갱신 불필요 — 라이브러리 업데이트(`pip install -U holidays`)만으로 최신화.
# 범위: 현재 연도 ± (이전 1년 ~ 이후 5년) = 7년 슬라이딩 윈도우.
# =====================================================
def _build_holidays_auto():
    """holidays 라이브러리로 한국·베트남 공휴일 자동 생성.
    근로자의 날(5/1)은 한국 근로기준법상 휴일이지만 공휴일이 아니어서 라이브러리에 없음 → 수동 추가.
    제헌절(7/17)은 2008년부터 공휴일 아님 → 라이브러리가 포함시키면 제외.
    """
    try:
        import holidays as _hol_lib
    except ImportError:
        # 라이브러리 미설치 시 빈 dict — UI 는 그대로 동작 (배지만 안 뜸)
        return {}, {}
    from datetime import date as _d
    cy = _d.today().year
    years = list(range(cy - 1, cy + 6))  # 7년치 (작년 ~ 5년 후)
    kr_dict, vn_dict = {}, {}
    try:
        kr_lib = _hol_lib.KR(years=years)
        for d, name in kr_lib.items():
            if "제헌절" in name:  # 2008년부터 공휴일 아님 — 제외
                continue
            kr_dict[d.isoformat()] = name
        # 근로자의 날 (근로기준법) — 라이브러리에 없음
        for y in years:
            kr_dict.setdefault(f"{y}-05-01", "근로자의 날")
    except Exception as _e:
        print(f"[holidays] KR 생성 실패: {_e}")
    try:
        vn_lib = _hol_lib.VN(years=years)
        for d, name in vn_lib.items():
            vn_dict[d.isoformat()] = name
    except Exception as _e:
        print(f"[holidays] VN 생성 실패: {_e}")
    return kr_dict, vn_dict

HOLIDAYS_KR, HOLIDAYS_VN = _build_holidays_auto()
print(f"[holidays] 자동 생성 완료 — KR {len(HOLIDAYS_KR)}건 / VN {len(HOLIDAYS_VN)}건")


# v5H215 (2026-05-08) — status → stage 자동 매핑 (단순화)
# 사용자가 선택한 세부 status 를 그대로 stage 로 표시 — 별도 일반화 라벨 없음.
# (예전 일괄 '제안작성' 라벨은 검사기 '제안서 해당없음' 케이스와 의미 충돌 → 폐기)
# 수주확정(mgmt_code 발급) 시점만 코드 흐름에서 stage='수주확정' 으로 별도 마킹 — 이건 기존 그대로.
def stage_from_status(status: str) -> str:
    return (status or "초기협의").strip() or "초기협의"

# v5H203: 공휴일 dict 을 Jinja 전역으로 노출 (knk_datepicker partial 에서 사용)
tpl.env.globals["KNK_HOLIDAYS_KR"] = HOLIDAYS_KR
tpl.env.globals["KNK_HOLIDAYS_VN"] = HOLIDAYS_VN

# v5H210: 공휴일 데이터 진단 엔드포인트 — 자동 계산 결과 확인용
@app.get("/_debug/holidays")
async def _debug_holidays():
    try:
        import holidays as _hl
        lib_ver = getattr(_hl, "__version__", "?")
    except ImportError:
        lib_ver = "미설치"
    years = sorted(set(d[:4] for d in HOLIDAYS_KR.keys()))
    return {
        "version": "v5H210",
        "library": f"holidays {lib_ver}",
        "auto_generated": True,
        "year_range": f"{years[0]} ~ {years[-1]}" if years else "(empty)",
        "kr_count": len(HOLIDAYS_KR),
        "vn_count": len(HOLIDAYS_VN),
        "kr_sample": dict(list(HOLIDAYS_KR.items())[:5]),
        "vn_sample": dict(list(HOLIDAYS_VN.items())[:5]),
    }


# =====================================================
# v5 H5 (2026-04-29 대표 일괄 승인) — 에러 페이지 핸들러 (404/500)
# 시안 G_standalone/master_G_error_v5H5.html 직접 이식
# =====================================================
from fastapi import HTTPException as _HTTPException
from starlette.exceptions import HTTPException as _StarletteHTTPException

@app.exception_handler(_StarletteHTTPException)
async def _v5_http_exception_handler(request: Request, exc: _StarletteHTTPException):
    code = exc.status_code
    if code in (404, 403, 500, 502, 503):
        try:
            return tpl.TemplateResponse(
                request=request, name="error.html",
                context={"status_code": code, "detail": str(exc.detail or ""),
                         "request": request, "user": get_user(request) if hasattr(request, 'session') else None,
                         "lang": "ko", "i": {}, "LANGS": LANGS,
                         "app_name": "HAIST WORKS"},
                status_code=code,
            )
        except Exception:
            pass
    return JSONResponse({"error": str(exc.detail), "status": code}, status_code=code)


# v5H158: 전역 Exception 핸들러 — 모든 미처리 예외를 가시화 (5xx 무음 차단)
@app.exception_handler(Exception)
async def _v5_global_exception_handler(request: Request, exc: Exception):
    import traceback as _tb
    _tb_str = _tb.format_exc()
    _err_msg = f"{type(exc).__name__}: {exc}"
    print("=" * 70)
    print(f"[v5H158 GLOBAL EXC] {request.method} {request.url.path}")
    print(f"  → {_err_msg}")
    print(_tb_str)
    print("=" * 70)
    # 사용자에게 친절한 HTML
    safe_path = str(request.url.path)
    return HTMLResponse(
        f"""<!DOCTYPE html><html lang="ko"><head><meta charset="UTF-8">
<title>오류 - KNK</title>
<style>
  body {{ font-family: 'Malgun Gothic', sans-serif; padding: 40px; background: #fef9f3; color: #2a2418; line-height: 1.6; }}
  .box {{ max-width: 900px; margin: 0 auto; background: #fff; border: 1px solid #f3d2d2; border-left: 6px solid #c00; border-radius: 10px; padding: 28px 32px; box-shadow: 0 4px 18px rgba(0,0,0,0.05); }}
  h1 {{ color: #c00; margin: 0 0 12px; font-size: 22px; }}
  .url {{ font-family: monospace; background: #f5f5f5; padding: 4px 10px; border-radius: 4px; font-size: 12.5px; color: #555; }}
  pre {{ background: #fff3f3; padding: 14px 16px; border: 1px solid #f99; border-radius: 6px; white-space: pre-wrap; word-break: break-all; font-size: 12px; color: #7a1f1f; max-height: 50vh; overflow: auto; margin-top: 14px; }}
  .actions {{ margin-top: 18px; display: flex; gap: 10px; }}
  .btn {{ padding: 9px 18px; border-radius: 6px; font-size: 13px; font-weight: 700; text-decoration: none; }}
  .btn-pri {{ background: #b91c1c; color: #fff; }}
  .btn-sec {{ background: #fff; color: #333; border: 1px solid #d6cdb6; }}
  .hint {{ background: #fff7e0; padding: 10px 14px; border-radius: 6px; border-left: 3px solid #d4a574; margin-top: 14px; font-size: 13px; color: #555; }}
</style></head><body>
<div class="box">
  <h1>⚠ 페이지 오류</h1>
  <div>요청: <span class="url">{request.method} {safe_path}</span></div>
  <div class="hint">💡 KNK_시작.bat 검은 창에 더 자세한 정보가 출력됩니다. 캡처해서 빅터에게 전달해주세요.</div>
  <pre>{_err_msg}

{_tb_str}</pre>
  <div class="actions">
    <a class="btn btn-sec" href="javascript:history.back()">← 뒤로</a>
    <a class="btn btn-pri" href="/home">🏠 홈</a>
  </div>
</div>
</body></html>""",
        status_code=500
    )


@app.on_event("startup")
def startup():
    init_db()
    seed_all()
    seed_sample_tasks(14)
    # v5H45 (2026-05-03 대표 지시) — 빈 페이지 자동 보충용 비즈니스 데이터 시드
    try:
        seed_business_data()
    except Exception as _e:
        print(f"[SEED-BIZ ERR] {_e}")
    # v5H50 (2026-05-03) — 최근 7일 task가 30건 미만이면 자동 보충 (dashboard/feed 데이터 신선도)
    try:
        added = seed_recent_tasks_topup()
        if added:
            print(f"[SEED-RECENT] +{added} fresh tasks for current week")
    except Exception as _e:
        print(f"[SEED-RECENT ERR] {_e}")
    # v5H58 (2026-05-03) — 고객사 등급 자동 재계산 (startup + 24h 주기)
    try:
        from . import customer_tier as _ct
        with db_session() as c:
            n = _ct.refresh_all_customer_tiers(c)
        print(f"[TIER] {n} customers tier auto-computed")
    except Exception as _e:
        print(f"[TIER ERR] {_e}")
    _start_tier_refresh_scheduler()
    # OPS-P1-A2 (B2 안 채택): 일일 미작성자 시스템 알림 스케줄러 시작
    _start_daily_reminder_scheduler()
    # v5H172 (2026-05-06): 통화 데이터 백필.
    #   orders.currency 컬럼은 DEFAULT 'KRW' 라 NULL 이 아니라 'KRW' 명시값이 들어있음.
    #   v5H171 이전 발행 SO 는 실제 프로젝트 통화와 무관하게 'KRW' 가 박혀 있으므로,
    #   "수금이력 0 건" 인 SO 만 안전하게 프로젝트 헤더 통화로 sync.
    try:
        with db_session() as c:
            ord_cols = {r[1] for r in c.execute("PRAGMA table_info(orders)").fetchall()}
            if "currency" in ord_cols:
                _r = c.execute("""
                    UPDATE orders
                       SET currency = (SELECT p.currency FROM projects p WHERE p.id = orders.project_id)
                     WHERE project_id IS NOT NULL
                       AND COALESCE(currency,'KRW') != COALESCE(
                             (SELECT p.currency FROM projects p WHERE p.id = orders.project_id),
                             'KRW')
                       AND COALESCE(
                             (SELECT p.currency FROM projects p WHERE p.id = orders.project_id),
                             '') != ''
                       AND NOT EXISTS (
                             SELECT 1 FROM receipts_payment rp WHERE rp.order_id = orders.id)
                """)
                if _r.rowcount:
                    print(f"[CCY-BACKFILL] orders.currency 백필 {_r.rowcount}건 (수금이력 없는 SO 만)")
                else:
                    print(f"[CCY-BACKFILL] 백필 대상 0건 (모든 SO 가 프로젝트 통화와 일치)")
            # order_items 도 동일 처리 (있으면)
            # v5H178b: 부모 orders.currency 변경분이 자식 order_items.currency 와
            # 불일치할 경우 — 수금이력 없는 라인은 부모 따라 재동기화 (NULL 만 채우는 게 아니라
            # mismatch 도 정리). 이로써 안지연3(USD) 처럼 v5H171 이전 발행 SO 의 KRW 잔재 제거.
            try:
                oi_cols = {r[1] for r in c.execute("PRAGMA table_info(order_items)").fetchall()}
                if "currency" in oi_cols:
                    _r2 = c.execute("""
                        UPDATE order_items
                           SET currency = (SELECT currency FROM orders WHERE orders.id = order_items.order_id)
                         WHERE COALESCE(currency,'') != COALESCE(
                                 (SELECT currency FROM orders WHERE orders.id = order_items.order_id),
                                 '')
                           AND NOT EXISTS (
                                 SELECT 1 FROM receipts_payment rp
                                  WHERE rp.order_id = order_items.order_id)
                    """)
                    if _r2.rowcount:
                        print(f"[CCY-BACKFILL] order_items.currency 부모 sync {_r2.rowcount}건")
            except Exception:
                pass
    except Exception as _e:
        print(f"[CCY-BACKFILL ERR] {_e}")
    # v5H178b (2026-05-06): 호기 라벨 중복 자동 정리 — 같은 SO 안에 동일 unit_label
    # 이 2개 이상이면 (예: '2호기' × 4) 첫 라벨의 시작 숫자를 추출해 순차 재번호.
    #   '2호기' × 4 → 2호기, 3호기, 4호기, 5호기
    #   '소모품-1차' × 3 → 소모품-1차, 소모품-2차, 소모품-3차
    # 안전 장치: 모든 라인이 *완전히 동일한* 라벨일 때만 작동. 부분 중복은 수기 정리.
    try:
        import re as _re_lbl
        with db_session() as c:
            sos = c.execute("""
                SELECT order_id, MIN(unit_label) AS lbl, COUNT(*) AS n
                  FROM order_items
                 WHERE unit_label IS NOT NULL AND unit_label != ''
                 GROUP BY order_id
                HAVING COUNT(DISTINCT unit_label) = 1 AND COUNT(*) > 1
            """).fetchall()
            fixed = 0
            for r in sos:
                oid = r["order_id"]
                lbl = r["lbl"]
                # 시작 숫자 + 접미사 분리 (예: '2호기' → 2 + '호기', '소모품-1차' → 1 + '차', prefix='소모품-')
                m = _re_lbl.match(r"^(.*?)(\d+)([^\d]*)$", lbl)
                if not m:
                    continue
                prefix, start_n, suffix = m.group(1), int(m.group(2)), m.group(3)
                items = c.execute(
                    "SELECT id FROM order_items WHERE order_id=? ORDER BY id ASC",
                    (oid,)
                ).fetchall()
                for i, it in enumerate(items):
                    new_lbl = f"{prefix}{start_n + i}{suffix}"
                    c.execute("UPDATE order_items SET unit_label=? WHERE id=?",
                              (new_lbl, it["id"]))
                # orders.unit_label / unit_qty / total_amount 재계산
                rows = c.execute(
                    "SELECT unit_label, amount FROM order_items WHERE order_id=? ORDER BY id ASC",
                    (oid,)
                ).fetchall()
                new_lbl_join = " · ".join(rr["unit_label"] for rr in rows)
                new_qty = len(rows)
                new_total = sum(float(rr["amount"] or 0) for rr in rows)
                c.execute(
                    "UPDATE orders SET unit_label=?, unit_qty=?, total_amount=? WHERE id=?",
                    (new_lbl_join, new_qty, new_total, oid)
                )
                fixed += 1
            if fixed:
                print(f"[LBL-FIX] 중복 호기 라벨 자동 재번호 {fixed}건 (orders.unit_qty/total_amount 동기화)")
    except Exception as _e:
        print(f"[LBL-FIX ERR] {_e}")
    # v5H181 (2026-05-06): customers.tier='신규' 비표준 → '일반' 으로 정리
    try:
        with db_session() as c:
            _r = c.execute("UPDATE customers SET tier='일반' WHERE tier='신규'")
            if _r.rowcount:
                print(f"[TIER-FIX] customers.tier='신규' → '일반' 정리 {_r.rowcount}건")
    except Exception as _e:
        print(f"[TIER-FIX ERR] {_e}")


# v5H58: 24시간마다 자동 재계산 (백그라운드 타이머)
def _tier_refresh_tick():
    import threading as _th
    try:
        from . import customer_tier as _ct
        with db_session() as c:
            n = _ct.refresh_all_customer_tiers(c)
        print(f"[TIER-AUTO] refreshed {n} customers")
    except Exception as e:
        print(f"[TIER-AUTO ERR] {e}")
    timer = _th.Timer(86400, _tier_refresh_tick)  # 24h
    timer.daemon = True
    timer.start()


def _start_tier_refresh_scheduler():
    import threading as _th
    timer = _th.Timer(86400, _tier_refresh_tick)
    timer.daemon = True
    timer.start()


# =====================================================
# OPS-P1-A2 [D-008] B2 안 — 일일 미작성자 시스템 알림 (16:30 평일)
# 외부 의존 0건 (stdlib threading.Timer). 메일·메신저 발송 X.
# notifications 테이블 INSERT만 → 다음 로그인 시 🔔 자동 표시.
# =====================================================
def _send_daily_missing_reminders():
    """오늘 일일업무카드를 작성하지 않은 활성 사용자에게 시스템 알림."""
    try:
        today_iso = date.today().isoformat()
        with db_session() as c:
            # 활성 + role 'admin'/'ceo' 제외 (대표·관리자는 일지 의무 X)
            missing = [dict(r) for r in c.execute(
                """SELECT u.id, u.name, u.rank
                   FROM users u
                   WHERE u.is_active = 1
                     AND u.role NOT IN ('admin','ceo')
                     AND NOT EXISTS (
                       SELECT 1 FROM tasks t
                       WHERE t.user_id = u.id AND t.work_date = ?
                     )""",
                (today_iso,)
            ).fetchall()]
        sent = 0
        for u in missing:
            try:
                ok = notify_user(
                    user_id=u["id"], type="TASK",
                    title="🔔 오늘 업무카드 미작성",
                    body=f"{u.get('name','')} {u.get('rank','') or ''}님, 오늘({today_iso}) 일일업무카드가 아직 등록되지 않았습니다.",
                    link="/daily",
                )
                if ok:
                    sent += 1
            except Exception:
                continue
        print(f"[DAILY-REMINDER] {today_iso} → 미작성 {len(missing)}명 / 알림 {sent}건 발송")
    except Exception as e:
        print(f"[DAILY-REMINDER ERR] {e}")


def _seconds_until_next_1630_weekday() -> float:
    """다음 평일(월~금) 16:30 까지 남은 초. 토·일이면 다음 월요일 16:30."""
    from datetime import datetime as _dt, timedelta as _td
    now = _dt.now()
    target = now.replace(hour=16, minute=30, second=0, microsecond=0)
    if target <= now:
        target = target + _td(days=1)
    # 주말 건너뛰기
    while target.weekday() >= 5:  # 5=토, 6=일
        target = target + _td(days=1)
    return max(60.0, (target - now).total_seconds())


def _daily_reminder_tick():
    """타이머 콜백 — 평일 16:30에 알림 발송 후 다음 일정 재예약."""
    import threading as _th
    from datetime import datetime as _dt
    try:
        if _dt.now().weekday() < 5:  # 평일만
            _send_daily_missing_reminders()
    except Exception as e:
        print(f"[DAILY-REMINDER tick ERR] {e}")
    # 다음 평일 16:30 재예약
    delay = _seconds_until_next_1630_weekday()
    timer = _th.Timer(delay, _daily_reminder_tick)
    timer.daemon = True
    timer.start()


def _start_daily_reminder_scheduler():
    """startup 1회 호출 — 첫 실행 시각 예약."""
    import threading as _th
    delay = _seconds_until_next_1630_weekday()
    print(f"[DAILY-REMINDER] 스케줄러 시작. 다음 발송까지 {int(delay)}초 ({int(delay/3600)}시간).")
    timer = _th.Timer(delay, _daily_reminder_tick)
    timer.daemon = True
    timer.start()


# =====================================================
# Phase 1 — 빅터 자연어 라우팅 + 메뉴 도움말 API (대표 결재 2026-04-29)
# =====================================================
@app.post("/victor/route")
async def victor_route_query(req: Request):
    """빅터 자연어 → 페이지 후보. 정확 매치 1건이면 redirect URL 반환.
    Body: {q: "견적서 작성"} → JSON.
    """
    u = get_user(req)
    if not u:
        print("[VICTOR-ROUTE] 401 login_required")
        return JSONResponse({"error": "login_required"}, 401)
    try:
        data = await req.json()
        q = (data.get("q") or "").strip()
    except Exception:
        q = ""
    print(f"[VICTOR-ROUTE] 요청 q='{q}' user={u.get('name','?')}")
    if not q:
        return JSONResponse({"ok": False, "msg": "검색어가 비어있습니다."})
    candidates = _menu.search(q, limit=5)
    print(f"[VICTOR-ROUTE] candidates={len(candidates)} → {[c['code'] for c in candidates]}")
    if not candidates:
        return JSONResponse({
            "ok": False,
            "msg": f"'{q}' 와 일치하는 메뉴를 찾지 못했습니다. M-XX-YY 코드 또는 메뉴 이름으로 시도해주세요.",
            "candidates": []
        })
    # 단일 매치 → 자동 이동 / 다중 매치 → 후보 표시
    if len(candidates) == 1:
        m = candidates[0]
        return JSONResponse({
            "ok": True,
            "auto_navigate": True,
            "code": m["code"], "label": m["label"], "path": m["path"], "tip": m["tip"]
        })
    return JSONResponse({
        "ok": True,
        "auto_navigate": False,
        "msg": f"'{q}' 후보 {len(candidates)}개 — 클릭하여 이동:",
        "candidates": [
            {"code": m["code"], "label": m["label"], "path": m["path"], "tip": m["tip"]}
            for m in candidates
        ],
    })


@app.get("/api/menu-help")
async def api_menu_help(req: Request, path: str = "", code: str = ""):
    """현재 페이지의 M-코드·라벨·도움말 조회 (❓ 버튼이 호출)."""
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "login_required"}, 401)
    m: dict | None = None
    if code:
        m = _menu.by_code(code)
    elif path:
        m = _menu.by_path(path.split("?")[0])
    if not m:
        return JSONResponse({"ok": False, "msg": "이 화면의 도움말이 아직 등록되지 않았습니다."})
    return JSONResponse({
        "ok": True,
        "code": m["code"], "label": m["label"], "path": m["path"],
        "tip": m["tip"], "persona": m["persona"], "aliases": m["aliases"],
    })


@app.get("/api/menu-catalog")
async def api_menu_catalog(req: Request):
    """전체 카탈로그 — 사이드바·도움말 일괄 조회용."""
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "login_required"}, 401)
    return JSONResponse({"ok": True, "menus": _menu.all_menus(), "hubs": _menu.hubs()})


@app.post("/admin/daily-reminder/run-now")
async def admin_daily_reminder_run_now(req: Request):
    """관리자 수동 트리거 — 즉시 발송 (테스트/긴급용)."""
    u = require(req, ["admin", "ceo"])
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    _send_daily_missing_reminders()
    return JSONResponse({"ok": True, "msg": "일일 미작성자 알림 즉시 발송 완료"})


CATEGORIES = ["설계", "제조", "고객대응", "회의", "출장", "검토", "개발", "구매", "품질", "기타"]
STATUSES = ["진행중", "완료", "지연", "대기", "보류"]


# =====================================================
# HELPERS
# =====================================================

# Victor 사이드 도크 — 현재 페이지 기반 맥락 칩 (제안 #08)
VICTOR_CONTEXT_CHIPS = {
    "/home":        ["내 할 일", "오늘 등록", "팀 현황"],
    "/daily":       ["오늘 업무 입력", "어제 업무", "이번 주 보고"],
    "/calendar":    ["이번 달 일정", "내 휴가", "회의 등록"],
    "/notifications": ["최신 알림", "읽지 않은 알림", "내 관련 변경"],
    "/feed":        ["오늘 팀 현황", "지연 공정", "최근 변경"],
    "/now":         ["실시간 누가 무엇을", "지연 공정", "최근 티켓"],
    "/progress":    ["지연 공정", "내 미완료 공정", "납기 임박"],
    "/tickets":     ["미처리 티켓", "내가 받은 요청", "티켓 등록 방법"],
    "/issues":      ["미해결 이슈", "긴급 이슈", "이슈 등록 방법"],
    "/changes":     ["미확인 변경", "내 관련 변경", "긴급 변경"],
    "/dashboard":   ["이번달 매출", "지연 공정", "미처리 업무"],
    "/bottlenecks": ["병목 상세", "담당자 배정", "해결 기록"],
    "/admin":       ["사용자 추가", "권한 변경", "Excel 내보내기"],
    "/sales":       ["이번달 수주", "고객사 Top 5", "영업 단계별"],
    "/logistics":   ["안전재고 미달", "미입고 발주", "재고 금액"],
    "/parts":       ["부품 검색", "FIFO 원가", "공급사 단가"],
    "/po":          ["발주 등록", "미입고 발주", "공급사 리드타임"],
    "/stock/movements": ["최근 출고", "재고 실사", "수불부"],
    "/rates":       ["오늘 환율", "최근 갱신", "통화별 추세"],
    "/board/company": ["긴급 공지", "내 관련 글", "새 글쓰기"],
    "/board/team": ["팀 공지", "내 팀 글", "승인 대기"],
}

def _victor_chips_for_path(path: str):
    """현재 URL 경로에 맞는 맥락 칩 반환. 매칭 없으면 기본 칩."""
    if not path:
        return VICTOR_CONTEXT_CHIPS["/home"]
    # exact match 먼저
    if path in VICTOR_CONTEXT_CHIPS:
        return VICTOR_CONTEXT_CHIPS[path]
    # prefix match (/changes/123 → /changes)
    for key in VICTOR_CONTEXT_CHIPS:
        if path.startswith(key + "/"):
            return VICTOR_CONTEXT_CHIPS[key]
    return VICTOR_CONTEXT_CHIPS["/home"]


# C안 §4 — 워크스페이스 스위처 (시안 12B 헤더)
WORKSPACES = [
    {"key": "hub",   "name": "통합",          "desc": "HAIST WORKS 메인 (업무·진행·이슈·요청)", "icon": "🏢", "href": "/home",      "external": False},
    {"key": "sales", "name": "매출·영업 센터", "desc": "Sales Hub · 영업·수주·고객사",          "icon": "📈", "href": "/sales",     "external": False},
    {"key": "logi",  "name": "자재·구매 센터", "desc": "Logistics Hub · 자재·구매·재고",         "icon": "📦", "href": "/logistics", "external": False},
]

def workspaces_for(user):
    """권한 기반 워크스페이스 목록 (시안 12B 상단 ws-switcher 용).

    2026-04-28 R/W 분리: VIEW 권한자에게도 메뉴 노출 (읽기 전용 진입).
    쓰기 권한 없으면 들어가서 등록 버튼이 비활성/숨김으로 표시.
    """
    if not user:
        return [WORKSPACES[0]]
    out = [WORKSPACES[0]]
    if can_view_sales(user):
        out.append(WORKSPACES[1])
    if can_view_logistics(user):
        out.append(WORKSPACES[2])
    return out

def current_workspace_for(path: str):
    """현재 path 기반 워크스페이스 매핑 (2026-04-28 G8 보강 — 사이드바 모든 sales/logi 경로 정합)"""
    p = path or ""
    # 매출·영업 센터 (Sales Hub)
    #   /sales* · /customer* · /export* (수출입·FTA) · /quotation* · /order*
    if (p.startswith("/sales") or p.startswith("/customer")
        or p.startswith("/export") or p.startswith("/quotation")
        or p.startswith("/order")):
        return WORKSPACES[1]
    # 자재·구매 센터 (Logistics Hub)
    #   /logistics* · /parts* · /po* · /stock* · /supplier* · /rates · /fx*
    if (p.startswith("/logistics") or p.startswith("/parts")
        or p.startswith("/po") or p.startswith("/stock")
        or p.startswith("/supplier") or p.startswith("/rates")
        or p.startswith("/fx")):
        return WORKSPACES[2]
    # 그 외 = 통합 (Hub)
    return WORKSPACES[0]


# =====================================================
# Phase 1 — 메뉴 식별번호 헬퍼 (대표 결재 2026-04-29)
# =====================================================
def _current_menu_for(path: str) -> dict | None:
    """현재 URL → 카탈로그 매칭 결과 (없으면 None)."""
    try:
        # 쿼리스트링 제거
        clean = (path or "").split("?")[0]
        return _menu.by_path(clean)
    except Exception:
        return None


# v5H35 (2026-05-02) — 테스트용 가짜 날짜 헬퍼 (세션 기반, 개인별)
def fake_today_iso(request) -> str:
    """세션에 fake_today 가 설정되면 그 값, 아니면 실제 today (ISO)."""
    try:
        if hasattr(request, "session"):
            fk = request.session.get("fake_today")
            if fk:
                return str(fk)
    except Exception:
        pass
    return date.today().isoformat()


def ctx(request, name, **kwargs):
    # 사용자 언어 결정
    user = kwargs.get("user")
    lang = "ko"
    if user and isinstance(user, dict):
        lang = user.get("lang") or "ko"
    elif hasattr(request, "session"):
        lang = request.session.get("lang", "ko")

    # 번역 사전 생성
    i = get_all_translations(lang)

    # CX23c v4 마스트헤드 컨텍스트 (대표 승인 2026-04-28)
    _today_d = date.today()
    _wkdays = ['월', '화', '수', '목', '금', '토', '일']
    _today_kor = f"{_today_d.year} · {_today_d.month}월 {_today_d.day}일 {_wkdays[_today_d.weekday()]}요일"
    _edition_no = _today_d.timetuple().tm_yday  # 일년 중 N번째 발행

    # v5H35: 테스트 모드 가짜 날짜 (세션 fake_today 가 설정된 경우만)
    _today_iso = fake_today_iso(request)
    _fake_active = False
    try:
        if hasattr(request, "session"):
            _fake_active = bool(request.session.get("fake_today"))
    except Exception:
        pass

    base = {
        "categories": CATEGORIES,
        "statuses": STATUSES,
        "today": _today_iso,
        "real_today": date.today().isoformat(),
        "fake_date_active": _fake_active,
        "now": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "today_kor": _today_kor,            # CX23c 마스트헤드용
        "edition_no": _edition_no,          # CX23c VOL/NO 표시
        "lang": lang,
        "i": i,
        "LANGS": LANGS,
        # HAIST WORKS 브랜드 (통합 후)
        "app_name": "HAIST WORKS",
        "app_subtitle": "KNK 통합 업무 플랫폼",
        "brand_slogan": "Human & AI create the Best",
        # 하이웍스 외부 시스템 URL (admin 설정에서 변경 가능)
        "hiworks_approval_url": get_setting("hiworks_approval_url", "https://office.hiworks.com/"),
        "hiworks_mail_url":     get_setting("hiworks_mail_url",     "https://mail.hiworks.com/"),
        "hiworks_domain":       get_setting("hiworks_domain", ""),
        # Victor 도크 맥락 칩 (제안 #08)
        "victor_chips":         _victor_chips_for_path(str(request.url.path) if hasattr(request, "url") else ""),
        # C안 v2 §2-4 — 워크스페이스 스위처 (uppercase + lowercase 양쪽 노출)
        "workspaces":          workspaces_for(user) if user else [],
        "current_workspace":   current_workspace_for(str(request.url.path) if hasattr(request, "url") else ""),
        "WORKSPACES":          workspaces_for(user) if user else [],
        # Phase 1 (대표 결재 2026-04-29): 메뉴 식별번호 — 현재 페이지 M-코드 + 도움말
        "current_menu":        _current_menu_for(str(request.url.path) if hasattr(request, "url") else ""),
        # v5H208: 공휴일 데이터를 context 로도 직접 전달 (globals 미적용 환경 안전망)
        "KNK_HOLIDAYS_KR":     HOLIDAYS_KR,
        "KNK_HOLIDAYS_VN":     HOLIDAYS_VN,
    }
    # 글로벌 알림 카운트 (로그인 상태일 때만)
    uid = request.session.get("user_id") if hasattr(request, "session") else None
    if uid:
        try:
            base["unread_notif"] = count_unread(uid)
        except Exception:
            base["unread_notif"] = 0
    else:
        base["unread_notif"] = 0
    # v5H72: 권한 헬퍼 — 템플릿에서 사용
    base["can_sales"]      = can_use_sales(user) if user else False
    base["can_sales_del"]  = can_delete_sales(user) if user else False  # 삭제 전용 (더 엄격)
    base["can_logi"]       = can_use_logistics(user) if user else False
    base["is_admin"]       = bool(user and user.get("role") in ("admin", "ceo"))
    base.update(kwargs)
    # v5H178b: HTML 페이지는 캐시 금지 — 인라인 편집 후 다른 탭/뒤로가기에서 stale 데이터 방지
    resp = tpl.TemplateResponse(request=request, name=name, context=base)
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


def get_user(req: Request):
    uid = req.session.get("user_id")
    if not uid:
        return None
    with db_session() as c:
        row = c.execute(
            """SELECT u.*, t.name AS team_name, t.code AS team_code, t.is_lab AS team_is_lab,
                      t.sector AS team_sector
               FROM users u LEFT JOIN teams t ON u.team_id=t.id
               WHERE u.id=? AND u.is_active=1""",
            (uid,),
        ).fetchone()
        return dict(row) if row else None


def require(req: Request, roles=None):
    u = get_user(req)
    if not u:
        return None
    if roles and u["role"] not in roles:
        return None
    return u


def role_home(user) -> str:
    """ESC-02 (감사보고_04 2026-04-22): role별 홈 URL 반환.
    require() 실패 + 로그인 상태인 경우 /login 대신 적절한 홈으로 리다이렉트.
    - ceo/admin/executive → /dashboard
    - leader → /team
    - member/그 외 → /home
    """
    if not user:
        return "/login"
    role = (user.get("role") or "member") if isinstance(user, dict) else str(user["role"])
    if role in ("ceo", "admin", "executive"):
        return "/dashboard"
    elif role == "leader":
        return "/team"
    else:
        return "/home"


def can_use_logistics(user) -> bool:
    """HAIST WORKS 물류 모듈 접근 권한.
    - admin / ceo / executive: 항상 허용
    - team_id == 7 (제조팀): 읽기 허용 (2026-04-22 대표 결재 D01-02 안B)
    - 그 외: users.can_use_logistics 플래그가 1일 때만
    """
    if not user:
        return False
    role = user.get("role") if isinstance(user, dict) else user["role"]
    if role in ("admin", "ceo", "executive"):
        return True
    # 제조팀(team_id=7) 읽기 허용 — 쓰기·발주·단가수정은 구매팀 권한 그대로 유지
    team_id = user.get("team_id") if isinstance(user, dict) else user["team_id"]
    if team_id == 7:
        return True
    flag = user.get("can_use_logistics") if isinstance(user, dict) else user["can_use_logistics"]
    return bool(flag)


def can_view_sales(user) -> bool:
    """매출·영업 **읽기 전용** 권한 (2026-04-28 대표 결재 — R/W 분리).
    - admin / ceo / executive / leader: 항상 허용 (전사 매출 현황 조회)
    - team_id 1·2·3 (영업·검사기·품질): 항상 허용
    - 그 외: users.can_view_sales 플래그 1
    - 쓰기 권한자(can_use_sales=1)는 자동 포함
    """
    if not user:
        return False
    role = user.get("role") if isinstance(user, dict) else user["role"]
    if role in ("admin", "ceo", "executive", "leader"):
        return True
    team_id = user.get("team_id") if isinstance(user, dict) else user["team_id"]
    if team_id in (1, 2, 3):
        return True
    try:
        if user.get("can_view_sales") or user.get("can_use_sales"):
            return True
    except (KeyError, IndexError):
        pass
    return False


def can_view_logistics(user) -> bool:
    """자재·구매 **읽기 전용** 권한 (2026-04-28 대표 결재 — R/W 분리).
    실무자가 부품·재고·단가·구매처를 조회해야 할 때 폭넓게 허용.
    - admin / ceo / executive / leader: 항상 허용
    - team_id 1,2,3,7,8,9,10 (영업·검사기·품질·생산1·생산2·가공·구매): 항상 허용
    - 그 외: users.can_view_logistics 플래그 1
    - 쓰기 권한자(can_use_logistics=1)는 자동 포함
    """
    if not user:
        return False
    role = user.get("role") if isinstance(user, dict) else user["role"]
    if role in ("admin", "ceo", "executive", "leader"):
        return True
    team_id = user.get("team_id") if isinstance(user, dict) else user["team_id"]
    if team_id in (1, 2, 3, 7, 8, 9, 10):
        return True
    try:
        if user.get("can_view_logistics") or user.get("can_use_logistics"):
            return True
    except (KeyError, IndexError):
        pass
    return False


def can_delete_sales(user) -> bool:
    """v5H72: 매출·영업 **삭제** 권한 — can_use_sales 보다 더 엄격.
    - admin / ceo / executive: 자동 허용
    - 기술영업팀(team_id=1) 팀장(leader): 자동 허용
    - 기타 멤버: users.can_use_sales=1 명시적으로 부여 받은 자만
    대표 정의: '기술영업팀에서 등록 권한을 받은 사용자만'."""
    if not user:
        return False
    role = (user.get("role") if isinstance(user, dict) else user["role"]) or ""
    if role in ("admin", "ceo", "executive"):
        return True
    team_id = user.get("team_id") if isinstance(user, dict) else None
    if role == "leader" and team_id == 1:
        return True
    # member 는 명시적 위임 받은 자만
    try:
        return bool(user.get("can_use_sales"))
    except Exception:
        return False


def can_use_sales(user) -> bool:
    """매출·영업 **쓰기** (등록·편집·견적·수주) 권한 (Plan Y S1).
    - admin / ceo / executive: 항상 허용
    - 영업팀·관리팀 leader/member: 항상 허용 (현장 입력자)
    - 그 외: users.can_use_sales 플래그가 1일 때만
    회귀 폴백: can_use_sales 컬럼이 없거나 미설정인 경우 can_use_logistics 로 graceful 폴백.
    """
    if not user:
        return False
    role = user.get("role") if isinstance(user, dict) else user["role"]
    if role in ("admin", "ceo", "executive"):
        return True
    # team_id 1·2·3 (대표직속·영업·관리) 폭넓게 허용 — 추후 팀장 위임 UI 로 정밀화 (S1-2)
    team_id = user.get("team_id") if isinstance(user, dict) else user["team_id"]
    if team_id in (1, 2, 3):
        return True
    try:
        flag = user.get("can_use_sales") if isinstance(user, dict) else user["can_use_sales"]
        if flag:
            return True
    except (KeyError, IndexError):
        pass
    # 회귀 폴백: 기존 logistics 권한자도 일단 허용 (S1 안전 모드, S2에서 분리 강화)
    return can_use_logistics(user)


def status_color(s):
    return {
        "진행중": ("#FFF8E1", "#F57F17"),
        "완료":   ("#E8F5E9", "#2E7D32"),
        "지연":   ("#FFEBEE", "#A5282C"),
        "대기":   ("#F5F5F5", "#4A4A4A"),
        "보류":   ("#F5F5F5", "#4A4A4A"),
    }.get(s, ("#F5F5F5", "#4A4A4A"))


def fetch_projects(c):
    return [dict(r) for r in c.execute(
        """SELECT p.id, p.code, p.name, p.type, c.name AS customer_name
           FROM projects p LEFT JOIN customers c ON p.customer_id=c.id
           WHERE p.status='진행중' ORDER BY p.id"""
    ).fetchall()]


def fetch_customers(c):
    return [dict(r) for r in c.execute(
        "SELECT id, name, tier FROM customers ORDER BY tier DESC, name"
    ).fetchall()]


# =====================================================
# AUTH
# =====================================================
@app.get("/login", response_class=HTMLResponse)
async def login_page(req: Request):
    return ctx(req, "login.html", error=None)


@app.post("/login")
async def login_post(req: Request, login_id: str = Form(...), password: str = Form(...)):
    with db_session() as c:
        u = c.execute(
            "SELECT * FROM users WHERE login_id=? AND password=? AND is_active=1",
            (login_id.strip(), hash_pw(password)),
        ).fetchone()
        if u:
            req.session["user_id"] = u["id"]
            r = u["role"]
            if r == "admin":
                return RedirectResponse("/dashboard", 303)
            if r == "ceo":
                return RedirectResponse("/dashboard", 303)
            if r in ("leader", "executive"):
                return RedirectResponse("/team", 303)
            return RedirectResponse("/daily", 303)
    return ctx(req, "login.html", error="아이디 또는 비밀번호가 올바르지 않습니다.")


@app.get("/logout")
async def logout(req: Request):
    req.session.clear()
    return RedirectResponse("/login", 303)


# =====================================================
# ROOT
# =====================================================
@app.get("/")
async def root(req: Request):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    return RedirectResponse("/home", 303)


# =====================================================
# HOME — 직관적 단일 페이지 (역할별 자동 분기)
# =====================================================
@app.get("/home", response_class=HTMLResponse)
@app.get("/home/{sel_date}", response_class=HTMLResponse)
async def home_page(req: Request, sel_date: str = "", tab: str = "",
                    no_perm: str = ""):  # D01-NEW-BANNER: 권한 없음 안내 파라미터
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if not sel_date:
        sel_date = date.today().isoformat()
    # 05 디자인팀: 3탭 분할 (내 업무 / 우리 팀 / 전사)
    # role 기반 기본값 — leader 는 team, 경영진은 all, 그 외는 my
    if tab not in ("my", "team", "all"):
        role = (u.get("role") or "member").lower()
        if role in ("ceo", "admin", "executive"):
            tab = "all"
        elif role == "leader":
            tab = "team"
        else:
            tab = "my"

    prev_d = (datetime.strptime(sel_date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
    next_d = (datetime.strptime(sel_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")

    with db_session() as c:
        teams = [dict(r) for r in c.execute(
            """SELECT t.*, u.name AS leader_name, u.rank AS leader_rank
               FROM teams t LEFT JOIN users u ON t.leader_id=u.id
               ORDER BY t.display_order"""
        ).fetchall()]

        projects = fetch_projects(c)
        customers = fetch_customers(c)

        # 전체 사용자 수
        total_users = c.execute(
            "SELECT COUNT(*) FROM users WHERE is_active=1 AND role!='admin'"
        ).fetchone()[0]
        today_reporters = c.execute(
            "SELECT COUNT(DISTINCT user_id) FROM tasks WHERE work_date=?",
            (sel_date,),
        ).fetchone()[0]
        participation_rate = round(today_reporters * 100 / total_users) if total_users else 0

        # 팀별 데이터 구축
        team_data = []
        for tm in teams:
            members = [dict(r) for r in c.execute(
                """SELECT id, name, rank, role FROM users
                   WHERE team_id=? AND is_active=1
                   ORDER BY CASE role WHEN 'ceo' THEN 0 WHEN 'executive' THEN 1
                            WHEN 'leader' THEN 2 ELSE 3 END, id""",
                (tm["id"],),
            ).fetchall()]
            mids = [m["id"] for m in members]
            if not mids:
                continue
            ph = ",".join("?" * len(mids))
            tasks = [dict(r) for r in c.execute(
                f"""SELECT t.*, u.name AS user_name, u.rank AS user_rank,
                           p.name AS project_name, p.code AS project_code,
                           cu.name AS customer_name,
                           (SELECT COUNT(*) FROM task_comments WHERE task_id=t.id) AS comment_count
                    FROM tasks t JOIN users u ON t.user_id=u.id
                    LEFT JOIN projects p ON t.project_id=p.id
                    LEFT JOIN customers cu ON t.customer_id=cu.id
                    WHERE t.user_id IN ({ph}) AND t.work_date=?
                    ORDER BY u.id, CASE t.status WHEN '지연' THEN 0 WHEN '진행중' THEN 1
                             WHEN '대기' THEN 2 WHEN '보류' THEN 3 ELSE 4 END, t.id""",
                mids + [sel_date],
            ).fetchall()]

            reported = len(set(t["user_id"] for t in tasks))
            delay_count = len([t for t in tasks if t["status"] == "지연"])
            progress_count = len([t for t in tasks if t["status"] == "진행중"])
            done_count = len([t for t in tasks if t["status"] == "완료"])

            # 신호등: 지연 있으면 빨강, 참여 50% 미만 노랑, 나머지 초록
            if delay_count > 0:
                signal = "red"
            elif reported < len(members) * 0.5:
                signal = "yellow"
            else:
                signal = "green"

            # 멤버별 그룹
            member_tasks = {}
            for m in members:
                mt = [t for t in tasks if t["user_id"] == m["id"]]
                member_tasks[m["id"]] = mt

            team_data.append({
                "team": tm,
                "members": members,
                "tasks": tasks,
                "member_tasks": member_tasks,
                "reported": reported,
                "total_members": len(members),
                "delay": delay_count,
                "progress": progress_count,
                "done": done_count,
                "total": len(tasks),
                "signal": signal,
            })

        # 내 업무 (직원/팀장 공통)
        my_tasks = [dict(r) for r in c.execute(
            """SELECT t.*, p.name AS project_name, p.code AS project_code,
                      cu.name AS customer_name,
                      (SELECT COUNT(*) FROM task_comments WHERE task_id=t.id) AS comment_count
               FROM tasks t
               LEFT JOIN projects p ON t.project_id=p.id
               LEFT JOIN customers cu ON t.customer_id=cu.id
               WHERE t.user_id=? AND t.work_date=?
               ORDER BY CASE t.status WHEN '지연' THEN 0 WHEN '진행중' THEN 1
                        WHEN '대기' THEN 2 WHEN '보류' THEN 3 ELSE 4 END, t.id""",
            (u["id"], sel_date),
        ).fetchall()]

        # 어제 미완료 (이월 후보)
        yday = (datetime.strptime(sel_date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
        pending_yday = [dict(r) for r in c.execute(
            """SELECT t.*, p.name AS project_name, cu.name AS customer_name
               FROM tasks t LEFT JOIN projects p ON t.project_id=p.id
               LEFT JOIN customers cu ON t.customer_id=cu.id
               WHERE t.user_id=? AND t.work_date=? AND t.status IN ('진행중','지연','대기')
               AND NOT EXISTS (SELECT 1 FROM tasks t2 WHERE t2.carry_from_id=t.id AND t2.work_date=?)
               ORDER BY t.id""",
            (u["id"], yday, sel_date),
        ).fetchall()]

        # 전사 요약
        all_delay = sum(td["delay"] for td in team_data)
        all_tasks = sum(td["total"] for td in team_data)

    # HAIST WORKS — 물류 KPI (간단 집계)
    try:
        from . import database as _logi_db
        logi_parts_stats = _logi_db.parts_count()
        logi_proj_stats = _logi_db.projects_count_logi()
    except Exception:
        logi_parts_stats = {"total": 0, "active": 0, "by_div": {}}
        logi_proj_stats = {"total": 0, "with_code": 0, "in_progress": 0,
                           "by_div": {}, "by_stage": {}}

    # 1순위 신규 기능 카운트 (변경/티켓/진행률 지연)
    hw_counts = {"changes_unread": 0, "tickets_pending": 0, "phases_delayed": 0,
                 "changes_recent": 0}
    try:
        from .database import (change_unread_count, change_recent_count,
                                tickets_count_for_user, progress_summary_for_user)
        hw_counts["changes_unread"] = change_unread_count(u["id"])
        hw_counts["changes_recent"] = change_recent_count(days=1)
        tk = tickets_count_for_user(u["id"], u.get("team_id"))
        hw_counts["tickets_pending"] = tk["my_open"] + tk["recv_pending"]
        pg = progress_summary_for_user(u["id"], u.get("team_id"))
        hw_counts["phases_delayed"] = pg["delayed"]
        hw_counts["phases_my_open"] = pg["my_open"]
    except Exception as e:
        print(f"[HW COUNTS ERROR] {e}")

    # 힐링 #12 §8-bis — 민감 데이터 권한 분기 (매출 컨텍스트 이중 방어)
    # 대표 지적 2026-04-24: "매출을 전직원 공개는 아닌 것 같은데…"
    # 1) UI 조건 분기 (home.html Jinja) · 2) 컨텍스트 분기 (여기) · 3) 라우트 권한 (별도)
    role = (u.get("role") or "member").lower()
    is_executive = role in ("ceo", "admin", "executive")
    is_leader_plus = role in ("ceo", "admin", "executive", "leader")

    monthly_revenue = None
    yoy_delta = None
    if is_executive:
        # 경영진만 매출 지표 컨텍스트 수신
        try:
            from datetime import date as _d
            ym = _d.today().strftime("%Y-%m")
            with db_session() as c:
                r = c.execute(
                    "SELECT COALESCE(SUM(order_amount),0) AS t "
                    "FROM projects WHERE order_date LIKE ? AND order_amount>0",
                    (f"{ym}%",)).fetchone()
                monthly_revenue = r["t"] if r else 0
                # YoY 전년 동월 대비
                last_year_ym = f"{_d.today().year - 1}-{_d.today().strftime('%m')}"
                r2 = c.execute(
                    "SELECT COALESCE(SUM(order_amount),0) AS t "
                    "FROM projects WHERE order_date LIKE ? AND order_amount>0",
                    (f"{last_year_ym}%",)).fetchone()
                last = r2["t"] if r2 else 0
                if last > 0:
                    yoy_delta = round((monthly_revenue - last) / last * 100, 1)
        except Exception as e:
            print(f"[REVENUE KPI ERROR] {e}")

    # 시간대별 인사말 (힐링 원칙서 §7-1)
    greeting_bucket = "default"
    try:
        _h = datetime.now().hour
        _n = u.get("name", "")
        if 6 <= _h < 11:
            greeting = f"좋은 아침입니다, {_n}님 ☀️"; greeting_bucket = "morning"
        elif 11 <= _h < 14:
            greeting = f"점심은 드셨나요, {_n}님? 잠깐 쉬어가요"; greeting_bucket = "lunch"
        elif 14 <= _h < 18:
            greeting = f"오후도 힘내세요, {_n}님 🌿"; greeting_bucket = "afternoon"
        elif 18 <= _h < 22:
            greeting = f"오늘도 수고하셨어요, {_n}님"; greeting_bucket = "evening"
        else:
            greeting = f"늦은 시간까지 애쓰시네요, {_n}님"; greeting_bucket = "night"
    except Exception:
        greeting = f"오늘도 평안하세요, {u.get('name','')}님"

    return ctx(
        req, "home.html",
        user=u, sel_date=sel_date, prev_date=prev_d, next_date=next_d,
        team_data=team_data, my_tasks=my_tasks, pending_yday=pending_yday,
        projects=projects, customers=customers,
        participation_rate=participation_rate,
        today_reporters=today_reporters, total_users=total_users,
        all_delay=all_delay, all_tasks=all_tasks,
        logi_parts_stats=logi_parts_stats, logi_proj_stats=logi_proj_stats,
        hw_counts=hw_counts,
        tab=tab,           # 05 디자인팀 3탭
        no_perm=no_perm,   # D01-NEW-BANNER: 권한 없음 안내 배너
        # 힐링 #12 §8-bis 권한 분기 컨텍스트
        monthly_revenue=monthly_revenue,
        yoy_delta=yoy_delta,
        is_executive=is_executive,
        is_leader_plus=is_leader_plus,
        greeting=greeting,
        greeting_bucket=greeting_bucket,  # QA-H6 패치 ④
    )


# =====================================================
# DAILY — 개인 일일 업무카드
# =====================================================
@app.get("/daily", response_class=HTMLResponse)
@app.get("/daily/{sel_date}", response_class=HTMLResponse)
async def daily_page(req: Request, sel_date: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if not sel_date:
        sel_date = date.today().isoformat()

    prev_d = (datetime.strptime(sel_date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
    next_d = (datetime.strptime(sel_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")

    with db_session() as c:
        tasks = [dict(r) for r in c.execute(
            """SELECT t.*, p.name AS project_name, p.code AS project_code,
                      c.name AS customer_name,
                      (SELECT COUNT(*) FROM task_comments WHERE task_id=t.id) AS comment_count,
                      (SELECT MAX(is_ceo_request) FROM task_comments WHERE task_id=t.id) AS has_ceo_req
               FROM tasks t
               LEFT JOIN projects p ON t.project_id=p.id
               LEFT JOIN customers c ON t.customer_id=c.id
               WHERE t.user_id=? AND t.work_date=?
               ORDER BY t.status, t.id""",
            (u["id"], sel_date),
        ).fetchall()]
        # 각 카드에 댓글 첨부
        for t in tasks:
            t["comments"] = [dict(r) for r in c.execute(
                """SELECT tc.*, u.name AS author_name, u.rank AS author_rank, u.role AS author_role
                   FROM task_comments tc JOIN users u ON tc.author_id=u.id
                   WHERE tc.task_id=? ORDER BY tc.created_at""",
                (t["id"],),
            ).fetchall()]

        # 어제 미완료 (carry-forward 후보)
        yday = (datetime.strptime(sel_date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
        pending_yday = [dict(r) for r in c.execute(
            """SELECT t.*, p.name AS project_name, c.name AS customer_name
               FROM tasks t
               LEFT JOIN projects p ON t.project_id=p.id
               LEFT JOIN customers c ON t.customer_id=c.id
               WHERE t.user_id=? AND t.work_date=? AND t.status IN ('진행중','지연','대기')
               AND NOT EXISTS (
                   SELECT 1 FROM tasks t2 WHERE t2.carry_from_id=t.id AND t2.work_date=?
               )
               ORDER BY t.id""",
            (u["id"], yday, sel_date),
        ).fetchall()]

        projects = fetch_projects(c)
        customers = fetch_customers(c)

        # 통계 - 이번 주
        wk_mon = (datetime.strptime(sel_date, "%Y-%m-%d") -
                  timedelta(days=datetime.strptime(sel_date, "%Y-%m-%d").weekday())).strftime("%Y-%m-%d")
        wk_sun = (datetime.strptime(wk_mon, "%Y-%m-%d") + timedelta(days=6)).strftime("%Y-%m-%d")
        week_stats = c.execute(
            """SELECT COUNT(*) AS total,
                      SUM(CASE WHEN status='완료' THEN 1 ELSE 0 END) AS done,
                      SUM(CASE WHEN status='진행중' THEN 1 ELSE 0 END) AS progress,
                      SUM(CASE WHEN status='지연' THEN 1 ELSE 0 END) AS delay,
                      COALESCE(SUM(hours),0) AS hours
               FROM tasks WHERE user_id=? AND work_date>=? AND work_date<=?""",
            (u["id"], wk_mon, wk_sun),
        ).fetchone()
        week_stats = dict(week_stats)

    return ctx(
        req, "daily.html",
        user=u, tasks=tasks, sel_date=sel_date,
        prev_date=prev_d, next_date=next_d,
        pending_yday=pending_yday,
        projects=projects, customers=customers,
        week_stats=week_stats, week_range=f"{wk_mon} ~ {wk_sun}",
    )


_TASK_STATUSES = ("진행중", "완료", "지연", "보류", "취소")
_TASK_CATEGORIES = ("영업", "구매", "생산", "품질", "기술", "관리", "출장", "회의", "교육", "기타")


def _validate_task_payload(d: dict) -> tuple:
    """v5H121: tasks API 정합성 검증. (title, category, status, hours) 반환.
    실패 시 ValueError(친절 한국어). 백워드 호환: 미지정/구식 값은 기본값으로 폴백."""
    title = (d.get("title") or "").strip()
    if not title:
        raise ValueError("제목은 필수입니다.")
    if len(title) > 200:
        raise ValueError("제목은 200자 이내로 입력하세요.")
    category = d.get("category") or "기타"
    if category not in _TASK_CATEGORIES:
        category = "기타"  # 폴백 — 구식 값 보호
    status = d.get("status") or "진행중"
    if status not in _TASK_STATUSES:
        status = "진행중"  # 폴백
    try:
        hours = float(d.get("hours") or 0)
    except (TypeError, ValueError):
        hours = 0.0
    if hours < 0:
        raise ValueError("공수는 0 이상이어야 합니다.")
    if hours > 24:
        raise ValueError("공수는 1일 24시간을 초과할 수 없습니다.")
    return title, category, status, hours


@app.post("/api/task")
async def api_create_task(req: Request):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    d = await req.json()
    # v5H121: 정합성 검증 + 화이트리스트
    try:
        title, category, status, hours = _validate_task_payload(d)
    except ValueError as ve:
        return JSONResponse({"ok": False, "error": str(ve)}, 400)
    # v5H28-29: project/customer — FK(매출 프로젝트·고객사) 또는 label(자유 텍스트, 사내업무 등) 자동 분기
    _pid = d.get("project_id") or None
    _plabel = (d.get("project_label") or "").strip() or None
    _cid = d.get("customer_id") or None
    _clabel = (d.get("customer_label") or "").strip() or None
    with db_session() as c:
        cur = c.execute(
            """INSERT INTO tasks(user_id, work_date, title, category, project_id, project_label,
                                  customer_id, customer_label,
                                  status, hours, notes, next_plan, due_date)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                u["id"],
                d.get("work_date") or date.today().isoformat(),
                title,
                category,
                _pid,
                _plabel if not _pid else None,  # FK 있으면 라벨 무시
                _cid,
                _clabel if not _cid else None,
                status,
                hours,
                d.get("notes") or "",
                (d.get("next_plan") or "").strip(),
                d.get("due_date") or None,
            ),
        )
        new_id = cur.lastrowid
        log_activity(c, u["id"], "task_create",
                     title=f"{u['name']} 신규 카드: {(d.get('title') or '')[:60]}",
                     body=(d.get("notes") or "")[:200],
                     task_id=new_id,
                     project_id=d.get("project_id") or None,
                     team_id=u.get("team_id"))
    # 지연으로 생성된 경우도 알림
    if (d.get("status") or "") == "지연":
        notify_status_change(new_id, u["id"], "", "지연")
    return JSONResponse({"ok": True, "id": new_id})


@app.put("/api/task/{tid}")
async def api_update_task(req: Request, tid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    d = await req.json()
    # v5H121: 정합성 검증
    try:
        title, category, new_status, hours = _validate_task_payload(d)
    except ValueError as ve:
        return JSONResponse({"ok": False, "error": str(ve)}, 400)
    with db_session() as c:
        prev = c.execute("SELECT user_id, status, title, project_id FROM tasks WHERE id=?", (tid,)).fetchone()
        if not prev or prev["user_id"] != u["id"]:
            return JSONResponse({"error": "권한 없음"}, 403)
        c.execute(
            """UPDATE tasks SET title=?, category=?, project_id=?, customer_id=?,
                               status=?, hours=?, notes=?, next_plan=?, due_date=?,
                               updated_at=datetime('now','localtime')
               WHERE id=?""",
            (
                title,
                category,
                d.get("project_id") or None,
                d.get("customer_id") or None,
                new_status,
                hours,
                d.get("notes") or "",
                (d.get("next_plan") or "").strip(),
                d.get("due_date") or None,
                tid,
            ),
        )
        if prev["status"] != new_status:
            log_activity(c, u["id"], "task_status",
                         title=f"{u['name']}: {prev['title'][:50]} — {prev['status']} → {new_status}",
                         task_id=tid, project_id=prev["project_id"], team_id=u.get("team_id"))
    if prev["status"] != new_status and new_status == "지연":
        notify_status_change(tid, u["id"], prev["status"], new_status)
    return JSONResponse({"ok": True})


@app.delete("/api/task/{tid}")
async def api_delete_task(req: Request, tid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    try:
        with db_session() as c:
            # v5H36: 실제 권한 확인 — 카드 존재 여부 + 작성자/관리자 검증
            row = c.execute("SELECT user_id FROM tasks WHERE id=?", (tid,)).fetchone()
            if not row:
                return JSONResponse({"ok": False, "error": "카드를 찾을 수 없습니다"}, 404)
            is_owner = (int(row["user_id"]) == int(u["id"]))
            is_admin = u.get("role","") in ("ceo","admin")
            if not (is_owner or is_admin):
                return JSONResponse({"ok": False, "error": "작성자만 삭제 가능"}, 403)
            # v5H37: carry_from_id 자기참조 FK가 CASCADE 아님 → 자식(이월된 카드)의 참조 NULL 처리 후 삭제
            c.execute("UPDATE tasks SET carry_from_id=NULL WHERE carry_from_id=?", (tid,))
            # activity_logs / notifications 등 task_id 참조도 안전하게 정리 (CASCADE 미설정 가능성)
            try: c.execute("UPDATE activity_logs SET task_id=NULL WHERE task_id=?", (tid,))
            except Exception: pass
            try: c.execute("UPDATE notifications SET task_id=NULL WHERE task_id=?", (tid,))
            except Exception: pass
            # 본 카드 삭제 (task_comments/task_reactions 등은 ON DELETE CASCADE)
            c.execute("DELETE FROM tasks WHERE id=?", (tid,))
        return JSONResponse({"ok": True})
    except Exception as e:
        # FK violation 또는 기타 오류 → 명확한 메시지
        import traceback; traceback.print_exc()
        return JSONResponse({"ok": False, "error": f"DB 오류: {type(e).__name__}: {str(e)[:200]}"}, 500)


@app.post("/api/task/{tid}/status")
async def api_quick_status(req: Request, tid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    d = await req.json()
    new_status = d.get("status") or "진행중"
    with db_session() as c:
        prev = c.execute("SELECT status, title, user_id, project_id FROM tasks WHERE id=?", (tid,)).fetchone()
        if not prev or prev["user_id"] != u["id"]:
            return JSONResponse({"error":"권한 없음"}, 403)
        c.execute(
            "UPDATE tasks SET status=?, updated_at=datetime('now','localtime') WHERE id=?",
            (new_status, tid),
        )
        if prev["status"] != new_status:
            log_activity(c, u["id"], "task_status",
                         title=f"{u['name']}: {prev['title'][:50]} — {prev['status']} → {new_status}",
                         task_id=tid, project_id=prev["project_id"], team_id=u.get("team_id"))
    if prev["status"] != new_status and new_status == "지연":
        notify_status_change(tid, u["id"], prev["status"], new_status)
    return JSONResponse({"ok": True})


@app.post("/api/carry-forward")
async def api_carry_forward(req: Request):
    """어제 미완료 카드 → 오늘로 이월 (전부 또는 선택)"""
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    d = await req.json()
    target = d.get("date") or date.today().isoformat()
    ids = d.get("ids") or []
    with db_session() as c:
        if not ids:
            yday = (datetime.strptime(target, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
            rows = c.execute(
                """SELECT id FROM tasks WHERE user_id=? AND work_date=?
                   AND status IN ('진행중','지연','대기')""",
                (u["id"], yday),
            ).fetchall()
            ids = [r["id"] for r in rows]
        cnt = 0
        for src_id in ids:
            src = c.execute("SELECT * FROM tasks WHERE id=? AND user_id=?",
                            (src_id, u["id"])).fetchone()
            if not src:
                continue
            exists = c.execute(
                "SELECT id FROM tasks WHERE carry_from_id=? AND work_date=?",
                (src_id, target),
            ).fetchone()
            if exists:
                continue
            c.execute(
                """INSERT INTO tasks(user_id, work_date, title, category, project_id,
                                      customer_id, status, hours, notes, due_date, carry_from_id)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    u["id"], target, src["title"], src["category"], src["project_id"],
                    src["customer_id"], "진행중", 0, src["notes"], src["due_date"], src_id,
                ),
            )
            cnt += 1
    return JSONResponse({"ok": True, "count": cnt})


# =====================================================
# SUMMARY — 통합 요약 (일/주/월 × 개인/부서/전사)
# =====================================================
@app.get("/summary", response_class=HTMLResponse)
async def summary_page(req: Request, period: str = "weekly", scope: str = "me", ref: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if period not in ("daily", "weekly", "monthly"):
        period = "weekly"
    if scope not in ("me", "team", "all"):
        scope = "me"
    # 권한: 일반 직원은 me/team만, all은 ceo/admin/executive
    if scope == "all" and u["role"] not in ("ceo", "admin", "executive"):
        scope = "team" if u.get("team_id") else "me"
    if scope == "team" and not u.get("team_id"):
        scope = "me"

    today = date.today()
    if not ref:
        ref = today.isoformat()
    rd = datetime.strptime(ref, "%Y-%m-%d").date()

    # 기간 범위 계산
    if period == "daily":
        frm = to = ref
        prev_ref = (rd - timedelta(days=1)).isoformat()
        next_ref = (rd + timedelta(days=1)).isoformat()
        period_label = ref
    elif period == "weekly":
        mon = rd - timedelta(days=rd.weekday())
        sun = mon + timedelta(days=6)
        frm, to = mon.isoformat(), sun.isoformat()
        prev_ref = (mon - timedelta(days=7)).isoformat()
        next_ref = (mon + timedelta(days=7)).isoformat()
        period_label = f"{frm} ~ {to}"
    else:  # monthly
        first = rd.replace(day=1)
        next_month_first = (first + timedelta(days=32)).replace(day=1)
        last = next_month_first - timedelta(days=1)
        frm, to = first.isoformat(), last.isoformat()
        prev_first = (first - timedelta(days=1)).replace(day=1)
        prev_ref = prev_first.isoformat()
        next_ref = next_month_first.isoformat()
        period_label = first.strftime("%Y년 %m월")

    # 스코프 필터 SQL 조각
    if scope == "me":
        scope_filter = "AND t.user_id = ?"
        scope_args = (u["id"],)
        scope_label = f"{u['name']} {u['rank']}"
    elif scope == "team":
        scope_filter = "AND uu.team_id = ?"
        scope_args = (u["team_id"],)
        scope_label = u["team_name"] or "(부서 미배정)"
    else:
        scope_filter = ""
        scope_args = ()
        scope_label = "전사"

    with db_session() as c:
        # 통계
        sql = f"""SELECT COUNT(*) AS total,
                         SUM(CASE WHEN t.status='완료' THEN 1 ELSE 0 END) AS done,
                         SUM(CASE WHEN t.status='진행중' THEN 1 ELSE 0 END) AS progress,
                         SUM(CASE WHEN t.status='지연' THEN 1 ELSE 0 END) AS delay,
                         SUM(CASE WHEN t.status='대기' THEN 1 ELSE 0 END) AS waiting,
                         COALESCE(SUM(t.hours),0) AS hours
                  FROM tasks t JOIN users uu ON t.user_id = uu.id
                  WHERE t.work_date BETWEEN ? AND ? {scope_filter}"""
        stats = dict(c.execute(sql, (frm, to) + scope_args).fetchone())
        for k in stats:
            stats[k] = stats[k] or 0
        completion_rate = round(stats["done"] * 100 / stats["total"]) if stats["total"] else 0

        # 카드 목록 (내러티브용)
        cards_sql = f"""SELECT t.*, uu.name AS user_name, uu.rank, tm.name AS team_name,
                              p.name AS project_name, cu.name AS customer_name
                       FROM tasks t JOIN users uu ON t.user_id = uu.id
                       LEFT JOIN teams tm ON uu.team_id = tm.id
                       LEFT JOIN projects p ON t.project_id=p.id
                       LEFT JOIN customers cu ON t.customer_id=cu.id
                       WHERE t.work_date BETWEEN ? AND ? {scope_filter}
                       ORDER BY CASE t.status WHEN '지연' THEN 0 WHEN '진행중' THEN 1
                                              WHEN '대기' THEN 2 ELSE 3 END, t.work_date DESC, t.id"""
        cards = [dict(r) for r in c.execute(cards_sql, (frm, to) + scope_args).fetchall()]

        # 팀별/사용자별/프로젝트별 집계
        team_agg = [dict(r) for r in c.execute(
            f"""SELECT tm.name AS team_name, COUNT(*) AS cnt,
                       SUM(CASE WHEN t.status='완료' THEN 1 ELSE 0 END) AS done,
                       SUM(CASE WHEN t.status='지연' THEN 1 ELSE 0 END) AS delay,
                       COALESCE(SUM(t.hours),0) AS hours
                FROM tasks t JOIN users uu ON t.user_id = uu.id
                LEFT JOIN teams tm ON uu.team_id = tm.id
                WHERE t.work_date BETWEEN ? AND ? {scope_filter}
                GROUP BY tm.name ORDER BY cnt DESC""",
            (frm, to) + scope_args,
        ).fetchall()]

        user_agg = [dict(r) for r in c.execute(
            f"""SELECT uu.name AS user_name, uu.rank, COUNT(*) AS cnt,
                       SUM(CASE WHEN t.status='완료' THEN 1 ELSE 0 END) AS done,
                       COALESCE(SUM(t.hours),0) AS hours
                FROM tasks t JOIN users uu ON t.user_id = uu.id
                WHERE t.work_date BETWEEN ? AND ? {scope_filter}
                GROUP BY uu.id ORDER BY cnt DESC LIMIT 20""",
            (frm, to) + scope_args,
        ).fetchall()]

        project_agg = [dict(r) for r in c.execute(
            f"""SELECT p.name AS project_name, p.code AS project_code, COUNT(*) AS cnt,
                       COALESCE(SUM(t.hours),0) AS hours
                FROM tasks t JOIN users uu ON t.user_id = uu.id
                JOIN projects p ON t.project_id = p.id
                WHERE t.work_date BETWEEN ? AND ? {scope_filter}
                GROUP BY p.id ORDER BY cnt DESC LIMIT 15""",
            (frm, to) + scope_args,
        ).fetchall()]

        # 월간일 때만 전월 대비 비교
        prev_stats = None
        delta = None
        if period == "monthly":
            prev_first = (rd.replace(day=1) - timedelta(days=1)).replace(day=1)
            prev_last = rd.replace(day=1) - timedelta(days=1)
            prev_stats = dict(c.execute(sql, (prev_first.isoformat(), prev_last.isoformat()) + scope_args).fetchone())
            for k in prev_stats:
                prev_stats[k] = prev_stats[k] or 0
            def pct(cur, prev):
                if not prev:
                    return None
                return round((cur - prev) * 100 / prev, 1)
            delta = {
                "total": pct(stats["total"], prev_stats["total"]),
                "done": pct(stats["done"], prev_stats["done"]),
                "delay": pct(stats["delay"], prev_stats["delay"]),
                "hours": pct(stats["hours"], prev_stats["hours"]),
            }

        # 일별 추이 (월간일 때)
        daily_trend = []
        if period == "monthly":
            daily_trend = [dict(r) for r in c.execute(
                f"""SELECT t.work_date AS d, COUNT(*) AS cnt,
                           SUM(CASE WHEN t.status='완료' THEN 1 ELSE 0 END) AS done,
                           COALESCE(SUM(t.hours),0) AS hours
                    FROM tasks t JOIN users uu ON t.user_id = uu.id
                    WHERE t.work_date BETWEEN ? AND ? {scope_filter}
                    GROUP BY t.work_date ORDER BY t.work_date""",
                (frm, to) + scope_args,
            ).fetchall()]

    return ctx(req, "summary.html", user=u,
               period=period, scope=scope, ref=ref,
               period_label=period_label, scope_label=scope_label,
               frm=frm, to=to, prev_ref=prev_ref, next_ref=next_ref,
               stats=stats, completion_rate=completion_rate,
               cards=cards[:80], total_cards=len(cards),
               team_agg=team_agg, user_agg=user_agg, project_agg=project_agg,
               prev_stats=prev_stats, delta=delta, daily_trend=daily_trend,
               can_all=u["role"] in ("ceo","admin","executive"),
               can_team=u.get("team_id") is not None,
               active="summary")


# =====================================================
# COMMENTS — 업무카드 댓글/요청사항
# =====================================================
@app.get("/api/task/{tid}/comments")
async def api_list_comments(req: Request, tid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    return JSONResponse({"ok": True, "comments": get_task_comments(tid)})


@app.post("/api/task/{tid}/comment")
async def api_add_comment(req: Request, tid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    # v5H122 (2026-05-04) Task 서브 API 페이로드 검증 강화
    try:
        d = await req.json()
    except Exception:
        return JSONResponse({"error": "잘못된 요청 형식입니다 (JSON 필요)"}, 400)
    body = (d.get("body") or "").strip()
    if not body:
        return JSONResponse({"error": "댓글 내용을 입력하세요"}, 400)
    if len(body) > 2000:
        return JSONResponse({"error": f"댓글은 최대 2000자까지 입력할 수 있습니다 (현재 {len(body)}자)"}, 400)
    # task 존재 확인 — 없는 카드에 댓글 달리는 것 차단
    try:
        with db_session() as _c:
            _t = _c.execute("SELECT id FROM tasks WHERE id=?", (tid,)).fetchone()
            if not _t:
                return JSONResponse({"error": "업무 카드를 찾을 수 없습니다"}, 404)
    except Exception:
        pass  # 폴백: 기존 동작
    parent_id = d.get("parent_id") or None
    if parent_id is not None:
        try:
            parent_id = int(parent_id)
        except Exception:
            return JSONResponse({"error": "parent_id 형식 오류"}, 400)
    cid = add_comment(tid, u["id"], body, parent_id)
    return JSONResponse({"ok": True, "id": cid})


@app.delete("/api/comment/{cid}")
async def api_delete_comment(req: Request, cid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    ok = delete_comment(cid, u["id"])
    return JSONResponse({"ok": ok})


# =====================================================
# TASK DETAIL — 모달용 단일 카드 정보 (어디서든 호출)
# =====================================================
# v5H35 (2026-05-02) — 테스트용 가짜 날짜 설정 (세션 기반, ceo/admin/leader 만)
@app.post("/dev/fake-date")
async def dev_set_fake_date(req: Request):
    u = get_user(req)
    if not u or u.get("role","") not in ("ceo","admin","leader","executive"):
        return JSONResponse({"error":"권한 없음"}, 403)
    d = await req.json()
    fk = (d.get("date") or "").strip()
    if fk in ("", "clear", "today"):
        req.session.pop("fake_today", None)
        return JSONResponse({"ok": True, "active": False, "today": date.today().isoformat()})
    # YYYY-MM-DD 검증
    try:
        datetime.strptime(fk, "%Y-%m-%d")
    except Exception:
        return JSONResponse({"error":"형식 오류 (YYYY-MM-DD)"}, 400)
    req.session["fake_today"] = fk
    return JSONResponse({"ok": True, "active": True, "today": fk, "real": date.today().isoformat()})


# v5H30 (2026-05-02) — 일일업무 카드 상세 페이지 (대표 지시: 상세 클릭 시 페이지 부재)
@app.get("/task/{tid}", response_class=HTMLResponse)
async def task_detail_page(request: Request, tid: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        t = c.execute(
            """SELECT t.*, u.name AS user_name, u.rank AS user_rank,
                      p.name AS project_name, p.code AS project_code, p.id AS project_id_link,
                      cs.name AS customer_name, cs.id AS customer_id_link,
                      tm.name AS team_name
               FROM tasks t LEFT JOIN users u ON t.user_id=u.id
               LEFT JOIN projects p ON t.project_id=p.id
               LEFT JOIN customers cs ON t.customer_id=cs.id
               LEFT JOIN teams tm ON u.team_id=tm.id
               WHERE t.id=?""", (tid,)
        ).fetchone()
        if not t:
            return HTMLResponse("<h1>업무 카드를 찾을 수 없습니다</h1><a href='/daily'>← 일일업무로</a>", 404)
        task = dict(t)
        # v5H32: 동일 프로젝트(또는 라벨)의 누적 공수·카드 수 계산 — '며칠/몇주 걸리는 업무' 추적용
        cum = {"hours": 0.0, "cards": 0, "first": None, "last": None}
        if task.get("project_id"):
            row = c.execute("SELECT COALESCE(SUM(hours),0) AS h, COUNT(*) AS n, MIN(work_date) AS f, MAX(work_date) AS l FROM tasks WHERE project_id=? AND user_id=?", (task["project_id"], task["user_id"])).fetchone()
            if row: cum = {"hours": float(row["h"] or 0), "cards": int(row["n"] or 0), "first": row["f"], "last": row["l"]}
        elif task.get("project_label"):
            row = c.execute("SELECT COALESCE(SUM(hours),0) AS h, COUNT(*) AS n, MIN(work_date) AS f, MAX(work_date) AS l FROM tasks WHERE project_label=? AND user_id=?", (task["project_label"], task["user_id"])).fetchone()
            if row: cum = {"hours": float(row["h"] or 0), "cards": int(row["n"] or 0), "first": row["f"], "last": row["l"]}
    comments = get_task_comments(tid)
    reactions = get_reactions(tid)
    return ctx(request, "task_detail.html",
               user=u, active="daily",
               task=task, comments=comments, reactions=reactions, cum=cum)


@app.get("/api/task/{tid}/detail")
async def api_task_detail(req: Request, tid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error":"로그인 필요"}, 401)
    with db_session() as c:
        t = c.execute(
            """SELECT t.*, u.name AS user_name, u.rank AS user_rank,
                      p.name AS project_name, cs.name AS customer_name,
                      tm.name AS team_name
               FROM tasks t LEFT JOIN users u ON t.user_id=u.id
               LEFT JOIN projects p ON t.project_id=p.id
               LEFT JOIN customers cs ON t.customer_id=cs.id
               LEFT JOIN teams tm ON u.team_id=tm.id
               WHERE t.id=?""", (tid,)
        ).fetchone()
        if not t:
            return JSONResponse({"error":"카드 없음"}, 404)
    comments = get_task_comments(tid)
    reactions = get_reactions(tid)
    delegations = get_delegations(tid)
    return JSONResponse({"ok":True, "task":dict(t), "comments":comments,
                         "reactions":reactions, "delegations":delegations})


# =====================================================
# REACTIONS — 1-click 빠른 피드백
# =====================================================
@app.post("/api/task/{tid}/reaction")
async def api_reaction(req: Request, tid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error":"로그인 필요"}, 401)
    # v5H122 페이로드 검증
    try:
        d = await req.json()
    except Exception:
        return JSONResponse({"error":"잘못된 요청 형식입니다 (JSON 필요)"}, 400)
    kind = (d.get("kind") or "").strip()
    if not kind:
        return JSONResponse({"error":"반응 종류를 선택하세요 (ack/question/risk/ok)"}, 400)
    if kind not in ("ack", "question", "risk", "ok"):
        return JSONResponse({"error": f"지원하지 않는 반응입니다: {kind} (ack/question/risk/ok 중 하나)"}, 400)
    # task 존재 확인
    try:
        with db_session() as _c:
            if not _c.execute("SELECT id FROM tasks WHERE id=?", (tid,)).fetchone():
                return JSONResponse({"error":"업무 카드를 찾을 수 없습니다"}, 404)
    except Exception:
        pass
    res = add_reaction(tid, u["id"], kind)
    if not res:
        return JSONResponse({"error":"잘못된 반응"}, 400)
    return JSONResponse({"ok":True, "result":res, "reactions":get_reactions(tid)})


# =====================================================
# 번역 API
# =====================================================
@app.api_route("/api/set-lang", methods=["GET", "POST"])
async def api_set_lang(req: Request):
    """사용자 UI 언어 변경.
    - POST(JSON body): 기존 프런트 fetch 호출용 (base.html changeLang)
    - GET(쿼리스트링 ?lang=vi): 주소창·북마크·테스트 호출용
    """
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    # lang 값 획득: GET 은 쿼리, POST 는 JSON body
    lang = None
    if req.method == "POST":
        try:
            data = await req.json()
            lang = (data or {}).get("lang")
        except Exception:
            lang = None
    if not lang:
        lang = req.query_params.get("lang") or "ko"
    if lang not in LANGS:
        lang = "ko"
    with db_session() as c:
        c.execute("UPDATE users SET lang=? WHERE id=?", (lang, u["id"]))
    req.session["lang"] = lang
    # GET 이면 이전 페이지로 303 리다이렉트(주소창/북마크 UX), POST 는 JSON
    if req.method == "GET":
        back = req.headers.get("referer") or "/home"
        return RedirectResponse(back, status_code=303)
    return JSONResponse({"ok": True, "lang": lang})


@app.post("/api/translate")
async def api_translate(req: Request):
    """
    사내 사전 번역 (외부 API 0건).
    사이클 69 (2026-04-27): deep_translator + Google Translate URL 직접 호출 제거.
    오리엔테이션 §3.1 외부 자산 0건 정책 이행. 대표 14:19 "부분 처리" 결정.
    동작: app/i18n.py T 사전(약 472키 / ko·vi·en)에서 입력 텍스트와 매칭되는 항목을 찾아 응답.
    매칭 실패 시 "사내 사전 미등록" 안내 + 빅터AI 등록 기능(사이클 70+) 예고.
    """
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    d = await req.json()
    text = (d.get("text") or "").strip()
    target = d.get("target") or "vi"  # 기본: 베트남어
    if not text:
        return JSONResponse({"error": "텍스트 없음"}, 400)
    if target not in ("ko", "vi", "en"):
        return JSONResponse({"error": "지원 언어: ko/vi/en"}, 400)

    # 사내 사전 lookup — i18n.py T 사전 정/역방향 모두 시도
    try:
        from .i18n import T as _T
    except Exception:
        return JSONResponse({"ok": False,
                             "error": "사내 사전 모듈 로드 실패"}, 500)

    norm = text.strip()
    # 1) value(원문 문구) → 다른 언어 value 매칭
    for _key, entry in _T.items():
        if not isinstance(entry, dict):
            continue
        for _src_lang, _src_text in entry.items():
            if isinstance(_src_text, str) and _src_text.strip() == norm:
                tgt_text = entry.get(target)
                if tgt_text:
                    return JSONResponse({"ok": True,
                                         "translated": tgt_text,
                                         "target": target,
                                         "source": "사내 사전"})
    # 2) 매칭 실패 — 빅터AI 등록 안내
    return JSONResponse({
        "ok": False,
        "error": "사내 사전 미등록 — 빅터AI 단어 등록 기능(예정)으로 추가 가능합니다.",
        "source": "사내 사전",
    }, 200)


# =====================================================
# 업무 위임 (Delegation)
# =====================================================
@app.post("/api/task/{tid}/delegate")
async def api_delegate(req: Request, tid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error":"로그인 필요"}, 401)
    # v5H122 페이로드 검증
    try:
        d = await req.json()
    except Exception:
        return JSONResponse({"error":"잘못된 요청 형식입니다 (JSON 필요)"}, 400)
    to_id = d.get("to_user_id")
    msg = (d.get("message") or "").strip()
    if not to_id:
        return JSONResponse({"error":"위임 대상을 선택하세요"}, 400)
    try:
        to_id_i = int(to_id)
    except Exception:
        return JSONResponse({"error":"위임 대상 형식 오류 (사용자 ID 숫자 필요)"}, 400)
    if to_id_i == u["id"]:
        return JSONResponse({"error":"자기 자신에게는 위임할 수 없습니다"}, 400)
    if len(msg) > 2000:
        return JSONResponse({"error": f"위임 메시지는 최대 2000자까지 입력할 수 있습니다 (현재 {len(msg)}자)"}, 400)
    # task + 대상 사용자 존재 검증
    try:
        with db_session() as _c:
            if not _c.execute("SELECT id FROM tasks WHERE id=?", (tid,)).fetchone():
                return JSONResponse({"error":"업무 카드를 찾을 수 없습니다"}, 404)
            tgt = _c.execute("SELECT id, is_active FROM users WHERE id=?", (to_id_i,)).fetchone()
            if not tgt:
                return JSONResponse({"error":"위임 대상 사용자를 찾을 수 없습니다"}, 404)
            if not tgt["is_active"]:
                return JSONResponse({"error":"비활성 사용자에게는 위임할 수 없습니다"}, 400)
    except Exception:
        pass
    delegate_task(tid, u["id"], to_id_i, msg)
    return JSONResponse({"ok": True})


@app.post("/api/delegation/{did}/resolve")
async def api_resolve_delegation(req: Request, did: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error":"로그인 필요"}, 401)
    ok = resolve_delegation(did, u["id"])
    if not ok:
        return JSONResponse({"error":"본인만 완료 처리 가능"}, 403)
    return JSONResponse({"ok": True})


# =====================================================
# @멘션 자동완성
# =====================================================
@app.get("/api/users/search")
async def api_user_search(req: Request, q: str = ""):
    u = get_user(req)
    if not u:
        return JSONResponse({"error":"로그인 필요"}, 401)
    return JSONResponse({"ok":True, "users":get_user_search(q.strip(), 8)})


# =====================================================
# SIDEBAR TREE — Notion 스타일 좌측 트리 데이터
# =====================================================
@app.get("/api/sidebar-tree")
async def api_sidebar_tree(req: Request):
    """좌측 사이드바에 표시할 트리 구조:
       팀 목록 > 하위 프로젝트 > (선택시 해당 페이지 이동)
       + 오늘 카드 수, 지연 카운트"""
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    today_s = date.today().isoformat()
    with db_session() as c:
        teams = [dict(r) for r in c.execute(
            """SELECT t.id, t.name, t.code, t.is_lab, t.display_order,
                      u.name AS leader_name
               FROM teams t LEFT JOIN users u ON t.leader_id=u.id
               ORDER BY t.display_order"""
        ).fetchall()]
        for t in teams:
            stats = c.execute(
                """SELECT COUNT(*) AS total,
                          SUM(CASE WHEN tk.status='지연' THEN 1 ELSE 0 END) AS delay
                   FROM tasks tk JOIN users u ON tk.user_id=u.id
                   WHERE u.team_id=? AND tk.work_date=?""",
                (t["id"], today_s)
            ).fetchone()
            t["today_count"] = stats["total"] or 0
            t["delay_count"] = stats["delay"] or 0
            # 팀 하위 활성 프로젝트
            projects = [dict(r) for r in c.execute(
                """SELECT DISTINCT p.id, p.name, p.status
                   FROM projects p
                   JOIN tasks tk ON tk.project_id=p.id
                   JOIN users u ON tk.user_id=u.id
                   WHERE u.team_id=? AND p.status IN ('active','진행중','planning')
                   ORDER BY p.name""",
                (t["id"],)
            ).fetchall()]
            t["projects"] = projects
        # 즐겨찾기 / 내 카드 요약
        my_today = c.execute(
            "SELECT COUNT(*) FROM tasks WHERE user_id=? AND work_date=?",
            (u["id"], today_s)
        ).fetchone()[0]
    return JSONResponse({"ok": True, "teams": teams, "my_today": my_today,
                         "user_role": u["role"], "user_team_id": u.get("team_id")})


# =====================================================
# ACTIVITIES — 실시간 활동 피드
# =====================================================
@app.get("/api/activities")
async def api_activities(req: Request, scope: str = "all", since_id: int = 0):
    u = get_user(req)
    if not u:
        return JSONResponse({"error":"로그인 필요"}, 401)
    team_id = None
    actor_id = None
    if scope == "team" and u.get("team_id"):
        team_id = u["team_id"]
    elif scope == "me":
        actor_id = u["id"]
    items = get_activities(limit=80, team_id=team_id, actor_id=actor_id, since_id=since_id)
    return JSONResponse({"ok":True, "items":items})


@app.get("/now", response_class=HTMLResponse)
async def now_feed(req: Request, scope: str = "all"):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    team_id = None; actor_id = None
    if scope == "team" and u.get("team_id"):
        team_id = u["team_id"]
    elif scope == "me":
        actor_id = u["id"]
    items = get_activities(limit=80, team_id=team_id, actor_id=actor_id)
    return ctx(req, "now.html", user=u, items=items, scope=scope,
               can_team=u.get("team_id") is not None, active="now")


# =====================================================
# 통합 검색 (글로벌 검색 강화 — 2026-04-26)
# 카테고리: 카드/댓글/회고 + 수주/고객/부품/이슈/티켓/사용자/게시판/수출입/재고
# 외부 검색엔진 0건 (Elasticsearch 등 절대 금지) · LIKE parameter binding 절대.
# =====================================================
SEARCH_CAT_LABELS = {
    "tasks": "📋 카드", "comments": "💬 댓글", "retros": "📖 회고",
    "orders": "📦 수주", "customers": "🤝 고객", "parts": "🔩 부품",
    "issues": "⚠️ 이슈", "tickets": "🎫 티켓", "users": "👤 사용자",
    "boards": "📰 게시판", "exports": "🚢 수출입", "audits": "📊 재고실사",
}


@app.get("/search", response_class=HTMLResponse)
async def search_page(req: Request, q: str = "", cat: str = "all"):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    cat = (cat or "all").strip().lower()
    cats = None if cat == "all" else [cat]
    res = {"tasks":[], "comments":[], "retros":[]}
    res_global = {}
    if q and len(q.strip()) >= 1:
        # 카드/댓글/회고 (기존 search_all)
        if cat in ("all", "tasks", "comments", "retros"):
            full = search_all(q, 50)
            if cat == "all":
                res = full
            elif cat == "tasks":
                res = {"tasks": full.get("tasks", []), "comments": [], "retros": []}
            elif cat == "comments":
                res = {"tasks": [], "comments": full.get("comments", []), "retros": []}
            elif cat == "retros":
                res = {"tasks": [], "comments": [], "retros": full.get("retros", [])}
        # 글로벌 9개 카테고리
        global_cats = list(GLOBAL_SEARCH_CATEGORIES.keys())
        if cat == "all":
            res_global = global_search(q, None, 5)
        elif cat in global_cats:
            res_global = global_search(q, [cat], 5)
    return ctx(req, "search.html", user=u, q=q, cat=cat, res=res,
               res_global=res_global, cat_labels=SEARCH_CAT_LABELS, active="search")


@app.post("/search/suggest")
async def api_search_suggest(req: Request):
    """헤더 검색창 자동완성 — 카테고리별 상위 3건씩, 라벨만."""
    u = get_user(req)
    if not u:
        return JSONResponse({"error":"로그인 필요"}, 401)
    try:
        d = await req.json()
    except Exception:
        d = {}
    q = (d.get("q") or "").strip()
    if len(q) < 2:
        return JSONResponse({"ok": True, "items": []})
    items = []
    # 핵심 5개 카테고리 (자동완성 부담 최소)
    suggest_cats = ["orders", "customers", "parts", "issues", "users"]
    res = global_search(q, suggest_cats, 3)
    for cat in suggest_cats:
        for r in res.get(cat, []):
            label = ""
            if cat == "orders":
                label = f"{r.get('order_no','')} · {r.get('customer_name','') or ''}"
                link = "/sales/orders"
            elif cat == "customers":
                label = r.get("name", "")
                link = f"/customer/{r['id']}"
            elif cat == "parts":
                label = f"{r.get('part_no','')} · {r.get('part_name','')}"
                link = f"/parts/{r['id']}"
            elif cat == "issues":
                label = f"{r.get('issue_no','') or ''} · {r.get('title','')}"
                link = f"/issues/{r['id']}"
            elif cat == "users":
                label = f"{r.get('name','')} · {r.get('rank','') or r.get('team_name','') or ''}"
                link = "/search?cat=users&q=" + q
            else:
                label = str(r.get("id", ""))
                link = "/search?q=" + q
            items.append({
                "cat": cat,
                "cat_label": SEARCH_CAT_LABELS.get(cat, cat),
                "label": label.strip(" ·"),
                "link": link,
            })
    return JSONResponse({"ok": True, "q": q, "items": items[:15]})


# =====================================================
# 프로젝트 회고 (Retro)
# =====================================================
@app.post("/api/project/{pid}/retro")
async def api_retro_save(req: Request, pid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error":"로그인 필요"}, 401)
    d = await req.json()
    rid = upsert_retro(pid, u["id"],
                       (d.get("went_well") or "").strip(),
                       (d.get("went_bad") or "").strip(),
                       (d.get("next_action") or "").strip(),
                       (d.get("risk_note") or "").strip())
    log_activity_standalone(u["id"], "retro",
                            title=f"{u['name']} 프로젝트 회고 작성",
                            project_id=pid, team_id=u.get("team_id"))
    return JSONResponse({"ok":True, "id":rid})


# =====================================================
# 코크핏 (팀장/CEO 라이브 시야)
# =====================================================
@app.get("/cockpit", response_class=HTMLResponse)
async def cockpit_page(req: Request):
    # Plan Y S1 대표 승인 2026-04-24: /cockpit 은 /dashboard 와 기능 중복 → 301 합병
    # 기존 기능(조종석 지표)은 /dashboard 내 "코크핏" 탭으로 제공 예정 (S2).
    # 기존 북마크 보호를 위해 라우트는 유지하되 리다이렉트로 전환.
    return RedirectResponse("/dashboard?view=cockpit", 301)


@app.get("/_cockpit_legacy", response_class=HTMLResponse)
async def _cockpit_legacy_unused(req: Request):
    """[DEPRECATED 2026-04-24] 합병 이전 구 /cockpit 로직 — 복원용 보관."""
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if u["role"] not in ("leader","executive","ceo","admin"):
        return RedirectResponse("/dashboard", 303)
    today_iso = date.today().isoformat()
    week_ago = (date.today() - timedelta(days=7)).isoformat()
    is_global = u["role"] in ("ceo","admin","executive")
    team_filter = ""
    params = []
    if not is_global and u.get("team_id"):
        team_filter = "AND u.team_id = ?"
        params = [u["team_id"]]

    with db_session() as c:
        # 우리팀(또는 전사) 멤버
        members = c.execute(
            f"""SELECT u.id, u.name, u.rank, tm.name AS team_name,
                       (SELECT COUNT(*) FROM tasks WHERE user_id=u.id AND work_date=?) AS today_cn,
                       (SELECT COUNT(*) FROM tasks WHERE user_id=u.id AND status='지연') AS delay_cn,
                       (SELECT COUNT(*) FROM tasks WHERE user_id=u.id AND status='진행중') AS prog_cn,
                       (SELECT COUNT(*) FROM tasks t JOIN task_comments tc ON tc.task_id=t.id
                        WHERE t.user_id=u.id AND tc.is_ceo_request=1
                        AND NOT EXISTS (SELECT 1 FROM task_comments tc2
                                        WHERE tc2.task_id=tc.task_id AND tc2.id>tc.id)) AS unanswered_ceo
                FROM users u LEFT JOIN teams tm ON u.team_id=tm.id
                WHERE u.is_active=1 AND u.role NOT IN ('admin') {team_filter}
                ORDER BY tm.display_order, u.id""",
            tuple([today_iso] + params)
        ).fetchall()

        # 막힌 카드 (지연 + 코멘트 미해결)
        stuck = c.execute(
            f"""SELECT t.id, t.title, t.work_date, u.name AS owner_name, tm.name AS team_name,
                       (SELECT COUNT(*) FROM task_comments WHERE task_id=t.id) AS cn
                FROM tasks t JOIN users u ON t.user_id=u.id
                LEFT JOIN teams tm ON u.team_id=tm.id
                WHERE t.status='지연' {team_filter}
                ORDER BY t.work_date LIMIT 30""",
            tuple(params)
        ).fetchall()

        # 미작성자 (오늘)
        missing = c.execute(
            f"""SELECT u.id, u.name, u.rank, tm.name AS team_name
                FROM users u LEFT JOIN teams tm ON u.team_id=tm.id
                WHERE u.is_active=1 AND u.role NOT IN ('admin','ceo')
                AND NOT EXISTS (SELECT 1 FROM tasks WHERE user_id=u.id AND work_date=?)
                {team_filter} ORDER BY tm.display_order, u.name""",
            tuple([today_iso] + params)
        ).fetchall()

    bn = detect_bottlenecks() if is_global else []
    return ctx(req, "cockpit.html", user=u,
               members=[dict(r) for r in members],
               stuck=[dict(r) for r in stuck],
               missing=[dict(r) for r in missing],
               bottlenecks=bn,
               is_global=is_global,
               active="cockpit")


# =====================================================
# 병목 자동 감지 페이지
# =====================================================
@app.get("/bottlenecks", response_class=HTMLResponse)
async def bottlenecks_page(req: Request):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if u["role"] not in ("leader","executive","ceo","admin"):
        return RedirectResponse("/dashboard", 303)
    items = detect_bottlenecks()
    return ctx(req, "bottlenecks.html", user=u, items=items, active="bottlenecks")


# =====================================================
# NOTIFICATIONS — 알림 (벨 아이콘 + 드롭다운)
# =====================================================
@app.get("/api/notifications")
async def api_notifications(req: Request, only_unread: int = 0):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    items = get_notifications(u["id"], only_unread=bool(only_unread), limit=30)
    return JSONResponse({"ok": True, "items": items, "unread": count_unread(u["id"])})


@app.post("/api/notification/{nid}/read")
async def api_notif_read(req: Request, nid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    mark_notification_read(nid, u["id"])
    return JSONResponse({"ok": True})


@app.post("/api/notifications/read-all")
async def api_notif_read_all(req: Request):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    mark_all_read(u["id"])
    return JSONResponse({"ok": True})


@app.get("/notifications", response_class=HTMLResponse)
async def notifications_page(req: Request):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    items = get_notifications(u["id"], limit=100)
    return ctx(req, "notifications.html", user=u, items=items, active="notifications")


# 통합 알림 — 비-/api/ 경로 (사이클 2026-04-26 알림시스템-통합)
@app.post("/notifications/{nid}/read")
async def notifications_read(req: Request, nid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "auth"}, 401)
    mark_notification_read(nid, u["id"])
    return JSONResponse({"ok": True})


@app.post("/notifications/read-all")
async def notifications_read_all(req: Request):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "auth"}, 401)
    mark_all_read(u["id"])
    return JSONResponse({"ok": True})


@app.get("/notifications/badge")
async def notifications_badge(req: Request):
    """헤더 배지 카운트 (UNREAD)."""
    u = get_user(req)
    if not u:
        return JSONResponse({"count": 0})
    return JSONResponse({"count": count_unread(u["id"])})


# =====================================================
# HISTORY — 개인 히스토리 (내가 한 일 검색/조회)
# =====================================================
@app.get("/history", response_class=HTMLResponse)
async def history_page(req: Request, q: str = "", frm: str = "", to: str = "", status: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if not frm:
        frm = (date.today() - timedelta(days=30)).isoformat()
    if not to:
        to = date.today().isoformat()
    # 유효한 status만 필터 허용
    valid_statuses = {"완료", "진행중", "지연", "대기", "보류"}
    if status and status not in valid_statuses:
        status = ""
    with db_session() as c:
        sql = """SELECT t.*, p.name AS project_name, c.name AS customer_name
                 FROM tasks t
                 LEFT JOIN projects p ON t.project_id=p.id
                 LEFT JOIN customers c ON t.customer_id=c.id
                 WHERE t.user_id=? AND t.work_date>=? AND t.work_date<=?"""
        params = [u["id"], frm, to]
        if q:
            sql += " AND (t.title LIKE ? OR t.notes LIKE ?)"
            params += [f"%{q}%", f"%{q}%"]
        if status:
            sql += " AND t.status=?"
            params.append(status)
        sql += " ORDER BY t.work_date DESC, t.id DESC"
        tasks = [dict(r) for r in c.execute(sql, params).fetchall()]

        summary = c.execute(
            """SELECT COUNT(*) AS total,
                      SUM(CASE WHEN status='완료' THEN 1 ELSE 0 END) AS done,
                      COALESCE(SUM(hours),0) AS hours
               FROM tasks WHERE user_id=? AND work_date>=? AND work_date<=?""",
            (u["id"], frm, to),
        ).fetchone()
        summary = dict(summary)

        by_category = [dict(r) for r in c.execute(
            """SELECT category, COUNT(*) AS cnt, COALESCE(SUM(hours),0) AS hrs
               FROM tasks WHERE user_id=? AND work_date>=? AND work_date<=?
               GROUP BY category ORDER BY cnt DESC""",
            (u["id"], frm, to),
        ).fetchall()]

    return ctx(req, "history.html",
               user=u, tasks=tasks, q=q, frm=frm, to=to, status=status,
               summary=summary, by_category=by_category)


# =====================================================
# TEAM — 팀장 뷰 + 팀원 권한 위임 UI (Plan Y S1)
# =====================================================
@app.get("/team/permissions", response_class=HTMLResponse)
async def team_permissions_index(req: Request):
    """팀원 권한 관리 인덱스 — team_id 없는 사용자(CEO/admin/임원) 진입점.
    - leader: 본인 팀으로 redirect
    - 그 외 권한자: 14팀 카드 인덱스 표시
    """
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if u["role"] not in ("leader", "executive", "ceo", "admin"):
        return RedirectResponse("/home", 303)
    if u["role"] == "leader" and u.get("team_id"):
        return RedirectResponse(f"/team/{u['team_id']}/permissions", 303)
    # CEO/임원/admin → 첫 활성 팀으로 redirect (인덱스도 페이지 구조 동일)
    with db_session() as c:
        first = c.execute(
            "SELECT id FROM teams ORDER BY display_order LIMIT 1"
        ).fetchone()
    if first:
        return RedirectResponse(f"/team/{first['id']}/permissions", 303)
    return RedirectResponse("/home", 303)


@app.get("/team/{team_id:int}/permissions", response_class=HTMLResponse)
async def team_permissions_page(req: Request, team_id: int):
    """Plan Y S1: 팀장 권한 위임 UI — 팀원별 4~5개 권한 토글 (3 클릭 원칙).
    - 팀장: 본인 팀만 관리 가능
    - CEO/executive/admin: 전 팀 관리 가능
    """
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if u["role"] not in ("leader", "executive", "ceo", "admin"):
        return RedirectResponse("/home", 303)
    # 팀장은 본인 팀만 (감사 가드)
    if u["role"] == "leader" and u.get("team_id") != team_id:
        return RedirectResponse(f"/team/{u.get('team_id')}/permissions", 303)

    with db_session() as c:
        team = c.execute("SELECT * FROM teams WHERE id=?", (team_id,)).fetchone()
        if not team:
            return RedirectResponse("/home", 303)
        team = dict(team)
        members = [dict(r) for r in c.execute(
            """SELECT id, name, rank, role,
                      can_use_sales, can_use_logistics,
                      can_view_sales, can_view_logistics,
                      can_edit_changes, can_close_tickets, is_admin
               FROM users
               WHERE team_id=? AND is_active=1
               ORDER BY
                 CASE role WHEN 'executive' THEN 0 WHEN 'leader' THEN 1 ELSE 2 END,
                 id""",
            (team_id,)
        ).fetchall()]
        # KPI: 권한별 부여 인원 (현재 팀) — 2026-04-28 view/use 분리
        kpi = {
            "total":      len(members),
            "view_sales": sum(1 for m in members if m.get("can_view_sales") or m.get("can_use_sales")),
            "sales":      sum(1 for m in members if m.get("can_use_sales")),
            "view_logi":  sum(1 for m in members if m.get("can_view_logistics") or m.get("can_use_logistics")),
            "logi":       sum(1 for m in members if m.get("can_use_logistics")),
            "chg":        sum(1 for m in members if m.get("can_edit_changes")),
            "tkt":        sum(1 for m in members if m.get("can_close_tickets")),
            "adm":        sum(1 for m in members if m.get("is_admin")),
            "seed":       sum(1 for m in members if m.get("role") in ("ceo", "executive")),
        }
        # 최근 7일 권한 변경 이력 (notifications 에서 "권한 변경" 항목)
        try:
            recent_changes = [dict(r) for r in c.execute(
                """SELECT n.title, n.body, n.created_at,
                          u.name AS actor_name, u.rank AS actor_rank
                   FROM notifications n
                   LEFT JOIN users u ON u.id = n.user_id
                   WHERE n.title='권한 변경'
                     AND n.body LIKE ?
                     AND date(n.created_at) >= date('now', '-7 days')
                   ORDER BY n.created_at DESC LIMIT 10""",
                (f"%team_id={team_id}%",)
            ).fetchall()]
        except Exception:
            recent_changes = []
        # 전 팀 목록 (CEO/임원/admin 만)
        all_teams = []
        if u["role"] in ("ceo", "admin", "executive"):
            all_teams = [dict(r) for r in c.execute(
                """SELECT t.*, (SELECT COUNT(*) FROM users u WHERE u.team_id=t.id AND u.is_active=1) AS member_count
                   FROM teams t ORDER BY t.display_order"""
            ).fetchall()]
    is_admin_actor = u["role"] in ("ceo", "admin")
    return ctx(req, "admin_team_perms.html", user=u, active="team_perms",
               team=team, members=members, kpi=kpi,
               recent_changes=recent_changes, all_teams=all_teams,
               is_admin_actor=is_admin_actor)


@app.post("/team/{team_id:int}/permissions")
async def team_permissions_save(req: Request, team_id: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if u["role"] not in ("leader", "executive", "ceo", "admin"):
        return RedirectResponse("/home", 303)
    if u["role"] == "leader" and u.get("team_id") != team_id:
        return RedirectResponse("/home", 303)

    form = await req.form()
    is_admin_actor = u["role"] in ("ceo", "admin")
    saved = 0
    with db_session() as c:
        members = c.execute(
            "SELECT id, role FROM users WHERE team_id=? AND is_active=1", (team_id,)
        ).fetchall()
        for m in members:
            mid = m["id"]
            if not form.get(f"touch_{mid}"):
                continue
            # CEO/임원은 시드 유지 (해제 불가)
            if m["role"] in ("ceo", "executive"):
                continue
            v_sales = 1 if form.get(f"vsales_{mid}") else 0
            sales   = 1 if form.get(f"sales_{mid}")  else 0
            v_logi  = 1 if form.get(f"vlogi_{mid}")  else 0
            logi    = 1 if form.get(f"logi_{mid}")   else 0
            chg     = 1 if form.get(f"chg_{mid}")    else 0
            tkt     = 1 if form.get(f"tkt_{mid}")    else 0
            # write implies read (등록 권한 있으면 보기 자동 ON)
            if sales: v_sales = 1
            if logi:  v_logi = 1
            # is_admin 은 ceo/admin만 변경 가능
            if is_admin_actor:
                adm = 1 if form.get(f"adm_{mid}") else 0
                c.execute(
                    "UPDATE users SET can_view_sales=?, can_use_sales=?, "
                    "can_view_logistics=?, can_use_logistics=?, "
                    "can_edit_changes=?, can_close_tickets=?, is_admin=? WHERE id=?",
                    (v_sales, sales, v_logi, logi, chg, tkt, adm, mid)
                )
            else:
                c.execute(
                    "UPDATE users SET can_view_sales=?, can_use_sales=?, "
                    "can_view_logistics=?, can_use_logistics=?, "
                    "can_edit_changes=?, can_close_tickets=? WHERE id=?",
                    (v_sales, sales, v_logi, logi, chg, tkt, mid)
                )
            saved += 1
        # 감사 로그: notification 자동 기록
        try:
            c.execute(
                "INSERT INTO notifications(user_id, title, body, created_at) VALUES(?,?,?,?)",
                (u["id"], "권한 변경", f"{saved}명의 권한을 업데이트하셨습니다 (team_id={team_id})",
                 datetime.now().isoformat(timespec="seconds"))
            )
        except Exception:
            pass
    return RedirectResponse(f"/team/{team_id}/permissions?saved={saved}", 303)


@app.get("/team", response_class=HTMLResponse)
@app.get("/team/{sel_date}", response_class=HTMLResponse)
async def team_page(req: Request, sel_date: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if u["role"] not in ("leader", "executive", "ceo", "admin"):
        return RedirectResponse("/daily", 303)
    if not sel_date:
        sel_date = date.today().isoformat()

    # CEO/admin은 쿼리로 team 선택 가능, 기본 전체
    with db_session() as c:
        teams = [dict(r) for r in c.execute(
            "SELECT * FROM teams ORDER BY display_order"
        ).fetchall()]

        target_team_id = req.query_params.get("team_id")
        if u["role"] in ("leader", "executive"):
            target_team_id = u["team_id"]
        elif target_team_id:
            target_team_id = int(target_team_id)
        else:
            target_team_id = teams[0]["id"] if teams else None

        members = []
        if target_team_id:
            members = [dict(r) for r in c.execute(
                """SELECT id, name, rank, role FROM users
                   WHERE team_id=? AND is_active=1
                   ORDER BY CASE role
                        WHEN 'ceo' THEN 0 WHEN 'executive' THEN 1
                        WHEN 'leader' THEN 2 ELSE 3 END, id""",
                (target_team_id,),
            ).fetchall()]

        mids = [m["id"] for m in members]
        day_tasks = []
        week_tasks = []
        if mids:
            ph = ",".join("?" * len(mids))
            day_tasks = [dict(r) for r in c.execute(
                f"""SELECT t.*, u.name AS user_name, u.rank AS user_rank,
                           p.name AS project_name, cu.name AS customer_name
                    FROM tasks t JOIN users u ON t.user_id=u.id
                    LEFT JOIN projects p ON t.project_id=p.id
                    LEFT JOIN customers cu ON t.customer_id=cu.id
                    WHERE t.user_id IN ({ph}) AND t.work_date=?
                    ORDER BY u.id, t.status, t.id""",
                mids + [sel_date],
            ).fetchall()]

            # 이번 주
            d0 = datetime.strptime(sel_date, "%Y-%m-%d")
            mon = (d0 - timedelta(days=d0.weekday())).strftime("%Y-%m-%d")
            sun = (d0 - timedelta(days=d0.weekday()) + timedelta(days=6)).strftime("%Y-%m-%d")
            week_tasks = [dict(r) for r in c.execute(
                f"""SELECT t.*, u.name AS user_name FROM tasks t JOIN users u ON t.user_id=u.id
                    WHERE t.user_id IN ({ph}) AND t.work_date>=? AND t.work_date<=?""",
                mids + [mon, sun],
            ).fetchall()]

            # 카드 표면 메타(댓글/리액션 카운트) 일괄 조회
            meta = get_meta_bulk([t["id"] for t in day_tasks])
            for t in day_tasks:
                m = meta.get(t["id"], {})
                t["meta_comments"] = m.get("comments", 0)
                t["meta_ack"] = m.get("ack", 0)
                t["meta_question"] = m.get("question", 0)
                t["meta_risk"] = m.get("risk", 0)
                t["meta_ok"] = m.get("ok", 0)
                t["meta_last_comment"] = m.get("last_comment", "")
                t["meta_has_activity"] = bool(
                    t["meta_comments"] or t["meta_ack"] or t["meta_question"]
                    or t["meta_risk"] or t["meta_ok"]
                )

        # 팀 전체 통계 (오늘)
        stats_by_user = {}
        for m in members:
            ut = [t for t in day_tasks if t["user_id"] == m["id"]]
            stats_by_user[m["id"]] = {
                "total": len(ut),
                "done": len([t for t in ut if t["status"] == "완료"]),
                "progress": len([t for t in ut if t["status"] == "진행중"]),
                "delay": len([t for t in ut if t["status"] == "지연"]),
                "hours": sum(t["hours"] or 0 for t in ut),
                "reported": 1 if ut else 0,
            }

        team_summary = {
            "members": len(members),
            "reported": sum(1 for s in stats_by_user.values() if s["reported"]),
            "total": len(day_tasks),
            "done": len([t for t in day_tasks if t["status"] == "완료"]),
            "delay": len([t for t in day_tasks if t["status"] == "지연"]),
            "progress": len([t for t in day_tasks if t["status"] == "진행중"]),
            "week_total": len(week_tasks),
            "week_done": len([t for t in week_tasks if t["status"] == "완료"]),
        }

    prev_d = (datetime.strptime(sel_date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
    next_d = (datetime.strptime(sel_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")

    return ctx(req, "team.html",
               user=u, teams=teams, target_team_id=target_team_id,
               members=members, day_tasks=day_tasks, stats_by_user=stats_by_user,
               team_summary=team_summary, sel_date=sel_date,
               prev_date=prev_d, next_date=next_d)


# =====================================================
# DASHBOARD — 대표이사 전사 뷰
# 2026-04-26 CEO 통합 대시보드 — /ceo 별칭 추가 (CEO·admin·executive 전용)
# =====================================================
@app.get("/ceo", response_class=HTMLResponse)
async def ceo_dashboard_alias(req: Request):
    """CEO 전용 통합 대시보드 — /dashboard 로 위임 (권한 동일)."""
    return RedirectResponse("/dashboard", 303)


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard_page(req: Request):
    # ESC-02: 로그인 상태 구분 — 미인증→/login, 권한 부족→role 홈
    u_any = get_user(req)
    if not u_any:
        return RedirectResponse("/login", 303)
    u = require(req, ["ceo", "admin", "executive"])
    if not u:
        # Plan Y S1 회귀 #1: leader 가 /dashboard 직접 접근 → /team 폴백
        # (이전: role_home 호출 후 leader 도 /dashboard 로 무한 루프 가능성 존재)
        # OPS-P1-G1 [A-008]: 권한 차단 무음 → no_perm 배너 안내
        target = role_home(u_any)
        if target == "/dashboard":
            target = "/home"  # 안전 폴백
        # 폴백 대상에 ?no_perm=dashboard 부착 (이미 쿼리 있으면 & 처리)
        sep = "&" if "?" in target else "?"
        return RedirectResponse(f"{target}{sep}no_perm=dashboard", 303)
    today = date.today()
    mon = (today - timedelta(days=today.weekday())).isoformat()
    sun = (today - timedelta(days=today.weekday()) + timedelta(days=6)).isoformat()
    today_s = today.isoformat()

    with db_session() as c:
        teams = [dict(r) for r in c.execute(
            """SELECT t.*, u.name AS leader_name, u.rank AS leader_rank
               FROM teams t LEFT JOIN users u ON t.leader_id=u.id
               ORDER BY t.display_order"""
        ).fetchall()]

        for t in teams:
            mc = c.execute(
                "SELECT COUNT(*) FROM users WHERE team_id=? AND is_active=1",
                (t["id"],),
            ).fetchone()[0]
            t["member_count"] = mc
            s = c.execute(
                """SELECT COUNT(*) AS total,
                          SUM(CASE WHEN status='완료' THEN 1 ELSE 0 END) AS done,
                          SUM(CASE WHEN status='진행중' THEN 1 ELSE 0 END) AS progress,
                          SUM(CASE WHEN status='지연' THEN 1 ELSE 0 END) AS delay,
                          COALESCE(SUM(hours),0) AS hours
                   FROM tasks tk JOIN users u ON tk.user_id=u.id
                   WHERE u.team_id=? AND tk.work_date>=? AND tk.work_date<=?""",
                (t["id"], mon, sun),
            ).fetchone()
            t["week_stats"] = {k: (s[k] or 0) for k in s.keys()}
            td = c.execute(
                """SELECT COUNT(*) AS total FROM tasks tk JOIN users u ON tk.user_id=u.id
                   WHERE u.team_id=? AND tk.work_date=?""",
                (t["id"], today_s),
            ).fetchone()
            t["today_count"] = td["total"] or 0
            # 참여율: 오늘 카드 작성자 수 / 팀원 수
            rp = c.execute(
                """SELECT COUNT(DISTINCT tk.user_id) FROM tasks tk JOIN users u ON tk.user_id=u.id
                   WHERE u.team_id=? AND tk.work_date=?""",
                (t["id"], today_s),
            ).fetchone()[0]
            t["reported"] = rp
            t["participation"] = round(rp * 100 / mc) if mc else 0

        total_stats = c.execute(
            """SELECT COUNT(*) AS total,
                      SUM(CASE WHEN status='완료' THEN 1 ELSE 0 END) AS done,
                      SUM(CASE WHEN status='지연' THEN 1 ELSE 0 END) AS delay,
                      SUM(CASE WHEN status='진행중' THEN 1 ELSE 0 END) AS progress,
                      COALESCE(SUM(hours),0) AS hours
               FROM tasks WHERE work_date>=? AND work_date<=?""",
            (mon, sun),
        ).fetchone()
        total_stats = {k: (total_stats[k] or 0) for k in total_stats.keys()}

        total_users = c.execute(
            "SELECT COUNT(*) FROM users WHERE is_active=1 AND role!='admin'"
        ).fetchone()[0]
        today_reporters = c.execute(
            "SELECT COUNT(DISTINCT user_id) FROM tasks WHERE work_date=?",
            (today_s,),
        ).fetchone()[0]
        participation_rate = round(today_reporters * 100 / total_users) if total_users else 0

        # v5H50: 이번 주 지연 0건일 때 최근 30일로 자동 확장 (시드 데이터 노출 보장)
        delays = [dict(r) for r in c.execute(
            """SELECT t.*, u.name AS user_name, tm.name AS team_name,
                      p.name AS project_name, cu.name AS customer_name
               FROM tasks t JOIN users u ON t.user_id=u.id
               LEFT JOIN teams tm ON u.team_id=tm.id
               LEFT JOIN projects p ON t.project_id=p.id
               LEFT JOIN customers cu ON t.customer_id=cu.id
               WHERE t.status='지연' AND t.work_date>=? AND t.work_date<=?
               ORDER BY t.work_date DESC LIMIT 10""",
            (mon, sun),
        ).fetchall()]
        if not delays:
            from_30d = (today - timedelta(days=30)).isoformat()
            delays = [dict(r) for r in c.execute(
                """SELECT t.*, u.name AS user_name, tm.name AS team_name,
                          p.name AS project_name, cu.name AS customer_name
                   FROM tasks t JOIN users u ON t.user_id=u.id
                   LEFT JOIN teams tm ON u.team_id=tm.id
                   LEFT JOIN projects p ON t.project_id=p.id
                   LEFT JOIN customers cu ON t.customer_id=cu.id
                   WHERE t.status='지연' AND t.work_date>=?
                   ORDER BY t.work_date DESC LIMIT 10""",
                (from_30d,),
            ).fetchall()]

        customers = [dict(r) for r in c.execute(
            """SELECT cu.name AS customer_name, COUNT(*) AS cnt,
                      COALESCE(SUM(t.hours),0) AS hours
               FROM tasks t JOIN customers cu ON t.customer_id=cu.id
               WHERE t.work_date>=? AND t.work_date<=?
               GROUP BY cu.name ORDER BY cnt DESC LIMIT 10""",
            (mon, sun),
        ).fetchall()]

        # 내러티브: 팀별 진행중 핵심 카드 3건씩 + 다음 계획
        # v5H51: 오늘이 주말이거나 비어있으면 가장 최근 평일로 폴백
        narr_date = today_s
        narr_check = c.execute(
            "SELECT COUNT(*) FROM tasks WHERE work_date=? "
            "AND status IN ('진행중','지연')",
            (today_s,)
        ).fetchone()[0]
        if narr_check == 0:
            r = c.execute(
                "SELECT MAX(work_date) FROM tasks "
                "WHERE work_date<=? AND status IN ('진행중','지연')",
                (today_s,)
            ).fetchone()
            if r and r[0]:
                narr_date = r[0]
        narratives = []
        for t in teams:
            cards = [dict(r) for r in c.execute(
                """SELECT t.title, t.status, t.next_plan, u.name AS user_name, u.rank,
                          p.name AS project_name
                   FROM tasks t JOIN users u ON t.user_id=u.id
                   LEFT JOIN projects p ON t.project_id=p.id
                   WHERE u.team_id=? AND t.work_date=?
                     AND t.status IN ('진행중','지연')
                   ORDER BY CASE t.status WHEN '지연' THEN 0 ELSE 1 END, t.id
                   LIMIT 3""",
                (t["id"], narr_date),
            ).fetchall()]
            if cards:
                narratives.append({"team": t, "cards": cards})

    # 2026-04-26 CEO 통합 대시보드 — 9 KPI + 알림 + 빠른 액션
    kpis_raw = ceo_dashboard_kpis(user_id=u["id"])
    with db_session() as c:
        unread_notifs = [dict(r) for r in c.execute(
            "SELECT id, kind, title, body, link, created_at "
            "FROM notifications WHERE user_id=? AND is_read=0 "
            "ORDER BY created_at DESC LIMIT 5",
            (u["id"],),
        ).fetchall()]

    # 2026-04-27 사이클53 — 안전재고 알림 위젯 (6번째 패널)
    # recommend_reorders(limit=5) 직접 호출 → safety_top5 컨텍스트 전달
    from .database import recommend_reorders as _ceo_recommend_reorders
    try:
        safety_top5 = _ceo_recommend_reorders(limit=5) or []
    except Exception:
        safety_top5 = []

    # ─────────────────────────────────────────────────────────────
    # v5H45 (2026-05-03 대표 대시보드 비어있음 자동 수정):
    # 템플릿(dashboard.html)이 기대하는 형식과 핸들러 데이터 형식이 어긋나서
    # narratives / ceo_kpis / teams 신호등이 화면에 0건으로 노출되던 현상 해결.
    # ─────────────────────────────────────────────────────────────
    # (1) teams: 템플릿 별칭(team_id/team_name/total_members/total/done/progress/delay/signal) 부여
    for t in teams:
        ws = t.get("week_stats") or {}
        t["team_id"] = t.get("id")
        t["team_name"] = t.get("name")
        t["total_members"] = t.get("member_count", 0)
        t["total"] = ws.get("total", 0)
        t["done"] = ws.get("done", 0)
        t["progress"] = ws.get("progress", 0)
        t["delay"] = ws.get("delay", 0)
        delay_n = t["delay"] or 0
        part = t.get("participation") or 0
        if delay_n >= 3 or part < 50:
            t["signal"] = "red"
        elif delay_n >= 1 or part < 80:
            t["signal"] = "yellow"
        else:
            t["signal"] = "green"

    # (2) narratives: {team, cards} → {icon, title, body}
    narratives_view = []
    for n in narratives:
        tm = n.get("team") or {}
        cards = n.get("cards") or []
        if not cards:
            continue
        delay_in = sum(1 for c in cards if c.get("status") == "지연")
        icon = "🔴" if delay_in else "🟡"
        title = f"{tm.get('name','')} 팀 · 진행 {len(cards)-delay_in}건 / 지연 {delay_in}건"
        body_parts = []
        for cd in cards[:3]:
            who = cd.get("user_name", "")
            rk = cd.get("rank") or ""
            tit = (cd.get("title") or "").strip()
            np = (cd.get("next_plan") or "").strip()
            chunk = f"[{who} {rk}] {tit}" if rk else f"[{who}] {tit}"
            if np:
                chunk += f" → {np[:40]}"
            body_parts.append(chunk)
        narratives_view.append({"icon": icon, "title": title,
                                "body": " · ".join(body_parts)})

    # 데이터가 한 건도 없을 때는 친절한 빈 상태 안내(빈 페이지 방지)
    if not narratives_view:
        narratives_view = [{
            "icon": "💡",
            "title": "오늘의 인사이트",
            "body": "오늘 진행 중·지연 업무가 보고되지 않았습니다. 좌측 사이드바 ‘오늘의 업무’에서 카드를 작성하면 이 영역에 자동으로 요약됩니다."
        }]

    # (3) ceo_kpis dict → 카드 리스트 (label/value/trend/note)
    def _won(n):
        try:
            n = float(n or 0)
        except Exception:
            return "0"
        if n >= 1e8:
            return f"{n/1e8:,.1f}억"
        if n >= 1e4:
            return f"{n/1e4:,.0f}만"
        return f"{n:,.0f}"

    s = kpis_raw.get("sales", {})
    st = kpis_raw.get("stock", {})
    ex = kpis_raw.get("exports", {})
    qm = kpis_raw.get("qms", {})
    wk = kpis_raw.get("weekly", {})
    gt = kpis_raw.get("gantt", {})
    fx = kpis_raw.get("fx", {})
    growth = s.get("growth", 0) or 0
    rate = wk.get("rate", 0) or 0
    ceo_kpis_view = [
        {"label": "이번달 매출", "value": f"{_won(s.get('month',0))}원",
         "trend": "up" if growth > 0 else ("alert" if growth < 0 else "neutral"),
         "note": f"전월비 {growth:+.1f}%" if s.get("prev_month") else "전월 데이터 없음"},
        {"label": "미수금 잔액", "value": f"{_won(s.get('unpaid',0))}원",
         "trend": "alert" if (s.get("unpaid") or 0) > 0 else "neutral",
         "note": "발행 INVOICE − 수금"},
        {"label": "활성 재고 품목", "value": f"{int(st.get('on_hand_kinds',0)):,}",
         "trend": "alert" if (st.get("qc_open") or 0) > 0 else "neutral",
         "note": f"QC 부적합 {int(st.get('qc_open',0))}건"},
        {"label": "수출 오더 진행", "value": f"{int(ex.get('in_progress',0)):,}건",
         "trend": "up" if (ex.get("ship_soon") or 0) > 0 else "neutral",
         "note": f"7일내 출하 {int(ex.get('ship_soon',0))}건"},
        {"label": "품질 미해결", "value": f"{int(qm.get('open',0)):,}건",
         "trend": "alert" if (qm.get("sla_violation") or 0) > 0 else "neutral",
         "note": f"SLA위반 {int(qm.get('sla_violation',0))}건"},
        {"label": "이번주 완료율", "value": f"{int(rate)}%",
         "trend": "up" if rate >= 70 else ("alert" if rate < 50 else "neutral"),
         "note": f"완료 {int(wk.get('completed',0))} · 지연 {int(wk.get('delayed',0))}"},
        {"label": "지연 프로젝트", "value": f"{int(gt.get('delayed_projects',0)):,}건",
         "trend": "alert" if (gt.get("delayed_projects") or 0) > 0 else "neutral",
         "note": "end_date 경과"},
        {"label": "USD 환율", "value": f"{(fx.get('usd_rate') or 0):,.1f}원",
         "trend": "neutral",
         "note": f"활성 알림 {int(fx.get('alerts_active',0))}건"},
    ]

    return ctx(req, "dashboard.html",
               user=u, teams=teams, total_stats=total_stats,
               mon=mon, sun=sun, today_s=today_s,
               participation_rate=participation_rate,
               today_reporters=today_reporters, total_users=total_users,
               delays=delays, customers=customers,
               narratives=narratives_view,
               ceo_kpis=ceo_kpis_view, unread_notifs=unread_notifs,
               safety_top5=safety_top5)


# =====================================================
# TEAM DAILY SUMMARY — 팀장 "오늘의 한 줄"
# =====================================================
@app.post("/api/team-summary")
async def api_team_summary(req: Request):
    u = require(req, ["leader", "executive", "ceo", "admin"])
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    d = await req.json()
    tid = d.get("team_id") or u.get("team_id")
    wdate = d.get("work_date") or date.today().isoformat()
    headline = (d.get("headline") or "").strip()
    notes = d.get("notes") or ""
    if not headline or not tid:
        return JSONResponse({"error": "내용/팀 필수"}, 400)
    with db_session() as c:
        ex = c.execute(
            "SELECT id FROM team_summaries WHERE team_id=? AND work_date=?",
            (tid, wdate),
        ).fetchone()
        if ex:
            c.execute(
                """UPDATE team_summaries SET headline=?, notes=?, author_id=?,
                       updated_at=datetime('now','localtime') WHERE id=?""",
                (headline, notes, u["id"], ex["id"]),
            )
        else:
            c.execute(
                """INSERT INTO team_summaries(team_id, work_date, author_id, headline, notes)
                   VALUES(?,?,?,?,?)""",
                (tid, wdate, u["id"], headline, notes),
            )
    return JSONResponse({"ok": True})


# =====================================================
# WEEKLY — 주간 자동 요약
# =====================================================
# v5H46: /weekly/team 별칭 — /weekly/{wk_mon}가 'team'을 날짜로 파싱하던 충돌 해결
@app.get("/weekly/team", response_class=HTMLResponse)
async def weekly_team_alias(req: Request):
    return RedirectResponse("/weekly", 303)


@app.get("/weekly", response_class=HTMLResponse)
@app.get("/weekly/{wk_mon}", response_class=HTMLResponse)
async def weekly_page(req: Request, wk_mon: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if not wk_mon or wk_mon == "team":
        td = date.today()
        wk_mon = (td - timedelta(days=td.weekday())).isoformat()
    try:
        mon = datetime.strptime(wk_mon, "%Y-%m-%d").date()
    except ValueError:
        td = date.today()
        mon = td - timedelta(days=td.weekday())
    sun = mon + timedelta(days=6)
    prev_mon = (mon - timedelta(days=7)).isoformat()
    next_mon = (mon + timedelta(days=7)).isoformat()

    with db_session() as c:
        # 개인 요약
        my_tasks = [dict(r) for r in c.execute(
            """SELECT t.*, p.name AS project_name, cu.name AS customer_name
               FROM tasks t LEFT JOIN projects p ON t.project_id=p.id
               LEFT JOIN customers cu ON t.customer_id=cu.id
               WHERE t.user_id=? AND t.work_date>=? AND t.work_date<=?
               ORDER BY t.work_date, t.id""",
            (u["id"], mon.isoformat(), sun.isoformat()),
        ).fetchall()]
        # 주간보고 2차 — KPI 8개 (오버타임/정시완료율 포함)
        _hours_total = sum(t["hours"] or 0 for t in my_tasks)
        _done_cnt = sum(1 for t in my_tasks if t["status"] == "완료")
        _overtime = max(0.0, _hours_total - 40.0)  # 주 40h 초과분
        _on_time_rate = round((_done_cnt / len(my_tasks) * 100), 0) if my_tasks else 0
        my_stats = {
            "total": len(my_tasks),
            "done": _done_cnt,
            "progress": sum(1 for t in my_tasks if t["status"] == "진행중"),
            "delay": sum(1 for t in my_tasks if t["status"] == "지연"),
            "hours": round(_hours_total, 1),
            "overtime": round(_overtime, 1),
            "on_time_rate": int(_on_time_rate),
        }
        # 분류별
        my_by_cat = {}
        for t in my_tasks:
            k = t["category"] or "기타"
            my_by_cat.setdefault(k, {"cnt": 0, "hours": 0})
            my_by_cat[k]["cnt"] += 1
            my_by_cat[k]["hours"] += t["hours"] or 0

        # 팀 요약 (팀장/임원/CEO/admin만)
        team_data = None
        if u["role"] in ("leader", "executive", "ceo", "admin"):
            tid = req.query_params.get("team_id")
            if u["role"] in ("leader", "executive"):
                tid = u["team_id"]
            elif tid:
                tid = int(tid)
            if tid:
                t_row = c.execute("SELECT * FROM teams WHERE id=?", (tid,)).fetchone()
                members = [dict(r) for r in c.execute(
                    "SELECT id, name, rank FROM users WHERE team_id=? AND is_active=1", (tid,)
                ).fetchall()]
                mids = [m["id"] for m in members]
                t_tasks = []
                if mids:
                    ph = ",".join("?" * len(mids))
                    t_tasks = [dict(r) for r in c.execute(
                        f"""SELECT tk.*, u.name AS user_name, p.name AS project_name,
                                  cu.name AS customer_name
                           FROM tasks tk JOIN users u ON tk.user_id=u.id
                           LEFT JOIN projects p ON tk.project_id=p.id
                           LEFT JOIN customers cu ON tk.customer_id=cu.id
                           WHERE tk.user_id IN ({ph}) AND tk.work_date>=? AND tk.work_date<=?
                           ORDER BY u.id, tk.work_date""",
                        mids + [mon.isoformat(), sun.isoformat()],
                    ).fetchall()]
                # 팀원별 집계
                per_user = {}
                for m in members:
                    per_user[m["id"]] = {"name": m["name"], "rank": m["rank"],
                                          "total": 0, "done": 0, "hours": 0, "tasks": []}
                for t in t_tasks:
                    per_user.setdefault(t["user_id"], {"name": t["user_name"], "rank": "",
                                                        "total": 0, "done": 0, "hours": 0, "tasks": []})
                    per_user[t["user_id"]]["total"] += 1
                    if t["status"] == "완료":
                        per_user[t["user_id"]]["done"] += 1
                    per_user[t["user_id"]]["hours"] += t["hours"] or 0
                    per_user[t["user_id"]]["tasks"].append(t)
                # 프로젝트별
                pj_agg = {}
                for t in t_tasks:
                    k = t["project_name"] or "(기타)"
                    pj_agg.setdefault(k, {"cnt": 0, "hours": 0, "done": 0})
                    pj_agg[k]["cnt"] += 1
                    pj_agg[k]["hours"] += t["hours"] or 0
                    if t["status"] == "완료":
                        pj_agg[k]["done"] += 1
                # 고객사별
                cu_agg = {}
                for t in t_tasks:
                    k = t["customer_name"] or "(기타)"
                    cu_agg.setdefault(k, {"cnt": 0, "hours": 0})
                    cu_agg[k]["cnt"] += 1
                    cu_agg[k]["hours"] += t["hours"] or 0
                team_data = {
                    "team": dict(t_row),
                    "members": members,
                    "per_user": per_user,
                    "pj_agg": sorted(pj_agg.items(), key=lambda x: -x[1]["cnt"])[:15],
                    "cu_agg": sorted(cu_agg.items(), key=lambda x: -x[1]["cnt"])[:15],
                    "total": len(t_tasks),
                    "done": sum(1 for t in t_tasks if t["status"] == "완료"),
                    "delay": sum(1 for t in t_tasks if t["status"] == "지연"),
                    "hours": round(sum(t["hours"] or 0 for t in t_tasks), 1),
                }

        # 전사 요약 (CEO/admin/executive)
        all_data = None
        if u["role"] in ("ceo", "admin", "executive"):
            all_tasks = [dict(r) for r in c.execute(
                """SELECT tk.*, u.name AS user_name, tm.name AS team_name, tm.code AS team_code,
                          p.name AS project_name, cu.name AS customer_name
                   FROM tasks tk JOIN users u ON tk.user_id=u.id
                   LEFT JOIN teams tm ON u.team_id=tm.id
                   LEFT JOIN projects p ON tk.project_id=p.id
                   LEFT JOIN customers cu ON tk.customer_id=cu.id
                   WHERE tk.work_date>=? AND tk.work_date<=?""",
                (mon.isoformat(), sun.isoformat()),
            ).fetchall()]
            by_team = {}
            for t in all_tasks:
                k = t["team_name"] or "(미배정)"
                by_team.setdefault(k, {"code": t["team_code"], "cnt": 0, "done": 0, "hours": 0})
                by_team[k]["cnt"] += 1
                if t["status"] == "완료":
                    by_team[k]["done"] += 1
                by_team[k]["hours"] += t["hours"] or 0
            by_cust = {}
            for t in all_tasks:
                if not t["customer_name"]:
                    continue
                by_cust.setdefault(t["customer_name"], {"cnt": 0, "hours": 0})
                by_cust[t["customer_name"]]["cnt"] += 1
                by_cust[t["customer_name"]]["hours"] += t["hours"] or 0
            all_data = {
                "total": len(all_tasks),
                "done": sum(1 for t in all_tasks if t["status"] == "완료"),
                "delay": sum(1 for t in all_tasks if t["status"] == "지연"),
                "hours": round(sum(t["hours"] or 0 for t in all_tasks), 1),
                "by_team": sorted(by_team.items(), key=lambda x: -x[1]["cnt"]),
                "by_cust": sorted(by_cust.items(), key=lambda x: -x[1]["cnt"])[:10],
            }

        teams_all = [dict(r) for r in c.execute(
            "SELECT * FROM teams ORDER BY display_order"
        ).fetchall()]

        # 갭서베이 Top10 #5 — 개인 8주 트렌드 (주간보고 2차 — 4→8주 확장 · VIEW 자동 집계)
        my_trend = []
        for i in range(7, -1, -1):
            wk_s = (mon - timedelta(days=7 * i)).isoformat()
            row = c.execute(
                """SELECT total_tasks, completed, in_progress, delayed, total_hours
                   FROM weekly_summary
                   WHERE user_id=? AND week_start=?""",
                (u["id"], wk_s),
            ).fetchone()
            if row:
                my_trend.append({
                    "week_start": wk_s,
                    "total": row["total_tasks"],
                    "done": row["completed"],
                    "progress": row["in_progress"],
                    "delay": row["delayed"],
                    "hours": row["total_hours"] or 0,
                })
            else:
                my_trend.append({"week_start": wk_s, "total": 0, "done": 0,
                                 "progress": 0, "delay": 0, "hours": 0})

        # 부서별 비교 (VIEW 자동 집계 · 팀장+ 권한일 때만 의미)
        dept_compare = []
        if u["role"] in ("leader", "executive", "ceo", "admin"):
            dept_compare = [dict(r) for r in c.execute(
                """SELECT tm.id AS team_id, tm.name AS team_name, tm.code AS team_code,
                          COALESCE(SUM(ws.total_tasks), 0)  AS total,
                          COALESCE(SUM(ws.completed), 0)    AS done,
                          COALESCE(SUM(ws.delayed), 0)      AS delay,
                          COALESCE(SUM(ws.total_hours), 0)  AS hours
                   FROM teams tm
                   LEFT JOIN weekly_summary ws
                          ON ws.team_id = tm.id AND ws.week_start = ?
                   GROUP BY tm.id, tm.name, tm.code
                   ORDER BY tm.display_order""",
                (mon.isoformat(),),
            ).fetchall()]

    return ctx(req, "weekly.html",
               user=u, my_tasks=my_tasks, my_stats=my_stats, my_by_cat=my_by_cat,
               team_data=team_data, all_data=all_data,
               my_trend=my_trend, dept_compare=dept_compare,
               wk_mon=mon.isoformat(), wk_sun=sun.isoformat(),
               prev_mon=prev_mon, next_mon=next_mon, teams_all=teams_all,
               active="weekly")


# 팀장 전용 — 부서별 집계 보기 (VIEW 자동 집계)
@app.get("/weekly/team", response_class=HTMLResponse)
async def weekly_team_page(req: Request, wk_mon: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if u["role"] not in ("leader", "executive", "ceo", "admin"):
        return RedirectResponse("/weekly", 303)
    if not wk_mon:
        td = date.today()
        wk_mon = (td - timedelta(days=td.weekday())).isoformat()
    return RedirectResponse(f"/weekly/{wk_mon}?scope=team", 303)


# 경영진 전용 — 전사 집계 보기
@app.get("/weekly/company", response_class=HTMLResponse)
async def weekly_company_page(req: Request, wk_mon: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if u["role"] not in ("ceo", "admin", "executive"):
        return RedirectResponse("/weekly", 303)
    if not wk_mon:
        td = date.today()
        wk_mon = (td - timedelta(days=td.weekday())).isoformat()
    return RedirectResponse(f"/weekly/{wk_mon}?scope=company", 303)


# 수동 재집계 트리거 (VIEW 기반이라 SQLite 의 ANALYZE 만 호출 — 캐시 갱신용)
@app.post("/weekly/refresh")
async def weekly_refresh(req: Request):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "auth"}, 401)
    with db_session() as c:
        try:
            c.execute("ANALYZE weekly_summary")
        except Exception:
            pass
    # OPS-P1-A7 [D-017]: 클라이언트 캐시 무효화 헤더
    return JSONResponse(
        {"ok": True, "msg": "재집계 완료"},
        headers={"Cache-Control": "no-cache, no-store, must-revalidate",
                 "Pragma": "no-cache", "Expires": "0"},
    )


# 주간보고 2차 — 마감 알림 (토 18시 발동 가정)
@app.post("/weekly/notify")
async def weekly_notify(req: Request):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "auth"}, 401)
    td = date.today()
    wk_mon = (td - timedelta(days=td.weekday())).isoformat()
    sun = (td - timedelta(days=td.weekday()) + timedelta(days=6)).isoformat()
    sent = 0
    with db_session() as c:
        # 본인 + 부서장 두 사람에게 INSERT
        targets = [u["id"]]
        if u.get("team_id"):
            row = c.execute("SELECT leader_id FROM teams WHERE id=?", (u["team_id"],)).fetchone()
            if row and row["leader_id"] and row["leader_id"] != u["id"]:
                targets.append(row["leader_id"])
    # 알림시스템 통합 (사이클 2026-04-26) — notify_user 단일 헬퍼 사용 (1시간 중복 방지 내장)
    for uid in targets:
        if notify_user(
            uid, "WEEKLY",
            f"📊 주간보고 마감 임박 ({wk_mon}~{sun})",
            body=wk_mon, link=f"/weekly/{wk_mon}",
        ):
            sent += 1
    return JSONResponse({"ok": True, "sent": sent, "wk_mon": wk_mon})


# 주간보고 2차 — CSV 다운로드 (csv 모듈만 · 외부 라이브러리 0)
@app.get("/weekly/export.csv")
async def weekly_export_csv(req: Request, wk_mon: str = "", scope: str = "me"):
    import csv as _csv
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if not wk_mon:
        td = date.today()
        wk_mon = (td - timedelta(days=td.weekday())).isoformat()
    mon = datetime.strptime(wk_mon, "%Y-%m-%d").date()
    sun = mon + timedelta(days=6)

    buf = io.StringIO()
    buf.write("﻿")  # UTF-8 BOM (엑셀 한글 호환)
    w = _csv.writer(buf)
    w.writerow(["일자", "이름", "팀", "제목", "분류", "프로젝트", "고객사", "상태", "공수"])
    with db_session() as c:
        if scope == "team" and u["role"] in ("leader", "executive", "ceo", "admin"):
            tid = u["team_id"]
            rows = c.execute(
                """SELECT t.work_date, u.name AS uname, tm.name AS tname,
                          t.title, t.category, p.name AS pname, cu.name AS cname,
                          t.status, t.hours
                   FROM tasks t JOIN users u ON t.user_id=u.id
                   LEFT JOIN teams tm ON u.team_id=tm.id
                   LEFT JOIN projects p ON t.project_id=p.id
                   LEFT JOIN customers cu ON t.customer_id=cu.id
                   WHERE u.team_id=? AND t.work_date>=? AND t.work_date<=?
                   ORDER BY t.work_date, u.name""",
                (tid, mon.isoformat(), sun.isoformat())
            ).fetchall()
        else:
            rows = c.execute(
                """SELECT t.work_date, u.name AS uname, tm.name AS tname,
                          t.title, t.category, p.name AS pname, cu.name AS cname,
                          t.status, t.hours
                   FROM tasks t JOIN users u ON t.user_id=u.id
                   LEFT JOIN teams tm ON u.team_id=tm.id
                   LEFT JOIN projects p ON t.project_id=p.id
                   LEFT JOIN customers cu ON t.customer_id=cu.id
                   WHERE t.user_id=? AND t.work_date>=? AND t.work_date<=?
                   ORDER BY t.work_date""",
                (u["id"], mon.isoformat(), sun.isoformat())
            ).fetchall()
        for r in rows:
            w.writerow([r["work_date"], r["uname"], r["tname"] or "",
                        r["title"], r["category"] or "", r["pname"] or "",
                        r["cname"] or "", r["status"], r["hours"] or 0])
    fn = f"weekly_{scope}_{wk_mon}.csv"
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode("utf-8")),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={fn}"}
    )


# 주간보고 2차 — 두 주 비교 (선택)
@app.get("/weekly/compare/{wk1}/{wk2}", response_class=HTMLResponse)
async def weekly_compare(req: Request, wk1: str, wk2: str):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        def _agg(wk):
            row = c.execute(
                """SELECT total_tasks, completed, in_progress, delayed, total_hours
                   FROM weekly_summary
                   WHERE user_id=? AND week_start=?""",
                (u["id"], wk)
            ).fetchone()
            if row:
                return {"wk": wk, "total": row["total_tasks"], "done": row["completed"],
                        "progress": row["in_progress"], "delay": row["delayed"],
                        "hours": row["total_hours"] or 0}
            return {"wk": wk, "total": 0, "done": 0, "progress": 0, "delay": 0, "hours": 0}
        a, b = _agg(wk1), _agg(wk2)
    diff = {
        "total": b["total"] - a["total"],
        "done": b["done"] - a["done"],
        "delay": b["delay"] - a["delay"],
        "hours": round(b["hours"] - a["hours"], 1),
    }
    html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><title>주간 비교 {wk1} vs {wk2}</title>
<style>body{{font-family:'맑은 고딕',sans-serif;padding:24px;color:#333}}
table{{border-collapse:collapse;margin-top:12px}}th,td{{padding:8px 14px;border:1px solid #d0d0d0}}
th{{background:#A5282C;color:#fff}}.dn{{color:#16a34a}}.up{{color:#dc2626}}</style></head>
<body><h1>주간 비교 — {u['name']}</h1>
<p>{wk1} → {wk2}</p>
<table><tr><th>항목</th><th>{wk1}</th><th>{wk2}</th><th>차이</th></tr>
<tr><td>총 카드</td><td>{a['total']}</td><td>{b['total']}</td><td class="{'up' if diff['total']>0 else 'dn'}">{diff['total']:+d}</td></tr>
<tr><td>완료</td><td>{a['done']}</td><td>{b['done']}</td><td class="{'up' if diff['done']>0 else 'dn'}">{diff['done']:+d}</td></tr>
<tr><td>지연</td><td>{a['delay']}</td><td>{b['delay']}</td><td class="{'dn' if diff['delay']<0 else 'up'}">{diff['delay']:+d}</td></tr>
<tr><td>공수(h)</td><td>{a['hours']:.1f}</td><td>{b['hours']:.1f}</td><td>{diff['hours']:+.1f}</td></tr>
</table>
<p><a href="/weekly">← 주간 요약</a></p></body></html>"""
    return HTMLResponse(html)


# =====================================================
# PROJECT DETAIL
# =====================================================
@app.get("/project/{pid}", response_class=HTMLResponse)
async def project_detail(req: Request, pid: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        p = c.execute(
            """SELECT p.*, cu.name AS customer_name FROM projects p
               LEFT JOIN customers cu ON p.customer_id=cu.id WHERE p.id=?""",
            (pid,),
        ).fetchone()
        if not p:
            return RedirectResponse("/", 303)
        p = dict(p)
        # 참여 전체 카드
        tasks = [dict(r) for r in c.execute(
            """SELECT tk.*, u.name AS user_name, u.rank AS user_rank,
                      tm.name AS team_name, tm.code AS team_code
               FROM tasks tk JOIN users u ON tk.user_id=u.id
               LEFT JOIN teams tm ON u.team_id=tm.id
               WHERE tk.project_id=? ORDER BY tk.work_date DESC, tk.id DESC""",
            (pid,),
        ).fetchall()]
        stats = {
            "total": len(tasks),
            "done": sum(1 for t in tasks if t["status"] == "완료"),
            "progress": sum(1 for t in tasks if t["status"] == "진행중"),
            "delay": sum(1 for t in tasks if t["status"] == "지연"),
            "hours": round(sum(t["hours"] or 0 for t in tasks), 1),
        }
        # 팀별 집계
        by_team = {}
        for t in tasks:
            k = t["team_name"] or "(미배정)"
            by_team.setdefault(k, {"cnt": 0, "hours": 0, "members": set()})
            by_team[k]["cnt"] += 1
            by_team[k]["hours"] += t["hours"] or 0
            by_team[k]["members"].add(t["user_name"])
        by_team_list = [(k, {"cnt": v["cnt"], "hours": v["hours"], "members": sorted(v["members"])})
                        for k, v in sorted(by_team.items(), key=lambda x: -x[1]["cnt"])]
        # 참여자 개별
        by_user = {}
        for t in tasks:
            k = f"{t['user_name']} ({t['team_name'] or '-'})"
            by_user.setdefault(k, {"cnt": 0, "hours": 0, "last": None})
            by_user[k]["cnt"] += 1
            by_user[k]["hours"] += t["hours"] or 0
            if not by_user[k]["last"] or t["work_date"] > by_user[k]["last"]:
                by_user[k]["last"] = t["work_date"]
        by_user_list = sorted(by_user.items(), key=lambda x: -x[1]["cnt"])[:20]

        # 타임라인용 — 일자별 그룹화
        timeline = {}
        for t in tasks:
            timeline.setdefault(t["work_date"], []).append(t)
        timeline_list = sorted(timeline.items(), reverse=True)
        # 댓글 통합
        all_comments = [dict(r) for r in c.execute(
            """SELECT tc.*, u.name AS author_name, u.rank AS author_rank,
                      tk.title AS task_title, tk.id AS tid
               FROM task_comments tc
               JOIN tasks tk ON tc.task_id=tk.id
               JOIN users u ON tc.author_id=u.id
               WHERE tk.project_id=? ORDER BY tc.created_at DESC LIMIT 30""",
            (pid,)
        ).fetchall()]
    retro = get_retro(pid)
    # v5H68: 연결된 수주(SO) 목록
    # v5H94: 페이지 로드 시 projects.order_amount 자동 자가치유
    #   = SUM(orders.total_amount). SO 있으면 합계가 단일 진실 소스.
    # v5H101: due_date / order_date 도 SO 기준으로 자가치유
    #   - due_date  = MAX(orders.due_date)  (가장 늦은 납기 = 전체 일정 기준)
    #   - order_date = MIN(orders.order_date) (최초 발주일)
    project_orders = []
    try:
        with db_session() as c2:
            # v5H104: 각 SO 의 unit_qty 를 실제 items count 와 일치시킴
            #   (qty=2 인데 items=1 인 모순 데이터 자동 정리)
            try:
                so_ids = [r[0] for r in c2.execute(
                    "SELECT id FROM orders WHERE project_id=?", (pid,)
                ).fetchall()]
                for _oid in so_ids:
                    _ic = c2.execute(
                        "SELECT COUNT(*) FROM order_items WHERE order_id=?", (_oid,)
                    ).fetchone()[0] or 0
                    if _ic > 0:
                        # items 가 있으면 그 개수에 맞추기
                        c2.execute(
                            "UPDATE orders SET unit_qty=? WHERE id=? AND COALESCE(unit_qty,1) <> ?",
                            (_ic, _oid, _ic)
                        )
                    # items 가 0 이면 건드리지 않음 (사용자가 추가 발주만 받고
                    # 아직 호기 라인 안 쪼갠 케이스 보존)
                # SO total_amount 도 items_sum 과 정합 (items 가 있을 때만)
                import re as _re
                _hogi_re = _re.compile(r"^\d+호기$")
                for _oid in so_ids:
                    row = c2.execute(
                        "SELECT COALESCE(SUM(amount),0), COUNT(*) "
                        "FROM order_items WHERE order_id=?", (_oid,)
                    ).fetchone()
                    _isum = float(row[0] or 0)
                    _icnt = int(row[1] or 0)
                    if _icnt > 0:
                        cur_t = c2.execute(
                            "SELECT total_amount FROM orders WHERE id=?", (_oid,)
                        ).fetchone()
                        if cur_t and abs(float(cur_t[0] or 0) - _isum) > 0.5:
                            c2.execute(
                                "UPDATE orders SET total_amount=? WHERE id=?",
                                (_isum, _oid)
                            )

                # v5H109: 라벨 자동 재번호 — 관리코드(프로젝트) 전체 기준
                # 같은 관리코드 안에서 SO 가 달라도 호기는 1·2·3·... 연속 번호
                try:
                    all_items = c2.execute(
                        "SELECT oi.id, oi.unit_label, oi.order_id "
                        "FROM order_items oi JOIN orders o ON o.id = oi.order_id "
                        "WHERE o.project_id=? "
                        "ORDER BY o.order_date ASC, o.id ASC, oi.id ASC",
                        (pid,)
                    ).fetchall()
                    labels = [(r[0], (r[1] or ""), r[2]) for r in all_items]
                    all_pat = labels and all(_hogi_re.match(lbl) for _, lbl, _ in labels)
                    need_fix = any(
                        lbl != f"{i+1}호기"
                        for i, (_, lbl, _) in enumerate(labels)
                    )
                    if all_pat and need_fix:
                        for i, (iid, _, _) in enumerate(labels):
                            c2.execute(
                                "UPDATE order_items SET unit_label=? WHERE id=?",
                                (f"{i+1}호기", iid)
                            )
                        # orders.unit_label 도 SO 별 새 라벨 합쳐 갱신
                        for _oid in so_ids:
                            sub = [
                                f"{i+1}호기"
                                for i, (_, _, oord) in enumerate(labels) if oord == _oid
                            ]
                            if sub:
                                c2.execute(
                                    "UPDATE orders SET unit_label=? WHERE id=?",
                                    (" · ".join(sub), _oid)
                                )
                except Exception:
                    pass
            except Exception:
                pass

            # v5H130: WON status + SO 0건 + order_amount > 0 → 백업 안전망
            #   인라인/폼 경로 모두 놓친 케이스를 상세 페이지 진입 시 보강.
            #   사용자 시나리오(009T2605)에서 v5H87 quick-status 누락으로
            #   SO가 발행 안 된 상태가 발견된 결함의 최후 차단막.
            try:
                _pre_sos = c2.execute(
                    "SELECT 1 FROM orders WHERE project_id=? LIMIT 1", (pid,)
                ).fetchone()
                _p_status = (p.get("status") if isinstance(p, dict) else p["status"]) or ""
                _p_amt = float(p.get("order_amount") or 0) if isinstance(p, dict) else float(p["order_amount"] or 0)
                # v5H142: NEW_EQUIP 만 자동 SO 발행 (소모품/수리는 별도 도메인 사용)
                _ptype_h = (p.get("project_type") if isinstance(p, dict) else (p["project_type"] if "project_type" in p.keys() else "")) or "NEW_EQUIP"
                if (not _pre_sos and _p_status in _logi.WON_STATUSES and _p_amt > 0
                    and _ptype_h == "NEW_EQUIP"):
                    res = _pwf.confirm_order_multi(
                        c2, int(pid),
                        units=[{
                            "label": _logi.project_unit_label(_ptype_h, 1),
                            "amount": _p_amt,
                            "due_date": (p.get("due_date") if isinstance(p, dict) else p["due_date"]) or "",
                            "ship_to": "",
                            "note": "",
                        }],
                        order_date=(p.get("order_date") if isinstance(p, dict) else p["order_date"]) or "",
                        created_by=u.get("id") or 0,
                        po_number=(p.get("customer_po") if isinstance(p, dict) else (p["customer_po"] if "customer_po" in p.keys() else "")) or "",
                    )
                    if res and res.get("ok"):
                        grp = (res.get("groups") or [{}])[0]
                        _auto_no = grp.get("so_no") or res.get("so_no") or ""
                        _logi.log_project_change(
                            c2, pid, u.get("id"), "수주발행(자동)",
                            "", _auto_no or _logi.project_unit_label(_ptype_h, 1),
                            note=f"v5H130 자가치유 — 상세 진입 시 누락된 SO 자동 발행 ({_p_amt:,.0f})"
                        )
            except Exception:
                pass

            project_orders = _pwf.get_project_orders(c2, pid)
            if project_orders:
                _so_sum = sum(float(o.get("total_amount") or 0) for o in project_orders)
                _curr = float(p.get("order_amount") or 0) if isinstance(p, dict) else float((p["order_amount"] or 0))
                if abs(_so_sum - _curr) > 0.5:
                    c2.execute("UPDATE projects SET order_amount=? WHERE id=?",
                               (_so_sum, pid))
                    # v5H130: 자가치유 변경 이력 기록 (이전 누락분)
                    try:
                        _logi.log_project_change(
                            c2, pid, u.get("id"), "수주액(자가치유)",
                            f"{_curr:,.0f}", f"{_so_sum:,.0f}",
                            note="v5H130 SO 합계 기준 자동 정정"
                        )
                    except Exception:
                        pass
                    if isinstance(p, dict):
                        p["order_amount"] = _so_sum
                # 납기 자가치유 (MAX)
                _due_dates = [o.get("due_date") for o in project_orders if o.get("due_date")]
                if _due_dates:
                    _max_due = max(_due_dates)
                    _proj_due = (p.get("due_date") if isinstance(p, dict) else p["due_date"]) or ""
                    if _max_due != _proj_due:
                        c2.execute("UPDATE projects SET due_date=? WHERE id=?",
                                   (_max_due, pid))
                        if isinstance(p, dict):
                            p["due_date"] = _max_due
                # 발주일 자가치유 (MIN)
                _ord_dates = [o.get("order_date") for o in project_orders if o.get("order_date")]
                if _ord_dates:
                    _min_ord = min(_ord_dates)
                    _proj_ord = (p.get("order_date") if isinstance(p, dict) else p["order_date"]) or ""
                    if _min_ord != _proj_ord:
                        c2.execute("UPDATE projects SET order_date=? WHERE id=?",
                                   (_min_ord, pid))
                        if isinstance(p, dict):
                            p["order_date"] = _min_ord
    except Exception:
        pass
    # v5H101: 프로젝트 변경 이력 (최근 30건)
    project_history_logs = []
    try:
        project_history_logs = _logi.get_project_history(pid, limit=30)
    except Exception:
        pass
    # v5H111: 사이드패널 호기 분해용 — 모든 SO 의 units 를 라벨 숫자 순 정렬
    # v5H133: 표시 순서를 내림차순(최근 호기 → 1호기)으로 반전 (대표 요청)
    all_units_sorted = []
    try:
        import re as _re_n
        _flat = []
        for _so in (project_orders or []):
            for _u in (_so.get("units") or []):
                _flat.append(dict(_u))
        def _sort_key(u):
            lbl = (u.get("unit_label") or "")
            m = _re_n.match(r"^(\d+)", lbl)
            return (int(m.group(1)) if m else 9999, lbl)
        all_units_sorted = sorted(_flat, key=_sort_key, reverse=True)
    except Exception:
        pass
    # v5H124: 프로젝트 SO 통화 혼합 감지 + 통화별 합계 분리
    currency_mix = []
    currency_warning = None
    primary_so_currency = "KRW"
    try:
        from collections import Counter as _C
        ccy_amount = {}
        ccy_count = {}
        for _so in (project_orders or []):
            _c = (_so.get("currency") or "KRW")
            ccy_amount[_c] = ccy_amount.get(_c, 0) + float(_so.get("total_amount") or 0)
            ccy_count[_c] = ccy_count.get(_c, 0) + 1
        if ccy_amount:
            currency_mix = sorted(
                [{"currency": k, "total": v, "cnt": ccy_count.get(k, 0)} for k, v in ccy_amount.items()],
                key=lambda r: -r["total"],
            )
            primary_so_currency = currency_mix[0]["currency"]
            if len(currency_mix) > 1:
                _list = ", ".join(f"{r['currency']}({r['cnt']}건)" for r in currency_mix)
                currency_warning = (
                    f"⚠ 통화 혼합 — {_list}. KPI 수주액은 주 통화({primary_so_currency}) 기준이며, "
                    "외화 SO 는 환산 없이 별도 표기됩니다."
                )
    except Exception:
        pass
    # v5H136 (2026-05-05): 이 프로젝트(장비)에 연결된 PO 라인 = 소모품·부품 사용 이력
    consumables = {"rows": [], "total_amount": 0, "total_qty": 0, "count": 0}
    try:
        consumables = _logi.get_project_consumables(pid, limit=200)
    except Exception:
        pass
    # v5H142 (2026-05-05): 이 프로젝트에 연결된 소모품 발주(consumable_order_items) 라인
    consumable_orders = {"rows": [], "total_amount": 0, "total_qty": 0, "count": 0}
    try:
        from . import consumables as _co_mod
        consumable_orders = _co_mod.get_project_consumable_orders(pid, limit=200)
    except Exception:
        pass
    # v5H137 (2026-05-05): 부모 프로젝트 정보 — CONSUMABLE/SERVICE 인 경우 안내 배너용
    parent_project = None
    try:
        _pp_id = p.get("parent_project_id") if isinstance(p, dict) else (p["parent_project_id"] if "parent_project_id" in p.keys() else None)
        if _pp_id:
            with db_session() as _cp:
                _pp = _cp.execute(
                    "SELECT id, mgmt_code, name, customer_name FROM projects WHERE id=?",
                    (int(_pp_id),)
                ).fetchone()
                if _pp:
                    parent_project = dict(_pp)
    except Exception:
        parent_project = None
    # v5H141 (2026-05-05): 자식 프로젝트(소모품/수리) 목록 — 부모 상세에 노출
    child_projects = []
    try:
        child_projects = _logi.get_child_projects(pid, limit=200)
    except Exception:
        child_projects = []
    # v5H200: 호기 상태로부터 종합 표시 상태 산출 (A안)
    # v5H214: 호기 0건(수주확정 전)이면 fallback 으로 stage 가 아닌 status 사용 — 사용자가 선택한 세부 상태 노출
    try:
        with db_session() as _cdc:
            _fb_status = (p.get("status") if isinstance(p, dict) else (p["status"] if "status" in p.keys() else "")) or ""
            _fb_stage  = (p.get("stage")  if isinstance(p, dict) else (p["stage"]  if "stage"  in p.keys() else "")) or ""
            project_display_status = _pwf.compute_project_display_status(
                _cdc, pid, fallback_stage=(_fb_status or _fb_stage)
            )
    except Exception:
        project_display_status = {"label": "—", "tone": "muted",
                                   "dist": {"진행중":0,"납품완료":0,"취소":0,"보류":0},
                                   "total": 0, "done": 0, "ratio_text": "",
                                   "has_canceled": False, "has_held": False}
    return ctx(req, "project_detail.html",
               project_display_status=project_display_status,
               user=u, p=p, tasks=tasks[:50], stats=stats,
               by_team=by_team_list, by_user=by_user_list, total_tasks=len(tasks),
               timeline=timeline_list[:30], all_comments=all_comments, retro=retro,
               project_orders=project_orders,
               STATUSES=_logi.LOGI_STATUSES,
               project_history=project_history_logs,
               all_units_sorted=all_units_sorted,
               currency_mix=currency_mix,
               currency_warning=currency_warning,
               primary_so_currency=primary_so_currency,
               consumables=consumables,
               consumable_orders=consumable_orders,
               PROJECT_TYPES=_logi.PROJECT_TYPES,
               PROJECT_TYPE_LABELS=_logi.PROJECT_TYPE_LABELS,
               parent_project=parent_project,
               child_projects=child_projects)


# =====================================================
# CUSTOMER LIST (2026-04-28 신설 — 사이드바 /customers 링크 broken 수정)
# =====================================================
@app.get("/customers", response_class=HTMLResponse)
async def customers_list(req: Request):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        try:
            # v5H58: 자동 산정 등급 점수(tier_score) 우선 정렬
            # v5H181: manager_name / phone / email / is_active / address 도 SELECT
            #         (이전엔 누락되어 list 페이지에서 빈 칸으로 보이던 버그)
            rows = c.execute(
                """SELECT cu.id, cu.name, cu.tier, cu.note, cu.biz_no, cu.ceo_name,
                          cu.manager_name, cu.phone, cu.email, cu.address,
                          COALESCE(cu.is_active, 1) AS is_active,
                          COALESCE(cu.tier_score, 0) AS tier_score,
                          cu.tier_computed_at,
                          COUNT(DISTINCT p.id) AS proj_count,
                          COALESCE(SUM(p.order_amount), 0) AS total_amount,
                          MAX(p.order_date) AS last_order
                   FROM customers cu
                   LEFT JOIN projects p ON p.customer_id = cu.id
                   GROUP BY cu.id
                   ORDER BY tier_score DESC, total_amount DESC, cu.name"""
            ).fetchall()
            customers = [dict(r) for r in rows]
        except Exception:
            customers = [dict(r) for r in c.execute(
                "SELECT id, name, tier, note FROM customers ORDER BY tier DESC, name"
            ).fetchall()]
    return ctx(req, "customers_list.html", user=u, active="customers", customers=customers)


# =====================================================
# CUSTOMER DETAIL
# =====================================================
@app.get("/customer/{cid}", response_class=HTMLResponse)
async def customer_detail(req: Request, cid: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        cu = c.execute("SELECT * FROM customers WHERE id=?", (cid,)).fetchone()
        if not cu:
            return RedirectResponse("/", 303)
        cu = dict(cu)
        # 고객사 프로젝트
        pjts = [dict(r) for r in c.execute(
            "SELECT * FROM projects WHERE customer_id=? ORDER BY id DESC", (cid,),
        ).fetchall()]
        # 최근 2주 카드
        since = (date.today() - timedelta(days=30)).isoformat()
        tasks = [dict(r) for r in c.execute(
            """SELECT tk.*, u.name AS user_name, tm.name AS team_name, p.name AS project_name
               FROM tasks tk JOIN users u ON tk.user_id=u.id
               LEFT JOIN teams tm ON u.team_id=tm.id
               LEFT JOIN projects p ON tk.project_id=p.id
               WHERE tk.customer_id=? AND tk.work_date>=? ORDER BY tk.work_date DESC""",
            (cid, since),
        ).fetchall()]
        stats = {
            "total": len(tasks),
            "done": sum(1 for t in tasks if t["status"] == "완료"),
            "delay": sum(1 for t in tasks if t["status"] == "지연"),
            "hours": round(sum(t["hours"] or 0 for t in tasks), 1),
        }
        by_team = {}
        for t in tasks:
            k = t["team_name"] or "(미배정)"
            by_team.setdefault(k, {"cnt": 0, "hours": 0})
            by_team[k]["cnt"] += 1
            by_team[k]["hours"] += t["hours"] or 0
        by_team_list = sorted(by_team.items(), key=lambda x: -x[1]["cnt"])
        # v5H56: 다중 담당자 (역할별 정렬)
        contacts = []
        try:
            contacts = [dict(r) for r in c.execute(
                "SELECT * FROM customer_contacts WHERE customer_id=? "
                "ORDER BY is_primary DESC, "
                "CASE role WHEN '영업' THEN 1 WHEN '구매' THEN 2 "
                "          WHEN '세금계산서' THEN 3 WHEN '품질' THEN 4 "
                "          WHEN '기술' THEN 5 WHEN '결재' THEN 6 ELSE 9 END, id",
                (cid,)
            ).fetchall()]
        except Exception:
            contacts = []

    # v5H58: 등급 점수 breakdown
    from . import customer_tier as _ct
    tier_breakdown = _ct.parse_breakdown(cu.get("tier_breakdown"))
    # v5H113: 고객사 변경 이력 (최근 50건)
    try:
        customer_history = _logi.get_customer_history(cid, limit=50)
    except Exception:
        customer_history = []
    return ctx(req, "customer_detail.html",
               user=u, cu=cu, pjts=pjts, tasks=tasks[:80],
               stats=stats, by_team=by_team_list, total_tasks=len(tasks),
               contacts=contacts, tier_breakdown=tier_breakdown,
               customer_history=customer_history)


# =====================================================
# CALENDAR (개인 월간 뷰)
# =====================================================
@app.get("/calendar", response_class=HTMLResponse)
@app.get("/calendar/{month}", response_class=HTMLResponse)
async def calendar_page(req: Request, month: str = "", scope: str = "me"):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if not month:
        month = date.today().strftime("%Y-%m")
    # 리더 이상이 아니면 me로 강제
    if scope == "team" and u["role"] not in ("leader", "executive", "ceo", "admin"):
        scope = "me"
    y, m = int(month[:4]), int(month[5:7])
    with db_session() as c:
        if scope == "team" and u.get("team_id"):
            rows = [dict(r) for r in c.execute(
                """SELECT work_date, status, COUNT(*) AS cnt, COALESCE(SUM(hours),0) AS hours
                   FROM tasks t JOIN users u ON t.user_id=u.id
                   WHERE u.team_id=? AND work_date LIKE ?
                   GROUP BY work_date, status""",
                (u["team_id"], f"{month}%"),
            ).fetchall()]
        else:
            rows = [dict(r) for r in c.execute(
                """SELECT work_date, status, COUNT(*) AS cnt, COALESCE(SUM(hours),0) AS hours
                   FROM tasks WHERE user_id=? AND work_date LIKE ?
                   GROUP BY work_date, status""",
                (u["id"], f"{month}%"),
            ).fetchall()]
    by_date = {}
    for r in rows:
        by_date.setdefault(r["work_date"], {"total": 0, "done": 0, "progress": 0,
                                              "delay": 0, "hours": 0})
        by_date[r["work_date"]]["total"] += r["cnt"]
        by_date[r["work_date"]]["hours"] += r["hours"]
        if r["status"] == "완료":
            by_date[r["work_date"]]["done"] += r["cnt"]
        elif r["status"] == "진행중":
            by_date[r["work_date"]]["progress"] += r["cnt"]
        elif r["status"] == "지연":
            by_date[r["work_date"]]["delay"] += r["cnt"]

    # v5H202: 일·월시작 캘린더(일=0, 월=1, ..., 토=6) + 공휴일 표기 (수주관리 기준 통일)
    import calendar as _cal_mod
    _ccal = _cal_mod.Calendar(firstweekday=6)  # Sunday-first
    weeks = []
    today_iso = date.today().isoformat()
    for week in _ccal.monthdatescalendar(y, m):
        row = []
        for dt in week:
            ds = dt.isoformat()
            in_month = (dt.month == m)
            row.append({
                "day": dt.day if in_month else 0,
                "date": ds if in_month else "",
                "d": by_date.get(ds) if in_month else None,
                "is_today": (ds == today_iso) if in_month else False,
                "wd": dt.weekday(),  # Mon=0..Sun=6
                "is_sun": dt.weekday() == 6,
                "is_sat": dt.weekday() == 5,
                "hol_kr": HOLIDAYS_KR.get(ds) if in_month else None,
                "hol_vn": HOLIDAYS_VN.get(ds) if in_month else None,
            })
        if any(c["day"] > 0 for c in row):
            weeks.append(row)

    prev_m = f"{y-1 if m==1 else y}-{12 if m==1 else m-1:02d}"
    next_m = f"{y+1 if m==12 else y}-{1 if m==12 else m+1:02d}"
    month_total = sum(d["total"] for d in by_date.values())
    month_done = sum(d["done"] for d in by_date.values())
    month_hours = round(sum(d["hours"] for d in by_date.values()), 1)

    return ctx(req, "calendar.html",
               user=u, weeks=weeks, month=month, prev_m=prev_m, next_m=next_m,
               month_total=month_total, month_done=month_done, month_hours=month_hours,
               scope=scope, can_team=u["role"] in ("leader","executive","ceo","admin") and u.get("team_id") is not None,
               active="calendar")


# =====================================================
# FEED (부서간 오늘 피드)
# =====================================================
@app.get("/feed", response_class=HTMLResponse)
@app.get("/feed/{sel_date}", response_class=HTMLResponse)
async def feed_page(req: Request, sel_date: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if not sel_date:
        sel_date = date.today().isoformat()
    with db_session() as c:
        # v5H49: 오늘 카드가 없으면 가장 최근 작성일로 자동 폴백 — 빈 피드 방지
        today_n = c.execute(
            "SELECT COUNT(*) FROM tasks WHERE work_date=?", (sel_date,)
        ).fetchone()[0]
        if today_n == 0:
            r = c.execute(
                "SELECT MAX(work_date) FROM tasks WHERE work_date<=?", (sel_date,)
            ).fetchone()
            if r and r[0]:
                sel_date = r[0]
        # 팀별 오늘 한 줄 요약
        summaries = [dict(r) for r in c.execute(
            """SELECT ts.*, t.name AS team_name, t.code AS team_code, t.is_lab,
                      u.name AS author_name
               FROM team_summaries ts JOIN teams t ON ts.team_id=t.id
               LEFT JOIN users u ON ts.author_id=u.id
               WHERE ts.work_date=? ORDER BY t.display_order""",
            (sel_date,),
        ).fetchall()]
        # 전 팀 오늘 카드 (요약)
        tasks = [dict(r) for r in c.execute(
            """SELECT tk.*, u.name AS user_name, u.rank AS user_rank, u.team_id AS team_id,
                      t.name AS team_name, t.code AS team_code,
                      p.name AS project_name, cu.name AS customer_name
               FROM tasks tk JOIN users u ON tk.user_id=u.id
               LEFT JOIN teams t ON u.team_id=t.id
               LEFT JOIN projects p ON tk.project_id=p.id
               LEFT JOIN customers cu ON tk.customer_id=cu.id
               WHERE tk.work_date=? ORDER BY t.display_order, u.id, tk.id""",
            (sel_date,),
        ).fetchall()]
        # 메타데이터(댓글수+리액션수) 일괄 조회 → 카드 표면 배지로 표시
        meta = get_meta_bulk([t["id"] for t in tasks])
        # 현재 사용자 멘션/응답대기 카운트 (나를 @멘션한 댓글이 달린 카드)
        mentioned_ids = set()
        if tasks:
            ph = ",".join("?" * len(tasks))
            mrows = c.execute(
                f"""SELECT DISTINCT tc.task_id
                    FROM comment_mentions cm
                    JOIN task_comments tc ON cm.comment_id=tc.id
                    WHERE cm.user_id=? AND tc.task_id IN ({ph})""",
                (u["id"],) + tuple(t["id"] for t in tasks)
            ).fetchall()
            mentioned_ids = {r["task_id"] for r in mrows}
        for t in tasks:
            m = meta.get(t["id"], {})
            t["meta_comments"] = m.get("comments", 0)
            t["meta_ack"] = m.get("ack", 0)
            t["meta_question"] = m.get("question", 0)
            t["meta_risk"] = m.get("risk", 0)
            t["meta_ok"] = m.get("ok", 0)
            t["meta_last_comment"] = m.get("last_comment", "")
            t["meta_mentioned_me"] = 1 if t["id"] in mentioned_ids else 0
            t["meta_has_activity"] = bool(
                t["meta_comments"] or t["meta_ack"] or t["meta_question"]
                or t["meta_risk"] or t["meta_ok"]
            )
            # 정렬 키: 지연(0) → 리스크(1) → 멘션(2) → 활동(3) → 진행중(4) → 대기(5) → 완료(6)
            if t["status"] == "지연":
                sk = 0
            elif t["meta_risk"]:
                sk = 1
            elif t["meta_mentioned_me"]:
                sk = 2
            elif t["meta_has_activity"]:
                sk = 3
            elif t["status"] == "진행중":
                sk = 4
            elif t["status"] in ("대기", "보류"):
                sk = 5
            else:
                sk = 6
            t["_sort"] = sk

        by_team = {}
        for t in tasks:
            k = t["team_name"] or "(미배정)"
            by_team.setdefault(k, {"code": t["team_code"], "tasks": [],
                                    "has_urgent": False})
            by_team[k]["tasks"].append(t)
            if t["_sort"] <= 2:  # 지연/리스크/멘션 중 하나라도 있으면 긴급팀
                by_team[k]["has_urgent"] = True
        # 각 팀 내부 정렬: 우선순위 → 담당자 → id
        for k, d in by_team.items():
            d["tasks"].sort(key=lambda x: (x["_sort"], x.get("user_name") or "", x["id"]))
        # v5H49: 템플릿이 td.team_name/total/done/delay/summaries 를 기대 → dict 리스트로 변환
        by_team_list = []
        for nm, info in sorted(by_team.items(), key=lambda x: x[1]["code"] or "99"):
            ts = info.get("tasks") or []
            by_team_list.append({
                "team_name": nm,
                "code": info.get("code"),
                "has_urgent": info.get("has_urgent", False),
                "tasks": ts,
                "summaries": ts[:5],
                "total": len(ts),
                "done": sum(1 for t in ts if t.get("status") == "완료"),
                "progress": sum(1 for t in ts if t.get("status") == "진행중"),
                "delay": sum(1 for t in ts if t.get("status") == "지연"),
            })

        # 전체 카운트 (필터칩용)
        all_total = len(tasks)
        all_done = sum(1 for t in tasks if t["status"] == "완료")
        all_progress = sum(1 for t in tasks if t["status"] == "진행중")
        all_delay = sum(1 for t in tasks if t["status"] == "지연")
        all_wait = sum(1 for t in tasks if t["status"] in ("대기", "보류"))
        all_risk = sum(1 for t in tasks if t["meta_risk"])
        all_mentioned = sum(1 for t in tasks if t["meta_mentioned_me"])
    prev_d = (datetime.strptime(sel_date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
    next_d = (datetime.strptime(sel_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
    # 멀티뷰용 flat JSON (칸반/캘린더)
    import json as _json
    flat_tasks = _json.dumps([{
        "id": t["id"], "title": t["title"], "status": t["status"],
        "user_name": t["user_name"], "user_rank": t.get("user_rank") or "",
        "team_name": t.get("team_name") or "", "project_name": t.get("project_name") or "",
        "customer_name": t.get("customer_name") or "",
        "hours": t.get("hours") or 0,
        "cm": t.get("meta_comments", 0), "risk": t.get("meta_risk", 0),
        "question": t.get("meta_question", 0), "ack": t.get("meta_ack", 0),
    } for t in tasks], ensure_ascii=False)
    # 역할 기반 기본 펼침 로직용 — 내 팀 id
    my_team_id = u.get("team_id")
    return ctx(req, "feed.html",
               user=u, summaries=summaries, by_team=by_team_list,
               sel_date=sel_date, prev_date=prev_d, next_date=next_d,
               all_total=all_total, all_done=all_done, all_progress=all_progress,
               all_delay=all_delay, all_wait=all_wait,
               all_risk=all_risk, all_mentioned=all_mentioned,
               my_team_id=my_team_id, flat_tasks_json=flat_tasks,
               sel_date_obj=sel_date, active="feed")


# =====================================================
# ADMIN — 관리자 페이지
# =====================================================
@app.get("/admin", response_class=HTMLResponse)
async def admin_page(req: Request):
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        teams = [dict(r) for r in c.execute(
            """SELECT t.*, u.name AS leader_name,
                      (SELECT COUNT(*) FROM users WHERE team_id=t.id AND is_active=1) AS member_count
               FROM teams t LEFT JOIN users u ON t.leader_id=u.id
               ORDER BY t.display_order"""
        ).fetchall()]
        users = [dict(r) for r in c.execute(
            """SELECT u.*, t.name AS team_name FROM users u
               LEFT JOIN teams t ON u.team_id=t.id
               WHERE u.role!='admin' ORDER BY t.display_order, u.role DESC, u.id"""
        ).fetchall()]
        projects = [dict(r) for r in c.execute(
            """SELECT p.*, cu.name AS customer_name FROM projects p
               LEFT JOIN customers cu ON p.customer_id=cu.id ORDER BY p.id DESC"""
        ).fetchall()]
        customers = [dict(r) for r in c.execute("SELECT * FROM customers ORDER BY tier DESC, id").fetchall()]
    return ctx(req, "admin.html",
               user=u, teams=teams, users=users, projects=projects, customers=customers,
               active="admin")


@app.post("/api/admin/user")
async def api_admin_user(req: Request):
    u = require(req, ["admin"])
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    d = await req.json()
    with db_session() as c:
        if d.get("id"):
            c.execute(
                "UPDATE users SET name=?, team_id=?, rank=?, role=?, is_active=? WHERE id=?",
                (d["name"], d.get("team_id") or None, d.get("rank", ""),
                 d.get("role", "member"), int(d.get("is_active", 1)), d["id"]),
            )
            if d.get("password"):
                c.execute("UPDATE users SET password=? WHERE id=?", (hash_pw(d["password"]), d["id"]))
        else:
            c.execute(
                """INSERT INTO users(name, login_id, password, team_id, rank, role)
                   VALUES(?,?,?,?,?,?)""",
                (d["name"], d["login_id"], hash_pw(d.get("password", "knk1234")),
                 d.get("team_id") or None, d.get("rank", ""), d.get("role", "member")),
            )
    return JSONResponse({"ok": True})


@app.post("/api/admin/project")
async def api_admin_project(req: Request):
    u = require(req, ["admin", "ceo"])
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    d = await req.json()
    with db_session() as c:
        if d.get("id"):
            c.execute(
                "UPDATE projects SET code=?, name=?, customer_id=?, type=?, status=? WHERE id=?",
                (d.get("code", ""), d["name"], d.get("customer_id") or None,
                 d.get("type", ""), d.get("status", "진행중"), d["id"]),
            )
        else:
            c.execute(
                "INSERT INTO projects(code, name, customer_id, type, status) VALUES(?,?,?,?,?)",
                (d.get("code", ""), d["name"], d.get("customer_id") or None,
                 d.get("type", ""), d.get("status", "진행중")),
            )
    return JSONResponse({"ok": True})


@app.post("/api/admin/customer")
async def api_admin_customer(req: Request):
    u = require(req, ["admin", "ceo"])
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    d = await req.json()
    with db_session() as c:
        if d.get("id"):
            c.execute(
                "UPDATE customers SET name=?, tier=?, note=? WHERE id=?",
                (d["name"], d.get("tier", "일반"), d.get("note", ""), d["id"]),
            )
        else:
            c.execute(
                "INSERT INTO customers(name, tier, note) VALUES(?,?,?)",
                (d["name"], d.get("tier", "일반"), d.get("note", "")),
            )
    return JSONResponse({"ok": True})


# =====================================================
# MGMT CODE IMPORT — 관리코드발행목록.xls 업로드
# =====================================================
@app.post("/api/admin/import-mgmt")
async def api_admin_import_mgmt(req: Request, file: UploadFile = File(...)):
    u = require(req, ["admin", "ceo"])
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    fn = (file.filename or "").lower()
    # 사이클 71: pandas 제거(대표 결정 2026-04-27) → .csv 만 허용. .xls/.xlsx는 안내 메시지.
    if fn.endswith(".xls") or fn.endswith(".xlsx"):
        return JSONResponse({"error": (
            ".xls/.xlsx 직접 파싱은 사이클 71(2026-04-27)에서 폐기되었습니다 (외부 의존 0 정책). "
            "Excel/LibreOffice에서 .csv (UTF-8) 로 저장 후 업로드해 주세요."
        )}, 400)
    if not fn.endswith(".csv"):
        return JSONResponse({"error": "csv 파일만 업로드 가능 (.xls/.xlsx 는 .csv 변환 후 업로드)"}, 400)
    try:
        data = await file.read()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".csv") as tf:
            tf.write(data)
            tmp_path = tf.name
        try:
            rows = parse_mgmt_csv(tmp_path)
            result = import_mgmt_rows(rows)
            result["parsed"] = len(rows)
            result["filename"] = file.filename
            return JSONResponse({"ok": True, "result": result})
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass
    except Exception as e:
        return JSONResponse({"error": f"파싱 실패: {type(e).__name__}: {e}"}, 500)


# =====================================================
# INITIAL PASSWORD REGENERATION (A안 킥오프)
# =====================================================
@app.post("/api/admin/regenerate-passwords")
async def api_regen_pw(req: Request):
    u = require(req, ["admin", "ceo"])
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    rows = regenerate_user_passwords()
    out_dir = os.path.join(BASE, "data")
    os.makedirs(out_dir, exist_ok=True)
    # 사이클 68 (2026-04-27): openpyxl 제거 → CSV (UTF-8 BOM · 대표 결정 이행)
    out_path = os.path.join(out_dir, "초기비밀번호_배포용.csv")
    build_password_csv(rows, out_path)
    return JSONResponse({"ok": True, "count": len(rows),
                         "download": "/admin/download-passwords"})


@app.get("/admin/health", response_class=HTMLResponse)
async def admin_health_page(req: Request):
    """건전성 점검 — 어떤 기능이 진짜 동작하는지 한눈에"""
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    from .database import health_check
    checks = health_check()
    # 레벨별 그룹
    by_level = {"core": [], "data": [], "external": [], "ops": []}
    for c in checks:
        by_level.setdefault(c.get("level", "ops"), []).append(c)
    # 상태별 카운트
    counts = {"ok": 0, "warn": 0, "error": 0, "info": 0}
    for c in checks:
        counts[c.get("status", "info")] += 1
    return ctx(req, "admin_health.html", user=u,
               checks=checks, by_level=by_level, counts=counts,
               active="admin")


# =====================================================
# 사이클 73 (2026-04-27): 사용 가이드 페이지
# 대표 16:20 직접 지시 — "내가 직접 사용해볼려고 하는데 사용 가이드가 있어야할것 같은데"
# 모든 로그인 사용자 접근 가능 (admin/ceo/sales/finance/etc.)
# =====================================================
@app.get("/guide", response_class=HTMLResponse)
async def guide_page(req: Request):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    return ctx(req, "guide.html", user=u, active="guide")


# =====================================================
# 외부자산 점검 (대표 직접 판단용 spike, 2026-04-27)
# 출처: 00_HAIST_WORKS_감사팀/_TO_09팀장_2026-04-27_긴급감사_openpyxl외부자산.md
# =====================================================
EXTERNAL_ASSETS_REVIEW = [
    {
        "name": "openpyxl",
        "type": "PyPI 라이브러리 (엑셀 파일 생성)",
        "status": "removed_2026-04-27_cycle68",
        "usage": [
            {"file": "app/main.py", "lines": "(제거됨)",
             "purpose": "사이클 68: 주간 요약 CSV 라우트로 대체 (line 3300~)",
             "deprecated": True},
            {"file": "app/database.py", "lines": "2271~2287",
             "purpose": "사이클 68: build_password_csv 로 교체 (csv 표준 모듈)",
             "deprecated": True},
            {"file": "scripts/migrate_baby_v2.py", "lines": "79",
             "purpose": "baby Excel → web SQLite 마이그레이션 (운영 외 1회성, baby 폐기 시 자연 제거)"},
            {"file": "scripts/baby_web_sync_check.py", "lines": "38",
             "purpose": "baby/web 정합성 체크 (운영 외 스크립트, baby 폐기 시 자연 제거)"},
        ],
        "alternatives": [
            "CSV 다운로드 (csv 표준 모듈) — 사이클 68 적용",
            "HTML 인쇄 view (브라우저 인쇄 기능)",
        ],
        "impact_summary": (
            "사이클 68 (2026-04-27 대표 결정 Remove) 적용 완료. "
            "운영 라우트에서 openpyxl import 0건. "
            "scripts/ 2개 파일은 baby 폐기 시 자연 제거 예정."
        ),
        "risk_security": 1,
        "risk_dependency": 3,
        "introduced_at": "2026-04-15 (commit 546fb23, 초기 구축)",
        "duration": "12일 + 55사이클 누적",
    },
    {
        "name": "pandas",
        "type": "PyPI 라이브러리 (데이터 분석, numpy 의존)",
        "status": "removed_2026-04-27_cycle71",
        "usage": [
            {"file": "app/database.py", "lines": "(제거됨)",
             "purpose": "사이클 71: parse_mgmt_xls는 NotImplementedError 안내, parse_mgmt_csv (csv 표준 모듈) 신설",
             "deprecated": True},
            {"file": "app/main.py", "lines": "(라우트 수정)",
             "purpose": "사이클 71: /api/admin/import-mgmt 가 .csv 만 수용, .xls/.xlsx 업로드는 변환 안내 400",
             "deprecated": True},
        ],
        "alternatives": [
            "외부 환경에서 .xls → .csv 변환 후 csv 표준 모듈로 처리 — 사이클 71 적용",
            "parse_mgmt_csv (database.py) — 표준 csv 모듈 기반 새 헬퍼",
        ],
        "impact_summary": (
            "사이클 71 (2026-04-27 대표 결정 Remove 2회 제출 14:20:25 + 14:22:08) 적용 완료. "
            "pandas import 0건. requirements.txt 도 0건 (이전부터 미명시였음). "
            "관리코드 마이그레이션은 사용자가 .xls → .csv 변환 후 업로드하는 1회성 흐름으로 전환."
        ),
        "risk_security": 1,
        "risk_dependency": 4,
        "introduced_at": "2026-04-15 (commit 546fb23, 초기 구축)",
        "duration": "12일 + 55사이클 누적",
    },
    {
        "name": "deep_translator",
        "type": "PyPI 라이브러리 + 외부 번역 API (Google Translate / MyMemory)",
        "status": "removed_2026-04-27_cycle69",
        "usage": [
            {"file": "app/main.py", "lines": "(제거됨)",
             "purpose": "사이클 69: api_translate() 사내 i18n 사전 lookup으로 교체",
             "deprecated": True},
            {"file": "requirements.txt", "lines": "(제거됨)",
             "purpose": "사이클 69: deep-translator 의존성 제거",
             "deprecated": True},
        ],
        "alternatives": [
            "사내 i18n 사전 lookup (app/i18n.py 약 472키 ko/vi/en) — 사이클 69 적용",
            "빅터AI 단어 등록 기능 (사이클 70+ 신설 예정, 대표 결재 후)",
        ],
        "impact_summary": (
            "사이클 69 (2026-04-27 대표 결정 Partial 부분처리) 적용 완료. "
            "외부 API 호출 0건. "
            "/api/translate 라우트는 i18n.py T 사전 lookup 기반으로 동작 (외부 자산 0)."
        ),
        "risk_security": 4,
        "risk_dependency": 3,
        "introduced_at": "2026-04-15 (commit 546fb23, 초기 구축)",
        "duration": "12일 + 55사이클 누적",
    },
]


@app.get("/admin/external-assets", response_class=HTMLResponse)
async def admin_external_assets(req: Request):
    """외부자산 점검 — 대표 직접 판단 (spike 2026-04-27)"""
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    saved = req.query_params.get("saved", "")
    return ctx(req, "external_assets_review.html", user=u,
               assets=EXTERNAL_ASSETS_REVIEW,
               saved=saved,
               active="admin")


@app.post("/admin/external-assets/decision")
async def admin_external_assets_decision(req: Request):
    """대표 결정 제출 → 99_DISPATCH 폴더에 결정 .md 자동 추가"""
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    form = await req.form()
    asset_name = (form.get("asset_name") or "").strip()
    decision = (form.get("decision") or "").strip()
    note = (form.get("note") or "").strip()

    # 입력 화이트리스트 (XSS/주입 방지)
    allowed_assets = {"openpyxl", "pandas", "deep_translator"}
    allowed_decisions = {"remove", "keep", "partial"}
    if asset_name not in allowed_assets or decision not in allowed_decisions:
        return RedirectResponse("/admin/external-assets?saved=err", 303)

    # 99_DISPATCH 폴더에 결정 추가 (append)
    dispatch_dir = os.path.join(BASE, "..", "99_DISPATCH")
    dispatch_dir = os.path.abspath(dispatch_dir)
    try:
        os.makedirs(dispatch_dir, exist_ok=True)
    except Exception:
        pass
    out_path = os.path.join(dispatch_dir, "외부자산_결정_2026-04-27.md")

    decision_label = {"remove": "제거(Remove)", "keep": "유지(Keep)",
                       "partial": "부분 처리(Partial)"}[decision]
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # 메모는 마크다운 안전을 위해 라인 단위로 정리
    safe_note = "\n".join(("> " + ln) for ln in note.splitlines()) if note else "> (메모 없음)"

    block = (
        f"\n---\n\n"
        f"## {asset_name} — {decision_label}\n\n"
        f"- 결정 시각: {ts}\n"
        f"- 결정자: {u.get('name','?')} (id={u.get('id','?')}, role={u.get('role','?')})\n"
        f"- 자산: `{asset_name}`\n"
        f"- 결정: **{decision_label}**\n"
        f"- 메모:\n{safe_note}\n"
    )
    header_needed = not os.path.exists(out_path)
    try:
        with open(out_path, "a", encoding="utf-8") as f:
            if header_needed:
                f.write(
                    "# 외부자산 점검 — 대표 결정 기록\n\n"
                    "> 자동 생성 (POST /admin/external-assets/decision)\n"
                    "> spike 발주: 2026-04-27 (대표 직접 지시 12:30)\n"
                    "> 본 파일은 09 팀장(빅터)이 모니터링하여 후속 spike 발주\n"
                )
            f.write(block)
    except Exception:
        return RedirectResponse("/admin/external-assets?saved=err", 303)

    return RedirectResponse(
        f"/admin/external-assets?saved={asset_name}", 303)


# =====================================================
# 회사 정보 admin (사이클 61 U6 — 대표가 직접 입력 / 외부자산 0건)
# 출처: 04 _TO_09팀장_2026-04-27_사이클58_UX재검증.md S2 U6
# 저장: app_settings 테이블 (key = "company_*")
# =====================================================
@app.get("/admin/company-info", response_class=HTMLResponse)
async def admin_company_info_page(req: Request):
    """회사 식별 정보 입력 페이지 — 견적서 헤더용 (대표 직접 입력)."""
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    # OPS-P0-4: saved 정수화 (문자열 "0" truthy 방지)
    try:
        saved = int(req.query_params.get("saved", "0") or "0")
    except (TypeError, ValueError):
        saved = 0
    # 현재 입력된 값 (default 미적용 raw)
    current = {key: (get_setting(key, "") or "") for key, _l, _d in COMPANY_INFO_KEYS}
    return ctx(req, "admin_company_info.html", user=u,
               keys=COMPANY_INFO_KEYS,
               current=current,
               saved=saved,
               active="admin")


@app.post("/admin/company-info")
async def admin_company_info_save(req: Request):
    """회사 정보 저장 → app_settings UPSERT."""
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    form = await req.form()
    # 화이트리스트 (XSS/주입 방지) — 정의된 키만 저장
    allowed_keys = {key for key, _l, _d in COMPANY_INFO_KEYS}
    saved = 0
    for k, v in form.items():
        if not k.startswith("k_"):
            continue
        key = k[2:].strip()
        if key not in allowed_keys:
            continue
        set_setting(key, (v or "").strip(), user_id=u["id"])
        saved += 1
    return RedirectResponse(f"/admin/company-info?saved={saved}", 303)


@app.get("/admin/settings", response_class=HTMLResponse)
async def admin_settings_page(req: Request):
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    settings = get_settings_all()
    return ctx(req, "admin_settings.html", user=u, settings=settings, active="admin")


@app.post("/admin/settings")
async def admin_settings_save(req: Request):
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    form = await req.form()
    # form 키는 "k_<key>" 형식 (XSS/주입 방지)
    saved = 0
    for k, v in form.items():
        if not k.startswith("k_"):
            continue
        key = k[2:].strip()
        if not key:
            continue
        set_setting(key, (v or "").strip(), user_id=u["id"])
        saved += 1
    return RedirectResponse(f"/admin/settings?saved={saved}", 303)


# =====================================================
# 사이클 79 (DEC-2 · 2026-04-27) — 인사총무 하이웍스 연동 카드
# 인사총무 업무는 하이웍스 외부 시스템 사용. HAIST WORKS는 연동 카드 + 외부 링크 + 본인 KNK 정보 조회만.
# 외부 자산 0건 (iframe import 없음, 외부 링크만 노출)
# =====================================================
HIWORKS_LINK_KEYS = [
    ("hiworks_main_url",       "하이웍스 메인",       "https://office.hiworks.com/"),
    ("hiworks_attendance_url", "출퇴근 / 근태",       ""),
    ("hiworks_leave_url",      "휴가 신청",           ""),
    ("hiworks_payroll_url",    "급여 명세서",         ""),
    ("hiworks_profile_url",    "인사 정보",           ""),
]


@app.get("/hr/hiworks", response_class=HTMLResponse)
async def hr_hiworks_page(req: Request):
    """인사총무 하이웍스 연동 페이지 — 모든 로그인 사용자, 본인 KNK 정보만 노출."""
    u = require(req)
    if not u:
        return RedirectResponse("/login", 303)
    # admin_settings 에 입력된 외부 링크 (미입력 시 빈 문자열)
    links = [{"key": k, "label": label, "url": (get_setting(k, "") or "").strip(), "default": default}
             for k, label, default in HIWORKS_LINK_KEYS]
    # 본인 KNK 정보 (다른 직원 정보 노출 금지 — 개인정보 보호)
    me = {
        "name":       u.get("name", ""),
        "login_id":   u.get("login_id", ""),
        "team_name":  u.get("team_name", "") or "(미지정)",
        "rank":       u.get("rank", "") or "(미지정)",
        "role":       u.get("role", "member"),
        "email":      u.get("email", "") or "",
        "hire_date":  u.get("created_at", "") or "",
    }
    return ctx(req, "hr_hiworks.html", user=u,
               links=links, me=me, active="hr_hiworks")


@app.get("/admin/hiworks-settings", response_class=HTMLResponse)
async def admin_hiworks_settings_page(req: Request):
    """하이웍스 연동 URL 설정 (admin/ceo only)."""
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    saved = req.query_params.get("saved", "")
    current = {k: (get_setting(k, "") or "") for k, _l, _d in HIWORKS_LINK_KEYS}
    return ctx(req, "admin_hiworks_settings.html", user=u,
               keys=HIWORKS_LINK_KEYS, current=current,
               saved=saved, active="admin")


@app.post("/admin/hiworks-settings")
async def admin_hiworks_settings_save(req: Request):
    """하이웍스 URL 저장 → app_settings UPSERT (화이트리스트만)."""
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    form = await req.form()
    allowed = {k for k, _l, _d in HIWORKS_LINK_KEYS}
    saved = 0
    for k, v in form.items():
        if not k.startswith("k_"):
            continue
        key = k[2:].strip()
        if key not in allowed:
            continue
        set_setting(key, (v or "").strip(), user_id=u["id"])
        saved += 1
    return RedirectResponse(f"/admin/hiworks-settings?saved={saved}", 303)


@app.get("/admin/download-passwords")
async def dl_passwords(req: Request):
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    # 사이클 68 (2026-04-27): openpyxl 제거 → CSV 다운로드 (UTF-8 BOM · 대표 결정 이행)
    out_path = os.path.join(BASE, "data", "초기비밀번호_배포용.csv")
    # 구버전 .xlsx 파일이 남아있다면 호환을 위해 폴백
    if not os.path.exists(out_path):
        legacy_xlsx = os.path.join(BASE, "data", "초기비밀번호_배포용.xlsx")
        if os.path.exists(legacy_xlsx):
            return JSONResponse({"error": "구버전 .xlsx 파일이 존재합니다. [비밀번호 재생성]을 1회 다시 실행하면 .csv로 자동 갱신됩니다."}, 410)
        return JSONResponse({"error": "먼저 재생성을 수행하세요"}, 404)
    with open(out_path, "rb") as f:
        data = f.read()
    from urllib.parse import quote
    fn = "KNK_초기비밀번호_배포용.csv"
    headers = {
        "Content-Disposition": f"attachment; filename=KNK_passwords.csv; filename*=UTF-8''{quote(fn)}"
    }
    return StreamingResponse(
        io.BytesIO(data),
        media_type="text/csv; charset=utf-8-sig",
        headers=headers,
    )


# =====================================================
# PROFILE — 본인 프로필 (정보 + 활동 30일 + 권한 매트릭스 1인)
# =====================================================
def _profile_payload(c, uid: int):
    """본인 프로필 페이로드 — 활동/권한/위임 토큰 (단일 사용자)."""
    out = {
        "tasks_30d": 0, "tasks_open": 0, "tasks_done": 0,
        "comments_30d": 0, "notifs_30d": 0, "notifs_unread": 0,
        "recent_tasks": [], "recent_comments": [], "recent_acts": [],
        "perms_direct": [], "perms_group": [], "perms_deleg": [],
        "tokens_received": [], "tokens_granted": [],
    }
    try:
        # 활동 집계 (최근 30일)
        out["tasks_30d"] = (c.execute(
            "SELECT COUNT(*) AS n FROM tasks WHERE user_id=? AND date(work_date) >= date('now','-30 day','localtime')",
            (uid,)).fetchone() or {"n": 0})["n"]
        out["tasks_open"] = (c.execute(
            "SELECT COUNT(*) AS n FROM tasks WHERE user_id=? AND status IN ('진행중','대기','지연','보류')",
            (uid,)).fetchone() or {"n": 0})["n"]
        out["tasks_done"] = (c.execute(
            "SELECT COUNT(*) AS n FROM tasks WHERE user_id=? AND status='완료' AND date(work_date) >= date('now','-30 day','localtime')",
            (uid,)).fetchone() or {"n": 0})["n"]
        out["comments_30d"] = (c.execute(
            "SELECT COUNT(*) AS n FROM task_comments WHERE author_id=? AND date(created_at) >= date('now','-30 day','localtime')",
            (uid,)).fetchone() or {"n": 0})["n"]
        out["notifs_30d"] = (c.execute(
            "SELECT COUNT(*) AS n FROM notifications WHERE user_id=? AND date(created_at) >= date('now','-30 day','localtime')",
            (uid,)).fetchone() or {"n": 0})["n"]
        out["notifs_unread"] = (c.execute(
            "SELECT COUNT(*) AS n FROM notifications WHERE user_id=? AND is_read=0",
            (uid,)).fetchone() or {"n": 0})["n"]
        # 최근 task 10건
        out["recent_tasks"] = [dict(r) for r in c.execute(
            "SELECT id, work_date, title, status, COALESCE(category,'') AS category "
            "FROM tasks WHERE user_id=? ORDER BY work_date DESC, id DESC LIMIT 10",
            (uid,)).fetchall()]
        # 최근 댓글 10건 (task 제목 join)
        out["recent_comments"] = [dict(r) for r in c.execute(
            "SELECT tc.id, tc.task_id, tc.body, tc.created_at, COALESCE(t.title,'') AS task_title "
            "FROM task_comments tc LEFT JOIN tasks t ON t.id=tc.task_id "
            "WHERE tc.author_id=? ORDER BY tc.created_at DESC LIMIT 10",
            (uid,)).fetchall()]
    except Exception:
        pass
    # activities 본인 행 (선택)
    try:
        out["recent_acts"] = [dict(r) for r in c.execute(
            "SELECT id, kind, title, created_at FROM activities WHERE actor_id=? ORDER BY id DESC LIMIT 10",
            (uid,)).fetchall()]
    except Exception:
        out["recent_acts"] = []
    # 권한 — 직접/그룹/위임 (단일 사용자 매트릭스)
    try:
        out["perms_direct"] = [dict(r) for r in c.execute(
            "SELECT p.id, COALESCE(p.resource||'.'||p.action, p.name) AS label "
            "FROM user_permissions up JOIN permissions p ON p.id=up.permission_id "
            "WHERE up.user_id=? ORDER BY label LIMIT 60", (uid,)).fetchall()]
    except Exception:
        out["perms_direct"] = []
    try:
        out["perms_group"] = [dict(r) for r in c.execute(
            "SELECT DISTINCT p.id, COALESCE(p.resource||'.'||p.action, p.name) AS label, g.name AS group_name "
            "FROM user_groups ug "
            "JOIN group_permissions gp ON gp.group_id=ug.group_id "
            "JOIN permissions p ON p.id=gp.permission_id "
            "JOIN permission_groups g ON g.id=ug.group_id "
            "WHERE ug.user_id=? ORDER BY label LIMIT 60", (uid,)).fetchall()]
    except Exception:
        out["perms_group"] = []
    try:
        out["perms_deleg"] = [dict(r) for r in c.execute(
            "SELECT dt.id, COALESCE(p.resource||'.'||p.action, p.name) AS label, "
            "       dt.expires_at, dt.status, uf.name AS from_name "
            "FROM delegation_tokens dt "
            "JOIN permissions p ON p.id=dt.permission_id "
            "LEFT JOIN users uf ON uf.id=dt.from_user "
            "WHERE dt.to_user=? AND dt.status='ACTIVE' ORDER BY dt.id DESC LIMIT 30",
            (uid,)).fetchall()]
        out["tokens_received"] = out["perms_deleg"]
        out["tokens_granted"] = [dict(r) for r in c.execute(
            "SELECT dt.id, COALESCE(p.resource||'.'||p.action, p.name) AS label, "
            "       dt.expires_at, dt.status, ut.name AS to_name "
            "FROM delegation_tokens dt "
            "JOIN permissions p ON p.id=dt.permission_id "
            "LEFT JOIN users ut ON ut.id=dt.to_user "
            "WHERE dt.from_user=? ORDER BY dt.id DESC LIMIT 30",
            (uid,)).fetchall()]
    except Exception:
        out["perms_deleg"] = []
        out["tokens_received"] = []
        out["tokens_granted"] = []
    # task_delegations (위임 받은 task)
    try:
        out["task_delegs_in"] = [dict(r) for r in c.execute(
            "SELECT td.id, td.task_id, td.message, td.status, td.created_at, "
            "       COALESCE(t.title,'') AS task_title, uf.name AS from_name "
            "FROM task_delegations td LEFT JOIN tasks t ON t.id=td.task_id "
            "LEFT JOIN users uf ON uf.id=td.from_user_id "
            "WHERE td.to_user_id=? ORDER BY td.id DESC LIMIT 10",
            (uid,)).fetchall()]
    except Exception:
        out["task_delegs_in"] = []
    return out


@app.get("/profile", response_class=HTMLResponse)
async def profile_page(req: Request, msg: str = "", err: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        pdata = _profile_payload(c, u["id"])
    return ctx(req, "profile.html", user=u, msg=msg, err=err, active="profile", **pdata)


@app.get("/me", response_class=HTMLResponse)
async def me_alias(req: Request, msg: str = "", err: str = ""):
    """본인 프로필 별칭 (/me → /profile)."""
    return await profile_page(req, msg=msg, err=err)


@app.post("/profile/change-password")
async def change_password(req: Request,
                          current_pw: str = Form(...),
                          new_pw: str = Form(...),
                          confirm_pw: str = Form(...)):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if new_pw != confirm_pw:
        return RedirectResponse("/profile?err=새 비밀번호가 일치하지 않습니다", 303)
    if len(new_pw) < 6:
        return RedirectResponse("/profile?err=비밀번호는 6자 이상", 303)
    with db_session() as c:
        row = c.execute("SELECT password FROM users WHERE id=?", (u["id"],)).fetchone()
        if not row or row["password"] != hash_pw(current_pw):
            return RedirectResponse("/profile?err=현재 비밀번호가 맞지 않습니다", 303)
        c.execute("UPDATE users SET password=? WHERE id=?",
                  (hash_pw(new_pw), u["id"]))
        # v5H115: 본인 비번 변경도 user_history 에 기록 (감사 추적)
        try:
            c.execute(
                "INSERT INTO user_history(user_id, changed_by, field, "
                "old_value, new_value, note) VALUES(?,?,?,?,?,?)",
                (u["id"], u["id"], "비밀번호", "***", "***", "본인 변경")
            )
        except Exception:
            pass
    return RedirectResponse("/profile?msg=비밀번호가 변경되었습니다", 303)


@app.post("/me")
async def me_update(req: Request,
                    email: str = Form(""),
                    lang: str = Form("")):
    """본인 프로필 수정 — email / lang 만 (본인 한정).
    phone/dept 컬럼은 스키마에 따라 선택 적용 (PRAGMA로 존재 시만).
    """
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    email = (email or "").strip()
    lang = (lang or "").strip()
    sets, vals = [], []
    if email:
        if "@" in email and len(email) <= 120:
            sets.append("email=?"); vals.append(email)
    if lang in ("ko", "en", "vi", "zh"):
        sets.append("lang=?"); vals.append(lang)
    # phone / dept (스키마에 존재할 때만)
    try:
        with db_session() as c:
            ucols = [r[1] for r in c.execute("PRAGMA table_info(users)").fetchall()]
    except Exception:
        ucols = []
    # form 에서 추가 필드 (있을 때만)
    form = await req.form()
    if "phone" in ucols and form.get("phone") is not None:
        ph = (form.get("phone") or "").strip()[:30]
        sets.append("phone=?"); vals.append(ph)
    if "dept" in ucols and form.get("dept") is not None:
        dp = (form.get("dept") or "").strip()[:60]
        sets.append("dept=?"); vals.append(dp)
    if not sets:
        return RedirectResponse("/profile?err=변경할 항목이 없습니다", 303)
    vals.append(u["id"])
    with db_session() as c:
        c.execute(f"UPDATE users SET {', '.join(sets)} WHERE id=?", vals)
    return RedirectResponse("/profile?msg=프로필이 갱신되었습니다", 303)


# =====================================================
# REMINDERS — 팀장용 오늘 미작성자 리스트
# =====================================================
# v5H46 (2026-05-03): 사이드바·내부 링크 정합 별칭 — 6개 404 → 정상 페이지 리다이렉트
@app.get("/reminders", response_class=HTMLResponse)
async def reminders_alias(req: Request, sel_date: str = ""):
    return await reminders_page(req, sel_date)


@app.get("/board", response_class=HTMLResponse)
async def board_alias(req: Request):
    return RedirectResponse("/board/company", 303)


@app.get("/qms/dashboard", response_class=HTMLResponse)
async def qms_dashboard_alias(req: Request):
    return RedirectResponse("/qms", 303)


@app.get("/qc/reports", response_class=HTMLResponse)
async def qc_reports_alias(req: Request):
    return RedirectResponse("/qc/inspection-reports", 303)


@app.get("/admin/fx", response_class=HTMLResponse)
async def admin_fx_alias(req: Request):
    return RedirectResponse("/fx/rates", 303)


@app.get("/admin/permissions/report", response_class=HTMLResponse)
async def admin_perm_report_alias(req: Request):
    return RedirectResponse("/admin/permissions/report/expiring", 303)


@app.get("/stock", response_class=HTMLResponse)
async def stock_root_alias(req: Request):
    return RedirectResponse("/stock/balances", 303)


@app.get("/receipts", response_class=HTMLResponse)
async def receipts_alias(req: Request):
    return RedirectResponse("/stock/receipts", 303)


# v5H48 (2026-05-03): 사이드바·내부 폼 링크가 미존재 라우트 가리키는 8건 → 인접 페이지로 별칭
# v5H52b: 관리자 사용자/팀 신규 폼 + POST 핸들러 (실제 작동)
@app.get("/admin/users/new", response_class=HTMLResponse)
async def admin_users_new_form(req: Request):
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        teams = [dict(r) for r in c.execute(
            "SELECT id, code, name FROM teams ORDER BY display_order").fetchall()]
    return ctx(req, "admin_user_form.html", user=u, active="admin",
               target_user=None, teams=teams)


@app.post("/admin/users/new")
async def admin_users_new_submit(req: Request):
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    form = await req.form()
    name = (form.get("name") or "").strip()
    login_id = (form.get("login_id") or "").strip()
    password = (form.get("password") or "knk1234").strip()
    role = (form.get("role") or "member").strip()
    if not name or not login_id:
        return RedirectResponse("/admin/users/new?error=required", 303)
    with db_session() as c:
        ex = c.execute("SELECT id FROM users WHERE login_id=?", (login_id,)).fetchone()
        if ex:
            return RedirectResponse("/admin/users/new?error=duplicate", 303)
        team_id = form.get("team_id") or None
        c.execute(
            "INSERT INTO users(name, login_id, password, role, team_id, rank, "
            "email, is_active) VALUES(?,?,?,?,?,?,?,?)",
            (name, login_id, hash_pw(password), role, team_id,
             form.get("rank", ""), form.get("email", ""),
             int(form.get("is_active") or 1))
        )
    return RedirectResponse("/admin", 303)


@app.get("/admin/users/{uid:int}/edit", response_class=HTMLResponse)
async def admin_users_edit_form(req: Request, uid: int):
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        row = c.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
        if not row:
            return RedirectResponse("/admin", 303)
        teams = [dict(r) for r in c.execute(
            "SELECT id, code, name FROM teams ORDER BY display_order").fetchall()]
    # v5H114: 사용자 변경 이력 카드
    try:
        user_history = _logi.get_user_history(uid, limit=50)
    except Exception:
        user_history = []
    return ctx(req, "admin_user_form.html", user=u, active="admin",
               target_user=dict(row), teams=teams,
               user_history=user_history)


@app.post("/admin/users/{uid:int}/edit")
async def admin_users_edit_submit(req: Request, uid: int):
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    form = await req.form()
    name = (form.get("name") or "").strip()
    if not name:
        return RedirectResponse(f"/admin/users/{uid}/edit?error=required", 303)
    with db_session() as c:
        # v5H113 LOW#18: 변경 전 스냅샷
        old_row = c.execute(
            "SELECT name, role, team_id, rank, email, is_active "
            "FROM users WHERE id=?", (uid,)
        ).fetchone()
        old_data = dict(old_row) if old_row else {}
        # 비번 변경된 경우만 갱신
        new_pw = (form.get("password") or "").strip()
        if new_pw:
            c.execute(
                "UPDATE users SET name=?, role=?, team_id=?, rank=?, email=?, "
                "is_active=?, password=? WHERE id=?",
                (name, form.get("role", "member"),
                 form.get("team_id") or None,
                 form.get("rank", ""), form.get("email", ""),
                 int(form.get("is_active") or 1), hash_pw(new_pw), uid)
            )
        else:
            c.execute(
                "UPDATE users SET name=?, role=?, team_id=?, rank=?, email=?, "
                "is_active=? WHERE id=?",
                (name, form.get("role", "member"),
                 form.get("team_id") or None,
                 form.get("rank", ""), form.get("email", ""),
                 int(form.get("is_active") or 1), uid)
            )
        # v5H113 LOW#18: user_history diff 기록
        new_data = {
            "name": name,
            "role": form.get("role", "member"),
            "team_id": form.get("team_id") or None,
            "rank": form.get("rank", ""),
            "email": form.get("email", ""),
            "is_active": int(form.get("is_active") or 1),
        }
        _label_map = {"name":"이름","role":"권한","team_id":"팀","rank":"직급","email":"이메일","is_active":"활성"}
        for _k, _label in _label_map.items():
            ov = "" if old_data.get(_k) is None else str(old_data.get(_k))
            nv = "" if new_data.get(_k) is None else str(new_data.get(_k))
            if ov == nv:
                continue
            try:
                c.execute(
                    "INSERT INTO user_history(user_id, changed_by, field, "
                    "old_value, new_value, note) VALUES(?,?,?,?,?,?)",
                    (uid, u.get("id"), _label, ov, nv, "")
                )
            except Exception:
                pass
        if new_pw:
            try:
                c.execute(
                    "INSERT INTO user_history(user_id, changed_by, field, "
                    "old_value, new_value, note) VALUES(?,?,?,?,?,?)",
                    (uid, u.get("id"), "비밀번호", "***", "***", "비밀번호 재설정")
                )
            except Exception:
                pass
    return RedirectResponse("/admin", 303)


@app.get("/admin/teams/new", response_class=HTMLResponse)
async def admin_teams_new_form(req: Request):
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        users = [dict(r) for r in c.execute(
            "SELECT id, name, rank FROM users WHERE is_active=1 "
            "AND role IN ('leader','executive','member') ORDER BY name").fetchall()]
    return ctx(req, "admin_team_form.html", user=u, active="admin",
               team=None, users=users)


@app.post("/admin/teams/new")
async def admin_teams_new_submit(req: Request):
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    form = await req.form()
    code = (form.get("code") or "").strip()
    name = (form.get("name") or "").strip()
    if not code or not name:
        return RedirectResponse("/admin/teams/new?error=required", 303)
    with db_session() as c:
        ex = c.execute("SELECT id FROM teams WHERE code=?", (code,)).fetchone()
        if ex:
            return RedirectResponse("/admin/teams/new?error=duplicate", 303)
        c.execute(
            "INSERT INTO teams(code, name, sector, display_order, is_lab, leader_id) "
            "VALUES(?,?,?,?,?,?)",
            (code, name, form.get("sector", "공통"),
             int(form.get("display_order") or 99),
             int(form.get("is_lab") or 0),
             form.get("leader_id") or None)
        )
    return RedirectResponse("/admin", 303)


@app.get("/admin/teams/{tid:int}/edit", response_class=HTMLResponse)
async def admin_teams_edit_form(req: Request, tid: int):
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        row = c.execute("SELECT * FROM teams WHERE id=?", (tid,)).fetchone()
        if not row:
            return RedirectResponse("/admin", 303)
        users = [dict(r) for r in c.execute(
            "SELECT id, name, rank FROM users WHERE is_active=1 "
            "AND role IN ('leader','executive','member') ORDER BY name").fetchall()]
    # v5H114: 팀 변경 이력 카드
    try:
        team_history = _logi.get_team_history(tid, limit=50)
    except Exception:
        team_history = []
    return ctx(req, "admin_team_form.html", user=u, active="admin",
               team=dict(row), users=users,
               team_history=team_history)


@app.post("/admin/teams/{tid:int}/edit")
async def admin_teams_edit_submit(req: Request, tid: int):
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    form = await req.form()
    name = (form.get("name") or "").strip()
    if not name:
        return RedirectResponse(f"/admin/teams/{tid}/edit?error=required", 303)
    with db_session() as c:
        # v5H113 LOW#19: 변경 전 스냅샷
        old_row = c.execute(
            "SELECT code, name, sector, display_order, is_lab, leader_id "
            "FROM teams WHERE id=?", (tid,)
        ).fetchone()
        old_data = dict(old_row) if old_row else {}
        new_data = {
            "code": form.get("code", ""),
            "name": name,
            "sector": form.get("sector", "공통"),
            "display_order": int(form.get("display_order") or 99),
            "is_lab": int(form.get("is_lab") or 0),
            "leader_id": form.get("leader_id") or None,
        }
        c.execute(
            "UPDATE teams SET code=?, name=?, sector=?, display_order=?, "
            "is_lab=?, leader_id=? WHERE id=?",
            (new_data["code"], new_data["name"], new_data["sector"],
             new_data["display_order"], new_data["is_lab"],
             new_data["leader_id"], tid)
        )
        # v5H113 LOW#19: team_history diff 기록
        _label_map = {
            "code":"코드","name":"이름","sector":"섹터",
            "display_order":"표시순","is_lab":"연구소","leader_id":"팀장"
        }
        for _k, _label in _label_map.items():
            ov = "" if old_data.get(_k) is None else str(old_data.get(_k))
            nv = "" if new_data.get(_k) is None else str(new_data.get(_k))
            if ov == nv:
                continue
            try:
                c.execute(
                    "INSERT INTO team_history(team_id, changed_by, field, "
                    "old_value, new_value, note) VALUES(?,?,?,?,?,?)",
                    (tid, u.get("id"), _label, ov, nv, "")
                )
            except Exception:
                pass
    return RedirectResponse("/admin", 303)


# =====================================================
# v5H57 (2026-05-03 대표 지시) — 사업자등록증 자동 파서
# PDF/JPG/PNG 업로드 → 사업자번호/상호/대표자/주소 자동 추출
# 외부 API 0건 (pdfplumber + Tesseract OCR 로컬 실행)
# =====================================================
from . import biz_doc as _biz_doc
import tempfile as _tempfile

_CUSTOMER_FILES_DIR = os.path.join(BASE, "data", "customer_files")
os.makedirs(_CUSTOMER_FILES_DIR, exist_ok=True)


# v5H68 (2026-05-03): 프로젝트 라이프사이클 — 수주확정/추가발주 워크플로우
from . import project_workflow as _pwf


@app.post("/projects/{pid:int}/confirm-order")
async def projects_confirm_order(req: Request, pid: int):
    """프로젝트 수주 확정 — 관리번호 자동 발급 + 수주번호(SO) 발행.
    v5H196: 모달 폐기. 프로젝트 등록값(order_date/due_date/order_amount/
            unit_qty/unit_price/customer_po) 기준으로 자동 발행.
            unit_qty=N 면 호기 N개(1호기, 2호기 ... N호기) 자동 생성."""
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    with db_session() as c:
        # 프로젝트 등록값 조회
        proj = c.execute(
            "SELECT order_date, due_date, order_amount, customer_po, "
            "       COALESCE(unit_qty,1) AS unit_qty, unit_price, "
            "       COALESCE(currency,'KRW') AS currency, "
            "       COALESCE(project_type,'NEW_EQUIP') AS project_type "
            "FROM projects WHERE id=?", (pid,)
        ).fetchone()
        if not proj:
            return JSONResponse({"ok": False, "message": "프로젝트 없음"}, 404)
        _qty = max(1, min(100, int(proj["unit_qty"] or 1)))
        try:
            _up = float(proj["unit_price"]) if proj["unit_price"] is not None else 0.0
        except Exception:
            _up = 0.0
        _amt_total = float(proj["order_amount"] or 0)
        if _up <= 0 and _qty > 0:
            _up = _amt_total / _qty
        # 호기 라벨: project_type 기준 (NEW_EQUIP→N호기 / OTHER→N건 등)
        _ptype = (proj["project_type"] or "NEW_EQUIP").upper()
        _units = []
        for i in range(_qty):
            _units.append({
                "label": _logi.project_unit_label(_ptype, i + 1),
                "amount": _up,
                "due_date": proj["due_date"] or "",
                "ship_to": "",
                "currency": proj["currency"] or "KRW",
                "note": "",
            })
        res = _pwf.confirm_order_multi(
            c, pid,
            units=_units,
            order_date=proj["order_date"] or "",
            created_by=u.get("id") or 0,
            po_number=proj["customer_po"] or "",
        )
    if not res or not res.get("ok"):
        return JSONResponse(res or {"ok": False, "message": "발행 실패"}, 500)
    # 결과: groups 안의 첫 번째 SO 정보 + 기본 응답
    grp = (res.get("groups") or [{}])[0]
    return JSONResponse({
        "ok": True,
        "mgmt_code": res.get("mgmt_code") or grp.get("mgmt_code"),
        "so_no": grp.get("so_no") or res.get("so_no"),
        "qty": _qty,
        "message": f"수주 확정 완료 — {_qty}대 호기 자동 발행",
    })


# v5H73: 본인 비밀번호 검증 API (삭제 등 위험 액션 전 재인증)
@app.post("/auth/verify-password")
async def auth_verify_password(req: Request):
    """현재 로그인 사용자의 비밀번호를 검증.
    Body: password=...
    응답: {ok: True} 또는 {ok: False, error: '...'}"""
    u = get_user(req)
    if not u:
        return JSONResponse({"ok": False, "error": "로그인 필요"}, 401)
    form = await req.form()
    pw = (form.get("password") or "").strip()
    if not pw:
        return JSONResponse({"ok": False, "error": "비밀번호를 입력하세요"})
    # DB 에서 사용자 password 해시 다시 조회 (세션 객체에 password 없을 수 있음)
    with db_session() as c:
        row = c.execute("SELECT password FROM users WHERE id=?", (u["id"],)).fetchone()
    if not row:
        return JSONResponse({"ok": False, "error": "사용자 없음"}, 404)
    if hash_pw(pw) != row["password"]:
        return JSONResponse({"ok": False, "error": "비밀번호가 일치하지 않습니다"})
    return JSONResponse({"ok": True})


@app.post("/projects/{pid:int}/header-edit")
async def projects_header_edit(req: Request, pid: int):
    """v5H177: 프로젝트 상세 헤더 인라인 편집.
    필드별 단건 PATCH — name / customer_id / currency / fx_rate / is_export / note / order_date / due_date.
    Tier 1 필드(mgmt_code, biz_div, project_type, po_type, created_*) 는 수정 불가.
    """
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    if not can_use_sales(u):
        return JSONResponse({"error": "권한 없음"}, 403)
    form = await req.form()
    field = (form.get("field") or "").strip()
    raw_value = form.get("value", "")
    LOCKED = {"mgmt_code", "biz_div", "project_type", "po_type",
              "created_at", "created_by", "id", "code"}
    EDITABLE = {"name", "customer_id", "currency", "fx_rate", "amount_krw",
                "is_export", "note", "order_date", "due_date", "model"}
    if field in LOCKED:
        return JSONResponse({"ok": False, "message": f"'{field}' 은 영구 잠금 필드"}, 400)
    if field not in EDITABLE:
        return JSONResponse({"ok": False, "message": f"허용되지 않은 필드: {field}"}, 400)
    # 값 정규화
    val: object = (raw_value or "").strip() if isinstance(raw_value, str) else raw_value
    if field == "currency":
        val = (val or "KRW").upper()
        if val not in ("KRW", "USD", "VND", "JPY", "CNY", "EUR"):
            return JSONResponse({"ok": False, "message": "허용 통화: KRW/USD/VND/JPY/CNY/EUR"}, 400)
    elif field == "fx_rate":
        try:
            val = float(val) if val else None
            if val is not None and val <= 0: val = None
        except ValueError:
            return JSONResponse({"ok": False, "message": "환율 형식 오류"}, 400)
    elif field == "amount_krw":
        try:
            val = float(str(val).replace(",", "")) if val else None
        except ValueError:
            val = None
    elif field == "customer_id":
        try:
            val = int(val) if val else None
        except ValueError:
            return JSONResponse({"ok": False, "message": "고객 ID 형식 오류"}, 400)
    elif field == "is_export":
        val = "1" if str(val).strip() in ("1", "true", "True", "on", "yes") else "0"
    # DB 업데이트 + 이력 기록
    with db_session() as c:
        old_row = c.execute(f"SELECT {field} FROM projects WHERE id=?", (pid,)).fetchone()
        if not old_row:
            return JSONResponse({"ok": False, "message": "프로젝트 없음"}, 404)
        old_val = old_row[0]
        if (old_val or "") == (val or ""):
            return JSONResponse({"ok": True, "message": "변동 없음", "value": val})
        c.execute(f"UPDATE projects SET {field}=? WHERE id=?", (val, pid))
        # v5H187: 통화 변경 시 자식(orders, order_items) 도 즉시 cascade.
        #   기존엔 startup 백필에만 의존 → 인라인 편집 후 즉시 반영 안되던 결함.
        #   안전: 수금이력 없는 SO 만 변경 (이미 수금된 외화는 그대로 유지).
        cascade_info = None
        if field == "currency" and val:
            try:
                _ord_cols = {r2[1] for r2 in c.execute("PRAGMA table_info(orders)").fetchall()}
                if "currency" in _ord_cols:
                    _r1 = c.execute("""
                        UPDATE orders SET currency = ?
                         WHERE project_id = ?
                           AND COALESCE(currency,'KRW') != ?
                           AND NOT EXISTS (
                                 SELECT 1 FROM receipts_payment rp WHERE rp.order_id = orders.id)
                    """, (val, pid, val))
                    n_orders = _r1.rowcount
                else:
                    n_orders = 0
                _oi_cols = {r2[1] for r2 in c.execute("PRAGMA table_info(order_items)").fetchall()}
                if "currency" in _oi_cols:
                    _r2 = c.execute("""
                        UPDATE order_items SET currency = ?
                         WHERE order_id IN (SELECT id FROM orders WHERE project_id = ?)
                           AND COALESCE(currency,'') != ?
                    """, (val, pid, val))
                    n_items = _r2.rowcount
                else:
                    n_items = 0
                if n_orders or n_items:
                    cascade_info = f"SO {n_orders}건 + 호기 {n_items}건 통화 동기화"
            except Exception:
                pass
        # 변경 이력 (project_history 가 있으면)
        try:
            _note = f"인라인 편집 ({field})"
            if cascade_info:
                _note += f" · {cascade_info}"
            c.execute(
                "INSERT INTO project_history(project_id, changed_by, field, old_value, new_value, note) "
                "VALUES(?,?,?,?,?,?)",
                (pid, u.get("id"), field, str(old_val or ""), str(val or ""), _note)
            )
        except Exception:
            pass
    return JSONResponse({"ok": True, "field": field, "value": val,
                          "cascade": cascade_info})


@app.post("/projects/{pid:int}/presales-edit")
async def projects_presales_edit(req: Request, pid: int):
    """v5H212: 수주 전 내역 인라인 편집 — proposal/quotation 의 submitted/memo 단건 PATCH.
    필드: proposal_submitted | proposal_memo | quotation_submitted | quotation_memo
    """
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    if not can_use_sales(u):
        return JSONResponse({"error": "권한 없음"}, 403)
    form = await req.form()
    field = (form.get("field") or "").strip()
    raw_value = form.get("value", "")
    ALLOWED = {"proposal_submitted", "proposal_memo",
               "quotation_submitted", "quotation_memo"}
    if field not in ALLOWED:
        return JSONResponse({"ok": False, "message": f"허용되지 않은 필드: {field}"}, 400)
    if field.endswith("_submitted"):
        val: object = 1 if str(raw_value).strip() in ("1", "true", "True", "on", "yes") else 0
    else:
        val = (raw_value or "").strip() if isinstance(raw_value, str) else (raw_value or "")
        if val == "":
            val = None
    with db_session() as c:
        old_row = c.execute(f"SELECT {field} FROM projects WHERE id=?", (pid,)).fetchone()
        if not old_row:
            return JSONResponse({"ok": False, "message": "프로젝트 없음"}, 404)
        old_val = old_row[0]
        if (old_val or 0 if field.endswith("_submitted") else (old_val or "")) == (val or 0 if field.endswith("_submitted") else (val or "")):
            return JSONResponse({"ok": True, "message": "변동 없음", "value": val})
        c.execute(f"UPDATE projects SET {field}=? WHERE id=?", (val, pid))
        # 변경 이력 — 사람이 읽기 쉬운 라벨로
        _label_map = {
            "proposal_submitted": "제안서 제출 여부",
            "proposal_memo":      "제안서 메모",
            "quotation_submitted":"견적서 제출 여부",
            "quotation_memo":     "견적서 메모",
        }
        try:
            if field.endswith("_submitted"):
                ov = "제출완료" if old_val else "미제출"
                nv = "제출완료" if val else "미제출"
            else:
                ov = (old_val or "")[:80]
                nv = (val or "")[:80] if val else ""
            c.execute(
                "INSERT INTO project_history(project_id, changed_by, field, old_value, new_value, note) "
                "VALUES(?,?,?,?,?,?)",
                (pid, u.get("id"), _label_map.get(field, field), str(ov), str(nv), "수주 전 내역 편집")
            )
        except Exception:
            pass
    return JSONResponse({"ok": True, "field": field, "value": val})


@app.post("/projects/{pid:int}/quick-status")
async def projects_quick_status(req: Request, pid: int):
    """v5H97: 프로젝트 상태 인라인 변경 (상세 페이지에서 클릭 한 번).
    v5H130 (2026-05-05): WON_STATUSES 진입 시 SO 자동 발행 (1호기) — form-POST
    경로(projects/new, projects/{pid}/edit)와 동일한 v5H87 안전망. 인라인
    경로에서 SO 누락 → '추가 발주' 버튼이 1호기를 덮어쓰던 결함을 차단."""
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    if not can_use_sales(u):
        return JSONResponse({"error": "권한 없음"}, 403)
    form = await req.form()
    new_status = (form.get("status") or "").strip()
    if new_status not in _logi.LOGI_STATUSES:
        return JSONResponse({"ok": False, "message": "유효하지 않은 상태"}, 400)
    auto_so_issued = False
    auto_so_no = None
    try:
      with db_session() as c:
        cur = c.execute(
            "SELECT status, mgmt_code, biz_div, order_amount, order_date, "
            "due_date, customer_po, "
            "COALESCE(unit_qty,1) AS unit_qty, unit_price, "
            "COALESCE(currency,'KRW') AS currency, "
            "COALESCE(project_type,'NEW_EQUIP') AS project_type "
            "FROM projects WHERE id=?", (pid,)
        ).fetchone()
        if not cur:
            return JSONResponse({"ok": False, "message": "프로젝트 없음"}, 404)
        old_status = cur["status"] or ""
        # v5H143: project_type 먼저 확정 (이전엔 분기 이후에 정의돼 NameError → HTTP 500 결함 1)
        try:
            _cur_ptype = (cur["project_type"] if "project_type" in cur.keys() else "") or "NEW_EQUIP"
        except Exception:
            _cur_ptype = "NEW_EQUIP"
        # v5H214: status 변경 시 stage 도 자동 동기화 (수주확정 코드 발급 분기 전에 선반영)
        _new_stage = stage_from_status(new_status)
        c.execute("UPDATE projects SET status=?, stage=?, updated_at=? WHERE id=?",
                  (new_status, _new_stage, _logi._logi_now() if hasattr(_logi, "_logi_now") else None, pid))
        # v5H101: 변경 이력
        _logi.log_project_change(c, pid, u.get("id"), "상태", old_status, new_status,
                                  note="인라인 빠른 변경")
        # status 가 won 으로 바뀌었는데 mgmt_code 없으면 발급
        # v5H142: NEW_EQUIP(T/M) 만 관리번호 발급
        # v5H150: OTHER 는 'K' prefix 로 발급
        _need_code_qs = (
            new_status in _logi.WON_STATUSES and not cur["mgmt_code"]
            and (
                (cur["biz_div"] in ("T", "M") and _cur_ptype == "NEW_EQUIP")
                or _cur_ptype == "OTHER"
            )
        )
        if _need_code_qs:
            try:
                _prefix_qs = "K" if _cur_ptype == "OTHER" else cur["biz_div"]
                code = _logi.generate_mgmt_code(_prefix_qs)
                c.execute("UPDATE projects SET mgmt_code=?, stage='수주확정' WHERE id=?",
                          (code, pid))
            except Exception:
                pass
        # v5H130: WON_STATUSES + SO 0건 + order_amount > 0 → 자동 SO 발행
        # v5H132: unit_qty N → N개 호기 라인 (단가=unit_price 또는 amt/qty)
        # v5H142: NEW_EQUIP 만 자동 SO 발행 (소모품/수리는 consumable_orders 도메인)
        if new_status in _logi.WON_STATUSES and _cur_ptype == "NEW_EQUIP":
            try:
                amt0 = float(cur["order_amount"] or 0)
                if amt0 > 0:
                    exists = c.execute(
                        "SELECT 1 FROM orders WHERE project_id=? LIMIT 1", (pid,)
                    ).fetchone()
                    if not exists:
                        try:
                            _qty0 = max(1, min(100, int(cur["unit_qty"] or 1)))
                        except Exception:
                            _qty0 = 1
                        try:
                            _up0 = float(cur["unit_price"]) if cur["unit_price"] is not None else 0.0
                        except Exception:
                            _up0 = 0.0
                        if _up0 <= 0:
                            _up0 = amt0 / _qty0
                        # v5H137: project_type 기준 라벨
                        _ptype_q = cur["project_type"] or "NEW_EQUIP"
                        _units_list = [{
                            "label": _logi.project_unit_label(_ptype_q, i + 1),
                            "amount": _up0,
                            "due_date": cur["due_date"] or "",
                            "ship_to": "",
                            "note": "",
                        } for i in range(_qty0)]
                        res = _pwf.confirm_order_multi(
                            c, int(pid),
                            units=_units_list,
                            order_date=cur["order_date"] or "",
                            created_by=u.get("id") or 0,
                            po_number=cur["customer_po"] or "",
                        )
                        if res and res.get("ok"):
                            auto_so_issued = True
                            grp = (res.get("groups") or [{}])[0]
                            auto_so_no = grp.get("so_no") or res.get("so_no")
                            # v5H137: project_type 라벨
                            _lbl = (_logi.project_unit_label(_ptype_q, _qty0)
                                    if _qty0 > 1
                                    else _logi.project_unit_label(_ptype_q, 1))
                            # v5H183: SO 발행 상세 — 통화·납기·단가·합계 풀 로깅
                            _ccy_h = (cur["currency"] if "currency" in cur.keys() else None) or "KRW"
                            _due_h = cur["due_date"] or "(미지정)"
                            _logi.log_project_change(
                                c, pid, u.get("id"), "수주발행(자동)",
                                "", auto_so_no or _lbl,
                                note=(f"진행중 진입 시 {_qty0}대 SO 자동 발행 — "
                                      f"단가 {_up0:,.0f} {_ccy_h} × {_qty0}대 = {amt0:,.0f} {_ccy_h} · "
                                      f"납기 {_due_h} · 라벨 {_lbl}")
                            )
                            # 각 호기 라인 개별 기록
                            for _i in range(_qty0):
                                _i_lbl = _logi.project_unit_label(_ptype_q, _i + 1)
                                _logi.log_project_change(
                                    c, pid, u.get("id"), f"호기 추가({_i_lbl})",
                                    "", f"{_up0:,.0f} {_ccy_h}",
                                    note=f"SO {auto_so_no} · 납기 {_due_h}"
                                )
            except Exception:
                pass
    except Exception as _qs_err:
        # v5H143: HTTP 500 대신 친절 JSON (모달/토스트 표시 가능)
        return JSONResponse({"ok": False,
                             "message": f"상태 변경 중 오류: {type(_qs_err).__name__} — {str(_qs_err)[:160]}"},
                            500)
    msg = f"상태 → {new_status}"
    if auto_so_issued:
        msg += f" · SO {auto_so_no or ''} 자동 발행"
    return JSONResponse({"ok": True, "message": msg,
                         "auto_so_issued": auto_so_issued,
                         "auto_so_no": auto_so_no})


@app.post("/sales/orders/items/{iid:int}/edit")
async def sales_order_item_edit(req: Request, iid: int):
    """v5H93: 호기 라인(order_items) 인라인 편집 — 라벨/금액/비고."""
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    if not can_use_sales(u):
        return JSONResponse({"error": "권한 없음"}, 403)
    form = await req.form()
    label = (form.get("label") or "").strip()
    raw_a = (form.get("amount") or "0").strip().replace(",", "")
    note  = (form.get("note") or "").strip()
    # v5H177: 호기별 발주일/납기/납품처 override
    u_order = (form.get("order_date") or "").strip() or None
    u_due = (form.get("due_date") or "").strip() or None
    u_ship = (form.get("ship_to") or "").strip() or None
    # v5H189: 호기별 통화 override (KRW/USD/VND/JPY/CNY/EUR)
    u_cur_raw = (form.get("currency") or "").strip().upper()
    u_cur = u_cur_raw if u_cur_raw in ("KRW","USD","VND","JPY","CNY","EUR") else None
    # v5H197: 호기별 거래구분 (0=내수, 1=수출, "" = 프로젝트 상속)
    u_iex_raw = (form.get("is_export") or "").strip()
    if u_iex_raw in ("0", "1"):
        u_iex = int(u_iex_raw)
    else:
        u_iex = None  # 비전송 = 변경 안함
    try:
        amt = float(raw_a) if raw_a else 0
    except ValueError:
        return JSONResponse({"ok": False, "message": "금액 형식 오류"}, 400)
    if amt < 0:
        return JSONResponse({"ok": False, "message": "금액은 0 이상이어야 합니다"}, 400)
    with db_session() as c:
        # v5H100 / v5H183: 변경 전 값 캡처 (이력 기록용) — order_date/due_date/ship_to 도 포함
        it = c.execute(
            "SELECT oi.order_id, oi.unit_label AS old_lbl, oi.amount AS old_amt, "
            "       oi.line_note AS old_note, "
            "       oi.order_date AS old_oi_ord, oi.due_date AS old_oi_due, "
            "       oi.ship_to AS old_oi_ship, "
            "       o.status, o.project_id, o.total_amount AS old_total, "
            "       o.order_no AS old_order_no, "
            "       COALESCE(o.currency,'KRW') AS o_ccy, "
            "       o.order_date AS o_order_date, o.due_date AS o_due_date, "
            "       o.ship_to AS o_ship "
            "FROM order_items oi JOIN orders o ON o.id = oi.order_id "
            "WHERE oi.id=?", (iid,)
        ).fetchone()
        if not it:
            return JSONResponse({"ok": False, "message": "라인을 찾을 수 없습니다"}, 404)
        st = (it["status"] or "").upper()
        if st in ("SHIPPED", "INVOICED", "PAID", "CANCELLED"):
            return JSONResponse({
                "ok": False,
                "message": f"{st} 상태 SO 의 호기는 수정 불가"
            }, 400)
        # v5H177: order_items 에 override 컬럼이 있을 때만 같이 업데이트 (백워드 호환)
        try:
            _oicols = {r2[1] for r2 in c.execute("PRAGMA table_info(order_items)").fetchall()}
        except Exception:
            _oicols = set()
        if {"order_date","due_date","ship_to"} <= _oicols:
            # SO 그룹값(부모) 조회 — 동일하면 override NULL 로 저장 (상속)
            _so = c.execute(
                "SELECT order_date, due_date, ship_to, COALESCE(currency,'KRW') AS currency FROM orders WHERE id=?",
                (it["order_id"],)
            ).fetchone()
            _so_d = dict(_so) if _so else {}
            ov_o = u_order if (u_order and u_order != (_so_d.get("order_date") or "")) else None
            ov_d = u_due if (u_due and u_due != (_so_d.get("due_date") or "")) else None
            ov_s = u_ship if (u_ship and u_ship != (_so_d.get("ship_to") or "")) else None
            # v5H189: 통화 — SO 부모와 동일하면 NULL(상속), 다르면 override
            ov_c = None
            if "currency" in _oicols and u_cur:
                ov_c = u_cur if u_cur != (_so_d.get("currency") or "KRW") else None
            # v5H197: 거래구분 — 폼에서 받은 그대로 (NULL = 변경 안 함이 아니라 명시 입력만)
            cols_set = ["unit_label=?", "unit_price=?", "amount=?", "line_note=?",
                        "order_date=?", "due_date=?", "ship_to=?"]
            vals_set = [label or None, amt, amt, note or None, ov_o, ov_d, ov_s]
            if "currency" in _oicols:
                cols_set.append("currency=?"); vals_set.append(ov_c)
            if "is_export" in _oicols and u_iex is not None:
                cols_set.append("is_export=?"); vals_set.append(u_iex)
            vals_set.append(iid)
            c.execute(
                f"UPDATE order_items SET {','.join(cols_set)} WHERE id=?",
                tuple(vals_set)
            )
        else:
            c.execute(
                "UPDATE order_items SET unit_label=?, unit_price=?, amount=?, line_note=? WHERE id=?",
                (label or None, amt, amt, note or None, iid)
            )
        # SO total_amount = SUM(items.amount), unit_label 재구성
        oid = it["order_id"]
        rows = c.execute(
            "SELECT unit_label, amount FROM order_items WHERE order_id=? ORDER BY id ASC",
            (oid,)
        ).fetchall()
        new_total = sum(float(r["amount"] or 0) for r in rows)
        new_label = " · ".join((r["unit_label"] or f"호기 {i+1}") for i, r in enumerate(rows))
        c.execute("UPDATE orders SET total_amount=?, unit_label=? WHERE id=?",
                  (new_total, new_label, oid))
        # 프로젝트 합계 동기화
        try:
            pid = it["project_id"]
            if pid:
                row = c.execute(
                    "SELECT COALESCE(SUM(total_amount),0) FROM orders WHERE project_id=?",
                    (pid,)
                ).fetchone()
                c.execute("UPDATE projects SET order_amount=? WHERE id=?",
                          (float(row[0] or 0), pid))
        except Exception:
            pass
        # v5H100: 변경 이력 기록 (before → after) — order_status_history
        try:
            old_lbl = it["old_lbl"] or "—"
            old_amt = float(it["old_amt"] or 0)
            old_total = float(it["old_total"] or 0)
            parts = [f"호기 라인 수정 [{old_lbl}"]
            if old_lbl != (label or "—"):
                parts.append(f"→ {label or '—'}")
            parts.append("]")
            change_msgs = []
            if abs(old_amt - amt) > 0.5:
                change_msgs.append(f"단가 {old_amt:,.0f} → {amt:,.0f}")
            old_note_v = it["old_note"] or ""
            if old_note_v != (note or ""):
                change_msgs.append(f"비고 변경")
            if abs(old_total - new_total) > 0.5:
                change_msgs.append(f"SO 합계 {old_total:,.0f} → {new_total:,.0f}")
            if not change_msgs:
                change_msgs.append("(변동 없음)")
            note_msg = " ".join(parts) + " · " + " / ".join(change_msgs)
            c.execute(
                "INSERT INTO order_status_history(order_id, from_status, to_status, "
                "changed_by, note) VALUES(?,?,?,?,?)",
                (oid, st, st, u.get("id"), note_msg)
            )
        except Exception:
            pass
        # v5H183: project_history 에도 호기별 변경 상세 기록 (이력 페이지 노출용)
        try:
            _pid = it["project_id"]
            if _pid:
                _ccy = it["o_ccy"] or "KRW"
                _order_no = it["old_order_no"] or ""
                _new_lbl = (label or old_lbl)
                _detail_changes = []
                if abs(old_amt - amt) > 0.5:
                    _detail_changes.append(f"단가 {old_amt:,.0f} → {amt:,.0f} {_ccy}")
                # 호기별 발주일 / 납기 / 납품처 (override) 변경 검출
                _so_ord = (it["o_order_date"] or "")
                _so_due = (it["o_due_date"] or "")
                _so_sh = (it["o_ship"] or "")
                _old_eff_ord = (it["old_oi_ord"] or _so_ord) or ""
                _old_eff_due = (it["old_oi_due"] or _so_due) or ""
                _old_eff_sh  = (it["old_oi_ship"] or _so_sh) or ""
                _new_eff_ord = (u_order or _so_ord) or ""
                _new_eff_due = (u_due or _so_due) or ""
                _new_eff_sh  = (u_ship or _so_sh) or ""
                if _old_eff_ord != _new_eff_ord:
                    _detail_changes.append(f"발주일 {_old_eff_ord or '(미)'} → {_new_eff_ord or '(미)'}")
                if _old_eff_due != _new_eff_due:
                    _detail_changes.append(f"납기 {_old_eff_due or '(미)'} → {_new_eff_due or '(미)'}")
                if _old_eff_sh != _new_eff_sh:
                    _detail_changes.append(f"납품처 {_old_eff_sh or '(미)'} → {_new_eff_sh or '(미)'}")
                if (it["old_note"] or "") != (note or ""):
                    _detail_changes.append(f"비고 변경")
                if (old_lbl or "") != (label or ""):
                    _detail_changes.append(f"라벨 {old_lbl or '—'} → {label or '—'}")
                # v5H189: 호기별 통화 변경 감지
                if u_cur and u_cur != _ccy:
                    _detail_changes.append(f"통화 {_ccy} → {u_cur}")
                if _detail_changes:
                    _logi.log_project_change(
                        c, _pid, u.get("id"),
                        f"호기 수정({_new_lbl})",
                        "", " / ".join(_detail_changes),
                        note=f"SO {_order_no}"
                    )
        except Exception:
            pass
    return JSONResponse({"ok": True, "message": "라인 수정 완료"})


UNIT_STATUSES = ("진행중", "납품완료", "취소", "보류")


@app.post("/sales/orders/items/{iid:int}/status")
async def sales_order_item_status(req: Request, iid: int):
    """v5H186: 호기 라인 상태 변경 (개별)."""
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    if not can_use_sales(u):
        return JSONResponse({"error": "권한 없음"}, 403)
    form = await req.form()
    new_status = (form.get("status") or "").strip()
    if new_status not in UNIT_STATUSES:
        return JSONResponse({"ok": False,
                              "message": f"허용: {' / '.join(UNIT_STATUSES)}"}, 400)
    with db_session() as c:
        try:
            _oicols = {r[1] for r in c.execute("PRAGMA table_info(order_items)").fetchall()}
        except Exception:
            _oicols = set()
        if "unit_status" not in _oicols:
            return JSONResponse({"ok": False,
                                  "message": "unit_status 컬럼 미생성 — 서버 재시작 필요"}, 500)
        it = c.execute(
            "SELECT oi.unit_label, oi.unit_status, o.project_id, o.order_no "
            "FROM order_items oi JOIN orders o ON o.id = oi.order_id "
            "WHERE oi.id=?", (iid,)
        ).fetchone()
        if not it:
            return JSONResponse({"ok": False, "message": "라인 없음"}, 404)
        old_st = it["unit_status"] or "진행중"
        if old_st == new_status:
            return JSONResponse({"ok": True, "message": "변동 없음"})
        c.execute("UPDATE order_items SET unit_status=? WHERE id=?",
                  (new_status, iid))
        # 이력 기록
        try:
            _logi.log_project_change(
                c, it["project_id"], u.get("id"),
                f"호기 상태({it['unit_label'] or '—'})",
                old_st, new_status,
                note=f"SO {it['order_no']} · 인라인 변경"
            )
        except Exception:
            pass
    return JSONResponse({"ok": True, "old": old_st, "new": new_status})


@app.post("/projects/{pid:int}/units/bulk-status")
async def projects_units_bulk_status(req: Request, pid: int):
    """v5H186: 프로젝트 전체 호기 라인 상태 일괄 변경.
    옵션: order_id 지정 시 해당 SO 호기들만, 미지정 시 프로젝트 전체."""
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    if not can_use_sales(u):
        return JSONResponse({"error": "권한 없음"}, 403)
    form = await req.form()
    new_status = (form.get("status") or "").strip()
    if new_status not in UNIT_STATUSES:
        return JSONResponse({"ok": False,
                              "message": f"허용: {' / '.join(UNIT_STATUSES)}"}, 400)
    scope_oid = form.get("order_id")
    try:
        scope_oid_i = int(scope_oid) if scope_oid else None
    except (TypeError, ValueError):
        scope_oid_i = None
    with db_session() as c:
        try:
            _oicols = {r[1] for r in c.execute("PRAGMA table_info(order_items)").fetchall()}
        except Exception:
            _oicols = set()
        if "unit_status" not in _oicols:
            return JSONResponse({"ok": False,
                                  "message": "unit_status 컬럼 미생성"}, 500)
        if scope_oid_i:
            r = c.execute(
                "UPDATE order_items SET unit_status=? "
                "WHERE order_id=? AND COALESCE(unit_status,'진행중') != ?",
                (new_status, scope_oid_i, new_status)
            )
        else:
            r = c.execute(
                "UPDATE order_items SET unit_status=? "
                "WHERE order_id IN (SELECT id FROM orders WHERE project_id=?) "
                "  AND COALESCE(unit_status,'진행중') != ?",
                (new_status, pid, new_status)
            )
        # 이력 기록
        try:
            scope_lbl = (f"SO #{scope_oid_i}" if scope_oid_i else "전체 호기")
            _logi.log_project_change(
                c, pid, u.get("id"),
                f"호기 상태 일괄 변경 ({scope_lbl})",
                "", new_status,
                note=f"{r.rowcount}건 변경"
            )
        except Exception:
            pass
    return JSONResponse({"ok": True, "changed": r.rowcount,
                          "message": f"{r.rowcount}건 변경 완료"})


@app.post("/sales/orders/items/{iid:int}/delete")
async def sales_order_item_delete(req: Request, iid: int):
    """v5H93: 호기 라인 삭제 — 1건만 제거. unit_qty/total_amount 자동 재계산."""
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    if not can_use_sales(u):
        return JSONResponse({"error": "권한 없음"}, 403)
    with db_session() as c:
        # v5H100: 삭제 전 값 캡처
        it = c.execute(
            "SELECT oi.order_id, oi.unit_label AS lbl, oi.amount AS amt, oi.line_note AS lnote, "
            "       o.status, o.project_id, o.total_amount AS old_total "
            "FROM order_items oi JOIN orders o ON o.id = oi.order_id "
            "WHERE oi.id=?", (iid,)
        ).fetchone()
        if not it:
            return JSONResponse({"ok": False, "message": "라인을 찾을 수 없습니다"}, 404)
        st = (it["status"] or "").upper()
        if st in ("SHIPPED", "INVOICED", "PAID", "CANCELLED"):
            return JSONResponse({
                "ok": False,
                "message": f"{st} 상태 SO 의 호기는 삭제 불가"
            }, 400)
        oid = it["order_id"]
        del_lbl = it["lbl"] or "—"
        del_amt = float(it["amt"] or 0)
        old_total = float(it["old_total"] or 0)
        c.execute("DELETE FROM order_items WHERE id=?", (iid,))
        # SO 재계산
        rows = c.execute(
            "SELECT unit_label, amount FROM order_items WHERE order_id=? ORDER BY id ASC",
            (oid,)
        ).fetchall()
        new_total = sum(float(r["amount"] or 0) for r in rows)
        new_qty = max(1, len(rows))
        new_label = " · ".join((r["unit_label"] or f"호기 {i+1}") for i, r in enumerate(rows))
        c.execute("UPDATE orders SET total_amount=?, unit_qty=?, unit_label=? WHERE id=?",
                  (new_total, new_qty, new_label, oid))
        # 프로젝트 합계 동기화
        try:
            pid = it["project_id"]
            if pid:
                row = c.execute(
                    "SELECT COALESCE(SUM(total_amount),0) FROM orders WHERE project_id=?",
                    (pid,)
                ).fetchone()
                c.execute("UPDATE projects SET order_amount=? WHERE id=?",
                          (float(row[0] or 0), pid))
        except Exception:
            pass
        # v5H100: 삭제 이력
        try:
            note_msg = (f"호기 라인 삭제 [{del_lbl}] · 단가 {del_amt:,.0f} 제거 · "
                        f"SO 합계 {old_total:,.0f} → {new_total:,.0f}")
            c.execute(
                "INSERT INTO order_status_history(order_id, from_status, to_status, "
                "changed_by, note) VALUES(?,?,?,?,?)",
                (oid, st, st, u.get("id"), note_msg)
            )
        except Exception:
            pass
    return JSONResponse({"ok": True, "message": "라인 삭제 완료"})


@app.post("/sales/orders/{oid:int}/add-unit")
async def sales_orders_add_unit(req: Request, oid: int):
    """v5H90: 기존 SO 에 호기 라인 추가 (단가가 다른 추가 발주 대응).
    동일 SO 안에서 호기별 단가가 다를 수 있게 order_items 새 행 INSERT
    + orders.unit_qty++/total_amount += amount 갱신."""
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    if not can_use_sales(u):
        return JSONResponse({"error": "권한 없음"}, 403)
    form = await req.form()
    label = (form.get("label") or "").strip()
    raw_a = (form.get("amount") or "0").strip().replace(",", "")
    note  = (form.get("note") or "").strip()
    cur_v = (form.get("currency") or "").strip().upper() or "KRW"
    if cur_v not in ("KRW", "USD", "VND"):
        cur_v = "KRW"
    # v5H110: 수량 — N대 한 번에 추가 (모두 같은 단가, 라벨은 자동 증가)
    raw_qty = (form.get("qty") or "1").strip()
    try:
        bulk_qty = int(float(raw_qty))
    except ValueError:
        bulk_qty = 1
    if bulk_qty < 1:
        bulk_qty = 1
    if bulk_qty > 100:
        return JSONResponse({"ok": False, "message": "한 번에 100대 초과 불가"}, 400)
    try:
        amt = float(raw_a) if raw_a else 0
    except ValueError:
        return JSONResponse({"ok": False, "message": "금액 형식 오류"}, 400)
    if amt <= 0:
        return JSONResponse({"ok": False, "message": "금액은 0보다 커야 합니다"}, 400)
    with db_session() as c:
        cur = c.execute(
            "SELECT status, project_id, unit_qty, unit_label, total_amount "
            "FROM orders WHERE id=?", (oid,)
        ).fetchone()
        if not cur:
            return JSONResponse({"ok": False, "message": "수주를 찾을 수 없습니다"}, 404)
        st = (cur["status"] or "").upper()
        if st in ("SHIPPED", "INVOICED", "PAID", "CANCELLED"):
            return JSONResponse({
                "ok": False,
                "message": f"{st} 상태 SO 는 호기 추가 불가"
            }, 400)
        # v5H109: 자동 라벨은 프로젝트 전체 호기 라인 개수 기준 (관리코드 우선)
        try:
            items_in_so = c.execute(
                "SELECT COUNT(*) FROM order_items WHERE order_id=?", (oid,)
            ).fetchone()[0] or 0
            project_id_local = cur["project_id"]
            items_in_project = c.execute(
                "SELECT COUNT(*) FROM order_items oi "
                "JOIN orders o ON o.id = oi.order_id WHERE o.project_id=?",
                (project_id_local,)
            ).fetchone()[0] or 0
        except Exception:
            items_in_so = 0
            items_in_project = 0
        # v5H110: 시작 라벨에서 숫자 추출 → bulk 시 N개 자동 증가
        import re as _re
        base_no = None
        if label:
            m = _re.match(r"^(\d+)호기$", label)
            if m:
                base_no = int(m.group(1))
        if base_no is None:
            base_no = items_in_project + 1
        # bulk_qty 만큼 라벨 생성: ['5호기', '6호기', ...]
        labels_bulk = [f"{base_no + i}호기" for i in range(bulk_qty)]
        # bulk 0번째가 사용자 입력 커스텀 라벨이면 그대로 첫 라벨 사용
        if label and not _re.match(r"^\d+호기$", label):
            labels_bulk[0] = label  # 커스텀 라벨 유지, 나머지는 N호기 형식

        new_qty = items_in_so + bulk_qty
        new_total = float(cur["total_amount"] or 0) + amt * bulk_qty
        old_lbl = (cur["unit_label"] or "").strip()
        new_lbl = (old_lbl + " · " + " · ".join(labels_bulk)) if old_lbl else " · ".join(labels_bulk)

        # order_items INSERT — bulk_qty 개 모두
        for _lbl in labels_bulk:
            try:
                c.execute(
                    "INSERT INTO order_items(order_id, qty, unit_price, amount, "
                    "unit_label, line_note) VALUES(?,1,?,?,?,?)",
                    (oid, amt, amt, _lbl, note)
                )
            except Exception:
                pass
        # orders 누적 (+ currency 필요 시 갱신)
        try:
            c.execute(
                "UPDATE orders SET unit_qty=?, total_amount=?, unit_label=?, currency=? WHERE id=?",
                (new_qty, new_total, new_lbl, cur_v, oid)
            )
        except Exception:
            c.execute(
                "UPDATE orders SET unit_qty=?, total_amount=?, unit_label=? WHERE id=?",
                (new_qty, new_total, new_lbl, oid)
            )
        # 프로젝트 합계 동기화
        try:
            pid = cur["project_id"]
            if pid:
                row = c.execute(
                    "SELECT COALESCE(SUM(total_amount),0) FROM orders WHERE project_id=?",
                    (pid,)
                ).fetchone()
                c.execute("UPDATE projects SET order_amount=? WHERE id=?",
                          (float(row[0] or 0), pid))
        except Exception:
            pass
        # 이력 (v5H100: SO 합계 변동 명시 / v5H110: bulk 표시)
        try:
            old_total = float(cur["total_amount"] or 0)
            old_qty = int(cur["unit_qty"] or 1)
            range_str = (labels_bulk[0] if len(labels_bulk) == 1
                         else f"{labels_bulk[0]}~{labels_bulk[-1]}")
            note_msg = (f"호기 추가 [{range_str}] · {bulk_qty}대 × 단가 {amt:,.0f} {cur_v} · "
                        f"수량 {old_qty} → {new_qty}대 · "
                        f"SO 합계 {old_total:,.0f} → {new_total:,.0f}")
            if note:
                note_msg += f" ({note})"
            c.execute(
                "INSERT INTO order_status_history(order_id, from_status, to_status, "
                "changed_by, note) VALUES(?,?,?,?,?)",
                (oid, st, st, u.get("id"), note_msg)
            )
        except Exception:
            pass
    msg = (f"호기 '{labels_bulk[0]}' 추가 완료" if bulk_qty == 1
           else f"호기 {labels_bulk[0]}~{labels_bulk[-1]} ({bulk_qty}대) 추가 완료")
    return JSONResponse({"ok": True, "message": msg})


@app.post("/sales/orders/{oid:int}/quick-edit")
async def sales_orders_quick_edit(req: Request, oid: int):
    """v5H84: SO 인라인 빠른 수정 — 호기수(unit_qty) / 금액(total_amount) 즉시 변경.
    프로젝트 상세에서 인라인 편집용. 출하/송장/취소 SO 는 거부."""
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    if not can_use_sales(u):
        return JSONResponse({"error": "권한 없음"}, 403)
    form = await req.form()
    raw_q = (form.get("unit_qty") or "").strip()
    raw_a = (form.get("total_amount") or "").strip().replace(",", "")
    raw_c = (form.get("currency") or "").strip().upper()
    raw_due = (form.get("due_date") or "").strip()
    raw_ord = (form.get("order_date") or "").strip()
    raw_ship = (form.get("ship_to") or "").strip()
    sets, vals = [], []
    if raw_c in ("KRW", "USD", "VND"):
        sets.append("currency=?"); vals.append(raw_c)
    if raw_q:
        try:
            q = int(float(raw_q))
            if q < 1:
                return JSONResponse({"ok": False, "message": "호기 수는 1 이상이어야 합니다"}, 400)
            sets.append("unit_qty=?"); vals.append(q)
        except ValueError:
            return JSONResponse({"ok": False, "message": "호기 수 형식 오류"}, 400)
    if raw_a:
        try:
            a = float(raw_a)
            if a < 0:
                return JSONResponse({"ok": False, "message": "금액은 0 이상이어야 합니다"}, 400)
            sets.append("total_amount=?"); vals.append(a)
        except ValueError:
            return JSONResponse({"ok": False, "message": "금액 형식 오류"}, 400)
    # v5H106: SO 헤더 필드 (납기/발주일/납품처) 도 인라인 편집
    if "due_date" in form:
        sets.append("due_date=?"); vals.append(raw_due or None)
    if "order_date" in form:
        sets.append("order_date=?"); vals.append(raw_ord or None)
    if "ship_to" in form:
        sets.append("ship_to=?"); vals.append(raw_ship or None)
    if not sets:
        return JSONResponse({"ok": False, "message": "수정 항목이 없습니다"}, 400)
    with db_session() as c:
        cur = c.execute("SELECT status, project_id FROM orders WHERE id=?", (oid,)).fetchone()
        if not cur:
            return JSONResponse({"ok": False, "message": "수주를 찾을 수 없습니다"}, 404)
        st = (cur["status"] or "").upper()
        if st in ("SHIPPED", "INVOICED", "PAID", "CANCELLED"):
            return JSONResponse({
                "ok": False,
                "message": f"이미 {st} 상태인 SO 는 인라인 수정 불가 (필요 시 삭제 후 재발급)"
            }, 400)
        # v5H91: 호기 라인 vs 신규 호기수/금액 정합성 검증
        try:
            items = c.execute(
                "SELECT COUNT(*) AS n, COALESCE(SUM(amount),0) AS s "
                "FROM order_items WHERE order_id=?", (oid,)
            ).fetchone()
            items_n = int(items["n"] or 0)
            items_s = float(items["s"] or 0)
        except Exception:
            items_n, items_s = 0, 0.0
        if items_n > 0:
            if raw_q:
                new_q = int(float(raw_q))
                if new_q != items_n:
                    return JSONResponse({
                        "ok": False,
                        "message": (f"호기 라인은 {items_n}건인데 호기수를 {new_q}대로 바꿀 수 없습니다.\n"
                                    f"먼저 '➕ 호기' 로 부족한 호기({new_q - items_n}대)를 추가하거나,\n"
                                    f"불필요한 호기 라인을 정리한 뒤 다시 시도하세요.")
                    }, 400)
            if raw_a:
                new_a = float(raw_a)
                if abs(new_a - items_s) > 0.5:
                    return JSONResponse({
                        "ok": False,
                        "message": (f"호기 단가 합계는 {items_s:,.0f}원인데 합계를 {new_a:,.0f}원으로 바꿀 수 없습니다.\n"
                                    f"차액 {new_a - items_s:,.0f}원이 발생합니다.\n"
                                    f"호기별 단가를 직접 수정하거나, '➕ 호기' 로 차액만큼의 호기를 추가하세요.")
                    }, 400)
        vals.append(oid)
        c.execute(f"UPDATE orders SET {', '.join(sets)} WHERE id=?", vals)
        # 프로젝트 합계 동기화 (project.order_amount = SUM 모든 SO)
        try:
            pid = cur["project_id"]
            if pid:
                row = c.execute(
                    "SELECT COALESCE(SUM(total_amount),0) FROM orders WHERE project_id=?",
                    (pid,)
                ).fetchone()
                c.execute("UPDATE projects SET order_amount=? WHERE id=?",
                          (float(row[0] or 0), pid))
        except Exception:
            pass
        # 변경 이력
        try:
            change_desc = []
            if "unit_qty=?" in ",".join(sets):
                change_desc.append(f"호기수→{int(form.get('unit_qty') or 0)}")
            if "total_amount=?" in ",".join(sets):
                change_desc.append(f"금액→{float(raw_a or 0):,.0f}")
            c.execute(
                "INSERT INTO order_status_history(order_id, from_status, to_status, "
                "changed_by, note) VALUES(?,?,?,?,?)",
                (oid, st, st, u.get("id"),
                 "인라인 수정: " + " / ".join(change_desc))
            )
        except Exception:
            pass
    return JSONResponse({"ok": True, "message": "저장 완료"})


@app.post("/sales/orders/{oid:int}/delete")
async def sales_orders_delete(req: Request, oid: int):
    """v5H70/v5H72: 수주 삭제 — 기술영업팀 등록권한자(can_use_sales)만 가능."""
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    if not can_delete_sales(u):
        return JSONResponse({
            "error": "권한 없음",
            "message": "수주 삭제는 기술영업팀 팀장 또는 위임받은 등록권한자(can_use_sales=1)만 가능합니다.\n영업팀 팀장에게 권한 신청 후 다시 시도하세요."
        }, 403)
    with db_session() as c:
        res = _pwf.delete_order(c, oid, restore_project=True)
    return JSONResponse(res)


@app.post("/projects/{pid:int}/add-followup-order")
async def projects_add_followup(req: Request, pid: int):
    """추가 발주 — 동일 관리번호 + 신규 수주번호 발행 (KNK 추적 표준)."""
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    form = await req.form()
    order_date = (form.get("order_date") or "").strip()
    raw_amt = (form.get("total_amount") or "0").replace(",", "")
    try:
        total = float(raw_amt) if raw_amt else 0
    except ValueError:
        total = 0
    due_date = (form.get("due_date") or "").strip()
    po_number = (form.get("po_number") or "").strip()
    note = (form.get("note") or "").strip()
    # v5H131: qty (1~100) — N대 일괄 등록. 미전달 시 1.
    raw_qty = (form.get("qty") or "1").strip()
    try:
        qty = int(float(raw_qty))
    except (TypeError, ValueError):
        qty = 1
    if qty < 1:
        qty = 1
    if qty > 100:
        return JSONResponse({"ok": False, "message": "한 번에 100대 초과 불가"}, 400)
    # v5H142: so_type — EQUIPMENT(기본) / CONSUMABLE / SERVICE / OTHER
    so_type = (form.get("so_type") or "EQUIPMENT").strip().upper()
    if so_type not in ("EQUIPMENT", "CONSUMABLE", "SERVICE", "OTHER"):
        so_type = "EQUIPMENT"
    # v5H178: 통화 + 납품처 (미전달 시 프로젝트 헤더 통화 사용)
    currency = (form.get("currency") or "").strip().upper()
    if currency not in ("KRW", "USD", "VND", "JPY", "CNY", "EUR"):
        currency = ""  # 빈 값 → add_followup_order 가 프로젝트 헤더 통화 사용
    ship_to = (form.get("ship_to") or "").strip()
    with db_session() as c:
        res = _pwf.add_followup_order(c, pid, order_date=order_date,
                                        total_amount=total, due_date=due_date,
                                        created_by=u.get("id"),
                                        po_number=po_number, note=note, qty=qty,
                                        so_type=so_type, currency=currency,
                                        ship_to=ship_to)
    return JSONResponse(res)


# v5H67: 공통 엑셀 내보내기 헬퍼 — 모든 모듈에서 재사용
def _make_xlsx_response(sheets: list, filename_prefix: str):
    """다중 시트 엑셀 다운로드 응답 생성.
    sheets: [{name, headers, rows}, ...]
      - name: 시트명 (한글 OK)
      - headers: 컬럼 헤더 리스트
      - rows: 데이터 행 리스트 (각 행은 list 또는 tuple)
    filename_prefix: 파일명 접두 (예: '공급사')
    파일명 자동: '{prefix}_{YYYY-MM-DD}.xlsx'"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return JSONResponse({"error": "openpyxl 미설치"}, 500)
    wb = Workbook()
    knk_fill = PatternFill("solid", fgColor="A5282C")
    white = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    first = True
    for sh in sheets:
        if first:
            ws = wb.active
            ws.title = sh["name"][:31]  # Excel sheet name 31자 제한
            first = False
        else:
            ws = wb.create_sheet(sh["name"][:31])
        ws.append(sh["headers"])
        for col_idx, _ in enumerate(sh["headers"], 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = white
            cell.fill = knk_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for row in sh["rows"]:
            ws.append(list(row))
        # 컬럼 너비 자동
        for col in ws.columns:
            max_len = 8
            col_letter = col[0].column_letter
            for cell in col:
                v = str(cell.value) if cell.value is not None else ""
                disp_len = sum(2 if ord(c) > 0x7F else 1 for c in v)
                max_len = max(max_len, min(50, disp_len + 2))
            ws.column_dimensions[col_letter].width = max_len
        ws.freeze_panes = "A2"
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote
    fname = f"{filename_prefix}_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )


# v5H66: 고객사 전체 엑셀 내보내기 (회사정보 + 담당자 2시트)
@app.get("/customers/export.xlsx")
async def customers_export_xlsx(req: Request):
    """전체 고객사 + 담당자 데이터를 .xlsx 로 다운로드.
    시트1: 고객사 (회사정보 + 자동 등급 + 점수)
    시트2: 담당자 (고객사명 + 부서/이름/전화/이메일/특징)"""
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return JSONResponse({"error": "openpyxl 미설치 — pip install openpyxl"}, 500)
    wb = Workbook()

    # === 시트 1: 고객사 ===
    ws1 = wb.active
    ws1.title = "고객사"
    headers1 = ["ID", "고객사명", "등급", "점수",
                "사업자번호", "대표자", "대표 전화", "대표 이메일",
                "주소", "활성", "비고", "프로젝트수",
                "수주합계(원)", "최근거래일", "등급계산일"]
    ws1.append(headers1)
    # 헤더 스타일
    knk_fill = PatternFill("solid", fgColor="A5282C")
    white = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, _ in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.font = white
        cell.fill = knk_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    with db_session() as c:
        rows = c.execute(
            """SELECT cu.id, cu.name, cu.tier, cu.tier_score,
                      cu.biz_no, cu.ceo_name, cu.phone, cu.email, cu.address,
                      cu.is_active, cu.note, cu.tier_computed_at,
                      COUNT(DISTINCT p.id) AS proj_count,
                      COALESCE(SUM(p.order_amount), 0) AS total_amount,
                      MAX(p.order_date) AS last_order
               FROM customers cu
               LEFT JOIN projects p ON p.customer_id = cu.id
               GROUP BY cu.id
               ORDER BY cu.tier_score DESC, cu.name"""
        ).fetchall()
        for r in rows:
            d = dict(r)
            ws1.append([
                d.get("id"), d.get("name"), d.get("tier"), d.get("tier_score") or 0,
                d.get("biz_no"), d.get("ceo_name"), d.get("phone"), d.get("email"),
                d.get("address"),
                "활성" if (d.get("is_active") != 0) else "비활성",
                d.get("note"), d.get("proj_count") or 0,
                d.get("total_amount") or 0, d.get("last_order"),
                d.get("tier_computed_at"),
            ])
        # === 시트 2: 담당자 ===
        ws2 = wb.create_sheet("담당자")
        headers2 = ["고객사명", "부서", "이름", "직위",
                    "전화", "휴대폰", "이메일", "주담당", "특징/메모"]
        ws2.append(headers2)
        for col_idx, _ in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = white
            cell.fill = knk_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        contacts = c.execute(
            """SELECT cu.name AS cust_name, cc.department, cc.name, cc.position,
                      cc.phone, cc.mobile, cc.email, cc.is_primary, cc.note
               FROM customer_contacts cc
               JOIN customers cu ON cu.id = cc.customer_id
               ORDER BY cu.tier_score DESC, cu.name, cc.is_primary DESC, cc.id"""
        ).fetchall()
        for r in contacts:
            d = dict(r)
            ws2.append([
                d.get("cust_name"), d.get("department"), d.get("name"),
                d.get("position"), d.get("phone"), d.get("mobile"),
                d.get("email"),
                "★" if d.get("is_primary") else "",
                d.get("note"),
            ])

    # 컬럼 너비 자동 (대략)
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = 8
            col_letter = col[0].column_letter
            for cell in col:
                v = str(cell.value) if cell.value is not None else ""
                # 한글은 2칸으로 추정
                disp_len = sum(2 if ord(c) > 0x7F else 1 for c in v)
                max_len = max(max_len, min(50, disp_len + 2))
            ws.column_dimensions[col_letter].width = max_len
        ws.freeze_panes = "A2"  # 헤더 고정

    # 메모리 → BytesIO → StreamingResponse
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote
    fname = f"고객사_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}",
        },
    )


# v5H67: 12개 모듈 엑셀 내보내기 일괄 추가 (대표 지시 — 모든 데이터 엑셀화)
def _xls_str(v):
    """엑셀 저장용 안전 변환 (None → '', dict/list → str)."""
    if v is None: return ""
    if isinstance(v, (int, float, str)): return v
    return str(v)


@app.get("/suppliers/export.xlsx")
async def suppliers_export_xlsx(req: Request):
    u = get_user(req)
    if not u: return RedirectResponse("/login", 303)
    with db_session() as c:
        rows = [dict(r) for r in c.execute(
            "SELECT id, name, code, contact, email, phone, country, currency, "
            "payment_terms, COALESCE(is_active,1) AS is_active, note "
            "FROM suppliers ORDER BY name").fetchall()]
    headers = ["ID","공급사명","코드","담당자","이메일","전화","국가","통화",
               "결제조건","활성","비고"]
    data = [[r["id"], r["name"], r["code"], r["contact"], r["email"], r["phone"],
             r["country"], r["currency"], r["payment_terms"],
             "활성" if r["is_active"] else "비활성", r["note"]] for r in rows]
    return _make_xlsx_response(
        [{"name": "공급사", "headers": headers, "rows": data}], "공급사")


@app.get("/parts/export.xlsx")
async def parts_export_xlsx(req: Request):
    u = get_user(req)
    if not u: return RedirectResponse("/login", 303)
    with db_session() as c:
        rows = [dict(r) for r in c.execute(
            "SELECT p.id, p.part_no, p.part_name, p.spec, p.maker, p.origin, "
            "p.unit, p.currency, p.std_price, p.biz_div, p.category, "
            "COALESCE(p.is_active,1) AS is_active, p.note, "
            "COALESCE(sb.on_hand, 0) AS on_hand "
            "FROM parts p LEFT JOIN stock_balances sb ON sb.part_id=p.id "
            "ORDER BY p.part_no").fetchall()]
    headers = ["ID","품번","품명","규격","제조사","원산지","단위","통화",
               "표준단가","사업부","분류","활성","비고","현재재고"]
    data = [[r["id"], r["part_no"], r["part_name"], r["spec"], r["maker"],
             r["origin"], r["unit"], r["currency"], r["std_price"],
             r["biz_div"], r["category"],
             "활성" if r["is_active"] else "비활성",
             r["note"], r["on_hand"]] for r in rows]
    return _make_xlsx_response(
        [{"name": "부품", "headers": headers, "rows": data}], "부품")


@app.get("/sales/orders/export.xlsx")
async def sales_orders_export_xlsx(req: Request):
    u = get_user(req)
    if not u: return RedirectResponse("/login", 303)
    with db_session() as c:
        # 헤더
        head_rows = [dict(r) for r in c.execute(
            "SELECT o.id, o.order_no, o.order_date, o.due_date, o.status, "
            "o.total_amount, COALESCE(cu.name,'') AS customer_name "
            "FROM orders o LEFT JOIN customers cu ON cu.id=o.customer_id "
            "ORDER BY o.id DESC").fetchall()]
        # 라인
        line_rows = [dict(r) for r in c.execute(
            "SELECT o.order_no, p.part_no, p.part_name, oi.qty, oi.unit_price, oi.amount "
            "FROM order_items oi "
            "LEFT JOIN orders o ON o.id=oi.order_id "
            "LEFT JOIN parts p ON p.id=oi.part_id "
            "ORDER BY o.id DESC, oi.id").fetchall()]
    sheets = [
        {"name": "수주", "headers": ["ID","수주번호","주문일","납기","상태","수주액(원)","고객사"],
         "rows": [[r["id"], r["order_no"], r["order_date"], r["due_date"],
                   r["status"], r["total_amount"], r["customer_name"]] for r in head_rows]},
        {"name": "수주라인", "headers": ["수주번호","품번","품명","수량","단가","금액"],
         "rows": [[r["order_no"], r["part_no"], r["part_name"], r["qty"],
                   r["unit_price"], r["amount"]] for r in line_rows]},
    ]
    return _make_xlsx_response(sheets, "수주")


@app.get("/sales/quotations/export.xlsx")
async def sales_quotations_export_xlsx(req: Request):
    u = get_user(req)
    if not u: return RedirectResponse("/login", 303)
    with db_session() as c:
        head_rows = [dict(r) for r in c.execute(
            "SELECT q.id, q.quote_no, q.created_at, q.valid_until, q.version, "
            "q.status, q.total_amount, COALESCE(cu.name,'') AS customer_name "
            "FROM quotations q LEFT JOIN customers cu ON cu.id=q.customer_id "
            "ORDER BY q.id DESC").fetchall()]
        line_rows = [dict(r) for r in c.execute(
            "SELECT q.quote_no, qi.line_no, qi.item_name, qi.qty, qi.unit, "
            "qi.unit_price, qi.total_price, qi.note "
            "FROM quotation_items qi "
            "LEFT JOIN quotations q ON q.id=qi.quotation_id "
            "ORDER BY q.id DESC, qi.line_no").fetchall()]
    sheets = [
        {"name": "견적", "headers": ["ID","견적번호","작성일","유효기한","버전","상태","금액(원)","고객사"],
         "rows": [[r["id"], r["quote_no"], r["created_at"], r["valid_until"],
                   r["version"], r["status"], r["total_amount"], r["customer_name"]]
                  for r in head_rows]},
        {"name": "견적라인", "headers": ["견적번호","#","품목명","수량","단위","단가","금액","비고"],
         "rows": [[r["quote_no"], r["line_no"], r["item_name"], r["qty"],
                   r["unit"], r["unit_price"], r["total_price"], r["note"]]
                  for r in line_rows]},
    ]
    return _make_xlsx_response(sheets, "견적")


@app.get("/po/export.xlsx")
async def po_export_xlsx(req: Request):
    u = get_user(req)
    if not u: return RedirectResponse("/login", 303)
    with db_session() as c:
        head_rows = [dict(r) for r in c.execute(
            "SELECT po.id, po.po_number, po.order_date, po.expected_date, "
            "po.status, po.currency, po.total_amount, "
            "COALESCE(s.name,'') AS supplier_name, "
            "COALESCE(p.name,'') AS project_name "
            "FROM purchase_orders po "
            "LEFT JOIN suppliers s ON s.id=po.supplier_id "
            "LEFT JOIN projects p ON p.id=po.project_id "
            "ORDER BY po.id DESC").fetchall()]
        line_rows = [dict(r) for r in c.execute(
            "SELECT po.po_number, pi.line_no, pi.part_no_snapshot, pi.part_name_snapshot, "
            "pi.unit, pi.quantity, pi.unit_price, pi.amount, pi.received_qty "
            "FROM po_items pi LEFT JOIN purchase_orders po ON po.id=pi.po_id "
            "ORDER BY po.id DESC, pi.line_no").fetchall()]
    sheets = [
        {"name": "발주", "headers": ["ID","발주번호","발주일","예상입고","상태","통화","합계","공급사","프로젝트"],
         "rows": [[r["id"], r["po_number"], r["order_date"], r["expected_date"],
                   r["status"], r["currency"], r["total_amount"],
                   r["supplier_name"], r["project_name"]] for r in head_rows]},
        {"name": "발주라인", "headers": ["발주번호","#","품번","품명","단위","수량","단가","금액","입고수량"],
         "rows": [[r["po_number"], r["line_no"], r["part_no_snapshot"],
                   r["part_name_snapshot"], r["unit"], r["quantity"],
                   r["unit_price"], r["amount"], r["received_qty"]] for r in line_rows]},
    ]
    return _make_xlsx_response(sheets, "발주")


@app.get("/stock/movements/export.xlsx")
async def stock_movements_export_xlsx(req: Request):
    u = get_user(req)
    if not u: return RedirectResponse("/login", 303)
    with db_session() as c:
        rows = [dict(r) for r in c.execute(
            "SELECT sm.id, sm.movement_no, sm.occurred_at, p.part_no, p.part_name, "
            "sm.kind, sm.quantity, sm.unit, sm.unit_price, sm.amount, "
            "sm.lot_no, po.po_number, prj.name AS project_name, sm.reason "
            "FROM stock_movements sm "
            "LEFT JOIN parts p ON p.id=sm.part_id "
            "LEFT JOIN purchase_orders po ON po.id=sm.po_id "
            "LEFT JOIN projects prj ON prj.id=sm.project_id "
            "ORDER BY sm.occurred_at DESC LIMIT 5000").fetchall()]
    headers = ["ID","수불번호","일시","품번","품명","유형","수량","단위",
               "단가","금액","로트","발주번호","프로젝트","사유"]
    data = [[r["id"], r["movement_no"], r["occurred_at"], r["part_no"], r["part_name"],
             r["kind"], r["quantity"], r["unit"], r["unit_price"], r["amount"],
             r["lot_no"], r["po_number"], r["project_name"], r["reason"]] for r in rows]
    return _make_xlsx_response(
        [{"name": "수불부", "headers": headers, "rows": data}], "수불부")


@app.get("/stock/balances/export.xlsx")
async def stock_balances_export_xlsx(req: Request):
    u = get_user(req)
    if not u: return RedirectResponse("/login", 303)
    with db_session() as c:
        rows = [dict(r) for r in c.execute(
            "SELECT sb.part_id, sb.part_no, sb.part_name, sb.on_hand, sb.unit, "
            "sb.last_movement_at, p.std_price, p.spec "
            "FROM stock_balances sb LEFT JOIN parts p ON p.id=sb.part_id "
            "ORDER BY sb.part_no").fetchall()]
    headers = ["품번","품명","규격","재고수량","단위","표준단가","재고가치(원)","최종거래일"]
    data = [[r["part_no"], r["part_name"], r["spec"], r["on_hand"], r["unit"],
             r["std_price"], (r["on_hand"] or 0) * (r["std_price"] or 0),
             r["last_movement_at"]] for r in rows]
    return _make_xlsx_response(
        [{"name": "재고잔고", "headers": headers, "rows": data}], "재고잔고")


@app.get("/issues/export.xlsx")
async def issues_export_xlsx(req: Request):
    u = get_user(req)
    if not u: return RedirectResponse("/login", 303)
    with db_session() as c:
        rows = [dict(r) for r in c.execute(
            "SELECT i.id, i.issue_no, i.title, i.severity, i.issue_type, i.status, "
            "i.customer_name, i.biz_div, i.occurred_at, i.description, "
            "i.root_cause, i.action_taken, i.prevention, i.cost_estimate, "
            "i.resolved_at, t.name AS owner_team_name, u.name AS owner_user_name "
            "FROM issues i LEFT JOIN teams t ON t.id=i.owner_team_id "
            "LEFT JOIN users u ON u.id=i.owner_user_id "
            "ORDER BY i.id DESC").fetchall()]
    headers = ["ID","이슈번호","제목","심각도","유형","상태","고객사","사업부",
               "발생일","증상","원인","조치","재발방지","비용(원)","해결일",
               "책임팀","담당자"]
    data = [[r["id"], r["issue_no"], r["title"], r["severity"], r["issue_type"],
             r["status"], r["customer_name"], r["biz_div"], r["occurred_at"],
             r["description"], r["root_cause"], r["action_taken"],
             r["prevention"], r["cost_estimate"], r["resolved_at"],
             r["owner_team_name"], r["owner_user_name"]] for r in rows]
    return _make_xlsx_response(
        [{"name": "이슈", "headers": headers, "rows": data}], "이슈")


@app.get("/tickets/export.xlsx")
async def tickets_export_xlsx(req: Request):
    u = get_user(req)
    if not u: return RedirectResponse("/login", 303)
    with db_session() as c:
        rows = [dict(r) for r in c.execute(
            "SELECT t.id, t.ticket_no, t.category, t.title, t.description, "
            "t.urgency, t.status, t.due_date, t.completed_at, "
            "t.hours_estimated, t.hours_actual, "
            "ru.name AS requester_name, tm.name AS recipient_team_name, "
            "ru2.name AS recipient_user_name "
            "FROM tickets t "
            "LEFT JOIN users ru ON ru.id=t.requester_id "
            "LEFT JOIN teams tm ON tm.id=t.recipient_team_id "
            "LEFT JOIN users ru2 ON ru2.id=t.recipient_user_id "
            "ORDER BY t.id DESC").fetchall()]
    headers = ["ID","티켓번호","분류","제목","설명","긴급도","상태","기한","완료일",
               "예상공수","실공수","요청자","수신팀","수신자"]
    data = [[r["id"], r["ticket_no"], r["category"], r["title"], r["description"],
             r["urgency"], r["status"], r["due_date"], r["completed_at"],
             r["hours_estimated"], r["hours_actual"],
             r["requester_name"], r["recipient_team_name"],
             r["recipient_user_name"]] for r in rows]
    return _make_xlsx_response(
        [{"name": "티켓", "headers": headers, "rows": data}], "티켓")


@app.get("/changes/export.xlsx")
async def changes_export_xlsx(req: Request):
    u = get_user(req)
    if not u: return RedirectResponse("/login", 303)
    with db_session() as c:
        rows = [dict(r) for r in c.execute(
            "SELECT ch.id, ch.change_no, ch.change_type, ch.biz_div, "
            "ch.target_label, ch.title, ch.description, "
            "ch.before_value, ch.after_value, ch.urgency, ch.status, "
            "ch.notified_at, ch.completed_at, u.name AS author_name "
            "FROM changes ch LEFT JOIN users u ON u.id=ch.author_id "
            "ORDER BY ch.id DESC").fetchall()]
    headers = ["ID","변경번호","유형","사업부","대상","제목","설명","이전","이후",
               "긴급도","상태","공지일","완료일","작성자"]
    data = [[r["id"], r["change_no"], r["change_type"], r["biz_div"],
             r["target_label"], r["title"], r["description"],
             r["before_value"], r["after_value"], r["urgency"], r["status"],
             r["notified_at"], r["completed_at"], r["author_name"]] for r in rows]
    return _make_xlsx_response(
        [{"name": "변경공지", "headers": headers, "rows": data}], "변경공지")


@app.get("/production/work-orders/export.xlsx")
async def wo_export_xlsx(req: Request):
    u = get_user(req)
    if not u: return RedirectResponse("/login", 303)
    with db_session() as c:
        rows = [dict(r) for r in c.execute(
            "SELECT wo.id, wo.wo_no, wo.qty, wo.status, "
            "wo.planned_start, wo.planned_end, wo.actual_end, "
            "p.part_no, p.part_name, "
            "wo.assigned_name, wo.created_by_name, wo.specifications, wo.remarks "
            "FROM work_orders wo LEFT JOIN parts p ON p.id=wo.part_id "
            "ORDER BY wo.id DESC").fetchall()]
    headers = ["ID","WO번호","수량","상태","계획시작","계획종료","실제완료",
               "품번","품명","작업자","작성자","사양","비고"]
    data = [[r["id"], r["wo_no"], r["qty"], r["status"], r["planned_start"],
             r["planned_end"], r["actual_end"], r["part_no"], r["part_name"],
             r["assigned_name"], r["created_by_name"],
             r["specifications"], r["remarks"]] for r in rows]
    return _make_xlsx_response(
        [{"name": "작업지시", "headers": headers, "rows": data}], "작업지시")


@app.get("/qc/inspection-reports/export.xlsx")
async def qc_export_xlsx(req: Request):
    u = get_user(req)
    if not u: return RedirectResponse("/login", 303)
    with db_session() as c:
        rows = [dict(r) for r in c.execute(
            "SELECT qr.id, qr.report_no, qr.customer_name, qr.machine_model, "
            "qr.machine_serial, qr.inspection_date, qr.inspector_name, "
            "qr.qa_manager_name, qr.overall, qr.status, qr.issued_at, qr.remarks "
            "FROM qc_inspection_reports qr ORDER BY qr.id DESC").fetchall()]
    headers = ["ID","보고서번호","고객사","장비모델","시리얼","검사일",
               "검사자","QA책임자","결과","상태","발급일","비고"]
    data = [[r["id"], r["report_no"], r["customer_name"], r["machine_model"],
             r["machine_serial"], r["inspection_date"], r["inspector_name"],
             r["qa_manager_name"], r["overall"], r["status"], r["issued_at"],
             r["remarks"]] for r in rows]
    return _make_xlsx_response(
        [{"name": "QC검사보고서", "headers": headers, "rows": data}], "QC검사보고서")


@app.get("/projects/export.xlsx")
async def projects_export_xlsx(req: Request):
    """v5H72: 프로젝트 엑셀 — 수주번호/단계/사업부라벨/누적공수/PM 등 풍부한 정보."""
    u = get_user(req)
    if not u: return RedirectResponse("/login", 303)
    with db_session() as c:
        # 프로젝트 본체 + 누적 업무 통계 + 수주 합계
        rows = [dict(r) for r in c.execute(
            """SELECT p.id, p.mgmt_code, p.code, p.name, p.biz_div, p.type,
                      p.stage, p.status, p.po_type,
                      p.order_amount, p.order_date, p.due_date,
                      p.start_date, p.end_date, p.customer_po,
                      COALESCE(p.logi_note,'') AS note, p.created_at,
                      COALESCE(cu.name, p.customer_name, '') AS customer_name,
                      COALESCE(u.name, p.pm_name, '') AS pm_name,
                      COALESCE(us.name, p.sales_name, '') AS sales_name,
                      (SELECT COUNT(*) FROM tasks tk WHERE tk.project_id=p.id) AS task_count,
                      (SELECT COALESCE(SUM(tk.hours),0) FROM tasks tk WHERE tk.project_id=p.id) AS total_hours,
                      (SELECT COUNT(*) FROM orders o WHERE o.project_id=p.id) AS so_count,
                      (SELECT COALESCE(SUM(o.total_amount),0) FROM orders o WHERE o.project_id=p.id) AS so_total
               FROM projects p
               LEFT JOIN customers cu ON cu.id=p.customer_id
               LEFT JOIN users u ON u.id=p.pm_id
               LEFT JOIN users us ON us.id=p.lead_user_id
               ORDER BY p.id DESC""").fetchall()]
        # 각 프로젝트의 수주번호 목록
        so_by_proj = {}
        for r in c.execute(
            "SELECT project_id, order_no FROM orders WHERE project_id IS NOT NULL "
            "ORDER BY project_id, order_date DESC, id DESC"
        ).fetchall():
            so_by_proj.setdefault(r["project_id"], []).append(r["order_no"])

    biz_label = lambda b: "검사기" if b == "T" else ("자동화" if b == "M" else (b or ""))

    headers = [
        "ID", "관리코드", "수주번호(들)", "프로젝트명", "사업부",
        "고객사", "단계", "상태", "PO유형",
        "수주액(원)", "수주합계(원)", "수주건수",
        "수주일", "납기", "시작일", "종료일",
        "PM", "영업담당", "고객 PO", "업무수", "누적공수(h)",
        "비고", "등록일",
    ]
    data = []
    for r in rows:
        sos = so_by_proj.get(r["id"], [])
        data.append([
            r["id"], r["mgmt_code"] or r["code"] or "",
            ", ".join(sos) if sos else "",
            r["name"], biz_label(r["biz_div"]),
            r["customer_name"],
            r["stage"] or "", r["status"] or "", r["po_type"] or "",
            r["order_amount"] or 0, r["so_total"] or 0, r["so_count"] or 0,
            r["order_date"] or "", r["due_date"] or "",
            r["start_date"] or "", r["end_date"] or "",
            r["pm_name"] or "", r["sales_name"] or "",
            r["customer_po"] or "",
            r["task_count"] or 0, round(r["total_hours"] or 0, 1),
            r["note"] or "", r["created_at"] or "",
        ])
    return _make_xlsx_response(
        [{"name": "프로젝트", "headers": headers, "rows": data}], "프로젝트")


@app.get("/customers/ocr-status")
async def customers_ocr_status(req: Request):
    """OCR 설치 상태 확인 — Tesseract + 언어 데이터."""
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    return JSONResponse({
        "tesseract": _biz_doc.has_tesseract(),
        "korean": _biz_doc.has_korean(),
        "langs": _biz_doc.get_installed_langs(),
    })


@app.post("/customers/install-korean-ocr")
async def customers_install_korean_ocr(req: Request):
    """한국어 OCR 데이터 자동 다운로드 (사용자 명시적 트리거)."""
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    res = _biz_doc.download_korean_traineddata()
    return JSONResponse(res)


@app.post("/customers/parse-biz")
async def customers_parse_biz(req: Request, file: UploadFile = File(...)):
    """사업자등록증 파일(PDF/JPG/PNG) 업로드 → 자동 파싱 → JSON 반환.
    UI 에서 비동기 호출 → 응답 받은 필드를 폼에 자동 채움."""
    u = get_user(req)
    if not u:
        return JSONResponse({"ok": False, "message": "로그인 필요"}, 401)
    fn = file.filename or "biz_doc"
    ext = os.path.splitext(fn)[1].lower()
    if ext not in (".pdf", ".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff"):
        return JSONResponse({
            "ok": False,
            "message": f"지원하지 않는 파일 형식입니다: {ext}\nPDF / JPG / PNG 만 가능합니다.",
        })
    # 임시 파일 저장 (파싱용)
    with _tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name
    try:
        result = _biz_doc.parse_file(tmp_path, fn)
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass
    return JSONResponse(result)


@app.post("/customers/{cid:int}/recompute-tier")
async def customers_recompute_tier(req: Request, cid: int):
    """v5H58: 단일 거래처 등급 수동 재계산 (사용자 트리거)."""
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    from . import customer_tier as _ct
    with db_session() as c:
        res = _ct.refresh_customer_tier(c, cid)
    return JSONResponse({"ok": True, **res})


@app.post("/admin/recompute-all-tiers")
async def admin_recompute_all_tiers(req: Request):
    """v5H58: 전체 거래처 등급 강제 재계산 (관리자)."""
    u = require(req, ["admin", "ceo"])
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    from . import customer_tier as _ct
    with db_session() as c:
        n = _ct.refresh_all_customer_tiers(c)
    return JSONResponse({"ok": True, "refreshed": n})


def _customer_contacts_from_form(form) -> list[dict]:
    """v5H56/v5H66: 폼에서 contact_*_N 패턴으로 다중 담당자 수집.
    v5H66: 'note' (특징) 필드 추가, 'role' 은 hidden 으로 '기타' 기본."""
    indices = sorted({
        int(k.split("_")[-1]) for k in form.keys()
        if k.startswith("contact_name_") and k.split("_")[-1].isdigit()
    })
    out = []
    for idx in indices:
        nm = (form.get(f"contact_name_{idx}") or "").strip()
        if not nm:
            continue
        out.append({
            "role":       (form.get(f"contact_role_{idx}") or "기타").strip(),
            "department": (form.get(f"contact_dept_{idx}") or "").strip(),
            "name":       nm,
            "position":   (form.get(f"contact_position_{idx}") or "").strip(),
            "phone":      (form.get(f"contact_phone_{idx}") or "").strip(),
            "mobile":     (form.get(f"contact_mobile_{idx}") or "").strip(),
            "email":      (form.get(f"contact_email_{idx}") or "").strip(),
            "is_primary": 1 if form.get(f"contact_primary_{idx}") == "1" else 0,
            "note":       (form.get(f"contact_note_{idx}") or "").strip(),
        })
    return out


@app.get("/customers/new", response_class=HTMLResponse)
async def customers_new_form(req: Request):
    """v5H52/v5H56: 고객사 신규 등록 폼 (다중 담당자 지원)."""
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    return ctx(req, "customer_form.html", user=u, active="sales",
               customer=None, contacts=[])


@app.post("/customers/new")
async def customers_new_submit(req: Request):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    form = await req.form()
    name = (form.get("name") or "").strip()
    if not name:
        return RedirectResponse("/customers/new?error=name_required", status_code=303)
    # v5H58: 등급은 자동 산정 (사용자 입력 무시) — 신규는 일단 '신규' 로 시작
    fields = {
        "name": name,
        "tier": "신규",
        "biz_no": (form.get("biz_no") or "").strip(),
        "ceo_name": (form.get("ceo_name") or "").strip(),
        "phone": (form.get("phone") or "").strip(),
        "email": (form.get("email") or "").strip(),
        "address": (form.get("address") or "").strip(),
        "is_active": int(form.get("is_active") or 1),
        "note": (form.get("note") or "").strip(),
    }
    contacts = _customer_contacts_from_form(form)
    with db_session() as c:
        existing = c.execute("SELECT id FROM customers WHERE name=?", (name,)).fetchone()
        if existing:
            return RedirectResponse(
                f"/customers?error=duplicate&id={existing['id']}", status_code=303)
        cols = ",".join(fields.keys())
        ph = ",".join(["?"] * len(fields))
        c.execute(f"INSERT INTO customers({cols}) VALUES({ph})", tuple(fields.values()))
        new_id = c.execute("SELECT last_insert_rowid()").fetchone()[0]
        for ct in contacts:
            c.execute(
                "INSERT INTO customer_contacts(customer_id, role, department, "
                "name, position, phone, mobile, email, is_primary, note) "
                "VALUES(?,?,?,?,?,?,?,?,?,?)",
                (new_id, ct["role"], ct["department"], ct["name"], ct["position"],
                 ct["phone"], ct["mobile"], ct["email"], ct["is_primary"],
                 ct.get("note", ""))
            )
        # v5H58: 즉시 등급 자동 산정
        from . import customer_tier as _ct
        _ct.refresh_customer_tier(c, new_id)
    return RedirectResponse(f"/customer/{new_id}", status_code=303)


@app.get("/customers/{cid:int}/edit", response_class=HTMLResponse)
async def customers_edit_form(req: Request, cid: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        row = c.execute("SELECT * FROM customers WHERE id=?", (cid,)).fetchone()
        if not row:
            return RedirectResponse("/customers", 303)
        contacts = [dict(r) for r in c.execute(
            "SELECT * FROM customer_contacts WHERE customer_id=? "
            "ORDER BY is_primary DESC, id", (cid,)
        ).fetchall()]
    return ctx(req, "customer_form.html", user=u, active="sales",
               customer=dict(row), contacts=contacts)


@app.post("/customers/{cid:int}/edit")
async def customers_edit_submit(req: Request, cid: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    form = await req.form()
    name = (form.get("name") or "").strip()
    if not name:
        return RedirectResponse(f"/customers/{cid}/edit?error=name_required", status_code=303)
    contacts = _customer_contacts_from_form(form)
    with db_session() as c:
        # v5H112: name 변경 감지 → projects.customer_name 일괄 동기화 + 이력
        # v5H113: 핵심 7필드 전체 변경 이력화
        old_row = c.execute(
            "SELECT name, biz_no, ceo_name, phone, email, address, is_active, note "
            "FROM customers WHERE id=?", (cid,)
        ).fetchone()
        old_data = dict(old_row) if old_row else {}
        old_name = (old_data.get("name") or "")
        new_is_active = int(form.get("is_active") or 1)
        new_data = {
            "name": name,
            "biz_no": form.get("biz_no", ""),
            "ceo_name": form.get("ceo_name", ""),
            "phone": form.get("phone", ""),
            "email": form.get("email", ""),
            "address": form.get("address", ""),
            "is_active": new_is_active,
            "note": form.get("note", ""),
        }
        # v5H58: 등급(tier) 은 사용자 입력 받지 않음 — 기존 값 유지, 자동 재계산이 갱신
        c.execute(
            "UPDATE customers SET name=?, biz_no=?, ceo_name=?, "
            "phone=?, email=?, address=?, is_active=?, note=? "
            "WHERE id=?",
            (new_data["name"], new_data["biz_no"], new_data["ceo_name"],
             new_data["phone"], new_data["email"], new_data["address"],
             new_data["is_active"], new_data["note"], cid)
        )
        # v5H112: name 변경 시 자식(projects) 동기화
        if old_name and old_name != name:
            try:
                c.execute(
                    "UPDATE projects SET customer_name=? WHERE customer_id=?",
                    (name, cid),
                )
            except Exception:
                pass
        # v5H113: 7필드 모두 diff → customer_history 기록
        _label_map = {
            "name": "고객사명", "biz_no": "사업자번호", "ceo_name": "대표자",
            "phone": "전화", "email": "이메일", "address": "주소",
            "is_active": "활성여부", "note": "비고",
        }
        for _key, _label in _label_map.items():
            ov = old_data.get(_key)
            nv = new_data.get(_key)
            ov_s = "" if ov is None else str(ov)
            nv_s = "" if nv is None else str(nv)
            if ov_s == nv_s:
                continue
            _note = ""
            if _key == "name":
                _note = "고객사명 변경 → projects.customer_name 동기화"
            try:
                c.execute(
                    "INSERT INTO customer_history(customer_id, changed_by, field, "
                    "old_value, new_value, note) VALUES(?,?,?,?,?,?)",
                    (cid, u.get("id"), _label, ov_s, nv_s, _note),
                )
            except Exception:
                pass
        # 담당자: 전체 삭제 후 재삽입 (간단·신뢰)
        c.execute("DELETE FROM customer_contacts WHERE customer_id=?", (cid,))
        for ct in contacts:
            c.execute(
                "INSERT INTO customer_contacts(customer_id, role, department, "
                "name, position, phone, mobile, email, is_primary) "
                "VALUES(?,?,?,?,?,?,?,?,?)",
                (cid, ct["role"], ct["department"], ct["name"], ct["position"],
                 ct["phone"], ct["mobile"], ct["email"], ct["is_primary"])
            )
        # v5H58: 수정 후에도 등급 즉시 재계산
        from . import customer_tier as _ct
        _ct.refresh_customer_tier(c, cid)
    return RedirectResponse(f"/customer/{cid}", status_code=303)


@app.get("/orders/new", response_class=HTMLResponse)
async def _alias_orders_new(req: Request):
    return RedirectResponse("/sales/quotations", 303)


@app.get("/sales/quotes/new", response_class=HTMLResponse)
async def sales_quotes_new_form(req: Request):
    """v5H52: 견적 신규 작성 폼 (라인 동적 추가, 실시간 합계, 콤마 포맷)."""
    u = _s1_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        customers = [dict(r) for r in c.execute(
            "SELECT id, name FROM customers WHERE COALESCE(is_active,1)=1 ORDER BY name"
        ).fetchall()]
    return ctx(req, "sales_quote_form.html", user=u, active="sales_quotations",
               customers=customers)


@app.post("/sales/quotes/new")
async def sales_quotes_new_submit(req: Request):
    """v5H52: 견적 + 라인 일괄 INSERT. 라인은 line_name_N/line_qty_N/line_price_N 패턴."""
    u = _s1_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    form = await req.form()
    customer_id = form.get("customer_id") or None
    valid_until = form.get("valid_until") or None
    version = int(form.get("version") or 1)
    note = form.get("note") or ""
    # 합계 (콤마 제거)
    total_raw = (form.get("total_amount") or "0").replace(",", "")
    try:
        total_amount = float(total_raw) if total_raw else 0
    except ValueError:
        total_amount = 0
    # 라인 수집
    line_indices = sorted({
        int(k.split("_")[-1]) for k in form.keys()
        if k.startswith("line_name_") and k.split("_")[-1].isdigit()
    })
    with db_session() as c:
        ym = datetime.now().strftime("%Y%m")
        seq_row = c.execute(
            "SELECT COUNT(*) FROM quotations WHERE quote_no LIKE ?",
            (f"QT-{ym}-%",),
        ).fetchone()
        seq = (seq_row[0] if seq_row else 0) + 1
        quote_no = f"QT-{ym}-{seq:04d}"
        cur = c.execute(
            "INSERT INTO quotations(quote_no, customer_id, total_amount, "
            "valid_until, version, status, created_by) "
            "VALUES(?,?,?,?,?,'DRAFT',?)",
            (quote_no, customer_id, total_amount, valid_until, version, u.get("id"))
        )
        qid = cur.lastrowid
        for ln, idx in enumerate(line_indices, start=1):
            name = (form.get(f"line_name_{idx}") or "").strip()
            if not name:
                continue
            qty_raw = (form.get(f"line_qty_{idx}") or "0").replace(",", "")
            price_raw = (form.get(f"line_price_{idx}") or "0").replace(",", "")
            try:
                qty = float(qty_raw)
                price = float(price_raw)
            except ValueError:
                qty = price = 0
            c.execute(
                "INSERT INTO quotation_items(quotation_id, line_no, item_name, "
                "qty, unit, unit_price, total_price) VALUES(?,?,?,?,?,?,?)",
                (qid, ln, name, qty,
                 form.get(f"line_unit_{idx}") or "EA",
                 price, qty * price)
            )
    return RedirectResponse(f"/sales/quotes/{qid}", status_code=303)


@app.get("/sales/quotes/{qid:int}", response_class=HTMLResponse)
async def sales_quote_detail(req: Request, qid: int):
    """견적 상세 — 라인 + 합계 + 인쇄/수주전환 액션."""
    u = _s1_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        row = c.execute(
            """SELECT q.*, COALESCE(cu.name,'-') AS customer_name,
                      uc.name AS created_by_name
               FROM quotations q
               LEFT JOIN customers cu ON cu.id = q.customer_id
               LEFT JOIN users uc ON uc.id = q.created_by
               WHERE q.id=?""",
            (qid,)
        ).fetchone()
        if not row:
            return RedirectResponse("/sales/quotations", 303)
        quote = dict(row)
        items = [dict(r) for r in c.execute(
            "SELECT * FROM quotation_items WHERE quotation_id=? ORDER BY line_no",
            (qid,)
        ).fetchall()]
        # v5H127: 자가치유 — quotations.total_amount vs SUM(quotation_items.total_price)
        # PO/parts 패턴과 동일. 1원 미만 차이는 무시.
        quote_mismatch = None
        try:
            line_sum = sum(float(it.get("total_price") or 0) for it in items)
            cur_total = float(quote.get("total_amount") or 0)
            if abs(line_sum - cur_total) >= 1.0:
                c.execute(
                    "UPDATE quotations SET total_amount=? WHERE id=?",
                    (round(line_sum, 2), qid),
                )
                quote_mismatch = {"old": cur_total, "new": round(line_sum, 2)}
                quote["total_amount"] = round(line_sum, 2)
        except Exception:
            quote_mismatch = None
    # v5H114: 견적 변경 이력 카드
    try:
        quotation_history = _logi.get_quotation_history(qid, limit=50)
    except Exception:
        quotation_history = []
    return ctx(req, "sales_quote_detail.html", user=u, active="sales_quotations",
               quote=quote, items=items, quote_mismatch=quote_mismatch,
               quotation_history=quotation_history)


@app.get("/sales/shipments/new", response_class=HTMLResponse)
async def _alias_sales_shipments_new(req: Request):
    return RedirectResponse("/sales/shipments-receipts", 303)


@app.get("/stock/audit/new", response_class=HTMLResponse)
async def _alias_stock_audit_new(req: Request):
    return RedirectResponse("/stock/audits", 303)


@app.get("/stock/audits/new", response_class=HTMLResponse)
async def _alias_stock_audits_new_get(req: Request):
    return RedirectResponse("/stock/audits", 303)


@app.get("/qc/reports/new", response_class=HTMLResponse)
async def _alias_qc_reports_new(req: Request):
    return RedirectResponse("/qc/inspection-reports/new", 303)


@app.get("/fta/certificates/new", response_class=HTMLResponse)
async def _alias_fta_certificates_new(req: Request):
    return RedirectResponse("/export/fta/new", 303)


@app.get("/admin/permissions/groups/new", response_class=HTMLResponse)
async def _alias_admin_perm_groups_new(req: Request):
    return RedirectResponse("/admin/permissions/groups", 303)


@app.get("/parts/import", response_class=HTMLResponse)
async def _alias_parts_import(req: Request):
    return RedirectResponse("/parts", 303)


@app.get("/projects/{pid:int}", response_class=HTMLResponse)
async def projects_detail_alias(req: Request, pid: int):
    return RedirectResponse(f"/project/{pid}", 303)


@app.get("/sales/orders/{oid:int}", response_class=HTMLResponse)
async def sales_order_detail(req: Request, oid: int):
    """수주 상세 — 라인/송장/수금/이력 (v5H47 신규)."""
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        # v5H96: 프로젝트 정보(관리코드/프로젝트명/모델/사업부) 함께 조인
        row = c.execute(
            """SELECT o.*, cu.name AS customer_name,
                      uc.name AS created_by_name,
                      pj.id AS project_id, pj.mgmt_code AS mgmt_code,
                      pj.name AS project_name, pj.biz_div AS biz_div,
                      pj.model_name AS model_name
               FROM orders o
               LEFT JOIN customers cu ON cu.id = o.customer_id
               LEFT JOIN users uc ON uc.id = o.created_by
               LEFT JOIN projects pj ON pj.id = o.project_id
               WHERE o.id=?""",
            (oid,)
        ).fetchone()
        if not row:
            return RedirectResponse("/sales/orders", 303)
        order = dict(row)
        # v5H96: 호기 라인은 part_id 없음 → unit_label/line_note 를 표기 소스로
        items = [dict(r) for r in c.execute(
            """SELECT oi.*, p.part_no, p.part_name
               FROM order_items oi
               LEFT JOIN parts p ON p.id = oi.part_id
               WHERE oi.order_id=? ORDER BY oi.id""",
            (oid,)
        ).fetchall()]
        # 표시용 정규화: part_no/part_name 비면 unit_label / line_note 로 폴백
        for it in items:
            if not it.get("part_no") and it.get("unit_label"):
                it["part_no"] = "—"
            if not it.get("part_name"):
                it["part_name"] = it.get("unit_label") or "호기"
        # v5H133: 호기 표시 순서를 내림차순(최근 호기 → 1호기)으로 반전 (대표 요청)
        try:
            import re as _re_so
            def _it_key(_it):
                _lbl = (_it.get("unit_label") or "")
                _m = _re_so.match(r"^(\d+)", _lbl)
                return (int(_m.group(1)) if _m else 9999, _lbl, _it.get("id") or 0)
            items.sort(key=_it_key, reverse=True)
        except Exception:
            pass
        invoices_ = [dict(r) for r in c.execute(
            "SELECT * FROM invoices WHERE order_id=? ORDER BY id DESC", (oid,)
        ).fetchall()]
        receipts_p = [dict(r) for r in c.execute(
            "SELECT * FROM receipts_payment WHERE order_id=? ORDER BY id DESC", (oid,)
        ).fetchall()]
        history = [dict(r) for r in c.execute(
            """SELECT h.*, u.name AS changed_by_name
               FROM order_status_history h
               LEFT JOIN users u ON u.id = h.changed_by
               WHERE h.order_id=? ORDER BY h.id DESC""",
            (oid,)
        ).fetchall()]
    # 핵심 KPI 계산
    invoiced = sum(float(i.get("total_amount") or 0) for i in invoices_)
    # v5H124: 통화 일치 수금만 합산 (수주 통화 기준). 다른 통화 수금은 ⚠ 표시
    _order_ccy = (order.get("currency") or "KRW")
    received = sum(
        float(r.get("amount") or 0) for r in receipts_p
        if (r.get("currency") or "KRW") == _order_ccy
    )
    received_other_ccy = [
        r for r in receipts_p if (r.get("currency") or "KRW") != _order_ccy
    ]
    # v5H124: shipments OVER 자가치유 감지 — 누적 출하 > 수주 수량이면 warn
    ship_over_warn = None
    try:
        with db_session() as _c2:
            _qty_row = _c2.execute(
                "SELECT COALESCE(o.unit_qty,(SELECT SUM(quantity) FROM order_items WHERE order_id=o.id),0) "
                "FROM orders o WHERE o.id=?",
                (oid,),
            ).fetchone()
            _ord_qty = float(_qty_row[0] or 0) if _qty_row else 0.0
            _ship_sum = float(_c2.execute(
                "SELECT COALESCE(SUM(shipped_qty),0) FROM shipments WHERE order_id=?",
                (oid,),
            ).fetchone()[0] or 0)
            if _ord_qty > 0 and _ship_sum > _ord_qty + 0.0001:
                ship_over_warn = {
                    "ord_qty": _ord_qty, "ship_sum": _ship_sum,
                    "over": _ship_sum - _ord_qty,
                }
    except Exception:
        pass
    # v5H102: 연결 프로젝트의 변경 이력도 함께 노출 (SO 상세에서 확인 편의)
    project_history_logs = []
    try:
        if order.get("project_id"):
            project_history_logs = _logi.get_project_history(order["project_id"], limit=30)
    except Exception:
        pass
    return ctx(req, "sales_order_detail.html", user=u, active="sales",
               order=order, items=items, invoices=invoices_,
               receipts=receipts_p, history=history,
               total_invoiced=invoiced, total_received=received,
               outstanding=max(0, invoiced - received),
               received_other_ccy=received_other_ccy,
               ship_over_warn=ship_over_warn,
               project_history=project_history_logs)


@app.get("/admin/reminders", response_class=HTMLResponse)
async def reminders_page(req: Request, sel_date: str = ""):
    u = require(req, ["leader", "executive", "ceo", "admin"])
    if not u:
        return RedirectResponse("/login", 303)
    if not sel_date:
        sel_date = date.today().isoformat()
    with db_session() as c:
        teams = [dict(r) for r in c.execute(
            """SELECT t.*, ul.name AS leader_name, ul.email AS leader_email
               FROM teams t LEFT JOIN users ul ON t.leader_id=ul.id
               ORDER BY t.display_order"""
        ).fetchall()]
        # 팀장이면 자기 팀만 표시
        if u["role"] == "leader":
            teams = [t for t in teams if t["id"] == u["team_id"]]
        for t in teams:
            members = [dict(r) for r in c.execute(
                """SELECT u.id, u.name, u.rank, u.email
                   FROM users u WHERE u.team_id=? AND u.is_active=1
                   AND u.role!='admin' ORDER BY u.id""",
                (t["id"],),
            ).fetchall()]
            reported_ids = {r["user_id"] for r in c.execute(
                """SELECT DISTINCT user_id FROM tasks
                   WHERE work_date=? AND user_id IN
                   (SELECT id FROM users WHERE team_id=?)""",
                (sel_date, t["id"]),
            ).fetchall()}
            t["members"] = members
            t["missing"] = [m for m in members if m["id"] not in reported_ids]
            t["reported_count"] = len(members) - len(t["missing"])
            t["total"] = len(members)
            t["participation"] = round(t["reported_count"] * 100 / max(t["total"], 1))
    return ctx(req, "reminders.html", user=u, teams=teams,
               sel_date=sel_date, active="reminders")


# =====================================================
# EXPORT — 주간 요약 CSV (사이클 68: openpyxl 제거 → csv 표준 모듈)
# 대표 결정 (2026-04-27 14:20:16): openpyxl 제거(Remove)
# 출처: 99_DISPATCH/외부자산_결정_2026-04-27.md
# 단일 .csv 안에 [전사 요약] 섹션 + [카드 상세] 섹션 2개를 빈 줄로 분리해 출력.
# UTF-8 BOM 으로 Excel/구글시트 한글 호환.
# =====================================================
@app.get("/export/weekly")
async def export_weekly(req: Request, wk_mon: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    import csv as _csv
    if not wk_mon:
        td = date.today()
        wk_mon = (td - timedelta(days=td.weekday())).isoformat()
    mon = datetime.strptime(wk_mon, "%Y-%m-%d").date()
    sun = mon + timedelta(days=6)

    buf = io.StringIO()
    buf.write("﻿")  # UTF-8 BOM (Excel 한글 호환)
    w = _csv.writer(buf)

    # 헤더 메타
    w.writerow([f"㈜케이엔케이 주간 업무 요약  {mon} ~ {sun}"])
    w.writerow([])

    with db_session() as c:
        # 섹션 1: 전사 요약
        w.writerow(["[전사 요약]"])
        w.writerow(["팀코드", "팀명", "팀장", "인원", "카드수", "완료", "지연", "공수(h)"])
        teams = [dict(r) for r in c.execute(
            """SELECT t.*, u.name AS leader_name,
                      (SELECT COUNT(*) FROM users WHERE team_id=t.id AND is_active=1) AS mc
               FROM teams t LEFT JOIN users u ON t.leader_id=u.id
               ORDER BY t.display_order"""
        ).fetchall()]
        for t in teams:
            s = c.execute(
                """SELECT COUNT(*) AS total,
                          SUM(CASE WHEN status='완료' THEN 1 ELSE 0 END) AS done,
                          SUM(CASE WHEN status='지연' THEN 1 ELSE 0 END) AS delay,
                          COALESCE(SUM(hours),0) AS hours
                   FROM tasks tk JOIN users u ON tk.user_id=u.id
                   WHERE u.team_id=? AND tk.work_date>=? AND tk.work_date<=?""",
                (t["id"], mon.isoformat(), sun.isoformat()),
            ).fetchone()
            w.writerow([t["code"], t["name"], t["leader_name"] or "-", t["mc"],
                        s["total"] or 0, s["done"] or 0, s["delay"] or 0,
                        round(s["hours"] or 0, 1)])

        # 섹션 사이 공백
        w.writerow([])
        w.writerow([])

        # 섹션 2: 카드 상세
        w.writerow(["[카드 상세]"])
        w.writerow(["날짜", "팀", "이름", "직급", "제목", "분류", "프로젝트",
                    "고객사", "상태", "공수(h)"])
        rows = c.execute(
            """SELECT tk.work_date, t.name AS team_name, u.name AS user_name, u.rank,
                      tk.title, tk.category, p.name AS pj, cu.name AS cu,
                      tk.status, tk.hours
               FROM tasks tk JOIN users u ON tk.user_id=u.id
               LEFT JOIN teams t ON u.team_id=t.id
               LEFT JOIN projects p ON tk.project_id=p.id
               LEFT JOIN customers cu ON tk.customer_id=cu.id
               WHERE tk.work_date>=? AND tk.work_date<=?
               ORDER BY tk.work_date, t.display_order, u.id""",
            (mon.isoformat(), sun.isoformat()),
        ).fetchall()
        for r in rows:
            w.writerow([r["work_date"], r["team_name"] or "", r["user_name"] or "",
                        r["rank"] or "", r["title"] or "", r["category"] or "",
                        r["pj"] or "", r["cu"] or "", r["status"] or "",
                        r["hours"] or 0])

    fname = f"KNK_주간요약_{mon}_{sun}.csv"
    from urllib.parse import quote
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode("utf-8")),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )


# =====================================================
# HAIST WORKS — 변경 Inform 시스템 (1순위 ② / 6팀 + 제조2 사고 사례)
# 설계: HAIST_WORKS/_DESIGN_변경_Inform.md
# =====================================================
from .database import (changes_list, change_get, change_create,
                        change_get_impacts, change_get_reads,
                        change_mark_read, change_ack, change_delete,
                        change_unread_count, change_recent_count,
                        CHANGE_TYPES, CHANGE_URGENCIES, CHANGE_STATUSES,
                        CHANGE_SOURCES,
                        detect_impact_teams, hiworks_notify)


@app.get("/changes", response_class=HTMLResponse)
async def changes_page(req: Request, q: str = "", change_type: str = "",
                       urgency: str = "", status: str = "", scope: str = "all"):
    """변경 목록 (필터 + scope=me/all)"""
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    scope_user_id = u["id"] if scope == "me" else None
    rows = changes_list(q=q, change_type=change_type, urgency=urgency,
                         status=status, scope_user_id=scope_user_id)
    return ctx(req, "changes_list.html", user=u, active="changes",
               changes=rows, q=q, change_type=change_type, urgency=urgency,
               status=status, scope=scope,
               CHANGE_TYPES=CHANGE_TYPES, CHANGE_URGENCIES=CHANGE_URGENCIES,
               CHANGE_STATUSES=CHANGE_STATUSES)


@app.get("/changes/new", response_class=HTMLResponse)
async def changes_new_form(req: Request):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    # 활성 프로젝트 (관리코드 발급된 것)
    with db_session() as c:
        projects = c.execute(
            """SELECT id, mgmt_code, name, biz_div, customer_name
               FROM projects
               WHERE mgmt_code IS NOT NULL AND mgmt_code != ''
               ORDER BY id DESC LIMIT 200"""
        ).fetchall()
    return ctx(req, "change_form.html", user=u, active="changes",
               change=None, projects=projects,
               CHANGE_TYPES=CHANGE_TYPES, CHANGE_URGENCIES=CHANGE_URGENCIES,
               CHANGE_SOURCES=CHANGE_SOURCES)


@app.post("/changes/new")
async def changes_new_submit(
    req: Request,
    change_type: str = Form(...),
    biz_div: str = Form(""),
    target_kind: str = Form(""),
    target_label: str = Form(""),
    project_id: str = Form(""),
    title: str = Form(...),
    description: str = Form(""),
    before_value: str = Form(""),
    after_value: str = Form(""),
    urgency: str = Form("일반"),
    source: str = Form("수동"),
    source_ref: str = Form(""),
    approval_url: str = Form(""),
):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)

    # v5H115 #C1: 제목/유형 필수 + urgency 화이트리스트
    title = (title or "").strip()
    change_type = (change_type or "").strip()
    if not title or not change_type:
        from urllib.parse import quote as _q
        return RedirectResponse("/changes/new?error=" + _q("제목·변경유형은 필수입니다"), 303)
    if urgency not in ("일반", "긴급", "낮음", "보통", "높음"):
        urgency = "일반"

    # 프로젝트 ID에서 사업부 자동 추출 (없으면 폼 입력 사용)
    pid = None
    if project_id:
        try:
            pid = int(project_id)
            with db_session() as c:
                p = c.execute(
                    "SELECT biz_div, mgmt_code, name FROM projects WHERE id=?", (pid,)
                ).fetchone()
            if p:
                if not biz_div:
                    biz_div = p["biz_div"] or ""
                if not target_label:
                    target_label = f"{p['mgmt_code']} {p['name']}"
        except (ValueError, TypeError):
            pid = None

    cid, change_no = change_create({
        "change_type": change_type, "biz_div": biz_div,
        "target_kind": target_kind, "target_label": target_label,
        "project_id": pid, "title": title, "description": description,
        "before_value": before_value, "after_value": after_value,
        "urgency": urgency,
        "source": source, "source_ref": source_ref,
        "approval_url": approval_url.strip() or None,
    }, author_id=u["id"])

    # 알림 발송 (web 게시판 자동 글 + 하이웍스 메신저)
    try:
        notify_change_impacts(cid, change_no, change_type, title, urgency, u)
    except Exception as e:
        print(f"[NOTIFY ERROR] {e}")

    return RedirectResponse(f"/changes/{cid}", 303)


def notify_change_impacts(cid, change_no, change_type, title, urgency, author):
    """변경 등록 후 통합 알림 발송 — 영향 강도별 차별화 (알림 피로 방지)
    - high: 즉시 카톡 푸시 + web 알림 + 게시판
    - medium: web 알림 + 게시판 (카톡 안 보냄)
    - low: 게시판 글에만 노출 (직접 알림 X)
    - 긴급(urgency=긴급): 강도 무시하고 모두 high로 격상
    """
    from .database import get_impact_intensity
    impacts = change_get_impacts(cid)
    icon = "🔴" if urgency == "긴급" else "🟡"
    biz_div = ""  # 변경 본체에서 가져와야 함
    with db_session() as c:
        row = c.execute("SELECT biz_div FROM changes WHERE id=?", (cid,)).fetchone()
        biz_div = row["biz_div"] if row else ""

    # 영향 부서별 강도 분류
    high_teams, medium_teams, low_teams = [], [], []
    for imp in impacts:
        if not imp.get("team_name"):
            continue
        intensity = "high" if urgency == "긴급" else \
                    get_impact_intensity(change_type, biz_div, imp["team_name"])
        if intensity == "high":
            high_teams.append(imp)
        elif intensity == "medium":
            medium_teams.append(imp)
        else:
            low_teams.append(imp)

    # 1. high 강도 → 하이웍스 메신저 즉시 푸시 (토큰 없으면 silent skip)
    for imp in high_teams:
        hiworks_notify(
            channel_id=f"team_{imp['impact_team_id']}",
            text=f"{icon} [{change_type}·직접영향] {title}\n변경번호: {change_no}\n작성자: {author['name']}\n→ /changes/{cid}",
        )

    # 2. 게시판 자동 글 — 모든 영향 부서 표시 (강도별 분류)
    try:
        bid = board_get_or_create_company()
        high_names = ", ".join([imp["team_name"] for imp in high_teams]) or "없음"
        medium_names = ", ".join([imp["team_name"] for imp in medium_teams]) or "없음"
        low_names = ", ".join([imp["team_name"] for imp in low_teams]) or "없음"
        body = (f"종류: {change_type}\n"
                f"긴급도: {urgency}\n"
                f"━━ 영향 강도 ━━\n"
                f"🔴 직접 영향: {high_names}\n"
                f"🟡 일정 영향: {medium_names}\n"
                f"⚪ 참고: {low_names}\n\n"
                f"변경번호: {change_no}\n"
                f"상세: /changes/{cid}")
        board_post_create(bid, author["id"],
                          f"[변경공지] {title}", body, category="공지",
                          approval_status="approved")
    except Exception as e:
        print(f"[BOARD POST ERROR] {e}")


@app.get("/changes/{cid}", response_class=HTMLResponse)
async def changes_detail(req: Request, cid: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    change = change_get(cid)
    if not change:
        return RedirectResponse("/changes", 303)
    impacts = change_get_impacts(cid)
    reads = change_get_reads(cid)
    # 자동 read 기록 (영향자인 경우만)
    is_impacted = any(
        r["user_id"] == u["id"] for r in reads
    )
    if is_impacted:
        change_mark_read(cid, u["id"])
    # 내 ack 상태
    my_ack = next((r for r in reads if r["user_id"] == u["id"]), None)
    return ctx(req, "change_detail.html", user=u, active="changes",
               change=change, impacts=impacts, reads=reads,
               is_impacted=is_impacted, my_ack=my_ack,
               is_author=(change["author_id"] == u["id"]))


@app.post("/changes/{cid}/ack")
async def changes_ack(req: Request, cid: int, note: str = Form("")):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    change_ack(cid, u["id"], note)
    return RedirectResponse(f"/changes/{cid}", 303)


@app.post("/changes/{cid}/delete")
async def changes_delete_submit(req: Request, cid: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    change = change_get(cid)
    if change and (change["author_id"] == u["id"] or u["role"] in ("admin", "ceo")):
        change_delete(cid)
    return RedirectResponse("/changes", 303)


@app.get("/api/changes/unread")
async def api_changes_unread(req: Request):
    u = get_user(req)
    if not u:
        return JSONResponse({"count": 0})
    return JSONResponse({"count": change_unread_count(u["id"])})


@app.get("/api/changes/recent")
async def api_changes_recent(req: Request, scope: str = "me", days: int = 1):
    u = get_user(req)
    if not u:
        return JSONResponse({"count": 0})
    uid = u["id"] if scope == "me" else None
    return JSONResponse({"count": change_recent_count(uid, days)})


# =====================================================
# HAIST WORKS — 진행률 대시보드 (1순위 ① / 8팀 공통 요구)
# =====================================================
from .database import (PHASE_DEFS, PHASE_CODE_TO_LABEL, PHASE_STATUSES,
                        ensure_phases_for_project, progress_matrix,
                        project_phases_get, project_phase_update,
                        progress_summary_for_user)


@app.get("/progress", response_class=HTMLResponse)
async def progress_dashboard(req: Request, biz_div: str = "", customer: str = "",
                              status: str = "", limit: int = 50):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    matrix = progress_matrix(biz_div=biz_div, customer=customer, status=status, limit=limit)
    # v5H46: 템플릿이 r.phases.get(phase_code) 형태로 접근 → dict 변환 (None 보호)
    for r in matrix:
        ph_list = r.get("phases") or []
        if not isinstance(ph_list, list):
            r["phases"] = {}
            continue
        r["phases"] = {p.get("phase_code"): p for p in ph_list if p and p.get("phase_code")}
    return ctx(req, "progress_matrix.html", user=u, active="progress",
               matrix=matrix, biz_div=biz_div, customer=customer, status=status,
               PHASE_DEFS=PHASE_DEFS, PHASE_STATUSES=PHASE_STATUSES)


@app.get("/progress/{project_id}", response_class=HTMLResponse)
async def progress_project_detail(req: Request, project_id: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        proj = c.execute(
            """SELECT p.*, p.name AS project_name FROM projects p WHERE p.id=?""",
            (project_id,)
        ).fetchone()
    if not proj:
        return RedirectResponse("/progress", 303)
    project = dict(proj)
    phases = project_phases_get(project_id)
    return ctx(req, "progress_detail.html", user=u, active="progress",
               project=project, phases=phases,
               PHASE_DEFS=PHASE_DEFS, PHASE_CODE_TO_LABEL=PHASE_CODE_TO_LABEL,
               PHASE_STATUSES=PHASE_STATUSES)


@app.post("/progress/phase/{phase_id}/update")
async def progress_phase_update(
    req: Request,
    phase_id: int,
    status: str = Form(""),
    progress_pct: str = Form(""),
    note: str = Form(""),
    project_id: str = Form(""),
):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    data = {}
    if status:
        data["status"] = status
        # 완료 시 자동 100% / 진행 시 0%면 50%로
        if status == "완료":
            data["progress_pct"] = 100
            data["actual_end"] = _logi_now()[:10]
        elif status == "진행":
            try:
                cur = float(progress_pct) if progress_pct else 0
                if cur == 0:
                    data["progress_pct"] = 50
            except (ValueError, TypeError):
                pass
            data["actual_start"] = _logi_now()[:10]
    if progress_pct:
        try:
            data["progress_pct"] = float(progress_pct)
        except (ValueError, TypeError):
            pass
    if note:
        data["note"] = note
    project_phase_update(phase_id, data, u["id"])
    return RedirectResponse(f"/progress/{project_id}" if project_id else "/progress", 303)


@app.get("/api/progress/summary")
async def api_progress_summary(req: Request):
    u = get_user(req)
    if not u:
        return JSONResponse({"my_open": 0, "delayed": 0})
    return JSONResponse(progress_summary_for_user(u["id"], u.get("team_id")))


# =====================================================
# HAIST WORKS — 요청 티켓 시스템 (1순위 ③ / 10팀 카톡 누락 해결)
# =====================================================
from .database import (tickets_list, ticket_get, ticket_create,
                        ticket_change_status, ticket_comments_list,
                        ticket_add_comment, ticket_delete,
                        tickets_count_for_user, route_ticket_team,
                        TICKET_CATEGORIES, TICKET_URGENCIES, TICKET_STATUSES,
                        TICKET_SOURCES)


# ─── 티켓 권한 가드 (2026-04-28 대표 결재 (나)안) ───────────────
def _can_edit_ticket(u, ticket) -> bool:
    """티켓 본문 편집·상태변경 권한:
    (a) 작성자 본인
    (b) 작성자와 같은 팀원 (휴가·출장 시 대리)
    (c) 수신팀의 can_close_tickets 보유자
    (d) CEO/admin/임원 (전사)
    """
    if not u or not ticket:
        return False
    role = (u.get("role") or "").lower()
    if role in ("ceo", "admin", "executive"):
        return True
    uid = u.get("id")
    # 작성자 본인
    if ticket.get("requester_id") == uid:
        return True
    utid = u.get("team_id")
    # 작성자와 같은 팀원
    try:
        with db_session() as c:
            r = c.execute(
                "SELECT team_id FROM users WHERE id=?", (ticket.get("requester_id"),)
            ).fetchone()
            if r and utid and r["team_id"] == utid:
                return True
    except Exception:
        pass
    # 수신팀의 can_close_tickets 보유자
    if (utid and ticket.get("recipient_team_id") == utid
            and bool(u.get("can_close_tickets"))):
        return True
    return False


def _can_delete_ticket(u, ticket) -> bool:
    """티켓 삭제: 작성자 본인 + CEO/admin (전사)."""
    if not u or not ticket:
        return False
    role = (u.get("role") or "").lower()
    if role in ("ceo", "admin"):
        return True
    return ticket.get("requester_id") == u.get("id")


@app.get("/tickets", response_class=HTMLResponse)
async def tickets_page(req: Request, scope: str = "", q: str = "",
                       category: str = "", urgency: str = "", status: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    # v5H47: scope 미지정 시 ceo/admin/executive 는 'all' (전사) 기본 — 빈 화면 방지
    if not scope:
        scope = "all" if u.get("role") in ("ceo", "admin", "executive") else "me"
    suid = u["id"] if scope == "me" else None
    stid = u.get("team_id") if scope == "team" else None
    rows = tickets_list(scope_user_id=suid, scope_team_id=stid,
                        status=status, category=category, urgency=urgency, q=q)
    return ctx(req, "tickets_list.html", user=u, active="tickets",
               tickets=rows, scope=scope, q=q,
               category=category, urgency=urgency, status=status,
               TICKET_CATEGORIES=TICKET_CATEGORIES,
               TICKET_URGENCIES=TICKET_URGENCIES,
               TICKET_STATUSES=TICKET_STATUSES)


@app.get("/tickets/new", response_class=HTMLResponse)
async def tickets_new_form(req: Request):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        projects = c.execute(
            """SELECT id, mgmt_code, name, biz_div, customer_name FROM projects
               WHERE mgmt_code IS NOT NULL AND mgmt_code != ''
               ORDER BY id DESC LIMIT 200"""
        ).fetchall()
        teams = c.execute("SELECT id, name FROM teams ORDER BY display_order").fetchall()
    return ctx(req, "ticket_form.html", user=u, active="tickets",
               ticket=None, projects=projects, teams=teams,
               TICKET_CATEGORIES=TICKET_CATEGORIES,
               TICKET_URGENCIES=TICKET_URGENCIES)


@app.post("/tickets/new")
async def tickets_new_submit(
    req: Request,
    category: str = Form(...),
    title: str = Form(...),
    description: str = Form(""),
    biz_div: str = Form(""),
    project_id: str = Form(""),
    target_label: str = Form(""),
    recipient_team_id: str = Form(""),
    urgency: str = Form("일반"),
    due_date: str = Form(""),
    hours_estimated: str = Form(""),
    source: str = Form("web"),
    approval_url: str = Form(""),
):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)

    # v5H115 #T1: 제목/카테고리 필수 + urgency 화이트리스트
    title = (title or "").strip()
    category = (category or "").strip()
    if not title or not category:
        from urllib.parse import quote as _q
        return RedirectResponse("/tickets?error=" + _q("제목·카테고리는 필수입니다"), 303)
    if urgency not in ("일반", "긴급", "낮음", "보통"):
        urgency = "일반"

    pid = None
    if project_id:
        try:
            pid = int(project_id)
            with db_session() as c:
                p = c.execute(
                    "SELECT biz_div, mgmt_code, name FROM projects WHERE id=?", (pid,)
                ).fetchone()
            if p:
                if not biz_div:
                    biz_div = p["biz_div"] or ""
                if not target_label:
                    target_label = f"{p['mgmt_code']} {p['name']}"
        except (ValueError, TypeError):
            pid = None

    rtid = None
    if recipient_team_id:
        try:
            rtid = int(recipient_team_id)
        except (ValueError, TypeError):
            rtid = None

    # OPS-P1-A1 [D-003]: 카테고리 기반 자동 라우팅 — 사용자 미선택 시 TICKET_ROUTING 사용
    if rtid is None:
        try:
            from app.database import TICKET_ROUTING  # noqa: F811
        except ImportError:
            TICKET_ROUTING = {}
        target_team_name = TICKET_ROUTING.get(category)
        if target_team_name:
            with db_session() as c:
                row = c.execute(
                    "SELECT id FROM teams WHERE name=? LIMIT 1", (target_team_name,)
                ).fetchone()
            if row:
                rtid = row["id"] if hasattr(row, "keys") else row[0]

    tid, ticket_no = ticket_create({
        "category": category, "title": title, "description": description,
        "biz_div": biz_div, "project_id": pid,
        "target_label": target_label, "recipient_team_id": rtid,
        "urgency": urgency, "due_date": due_date,
        "hours_estimated": hours_estimated, "source": source,
        "approval_url": approval_url,
    }, requester_id=u["id"])

    # 하이웍스 메신저 푸시 (수신 부서) — 토큰 없으면 silent skip
    try:
        ticket = ticket_get(tid)
        if ticket and ticket.get("recipient_team_id"):
            icon = "🔴" if urgency == "긴급" else "🎫"
            hiworks_notify(
                channel_id=f"team-{ticket['recipient_team_id']}",
                text=f"{icon} [{category}] {title}\n티켓: {ticket_no}\n요청: {u['name']}\n→ http://localhost:8081/tickets/{tid}",
            )
    except Exception as e:
        print(f"[TICKET NOTIFY ERROR] {e}")

    return RedirectResponse(f"/tickets/{tid}", 303)


@app.get("/tickets/{tid}", response_class=HTMLResponse)
async def tickets_detail(req: Request, tid: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    ticket = ticket_get(tid)
    if not ticket:
        return RedirectResponse("/tickets", 303)
    comments = ticket_comments_list(tid)
    is_requester = ticket["requester_id"] == u["id"]
    is_recipient = (ticket.get("recipient_user_id") == u["id"]) or \
                   (ticket.get("recipient_team_id") == u.get("team_id"))
    can_edit = _can_edit_ticket(u, ticket)
    can_delete = _can_delete_ticket(u, ticket)
    # 편집 폼용 전팀 목록
    with db_session() as c:
        all_teams_for_edit = [dict(r) for r in c.execute(
            "SELECT id, name FROM teams ORDER BY display_order"
        ).fetchall()]
    return ctx(req, "ticket_detail.html", user=u, active="tickets",
               ticket=ticket, comments=comments,
               is_requester=is_requester, is_recipient=is_recipient,
               can_edit=can_edit, can_delete=can_delete,
               all_teams_for_edit=all_teams_for_edit,
               TICKET_STATUSES=TICKET_STATUSES,
               TICKET_URGENCIES=TICKET_URGENCIES)


@app.post("/tickets/{tid}/status")
async def tickets_status_change(req: Request, tid: int,
                                 new_status: str = Form(...),
                                 note: str = Form("")):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    ticket = ticket_get(tid)
    if not _can_edit_ticket(u, ticket):
        return RedirectResponse(f"/tickets/{tid}?err=no_perm", 303)
    ticket_change_status(tid, new_status, u["id"], note)
    return RedirectResponse(f"/tickets/{tid}", 303)


# ─── 티켓 본문 편집 (2026-04-28 신설 · (나)안 · 이력 자동) ───────────
@app.post("/tickets/{tid}/edit")
async def tickets_edit_submit(
    req: Request, tid: int,
    title: str = Form(""),
    description: str = Form(""),
    due_date: str = Form(""),
    urgency: str = Form(""),
    recipient_team_id: str = Form(""),
    target_label: str = Form(""),
    hours_estimated: str = Form(""),
):
    from .database import ticket_edit as _ticket_edit
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    ticket = ticket_get(tid)
    if not ticket:
        return RedirectResponse("/tickets", 303)
    if not _can_edit_ticket(u, ticket):
        return RedirectResponse(f"/tickets/{tid}?err=no_perm", 303)
    changes = {
        "title": title, "description": description,
        "due_date": due_date, "urgency": urgency,
        "recipient_team_id": recipient_team_id or None,
        "target_label": target_label, "hours_estimated": hours_estimated,
    }
    n = _ticket_edit(tid, changes, editor_id=u["id"],
                     editor_name=f"{u.get('name','')} {u.get('rank','') or ''}".strip())
    return RedirectResponse(f"/tickets/{tid}?edited={n}", 303)


@app.post("/tickets/{tid}/comment")
async def tickets_add_comment(req: Request, tid: int, body: str = Form(...)):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    ticket_add_comment(tid, u["id"], body)
    # OPS-P1-A3 [D-018]: 댓글 후 티켓 관여자(요청자/담당자/팀장) 알림
    try:
        ticket = ticket_get(tid)
        if ticket:
            recipients = set()
            if ticket.get("requester_id") and ticket["requester_id"] != u["id"]:
                recipients.add(ticket["requester_id"])
            if ticket.get("assignee_id") and ticket["assignee_id"] != u["id"]:
                recipients.add(ticket["assignee_id"])
            preview = (body or "").strip().replace("\n", " ")[:40]
            for rid in recipients:
                try:
                    notify_user(
                        user_id=rid,
                        type="ticket_comment",
                        title=f"💬 티켓 댓글 — {ticket.get('ticket_no','')}",
                        body=f"{u.get('name','')}: {preview}",
                        link=f"/tickets/{tid}",
                    )
                except Exception:
                    pass
            # 수신팀 메신저 푸시 (토큰 없으면 silent)
            if ticket.get("recipient_team_id"):
                try:
                    hiworks_notify(
                        channel_id=f"team-{ticket['recipient_team_id']}",
                        text=f"💬 [{ticket.get('ticket_no','')}] {u.get('name','')}: {preview}",
                    )
                except Exception:
                    pass
    except Exception as e:
        print(f"[TICKET COMMENT NOTIFY ERROR] {e}")
    return RedirectResponse(f"/tickets/{tid}", 303)


@app.post("/tickets/{tid}/delete")
async def tickets_delete_submit(req: Request, tid: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    ticket = ticket_get(tid)
    if not ticket:
        return RedirectResponse("/tickets", 303)
    if not _can_delete_ticket(u, ticket):
        return RedirectResponse(f"/tickets/{tid}?err=no_perm_delete", 303)
    ticket_delete(tid)
    return RedirectResponse("/tickets", 303)


@app.get("/api/tickets/count")
async def api_tickets_count(req: Request):
    u = get_user(req)
    if not u:
        return JSONResponse({"my_open": 0, "recv_pending": 0})
    return JSONResponse(tickets_count_for_user(u["id"], u.get("team_id")))


# =====================================================
# HAIST WORKS — 이슈·AS DB (3순위 ⑦)
# 고객사 이슈 추적 → 원인분석 → 재발방지 학습 → 변경 연계
# =====================================================
from .database import (issues_list, issue_get, issue_create, issue_update,
                        issue_delete, issue_logs_get, issues_kpi,
                        ISSUE_SEVERITIES, ISSUE_TYPES, ISSUE_STATUSES,
                        route_issue_team)


@app.get("/issues", response_class=HTMLResponse)
async def issues_page(req: Request, scope: str = "open", q: str = "",
                      status: str = "", severity: str = "", issue_type: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    items = issues_list(scope=scope, user_id=u["id"], team_id=u.get("team_id"),
                        status=status, severity=severity, issue_type=issue_type, q=q)
    kpi = issues_kpi()
    return ctx(req, "issues_list.html", user=u, items=items, kpi=kpi,
               scope=scope, q=q, status=status, severity=severity, issue_type=issue_type,
               ISSUE_STATUSES=ISSUE_STATUSES, ISSUE_SEVERITIES=ISSUE_SEVERITIES,
               ISSUE_TYPES=ISSUE_TYPES, active="issues")


@app.get("/issues/new", response_class=HTMLResponse)
async def issues_new_form(req: Request, project_id: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        teams = [dict(r) for r in c.execute(
            "SELECT id, name FROM teams ORDER BY display_order"
        ).fetchall()]
        projects = [dict(r) for r in c.execute(
            """SELECT id, mgmt_code, name, biz_div, customer_id
               FROM projects WHERE status IN ('active','진행중','planning')
               ORDER BY mgmt_code DESC LIMIT 200"""
        ).fetchall()]
        customers = [dict(r) for r in c.execute(
            "SELECT id, name FROM customers ORDER BY tier DESC, name"
        ).fetchall()]
    return ctx(req, "issue_form.html", user=u, teams=teams, projects=projects,
               customers=customers, default_project_id=project_id,
               ISSUE_SEVERITIES=ISSUE_SEVERITIES, ISSUE_TYPES=ISSUE_TYPES,
               active="issues")


@app.post("/issues/new")
async def issues_new_submit(
    req: Request,
    title: str = Form(...),
    severity: str = Form("중"),
    issue_type: str = Form("AS"),
    biz_div: str = Form(""),
    project_id: str = Form(""),
    customer_id: str = Form(""),
    customer_name: str = Form(""),
    occurred_at: str = Form(""),
    detected_by: str = Form(""),
    description: str = Form(""),
    owner_team_id: str = Form(""),
):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    # v5H115 #I1/#I2: 제목 필수 + severity/issue_type 화이트리스트
    title = (title or "").strip()
    if not title:
        from urllib.parse import quote as _q
        return RedirectResponse("/issues?error=" + _q("제목은 필수입니다"), 303)
    if severity not in ISSUE_SEVERITIES:
        severity = "중"
    if issue_type not in ISSUE_TYPES:
        issue_type = "기타"
    pid = int(project_id) if project_id and project_id.isdigit() else None
    cid = int(customer_id) if customer_id and customer_id.isdigit() else None
    otid = int(owner_team_id) if owner_team_id and owner_team_id.isdigit() else None
    iid, issue_no = issue_create({
        "title": title, "severity": severity, "issue_type": issue_type,
        "biz_div": biz_div, "project_id": pid, "customer_id": cid,
        "customer_name": customer_name,
        "occurred_at": occurred_at, "detected_by": detected_by,
        "description": description, "owner_team_id": otid,
    }, created_by=u["id"])

    # 치명/심각 이슈 → 하이웍스 메신저 즉시 푸시
    if severity in ("치명", "심각"):
        try:
            issue = issue_get(iid)
            recip = f"team-{issue['owner_team_id']}" if issue and issue.get("owner_team_id") else None
            icon = "🚨" if severity == "치명" else "⚠️"
            hiworks_notify(
                channel_id=recip or "all",
                text=(f"{icon} [{severity}·{issue_type}] {title}\n"
                      f"이슈번호: {issue_no}\n고객사: {customer_name or '-'}\n"
                      f"발견: {detected_by or u['name']}\n→ /issues/{iid}"),
            )
        except Exception as e:
            print(f"[ISSUE NOTIFY ERROR] {e}")

    return RedirectResponse(f"/issues/{iid}", 303)


@app.get("/issues/{iid}", response_class=HTMLResponse)
async def issues_detail(req: Request, iid: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    issue = issue_get(iid)
    if not issue:
        return RedirectResponse("/issues", 303)
    logs = issue_logs_get(iid)
    with db_session() as c:
        teams = [dict(r) for r in c.execute(
            "SELECT id, name FROM teams ORDER BY display_order"
        ).fetchall()]
    is_owner = (issue["owner_user_id"] == u["id"]
                or (issue["owner_team_id"] and issue["owner_team_id"] == u.get("team_id")))
    can_edit = is_owner or issue["created_by"] == u["id"] or u["role"] in ("admin", "ceo")
    return ctx(req, "issue_detail.html", user=u, issue=issue, logs=logs,
               teams=teams, can_edit=can_edit,
               ISSUE_STATUSES=ISSUE_STATUSES, ISSUE_SEVERITIES=ISSUE_SEVERITIES,
               ISSUE_TYPES=ISSUE_TYPES, active="issues")


@app.post("/issues/{iid}/update")
async def issues_update_submit(
    req: Request, iid: int,
    status: str = Form(""),
    root_cause: str = Form(""),
    action_taken: str = Form(""),
    prevention: str = Form(""),
    owner_team_id: str = Form(""),
    cost_estimate: str = Form(""),
    related_change_id: str = Form(""),
    comment: str = Form(""),
    note: str = Form(""),
):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    data = {"note": note, "comment": comment.strip() or None}
    if status:
        data["status"] = status
    if root_cause.strip():
        data["root_cause"] = root_cause.strip()
    if action_taken.strip():
        data["action_taken"] = action_taken.strip()
    if prevention.strip():
        data["prevention"] = prevention.strip()
    if owner_team_id and owner_team_id.isdigit():
        data["owner_team_id"] = int(owner_team_id)
    if cost_estimate:
        try:
            _ce = float(cost_estimate)
            # v5H115 #I3: 음수 비용 차단
            if _ce >= 0:
                data["cost_estimate"] = _ce
        except ValueError:
            pass
    # v5H115 #I2: 상태 화이트리스트 (변경 시)
    if status and status not in ISSUE_STATUSES:
        data.pop("status", None)
    if related_change_id and related_change_id.isdigit():
        data["related_change_id"] = int(related_change_id)
    issue_update(iid, data, user_id=u["id"])
    return RedirectResponse(f"/issues/{iid}", 303)


@app.post("/issues/{iid}/delete")
async def issues_delete_submit(req: Request, iid: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    issue = issue_get(iid)
    if issue and (issue["created_by"] == u["id"] or u["role"] in ("admin", "ceo")):
        issue_delete(iid)
    return RedirectResponse("/issues", 303)


@app.get("/api/issues/count")
async def api_issues_count(req: Request):
    u = get_user(req)
    if not u:
        return JSONResponse({"open": 0, "critical": 0})
    kpi = issues_kpi()
    return JSONResponse({"open": kpi["open"], "critical": kpi["critical"]})


# =====================================================
# HAIST WORKS — 게시판 라우트
# =====================================================
from .database import (board_get_or_create_company, board_get_or_create_team,
                        board_posts_list, board_posts_pending, board_post_get,
                        board_post_create, board_post_update, board_post_delete,
                        board_post_approve, board_post_reject, board_post_toggle_pin,
                        board_post_increment_view,
                        board_comments_list, board_comment_create, board_comment_delete,
                        BOARD_CATEGORIES)


def _is_team_leader(user, team_id):
    """사용자가 해당 팀의 팀장인지 확인"""
    if not user:
        return False
    if user.get("role") in ("admin", "ceo", "executive"):
        return True
    if user.get("role") == "leader" and user.get("team_id") == team_id:
        return True
    return False


@app.get("/board/company", response_class=HTMLResponse)
async def board_company(req: Request):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    bid = board_get_or_create_company()
    posts = board_posts_list(bid)
    can_write = u["role"] in ("admin", "ceo", "executive")
    return ctx(req, "board_list.html", user=u, active="board_company",
               board_id=bid, board_name="전사 게시판", board_type="company",
               posts=posts, can_write=can_write, is_leader=False,
               pending_count=0, CATEGORIES=BOARD_CATEGORIES)


@app.get("/board/team", response_class=HTMLResponse)
async def board_team_my(req: Request):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    tid = u.get("team_id")
    # 팀 미소속 (admin/ceo 등) → 전체 부서 게시판 목록
    if not tid:
        with db_session() as c:
            teams = c.execute(
                """SELECT t.id, t.name, t.sector,
                          (SELECT COUNT(*) FROM board_posts bp
                           JOIN boards b ON bp.board_id=b.id
                           WHERE b.type='team' AND b.team_id=t.id
                           AND bp.approval_status='approved') AS post_count,
                          (SELECT COUNT(*) FROM board_posts bp
                           JOIN boards b ON bp.board_id=b.id
                           WHERE b.type='team' AND b.team_id=t.id
                           AND bp.approval_status='pending') AS pending_count
                   FROM teams t ORDER BY t.display_order"""
            ).fetchall()
        return ctx(req, "board_teams.html", user=u, active="board_team",
                   teams=teams)
    bid = board_get_or_create_team(tid)
    is_leader = _is_team_leader(u, tid)
    posts = board_posts_list(bid, include_pending=is_leader)
    pending = board_posts_pending(bid) if is_leader else []
    return ctx(req, "board_list.html", user=u, active="board_team",
               board_id=bid, board_name=f"{u.get('team_name','')} 게시판",
               board_type="team", team_id=tid,
               posts=posts, can_write=True, is_leader=is_leader,
               pending_count=len(pending), pending_posts=pending,
               CATEGORIES=BOARD_CATEGORIES)


@app.get("/board/team/{team_id}", response_class=HTMLResponse)
async def board_team_specific(req: Request, team_id: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    bid = board_get_or_create_team(team_id)
    is_leader = _is_team_leader(u, team_id)
    with db_session() as c:
        t = c.execute("SELECT name FROM teams WHERE id=?", (team_id,)).fetchone()
    team_name = t["name"] if t else f"팀{team_id}"
    posts = board_posts_list(bid, include_pending=is_leader)
    pending = board_posts_pending(bid) if is_leader else []
    can_write = (u.get("team_id") == team_id) or u["role"] in ("admin", "ceo", "executive")
    return ctx(req, "board_list.html", user=u, active="board_team",
               board_id=bid, board_name=f"{team_name} 게시판",
               board_type="team", team_id=team_id,
               posts=posts, can_write=can_write, is_leader=is_leader,
               pending_count=len(pending), pending_posts=pending,
               CATEGORIES=BOARD_CATEGORIES)


@app.get("/board/new", response_class=HTMLResponse)
async def board_new_form(req: Request, board_id: int = 0):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    return ctx(req, "board_form.html", user=u, active="board",
               board_id=board_id, post=None, CATEGORIES=BOARD_CATEGORIES)


@app.post("/board/new")
async def board_new_submit(req: Request,
                           board_id: int = Form(...),
                           title: str = Form(...),
                           body: str = Form(""),
                           category: str = Form("일반")):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    # v5H115 #B1: 제목 필수 + 카테고리 화이트리스트
    title = (title or "").strip()
    if not title:
        from urllib.parse import quote as _q
        return RedirectResponse(f"/board/new?board_id={board_id}&error="
                                + _q("제목은 필수입니다"), 303)
    if BOARD_CATEGORIES and category not in BOARD_CATEGORIES:
        category = "일반" if "일반" in BOARD_CATEGORIES else BOARD_CATEGORIES[0]
    # 전사 게시판: admin/ceo/executive만 → 바로 approved
    # 부서 게시판: 팀장/경영진 → approved, 일반 부서원 → pending
    with db_session() as c:
        board = c.execute("SELECT type, team_id FROM boards WHERE id=?", (board_id,)).fetchone()
    if not board:
        return RedirectResponse("/home", 303)

    if board["type"] == "company":
        status = "approved"
        redirect_url = "/board/company"
    else:
        if _is_team_leader(u, board["team_id"]):
            status = "approved"
        else:
            status = "pending"
        redirect_url = f"/board/team/{board['team_id']}"

    board_post_create(board_id, u["id"], title, body, category, status)
    return RedirectResponse(redirect_url, 303)


@app.get("/board/post/{post_id}", response_class=HTMLResponse)
async def board_post_detail(req: Request, post_id: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    post = board_post_get(post_id)
    if not post:
        return RedirectResponse("/home", 303)
    board_post_increment_view(post_id)
    comments = board_comments_list(post_id)
    is_leader = _is_team_leader(u, post.get("board_team_id"))
    is_author = post["author_id"] == u["id"]
    return ctx(req, "board_detail.html", user=u, active="board",
               post=post, comments=comments,
               is_leader=is_leader, is_author=is_author)


@app.post("/board/post/{post_id}/comment")
async def board_add_comment(req: Request, post_id: int, body: str = Form(...)):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    board_comment_create(post_id, u["id"], body)
    return RedirectResponse(f"/board/post/{post_id}", 303)


@app.post("/board/post/{post_id}/approve")
async def board_approve(req: Request, post_id: int):
    # OPS-P1-G2 [D-006]: 게시글 승인 권한 명시 검증 (leader/admin/ceo 만)
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    post = board_post_get(post_id)
    if not post:
        return RedirectResponse("/home", 303)
    is_team_leader = _is_team_leader(u, post.get("board_team_id"))
    is_priv = u.get("role") in ("admin", "ceo")
    if not (is_team_leader or is_priv):
        raise HTTPException(403, "게시글 승인 권한이 없습니다 (팀장/관리자/대표 전용)")
    board_post_approve(post_id, u["id"])
    return RedirectResponse(f"/board/team/{post['board_team_id']}", 303)


@app.post("/board/post/{post_id}/reject")
async def board_reject(req: Request, post_id: int, reason: str = Form("")):
    # OPS-P1-G2 [D-006]: 게시글 반려 권한 명시 검증
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    post = board_post_get(post_id)
    if not post:
        return RedirectResponse("/home", 303)
    is_team_leader = _is_team_leader(u, post.get("board_team_id"))
    is_priv = u.get("role") in ("admin", "ceo")
    if not (is_team_leader or is_priv):
        raise HTTPException(403, "게시글 반려 권한이 없습니다 (팀장/관리자/대표 전용)")
    board_post_reject(post_id, u["id"], reason)
    return RedirectResponse(f"/board/team/{post['board_team_id']}", 303)


@app.post("/board/post/{post_id}/pin")
async def board_pin(req: Request, post_id: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    post = board_post_get(post_id)
    if post and _is_team_leader(u, post.get("board_team_id")):
        board_post_toggle_pin(post_id)
    redir = f"/board/team/{post['board_team_id']}" if post and post["board_type"] == "team" else "/board/company"
    return RedirectResponse(redir, 303)


@app.post("/board/post/{post_id}/delete")
async def board_delete(req: Request, post_id: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    post = board_post_get(post_id)
    if post and (post["author_id"] == u["id"] or _is_team_leader(u, post.get("board_team_id"))):
        board_post_delete(post_id)
    redir = f"/board/team/{post['board_team_id']}" if post and post["board_type"] == "team" else "/board/company"
    return RedirectResponse(redir, 303)


@app.post("/board/comment/{cid}/delete")
async def board_del_comment(req: Request, cid: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        cm = c.execute("SELECT bc.post_id, bc.author_id FROM board_comments bc WHERE bc.id=?", (cid,)).fetchone()
    if cm and (cm["author_id"] == u["id"] or u["role"] in ("admin", "ceo")):
        board_comment_delete(cid)
        return RedirectResponse(f"/board/post/{cm['post_id']}", 303)
    return RedirectResponse("/home", 303)


# =====================================================
# HAIST WORKS — 물류 모듈 라우트
# =====================================================
from . import database as _logi


# ── 관리자: 물류 권한 토글 ─────────────────────────────
@app.post("/api/admin/user-logistics")
async def admin_toggle_logistics(request: Request,
                                 user_id: int = Form(...),
                                 can_use_logistics: str = Form("0")):
    me = require(request, roles=["admin", "ceo"])
    if not me:
        return JSONResponse({"ok": False, "error": "관리자 권한 필요"}, status_code=403)
    flag = 1 if can_use_logistics == "1" else 0
    with db_session() as c:
        c.execute(
            "UPDATE users SET can_use_logistics = ? WHERE id = ?",
            (flag, user_id),
        )
    return JSONResponse({"ok": True, "user_id": user_id, "can_use_logistics": flag})


# ── 자재·구매 홈 (기존 /logistics — 명칭 통일) ──────────────
@app.get("/logistics", response_class=HTMLResponse)
async def logi_dashboard(request: Request):
    """자재·구매 센터 (부품·공급사·발주·입출고·수불·환율)"""
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_view_logistics(u):
        return RedirectResponse("/home", 303)
    parts_stats = _logi.parts_count()
    from .database import stock_kpi as _stock_kpi
    try:
        s_kpi = _stock_kpi()
    except Exception:
        s_kpi = None
    return ctx(request, "logistics_home.html",
               user=u, active="logistics",
               parts_stats=parts_stats,
               stock_kpi=s_kpi)


# ── 매출·영업 홈 (신규 · 2026-04-21 도메인 분리) ──────────────
@app.get("/sales", response_class=HTMLResponse)
async def sales_dashboard(request: Request):
    """매출·영업 센터 (프로젝트·관리코드·고객사·수주·매출 KPI)"""
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    # Plan Y S1 회귀 #2: /sales 는 can_use_sales (도메인 분리). /logistics 와 권한 독립
    if not can_view_sales(u):
        return RedirectResponse("/home", 303)
    proj_stats = _logi.projects_count_logi()
    # 매출 KPI (Victor sales 핸들러와 동일 로직)
    today = date.today()
    ym = today.strftime("%Y-%m")
    year = today.year
    with db_session() as c:
        month = c.execute(
            """SELECT COALESCE(SUM(order_amount),0) AS total, COUNT(*) AS cnt
               FROM projects WHERE order_date LIKE ? AND order_amount > 0""",
            (f"{ym}%",)).fetchone()
        ytd = c.execute(
            """SELECT COALESCE(SUM(order_amount),0) AS total, COUNT(*) AS cnt
               FROM projects WHERE order_date LIKE ? AND order_amount > 0""",
            (f"{year}%",)).fetchone()
        by_biz = [dict(r) for r in c.execute(
            """SELECT biz_div, COALESCE(SUM(order_amount),0) AS total, COUNT(*) AS cnt
               FROM projects WHERE order_date LIKE ? AND biz_div IN ('T','M')
               GROUP BY biz_div""",
            (f"{year}%",)).fetchall()]
        by_stage = [dict(r) for r in c.execute(
            """SELECT stage, COUNT(*) AS cnt,
                      COALESCE(SUM(order_amount),0) AS amount
               FROM projects WHERE stage IS NOT NULL AND stage != ''
               GROUP BY stage ORDER BY cnt DESC"""
        ).fetchall()]
        recent = [dict(r) for r in c.execute(
            """SELECT id, mgmt_code, name, customer_name, stage, order_amount,
                      order_date, due_date, biz_div
               FROM projects
               WHERE mgmt_code IS NOT NULL AND mgmt_code != ''
               ORDER BY id DESC LIMIT 10""").fetchall()]
        customers_top = [dict(r) for r in c.execute(
            """SELECT customer_name, COUNT(*) AS cnt,
                      COALESCE(SUM(order_amount),0) AS total
               FROM projects
               WHERE customer_name IS NOT NULL AND customer_name != ''
                 AND order_date LIKE ?
               GROUP BY customer_name
               ORDER BY total DESC LIMIT 5""",
            (f"{year}%",)).fetchall()]
    sales_kpi = {
        "month_total": month["total"], "month_cnt": month["cnt"],
        "ytd_total": ytd["total"], "ytd_cnt": ytd["cnt"],
        "by_biz": by_biz, "by_stage": by_stage,
        "recent": recent, "top_customers": customers_top,
        "ym": ym, "year": year,
    }
    return ctx(request, "sales_home.html",
               user=u, active="sales",
               proj_stats=proj_stats, sales_kpi=sales_kpi)


# ── 부품 마스터 (parts) ────────────────────────────────
@app.get("/parts", response_class=HTMLResponse)
async def parts_list_page(request: Request, q: str = "", biz_div: str = "",
                          category: str = ""):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_view_logistics(u):
        return RedirectResponse("/home", 303)
    rows = _logi.parts_list(q=q, biz_div=biz_div, category=category)
    return ctx(request, "parts.html",
               user=u, active="parts",
               parts=rows, q=q, biz_div=biz_div, category=category)


@app.get("/parts/new", response_class=HTMLResponse)
async def parts_new_form(request: Request):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    return ctx(request, "part_form.html", user=u, active="parts", part=None)


@app.post("/parts/new")
async def parts_new_submit(
    request: Request,
    part_no: str = Form(...), part_name: str = Form(...),
    spec: str = Form(""), maker: str = Form(""), origin: str = Form(""),
    unit: str = Form("EA"), currency: str = Form("KRW"),
    std_price: str = Form("0"), biz_div: str = Form(""),
    category: str = Form(""), note: str = Form(""), is_active: str = Form("1"),
    safety_stock: str = Form("0"), location: str = Form(""),
):
    _u = get_user(request)
    if not _u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(_u):
        return RedirectResponse("/home", 303)
    # v5H113: 검증 실패 친절한 에러
    try:
        _logi.parts_create({
            "part_no": part_no, "part_name": part_name, "spec": spec,
            "maker": maker, "origin": origin, "unit": unit,
            "currency": currency, "std_price": std_price,
            "biz_div": biz_div, "category": category, "note": note,
            "is_active": is_active,
            "safety_stock": safety_stock, "location": location,
        })
    except ValueError as ve:
        from urllib.parse import quote
        return RedirectResponse(f"/parts/new?error={quote(str(ve))}", status_code=303)
    return RedirectResponse("/parts", status_code=303)


@app.get("/parts/{pid}/edit", response_class=HTMLResponse)
async def parts_edit_form(request: Request, pid: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    p = _logi.parts_get(pid)
    if not p:
        return RedirectResponse("/parts", status_code=303)
    # v5H129: 첨부 목록 같이 표시
    try:
        attachments = part_attachments_list(pid)
    except Exception:
        attachments = []
    return ctx(request, "part_form.html", user=u, active="parts",
               part=p, attachments=attachments)


@app.post("/parts/{pid}/edit")
async def parts_edit_submit(
    request: Request,
    pid: int, part_no: str = Form(...), part_name: str = Form(...),
    spec: str = Form(""), maker: str = Form(""), origin: str = Form(""),
    unit: str = Form("EA"), currency: str = Form("KRW"),
    std_price: str = Form("0"), biz_div: str = Form(""),
    category: str = Form(""), note: str = Form(""), is_active: str = Form("1"),
    safety_stock: str = Form("0"), location: str = Form(""),
):
    _u = get_user(request)
    if not _u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(_u):
        return RedirectResponse("/home", 303)
    # v5H113: 검증 실패 친절한 에러
    try:
        _logi.parts_update(pid, {
            "part_no": part_no, "part_name": part_name, "spec": spec,
            "maker": maker, "origin": origin, "unit": unit,
            "currency": currency, "std_price": std_price,
            "biz_div": biz_div, "category": category, "note": note,
            "is_active": is_active,
            "safety_stock": safety_stock, "location": location,
        })
    except ValueError as ve:
        from urllib.parse import quote
        return RedirectResponse(f"/parts/{pid}/edit?error={quote(str(ve))}", status_code=303)
    return RedirectResponse("/parts", status_code=303)


@app.post("/parts/{pid}/delete")
async def parts_delete_submit(request: Request, pid: int):
    _u = get_user(request)
    if not _u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(_u):
        return RedirectResponse("/home", 303)
    # v5H112: cascade 안전망 + JSON 에러 (v5H98 패턴)
    try:
        _logi.parts_delete(pid)
    except Exception as e:
        return JSONResponse(
            {"ok": False, "error": f"자재 삭제 실패: {e}"}, status_code=400
        )
    return RedirectResponse("/parts", status_code=303)


# ── 프로젝트 / 관리코드 발행대장 ─────────────────────────
@app.get("/projects", response_class=HTMLResponse)
async def projects_list_page(request: Request, q: str = "", biz_div: str = "",
                             stage: str = "", status: str = "",
                             project_type: str = ""):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_view_sales(u):
        return RedirectResponse("/home", 303)
    # v5H218: biz_div(T/M) = 진행 사업부, project_type 으로 유형(검사기/자동화/기타/소모품) 구분
    rows: list = []
    _proj_rows = _logi.projects_list_logi(q=q, biz_div=biz_div, stage=stage, status=status,
                                           project_type=project_type)
    # v5H217b: sqlite3.Row 는 .get() 미지원 → dict 변환
    rows = [dict(r) for r in _proj_rows]
    try:
        with db_session() as _ds:
            for r in rows:
                r["_kind"] = "project"
                _pid = r.get("id")
                if _pid:
                    r["_disp_status"] = _pwf.compute_project_display_status(
                        _ds, int(_pid),
                        fallback_stage=(r.get("status") or r.get("stage") or "")
                    )
    except Exception:
        pass
    # v5H217+v5H218: 소모품 묶음(consumable_orders) 도 통합. biz_div 필터 시 consumable.biz_div 매칭
    if project_type in ("", "CONSUMABLE"):
        try:
            from . import consumables as _co_mod
            _co_status_map = {"DRAFT": "초기협의", "QUOTED": "견적발행",
                               "CONFIRMED": "진행중", "SHIPPED": "납품완료",
                               "PAID": "납품완료", "CANCELLED": "취소"}
            _co_rows = _co_mod.co_list(status="", q=q, limit=500)
            for cr in _co_rows:
                _co_biz = (cr.get("biz_div") or "").strip().upper() if cr.get("biz_div") else ""
                # 사업부 필터 적용 (T/M)
                if biz_div and biz_div in ("T", "M") and _co_biz != biz_div:
                    continue
                _st_logi = _co_status_map.get(cr.get("status") or "DRAFT", "초기협의")
                if status and _st_logi != status:
                    continue
                _name_disp = (cr.get("customer_name") or "고객사 미정") + " 소모품 발주 (" + (cr.get("co_no") or f"#{cr.get('id')}") + ")"
                rows.append({
                    "_kind": "consumable",
                    "id": cr.get("id"),
                    "mgmt_code": cr.get("mgmt_code") or "—",
                    "name": _name_disp,
                    "project_name": _name_disp,
                    "customer_name": cr.get("customer_name") or "—",
                    "biz_div": _co_biz or None,  # T/M (미선택 시 None → '미분류' 노출)
                    "project_type": "CONSUMABLE",
                    "status": _st_logi,
                    "stage": _st_logi,
                    "order_amount": cr.get("total_amount") or 0,
                    "order_date": cr.get("order_date") or "",
                    "_co_no": cr.get("co_no"),
                })
        except Exception as _e:
            print(f"[v5H218] 소모품 병합 실패: {_e}")
    # 정렬: 발주일 desc → 관리코드 desc
    rows.sort(key=lambda r: (str(r.get("order_date") or ""), str(r.get("mgmt_code") or "")), reverse=True)
    return ctx(request, "projects.html",
               user=u, active="sales_projects",
               projects=rows, q=q, biz_div=biz_div, stage=stage, status=status,
               project_type=project_type,
               STAGES=_logi.STAGES, STATUSES=_logi.LOGI_STATUSES,
               PROJECT_TYPES=_logi.PROJECT_TYPES,
               PROJECT_TYPE_LABELS=_logi.PROJECT_TYPE_LABELS)


@app.get("/projects/new", response_class=HTMLResponse)
async def projects_new_form(request: Request,
                            type: str = "", biz_div: str = ""):
    """v5H148 (2026-05-05) — 대표 직접 지시: 프로젝트 등록 진입점 통합.
    - type 파라미터 없으면 4-카드 chooser (T검사기/M자동화/기타/소모품)
    - type 있으면 기존 폼 + biz_div/project_type 사전 선택
    백워드 호환: 직접 ?type=NEW_EQUIP 진입 정상 동작."""
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_sales(u):
        return RedirectResponse("/home", 303)
    # type 없으면 chooser 페이지
    _t = (type or "").strip().upper()
    if not _t:
        return ctx(request, "project_new_chooser.html",
                   user=u, active="sales_projects")
    # type 정규화 (CONSUMABLE/SERVICE 는 더이상 폼에서 노출 안 함 → OTHER)
    if _t not in ("NEW_EQUIP", "OTHER"):
        _t = "OTHER" if _t in ("CONSUMABLE", "SERVICE") else "NEW_EQUIP"
    _bd = (biz_div or "").strip().upper()
    if _bd not in ("T", "M"):
        _bd = ""
    # 폼이 project.* 로 prefill 을 읽으므로 가벼운 placeholder 객체 전달
    _preset = {
        "project_type": _t,
        "biz_div": _bd,
        "name": "", "customer_name": "", "po_type": "",
        "mgmt_code": "", "stage": "제안작성", "currency": "KRW",
        "id": None, "parent_project_id": None,
    }
    return ctx(request, "project_form.html",
               user=u, active="sales_projects",
               project=None,
               preset=_preset,
               STAGES=_logi.STAGES, STATUSES=_logi.LOGI_STATUSES,
               PO_TYPES=_logi.PO_TYPES,
               customers=_logi.customers_for_picker())


@app.post("/projects/new")
async def projects_new_submit(request: Request):
    """v5H52: project_form.html 의 실제 필드명(name/customer_name 등)과
    정합. 콤마 들어간 KRW 입력값도 자동 정리."""
    _u = get_user(request)
    if not _u:
        return RedirectResponse("/login", 303)
    if not can_use_sales(_u):
        return RedirectResponse("/home", 303)
    form = await request.form()
    # 템플릿/대체 필드명 둘 다 허용 (호환)
    project_name = (form.get("name") or form.get("project_name") or "").strip()
    customer = (form.get("customer_name") or form.get("customer") or "").strip()
    biz_div = (form.get("biz_div") or "").strip()
    # v5H192: 필수 필드 종합 검증 (비고 제외) — 빈 항목이 있으면 즉시 안내
    # v5H193: 검증 실패 redirect 시 type/biz_div 보존 (chooser 로 빠지지 않도록)
    _ptype_form = (form.get("project_type") or "NEW_EQUIP").strip().upper()
    if _ptype_form not in ("NEW_EQUIP", "OTHER", "CONSUMABLE", "SERVICE"):
        _ptype_form = "NEW_EQUIP"
    _back_qs = f"type={_ptype_form}"
    if biz_div:
        _back_qs += f"&biz_div={biz_div}"
    def _err_redirect(code: str, extra: str = "") -> RedirectResponse:
        url = f"/projects/new?{_back_qs}&error={code}"
        if extra:
            url += "&" + extra
        return RedirectResponse(url, status_code=303)
    if not project_name:
        return _err_redirect("name_required")
    # v5H218: 진행 사업부는 항상 T 또는 M 이어야 함 (K 기타는 chooser 에서 K 로 들어와도 폼에서 T/M 으로 보정)
    if biz_div not in ("T", "M"):
        return _err_redirect("biz_div_required")
    if not customer:
        return _err_redirect("customer_required")
    # 고객사 화이트리스트 검증
    with db_session() as _cc:
        _ok = _cc.execute(
            "SELECT 1 FROM customers WHERE name=? LIMIT 1", (customer,)
        ).fetchone()
    if not _ok:
        from urllib.parse import quote as _q
        return _err_redirect("customer_not_registered", f"cust={_q(customer)}")
    # v5H211: 단계별 차등 검증 — 진행중/납품완료(또는 즉시 수주확정)면 strict, 그 전 단계는 단가·발주일·납기·환율 옵션
    # 제안 단계(초기협의/제안서전달/견적발행/수주예정/보류/기타)는 추상적 견적이라 강제 입력 부담 제거
    po_type_v = (form.get("po_type") or "").strip()
    order_date_v = (form.get("order_date") or "").strip()
    due_date_v = (form.get("due_date") or "").strip()
    raw_price_v = (form.get("unit_price") or "0").strip().replace(",", "")
    try:
        _up_v = float(raw_price_v) if raw_price_v else 0
    except ValueError:
        _up_v = 0
    _status_v = (form.get("status") or "초기협의").strip()
    _confirm_now_v = (form.get("confirm_now") or "").strip() in ("1", "on", "true", "yes")
    _strict = _status_v in ("진행중", "납품완료") or _confirm_now_v
    # PO유형은 항상 필수 (관리코드 산출 키)
    if not po_type_v:
        return _err_redirect("po_type_required")
    if _strict:
        if not order_date_v:
            return _err_redirect("order_date_required")
        if not due_date_v:
            return _err_redirect("due_date_required")
        if _up_v <= 0:
            return _err_redirect("unit_price_required")
    # 외화 시 기준환율 — strict 단계에서만 필수 (제안 단계는 환율 미정 OK)
    _ccy_v = (form.get("currency") or "KRW").strip().upper()
    if _strict and _ccy_v != "KRW":
        _fx_raw = (form.get("fx_rate") or "").strip().replace(",", "")
        try:
            _fx_v = float(_fx_raw) if _fx_raw else 0
        except ValueError:
            _fx_v = 0
        if _fx_v <= 0:
            return _err_redirect("fx_rate_required", f"ccy={_ccy_v}")
    # 콤마 제거 후 숫자로
    raw_amt = (form.get("order_amount") or "0").strip().replace(",", "")
    try:
        amt = float(raw_amt) if raw_amt else 0
    except ValueError:
        amt = 0
    # v5H132: 단가 × 수량 = 금액 (폼이 hidden order_amount 를 미리 계산해 보내지만 서버 재검증)
    raw_price = (form.get("unit_price") or "0").strip().replace(",", "")
    try:
        unit_price = float(raw_price) if raw_price else 0.0
    except ValueError:
        unit_price = 0.0
    raw_qty = (form.get("unit_qty") or "1").strip()
    try:
        unit_qty = int(float(raw_qty))
    except (TypeError, ValueError):
        unit_qty = 1
    if unit_qty < 1:
        unit_qty = 1
    if unit_qty > 100:
        unit_qty = 100
    # 폼 hidden order_amount 가 있으면 우선 사용, 없으면 단가×수량으로 계산.
    # 0 이거나 단가>0 이면 항상 단가×수량으로 재계산 (자가치유)
    if unit_price > 0:
        amt = unit_price * unit_qty
    elif amt > 0 and unit_qty >= 1:
        # 폴백: 단가 미입력이면 amt/qty 를 단가로 추정
        unit_price = amt / unit_qty
    # v5H77: '수주확정 동시발급' 체크 — 등록 직후 confirm_order 호출
    confirm_now = (form.get("confirm_now") or "").strip() in ("1", "on", "true", "yes")
    # v5H214: stage 는 status 에서 자동 매핑 (사용자는 status 만 선택)
    _status_form = (form.get("status") or "초기협의").strip()
    stage_val = stage_from_status(_status_form)
    if confirm_now:
        # 수주 확정과 함께 등록하면 stage 를 즉시 '수주확정' 으로
        stage_val = "수주확정"
    # v5H137: 프로젝트 유형 + 부모 프로젝트 — 폼에서 받음 (백워드 호환: 미전달 → NEW_EQUIP/None)
    _ptype = (form.get("project_type") or "NEW_EQUIP").strip().upper()
    if _ptype not in _logi.PROJECT_TYPES:
        _ptype = "NEW_EQUIP"
    _parent_id_raw = (form.get("parent_project_id") or "").strip()
    _parent_id = int(_parent_id_raw) if _parent_id_raw.isdigit() else None
    # parent_project_id 는 CONSUMABLE/SERVICE 일 때만 의미 있음 (그 외엔 무시)
    if _ptype not in ("CONSUMABLE", "SERVICE"):
        _parent_id = None
    new_pid, _new_code = _logi.projects_create_logi({
        "_changed_by": _u.get("id"),
        "biz_div": biz_div, "project_name": project_name, "customer": customer,
        "model": form.get("model", ""),
        "stage": stage_val,
        "po_type": form.get("po_type", "신규") or "신규",
        "status": form.get("status", "초기협의") or "초기협의",
        "customer_po": form.get("customer_po", ""),
        "currency": (form.get("currency", "KRW") or "KRW").upper(),
        "is_export": form.get("is_export", "0"),
        "order_amount": amt,
        "unit_qty": unit_qty,
        "unit_price": unit_price if unit_price > 0 else None,
        "order_date": form.get("order_date", ""),
        "due_date": form.get("due_date", ""),
        # v5H201: 제안 단계 일정 (수주확정 전 스케줄용)
        "proposal_date": form.get("proposal_date", ""),
        "quotation_date": form.get("quotation_date", ""),
        "pm": form.get("pm", ""), "sales": form.get("sales", ""),
        "note": form.get("note", ""),
        "project_type": _ptype,
        "parent_project_id": _parent_id,
    })
    if confirm_now and new_pid:
        # v5H81: 호기 라인 — (납기, 납품지) 그룹화 키 포함
        # v5H92: 통화(currency) 도 라인별로 받음 (KRW/USD)
        labels = form.getlist("unit_label[]")
        amounts = form.getlist("unit_amount[]")
        dues = form.getlist("unit_due[]")
        ships = form.getlist("unit_ship[]")
        notes_u = form.getlist("unit_note[]")
        currs = form.getlist("unit_currency[]")
        units = []
        for i in range(max(len(labels), len(amounts))):
            # v5H137: 폼에서 라벨 비어 있으면 project_type 기준 자동 라벨
            lbl = (labels[i] if i < len(labels) else "").strip() or _logi.project_unit_label(_ptype, i + 1)
            raw = (amounts[i] if i < len(amounts) else "0").strip().replace(",", "")
            try:
                u_amt = float(raw) if raw else 0
            except ValueError:
                u_amt = 0
            u_due = (dues[i] if i < len(dues) else "").strip()
            u_ship = (ships[i] if i < len(ships) else "").strip()
            u_note = (notes_u[i] if i < len(notes_u) else "").strip()
            # v5H171: 화이트리스트 KRW/USD/VND/JPY/CNY/EUR (폼 옵션과 일치, VND 묵음 손실 버그 수정)
            # 헤더 통화로 대체(빈 셀 시) — 사용자가 헤더에서 USD 선택했으면 호기 라인도 USD 가 default
            _hdr_cur = (form.get("currency", "KRW") or "KRW").upper()
            u_cur = (currs[i] if i < len(currs) else "").strip().upper() or _hdr_cur
            if u_cur not in ("KRW", "USD", "VND", "JPY", "CNY", "EUR"):
                u_cur = _hdr_cur if _hdr_cur in ("KRW","USD","VND","JPY","CNY","EUR") else "KRW"
            if not (labels[i] if i < len(labels) else "") and u_amt == 0:
                continue
            units.append({"label": lbl, "amount": u_amt,
                          "due_date": u_due, "ship_to": u_ship,
                          "note": u_note, "currency": u_cur})
        try:
            with db_session() as c:
                if units:
                    _pwf.confirm_order_multi(
                        c, int(new_pid), units=units,
                        order_date=form.get("order_date", ""),
                        created_by=_u.get("id") or 0,
                        po_number=form.get("customer_po", ""),
                    )
                else:
                    # 호기 라인이 비어 있으면 기존 단일 SO 흐름 (하위호환)
                    _pwf.confirm_order(
                        c, int(new_pid),
                        order_date=form.get("order_date", ""),
                        total_amount=amt,
                        due_date=form.get("due_date", ""),
                        created_by=_u.get("id") or 0,
                        po_number=form.get("customer_po", ""),
                        note=form.get("note", ""),
                    )
        except Exception:
            return RedirectResponse(
                f"/project/{new_pid}?warn=so_issue_failed", status_code=303
            )
        return RedirectResponse(f"/project/{new_pid}", status_code=303)
    # v5H87: confirm_now 미체크 + status 가 won (진행중/납품완료) 인 경우
    # 자동으로 SO 발행 (관리코드만 있고 SO 없는 모순 상태 방지)
    # v5H132: 수량 N → N개 호기 라인 자동 생성 (단가=unit_price)
    # v5H142: NEW_EQUIP 만 자동 SO 발행 (소모품/수리는 consumable_orders 도메인)
    status_val = (form.get("status") or "").strip()
    if not confirm_now and new_pid and status_val in _logi.WON_STATUSES and _ptype == "NEW_EQUIP":
        try:
            with db_session() as c:
                # 이미 SO 가 있으면 발행 안 함 (재진입 방지)
                exists = c.execute(
                    "SELECT 1 FROM orders WHERE project_id=? LIMIT 1", (new_pid,)
                ).fetchone()
                if not exists:
                    # v5H132: 수량 N → N개 호기 (각 단가=unit_price)
                    # v5H137: 라벨 패턴은 project_type 기준 (NEW_EQUIP→호기, CONSUMABLE→회차, SERVICE→차, OTHER→건)
                    _per_unit = unit_price if unit_price > 0 else (amt / max(1, unit_qty))
                    _units_list = [{
                        "label": _logi.project_unit_label(_ptype, i + 1),
                        "amount": _per_unit,
                        "due_date": form.get("due_date", ""),
                        "ship_to": "",
                        "note": "",
                    } for i in range(max(1, unit_qty))]
                    _pwf.confirm_order_multi(
                        c, int(new_pid),
                        units=_units_list,
                        order_date=form.get("order_date", ""),
                        created_by=_u.get("id") or 0,
                        po_number=form.get("customer_po", ""),
                    )
        except Exception:
            return RedirectResponse(
                f"/project/{new_pid}?warn=so_issue_failed", status_code=303
            )
        return RedirectResponse(f"/project/{new_pid}", status_code=303)
    return RedirectResponse("/projects", status_code=303)


# =====================================================
# v5H152 (2026-05-05): 프로젝트 엑셀 일괄 등록
#   - GET  /projects/import-template  → 양식 .xlsx 다운로드
#   - POST /projects/import-xlsx      → 업로드 + 파싱 + 미리보기 JSON
#   - POST /projects/import-confirm   → 확정 INSERT (행 배열 수신)
# =====================================================
PROJ_IMPORT_HEADERS = [
    "프로젝트명", "관리코드", "PO유형", "고객사명", "모델명",
    "거래구분", "발주일", "납기일", "단가", "수량",
    "통화", "상태", "PM", "영업담당", "납품처", "비고"
]
PROJ_IMPORT_CURRENCIES = {"KRW", "USD", "VND"}
PROJ_IMPORT_STATUSES = {
    "초기협의", "제안서전달", "견적발행", "수주예정",
    "진행중", "납품완료", "취소", "보류"
}


def _proj_import_parse_xlsx(file_bytes: bytes) -> list[dict]:
    """업로드된 엑셀을 파싱해 row dict 리스트 반환.
    각 dict: {sheet, row_no, biz_div, name, customer_name, ..., _errors: [...]}.
    빈 프로젝트명 행 / 예제 row(4) 자동 스킵."""
    from openpyxl import load_workbook
    from datetime import datetime, date
    import io as _io
    wb = load_workbook(_io.BytesIO(file_bytes), data_only=True, read_only=True)
    out = []
    # 사업부 매핑: 시트명 → biz_div
    sheet_div_map = {
        "T_검사기": "T",
        "M_자동화": "M",
    }
    # 등록된 고객사 캐시 (대소문자 무시 비교용은 안 함 — 정확 매칭)
    customer_set = set()
    try:
        with db_session() as _cc:
            for r in _cc.execute("SELECT name FROM customers"):
                if r and r[0]:
                    customer_set.add(r[0])
    except Exception:
        pass

    def _to_str(v) -> str:
        if v is None:
            return ""
        if isinstance(v, (datetime, date)):
            return v.strftime("%Y-%m-%d")
        return str(v).strip()

    def _to_date(v) -> str:
        if v is None or v == "":
            return ""
        if isinstance(v, (datetime, date)):
            return v.strftime("%Y-%m-%d")
        s = str(v).strip()
        if not s:
            return ""
        # YYYY-MM-DD / YYYY/MM/DD / YYYY.MM.DD 허용
        for sep in ("-", "/", "."):
            parts = s.replace(sep, "-").split("-")
            if len(parts) == 3:
                try:
                    y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
                    return f"{y:04d}-{m:02d}-{d:02d}"
                except ValueError:
                    continue
        return s  # 파싱 실패 시 원문 (검증에서 잡음)

    for sh_name in wb.sheetnames:
        if sh_name not in sheet_div_map:
            continue  # 안내 시트 등 스킵
        biz_div = sheet_div_map[sh_name]
        ws = wb[sh_name]
        # 데이터 row 5+ (row1=제목, row2=안내, row3=헤더, row4=예제)
        row_no = 0
        for r in ws.iter_rows(min_row=5, values_only=True):
            row_no += 1
            actual_row = row_no + 4  # 사용자에게 표시할 엑셀 행번호
            if not r or all((c is None or str(c).strip() == "") for c in r):
                continue
            # 16개 컬럼 패딩
            cells = list(r) + [None] * (16 - len(r)) if len(r) < 16 else list(r[:16])
            name = _to_str(cells[0])
            if not name:
                continue  # 프로젝트명 빈 행 스킵
            mgmt_code = _to_str(cells[1])
            po_type = _to_str(cells[2]) or "신규"
            customer = _to_str(cells[3])
            model = _to_str(cells[4])
            trade_raw = _to_str(cells[5])
            order_date = _to_date(cells[6])
            due_date = _to_date(cells[7])
            unit_price_raw = _to_str(cells[8])
            unit_qty_raw = _to_str(cells[9])
            currency = _to_str(cells[10]).upper() or "KRW"
            status = _to_str(cells[11]) or "초기협의"
            pm = _to_str(cells[12])
            sales = _to_str(cells[13])
            ship_to = _to_str(cells[14])
            note = _to_str(cells[15])
            errors = []
            # 거래구분
            is_export = 1 if trade_raw in ("수출", "export", "1") else 0
            # 단가
            try:
                unit_price = float(str(unit_price_raw).replace(",", "")) if unit_price_raw else 0.0
            except ValueError:
                unit_price = 0.0
                errors.append(f"단가 형식 오류: '{unit_price_raw}'")
            if unit_price < 0:
                errors.append("단가는 0 이상이어야 합니다")
                unit_price = 0.0
            # 수량
            try:
                unit_qty = int(float(str(unit_qty_raw).replace(",", ""))) if unit_qty_raw else 1
            except ValueError:
                unit_qty = 1
                errors.append(f"수량 형식 오류: '{unit_qty_raw}'")
            if unit_qty < 1 or unit_qty > 100:
                errors.append(f"수량 범위(1~100) 오류: {unit_qty}")
                unit_qty = max(1, min(100, unit_qty))
            # 통화
            if currency not in PROJ_IMPORT_CURRENCIES:
                errors.append(f"통화 화이트리스트 위반: '{currency}' (KRW/USD/VND)")
                currency = "KRW"
            # 상태
            if status not in PROJ_IMPORT_STATUSES:
                errors.append(f"상태 화이트리스트 위반: '{status}'")
                status = "초기협의"
            # 고객사
            cust_warn = ""
            if customer and customer not in customer_set:
                cust_warn = f"미등록 고객사 (텍스트로 저장): '{customer}'"
                errors.append(cust_warn)
            # 날짜 파싱 검증 (YYYY-MM-DD 패턴)
            for dn, dv in (("발주일", order_date), ("납기일", due_date)):
                if dv:
                    parts = dv.split("-")
                    if not (len(parts) == 3 and all(p.isdigit() for p in parts)):
                        errors.append(f"{dn} 날짜 파싱 실패: '{dv}'")
            out.append({
                "sheet": sh_name,
                "biz_div": biz_div,
                "row_no": actual_row,
                "name": name,
                "mgmt_code_input": mgmt_code,
                "po_type": po_type,
                "customer_name": customer,
                "model_name": model,
                "is_export": is_export,
                "trade_label": "수출" if is_export else "내수",
                "order_date": order_date,
                "due_date": due_date,
                "unit_price": unit_price,
                "unit_qty": unit_qty,
                "amount": unit_price * unit_qty,
                "currency": currency,
                "status": status,
                "pm_name": pm,
                "sales_name": sales,
                "ship_to": ship_to,
                "note": note,
                "_errors": errors,
                "_is_warning_only": (len(errors) == 1 and errors[0] == cust_warn),
            })
    return out


@app.get("/projects/import-template")
async def projects_import_template(request: Request):
    """양식 엑셀 파일 다운로드 (v5H152)."""
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_sales(u):
        return RedirectResponse("/home", 303)
    from pathlib import Path as _Path
    p = _Path(__file__).parent / "static" / "templates" / "프로젝트_일괄등록_양식.xlsx"
    if not p.exists():
        return JSONResponse({"error": "양식 파일을 찾을 수 없습니다"}, 404)
    return FileResponse(
        str(p),
        filename="KNK_프로젝트_일괄등록_양식.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/projects/import-xlsx")
async def projects_import_xlsx(request: Request, xlsx: UploadFile = File(...)):
    """엑셀 업로드 → 파싱 → 미리보기 JSON 반환 (v5H152)."""
    u = get_user(request)
    if not u:
        return JSONResponse({"ok": False, "error": "login_required"}, 401)
    if not can_use_sales(u):
        return JSONResponse({"ok": False, "error": "permission_denied"}, 403)
    try:
        body = await xlsx.read()
        rows = _proj_import_parse_xlsx(body)
    except Exception as e:
        return JSONResponse({"ok": False, "error": f"파싱 실패: {e}"}, 400)
    valid_count = sum(1 for r in rows if not r["_errors"])
    error_count = sum(1 for r in rows if r["_errors"])
    return JSONResponse({
        "ok": True,
        "rows": rows,
        "total": len(rows),
        "valid_count": valid_count,
        "error_count": error_count,
    })


@app.post("/projects/import-confirm")
async def projects_import_confirm(request: Request):
    """미리보기 후 사용자 확정 → 일괄 INSERT (v5H152).
    body JSON: {rows: [{...}, ...]}.
    각 row 는 projects_create_logi 호출. project_type='NEW_EQUIP' 강제."""
    u = get_user(request)
    if not u:
        return JSONResponse({"ok": False, "error": "login_required"}, 401)
    if not can_use_sales(u):
        return JSONResponse({"ok": False, "error": "permission_denied"}, 403)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"ok": False, "error": "invalid_json"}, 400)
    rows = body.get("rows") or []
    if not isinstance(rows, list) or not rows:
        return JSONResponse({"ok": False, "error": "no_rows"}, 400)
    created = []
    failed = []
    for r in rows:
        try:
            # 미등록 고객사 경고만 있는 행도 통과 (텍스트로 저장)
            biz_div = (r.get("biz_div") or "").upper()
            if biz_div not in ("T", "M"):
                failed.append({"row_no": r.get("row_no"), "name": r.get("name"),
                               "error": f"biz_div 미상: {biz_div}"})
                continue
            name = (r.get("name") or "").strip()
            if not name:
                failed.append({"row_no": r.get("row_no"), "name": "",
                               "error": "프로젝트명 누락"})
                continue
            unit_price = float(r.get("unit_price") or 0)
            unit_qty = int(r.get("unit_qty") or 1)
            amt = unit_price * unit_qty
            new_pid, new_code = _logi.projects_create_logi({
                "_changed_by": u.get("id"),
                "biz_div": biz_div,
                "project_name": name,
                "customer": r.get("customer_name") or "",
                "model": r.get("model_name") or "",
                "stage": "수주확정" if (r.get("status") in _logi.WON_STATUSES) else "제안작성",
                "po_type": r.get("po_type") or "신규",
                "status": r.get("status") or "초기협의",
                "customer_po": "",
                "currency": (r.get("currency") or "KRW").upper(),
                "is_export": int(r.get("is_export") or 0),
                "order_amount": amt,
                "unit_qty": unit_qty,
                "unit_price": unit_price if unit_price > 0 else None,
                "order_date": r.get("order_date") or "",
                "due_date": r.get("due_date") or "",
                "pm": r.get("pm_name") or "",
                "sales": r.get("sales_name") or "",
                "note": r.get("note") or "",
                "project_type": "NEW_EQUIP",  # 양식이 검사기/자동화 전용
                "parent_project_id": None,
            })
            # 상태가 won 이면 자동 SO 발행 (단일 호기, 사용자 폴백 경로 모방)
            status_v = r.get("status") or ""
            if new_pid and status_v in _logi.WON_STATUSES:
                try:
                    with db_session() as c:
                        exists = c.execute(
                            "SELECT 1 FROM orders WHERE project_id=? LIMIT 1",
                            (new_pid,)
                        ).fetchone()
                        if not exists:
                            per = unit_price if unit_price > 0 else (amt / max(1, unit_qty))
                            units_list = [{
                                "label": _logi.project_unit_label("NEW_EQUIP", i + 1),
                                "amount": per,
                                "due_date": r.get("due_date") or "",
                                "ship_to": r.get("ship_to") or "",
                                "note": "",
                            } for i in range(max(1, unit_qty))]
                            _pwf.confirm_order_multi(
                                c, int(new_pid),
                                units=units_list,
                                order_date=r.get("order_date") or "",
                                created_by=u.get("id") or 0,
                                po_number="",
                            )
                except Exception:
                    pass  # SO 실패는 등록 자체는 성공으로 (편집 화면에서 자가치유)
            created.append({"row_no": r.get("row_no"), "id": new_pid,
                            "mgmt_code": new_code, "name": name})
        except Exception as e:
            failed.append({"row_no": r.get("row_no"), "name": r.get("name"),
                           "error": str(e)})
    return JSONResponse({
        "ok": True,
        "created_count": len(created),
        "failed_count": len(failed),
        "created": created,
        "failed": failed,
    })


# =====================================================
# v5H153 (2026-05-05): 고객사 엑셀 일괄 등록
#   - GET  /customers/import-template  → 양식 .xlsx 다운로드
#   - POST /customers/import-xlsx      → 업로드 + 파싱 + 검증 + 미리보기 JSON
#   - POST /customers/import-confirm   → 확정 INSERT/UPDATE (UPSERT by name)
#  양식: app/static/templates/고객사_일괄등록_양식.xlsx
#  시트 '고객사' (row3=헤더, row4-6=예제, row7+=입력)
#  컬럼: 고객사명/사업자등록번호/대표자명/담당자명/전화번호/이메일/주소/등급/활성/비고
# =====================================================
CUST_IMPORT_HEADERS = [
    "고객사명", "사업자등록번호", "대표자명", "담당자명",
    "전화번호", "이메일", "주소", "등급", "활성", "비고"
]
CUST_IMPORT_TIERS = {"A", "B", "C", "VIP", ""}


def _cust_import_parse_xlsx(file_bytes: bytes) -> list[dict]:
    """고객사 일괄 등록 양식 파싱.
    시트 '고객사' row7+ 데이터, 빈 고객사명 행 자동 스킵.
    예제 row 4-6 도 고객사명이 '예) ...' 또는 자연 스킵 대상."""
    from openpyxl import load_workbook
    import io as _io
    import re as _re
    wb = load_workbook(_io.BytesIO(file_bytes), data_only=True, read_only=True)
    if "고객사" not in wb.sheetnames:
        raise ValueError("'고객사' 시트를 찾을 수 없습니다 (양식 파일 확인)")
    ws = wb["고객사"]

    def _to_str(v) -> str:
        if v is None:
            return ""
        return str(v).strip()

    # 기존 고객사 인덱스 (이름 / 사업자번호 정규화) 캐시
    name_map: dict[str, int] = {}
    bizno_map: dict[str, tuple[int, str]] = {}
    try:
        with db_session() as _cc:
            for r in _cc.execute("SELECT id, name, biz_no FROM customers"):
                if r and r[0] and r[1]:
                    name_map[r[1]] = r[0]
                    bn = _re.sub(r"\D", "", r[2] or "")
                    if bn:
                        bizno_map[bn] = (r[0], r[1])
    except Exception:
        pass

    email_re = _re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
    out = []
    row_no = 0
    # row7+ 데이터 (row1-2=제목/안내, row3=헤더, row4-6=예제)
    for r in ws.iter_rows(min_row=7, values_only=True):
        row_no += 1
        actual_row = row_no + 6
        if not r or all((c is None or str(c).strip() == "") for c in r):
            continue
        cells = list(r) + [None] * (10 - len(r)) if len(r) < 10 else list(r[:10])
        name = _to_str(cells[0])
        if not name:
            continue
        # 예제 행 잔존 방어
        if name.startswith("예)") or name.startswith("예 )"):
            continue
        biz_no_raw = _to_str(cells[1])
        ceo = _to_str(cells[2])
        manager = _to_str(cells[3])
        phone = _to_str(cells[4])
        email = _to_str(cells[5])
        address = _to_str(cells[6])
        tier = _to_str(cells[7]).upper()
        active_raw = _to_str(cells[8])
        note = _to_str(cells[9])

        errors = []
        warnings = []

        # 사업자번호: 대시 제거 후 10자리 숫자 검증 (입력 있을 때만)
        biz_no_norm = _re.sub(r"\D", "", biz_no_raw) if biz_no_raw else ""
        biz_no_display = biz_no_raw
        if biz_no_raw:
            if len(biz_no_norm) != 10 or not biz_no_norm.isdigit():
                errors.append(f"사업자번호 형식 오류(10자리 숫자): '{biz_no_raw}'")
            else:
                # 표준 표기 999-99-99999
                biz_no_display = (
                    f"{biz_no_norm[:3]}-{biz_no_norm[3:5]}-{biz_no_norm[5:]}"
                )

        # 이메일
        if email and not email_re.match(email):
            errors.append(f"이메일 형식 오류: '{email}'")

        # 등급
        if tier not in CUST_IMPORT_TIERS:
            errors.append(f"등급 화이트리스트 위반: '{tier}' (A/B/C/VIP/공란)")
            tier = ""

        # 활성: 1/0/공란 → 공란은 1
        if active_raw == "":
            is_active = 1
        elif active_raw in ("1", "활성", "Y", "y", "True", "true"):
            is_active = 1
        elif active_raw in ("0", "비활성", "N", "n", "False", "false"):
            is_active = 0
        else:
            errors.append(f"활성 값 오류(1/0): '{active_raw}'")
            is_active = 1

        # UPSERT 모드 판정
        existing_id = name_map.get(name)
        action = "update" if existing_id else "create"

        # 사업자번호 중복 (다른 이름) → 경고만
        if biz_no_norm and biz_no_norm in bizno_map:
            other_id, other_name = bizno_map[biz_no_norm]
            if other_name != name:
                warnings.append(
                    f"사업자번호 중복: 기존 '{other_name}' (#{other_id}) 와 동일"
                )

        out.append({
            "row_no": actual_row,
            "name": name,
            "biz_no": biz_no_display,
            "biz_no_norm": biz_no_norm,
            "ceo_name": ceo,
            "manager_name": manager,
            "phone": phone,
            "email": email,
            "address": address,
            "tier": tier,
            "is_active": is_active,
            "note": note,
            "_existing_id": existing_id,
            "_action": action,
            "_errors": errors,
            "_warnings": warnings,
        })
    return out


@app.get("/customers/import-template")
async def customers_import_template(request: Request):
    """고객사 일괄 등록 양식 다운로드 (v5H153)."""
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_sales(u):
        return RedirectResponse("/home", 303)
    from pathlib import Path as _Path
    p = _Path(__file__).parent / "static" / "templates" / "고객사_일괄등록_양식.xlsx"
    if not p.exists():
        return JSONResponse({"error": "양식 파일을 찾을 수 없습니다"}, 404)
    return FileResponse(
        str(p),
        filename="KNK_고객사_일괄등록_양식.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/customers/import-xlsx")
async def customers_import_xlsx(request: Request, xlsx: UploadFile = File(...)):
    """고객사 엑셀 업로드 → 파싱 + 검증 → 미리보기 JSON 반환 (v5H153)."""
    u = get_user(request)
    if not u:
        return JSONResponse({"ok": False, "error": "login_required"}, 401)
    if not can_use_sales(u):
        return JSONResponse({"ok": False, "error": "permission_denied"}, 403)
    try:
        body = await xlsx.read()
        rows = _cust_import_parse_xlsx(body)
    except Exception as e:
        return JSONResponse({"ok": False, "error": f"파싱 실패: {e}"}, 400)
    create_count = sum(1 for r in rows if r["_action"] == "create" and not r["_errors"])
    update_count = sum(1 for r in rows if r["_action"] == "update" and not r["_errors"])
    error_count = sum(1 for r in rows if r["_errors"])
    warning_count = sum(1 for r in rows if r["_warnings"] and not r["_errors"])
    return JSONResponse({
        "ok": True,
        "rows": rows,
        "total": len(rows),
        "create_count": create_count,
        "update_count": update_count,
        "error_count": error_count,
        "warning_count": warning_count,
    })


@app.post("/customers/import-confirm")
async def customers_import_confirm(request: Request):
    """미리보기 확정 → INSERT/UPDATE 실행 (v5H153).
    body JSON: {rows: [...]}.
    UPDATE 시 빈 칸이 아닌 필드만 갱신(기존 데이터 보호)."""
    u = get_user(request)
    if not u:
        return JSONResponse({"ok": False, "error": "login_required"}, 401)
    if not can_use_sales(u):
        return JSONResponse({"ok": False, "error": "permission_denied"}, 403)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"ok": False, "error": "invalid_json"}, 400)
    rows = body.get("rows") or []
    if not isinstance(rows, list) or not rows:
        return JSONResponse({"ok": False, "error": "no_rows"}, 400)

    created, updated, failed = [], [], []
    try:
        from . import customer_tier as _ct
    except Exception:
        _ct = None

    with db_session() as c:
        for r in rows:
            try:
                name = (r.get("name") or "").strip()
                if not name:
                    failed.append({"row_no": r.get("row_no"), "name": "",
                                   "error": "고객사명 누락"})
                    continue
                biz_no = (r.get("biz_no") or "").strip()
                ceo = (r.get("ceo_name") or "").strip()
                manager = (r.get("manager_name") or "").strip()
                phone = (r.get("phone") or "").strip()
                email = (r.get("email") or "").strip()
                address = (r.get("address") or "").strip()
                tier = (r.get("tier") or "").strip().upper()
                is_active = int(r.get("is_active") if r.get("is_active") is not None else 1)
                note = (r.get("note") or "").strip()

                # 이름으로 기존 행 재조회 (preview 후 동시 변경 가능성 방어)
                existing = c.execute(
                    "SELECT id FROM customers WHERE name=?", (name,)
                ).fetchone()
                if existing:
                    cid = existing["id"] if hasattr(existing, "keys") else existing[0]
                    # 빈 칸이 아닌 필드만 UPDATE
                    sets, vals = [], []
                    for col, v in (
                        ("biz_no", biz_no), ("ceo_name", ceo),
                        ("manager_name", manager), ("phone", phone),
                        ("email", email), ("address", address),
                        ("note", note),
                    ):
                        if v:
                            sets.append(f"{col}=?")
                            vals.append(v)
                    # 활성은 명시값 항상 반영
                    sets.append("is_active=?")
                    vals.append(is_active)
                    if sets:
                        vals.append(cid)
                        c.execute(
                            f"UPDATE customers SET {','.join(sets)} WHERE id=?",
                            tuple(vals)
                        )
                    if _ct:
                        try:
                            _ct.refresh_customer_tier(c, cid)
                        except Exception:
                            pass
                    updated.append({"row_no": r.get("row_no"),
                                    "id": cid, "name": name})
                else:
                    fields = {
                        "name": name,
                        # v5H181: '신규' 는 비표준 — 비어있으면 DB DEFAULT '일반' 사용 위해 미포함
                        "tier": tier or "일반",
                        "biz_no": biz_no,
                        "ceo_name": ceo,
                        "manager_name": manager,
                        "phone": phone,
                        "email": email,
                        "address": address,
                        "is_active": is_active,
                        "note": note,
                    }
                    cols = ",".join(fields.keys())
                    ph = ",".join(["?"] * len(fields))
                    c.execute(
                        f"INSERT INTO customers({cols}) VALUES({ph})",
                        tuple(fields.values())
                    )
                    new_id = c.execute("SELECT last_insert_rowid()").fetchone()[0]
                    if _ct:
                        try:
                            _ct.refresh_customer_tier(c, new_id)
                        except Exception:
                            pass
                    created.append({"row_no": r.get("row_no"),
                                    "id": new_id, "name": name})
            except Exception as e:
                failed.append({"row_no": r.get("row_no"),
                               "name": r.get("name"), "error": str(e)})

    return JSONResponse({
        "ok": True,
        "created_count": len(created),
        "updated_count": len(updated),
        "failed_count": len(failed),
        "created": created,
        "updated": updated,
        "failed": failed,
    })


@app.get("/projects/{pid}/edit", response_class=HTMLResponse)
async def projects_edit_form(request: Request, pid: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_sales(u):
        return RedirectResponse("/home", 303)
    # v5H101: 편집 폼 진입 시 SO 기준으로 자가치유
    #   납기/발주일/수주액 = SO 합계·MAX·MIN 으로 정정 후 폼 노출
    try:
        with db_session() as c2:
            sos = _pwf.get_project_orders(c2, pid)
            if sos:
                so_sum = sum(float(o.get("total_amount") or 0) for o in sos)
                dues = [o.get("due_date") for o in sos if o.get("due_date")]
                ords = [o.get("order_date") for o in sos if o.get("order_date")]
                heal_pairs = [("order_amount", so_sum)]
                if dues:
                    heal_pairs.append(("due_date", max(dues)))
                if ords:
                    heal_pairs.append(("order_date", min(ords)))
                for col, val in heal_pairs:
                    try:
                        c2.execute(f"UPDATE projects SET {col}=? WHERE id=?", (val, pid))
                    except Exception:
                        pass
    except Exception:
        pass
    p = _logi.projects_get_logi(pid)
    if not p:
        return RedirectResponse("/projects", status_code=303)
    # v5H174: 등록자/등록일시 표시용 — created_by → 사용자명 lookup
    p = dict(p)
    try:
        if p.get("created_by"):
            with db_session() as _c:
                _u_row = _c.execute(
                    "SELECT name FROM users WHERE id=?", (p.get("created_by"),)
                ).fetchone()
                if _u_row:
                    p["created_by_name"] = _u_row[0]
    except Exception:
        pass
    # v5H103: SO 존재 여부 → 폼 수주액 readonly 안내용
    has_orders = False
    try:
        with db_session() as c3:
            row = c3.execute(
                "SELECT COUNT(*) FROM orders WHERE project_id=?", (pid,)
            ).fetchone()
            has_orders = (row[0] or 0) > 0
    except Exception:
        pass
    return ctx(request, "project_form.html",
               user=u, active="sales_projects",
               project=p,
               STAGES=_logi.STAGES, STATUSES=_logi.LOGI_STATUSES,
               PO_TYPES=_logi.PO_TYPES,
               PROJECT_TYPES=_logi.PROJECT_TYPES,
               PROJECT_TYPE_LABELS=_logi.PROJECT_TYPE_LABELS,
               customers=_logi.customers_for_picker(),
               has_orders=has_orders)


@app.post("/projects/{pid}/edit")
async def projects_edit_submit(request: Request, pid: int):
    """v5H70: 폼 실제 필드명(name/customer_name)과 정합 + 콤마 자동 정리.
    v5H52 의 /projects/new 와 같은 패턴 — edit 도 동일 수정."""
    _u = get_user(request)
    if not _u:
        return RedirectResponse("/login", 303)
    if not can_use_sales(_u):
        return RedirectResponse("/home", 303)
    form = await request.form()
    # v5H192: 잠금 해제 검증 — 폼에 unlock_verified=1 + password 가 함께 와야 수정 허용
    if (form.get("unlock_verified") or "") != "1":
        return RedirectResponse(f"/projects/{pid}/edit?error=password_required", status_code=303)
    _pw = (form.get("password") or "").strip()
    if not _pw:
        return RedirectResponse(f"/projects/{pid}/edit?error=password_required", status_code=303)
    with db_session() as _pwc:
        _row = _pwc.execute("SELECT password FROM users WHERE id=?", (_u.get("id"),)).fetchone()
    if not _row or hash_pw(_pw) != _row["password"]:
        return RedirectResponse(f"/projects/{pid}/edit?error=password_invalid", status_code=303)
    project_name = (form.get("name") or form.get("project_name") or "").strip()
    customer = (form.get("customer_name") or form.get("customer") or "").strip()
    biz_div = (form.get("biz_div") or "").strip()
    # v5H192: 필수 필드 검증 (비고 제외)
    if not project_name:
        return RedirectResponse(f"/projects/{pid}/edit?error=name_required", status_code=303)
    if not customer:
        return RedirectResponse(f"/projects/{pid}/edit?error=customer_required", status_code=303)
    if customer:
        with db_session() as _cc:
            _ok = _cc.execute(
                "SELECT 1 FROM customers WHERE name=? LIMIT 1", (customer,)
            ).fetchone()
        if not _ok:
            from urllib.parse import quote as _q
            return RedirectResponse(
                f"/projects/{pid}/edit?error=customer_not_registered&cust={_q(customer)}",
                status_code=303
            )
    raw_amt = (form.get("order_amount") or "0").strip().replace(",", "")
    try:
        amt = float(raw_amt) if raw_amt else 0
    except ValueError:
        amt = 0
    # v5H132: 단가/수량 받아서 amt 재계산
    raw_price = (form.get("unit_price") or "0").strip().replace(",", "")
    try:
        unit_price = float(raw_price) if raw_price else 0.0
    except ValueError:
        unit_price = 0.0
    raw_qty = (form.get("unit_qty") or "1").strip()
    try:
        unit_qty = int(float(raw_qty))
    except (TypeError, ValueError):
        unit_qty = 1
    if unit_qty < 1:
        unit_qty = 1
    if unit_qty > 100:
        unit_qty = 100
    if unit_price > 0:
        amt = unit_price * unit_qty
    elif amt > 0 and unit_qty >= 1:
        unit_price = amt / unit_qty
    status_val = form.get("status", "초기협의") or "초기협의"
    # v5H137: 프로젝트 유형 + 부모 프로젝트
    _ptype = (form.get("project_type") or "NEW_EQUIP").strip().upper()
    if _ptype not in _logi.PROJECT_TYPES:
        _ptype = "NEW_EQUIP"
    _parent_id_raw = (form.get("parent_project_id") or "").strip()
    _parent_id = int(_parent_id_raw) if _parent_id_raw.isdigit() else None
    if _ptype not in ("CONSUMABLE", "SERVICE"):
        _parent_id = None
    _logi.projects_update_logi(pid, {
        "_changed_by": _u.get("id"),
        "biz_div": biz_div, "project_name": project_name, "customer": customer,
        "model": form.get("model", ""),
        # v5H214: stage 는 status 에서 자동 매핑
        "stage": stage_from_status(status_val),
        "po_type": form.get("po_type", "신규") or "신규",
        "status": status_val,
        "customer_po": form.get("customer_po", ""),
        "currency": (form.get("currency", "KRW") or "KRW").upper(),
        # v5H171: 수정 시에도 fx_rate / amount_krw 저장
        "fx_rate": (lambda v: float(v) if (v or "").strip() else None)(form.get("fx_rate", "")),
        "amount_krw": (lambda v: float(v) if (v or "").strip() else None)(form.get("amount_krw", "")),
        "is_export": form.get("is_export", "0"),
        "order_amount": amt,
        "unit_qty": unit_qty,
        "unit_price": unit_price if unit_price > 0 else None,
        "order_date": form.get("order_date", ""),
        "due_date": form.get("due_date", ""),
        # v5H201: 제안 단계 일정 (수주확정 전 스케줄용)
        "proposal_date": form.get("proposal_date", ""),
        "quotation_date": form.get("quotation_date", ""),
        "pm": form.get("pm", ""), "sales": form.get("sales", ""),
        "note": form.get("note", ""),
        "project_type": _ptype,
        "parent_project_id": _parent_id,
    })
    # v5H87: 수정 후 status 가 won 인데 SO 가 아직 없으면 자동 발행
    # v5H132: 수량 N → N개 호기 라인
    # v5H142: NEW_EQUIP 만 자동 SO 발행
    if status_val in _logi.WON_STATUSES and _ptype == "NEW_EQUIP":
        try:
            with db_session() as c:
                exists = c.execute(
                    "SELECT 1 FROM orders WHERE project_id=? LIMIT 1", (pid,)
                ).fetchone()
                if not exists:
                    _per_unit = unit_price if unit_price > 0 else (amt / max(1, unit_qty))
                    # v5H137: project_type 기준 라벨
                    _units_list = [{
                        "label": _logi.project_unit_label(_ptype, i + 1),
                        "amount": _per_unit,
                        "due_date": form.get("due_date", ""),
                        "ship_to": "",
                        "note": "",
                    } for i in range(max(1, unit_qty))]
                    _pwf.confirm_order_multi(
                        c, int(pid),
                        units=_units_list,
                        order_date=form.get("order_date", ""),
                        created_by=_u.get("id") or 0,
                        po_number=form.get("customer_po", ""),
                    )
        except Exception:
            pass
    return RedirectResponse(f"/project/{pid}", status_code=303)


@app.post("/projects/{pid}/delete")
async def projects_delete_submit(request: Request, pid: int):
    """v5H72: 프로젝트 삭제 — 더 엄격한 can_delete_sales 권한.
    v5H98: 삭제 실패 시 구체적 FK/sqlite 에러 메시지 JSON 으로 반환."""
    _u = get_user(request)
    if not _u:
        return RedirectResponse("/login", 303)
    if not can_delete_sales(_u):
        return JSONResponse({"error": "권한 없음",
                              "message": "프로젝트 삭제는 영업팀 팀장 또는 위임받은 등록권한자만 가능합니다."}, 403)
    try:
        _logi.projects_delete_logi(pid)
    except Exception as e:
        import traceback
        err_msg = str(e)
        tb = traceback.format_exc()
        # 어느 테이블이 막는지 힌트 추출 (FOREIGN KEY constraint failed)
        hint = ""
        if "FOREIGN KEY" in err_msg or "foreign key" in err_msg.lower():
            hint = ("관련 자식 데이터가 남아 있어 삭제 차단됨. "
                    "잠시 후 다시 시도하거나 시스템 관리자에게 보고하세요.")
        return JSONResponse({
            "ok": False,
            "error": "삭제 실패",
            "message": f"{err_msg}\n{hint}",
            "trace": tb[-500:],  # 마지막 500자만
        }, 500)
    return JSONResponse({"ok": True, "message": "프로젝트 삭제 완료"})


@app.post("/projects/bulk-delete")
async def projects_bulk_delete(request: Request):
    """v5H179: 다건 프로젝트 일괄 삭제. ids=콤마구분 또는 ids[]=반복.
    개별 실패는 무시하고 계속 진행 — 결과 집계 반환."""
    _u = get_user(request)
    if not _u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    if not can_delete_sales(_u):
        return JSONResponse({"error": "권한 없음",
                              "message": "프로젝트 삭제는 영업팀 팀장 또는 위임받은 등록권한자만 가능합니다."}, 403)
    form = await request.form()
    raw = form.getlist("ids[]") or form.getlist("ids")
    if not raw and form.get("ids"):
        raw = [s.strip() for s in str(form.get("ids")).split(",") if s.strip()]
    pids: list[int] = []
    for v in raw:
        try:
            pids.append(int(v))
        except (TypeError, ValueError):
            continue
    if not pids:
        return JSONResponse({"ok": False, "message": "삭제할 ID 가 없습니다"}, 400)
    if len(pids) > 200:
        return JSONResponse({"ok": False, "message": "한 번에 200건 초과 불가"}, 400)
    ok, fail = [], []
    for pid in pids:
        try:
            _logi.projects_delete_logi(pid)
            ok.append(pid)
        except Exception as e:
            fail.append({"id": pid, "error": str(e)[:200]})
    return JSONResponse({
        "ok": True,
        "deleted": len(ok),
        "failed": len(fail),
        "fail_details": fail[:20],  # 처음 20건만
        "message": f"{len(ok)}건 삭제 완료" + (f", {len(fail)}건 실패" if fail else ""),
    })


@app.post("/customers/bulk-delete")
async def customers_bulk_delete(request: Request):
    """v5H180: 다건 고객사 일괄 삭제.
    안전 정책:
      - can_delete_sales 권한 필요 (영업팀 팀장/등록권한자)
      - 프로젝트(또는 SO/견적)가 연결된 고객사는 삭제 거부 (FK 안전)
      - 한 번에 200건 제한
      - 개별 실패는 집계 후 계속 진행
    """
    _u = get_user(request)
    if not _u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    if not can_delete_sales(_u):
        return JSONResponse({
            "error": "권한 없음",
            "message": "고객사 삭제는 영업팀 팀장 또는 위임받은 등록권한자만 가능합니다."
        }, 403)
    form = await request.form()
    raw = form.getlist("ids[]") or form.getlist("ids")
    if not raw and form.get("ids"):
        raw = [s.strip() for s in str(form.get("ids")).split(",") if s.strip()]
    cids: list[int] = []
    for v in raw:
        try:
            cids.append(int(v))
        except (TypeError, ValueError):
            continue
    if not cids:
        return JSONResponse({"ok": False, "message": "삭제할 ID 가 없습니다"}, 400)
    if len(cids) > 200:
        return JSONResponse({"ok": False, "message": "한 번에 200건 초과 불가"}, 400)
    ok, fail = [], []
    with db_session() as c:
        for cid in cids:
            try:
                # FK 검증: 프로젝트/SO/견적/PO 연결 여부 확인
                proj_n = (c.execute(
                    "SELECT COUNT(*) FROM projects WHERE customer_id=?", (cid,)
                ).fetchone() or [0])[0]
                ord_n = 0
                try:
                    ord_n = (c.execute(
                        "SELECT COUNT(*) FROM orders WHERE customer_id=?", (cid,)
                    ).fetchone() or [0])[0]
                except Exception:
                    pass
                if proj_n or ord_n:
                    parts = []
                    if proj_n: parts.append(f"프로젝트 {proj_n}건")
                    if ord_n: parts.append(f"수주 {ord_n}건")
                    fail.append({"id": cid,
                                 "error": f"연결 데이터 존재 — {', '.join(parts)}. 먼저 정리 필요"})
                    continue
                # 고객사 담당자 (있으면) → cascade 삭제
                try:
                    c.execute("DELETE FROM customer_contacts WHERE customer_id=?", (cid,))
                except Exception:
                    pass
                c.execute("DELETE FROM customers WHERE id=?", (cid,))
                ok.append(cid)
            except Exception as e:
                fail.append({"id": cid, "error": str(e)[:200]})
    return JSONResponse({
        "ok": True,
        "deleted": len(ok),
        "failed": len(fail),
        "fail_details": fail[:20],
        "message": f"{len(ok)}건 삭제 완료" + (f", {len(fail)}건 실패" if fail else ""),
    })


# =====================================================
# HAIST WORKS — 공급사 (suppliers) 라우트
# =====================================================
@app.get("/suppliers", response_class=HTMLResponse)
async def suppliers_list_page(request: Request, q: str = ""):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_view_logistics(u):
        return RedirectResponse("/home", 303)
    rows = _logi.suppliers_list(q=q)
    return ctx(request, "suppliers.html",
               user=u, active="suppliers",
               suppliers=rows, q=q)


@app.get("/suppliers/new", response_class=HTMLResponse)
async def suppliers_new_form(request: Request):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    return ctx(request, "supplier_form.html",
               user=u, active="suppliers", supplier=None,
               PAYMENT_TERMS=_logi.PAYMENT_TERMS)


@app.post("/suppliers/new")
async def suppliers_new_submit(
    request: Request,
    name: str = Form(...), code: str = Form(""), contact: str = Form(""),
    email: str = Form(""), phone: str = Form(""), country: str = Form(""),
    currency: str = Form("KRW"), payment_terms: str = Form(""),
    note: str = Form(""), is_active: str = Form("1"),
):
    _u = get_user(request)
    if not _u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(_u):
        return RedirectResponse("/home", 303)
    # v5H113: 검증 실패 친절한 에러
    try:
        _logi.supplier_create({
            "name": name, "code": code, "contact": contact, "email": email,
            "phone": phone, "country": country, "currency": currency,
            "payment_terms": payment_terms, "note": note, "is_active": is_active,
        })
    except ValueError as ve:
        from urllib.parse import quote
        return RedirectResponse(f"/suppliers/new?error={quote(str(ve))}", status_code=303)
    return RedirectResponse("/suppliers", status_code=303)


@app.get("/suppliers/{sid}/edit", response_class=HTMLResponse)
async def suppliers_edit_form(request: Request, sid: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    s = _logi.supplier_get(sid)
    if not s:
        return RedirectResponse("/suppliers", 303)
    # 리드타임 통계 자동 계산 (동적 변수 ⑤)
    leadtime = supplier_leadtime_stats(sid)
    return ctx(request, "supplier_form.html",
               user=u, active="suppliers", supplier=s,
               leadtime=leadtime,
               PAYMENT_TERMS=_logi.PAYMENT_TERMS)


@app.post("/suppliers/{sid}/edit")
async def suppliers_edit_submit(
    request: Request, sid: int,
    name: str = Form(...), code: str = Form(""), contact: str = Form(""),
    email: str = Form(""), phone: str = Form(""), country: str = Form(""),
    currency: str = Form("KRW"), payment_terms: str = Form(""),
    note: str = Form(""), is_active: str = Form("1"),
):
    _u = get_user(request)
    if not _u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(_u):
        return RedirectResponse("/home", 303)
    # v5H113: 검증 실패 친절한 에러
    try:
        _logi.supplier_update(sid, {
            "name": name, "code": code, "contact": contact, "email": email,
            "phone": phone, "country": country, "currency": currency,
            "payment_terms": payment_terms, "note": note, "is_active": is_active,
        })
    except ValueError as ve:
        from urllib.parse import quote
        return RedirectResponse(f"/suppliers/{sid}/edit?error={quote(str(ve))}", status_code=303)
    return RedirectResponse("/suppliers", status_code=303)


@app.post("/suppliers/{sid}/delete")
async def suppliers_delete_submit(request: Request, sid: int):
    _u = get_user(request)
    if not _u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(_u):
        return RedirectResponse("/home", 303)
    # v5H112: cascade 안전망 + JSON 에러
    try:
        _logi.supplier_delete(sid)
    except Exception as e:
        return JSONResponse(
            {"ok": False, "error": f"공급사 삭제 실패: {e}"}, status_code=400
        )
    return RedirectResponse("/suppliers", status_code=303)


# =====================================================
# HAIST WORKS — 발주 (purchase_orders) 라우트
# =====================================================
def _parse_po_lines_from_form(form, with_id: bool = False):
    """v5H113: PO 폼 라인 파싱 — 두 가지 패턴 지원
       (A) item_part_id[]/item_qty[]/item_price[]/item_delivery[]/item_note[]/item_id[]
       (B) line_name_N/line_code_N/line_qty_N/line_price_N/line_id_N (현재 po_form.html)
       — A 우선, A 비어있으면 B 폴백. B 는 part_no/part_name 스냅샷으로 전달.
    """
    items = []
    # 패턴 A
    part_ids = form.getlist("item_part_id")
    qtys_a = form.getlist("item_qty")
    prices_a = form.getlist("item_price")
    delivs_a = form.getlist("item_delivery")
    notes_a = form.getlist("item_note")
    ids_a = form.getlist("item_id")
    if part_ids or any((q for q in qtys_a if q)):
        for i, pid in enumerate(part_ids):
            qv = qtys_a[i] if i < len(qtys_a) else ""
            if not pid and not qv:
                continue
            row = {
                "part_id": pid,
                "quantity": qv or "0",
                "unit_price": prices_a[i] if i < len(prices_a) else "0",
                "delivery_date": delivs_a[i] if i < len(delivs_a) else "",
                "note": notes_a[i] if i < len(notes_a) else "",
            }
            if with_id:
                row["id"] = ids_a[i] if i < len(ids_a) else ""
            items.append(row)
        if items:
            return items
    # 패턴 B (인덱스 기반) — 현재 po_form.html
    indices = sorted({
        int(k.split("_")[-1])
        for k in form.keys()
        if (k.startswith("line_name_") or k.startswith("line_qty_")
            or k.startswith("line_price_") or k.startswith("line_code_")
            or k.startswith("line_id_"))
        and k.split("_")[-1].isdigit()
    })
    for idx in indices:
        name = (form.get(f"line_name_{idx}") or "").strip()
        code = (form.get(f"line_code_{idx}") or "").strip()
        qv = (form.get(f"line_qty_{idx}") or "").strip()
        pv = (form.get(f"line_price_{idx}") or "").strip()
        # 빈 라인 스킵
        try:
            q_num = float(qv or 0)
        except Exception:
            q_num = 0
        if not name and not code and q_num <= 0:
            continue
        # v5H114: 폼이 datalist 매칭 시 line_part_id_N 도 함께 전달 (FK 자동매핑)
        part_id_b = (form.get(f"line_part_id_{idx}") or "").strip()
        row = {
            "part_id": part_id_b,
            "part_no_snapshot": code,
            "part_name_snapshot": name,
            "quantity": qv or "0",
            "unit_price": pv or "0",
            "delivery_date": "",
            "note": "",
        }
        if with_id:
            row["id"] = (form.get(f"line_id_{idx}") or "").strip()
        items.append(row)
    return items


@app.get("/po", response_class=HTMLResponse)
async def po_list_page(request: Request, q: str = "", status: str = ""):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_view_logistics(u):
        return RedirectResponse("/home", 303)
    rows = _logi.po_list(q=q, status=status)
    return ctx(request, "po_list.html",
               user=u, active="po",
               orders=rows, q=q, status=status,
               PO_STATUSES=_logi.PO_STATUSES)


@app.get("/po/new", response_class=HTMLResponse)
async def po_new_form(request: Request):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    # 드롭다운용 데이터
    suppliers = _logi.suppliers_list(active_only=True)
    # 관리코드 발급된 프로젝트만 선택 가능
    projects = _logi.projects_list_logi()
    projects_with_code = [p for p in projects if p["mgmt_code"]]
    parts_all = _logi.parts_list()
    return ctx(request, "po_form.html",
               user=u, active="po", po=None, items=[],
               suppliers=suppliers, projects=projects_with_code,
               parts_all=parts_all,
               PO_STATUSES=_logi.PO_STATUSES,
               SHIPPING_TERMS=_logi.SHIPPING_TERMS,
               PAYMENT_TERMS=_logi.PAYMENT_TERMS,
               PO_TYPES_KIND=_logi.PO_TYPES_KIND)


@app.post("/po/new")
async def po_new_submit(request: Request):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    form = await request.form()
    header = {
        "project_id": form.get("project_id", ""),
        "supplier_id": form.get("supplier_id", ""),
        "order_date": form.get("order_date", ""),
        "expected_date": form.get("expected_date", ""),
        "currency": form.get("currency", "KRW"),
        "exchange_rate": form.get("exchange_rate", "1"),
        "status": form.get("status", "작성중"),
        "shipping_terms": form.get("shipping_terms", ""),
        "payment_terms": form.get("payment_terms", ""),
        "po_type": form.get("po_type", "일반"),
        "note": form.get("note", ""),
    }
    # 라인 파싱: item_part_id[], item_qty[], item_price[], item_delivery[], item_note[]
    items = _parse_po_lines_from_form(form)
    # v5H112: 정합성 검증 ValueError → 친절한 redirect
    try:
        po_id, po_num = _logi.po_create(header, items, created_by=u["id"])
    except ValueError as ve:
        from urllib.parse import quote
        return RedirectResponse(f"/po/new?error={quote(str(ve))}", status_code=303)
    except Exception as e:
        return JSONResponse({"ok": False, "error": f"발주 생성 실패: {e}"}, status_code=400)
    return RedirectResponse(f"/po/{po_id}", status_code=303)


@app.get("/po/{po_id}", response_class=HTMLResponse)
async def po_detail(request: Request, po_id: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    header, items = _logi.po_get(po_id)
    if not header:
        return RedirectResponse("/po", 303)
    # v5H117 P4: 자가치유 — header.total_amount vs SUM(items.qty*unit_price) 비교 후 동기화
    po_mismatch = False
    try:
        line_sum = sum(
            float(it.get("qty") or 0) * float(it.get("unit_price") or 0)
            for it in (items or [])
        )
        cur_total = float(header.get("total_amount") or 0) if isinstance(header, dict) else float(getattr(header, "total_amount", 0) or 0)
        # 헤더 부가세 등으로 의도적 차이 가능 — 1원 미만은 무시, 큰 차이만 보정
        if abs(line_sum - cur_total) >= 1.0:
            po_mismatch = True
            try:
                with db_session() as c:
                    c.execute(
                        "UPDATE purchase_orders SET total_amount=? WHERE id=?",
                        (round(line_sum, 2), po_id),
                    )
                if isinstance(header, dict):
                    header["total_amount"] = round(line_sum, 2)
                    header["_self_heal_amount"] = round(cur_total, 2)
            except Exception:
                pass
    except Exception:
        pass
    # v5H136 (2026-05-05): 각 라인에 연결된 프로젝트 (다대다) fetch
    items_with_links = []
    try:
        for it in (items or []):
            it_d = dict(it) if not isinstance(it, dict) else dict(it)
            iid = it_d.get("id")
            if iid:
                try:
                    it_d["links"] = _logi.get_po_item_links(int(iid))
                except Exception:
                    it_d["links"] = []
            else:
                it_d["links"] = []
            items_with_links.append(it_d)
    except Exception:
        items_with_links = items
    # 연결 폼용 프로젝트 목록 (mgmt_code 있는 활성 프로젝트만)
    link_projects = []
    try:
        link_projects = [p for p in _logi.projects_list_logi() if p.get("mgmt_code")]
    except Exception:
        pass
    return ctx(request, "po_detail.html",
               user=u, active="po", po=header, items=items_with_links,
               po_mismatch=po_mismatch, link_projects=link_projects)


@app.get("/po/{po_id}/edit", response_class=HTMLResponse)
async def po_edit_form(request: Request, po_id: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    header, items = _logi.po_get(po_id)
    if not header:
        return RedirectResponse("/po", 303)
    suppliers = _logi.suppliers_list(active_only=True)
    projects = [p for p in _logi.projects_list_logi() if p["mgmt_code"]]
    parts_all = _logi.parts_list()
    return ctx(request, "po_form.html",
               user=u, active="po", po=header, items=items,
               suppliers=suppliers, projects=projects,
               parts_all=parts_all,
               PO_STATUSES=_logi.PO_STATUSES,
               SHIPPING_TERMS=_logi.SHIPPING_TERMS,
               PAYMENT_TERMS=_logi.PAYMENT_TERMS,
               PO_TYPES_KIND=_logi.PO_TYPES_KIND)


@app.post("/po/{po_id}/edit")
async def po_edit_submit(request: Request, po_id: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    form = await request.form()
    header = {
        "project_id": form.get("project_id", ""),
        "supplier_id": form.get("supplier_id", ""),
        "order_date": form.get("order_date", ""),
        "expected_date": form.get("expected_date", ""),
        "currency": form.get("currency", "KRW"),
        "exchange_rate": form.get("exchange_rate", "1"),
        "status": form.get("status", "작성중"),
        "shipping_terms": form.get("shipping_terms", ""),
        "payment_terms": form.get("payment_terms", ""),
        "po_type": form.get("po_type", "일반"),
        "note": form.get("note", ""),
    }
    # v5H112: 라인 id 파싱 (UPSERT 위해) — 신규 라인은 빈 값
    items = _parse_po_lines_from_form(form, with_id=True)
    # v5H112: ValueError → 친절한 에러 (입고이력 보존 거부 포함)
    try:
        _logi.po_update(po_id, header, items)
    except ValueError as ve:
        from urllib.parse import quote
        return RedirectResponse(f"/po/{po_id}/edit?error={quote(str(ve))}", status_code=303)
    except Exception as e:
        return JSONResponse({"ok": False, "error": f"발주 수정 실패: {e}"}, status_code=400)
    return RedirectResponse(f"/po/{po_id}", status_code=303)


@app.post("/po/{po_id}/delete")
async def po_delete_submit(request: Request, po_id: int):
    _u = get_user(request)
    if not _u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(_u):
        return RedirectResponse("/home", 303)
    # v5H112: cascade 안전망 + JSON 에러
    try:
        _logi.po_delete(po_id)
    except Exception as e:
        return JSONResponse(
            {"ok": False, "error": f"발주 삭제 실패: {e}"}, status_code=400
        )
    return RedirectResponse("/po", status_code=303)


# =====================================================
# 입고·출고·수불부 (2026-04-20 물류 본질 완성)
# =====================================================
from .database import (stock_movement_create, po_receive, stock_issue,
                        stock_adjust, stock_movements_list, stock_kpi,
                        part_stock_history, part_fifo_layers, part_price_history,
                        # Top3 S2 3차 (2026-04-26): FIFO 강화 + ABC + 회전율
                        fifo_layers, abc_classification, stock_turnover,
                        MOVEMENT_KINDS, MOVEMENT_KIND_LABEL,
                        gen_movement_no,
                        exchange_rate_create, exchange_rates_list, exchange_rates_latest,
                        get_exchange_rate, CURRENCIES,
                        part_price_create, part_price_approve, part_prices_list,
                        part_active_price, PRICE_TYPES,
                        supplier_leadtime_stats,
                        # 재고 실사·조정 (Top10 #10 — 2026-04-26)
                        stock_audit_create, stock_audits_list, stock_audit_get,
                        stock_audit_item_upsert, stock_adjustments_list,
                        stock_adjustment_approve, stock_adjustment_reject,
                        # 재고실사 2차 (2026-04-26): 첨부 + close + 효과
                        audit_attachment_create, audit_attachments_list,
                        audit_attachment_get, stock_audit_close,
                        stock_audit_effect_summary,
                        # 자재 첨부 (v5H129 — 2026-05-04): 사진/도면
                        part_attachment_create, part_attachments_list,
                        part_attachment_get, part_attachment_delete,
                        # 환율·단가 강화 (Top10 #9 — 2026-04-26)
                        cost_simulation_create, cost_simulations_list,
                        price_change_log, price_change_history_list,
                        rate_alert_create, rate_alerts_list,
                        exchange_rates_csv_upload,
                        # 발견 3건 통합 (2026-04-26)
                        check_rate_alerts,
                        # 사이클 54 환율·단가 1차 (2026-04-27)
                        get_latest_fx_rate, convert_amount,
                        get_latest_part_price)


# =====================================================
# 환율 관리 (동적 변수 ③)
# =====================================================
@app.get("/rates", response_class=HTMLResponse)
async def rates_page(request: Request, currency: str = ""):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    items = exchange_rates_list(limit=200, currency=currency)
    latest = exchange_rates_latest()
    return ctx(request, "rates.html", user=u, items=items, latest=latest,
               currency=currency, CURRENCIES=CURRENCIES, active="rates")


@app.post("/rates/new")
async def rates_create_submit(
    request: Request,
    rate_date: str = Form(...),
    from_currency: str = Form(...),
    to_currency: str = Form("KRW"),
    rate: str = Form(...),
    source: str = Form("수동"),
    note: str = Form(""),
):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    try:
        exchange_rate_create({
            "rate_date": rate_date,
            "from_currency": from_currency,
            "to_currency": to_currency,
            "rate": float(rate),
            "source": source,
            "note": note,
        }, user_id=u["id"])
        # S3-1 옵션 A: 수동 등록 시점에도 자동 알림 발동 검사
        check_rate_alerts(from_currency, float(rate))
    except Exception as e:
        from urllib.parse import quote
        return RedirectResponse(f"/rates?error={quote(str(e))}", 303)
    return RedirectResponse("/rates?success=1", 303)


# =====================================================
# 적용일자 단가 (동적 변수 ②)
# =====================================================
@app.post("/parts/{pid}/prices/new")
async def parts_price_create_submit(
    request: Request, pid: int,
    supplier_id: str = Form(""),
    price_type: str = Form("견적"),
    unit_price: str = Form(...),
    currency: str = Form("KRW"),
    effective_from: str = Form(...),
    effective_to: str = Form(""),
    negotiated_at: str = Form(""),
    min_qty: str = Form("0"),
    max_qty: str = Form(""),
    note: str = Form(""),
):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    sid = int(supplier_id) if supplier_id.isdigit() else None
    new_price = float(unit_price)
    # S4-1 옵션 A: 직전 활성 단가 조회 (애플리케이션 레벨 훅)
    prev = part_active_price(pid, supplier_id=sid or 0) or {}
    old_price = prev.get("unit_price")
    try:
        part_price_create({
            "part_id": pid,
            "supplier_id": sid,
            "price_type": price_type,
            "unit_price": new_price,
            "currency": currency,
            "effective_from": effective_from,
            "effective_to": effective_to or None,
            "negotiated_at": negotiated_at or None,
            "min_qty": float(min_qty or 0),
            "max_qty": float(max_qty) if max_qty else None,
            "note": note,
        }, user_id=u["id"])
        # S4-1 옵션 A: price_change_history 자동 INSERT (변동률 자동 계산)
        try:
            price_change_log(pid, sid, old_price, new_price,
                             effective_from, u["id"], note=note or "")
        except Exception:
            pass  # 본 등록은 성공했으므로 훅 실패는 흡수
    except Exception as e:
        from urllib.parse import quote
        return RedirectResponse(f"/parts/{pid}?error={quote(str(e))}", 303)
    return RedirectResponse(f"/parts/{pid}?price_added=1", 303)


@app.post("/parts/prices/{price_id}/approve")
async def parts_price_approve_submit(request: Request, price_id: int):
    u = get_user(request)
    if not u or u["role"] not in ("admin", "ceo", "leader", "executive"):
        return JSONResponse({"error": "권한 없음"}, 401)
    with db_session() as c:
        row = c.execute("SELECT part_id FROM part_prices WHERE id=?", (price_id,)).fetchone()
    part_price_approve(price_id, user_id=u["id"])
    pid = row["part_id"] if row else 0
    return RedirectResponse(f"/parts/{pid}?approved=1", 303)


@app.get("/parts/{pid:int}", response_class=HTMLResponse)
async def parts_detail_page(request: Request, pid: int):
    """부품 상세 — FIFO 레이어, 공급사 단가 이력, 적용일자 단가, 입출고 이력 통합"""
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    part = _logi.parts_get(pid)
    if not part:
        return RedirectResponse("/parts", 303)
    part = dict(part)
    # v5H124 MED: stock_movements 합계 vs parts.stock_qty 정합성 자가치유
    # 직접 SQL UPDATE 등으로 어긋난 경우 페이지 진입 시 자동 보정 (v5H94 패턴 확장)
    stock_self_heal = None
    try:
        with db_session() as _c:
            _sm_total = _c.execute(
                "SELECT COALESCE(SUM(quantity),0) FROM stock_movements WHERE part_id=?",
                (pid,),
            ).fetchone()[0] or 0
            _cur_qty = float(part.get("stock_qty") or 0)
            if abs(float(_sm_total) - _cur_qty) > 0.0001:
                _c.execute(
                    "UPDATE parts SET stock_qty=?, updated_at=? WHERE id=?",
                    (float(_sm_total), datetime.now().isoformat(timespec="seconds"), pid),
                )
                stock_self_heal = {
                    "before": _cur_qty, "after": float(_sm_total),
                    "delta": float(_sm_total) - _cur_qty,
                }
                part["stock_qty"] = float(_sm_total)
    except Exception:
        pass
    layers = part_fifo_layers(pid)
    price_hist = part_price_history(pid, limit=30)
    recent_moves = part_stock_history(pid, limit=30)
    stock_value = sum((l.get("remaining_qty") or 0) * (l.get("unit_price") or 0) for l in layers)
    # 적용일자 단가 (동적 변수 ②)
    managed_prices = part_prices_list(pid)
    active_price = part_active_price(pid)
    # 공급사 선택지 (단가 등록 폼용)
    with db_session() as c:
        suppliers = [dict(r) for r in c.execute(
            "SELECT id, name FROM suppliers WHERE is_active=1 ORDER BY name"
        ).fetchall()]
    # v5H129 (2026-05-04): 자재 첨부(사진/도면) 목록
    try:
        attachments = part_attachments_list(pid)
    except Exception:
        attachments = []
    # v5H136 (2026-05-05): 자재가 어떤 프로젝트(장비)에서 얼마나 쓰였는지 (소모품 식별)
    project_usage = []
    try:
        project_usage = _logi.get_part_project_usage(pid, limit=50)
    except Exception:
        pass
    return ctx(request, "part_detail.html", user=u,
               part=part, layers=layers,
               price_history=price_hist["history"],
               by_supplier=price_hist["by_supplier"],
               managed_prices=managed_prices,
               active_price=active_price,
               recent_moves=recent_moves,
               stock_value=stock_value,
               stock_self_heal=stock_self_heal,
               suppliers=suppliers,
               attachments=attachments,
               project_usage=project_usage,
               CURRENCIES=CURRENCIES, PRICE_TYPES=PRICE_TYPES,
               active="parts")


@app.get("/po/{po_id}/receive", response_class=HTMLResponse)
async def po_receive_form(request: Request, po_id: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    header, items = _logi.po_get(po_id)
    if not header:
        return RedirectResponse("/po", 303)
    po_ctx = dict(header)
    po_ctx["items"] = [dict(r) for r in items]
    return ctx(request, "po_receive.html", user=u, po=po_ctx, active="po")


@app.post("/po/{po_id}/receive")
async def po_receive_submit(request: Request, po_id: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    form = await request.form()
    occurred = form.get("occurred_at", "") or ""
    item_ids = form.getlist("po_item_id")
    qtys = form.getlist("receive_qty")
    notes = form.getlist("item_note")
    lots = form.getlist("lot_no")
    expiries = form.getlist("expiry_date")
    lines = []
    for i, iid in enumerate(item_ids):
        try:
            q = float(qtys[i]) if i < len(qtys) and qtys[i] else 0
        except ValueError:
            q = 0
        if q > 0:
            lines.append({
                "po_item_id": int(iid),
                "receive_qty": q,
                "note": notes[i] if i < len(notes) else "",
                "lot_no": lots[i] if i < len(lots) else "",
                "expiry_date": expiries[i] if i < len(expiries) else "",
            })
    if not lines:
        return RedirectResponse(f"/po/{po_id}/receive?error=empty", 303)
    result = po_receive(po_id, lines, u["id"], occurred_at=occurred)
    if not result.get("ok"):
        return RedirectResponse(f"/po/{po_id}/receive?error=1", 303)
    return RedirectResponse(f"/po/{po_id}?received={result['count']}", 303)


@app.get("/stock/issue", response_class=HTMLResponse)
async def stock_issue_form(request: Request, part_id: str = ""):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_view_logistics(u):
        return RedirectResponse("/home", 303)
    with db_session() as c:
        parts = c.execute(
            """SELECT id, part_no, part_name, unit, stock_qty, std_price, safety_stock
               FROM parts WHERE is_active=1 ORDER BY part_name"""
        ).fetchall()
        projects = c.execute(
            """SELECT id, mgmt_code, name FROM projects
               WHERE status IN ('active','진행중','planning','수주확정','납품')
               ORDER BY mgmt_code DESC LIMIT 200"""
        ).fetchall()
        customers = c.execute(
            "SELECT id, name FROM customers ORDER BY tier DESC, name"
        ).fetchall()
    return ctx(request, "stock_issue.html", user=u,
               parts=[dict(r) for r in parts],
               projects=[dict(r) for r in projects],
               customers=[dict(r) for r in customers],
               default_part_id=part_id, active="stock_issue")


@app.post("/stock/issue")
async def stock_issue_submit(
    request: Request,
    part_id: str = Form(...),
    quantity: str = Form(...),
    project_id: str = Form(""),
    customer_id: str = Form(""),
    unit_price: str = Form("0"),
    reason: str = Form("현장 출고"),
    location: str = Form(""),
    occurred_at: str = Form(""),
    note: str = Form(""),
):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_view_logistics(u):
        return RedirectResponse("/home", 303)
    try:
        pid = int(part_id)
        qty = float(quantity)
    except ValueError:
        return RedirectResponse("/stock/issue?error=invalid", 303)
    if qty <= 0:
        return RedirectResponse("/stock/issue?error=qty", 303)
    try:
        mid, mno = stock_issue({
            "part_id": pid,
            "quantity": qty,
            "project_id": int(project_id) if project_id.isdigit() else None,
            "customer_id": int(customer_id) if customer_id.isdigit() else None,
            "unit_price": float(unit_price or 0),
            "reason": reason,
            "location": location,
            "occurred_at": occurred_at or None,
            "note": note,
        }, u["id"])
    except ValueError as e:
        from urllib.parse import quote
        return RedirectResponse(f"/stock/issue?error={quote(str(e))}", 303)
    return RedirectResponse(f"/stock/movements?success={mno}", 303)


@app.get("/stock/adjust", response_class=HTMLResponse)
async def stock_adjust_form(request: Request, part_id: str = ""):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_view_logistics(u):
        return RedirectResponse("/home", 303)
    with db_session() as c:
        parts = c.execute(
            """SELECT id, part_no, part_name, unit, stock_qty
               FROM parts WHERE is_active=1 ORDER BY part_name"""
        ).fetchall()
    return ctx(request, "stock_adjust.html", user=u,
               parts=[dict(r) for r in parts],
               default_part_id=part_id, active="stock_adjust")


@app.post("/stock/adjust")
async def stock_adjust_submit(
    request: Request,
    part_id: str = Form(...),
    quantity: str = Form(...),
    reason: str = Form(...),
    note: str = Form(""),
):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_view_logistics(u):
        return RedirectResponse("/home", 303)
    try:
        pid = int(part_id)
        qty = float(quantity)
    except ValueError:
        return RedirectResponse("/stock/adjust?error=invalid", 303)
    # v5H113 LOW#17: 친절한 에러
    try:
        stock_adjust({
            "part_id": pid,
            "quantity": qty,
            "reason": reason,
            "note": note,
        }, u["id"])
    except ValueError as ve:
        from urllib.parse import quote
        return RedirectResponse(f"/stock/adjust?error={quote(str(ve))}", 303)
    return RedirectResponse("/stock/movements?success=adjust", 303)


# =====================================================
# 재고 실사·조정 워크플로 (Top10 #10 — 2026-04-26 P4 자재팀 분기 1회)
# - 자재팀(can_use_logistics) + admin/ceo: 실사 진행 가능
# - 조정 승인: admin/ceo/executive 또는 team_id==10(구매팀) leader (자재팀장)
# =====================================================
def _audit_guard(u) -> bool:
    """실사 화면 접근 권한 — 물류 권한자."""
    return bool(u) and can_use_logistics(u)


def _audit_approve_guard(u) -> bool:
    """조정 승인 권한 — admin/ceo/executive 또는 자재팀장(team_id=10 leader)."""
    if not u:
        return False
    role = u.get("role") if isinstance(u, dict) else u["role"]
    if role in ("admin", "ceo", "executive"):
        return True
    team_id = u.get("team_id") if isinstance(u, dict) else u["team_id"]
    if team_id == 10 and role == "leader":
        return True
    return False


@app.get("/stock/audits", response_class=HTMLResponse)
async def stock_audits_page(request: Request):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not _audit_guard(u):
        return RedirectResponse("/home", 303)
    audits = stock_audits_list(limit=50)
    return ctx(request, "stock_audits.html", user=u, audits=audits,
               can_approve=_audit_approve_guard(u), active="stock")


@app.post("/stock/audits/new")
async def stock_audits_new(request: Request, note: str = Form("")):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not _audit_guard(u):
        return RedirectResponse("/home", 303)
    aid, ano = stock_audit_create(led_by=u["id"], note=note)
    return RedirectResponse(f"/stock/audits/{aid}?success={ano}", 303)


@app.get("/stock/audits/{audit_id:int}", response_class=HTMLResponse)
async def stock_audit_detail(request: Request, audit_id: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not _audit_guard(u):
        return RedirectResponse("/home", 303)
    audit = stock_audit_get(audit_id)
    if not audit:
        return RedirectResponse("/stock/audits", 303)
    with db_session() as c:
        parts = c.execute(
            """SELECT id, part_no, part_name, unit, stock_qty
               FROM parts WHERE is_active=1 ORDER BY part_name LIMIT 500"""
        ).fetchall()
        # close 가능 조건 사전 계산
        pending_lines = c.execute(
            "SELECT COUNT(*) FROM stock_audit_items WHERE audit_id=? AND status='PENDING'",
            (audit_id,),
        ).fetchone()[0]
        pending_adjs = c.execute(
            """SELECT COUNT(*) FROM stock_adjustments adj
               JOIN stock_audit_items ai ON adj.audit_item_id=ai.id
               WHERE ai.audit_id=? AND adj.status='PENDING'""",
            (audit_id,),
        ).fetchone()[0]
    can_close = (audit["status"] != "CLOSED" and pending_lines == 0
                 and pending_adjs == 0 and (audit["items"] or []))
    effect = stock_audit_effect_summary(audit_id)
    return ctx(request, "stock_audit.html", user=u, mode="detail",
               audits=[], audit=audit, parts=[dict(r) for r in parts],
               can_approve=_audit_approve_guard(u),
               can_close=bool(can_close), pending_lines=pending_lines,
               pending_adjs=pending_adjs, effect=effect, active="stock")


@app.post("/stock/audits/{audit_id}/items")
async def stock_audit_item_add(
    request: Request, audit_id: int,
    part_id: str = Form(...),
    counted_qty: str = Form(...),
    variance_reason: str = Form(""),
):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not _audit_guard(u):
        return RedirectResponse("/home", 303)
    try:
        pid = int(part_id)
        cq = float(counted_qty)
    except ValueError:
        return RedirectResponse(f"/stock/audits/{audit_id}?error=invalid", 303)
    stock_audit_item_upsert(audit_id, pid, cq, variance_reason.strip(), u["id"])
    return RedirectResponse(f"/stock/audits/{audit_id}?success=line", 303)


@app.get("/stock/adjustments", response_class=HTMLResponse)
async def stock_adjustments_page(request: Request, status: str = "PENDING"):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not _audit_guard(u):
        return RedirectResponse("/home", 303)
    items = stock_adjustments_list(status=status, limit=200)
    # 첨부 카운트 (라인 옆 표시용)
    att_counts = {}
    if items:
        with db_session() as c:
            ids = [int(x["id"]) for x in items]
            qmarks = ",".join("?" * len(ids))
            rows = c.execute(
                f"SELECT adjustment_id, COUNT(*) AS cnt FROM audit_attachments "
                f"WHERE adjustment_id IN ({qmarks}) GROUP BY adjustment_id",
                ids,
            ).fetchall()
            att_counts = {r["adjustment_id"]: r["cnt"] for r in rows}
    return ctx(request, "stock_adjustment.html", user=u,
               adjustments=items, filter_status=status, att_counts=att_counts,
               can_approve=_audit_approve_guard(u), active="stock")


@app.post("/stock/adjustments/{adj_id}/approve")
async def stock_adjustments_approve(request: Request, adj_id: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not _audit_approve_guard(u):
        return RedirectResponse("/stock/adjustments?error=denied", 303)
    try:
        _mid, mno = stock_adjustment_approve(adj_id, u["id"])
    except ValueError as e:
        return RedirectResponse(f"/stock/adjustments?error={e}", 303)
    return RedirectResponse(f"/stock/adjustments?success={mno}", 303)


@app.post("/stock/adjustments/{adj_id}/reject")
async def stock_adjustments_reject(request: Request, adj_id: int,
                                   note: str = Form("")):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not _audit_approve_guard(u):
        return RedirectResponse("/stock/adjustments?error=denied", 303)
    stock_adjustment_reject(adj_id, u["id"], note=note.strip())
    return RedirectResponse("/stock/adjustments?success=rejected", 303)


# =====================================================
# 재고실사 2차 (2026-04-26): 증명서 첨부 + close 워크플로
# 외부 파일 저장소 0건 — 로컬 디스크 ./uploads/audits/<adj_id>/<file>
# =====================================================
_AUDIT_UPLOAD_ROOT = os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads", "audits")
_AUDIT_ALLOWED_EXT = {".pdf", ".jpg", ".jpeg", ".png", ".xlsx"}
_AUDIT_MAX_BYTES = 10 * 1024 * 1024  # 10MB


def _audit_safe_filename(name: str) -> str:
    """path traversal 방지 — basename만 + 영숫자/._- 외 _로 치환."""
    base = os.path.basename((name or "").replace("\\", "/"))
    base = base.lstrip(".") or "file"
    out = []
    for ch in base:
        if ch.isalnum() or ch in "._-":
            out.append(ch)
        else:
            out.append("_")
    safe = "".join(out)[:120]
    return safe or "file"


@app.post("/stock/adjustments/{adj_id}/attach")
async def stock_adjustment_attach(request: Request, adj_id: int,
                                  file: UploadFile = File(...)):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not _audit_guard(u):
        return RedirectResponse("/home", 303)
    raw = await file.read()
    if len(raw) > _AUDIT_MAX_BYTES:
        return RedirectResponse(f"/stock/adjustments?error=size_over_10MB", 303)
    if len(raw) == 0:
        return RedirectResponse("/stock/adjustments?error=empty_file", 303)
    safe_name = _audit_safe_filename(file.filename or "upload")
    ext = os.path.splitext(safe_name)[1].lower()
    if ext not in _AUDIT_ALLOWED_EXT:
        return RedirectResponse("/stock/adjustments?error=ext_not_allowed", 303)
    target_dir = os.path.join(_AUDIT_UPLOAD_ROOT, str(int(adj_id)))
    os.makedirs(target_dir, exist_ok=True)
    # 동명파일 충돌 회피 — 타임스탬프 prefix
    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    final_name = f"{ts}_{safe_name}"
    final_path = os.path.join(target_dir, final_name)
    # path traversal 2차 검증 — 정규화 후 root 안에 있는지
    abs_root = os.path.abspath(_AUDIT_UPLOAD_ROOT)
    abs_final = os.path.abspath(final_path)
    if not abs_final.startswith(abs_root + os.sep):
        return RedirectResponse("/stock/adjustments?error=path_invalid", 303)
    with open(final_path, "wb") as f:
        f.write(raw)
    audit_attachment_create(adj_id, abs_final, safe_name, u["id"])
    return RedirectResponse(f"/stock/adjustments/{adj_id}/attachments?success=uploaded", 303)


@app.get("/stock/adjustments/{adj_id}/attachments", response_class=HTMLResponse)
async def stock_adjustment_attachments_page(request: Request, adj_id: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not _audit_guard(u):
        return RedirectResponse("/home", 303)
    atts = audit_attachments_list(adj_id)
    return ctx(request, "stock_adjustment.html", user=u,
               adjustments=[], filter_status="ATTACH", att_counts={},
               attach_view=True, attach_adj_id=adj_id, attachments=atts,
               can_approve=_audit_approve_guard(u), active="stock")


@app.get("/stock/adjustments/{adj_id}/attachments/{att_id}/download")
async def stock_adjustment_attachment_download(request: Request, adj_id: int, att_id: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not _audit_guard(u):
        return RedirectResponse("/home", 303)
    rec = audit_attachment_get(att_id)
    if not rec or int(rec["adjustment_id"]) != int(adj_id):
        return RedirectResponse(f"/stock/adjustments/{adj_id}/attachments?error=not_found", 303)
    fp = rec["file_path"]
    abs_root = os.path.abspath(_AUDIT_UPLOAD_ROOT)
    abs_fp = os.path.abspath(fp)
    if not abs_fp.startswith(abs_root + os.sep) or not os.path.exists(abs_fp):
        return RedirectResponse(f"/stock/adjustments/{adj_id}/attachments?error=file_missing", 303)
    return FileResponse(abs_fp, filename=rec.get("file_name") or os.path.basename(abs_fp))


@app.post("/stock/audits/{audit_id}/close")
async def stock_audits_close_route(request: Request, audit_id: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not _audit_approve_guard(u):
        return RedirectResponse(f"/stock/audits/{audit_id}?error=denied", 303)
    ok, msg = stock_audit_close(audit_id)
    if not ok:
        return RedirectResponse(f"/stock/audits/{audit_id}?error={msg}", 303)
    return RedirectResponse(f"/stock/audits/{audit_id}?success=closed", 303)
# ===== /재고실사 2차 =====


# =====================================================
# 자재 첨부 (v5H129 — 2026-05-04): 사진/도면 업로드
# 클라이언트 측 Canvas API 로 이미지 압축 후 업로드 → 용량 90%+ 절감
# 외부 저장소 0건 — 로컬 ./uploads/parts/<part_id>/<file>
# =====================================================
_PARTS_UPLOAD_ROOT = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                                  "uploads", "parts")
_PARTS_ALLOWED_EXT = {".jpg", ".jpeg", ".png", ".webp", ".pdf", ".dwg", ".dxf", ".gif"}
_PARTS_MAX_BYTES = 20 * 1024 * 1024  # 20MB (원본 — 클라이언트에서 더 줄여 보내라고 안내)
_PARTS_KIND_WHITELIST = {"photo", "drawing", "spec"}


def _parts_safe_filename(name: str) -> str:
    """path traversal 방지 — basename + 영숫자/한글/._- 허용 + 그 외 _."""
    base = os.path.basename((name or "").replace("\\", "/"))
    base = base.lstrip(".") or "file"
    out = []
    for ch in base:
        # 영숫자, 한글(가-힣, ㄱ-ㅎ, ㅏ-ㅣ), . _ - 허용
        if ch.isalnum() or ch in "._-" or ('가' <= ch <= '힣'):
            out.append(ch)
        else:
            out.append("_")
    safe = "".join(out)[:120]
    return safe or "file"


@app.post("/parts/{pid}/attach")
async def parts_attach_upload(request: Request, pid: int,
                              file: UploadFile = File(...),
                              kind: str = Form("photo")):
    """v5H129: 자재 사진/도면 첨부 업로드. 클라이언트가 이미 압축한 파일을 받음."""
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    # 자재 존재 확인
    p = _logi.parts_get(pid)
    if not p:
        return RedirectResponse("/parts?error=part_not_found", 303)
    raw = await file.read()
    if len(raw) == 0:
        return RedirectResponse(f"/parts/{pid}/edit?error=빈+파일입니다", 303)
    if len(raw) > _PARTS_MAX_BYTES:
        return RedirectResponse(
            f"/parts/{pid}/edit?error=파일이+너무+큽니다+%2820MB+초과%29.+사진은+자동압축되며+도면은+10MB+이하+권장",
            303)
    safe_name = _parts_safe_filename(file.filename or "upload")
    ext = os.path.splitext(safe_name)[1].lower()
    if ext not in _PARTS_ALLOWED_EXT:
        return RedirectResponse(
            f"/parts/{pid}/edit?error=허용되지+않은+확장자입니다+%28jpg%2Fpng%2Fwebp%2Fpdf%2Fdwg%2Fdxf%29",
            303)
    kind_v = (kind or "photo").strip().lower()
    if kind_v not in _PARTS_KIND_WHITELIST:
        kind_v = "photo"
    target_dir = os.path.join(_PARTS_UPLOAD_ROOT, str(int(pid)))
    os.makedirs(target_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    final_name = f"{ts}_{safe_name}"
    final_path = os.path.join(target_dir, final_name)
    # path traversal 2차 검증
    abs_root = os.path.abspath(_PARTS_UPLOAD_ROOT)
    abs_final = os.path.abspath(final_path)
    if not abs_final.startswith(abs_root + os.sep):
        return RedirectResponse(f"/parts/{pid}/edit?error=잘못된+경로", 303)
    with open(final_path, "wb") as fp:
        fp.write(raw)
    mime = (file.content_type or "").split(";")[0].strip() or "application/octet-stream"
    part_attachment_create(pid, abs_final, safe_name, len(raw), mime, kind_v, u["id"])
    return RedirectResponse(f"/parts/{pid}/edit?success=첨부+완료", 303)


@app.get("/parts/{pid}/attachments/{aid}")
async def parts_attachment_download(request: Request, pid: int, aid: int):
    """v5H129: 자재 첨부 다운로드 (이미지는 inline 표시 가능)."""
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    rec = part_attachment_get(aid)
    if not rec or int(rec["part_id"]) != int(pid):
        return RedirectResponse(f"/parts/{pid}/edit?error=첨부+없음", 303)
    fp = rec["file_path"]
    abs_root = os.path.abspath(_PARTS_UPLOAD_ROOT)
    abs_fp = os.path.abspath(fp)
    if not abs_fp.startswith(abs_root + os.sep) or not os.path.exists(abs_fp):
        return RedirectResponse(f"/parts/{pid}/edit?error=파일+소실", 303)
    return FileResponse(
        abs_fp,
        filename=rec.get("file_name") or os.path.basename(abs_fp),
        media_type=rec.get("mime_type") or "application/octet-stream",
    )


@app.post("/parts/{pid}/attachments/{aid}/delete")
async def parts_attachment_delete(request: Request, pid: int, aid: int):
    """v5H129: 자재 첨부 삭제 (DB + 디스크)."""
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    rec = part_attachment_get(aid)
    if not rec or int(rec["part_id"]) != int(pid):
        return RedirectResponse(f"/parts/{pid}/edit?error=첨부+없음", 303)
    # 디스크 파일 삭제 (path traversal 검증)
    fp = rec["file_path"]
    abs_root = os.path.abspath(_PARTS_UPLOAD_ROOT)
    abs_fp = os.path.abspath(fp)
    if abs_fp.startswith(abs_root + os.sep) and os.path.exists(abs_fp):
        try:
            os.remove(abs_fp)
        except OSError:
            pass
    part_attachment_delete(aid)
    return RedirectResponse(f"/parts/{pid}/edit?success=첨부+삭제됨", 303)
# ===== /자재 첨부 =====


@app.get("/stock/movements", response_class=HTMLResponse)
async def stock_movements_page(
    request: Request,
    part_id: str = "",
    kind: str = "",
    since: str = "",
    until: str = "",
    po_id: str = "",
    project_id: str = "",
    q: str = "",
):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_view_logistics(u):
        return RedirectResponse("/home", 303)
    items = stock_movements_list(
        part_id=int(part_id) if part_id.isdigit() else 0,
        kind=kind,
        since=since, until=until,
        po_id=int(po_id) if po_id.isdigit() else 0,
        project_id=int(project_id) if project_id.isdigit() else 0,
        q=q,
        limit=300,
    )
    kpi = stock_kpi()
    part_filter = None
    if part_id.isdigit():
        part_filter = _logi.parts_get(int(part_id))
    return ctx(request, "stock_movements.html", user=u, items=items, kpi=kpi,
               part_filter=part_filter, filter_kind=kind, filter_since=since,
               filter_until=until, q=q, MOVEMENT_KIND_LABEL=MOVEMENT_KIND_LABEL,
               active="stock")


# =====================================================
# HAIST Victor — 사내 AI 컨시어지 (Phase 1)
# 자연어 질문 → 데이터/페이지 자동 라우팅
# =====================================================
from .victor import ask as victor_ask


@app.post("/api/victor/ask")
async def api_victor_ask(req: Request):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    try:
        data = await req.json()
        query = (data.get("query") or "").strip()
    except Exception:
        query = ""
    result = victor_ask(query, u, db_session)
    return JSONResponse({"ok": True, "result": result})


@app.get("/api/victor/ask")
async def api_victor_ask_get(req: Request, q: str = ""):
    """GET 방식 지원 (간단한 테스트/디버깅용)"""
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    result = victor_ask(q, u, db_session)
    return JSONResponse({"ok": True, "result": result})


# =====================================================
# TOP3 S3 — 권한 위임 1차 라우트 골격 (2026-04-25)
# 시안: 05_HAIST_WORKS_디자인팀/_TO_01_정식시안_Top3_S3_v1.md
# 1차 = 골격만 (UI 본문·토큰 발행 로직은 다음 사이클).
# 권한 분기: CEO·admin only — 평직원/팀장 차단.
# =====================================================
@app.get("/admin/permissions", response_class=HTMLResponse)
async def admin_permissions_page(req: Request):
    """권한 위임 메인 — 보내기/회수 2탭 (시안 §0)"""
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    return ctx(req, "admin_permissions.html", user=u, active="admin")


@app.get("/admin/permissions/grant", response_class=HTMLResponse)
async def admin_permissions_grant_page(req: Request):
    """보내기 탭 — 위임 발송 폼 (시안 §1-A)"""
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    return ctx(req, "admin_permissions_grant.html", user=u, active="admin")


@app.post("/admin/permissions/grant")
async def admin_permissions_grant_submit(req: Request):
    """위임 토큰 발행 (S3 2차 본문 · 시안 §1-A 5필드)
    트랜잭션: delegation_tokens INSERT + delegation_audit INSERT (audit 누락 0건)
    RBAC 컬럼 분리: resource + action 셀렉터 → permissions 조회 (없으면 자동 INSERT)
    """
    u = require(req, ["admin", "ceo"])
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    to_user_q  = (form.get("to_user") or "").strip()
    resource   = (form.get("resource") or "").strip()
    action     = (form.get("action") or "").strip()
    expires_at = (form.get("expires_at") or "").strip()
    reason     = (form.get("reason") or "").strip()
    can_redel  = 1 if form.get("can_redelegate") else 0
    if not (to_user_q and resource and action and expires_at):
        return JSONResponse({"error": "필수 항목 누락"}, 400)
    with db_session() as c:
        # 위임받는 자 조회 (이름 또는 이메일)
        tu = c.execute(
            "SELECT id, name FROM users WHERE name=? OR login_id=? LIMIT 1",
            (to_user_q, to_user_q)
        ).fetchone()
        if not tu:
            return JSONResponse({"error": "대상 사용자 없음"}, 404)
        # 권한 카탈로그 조회/INSERT (resource/action/scope 신규 컬럼 사용 — RBAC 분리)
        prow = c.execute(
            "SELECT id FROM permissions WHERE resource=? AND action=? LIMIT 1",
            (resource, action)
        ).fetchone()
        if prow:
            perm_id = prow["id"]
        else:
            # 신규 권한 자동 등록 (name = 'resource.action' 호환 유지)
            c.execute(
                "INSERT INTO permissions(name, resource, action, scope, description) VALUES(?,?,?,?,?)",
                (f"{resource}.{action}", resource, action, resource, f"{resource} {action}")
            )
            perm_id = c.execute("SELECT last_insert_rowid()").fetchone()[0]
        # delegation_tokens INSERT
        c.execute(
            "INSERT INTO delegation_tokens(from_user, to_user, permission_id, expires_at, can_redelegate, status) "
            "VALUES(?,?,?,?,?,'ACTIVE')",
            (u["id"], tu["id"], perm_id, expires_at, can_redel)
        )
        token_id = c.execute("SELECT last_insert_rowid()").fetchone()[0]
        # delegation_audit INSERT (트랜잭션 무결성 — audit 누락 0건)
        c.execute(
            "INSERT INTO delegation_audit(token_id, action, actor, details) VALUES(?,?,?,?)",
            (token_id, "GRANT", u["id"], f"{resource}.{action} → {tu['name']} / 만료 {expires_at} / 사유: {reason or '-'}")
        )
        _grant_target = tu["id"]
    # 알림시스템 통합 (사이클 2026-04-26) — 위임 받는 자에게 PERMISSION 알림
    notify_user(
        _grant_target, "PERMISSION",
        f"🔑 권한 위임 — {resource}.{action}",
        body=f"{u.get('name','')} 님이 권한을 위임했습니다 (만료 {expires_at})",
        link="/admin/permissions",
    )
    return RedirectResponse("/admin/permissions", 303)


@app.get("/admin/permissions/revoke", response_class=HTMLResponse)
async def admin_permissions_revoke_page(req: Request):
    """회수 탭 — 위임 카드 리스트 (시안 §6-1)"""
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    return ctx(req, "admin_permissions_revoke.html", user=u, active="admin")


@app.post("/admin/permissions/revoke")
async def admin_permissions_revoke_submit(req: Request):
    """위임 토큰 회수 + Cascade (S3 2차 본문 · 시안 §6-1)
    트랜잭션: 본 토큰 UPDATE status=REVOKED + 하위 재위임 토큰 Cascade UPDATE
              + delegation_audit INSERT (각 회수마다 1행, immutable)
    2단계 확인: confirm_text == '회수합니다' 인 경우만 실행
    """
    u = require(req, ["admin", "ceo"])
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    token_id     = form.get("token_id")
    confirm_text = (form.get("confirm_text") or "").strip()
    if not token_id:
        return JSONResponse({"error": "token_id 필수"}, 400)
    if confirm_text != "회수합니다":
        return JSONResponse({"error": "2단계 확인 실패 (회수합니다 입력 필요)"}, 400)
    with db_session() as c:
        # 본 토큰 회수
        row = c.execute("SELECT to_user FROM delegation_tokens WHERE id=?", (token_id,)).fetchone()
        if not row:
            return JSONResponse({"error": "토큰 없음"}, 404)
        c.execute("UPDATE delegation_tokens SET status='REVOKED' WHERE id=? AND status='ACTIVE'", (token_id,))
        c.execute(
            "INSERT INTO delegation_audit(token_id, action, actor, details) VALUES(?,?,?,?)",
            (token_id, "REVOKE", u["id"], f"본 토큰 회수 (token_id={token_id})")
        )
        _revoke_target = row["to_user"]
        # Cascade — 본 토큰 수령자가 재위임한 ACTIVE 하위 토큰 회수
        children = c.execute(
            "SELECT id FROM delegation_tokens WHERE from_user=? AND status='ACTIVE'",
            (row["to_user"],)
        ).fetchall()
        _cascade_targets = []
        for ch in children:
            c.execute("UPDATE delegation_tokens SET status='REVOKED' WHERE id=?", (ch["id"],))
            c.execute(
                "INSERT INTO delegation_audit(token_id, action, actor, details) VALUES(?,?,?,?)",
                (ch["id"], "REVOKE", u["id"], f"Cascade 회수 (parent_token_id={token_id})")
            )
            ct = c.execute(
                "SELECT to_user FROM delegation_tokens WHERE id=?", (ch["id"],)
            ).fetchone()
            if ct and ct["to_user"]:
                _cascade_targets.append(ct["to_user"])
    # 알림시스템 통합 (사이클 2026-04-26) — 회수 대상자에게 PERMISSION 알림
    notify_user(
        _revoke_target, "PERMISSION",
        "🔒 권한 회수",
        body=f"위임 권한이 회수되었습니다 (token_id={token_id})",
        link="/admin/permissions",
    )
    for tgt in _cascade_targets:
        notify_user(
            tgt, "PERMISSION", "🔒 권한 회수 (Cascade)",
            body=f"상위 위임 회수에 따라 하위 권한이 회수되었습니다 (parent_token_id={token_id})",
            link="/admin/permissions",
        )
    return RedirectResponse("/admin/permissions", 303)


def _audit_query(c, action: str = "", date_from: str = "", date_to: str = "",
                 actor: str = "", target: str = "", q: str = "", limit: int = 200):
    """S3 4차 — 감사 로그 검색·필터 공용 빌더.
    - action: GRANT/DELEGATE/REVOKE/EXPIRE/REDELEGATE (whitelist)
    - date_from/date_to: YYYY-MM-DD (timestamp 부분일치)
    - actor/target: 사용자명 LIKE (actor=u.name, target=tu.name via dt.to_user)
    - q: resource·action·token_id·details LIKE
    """
    sql = (
        "SELECT da.id, da.timestamp, da.action, da.details, da.actor, da.token_id, "
        "       u.name AS actor_name, tu.name AS target_name, "
        "       COALESCE(p.resource||'.'||p.action, p.name) AS perm_label, "
        "       p.resource AS perm_resource, p.action AS perm_action "
        "FROM delegation_audit da "
        "LEFT JOIN users u ON u.id = da.actor "
        "LEFT JOIN delegation_tokens dt ON dt.id = da.token_id "
        "LEFT JOIN users tu ON tu.id = dt.to_user "
        "LEFT JOIN permissions p ON p.id = dt.permission_id "
    )
    where, params = [], []
    if action in ("GRANT", "DELEGATE", "REVOKE", "EXPIRE", "REDELEGATE"):
        where.append("da.action=?"); params.append(action)
    if date_from:
        where.append("da.timestamp>=?"); params.append(date_from + " 00:00:00")
    if date_to:
        where.append("da.timestamp<=?"); params.append(date_to + " 23:59:59")
    if actor:
        where.append("u.name LIKE ?"); params.append(f"%{actor}%")
    if target:
        where.append("tu.name LIKE ?"); params.append(f"%{target}%")
    if q:
        where.append("(p.resource LIKE ? OR p.action LIKE ? OR CAST(da.token_id AS TEXT)=? OR da.details LIKE ?)")
        params.extend([f"%{q}%", f"%{q}%", q, f"%{q}%"])
    if where:
        sql += "WHERE " + " AND ".join(where) + " "
    sql += f"ORDER BY da.timestamp DESC, da.id DESC LIMIT {int(limit)}"
    try:
        return [dict(r) for r in c.execute(sql, params).fetchall()]
    except Exception:
        return []


@app.get("/admin/permissions/audit", response_class=HTMLResponse)
async def admin_permissions_audit_page(req: Request, action: str = "",
                                        date_from: str = "", date_to: str = "",
                                        actor: str = "", target: str = "", q: str = ""):
    """감사 로그 — 시간역순 타임라인 (시안 §6-2, append-only).
    S3 4차 — 액션·기간·actor·target·검색(q) 필터 강화.
    """
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    af = action if action in ("GRANT", "DELEGATE", "REVOKE", "EXPIRE", "REDELEGATE") else ""
    with db_session() as c:
        rows = _audit_query(c, af, date_from, date_to, actor, target, q, 200)
    return ctx(req, "admin_permissions_audit.html", user=u, active="admin",
               audit_rows=rows, action_filter=af,
               date_from=date_from, date_to=date_to,
               actor_q=actor, target_q=target, q=q)


@app.get("/admin/permissions/audit.csv")
async def admin_permissions_audit_csv(req: Request, action: str = "",
                                       date_from: str = "", date_to: str = "",
                                       actor: str = "", target: str = "", q: str = ""):
    """감사 로그 CSV 다운로드 — csv 모듈 단독 (외부 라이브러리 0).
    동일 필터 재사용 (LIMIT 5000 으로 상향)."""
    import csv as _csv
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    af = action if action in ("GRANT", "DELEGATE", "REVOKE", "EXPIRE", "REDELEGATE") else ""
    with db_session() as c:
        rows = _audit_query(c, af, date_from, date_to, actor, target, q, 5000)
    buf = io.StringIO()
    buf.write("﻿")  # UTF-8 BOM
    w = _csv.writer(buf)
    w.writerow(["id", "timestamp", "action", "actor", "target", "permission", "token_id", "details"])
    for r in rows:
        w.writerow([r.get("id"), r.get("timestamp"), r.get("action"),
                    r.get("actor_name") or r.get("actor") or "",
                    r.get("target_name") or "",
                    r.get("perm_label") or "",
                    r.get("token_id") or "",
                    (r.get("details") or "").replace("\n", " ")])
    fn = f"audit_{date.today().isoformat()}.csv"
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode("utf-8")),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={fn}"}
    )


# =====================================================
# TOP3 S3 — 권한 위임 3차 (2026-04-26)
# 시안: 05_HAIST_WORKS_디자인팀/_TO_01_정식시안_Top3_S3_v1.md
# 3차 = 권한 그룹 관리 + 매트릭스 보기 + 그룹 단위 위임.
# 권한 분기: CEO·admin only — 평직원/팀장 차단.
# =====================================================
@app.get("/admin/permissions/groups", response_class=HTMLResponse)
async def admin_permissions_groups_list(req: Request):
    """그룹 목록 — permission_groups + 권한 카운트 + 멤버 카운트 (시안 §5 그룹 상속)"""
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        try:
            rows = c.execute(
                "SELECT g.id, g.name, g.description, g.created_at, "
                "       (SELECT COUNT(*) FROM group_permissions gp WHERE gp.group_id=g.id) AS perm_count, "
                "       (SELECT COUNT(*) FROM user_groups ug WHERE ug.group_id=g.id) AS user_count "
                "FROM permission_groups g ORDER BY g.id ASC"
            ).fetchall()
            groups = [dict(r) for r in rows]
        except Exception:
            groups = []
    return ctx(req, "admin_permissions_groups.html", user=u, active="admin",
               groups=groups, group_detail=None, all_perms=[], all_users=[],
               group_perm_ids=set(), group_user_ids=set())


@app.post("/admin/permissions/groups")
async def admin_permissions_groups_create(req: Request):
    """그룹 신규 INSERT — name UNIQUE 가드"""
    u = require(req, ["admin", "ceo"])
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    name = (form.get("name") or "").strip()
    desc = (form.get("description") or "").strip()
    if not name:
        return JSONResponse({"error": "그룹명 필수"}, 400)
    with db_session() as c:
        exists = c.execute("SELECT id FROM permission_groups WHERE name=?", (name,)).fetchone()
        if exists:
            return RedirectResponse(f"/admin/permissions/groups/{exists['id']}", 303)
        c.execute(
            "INSERT INTO permission_groups(name, description) VALUES(?,?)",
            (name, desc)
        )
        gid = c.execute("SELECT last_insert_rowid()").fetchone()[0]
    return RedirectResponse(f"/admin/permissions/groups/{gid}", 303)


@app.get("/admin/permissions/groups/{group_id:int}", response_class=HTMLResponse)
async def admin_permissions_groups_detail(req: Request, group_id: int):
    """그룹 상세 — 그룹 정보 + 권한 + 멤버 + 추가 가능한 권한/사용자 목록"""
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        try:
            grow = c.execute(
                "SELECT id, name, description, created_at FROM permission_groups WHERE id=?",
                (group_id,)
            ).fetchone()
            if not grow:
                return RedirectResponse("/admin/permissions/groups", 303)
            group_detail = dict(grow)
            # 전체 그룹 목록 (좌측 사이드)
            grows = c.execute(
                "SELECT g.id, g.name, "
                "       (SELECT COUNT(*) FROM group_permissions gp WHERE gp.group_id=g.id) AS perm_count, "
                "       (SELECT COUNT(*) FROM user_groups ug WHERE ug.group_id=g.id) AS user_count "
                "FROM permission_groups g ORDER BY g.id ASC"
            ).fetchall()
            groups = [dict(r) for r in grows]
            # 전체 권한 카탈로그
            prows = c.execute(
                "SELECT id, COALESCE(resource||'.'||action, name) AS label, scope "
                "FROM permissions ORDER BY label"
            ).fetchall()
            all_perms = [dict(r) for r in prows]
            # 전체 사용자
            urows = c.execute(
                "SELECT id, name, login_id FROM users ORDER BY name LIMIT 200"
            ).fetchall()
            all_users = [dict(r) for r in urows]
            # 그룹에 이미 등록된 권한/사용자 ID
            gpids = {r["permission_id"] for r in c.execute(
                "SELECT permission_id FROM group_permissions WHERE group_id=?",
                (group_id,)
            ).fetchall()}
            guids = {r["user_id"] for r in c.execute(
                "SELECT user_id FROM user_groups WHERE group_id=?",
                (group_id,)
            ).fetchall()}
        except Exception:
            group_detail, groups, all_perms, all_users = None, [], [], []
            gpids, guids = set(), set()
    return ctx(req, "admin_permissions_groups.html", user=u, active="admin",
               groups=groups, group_detail=group_detail,
               all_perms=all_perms, all_users=all_users,
               group_perm_ids=gpids, group_user_ids=guids)


@app.post("/admin/permissions/groups/{group_id}/permissions")
async def admin_permissions_groups_perms(req: Request, group_id: int):
    """그룹↔권한 매핑 갱신 — checkbox 제출 후 전체 재기록 (트랜잭션)"""
    u = require(req, ["admin", "ceo"])
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    perm_ids = form.getlist("perm_id") if hasattr(form, "getlist") else form.getlist("perm_id")
    pids = []
    for v in perm_ids:
        try:
            pids.append(int(v))
        except (TypeError, ValueError):
            continue
    with db_session() as c:
        c.execute("DELETE FROM group_permissions WHERE group_id=?", (group_id,))
        for pid in pids:
            c.execute(
                "INSERT OR IGNORE INTO group_permissions(group_id, permission_id) VALUES(?,?)",
                (group_id, pid)
            )
    return RedirectResponse(f"/admin/permissions/groups/{group_id}", 303)


@app.post("/admin/permissions/groups/{group_id}/users")
async def admin_permissions_groups_users(req: Request, group_id: int):
    """그룹↔사용자 매핑 갱신 — checkbox 제출 후 전체 재기록 (트랜잭션)"""
    u = require(req, ["admin", "ceo"])
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    user_ids = form.getlist("user_id") if hasattr(form, "getlist") else form.getlist("user_id")
    uids = []
    for v in user_ids:
        try:
            uids.append(int(v))
        except (TypeError, ValueError):
            continue
    with db_session() as c:
        c.execute("DELETE FROM user_groups WHERE group_id=?", (group_id,))
        for uid in uids:
            c.execute(
                "INSERT OR IGNORE INTO user_groups(user_id, group_id) VALUES(?,?)",
                (uid, group_id)
            )
    return RedirectResponse(f"/admin/permissions/groups/{group_id}", 303)


@app.get("/admin/permissions/matrix", response_class=HTMLResponse)
async def admin_permissions_matrix(req: Request, dept: str = "", q: str = ""):
    """권한 매트릭스 — 사용자 vs 권한 (resource×action). 직접/그룹/위임 3색 (정적 CSS grid)
    - 부서 필터 + 검색 (이름/login_id 부분일치)
    - JS 0건 (서버 렌더링)
    """
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    dept_filter = (dept or "").strip()
    query = (q or "").strip()
    matrix = []
    perms = []
    depts = []
    with db_session() as c:
        try:
            # 권한 카탈로그
            prows = c.execute(
                "SELECT id, COALESCE(resource||'.'||action, name) AS label "
                "FROM permissions ORDER BY label LIMIT 30"
            ).fetchall()
            perms = [dict(r) for r in prows]
            # 부서 목록 (셀렉터)
            try:
                drows = c.execute("SELECT DISTINCT dept FROM users WHERE dept IS NOT NULL AND dept<>'' ORDER BY dept").fetchall()
                depts = [r["dept"] for r in drows]
            except Exception:
                depts = []
            # 사용자 목록
            usql = "SELECT id, name, login_id, COALESCE(dept,'') AS dept FROM users WHERE 1=1 "
            uparams = []
            if dept_filter:
                usql += "AND dept=? "
                uparams.append(dept_filter)
            if query:
                usql += "AND (name LIKE ? OR login_id LIKE ?) "
                uparams.extend([f"%{query}%", f"%{query}%"])
            usql += "ORDER BY name LIMIT 60"
            urows = c.execute(usql, uparams).fetchall()
            # 각 사용자 × 권한 셀: 직접(D) / 그룹(G) / 위임(T) 마크
            for ur in urows:
                uid = ur["id"]
                # 그룹 상속 권한
                ginh = {r["permission_id"] for r in c.execute(
                    "SELECT DISTINCT gp.permission_id FROM group_permissions gp "
                    "JOIN user_groups ug ON ug.group_id=gp.group_id WHERE ug.user_id=?",
                    (uid,)
                ).fetchall()}
                # 위임 토큰 (ACTIVE만)
                tdel = {r["permission_id"] for r in c.execute(
                    "SELECT permission_id FROM delegation_tokens WHERE to_user=? AND status='ACTIVE'",
                    (uid,)
                ).fetchall()}
                # 직접 권한 — user_permissions 가 별도로 없으면 빈셋 (스키마에 따라)
                try:
                    drect = {r["permission_id"] for r in c.execute(
                        "SELECT permission_id FROM user_permissions WHERE user_id=?",
                        (uid,)
                    ).fetchall()}
                except Exception:
                    drect = set()
                cells = []
                for p in perms:
                    pid = p["id"]
                    mark = ""
                    if pid in drect:
                        mark = "D"
                    elif pid in ginh:
                        mark = "G"
                    elif pid in tdel:
                        mark = "T"
                    cells.append(mark)
                matrix.append({
                    "user_id": uid, "name": ur["name"],
                    "login_id": ur["login_id"], "dept": ur["dept"],
                    "cells": cells,
                })
        except Exception:
            matrix, perms, depts = [], [], []
    return ctx(req, "admin_permissions_matrix.html", user=u, active="admin",
               matrix=matrix, perms=perms, depts=depts,
               dept_filter=dept_filter, query=query)


@app.post("/admin/permissions/grant-group")
async def admin_permissions_grant_group(req: Request):
    """그룹 단위 위임 — 그룹 전체 멤버에 동일 권한을 위임 토큰으로 발행
    트랜잭션: 멤버별 delegation_tokens INSERT + delegation_audit INSERT (1 트랜잭션)
    """
    u = require(req, ["admin", "ceo"])
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    group_id   = form.get("group_id")
    resource   = (form.get("resource") or "").strip()
    action     = (form.get("action") or "").strip()
    expires_at = (form.get("expires_at") or "").strip()
    reason     = (form.get("reason") or "").strip()
    can_redel  = 1 if form.get("can_redelegate") else 0
    if not (group_id and resource and action and expires_at):
        return JSONResponse({"error": "필수 항목 누락"}, 400)
    with db_session() as c:
        gr = c.execute("SELECT id, name FROM permission_groups WHERE id=?", (group_id,)).fetchone()
        if not gr:
            return JSONResponse({"error": "그룹 없음"}, 404)
        # 권한 카탈로그 조회/INSERT
        prow = c.execute(
            "SELECT id FROM permissions WHERE resource=? AND action=? LIMIT 1",
            (resource, action)
        ).fetchone()
        if prow:
            perm_id = prow["id"]
        else:
            c.execute(
                "INSERT INTO permissions(name, resource, action, scope, description) VALUES(?,?,?,?,?)",
                (f"{resource}.{action}", resource, action, resource, f"{resource} {action}")
            )
            perm_id = c.execute("SELECT last_insert_rowid()").fetchone()[0]
        # 그룹 멤버 조회
        members = c.execute(
            "SELECT user_id FROM user_groups WHERE group_id=?", (group_id,)
        ).fetchall()
        if not members:
            return JSONResponse({"error": "그룹 멤버 없음"}, 400)
        # 멤버별 토큰 + audit 발행
        for m in members:
            c.execute(
                "INSERT INTO delegation_tokens(from_user, to_user, permission_id, expires_at, can_redelegate, status) "
                "VALUES(?,?,?,?,?,'ACTIVE')",
                (u["id"], m["user_id"], perm_id, expires_at, can_redel)
            )
            tid = c.execute("SELECT last_insert_rowid()").fetchone()[0]
            c.execute(
                "INSERT INTO delegation_audit(token_id, action, actor, details) VALUES(?,?,?,?)",
                (tid, "GRANT", u["id"],
                 f"[그룹위임] {gr['name']} → user_id={m['user_id']} / {resource}.{action} / 만료 {expires_at} / 사유: {reason or '-'}")
            )
    return RedirectResponse(f"/admin/permissions/groups/{group_id}", 303)


# =====================================================
# TOP3 S3 — 권한 위임 4차 (2026-04-26): 권한 리포트 + 만료 정리
# 시안: 05_HAIST_WORKS_디자인팀/_TO_01_정식시안_Top3_S3_v1.md §7 운영 리포트
# 4차 = 사용자/그룹/만료임박 3리포트 + 만료 토큰 자동 정리(audit immutable).
# 권한: CEO·admin only.
# =====================================================
@app.get("/admin/permissions/report/users", response_class=HTMLResponse)
async def admin_permissions_report_users(req: Request):
    """사용자별 권한 카운트 + 활성 토큰 (시안 §7-1)."""
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        try:
            rows = c.execute(
                "SELECT u.id, u.name, u.role, t.name AS team_name, "
                "       (SELECT COUNT(*) FROM delegation_tokens dt "
                "          WHERE dt.to_user=u.id AND dt.status='ACTIVE') AS active_tokens, "
                "       (SELECT COUNT(*) FROM delegation_tokens dt "
                "          WHERE dt.from_user=u.id AND dt.status='ACTIVE') AS granted_tokens, "
                "       (SELECT COUNT(DISTINCT ug.group_id) FROM user_groups ug "
                "          WHERE ug.user_id=u.id) AS group_count "
                "FROM users u LEFT JOIN teams t ON t.id=u.team_id "
                "ORDER BY active_tokens DESC, u.name ASC LIMIT 500"
            ).fetchall()
            users = [dict(r) for r in rows]
        except Exception:
            users = []
    # v5H46: 템플릿이 summary/headers/rows/report_title 형식 기대
    summary = _perm_report_summary()
    rows_view = [[ux.get("name",""), ux.get("team_name") or "-", ux.get("role",""),
                  ux.get("active_tokens",0), ux.get("granted_tokens",0),
                  ux.get("group_count",0)] for ux in users]
    return ctx(req, "admin_permissions_report.html", user=u, active="admin",
               report_kind="users", users=users, groups=[], expiring=[],
               summary=summary, report_title="사용자별 권한 분포",
               headers=["이름","팀","역할","수신중","위임중","그룹수"],
               rows=rows_view)


@app.get("/admin/permissions/report/groups", response_class=HTMLResponse)
async def admin_permissions_report_groups(req: Request):
    """그룹별 멤버·권한 분포 (시안 §7-2)."""
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        try:
            rows = c.execute(
                "SELECT g.id, g.name, g.description, "
                "       (SELECT COUNT(*) FROM group_permissions gp WHERE gp.group_id=g.id) AS perm_count, "
                "       (SELECT COUNT(*) FROM user_groups ug WHERE ug.group_id=g.id) AS user_count "
                "FROM permission_groups g ORDER BY user_count DESC, g.name ASC"
            ).fetchall()
            groups = [dict(r) for r in rows]
        except Exception:
            groups = []
    summary = _perm_report_summary()
    rows_view = [[g.get("name",""), g.get("description") or "-",
                  g.get("perm_count",0), g.get("user_count",0)] for g in groups]
    return ctx(req, "admin_permissions_report.html", user=u, active="admin",
               report_kind="groups", users=[], groups=groups, expiring=[],
               summary=summary, report_title="그룹별 멤버·권한 분포",
               headers=["그룹명","설명","권한수","멤버수"],
               rows=rows_view)


@app.get("/admin/permissions/report/expiring", response_class=HTMLResponse)
async def admin_permissions_report_expiring(req: Request, days: int = 7):
    """만료 임박 토큰 (시안 §7-3) — 기본 7일 내 ACTIVE."""
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    try:
        d = max(1, min(int(days), 90))
    except Exception:
        d = 7
    cutoff = (date.today() + timedelta(days=d)).isoformat()
    with db_session() as c:
        try:
            rows = c.execute(
                "SELECT dt.id AS token_id, dt.expires_at, dt.status, "
                "       fu.name AS from_name, tu.name AS to_name, "
                "       COALESCE(p.resource||'.'||p.action, p.name) AS perm_label "
                "FROM delegation_tokens dt "
                "LEFT JOIN users fu ON fu.id=dt.from_user "
                "LEFT JOIN users tu ON tu.id=dt.to_user "
                "LEFT JOIN permissions p ON p.id=dt.permission_id "
                "WHERE dt.status='ACTIVE' AND dt.expires_at IS NOT NULL "
                "  AND dt.expires_at<=? "
                "ORDER BY dt.expires_at ASC LIMIT 500",
                (cutoff,)
            ).fetchall()
            expiring = [dict(r) for r in rows]
        except Exception:
            expiring = []
    summary = _perm_report_summary()
    rows_view = [[ex.get("expires_at","-"), ex.get("from_name","-"),
                  ex.get("to_name","-"), ex.get("perm_label","-"),
                  ex.get("status","-")] for ex in expiring]
    return ctx(req, "admin_permissions_report.html", user=u, active="admin",
               report_kind="expiring", users=[], groups=[], expiring=expiring,
               expiring_days=d,
               summary=summary,
               report_title=f"만료 임박 토큰 ({d}일 내)",
               headers=["만료일","From","To","권한","상태"],
               rows=rows_view)


def _perm_report_summary() -> dict:
    """권한 리포트 4 KPI 요약 — 모든 리포트 페이지 공통 헤더."""
    out = {"total_users": 0, "users_with_perms": 0,
           "total_perms": 0, "active_delegations": 0}
    with db_session() as c:
        try:
            out["total_users"] = c.execute(
                "SELECT COUNT(*) FROM users WHERE is_active=1"
            ).fetchone()[0]
        except Exception:
            pass
        try:
            out["users_with_perms"] = c.execute(
                "SELECT COUNT(DISTINCT to_user) FROM delegation_tokens "
                "WHERE status='ACTIVE'"
            ).fetchone()[0]
        except Exception:
            pass
        try:
            out["total_perms"] = c.execute(
                "SELECT COUNT(*) FROM permissions"
            ).fetchone()[0]
        except Exception:
            pass
        try:
            out["active_delegations"] = c.execute(
                "SELECT COUNT(*) FROM delegation_tokens WHERE status='ACTIVE'"
            ).fetchone()[0]
        except Exception:
            pass
    return out


@app.post("/admin/permissions/cleanup-expired")
async def admin_permissions_cleanup_expired(req: Request):
    """만료 토큰 자동 정리 — status='ACTIVE' AND expires_at<=now → status='EXPIRED'.
    트랜잭션: UPDATE + delegation_audit INSERT (각 토큰별 1행, immutable, audit 누락 0건).
    수동 트리거 (스케줄러 미구현).
    """
    u = require(req, ["admin", "ceo"])
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cleaned = 0
    with db_session() as c:
        try:
            rows = c.execute(
                "SELECT id FROM delegation_tokens "
                "WHERE status='ACTIVE' AND expires_at IS NOT NULL AND expires_at<=?",
                (now_str,)
            ).fetchall()
            ids = [r["id"] for r in rows]
            for tid in ids:
                c.execute(
                    "UPDATE delegation_tokens SET status='EXPIRED' WHERE id=? AND status='ACTIVE'",
                    (tid,)
                )
                c.execute(
                    "INSERT INTO delegation_audit(token_id, action, actor, details) VALUES(?,?,?,?)",
                    (tid, "EXPIRE", u["id"], f"수동 만료 정리 (cleanup-expired @ {now_str})")
                )
                cleaned += 1
        except Exception as e:
            return JSONResponse({"error": str(e)}, 500)
    return RedirectResponse(f"/admin/permissions/report/expiring?msg={cleaned}건+정리완료", 303)


# =====================================================
# TOP3 S2 — 재고 입출고 1차 라우트 골격 (2026-04-25)
# 시안: 05_HAIST_WORKS_디자인팀/_TO_01_정식시안_Top3_S2_v1.md
# 1차 = 골격만 (UI 본문 · INSERT/UPDATE 로직은 다음 사이클).
# 권한: P4 구매팀(can_use_logistics) 또는 admin/ceo.
# =====================================================
def _s2_guard(req: Request):
    """S2 권한 가드 — 구매팀 권한 OR admin/ceo. 없으면 None 반환."""
    u = get_user(req)
    if not u:
        return None
    if u.get("role") in ("admin", "ceo") or can_use_logistics(u):
        return u
    return None


@app.get("/stock/balances", response_class=HTMLResponse)
async def stock_balances_page(req: Request):
    """재고 잔고 — stock_balances VIEW 조회 (시안 §화면 영역 잔고)"""
    u = _s2_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        rows = c.execute(
            """SELECT sb.part_id, sb.part_no, sb.part_name, sb.on_hand, sb.unit,
                      sb.last_movement_at, p.spec, p.std_price
               FROM stock_balances sb
               LEFT JOIN parts p ON p.id = sb.part_id
               ORDER BY sb.part_no LIMIT 200"""
        ).fetchall()
    balances = []
    for r in rows:
        b = dict(r)
        b["qty"] = b.get("on_hand") or 0
        b["unit_price"] = b.get("std_price") or 0
        b["value"] = (b["qty"] * b["unit_price"]) if b["unit_price"] else 0
        b["safety_stock"] = 0  # 추후 part_safety_stock 테이블 연동
        balances.append(b)
    last_update = max((b.get("last_movement_at") or "") for b in balances) if balances else "-"
    return ctx(req, "stock_balances.html", user=u, active="stock",
               balances=balances, last_update=last_update or "-")


# Top3 S2 3차 (2026-04-26) — FIFO 레이어 상세 / ABC 분류 / 재고회전율 ==========
@app.get("/stock/balances/fifo/{part_id}", response_class=HTMLResponse)
async def stock_fifo_page(req: Request, part_id: int):
    """FIFO 레이어 상세 — 입고일·잔량·단가 시각화 (S2-3차)."""
    u = _s2_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    summary = fifo_layers(part_id)
    return ctx(req, "stock_fifo.html", user=u, active="stock", summary=summary)


@app.get("/stock/abc", response_class=HTMLResponse)
async def stock_abc_page(req: Request, days: int = 90, top: int = 50):
    """ABC 분류 — 최근 N일 출고 매출 누적 비중 기준 (A 80%, B 95%, C 나머지)."""
    u = _s2_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    items = abc_classification(days=days)
    a_cnt = sum(1 for r in items if r["abc_class"] == "A")
    b_cnt = sum(1 for r in items if r["abc_class"] == "B")
    c_cnt = sum(1 for r in items if r["abc_class"] == "C")
    return ctx(req, "stock_abc.html", user=u, active="stock",
               items=items[:top], total=len(items),
               a_count=a_cnt, b_count=b_cnt, c_count=c_cnt, days=days)


@app.get("/stock/turnover", response_class=HTMLResponse)
async def stock_turnover_page(req: Request, days: int = 90):
    """재고 회전율 — 출고량/평균재고 (FAST≥2 / NORMAL 0.5~2 / SLOW<0.5)."""
    u = _s2_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    items = stock_turnover(days=days)
    fast = sum(1 for r in items if r["band"] == "FAST")
    normal = sum(1 for r in items if r["band"] == "NORMAL")
    slow = sum(1 for r in items if r["band"] == "SLOW")
    return ctx(req, "stock_turnover.html", user=u, active="stock",
               items=items, fast=fast, normal=normal, slow=slow, days=days)


@app.get("/stock/receipts", response_class=HTMLResponse)
async def stock_receipts_page(req: Request):
    """입고 목록 — receipts 테이블 (시안 §화면 영역 입고 GR)"""
    u = _s2_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        rows = c.execute(
            """SELECT id, po_id, total_qty, qc_inspection_id, status, received_at, note
               FROM receipts ORDER BY id DESC LIMIT 100"""
        ).fetchall()
    return ctx(req, "stock_receipts.html", user=u, active="stock",
               receipts=[dict(r) for r in rows])


@app.post("/stock/receipts")
async def stock_receipts_submit(req: Request):
    """입고 등록 (Top3-S2-2차) — INSERT receipts + INSERT stock_movements{kind=IN, qty +}.
    receipts.status=PENDING (검수 대기). 트랜잭션 무결성: 두 INSERT 동일 db_session.
    """
    u = _s2_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    try:
        po_id = int(form.get("po_id") or 0) or None
        part_id = int(form.get("part_id") or 0)
        qty = float(form.get("qty") or 0)
        if part_id <= 0 or qty <= 0:
            raise ValueError("part_id/qty invalid")
    except Exception:
        return RedirectResponse("/stock/receipts?error=invalid", 303)
    # 1) receipts INSERT
    with db_session() as c:
        c.execute(
            """INSERT INTO receipts (po_id, received_by, total_qty, status, note)
               VALUES (?,?,?,?,?)""",
            (po_id, u.get("id"), qty, "PENDING", "S2-2차 GR")
        )
        gr_id = c.execute("SELECT last_insert_rowid()").fetchone()[0]
    # 2) stock_movements INSERT (kind=IN, qty +) — balance VIEW 자동 반영
    try:
        from .database import stock_movement_create
        stock_movement_create({
            "part_id": part_id, "kind": "IN", "quantity": qty,
            "po_id": po_id, "reason": f"GR-{gr_id}", "note": "Top3-S2-2차 입고"
        }, u.get("id") or 0)
    except Exception as e:
        return RedirectResponse(f"/stock/receipts?error=mv:{e}", 303)
    return RedirectResponse(f"/stock/receipts?success=GR-{gr_id}", 303)


# 라우트 등록 순서 보정 (04 V10 권고): /stock/qc/disposition 을 /stock/qc/{po_item_id} 위로
@app.get("/stock/qc/disposition/{qc_id}", response_class=HTMLResponse)
async def stock_qc_disposition_page(req: Request, qc_id: int):
    """부적합 처리 모달 — RETURN/SPECIAL_ACCEPT/SCRAP (시안 §데이터 모델 qc_disposition)"""
    u = _s2_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    return ctx(req, "stock_qc.html", user=u, qc_id=qc_id, mode="disposition", active="stock")


@app.post("/stock/qc/disposition")
async def stock_qc_disposition_submit(req: Request):
    """부적합 처리 — INSERT qc_disposition + UPDATE qc_inspections.status (FAIL 분기 확정)"""
    u = _s2_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    try:
        qc_id = int(form.get("qc_inspection_id") or 0)
        action = (form.get("action") or "").upper()
        note = (form.get("note") or "").strip()
        if qc_id <= 0 or action not in ("RETURN", "SPECIAL_ACCEPT", "SCRAP"):
            raise ValueError("invalid")
    except Exception:
        return RedirectResponse("/stock/receipts?error=disp_invalid", 303)
    with db_session() as c:
        c.execute(
            """INSERT INTO qc_disposition (qc_inspection_id, action, decided_by, note)
               VALUES (?,?,?,?)""",
            (qc_id, action, u.get("id"), note or None)
        )
        # SPECIAL_ACCEPT는 부분 사용 가능 → status 유지, RETURN/SCRAP는 FAIL 확정
        if action in ("RETURN", "SCRAP"):
            c.execute("UPDATE qc_inspections SET status='FAIL' WHERE id=?", (qc_id,))
    return RedirectResponse(f"/stock/receipts?success=disp-{action}", 303)


@app.get("/stock/qc/{po_item_id}", response_class=HTMLResponse)
async def stock_qc_page(req: Request, po_item_id: int):
    """검수 화면 — qc_inspections 작성 (시안 §화면 영역 QC 우 패널)"""
    u = _s2_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    return ctx(req, "stock_qc.html", user=u, po_item_id=po_item_id, active="stock")


@app.post("/stock/qc/{po_item_id}")
async def stock_qc_submit(req: Request, po_item_id: int):
    """검수 결과 등록 (Top3-S2-2차) — INSERT qc_inspections + UPDATE receipts.qc_inspection_id.
    status 분기: PASS / PARTIAL / HOLD / FAIL. FAIL/PARTIAL 시 부적합 모달로 redirect.
    """
    u = _s2_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    try:
        pass_qty = float(form.get("pass_qty") or 0)
        fail_qty = float(form.get("fail_qty") or 0)
        status = (form.get("status") or "PENDING").upper()
        if status not in ("PASS", "PARTIAL", "HOLD", "FAIL"):
            status = "PENDING"
        fail_reason = (form.get("fail_reason") or "").strip() or None
    except Exception:
        return RedirectResponse(f"/stock/qc/{po_item_id}?error=invalid", 303)
    # OPS-V5 [정적점검]: 음수 차단 + FAIL/PARTIAL 사유 필수 (감사 추적)
    if pass_qty < 0 or fail_qty < 0:
        return RedirectResponse(f"/stock/qc/{po_item_id}?error=qty_negative", 303)
    if status in ("FAIL", "PARTIAL") and not fail_reason:
        return RedirectResponse(f"/stock/qc/{po_item_id}?error=fail_reason_required", 303)
    # po_item_id 가 receipts.id 로 들어올 수 있으므로 둘다 시도 (UI 단순화)
    with db_session() as c:
        c.execute(
            """INSERT INTO qc_inspections
               (po_item_id, receipt_id, inspector_id, pass_qty, fail_qty, fail_reason, status)
               VALUES (?,?,?,?,?,?,?)""",
            (po_item_id, po_item_id, u.get("id"), pass_qty, fail_qty, fail_reason, status)
        )
        qc_id = c.execute("SELECT last_insert_rowid()").fetchone()[0]
        # receipts UPDATE: qc_inspection_id 연결 + status 동기화
        c.execute(
            "UPDATE receipts SET qc_inspection_id=?, status=? WHERE id=?",
            (qc_id, status, po_item_id)
        )
    # FAIL/PARTIAL → 부적합 모달
    if status in ("FAIL", "PARTIAL"):
        return RedirectResponse(f"/stock/qc/disposition/{qc_id}", 303)
    return RedirectResponse(f"/stock/receipts?success=qc-{status}", 303)


@app.get("/stock/issues", response_class=HTMLResponse)
async def stock_issues_page(req: Request):
    """출고 목록 — issues_out 테이블 (시안 §화면 영역 출고 GI)"""
    u = _s2_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        rows = c.execute(
            """SELECT id, part_id, qty, purpose, status, requested_at, issued_at
               FROM issues_out ORDER BY id DESC LIMIT 100"""
        ).fetchall()
    return ctx(req, "stock_issues.html", user=u, active="stock",
               issues=[dict(r) for r in rows])


@app.post("/stock/issues")
async def stock_issues_submit(req: Request):
    """출고 등록 (Top3-S2-2차) — INSERT issues_out · status=PENDING.
    실제 재고 차감은 /stock/issues/{id}/approve 에서 stock_movements 동시 INSERT.
    """
    u = _s2_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    try:
        part_id = int(form.get("part_id") or 0)
        qty = float(form.get("qty") or 0)
        purpose = (form.get("purpose") or "").strip() or None
        if part_id <= 0 or qty <= 0:
            raise ValueError("invalid")
    except Exception:
        return RedirectResponse("/stock/issues?error=invalid", 303)
    with db_session() as c:
        c.execute(
            """INSERT INTO issues_out (part_id, requester_id, qty, purpose, status)
               VALUES (?,?,?,?,?)""",
            (part_id, u.get("id"), qty, purpose, "PENDING")
        )
        gi_id = c.execute("SELECT last_insert_rowid()").fetchone()[0]
    return RedirectResponse(f"/stock/issues?success=GI-{gi_id}", 303)


@app.post("/stock/issues/{issue_id}/approve")
async def stock_issues_approve(req: Request, issue_id: int):
    """출고 승인·실행 (Top3-S2-2차) — UPDATE issues_out.status=ISSUED + INSERT stock_movements{kind=OUT, qty -}.
    트랜잭션 무결성: stock_movement_create 가 자체 db_session 으로 balance VIEW 즉시 반영.
    """
    u = _s2_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    # 1) issues_out 조회 + 승인
    with db_session() as c:
        row = c.execute(
            "SELECT id, part_id, qty, status FROM issues_out WHERE id=?", (issue_id,)
        ).fetchone()
        if not row:
            return RedirectResponse("/stock/issues?error=notfound", 303)
        if row["status"] != "PENDING":
            return RedirectResponse(f"/stock/issues?error=already-{row['status']}", 303)
        c.execute(
            "UPDATE issues_out SET status='ISSUED', approver_id=?, issued_at=datetime('now','localtime') WHERE id=?",
            (u.get("id"), issue_id)
        )
        part_id = row["part_id"]
        qty = float(row["qty"] or 0)
    # 2) stock_movements INSERT (kind=OUT, qty -) — balance VIEW 자동 반영 (FIFO)
    try:
        from .database import stock_movement_create
        stock_movement_create({
            "part_id": part_id, "kind": "OUT", "quantity": qty,
            "reason": f"GI-{issue_id}", "note": "Top3-S2-2차 출고 승인"
        }, u.get("id") or 0)
    except Exception as e:
        return RedirectResponse(f"/stock/issues?error=mv:{e}", 303)
    return RedirectResponse(f"/stock/issues?success=GI-{issue_id}-ISSUED", 303)


# =====================================================
# TOP3 S1 — 매출 라이프사이클 1차 라우트 골격 (2026-04-25)
# 시안: 05_HAIST_WORKS_디자인팀/_TO_01_정식시안_Top3_S1_v1.md
# 1차 = 골격만 (UI 본문 · INSERT/UPDATE 로직은 다음 사이클).
# 권한: P2 영업팀(can_use_sales) 또는 admin/ceo. 평직원 차단.
# 4탭: 견적QT(탭1) / 수주SO(탭2) / 생산WO(탭3) / 출하DO·수금RC(탭4)
# 9 enum: DRAFT/QUOTED/CONFIRMED/IN_PRODUCTION/READY_TO_SHIP/SHIPPED/INVOICED/PAID/CANCELLED
#         (database.py orders.status CHECK constraint · invoices 추가 정합)
# =====================================================
def _s1_guard(req: Request):
    """S1 권한 가드 — 영업팀 권한 OR admin/ceo. 없으면 None 반환."""
    u = get_user(req)
    if not u:
        return None
    if u.get("role") in ("admin", "ceo") or can_use_sales(u):
        return u
    return None


@app.get("/sales/quotations", response_class=HTMLResponse)
async def sales_quotations_page(req: Request):
    """견적 탭 (시안 §1 탭1 QT) — quotations 리스트 + Empty State"""
    u = _s1_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        rows = c.execute(
            """SELECT q.id, q.quote_no, q.customer_id,
                      COALESCE(cu.name,'-') AS customer_name,
                      q.total_amount, q.valid_until, q.version, q.status,
                      q.created_at
               FROM quotations q
               LEFT JOIN customers cu ON cu.id = q.customer_id
               ORDER BY q.id DESC LIMIT 200"""
        ).fetchall()
        items = [dict(r) for r in rows]
        # 사이클 61 U5 — 각 견적의 라인 사전 로드 (삭제 UI 노출용)
        lines_by_quote = {}
        if items:
            quote_ids = [it["id"] for it in items]
            placeholders = ",".join(["?"] * len(quote_ids))
            line_rows = c.execute(
                f"""SELECT id, quotation_id, line_no, item_name, qty, unit,
                          unit_price, total_price, note
                   FROM quotation_items
                   WHERE quotation_id IN ({placeholders})
                   ORDER BY quotation_id, line_no""",
                quote_ids,
            ).fetchall()
            for lr in line_rows:
                lines_by_quote.setdefault(lr["quotation_id"], []).append(dict(lr))
    return ctx(req, "sales_quotations.html", user=u, active="sales_quotations",
               tab="quotations", items=items, lines_by_quote=lines_by_quote)


@app.post("/sales/quotations")
async def sales_quotations_create(req: Request):
    """견적 생성 (Top3-S1-2차 — quotations INSERT, status=DRAFT)"""
    u = _s1_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    customer_id = form.get("customer_id") or None
    # v5H113 M9: 헤더 단독 생성 시 total_amount 임의 입력 방지 — 0 으로 고정 (라인 추가 시 자동 합산)
    total_amount = 0.0
    valid_until = form.get("valid_until") or None
    version = int(form.get("version") or 1)
    with db_session() as c:
        # quote_no = QT-YYYYMM-#### (월별 시퀀스)
        ym = datetime.now().strftime("%Y%m")
        row = c.execute(
            "SELECT COUNT(*) FROM quotations WHERE quote_no LIKE ?",
            (f"QT-{ym}-%",),
        ).fetchone()
        seq = (row[0] if row else 0) + 1
        quote_no = f"QT-{ym}-{seq:04d}"
        cur = c.execute(
            """INSERT INTO quotations
               (quote_no, customer_id, total_amount, valid_until, version,
                status, created_by)
               VALUES (?,?,?,?,?,'DRAFT',?)""",
            (quote_no, customer_id, total_amount, valid_until, version, u.get("id")),
        )
        return JSONResponse({"ok": True, "quote_id": cur.lastrowid, "quote_no": quote_no})


# =====================================================
# 회사 식별 정보 (사이클 61 — 견적서 헤더용 / 외부자산 0건)
# 실제 값은 /admin/company-info 페이지에서 대표가 직접 입력 → app_settings 테이블
# 미입력 시 placeholder 텍스트 노출 ("[등록 대기]")
# =====================================================
COMPANY_INFO_KEYS = [
    ("company_name_ko",  "회사명 (국문)",   "주식회사 케이엔케이 (KNK)"),
    ("company_name_en",  "Company (EN)",   "KNK Co., Ltd."),
    ("company_biz_no",   "사업자등록번호",  ""),
    ("company_address",  "회사 주소",       ""),
    ("company_address_en", "Address (EN)",  ""),
    ("company_tel",      "대표 전화",       ""),
    ("company_fax",      "팩스",            ""),
    ("company_email",    "대표 이메일",     ""),
    ("company_ceo_ko",   "대표자 (국문)",   "김정락"),
    ("company_ceo_en",   "CEO (EN)",       "Jeongrak Kim"),
]

def _company_info_dict() -> dict:
    """app_settings 에서 회사 정보 조회. 미입력 키는 default 값 반환."""
    out = {}
    for key, label, default in COMPANY_INFO_KEYS:
        val = (get_setting(key, "") or "").strip()
        out[key] = val if val else default
    return out


# =====================================================
# Sales 견적 라인 + 인쇄 + 수주 전환 (사이클 58 — 2차 보강)
# =====================================================

@app.get("/sales/quotations/{quote_id}/print", response_class=HTMLResponse)
async def sales_quotation_print(req: Request, quote_id: int, lang: str = ""):
    """견적서 인쇄 view — 1페이지 A4, @media print 인라인 CSS, 외부 자산 0건.
    사이클 63 P11 — 다국어 지원 (?lang=ko|en|vi 쿼리. 미지정 시 사용자 lang)."""
    u = _s1_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        h = c.execute(
            """SELECT q.id, q.quote_no, q.customer_id,
                      COALESCE(cu.name,'-')   AS customer_name,
                      COALESCE(cu.tier,'')    AS customer_tier,
                      COALESCE(cu.note,'')    AS customer_note,
                      q.total_amount, q.valid_until, q.version, q.status,
                      q.created_at,
                      COALESCE(us.name,'-')   AS owner_name
               FROM quotations q
               LEFT JOIN customers cu ON cu.id = q.customer_id
               LEFT JOIN users     us ON us.id = q.created_by
               WHERE q.id = ?""",
            (quote_id,),
        ).fetchone()
        if not h:
            return RedirectResponse("/sales/quotations", 303)
        header = dict(h)
    # 라인 (헬퍼 호출 — 자체 db_session)
    from .database import get_quotation_items as _gqi
    lines = _gqi(quote_id)
    company = _company_info_dict()  # 사이클 61 U6 — 회사 식별 정보
    # 사이클 63 — 견적서 인쇄 언어 override (?lang=ko|en|vi 만 허용. 그 외 무시)
    # ctx() 의 base.update(kwargs) 가 마지막에 실행 → 여기서 lang/i 덮어쓰기 가능.
    # print_lang 미지정 시 사용자 lang 그대로 사용 (ctx 기본 동작).
    extra = {"header": header, "lines": lines, "company": company,
             "quote_id": quote_id}
    if lang in ("ko", "en", "vi"):
        extra["lang"] = lang
        extra["i"] = get_all_translations(lang)
    return ctx(req, "quotation_print.html", user=u, active="sales", **extra)


@app.post("/sales/quotations/{quote_id}/items")
async def sales_quotation_item_add(req: Request, quote_id: int):
    """견적 라인 추가 — quotation_items INSERT + quotations.total_amount 재합계."""
    u = _s1_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    item_name = (form.get("item_name") or "").strip()
    if not item_name:
        return JSONResponse({"error": "품목명 누락"}, 400)
    part_id = form.get("part_id") or None
    try:
        part_id = int(part_id) if part_id else None
    except (TypeError, ValueError):
        part_id = None
    qty = float(form.get("qty") or 0)
    unit = (form.get("unit") or "EA").strip()
    unit_price = float(form.get("unit_price") or 0)
    note = (form.get("note") or "").strip()
    total_price = qty * unit_price
    with db_session() as c:
        # 견적 존재 검증
        h = c.execute("SELECT id, status FROM quotations WHERE id=?", (quote_id,)).fetchone()
        if not h:
            return JSONResponse({"error": "견적 없음"}, 404)
        # v5H112: SO 발행된(CONFIRMED) 견적은 라인 편집 거부 — SSOT 보호 (lock 정책)
        if (h["status"] or "").upper() == "CONFIRMED":
            return JSONResponse(
                {"error": "수주 확정된 견적은 라인을 추가할 수 없습니다 (SSOT 보호). "
                          "변경이 필요하면 새 견적 버전을 작성해주세요."}, 400,
            )
        # 다음 line_no 채번
        row = c.execute(
            "SELECT COALESCE(MAX(line_no),0)+1 FROM quotation_items WHERE quotation_id=?",
            (quote_id,),
        ).fetchone()
        line_no = row[0] if row else 1
        cur = c.execute(
            """INSERT INTO quotation_items
               (quotation_id, line_no, part_id, item_name, qty, unit,
                unit_price, total_price, note)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (quote_id, line_no, part_id, item_name, qty, unit,
             unit_price, total_price, note or None),
        )
        # 헤더 total_amount 재합계
        s = c.execute(
            "SELECT COALESCE(SUM(total_price),0) FROM quotation_items WHERE quotation_id=?",
            (quote_id,),
        ).fetchone()
        c.execute(
            "UPDATE quotations SET total_amount=? WHERE id=?",
            (float(s[0] or 0), quote_id),
        )
        # v5H113 M10: 라인 추가 이력
        try:
            c.execute(
                "INSERT INTO quotation_history(quotation_id, changed_by, field, "
                "old_value, new_value, note) VALUES(?,?,?,?,?,?)",
                (quote_id, u.get("id"), "라인추가", "",
                 f"{item_name} · 수량 {qty} · 단가 {unit_price:,.0f}",
                 f"line_no={line_no} · 합계 {float(s[0] or 0):,.0f}")
            )
        except Exception:
            pass
    return RedirectResponse(f"/sales/quotations", 303)


@app.post("/sales/quotations/{quote_id}/items/{item_id}/delete")
async def sales_quotation_item_delete(req: Request, quote_id: int, item_id: int):
    """견적 라인 제거 + total_amount 재합계."""
    u = _s1_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    with db_session() as c:
        # v5H112: CONFIRMED 견적은 라인 삭제 거부 (SSOT 보호)
        st = c.execute("SELECT status FROM quotations WHERE id=?", (quote_id,)).fetchone()
        if st and (st["status"] or "").upper() == "CONFIRMED":
            return JSONResponse(
                {"error": "수주 확정된 견적은 라인을 삭제할 수 없습니다 (SSOT 보호)."}, 400,
            )
        # v5H113 M10: 삭제 전 라인 정보 캡처
        old_line = c.execute(
            "SELECT item_name, qty, unit_price, total_price FROM quotation_items "
            "WHERE id=? AND quotation_id=?",
            (item_id, quote_id),
        ).fetchone()
        c.execute(
            "DELETE FROM quotation_items WHERE id=? AND quotation_id=?",
            (item_id, quote_id),
        )
        s = c.execute(
            "SELECT COALESCE(SUM(total_price),0) FROM quotation_items WHERE quotation_id=?",
            (quote_id,),
        ).fetchone()
        c.execute(
            "UPDATE quotations SET total_amount=? WHERE id=?",
            (float(s[0] or 0), quote_id),
        )
        # v5H113 M10: 라인 삭제 이력
        if old_line:
            try:
                c.execute(
                    "INSERT INTO quotation_history(quotation_id, changed_by, field, "
                    "old_value, new_value, note) VALUES(?,?,?,?,?,?)",
                    (quote_id, u.get("id"), "라인삭제",
                     f"{old_line['item_name']} · 수량 {old_line['qty']} · 단가 {old_line['unit_price']:,.0f}",
                     "",
                     f"합계 {float(s[0] or 0):,.0f}")
                )
            except Exception:
                pass
    return RedirectResponse(f"/sales/quotations", 303)


@app.post("/sales/quotations/{quote_id}/convert-to-order")
async def sales_quotation_convert_to_order(req: Request, quote_id: int):
    """수주 전환 — quotations.CONFIRMED + clone_quotation_to_order 헬퍼 + history."""
    u = _s1_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    due_date = form.get("due_date") or None
    from .database import clone_quotation_to_order as _clone
    with db_session() as c:
        q = c.execute(
            "SELECT id, status FROM quotations WHERE id=?", (quote_id,)
        ).fetchone()
        if not q:
            return JSONResponse({"error": "견적 없음"}, 404)
        # 견적 상태 CONFIRMED 전환
        c.execute(
            "UPDATE quotations SET status='CONFIRMED' WHERE id=?", (quote_id,)
        )
    # 헬퍼는 자체 db_session — 위 트랜잭션 커밋 후 호출
    order_id, order_no = _clone(quote_id, due_date=due_date,
                                 created_by=u.get("id") or 0)
    if not order_id:
        return JSONResponse({"error": "수주 전환 실패"}, 500)
    # 상태 이력
    with db_session() as c:
        c.execute(
            """INSERT INTO order_status_history
               (order_id, from_status, to_status, changed_by, note)
               VALUES (?,?,?,?,?)""",
            (order_id, "DRAFT", "CONFIRMED", u.get("id"),
             "견적→수주 전환 (라인 자동 복제)"),
        )
        # v5H113 M11: customer_id NULL 폴백 — projects 연계가 있으면 backfill
        try:
            new_order = c.execute(
                "SELECT customer_id FROM orders WHERE id=?", (order_id,)
            ).fetchone()
            if new_order and not new_order["customer_id"]:
                # quotations 테이블에 project_id 가 있을 수 있음
                qrow = c.execute(
                    "SELECT customer_id FROM quotations WHERE id=?", (quote_id,)
                ).fetchone()
                fallback_cid = (qrow["customer_id"] if qrow else None)
                if fallback_cid:
                    c.execute(
                        "UPDATE orders SET customer_id=? WHERE id=?",
                        (fallback_cid, order_id)
                    )
                    c.execute(
                        """INSERT INTO order_status_history
                           (order_id, from_status, to_status, changed_by, note)
                           VALUES (?,?,?,?,?)""",
                        (order_id, "CONFIRMED", "CONFIRMED", u.get("id"),
                         f"customer_id NULL 폴백 → 견적의 customer_id={fallback_cid} 로 backfill")
                    )
                else:
                    c.execute(
                        """INSERT INTO order_status_history
                           (order_id, from_status, to_status, changed_by, note)
                           VALUES (?,?,?,?,?)""",
                        (order_id, "CONFIRMED", "CONFIRMED", u.get("id"),
                         "⚠ customer_id NULL — 견적·수주에 고객사 미지정. 수주 상세에서 지정 필요.")
                    )
        except Exception:
            pass
    return RedirectResponse("/sales/orders", 303)


@app.get("/sales/orders", response_class=HTMLResponse)
async def sales_orders_page(req: Request, biz: str = "", due_date: str = "", cal_ym: str = ""):
    """수주 탭 — orders 리스트.
    v5H166: cal_ym=YYYY-MM 으로 캘린더 표시 월 변경 (기본: 이번달)."""
    u = _s1_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    biz = (biz or "").upper().strip()
    if biz not in ("", "T", "M", "K", "C"):
        biz = ""
    with db_session() as c:
        try:
            ocols = {r[1] for r in c.execute("PRAGMA table_info(orders)").fetchall()}
        except Exception:
            ocols = set()
        # 4 탭 카운트 (안전: 실패해도 0)
        tab_counts = {"all": 0, "T": 0, "M": 0, "K": 0, "C": 0}
        try:
            tab_counts["all"] = (c.execute("SELECT COUNT(*) FROM orders").fetchone() or [0])[0]
            for div in ("T", "M", "K"):
                row = c.execute(
                    "SELECT COUNT(*) FROM orders o LEFT JOIN projects p ON p.id=o.project_id "
                    "WHERE COALESCE(p.biz_div,'')=?", (div,)
                ).fetchone()
                tab_counts[div] = row[0] if row else 0
            row = c.execute("SELECT COUNT(*) FROM consumable_orders").fetchone()
            tab_counts["C"] = row[0] if row else 0
        except Exception:
            pass

        items = []
        if biz == "C":
            # 소모품 탭 — consumable_orders 표시 (다른 스키마, 같은 표 형태로 정규화)
            try:
                crows = c.execute(
                    """SELECT co.id, co.co_no AS order_no,
                              COALESCE(cu.name, co.customer_name, '-') AS customer_name,
                              co.total_amount, co.due_date, co.status, co.order_date,
                              0 AS tax_invoice_issued,
                              NULL AS tax_invoice_no, NULL AS tax_invoice_date, NULL AS tax_invoice_note,
                              NULL AS ship_to, NULL AS unit_qty, NULL AS unit_label,
                              co.currency AS currency,
                              NULL AS project_id, co.mgmt_code AS mgmt_code,
                              NULL AS project_name, NULL AS biz_div,
                              NULL AS model_name, NULL AS po_type
                       FROM consumable_orders co
                       LEFT JOIN customers cu ON cu.id=co.customer_id
                       ORDER BY co.id DESC LIMIT 200"""
                ).fetchall()
                items = [dict(r) for r in crows]
            except Exception:
                items = []
        else:
            extra = []
            for cn in ("ship_to", "unit_qty", "unit_label", "currency"):
                if cn in ocols:
                    extra.append(f"o.{cn}")
            pj_extra = (
                ", p.id AS project_id, p.mgmt_code AS mgmt_code, "
                "p.name AS project_name, p.biz_div AS biz_div, "
                "p.model_name AS model_name, p.po_type AS po_type"
            )
            sel_extra = (", " + ", ".join(extra)) if extra else ""
            where_biz = ""
            params = []
            if biz in ("T", "M", "K"):
                where_biz = " WHERE COALESCE(p.biz_div,'')=?"
                params = [biz]
            rows = c.execute(
                f"""SELECT o.id, o.order_no, o.customer_id,
                          COALESCE(cu.name, p.customer_name, pcu.name, '-') AS customer_name,
                          o.total_amount, o.due_date, o.status,
                          o.order_date,
                          COALESCE(o.tax_invoice_issued,0) AS tax_invoice_issued,
                          o.tax_invoice_no, o.tax_invoice_date, o.tax_invoice_note
                          {sel_extra}
                          {pj_extra}
                   FROM orders o
                   LEFT JOIN customers cu  ON cu.id  = o.customer_id
                   LEFT JOIN projects p    ON p.id   = o.project_id
                   LEFT JOIN customers pcu ON pcu.id = p.customer_id
                   {where_biz}
                   ORDER BY o.id DESC LIMIT 200""",
                params
            ).fetchall()
            items = [dict(r) for r in rows]
    # v5H161 Phase 2: KPI 산출 (선택 탭 한정)
    from datetime import date as _d, timedelta as _td
    _today = _d.today()
    _today_iso = _today.isoformat()
    _week_start = (_today - _td(days=_today.weekday())).isoformat()
    _week_end = (_today + _td(days=(6 - _today.weekday()))).isoformat()
    _month_start = _today.replace(day=1).isoformat()
    if _today.month == 12:
        _next_m = _today.replace(year=_today.year + 1, month=1, day=1)
    else:
        _next_m = _today.replace(month=_today.month + 1, day=1)
    _month_end = (_next_m - _td(days=1)).isoformat()

    def _amt(it):
        try:
            return float(it.get("total_amount") or 0)
        except Exception:
            return 0.0

    kpi = {"week_count": 0, "week_amount": 0.0,
           "month_count": 0, "month_amount": 0.0,
           "wait_ship": 0, "wait_invoice": 0, "wait_payment": 0,
           "outstanding": 0.0,
           "overdue_count": 0}
    for it in items:
        st = (it.get("status") or "").upper()
        dd = (it.get("due_date") or "")[:10]
        amt = _amt(it)
        if st == "CANCELLED":
            continue
        if dd and _week_start <= dd <= _week_end:
            kpi["week_count"] += 1
            kpi["week_amount"] += amt
        if dd and _month_start <= dd <= _month_end:
            kpi["month_count"] += 1
            kpi["month_amount"] += amt
        if st == "CONFIRMED":
            kpi["wait_ship"] += 1
            kpi["outstanding"] += amt
            if dd and dd < _today_iso:
                kpi["overdue_count"] += 1
        elif st == "SHIPPED":
            kpi["wait_invoice"] += 1
            kpi["outstanding"] += amt
        elif st == "INVOICED":
            kpi["wait_payment"] += 1
            kpi["outstanding"] += amt

    # v5H202: 공휴일 상수는 모듈 최상단에서 공유 (HOLIDAYS_KR / HOLIDAYS_VN)

    # v5H162 Phase 3: 캘린더 buckets (이번달+다음달)
    cal_buckets = {}
    for it in items:
        dd = (it.get("due_date") or "")[:10]
        if not dd:
            continue
        if dd < _month_start or dd > (_next_m.replace(month=(_next_m.month % 12 + 1) if _next_m.month != 12 else 1,
                                                       year=_next_m.year if _next_m.month != 12 else _next_m.year + 1,
                                                       day=1) - _td(days=1)).isoformat():
            continue
        b = cal_buckets.setdefault(dd, {"count": 0, "amount": 0.0, "overdue": 0})
        b["count"] += 1
        b["amount"] += _amt(it)
        st = (it.get("status") or "").upper()
        if dd < _today_iso and st in ("CONFIRMED", "DRAFT"):
            b["overdue"] += 1

    def _build_month(year, month):
        import calendar as _cal
        cal = _cal.Calendar(firstweekday=6)  # Sunday-first
        weeks = []
        for week in cal.monthdatescalendar(year, month):
            row = []
            for dt in week:
                in_month = (dt.month == month)
                iso = dt.isoformat()
                b = cal_buckets.get(iso) or {"count": 0, "amount": 0.0, "overdue": 0}
                d_left = (dt - _today).days
                if b["count"] == 0:
                    cls = "empty"
                elif b["overdue"] > 0:
                    cls = "overdue"
                elif d_left < 0:
                    cls = "past"
                elif d_left <= 3:
                    cls = "d3"
                elif d_left <= 7:
                    cls = "d7"
                else:
                    cls = "future"
                hol_kr = HOLIDAYS_KR.get(iso)
                hol_vn = HOLIDAYS_VN.get(iso)
                row.append({"date": iso, "day": dt.day, "in_month": in_month,
                            "count": b["count"], "amount": b["amount"],
                            "cls": cls, "is_today": (iso == _today_iso),
                            "is_selected": (iso == due_date),
                            "hol_kr": hol_kr, "hol_vn": hol_vn,
                            "is_sun": (dt.weekday() == 6),
                            "is_sat": (dt.weekday() == 5)})
            weeks.append(row)
        return {"year": year, "month": month, "weeks": weeks}

    # v5H166: cal_ym 파라미터로 시작 월 결정 (기본: 오늘 월)
    _start_y, _start_m = _today.year, _today.month
    if cal_ym:
        try:
            _start_y, _start_m = int(cal_ym[:4]), int(cal_ym[5:7])
            if not (1 <= _start_m <= 12):
                _start_y, _start_m = _today.year, _today.month
        except Exception:
            _start_y, _start_m = _today.year, _today.month
    if _start_m == 12:
        _start2_y, _start2_m = _start_y + 1, 1
    else:
        _start2_y, _start2_m = _start_y, _start_m + 1
    if _start_m == 1:
        _prev_y, _prev_m = _start_y - 1, 12
    else:
        _prev_y, _prev_m = _start_y, _start_m - 1
    # v5H168: 한 달씩 이동 (이전/다음 = ±1개월)
    if _start_m == 12:
        _next_y, _next_m_nav = _start_y + 1, 1
    else:
        _next_y, _next_m_nav = _start_y, _start_m + 1

    cal_months = []
    try:
        cal_months = [
            _build_month(_start_y, _start_m),
            _build_month(_start2_y, _start2_m),
        ]
    except Exception:
        cal_months = []
    cal_nav = {
        "prev": f"{_prev_y:04d}-{_prev_m:02d}",
        "next": f"{_next_y:04d}-{_next_m_nav:02d}",
        "today": f"{_today.year:04d}-{_today.month:02d}",
        "current": f"{_start_y:04d}-{_start_m:02d}",
        "is_today_view": (_start_y == _today.year and _start_m == _today.month),
    }

    # v5H164: 임박 납기 리스트 (필터 적용 전 — 전체 기준)
    upcoming = []
    try:
        for it in items:
            dd = (it.get("due_date") or "")[:10]
            st = (it.get("status") or "").upper()
            if not dd or st in ("PAID", "CANCELLED"):
                continue
            from datetime import datetime as _dt
            try:
                ddate = _dt.strptime(dd, "%Y-%m-%d").date()
                d_left = (ddate - _today).days
            except Exception:
                continue
            # 표시 대상: 오버듀(미출하) + 60일 이내
            if st in ("CONFIRMED", "DRAFT") and d_left < 0:
                cls = "overdue"
            elif d_left < 0:
                continue  # 출하 완료된 지난 건 제외
            elif d_left <= 3:
                cls = "d3"
            elif d_left <= 7:
                cls = "d7"
            elif d_left <= 30:
                cls = "d30"
            elif d_left <= 60:
                cls = "future"
            else:
                continue
            upcoming.append({**it, "d_left": d_left, "_cls": cls})
        # 정렬: 오버듀 먼저(가장 오래된), 그 다음 임박순
        upcoming.sort(key=lambda x: (0 if x["d_left"] < 0 else 1, x["d_left"]))
        upcoming = upcoming[:60]  # v5H168: 상위 60건 (리스트 더 많이 표시)
    except Exception:
        upcoming = []

    # 필터 적용 (캘린더 날짜 클릭 시)
    if due_date:
        items = [it for it in items if (it.get("due_date") or "")[:10] == due_date[:10]]

    return ctx(req, "sales_orders.html", user=u, active="sales_orders",
               tab="orders", items=items,
               biz=biz, tab_counts=tab_counts, is_consumable=(biz == "C"),
               kpi=kpi, today_iso=_today_iso,
               cal_months=cal_months, due_date=due_date,
               upcoming=upcoming, cal_nav=cal_nav)


@app.post("/sales/orders")
async def sales_orders_confirm(req: Request):
    """수주 확정 (Top3-S1-2차 — quotation.CONFIRMED + orders INSERT + history)"""
    u = _s1_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    quote_id = form.get("quote_id") or None
    due_date = form.get("due_date") or None
    if not quote_id:
        return JSONResponse({"error": "quote_id 누락"}, 400)
    with db_session() as c:
        q = c.execute(
            "SELECT customer_id, total_amount FROM quotations WHERE id=?",
            (quote_id,),
        ).fetchone()
        if not q:
            return JSONResponse({"error": "견적 없음"}, 404)
        # 견적 상태 CONFIRMED 로 전환
        c.execute(
            "UPDATE quotations SET status='CONFIRMED' WHERE id=?", (quote_id,)
        )
        # 수주 헤더 INSERT (status=CONFIRMED)
        ym = datetime.now().strftime("%Y%m")
        row = c.execute(
            "SELECT COUNT(*) FROM orders WHERE order_no LIKE ?",
            (f"SO-{ym}-%",),
        ).fetchone()
        seq = (row[0] if row else 0) + 1
        order_no = f"SO-{ym}-{seq:04d}"
        cur = c.execute(
            """INSERT INTO orders
               (order_no, quote_id, customer_id, order_date, due_date,
                total_amount, status, created_by)
               VALUES (?,?,?,?,?,?,'CONFIRMED',?)""",
            (order_no, quote_id, q[0], date.today().isoformat(),
             due_date, q[1] or 0, u.get("id")),
        )
        order_id = cur.lastrowid
        # 상태 이력 (DRAFT → CONFIRMED)
        c.execute(
            """INSERT INTO order_status_history
               (order_id, from_status, to_status, changed_by, note)
               VALUES (?,?,?,?,?)""",
            (order_id, "DRAFT", "CONFIRMED", u.get("id"), "견적→수주 전환"),
        )
        return JSONResponse({"ok": True, "order_id": order_id, "order_no": order_no})


@app.post("/sales/orders/{order_id}/tax-invoice")
async def sales_order_tax_invoice_update(order_id: int, req: Request):
    """세금계산서 발행여부 체크 (사이클 78 DEC-1)
    KNK 내부 회계 시스템에서 실제 발행 후, HAIST WORKS 는 발행 여부만 기록.
    권한: admin/ceo 또는 영업 (_s1_guard)."""
    u = _s1_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    issued_raw = (form.get("tax_invoice_issued") or "").strip().lower()
    issued = 1 if issued_raw in ("1", "on", "true", "yes") else 0
    inv_no = (form.get("tax_invoice_no") or "").strip() or None
    inv_date = (form.get("tax_invoice_date") or "").strip() or None
    inv_note = (form.get("tax_invoice_note") or "").strip() or None
    with db_session() as c:
        o = c.execute("SELECT id FROM orders WHERE id=?", (order_id,)).fetchone()
        if not o:
            return JSONResponse({"error": "수주 없음"}, 404)
        c.execute(
            """UPDATE orders
               SET tax_invoice_issued = ?,
                   tax_invoice_no     = ?,
                   tax_invoice_date   = ?,
                   tax_invoice_note   = ?
               WHERE id = ?""",
            (issued, inv_no, inv_date, inv_note, order_id),
        )
    return RedirectResponse("/sales/orders", 303)


@app.get("/sales/production", response_class=HTMLResponse)
async def sales_production_page(req: Request):
    """생산지시 탭 (시안 §1 탭3 WO) — production_orders 리스트 + 진행률"""
    u = _s1_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        rows = c.execute(
            """SELECT p.id, p.order_id, o.order_no,
                      p.planned_start, p.planned_end,
                      p.actual_start, p.actual_end, p.status
               FROM production_orders p
               LEFT JOIN orders o ON o.id = p.order_id
               ORDER BY p.id DESC LIMIT 200"""
        ).fetchall()
        items = [dict(r) for r in rows]
    return ctx(req, "sales_production.html", user=u, active="sales_production",
               tab="production", items=items)


@app.post("/sales/production/start")
async def sales_production_start(req: Request):
    """생산 시작 (Top3-S1-2차 — orders.IN_PRODUCTION + production_orders + history)"""
    u = _s1_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    order_id = form.get("order_id") or None
    planned_start = form.get("planned_start") or date.today().isoformat()
    planned_end = form.get("planned_end") or None
    if not order_id:
        return JSONResponse({"error": "order_id 누락"}, 400)
    with db_session() as c:
        o = c.execute("SELECT status FROM orders WHERE id=?", (order_id,)).fetchone()
        if not o:
            return JSONResponse({"error": "수주 없음"}, 404)
        prev_status = o[0]
        c.execute(
            "UPDATE orders SET status='IN_PRODUCTION' WHERE id=?", (order_id,)
        )
        cur = c.execute(
            """INSERT INTO production_orders
               (order_id, planned_start, planned_end, actual_start, status)
               VALUES (?,?,?,?,'IN_PRODUCTION')""",
            (order_id, planned_start, planned_end, datetime.now().isoformat(timespec="seconds")),
        )
        c.execute(
            """INSERT INTO order_status_history
               (order_id, from_status, to_status, changed_by, note)
               VALUES (?,?,?,?,?)""",
            (order_id, prev_status, "IN_PRODUCTION", u.get("id"), "생산 시작"),
        )
        return JSONResponse({"ok": True, "production_id": cur.lastrowid})


@app.get("/sales/shipments-receipts", response_class=HTMLResponse)
async def sales_shipments_receipts_page(req: Request):
    """출하·수금 탭 (시안 §1 탭4 DO+INV+RC) — shipments + receipts_payment 통합 라인"""
    u = _s1_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        # 수주별 통합 라인: 출하 합계 + 수금 합계 + 세금계산서 발행 여부
        # v5H124: currency 인지 — 수주 통화와 동일 통화 수금만 합산해 paid_total 산출
        try:
            rows = c.execute(
                """SELECT o.id AS order_id, o.order_no, o.total_amount, o.status,
                          COALESCE(cu.name,'-') AS customer_name,
                          COALESCE(o.currency,'KRW') AS currency,
                          (SELECT COALESCE(SUM(s.shipped_qty),0)
                             FROM shipments s WHERE s.order_id = o.id) AS shipped_qty_sum,
                          (SELECT COALESCE(SUM(r.amount),0)
                             FROM receipts_payment r
                             WHERE r.order_id = o.id
                               AND COALESCE(r.currency,'KRW')=COALESCE(o.currency,'KRW')
                          ) AS paid_total,
                          (SELECT COUNT(*) FROM invoices i
                             WHERE i.order_id = o.id AND i.status='ISSUED') AS invoice_issued
                   FROM orders o
                   LEFT JOIN customers cu ON cu.id = o.customer_id
                   WHERE o.status IN ('IN_PRODUCTION','READY_TO_SHIP','SHIPPED','INVOICED','PAID')
                   ORDER BY o.id DESC LIMIT 200"""
            ).fetchall()
        except Exception:
            rows = c.execute(
                """SELECT o.id AS order_id, o.order_no, o.total_amount, o.status,
                          COALESCE(cu.name,'-') AS customer_name,
                          'KRW' AS currency,
                          (SELECT COALESCE(SUM(s.shipped_qty),0)
                             FROM shipments s WHERE s.order_id = o.id) AS shipped_qty_sum,
                          (SELECT COALESCE(SUM(r.amount),0)
                             FROM receipts_payment r WHERE r.order_id = o.id) AS paid_total,
                          (SELECT COUNT(*) FROM invoices i
                             WHERE i.order_id = o.id AND i.status='ISSUED') AS invoice_issued
                   FROM orders o
                   LEFT JOIN customers cu ON cu.id = o.customer_id
                   WHERE o.status IN ('IN_PRODUCTION','READY_TO_SHIP','SHIPPED','INVOICED','PAID')
                   ORDER BY o.id DESC LIMIT 200"""
            ).fetchall()
        items = [dict(r) for r in rows]
    return ctx(req, "sales_shipments_receipts.html", user=u, active="sales_shipments",
               tab="shipments", items=items)


@app.post("/sales/shipments")
async def sales_shipments_create(req: Request):
    """출하 등록 (Top3-S1-2차 — shipments INSERT + orders.SHIPPED + history · 1:N 부분출하)"""
    u = _s1_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    order_id = form.get("order_id") or None
    shipped_qty = float(form.get("shipped_qty") or 0)
    tracking = form.get("tracking") or None
    if not order_id:
        return JSONResponse({"error": "order_id 누락"}, 400)
    # v5H112: 출하 수량 검증
    if shipped_qty <= 0:
        return JSONResponse(
            {"error": f"출하 수량은 0보다 커야 합니다 (입력값: {shipped_qty})."}, 400
        )
    with db_session() as c:
        o = c.execute("SELECT status FROM orders WHERE id=?", (order_id,)).fetchone()
        if not o:
            return JSONResponse({"error": "수주 없음"}, 404)
        prev_status = o[0]
        # v5H112: 출하 OVER 차단 — 누적 출하 합계 vs 수주 수량 (orders.unit_qty 또는 order_items SUM)
        try:
            ord_qty_row = c.execute(
                "SELECT COALESCE(unit_qty, "
                "(SELECT SUM(quantity) FROM order_items WHERE order_id=?), 0) AS oq "
                "FROM orders WHERE id=?",
                (order_id, order_id),
            ).fetchone()
            ord_qty = (ord_qty_row[0] or 0) if ord_qty_row else 0
            shipped_so_far_row = c.execute(
                "SELECT COALESCE(SUM(shipped_qty),0) FROM shipments WHERE order_id=?",
                (order_id,),
            ).fetchone()
            shipped_so_far = shipped_so_far_row[0] or 0
            if ord_qty > 0 and (shipped_so_far + shipped_qty) > ord_qty + 0.0001:
                return JSONResponse(
                    {"error": (
                        f"출하 수량 초과 — 수주 {ord_qty} / 기존 출하 {shipped_so_far} / "
                        f"이번 출하 {shipped_qty} = 누적 {shipped_so_far + shipped_qty}. "
                        "수주 수량 내로 입력해주세요."
                    )}, 400,
                )
        except Exception:
            pass  # 검증 실패 시 통과 (backward compatible)
        cur = c.execute(
            """INSERT INTO shipments
               (order_id, shipped_at, shipped_qty, shipped_by, tracking)
               VALUES (?,?,?,?,?)""",
            (order_id, datetime.now().isoformat(timespec="seconds"),
             shipped_qty, u.get("id"), tracking),
        )
        # SHIPPED 으로 전환 (READY_TO_SHIP 또는 IN_PRODUCTION → SHIPPED)
        if prev_status != "SHIPPED":
            c.execute("UPDATE orders SET status='SHIPPED' WHERE id=?", (order_id,))
            c.execute(
                """INSERT INTO order_status_history
                   (order_id, from_status, to_status, changed_by, note)
                   VALUES (?,?,?,?,?)""",
                (order_id, prev_status, "SHIPPED", u.get("id"),
                 f"출하 등록 (수량 {shipped_qty})"),
            )
        _ship_id = cur.lastrowid
    # 알림시스템 통합 (사이클 2026-04-26) — 출하 담당자에게 SALES 알림 (1시간 중복 방지 내장)
    notify_user(
        u.get("id"), "SALES",
        f"🚚 출하 등록 — 수주 {order_id}",
        body=f"수량 {shipped_qty} / 송장 {tracking or '-'}",
        link=f"/sales/orders/{order_id}",
    )
    return JSONResponse({"ok": True, "shipment_id": _ship_id})


@app.post("/sales/receipts")
async def sales_receipts_create(req: Request):
    """수금 등록 (Top3-S1-2차 — receipts_payment INSERT + 합계 비교 → PAID/유지 + history)
    PARTIAL_RECEIPT 별도 enum 폐기 — 합계 < 수주금액이면 SHIPPED 유지, 합계 >= 이면 PAID."""
    u = _s1_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    order_id = form.get("order_id") or None
    amount = float(form.get("amount") or 0)
    method = form.get("method") or None
    note = form.get("note") or None
    # v5H124: 통화 입력 (화이트리스트, 미입력 시 KRW)
    _ALLOWED_CCY = {"KRW", "USD", "VND", "JPY", "CNY", "EUR"}
    currency = (form.get("currency") or "KRW").strip().upper()
    if currency not in _ALLOWED_CCY:
        return JSONResponse(
            {"error": f"허용되지 않는 통화: {currency}. (KRW/USD/VND/JPY/CNY/EUR 중 선택)"},
            400,
        )
    fx_rate_raw = form.get("fx_rate")
    fx_rate = None
    if fx_rate_raw:
        try:
            fx_rate = float(fx_rate_raw)
            if fx_rate <= 0:
                fx_rate = None
        except Exception:
            fx_rate = None
    # v5H126: fx_rate 미입력 + 비KRW 통화 → exchange_rates 에서 자동 snapshot
    if fx_rate is None and currency != "KRW":
        try:
            from .database import _get_active_fx_rate as _gafx
            _auto = _gafx(currency, ref_date=date.today().isoformat())
            if _auto and _auto > 0:
                fx_rate = _auto
        except Exception:
            pass
    if not order_id:
        return JSONResponse({"error": "order_id 누락"}, 400)
    # v5H112: 음수/0 수금 차단 — 친절한 에러
    if amount <= 0:
        return JSONResponse(
            {"error": f"수금액은 0보다 커야 합니다 (입력값: {amount}). 양수로 입력해주세요."},
            400,
        )
    with db_session() as c:
        # v5H124: orders.currency 함께 조회 (수주 통화 vs 수금 통화 mismatch 검출)
        try:
            o = c.execute(
                "SELECT status, total_amount, COALESCE(currency,'KRW') FROM orders WHERE id=?",
                (order_id,),
            ).fetchone()
        except Exception:
            o = c.execute(
                "SELECT status, total_amount, 'KRW' FROM orders WHERE id=?", (order_id,)
            ).fetchone()
        if not o:
            return JSONResponse({"error": "수주 없음"}, 404)
        prev_status, total, order_ccy = o[0], (o[1] or 0), (o[2] or "KRW")
        # 통화 mismatch 시 audit 경고 (차단 X — 환차 발생 가능, FYI)
        ccy_warn = None
        if currency != order_ccy:
            ccy_warn = (
                f"⚠ 통화 불일치 — 수주 {order_ccy} vs 수금 {currency} "
                f"(환산 누락 시 미수금 계산 오차 위험)"
            )
        # v5H124: receipts_payment 에 currency / fx_rate 가 있으면 함께 INSERT, 없으면 백워드
        rpcols2 = set()
        try:
            rpcols2 = {r[1] for r in c.execute("PRAGMA table_info(receipts_payment)").fetchall()}
        except Exception:
            pass
        if "currency" in rpcols2 and "fx_rate" in rpcols2:
            cur = c.execute(
                """INSERT INTO receipts_payment
                   (order_id, received_at, amount, method, received_by, note, currency, fx_rate)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (order_id, datetime.now().isoformat(timespec="seconds"),
                 amount, method, u.get("id"), note, currency, fx_rate),
            )
        else:
            cur = c.execute(
                """INSERT INTO receipts_payment
                   (order_id, received_at, amount, method, received_by, note)
                   VALUES (?,?,?,?,?,?)""",
                (order_id, datetime.now().isoformat(timespec="seconds"),
                 amount, method, u.get("id"), note),
            )
        # v5H124: doc_audit_log — 수금 등록 + 통화 mismatch 경고 합성
        try:
            _audit_note = f"amount={amount} {currency}"
            if ccy_warn:
                _audit_note += f" | {ccy_warn}"
            _doc_audit_log(c, "receipts_payment", cur.lastrowid, "RECEIVE",
                           u.get("id"), _audit_note)
        except Exception:
            pass
        # 누적 수금 합계 → PAID 분기 (v5H124 — 수주 통화와 동일 통화만 합산해 PAID 판정)
        if "currency" in rpcols2:
            row = c.execute(
                """SELECT COALESCE(SUM(amount),0) FROM receipts_payment
                   WHERE order_id=? AND COALESCE(currency,'KRW')=?""",
                (order_id, order_ccy),
            ).fetchone()
        else:
            row = c.execute(
                "SELECT COALESCE(SUM(amount),0) FROM receipts_payment WHERE order_id=?",
                (order_id,),
            ).fetchone()
        paid_total = row[0] or 0
        # 통화가 같을 때만 PAID 자동 전환 (이종 통화는 수동 정산 권장)
        new_status = (
            "PAID" if (currency == order_ccy and paid_total >= total and total > 0)
            else prev_status
        )
        if new_status != prev_status:
            c.execute(
                "UPDATE orders SET status=? WHERE id=?", (new_status, order_id)
            )
            c.execute(
                """INSERT INTO order_status_history
                   (order_id, from_status, to_status, changed_by, note)
                   VALUES (?,?,?,?,?)""",
                (order_id, prev_status, new_status, u.get("id"),
                 f"수금 누적 {paid_total}/{total}"),
            )
        return JSONResponse({
            "ok": True, "receipt_id": cur.lastrowid,
            "paid_total": paid_total, "status": new_status,
            "currency": currency, "order_currency": order_ccy,
            "warning": ccy_warn,
        })


# =====================================================
# Top3 S1 3차 — 매출 대시 강화 + 매출 예측 (2026-04-26)
# 외부 차트·numpy 0건. _linear_regression (line 6572) 재사용.
# G1~G5 핫패치 보존, v2 본체 무수정.
# =====================================================

def _sales_monthly_series(c, months: int = 12):
    """최근 N개월 매출 시계열 (orders.order_date + total_amount).
    반환: [{ym, total, cnt}, ...] (오름차순). 빈 달은 0 채움.
    v5H123 (2026-05-04): 통화 혼합 집계 결함 수정 — currency 컬럼 존재 시 KRW 만 합산.
    USD/VND 등은 환율 미반영 단순 합산이 의미 없으므로 KRW 단독 기준이 정확. 폴백: 컬럼 미존재 시 기존 동작."""
    today = date.today()
    out = []
    # currency 컬럼 존재 여부 확인 (v5H92 이후)
    has_currency = False
    try:
        cols = [r[1] for r in c.execute("PRAGMA table_info(orders)").fetchall()]
        has_currency = "currency" in cols
    except Exception:
        pass
    for i in range(months - 1, -1, -1):
        y = today.year
        m = today.month - i
        while m <= 0:
            m += 12
            y -= 1
        ym = f"{y:04d}-{m:02d}"
        if has_currency:
            row = c.execute(
                """SELECT COALESCE(SUM(total_amount),0) AS total, COUNT(*) AS cnt
                   FROM orders WHERE order_date LIKE ?
                     AND status NOT IN ('CANCELLED','DRAFT')
                     AND COALESCE(currency,'KRW')='KRW'""",
                (f"{ym}%",),
            ).fetchone()
        else:
            row = c.execute(
                """SELECT COALESCE(SUM(total_amount),0) AS total, COUNT(*) AS cnt
                   FROM orders WHERE order_date LIKE ?
                     AND status NOT IN ('CANCELLED','DRAFT')""",
                (f"{ym}%",),
            ).fetchone()
        out.append({"ym": ym, "total": row[0] or 0, "cnt": row[1] or 0})
    return out


def _sales_forecast(series, horizon: int = 3):
    """선형회귀 → 다음 N개월 예측. _linear_regression 재사용 (numpy 0).
    반환: {points:[{ym,total,is_pred}], slope, intercept, r2, horizon, end_ym}."""
    if not series or len(series) < 2:
        return None
    xs = list(range(len(series)))
    ys = [float(s["total"]) for s in series]
    slope, intercept, r2 = _linear_regression(xs, ys)
    pts = [{"ym": s["ym"], "total": s["total"], "is_pred": False} for s in series]
    last_ym = series[-1]["ym"]
    ly, lm = int(last_ym[:4]), int(last_ym[5:7])
    for k in range(1, horizon + 1):
        lm += 1
        if lm > 12:
            lm -= 12
            ly += 1
        x = len(series) - 1 + k
        pred = max(0.0, slope * x + intercept)
        pts.append({"ym": f"{ly:04d}-{lm:02d}", "total": pred, "is_pred": True})
    return {"points": pts, "slope": slope, "intercept": intercept,
            "r_squared": r2, "horizon": horizon, "end_ym": pts[-1]["ym"],
            "end_total": pts[-1]["total"], "sample_n": len(series)}


def _sales_dashboard_ctx(c):
    """대시보드 + 예측 공통 컨텍스트 (KPI 8 + 차트 데이터 + 파이프라인).
    v5H123 (2026-05-04): 통화 혼합 집계 결함 수정 — orders.currency 컬럼 존재 시
    모든 SUM(total_amount) 쿼리를 KRW 단독으로 제한. 외화(USD/VND/JPY 등)는
    환율 미반영 단순 합산이 잘못된 KPI 를 만들기 때문 (대표 보고: 수주액 정합성).
    by_currency 분포는 별도 ctx 로 전달해 템플릿에서 펼침 가능."""
    today = date.today()
    ym = today.strftime("%Y-%m")
    # 전월
    py, pm = today.year, today.month - 1
    if pm <= 0:
        pm += 12; py -= 1
    prev_ym = f"{py:04d}-{pm:02d}"
    # currency 컬럼 존재 여부 (v5H92 이후)
    has_currency = False
    try:
        cols = [r[1] for r in c.execute("PRAGMA table_info(orders)").fetchall()]
        has_currency = "currency" in cols
    except Exception:
        pass
    krw_filter = " AND COALESCE(currency,'KRW')='KRW'" if has_currency else ""
    month_total = c.execute(
        f"""SELECT COALESCE(SUM(total_amount),0) FROM orders
           WHERE order_date LIKE ? AND status NOT IN ('CANCELLED','DRAFT'){krw_filter}""",
        (f"{ym}%",),
    ).fetchone()[0] or 0
    prev_total = c.execute(
        f"""SELECT COALESCE(SUM(total_amount),0) FROM orders
           WHERE order_date LIKE ? AND status NOT IN ('CANCELLED','DRAFT'){krw_filter}""",
        (f"{prev_ym}%",),
    ).fetchone()[0] or 0
    mom = ((month_total - prev_total) / prev_total * 100.0) if prev_total > 0 else 0.0
    # 통화별 분포 (이번달) — 다통화 거래 가시화
    by_currency = []
    if has_currency:
        try:
            by_currency = [dict(r) for r in c.execute(
                """SELECT COALESCE(currency,'KRW') AS currency,
                          COALESCE(SUM(total_amount),0) AS total,
                          COUNT(*) AS cnt
                   FROM orders WHERE order_date LIKE ?
                     AND status NOT IN ('CANCELLED','DRAFT')
                   GROUP BY COALESCE(currency,'KRW') ORDER BY total DESC""",
                (f"{ym}%",),
            ).fetchall()]
        except Exception:
            by_currency = []
    # 수금률 = 수금 합계 / INVOICED+PAID 수주 합계 (KRW 한정)
    inv_total = c.execute(
        f"""SELECT COALESCE(SUM(total_amount),0) FROM orders
           WHERE status IN ('INVOICED','PAID','SHIPPED'){krw_filter}"""
    ).fetchone()[0] or 0
    rcv_total = c.execute(
        "SELECT COALESCE(SUM(amount),0) FROM receipts_payment"
    ).fetchone()[0] or 0
    rcv_rate = (rcv_total / inv_total * 100.0) if inv_total > 0 else 0.0
    # 평균 결제 일수 (issue_date → 첫 수금)
    avg_days = 0.0
    rows = c.execute(
        """SELECT i.order_id, MIN(i.issue_date) AS iss,
                  MIN(rp.received_at) AS rcv
           FROM invoices i
           LEFT JOIN receipts_payment rp ON rp.order_id = i.order_id
           WHERE i.issue_date IS NOT NULL AND rp.received_at IS NOT NULL
           GROUP BY i.order_id LIMIT 200"""
    ).fetchall()
    if rows:
        days_list = []
        for r in rows:
            try:
                d1 = datetime.strptime(r[1][:10], "%Y-%m-%d")
                d2 = datetime.strptime(r[2][:10], "%Y-%m-%d")
                days_list.append((d2 - d1).days)
            except Exception:
                pass
        if days_list:
            avg_days = sum(days_list) / len(days_list)
    # 미수금 = INVOICED 수주 합 - 수금
    unpaid = max(0.0, inv_total - rcv_total)
    active_orders = c.execute(
        """SELECT COUNT(*) FROM orders
           WHERE status IN ('CONFIRMED','IN_PRODUCTION','READY_TO_SHIP')"""
    ).fetchone()[0] or 0
    in_prod = c.execute(
        "SELECT COUNT(*) FROM production_orders WHERE status='IN_PRODUCTION'"
    ).fetchone()[0] or 0
    # 출하 임박 = due_date 7일 이내 + status IN_PRODUCTION/READY_TO_SHIP
    soon_end = (today + timedelta(days=7)).isoformat()
    ship_soon = c.execute(
        """SELECT COUNT(*) FROM orders
           WHERE due_date IS NOT NULL AND due_date <= ? AND due_date >= ?
             AND status IN ('IN_PRODUCTION','READY_TO_SHIP','CONFIRMED')""",
        (soon_end, today.isoformat()),
    ).fetchone()[0] or 0
    # 파이프라인 9 status 분포
    pipeline = {s: 0 for s in ["DRAFT", "QUOTED", "CONFIRMED", "IN_PRODUCTION",
                                "READY_TO_SHIP", "SHIPPED", "INVOICED", "PAID", "CANCELLED"]}
    for r in c.execute(
        "SELECT status, COUNT(*) AS cnt FROM orders GROUP BY status"
    ).fetchall():
        if r[0] in pipeline:
            pipeline[r[0]] = r[1]
    # 거래처 Top 5 (Pareto) — v5H123 KRW 단독 (통화 혼합 방지)
    if has_currency:
        top_customers = [dict(r) for r in c.execute(
            """SELECT COALESCE(cu.name,'-') AS name,
                      COUNT(*) AS cnt, COALESCE(SUM(o.total_amount),0) AS total
               FROM orders o LEFT JOIN customers cu ON cu.id = o.customer_id
               WHERE o.status NOT IN ('CANCELLED','DRAFT')
                 AND COALESCE(o.currency,'KRW')='KRW'
               GROUP BY o.customer_id ORDER BY total DESC LIMIT 5"""
        ).fetchall()]
    else:
        top_customers = [dict(r) for r in c.execute(
            """SELECT COALESCE(cu.name,'-') AS name,
                      COUNT(*) AS cnt, COALESCE(SUM(o.total_amount),0) AS total
               FROM orders o LEFT JOIN customers cu ON cu.id = o.customer_id
               WHERE o.status NOT IN ('CANCELLED','DRAFT')
               GROUP BY o.customer_id ORDER BY total DESC LIMIT 5"""
        ).fetchall()]
    grand_total = sum(t["total"] for t in top_customers) or 1
    cum = 0.0
    for t in top_customers:
        cum += t["total"]
        t["pct"] = (t["total"] / grand_total * 100.0)
        t["cum_pct"] = (cum / grand_total * 100.0)
    # v5H126: 통화별 미수금 (수주 INVOICED+SHIPPED+PAID - 수금) — 외화 가시화
    outstanding_by_currency = []
    if has_currency:
        try:
            # receipts_payment.currency 컬럼 존재 검사
            rpcols = {r[1] for r in c.execute("PRAGMA table_info(receipts_payment)").fetchall()}
            rcv_ccy = "COALESCE(currency,'KRW')" if "currency" in rpcols else "'KRW'"
            inv_rows = c.execute(
                """SELECT COALESCE(currency,'KRW') AS ccy,
                          COALESCE(SUM(total_amount),0) AS inv
                   FROM orders
                   WHERE status IN ('INVOICED','PAID','SHIPPED')
                   GROUP BY COALESCE(currency,'KRW')"""
            ).fetchall()
            rcv_rows = c.execute(
                f"""SELECT {rcv_ccy} AS ccy,
                           COALESCE(SUM(amount),0) AS rcv
                    FROM receipts_payment
                    GROUP BY {rcv_ccy}"""
            ).fetchall()
            _rcv_map = {r[0]: float(r[1] or 0) for r in rcv_rows}
            for r in inv_rows:
                ccy = r[0]
                inv = float(r[1] or 0)
                rcv = _rcv_map.get(ccy, 0.0)
                out = max(0.0, inv - rcv)
                if inv > 0 or rcv > 0:
                    outstanding_by_currency.append({
                        "currency": ccy, "invoiced": inv,
                        "received": rcv, "outstanding": out,
                    })
            outstanding_by_currency.sort(key=lambda x: -x["outstanding"])
        except Exception:
            outstanding_by_currency = []
    series = _sales_monthly_series(c, 12)
    chart_max = max((s["total"] for s in series), default=1) or 1
    return {
        "kpi": {
            "month_total": month_total, "mom": mom,
            "rcv_rate": rcv_rate, "avg_days": avg_days,
            "unpaid": unpaid, "active_orders": active_orders,
            "in_prod": in_prod, "ship_soon": ship_soon,
        },
        "ym": ym, "prev_ym": prev_ym,
        "pipeline": pipeline,
        "top_customers": top_customers,
        "series": series, "chart_max": chart_max,
        "by_currency": by_currency,  # v5H123 통화별 분포 (KRW 외 외화 노출용)
        "outstanding_by_currency": outstanding_by_currency,  # v5H126 통화별 미수금
        "primary_currency": "KRW",   # v5H123 KPI 통화 기준 명시
    }


@app.get("/sales/dashboard", response_class=HTMLResponse)
async def sales_dashboard_v3(req: Request):
    """Top3-S1-3차 강화 대시 — KPI 8 + 월별 차트 + Pareto Top5 + 파이프라인 9."""
    u = _s1_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        ctx_data = _sales_dashboard_ctx(c)
    # v5H49: 템플릿이 평면 키(kpi_month_total 등) 기대 → 평면화
    k = ctx_data.get("kpi", {})
    flat = {
        "kpi_month_total":  k.get("month_total", 0),
        "kpi_month_cnt":    k.get("active_orders", 0),
        "kpi_ytd_total":    sum(s["total"] for s in ctx_data.get("series", [])),
        "kpi_ytd_cnt":      sum(s["cnt"] for s in ctx_data.get("series", [])),
        "kpi_avg_amount":   round(k.get("month_total", 0) / max(1, k.get("active_orders", 1))),
        "kpi_outstanding":  k.get("unpaid", 0),
        "kpi_mom":          k.get("mom", 0),
        "kpi_rcv_rate":     k.get("rcv_rate", 0),
        "kpi_avg_days":     k.get("avg_days", 0),
        "kpi_in_prod":      k.get("in_prod", 0),
        "kpi_ship_soon":    k.get("ship_soon", 0),
    }
    ctx_data.update(flat)
    # v5H51: pareto_top5 / pipeline 형식 정합 (템플릿 정합)
    tcs = ctx_data.get("top_customers", [])
    ctx_data["pareto_top5"] = [{
        "customer_name": tc.get("name", "-"),
        "total":  tc.get("total", 0),
        "cnt":    tc.get("cnt", 0),
        "share":  tc.get("pct", 0),
    } for tc in tcs[:5]]
    pipe_dict = ctx_data.get("pipeline", {})
    if isinstance(pipe_dict, dict):
        ctx_data["pipeline"] = [
            {"stage": stage, "cnt": (info.get("cnt", 0) if isinstance(info, dict) else int(info or 0)),
             "amount": (info.get("amount", 0) if isinstance(info, dict) else 0)}
            for stage, info in pipe_dict.items()
        ]
    return ctx(req, "sales_dashboard.html", user=u, active="sales_dashboard",
               tab="dashboard", **ctx_data)


@app.get("/sales/forecast", response_class=HTMLResponse)
async def sales_forecast_page(req: Request):
    """Top3-S1-3차 매출 예측 — 최근 12개월 → 향후 3개월 (선형회귀, R²)."""
    u = _s1_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        series = _sales_monthly_series(c, 12)
    fc_row = None  # 매출 전용 저장 테이블은 다음 사이클 (read-only)
    forecast = _sales_forecast(series, horizon=3)
    chart_max = max(
        max((s["total"] for s in series), default=1),
        max((p["total"] for p in (forecast["points"] if forecast else [])), default=1),
    ) or 1
    return ctx(req, "sales_forecast.html", user=u, active="sales",
               tab="forecast", series=series, forecast=forecast,
               chart_max=chart_max, saved_forecast=fc_row)


@app.post("/sales/forecast/refresh")
async def sales_forecast_refresh(req: Request):
    """Top3-S1-3차 예측 재계산 트리거 — JSON {ok, end_ym, end_total, r2, sample_n}."""
    u = _s1_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    with db_session() as c:
        series = _sales_monthly_series(c, 12)
    fc = _sales_forecast(series, horizon=3)
    if not fc:
        return JSONResponse({"error": "insufficient_data",
                              "sample_n": len(series)}, status_code=400)
    return JSONResponse({
        "ok": True, "end_ym": fc["end_ym"],
        "end_total": fc["end_total"], "r_squared": fc["r_squared"],
        "slope": fc["slope"], "sample_n": fc["sample_n"],
    })


# =====================================================
# Top3 S1 4차 — 미수금 추적 + 수금 알림 자동 (2026-04-26)
# 헬퍼 _outstanding_receivables / check_receivable_alerts.
# 알림 통합 (사이클 2026-04-26 notify_user SALES) 활용 · 1시간 중복 방지 내장.
# G1~G5 핫패치 보존 · v2 본체 무수정 · 외부 자산 0건.
# =====================================================

def _parse_terms_days(terms: str) -> int:
    """payment_terms 문자열 → 일수. NET30/30일/선금 등 휴리스틱.
    매칭 실패 시 30 (기본 NET30)."""
    if not terms:
        return 30
    s = str(terms).upper().replace(" ", "")
    # NET#, #일, #DAYS
    import re
    m = re.search(r"(\d{1,3})", s)
    if m:
        try:
            d = int(m.group(1))
            if 0 < d <= 365:
                return d
        except Exception:
            pass
    if "선금" in str(terms) or "CASH" in s or "현금" in str(terms):
        return 0
    return 30


def _outstanding_receivables(c, only_overdue: bool = False):
    """미수금 건별 집계 — orders.total_amount - SUM(receipts_payment.amount).
    연체일 = today - (order_date + payment_terms.terms days).
    등급: CURRENT(미만기) / D-30 / D-60 / D-90+ (연체일 기준).
    Returns: list of dicts (overdue desc).
    """
    today = date.today()
    rows = c.execute(
        """SELECT o.id AS order_id, o.order_no, o.order_date, o.due_date,
                  o.total_amount, o.status, o.customer_id,
                  COALESCE(cu.name,'-') AS customer_name,
                  COALESCE((SELECT SUM(amount) FROM receipts_payment rp
                            WHERE rp.order_id=o.id), 0) AS paid_total,
                  COALESCE((SELECT terms FROM payment_terms pt
                            WHERE pt.customer_id=o.customer_id
                            ORDER BY pt.id DESC LIMIT 1), '') AS terms
           FROM orders o
           LEFT JOIN customers cu ON cu.id = o.customer_id
           WHERE o.status IN ('SHIPPED','INVOICED')
             AND o.total_amount > 0"""
    ).fetchall()
    out = []
    for r in rows:
        d = dict(r)
        outstanding = (d["total_amount"] or 0) - (d["paid_total"] or 0)
        if outstanding <= 0:
            continue
        # 만기일: order_date + terms일
        days_terms = _parse_terms_days(d["terms"])
        try:
            od = datetime.strptime((d["order_date"] or today.isoformat())[:10], "%Y-%m-%d").date()
        except Exception:
            od = today
        due = od + timedelta(days=days_terms)
        overdue_days = (today - due).days  # 음수 = 만기 이전
        if only_overdue and overdue_days <= 0:
            continue
        if overdue_days <= 0:
            grade = "CURRENT"
        elif overdue_days <= 30:
            grade = "D-30"
        elif overdue_days <= 60:
            grade = "D-60"
        else:
            grade = "D-90+"
        d["outstanding"] = outstanding
        d["due_date_calc"] = due.isoformat()
        d["overdue_days"] = overdue_days
        d["grade"] = grade
        d["terms_days"] = days_terms
        out.append(d)
    out.sort(key=lambda x: (-x["overdue_days"], -x["outstanding"]))
    return out


def _outstanding_summary(items):
    """등급별 집계 KPI."""
    grades = {"CURRENT": 0.0, "D-30": 0.0, "D-60": 0.0, "D-90+": 0.0}
    counts = {"CURRENT": 0, "D-30": 0, "D-60": 0, "D-90+": 0}
    total = 0.0
    overdue_total = 0.0
    for it in items:
        g = it["grade"]
        amt = it["outstanding"]
        grades[g] = grades.get(g, 0.0) + amt
        counts[g] = counts.get(g, 0) + 1
        total += amt
        if g != "CURRENT":
            overdue_total += amt
    return {
        "by_grade": grades, "counts": counts,
        "total": total, "overdue_total": overdue_total,
        "n_total": len(items),
        "overdue_rate": (overdue_total / total * 100.0) if total > 0 else 0.0,
    }


def check_receivable_alerts():
    """수금 알림 트리거 — 만기 임박(D-7) + 연체(D+1, D+30, D+60).
    notify_user(SALES, ...) 사용 (1시간 중복 방지 내장).
    수신자: orders.created_by (수주 등록자) → can_use_sales 폴백.
    Returns: {sent: int, skipped: int, items: int}.
    """
    sent = 0; skipped = 0; total_items = 0
    with db_session() as c:
        items = _outstanding_receivables(c)
        total_items = len(items)
        # 영업 권한자 폴백 (주된 알림 대상 미상시)
        sales_uids = [r[0] for r in c.execute(
            "SELECT id FROM users WHERE can_use_sales=1 OR role IN ('admin','ceo')"
        ).fetchall()]
        for it in items:
            ov = it["overdue_days"]
            order_no = it.get("order_no") or f"#{it['order_id']}"
            cust = it.get("customer_name") or "-"
            outstanding = it.get("outstanding") or 0
            # 발송 조건: 만기 7일 임박 OR 연체 1/30/60일 도달
            tag = None
            if ov == -7:
                tag = "만기 임박 (D-7)"
            elif ov == 1:
                tag = "연체 1일"
            elif ov == 30:
                tag = "연체 30일"
            elif ov == 60:
                tag = "연체 60일"
            if not tag:
                continue
            title = f"💰 미수금 {tag} — {order_no}"
            body = (f"거래처: {cust} / 미수금: {int(outstanding):,}원 / "
                    f"등급: {it['grade']} / 만기: {it['due_date_calc']}")
            link = f"/sales/orders/{it['order_id']}"
            # 우선 created_by, 폴백 영업권한자 전체
            recipients = []
            cb = c.execute(
                "SELECT created_by FROM orders WHERE id=?", (it["order_id"],)
            ).fetchone()
            if cb and cb[0]:
                recipients.append(cb[0])
            else:
                recipients.extend(sales_uids)
            for uid in recipients:
                if notify_user(uid, "SALES", title, body=body, link=link):
                    sent += 1
                else:
                    skipped += 1
    return {"sent": sent, "skipped": skipped, "items": total_items}


@app.get("/sales/outstanding", response_class=HTMLResponse)
async def sales_outstanding_page(req: Request):
    """Top3-S1-4차 미수금 대시 — 등급별 집계 + 상세 (D-30/D-60/D-90+/CURRENT)."""
    u = _s1_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        items = _outstanding_receivables(c)
    summary = _outstanding_summary(items)
    return ctx(req, "sales_outstanding.html", user=u, active="sales_outstanding",
               tab="outstanding", items=items, summary=summary)


@app.get("/sales/aging", response_class=HTMLResponse)
async def sales_aging_page(req: Request):
    """Top3-S1-4차 연체 분석 — 거래처별 연체 매트릭스 (히트맵 테이블)."""
    u = _s1_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        items = _outstanding_receivables(c)
    # 거래처 × 등급 매트릭스
    matrix = {}
    for it in items:
        cust = it.get("customer_name") or "-"
        if cust not in matrix:
            matrix[cust] = {"CURRENT": 0.0, "D-30": 0.0, "D-60": 0.0,
                             "D-90+": 0.0, "total": 0.0, "max_overdue": 0}
        matrix[cust][it["grade"]] += it["outstanding"]
        matrix[cust]["total"] += it["outstanding"]
        if it["overdue_days"] > matrix[cust]["max_overdue"]:
            matrix[cust]["max_overdue"] = it["overdue_days"]
    rows = sorted(
        [{"customer": k, **v} for k, v in matrix.items()],
        key=lambda x: (-x["max_overdue"], -x["total"]),
    )
    summary = _outstanding_summary(items)
    return ctx(req, "sales_aging.html", user=u, active="sales",
               tab="aging", rows=rows, summary=summary)


@app.post("/sales/alerts/check")
async def sales_alerts_check(req: Request):
    """Top3-S1-4차 수금 알림 수동 트리거 — JSON {sent, skipped, items}."""
    u = _s1_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    result = check_receivable_alerts()
    return JSONResponse({"ok": True, **result})


# =====================================================
# 수출입 서류 — P11 베트남 수출 실무자 1차 라우트 골격 (2026-04-25)
# 한국 수출 표준 4단계: CI(상업송장) / PL(패킹리스트) / BL(선하증권) / 관세신고
# 1차 = 골격만 (UI 본문 · 외부 운송사 API 미도입 · 다음 사이클).
# 권한: P11(team_id=12 베트남법인) OR admin/ceo/executive OR can_use_sales.
# 매출 자동 채움: orders 테이블 참조 → export_orders INSERT 시 자동 조회.
# =====================================================
def _export_guard(req: Request):
    """수출입 권한 가드 — 베트남법인(team_id=12) OR admin/ceo/executive OR 영업권한자."""
    u = get_user(req)
    if not u:
        return None
    role = u.get("role") if isinstance(u, dict) else u["role"]
    if role in ("admin", "ceo", "executive"):
        return u
    team_id = u.get("team_id") if isinstance(u, dict) else u["team_id"]
    if team_id == 12:  # 12 베트남법인 (P11)
        return u
    if can_use_sales(u):
        return u
    return None


@app.get("/export", response_class=HTMLResponse)
async def export_home(req: Request):
    """수출 메인 (수주 목록 + 진행 상태 KPI) — 2차 본문."""
    u = _export_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        rows = c.execute(
            """SELECT eo.id, eo.buyer, eo.shipping_terms, eo.payment_terms,
                      eo.port_of_loading, eo.port_of_discharge, eo.status,
                      eo.created_at, eo.order_id,
                      COALESCE(o.order_no,'-') AS order_no,
                      COALESCE(o.total_amount,0) AS order_amount
               FROM export_orders eo
               LEFT JOIN orders o ON o.id = eo.order_id
               ORDER BY eo.id DESC LIMIT 200"""
        ).fetchall()
        items = [dict(r) for r in rows]
        # 상태별 카운트 (KPI)
        st_rows = c.execute(
            "SELECT status, COUNT(*) AS n FROM export_orders GROUP BY status"
        ).fetchall()
        st_map = {r["status"]: r["n"] for r in st_rows}
    return ctx(req, "export_home.html", user=u, active="export", items=items,
               st_draft=st_map.get("DRAFT", 0),
               st_ci=st_map.get("CI_ISSUED", 0),
               st_pl=st_map.get("PL_READY", 0),
               st_ship=st_map.get("SHIPPED", 0),
               st_clr=st_map.get("CLEARED", 0))


@app.get("/export/orders/new", response_class=HTMLResponse)
async def export_order_new_form(req: Request):
    """수출 수주 등록 폼 — 1차 골격 (UI 본문 다음 사이클)."""
    u = _export_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        orders = [dict(r) for r in c.execute(
            """SELECT id, order_no, customer_id, total_amount, order_date
               FROM orders ORDER BY id DESC LIMIT 100"""
        ).fetchall()]
    return ctx(req, "export_order_form.html", user=u, active="export",
               orders=orders)


@app.post("/export/orders")
async def export_order_create(req: Request):
    """수출 수주 INSERT — 매출 자동 채움 가설 핸들러.
    order_id 받으면 orders 에서 customer / total_amount / order_date 자동 조회."""
    u = _export_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    order_id = form.get("order_id") or None
    buyer = form.get("buyer") or None
    shipping_terms = form.get("shipping_terms") or None
    payment_terms = form.get("payment_terms") or None
    port_of_loading = form.get("port_of_loading") or "BUSAN"
    port_of_discharge = form.get("port_of_discharge") or None
    with db_session() as c:
        # 매출 자동 채움 — order_id 검증 (orders 존재 확인)
        if order_id:
            o = c.execute(
                "SELECT id, customer_id, total_amount, order_date FROM orders WHERE id=?",
                (order_id,)
            ).fetchone()
            if not o:
                return JSONResponse({"error": "수주 없음(order_id)"}, 404)
        cur = c.execute(
            """INSERT INTO export_orders
               (order_id, buyer, shipping_terms, payment_terms,
                port_of_loading, port_of_discharge, status, created_by)
               VALUES (?,?,?,?,?,?,'DRAFT',?)""",
            (order_id, buyer, shipping_terms, payment_terms,
             port_of_loading, port_of_discharge, u.get("id")),
        )
        return JSONResponse({"ok": True, "export_order_id": cur.lastrowid})


@app.get("/export/orders/{eo_id}", response_class=HTMLResponse)
async def export_order_detail(req: Request, eo_id: int):
    """수출 수주 상세 (CI/PL/BL/관세 탭) — 1차 골격."""
    u = _export_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        # v5H126: orders.currency 함께 조회 (CI 통화 mismatch 환산표용)
        try:
            eo = c.execute(
                """SELECT eo.*, COALESCE(o.order_no,'-') AS order_no,
                          COALESCE(o.total_amount,0) AS order_amount,
                          COALESCE(o.order_date,'-') AS order_date,
                          COALESCE(o.currency,'KRW') AS order_currency,
                          COALESCE(cu.name,'-') AS customer_name
                   FROM export_orders eo
                   LEFT JOIN orders o ON o.id = eo.order_id
                   LEFT JOIN customers cu ON cu.id = o.customer_id
                   WHERE eo.id = ?""",
                (eo_id,)
            ).fetchone()
        except Exception:
            eo = c.execute(
                """SELECT eo.*, COALESCE(o.order_no,'-') AS order_no,
                          COALESCE(o.total_amount,0) AS order_amount,
                          COALESCE(o.order_date,'-') AS order_date,
                          'KRW' AS order_currency,
                          COALESCE(cu.name,'-') AS customer_name
                   FROM export_orders eo
                   LEFT JOIN orders o ON o.id = eo.order_id
                   LEFT JOIN customers cu ON cu.id = o.customer_id
                   WHERE eo.id = ?""",
                (eo_id,)
            ).fetchone()
        if not eo:
            return RedirectResponse("/export", 303)
        eo = dict(eo)
        ci = [dict(r) for r in c.execute(
            "SELECT * FROM commercial_invoices WHERE export_order_id=? ORDER BY id DESC",
            (eo_id,)).fetchall()]
        pl = [dict(r) for r in c.execute(
            "SELECT * FROM packing_lists WHERE export_order_id=? ORDER BY id DESC",
            (eo_id,)).fetchall()]
        bl = [dict(r) for r in c.execute(
            "SELECT * FROM bills_of_lading WHERE export_order_id=? ORDER BY id DESC",
            (eo_id,)).fetchall()]
        cu_decl = [dict(r) for r in c.execute(
            "SELECT * FROM customs_declarations WHERE export_order_id=? ORDER BY id DESC",
            (eo_id,)).fetchall()]
    # v5H126: CI ↔ 수주 통화 mismatch → 환산표 (KRW 기준 차액)
    ci_fx_table = []
    try:
        from .database import _get_active_fx_rate as _gafx
        order_ccy = (eo.get("order_currency") or "KRW").upper()
        order_amt = float(eo.get("order_amount") or 0)
        for _ci in ci:
            ci_ccy = (_ci.get("currency") or "USD").upper()
            ci_amt = float(_ci.get("total_amount") or 0)
            if ci_ccy == order_ccy:
                continue  # 동일 통화면 환산표 불필요
            ref = (_ci.get("issue_date") or date.today().isoformat())[:10]
            ord_fx = _gafx(order_ccy, ref) or 1.0
            ci_fx = _gafx(ci_ccy, ref) or 1.0
            ord_krw = order_amt * ord_fx
            ci_krw = ci_amt * ci_fx
            ci_fx_table.append({
                "ci_id": _ci.get("id"),
                "invoice_no": _ci.get("invoice_no"),
                "issue_date": ref,
                "order_currency": order_ccy, "order_amount": order_amt,
                "order_fx": ord_fx, "order_krw": ord_krw,
                "ci_currency": ci_ccy, "ci_amount": ci_amt,
                "ci_fx": ci_fx, "ci_krw": ci_krw,
                "diff_krw": ci_krw - ord_krw,
                "fx_missing": (ord_fx == 1.0 and order_ccy != "KRW")
                              or (ci_fx == 1.0 and ci_ccy != "KRW"),
            })
    except Exception:
        ci_fx_table = []
    return ctx(req, "export_order_detail.html", user=u, active="export",
               eo=eo, ci=ci, pl=pl, bl=bl, customs=cu_decl,
               ci_fx_table=ci_fx_table)


@app.get("/export/ci/{eo_id}", response_class=HTMLResponse)
async def export_ci_form(req: Request, eo_id: int):
    """CI 작성/조회 (매출 자동 채움) — 2차 본문."""
    u = _export_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        eo = c.execute(
            """SELECT eo.*, COALESCE(o.total_amount,0) AS order_amount
               FROM export_orders eo
               LEFT JOIN orders o ON o.id = eo.order_id
               WHERE eo.id = ?""", (eo_id,)).fetchone()
        if not eo:
            return RedirectResponse("/export", 303)
        eo = dict(eo)
        items = [dict(r) for r in c.execute(
            "SELECT * FROM commercial_invoices WHERE export_order_id=? ORDER BY id DESC",
            (eo_id,)).fetchall()]
    return ctx(req, "export_ci.html", user=u, active="export",
               eo=eo, items=items, today=date.today().isoformat())


def _next_seq_no(c, table: str, col: str, prefix: str) -> str:
    """월별 시퀀스 발급 — `{prefix}-YYYYMM-####` (idempotent · race 방어 transaction 내).
    예: CI-202604-0001, PL-202604-0023."""
    ym = datetime.now().strftime("%Y%m")
    pat = f"{prefix}-{ym}-%"
    row = c.execute(
        f"SELECT {col} FROM {table} WHERE {col} LIKE ? "
        f"ORDER BY {col} DESC LIMIT 1", (pat,)
    ).fetchone()
    if row and row[col]:
        try:
            seq = int(str(row[col]).rsplit("-", 1)[-1]) + 1
        except Exception:
            seq = 1
    else:
        seq = 1
    return f"{prefix}-{ym}-{seq:04d}"


@app.post("/export/ci")
async def export_ci_create(req: Request):
    """CI INSERT — 2차 본문.
    invoice_no = `CI-YYYYMM-####` 자동. total_amount = 폼 우선, 미입력 시 orders 자동 채움.
    동시 UPDATE export_orders.status DRAFT/BOOKED → CI_ISSUED."""
    u = _export_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    eo_id = form.get("export_order_id")
    if not eo_id:
        return JSONResponse({"error": "export_order_id 필수"}, 400)
    issue_date = form.get("issue_date") or date.today().isoformat()
    currency = form.get("currency") or "USD"
    raw_amt = form.get("total_amount")
    with db_session() as c:
        # v5H123: orders.currency 도 함께 조회 (통화 일관성 검증)
        try:
            eo = c.execute(
                """SELECT eo.id, eo.status, COALESCE(o.total_amount,0) AS auto_amt,
                          COALESCE(o.currency,'KRW') AS order_currency
                   FROM export_orders eo LEFT JOIN orders o ON o.id=eo.order_id
                   WHERE eo.id=?""", (eo_id,)).fetchone()
        except Exception:
            # currency 컬럼 미존재 폴백
            eo = c.execute(
                """SELECT eo.id, eo.status, COALESCE(o.total_amount,0) AS auto_amt,
                          'KRW' AS order_currency
                   FROM export_orders eo LEFT JOIN orders o ON o.id=eo.order_id
                   WHERE eo.id=?""", (eo_id,)).fetchone()
        if not eo:
            return JSONResponse({"error": "수출 수주 없음"}, 404)
        # v5H117: 발행 사후 변경 차단 — DRAFT/BOOKED/CI_ISSUED 외 거부
        if eo["status"] not in ("DRAFT", "BOOKED", "CI_ISSUED"):
            from urllib.parse import quote as _q
            return RedirectResponse(
                f"/export/orders/{eo_id}?error=" + _q(f"이미 {eo['status']} 상태 — CI 추가 발행 불가"), 303
            )
        # 매출 자동 채움 — 폼 미입력 시 orders.total_amount
        try:
            total_amount = float(raw_amt) if raw_amt not in (None, "") else float(eo["auto_amt"] or 0)
        except Exception:
            total_amount = float(eo["auto_amt"] or 0)
        # v5H117: 음수 금액 차단
        if total_amount < 0:
            from urllib.parse import quote as _q
            return RedirectResponse(
                f"/export/orders/{eo_id}?error=" + _q("CI 금액은 음수일 수 없습니다"), 303
            )
        # v5H123: CI 통화 vs orders 통화 불일치 친절 경고 (차단 대신 redirect 안내 — 수출 시 환산이 일반적)
        try:
            order_cur = (eo["order_currency"] or "KRW").upper()
            ci_cur = (currency or "USD").upper()
            if order_cur and ci_cur and order_cur != ci_cur:
                # 차단하지 않음 — 수출은 KRW 수주를 USD CI 로 발행하는 사례 정상.
                # 단, 로그에 명시 기록 (감사 추적용)
                _ci_currency_warn = (
                    f"⚠ 통화 불일치: 수주={order_cur} / CI={ci_cur} (환산 확인 필요)"
                )
            else:
                _ci_currency_warn = None
        except Exception:
            _ci_currency_warn = None
        invoice_no = _next_seq_no(c, "commercial_invoices", "invoice_no", "CI")
        cur = c.execute(
            """INSERT INTO commercial_invoices
               (invoice_no, export_order_id, issue_date, total_amount, currency,
                signed_by, status)
               VALUES (?,?,?,?,?,?, 'ISSUED')""",
            (invoice_no, eo_id, issue_date, total_amount, currency, u.get("id")),
        )
        ci_id = cur.lastrowid
        # 상태 전이: DRAFT/BOOKED → CI_ISSUED
        if eo["status"] in ("DRAFT", "BOOKED"):
            c.execute(
                "UPDATE export_orders SET status='CI_ISSUED' WHERE id=?", (eo_id,)
            )
        # v5H117: 감사로그 (FTA H1 패턴) + v5H123 통화 불일치 경고 합성
        _audit_note = f"CI {invoice_no} ISSUED amount={total_amount} {currency}"
        if _ci_currency_warn:
            _audit_note += f" | {_ci_currency_warn}"
        _doc_audit_log(c, "commercial_invoice", ci_id, "ISSUE", u.get("id"), _audit_note)
    return RedirectResponse(f"/export/orders/{eo_id}", 303)


@app.get("/export/pl/{eo_id}", response_class=HTMLResponse)
async def export_pl_form(req: Request, eo_id: int):
    """PL 작성/조회 — 1차 골격."""
    u = _export_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        eo = c.execute(
            "SELECT * FROM export_orders WHERE id=?", (eo_id,)).fetchone()
        if not eo:
            return RedirectResponse("/export", 303)
        eo = dict(eo)
        items = [dict(r) for r in c.execute(
            "SELECT * FROM packing_lists WHERE export_order_id=? ORDER BY id DESC",
            (eo_id,)).fetchall()]
    return ctx(req, "export_pl.html", user=u, active="export",
               eo=eo, items=items)


@app.post("/export/pl")
async def export_pl_create(req: Request):
    """PL INSERT — 2차 본문 (헤더 + 다중 라인 packing_items).
    pl_no = `PL-YYYYMM-####`. 라인 합계는 폼 total_* 우선, 라인만 있으면 서버 합산.
    동시 UPDATE export_orders.status CI_ISSUED → PL_READY."""
    u = _export_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    eo_id = form.get("export_order_id")
    if not eo_id:
        return JSONResponse({"error": "export_order_id 필수"}, 400)
    # 다중 라인 — line_qty 등 getlist
    line_parts = form.getlist("line_part")
    line_qtys = form.getlist("line_qty")
    line_pkgs = form.getlist("line_pkg")
    line_ws = form.getlist("line_weight")
    line_vs = form.getlist("line_volume")
    # 헤더 합계 (폼 → 미입력 시 서버 재합산)
    def _f(x, d=0.0):
        try: return float(x) if x not in (None, "") else d
        except: return d
    tot_pkg = int(_f(form.get("total_packages")))
    tot_w = _f(form.get("total_weight"))
    tot_v = _f(form.get("total_volume"))
    if tot_pkg == 0 and line_qtys:
        tot_pkg = int(sum(_f(q) for q in line_qtys))
    if tot_w == 0 and line_ws:
        tot_w = sum(_f(w) for w in line_ws)
    if tot_v == 0 and line_vs:
        tot_v = sum(_f(v) for v in line_vs)
    # v5H117: 음수 헤더 차단
    if tot_pkg < 0 or tot_w < 0 or tot_v < 0:
        from urllib.parse import quote as _q
        return RedirectResponse(
            f"/export/pl/{eo_id}?error=" + _q("PL 헤더 합계는 음수일 수 없습니다"), 303
        )
    # v5H117: 라인 음수/0 사전 검증 — 잘못된 입력은 명시적 에러
    line_errors = []
    valid_lines = 0
    for i in range(max(len(line_parts), len(line_qtys))):
        q = _f(line_qtys[i] if i < len(line_qtys) else 0)
        w = _f(line_ws[i] if i < len(line_ws) else 0)
        v = _f(line_vs[i] if i < len(line_vs) else 0)
        if q < 0 or w < 0 or v < 0:
            line_errors.append(f"라인 {i+1}: 음수 값 불가")
            continue
        if q > 0:
            valid_lines += 1
    if line_errors:
        from urllib.parse import quote as _q
        return RedirectResponse(
            f"/export/pl/{eo_id}?error=" + _q(" / ".join(line_errors[:3])), 303
        )
    if valid_lines == 0:
        from urllib.parse import quote as _q
        return RedirectResponse(
            f"/export/pl/{eo_id}?error=" + _q("PL 라인 1건 이상 필요합니다"), 303
        )
    with db_session() as c:
        eo = c.execute(
            "SELECT id, status FROM export_orders WHERE id=?", (eo_id,)).fetchone()
        if not eo:
            return JSONResponse({"error": "수출 수주 없음"}, 404)
        # v5H117: 발행 사후 변경 차단
        if eo["status"] in ("SHIPPED", "CLEARED", "CANCELLED"):
            from urllib.parse import quote as _q
            return RedirectResponse(
                f"/export/orders/{eo_id}?error=" + _q(f"이미 {eo['status']} 상태 — PL 추가 발행 불가"), 303
            )
        pl_no = _next_seq_no(c, "packing_lists", "pl_no", "PL")
        cur = c.execute(
            """INSERT INTO packing_lists
               (pl_no, export_order_id, total_packages, total_weight, total_volume)
               VALUES (?,?,?,?,?)""",
            (pl_no, eo_id, tot_pkg, tot_w, tot_v),
        )
        pl_id = cur.lastrowid
        # 라인 INSERT (다중)
        n = max(len(line_parts), len(line_qtys))
        for i in range(n):
            qty = _f(line_qtys[i] if i < len(line_qtys) else 0)
            if qty <= 0:
                continue
            c.execute(
                """INSERT INTO packing_items
                   (pl_id, part_id, qty, package_type, weight, volume)
                   VALUES (?, NULL, ?, ?, ?, ?)""",
                (pl_id, qty,
                 (line_pkgs[i] if i < len(line_pkgs) else "CARTON"),
                 _f(line_ws[i] if i < len(line_ws) else 0),
                 _f(line_vs[i] if i < len(line_vs) else 0)),
            )
        # 상태 전이: CI_ISSUED → PL_READY (DRAFT/BOOKED 도 허용 — 동시 진행)
        if eo["status"] in ("DRAFT", "BOOKED", "CI_ISSUED"):
            c.execute(
                "UPDATE export_orders SET status='PL_READY' WHERE id=?", (eo_id,)
            )
        _doc_audit_log(c, "packing_list", pl_id, "ISSUE", u.get("id"),
                       f"PL {pl_no} lines={valid_lines} pkg={tot_pkg} w={tot_w}")
    return RedirectResponse(f"/export/orders/{eo_id}", 303)


@app.get("/export/bl/{eo_id}", response_class=HTMLResponse)
async def export_bl_form(req: Request, eo_id: int):
    """BL · 관세 통합 폼 — 2차 본문 (CI 자동 채움 → declared_value)."""
    u = _export_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        eo = c.execute(
            "SELECT * FROM export_orders WHERE id=?", (eo_id,)).fetchone()
        if not eo:
            return RedirectResponse("/export", 303)
        eo = dict(eo)
        bl = [dict(r) for r in c.execute(
            "SELECT * FROM bills_of_lading WHERE export_order_id=? ORDER BY id DESC",
            (eo_id,)).fetchall()]
        customs = [dict(r) for r in c.execute(
            "SELECT * FROM customs_declarations WHERE export_order_id=? ORDER BY id DESC",
            (eo_id,)).fetchall()]
        # CI 자동 채움 — 가장 최근 ISSUED CI 의 total_amount
        ci_row = c.execute(
            """SELECT total_amount FROM commercial_invoices
               WHERE export_order_id=? AND status='ISSUED'
               ORDER BY id DESC LIMIT 1""", (eo_id,)).fetchone()
        ci_amount = ci_row["total_amount"] if ci_row else 0
    return ctx(req, "export_bl_customs.html", user=u, active="export",
               eo=eo, bl=bl, customs=customs, ci_amount=ci_amount,
               today=date.today().isoformat())


@app.post("/export/bl")
async def export_bl_create(req: Request):
    """BL INSERT — 2차 본문 (외부 운송사 API 미사용 · 수동 입력 그대로 저장).
    동시 UPDATE export_orders.status PL_READY/CI_ISSUED → SHIPPED."""
    u = _export_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    eo_id = form.get("export_order_id")
    bl_no = (form.get("bl_no") or "").strip()
    if not eo_id or not bl_no:
        return JSONResponse({"error": "export_order_id / bl_no 필수"}, 400)
    with db_session() as c:
        eo = c.execute(
            "SELECT id, status FROM export_orders WHERE id=?", (eo_id,)).fetchone()
        if not eo:
            return JSONResponse({"error": "수출 수주 없음"}, 404)
        # v5H117: CLEARED/CANCELLED 후 추가 발행 차단
        if eo["status"] in ("CLEARED", "CANCELLED"):
            from urllib.parse import quote as _q
            return RedirectResponse(
                f"/export/orders/{eo_id}?error=" + _q(f"이미 {eo['status']} 상태 — BL 추가 발행 불가"), 303
            )
        cur = c.execute(
            """INSERT INTO bills_of_lading
               (bl_no, export_order_id, shipping_company, vessel,
                departure_date, arrival_date, tracking_no, status)
               VALUES (?,?,?,?,?,?,?, 'ISSUED')""",
            (bl_no, eo_id,
             form.get("shipping_company") or None,
             form.get("vessel") or None,
             form.get("departure_date") or None,
             form.get("arrival_date") or None,
             form.get("tracking_no") or None),
        )
        bl_id = cur.lastrowid
        # 상태 전이: PL_READY/CI_ISSUED/BOOKED → SHIPPED
        if eo["status"] in ("DRAFT", "BOOKED", "CI_ISSUED", "PL_READY"):
            c.execute(
                "UPDATE export_orders SET status='SHIPPED' WHERE id=?", (eo_id,)
            )
        _doc_audit_log(c, "bill_of_lading", bl_id, "ISSUE", u.get("id"),
                       f"BL {bl_no} eo={eo_id}")
    return RedirectResponse(f"/export/orders/{eo_id}", 303)


@app.get("/export/customs/{eo_id}", response_class=HTMLResponse)
async def export_customs_view(req: Request, eo_id: int):
    """관세 신고 조회 — 1차 골격 (BL 템플릿 재사용)."""
    u = _export_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    return RedirectResponse(f"/export/bl/{eo_id}", 303)


@app.post("/export/customs")
async def export_customs_create(req: Request):
    """관세 신고 INSERT — 2차 본문.
    declared_value = 폼 우선, 미입력 시 최신 CI total_amount 자동 채움.
    동시 UPDATE export_orders.status SHIPPED → CLEARED."""
    u = _export_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    eo_id = form.get("export_order_id")
    hs_code = (form.get("hs_code") or "").strip()
    if not eo_id or not hs_code:
        return JSONResponse({"error": "export_order_id / hs_code 필수"}, 400)
    declaration_no = (form.get("declaration_no") or "").strip() or None
    fta_origin = form.get("fta_origin") or None
    raw_dv = form.get("declared_value")
    with db_session() as c:
        eo = c.execute(
            "SELECT id, status FROM export_orders WHERE id=?", (eo_id,)).fetchone()
        if not eo:
            return JSONResponse({"error": "수출 수주 없음"}, 404)
        # CI 자동 채움
        try:
            declared_value = float(raw_dv) if raw_dv not in (None, "") else 0.0
        except Exception:
            declared_value = 0.0
        if declared_value < 0:
            from urllib.parse import quote as _q
            return RedirectResponse(
                f"/export/bl/{eo_id}?error=" + _q("관세 신고 가액은 음수 불가"), 303
            )
        # v5H117: CLEARED/CANCELLED 후 추가 신고 차단
        if eo["status"] in ("CLEARED", "CANCELLED"):
            from urllib.parse import quote as _q
            return RedirectResponse(
                f"/export/orders/{eo_id}?error=" + _q(f"이미 {eo['status']} 상태 — 관세 신고 추가 불가"), 303
            )
        if declared_value <= 0:
            ci_row = c.execute(
                """SELECT total_amount FROM commercial_invoices
                   WHERE export_order_id=? AND status='ISSUED'
                   ORDER BY id DESC LIMIT 1""", (eo_id,)).fetchone()
            if ci_row:
                declared_value = float(ci_row["total_amount"] or 0)
        cur = c.execute(
            """INSERT INTO customs_declarations
               (declaration_no, export_order_id, hs_code, fta_origin,
                declared_value, cleared_at, status)
               VALUES (?,?,?,?,?, datetime('now','localtime'), 'CLEARED')""",
            (declaration_no, eo_id, hs_code, fta_origin, declared_value),
        )
        cd_id = cur.lastrowid
        # 상태 전이: SHIPPED → CLEARED
        if eo["status"] in ("DRAFT", "BOOKED", "CI_ISSUED", "PL_READY", "SHIPPED"):
            c.execute(
                "UPDATE export_orders SET status='CLEARED' WHERE id=?", (eo_id,)
            )
        _doc_audit_log(c, "customs_declaration", cd_id, "CLEAR", u.get("id"),
                       f"hs={hs_code} value={declared_value}")
    return RedirectResponse(f"/export/orders/{eo_id}", 303)


# =====================================================
# 수출입 P11 3차 — 인쇄용 view + 일정 자동 알림 (2026-04-26)
# 외부 PDF 라이브러리 0건. HTML 인쇄 layout(@media print) 만 사용.
# G1~G5 핫패치 보존, v2 본체 무수정.
# =====================================================
@app.get("/export/ci/{ci_id}/print", response_class=HTMLResponse)
async def export_ci_print(req: Request, ci_id: int):
    """CI 인쇄용 (한국어 + 영어 양식 · 헤더/사이드바 숨김)."""
    u = _export_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        ci = c.execute(
            """SELECT ci.*, eo.buyer, eo.shipping_terms, eo.payment_terms,
                      eo.port_of_loading, eo.port_of_discharge,
                      COALESCE(o.order_no,'-') AS order_no,
                      COALESCE(cu.name,'-') AS customer_name
               FROM commercial_invoices ci
               JOIN export_orders eo ON eo.id = ci.export_order_id
               LEFT JOIN orders o ON o.id = eo.order_id
               LEFT JOIN customers cu ON cu.id = o.customer_id
               WHERE ci.id = ?""", (ci_id,)).fetchone()
        if not ci:
            return RedirectResponse("/export", 303)
        ci = dict(ci)
    return tpl.TemplateResponse(
        "export_ci_print.html",
        {"request": req, "ci": ci, "today": date.today().isoformat()})


@app.get("/export/pl/{pl_id}/print", response_class=HTMLResponse)
async def export_pl_print(req: Request, pl_id: int):
    """PL 인쇄용 (라인 합계 · 헤더/사이드바 숨김)."""
    u = _export_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        pl = c.execute(
            """SELECT pl.*, eo.buyer, eo.shipping_terms,
                      eo.port_of_loading, eo.port_of_discharge,
                      COALESCE(o.order_no,'-') AS order_no
               FROM packing_lists pl
               JOIN export_orders eo ON eo.id = pl.export_order_id
               LEFT JOIN orders o ON o.id = eo.order_id
               WHERE pl.id = ?""", (pl_id,)).fetchone()
        if not pl:
            return RedirectResponse("/export", 303)
        pl = dict(pl)
        lines = [dict(r) for r in c.execute(
            "SELECT * FROM packing_items WHERE pl_id=? ORDER BY id ASC",
            (pl_id,)).fetchall()]
    return tpl.TemplateResponse(
        "export_pl_print.html",
        {"request": req, "pl": pl, "lines": lines,
         "today": date.today().isoformat()})


@app.get("/export/bl/{bl_id}/print", response_class=HTMLResponse)
async def export_bl_print(req: Request, bl_id: int):
    """BL/관세 인쇄용 (헤더/사이드바 숨김)."""
    u = _export_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        bl = c.execute(
            """SELECT bl.*, eo.buyer, eo.shipping_terms,
                      eo.port_of_loading, eo.port_of_discharge,
                      COALESCE(o.order_no,'-') AS order_no
               FROM bills_of_lading bl
               JOIN export_orders eo ON eo.id = bl.export_order_id
               LEFT JOIN orders o ON o.id = eo.order_id
               WHERE bl.id = ?""", (bl_id,)).fetchone()
        if not bl:
            return RedirectResponse("/export", 303)
        bl = dict(bl)
        customs = [dict(r) for r in c.execute(
            "SELECT * FROM customs_declarations WHERE export_order_id=? ORDER BY id DESC",
            (bl["export_order_id"],)).fetchall()]
    return tpl.TemplateResponse(
        "export_bl_print.html",
        {"request": req, "bl": bl, "customs": customs,
         "today": date.today().isoformat()})


def check_export_alerts():
    """수출입 일정 자동 알림 — 출하 D-3 / CI 만료(90일) / 관세 신고 임박.
    notify_user 1시간 중복 방지 내장. 매 호출 idempotent.
    Returns: {'shipping': n, 'ci_expire': n, 'customs': n} 카운트."""
    fired = {"shipping": 0, "ci_expire": 0, "customs": 0}
    today = date.today()
    d3 = (today + timedelta(days=3)).isoformat()
    today_iso = today.isoformat()
    d90 = (today - timedelta(days=90)).isoformat()
    with db_session() as c:
        # 베트남법인(team_id=12) + admin/ceo/executive + 영업권 사용자에게 송부
        recipients = [r["id"] for r in c.execute(
            "SELECT id FROM users WHERE is_active=1 AND "
            "(team_id=12 OR role IN ('admin','ceo','executive'))"
        ).fetchall()]
        # 1) 출하 임박 D-3 — bills_of_lading.departure_date
        bls = c.execute(
            """SELECT bl.id, bl.bl_no, bl.departure_date, bl.export_order_id
               FROM bills_of_lading bl
               WHERE bl.departure_date IS NOT NULL
                 AND bl.departure_date BETWEEN ? AND ?
                 AND bl.status IN ('DRAFT','ISSUED')""",
            (today_iso, d3)).fetchall()
        for bl in bls:
            for uid in recipients:
                if notify_user(uid, "EXPORT",
                               f"출하 임박 D-3 — B/L {bl['bl_no']}",
                               f"출항일 {bl['departure_date']}",
                               f"/export/orders/{bl['export_order_id']}"):
                    fired["shipping"] += 1
        # 2) CI 만료 — issue_date + 90일 경과 ISSUED 상태
        cis = c.execute(
            """SELECT id, invoice_no, issue_date, export_order_id
               FROM commercial_invoices
               WHERE status='ISSUED' AND issue_date IS NOT NULL
                 AND issue_date <= ?""", (d90,)).fetchall()
        for ci in cis:
            for uid in recipients:
                if notify_user(uid, "EXPORT",
                               f"CI 만료 — {ci['invoice_no']} (90일 경과)",
                               f"발급일 {ci['issue_date']}",
                               f"/export/orders/{ci['export_order_id']}"):
                    fired["ci_expire"] += 1
        # 3) 관세 신고 임박 — SHIPPED 인데 미신고
        ships = c.execute(
            """SELECT eo.id FROM export_orders eo
               WHERE eo.status='SHIPPED'
                 AND NOT EXISTS (SELECT 1 FROM customs_declarations cd
                                 WHERE cd.export_order_id=eo.id
                                   AND cd.status IN ('SUBMITTED','CLEARED'))"""
        ).fetchall()
        for eo in ships:
            for uid in recipients:
                if notify_user(uid, "EXPORT",
                               f"관세 신고 미접수 — 수출 #{eo['id']}",
                               "출하 후 관세 신고가 누락되었습니다.",
                               f"/export/orders/{eo['id']}"):
                    fired["customs"] += 1
    return fired


@app.post("/export/alerts/check")
async def export_alerts_check(req: Request):
    """수출입 일정 알림 점검 트리거 (관리자/CEO 전용)."""
    u = _export_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    role = u.get("role") if isinstance(u, dict) else u["role"]
    if role not in ("admin", "ceo", "executive"):
        return JSONResponse({"error": "관리자 전용"}, 403)
    fired = check_export_alerts()
    return JSONResponse({"ok": True, "fired": fired})


# =====================================================
# HAIST WORKS — 진행 간트/번다운 1차 (2026-04-26 갭서베이 Top10 #4)
# DB: project_milestones / project_burndown_snapshots (idempotent)
# UI: progress_gantt.html / progress_burndown.html (외부 차트 라이브러리 0)
# 페르소나: P1 PM(주 2~3회) · P2 CEO 전사 대시
# =====================================================
def _progress_guard(req: Request, project_id: int = None):
    """진행 가드 — admin/ceo/executive OR PM(pm_id) OR 프로젝트 lead OR 동일 팀."""
    u = get_user(req)
    if not u:
        return None
    role = u.get("role") if isinstance(u, dict) else u["role"]
    if role in ("admin", "ceo", "executive"):
        return u
    if project_id is None:
        return u  # 전사 대시는 별도 분기 (대시 라우트에서 admin/ceo만 허용)
    with db_session() as c:
        proj = c.execute(
            "SELECT pm_id, lead_user_id FROM projects WHERE id=?", (project_id,)
        ).fetchone()
    if not proj:
        return None
    if u["id"] in (proj["pm_id"], proj["lead_user_id"]):
        return u
    if role in ("leader", "pm"):
        return u
    return None


def _burndown_compute(project_id: int):
    """프로젝트의 task 기반 번다운 스냅샷 데이터 계산. 외부 라이브러리 0."""
    with db_session() as c:
        total = c.execute(
            "SELECT COUNT(*) FROM tasks WHERE project_id=?", (project_id,)
        ).fetchone()[0]
        done = c.execute(
            "SELECT COUNT(*) FROM tasks WHERE project_id=? AND status='완료'",
            (project_id,),
        ).fetchone()[0]
        rem_h = c.execute(
            """SELECT COALESCE(SUM(hours), 0) FROM tasks
               WHERE project_id=? AND status != '완료'""",
            (project_id,),
        ).fetchone()[0]
    return {"total_tasks": total, "completed_tasks": done,
            "remaining_hours": float(rem_h or 0)}


@app.get("/progress/{project_id}/gantt", response_class=HTMLResponse)
async def progress_gantt(req: Request, project_id: int):
    """간트 차트 — project_phases 의 planned_start/planned_end 를 CSS bar 로 렌더."""
    u = _progress_guard(req, project_id)
    if not u:
        return RedirectResponse("/progress", 303)
    with db_session() as c:
        proj = c.execute(
            "SELECT p.*, p.name AS project_name FROM projects p WHERE p.id=?",
            (project_id,),
        ).fetchone()
        if not proj:
            return RedirectResponse("/progress", 303)
        phases = [dict(r) for r in c.execute(
            """SELECT id, phase_code, phase_order, status, progress_pct,
                      planned_start, planned_end, actual_start, actual_end, note
               FROM project_phases WHERE project_id=? ORDER BY phase_order""",
            (project_id,),
        ).fetchall()]
        milestones = [dict(r) for r in c.execute(
            """SELECT id, name, target_date, completed_at, status
               FROM project_milestones WHERE project_id=?
               ORDER BY target_date ASC""",
            (project_id,),
        ).fetchall()]
    today_str = datetime.now().strftime("%Y-%m-%d")
    return ctx(req, "progress_gantt.html", user=u, active="progress",
               project=dict(proj), phases=phases, milestones=milestones,
               today_str=today_str,
               PHASE_CODE_TO_LABEL=PHASE_CODE_TO_LABEL)


def _linear_regression(xs, ys):
    """순수 Python 선형 회귀 (slope, intercept, r_squared). 외부 라이브러리 0."""
    n = len(xs)
    if n < 2:
        return 0.0, 0.0, 0.0
    sx = sum(xs); sy = sum(ys)
    mx = sx / n; my = sy / n
    num = 0.0; den = 0.0; sst = 0.0
    for i in range(n):
        dx = xs[i] - mx; dy = ys[i] - my
        num += dx * dy
        den += dx * dx
        sst += dy * dy
    if den == 0:
        return 0.0, my, 0.0
    slope = num / den
    intercept = my - slope * mx
    if sst == 0:
        return slope, intercept, 1.0
    ssr = 0.0
    for i in range(n):
        pred = slope * xs[i] + intercept
        ssr += (ys[i] - pred) ** 2
    r2 = max(0.0, 1.0 - ssr / sst)
    return slope, intercept, r2


def _burndown_forecast(snaps):
    """스냅샷 리스트 → (forecast_date_str, slope, r2, days_to_zero) 또는 None.
    snaps: [{snap_date, total_tasks, completed_tasks, ...}, ...] (날짜 오름차순)"""
    if not snaps or len(snaps) < 2:
        return None
    base = datetime.strptime(snaps[0]["snap_date"], "%Y-%m-%d")
    xs = []; ys = []
    for s in snaps:
        d = datetime.strptime(s["snap_date"], "%Y-%m-%d")
        xs.append((d - base).days)
        ys.append(float(s["total_tasks"] - s["completed_tasks"]))
    slope, intercept, r2 = _linear_regression(xs, ys)
    if slope >= 0:
        return {"slope": slope, "intercept": intercept, "r_squared": r2,
                "forecast_date": None, "days_to_zero": None,
                "base_date": snaps[0]["snap_date"]}
    x_zero = -intercept / slope
    fdate = base + timedelta(days=int(round(x_zero)))
    return {"slope": slope, "intercept": intercept, "r_squared": r2,
            "forecast_date": fdate.strftime("%Y-%m-%d"),
            "days_to_zero": x_zero,
            "base_date": snaps[0]["snap_date"]}


@app.get("/progress/{project_id}/burndown", response_class=HTMLResponse)
async def progress_burndown(req: Request, project_id: int):
    """번다운 — 일별 스냅샷 + 회귀 기반 예측 종료일. 외부 차트 0건 (SVG 직접)."""
    u = _progress_guard(req, project_id)
    if not u:
        return RedirectResponse("/progress", 303)
    with db_session() as c:
        proj = c.execute(
            "SELECT p.*, p.name AS project_name FROM projects p WHERE p.id=?",
            (project_id,),
        ).fetchone()
        if not proj:
            return RedirectResponse("/progress", 303)
        snaps = [dict(r) for r in c.execute(
            """SELECT snap_date, total_tasks, completed_tasks, remaining_hours
               FROM project_burndown_snapshots WHERE project_id=?
               ORDER BY snap_date ASC""",
            (project_id,),
        ).fetchall()]
        fc_row = c.execute(
            """SELECT computed_at, sample_n, slope, intercept, r_squared,
                      forecast_date, planned_end
               FROM project_forecasts WHERE project_id=?""",
            (project_id,),
        ).fetchone()
    today_pt = _burndown_compute(project_id)
    forecast = _burndown_forecast(snaps)
    today_str = datetime.now().strftime("%Y-%m-%d")
    return ctx(req, "progress_burndown.html", user=u, active="progress",
               project=dict(proj), snaps=snaps, today_pt=today_pt,
               forecast=forecast, today_str=today_str,
               saved_forecast=dict(fc_row) if fc_row else None)


@app.post("/progress/{project_id}/forecast")
async def progress_forecast(req: Request, project_id: int):
    """선형 회귀로 예측 종료일 산출 후 project_forecasts UPSERT.
    응답 JSON: {forecast_date, current_slope, r_squared, sample_n}."""
    u = _progress_guard(req, project_id)
    if not u:
        return JSONResponse({"error": "forbidden"}, status_code=403)
    with db_session() as c:
        proj = c.execute(
            "SELECT id, end_date FROM projects WHERE id=?", (project_id,),
        ).fetchone()
        if not proj:
            return JSONResponse({"error": "not_found"}, status_code=404)
        snaps = [dict(r) for r in c.execute(
            """SELECT snap_date, total_tasks, completed_tasks
               FROM project_burndown_snapshots WHERE project_id=?
               ORDER BY snap_date ASC""",
            (project_id,),
        ).fetchall()]
    fc = _burndown_forecast(snaps)
    if not fc:
        return JSONResponse({"error": "insufficient_data",
                              "sample_n": len(snaps)}, status_code=400)
    with db_session() as c:
        c.execute(
            """INSERT INTO project_forecasts
                 (project_id, sample_n, slope, intercept, r_squared,
                  forecast_date, planned_end, computed_at)
               VALUES (?,?,?,?,?,?,?,datetime('now','localtime'))
               ON CONFLICT(project_id) DO UPDATE SET
                   sample_n=excluded.sample_n,
                   slope=excluded.slope,
                   intercept=excluded.intercept,
                   r_squared=excluded.r_squared,
                   forecast_date=excluded.forecast_date,
                   planned_end=excluded.planned_end,
                   computed_at=excluded.computed_at""",
            (project_id, len(snaps), fc["slope"], fc["intercept"],
             fc["r_squared"], fc["forecast_date"], proj["end_date"]),
        )
    return JSONResponse({
        "forecast_date": fc["forecast_date"],
        "current_slope": fc["slope"],
        "r_squared": fc["r_squared"],
        "sample_n": len(snaps),
        "planned_end": proj["end_date"],
    })


@app.get("/progress/{project_id}/milestones", response_class=HTMLResponse)
async def progress_milestones(req: Request, project_id: int):
    """마일스톤 페이지 — 프로젝트별 목록 + 달성률."""
    u = _progress_guard(req, project_id)
    if not u:
        return RedirectResponse("/progress", 303)
    with db_session() as c:
        proj = c.execute(
            "SELECT p.*, p.name AS project_name FROM projects p WHERE p.id=?",
            (project_id,),
        ).fetchone()
        if not proj:
            return RedirectResponse("/progress", 303)
        ms = [dict(r) for r in c.execute(
            """SELECT id, name, target_date, completed_at, status
               FROM project_milestones WHERE project_id=?
               ORDER BY target_date IS NULL, target_date ASC""",
            (project_id,),
        ).fetchall()]
    total = len(ms)
    done = sum(1 for m in ms if m["status"] == "DONE")
    pct = (done * 100.0 / total) if total else 0.0
    today_str = datetime.now().strftime("%Y-%m-%d")
    return ctx(req, "progress_milestones.html", user=u, active="progress",
               project=dict(proj), milestones=ms,
               total_ms=total, done_ms=done, pct_ms=pct,
               today_str=today_str)


@app.post("/progress/{project_id}/snapshot")
async def progress_snapshot(req: Request, project_id: int):
    """일별 스냅샷 트리거. 같은 날 중복 시 UPDATE (UNIQUE constraint)."""
    u = _progress_guard(req, project_id)
    if not u:
        return RedirectResponse("/progress", 303)
    pt = _burndown_compute(project_id)
    today = datetime.now().strftime("%Y-%m-%d")
    with db_session() as c:
        c.execute(
            """INSERT INTO project_burndown_snapshots
                 (project_id, snap_date, total_tasks, completed_tasks, remaining_hours)
               VALUES (?,?,?,?,?)
               ON CONFLICT(project_id, snap_date) DO UPDATE SET
                   total_tasks=excluded.total_tasks,
                   completed_tasks=excluded.completed_tasks,
                   remaining_hours=excluded.remaining_hours""",
            (project_id, today, pt["total_tasks"],
             pt["completed_tasks"], pt["remaining_hours"]),
        )
    return RedirectResponse(f"/progress/{project_id}/burndown", 303)


@app.get("/progress-dashboard", response_class=HTMLResponse)
async def progress_dashboard_company(req: Request):
    """전사 프로젝트 대시 — admin/ceo/executive 전용. /progress/{int} 충돌 회피용 하이픈 URL."""
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    role = u.get("role") if isinstance(u, dict) else u["role"]
    if role not in ("admin", "ceo", "executive"):
        return RedirectResponse("/progress", 303)
    with db_session() as c:
        rows = [dict(r) for r in c.execute(
            """SELECT p.id, p.name, p.mgmt_code, p.status, p.start_date, p.end_date,
                      (SELECT COUNT(*) FROM tasks t WHERE t.project_id=p.id) AS total_t,
                      (SELECT COUNT(*) FROM tasks t WHERE t.project_id=p.id
                          AND t.status='완료') AS done_t,
                      (SELECT COUNT(*) FROM project_milestones m
                          WHERE m.project_id=p.id AND m.status='DONE') AS done_ms,
                      (SELECT COUNT(*) FROM project_milestones m
                          WHERE m.project_id=p.id) AS total_ms
               FROM projects p
               WHERE p.status IN ('진행중','진행')
               ORDER BY (p.end_date IS NULL), p.end_date ASC LIMIT 50"""
        ).fetchall()]
    return ctx(req, "progress_dashboard.html", user=u, active="progress",
               projects=rows)


# =====================================================
# HAIST WORKS — QMS 강화 1차 (2026-04-26 갭서베이 Top10 #6)
# DB: qms_audit_log / corrective_actions / preventive_actions (idempotent)
#     + issues 5컬럼 ALTER ADD (sla_hours/detected_at/responded_at/recurrence_id/sla_breached)
# UI: qms_dashboard.html / qms_recurrence.html (외부 차트 라이브러리 0)
# 페르소나: P2 제조/품질팀 주2회 · P-CEO 분기 품질 KPI
# =====================================================
def _qms_guard(req: Request):
    """품질 가드 — admin/ceo/executive OR 품질팀(team name LIKE %품질%) OR can_use_quality=1."""
    u = get_user(req)
    if not u:
        return None
    role = u.get("role") if isinstance(u, dict) else u["role"]
    if role in ("admin", "ceo", "executive"):
        return u
    if u.get("can_use_quality"):
        return u
    # 팀명 LIKE '%품질%' (is_team("quality") 대용)
    tid = u.get("team_id")
    if tid:
        with db_session() as c:
            row = c.execute(
                "SELECT name FROM teams WHERE id=?", (tid,)
            ).fetchone()
        if row and "품질" in (row["name"] or ""):
            return u
    return None


def _qms_sla_status(issue: dict) -> dict:
    """SLA 현황 계산 — sla_hours 기반 elapsed/remaining/breached 계산. 외부 라이브러리 0."""
    sla_h = issue.get("sla_hours") or 24
    detected = issue.get("detected_at") or issue.get("occurred_at") or issue.get("created_at")
    if not detected:
        return {"sla_hours": sla_h, "elapsed_h": 0, "remaining_h": sla_h, "breached": False}
    try:
        det_dt = datetime.strptime(detected[:19], "%Y-%m-%d %H:%M:%S")
    except (ValueError, TypeError):
        try:
            det_dt = datetime.strptime(detected[:10], "%Y-%m-%d")
        except (ValueError, TypeError):
            return {"sla_hours": sla_h, "elapsed_h": 0, "remaining_h": sla_h, "breached": False}
    resolved = issue.get("resolved_at")
    if resolved:
        try:
            end_dt = datetime.strptime(resolved[:19], "%Y-%m-%d %H:%M:%S")
        except (ValueError, TypeError):
            end_dt = datetime.now()
    else:
        end_dt = datetime.now()
    elapsed_h = (end_dt - det_dt).total_seconds() / 3600.0
    remaining_h = sla_h - elapsed_h
    return {"sla_hours": sla_h, "elapsed_h": round(elapsed_h, 1),
            "remaining_h": round(remaining_h, 1), "breached": remaining_h < 0 and not resolved}


@app.get("/qms", response_class=HTMLResponse)
async def qms_dashboard(req: Request):
    """QMS 품질 대시 — severity/SLA/재발 KPI 통합."""
    u = _qms_guard(req)
    if not u:
        return RedirectResponse("/issues", 303)
    with db_session() as c:
        total = c.execute("SELECT COUNT(*) FROM issues").fetchone()[0]
        open_cnt = c.execute(
            "SELECT COUNT(*) FROM issues WHERE status NOT IN ('해결','종결')"
        ).fetchone()[0]
        critical = c.execute(
            "SELECT COUNT(*) FROM issues WHERE severity IN ('치명','심각','CRITICAL','HIGH') "
            "AND status NOT IN ('해결','종결')"
        ).fetchone()[0]
        breached = c.execute(
            "SELECT COUNT(*) FROM issues WHERE COALESCE(sla_breached,0)=1 "
            "AND status NOT IN ('해결','종결')"
        ).fetchone()[0]
        recur = c.execute(
            "SELECT COUNT(DISTINCT recurrence_id) FROM issues WHERE recurrence_id IS NOT NULL AND recurrence_id != ''"
        ).fetchone()[0]
        ca_open = c.execute(
            "SELECT COUNT(*) FROM corrective_actions WHERE status IN ('OPEN','IN_PROGRESS')"
        ).fetchone()[0]
        # SLA 위반 위험 목록 (open + 잔여 시간 < 4h or breached)
        rows = [dict(r) for r in c.execute(
            """SELECT id, issue_no, title, severity, status, sla_hours,
                      occurred_at, detected_at, resolved_at, owner_team_id, recurrence_id
               FROM issues WHERE status NOT IN ('해결','종결')
               ORDER BY (severity IN ('치명','심각','CRITICAL','HIGH')) DESC,
                        occurred_at ASC LIMIT 50"""
        ).fetchall()]
        for r in rows:
            r["sla"] = _qms_sla_status(r)
    kpi = {"total": total, "open": open_cnt, "critical": critical,
           "breached": breached, "recurrence_groups": recur, "ca_open": ca_open}
    return ctx(req, "qms_dashboard.html", user=u, active="qms",
               kpi=kpi, items=rows)


@app.get("/qms/issues/{iid}/sla", response_class=HTMLResponse)
async def qms_issue_sla(req: Request, iid: int):
    """단일 이슈 SLA 현황 — JSON 반환."""
    u = _qms_guard(req)
    if not u:
        return JSONResponse({"error": "forbidden"}, status_code=403)
    with db_session() as c:
        row = c.execute("SELECT * FROM issues WHERE id=?", (iid,)).fetchone()
    if not row:
        return JSONResponse({"error": "notfound"}, status_code=404)
    return JSONResponse(_qms_sla_status(dict(row)))


@app.post("/qms/issues/{iid}/corrective")
async def qms_corrective_add(req: Request, iid: int,
                              action: str = Form(...),
                              responsible: str = Form(""),
                              due_date: str = Form("")):
    """시정조치 추가 + 감사로그 기록."""
    u = _qms_guard(req)
    if not u:
        return RedirectResponse(f"/issues/{iid}", 303)
    resp_id = int(responsible) if responsible and responsible.isdigit() else None
    with db_session() as c:
        cur = c.execute(
            """INSERT INTO corrective_actions
                 (issue_id, action, responsible, due_date, created_by)
               VALUES (?,?,?,?,?)""",
            (iid, action.strip(), resp_id, due_date or None, u["id"]),
        )
        ca_id = cur.lastrowid
        c.execute(
            """INSERT INTO qms_audit_log (issue_id, action, actor, note)
               VALUES (?, 'corrective_added', ?, ?)""",
            (iid, u["id"], f"CA#{ca_id}: {action[:80]}"),
        )
    # 알림시스템 통합 (사이클 2026-04-26) — 시정조치 담당자에게 QMS 알림 (SLA due_date 표기)
    if resp_id:
        notify_user(
            resp_id, "QMS",
            f"⚙️ 시정조치 배정 — Issue #{iid} (CA#{ca_id})",
            body=f"기한 {due_date or '미지정'} / {action[:80]}",
            link=f"/issues/{iid}",
        )
    return RedirectResponse(f"/issues/{iid}", 303)


@app.post("/qms/issues/{iid}/preventive")
async def qms_preventive_add(req: Request, iid: int,
                              corrective_id: str = Form(...),
                              action: str = Form(...)):
    """예방조치 추가 (특정 corrective_action 에 종속) + 감사로그."""
    u = _qms_guard(req)
    if not u:
        return RedirectResponse(f"/issues/{iid}", 303)
    if not corrective_id.isdigit():
        return RedirectResponse(f"/issues/{iid}", 303)
    ca_id = int(corrective_id)
    with db_session() as c:
        cur = c.execute(
            """INSERT INTO preventive_actions
                 (corrective_id, action, created_by)
               VALUES (?,?,?)""",
            (ca_id, action.strip(), u["id"]),
        )
        pa_id = cur.lastrowid
        c.execute(
            """INSERT INTO qms_audit_log (issue_id, action, actor, note)
               VALUES (?, 'preventive_added', ?, ?)""",
            (iid, u["id"], f"PA#{pa_id} (CA#{ca_id}): {action[:80]}"),
        )
    return RedirectResponse(f"/issues/{iid}", 303)


@app.get("/qms/recurrence", response_class=HTMLResponse)
async def qms_recurrence(req: Request):
    """재발 추적 — 같은 root_cause/recurrence_id 그룹화 트리."""
    u = _qms_guard(req)
    if not u:
        return RedirectResponse("/issues", 303)
    with db_session() as c:
        # recurrence_id 명시 그룹
        groups_by_rid = [dict(r) for r in c.execute(
            """SELECT recurrence_id AS gid, COUNT(*) AS cnt,
                      MIN(occurred_at) AS first_at, MAX(occurred_at) AS last_at
               FROM issues
               WHERE recurrence_id IS NOT NULL AND recurrence_id != ''
               GROUP BY recurrence_id
               HAVING cnt >= 2
               ORDER BY cnt DESC, last_at DESC LIMIT 50"""
        ).fetchall()]
        # root_cause 텍스트 그룹 (recurrence_id 없을 때 fallback)
        groups_by_rc = [dict(r) for r in c.execute(
            """SELECT SUBSTR(root_cause,1,60) AS gid, COUNT(*) AS cnt,
                      MIN(occurred_at) AS first_at, MAX(occurred_at) AS last_at
               FROM issues
               WHERE root_cause IS NOT NULL AND TRIM(root_cause) != ''
                     AND (recurrence_id IS NULL OR recurrence_id = '')
               GROUP BY SUBSTR(root_cause,1,60)
               HAVING cnt >= 2
               ORDER BY cnt DESC, last_at DESC LIMIT 30"""
        ).fetchall()]
        # 각 그룹의 이슈 리스트 (recurrence_id 기준)
        details = {}
        for g in groups_by_rid:
            details[g["gid"]] = [dict(r) for r in c.execute(
                """SELECT id, issue_no, title, severity, status, occurred_at
                   FROM issues WHERE recurrence_id = ?
                   ORDER BY occurred_at DESC LIMIT 20""",
                (g["gid"],),
            ).fetchall()]
    return ctx(req, "qms_recurrence.html", user=u, active="qms",
               groups_rid=groups_by_rid, groups_rc=groups_by_rc, details=details)


# =====================================================
# QMS 2차 Pareto + CAPA 심화 (2026-04-26)
# Pareto: root_cause 빈도 + 누적 % (80/20 법칙)
# CAPA 라이프사이클: DRAFT → APPROVED → IN_PROGRESS → COMPLETED → VERIFIED
# 외부 차트 라이브러리 0건 (CSS bar + SVG cumulative line)
# =====================================================
def _qms_capa_guard(req: Request):
    """승인/검증 권한 — admin/ceo/executive OR 품질팀장(role 'leader')."""
    u = _qms_guard(req)
    if not u:
        return None
    role = u.get("role") if isinstance(u, dict) else u["role"]
    if role in ("admin", "ceo", "executive", "leader"):
        return u
    return None


def _capa_kpi(c) -> dict:
    """CAPA KPI 계산 — 평균 closure time / 검증 비율 / 부서별 분포 (외부 라이브러리 0)."""
    # 평균 closure time: created_at(DRAFT 진입) → completed_at (단위: 일)
    rows = c.execute(
        """SELECT created_at, completed_at FROM corrective_actions
           WHERE completed_at IS NOT NULL AND created_at IS NOT NULL"""
    ).fetchall()
    days = []
    for r in rows:
        try:
            t0 = datetime.strptime(r["created_at"][:19], "%Y-%m-%d %H:%M:%S")
            t1 = datetime.strptime(r["completed_at"][:19], "%Y-%m-%d %H:%M:%S")
            days.append((t1 - t0).total_seconds() / 86400.0)
        except (ValueError, TypeError):
            continue
    avg_closure = round(sum(days) / len(days), 1) if days else 0
    # 검증 비율: VERIFIED / COMPLETED
    completed = c.execute(
        "SELECT COUNT(*) FROM corrective_actions "
        "WHERE lifecycle_status IN ('COMPLETED','VERIFIED')"
    ).fetchone()[0]
    verified = c.execute(
        "SELECT COUNT(*) FROM corrective_actions WHERE lifecycle_status='VERIFIED'"
    ).fetchone()[0]
    verify_rate = round(100.0 * verified / completed, 1) if completed else 0
    # 부서별 분포 (issue.owner_team_id → teams.name 조인)
    dept_rows = [dict(r) for r in c.execute(
        """SELECT COALESCE(t.name,'(미지정)') AS dept, COUNT(ca.id) AS cnt
           FROM corrective_actions ca
           LEFT JOIN issues i ON i.id = ca.issue_id
           LEFT JOIN teams t ON t.id = i.owner_team_id
           GROUP BY t.name
           ORDER BY cnt DESC LIMIT 10"""
    ).fetchall()]
    return {"avg_closure_days": avg_closure, "verify_rate": verify_rate,
            "completed": completed, "verified": verified, "by_dept": dept_rows,
            "sample_size": len(days)}


@app.get("/qms/pareto", response_class=HTMLResponse)
async def qms_pareto(req: Request):
    """Pareto 차트 — root_cause 빈도 + 누적 % (80/20 법칙)."""
    u = _qms_guard(req)
    if not u:
        return RedirectResponse("/issues", 303)
    with db_session() as c:
        rows = [dict(r) for r in c.execute(
            """SELECT SUBSTR(root_cause,1,40) AS cause, COUNT(*) AS cnt
               FROM issues
               WHERE root_cause IS NOT NULL AND TRIM(root_cause) != ''
               GROUP BY SUBSTR(root_cause,1,40)
               ORDER BY cnt DESC LIMIT 20"""
        ).fetchall()]
    total = sum(r["cnt"] for r in rows) or 1
    cum = 0
    p80_idx = -1
    for i, r in enumerate(rows):
        cum += r["cnt"]
        r["pct"] = round(100.0 * r["cnt"] / total, 1)
        r["cum_pct"] = round(100.0 * cum / total, 1)
        if p80_idx < 0 and r["cum_pct"] >= 80.0:
            p80_idx = i
    max_cnt = rows[0]["cnt"] if rows else 1
    summary = {
        "total_issues": total, "distinct_causes": len(rows),
        "p80_cutoff_idx": p80_idx + 1 if p80_idx >= 0 else 0,
        "p80_pct_of_causes": round(100.0 * (p80_idx + 1) / len(rows), 1) if rows and p80_idx >= 0 else 0,
    }
    return ctx(req, "qms_pareto.html", user=u, active="qms",
               rows=rows, summary=summary, max_cnt=max_cnt)


@app.get("/qms/capa", response_class=HTMLResponse)
async def qms_capa_dashboard(req: Request):
    """CAPA 라이프사이클 관리 — DRAFT/APPROVED/IN_PROGRESS/COMPLETED/VERIFIED 분포 + KPI."""
    u = _qms_guard(req)
    if not u:
        return RedirectResponse("/issues", 303)
    with db_session() as c:
        # 라이프사이클 카운트 (corrective)
        ca_buckets = {"DRAFT": 0, "APPROVED": 0, "IN_PROGRESS": 0, "COMPLETED": 0, "VERIFIED": 0}
        for r in c.execute(
            "SELECT COALESCE(lifecycle_status,'DRAFT') AS s, COUNT(*) AS n "
            "FROM corrective_actions GROUP BY lifecycle_status"
        ).fetchall():
            ca_buckets[r["s"] if r["s"] in ca_buckets else "DRAFT"] = r["n"]
        # 라이프사이클 카운트 (preventive)
        pa_buckets = {"DRAFT": 0, "APPROVED": 0, "IN_PROGRESS": 0, "COMPLETED": 0, "VERIFIED": 0}
        for r in c.execute(
            "SELECT COALESCE(lifecycle_status,'DRAFT') AS s, COUNT(*) AS n "
            "FROM preventive_actions GROUP BY lifecycle_status"
        ).fetchall():
            pa_buckets[r["s"] if r["s"] in pa_buckets else "DRAFT"] = r["n"]
        # CAPA 목록 (최근 30개, 진행중 우선)
        items = [dict(r) for r in c.execute(
            """SELECT ca.id, ca.action, ca.due_date, ca.created_at,
                      ca.completed_at, ca.verified_at,
                      COALESCE(ca.lifecycle_status,'DRAFT') AS lifecycle_status,
                      i.id AS issue_id, i.issue_no, i.title AS issue_title
               FROM corrective_actions ca
               LEFT JOIN issues i ON i.id = ca.issue_id
               ORDER BY (ca.lifecycle_status='VERIFIED'),
                        (ca.lifecycle_status='COMPLETED'),
                        ca.created_at DESC
               LIMIT 30"""
        ).fetchall()]
        kpi = _capa_kpi(c)
    # v5H49: 템플릿 호환 — ca.OPEN/ca.CLOSED 형식도 함께 노출 (legacy)
    ca_view = dict(ca_buckets)
    pa_view = dict(pa_buckets)
    ca_view["OPEN"] = ca_buckets.get("IN_PROGRESS", 0) + ca_buckets.get("APPROVED", 0) + ca_buckets.get("DRAFT", 0)
    ca_view["CLOSED"] = ca_buckets.get("COMPLETED", 0) + ca_buckets.get("VERIFIED", 0)
    pa_view["OPEN"] = pa_buckets.get("IN_PROGRESS", 0) + pa_buckets.get("APPROVED", 0) + pa_buckets.get("DRAFT", 0)
    pa_view["CLOSED"] = pa_buckets.get("COMPLETED", 0) + pa_buckets.get("VERIFIED", 0)
    # 템플릿이 {% set ca = ca_buckets %} 로 ca_buckets 만 읽음 → ca_buckets 에 직접 OPEN/CLOSED 추가
    ca_buckets["OPEN"] = ca_view["OPEN"]
    ca_buckets["CLOSED"] = ca_view["CLOSED"]
    pa_buckets["OPEN"] = pa_view["OPEN"]
    pa_buckets["CLOSED"] = pa_view["CLOSED"]
    return ctx(req, "qms_capa.html", user=u, active="qms",
               ca=ca_view, pa=pa_view,
               ca_buckets=ca_buckets, pa_buckets=pa_buckets,
               items=items, kpi=kpi)


def _doc_audit_log(c, doc_type: str, doc_id: int, action: str,
                   actor_id, note: str = ""):
    """문서 라이프사이클 감사로그 (v5H116 신설).

    경량 통합 테이블 — FTA 증명서 / QC 출하성적서 / 작업지시서의 발급/발행 액션 추적.
    테이블이 없으면 자동 생성 (idempotent). 실패는 본 트랜잭션을 깨지 않도록 흡수."""
    try:
        c.execute(
            """CREATE TABLE IF NOT EXISTS doc_audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                doc_type TEXT NOT NULL,
                doc_id INTEGER NOT NULL,
                action TEXT NOT NULL,
                actor_id INTEGER,
                note TEXT,
                logged_at TEXT DEFAULT (datetime('now','localtime'))
            )"""
        )
        c.execute(
            "CREATE INDEX IF NOT EXISTS idx_doc_audit_log_doc "
            "ON doc_audit_log(doc_type, doc_id)"
        )
        c.execute(
            "INSERT INTO doc_audit_log (doc_type, doc_id, action, actor_id, note) "
            "VALUES (?,?,?,?,?)",
            (doc_type, int(doc_id), action, actor_id, note or ""),
        )
    except Exception:
        pass


def _capa_transition(req: Request, table: str, cid: int, target: str,
                      need_admin: bool, note: str = "") -> bool:
    """CAPA 라이프사이클 전이 헬퍼 — 가드/UPDATE/감사로그 통합. table='corrective_actions' or 'preventive_actions'."""
    u = _qms_capa_guard(req) if need_admin else _qms_guard(req)
    if not u:
        return False
    if table not in ("corrective_actions", "preventive_actions"):
        return False
    col_map = {
        "APPROVED":    ("approved_by", "approved_at"),
        "COMPLETED":   (None, "completed_at"),
        "VERIFIED":    ("verified_by", "verified_at"),
        "IN_PROGRESS": (None, None),
    }
    if target not in col_map:
        return False
    actor_col, ts_col = col_map[target]
    sets = ["lifecycle_status = ?"]
    vals = [target]
    if actor_col:
        sets.append(f"{actor_col} = ?")
        vals.append(u["id"])
    if ts_col:
        sets.append(f"{ts_col} = datetime('now','localtime')")
    if target == "VERIFIED" and note:
        sets.append("effectiveness_note = ?")
        vals.append(note.strip())
    vals.append(cid)
    # v5H117 P3: 라이프사이클 전이 사전 검증 — 역행 차단
    valid_from = {
        "APPROVED":    ("DRAFT",),
        "IN_PROGRESS": ("APPROVED",),
        "COMPLETED":   ("IN_PROGRESS", "APPROVED"),
        "VERIFIED":    ("COMPLETED",),
    }
    with db_session() as c:
        try:
            cur_row = c.execute(
                f"SELECT lifecycle_status FROM {table} WHERE id=?", (cid,)
            ).fetchone()
            if not cur_row:
                return False
            cur_status = cur_row["lifecycle_status"] or "DRAFT"
            allowed = valid_from.get(target, ())
            if allowed and cur_status not in allowed:
                # 이미 동일 상태이거나 역행 — 거부
                return False
            c.execute(f"UPDATE {table} SET {', '.join(sets)} WHERE id=?", vals)
            # 감사로그 (issue_id 조회)
            if table == "corrective_actions":
                row = c.execute(
                    "SELECT issue_id FROM corrective_actions WHERE id=?", (cid,)
                ).fetchone()
            else:
                row = c.execute(
                    "SELECT ca.issue_id FROM preventive_actions pa "
                    "JOIN corrective_actions ca ON ca.id = pa.corrective_id "
                    "WHERE pa.id=?", (cid,)
                ).fetchone()
            if row:
                tag = "ca" if table == "corrective_actions" else "pa"
                c.execute(
                    """INSERT INTO qms_audit_log (issue_id, action, actor, note)
                       VALUES (?, ?, ?, ?)""",
                    (row["issue_id"], f"{tag}_{target.lower()}", u["id"],
                     f"#{cid} → {target}" + (f" · {note[:60]}" if note else "")),
                )
        except Exception:
            return False
    return True


@app.post("/qms/corrective/{cid}/approve")
async def qms_corrective_approve(req: Request, cid: int):
    """시정조치 승인 — DRAFT → APPROVED (admin/ceo/executive/leader)."""
    ok = _capa_transition(req, "corrective_actions", cid, "APPROVED", need_admin=True)
    return RedirectResponse("/qms/capa" if ok else "/qms", 303)


@app.post("/qms/corrective/{cid}/start")
async def qms_corrective_start(req: Request, cid: int):
    """시정조치 진행 — APPROVED → IN_PROGRESS (담당자)."""
    ok = _capa_transition(req, "corrective_actions", cid, "IN_PROGRESS", need_admin=False)
    return RedirectResponse("/qms/capa" if ok else "/qms", 303)


@app.post("/qms/corrective/{cid}/complete")
async def qms_corrective_complete(req: Request, cid: int):
    """시정조치 완료 — IN_PROGRESS → COMPLETED (담당자)."""
    ok = _capa_transition(req, "corrective_actions", cid, "COMPLETED", need_admin=False)
    return RedirectResponse("/qms/capa" if ok else "/qms", 303)


@app.post("/qms/corrective/{cid}/verify")
async def qms_corrective_verify(req: Request, cid: int,
                                  effectiveness_note: str = Form("")):
    """시정조치 효과 검증 — COMPLETED → VERIFIED (admin/leader)."""
    ok = _capa_transition(req, "corrective_actions", cid, "VERIFIED",
                          need_admin=True, note=effectiveness_note)
    return RedirectResponse("/qms/capa" if ok else "/qms", 303)


@app.post("/qms/preventive/{pid}/approve")
async def qms_preventive_approve(req: Request, pid: int):
    """예방조치 승인 — DRAFT → APPROVED."""
    ok = _capa_transition(req, "preventive_actions", pid, "APPROVED", need_admin=True)
    return RedirectResponse("/qms/capa" if ok else "/qms", 303)


@app.post("/qms/preventive/{pid}/complete")
async def qms_preventive_complete(req: Request, pid: int):
    """예방조치 완료 — IN_PROGRESS → COMPLETED."""
    ok = _capa_transition(req, "preventive_actions", pid, "COMPLETED", need_admin=False)
    return RedirectResponse("/qms/capa" if ok else "/qms", 303)


@app.post("/qms/preventive/{pid}/verify")
async def qms_preventive_verify(req: Request, pid: int,
                                  effectiveness_note: str = Form("")):
    """예방조치 효과 검증 — COMPLETED → VERIFIED."""
    ok = _capa_transition(req, "preventive_actions", pid, "VERIFIED",
                          need_admin=True, note=effectiveness_note)
    return RedirectResponse("/qms/capa" if ok else "/qms", 303)


# =====================================================
# 환율·단가 강화 1차 (2026-04-26 Top10 #9 P4 구매팀 월 1회)
# 외부 환율 API 미사용 (수동 입력 + CSV 업로드만)
# =====================================================
def _rates_guard(req: Request):
    """환율·단가 가드 — can_use_logistics 위임 (구매팀 + admin/ceo/executive)."""
    u = get_user(req)
    if not u:
        return None
    if not can_use_logistics(u):
        return None
    return u


@app.get("/rates/dashboard", response_class=HTMLResponse)
async def rates_dashboard_page(request: Request):
    u = _rates_guard(request)
    if not u:
        return RedirectResponse("/login", 303)
    items = exchange_rates_list(limit=60, currency="")
    latest = exchange_rates_latest()
    alerts = rate_alerts_list(active_only=True)
    # 트렌드: 통화별 최근 14일 (대시보드 KPI)
    with db_session() as c:
        trend_rows = [dict(r) for r in c.execute(
            """SELECT from_currency, rate_date, rate
               FROM exchange_rates
               WHERE to_currency='KRW' AND rate_date >= date('now','-30 days')
               ORDER BY rate_date DESC, from_currency"""
        ).fetchall()]
    return ctx(request, "rates_dashboard.html", user=u, items=items,
               latest=latest, alerts=alerts, trend=trend_rows,
               CURRENCIES=CURRENCIES, active="rates")


@app.post("/rates/upload")
async def rates_csv_upload(request: Request, csv_text: str = Form(...)):
    """CSV 일괄 업로드 — 외부 API 미호출. 헤더 필수: rate_date,from_currency(또는 from),to_currency(또는 to),rate.
    S2-1 헤더 가드: 첫 비주석 행을 헤더로 검증, BOM 제거, 잘못된 형식은 400 반환."""
    u = _rates_guard(request)
    if not u:
        return RedirectResponse("/login", 303)
    raw = csv_text or ""
    if raw.startswith("﻿"):  # S2-1: UTF-8 BOM 제거
        raw = raw.lstrip("﻿")
    lines = [ln.strip() for ln in raw.splitlines()]
    lines = [ln for ln in lines if ln and not ln.startswith("#")]
    if not lines:
        return JSONResponse({"error": "CSV 비어있음"}, 400)
    header_cols = [c.strip().lower().lstrip("﻿") for c in lines[0].split(",")]
    HEADER_ALIAS = {"from": "from_currency", "to": "to_currency",
                    "date": "rate_date", "currency": "from_currency"}
    norm_header = [HEADER_ALIAS.get(h, h) for h in header_cols]
    REQUIRED = {"rate_date", "from_currency", "rate"}
    missing = REQUIRED - set(norm_header)
    if missing:
        return JSONResponse(
            {"error": f"CSV 헤더 누락: {sorted(missing)} · 필수=rate_date,from_currency(또는 from),rate"},
            400,
        )
    idx = {k: norm_header.index(k) for k in norm_header}
    rows = []
    for ln in lines[1:]:
        cols = [c.strip() for c in ln.split(",")]
        if not any(cols):
            continue  # S2-1: 빈 행 정리
        try:
            rd = cols[idx["rate_date"]] if idx.get("rate_date", -1) >= 0 else ""
            fc = cols[idx["from_currency"]] if idx.get("from_currency", -1) >= 0 else ""
            rt = cols[idx["rate"]] if idx.get("rate", -1) >= 0 else ""
        except IndexError:
            continue
        if not (rd and fc and rt):
            continue  # S2-1: 필수 빈 값 정리
        rows.append({
            "rate_date": rd,
            "from_currency": fc,
            "to_currency": (cols[idx["to_currency"]] if idx.get("to_currency", -1) >= 0
                            and idx["to_currency"] < len(cols) and cols[idx["to_currency"]]
                            else "KRW"),
            "rate": rt,
            "source": (cols[idx["source"]] if idx.get("source", -1) >= 0
                       and idx["source"] < len(cols) else "CSV"),
            "note": (cols[idx["note"]] if idx.get("note", -1) >= 0
                     and idx["note"] < len(cols) else ""),
        })
    res = exchange_rates_csv_upload(rows, user_id=u["id"])
    # S3-1 옵션 A: 업로드 후 자동 알림 발동 검사 (통화별 평균 rate 사용)
    fired_total = 0
    by_cur: dict = {}
    for r in rows:
        try:
            by_cur.setdefault(r["from_currency"].upper(), []).append(float(r["rate"]))
        except Exception:
            pass
    for cur, vals in by_cur.items():
        if vals:
            fired_total += check_rate_alerts(cur, vals[-1])  # 최신 행 기준
    msg = f"OK={res['ok']}/NG={res['ng']}/FIRED={fired_total}"
    return RedirectResponse(f"/rates/dashboard?upload={msg}", 303)


@app.get("/rates/cost-sim/{part_id}", response_class=HTMLResponse)
async def rates_cost_sim_page(request: Request, part_id: int):
    u = _rates_guard(request)
    if not u:
        return RedirectResponse("/login", 303)
    part = _logi.parts_get(part_id)
    if not part:
        return RedirectResponse("/parts", 303)
    sims = cost_simulations_list(part_id, limit=20)
    latest = exchange_rates_latest()
    active = part_active_price(part_id) or {}
    return ctx(request, "rates_cost_sim.html", user=u,
               part=dict(part), sims=sims, latest=latest,
               active_price=active, CURRENCIES=CURRENCIES, active="rates")


@app.post("/rates/cost-sim")
async def rates_cost_sim_submit(
    request: Request,
    part_id: str = Form(...),
    base_currency: str = Form("USD"),
    target_currency: str = Form("KRW"),
    exchange_rate: str = Form(...),
    unit_price_base: str = Form(...),
    margin_pct: str = Form("0"),
    note: str = Form(""),
):
    u = _rates_guard(request)
    if not u:
        return RedirectResponse("/login", 303)
    # OPS-V3 [정적점검 결함]: 입력 검증 누락 → ValueError/0음수 차단
    try:
        pid = int(part_id)
        base = float(unit_price_base)
        rate = float(exchange_rate)
        margin = float(margin_pct or 0)
    except (ValueError, TypeError):
        return RedirectResponse("/rates?error=입력값+형식+오류", 303)
    if pid <= 0 or base <= 0 or rate <= 0:
        return RedirectResponse(f"/rates/cost-sim/{pid}?error=part_id/단가/환율은+양수여야+합니다", 303)
    if margin < -100:
        return RedirectResponse(f"/rates/cost-sim/{pid}?error=마진율은+-100%+이상이어야+합니다", 303)
    if base_currency == target_currency and abs(rate - 1.0) > 1e-9:
        # 같은 통화인데 환율 != 1.0 → 사용자 실수 가능 (경고만, 계속 진행)
        pass
    target = base * rate * (1.0 + margin / 100.0)
    try:
        cost_simulation_create({
            "part_id": pid,
            "base_currency": base_currency,
            "target_currency": target_currency,
            "exchange_rate": rate,
            "unit_price_base": base,
            "unit_price_target": target,
            "margin_pct": margin,
            "note": note,
        }, user_id=u["id"])
    except Exception as e:
        return RedirectResponse(f"/rates/cost-sim/{pid}?error={e}", 303)
    return RedirectResponse(f"/rates/cost-sim/{pid}?saved=1", 303)


@app.get("/rates/price-history/{part_id}", response_class=HTMLResponse)
async def rates_price_history_page(request: Request, part_id: int):
    u = _rates_guard(request)
    if not u:
        return RedirectResponse("/login", 303)
    part = _logi.parts_get(part_id)
    if not part:
        return RedirectResponse("/parts", 303)
    history = price_change_history_list(part_id, limit=80)
    return ctx(request, "rates_history.html", user=u,
               part=dict(part), history=history, active="rates")


@app.get("/rates/alerts", response_class=HTMLResponse)
async def rates_alerts_page(request: Request):
    u = _rates_guard(request)
    if not u:
        return RedirectResponse("/login", 303)
    alerts = rate_alerts_list(active_only=False)
    latest = exchange_rates_latest()
    return ctx(request, "rates_alerts.html", user=u, alerts=alerts,
               latest=latest, CURRENCIES=CURRENCIES, active="rates")


@app.post("/rates/alerts")
async def rates_alerts_submit(
    request: Request,
    target_currency: str = Form(...),
    threshold: str = Form(...),
    direction: str = Form("above"),
):
    u = _rates_guard(request)
    if not u:
        return RedirectResponse("/login", 303)
    try:
        rate_alert_create(target_currency, float(threshold), direction, u["id"])
    except Exception as e:
        return RedirectResponse(f"/rates/alerts?error={e}", 303)
    # 알림시스템 통합 (사이클 2026-04-26) — 등록자에게 RATE 알림
    notify_user(
        u["id"], "RATE",
        f"💱 환율 알림 등록 — {target_currency}",
        body=f"임계 {threshold} ({direction})",
        link="/rates/alerts",
    )
    return RedirectResponse("/rates/alerts?saved=1", 303)


# =====================================================
# 사이클 51 S2-4차 (2026-04-27) — 안전재고 / 재발주점 / 발주 추천 / 알림 자동
# 라우트 +4: GET /stock/safety · POST /stock/safety/{part_id}
#           GET /stock/reorder-recommendations · POST /stock/alerts/check
# 권한: _s2_guard (구매팀 또는 admin/ceo). 알림 트리거는 admin/ceo 한정.
# =====================================================
@app.get("/stock/safety", response_class=HTMLResponse)
async def stock_safety_page(req: Request, q: str = ""):
    """안전재고 설정 페이지 — parts 별 safety_stock / reorder_point / reorder_qty."""
    u = _s2_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    from .database import parts_safety_settings_list
    items = parts_safety_settings_list(q=q)
    return ctx(req, "stock_safety.html", user=u, active="stock",
               items=items, q=q)


@app.post("/stock/safety/{part_id}")
async def stock_safety_save(req: Request, part_id: int):
    """안전재고 등록/수정 — safety_stock / reorder_point / reorder_qty 일괄 갱신."""
    u = _s2_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    form = await req.form()
    from .database import parts_safety_update
    ok = parts_safety_update(
        part_id,
        safety_stock=form.get("safety_stock") or 0,
        reorder_point=form.get("reorder_point") or 0,
        reorder_qty=form.get("reorder_qty") or 0,
    )
    suffix = "saved=1" if ok else "error=1"
    q = form.get("q") or ""
    qs = f"?q={q}&{suffix}" if q else f"?{suffix}"
    return RedirectResponse(f"/stock/safety{qs}", 303)


@app.get("/stock/reorder-recommendations", response_class=HTMLResponse)
async def stock_reorder_page(req: Request, limit: int = 200):
    """발주 추천 — 재발주점 미달 부품 + 권장 발주량 + 우선순위(HIGH/MID/LOW)."""
    u = _s2_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    # OPS-V4 [정적점검]: limit 상한 가드 (메모리 폭주 방지)
    limit = max(1, min(int(limit or 200), 1000))
    from .database import recommend_reorders
    items = recommend_reorders(limit=limit)
    high = sum(1 for r in items if r["priority"] == "HIGH")
    mid = sum(1 for r in items if r["priority"] == "MID")
    low = sum(1 for r in items if r["priority"] == "LOW")
    return ctx(req, "stock_reorder.html", user=u, active="stock",
               items=items, high_cnt=high, mid_cnt=mid, low_cnt=low)


@app.post("/stock/alerts/check")
async def stock_alerts_check(req: Request):
    """알림 트리거 — 관리자 한정. check_stock_alerts() 실행 후 결과 리턴."""
    u = _s2_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    if u.get("role") not in ("admin", "ceo"):
        return JSONResponse({"error": "관리자 전용"}, 403)
    from .database import check_stock_alerts
    out = check_stock_alerts()
    return JSONResponse({
        "ok": True,
        "checked": out.get("checked", 0),
        "alerts_sent": out.get("alerts_sent", 0),
        "low_count": len(out.get("low_parts", [])),
    })


# =====================================================
# 사이클 54 환율·단가 1차 (2026-04-27)
# 외부 API 0건. 수동 입력 + 이력 보존 + 단가 자동 적용 흐름.
# 권한: admin / finance(team_id=3 관리팀) / ceo (+ logistics 권한자)
# =====================================================
def _fx_guard(user) -> bool:
    """환율 관리 권한 — admin/ceo/executive + 관리팀(finance) + logistics 권한자"""
    if not user:
        return False
    role = user.get("role") if isinstance(user, dict) else user["role"]
    if role in ("admin", "ceo", "executive"):
        return True
    team_id = user.get("team_id") if isinstance(user, dict) else user["team_id"]
    if team_id == 3:  # 관리팀 = finance
        return True
    return can_use_logistics(user)


@app.get("/fx/rates", response_class=HTMLResponse)
async def fx_rates_page(request: Request, currency: str = ""):
    """
    환율 목록 + 입력 폼 (사이클 54 1차).

    외부 도메인 코드 호환용 영문 표준 경로 환율 관리 endpoint.
    사내 사용자는 기존 /rates 사용 권장 (사이드바 링크).
    본 endpoint는 외부 시스템·API 통합용으로 별도 운영.
    사이클 55 (2026-04-27) S2-1 A안 적용.
    """
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not _fx_guard(u):
        return RedirectResponse("/home", 303)
    items = exchange_rates_list(limit=200, currency=currency)
    latest = exchange_rates_latest()
    return ctx(request, "fx_rates.html", user=u, items=items, latest=latest,
               currency=currency, CURRENCIES=CURRENCIES, active="fx_rates")


@app.post("/fx/rates")
async def fx_rates_create(
    request: Request,
    rate_date: str = Form(...),
    from_currency: str = Form(...),
    to_currency: str = Form("KRW"),
    rate: str = Form(...),
    source: str = Form("수동"),
    note: str = Form(""),
):
    """
    환율 신규 등록 — 같은 날짜+통화쌍 중복 시 UPSERT (exchange_rate_create 내부 처리).

    외부 도메인 코드 호환용 영문 표준 경로 환율 등록 endpoint.
    사내 사용자는 기존 /rates 사용 권장 (사이드바 링크).
    본 endpoint는 외부 시스템·API 통합용으로 별도 운영.
    사이클 55 (2026-04-27) S2-1 A안 적용.
    """
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not _fx_guard(u):
        return RedirectResponse("/home", 303)
    try:
        exchange_rate_create({
            "rate_date": rate_date,
            "from_currency": from_currency,
            "to_currency": to_currency,
            "rate": float(rate),
            "source": source,
            "note": note,
        }, user_id=u["id"])
    except Exception as e:
        from urllib.parse import quote
        return RedirectResponse(f"/fx/rates?error={quote(str(e))}", 303)
    return RedirectResponse("/fx/rates?success=1", 303)


@app.get("/parts/{part_id}/prices", response_class=HTMLResponse)
async def part_prices_page(request: Request, part_id: int):
    """부품 단가 이력 + 입력 폼 (사이클 54 1차)."""
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not (_fx_guard(u) or can_use_sales(u)):
        return RedirectResponse("/home", 303)
    part = _logi.parts_get(part_id)
    if not part:
        return RedirectResponse("/parts", 303)
    prices = part_prices_list(part_id)
    active = part_active_price(part_id)
    latest_cost = get_latest_part_price(part_id, price_type="cost")
    with db_session() as c:
        suppliers = [dict(r) for r in c.execute(
            "SELECT id, name FROM suppliers WHERE is_active=1 ORDER BY name"
        ).fetchall()]
    return ctx(request, "part_prices.html", user=u,
               part=dict(part), prices=prices, active_price=active,
               latest_cost=latest_cost, suppliers=suppliers,
               CURRENCIES=CURRENCIES, PRICE_TYPES=PRICE_TYPES,
               active="parts")


@app.post("/parts/{part_id}/prices")
async def part_prices_create(
    request: Request, part_id: int,
    supplier_id: str = Form(""),
    price_type: str = Form("견적"),
    unit_price: str = Form(...),
    currency: str = Form("KRW"),
    effective_from: str = Form(...),
    effective_to: str = Form(""),
    negotiated_at: str = Form(""),
    min_qty: str = Form("0"),
    max_qty: str = Form(""),
    note: str = Form(""),
):
    """
    부품 단가 신규 등록 (사이클 54 1차) — 이력 보존 (UPDATE 아닌 INSERT).
    사이클 55 S4-1 보강: negotiated_at / min_qty / max_qty + price_change_log 자동 INSERT.
    """
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not (_fx_guard(u) or can_use_sales(u)):
        return RedirectResponse("/home", 303)
    sid = int(supplier_id) if supplier_id and supplier_id.isdigit() else None
    new_price = float(unit_price)
    # S4-1: 직전 활성 단가 조회 (애플리케이션 레벨 훅)
    prev = part_active_price(part_id, supplier_id=sid or 0) or {}
    old_price = prev.get("unit_price")
    try:
        part_price_create({
            "part_id": part_id,
            "supplier_id": sid,
            "price_type": price_type,
            "unit_price": new_price,
            "currency": currency,
            "effective_from": effective_from,
            "effective_to": effective_to or None,
            "negotiated_at": negotiated_at or None,
            "min_qty": float(min_qty or 0),
            "max_qty": float(max_qty) if max_qty else None,
            "note": note,
        }, user_id=u["id"])
        # S4-1: price_change_history 자동 INSERT (변동률 자동 계산)
        try:
            price_change_log(part_id, sid, old_price, new_price,
                             effective_from, u["id"], note=note or "")
        except Exception:
            pass  # 본 등록은 성공했으므로 훅 실패는 흡수
    except Exception as e:
        from urllib.parse import quote
        return RedirectResponse(f"/parts/{part_id}/prices?error={quote(str(e))}", 303)
    return RedirectResponse(f"/parts/{part_id}/prices?success=1", 303)


@app.get("/api/parts/{pid}/active-price")
async def api_part_active_price(request: Request, pid: int, currency: str = "KRW"):
    """v5H116: PO 폼 단가 자동완성용 — 부품의 활성 단가 JSON 반환.

    parts.unit_price 폴백 대신 part_prices 의 최신 활성 row 사용.
    응답: {price, currency, supplier_id, supplier_name, price_type, source}
    매칭 실패 시 parts.unit_price 폴백 (source='parts.unit_price')."""
    u = get_user(request)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    cur = (currency or "KRW").strip().upper() or "KRW"
    ap = part_active_price(pid)
    if ap and (str(ap.get("currency") or "KRW").upper() == cur):
        return JSONResponse({
            "price": float(ap.get("unit_price") or 0),
            "currency": ap.get("currency") or "KRW",
            "supplier_id": ap.get("supplier_id"),
            "supplier_name": ap.get("supplier_name") or "",
            "price_type": ap.get("price_type") or "",
            "source": "part_prices",
        })
    # 폴백: parts.unit_price (있을 때만)
    try:
        with db_session() as c:
            row = c.execute(
                "SELECT unit_price, COALESCE(currency,'KRW') AS currency "
                "FROM parts WHERE id=?", (pid,)
            ).fetchone()
        if row:
            return JSONResponse({
                "price": float(row["unit_price"] or 0),
                "currency": row["currency"] or "KRW",
                "supplier_id": None,
                "supplier_name": "",
                "price_type": "",
                "source": "parts.unit_price",
            })
    except Exception:
        pass
    return JSONResponse({
        "price": 0, "currency": cur, "supplier_id": None,
        "supplier_name": "", "price_type": "", "source": "none",
    })


# =====================================================
# 사이클 75 — FTA 원산지증명서 (C/O) 발급 모듈 (2026-04-27)
# 04 시뮬 MISSING #1: 안지연 본업. KAFTA/KEUFTA/RCEP 5종 처리.
# 외부 PDF 라이브러리 0건 (HTML 인쇄 view + window.print() 사이클 63 패턴).
# v2 본체 미접촉 / 핫패치 130 보존 / SQL 파라미터 바인딩.
# =====================================================
from .database import (create_fta_certificate, get_fta_certificates,
                       get_fta_certificate)

FTA_TYPES = [
    ("KAFTA",  "한·아세안 FTA (Korea-ASEAN)"),
    ("KEUFTA", "한·EU FTA (Korea-EU)"),
    ("KCFTA",  "한·중 FTA (Korea-China)"),
    ("KVFTA",  "한·베트남 FTA (Korea-Vietnam)"),
    ("RCEP",   "역내포괄적경제동반자협정 (RCEP)"),
]
FTA_TYPE_CODES = {code for code, _label in FTA_TYPES}
ORIGIN_COUNTRIES = [("KR", "한국"), ("VN", "베트남"), ("CN", "중국")]


@app.get("/export/fta", response_class=HTMLResponse)
async def export_fta_list(req: Request):
    """원산지증명서 목록 (안지연 본업 · 04 시뮬 MISSING #1)."""
    u = _export_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    status = req.query_params.get("status") or None
    fta_type = req.query_params.get("fta_type") or None
    items = get_fta_certificates(status=status, fta_type=fta_type, limit=300)
    return ctx(req, "fta_list.html", user=u, active="export",
               items=items, FTA_TYPES=FTA_TYPES,
               filter_status=status or "", filter_fta_type=fta_type or "")


@app.get("/export/fta/new", response_class=HTMLResponse)
async def export_fta_new_form(req: Request):
    """원산지증명서 신규 발급 폼."""
    u = _export_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        customers = [dict(r) for r in c.execute(
            "SELECT id, name, COALESCE(tier,'') AS country FROM customers ORDER BY name LIMIT 200"
        ).fetchall()]
        export_orders = [dict(r) for r in c.execute(
            """SELECT eo.id, eo.buyer, COALESCE(o.order_no,'-') AS order_no
               FROM export_orders eo
               LEFT JOIN orders o ON o.id = eo.order_id
               ORDER BY eo.id DESC LIMIT 100"""
        ).fetchall()]
        parts_options = [dict(r) for r in c.execute(
            "SELECT id, part_name AS name, spec, unit FROM parts ORDER BY part_name LIMIT 300"
        ).fetchall()]
    return ctx(req, "fta_form.html", user=u, active="export",
               customers=customers, export_orders=export_orders,
               parts_options=parts_options,
               FTA_TYPES=FTA_TYPES, ORIGIN_COUNTRIES=ORIGIN_COUNTRIES,
               cert=None)


@app.post("/export/fta")
async def export_fta_create(req: Request):
    """원산지증명서 신규 등록. cert_no 자동 생성 (FTA-YYYY-####)."""
    u = _export_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    fta_type = (form.get("fta_type") or "").strip().upper()
    if fta_type not in FTA_TYPE_CODES:
        return JSONResponse({"error": "fta_type 필수"}, 400)
    customer_id_raw = form.get("customer_id")
    customer_id = int(customer_id_raw) if (customer_id_raw and customer_id_raw.isdigit()) else None
    customer_name = (form.get("customer_name") or "").strip() or None
    customer_address = (form.get("customer_address") or "").strip() or None
    customer_country = (form.get("customer_country") or "").strip() or None
    eo_raw = form.get("export_order_id")
    export_order_id = int(eo_raw) if (eo_raw and eo_raw.isdigit()) else None
    export_invoice_no = (form.get("export_invoice_no") or "").strip() or None
    export_date = (form.get("export_date") or "").strip() or None
    origin_country = (form.get("origin_country") or "KR").strip() or "KR"
    currency = (form.get("currency") or "USD").strip() or "USD"
    remarks = (form.get("remarks") or "").strip() or None
    # 라인 파싱 — line_part_id[], line_part_name[], line_hs[], line_qty[], line_unit[], line_unit_price[], line_origin[]
    part_ids = form.getlist("line_part_id") if hasattr(form, "getlist") else []
    part_names = form.getlist("line_part_name") if hasattr(form, "getlist") else []
    hs_codes = form.getlist("line_hs") if hasattr(form, "getlist") else []
    qtys = form.getlist("line_qty") if hasattr(form, "getlist") else []
    units = form.getlist("line_unit") if hasattr(form, "getlist") else []
    unit_prices = form.getlist("line_unit_price") if hasattr(form, "getlist") else []
    origins = form.getlist("line_origin") if hasattr(form, "getlist") else []
    items = []
    total_value = 0.0
    n = max(len(part_names), len(hs_codes), len(qtys))
    # v5H116: 라인 음수/0 검증 — FTA 증명서 정합성
    line_errors = []
    for i in range(n):
        pname = (part_names[i] if i < len(part_names) else "").strip()
        hs = (hs_codes[i] if i < len(hs_codes) else "").strip()
        if not pname and not hs:
            continue
        pid_raw = part_ids[i] if i < len(part_ids) else ""
        try:
            pid = int(pid_raw) if pid_raw and pid_raw.isdigit() else None
        except Exception:
            pid = None
        try:
            qv = float(qtys[i]) if i < len(qtys) and qtys[i] else 0.0
        except Exception:
            qv = 0.0
        try:
            up = float(unit_prices[i]) if i < len(unit_prices) and unit_prices[i] else 0.0
        except Exception:
            up = 0.0
        if qv <= 0:
            line_errors.append(f"라인 {i+1}: 수량은 0보다 커야 합니다")
            continue
        if up < 0:
            line_errors.append(f"라인 {i+1}: 단가는 음수일 수 없습니다")
            continue
        line_total = round(qv * up, 2)
        total_value += line_total
        items.append({
            "part_id": pid,
            "part_name": pname or None,
            "hs_code": hs or None,
            "qty": qv,
            "unit": (units[i] if i < len(units) else "") or None,
            "unit_price": up,
            "origin_country": (origins[i] if i < len(origins) else "") or origin_country,
            "total": line_total,
        })
    if line_errors:
        from urllib.parse import quote as _q
        return RedirectResponse(
            "/export/fta/new?error=" + _q(" / ".join(line_errors[:3])), 303
        )
    if not items:
        from urllib.parse import quote as _q
        return RedirectResponse(
            "/export/fta/new?error=" + _q("라인 1건 이상 필요합니다"), 303
        )
    issuer_id = u["id"] if isinstance(u, dict) else u["id"]
    issuer_name = (u.get("name") if isinstance(u, dict) else u["name"]) or None
    cert_id, cert_no = create_fta_certificate(
        fta_type=fta_type,
        customer_id=customer_id,
        customer_name=customer_name,
        customer_address=customer_address,
        customer_country=customer_country,
        export_order_id=export_order_id,
        export_invoice_no=export_invoice_no,
        export_date=export_date,
        origin_country=origin_country,
        total_value=total_value,
        currency=currency,
        issuer_id=issuer_id,
        issuer_name=issuer_name,
        items=items,
        remarks=remarks,
        created_by=issuer_id,
    )
    return RedirectResponse(f"/export/fta/{cert_id}", 303)


@app.get("/export/fta/{cert_id}", response_class=HTMLResponse)
async def export_fta_detail(req: Request, cert_id: int):
    """원산지증명서 상세 (라인 + 발급/인쇄 액션)."""
    u = _export_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    cert = get_fta_certificate(cert_id)
    if not cert:
        return RedirectResponse("/export/fta", 303)
    company = _company_info_dict()
    with db_session() as c:
        customers = [dict(r) for r in c.execute(
            "SELECT id, name, COALESCE(tier,'') AS country FROM customers ORDER BY name LIMIT 200"
        ).fetchall()]
        export_orders = [dict(r) for r in c.execute(
            """SELECT eo.id, eo.buyer, COALESCE(o.order_no,'-') AS order_no
               FROM export_orders eo
               LEFT JOIN orders o ON o.id = eo.order_id
               ORDER BY eo.id DESC LIMIT 100"""
        ).fetchall()]
        parts_options = [dict(r) for r in c.execute(
            "SELECT id, part_name AS name, spec, unit FROM parts ORDER BY part_name LIMIT 300"
        ).fetchall()]
    return ctx(req, "fta_form.html", user=u, active="export",
               cert=cert, company=company,
               customers=customers, export_orders=export_orders,
               parts_options=parts_options,
               FTA_TYPES=FTA_TYPES, ORIGIN_COUNTRIES=ORIGIN_COUNTRIES)


@app.post("/export/fta/{cert_id}/issue")
async def export_fta_issue(req: Request, cert_id: int):
    """원산지증명서 발급 확정 (DRAFT → ISSUED).

    v5H116: 존재/상태 검증 강화 + issued_by 기록 + 감사로그(qms_audit_log issue_id=None 사용 불가하므로
    order_status_history 패턴 대신 doc_audit_log 테이블 통합 — 없으면 즉시 생성)."""
    u = _export_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    cid = int(cert_id)
    with db_session() as c:
        row = c.execute(
            "SELECT id, status FROM fta_certificates WHERE id=?", (cid,)
        ).fetchone()
        if not row:
            return JSONResponse({"error": "원산지증명서 없음"}, 404)
        if row["status"] != "DRAFT":
            from urllib.parse import quote as _q
            return RedirectResponse(
                f"/export/fta/{cid}?error=" + _q(f"이미 {row['status']} 상태입니다"), 303
            )
        # issued_by 컬럼이 있으면 기록 (없으면 무시)
        try:
            c.execute(
                "UPDATE fta_certificates SET status='ISSUED', "
                "issued_at=datetime('now','localtime'), issued_by=? "
                "WHERE id=? AND status='DRAFT'",
                (u["id"], cid),
            )
        except Exception:
            c.execute(
                "UPDATE fta_certificates SET status='ISSUED', "
                "issued_at=datetime('now','localtime') WHERE id=? AND status='DRAFT'",
                (cid,),
            )
        _doc_audit_log(c, "fta_certificate", cid, "ISSUE", u["id"], "DRAFT → ISSUED")
    return RedirectResponse(f"/export/fta/{cid}", 303)


@app.get("/export/fta/{cert_id}/print", response_class=HTMLResponse)
async def export_fta_print(req: Request, cert_id: int):
    """원산지증명서 인쇄 view (HTML + window.print(), 외부 PDF 0건)."""
    u = _export_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    cert = get_fta_certificate(cert_id)
    if not cert:
        return RedirectResponse("/export/fta", 303)
    company = _company_info_dict()
    fta_label_map = {code: label for code, label in FTA_TYPES}
    return ctx(req, "fta_print.html", user=u,
               cert=cert, company=company,
               fta_label=fta_label_map.get(cert.get("fta_type"), cert.get("fta_type") or "FTA"))


# =====================================================
# 검사기 출하성적서 QC INSPECTION REPORT (사이클 76 · 2026-04-27)
# 김정록 본업 — 검사기 반복성 검증 + 출하성적서 작성 (04 시뮬 MISSING #2)
# 외부 PDF 라이브러리 0건 (HTML 인쇄 view + window.print() 사이클 75 패턴)
# v2 본체 미접촉 / 핫패치 130 보존 / SQL 파라미터 바인딩
# =====================================================
from .database import (create_qc_inspection_report, get_qc_inspection_reports,
                       get_qc_inspection_report, QC_STANDARD_ITEMS)

QC_OVERALL_OPTIONS = [
    ("PASS",              "합격 (PASS)"),
    ("CONDITIONAL_PASS",  "조건부 합격 (CONDITIONAL PASS)"),
    ("FAIL",              "불합격 (FAIL)"),
]


def _qcr_guard(req: Request):
    """검사기 출하성적서 가드 — admin/ceo/executive OR 품질팀 OR 검사기팀 OR can_use_quality."""
    u = get_user(req)
    if not u:
        return None
    role = u.get("role") if isinstance(u, dict) else u["role"]
    if role in ("admin", "ceo", "executive"):
        return u
    if u.get("can_use_quality"):
        return u
    tid = u.get("team_id")
    if tid:
        with db_session() as c:
            row = c.execute(
                "SELECT name FROM teams WHERE id=?", (tid,)
            ).fetchone()
        if row:
            tname = row["name"] or ""
            if "품질" in tname or "검사기" in tname or "QA" in tname.upper():
                return u
    return None


@app.get("/qc/inspection-reports", response_class=HTMLResponse)
async def qc_report_list(req: Request):
    """검사기 출하성적서 목록 (김정록 본업 · 04 시뮬 MISSING #2)."""
    u = _qcr_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    status = req.query_params.get("status") or None
    overall = req.query_params.get("overall") or None
    items = get_qc_inspection_reports(status=status, overall=overall, limit=300)
    return ctx(req, "qc_report_list.html", user=u, active="qc_reports",
               items=items, QC_OVERALL_OPTIONS=QC_OVERALL_OPTIONS,
               filter_status=status or "", filter_overall=overall or "")


@app.get("/qc/inspection-reports/new", response_class=HTMLResponse)
async def qc_report_new_form(req: Request):
    """검사기 출하성적서 신규 발급 폼."""
    u = _qcr_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        customers = [dict(r) for r in c.execute(
            "SELECT id, name, COALESCE(tier,'') AS country FROM customers ORDER BY name LIMIT 200"
        ).fetchall()]
        orders = [dict(r) for r in c.execute(
            """SELECT o.id, o.order_no, COALESCE(cu.name,'-') AS cust_name
               FROM orders o
               LEFT JOIN customers cu ON cu.id = o.customer_id
               ORDER BY o.id DESC LIMIT 100"""
        ).fetchall()]
        parts_options = [dict(r) for r in c.execute(
            "SELECT id, part_name AS name, spec, unit FROM parts ORDER BY part_name LIMIT 300"
        ).fetchall()]
    return ctx(req, "qc_report_form.html", user=u, active="qc_reports",
               customers=customers, orders=orders, parts_options=parts_options,
               QC_OVERALL_OPTIONS=QC_OVERALL_OPTIONS,
               QC_STANDARD_ITEMS=QC_STANDARD_ITEMS,
               report=None)


@app.post("/qc/inspection-reports")
async def qc_report_create(req: Request):
    """검사기 출하성적서 신규 등록. report_no 자동 생성 (QCR-YYYY-####)."""
    u = _qcr_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    customer_id_raw = form.get("customer_id")
    customer_id = int(customer_id_raw) if (customer_id_raw and customer_id_raw.isdigit()) else None
    customer_name = (form.get("customer_name") or "").strip() or None
    order_id_raw = form.get("order_id")
    order_id = int(order_id_raw) if (order_id_raw and order_id_raw.isdigit()) else None
    order_no = (form.get("order_no") or "").strip() or None
    part_id_raw = form.get("part_id")
    part_id = int(part_id_raw) if (part_id_raw and part_id_raw.isdigit()) else None
    machine_model = (form.get("machine_model") or "").strip() or None
    machine_serial = (form.get("machine_serial") or "").strip() or None
    inspection_date = (form.get("inspection_date") or "").strip() or None
    overall = (form.get("overall") or "PASS").strip().upper()
    if overall not in ("PASS", "FAIL", "CONDITIONAL_PASS"):
        overall = "PASS"
    qa_raw = form.get("qa_manager_id")
    qa_manager_id = int(qa_raw) if (qa_raw and qa_raw.isdigit()) else None
    qa_manager_name = (form.get("qa_manager_name") or "").strip() or None
    remarks = (form.get("remarks") or "").strip() or None
    # 라인 파싱 — line_item_name[], line_spec[], line_measured[], line_judgment[], line_remarks[]
    item_names = form.getlist("line_item_name") if hasattr(form, "getlist") else []
    specs = form.getlist("line_spec") if hasattr(form, "getlist") else []
    measureds = form.getlist("line_measured") if hasattr(form, "getlist") else []
    judgments = form.getlist("line_judgment") if hasattr(form, "getlist") else []
    line_remarks_l = form.getlist("line_remarks") if hasattr(form, "getlist") else []
    items = []
    fail_seen = False
    n = max(len(item_names), len(specs), len(measureds))
    for i in range(n):
        nm = (item_names[i] if i < len(item_names) else "").strip()
        if not nm:
            continue
        jdg = (judgments[i] if i < len(judgments) else "PASS").strip().upper() or "PASS"
        if jdg not in ("PASS", "FAIL", "NA"):
            jdg = "PASS"
        if jdg == "FAIL":
            fail_seen = True
        items.append({
            "item_name": nm,
            "spec_value": (specs[i] if i < len(specs) else "").strip() or None,
            "measured_value": (measureds[i] if i < len(measureds) else "").strip() or None,
            "judgment": jdg,
            "remarks": (line_remarks_l[i] if i < len(line_remarks_l) else "").strip() or None,
        })
    # overall 자동 보정 — 라인 중 FAIL 있고 사용자가 PASS 만 선택했으면 CONDITIONAL_PASS 로 보정
    if fail_seen and overall == "PASS":
        overall = "CONDITIONAL_PASS"
    inspector_id = u["id"] if isinstance(u, dict) else u["id"]
    inspector_name = (u.get("name") if isinstance(u, dict) else u["name"]) or None
    report_id, report_no = create_qc_inspection_report(
        customer_id=customer_id,
        customer_name=customer_name,
        order_id=order_id,
        order_no=order_no,
        part_id=part_id,
        machine_model=machine_model,
        machine_serial=machine_serial,
        inspection_date=inspection_date,
        inspector_id=inspector_id,
        inspector_name=inspector_name,
        qa_manager_id=qa_manager_id,
        qa_manager_name=qa_manager_name,
        overall=overall,
        items=items,
        remarks=remarks,
        created_by=inspector_id,
    )
    return RedirectResponse(f"/qc/inspection-reports/{report_id}", 303)


@app.get("/qc/inspection-reports/{report_id}", response_class=HTMLResponse)
async def qc_report_detail(req: Request, report_id: int):
    """검사기 출하성적서 상세 (라인 + 발급/인쇄 액션)."""
    u = _qcr_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    report = get_qc_inspection_report(report_id)
    if not report:
        return RedirectResponse("/qc/inspection-reports", 303)
    company = _company_info_dict()
    with db_session() as c:
        customers = [dict(r) for r in c.execute(
            "SELECT id, name, COALESCE(tier,'') AS country FROM customers ORDER BY name LIMIT 200"
        ).fetchall()]
        orders = [dict(r) for r in c.execute(
            """SELECT o.id, o.order_no, COALESCE(cu.name,'-') AS cust_name
               FROM orders o
               LEFT JOIN customers cu ON cu.id = o.customer_id
               ORDER BY o.id DESC LIMIT 100"""
        ).fetchall()]
        parts_options = [dict(r) for r in c.execute(
            "SELECT id, part_name AS name, spec, unit FROM parts ORDER BY part_name LIMIT 300"
        ).fetchall()]
    return ctx(req, "qc_report_form.html", user=u, active="qc_reports",
               report=report, company=company,
               customers=customers, orders=orders, parts_options=parts_options,
               QC_OVERALL_OPTIONS=QC_OVERALL_OPTIONS,
               QC_STANDARD_ITEMS=QC_STANDARD_ITEMS)


@app.post("/qc/inspection-reports/{report_id}/issue")
async def qc_report_issue(req: Request, report_id: int):
    """검사기 출하성적서 발급 확정 (DRAFT → ISSUED).

    v5H116: 존재/상태 검증 + 발급자 audit log."""
    u = _qcr_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    rid = int(report_id)
    with db_session() as c:
        row = c.execute(
            "SELECT id, status FROM qc_inspection_reports WHERE id=?", (rid,)
        ).fetchone()
        if not row:
            return JSONResponse({"error": "검사 성적서 없음"}, 404)
        if row["status"] != "DRAFT":
            from urllib.parse import quote as _q
            return RedirectResponse(
                f"/qc/inspection-reports/{rid}?error=" + _q(f"이미 {row['status']} 상태입니다"), 303
            )
        try:
            c.execute(
                "UPDATE qc_inspection_reports SET status='ISSUED', "
                "issued_at=datetime('now','localtime'), issued_by=? "
                "WHERE id=? AND status='DRAFT'",
                (u["id"], rid),
            )
        except Exception:
            c.execute(
                "UPDATE qc_inspection_reports SET status='ISSUED', "
                "issued_at=datetime('now','localtime') WHERE id=? AND status='DRAFT'",
                (rid,),
            )
        _doc_audit_log(c, "qc_inspection_report", rid, "ISSUE", u["id"], "DRAFT → ISSUED")
    return RedirectResponse(f"/qc/inspection-reports/{rid}", 303)


@app.get("/qc/inspection-reports/{report_id}/print", response_class=HTMLResponse)
async def qc_report_print(req: Request, report_id: int):
    """검사기 출하성적서 인쇄 view (HTML + window.print(), 외부 PDF 0건)."""
    u = _qcr_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    report = get_qc_inspection_report(report_id)
    if not report:
        return RedirectResponse("/qc/inspection-reports", 303)
    company = _company_info_dict()
    overall_label_map = {code: label for code, label in QC_OVERALL_OPTIONS}
    return ctx(req, "qc_report_print.html", user=u,
               report=report, company=company,
               overall_label=overall_label_map.get(report.get("overall"), report.get("overall") or "PASS"))


# =====================================================
# WORK ORDERS — 가공팀 작업지시서 (2026-04-27 사이클77)
# 윤영조·이수빈 본업 · 04 시뮬 MISSING #3 보완 · 외부 자산 0건
# =====================================================
from .database import (create_work_order, get_work_orders, get_work_order)

WO_STATUS_OPTIONS = [
    ("DRAFT",       "작성 중 (DRAFT)"),
    ("RELEASED",    "발행 (RELEASED)"),
    ("IN_PROGRESS", "진행 중 (IN_PROGRESS)"),
    ("COMPLETED",   "완료 (COMPLETED)"),
    ("CANCELLED",   "취소 (CANCELLED)"),
]

WO_STD_STEPS = [
    ("절삭", 60),
    ("연마", 30),
    ("검수", 15),
]


def _wo_guard(req: Request):
    """가공팀 작업지시서 가드 — admin/ceo/executive OR 가공팀(team_id=9) OR production 권한 OR leader."""
    u = get_user(req)
    if not u:
        return None
    role = u.get("role") if isinstance(u, dict) else u["role"]
    if role in ("admin", "ceo", "executive", "production"):
        return u
    tid = u.get("team_id")
    if tid:
        with db_session() as c:
            row = c.execute(
                "SELECT name FROM teams WHERE id=?", (tid,)
            ).fetchone()
        if row:
            tname = row["name"] or ""
            if "가공" in tname or "machining" in tname.lower():
                return u
    # 팀장이면 자기 팀이 가공팀인 경우 통과
    if u.get("is_leader"):
        return u
    return None


@app.get("/production/work-orders", response_class=HTMLResponse)
async def wo_list(req: Request):
    """가공팀 작업지시서 목록 (윤영조·이수빈 본업 · MISSING #3)."""
    u = _wo_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    status = req.query_params.get("status") or None
    items = get_work_orders(status=status, limit=300)
    return ctx(req, "wo_list.html", user=u, active="work_orders",
               items=items, WO_STATUS_OPTIONS=WO_STATUS_OPTIONS,
               filter_status=status or "")


@app.get("/production/work-orders/new", response_class=HTMLResponse)
async def wo_new_form(req: Request):
    """가공팀 작업지시서 신규 폼."""
    u = _wo_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    with db_session() as c:
        orders = [dict(r) for r in c.execute(
            """SELECT o.id, o.order_no, COALESCE(cu.name,'-') AS cust_name
               FROM orders o
               LEFT JOIN customers cu ON cu.id = o.customer_id
               ORDER BY o.id DESC LIMIT 100"""
        ).fetchall()]
        projects = [dict(r) for r in c.execute(
            "SELECT id, name FROM projects ORDER BY id DESC LIMIT 100"
        ).fetchall()]
        parts_options = [dict(r) for r in c.execute(
            "SELECT id, part_name AS name, spec, unit FROM parts ORDER BY part_name LIMIT 300"
        ).fetchall()]
        users_options = [dict(r) for r in c.execute(
            "SELECT id, name FROM users WHERE is_active=1 ORDER BY name LIMIT 200"
        ).fetchall()]
    return ctx(req, "wo_form.html", user=u, active="work_orders",
               orders=orders, projects=projects,
               parts_options=parts_options, users_options=users_options,
               WO_STATUS_OPTIONS=WO_STATUS_OPTIONS,
               WO_STD_STEPS=WO_STD_STEPS,
               wo=None)


@app.post("/production/work-orders")
async def wo_create(req: Request):
    """가공팀 작업지시서 신규 등록. wo_no 자동 생성 (WO-YYYY-####)."""
    u = _wo_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    form = await req.form()
    order_id_raw = form.get("order_id")
    order_id = int(order_id_raw) if (order_id_raw and order_id_raw.isdigit()) else None
    project_id_raw = form.get("project_id")
    project_id = int(project_id_raw) if (project_id_raw and project_id_raw.isdigit()) else None
    part_id_raw = form.get("part_id")
    part_id = int(part_id_raw) if (part_id_raw and part_id_raw.isdigit()) else None
    try:
        qty = float(form.get("qty") or 0)
    except Exception:
        qty = 0
    # v5H115 #W1: 수량 0 이하 차단
    if qty <= 0:
        from urllib.parse import quote as _q
        return RedirectResponse(
            "/production/work-orders?error=" + _q("작업지시 수량은 0보다 커야 합니다"), 303
        )
    assigned_raw = form.get("assigned_to")
    assigned_to = int(assigned_raw) if (assigned_raw and assigned_raw.isdigit()) else None
    assigned_name = (form.get("assigned_name") or "").strip() or None
    planned_start = (form.get("planned_start") or "").strip() or None
    planned_end = (form.get("planned_end") or "").strip() or None
    specifications = (form.get("specifications") or "").strip() or None
    remarks = (form.get("remarks") or "").strip() or None
    # 라인 파싱
    step_names = form.getlist("line_step") if hasattr(form, "getlist") else []
    durations = form.getlist("line_duration") if hasattr(form, "getlist") else []
    progresses = form.getlist("line_progress") if hasattr(form, "getlist") else []
    workers = form.getlist("line_worker") if hasattr(form, "getlist") else []
    line_remarks_l = form.getlist("line_remarks") if hasattr(form, "getlist") else []
    items = []
    n = max(len(step_names), len(durations), len(progresses))
    for i in range(n):
        nm = (step_names[i] if i < len(step_names) else "").strip()
        if not nm:
            continue
        items.append({
            "step_name": nm,
            "duration_min": durations[i] if i < len(durations) else 0,
            "progress": progresses[i] if i < len(progresses) else 0,
            "worker_name": (workers[i] if i < len(workers) else "").strip() or None,
            "remarks": (line_remarks_l[i] if i < len(line_remarks_l) else "").strip() or None,
        })
    creator_id = u["id"] if isinstance(u, dict) else u["id"]
    creator_name = (u.get("name") if isinstance(u, dict) else u["name"]) or None
    wo_id, wo_no = create_work_order(
        order_id=order_id,
        project_id=project_id,
        part_id=part_id,
        qty=qty,
        assigned_to=assigned_to,
        assigned_name=assigned_name,
        created_by=creator_id,
        created_by_name=creator_name,
        planned_start=planned_start,
        planned_end=planned_end,
        specifications=specifications,
        items=items,
        remarks=remarks,
    )
    return RedirectResponse(f"/production/work-orders/{wo_id}", 303)


@app.get("/production/work-orders/{wo_id}", response_class=HTMLResponse)
async def wo_detail(req: Request, wo_id: int):
    """가공팀 작업지시서 상세 (라인 + 진행률 + 발행/인쇄 액션)."""
    u = _wo_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    wo = get_work_order(wo_id)
    if not wo:
        return RedirectResponse("/production/work-orders", 303)
    return ctx(req, "wo_form.html", user=u, active="work_orders",
               wo=wo, WO_STATUS_OPTIONS=WO_STATUS_OPTIONS,
               WO_STD_STEPS=WO_STD_STEPS)


@app.post("/production/work-orders/{wo_id}/release")
async def wo_release(req: Request, wo_id: int):
    """가공팀 작업지시서 발행 확정 (DRAFT → RELEASED).

    v5H116: 존재/상태 검증 + 발행자 audit log."""
    u = _wo_guard(req)
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    woid = int(wo_id)
    with db_session() as c:
        row = c.execute(
            "SELECT id, status FROM work_orders WHERE id=?", (woid,)
        ).fetchone()
        if not row:
            return JSONResponse({"error": "작업지시서 없음"}, 404)
        if row["status"] != "DRAFT":
            from urllib.parse import quote as _q
            return RedirectResponse(
                f"/production/work-orders/{woid}?error=" + _q(f"이미 {row['status']} 상태입니다"), 303
            )
        try:
            c.execute(
                "UPDATE work_orders SET status='RELEASED', "
                "released_at=datetime('now','localtime'), released_by=? "
                "WHERE id=? AND status='DRAFT'",
                (u["id"], woid),
            )
        except Exception:
            c.execute(
                "UPDATE work_orders SET status='RELEASED' WHERE id=? AND status='DRAFT'",
                (woid,),
            )
        _doc_audit_log(c, "work_order", woid, "RELEASE", u["id"], "DRAFT → RELEASED")
    return RedirectResponse(f"/production/work-orders/{woid}", 303)


@app.get("/production/work-orders/{wo_id}/print", response_class=HTMLResponse)
async def wo_print(req: Request, wo_id: int):
    """가공팀 작업지시서 인쇄 view (HTML + window.print, 외부 PDF 0건)."""
    u = _wo_guard(req)
    if not u:
        return RedirectResponse("/home", 303)
    wo = get_work_order(wo_id)
    if not wo:
        return RedirectResponse("/production/work-orders", 303)
    company = _company_info_dict()
    return ctx(req, "wo_print.html", user=u,
               wo=wo, company=company)


# =====================================================
# v5H136 (2026-05-05) — PO 라인 ↔ 프로젝트 다대다 연결 라우트
# 검사기/장비 소모품 발주를 관리번호에 귀속 → 수리 이력·운영비 추적
# =====================================================
@app.post("/po/{po_id}/items/{iid}/link-project")
async def po_item_link_project_route(request: Request, po_id: int, iid: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    form = await request.form()
    try:
        project_id = int(form.get("project_id") or 0)
    except ValueError:
        project_id = 0
    if not project_id:
        return RedirectResponse(f"/po/{po_id}?error=프로젝트를+선택하세요", 303)
    allocated_qty = form.get("allocated_qty") or None
    allocation_pct = form.get("allocation_pct") or None
    note = (form.get("note") or "").strip() or None
    try:
        _logi.po_item_link_project(
            po_item_id=iid, project_id=project_id,
            allocated_qty=allocated_qty, allocation_pct=allocation_pct,
            note=note,
            user_id=(u.get("id") if isinstance(u, dict) else u["id"]),
        )
    except Exception as e:
        return RedirectResponse(f"/po/{po_id}?error=연결+실패:{e}", 303)
    return RedirectResponse(f"/po/{po_id}#item-{iid}", 303)


@app.post("/po/links/{link_id}/delete")
async def po_item_unlink_project_route(request: Request, link_id: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    po_id = None
    try:
        with db_session() as c:
            row = c.execute(
                """SELECT pi.po_id FROM po_item_project_links l
                   JOIN po_items pi ON l.po_item_id=pi.id
                   WHERE l.id=?""", (link_id,)
            ).fetchone()
            if row:
                po_id = row[0]
    except Exception:
        pass
    try:
        _logi.po_item_unlink_project(link_id)
    except Exception:
        pass
    if po_id:
        return RedirectResponse(f"/po/{po_id}", 303)
    ref = request.headers.get("referer", "/po")
    return RedirectResponse(ref, 303)


@app.get("/api/projects/search")
async def api_projects_search(request: Request, q: str = ""):
    """프로젝트 자동완성 — 관리번호 또는 이름 부분일치.
    PO 라인 → 프로젝트 연결 폼에서 사용."""
    u = get_user(request)
    if not u:
        return JSONResponse({"ok": False, "error": "auth"}, 401)
    if not can_use_logistics(u):
        return JSONResponse({"ok": False, "error": "forbidden"}, 403)
    q = (q or "").strip()
    rows = []
    try:
        with db_session() as c:
            if q:
                like = f"%{q}%"
                rs = c.execute(
                    """SELECT p.id, p.mgmt_code, p.name, p.equip_type, p.status,
                              p.customer_id, p.biz_div, p.model_name, p.po_type, p.is_export,
                              cu.name AS customer_name
                       FROM projects p
                       LEFT JOIN customers cu ON p.customer_id=cu.id
                       WHERE (p.mgmt_code LIKE ? OR p.name LIKE ?)
                         AND p.mgmt_code IS NOT NULL AND p.mgmt_code != ''
                       ORDER BY p.id DESC LIMIT 30""",
                    (like, like)
                ).fetchall()
            else:
                rs = c.execute(
                    """SELECT p.id, p.mgmt_code, p.name, p.equip_type, p.status,
                              p.customer_id, p.biz_div, p.model_name, p.po_type, p.is_export,
                              cu.name AS customer_name
                       FROM projects p
                       LEFT JOIN customers cu ON p.customer_id=cu.id
                       WHERE p.mgmt_code IS NOT NULL AND p.mgmt_code != ''
                       ORDER BY p.id DESC LIMIT 30"""
                ).fetchall()
            rows = [dict(r) for r in rs]
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, 500)
    return JSONResponse({"ok": True, "rows": rows})



# ═══════════════════════════════════════════════════════════════════════════════
# v5H142 (2026-05-05) — 소모품 발주 전용 도메인 (대표 직접 요청)
# 신규 검사기와 분리. 관리번호 발급 X. 엑셀 일괄 import + 이미지 자동 압축
# ═══════════════════════════════════════════════════════════════════════════════
from . import consumables as _co
import shutil


@app.get("/consumables", response_class=HTMLResponse)
async def consumables_list(request: Request, status: str = "", q: str = ""):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not (can_use_logistics(u) or can_use_sales(u)):
        return RedirectResponse("/home", 303)
    rows = _co.co_list(status=status, q=q, limit=300)
    groups: dict = {}
    for r in rows:
        ym = (r.get("order_date") or "")[:7] or "기타"
        groups.setdefault(ym, []).append(r)
    months = sorted(groups.keys(), reverse=True)
    return ctx(request, "consumables.html",
               user=u, active="consumables",
               rows=rows, months=months, groups=groups,
               status=status, q=q,
               STATUS_LABELS=_co.CO_STATUS_LABELS,
               STATUSES=_co.CO_STATUSES)


@app.get("/consumables/new", response_class=HTMLResponse)
async def consumables_new_form(request: Request):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not (can_use_logistics(u) or can_use_sales(u)):
        return RedirectResponse("/home", 303)
    return ctx(request, "consumable_form_upload.html",
               user=u, active="consumables",
               customers=_logi.customers_for_picker())


@app.post("/consumables/upload-xlsx")
async def consumables_upload_xlsx(request: Request,
                                   file: UploadFile = File(...),
                                   customer_name: str = Form(""),
                                   biz_div: str = Form(""),
                                   order_date: str = Form(""),
                                   due_date: str = Form(""),
                                   currency: str = Form("KRW"),
                                   note: str = Form("")):
    u = get_user(request)
    if not u:
        return JSONResponse({"ok": False, "error": "auth"}, 401)
    if not (can_use_logistics(u) or can_use_sales(u)):
        return JSONResponse({"ok": False, "error": "forbidden"}, 403)
    # v5H218: 진행 사업부 필수 (T/M)
    if biz_div not in ("T", "M"):
        return JSONResponse({"ok": False, "error": "biz_div_required",
                              "message": "진행 사업부(검사기/자동화)를 선택해주세요"}, 400)
    co_id, co_no = _co.co_create(
        customer_name=customer_name, biz_div=biz_div, order_date=order_date, due_date=due_date,
        currency=currency, note=note, source_file=file.filename or "",
        created_by=u.get("id"),
    )
    raw = await file.read()
    tmp_dir = tempfile.mkdtemp(prefix=f"co_{co_id}_")
    tmp_path = os.path.join(tmp_dir, file.filename or "upload.xlsx")
    with open(tmp_path, "wb") as f:
        f.write(raw)
    img_dir = _co.co_image_dir(co_id)
    parsed = _co.parse_consumable_xlsx(tmp_path, image_out_dir=img_dir)
    enriched = []
    for ln in parsed["lines"]:
        part_match = _co.match_part_by_name(ln.get("part_name", ""))
        proj_match = _co.match_project_by_model(ln.get("model_use", ""))
        imgs_v = parsed["images"].get(ln["line_no"], [])
        if not isinstance(imgs_v, list):
            imgs_v = []
        ln_out = {
            **ln,
            "part_match": part_match,
            "project_match": proj_match,
            "images": [{
                "thumb_url": _co.co_image_url(co_id, im["thumb"]),
                "full_url": _co.co_image_url(co_id, im["full"]),
                "thumb_file": im["thumb"], "full_file": im["full"],
                "orig_size": im.get("orig_size", 0),
                "compressed_size": im.get("compressed", 0),
            } for im in imgs_v],
        }
        enriched.append(ln_out)
    try:
        shutil.rmtree(tmp_dir, ignore_errors=True)
    except Exception:
        pass
    return JSONResponse({
        "ok": True,
        "co_id": co_id, "co_no": co_no,
        "header_row": parsed.get("header_row"),
        "col_map": parsed.get("col_map"),
        "lines": enriched,
        "image_count": parsed.get("image_count", 0),
        "error": parsed.get("error"),
    })


@app.post("/consumables/{co_id:int}/import-confirmed")
async def consumables_import_confirmed(request: Request, co_id: int):
    u = get_user(request)
    if not u:
        return JSONResponse({"ok": False, "error": "auth"}, 401)
    if not (can_use_logistics(u) or can_use_sales(u)):
        return JSONResponse({"ok": False, "error": "forbidden"}, 403)
    try:
        body = await request.json()
    except Exception:
        body = {}
    items_in = body.get("items") or []
    items_out = []
    for it in items_in:
        items_out.append({
            "line_no": it.get("line_no") or 0,
            "model_use": it.get("model_use") or "",
            "part_name": it.get("part_name") or "",
            "spec": it.get("spec") or "",
            "qty": it.get("qty") or 0,
            "unit": it.get("unit") or "EA",
            "unit_price": it.get("unit_price") or 0,
            "part_id": it.get("part_id"),
            "linked_project_id": it.get("linked_project_id"),
            "image_path": (_co.co_image_url(co_id, it["image_full_file"])
                            if it.get("image_full_file") else None),
            "image_thumb_path": (_co.co_image_url(co_id, it["image_thumb_file"])
                                  if it.get("image_thumb_file") else None),
            "note": it.get("note") or "",
        })
    n = _co.coi_bulk_insert(co_id, items_out)
    return JSONResponse({"ok": True, "co_id": co_id, "inserted": n})


@app.get("/consumables/{co_id:int}", response_class=HTMLResponse)
async def consumables_detail(request: Request, co_id: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not (can_use_logistics(u) or can_use_sales(u)):
        return RedirectResponse("/home", 303)
    co = _co.co_get(co_id)
    if not co:
        return RedirectResponse("/consumables", 303)
    items = _co.coi_list(co_id)
    return ctx(request, "consumable_detail.html",
               user=u, active="consumables",
               co=co, items=items,
               STATUS_LABELS=_co.CO_STATUS_LABELS,
               STATUSES=_co.CO_STATUSES)


@app.post("/consumables/{co_id:int}/items/{iid:int}/edit")
async def consumables_item_edit(request: Request, co_id: int, iid: int):
    u = get_user(request)
    if not u:
        return JSONResponse({"ok": False, "error": "auth"}, 401)
    if not (can_use_logistics(u) or can_use_sales(u)):
        return JSONResponse({"ok": False, "error": "forbidden"}, 403)
    form = await request.form()
    fields = {}
    for k in ("model_use", "part_name", "spec", "unit", "note"):
        if k in form:
            fields[k] = (form.get(k) or "").strip()
    for k in ("qty", "unit_price"):
        if k in form:
            try:
                fields[k] = float((form.get(k) or "0").replace(",", ""))
            except ValueError:
                fields[k] = 0
    for k in ("part_id", "linked_project_id"):
        if k in form:
            v = (form.get(k) or "").strip()
            fields[k] = int(v) if v.isdigit() else None
    _co.coi_update(iid, fields)
    co = _co.co_get(co_id)
    return JSONResponse({"ok": True, "total_amount": (co or {}).get("total_amount", 0)})


@app.post("/consumables/{co_id:int}/items/{iid:int}/link-project")
async def consumables_item_link_project(request: Request, co_id: int, iid: int):
    u = get_user(request)
    if not u:
        return JSONResponse({"ok": False, "error": "auth"}, 401)
    if not (can_use_logistics(u) or can_use_sales(u)):
        return JSONResponse({"ok": False, "error": "forbidden"}, 403)
    form = await request.form()
    pid_raw = (form.get("project_id") or "").strip()
    pid = int(pid_raw) if pid_raw.isdigit() else None
    _co.coi_update(iid, {"linked_project_id": pid})
    return JSONResponse({"ok": True, "linked_project_id": pid})


@app.post("/consumables/{co_id:int}/items/{iid:int}/match-part")
async def consumables_item_match_part(request: Request, co_id: int, iid: int):
    u = get_user(request)
    if not u:
        return JSONResponse({"ok": False, "error": "auth"}, 401)
    if not (can_use_logistics(u) or can_use_sales(u)):
        return JSONResponse({"ok": False, "error": "forbidden"}, 403)
    form = await request.form()
    pid_raw = (form.get("part_id") or "").strip()
    pid = int(pid_raw) if pid_raw.isdigit() else None
    _co.coi_update(iid, {"part_id": pid})
    return JSONResponse({"ok": True, "part_id": pid})


@app.post("/consumables/{co_id:int}/items/{iid:int}/delete")
async def consumables_item_delete(request: Request, co_id: int, iid: int):
    u = get_user(request)
    if not u:
        return JSONResponse({"ok": False, "error": "auth"}, 401)
    if not (can_use_logistics(u) or can_use_sales(u)):
        return JSONResponse({"ok": False, "error": "forbidden"}, 403)
    _co.coi_delete(iid)
    return JSONResponse({"ok": True})


@app.post("/consumables/{co_id:int}/status")
async def consumables_set_status(request: Request, co_id: int):
    u = get_user(request)
    if not u:
        return JSONResponse({"ok": False, "error": "auth"}, 401)
    if not (can_use_logistics(u) or can_use_sales(u)):
        return JSONResponse({"ok": False, "error": "forbidden"}, 403)
    form = await request.form()
    st = (form.get("status") or "").strip().upper()
    if st not in _co.CO_STATUSES:
        return JSONResponse({"ok": False, "error": "invalid status"}, 400)
    with db_session() as c:
        c.execute("UPDATE consumable_orders SET status=? WHERE id=?", (st, int(co_id)))
    return JSONResponse({"ok": True, "status": st})


# v5H145 (2026-05-05) — 관련부서 통보 발송
# 대표 의도: 영업이 소모품 발주 등록 → 자재구매팀·생산팀이 시스템 알림으로 즉시 인지
# 대상: can_use_logistics=1 (자재팀) + can_use_production=1 (생산팀이 있다면) + role IN admin/ceo
@app.post("/consumables/{co_id:int}/notify")
async def consumables_notify(request: Request, co_id: int):
    u = get_user(request)
    if not u:
        return JSONResponse({"ok": False, "error": "auth"}, 401)
    if not (can_use_logistics(u) or can_use_sales(u)):
        return JSONResponse({"ok": False, "error": "forbidden"}, 403)
    co = _co.co_get(co_id)
    if not co:
        return JSONResponse({"ok": False, "error": "not_found"}, 404)
    items = _co.coi_list(co_id) or []
    line_count = len(items)
    total = co.get("total_amount") or 0
    curr = co.get("currency") or "KRW"
    co_no = co.get("co_no") or f"CO-{co_id}"
    cust = co.get("customer_name") or "—"
    if curr == "KRW":
        amt_txt = f"{total:,.0f}원"
    else:
        amt_txt = f"{curr} {total:,.2f}"
    title = f"📦 소모품 발주 [{co_no}] 등록됨"
    body = (f"고객사: {cust} · 라인 {line_count}건 · 합계 {amt_txt}\n"
            f"등록자: {u.get('name') or u.get('username') or '—'}\n"
            f"검토 부탁드립니다.")
    link = f"/consumables/{co_id}"
    sender_id = u.get("id")
    sent_to = 0
    try:
        with db_session() as c:
            # 통보 대상: 자재구매(logistics) + admin/ceo. 발신자 본인은 제외.
            rows = c.execute(
                "SELECT id FROM users "
                "WHERE (COALESCE(can_use_logistics,0)=1 OR role IN ('admin','ceo')) "
                "  AND COALESCE(is_active,1)=1 "
                "  AND id != ?",
                (sender_id,)
            ).fetchall()
            for r in rows:
                uid = r[0] if not isinstance(r, dict) else r["id"]
                c.execute(
                    "INSERT INTO notifications(user_id, kind, title, body, link) "
                    "VALUES(?,?,?,?,?)",
                    (uid, "consumable_order", title, body, link)
                )
                sent_to += 1
    except Exception as e:
        return JSONResponse({"ok": False, "error": f"db_error: {e}"}, 500)
    return JSONResponse({"ok": True, "sent_to": sent_to, "co_no": co_no})


@app.post("/consumables/{co_id:int}/delete")
async def consumables_delete(request: Request, co_id: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not (can_use_logistics(u) or can_use_sales(u)):
        return RedirectResponse("/home", 303)
    _co.co_delete(co_id)
    return RedirectResponse("/consumables", 303)


@app.get("/api/parts/search")
async def api_parts_search(request: Request, q: str = ""):
    u = get_user(request)
    if not u:
        return JSONResponse({"ok": False, "error": "auth"}, 401)
    q = (q or "").strip()
    rows = []
    try:
        with db_session() as c:
            if q:
                like = f"%{q}%"
                rs = c.execute(
                    "SELECT id, part_no, part_name, std_price, unit FROM parts "
                    "WHERE (part_no LIKE ? OR part_name LIKE ?) "
                    "  AND COALESCE(is_active,1)=1 LIMIT 30",
                    (like, like)
                ).fetchall()
            else:
                rs = c.execute(
                    "SELECT id, part_no, part_name, std_price, unit FROM parts "
                    "WHERE COALESCE(is_active,1)=1 ORDER BY id DESC LIMIT 30"
                ).fetchall()
            rows = [dict(r) for r in rs]
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, 500)
    return JSONResponse({"ok": True, "rows": rows})