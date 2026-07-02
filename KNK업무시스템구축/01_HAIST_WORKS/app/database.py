"""
KNK 일일업무일지 v2 - Database Layer
Phase 1 MVP - Task Card 기반 구조

═══════════════════════════════════════════════════════════════════════════════
 database.py 목차 (TABLE OF CONTENTS) — v1.0 (2026-04-29 옵션 A 정리)
═══════════════════════════════════════════════════════════════════════════════

 PART A. 기반
   §A1.  DB 연결 / 헬퍼 (db_session, hash_pw)         L 1~30
   §A2.  SCHEMA — 모든 CREATE TABLE 정의              L 32

 PART B. 시드 데이터 / 초기화
   §B1.  ORG CHART 시드 (TEAMS, USERS)                L 1379
   §B2.  로그인 ID 생성 규칙                           L 1528
   §B3.  init_db / seed_all                           L 1541
   §B4.  샘플 데이터 (지난 14일 현실적 데이터)         L 2026
   §B5.  관리코드 발행목록 import                      L 2272
   §B6.  초기 비번 일괄 발급                           L 2442

 PART C. 댓글·알림·활동
   §C1.  COMMENTS & 멘션 파싱                         L 2500
   §C2.  활동 로그 / 반응 / 회고                       L 2580
   §C3.  알림시스템 통합 헬퍼 (notify_user 등)         L 3049
   §C4.  사내 메신저 push                             L 5621

 PART D. 자재구매 (Logistics)
   §D1.  parts / 관리코드 발행대장                      L 3095
   §D2.  발주 (suppliers + PO 헤더/라인)                L 3378
   §D3.  STOCK MOVEMENTS (수불부)                      L 3688
   §D4.  안전재고 알림 + 발주 추천                      L 3802
   §D5.  재고 실사·조정                                 L 4228
   §D6.  FIFO + ABC + 회전율                           L 4639
   §D7.  환율 관리 (FXLoader 패턴)                     L 4744
   §D8.  적용일자 단가 (price_history)                  L 4821
   §D9.  환율·단가 헬퍼                                L 4916
   §D10. 공급사 리드타임 health_check                  L 4997

 PART E. 변경·이슈·티켓·게시판
   §E1.  게시판 (boards / board_posts / board_comments) L 5261
   §E2.  변경 Inform 시스템                            L 5416
   §E3.  ISSUES · AS DB                                L 5426
   §E4.  요청 티켓 시스템                              L 5943
   §E5.  ISSUES 헬퍼                                   L 6247

 PART F. 분석·검색
   §F1.  진행률 대시보드 (PHASE_DEFS 등)                L 6497
   §F2.  글로벌 통합 검색                              L 6812
   §F3.  CEO 통합 대시보드 KPI                          L 7033

 PART G. 매출영업
   §G1.  Sales 견적 라인 헬퍼                          L 7224
   §G2.  단가 시뮬레이션 (cost_simulation)              L 6668

 PART H. 수출입
   §H1.  FTA 원산지증명서                              L 7298

 PART I. 품질관리
   §I1.  QC INSPECTION REPORT                          L 7432

 PART J. 생산
   §J1.  WORK ORDERS (가공작업지시서)                  L 7584

 PART K. 시스템 설정
   §K1.  app_settings (key-value 저장)                 L 5572

───────────────────────────────────────────────────────────────────────────────
 사용법:
  - 특정 영역 수정: §AN. 검색 (예: §D4. → 안전재고 알림)
  - 함수 검색: ^def 함수명
  - 스키마 검색: "CREATE TABLE 테이블명"
  - 행 수가 정확하지 않을 수 있음 (수정 시 변동) — _INDEX_코드구조.md 참조
═══════════════════════════════════════════════════════════════════════════════
"""
import sqlite3, os, hashlib, secrets
from contextlib import contextmanager

DB_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "knk.db")

# v5H226f-4: 비밀번호 해시 강화 (stdlib pbkdf2, 사용자별 솔트, 200k 반복)
# 기존 sha256 정적 솔트 → 레인보우 테이블 취약. 점진 마이그레이션 전략:
#   - 신규 생성/변경: pbkdf2 형식 ("pbkdf2$<iter>$<salt>$<hash>")
#   - 검증: 새 형식 우선, 미일치 시 legacy sha256 fallback (verify_pw 헬퍼)
#   - 기존 사용자는 다음 로그인/비번 변경 시 자동 업그레이드 (main 의 login 흐름 참조)
_PBKDF2_ITER = 200_000

def _legacy_hash(pw: str) -> str:
    """v5H226f 이전의 정적 솔트 sha256 — verify_pw 의 fallback 용도."""
    return hashlib.sha256(("knk-haist-" + pw).encode()).hexdigest()

def hash_pw(pw: str) -> str:
    """신규 비밀번호 해시 (pbkdf2-sha256, 사용자별 16바이트 솔트, 200k 반복).
    형식: pbkdf2$200000$<salt_hex>$<hash_hex>"""
    salt = secrets.token_bytes(16)
    h = hashlib.pbkdf2_hmac("sha256", pw.encode(), salt, _PBKDF2_ITER)
    return f"pbkdf2${_PBKDF2_ITER}${salt.hex()}${h.hex()}"

def verify_pw(pw: str, stored: str) -> bool:
    """비밀번호 검증 — pbkdf2 신규 형식 + legacy sha256 모두 지원.
    True 반환 후 호출자가 legacy 면 hash_pw() 결과로 UPDATE 권장(점진 업그레이드)."""
    if not stored:
        return False
    if stored.startswith("pbkdf2$"):
        try:
            _, iter_s, salt_hex, hash_hex = stored.split("$", 3)
            it = int(iter_s)
            salt = bytes.fromhex(salt_hex)
            h = hashlib.pbkdf2_hmac("sha256", pw.encode(), salt, it)
            return secrets.compare_digest(h.hex(), hash_hex)
        except Exception:
            return False
    # legacy sha256 fallback
    return secrets.compare_digest(stored, _legacy_hash(pw))

def is_legacy_hash(stored: str) -> bool:
    """저장된 해시가 legacy(sha256) 형식이면 True — 점진 업그레이드 판단용."""
    return bool(stored) and not stored.startswith("pbkdf2$")

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

@contextmanager
def db_session():
    conn = get_db()
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# =====================================================
# v5H119: 공통 cascade 안전망 헬퍼
# =====================================================
def _safe_delete_with_cascade(conn, table_name: str, row_id,
                              fk_column: str = "project_id",
                              explicit_children: list = None,
                              keep_tables: tuple = ()) -> dict:
    """sqlite_master 동적 스캔 → fk_column 보유 자식 테이블 SET NULL 시도 →
    실패 시 DELETE → 본행 DELETE.

    Args:
        conn: 활성 sqlite3 connection (db_session() 컨텍스트 내).
        table_name: 본행 테이블명 (예: 'parts', 'purchase_orders').
        row_id: 삭제 대상 PK (id 컬럼 기준).
        fk_column: 자식 테이블의 FK 컬럼명 (예: 'part_id', 'po_id', 'change_id').
        explicit_children: [(sql, params_tuple), ...] — 동적 스캔 전 명시 실행할 SQL.
        keep_tables: 동적 스캔에서 제외할 테이블명 튜플 (본인 + 자식 명시 등).

    Returns:
        {"ok": bool, "table_results": [{"table": str, "action": str}], "error": str|None}

    백워드 호환: 본 함수는 raise 하지 않으며 결과 dict 만 반환. 호출자는 dict["ok"]
    검사 후 폴백 가능. 본행 DELETE 실패 시에만 ok=False.
    """
    results = []
    err = None
    try:
        # 1. 명시적 자식 SQL
        for item in (explicit_children or []):
            try:
                sql, params = item
                conn.execute(sql, params)
                results.append({"table": "(explicit)", "action": sql[:60]})
            except Exception as e:
                results.append({"table": "(explicit)", "action": f"FAIL: {e}"})

        # 2. 동적 스캔
        skip = set(keep_tables) | {table_name, "sqlite_sequence"}
        try:
            tables = [r[0] for r in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' "
                "AND name NOT LIKE 'sqlite_%'"
            ).fetchall()]
        except Exception:
            tables = []
        for tbl in tables:
            if tbl in skip:
                continue
            try:
                cols = [r[1] for r in conn.execute(
                    f"PRAGMA table_info({tbl})").fetchall()]
            except Exception:
                continue
            if fk_column not in cols:
                continue
            try:
                conn.execute(
                    f"UPDATE {tbl} SET {fk_column}=NULL WHERE {fk_column}=?",
                    (row_id,))
                results.append({"table": tbl, "action": "SET NULL"})
            except Exception:
                try:
                    conn.execute(
                        f"DELETE FROM {tbl} WHERE {fk_column}=?", (row_id,))
                    results.append({"table": tbl, "action": "DELETE"})
                except Exception as e2:
                    results.append({"table": tbl, "action": f"FAIL: {e2}"})

        # 3. 본행 삭제
        conn.execute(f"DELETE FROM {table_name} WHERE id=?", (row_id,))
        return {"ok": True, "table_results": results, "error": None}
    except Exception as e:
        err = str(e)
        return {"ok": False, "table_results": results, "error": err}


# =====================================================
# SCHEMA
# =====================================================
SCHEMA = """
CREATE TABLE IF NOT EXISTS teams (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    code TEXT UNIQUE NOT NULL,
    name TEXT NOT NULL,
    leader_id INTEGER,
    is_lab INTEGER DEFAULT 0,
    sector TEXT,
    parent_team_id INTEGER REFERENCES teams(id),
    display_order INTEGER DEFAULT 0
);

CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    login_id TEXT UNIQUE NOT NULL,
    password TEXT NOT NULL,
    email TEXT,
    team_id INTEGER REFERENCES teams(id),
    rank TEXT,
    role TEXT DEFAULT 'member',
    is_active INTEGER DEFAULT 1,
    lang TEXT DEFAULT 'ko',
    created_at TEXT DEFAULT (datetime('now','localtime'))
);

CREATE TABLE IF NOT EXISTS customers (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT UNIQUE NOT NULL,
    tier TEXT DEFAULT '일반',
    note TEXT
);

-- v5H56 (2026-05-03 대표 지적) — 거래처 담당자 다대다
-- 한 거래처에 영업·구매·세금계산서·기술·품질·결재 등 여러 담당자 가능
CREATE TABLE IF NOT EXISTS customer_contacts (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    customer_id     INTEGER NOT NULL REFERENCES customers(id) ON DELETE CASCADE,
    role            TEXT NOT NULL,           -- 영업/구매/세금계산서/품질/기술/결재/기타
    department      TEXT,                    -- 부서명 (예: 구매2팀)
    name            TEXT NOT NULL,           -- 담당자 이름
    position        TEXT,                    -- 직위 (과장/팀장 등)
    phone           TEXT,
    mobile          TEXT,
    email           TEXT,
    is_primary      INTEGER DEFAULT 0,       -- 1 = 해당 역할의 주담당
    note            TEXT,
    created_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_ccontact_cust ON customer_contacts(customer_id);
CREATE INDEX IF NOT EXISTS idx_ccontact_role ON customer_contacts(role);

-- v5H227 (2026-05-31 대표 지시) — 전사 공유 연락처(주소록)
-- customer_contacts(거래처 종속)와 별개. 전 직원이 등록·검색·열람하는 공용 주소록.
-- 자동등록 3경로: 명함 사진(OCR=source 'card') / 메일 서명(source 'signature') / 직접입력('manual')
-- 수정·삭제 권한: 등록 본인 + 관리자. 열람·검색: 전 직원.
CREATE TABLE IF NOT EXISTS shared_contacts (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    name            TEXT NOT NULL,           -- 이름
    position        TEXT,                    -- 직책/직위 (예: 부장, 대표)
    company         TEXT,                    -- 회사명
    department      TEXT,                    -- 부서
    mobile          TEXT,                    -- 휴대폰
    phone           TEXT,                    -- 회사 전화
    fax             TEXT,                    -- 팩스
    email           TEXT,                    -- 이메일
    address         TEXT,                    -- 주소
    note            TEXT,                    -- 비고
    card_image      TEXT,                    -- 명함 사진 파일 경로 (/uploads/contacts/..)
    source          TEXT DEFAULT 'manual',   -- manual / card / signature
    customer_id     INTEGER REFERENCES customers(id) ON DELETE SET NULL,  -- 거래처 연결 (선택)
    visibility      TEXT DEFAULT 'private',   -- v5H226z580 공유 범위: private(개인)/dept(부서)/company(전사)
    share_team_id   INTEGER REFERENCES teams(id) ON DELETE SET NULL,      -- dept 공유 시 대상 팀
    created_by      INTEGER REFERENCES users(id),
    created_at      TEXT DEFAULT (datetime('now','localtime')),
    updated_by      INTEGER REFERENCES users(id),
    updated_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_scontact_name    ON shared_contacts(name);
CREATE INDEX IF NOT EXISTS idx_scontact_company ON shared_contacts(company);
CREATE INDEX IF NOT EXISTS idx_scontact_email   ON shared_contacts(email);
CREATE INDEX IF NOT EXISTS idx_scontact_mobile  ON shared_contacts(mobile);

-- v5H228 (2026-05-31 대표 지시) — 직원 본인 하이웍스 계정 SMTP 발송용 자격증명
-- 비밀번호는 Fernet 으로 암호화 보관(KNK_MAIL_KEY). 하이웍스 "앱 비밀번호" 사용 권장.
CREATE TABLE IF NOT EXISTS user_mail_creds (
    user_id           INTEGER PRIMARY KEY REFERENCES users(id) ON DELETE CASCADE,
    smtp_email        TEXT,
    smtp_password_enc TEXT,
    updated_at        TEXT DEFAULT (datetime('now','localtime'))
);

-- ─── KNK Mail (자체 메일 시스템 Phase 0 POC, 2026-06-13 대표 지시) ───
-- 받은 메일 보관함. Cloudflare Email Routing → /api/mail/inbound → 적재.
-- AI(ai_client)가 category(견적/발주/세금계산서/일반)·lang·summary 자동 채움.
CREATE TABLE IF NOT EXISTS mail_messages (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id     INTEGER REFERENCES users(id) ON DELETE CASCADE,  -- 받는 사람(소유자)
    direction   TEXT DEFAULT 'in',          -- in(받음) / out(보냄) / draft(임시저장)
    from_email  TEXT,
    from_name   TEXT,
    to_email    TEXT,                        -- 받은 주소(@knk-mailtest.com 등)
    cc          TEXT,
    bcc         TEXT,                        -- 숨은참조(보낸/임시저장 보관용)
    subject     TEXT,
    body_text   TEXT,                        -- 화면 표시용 본문(텍스트)
    body_html   TEXT,                        -- 원문 HTML 보관(표시는 text)
    category    TEXT DEFAULT '일반',         -- AI 분류: 견적/발주/세금계산서/일반
    lang        TEXT,                        -- AI 언어감지: ko/vi/en
    summary     TEXT,                        -- AI 5줄 요약
    is_read     INTEGER DEFAULT 0,
    is_starred  INTEGER DEFAULT 0,
    is_deleted  INTEGER DEFAULT 0,
    raw_size    INTEGER DEFAULT 0,
    received_at TEXT DEFAULT (datetime('now','localtime')),
    read_at     TEXT,
    deleted_at  TEXT,
    created_at  TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_mailmsg_user ON mail_messages(user_id, is_deleted, received_at DESC);
CREATE INDEX IF NOT EXISTS idx_mailmsg_cat  ON mail_messages(user_id, category);

CREATE TABLE IF NOT EXISTS mail_attachments (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    mail_id     INTEGER REFERENCES mail_messages(id) ON DELETE CASCADE,
    filename    TEXT,
    mime        TEXT,
    size        INTEGER DEFAULT 0,
    path        TEXT,                        -- NAS/uploads 저장 경로(미사용 시 NULL)
    content_id  TEXT,                        -- 인라인 이미지 cid 매핑(<...> 제거 보관)
    is_inline   INTEGER DEFAULT 0,           -- 1=본문 인라인 이미지(서명 로고 등)
    data        BLOB,                        -- 첨부 바이트(POC: DB 보관, 상한 초과 시 메타만)
    created_at  TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_mailatt_mail ON mail_attachments(mail_id);

-- 수신 주소 → WORKS 사용자 매핑(없으면 대표 기본). POC: test@ → 대표.
CREATE TABLE IF NOT EXISTS mail_aliases (
    address     TEXT PRIMARY KEY,            -- 소문자 정규화된 수신 주소
    user_id     INTEGER REFERENCES users(id) ON DELETE CASCADE,
    created_at  TEXT DEFAULT (datetime('now','localtime'))
);

-- KNK 대용량 첨부(2026-06-13 대표 지시): 큰 파일을 NAS 보관 + 다운로드 링크로 발송.
-- 파일은 data/mail_large/{token}/ (정적 비공개) — /dl/{token} 라우트만 서빙(만료·횟수 검사).
CREATE TABLE IF NOT EXISTS mail_large_files (
    id             INTEGER PRIMARY KEY AUTOINCREMENT,
    token          TEXT UNIQUE NOT NULL,       -- 추측 불가 다운로드 토큰
    filename       TEXT,
    mime           TEXT,
    size           INTEGER DEFAULT 0,
    path           TEXT,                        -- NAS 저장 경로
    uploaded_by    INTEGER REFERENCES users(id) ON DELETE SET NULL,
    download_count INTEGER DEFAULT 0,
    expires_at     TEXT,                        -- 만료 시각(localtime). NULL=무기한
    created_at     TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_mlf_token ON mail_large_files(token);

-- 메일 서명(2026-06-13 대표 지시·편의기능): 직원별 1회 저장 → 메일 쓸 때 자동 삽입.
CREATE TABLE IF NOT EXISTS mail_signatures (
    user_id    INTEGER PRIMARY KEY REFERENCES users(id) ON DELETE CASCADE,
    body       TEXT,
    updated_at TEXT DEFAULT (datetime('now','localtime'))
);

CREATE TABLE IF NOT EXISTS projects (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    code TEXT,
    name TEXT NOT NULL,
    customer_id INTEGER REFERENCES customers(id),
    type TEXT,
    status TEXT DEFAULT '진행중',
    pm_id INTEGER REFERENCES users(id),
    start_date TEXT,
    end_date TEXT,
    mgmt_code TEXT UNIQUE,
    equip_type TEXT,
    year_month TEXT,
    model_name TEXT,
    server_path TEXT,
    lead_user_id INTEGER REFERENCES users(id),
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_projects_mgmt ON projects(mgmt_code);
CREATE INDEX IF NOT EXISTS idx_projects_equip ON projects(equip_type);

CREATE TABLE IF NOT EXISTS tasks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER REFERENCES users(id),
    work_date TEXT NOT NULL,
    title TEXT NOT NULL,
    category TEXT,
    project_id INTEGER REFERENCES projects(id),
    customer_id INTEGER REFERENCES customers(id),
    status TEXT DEFAULT '진행중',
    hours REAL DEFAULT 0,
    notes TEXT,
    next_plan TEXT,
    due_date TEXT,
    carry_from_id INTEGER REFERENCES tasks(id),
    created_at TEXT DEFAULT (datetime('now','localtime')),
    updated_at TEXT DEFAULT (datetime('now','localtime'))
);

CREATE TABLE IF NOT EXISTS task_comments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    task_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
    author_id INTEGER NOT NULL REFERENCES users(id),
    body TEXT NOT NULL,
    is_ceo_request INTEGER DEFAULT 0,
    parent_id INTEGER REFERENCES task_comments(id),
    created_at TEXT DEFAULT (datetime('now','localtime'))
);

CREATE TABLE IF NOT EXISTS notifications (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL REFERENCES users(id),
    kind TEXT NOT NULL,
    title TEXT NOT NULL,
    body TEXT,
    link TEXT,
    task_id INTEGER REFERENCES tasks(id) ON DELETE CASCADE,
    comment_id INTEGER REFERENCES task_comments(id) ON DELETE CASCADE,
    is_read INTEGER DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);

CREATE TABLE IF NOT EXISTS team_summaries (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    team_id INTEGER REFERENCES teams(id),
    work_date TEXT NOT NULL,
    author_id INTEGER REFERENCES users(id),
    headline TEXT NOT NULL,
    notes TEXT,
    created_at TEXT DEFAULT (datetime('now','localtime')),
    updated_at TEXT DEFAULT (datetime('now','localtime')),
    UNIQUE(team_id, work_date)
);

CREATE INDEX IF NOT EXISTS idx_tasks_user_date ON tasks(user_id, work_date);
CREATE INDEX IF NOT EXISTS idx_tasks_date ON tasks(work_date);
CREATE INDEX IF NOT EXISTS idx_tasks_status ON tasks(status);
CREATE INDEX IF NOT EXISTS idx_tasks_project ON tasks(project_id);
CREATE INDEX IF NOT EXISTS idx_tasks_customer ON tasks(customer_id);
CREATE INDEX IF NOT EXISTS idx_users_team ON users(team_id);
CREATE INDEX IF NOT EXISTS idx_team_summaries_team_date ON team_summaries(team_id, work_date);
CREATE INDEX IF NOT EXISTS idx_comments_task ON task_comments(task_id);
CREATE INDEX IF NOT EXISTS idx_comments_author ON task_comments(author_id);
CREATE INDEX IF NOT EXISTS idx_notif_user_read ON notifications(user_id, is_read);
CREATE INDEX IF NOT EXISTS idx_notif_created ON notifications(created_at);

-- ============ Phase 1+ : 협업 신경망 ============
CREATE TABLE IF NOT EXISTS activities (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    actor_id INTEGER NOT NULL REFERENCES users(id),
    kind TEXT NOT NULL,            -- task_create / task_update / task_status / comment / reaction / retro
    task_id INTEGER REFERENCES tasks(id) ON DELETE CASCADE,
    project_id INTEGER REFERENCES projects(id) ON DELETE SET NULL,
    team_id INTEGER REFERENCES teams(id) ON DELETE SET NULL,
    title TEXT NOT NULL,
    body TEXT,
    meta TEXT,                     -- JSON
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_act_created ON activities(created_at DESC);
CREATE INDEX IF NOT EXISTS idx_act_team ON activities(team_id);
CREATE INDEX IF NOT EXISTS idx_act_actor ON activities(actor_id);
CREATE INDEX IF NOT EXISTS idx_act_task ON activities(task_id);

CREATE TABLE IF NOT EXISTS task_reactions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    task_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
    user_id INTEGER NOT NULL REFERENCES users(id),
    kind TEXT NOT NULL,            -- ack / question / risk / ok
    created_at TEXT DEFAULT (datetime('now','localtime')),
    UNIQUE(task_id, user_id, kind)
);
CREATE INDEX IF NOT EXISTS idx_react_task ON task_reactions(task_id);

CREATE TABLE IF NOT EXISTS project_retros (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
    author_id INTEGER NOT NULL REFERENCES users(id),
    went_well TEXT,
    went_bad TEXT,
    next_action TEXT,
    risk_note TEXT,
    created_at TEXT DEFAULT (datetime('now','localtime')),
    updated_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_retro_project ON project_retros(project_id);

-- 댓글 멘션
CREATE TABLE IF NOT EXISTS comment_mentions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    comment_id INTEGER NOT NULL REFERENCES task_comments(id) ON DELETE CASCADE,
    user_id INTEGER NOT NULL REFERENCES users(id),
    UNIQUE(comment_id, user_id)
);

CREATE TABLE IF NOT EXISTS task_delegations (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    task_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
    from_user_id INTEGER NOT NULL REFERENCES users(id),
    to_user_id INTEGER NOT NULL REFERENCES users(id),
    message TEXT,
    status TEXT DEFAULT 'pending',
    created_at TEXT DEFAULT (datetime('now','localtime')),
    resolved_at TEXT
);
CREATE INDEX IF NOT EXISTS idx_deleg_task ON task_delegations(task_id);
CREATE INDEX IF NOT EXISTS idx_deleg_to ON task_delegations(to_user_id, status);

-- ============ 물류 모듈 (HAIST WORKS) ============
-- 부품 마스터
CREATE TABLE IF NOT EXISTS parts (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    part_no     TEXT NOT NULL UNIQUE,
    part_name   TEXT NOT NULL,
    spec        TEXT,
    maker       TEXT,
    origin      TEXT,
    unit        TEXT DEFAULT 'EA',
    currency    TEXT DEFAULT 'KRW',
    std_price   REAL DEFAULT 0,
    biz_div     TEXT,                -- T 검사기 / M 자동화 / 공통
    category    TEXT,                -- 완성품/자재/소모품
    note        TEXT,
    is_active   INTEGER DEFAULT 1,
    created_at  TEXT DEFAULT (datetime('now','localtime')),
    updated_at  TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_parts_part_no ON parts(part_no);
CREATE INDEX IF NOT EXISTS idx_parts_biz_div ON parts(biz_div);
CREATE INDEX IF NOT EXISTS idx_parts_category ON parts(category);

-- v5H226z129 (2026-05-31): 자재 카탈로그 즐겨찾기 (사용자별 부품 북마크)
--   누락된 테이블 — catalog_page 가 사용하는데 정의가 없어 NAS 에서 'no such table' 오류 발생.
CREATE TABLE IF NOT EXISTS catalog_favorites (
    user_id     INTEGER NOT NULL,
    part_id     INTEGER NOT NULL,
    created_at  TEXT DEFAULT (datetime('now','localtime')),
    PRIMARY KEY (user_id, part_id)
);
CREATE INDEX IF NOT EXISTS idx_catfav_user ON catalog_favorites(user_id);

-- v5H226z130 (2026-05-31): 자재 요청서 (카탈로그 장바구니 → 요청 제출 + 신규자재 검토)
--   누락 테이블 보충 — main.py material_request_* 라우트가 사용하는데 NAS DB 에 미생성.
CREATE TABLE IF NOT EXISTS material_requests (
    id                INTEGER PRIMARY KEY AUTOINCREMENT,
    request_no        TEXT,
    request_type      TEXT DEFAULT 'bom',
    title             TEXT,
    project_id        INTEGER,
    requester_id      INTEGER,
    requester_team_id INTEGER,
    status            TEXT DEFAULT '제출',
    total_items       INTEGER DEFAULT 0,
    new_item_count    INTEGER DEFAULT 0,
    note              TEXT,
    created_at        TEXT DEFAULT (datetime('now','localtime')),
    updated_at        TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_matreq_requester ON material_requests(requester_id);
CREATE INDEX IF NOT EXISTS idx_matreq_status ON material_requests(status);

CREATE TABLE IF NOT EXISTS material_request_items (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    request_id          INTEGER NOT NULL,
    line_no             INTEGER,
    part_id             INTEGER,
    quantity            REAL DEFAULT 0,
    unit                TEXT DEFAULT 'EA',
    is_new_review       INTEGER DEFAULT 0,
    review_status       TEXT DEFAULT 'pending',
    requested_part_no   TEXT,
    requested_part_name TEXT,
    requested_spec      TEXT,
    requested_maker     TEXT,
    request_reason      TEXT,
    registered_part_id  INTEGER,
    review_note         TEXT,
    quote_supplier_id   INTEGER,
    quoted_price        REAL,
    quoted_lead_days    INTEGER,
    created_at          TEXT DEFAULT (datetime('now','localtime')),
    updated_at          TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_matreqitem_req ON material_request_items(request_id);
CREATE INDEX IF NOT EXISTS idx_matreqitem_newrev ON material_request_items(is_new_review, review_status);

-- 공급사 마스터 (발주 대상 거래처)
CREATE TABLE IF NOT EXISTS suppliers (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    name            TEXT NOT NULL,
    code            TEXT UNIQUE,
    contact         TEXT,
    email           TEXT,
    phone           TEXT,
    country         TEXT,
    currency        TEXT DEFAULT 'KRW',
    payment_terms   TEXT,             -- 선금/현금/30일/60일
    note            TEXT,
    is_active       INTEGER DEFAULT 1,
    created_at      TEXT DEFAULT (datetime('now','localtime')),
    updated_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_sup_name ON suppliers(name);

-- 발주 헤더
CREATE TABLE IF NOT EXISTS purchase_orders (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    po_number       TEXT NOT NULL UNIQUE,           -- PO-YYMMDD-NNN
    project_id      INTEGER REFERENCES projects(id),-- 관리코드 연결 (선택)
    supplier_id     INTEGER REFERENCES suppliers(id),
    order_date      TEXT NOT NULL,                   -- YYYY-MM-DD
    expected_date   TEXT,                            -- 예상 입고일
    currency        TEXT DEFAULT 'KRW',
    exchange_rate   REAL DEFAULT 1,
    total_amount    REAL DEFAULT 0,                  -- 라인 합계 (자동)
    status          TEXT DEFAULT '작성중',            -- 작성중/발주완료/부분입고/입고완료/취소
    shipping_terms  TEXT,                            -- EXW/FOB/CIF/DDP/국내
    payment_terms   TEXT,                            -- 선금/현금/30일/60일
    po_type         TEXT DEFAULT '일반',              -- 일반/긴급/정기
    created_by      INTEGER REFERENCES users(id),
    note            TEXT,
    created_at      TEXT DEFAULT (datetime('now','localtime')),
    updated_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_po_number ON purchase_orders(po_number);
CREATE INDEX IF NOT EXISTS idx_po_project ON purchase_orders(project_id);
CREATE INDEX IF NOT EXISTS idx_po_supplier ON purchase_orders(supplier_id);
CREATE INDEX IF NOT EXISTS idx_po_status ON purchase_orders(status);

-- 발주 라인
CREATE TABLE IF NOT EXISTS po_items (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    po_id               INTEGER NOT NULL REFERENCES purchase_orders(id) ON DELETE CASCADE,
    line_no             INTEGER DEFAULT 0,
    part_id             INTEGER REFERENCES parts(id),
    part_no_snapshot    TEXT,       -- 발주 시점 부품번호
    part_name_snapshot  TEXT,
    spec_snapshot       TEXT,
    unit                TEXT DEFAULT 'EA',
    quantity            REAL NOT NULL DEFAULT 0,
    unit_price          REAL NOT NULL DEFAULT 0,
    amount              REAL DEFAULT 0,    -- quantity * unit_price
    received_qty        REAL DEFAULT 0,    -- 누적 입고 수량 (4단계)
    delivery_date       TEXT,
    note                TEXT
);
CREATE INDEX IF NOT EXISTS idx_poitem_po ON po_items(po_id);
CREATE INDEX IF NOT EXISTS idx_poitem_part ON po_items(part_id);

-- v5H136 (2026-05-05): PO 라인 ↔ 프로젝트 다대다 연결
-- 목적: 검사기/장비 소모품 발주를 해당 장비(관리번호)에 귀속시켜
--   ① 자주 나가는 소모품 식별  ② 수리 이력 추적  ③ 장비별 운영비 집계
-- 공통 부품(연결 0건) 은 자연 폴백 — 기존 PO 데이터 무영향
CREATE TABLE IF NOT EXISTS po_item_project_links (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    po_item_id      INTEGER NOT NULL REFERENCES po_items(id) ON DELETE CASCADE,
    project_id      INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
    allocated_qty   REAL,                              -- 분배 수량 (NULL=라인 전체)
    allocation_pct  REAL,                              -- 또는 % (NULL=미사용)
    note            TEXT,
    created_at      TEXT DEFAULT (datetime('now','localtime')),
    created_by      INTEGER REFERENCES users(id)
);
CREATE INDEX IF NOT EXISTS idx_popl_po_item ON po_item_project_links(po_item_id);
CREATE INDEX IF NOT EXISTS idx_popl_project ON po_item_project_links(project_id);

-- =====================================================
-- v5H142 (2026-05-05): 소모품 발주 전용 도메인 (대표 직접 요청)
-- 신규 검사기와 분리 — 관리번호 발급 X, 엑셀 일괄 import + 이미지 자동 압축
-- =====================================================
CREATE TABLE IF NOT EXISTS consumable_orders (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    co_no           TEXT,                            -- CO-YYMMNN 자동 채번
    customer_id     INTEGER REFERENCES customers(id),
    customer_name   TEXT,
    order_date      TEXT,
    due_date        TEXT,
    currency        TEXT DEFAULT 'KRW',
    total_amount    REAL DEFAULT 0,
    status          TEXT DEFAULT 'DRAFT',            -- DRAFT/QUOTED/CONFIRMED/SHIPPED/PAID/CANCELLED
    note            TEXT,
    source_file     TEXT,                            -- 업로드 원본 파일명
    created_at      TEXT DEFAULT (datetime('now','localtime')),
    created_by      INTEGER REFERENCES users(id)
);
CREATE INDEX IF NOT EXISTS idx_co_cust ON consumable_orders(customer_id);
CREATE INDEX IF NOT EXISTS idx_co_status ON consumable_orders(status);
CREATE INDEX IF NOT EXISTS idx_co_date ON consumable_orders(order_date);

CREATE TABLE IF NOT EXISTS consumable_order_items (
    id                 INTEGER PRIMARY KEY AUTOINCREMENT,
    co_id              INTEGER NOT NULL REFERENCES consumable_orders(id) ON DELETE CASCADE,
    line_no            INTEGER,
    model_use          TEXT,                         -- 엑셀 MODEL USE
    part_id            INTEGER REFERENCES parts(id), -- 자재 마스터 매칭(선택)
    part_name          TEXT,                         -- 엑셀 SUPPLIER NAME (품명)
    spec               TEXT,
    qty                REAL,
    unit               TEXT DEFAULT 'EA',
    unit_price         REAL DEFAULT 0,
    amount             REAL DEFAULT 0,
    linked_project_id  INTEGER REFERENCES projects(id), -- 관리번호 연결(선택)
    note               TEXT,
    image_path         TEXT,                         -- 압축본 경로(사진/PICTURE)
    image_thumb_path   TEXT,                         -- 썸네일 경로(사진/PICTURE)
    image_loc_path     TEXT,                         -- v5H226z286: 사진위치(PICTURE LOCATION) 압축본
    image_loc_thumb_path TEXT,                       -- v5H226z286: 사진위치 썸네일
    equip_name         TEXT                          -- v5H226z288: 장비명
);
CREATE INDEX IF NOT EXISTS idx_coi_co ON consumable_order_items(co_id);
CREATE INDEX IF NOT EXISTS idx_coi_proj ON consumable_order_items(linked_project_id);
CREATE INDEX IF NOT EXISTS idx_coi_part ON consumable_order_items(part_id);

-- v5H226z298 (대표 지시): 소모품 수주 변경 이력(히스토리 탭) — 누가·언제·무엇을 어떻게 바꿨는지
CREATE TABLE IF NOT EXISTS consumable_history (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    co_id           INTEGER NOT NULL,
    item_id         INTEGER,                       -- 라인 변경이면 라인 id, 주문(헤더) 변경이면 NULL
    scope           TEXT,                          -- 'order' | 'line'
    field           TEXT,                          -- 컬럼명
    label           TEXT,                          -- 표시용 항목명(예: 고객사)
    old_value       TEXT,
    new_value       TEXT,
    changed_by      INTEGER,
    changed_by_name TEXT,
    changed_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_cohist_co ON consumable_history(co_id, id);

-- =====================================================
-- STOCK MOVEMENTS — 입출고 원장 (수불부) (2026-04-20)
-- 모든 재고 증감은 여기 기록 → parts.stock_qty는 합계 결과
-- =====================================================
CREATE TABLE IF NOT EXISTS stock_movements (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    movement_no     TEXT UNIQUE,                 -- SM-YYMMDD-NNN 자동 채번
    part_id         INTEGER NOT NULL REFERENCES parts(id),
    kind            TEXT NOT NULL,               -- IN(입고) / OUT(출고) / ADJUST(실사조정) / TRANSFER(이동)
    quantity        REAL NOT NULL,               -- 부호 있는 수량: IN=+, OUT=-, ADJUST=±
    unit            TEXT DEFAULT 'EA',
    unit_price      REAL DEFAULT 0,              -- 스냅샷 단가 (출고는 FIFO 가중평균)
    amount          REAL DEFAULT 0,              -- qty * unit_price (절대값 기준)
    -- FIFO·로트 추적 (2026-04-21, 자재_물류 리서치 §4·§6·§12 반영)
    remaining_qty   REAL DEFAULT 0,              -- IN 행 전용: 아직 소비되지 않은 수량 (FIFO)
    lot_no          TEXT,                        -- 로트/배치 번호 (공급사 lot or 자체 채번)
    expiry_date     TEXT,                        -- 유효기한 (선택)
    -- 참조 연결
    po_id           INTEGER REFERENCES purchase_orders(id),    -- 입고 시 발주 참조
    po_item_id      INTEGER REFERENCES po_items(id),           -- 입고 시 라인 참조
    project_id      INTEGER REFERENCES projects(id),           -- 출고 시 프로젝트
    customer_id     INTEGER REFERENCES customers(id),          -- 출고 시 고객사
    -- 메타
    reason          TEXT,                        -- 사유 (출고 목적, 조정 사유)
    location        TEXT,                        -- 창고/위치 (선택)
    occurred_at     TEXT NOT NULL,               -- 실제 발생 시각 (YYYY-MM-DD HH:MM)
    note            TEXT,
    created_by      INTEGER REFERENCES users(id),
    created_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_sm_part ON stock_movements(part_id);
CREATE INDEX IF NOT EXISTS idx_sm_kind ON stock_movements(kind);
CREATE INDEX IF NOT EXISTS idx_sm_po ON stock_movements(po_id);
CREATE INDEX IF NOT EXISTS idx_sm_project ON stock_movements(project_id);
CREATE INDEX IF NOT EXISTS idx_sm_occurred ON stock_movements(occurred_at);

-- =====================================================
-- 환율 · 시간별 단가 · 리드타임 (2026-04-21 동적 변수 리서치)
-- 근거: 자재관리_동적변수_대응시스템.md §3·§4·§6
-- =====================================================

-- 환율 테이블 — 날짜별 통화 환산 (FXLoader 패턴)
CREATE TABLE IF NOT EXISTS exchange_rates (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    rate_date       TEXT NOT NULL,                -- YYYY-MM-DD
    from_currency   TEXT NOT NULL,                -- USD / VND / JPY / CNY 등
    to_currency     TEXT NOT NULL DEFAULT 'KRW',  -- 기본 KRW
    rate            REAL NOT NULL,                -- 1 from = rate * to
    source          TEXT DEFAULT '수동',          -- 수동 / 한국은행 / 기타
    note            TEXT,
    created_by      INTEGER REFERENCES users(id),
    created_at      TEXT DEFAULT (datetime('now','localtime')),
    UNIQUE(rate_date, from_currency, to_currency)
);
CREATE INDEX IF NOT EXISTS idx_fx_date ON exchange_rates(rate_date);
CREATE INDEX IF NOT EXISTS idx_fx_from ON exchange_rates(from_currency);

-- 적용일자 기반 단가 관리 — 한국 ERP 표준 (Decomsoft/emaxit)
-- 기존 parts.std_price는 "기본 단가"로 유지, 이 테이블은 시간별 변동 관리
CREATE TABLE IF NOT EXISTS part_prices (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    part_id         INTEGER NOT NULL REFERENCES parts(id),
    supplier_id     INTEGER REFERENCES suppliers(id),  -- NULL = 공급사 무관 기본
    price_type      TEXT DEFAULT '견적',          -- 확정 / 가 / 견적 (한국 표준 3종)
    unit_price      REAL NOT NULL,
    currency        TEXT DEFAULT 'KRW',
    effective_from  TEXT NOT NULL,                -- 적용 시작일 (YYYY-MM-DD)
    effective_to    TEXT,                          -- 적용 종료일 (NULL = 무기한)
    negotiated_at   TEXT,                          -- 협의일자 (공급사와 합의)
    min_qty         REAL DEFAULT 0,                -- 수량 할인 준비 (④ 변수용 P1)
    max_qty         REAL,
    note            TEXT,
    approved_by     INTEGER REFERENCES users(id),
    approved_at     TEXT,                          -- NULL이면 대기중
    created_by      INTEGER REFERENCES users(id),
    created_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_ppr_part ON part_prices(part_id);
CREATE INDEX IF NOT EXISTS idx_ppr_supplier ON part_prices(supplier_id);
CREATE INDEX IF NOT EXISTS idx_ppr_effective ON part_prices(effective_from, effective_to);
CREATE INDEX IF NOT EXISTS idx_ppr_type ON part_prices(price_type);

-- ============ 게시판 (HAIST WORKS) ============
-- 게시판 마스터: 전사 / 부서별
CREATE TABLE IF NOT EXISTS boards (
    id        INTEGER PRIMARY KEY AUTOINCREMENT,
    name      TEXT NOT NULL,          -- '전사 게시판' / '검사기팀 게시판' 등
    type      TEXT NOT NULL,          -- 'company' 전사 / 'team' 부서
    team_id   INTEGER REFERENCES teams(id),  -- 부서 게시판일 때
    created_at TEXT DEFAULT (datetime('now','localtime')),
    UNIQUE(type, team_id)
);

-- 게시글
CREATE TABLE IF NOT EXISTS board_posts (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    board_id        INTEGER NOT NULL REFERENCES boards(id),
    author_id       INTEGER NOT NULL REFERENCES users(id),
    title           TEXT NOT NULL,
    body            TEXT,
    category        TEXT DEFAULT '일반',       -- 공지/일반/자료/질문
    is_pinned       INTEGER DEFAULT 0,         -- 고정 (관리자/팀장만)
    view_count      INTEGER DEFAULT 0,
    approval_status TEXT DEFAULT 'approved',    -- approved/pending/rejected
    approved_by     INTEGER REFERENCES users(id),
    approved_at     TEXT,
    reject_reason   TEXT,
    created_at      TEXT DEFAULT (datetime('now','localtime')),
    updated_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_bp_board ON board_posts(board_id);
CREATE INDEX IF NOT EXISTS idx_bp_author ON board_posts(author_id);
CREATE INDEX IF NOT EXISTS idx_bp_approval ON board_posts(approval_status);

-- 댓글
CREATE TABLE IF NOT EXISTS board_comments (
    id        INTEGER PRIMARY KEY AUTOINCREMENT,
    post_id   INTEGER NOT NULL REFERENCES board_posts(id) ON DELETE CASCADE,
    author_id INTEGER NOT NULL REFERENCES users(id),
    body      TEXT NOT NULL,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_bc_post ON board_comments(post_id);

-- ============ 변경 Inform 시스템 (HAIST WORKS) ============
-- 변경 사건 본체
CREATE TABLE IF NOT EXISTS changes (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    change_no       TEXT NOT NULL UNIQUE,         -- CHG-YYMMDD-NNN
    change_type     TEXT NOT NULL,                 -- 기구설계/전장설계/소프트웨어/BOM/도면/Concept/사양
    biz_div         TEXT,                          -- T 검사기 / M 자동화 (영향 부서 자동 판별용)
    target_kind     TEXT,                          -- project / part / document
    target_id       INTEGER,
    target_label    TEXT,                          -- 사람이 읽는 라벨 (예: "001T2604 검사기")
    project_id      INTEGER REFERENCES projects(id),
    title           TEXT NOT NULL,
    description     TEXT,
    before_value    TEXT,
    after_value     TEXT,
    attached_files  TEXT,                          -- JSON 배열
    urgency         TEXT DEFAULT '일반',            -- 일반/긴급/예약
    author_id       INTEGER NOT NULL REFERENCES users(id),
    status          TEXT DEFAULT '공지중',          -- 작성중/공지중/확인완료/취소
    notified_at     TEXT,
    completed_at    TEXT,
    created_at      TEXT DEFAULT (datetime('now','localtime')),
    updated_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_chg_no ON changes(change_no);
CREATE INDEX IF NOT EXISTS idx_chg_project ON changes(project_id);
CREATE INDEX IF NOT EXISTS idx_chg_status ON changes(status);
CREATE INDEX IF NOT EXISTS idx_chg_author ON changes(author_id);
CREATE INDEX IF NOT EXISTS idx_chg_created ON changes(created_at DESC);

-- 영향 부서/사용자
CREATE TABLE IF NOT EXISTS change_impacts (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    change_id       INTEGER NOT NULL REFERENCES changes(id) ON DELETE CASCADE,
    impact_kind     TEXT NOT NULL,                 -- team / user
    impact_team_id  INTEGER REFERENCES teams(id),
    impact_user_id  INTEGER REFERENCES users(id),
    auto_detected   INTEGER DEFAULT 1,              -- 1=자동, 0=수동 추가
    impact_reason   TEXT
);
CREATE INDEX IF NOT EXISTS idx_cimp_change ON change_impacts(change_id);
CREATE INDEX IF NOT EXISTS idx_cimp_team ON change_impacts(impact_team_id);

-- 확인 추적
CREATE TABLE IF NOT EXISTS change_reads (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    change_id       INTEGER NOT NULL REFERENCES changes(id) ON DELETE CASCADE,
    user_id         INTEGER NOT NULL REFERENCES users(id),
    read_at         TEXT,
    ack_at          TEXT,
    ack_note        TEXT,
    UNIQUE(change_id, user_id)
);
CREATE INDEX IF NOT EXISTS idx_cread_change ON change_reads(change_id);
CREATE INDEX IF NOT EXISTS idx_cread_user ON change_reads(user_id);

-- ============ 요청 티켓 시스템 (HAIST WORKS — 메신저 요청 누락 해결) ============
CREATE TABLE IF NOT EXISTS tickets (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    ticket_no       TEXT NOT NULL UNIQUE,           -- TKT-YYMMDD-NNN
    category        TEXT NOT NULL,                   -- 자재요청/긴급가공/MODIFY/검수요청/AS/기타
    title           TEXT NOT NULL,
    description     TEXT,
    requester_id    INTEGER NOT NULL REFERENCES users(id),
    recipient_team_id INTEGER REFERENCES teams(id),
    recipient_user_id INTEGER REFERENCES users(id),
    project_id      INTEGER REFERENCES projects(id),
    target_label    TEXT,
    urgency         TEXT DEFAULT '일반',
    status          TEXT DEFAULT '요청',
    source          TEXT DEFAULT 'web',
    source_chat_id  TEXT,
    due_date        TEXT,
    completed_at    TEXT,
    hours_estimated REAL,
    hours_actual    REAL,
    accept_note     TEXT,
    complete_note   TEXT,
    reject_reason   TEXT,
    created_at      TEXT DEFAULT (datetime('now','localtime')),
    updated_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_tkt_no ON tickets(ticket_no);
CREATE INDEX IF NOT EXISTS idx_tkt_requester ON tickets(requester_id);
CREATE INDEX IF NOT EXISTS idx_tkt_recipient_team ON tickets(recipient_team_id);
CREATE INDEX IF NOT EXISTS idx_tkt_status ON tickets(status);
CREATE INDEX IF NOT EXISTS idx_tkt_category ON tickets(category);

CREATE TABLE IF NOT EXISTS ticket_comments (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    ticket_id   INTEGER NOT NULL REFERENCES tickets(id) ON DELETE CASCADE,
    author_id   INTEGER NOT NULL REFERENCES users(id),
    body        TEXT NOT NULL,
    is_status_change INTEGER DEFAULT 0,
    created_at  TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_tcm_ticket ON ticket_comments(ticket_id);

-- ============ 진행률 대시보드 (HAIST WORKS — 1순위 ① / 8팀) ============
-- 프로젝트 × 12공정 진척률
CREATE TABLE IF NOT EXISTS project_phases (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id      INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
    phase_code      TEXT NOT NULL,                  -- order/concept/design/elec/sw/machining/buying/assembly/qc/ship/setup/knkvn
    phase_order     INTEGER NOT NULL,
    status          TEXT DEFAULT '예정',             -- 예정/진행/완료/지연/보류
    progress_pct    REAL DEFAULT 0,                  -- 0-100
    assignee_id     INTEGER REFERENCES users(id),
    assignee_team_id INTEGER REFERENCES teams(id),
    planned_start   TEXT,
    planned_end     TEXT,
    actual_start    TEXT,
    actual_end      TEXT,
    note            TEXT,
    updated_by      INTEGER REFERENCES users(id),
    updated_at      TEXT DEFAULT (datetime('now','localtime')),
    UNIQUE(project_id, phase_code)
);
CREATE INDEX IF NOT EXISTS idx_pp_project ON project_phases(project_id);
CREATE INDEX IF NOT EXISTS idx_pp_status ON project_phases(status);
CREATE INDEX IF NOT EXISTS idx_pp_assignee ON project_phases(assignee_id);

-- =====================================================
-- APP SETTINGS (key-value, admin/ceo만 변경) — 2026-04-20
-- 외부 그룹웨어 URL 등 운영 설정 저장
-- =====================================================
CREATE TABLE IF NOT EXISTS app_settings (
    key         TEXT PRIMARY KEY,
    value       TEXT,
    description TEXT,
    updated_at  TEXT DEFAULT (datetime('now','localtime')),
    updated_by  INTEGER REFERENCES users(id)
);

-- =====================================================
-- ISSUES · AS DB (3순위 ⑦) — 2026-04-20
-- 고객사 이슈/AS 발생 → 부서 추적 → 재발 방지 학습
-- =====================================================
CREATE TABLE IF NOT EXISTS issues (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    issue_no        TEXT UNIQUE,                       -- ISS-2604-001 자동 채번
    title           TEXT NOT NULL,
    severity        TEXT DEFAULT '중',                 -- 치명/심각/중/경
    issue_type      TEXT DEFAULT 'AS',                 -- AS/품질/설계결함/SW버그/기타
    status          TEXT DEFAULT '접수',                -- 접수/원인분석/조치중/해결/재발방지등록/종결
    customer_id     INTEGER REFERENCES customers(id),
    customer_name   TEXT,                              -- 백업 텍스트
    project_id      INTEGER REFERENCES projects(id),
    mgmt_code       TEXT,                              -- KNK PMS 8자리
    biz_div         TEXT,                              -- T/M
    occurred_at     TEXT,                              -- 발생일
    detected_by     TEXT,                              -- 발견자 (고객사 담당 or 사내)
    description     TEXT,                              -- 증상 설명
    root_cause      TEXT,                              -- 원인 분석 결과
    action_taken    TEXT,                              -- 조치 내역
    prevention      TEXT,                              -- 재발방지 대책
    owner_team_id   INTEGER REFERENCES teams(id),      -- 책임 부서
    owner_user_id   INTEGER REFERENCES users(id),      -- 담당자
    resolved_at     TEXT,                              -- 해결 시각
    cost_estimate   REAL DEFAULT 0,                    -- 손실/조치 비용 (원)
    related_change_id INTEGER REFERENCES changes(id),  -- 연관 변경 (재발방지로 변경 발생 시)
    created_by      INTEGER REFERENCES users(id),
    created_at      TEXT DEFAULT (datetime('now','localtime')),
    updated_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_iss_status ON issues(status);
CREATE INDEX IF NOT EXISTS idx_iss_owner_team ON issues(owner_team_id);
CREATE INDEX IF NOT EXISTS idx_iss_customer ON issues(customer_id);
CREATE INDEX IF NOT EXISTS idx_iss_project ON issues(project_id);
CREATE INDEX IF NOT EXISTS idx_iss_severity ON issues(severity);

-- 이슈 진행 로그 (코멘트/상태 변경 이력)
CREATE TABLE IF NOT EXISTS issue_logs (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    issue_id    INTEGER NOT NULL REFERENCES issues(id) ON DELETE CASCADE,
    user_id     INTEGER REFERENCES users(id),
    action      TEXT,                                  -- 코멘트/상태변경/원인추가/조치/재발방지
    content     TEXT,
    old_status  TEXT,
    new_status  TEXT,
    created_at  TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_ilog_issue ON issue_logs(issue_id);

-- =====================================================
-- TOP3 S3 — 권한 위임 1차 (DB 스키마, 2026-04-25)
-- 시안: 05_HAIST_WORKS_디자인팀/_TO_01_정식시안_Top3_S3_v1.md §데이터모델
-- 골격만 (UI/로직 다음 사이클). 모든 CREATE 는 idempotent 가드.
-- =====================================================

-- 권한 카탈로그 (resource·action 단위)
-- 2026-04-25 Top3-S3-2차 RBAC 컬럼 분리 (시안 §7 원안 정합):
--   기존: name(UNIQUE) / scope  (1차 단순화)
--   신규: resource / action / scope / description + UNIQUE(resource, action, scope)
--   옵션 B 채택 — name 컬럼 deprecated 유지 (호환성). ALTER TABLE ADD COLUMN 으로 보강.
CREATE TABLE IF NOT EXISTS permissions (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    name        TEXT UNIQUE NOT NULL,            -- [DEPRECATED 2026-04-25] 예: 'stock.write'
    scope       TEXT,                            -- 예: 'stock' / 'sales' / 'hr'
    created_at  TEXT DEFAULT (datetime('now','localtime'))
);

-- 권한 묶음 (Group Inheritance ③)
CREATE TABLE IF NOT EXISTS permission_groups (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    name         TEXT UNIQUE NOT NULL,
    description  TEXT,
    created_at   TEXT DEFAULT (datetime('now','localtime'))
);

-- 그룹↔권한 다대다
CREATE TABLE IF NOT EXISTS group_permissions (
    group_id      INTEGER NOT NULL REFERENCES permission_groups(id) ON DELETE CASCADE,
    permission_id INTEGER NOT NULL REFERENCES permissions(id) ON DELETE CASCADE,
    PRIMARY KEY (group_id, permission_id)
);

-- 사용자↔그룹 다대다
CREATE TABLE IF NOT EXISTS user_groups (
    user_id   INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
    group_id  INTEGER NOT NULL REFERENCES permission_groups(id) ON DELETE CASCADE,
    PRIMARY KEY (user_id, group_id)
);

-- 위임 토큰 (Delegation Token ④, 시안 §3 7테이블 中 delegations)
CREATE TABLE IF NOT EXISTS delegation_tokens (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    from_user       INTEGER NOT NULL REFERENCES users(id),
    to_user         INTEGER NOT NULL REFERENCES users(id),
    permission_id   INTEGER NOT NULL REFERENCES permissions(id),
    expires_at      TEXT,
    can_redelegate  INTEGER DEFAULT 0,           -- 0=OFF(기본) / 1=ON
    status          TEXT DEFAULT 'ACTIVE',       -- ACTIVE / EXPIRED / REVOKED
    created_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_dt_to ON delegation_tokens(to_user, status);
CREATE INDEX IF NOT EXISTS idx_dt_from ON delegation_tokens(from_user, status);

-- 감사 로그 (append-only, 시안 §3 permission_audit_log)
CREATE TABLE IF NOT EXISTS delegation_audit (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    token_id    INTEGER REFERENCES delegation_tokens(id),
    action      TEXT NOT NULL,                   -- GRANT / DELEGATE / REVOKE / EXPIRE
    actor       INTEGER REFERENCES users(id),
    timestamp   TEXT DEFAULT (datetime('now','localtime')),
    details     TEXT
);
CREATE INDEX IF NOT EXISTS idx_da_token ON delegation_audit(token_id);
CREATE INDEX IF NOT EXISTS idx_da_actor ON delegation_audit(actor, timestamp);

-- =====================================================
-- TOP3 S2 — 재고 입출고 1차 (2026-04-25)
-- 시안: 05_HAIST_WORKS_디자인팀/_TO_01_정식시안_Top3_S2_v1.md §데이터 모델
-- 1차 = 스키마 골격만 (실 INSERT/UPDATE 로직은 다음 사이클).
-- 모두 idempotent (CREATE TABLE/VIEW IF NOT EXISTS).
-- =====================================================

-- 입고 헤더 GR (시안 §데이터 모델 — goods_receipts 별칭)
CREATE TABLE IF NOT EXISTS receipts (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    po_id               INTEGER REFERENCES purchase_orders(id),
    received_at         TEXT DEFAULT (datetime('now','localtime')),
    received_by         INTEGER REFERENCES users(id),
    total_qty           REAL DEFAULT 0,
    qc_inspection_id    INTEGER,                     -- qc_inspections(id) 역참조
    status              TEXT DEFAULT 'PENDING',      -- PENDING/PASS/PARTIAL/FAIL
    note                TEXT
);
CREATE INDEX IF NOT EXISTS idx_receipts_po ON receipts(po_id);
CREATE INDEX IF NOT EXISTS idx_receipts_status ON receipts(status);

-- QC 검수 (시안 §데이터 모델 — PENDING/PASS/FAIL/HOLD/PARTIAL/CONCESSION)
CREATE TABLE IF NOT EXISTS qc_inspections (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    po_item_id      INTEGER REFERENCES po_items(id),
    receipt_id      INTEGER REFERENCES receipts(id),
    inspector_id    INTEGER REFERENCES users(id),
    inspected_at    TEXT DEFAULT (datetime('now','localtime')),
    pass_qty        REAL DEFAULT 0,
    fail_qty        REAL DEFAULT 0,
    fail_reason     TEXT,
    status          TEXT DEFAULT 'PENDING'           -- PENDING/PASS/FAIL/HOLD/PARTIAL
);
CREATE INDEX IF NOT EXISTS idx_qc_poitem ON qc_inspections(po_item_id);
CREATE INDEX IF NOT EXISTS idx_qc_status ON qc_inspections(status);

-- 부적합 처리 (시안 §데이터 모델 — RETURN/SPECIAL_ACCEPT/SCRAP)
CREATE TABLE IF NOT EXISTS qc_disposition (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    qc_inspection_id    INTEGER NOT NULL REFERENCES qc_inspections(id),
    action              TEXT NOT NULL,                -- RETURN/SPECIAL_ACCEPT/SCRAP
    decided_by          INTEGER REFERENCES users(id),
    decided_at          TEXT DEFAULT (datetime('now','localtime')),
    note                TEXT
);
CREATE INDEX IF NOT EXISTS idx_qcdisp_qc ON qc_disposition(qc_inspection_id);

-- 출고 GI (시안 §데이터 모델 — PENDING/APPROVED/ISSUED/REJECTED)
CREATE TABLE IF NOT EXISTS issues_out (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    part_id         INTEGER NOT NULL REFERENCES parts(id),
    requester_id    INTEGER REFERENCES users(id),
    approver_id     INTEGER REFERENCES users(id),
    requested_at    TEXT DEFAULT (datetime('now','localtime')),
    issued_at       TEXT,
    qty             REAL NOT NULL DEFAULT 0,
    purpose         TEXT,
    status          TEXT DEFAULT 'PENDING'           -- PENDING/APPROVED/ISSUED/REJECTED
);
CREATE INDEX IF NOT EXISTS idx_issues_part ON issues_out(part_id);
CREATE INDEX IF NOT EXISTS idx_issues_status ON issues_out(status);

-- 재고 잔고 VIEW (시안 §데이터 모델 — stock_balances MATERIALIZED VIEW · SQLite는 일반 VIEW)
-- on_hand = sum(IN) - sum(OUT) - sum(ADJUST 부호 합) · idempotent
-- 사이클 52 (2026-04-27) 결함 수정: parts.name → parts.part_name (실제 컬럼명 정합).
-- DROP VIEW IF EXISTS + CREATE VIEW 패턴으로 기존 DB도 자동 재생성 (마이그레이션 안전).
DROP VIEW IF EXISTS stock_balances;
CREATE VIEW stock_balances AS
SELECT  p.id                                  AS part_id,
        p.part_no                             AS part_no,
        p.part_name                           AS part_name,
        COALESCE(SUM(sm.quantity), 0)         AS on_hand,
        p.unit                                AS unit,
        MAX(sm.occurred_at)                   AS last_movement_at
FROM    parts p
LEFT JOIN stock_movements sm ON sm.part_id = p.id
GROUP BY p.id, p.part_no, p.part_name, p.unit;

-- =====================================================
-- TOP3 S1 — 매출 라이프사이클 (2026-04-25 Top3-S1-2차 시안 정합)
-- 시안: 05_HAIST_WORKS_디자인팀/_TO_01_정식시안_Top3_S1_v1.md §3
-- 시안 정합 enum 9종(8 + CANCELLED):
--   DRAFT / QUOTED / CONFIRMED / IN_PRODUCTION /
--   READY_TO_SHIP / SHIPPED / INVOICED / PAID / CANCELLED
-- (1차 CLOSED·PARTIAL_RECEIPT 제거 — 시안 우선 채택, PARTIAL_RECEIPT 는
--  receipts_payment 1:N 합계 기반 분기 처리)
-- 2차 추가 테이블: invoices(세금계산서), order_status_history(상태 이력)
-- idempotent 가드 (CREATE TABLE IF NOT EXISTS) 전건 적용.
-- 기존 customers 테이블은 보존 (변경 0).
-- =====================================================

-- 견적 (시안 §3 quotations — 탭 1 데이터 소스)
CREATE TABLE IF NOT EXISTS quotations (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    quote_no        TEXT UNIQUE,                             -- QT-YYYYMM-####
    customer_id     INTEGER REFERENCES customers(id),
    total_amount    REAL DEFAULT 0,
    valid_until     TEXT,                                    -- YYYY-MM-DD
    version         INTEGER DEFAULT 1,                       -- 시안 §1 탭1 Ver 노출
    status          TEXT DEFAULT 'DRAFT'
                    CHECK(status IN ('DRAFT','QUOTED','CONFIRMED','CANCELLED')),
    created_by      INTEGER REFERENCES users(id),
    created_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_quotations_customer ON quotations(customer_id);
CREATE INDEX IF NOT EXISTS idx_quotations_status ON quotations(status);

-- 견적 라인 (사이클 58 2차 보강 — quotation_items)
-- 1차에서 헤더(quotations)만 있었음. 2차에서 라인 분리 — 인쇄/수주전환의 데이터 소스.
CREATE TABLE IF NOT EXISTS quotation_items (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    quotation_id    INTEGER NOT NULL REFERENCES quotations(id),
    line_no         INTEGER DEFAULT 1,
    part_id         INTEGER REFERENCES parts(id),                -- NULL 허용 (자유 품목)
    item_name       TEXT,
    qty             REAL NOT NULL DEFAULT 0,
    unit            TEXT,                                        -- EA / SET / kg 등
    unit_price      REAL NOT NULL DEFAULT 0,
    total_price     REAL NOT NULL DEFAULT 0,                     -- qty * unit_price (호출자 산출)
    note            TEXT,
    created_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_quotation_items_quotation ON quotation_items(quotation_id);
CREATE INDEX IF NOT EXISTS idx_quotation_items_part ON quotation_items(part_id);

-- 수주 헤더 (시안 §3 orders — 탭 2 SO)
CREATE TABLE IF NOT EXISTS orders (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    order_no        TEXT UNIQUE,                             -- SO-YYYYMM-####
    quote_id        INTEGER REFERENCES quotations(id),
    customer_id     INTEGER REFERENCES customers(id),
    order_date      TEXT,                                    -- YYYY-MM-DD
    due_date        TEXT,                                    -- 시안 §1 탭2 납기
    total_amount    REAL DEFAULT 0,
    status          TEXT DEFAULT 'CONFIRMED'
                    CHECK(status IN ('DRAFT','QUOTED','CONFIRMED','IN_PRODUCTION',
                                     'READY_TO_SHIP','SHIPPED','INVOICED','PAID','CANCELLED')),
    created_by      INTEGER REFERENCES users(id),
    created_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_orders_customer ON orders(customer_id);
CREATE INDEX IF NOT EXISTS idx_orders_status ON orders(status);
CREATE INDEX IF NOT EXISTS idx_orders_due ON orders(due_date);

-- 수주 라인 (시안 §3 sales_order_items 별칭 order_items)
CREATE TABLE IF NOT EXISTS order_items (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    order_id        INTEGER NOT NULL REFERENCES orders(id),
    part_id         INTEGER REFERENCES parts(id),
    qty             REAL NOT NULL DEFAULT 0,
    unit_price      REAL NOT NULL DEFAULT 0,
    amount          REAL NOT NULL DEFAULT 0,                 -- qty * unit_price (트리거 X · 호출자 산출)
    allocated_qty   REAL DEFAULT 0                           -- 시안 §2 할당재고
);
CREATE INDEX IF NOT EXISTS idx_order_items_order ON order_items(order_id);
CREATE INDEX IF NOT EXISTS idx_order_items_part ON order_items(part_id);

-- v5H226z208 (2026-06-04 대표 지시): 프로젝트 품목 (새 규칙 — 관리번호=프로젝트, 그 아래 품목 N개)
--   제안 단계부터 품목을 보관(수주 SO 와 독립). 출고형태별 칸 규칙:
--     ASSEMBLY(ASS'Y) → 장비명(item_name)+모델명(model_name)
--     PARTS(부품)      → 품명(item_name)+모델명(model_name)
--     ETC(기타)        → 품명(item_name) (모델명 없음)
--   수주확정 시 이 품목이 전개된 SO 를 order_id 로 연결(NULL=제안단계·미전개).
CREATE TABLE IF NOT EXISTS project_items (
    id                 INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id         INTEGER NOT NULL REFERENCES projects(id),
    seq                INTEGER DEFAULT 1,                 -- 품목 순번(표시 순서)
    shipment_form      TEXT DEFAULT 'ASSEMBLY',           -- ASSEMBLY/PARTS/ETC
    item_name          TEXT,                              -- 장비명(ASS'Y) 또는 품명(부품/기타)
    model_name         TEXT,                              -- 모델명 (ASS'Y·부품; 기타는 비움)
    qty                INTEGER DEFAULT 1,                 -- 수량(대수/낱개)
    unit_price         REAL DEFAULT 0,
    amount             REAL DEFAULT 0,                    -- 명시 금액 또는 단가×수량
    currency           TEXT DEFAULT 'KRW',
    is_export          INTEGER DEFAULT 0,                 -- 0=내수,1=수출
    order_customer_id  INTEGER REFERENCES customers(id),  -- 발주처(구매대행사 등)
    secondary_customer TEXT,                              -- 2차 고객사(최종)
    final_amount       REAL,                              -- 최종 고객 매출(참고)
    order_date         TEXT,
    due_date           TEXT,
    ship_to            TEXT,
    note               TEXT,
    order_id           INTEGER REFERENCES orders(id),     -- 수주확정 시 전개된 SO (NULL=제안단계)
    created_at         TEXT DEFAULT (datetime('now','localtime')),
    updated_at         TEXT
);
CREATE INDEX IF NOT EXISTS idx_project_items_project ON project_items(project_id);
CREATE INDEX IF NOT EXISTS idx_project_items_order ON project_items(order_id);

-- v5H226z215 (2026-06-05 대표 지시): A/S 관리(수리 접수·처리) — 라이프밸류 전용 신설 기능.
--   고객이 제품을 택배로 보냄 → 수리 → 재발송 (RMA). 흐름: 접수 → 수리중 → 발송완료 (+보류/반송)
CREATE TABLE IF NOT EXISTS as_tickets (
    id                INTEGER PRIMARY KEY AUTOINCREMENT,
    as_no             TEXT UNIQUE,                          -- 접수번호 AS-YYMMDD-NNN
    received_date     TEXT,                                 -- 접수일 (YYYY-MM-DD)
    biz_div           TEXT DEFAULT 'L',                     -- 사업부 (라이프밸류 L 중심)
    -- 고객 (개인/회사) — customers 연결 또는 자유 입력
    customer_id       INTEGER REFERENCES customers(id),
    customer_name     TEXT,
    customer_phone    TEXT,
    customer_address  TEXT,
    -- 제품
    product_id        INTEGER,                              -- B2C 제품 마스터 연결(향후, NULL 허용)
    product_name      TEXT,
    model_name        TEXT,
    purchase_date     TEXT,                                 -- 구매일 (보증 판단)
    symptom           TEXT,                                 -- 증상(고장 내용)
    -- 택배 송장
    inbound_tracking  TEXT,                                 -- 받은 택배 송장(고객→본사)
    outbound_tracking TEXT,                                 -- 보낸 택배 송장(본사→고객)
    shipped_date      TEXT,                                 -- 발송일
    -- 수리
    is_paid           INTEGER DEFAULT 0,                    -- 0=무상, 1=유상
    repair_fee        REAL DEFAULT 0,                       -- 수리비(유상 시)
    repair_note       TEXT,                                 -- 수리 내역(작업·교체 부품)
    -- 상태/담당
    status            TEXT DEFAULT '접수'
                      CHECK(status IN ('접수','수리중','발송완료','보류','반송')),
    assignee_id       INTEGER REFERENCES users(id),         -- 담당자
    note              TEXT,
    created_by        INTEGER REFERENCES users(id),
    created_at        TEXT DEFAULT (datetime('now','localtime')),
    updated_at        TEXT
);
CREATE INDEX IF NOT EXISTS idx_as_tickets_status ON as_tickets(status);
CREATE INDEX IF NOT EXISTS idx_as_tickets_customer ON as_tickets(customer_id);
CREATE INDEX IF NOT EXISTS idx_as_tickets_received ON as_tickets(received_date);

-- A/S 상태 이력 (접수→수리중→발송완료 등 변경 추적)
CREATE TABLE IF NOT EXISTS as_status_history (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    as_id       INTEGER NOT NULL REFERENCES as_tickets(id),
    from_status TEXT,
    to_status   TEXT,
    changed_by  INTEGER REFERENCES users(id),
    note        TEXT,
    changed_at  TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_as_status_history_as ON as_status_history(as_id);

-- 생산지시 (시안 §3 work_orders 별칭 production_orders — 탭 3 WO)
CREATE TABLE IF NOT EXISTS production_orders (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    order_id        INTEGER NOT NULL REFERENCES orders(id),
    planned_start   TEXT,
    planned_end     TEXT,
    actual_start    TEXT,
    actual_end      TEXT,
    status          TEXT DEFAULT 'IN_PRODUCTION'
                    CHECK(status IN ('PLANNED','IN_PRODUCTION','READY_TO_SHIP','DONE','CANCELLED')),
    created_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_prodorders_order ON production_orders(order_id);
CREATE INDEX IF NOT EXISTS idx_prodorders_status ON production_orders(status);

-- 출하 (시안 §3 delivery_notes 별칭 shipments — 탭 4 DO · 1:N 부분출하 지원)
CREATE TABLE IF NOT EXISTS shipments (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    order_id        INTEGER NOT NULL REFERENCES orders(id),
    shipped_at      TEXT DEFAULT (datetime('now','localtime')),
    shipped_qty     REAL DEFAULT 0,
    shipped_by      INTEGER REFERENCES users(id),
    tracking        TEXT
);
CREATE INDEX IF NOT EXISTS idx_shipments_order ON shipments(order_id);

-- 수금 (시안 §3 receipts 별칭 receipts_payment — 기존 receipts(입고)와 분리 · 1:N 부분수금)
CREATE TABLE IF NOT EXISTS receipts_payment (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    order_id        INTEGER NOT NULL REFERENCES orders(id),
    received_at     TEXT DEFAULT (datetime('now','localtime')),
    amount          REAL NOT NULL DEFAULT 0,
    method          TEXT,                                    -- 현금/카드/이체/어음 등
    received_by     INTEGER REFERENCES users(id),
    note            TEXT
);
CREATE INDEX IF NOT EXISTS idx_receipts_payment_order ON receipts_payment(order_id);

-- 결제조건 (시안 §3 payment_terms — 거래처별 NET30 등)
CREATE TABLE IF NOT EXISTS payment_terms (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    customer_id     INTEGER REFERENCES customers(id),
    terms           TEXT,                                    -- NET30 / NET60 / 선금50% 등
    created_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_payment_terms_customer ON payment_terms(customer_id);

-- 세금계산서 (시안 §3 invoices — 탭4 INV · 2차 시안 정합 보정)
CREATE TABLE IF NOT EXISTS invoices (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    invoice_no      TEXT UNIQUE,                             -- INV-YYYYMM-####
    order_id        INTEGER NOT NULL REFERENCES orders(id),
    customer_id     INTEGER REFERENCES customers(id),
    issue_date      TEXT,                                    -- YYYY-MM-DD
    amount_excl_vat REAL DEFAULT 0,
    vat             REAL DEFAULT 0,
    total_amount    REAL DEFAULT 0,
    status          TEXT DEFAULT 'DRAFT'
                    CHECK(status IN ('DRAFT','ISSUED','CANCELLED')),
    created_by      INTEGER REFERENCES users(id),
    created_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_invoices_order ON invoices(order_id);
CREATE INDEX IF NOT EXISTS idx_invoices_issue_date ON invoices(issue_date);

-- 상태 이력 (시안 §2 order_status_history — 우측 타임라인 스텝퍼 데이터 소스)
CREATE TABLE IF NOT EXISTS order_status_history (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    order_id        INTEGER NOT NULL REFERENCES orders(id),
    from_status     TEXT,
    to_status       TEXT,
    changed_by      INTEGER REFERENCES users(id),
    changed_at      TEXT DEFAULT (datetime('now','localtime')),
    note            TEXT
);
CREATE INDEX IF NOT EXISTS idx_order_status_history_order ON order_status_history(order_id);
CREATE INDEX IF NOT EXISTS idx_order_status_history_changed_at ON order_status_history(changed_at);

-- =====================================================
-- 수출입 서류 — P11 베트남 수출 실무자 1차 (2026-04-25)
-- 한국 수출 표준 6테이블: EO / CI / PL / PL_ITEM / BL / CUSTOMS
-- 매출 자동 채움: orders 테이블 참조 (Top3 S1 완결 기반)
-- idempotent 가드 (CREATE TABLE IF NOT EXISTS) 전건 적용.
-- =====================================================

-- 수출 수주 헤더 (orders 와 1:1 — 매출 자동 채움 소스)
CREATE TABLE IF NOT EXISTS export_orders (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    order_id            INTEGER REFERENCES orders(id),
    buyer               TEXT,                                    -- 해외 바이어명
    shipping_terms      TEXT,                                    -- FOB / CIF / EXW / DDP 등
    payment_terms       TEXT,                                    -- T/T / L/C / D/P / D/A 등
    port_of_loading     TEXT,                                    -- 선적항 (예: BUSAN)
    port_of_discharge   TEXT,                                    -- 양륙항 (예: HAIPHONG)
    status              TEXT DEFAULT 'DRAFT'
                        CHECK(status IN ('DRAFT','BOOKED','CI_ISSUED','PL_READY','SHIPPED','CLEARED','CLOSED','CANCELLED')),
    created_by          INTEGER REFERENCES users(id),
    created_at          TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_export_orders_order ON export_orders(order_id);
CREATE INDEX IF NOT EXISTS idx_export_orders_status ON export_orders(status);

-- 상업송장 (CI — Commercial Invoice)
CREATE TABLE IF NOT EXISTS commercial_invoices (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    invoice_no          TEXT UNIQUE,                             -- CI-YYYYMM-####
    export_order_id     INTEGER NOT NULL REFERENCES export_orders(id),
    issue_date          TEXT,                                    -- YYYY-MM-DD
    total_amount        REAL DEFAULT 0,
    currency            TEXT DEFAULT 'USD',
    signed_by           INTEGER REFERENCES users(id),
    status              TEXT DEFAULT 'DRAFT'
                        CHECK(status IN ('DRAFT','ISSUED','CANCELLED')),
    created_at          TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_ci_eo ON commercial_invoices(export_order_id);
CREATE INDEX IF NOT EXISTS idx_ci_status ON commercial_invoices(status);

-- 패킹리스트 (PL — Packing List 헤더)
CREATE TABLE IF NOT EXISTS packing_lists (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    pl_no               TEXT UNIQUE,                             -- PL-YYYYMM-####
    export_order_id     INTEGER NOT NULL REFERENCES export_orders(id),
    total_packages      INTEGER DEFAULT 0,
    total_weight        REAL DEFAULT 0,                          -- kg (G/W)
    total_volume        REAL DEFAULT 0,                          -- CBM
    created_at          TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_pl_eo ON packing_lists(export_order_id);

-- 패킹 라인 (PL Item)
CREATE TABLE IF NOT EXISTS packing_items (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    pl_id               INTEGER NOT NULL REFERENCES packing_lists(id),
    part_id             INTEGER REFERENCES parts(id),
    qty                 REAL NOT NULL DEFAULT 0,
    package_type        TEXT,                                    -- CARTON / PALLET / WOODEN CASE 등
    weight              REAL DEFAULT 0,                          -- kg
    volume              REAL DEFAULT 0                           -- CBM
);
CREATE INDEX IF NOT EXISTS idx_pl_items_pl ON packing_items(pl_id);
CREATE INDEX IF NOT EXISTS idx_pl_items_part ON packing_items(part_id);

-- 선하증권 (B/L — Bill of Lading)
CREATE TABLE IF NOT EXISTS bills_of_lading (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    bl_no               TEXT UNIQUE,                             -- 운송사 발급 번호
    export_order_id     INTEGER NOT NULL REFERENCES export_orders(id),
    shipping_company    TEXT,                                    -- 운송사명 (외부 API 미사용 · 수동 입력)
    vessel              TEXT,                                    -- 선명 / 항공편명
    departure_date      TEXT,                                    -- YYYY-MM-DD
    arrival_date        TEXT,                                    -- YYYY-MM-DD (예정/실제)
    tracking_no         TEXT,                                    -- 운송사 추적번호
    status              TEXT DEFAULT 'DRAFT'
                        CHECK(status IN ('DRAFT','ISSUED','IN_TRANSIT','DELIVERED','CANCELLED')),
    created_at          TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_bl_eo ON bills_of_lading(export_order_id);
CREATE INDEX IF NOT EXISTS idx_bl_status ON bills_of_lading(status);

-- 관세 신고 (Customs Declaration)
CREATE TABLE IF NOT EXISTS customs_declarations (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    declaration_no      TEXT UNIQUE,                             -- 신고번호 (수동 입력)
    export_order_id     INTEGER NOT NULL REFERENCES export_orders(id),
    hs_code             TEXT,                                    -- HS 코드
    fta_origin          TEXT,                                    -- KR / VN / -- 등 (FTA 원산지)
    declared_value      REAL DEFAULT 0,                          -- 신고가액
    cleared_at          TEXT,                                    -- 통관 완료 일시
    status              TEXT DEFAULT 'DRAFT'
                        CHECK(status IN ('DRAFT','SUBMITTED','CLEARED','REJECTED','CANCELLED')),
    created_at          TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_customs_eo ON customs_declarations(export_order_id);
CREATE INDEX IF NOT EXISTS idx_customs_status ON customs_declarations(status);

-- =====================================================
-- FTA 원산지증명서 (사이클 75 · 2026-04-27 · 04 시뮬 MISSING #1)
-- 안지연 본업: 검사기·자동화장비 수출입 시 원산지증명서(C/O) 발급
-- KAFTA / KEUFTA / KCFTA / KVFTA / RCEP 5종. 외부 PDF 라이브러리 0건.
-- idempotent (CREATE TABLE IF NOT EXISTS).
-- =====================================================
CREATE TABLE IF NOT EXISTS fta_certificates (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    cert_no             TEXT UNIQUE,                              -- FTA-YYYY-####
    fta_type            TEXT,                                     -- KAFTA / KEUFTA / KCFTA / KVFTA / RCEP
    customer_id         INTEGER REFERENCES customers(id),         -- KNK 고객사 (있으면)
    customer_name       TEXT,                                     -- 거래처명 (스냅샷)
    customer_address    TEXT,                                     -- 거래처 주소
    customer_country    TEXT,                                     -- 거래처 국가
    export_order_id     INTEGER REFERENCES export_orders(id),     -- 매핑 (선택)
    export_invoice_no   TEXT,                                     -- CI 번호 (스냅샷)
    export_date         TEXT,                                     -- 수출일 YYYY-MM-DD
    origin_country      TEXT DEFAULT 'KR',                        -- 한국(KR) / 베트남(VN) / 중국(CN)
    total_value         REAL DEFAULT 0,                           -- 총액
    currency            TEXT DEFAULT 'USD',
    issuer_id           INTEGER REFERENCES users(id),             -- 발급자 (안지연 등)
    issuer_name         TEXT,                                     -- 발급자명 스냅샷
    issued_at           TEXT,                                     -- 발급일시
    status              TEXT DEFAULT 'DRAFT'
                        CHECK(status IN ('DRAFT','ISSUED','SENT','CANCELLED')),
    remarks             TEXT,                                     -- 비고
    created_by          INTEGER REFERENCES users(id),
    created_at          TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_fta_cert_customer ON fta_certificates(customer_id);
CREATE INDEX IF NOT EXISTS idx_fta_cert_status ON fta_certificates(status);
CREATE INDEX IF NOT EXISTS idx_fta_cert_type ON fta_certificates(fta_type);

CREATE TABLE IF NOT EXISTS fta_certificate_items (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    cert_id             INTEGER NOT NULL REFERENCES fta_certificates(id) ON DELETE CASCADE,
    line_no             INTEGER DEFAULT 1,
    part_id             INTEGER REFERENCES parts(id),
    part_name           TEXT,                                     -- 품명 스냅샷
    hs_code             TEXT,                                     -- HS 코드
    qty                 REAL DEFAULT 0,
    unit                TEXT,                                     -- EA / SET / KG 등
    unit_price          REAL DEFAULT 0,
    origin_country      TEXT,                                     -- 라인별 원산지
    total               REAL DEFAULT 0
);
CREATE INDEX IF NOT EXISTS idx_fta_item_cert ON fta_certificate_items(cert_id);
CREATE INDEX IF NOT EXISTS idx_fta_item_part ON fta_certificate_items(part_id);

-- =====================================================
-- 검사기 출하성적서 QC INSPECTION REPORT (사이클 76 · 2026-04-27 · 04 시뮬 MISSING #2)
-- 김정록 본업: 검사기 반복성 검증·문제점 파악·출하성적서 작성
-- 항목: 반복성 / 정확도 / 통신 / 외관 / 동작 / 안전 (코드측 표준 6종)
-- 외부 PDF 라이브러리 0건 (HTML 인쇄). idempotent.
-- =====================================================
CREATE TABLE IF NOT EXISTS qc_inspection_reports (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    report_no           TEXT UNIQUE,                              -- QCR-YYYY-####
    customer_id         INTEGER REFERENCES customers(id),
    customer_name       TEXT,                                     -- 거래처명 스냅샷
    order_id            INTEGER REFERENCES orders(id),            -- 매핑 (선택)
    order_no            TEXT,                                     -- 수주번호 스냅샷
    part_id             INTEGER REFERENCES parts(id),             -- 검사기 부품 매핑
    machine_model       TEXT,                                     -- 검사기 모델명 (예: HAIST-INS-VX1)
    machine_serial      TEXT,                                     -- 시리얼 번호
    inspection_date     TEXT,                                     -- 검사일 YYYY-MM-DD
    inspector_id        INTEGER REFERENCES users(id),             -- 검사자 (김정록 등)
    inspector_name      TEXT,                                     -- 검사자명 스냅샷
    qa_manager_id       INTEGER REFERENCES users(id),             -- QA 책임자 (서명자)
    qa_manager_name     TEXT,                                     -- QA 책임자명 스냅샷
    overall             TEXT DEFAULT 'PASS'
                        CHECK(overall IN ('PASS','FAIL','CONDITIONAL_PASS')),
    issued_at           TEXT,                                     -- 발급일시
    sent_at             TEXT,                                     -- 발송일시 (SENT 전이)
    status              TEXT DEFAULT 'DRAFT'
                        CHECK(status IN ('DRAFT','ISSUED','SENT','CANCELLED')),
    remarks             TEXT,                                     -- 비고
    created_by          INTEGER REFERENCES users(id),
    created_at          TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_qcr_customer ON qc_inspection_reports(customer_id);
CREATE INDEX IF NOT EXISTS idx_qcr_status ON qc_inspection_reports(status);
CREATE INDEX IF NOT EXISTS idx_qcr_order ON qc_inspection_reports(order_id);

CREATE TABLE IF NOT EXISTS qc_inspection_items (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    report_id           INTEGER NOT NULL REFERENCES qc_inspection_reports(id) ON DELETE CASCADE,
    line_no             INTEGER DEFAULT 1,
    item_name           TEXT NOT NULL,                            -- 반복성/정확도/통신/외관/동작/안전 등
    spec_value          TEXT,                                     -- 기준값 (≤0.5μm, 100±0.1mm 등)
    measured_value      TEXT,                                     -- 측정값
    judgment            TEXT DEFAULT 'PASS'
                        CHECK(judgment IN ('PASS','FAIL','NA')),
    remarks             TEXT
);
CREATE INDEX IF NOT EXISTS idx_qcr_item_report ON qc_inspection_items(report_id);

-- =====================================================
-- WORK ORDERS (2026-04-27 사이클77 — 가공팀 작업지시서 · 윤영조·이수빈 본업)
-- 가공 단계별 라인 보유 (절삭/연마/검수). production_orders(매출-생산 연계)와 별도 테이블.
-- 04 시뮬 MISSING #3 보완. 외부 자산 0건 (인쇄는 HTML window.print).
-- =====================================================
CREATE TABLE IF NOT EXISTS work_orders (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    wo_no               TEXT UNIQUE,                              -- WO-YYYY-####
    order_id            INTEGER REFERENCES orders(id),            -- 수주 기반 (선택)
    project_id          INTEGER REFERENCES projects(id),          -- 프로젝트 기반 (선택)
    part_id             INTEGER REFERENCES parts(id),             -- 가공 부품
    qty                 REAL DEFAULT 0,                           -- 가공 수량
    assigned_to         INTEGER REFERENCES users(id),             -- 작업자 (이수빈 등)
    assigned_name       TEXT,                                     -- 작업자명 스냅샷
    created_by          INTEGER REFERENCES users(id),             -- 작성자 (윤영조 등)
    created_by_name     TEXT,                                     -- 작성자명 스냅샷
    planned_start       TEXT,                                     -- YYYY-MM-DD
    planned_end         TEXT,
    actual_end          TEXT,                                     -- 실제 완료일
    specifications      TEXT,                                     -- 가공 사양 (자유 텍스트)
    status              TEXT DEFAULT 'DRAFT'
                        CHECK(status IN ('DRAFT','RELEASED','IN_PROGRESS','COMPLETED','CANCELLED')),
    remarks             TEXT,
    created_at          TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_work_orders_status ON work_orders(status);
CREATE INDEX IF NOT EXISTS idx_work_orders_order ON work_orders(order_id);
CREATE INDEX IF NOT EXISTS idx_work_orders_part ON work_orders(part_id);
CREATE INDEX IF NOT EXISTS idx_work_orders_assigned ON work_orders(assigned_to);

CREATE TABLE IF NOT EXISTS work_order_items (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    wo_id               INTEGER NOT NULL REFERENCES work_orders(id) ON DELETE CASCADE,
    line_no             INTEGER DEFAULT 1,
    step_name           TEXT NOT NULL,                            -- 절삭 / 연마 / 검수 등
    duration_min        INTEGER DEFAULT 0,                        -- 작업시간 분
    progress            INTEGER DEFAULT 0
                        CHECK(progress BETWEEN 0 AND 100),        -- 진행률 0~100%
    worker_id           INTEGER REFERENCES users(id),             -- 단계 작업자
    worker_name         TEXT,
    remarks             TEXT
);
CREATE INDEX IF NOT EXISTS idx_woi_wo ON work_order_items(wo_id);

-- =====================================================
-- PROJECT GANTT / BURNDOWN (2026-04-26 갭서베이 Top10 #4 — 1차)
-- 마일스톤 + 일별 번다운 스냅샷. idempotent · projects 테이블 미접촉.
-- =====================================================
CREATE TABLE IF NOT EXISTS project_milestones (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id      INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
    name            TEXT NOT NULL,
    target_date     TEXT,
    completed_at    TEXT,
    status          TEXT DEFAULT 'PLANNED'
                    CHECK(status IN ('PLANNED','IN_PROGRESS','DONE','DELAYED','CANCELLED')),
    created_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_pm_project ON project_milestones(project_id);
CREATE INDEX IF NOT EXISTS idx_pm_status ON project_milestones(status);

CREATE TABLE IF NOT EXISTS project_burndown_snapshots (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id      INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
    snap_date       TEXT NOT NULL,
    total_tasks     INTEGER DEFAULT 0,
    completed_tasks INTEGER DEFAULT 0,
    remaining_hours REAL DEFAULT 0,
    created_at      TEXT DEFAULT (datetime('now','localtime')),
    UNIQUE(project_id, snap_date)
);
CREATE INDEX IF NOT EXISTS idx_pbs_project ON project_burndown_snapshots(project_id);
CREATE INDEX IF NOT EXISTS idx_pbs_date ON project_burndown_snapshots(snap_date);

-- 2차 (2026-04-26): 회귀 기반 예측 결과 저장. 외부 라이브러리 0 (순수 Python).
CREATE TABLE IF NOT EXISTS project_forecasts (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id      INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
    computed_at     TEXT DEFAULT (datetime('now','localtime')),
    sample_n        INTEGER DEFAULT 0,
    slope           REAL DEFAULT 0,        -- 일당 잔여작업 감소량 (음수=감소)
    intercept       REAL DEFAULT 0,
    r_squared       REAL DEFAULT 0,
    forecast_date   TEXT,                  -- 추세선이 0에 도달하는 일자 (YYYY-MM-DD)
    planned_end     TEXT,                  -- 비교용 (projects.end_date 스냅샷)
    UNIQUE(project_id)
);
CREATE INDEX IF NOT EXISTS idx_pf_project ON project_forecasts(project_id);

-- =====================================================
-- QMS 강화 1차 (2026-04-26 갭서베이 Top10 #6 — P2 제조/품질팀 주2회)
-- 신규 테이블 3종 (감사로그/시정조치/예방조치) · issues 테이블 5컬럼 ALTER ADD
-- 페르소나: P2 품질팀(주2회 SLA·재발 추적) · P-CEO(품질 KPI 분기)
-- 모든 CREATE 는 idempotent · ALTER ADD 는 init_db() 마이그레이션 블록에서 가드
-- =====================================================
CREATE TABLE IF NOT EXISTS qms_audit_log (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    issue_id    INTEGER NOT NULL REFERENCES issues(id) ON DELETE CASCADE,
    action      TEXT NOT NULL,            -- 'created' / 'severity_changed' / 'sla_breach' / 'corrective_added' / 'preventive_added' / 'closed'
    actor       INTEGER REFERENCES users(id),
    note        TEXT,                     -- 변경 상세 (immutable: UPDATE/DELETE 안 함)
    timestamp   TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_qmsa_issue ON qms_audit_log(issue_id);
CREATE INDEX IF NOT EXISTS idx_qmsa_action ON qms_audit_log(action);

CREATE TABLE IF NOT EXISTS corrective_actions (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    issue_id        INTEGER NOT NULL REFERENCES issues(id) ON DELETE CASCADE,
    action          TEXT NOT NULL,
    responsible     INTEGER REFERENCES users(id),
    due_date        TEXT,
    completed_at    TEXT,
    status          TEXT DEFAULT 'OPEN'
                    CHECK(status IN ('OPEN','IN_PROGRESS','DONE','CANCELLED')),
    created_by      INTEGER REFERENCES users(id),
    created_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_ca_issue ON corrective_actions(issue_id);
CREATE INDEX IF NOT EXISTS idx_ca_status ON corrective_actions(status);

CREATE TABLE IF NOT EXISTS preventive_actions (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    corrective_id   INTEGER NOT NULL REFERENCES corrective_actions(id) ON DELETE CASCADE,
    action          TEXT NOT NULL,
    completed_at    TEXT,
    created_by      INTEGER REFERENCES users(id),
    created_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_pa_ca ON preventive_actions(corrective_id);

-- =====================================================
-- STOCK AUDITS — 재고 실사·조정 (Top10 #10 P4 자재팀 분기 1회) (2026-04-26)
-- 모든 CREATE 는 idempotent · ALTER ADD 는 init_db() 마이그레이션 블록에서 가드
-- =====================================================
CREATE TABLE IF NOT EXISTS stock_audits (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    audit_no        TEXT UNIQUE,                  -- AUD-YYYYMM-####
    start_date      TEXT NOT NULL,
    end_date        TEXT,
    status          TEXT DEFAULT 'OPEN'
                    CHECK(status IN ('OPEN','COUNTING','REVIEW','CLOSED','CANCELLED','DRAFT','IN_PROGRESS','FINALIZED')),
    led_by          INTEGER REFERENCES users(id), -- 자재팀장
    note            TEXT,
    created_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_sa_status ON stock_audits(status);
CREATE INDEX IF NOT EXISTS idx_sa_start ON stock_audits(start_date);

CREATE TABLE IF NOT EXISTS stock_audit_items (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    audit_id        INTEGER NOT NULL REFERENCES stock_audits(id) ON DELETE CASCADE,
    part_id         INTEGER NOT NULL REFERENCES parts(id),
    system_qty      REAL NOT NULL DEFAULT 0,  -- 시스템 수량 스냅샷 (실사 시점)
    counted_qty     REAL,                      -- 실측 수량
    variance        REAL DEFAULT 0,            -- counted - system
    variance_reason TEXT,                      -- 차이 사유 (필수 입력)
    status          TEXT DEFAULT 'PENDING'
                    CHECK(status IN ('PENDING','MATCHED','SHORT','OVER')),
    counted_by      INTEGER REFERENCES users(id),
    counted_at      TEXT,
    created_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_sai_audit ON stock_audit_items(audit_id);
CREATE INDEX IF NOT EXISTS idx_sai_part ON stock_audit_items(part_id);
CREATE INDEX IF NOT EXISTS idx_sai_status ON stock_audit_items(status);

CREATE TABLE IF NOT EXISTS stock_adjustments (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    audit_item_id   INTEGER NOT NULL REFERENCES stock_audit_items(id) ON DELETE CASCADE,
    adjusted_qty    REAL NOT NULL,             -- 조정 수량 (+증가/-감소)
    reason          TEXT NOT NULL,
    status          TEXT DEFAULT 'PENDING'
                    CHECK(status IN ('PENDING','APPROVED','REJECTED')),
    approved_by     INTEGER REFERENCES users(id),
    approved_at     TEXT,
    movement_id     INTEGER REFERENCES stock_movements(id), -- 승인 시 생성된 SM 참조
    note            TEXT,
    created_by      INTEGER REFERENCES users(id),
    created_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_sadj_item ON stock_adjustments(audit_item_id);
CREATE INDEX IF NOT EXISTS idx_sadj_status ON stock_adjustments(status);

CREATE TABLE IF NOT EXISTS audit_attachments (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    adjustment_id   INTEGER NOT NULL REFERENCES stock_adjustments(id) ON DELETE CASCADE,
    file_path       TEXT NOT NULL,
    file_name       TEXT,
    uploaded_by     INTEGER REFERENCES users(id),
    uploaded_at     TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_aatt_adj ON audit_attachments(adjustment_id);

-- =====================================================
-- 자재 첨부 (v5H129 — 2026-05-04): 사진/도면 (클라이언트 압축 후 저장)
-- 외부 저장소 0건 — 로컬 ./uploads/parts/<part_id>/<file>
-- =====================================================
CREATE TABLE IF NOT EXISTS part_attachments (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    part_id         INTEGER NOT NULL REFERENCES parts(id) ON DELETE CASCADE,
    file_path       TEXT NOT NULL,
    file_name       TEXT,
    file_size       INTEGER DEFAULT 0,
    mime_type       TEXT,
    kind            TEXT DEFAULT 'photo',
    uploaded_by     INTEGER REFERENCES users(id),
    uploaded_at     TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_patt_part ON part_attachments(part_id);

-- =====================================================
-- 환율·단가 강화 (2026-04-26 Top10 #9 P4 구매팀 월 1회) — idempotent
-- 외부 환율 API 미사용 (수동 + CSV 업로드만)
-- =====================================================
CREATE TABLE IF NOT EXISTS cost_simulations (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    part_id             INTEGER NOT NULL REFERENCES parts(id),
    simulated_at        TEXT NOT NULL,
    base_currency       TEXT NOT NULL DEFAULT 'USD',
    target_currency     TEXT NOT NULL DEFAULT 'KRW',
    exchange_rate       REAL NOT NULL,
    unit_price_base     REAL NOT NULL,
    unit_price_target   REAL NOT NULL,
    margin_pct          REAL DEFAULT 0,
    note                TEXT,
    created_by          INTEGER REFERENCES users(id),
    created_at          TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_csim_part ON cost_simulations(part_id);
CREATE INDEX IF NOT EXISTS idx_csim_at ON cost_simulations(simulated_at);

CREATE TABLE IF NOT EXISTS price_change_history (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    part_id         INTEGER NOT NULL REFERENCES parts(id),
    supplier_id     INTEGER REFERENCES suppliers(id),
    old_price       REAL,
    new_price       REAL NOT NULL,
    change_pct      REAL,
    effective_date  TEXT NOT NULL,
    note            TEXT,
    changed_by      INTEGER REFERENCES users(id),
    created_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_pch_part ON price_change_history(part_id);
CREATE INDEX IF NOT EXISTS idx_pch_supplier ON price_change_history(supplier_id);
CREATE INDEX IF NOT EXISTS idx_pch_date ON price_change_history(effective_date);

CREATE TABLE IF NOT EXISTS rate_alerts (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    target_currency TEXT NOT NULL,
    threshold       REAL NOT NULL,
    direction       TEXT NOT NULL DEFAULT 'above'
                    CHECK(direction IN ('above','below')),
    is_active       INTEGER DEFAULT 1,
    triggered_at    TEXT,
    created_by      INTEGER REFERENCES users(id),
    created_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_ralert_currency ON rate_alerts(target_currency);
CREATE INDEX IF NOT EXISTS idx_ralert_active ON rate_alerts(is_active);

-- =====================================================
-- WEEKLY SUMMARY VIEW (2026-04-26 갭서베이 Top10 #5 — 주간보고 자동 집계 1차)
-- 자동 집계: tasks 테이블에서 (user_id, week_start) 별 GROUP BY
-- week_start = 해당 주 월요일 (ISO weekday 1=Mon → date - (weekday-1) days)
-- SQLite 의 strftime('%w', date) : 0=Sun..6=Sat 이므로 (weekday+6)%7 로 월요일 보정
-- idempotent (CREATE VIEW IF NOT EXISTS) — 테이블 신설 0건
-- =====================================================
CREATE VIEW IF NOT EXISTS weekly_summary AS
SELECT
    t.user_id                                                       AS user_id,
    u.team_id                                                       AS team_id,
    DATE(t.work_date,
         '-' || ((CAST(strftime('%w', t.work_date) AS INTEGER) + 6) % 7) || ' days'
    )                                                               AS week_start,
    DATE(t.work_date,
         '-' || ((CAST(strftime('%w', t.work_date) AS INTEGER) + 6) % 7) || ' days',
         '+6 days'
    )                                                               AS week_end,
    COUNT(*)                                                        AS total_tasks,
    SUM(CASE WHEN t.status = '완료'  THEN 1 ELSE 0 END)               AS completed,
    SUM(CASE WHEN t.status = '진행중' THEN 1 ELSE 0 END)              AS in_progress,
    SUM(CASE WHEN t.status = '지연'  THEN 1 ELSE 0 END)               AS delayed,
    ROUND(COALESCE(SUM(t.hours), 0), 1)                              AS total_hours
FROM    tasks t
LEFT JOIN users u ON t.user_id = u.id
GROUP BY t.user_id, u.team_id, week_start;
"""

# =====================================================
# ORG CHART SEED DATA (2026.03.11 기준)
# =====================================================
TEAMS = [
    # (code, name, is_lab, sector, order)
    ("01", "기술영업팀", 0, "공통",  1),
    ("02", "검사기팀",   1, "검사기", 2),
    ("03", "품질팀",     0, "검사기", 3),
    ("04", "설계팀",     1, "공통",  4),
    ("05", "소프트웨어팀", 1, "자동화", 5),
    ("06", "전장설계팀", 1, "자동화", 6),
    ("07", "제조기술1팀", 0, "검사기", 7),
    ("08", "제조기술2팀", 0, "자동화", 8),
    ("09", "가공팀",     0, "공통",  9),
    ("10", "구매팀",     0, "공통", 10),
    ("11", "관리팀",     0, "공통", 11),
    ("12", "베트남법인", 0, "공통", 12),
    ("13", "개발혁신팀", 1, "검사기", 13),
    ("14", "라이프밸류팀", 1, "공통", 14),
]

# 사용자 시드: (이름, 팀코드, 직급, 역할)
# role: ceo / executive / leader / member
USERS = [
    # 경영진
    ("김동후", None, "대표이사", "ceo"),
    ("김정락", None, "대표이사", "ceo"),
    ("이한빈", "02", "상무",     "executive"),  # 연구소장 + 검사기팀장
    ("윤경호", "04", "상무",     "executive"),  # 설계팀장
    ("최홍광", None, "전무",     "executive"),  # 인사총괄
    ("최보현", "13", "상무",     "executive"),  # 개발혁신팀장

    # 01 기술영업팀
    ("이해림", "01", "이사",     "leader"),
    ("이현",   "01", "매니저", "member"),
    ("오경환", "01", "프로",   "member"),
    ("배승진", "01", "프로",   "member"),
    ("안지연", "01", "프로",   "member"),
    ("이새롬", "01", "프로",   "member"),

    # 02 검사기팀 (팀장 이한빈 상무 - 이미 등록)
    ("이치권", "02", "이사",   "member"),
    ("이성진", "02", "매니저", "member"),
    ("길희용", "02", "매니저", "member"),
    ("윤광훈", "02", "프로",   "member"),
    ("김태형", "02", "프로",   "member"),
    ("이서준", "02", "사원",   "member"),
    ("김지훈", "02", "사원",   "member"),
    ("지경숙", "02", "반장",   "member"),

    # 03 품질팀
    ("김정록", "03", "매니저", "leader"),
    ("정형진", "03", "사원",   "member"),

    # 04 설계팀 (팀장 윤경호 상무 - 이미 등록)
    ("이영준",  "04", "매니저", "member"),  # 검사기
    ("김범수",  "04", "프로",   "member"),
    ("안호재",  "04", "사원",   "member"),
    ("정민규",  "04", "매니저", "member"),  # 자동화
    ("이상천",  "04", "매니저", "member"),
    ("한재운",  "04", "매니저", "member"),
    ("신광용",  "04", "매니저", "member"),
    ("김선주",  "04", "매니저", "member"),
    ("김동현",  "04", "프로",   "member"),
    ("최현규",  "04", "프로",   "member"),
    ("김원",    "04", "사원",   "member"),

    # 05 소프트웨어팀
    ("이한중", "05", "매니저", "leader"),
    ("최창호", "05", "매니저", "member"),
    ("황정석", "05", "매니저", "member"),
    ("현종필", "05", "매니저", "member"),
    ("이정우", "05", "매니저", "member"),
    ("주진호", "05", "매니저", "member"),
    ("김동욱", "05", "매니저", "member"),
    ("차상권", "05", "매니저", "member"),
    ("이영준2","05", "매니저", "member"),
    ("이충희", "05", "프로",   "member"),
    ("박주창", "05", "프로",   "member"),
    ("김기운", "05", "프로",   "member"),

    # 06 전장설계팀
    ("김형렬", "06", "매니저", "leader"),
    ("박재석", "06", "사원",   "member"),

    # 07 제조기술1팀
    ("노충일", "07", "매니저", "leader"),
    ("마준영", "07", "사원",   "member"),
    ("이태우", "07", "사원",   "member"),
    ("금진호", "07", "사원",   "member"),

    # 08 제조기술2팀
    ("임택훈", "08", "매니저", "leader"),
    ("서재희", "08", "매니저", "member"),
    ("연태흠", "08", "매니저", "member"),
    ("방성기", "08", "프로",   "member"),
    ("김한성", "08", "프로",   "member"),
    ("박지현", "08", "사원",   "member"),
    ("나영훈", "08", "사원",   "member"),
    ("강대성", "08", "사원",   "member"),

    # 09 가공팀
    ("윤영조", "09", "매니저", "leader"),
    ("이수빈", "09", "매니저", "member"),
    ("이청명", "09", "사원",   "member"),

    # 10 구매팀
    ("정성진", "10", "매니저", "leader"),
    ("허동준", "10", "매니저", "member"),
    ("오용균", "10", "프로",   "member"),
    ("김선미", "10", "프로",   "member"),
    ("이홍규", "10", "프로",   "member"),
    ("박성준", "10", "사원",   "member"),
    ("란",     "10", "사원",   "member"),

    # 11 관리팀
    ("박지은", "11", "매니저", "leader"),
    ("엄혜린", "11", "매니저", "member"),
    ("엄주영", "11", "프로",   "member"),
    ("최혜연", "11", "프로",   "member"),

    # 12 베트남법인
    ("이용식", "12", "법인장", "leader"),
    ("땀",     "12", "부장",   "member"),
    ("박지만", "12", "부장",   "member"),
    ("탕",     "12", "차장",   "member"),
    ("쑤아잉", "12", "차장",   "member"),

    # 13 개발혁신팀 (팀장 최보현 상무 - 이미 등록)
    ("박승환", "13", "매니저", "member"),
    ("김수현", "13", "사원",   "member"),

    # 14 라이프밸류팀
    ("나재겸", "14", "매니저", "leader"),
    ("김기선", "14", "매니저", "member"),
    ("박성수", "14", "사원",   "member"),
]

# 주요 고객사
CUSTOMERS = [
    ("고객사A",    "주요", "1차 협력사"),
    ("고객사B",    "주요", "1차 협력사"),
    ("드림텍",     "주요", ""),
    ("한국성전",   "주요", ""),
    ("기타전장",   "일반", "전장사업 관련"),
    ("내부",       "일반", "사내 프로젝트"),
]


# =====================================================
# LOGIN ID 생성 규칙
# =====================================================
def make_login_id(name: str, dup_map: dict) -> str:
    base = name.lower().replace(" ", "")
    # 한글은 그대로 초성처리 대신 name 그대로 사용하되, 영문/숫자가 아니면 중복 방지만
    if base in dup_map:
        dup_map[base] += 1
        return f"{base}{dup_map[base]}"
    dup_map[base] = 1
    return base


# =====================================================
# v5H226z619 (대표 지시): 고객사명 UNIQUE 완화 — 같은 상호라도 종사업장번호가 다르면
#   별개 사업장으로 등록 허용(예: (주)드림텍 본사 0001 / 아산밸리 0005).
#   DB 의 customers.name UNIQUE 제약을 제거(중복검사는 앱 레벨에서 상호+종사업장으로).
#   ⚠ 프로덕션 스키마 변경 — 전용 autocommit 연결로 FK OFF 후, 현재 스키마 원문에서
#     UNIQUE 만 제거해 재구성. DB 파일 백업 + 행수 검증 + 검증 통과 후에만 교체(데이터 보존).
#   idempotent: name UNIQUE 가 이미 없으면 즉시 반환. init_db() 완료 후 1회 호출.
# =====================================================
def migrate_customers_name_unique():
    import sqlite3 as _sq, re as _re_u
    conn = _sq.connect(DB_PATH)
    conn.isolation_level = None   # autocommit — 트랜잭션 밖이라 PRAGMA foreign_keys 적용 가능
    try:
        row = conn.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='customers'").fetchone()
        csql = (row[0] if row else "") or ""
        if not _re_u.search(r'\bname\b\s+TEXT\s+UNIQUE', csql, _re_u.IGNORECASE):
            return  # 이미 완화됨
        # 1) DB 파일 백업(되돌림 대비)
        try:
            import shutil as _sh, datetime as _dtu
            bdir = os.path.join(os.path.dirname(DB_PATH), "backups")
            os.makedirs(bdir, exist_ok=True)
            _sh.copy2(DB_PATH, os.path.join(bdir, f"knk.db.bak_custuniq_{_dtu.datetime.now().strftime('%Y%m%d_%H%M%S')}"))
        except Exception:
            pass
        n_before = conn.execute("SELECT COUNT(*) FROM customers").fetchone()[0]
        # 2) 새 스키마 = 현재 원문에서 name 의 UNIQUE 만 제거 + 테이블명 변경
        newsql = _re_u.sub(r'(\bname\b\s+TEXT\s+)UNIQUE\s+', r'\1', csql, count=1, flags=_re_u.IGNORECASE)
        newsql = _re_u.sub(r'CREATE\s+TABLE\s+(IF\s+NOT\s+EXISTS\s+)?["\[`]?customers["\]`]?\s*\(',
                           'CREATE TABLE "customers_new" (', newsql, count=1, flags=_re_u.IGNORECASE)
        colnames = [r[1] for r in conn.execute("PRAGMA table_info(customers)").fetchall()]
        csv = ", ".join(f'"{x}"' for x in colnames)
        conn.execute("PRAGMA foreign_keys=OFF")
        conn.execute("DROP TABLE IF EXISTS customers_new")
        conn.execute(newsql)
        conn.execute(f"INSERT INTO customers_new ({csv}) SELECT {csv} FROM customers")
        n_after = conn.execute("SELECT COUNT(*) FROM customers_new").fetchone()[0]
        # 3) 행수 검증 — 다르면 중단(원본 보존)
        if n_after != n_before:
            conn.execute("DROP TABLE IF EXISTS customers_new")
            print(f"[z619] customers UNIQUE 완화 중단 — 행수 불일치 {n_before}->{n_after}(원본 보존)")
            return
        conn.execute("DROP TABLE customers")
        conn.execute("ALTER TABLE customers_new RENAME TO customers")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_customers_name ON customers(name)")
        print(f"[z619] customers.name UNIQUE 제거 완료 — 같은 상호 다른 종사업장 등록 허용 ({n_after}건 보존)")
    except Exception as e:
        try:
            conn.execute("DROP TABLE IF EXISTS customers_new")
        except Exception:
            pass
        print(f"[z619] customers UNIQUE 완화 스킵: {e}")
    finally:
        conn.close()


# =====================================================
# INIT + SEED
# =====================================================
def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    with db_session() as c:
        c.executescript(SCHEMA)
        # 마이그레이션: 기존 DB에 next_plan 컬럼이 없으면 추가
        cols = [r[1] for r in c.execute("PRAGMA table_info(tasks)").fetchall()]
        if "next_plan" not in cols:
            try:
                c.execute("ALTER TABLE tasks ADD COLUMN next_plan TEXT")
            except Exception:
                pass
        # v5H28 (2026-05-02): 자유 입력 프로젝트 라벨 (project_id 와 별개)
        # 사용자가 매출 프로젝트 외 사내·일반·개인 업무를 자유 텍스트로 등록 가능
        if "project_label" not in cols:
            try:
                c.execute("ALTER TABLE tasks ADD COLUMN project_label TEXT")
            except Exception:
                pass
        # v5H29 (2026-05-02): 자유 입력 고객사 라벨 ('사내업무' 등)
        if "customer_label" not in cols:
            try:
                c.execute("ALTER TABLE tasks ADD COLUMN customer_label TEXT")
            except Exception:
                pass
        # v5H52 (2026-05-03): 고객사 등록 폼 확장 — 사업자번호/대표/담당자/전화/이메일/주소
        # v5H68 (2026-05-03): orders.project_id 컬럼 추가 — 수주↔프로젝트 연결 (KNK 워크플로우)
        ocols = [r[1] for r in c.execute("PRAGMA table_info(orders)").fetchall()]
        if "project_id" not in ocols:
            try:
                c.execute("ALTER TABLE orders ADD COLUMN project_id INTEGER REFERENCES projects(id)")
                c.execute("CREATE INDEX IF NOT EXISTS idx_orders_project ON orders(project_id)")
            except Exception:
                pass
        # v5H226z678 (대표 지시): 수주(SO)별 담당자 — 같은 관리번호에 담당자가 다른 여러 주문(추가제작)을 구분.
        #   기존엔 담당자가 projects(프로젝트) 단위라 같은 관리번호 여러 수주가 화면에서 똑같아 보였음.
        for _ocol, _oddl in [
            ("cc_name",  "ALTER TABLE orders ADD COLUMN cc_name TEXT"),
            ("cc_dept",  "ALTER TABLE orders ADD COLUMN cc_dept TEXT"),
            ("cc_phone", "ALTER TABLE orders ADD COLUMN cc_phone TEXT"),
        ]:
            if _ocol not in ocols:
                try:
                    c.execute(_oddl)
                except Exception:
                    pass

        # v5H58 (2026-05-03): 등급 자동 산정 — tier_score / tier_computed_at
        cucols = [r[1] for r in c.execute("PRAGMA table_info(customers)").fetchall()]
        for col, ddl in [
            ("biz_no",          "ALTER TABLE customers ADD COLUMN biz_no TEXT"),
            ("ceo_name",        "ALTER TABLE customers ADD COLUMN ceo_name TEXT"),
            ("manager_name",    "ALTER TABLE customers ADD COLUMN manager_name TEXT"),
            ("phone",           "ALTER TABLE customers ADD COLUMN phone TEXT"),
            ("email",           "ALTER TABLE customers ADD COLUMN email TEXT"),
            ("address",         "ALTER TABLE customers ADD COLUMN address TEXT"),
            ("is_active",       "ALTER TABLE customers ADD COLUMN is_active INTEGER DEFAULT 1"),
            ("created_at",      "ALTER TABLE customers ADD COLUMN created_at TEXT DEFAULT (datetime('now','localtime'))"),
            ("tier_score",      "ALTER TABLE customers ADD COLUMN tier_score INTEGER DEFAULT 0"),
            ("tier_computed_at","ALTER TABLE customers ADD COLUMN tier_computed_at TEXT"),
            ("tier_breakdown",  "ALTER TABLE customers ADD COLUMN tier_breakdown TEXT"),
            # v5H226z147 (2026-06-02 대표 지시): 기존 시스템 거래처 항목 반영
            ("code",            "ALTER TABLE customers ADD COLUMN code TEXT"),            # 거래처코드(자동 C-0001)
            ("business_type",   "ALTER TABLE customers ADD COLUMN business_type TEXT"),  # 업태
            ("business_item",   "ALTER TABLE customers ADD COLUMN business_item TEXT"),  # 업종(종목)
            ("zipcode",         "ALTER TABLE customers ADD COLUMN zipcode TEXT"),        # 우편번호
            ("address_detail",  "ALTER TABLE customers ADD COLUMN address_detail TEXT"), # 상세주소
            ("fax",             "ALTER TABLE customers ADD COLUMN fax TEXT"),            # 팩스번호
            ("sub_biz_no",      "ALTER TABLE customers ADD COLUMN sub_biz_no TEXT"),     # 종사업장번호
            # v5H226z203 (2026-06-04 대표 지시): 해외 고객사 — 사업자번호 없는 해외 거래처 지원
            ("is_overseas",     "ALTER TABLE customers ADD COLUMN is_overseas INTEGER DEFAULT 0"),  # 1=해외
            ("country",         "ALTER TABLE customers ADD COLUMN country TEXT"),        # 국가(해외 시)
            # v5H226z744 (대표 지시): 결제조건 — 세금계산서 발행일 + N일 = 입금 예상일. 0/빈칸=미설정.
            ("pay_days",        "ALTER TABLE customers ADD COLUMN pay_days INTEGER DEFAULT 0"),  # 세금계산서 발행 후 N일 결제
        ]:
            if col not in cucols:
                try:
                    c.execute(ddl)
                except Exception:
                    pass
        # v5H226z580 (대표 지시 2026-06-22): 연락처 공유 범위 — 기본 개인(private)·부서(dept)·전사(company).
        #   기존 연락처는 전부 전사 공유였으므로 'company' 로 백필(기존 가시성 보존). 신규는 DEFAULT 'private'.
        try:
            sccols = [r[1] for r in c.execute("PRAGMA table_info(shared_contacts)").fetchall()]
            if sccols:
                if "visibility" not in sccols:
                    c.execute("ALTER TABLE shared_contacts ADD COLUMN visibility TEXT DEFAULT 'private'")
                    # ADD COLUMN 으로 기존 행이 'private' 가 되므로, 기존 전사공유 상태 보존 위해 company 로 백필
                    c.execute("UPDATE shared_contacts SET visibility='company'")
                if "share_team_id" not in sccols:
                    c.execute("ALTER TABLE shared_contacts ADD COLUMN share_team_id INTEGER")
        except Exception:
            pass
        # z407 (2026-06-13 대표 지시): 동시 편집 감지(낙관적 잠금)용 updated_at 보강.
        #   수주(orders)·호기(order_items)·소모품수주(consumable_orders)에 updated_at 추가 +
        #   기존 행은 created_at(있으면)으로 백필(없으면 now) → 잠금 기준(baseline) 확보. (additive·안전)
        for _tbl in ("orders", "order_items", "consumable_orders"):
            try:
                _tcols = [r[1] for r in c.execute(f"PRAGMA table_info({_tbl})").fetchall()]
                if not _tcols:
                    continue   # 테이블 없음 → 건너뜀
                if "updated_at" not in _tcols:
                    c.execute(f"ALTER TABLE {_tbl} ADD COLUMN updated_at TEXT")
                    if "created_at" in _tcols:
                        c.execute(f"UPDATE {_tbl} SET updated_at=created_at WHERE updated_at IS NULL")
                    else:
                        c.execute(f"UPDATE {_tbl} SET updated_at=datetime('now','localtime') WHERE updated_at IS NULL")
            except Exception:
                pass
        # z409: tasks·projects 는 updated_at 컬럼이 이미 있으나(위 루프 대상 아님) 기존 행이 NULL 일 수 있음
        #   → NULL 백필해 낙관적 잠금 baseline 확보(없으면 base_ts 빈값=보호 안 됨).
        for _bt in ("tasks", "projects"):
            try:
                c.execute(f"UPDATE {_bt} SET updated_at=COALESCE(updated_at, created_at, datetime('now','localtime')) WHERE updated_at IS NULL")
            except Exception:
                pass
        # z393 (2026-06-13 대표 지시): 메일 첨부 본문 저장 + 인라인 이미지(서명 로고) 표시
        try:
            macols = [r[1] for r in c.execute("PRAGMA table_info(mail_attachments)").fetchall()]
            for col, ddl in [
                ("content_id", "ALTER TABLE mail_attachments ADD COLUMN content_id TEXT"),
                ("is_inline",  "ALTER TABLE mail_attachments ADD COLUMN is_inline INTEGER DEFAULT 0"),
                ("data",       "ALTER TABLE mail_attachments ADD COLUMN data BLOB"),
            ]:
                if col not in macols:
                    try:
                        c.execute(ddl)
                    except Exception:
                        pass
        except Exception:
            pass

        # 메일 숨은참조(bcc) 컬럼 — 보낸/임시저장 보관용 (2026-06-13 아웃룩 편의기능)
        try:
            mmcols = [r[1] for r in c.execute("PRAGMA table_info(mail_messages)").fetchall()]
            if "bcc" not in mmcols:
                c.execute("ALTER TABLE mail_messages ADD COLUMN bcc TEXT")
        except Exception:
            pass

        # 마이그레이션: users에 lang 컬럼 추가
        ucols = [r[1] for r in c.execute("PRAGMA table_info(users)").fetchall()]
        if "lang" not in ucols:
            try:
                c.execute("ALTER TABLE users ADD COLUMN lang TEXT DEFAULT 'ko'")
            except Exception:
                pass
        # 마이그레이션: 물류 모듈 접근 권한 (HAIST WORKS)
        if "can_use_logistics" not in ucols:
            try:
                c.execute("ALTER TABLE users ADD COLUMN can_use_logistics INTEGER DEFAULT 0")
            except Exception:
                pass
        # 마이그레이션 (Plan Y S1): 관리자 권한 + 매출/구매 분리 권한
        # 2026-04-28 추가: 매출/자재 읽기 전용 권한 (대표 결재 — 등록은 못해도 조회는 폭넓게)
        for col, ddl in [
            ("is_admin",          "ALTER TABLE users ADD COLUMN is_admin INTEGER DEFAULT 0"),
            ("can_use_sales",     "ALTER TABLE users ADD COLUMN can_use_sales INTEGER DEFAULT 0"),
            ("can_edit_changes",  "ALTER TABLE users ADD COLUMN can_edit_changes INTEGER DEFAULT 0"),
            ("can_close_tickets", "ALTER TABLE users ADD COLUMN can_close_tickets INTEGER DEFAULT 0"),
            ("can_view_sales",     "ALTER TABLE users ADD COLUMN can_view_sales INTEGER DEFAULT 0"),
            ("can_view_logistics", "ALTER TABLE users ADD COLUMN can_view_logistics INTEGER DEFAULT 0"),
            # z746(대표 지시): 제작요청 통보 자동 라우팅용 담당 사업부(CSV: T,M,L,E). 비어있으면 '공통'(전 사업부 통보).
            ("biz_divs",           "ALTER TABLE users ADD COLUMN biz_divs TEXT DEFAULT ''"),
        ]:
            if col not in ucols:
                try:
                    c.execute(ddl)
                except Exception:
                    pass
        # Plan Y S1 — 권한 시드 (idempotent, 매 init 실행 / seed 게이트 외부)
        # CEO + executive 전권 자동 부여 (PS-CEO/PS-임원 페르소나 충족)
        try:
            c.execute(
                "UPDATE users SET can_use_logistics=1, can_use_sales=1, "
                "can_view_sales=1, can_view_logistics=1, "
                "is_admin=1, can_edit_changes=1, can_close_tickets=1 "
                "WHERE role IN ('ceo','executive') AND is_active=1"
            )
            # leader: 본인 영역 변경/티켓 처리 권한 + 매출/자재 보기 기본 ON
            c.execute(
                "UPDATE users SET can_edit_changes=1, can_close_tickets=1, "
                "can_view_sales=1, can_view_logistics=1 "
                "WHERE role='leader' AND is_active=1"
            )
            # 2026-04-28 자재 보기 시드 (실무자 폭넓게):
            #   - 영업팀(1), 검사기(2), 품질(3), 생산팀(7,8), 가공(9), 구매(10) 평직원 전원 자동
            #     (재고·부품·단가·구매사 조회 필요)
            c.execute(
                "UPDATE users SET can_view_logistics=1 "
                "WHERE role='member' AND is_active=1 AND team_id IN (1,2,3,7,8,9,10)"
            )
            # 매출 보기 시드:
            #   - 영업(1), 검사기(2), 품질(3) 평직원 (현재 폴백과 동일 범위)
            c.execute(
                "UPDATE users SET can_view_sales=1 "
                "WHERE role='member' AND is_active=1 AND team_id IN (1,2,3)"
            )
            # 쓰기 권한자는 보기도 자동 부여 (write implies read)
            c.execute(
                "UPDATE users SET can_view_sales=1 WHERE can_use_sales=1 AND is_active=1"
            )
            c.execute(
                "UPDATE users SET can_view_logistics=1 WHERE can_use_logistics=1 AND is_active=1"
            )
        except Exception:
            pass
        # 마이그레이션: changes에 source 컬럼 (Abram Scientific 모델 — 외부 도구 출처)
        try:
            ccols = [r[1] for r in c.execute("PRAGMA table_info(changes)").fetchall()]
            if ccols and "source" not in ccols:
                c.execute("ALTER TABLE changes ADD COLUMN source TEXT DEFAULT '수동'")
            if ccols and "source_ref" not in ccols:
                c.execute("ALTER TABLE changes ADD COLUMN source_ref TEXT")
            # B: 외부 그룹웨어 결재 URL 첨부 (자체 결재 X, 외부 결재 링크만)
            if ccols and "approval_url" not in ccols:
                c.execute("ALTER TABLE changes ADD COLUMN approval_url TEXT")
        except Exception:
            pass
        # 마이그레이션: tickets에 approval_url 추가
        try:
            tkcols = [r[1] for r in c.execute("PRAGMA table_info(tickets)").fetchall()]
            if tkcols and "approval_url" not in tkcols:
                c.execute("ALTER TABLE tickets ADD COLUMN approval_url TEXT")
        except Exception:
            pass
        # 마이그레이션: parts에 재고 컬럼 추가 (2026-04-20 — 입출고 기능)
        try:
            prcols = [r[1] for r in c.execute("PRAGMA table_info(parts)").fetchall()]
            if prcols and "stock_qty" not in prcols:
                c.execute("ALTER TABLE parts ADD COLUMN stock_qty REAL DEFAULT 0")
            if prcols and "safety_stock" not in prcols:
                c.execute("ALTER TABLE parts ADD COLUMN safety_stock REAL DEFAULT 0")
            if prcols and "location" not in prcols:
                c.execute("ALTER TABLE parts ADD COLUMN location TEXT")
            # 사이클 51 S2-4차 (2026-04-27): 안전재고 재발주점·권장 발주량
            if prcols and "reorder_point" not in prcols:
                c.execute("ALTER TABLE parts ADD COLUMN reorder_point REAL DEFAULT 0")
            if prcols and "reorder_qty" not in prcols:
                c.execute("ALTER TABLE parts ADD COLUMN reorder_qty REAL DEFAULT 0")
            # v5H226z83/z84 (2026-05-14): 매입단가 — 구매사 3곳 × 각 1~5차(날짜+단가) JSON
            # 형식: [{"supplier":"공원전기(주)","entries":[{"seq":1,"date":"2026-05-01","price":52750}, ...]}, ...]
            # 대표단가(std_price) = 전체 중 가장 최근 날짜의 단가
            if prcols and "purchase_prices" not in prcols:
                c.execute("ALTER TABLE parts ADD COLUMN purchase_prices TEXT")
            # v5H226z99 (2026-05-16): 자재 용도(purpose) — 카탈로그 검색·식별 보조용 자유 텍스트
            if prcols and "purpose" not in prcols:
                c.execute("ALTER TABLE parts ADD COLUMN purpose TEXT")
            # v5H226z110 (2026-05-18): 제조사 모델 번호 — part_no(내부 코드)와 별도 보관
            #   예: 제조사 모델 코드(SVM-200A 등). 원제조사 스펙 검색·호환성 확인용
            if prcols and "model_name" not in prcols:
                c.execute("ALTER TABLE parts ADD COLUMN model_name TEXT")
            # v5H226z110b (2026-05-19): 기본 공급사명 — 자재마다 주로 구매하는 공급사 이름
            #   발주 자동 매칭에 활용. suppliers.name 과 매칭 시도, 자유 텍스트도 허용
            if prcols and "default_supplier" not in prcols:
                c.execute("ALTER TABLE parts ADD COLUMN default_supplier TEXT")
            # v5H226z110c (2026-05-19): 제조사 담당자 풀세트 — 자재 단위로 관리 (제조사 마스터 미구축)
            if prcols and "maker_contact_name" not in prcols:
                c.execute("ALTER TABLE parts ADD COLUMN maker_contact_name TEXT")
            if prcols and "maker_contact_phone" not in prcols:
                c.execute("ALTER TABLE parts ADD COLUMN maker_contact_phone TEXT")
            if prcols and "maker_contact_email" not in prcols:
                c.execute("ALTER TABLE parts ADD COLUMN maker_contact_email TEXT")
            # v5H226z111 (2026-05-19): 구매사(협력사) 3개 슬롯 — 분류·재고 관리 박스로 이전
            #   기존 default_supplier 는 deprecated. vendor1·2·3 가 협력사 마스터에서 선택된 값
            if prcols and "vendor1" not in prcols:
                c.execute("ALTER TABLE parts ADD COLUMN vendor1 TEXT")
            if prcols and "vendor2" not in prcols:
                c.execute("ALTER TABLE parts ADD COLUMN vendor2 TEXT")
            if prcols and "vendor3" not in prcols:
                c.execute("ALTER TABLE parts ADD COLUMN vendor3 TEXT")
        except Exception:
            pass
        # v5H226z89 (2026-05-14): suppliers 에 사업자등록번호·대표자 컬럼 추가
        #   (customers 테이블과 동일 명명 — biz_no / ceo_name)
        # v5H226z90 (2026-05-14): 관리부서(manage_dept) + 담당자2 풀세트(contact2/phone2/email2)
        try:
            spcols = [r[1] for r in c.execute("PRAGMA table_info(suppliers)").fetchall()]
            if spcols and "biz_no" not in spcols:
                c.execute("ALTER TABLE suppliers ADD COLUMN biz_no TEXT")
            if spcols and "ceo_name" not in spcols:
                c.execute("ALTER TABLE suppliers ADD COLUMN ceo_name TEXT")
            if spcols and "manage_dept" not in spcols:
                c.execute("ALTER TABLE suppliers ADD COLUMN manage_dept TEXT")
            if spcols and "contact2" not in spcols:
                c.execute("ALTER TABLE suppliers ADD COLUMN contact2 TEXT")
            if spcols and "phone2" not in spcols:
                c.execute("ALTER TABLE suppliers ADD COLUMN phone2 TEXT")
            if spcols and "email2" not in spcols:
                c.execute("ALTER TABLE suppliers ADD COLUMN email2 TEXT")
            # v5H226z117 (2026-05-19) — 협력사 주소
            if spcols and "address" not in spcols:
                c.execute("ALTER TABLE suppliers ADD COLUMN address TEXT")
        except Exception:
            pass
        # 마이그레이션: stock_movements에 FIFO/lot 컬럼 (2026-04-21 리서치 반영)
        try:
            smcols = [r[1] for r in c.execute("PRAGMA table_info(stock_movements)").fetchall()]
            if smcols and "remaining_qty" not in smcols:
                c.execute("ALTER TABLE stock_movements ADD COLUMN remaining_qty REAL DEFAULT 0")
                # 기존 IN 행에 대해 remaining_qty = quantity (완전소비 가정 후 OUT 차감은 스킵)
                # → 처음 배포 시점에는 IN 전량이 아직 있다고 가정 (안전)
                c.execute(
                    "UPDATE stock_movements SET remaining_qty = quantity WHERE kind='IN'"
                )
            if smcols and "lot_no" not in smcols:
                c.execute("ALTER TABLE stock_movements ADD COLUMN lot_no TEXT")
            if smcols and "expiry_date" not in smcols:
                c.execute("ALTER TABLE stock_movements ADD COLUMN expiry_date TEXT")
        except Exception:
            pass
        # 마이그레이션 (2026-04-25 Top3-S3-2차): permissions RBAC 컬럼 분리
        # 시안 §7 원안 — name(UNIQUE) → resource/action/scope/description 분리
        # 옵션 B 채택: ADD COLUMN으로 신규 컬럼 추가 + name(deprecated)/scope 유지
        try:
            pcols = [r[1] for r in c.execute("PRAGMA table_info(permissions)").fetchall()]
            if pcols and "resource" not in pcols:
                c.execute("ALTER TABLE permissions ADD COLUMN resource TEXT")
            if pcols and "action" not in pcols:
                c.execute("ALTER TABLE permissions ADD COLUMN action TEXT")
            if pcols and "description" not in pcols:
                c.execute("ALTER TABLE permissions ADD COLUMN description TEXT")
            # UNIQUE(resource, action, scope) — 부분 인덱스 (NULL 행 제외)
            c.execute(
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_perm_ras "
                "ON permissions(resource, action, scope) "
                "WHERE resource IS NOT NULL AND action IS NOT NULL"
            )
        except Exception:
            pass
        # 시드: app_settings 기본값 (전자결재·메일 외부 연동)
        try:
            # v5H226z114 (2026-05-31): 익명화 후속 4건 완전 종료 — hiworks_* → groupware_*
            # - 하드코딩 URL 기본값 제거 (admin 페이지에서 입력 강제)
            # - notify_channel 옵션값 "hiworks" → "groupware" (구값은 코드에서 자동 호환)
            for k, v, desc in [
                # 외부 링크 (사이드바) — admin/groupware-settings 에서 입력
                ("groupware_approval_url", "", "전자결재 URL (사이드바 외부 링크)"),
                ("groupware_mail_url",     "", "메일 URL (사이드바 외부 링크)"),
                ("groupware_domain",       "", "회사 그룹웨어 도메인 (예: knk.co.kr)"),
                # 인사총무 외부 그룹웨어 연동 카드 — 외부 자산 0건, 외부 링크만
                ("groupware_main_url",        "", "그룹웨어 메인 (인사총무 카드 — 외부 링크)"),
                ("groupware_attendance_url",  "", "출퇴근/근태 URL (외부 링크)"),
                ("groupware_leave_url",       "", "휴가 신청 URL (외부 링크)"),
                ("groupware_payroll_url",     "", "급여 명세서 URL (외부 링크)"),
                ("groupware_profile_url",     "", "인사 정보 URL (외부 링크)"),
                # 그룹웨어 API 토큰 (관리자 페이지 > 환경설정 > API 관리에서 발급)
                ("groupware_messenger_token", "", "메신저 알림 API 토큰 — 변경/이슈 푸시용"),
                ("groupware_hr_token",        "", "인사관리 API 토큰 — 근태/조직 조회용 (선택)"),
                ("groupware_approval_token",  "", "전자결재 API 토큰 — 자동 기안용 (Phase 2, 선택)"),
                ("groupware_api_base",        "", "외부 그룹웨어 API base URL (예: https://api.your-groupware.com)"),
                # 알림 채널 마스터 스위치
                ("notify_channel",       "off",     "알림 채널: off (기본 — 토큰 미설정 시 안전) / groupware (메신저) / smtp (메일)"),
            ]:
                c.execute("INSERT OR IGNORE INTO app_settings(key, value, description) VALUES(?,?,?)", (k, v, desc))
            # 2026-04-22 대표 결재(D01-06): 외부 웹훅 연동 폐기 — 레거시 키 정리
            for _legacy_key in ('legacy_chat_webhook', 'kakaowork_webhook'):
                c.execute("DELETE FROM app_settings WHERE key=?", (_legacy_key,))
        except Exception:
            pass
        # 게시판 시드: 전사 게시판 + 팀별 게시판 자동 생성
        board_cnt = c.execute("SELECT COUNT(*) FROM boards").fetchone()[0]
        if board_cnt == 0:
            c.execute("INSERT OR IGNORE INTO boards (name, type, team_id) VALUES ('전사 게시판', 'company', NULL)")
            teams_rows = c.execute("SELECT id, name FROM teams ORDER BY display_order").fetchall()
            for t in teams_rows:
                c.execute("INSERT OR IGNORE INTO boards (name, type, team_id) VALUES (?, 'team', ?)",
                          (f"{t['name']} 게시판", t["id"]))
        # 마이그레이션: projects에 물류 모듈 컬럼 추가 (HAIST WORKS)
        pcols = [r[1] for r in c.execute("PRAGMA table_info(projects)").fetchall()]
        _logi_adds = [
            ("biz_div",       "TEXT"),                  # T 검사기 / M 자동화
            ("stage",         "TEXT DEFAULT '제안작성'"),
            ("po_type",       "TEXT DEFAULT '신규'"),
            ("customer_name", "TEXT"),                   # 텍스트 고객사 (customer_id FK와 병행)
            ("customer_po",   "TEXT"),
            ("currency",      "TEXT DEFAULT 'KRW'"),
            ("order_amount",  "REAL DEFAULT 0"),
            ("order_date",    "TEXT"),
            ("due_date",      "TEXT"),
            ("pm_name",       "TEXT"),                   # 텍스트 PM (pm_id FK와 병행)
            ("sales_name",    "TEXT"),
            ("logi_note",     "TEXT"),                   # 물류 비고
            ("updated_at",    "TEXT"),
            ("is_export",     "INTEGER DEFAULT 0"),       # v5H97: 0=내수, 1=수출
            ("unit_qty",      "INTEGER DEFAULT 1"),       # v5H132: 등록 시 호기 수량
            ("unit_price",    "REAL"),                     # v5H132: 1대 단가 (NULL → order_amount 폴백)
            # v5H137 (2026-05-05): 프로젝트 유형 분류 + 부모 프로젝트 연결 (대표 직접 요청)
            #   project_type: NEW_EQUIP(기본)/CONSUMABLE/SERVICE/OTHER
            #   parent_project_id: 소모품·수리 시 어느 장비(관리번호) 의 건인지 연결 (NULL 허용)
            ("project_type",       "TEXT DEFAULT 'NEW_EQUIP'"),
            ("parent_project_id",  "INTEGER"),
            # v5H154 (2026-05-05): 외화 수주 시 기준환율 + 원화 환산 보존 (대표 지시)
            ("fx_rate",            "REAL"),
            ("amount_krw",         "REAL"),
            # v5H201 (2026-05-07): 제안 단계 일정 — 수주확정 전 스케줄 관리용
            ("proposal_date",      "TEXT"),       # 제안서 (예정/실제) 일정. NULL=미정·해당없음
            ("quotation_date",     "TEXT"),       # 견적서 일정. NULL=미정·해당없음
            # v5H212 (2026-05-08): 수주 전 내역 — 제출 여부 + 자유 메모
            ("proposal_submitted", "INTEGER DEFAULT 0"),  # 0=미제출, 1=제출완료
            ("proposal_memo",      "TEXT"),                # 진행 메모 (자유 기록)
            ("quotation_submitted","INTEGER DEFAULT 0"),
            ("quotation_memo",     "TEXT"),
            # v5H226z (2026-05-08): 출고 형태 — ASSEMBLY(완제품, 호기)/PARTS(부품, PACKING LIST)
            ("shipment_form",      "TEXT DEFAULT 'ASSEMBLY'"),
            # v5H226z161 (2026-06-02 대표 지시): 수주 전 '예상 매출' 별도 보존.
            #   order_amount(확정, SO 합계로 자가치유)와 분리 → 예상 vs 확정 비교용. 자가치유가 건드리지 않음.
            ("expected_amount",    "REAL"),
            # v5H226z166 (2026-06-02 대표 지시): 고객사 담당자(프로젝트별) — 고객사 주소록(customer_contacts)과 양방향 연동
            ("cc_name",       "TEXT"),   # 담당자명
            ("cc_dept",       "TEXT"),   # 부서
            ("cc_position",   "TEXT"),   # 직책
            ("cc_phone",      "TEXT"),   # 연락처
            ("cc_email",      "TEXT"),   # 메일
            ("cc_note",       "TEXT"),   # 성향(메모)
            # v5H226z195 (2026-06-04 대표 지시): 2단계 발주(법인 경유) — 최종 고객 매출 참고관리.
            #   본사 매출 = 1차고객사(베트남법인) 발주액(=order_amount, 합계 포함).
            #   2차고객사(최종, 삼성 등) + 최종 고객 매출(참고, 합계 제외) 별도 보존.
            ("two_tier_order",     "INTEGER DEFAULT 0"),  # 1=2단계 발주(법인 경유)
            ("secondary_customer", "TEXT"),               # 2차 고객사(최종 고객)
            ("final_amount",       "REAL"),               # 최종 고객 매출(참고 — 합계 제외)
            # v5H226z211 (2026-06-05 대표 지시): 장비명 — 관리번호=장비 단위.
            #   같은 프로젝트명이라도 장비명이 다르면 관리번호가 다를 수 있음(프로젝트명 비유일).
            #   model_name=장비 모델명. 부품/기타는 이 장비에 딸린 납품물(project_items).
            ("equip_name",         "TEXT"),
            # v5H226z349 (대표 지시): 형태(제품/상품/기타) — 회계 성격 분류. NULL=미지정.
            ("form_type",          "TEXT"),
            # v5H226z375 (대표 지시): 각인 — 장비에 새기는 마킹 문구(제작요청서 항목). 예: 'S3908 ASSY'
            ("engraving",          "TEXT"),
        ]
        for col, decl in _logi_adds:
            if col not in pcols:
                try:
                    c.execute(f"ALTER TABLE projects ADD COLUMN {col} {decl}")
                except Exception:
                    pass
        # v5H226z349 (대표 지시): PO유형 'A/S' → '수리' 이름변경에 따른 기존 데이터 일괄 변환(멱등).
        try:
            c.execute("UPDATE projects SET po_type='수리' WHERE po_type='A/S'")
        except Exception:
            pass
        # v5H137: parent_project_id 인덱스 (소모품 → 부모 장비 역조회 가속)
        try:
            c.execute("CREATE INDEX IF NOT EXISTS idx_projects_parent ON projects(parent_project_id)")
        except Exception:
            pass

        # v5H226z208 (2026-06-04 대표 지시): project_items 백필 — 새 규칙(관리번호=프로젝트, 품목 N개).
        #   기존 데이터를 품목으로 무손실 표현. 이미 품목이 있는 프로젝트는 건너뜀(멱등).
        #   - 수주(SO) 있는 프로젝트: SO 별로 품목 1건(수량/단가/발주처/출고형태 매핑) + order_id 연결
        #   - 제안단계(수주 없음): 대표 모델로 품목 1건
        try:
            _ocols = {r2[1] for r2 in c.execute("PRAGMA table_info(orders)").fetchall()}
            _has_sotype = "so_type" in _ocols
            _has_ocur   = "currency" in _ocols
            _has_oship  = "ship_to" in _ocols
            _has_olabel = "unit_label" in _ocols
            _has_oqty   = "unit_qty" in _ocols
            def _sf_from_so_type(st, proj_sf):
                st = (st or "").upper()
                if st == "EQUIPMENT": return "ASSEMBLY"
                if st in ("CONSUMABLE", "PARTS_EXPORT"): return "PARTS"
                if st == "OTHER": return "ETC"
                return (proj_sf or "ASSEMBLY").upper()
            _bf_made = 0
            for prow in c.execute("SELECT * FROM projects").fetchall():
                p = dict(prow)
                pid = p["id"]
                if c.execute("SELECT COUNT(*) FROM project_items WHERE project_id=?", (pid,)).fetchone()[0]:
                    continue
                proj_sf = p.get("shipment_form") or "ASSEMBLY"
                ords = c.execute("SELECT * FROM orders WHERE project_id=? ORDER BY id", (pid,)).fetchall()
                if ords:
                    _seq = 0
                    for orow in ords:
                        o = dict(orow); _seq += 1
                        _oqty = int(o.get("unit_qty") or 1) if _has_oqty else 1
                        _amt  = float(o.get("total_amount") or 0)
                        _up   = (_amt / _oqty) if _oqty else _amt
                        _sf   = _sf_from_so_type(o.get("so_type") if _has_sotype else "", proj_sf)
                        _iname = (o.get("unit_label") if _has_olabel else None) or p.get("model_name") or p.get("name")
                        c.execute(
                            "INSERT INTO project_items(project_id, seq, shipment_form, item_name, model_name, "
                            "qty, unit_price, amount, currency, is_export, order_customer_id, "
                            "secondary_customer, final_amount, order_date, due_date, ship_to, note, order_id) "
                            "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                            (pid, _seq, _sf, _iname, p.get("model_name"),
                             _oqty, _up, _amt,
                             (o.get("currency") if _has_ocur else None) or p.get("currency") or "KRW",
                             int(p.get("is_export") or 0), o.get("customer_id"),
                             p.get("secondary_customer"), p.get("final_amount"),
                             o.get("order_date"), o.get("due_date"),
                             (o.get("ship_to") if _has_oship else None),
                             "[백필] 기존 수주 기준", o["id"]))
                        _bf_made += 1
                else:
                    _amt = float(p.get("order_amount") or 0) or float(p.get("expected_amount") or 0)
                    _qty = int(p.get("unit_qty") or 1)
                    _up  = p.get("unit_price")
                    if _up is None:
                        _up = (_amt / _qty) if _qty else _amt
                    c.execute(
                        "INSERT INTO project_items(project_id, seq, shipment_form, item_name, model_name, "
                        "qty, unit_price, amount, currency, is_export, secondary_customer, final_amount, "
                        "order_date, due_date, note) "
                        "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        (pid, 1, (proj_sf or "ASSEMBLY").upper(),
                         p.get("model_name") or p.get("name"), p.get("model_name"),
                         _qty, _up, _amt, p.get("currency") or "KRW",
                         int(p.get("is_export") or 0),
                         p.get("secondary_customer"), p.get("final_amount"),
                         p.get("order_date"), p.get("due_date"),
                         "[백필] 제안단계 대표품목"))
                    _bf_made += 1
            if _bf_made:
                print(f"[v5H226z208] project_items 백필 {_bf_made}건 생성")
        except Exception as _e:
            print(f"[v5H226z208] project_items 백필 스킵: {_e}")

        # v5H216 (2026-05-08): consumable_orders 에 mgmt_code 컬럼 추가 + 백필
        # 소모품 발주 묶음에도 'S' prefix 관리번호 부여 (예: 001S2605)
        # v5H218 (2026-05-08): biz_div 컬럼 추가 (진행 사업부 — 매출 집계 분류)
        try:
            cocols = {r2[1] for r2 in c.execute("PRAGMA table_info(consumable_orders)").fetchall()}
            if "mgmt_code" not in cocols:
                c.execute("ALTER TABLE consumable_orders ADD COLUMN mgmt_code TEXT")
                print("[v5H216] consumable_orders.mgmt_code 컬럼 추가됨")
            if "biz_div" not in cocols:
                c.execute("ALTER TABLE consumable_orders ADD COLUMN biz_div TEXT")
                print("[v5H218] consumable_orders.biz_div 컬럼 추가됨")
            # v5H226z264 (대표 지시): 작업 일정표 셀 메모(혼합) — 소모품도 부서·담당자·납품위치를
            #   원본에 저장할 수 있게 컬럼 추가(프로젝트의 cc_dept/cc_name/ship_to 와 대응). 추가형·무손실.
            for _coladd in ("cc_dept", "cc_name", "ship_to"):
                if _coladd not in cocols:
                    c.execute(f"ALTER TABLE consumable_orders ADD COLUMN {_coladd} TEXT")
                    print(f"[v5H226z264] consumable_orders.{_coladd} 컬럼 추가됨")
            # v5H226z285 (대표 지시): 소모품 양식에 '거래구분(내수/수출)' 추가 — 업로드 시 반영.
            if "is_export" not in cocols:
                c.execute("ALTER TABLE consumable_orders ADD COLUMN is_export INTEGER DEFAULT 0")
                print("[v5H226z285] consumable_orders.is_export 컬럼 추가됨")
            # v5H226z287 (대표 지시): 엑셀 정보란 전체 반영 — 2차고객사·연락처·모델명·장비명·영업담당자.
            for _coladd in ("secondary_customer", "cc_phone", "model_name", "equip_name", "sales_name"):
                if _coladd not in cocols:
                    c.execute(f"ALTER TABLE consumable_orders ADD COLUMN {_coladd} TEXT")
                    print(f"[v5H226z287] consumable_orders.{_coladd} 컬럼 추가됨")
            # v5H226z563 (대표 지시): 소모품도 엑셀 '형태'(제품/상품/기타) 반영 — 기존 '상품 고정' 폐기.
            if "form_type" not in cocols:
                c.execute("ALTER TABLE consumable_orders ADD COLUMN form_type TEXT")
                print("[v5H226z563] consumable_orders.form_type 컬럼 추가됨")
            # v5H226z367 (대표 지시): 소모품도 거래명세서·세금계산서 발행일자 추적 — 작업일정표 컬럼 연동(소모품 포함 결정)
            for _dc in ("statement_date", "tax_invoice_date"):
                if _dc not in cocols:
                    c.execute(f"ALTER TABLE consumable_orders ADD COLUMN {_dc} TEXT")
                    print(f"[v5H226z367] consumable_orders.{_dc} 컬럼 추가됨")
            # v5H226z467 (대표 지시): 소모품도 세금계산서 3차 분할(계약금/중도금/잔금) — 1차=기존 tax_invoice_date
            for _dc, _typ in (("tax_invoice_amt1", "REAL"), ("tax_invoice_date2", "TEXT"),
                              ("tax_invoice_amt2", "REAL"), ("tax_invoice_date3", "TEXT"),
                              ("tax_invoice_amt3", "REAL")):
                if _dc not in cocols:
                    c.execute(f"ALTER TABLE consumable_orders ADD COLUMN {_dc} {_typ}")
                    print(f"[v5H226z467] consumable_orders.{_dc} 컬럼 추가됨")
            # v5H226z286 (대표 지시): 엑셀의 사진 칸 2개 모두 반영 — 라인에 '사진위치' 이미지 컬럼 추가.
            _coicols = {r2[1] for r2 in c.execute("PRAGMA table_info(consumable_order_items)").fetchall()}
            for _lc in ("image_loc_path", "image_loc_thumb_path"):
                if _lc not in _coicols:
                    c.execute(f"ALTER TABLE consumable_order_items ADD COLUMN {_lc} TEXT")
                    print(f"[v5H226z286] consumable_order_items.{_lc} 컬럼 추가됨")
            # v5H226z288 (대표 지시): 미리보기 엑셀 일치 — 라인 장비명 저장
            if "equip_name" not in _coicols:
                c.execute("ALTER TABLE consumable_order_items ADD COLUMN equip_name TEXT")
                print("[v5H226z288] consumable_order_items.equip_name 컬럼 추가됨")
            # v5H225: 관리코드 prefix 정책 변경 — K → E (Etc.), S → C (Consumable)
            # 기존 데이터 자동 백필 + project_history 에 변경 이력 기록
            try:
                _migrated = 0
                for _old_p, _new_p in (("K", "E"), ("S", "C")):
                    rows_old = c.execute(
                        "SELECT id, mgmt_code FROM projects "
                        "WHERE mgmt_code IS NOT NULL AND length(mgmt_code) = 8 "
                        "AND substr(mgmt_code, 4, 1) = ?",
                        (_old_p,)
                    ).fetchall()
                    for r in rows_old:
                        old = r["mgmt_code"]
                        new = old[:3] + _new_p + old[4:]
                        c.execute("UPDATE projects SET mgmt_code=? WHERE id=?", (new, r["id"]))
                        try:
                            c.execute(
                                "INSERT INTO project_history(project_id, changed_by, field, old_value, new_value, note) "
                                "VALUES(?,?,?,?,?,?)",
                                (r["id"], None, "관리코드 prefix 정책 변경", old, new,
                                 f"v5H225: {_old_p} prefix → {_new_p} prefix 일괄 마이그레이션")
                            )
                        except Exception:
                            pass
                        _migrated += 1
                # legacy consumable_orders.mgmt_code (S → C)
                try:
                    co_rows = c.execute(
                        "SELECT id, mgmt_code FROM consumable_orders "
                        "WHERE mgmt_code IS NOT NULL AND length(mgmt_code) = 8 "
                        "AND substr(mgmt_code, 4, 1) = 'S'"
                    ).fetchall()
                    for r in co_rows:
                        old = r["mgmt_code"]
                        new = old[:3] + "C" + old[4:]
                        c.execute("UPDATE consumable_orders SET mgmt_code=? WHERE id=?", (new, r["id"]))
                        _migrated += 1
                except Exception:
                    pass
                if _migrated > 0:
                    print(f"[v5H225] 관리코드 prefix 백필: {_migrated}건 (K→E / S→C)")
            except Exception as _e:
                print(f"[v5H225] 백필 실패: {_e}")
            # v5H225: 불일치 검증 (E/C 새 prefix 기준)
            try:
                anomalies = c.execute("""
                    SELECT id, mgmt_code, biz_div, project_type FROM projects
                    WHERE mgmt_code IS NOT NULL AND mgmt_code != ''
                      AND length(mgmt_code) >= 4
                      AND (
                        (project_type='OTHER'      AND substr(mgmt_code,4,1) != 'E') OR
                        (project_type='CONSUMABLE' AND substr(mgmt_code,4,1) != 'C') OR
                        (project_type='NEW_EQUIP'  AND substr(mgmt_code,4,1) NOT IN ('T','M'))
                      )
                """).fetchall()
                if anomalies:
                    print(f"[v5H225] ⚠ 관리코드/유형 불일치 {len(anomalies)}건 감지 (자동 수정 안 함):")
                    for a in anomalies:
                        print(f"  · pid={a['id']} code={a['mgmt_code']} biz={a['biz_div']} type={a['project_type']}")
            except Exception:
                pass
            # 기존 묶음 중 mgmt_code NULL 인 행 백필
            # 같은 conn 안에서 generate_mgmt_code() 가 별도 conn 으로 검색해 미커밋 UPDATE 못 봄 → 직접 SQL 로 sequence 추적
            try:
                _missing = c.execute(
                    "SELECT id, COALESCE(order_date, created_at, date('now')) AS d "
                    "FROM consumable_orders WHERE mgmt_code IS NULL OR mgmt_code = '' "
                    "ORDER BY id ASC"
                ).fetchall()
                _backfilled = 0
                _seq_by_yymm = {}  # YYMM → 이번 백필 내 발급된 max sequence
                # 기존 DB 의 max sequence 도 미리 스캔
                for _row in c.execute(
                    "SELECT mgmt_code FROM consumable_orders WHERE mgmt_code IS NOT NULL"
                ).fetchall():
                    _code = _row["mgmt_code"] or ""
                    if len(_code) == 8 and _code[3] == "S":
                        try:
                            _seq_by_yymm[_code[4:]] = max(_seq_by_yymm.get(_code[4:], 0), int(_code[:3]))
                        except ValueError:
                            pass
                for _r in _missing:
                    try:
                        _d_iso = (_r["d"] or "")[:10]
                        if _d_iso and len(_d_iso) >= 7:
                            _y, _m = int(_d_iso[:4]), int(_d_iso[5:7])
                            _yymm = f"{_y % 100:02d}{_m:02d}"
                        else:
                            from datetime import date as _dt
                            _today = _dt.today()
                            _yymm = f"{_today.year % 100:02d}{_today.month:02d}"
                        _seq_by_yymm[_yymm] = _seq_by_yymm.get(_yymm, 0) + 1
                        _code_new = f"{_seq_by_yymm[_yymm]:03d}S{_yymm}"
                        c.execute("UPDATE consumable_orders SET mgmt_code=? WHERE id=?", (_code_new, _r["id"]))
                        _backfilled += 1
                    except Exception:
                        pass
                if _backfilled > 0:
                    print(f"[v5H216] 소모품 묶음 mgmt_code 백필: {_backfilled}건")
            except Exception as _e:
                print(f"[v5H216] 백필 실패: {_e}")
            # v5H226z262 (대표 지시·연결성 감사): 소모품 식별자 유일성 보장.
            #   근본원인 = projects.mgmt_code 는 UNIQUE 인데 consumable_orders.co_no/mgmt_code 는 무제약 →
            #   동시 업로드 race 시 같은 수주번호·관리코드가 '오류 없이' 중복 저장됨(추적·집계 오염).
            #   조치 = 부분 UNIQUE 인덱스로 DB 차원 차단. 기존 중복이 있으면 인덱스 생성이 실패하므로,
            #   먼저 중복을 탐지·로그(수동 정리 유도)하고 인덱스 생성은 try 로 감싸 startup 크래시 방지.
            try:
                for _c_lbl, _c_col in (("수주번호", "co_no"), ("관리코드", "mgmt_code")):
                    if _c_col not in cocols and _c_col != "co_no":
                        continue
                    _dups = c.execute(
                        f"SELECT {_c_col} AS v, COUNT(*) AS n FROM consumable_orders "
                        f"WHERE {_c_col} IS NOT NULL AND {_c_col} != '' "
                        f"GROUP BY {_c_col} HAVING COUNT(*) > 1"
                    ).fetchall()
                    if _dups:
                        print(f"[v5H226z262] ⚠ 소모품 {_c_lbl} 중복 {len(_dups)}종 감지(수동 정리 필요):")
                        for _dd in _dups:
                            print(f"    · {_dd['v']} × {_dd['n']}건")
                for _idx_sql in (
                    "CREATE UNIQUE INDEX IF NOT EXISTS uq_consumable_co_no "
                    "ON consumable_orders(co_no) WHERE co_no IS NOT NULL AND co_no != ''",
                    "CREATE UNIQUE INDEX IF NOT EXISTS uq_consumable_mgmt "
                    "ON consumable_orders(mgmt_code) WHERE mgmt_code IS NOT NULL AND mgmt_code != ''",
                ):
                    try:
                        c.execute(_idx_sql)
                    except Exception as _e_idx:
                        print(f"[v5H226z262] 소모품 UNIQUE 인덱스 생성 보류(기존 중복 존재 — 정리 후 재시도): {_e_idx}")
            except Exception as _e:
                print(f"[v5H226z262] 소모품 유일성 마이그레이션 스킵: {_e}")
        except Exception:
            pass

        # v5H226z264 (대표 지시): 작업 일정표 보드 — ① 진행기간 막대 '보드 전용 메모'(원본 무수정)
        #   ② 칸 너비·숨김을 '내 계정'에 저장. 둘 다 추가형·무손실.
        try:
            c.execute("""CREATE TABLE IF NOT EXISTS schedule_board_memos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ref_kind   TEXT NOT NULL,          -- 'project' | 'consumable'
                ref_id     INTEGER NOT NULL,
                memo       TEXT,
                updated_by INTEGER,
                updated_at TEXT,
                UNIQUE(ref_kind, ref_id)
            )""")
            # v5H226z712 (대표 지시): 날짜별 진행 메모 — 클릭한 그 날짜(ymd)에만 메모. (행 단위 schedule_board_memos 와 별개)
            c.execute("""CREATE TABLE IF NOT EXISTS schedule_board_day_memos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ref_kind   TEXT NOT NULL,          -- 'project' | 'consumable'
                ref_id     INTEGER NOT NULL,
                ymd        TEXT NOT NULL,          -- YYYY-MM-DD (클릭한 날짜)
                memo       TEXT,
                updated_by INTEGER,
                updated_at TEXT,
                UNIQUE(ref_kind, ref_id, ymd)
            )""")
            c.execute("""CREATE TABLE IF NOT EXISTS user_view_prefs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id   INTEGER NOT NULL,
                view_key  TEXT NOT NULL,           -- 예: 'schedule_board_cols'
                config    TEXT,                    -- JSON
                updated_at TEXT,
                UNIQUE(user_id, view_key)
            )""")
        except Exception as _e:
            print(f"[v5H226z264] 일정표 보드 표 생성 스킵: {_e}")

        # v5H226z518 (대표 지시): WORKS '지시형' 전환 — 업무 지시/요청 원장(work_order).
        #   '지시'=상급자→부하/부서·영업→부서(제작·프로젝트성, 잡무 제외) / '요청'=동료 수평(수락·반려 자유).
        #   '내 업무함'(일일업무 개편)·'일 시키기'·'내가 시킨 일'의 단일 저장소. 시간대=입력자 소속 기준.
        try:
            c.execute("""CREATE TABLE IF NOT EXISTS work_order (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                kind            TEXT NOT NULL DEFAULT '지시',   -- '지시' | '요청'
                title           TEXT NOT NULL,
                body            TEXT,
                created_by      INTEGER NOT NULL,              -- 지시/요청한 사람
                created_by_name TEXT,
                target_user_id  INTEGER,                       -- 사람 대상(택1)
                target_team_id  INTEGER,                       -- 부서 대상(택1)
                target_label    TEXT,                          -- 표시용(사람명/부서명 스냅샷)
                project_id      INTEGER,                       -- 연결 관리번호/프로젝트(선택)
                project_code    TEXT,                          -- 표시용 관리번호 스냅샷
                due_date        TEXT,
                priority        TEXT DEFAULT '보통',           -- 긴급|높음|보통|낮음
                status          TEXT DEFAULT '받음',           -- 받음|진행|완료|반려
                report          TEXT,                          -- 완료/반려 보고
                hours           REAL,                          -- 소요시간(h)
                done_at         TEXT,
                accepted_by     INTEGER,                       -- 부서 대상일 때 실제 집어든 사람
                accepted_by_name TEXT,
                created_at      TEXT,
                updated_at      TEXT
            )""")
            c.execute("CREATE INDEX IF NOT EXISTS idx_wo_target_user ON work_order(target_user_id, status)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_wo_target_team ON work_order(target_team_id, status)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_wo_created_by ON work_order(created_by, status)")
        except Exception as _e:
            print(f"[v5H226z518] work_order 표 생성 스킵: {_e}")

        # v5H226z270 (대표 지시): 작업 일정표 = 업무 원장 허브 — 단계 파이프라인 기록표.
        #   각 프로젝트/소모품의 단계별 진행을 '누가·언제' 와 함께 적재(추적). 단계는 STAGE_SPEC 화이트리스트.
        try:
            c.execute("""CREATE TABLE IF NOT EXISTS project_stage_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ref_kind   TEXT NOT NULL,          -- 'project' | 'consumable'
                ref_id     INTEGER NOT NULL,
                stage_key  TEXT NOT NULL,          -- input,quote,order,prod_request,progress,verify,label,pack,ship,tax_invoice
                sub_key    TEXT NOT NULL DEFAULT '', -- progress 세부: design,circuit,program,wiring,assembly,test (외엔 '')
                done       INTEGER DEFAULT 0,
                on_date    TEXT,                   -- 단계 일자(출하일·완료일 등)
                by_user    INTEGER,
                by_name    TEXT,                   -- 누른 사람 이름(표시용 스냅샷)
                memo       TEXT,
                extra      TEXT,                   -- JSON (출하:수량/국내수출 · 세금계산서:금액 등)
                updated_at TEXT,
                UNIQUE(ref_kind, ref_id, stage_key, sub_key)
            )""")
        except Exception as _e:
            print(f"[v5H226z270] 단계 기록표 생성 스킵: {_e}")
        # v5H226z477 (대표 지시): 단계를 '시작전→진행중→완료'로 — 착수(시작) 시각·담당자 + 완료 소요시간(h)
        for _c477, _t477 in (("started_at", "TEXT"), ("started_by", "INTEGER"),
                             ("started_by_name", "TEXT"), ("hours", "REAL")):
            try:
                c.execute(f"ALTER TABLE project_stage_log ADD COLUMN {_c477} {_t477}")
            except Exception:
                pass   # 이미 있으면 무시(idempotent)
        # v5H226z479 (대표 지시·Phase 2): 단계 '착수 가능' 부서 통보 1회 dedup 기록
        try:
            c.execute("""CREATE TABLE IF NOT EXISTS stage_notify_log (
                ref_kind   TEXT NOT NULL,
                ref_id     INTEGER NOT NULL,
                stage_key  TEXT NOT NULL,
                sub_key    TEXT NOT NULL DEFAULT '',
                notified_at TEXT,
                UNIQUE(ref_kind, ref_id, stage_key, sub_key)
            )""")
        except Exception as _e:
            print(f"[v5H226z479] stage_notify_log 생성 스킵: {_e}")

        # v5H226z488 (대표 지시): 세금계산서 '묶어 발행' — 여러 출하 건(여러 관리번호/여러 호기/같은 출하시점)을
        #   한 장의 세금계산서로 합산 발행. 장부(tax_invoices) 1장 + 묶인 항목(tax_invoice_lines).
        try:
            c.execute("""CREATE TABLE IF NOT EXISTS tax_invoices (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                ti_no         TEXT,                       -- 묶음번호 TI-YYMM-####
                issue_date    TEXT,                       -- 발행일 YYYY-MM-DD
                customer_name TEXT,
                customer_id   INTEGER,
                total_amount  REAL DEFAULT 0,
                currency      TEXT DEFAULT 'KRW',
                memo          TEXT,
                status        TEXT DEFAULT '발행',         -- 발행 / 취소
                created_by    INTEGER,
                created_by_name TEXT,
                created_at    TEXT DEFAULT (datetime('now','localtime')),
                updated_at    TEXT DEFAULT (datetime('now','localtime'))
            )""")
            c.execute("""CREATE TABLE IF NOT EXISTS tax_invoice_lines (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                ti_id       INTEGER NOT NULL REFERENCES tax_invoices(id) ON DELETE CASCADE,
                row_sig     TEXT,                          -- 작업일정표 행 식별('project:123'/'consumable:45'/'unit:7,12,30')
                ref_kind    TEXT,                          -- project / consumable / unit
                ref_id      INTEGER,                       -- 프로젝트·소모품 id (unit이면 프로젝트 id)
                unit_iids   TEXT,                          -- 호기(order_items) id CSV (whole이면 비움)
                mgmt_code   TEXT,                          -- 관리번호 스냅샷
                label       TEXT,                          -- 호기 라벨 스냅샷('1~3호기' 등)
                customer    TEXT,                          -- 고객사 스냅샷
                amount      REAL DEFAULT 0,
                tier        INTEGER DEFAULT 1              -- v5H226z489: 차수(1=계약금·2=중도금·3=잔금)
            )""")
            c.execute("CREATE INDEX IF NOT EXISTS idx_ti_lines_sig ON tax_invoice_lines(row_sig)")
            # v5H226z489 (대표 지시): 기존 배포 테이블에 tier(차수) 보강 — 묶음을 1/2/3차 단위로
            _tlc = {r[1] for r in c.execute("PRAGMA table_info(tax_invoice_lines)").fetchall()}
            if "tier" not in _tlc:
                c.execute("ALTER TABLE tax_invoice_lines ADD COLUMN tier INTEGER DEFAULT 1")
        except Exception as _e:
            print(f"[v5H226z488] tax_invoices 생성 스킵: {_e}")

        # v5H226z496 (대표 지시): 부서 일정표 — 부서가 작업일정표 날짜칸을 클릭해 진행/완료/변경 사항을 기록.
        #   아침회의에서 품목별로 "○일까지 완료→다음 부서로 넘김"을 달력칸에 적던 방식을 그대로. 단계(STAGE)와 무관·자유 부서.
        try:
            c.execute("""CREATE TABLE IF NOT EXISTS dept_schedule_log (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                ref_kind        TEXT NOT NULL,                 -- project / consumable
                ref_id          INTEGER NOT NULL,              -- 프로젝트·소모품 id
                log_date        TEXT NOT NULL,                 -- 날짜칸 YYYY-MM-DD
                dept            TEXT,                          -- 기록 부서
                log_type        TEXT DEFAULT '진행',            -- 진행 / 완료 / 변경 / 기타
                content         TEXT,
                created_by      INTEGER,
                created_by_name TEXT,
                created_at      TEXT DEFAULT (datetime('now','localtime')),
                updated_at      TEXT DEFAULT (datetime('now','localtime'))
            )""")
            c.execute("CREATE INDEX IF NOT EXISTS idx_dsl_ref ON dept_schedule_log(ref_kind, ref_id, log_date)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_dsl_date ON dept_schedule_log(log_date)")
            # v5H226z497 (대표 지시·Phase 2): '넘길 부서'(다음 담당) — 기록 시 지정하면 그 부서에 Eum+인앱 통보
            # v5H226z499 (대표 지시·Phase 3): unit_label = 호기별 기록(수량 여러 대=완제품). 빈값=품목 전체.
            _dslc = {r[1] for r in c.execute("PRAGMA table_info(dept_schedule_log)").fetchall()}
            if "handoff_to" not in _dslc:
                c.execute("ALTER TABLE dept_schedule_log ADD COLUMN handoff_to TEXT")
            if "unit_label" not in _dslc:
                c.execute("ALTER TABLE dept_schedule_log ADD COLUMN unit_label TEXT")
            # v5H226z570 (대표 지시·업무량 집계): 소요시간(시간) — 완료 기록 시 직접 입력(부서/개인 업무량 측정)
            if "hours" not in _dslc:
                c.execute("ALTER TABLE dept_schedule_log ADD COLUMN hours REAL")
        except Exception as _e:
            print(f"[v5H226z496] dept_schedule_log 생성 스킵: {_e}")

        # ───────── v5H226z578 (대표 지시): 부서 자체 연구개발(R&D) 과제 — 매출과 별개·관리번호 R 발번 ─────────
        try:
            c.execute("""CREATE TABLE IF NOT EXISTS rnd_projects (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                mgmt_code       TEXT,                          -- 001R2606 (연구과제 관리번호)
                title           TEXT NOT NULL,                 -- 과제명
                dept            TEXT,                          -- 담당 부서(라벨)
                owner_id        INTEGER,                       -- 담당자 user id
                owner_name      TEXT,                          -- 담당자명
                category        TEXT,                          -- 신제품/공정개선/지그·툴/자동화/기타
                purpose         TEXT,                          -- 배경·목적
                start_date      TEXT,
                target_date     TEXT,                          -- 목표 완료일
                priority        TEXT DEFAULT '보통',            -- 긴급/높음/보통/낮음
                status          TEXT DEFAULT '기획',            -- 기획/진행/시험검증/완료/보류/중단
                result_summary  TEXT,                          -- 결과 요약
                created_by      INTEGER,
                created_by_name TEXT,
                created_at      TEXT DEFAULT (datetime('now','localtime')),
                updated_at      TEXT DEFAULT (datetime('now','localtime'))
            )""")
            c.execute("CREATE INDEX IF NOT EXISTS idx_rnd_dept ON rnd_projects(dept)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_rnd_status ON rnd_projects(status)")
            c.execute("""CREATE TABLE IF NOT EXISTS rnd_logs (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                rnd_id          INTEGER NOT NULL,
                log_date        TEXT,                          -- YYYY-MM-DD
                log_type        TEXT DEFAULT '진행',            -- 진행/완료/변경/기타
                content         TEXT,
                hours           REAL,
                created_by      INTEGER,
                created_by_name TEXT,
                created_at      TEXT DEFAULT (datetime('now','localtime'))
            )""")
            c.execute("CREATE INDEX IF NOT EXISTS idx_rndlog_rid ON rnd_logs(rnd_id, id)")
        except Exception as _e:
            print(f"[v5H226z578] rnd_projects 생성 스킵: {_e}")

        # 마이그레이션 (수출입 P11 2차): export_orders.status CHECK 확장
        # — 1차에서 'DRAFT,BOOKED,SHIPPED,CLEARED,CLOSED,CANCELLED' 였던 것을
        #   라이프사이클 'DRAFT → CI_ISSUED → PL_READY → SHIPPED → CLEARED' 반영
        # — SQLite CHECK 제약은 ALTER 불가 → table 재생성 idempotent 가드
        try:
            row = c.execute(
                "SELECT sql FROM sqlite_master WHERE type='table' AND name='export_orders'"
            ).fetchone()
            if row and "CI_ISSUED" not in (row["sql"] or ""):
                c.executescript("""
                    BEGIN;
                    CREATE TABLE export_orders__new (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        order_id INTEGER REFERENCES orders(id),
                        buyer TEXT, shipping_terms TEXT, payment_terms TEXT,
                        port_of_loading TEXT, port_of_discharge TEXT,
                        status TEXT DEFAULT 'DRAFT'
                            CHECK(status IN ('DRAFT','BOOKED','CI_ISSUED','PL_READY','SHIPPED','CLEARED','CLOSED','CANCELLED')),
                        created_by INTEGER REFERENCES users(id),
                        created_at TEXT DEFAULT (datetime('now','localtime'))
                    );
                    INSERT INTO export_orders__new
                        SELECT id, order_id, buyer, shipping_terms, payment_terms,
                               port_of_loading, port_of_discharge, status,
                               created_by, created_at FROM export_orders;
                    DROP TABLE export_orders;
                    ALTER TABLE export_orders__new RENAME TO export_orders;
                    CREATE INDEX IF NOT EXISTS idx_export_orders_order ON export_orders(order_id);
                    CREATE INDEX IF NOT EXISTS idx_export_orders_status ON export_orders(status);
                    COMMIT;
                """)
        except Exception:
            pass

        # 마이그레이션 (QMS 강화 1차 · 갭서베이 Top10 #6 · 2026-04-26):
        # issues 테이블 +5 컬럼 (sla_hours/detected_at/responded_at/resolved_at_v2/recurrence_id)
        # users 테이블 +1 컬럼 (can_use_quality)
        # 기존 severity / resolved_at / root_cause / prevention 무수정 (재사용)
        try:
            icols = [r[1] for r in c.execute("PRAGMA table_info(issues)").fetchall()]
            for col, decl in [
                ("sla_hours",      "ALTER TABLE issues ADD COLUMN sla_hours INTEGER DEFAULT 24"),
                ("detected_at",    "ALTER TABLE issues ADD COLUMN detected_at TEXT"),
                ("responded_at",   "ALTER TABLE issues ADD COLUMN responded_at TEXT"),
                ("recurrence_id",  "ALTER TABLE issues ADD COLUMN recurrence_id TEXT"),
                ("sla_breached",   "ALTER TABLE issues ADD COLUMN sla_breached INTEGER DEFAULT 0"),
            ]:
                if col not in icols:
                    try:
                        c.execute(decl)
                    except Exception:
                        pass
            # 기존 인덱스 1개 추가
            try:
                c.execute("CREATE INDEX IF NOT EXISTS idx_iss_recurrence ON issues(recurrence_id)")
            except Exception:
                pass
        except Exception:
            pass

        try:
            ucols = [r[1] for r in c.execute("PRAGMA table_info(users)").fetchall()]
            if "can_use_quality" not in ucols:
                try:
                    c.execute("ALTER TABLE users ADD COLUMN can_use_quality INTEGER DEFAULT 0")
                except Exception:
                    pass
            # 품질팀 leader/member 자동 부여 (idempotent UPDATE)
            try:
                c.execute(
                    "UPDATE users SET can_use_quality=1 "
                    "WHERE team_id IN (SELECT id FROM teams WHERE name LIKE '%품질%') "
                    "AND is_active=1"
                )
                # admin/ceo/executive 전권 자동
                c.execute(
                    "UPDATE users SET can_use_quality=1 "
                    "WHERE role IN ('admin','ceo','executive') AND is_active=1"
                )
            except Exception:
                pass
        except Exception:
            pass

        # v5H226z550: customers 에 별칭(alias) 컬럼 — 일정표 등 화면 표시용 짧은 이름('주식회사 김정락'→'김정락').
        #   연결/매칭은 정식 name 으로 그대로, alias 는 표시 전용. (없으면 name 으로 폴백)
        try:
            _cucols = [r[1] for r in c.execute("PRAGMA table_info(customers)").fetchall()]
            if "alias" not in _cucols:
                try:
                    c.execute("ALTER TABLE customers ADD COLUMN alias TEXT")
                except Exception:
                    pass
            # v5H226z655 (대표 지시): resolve_customer_id 의 `OR alias=?`(상호/시스템명 매칭) 가속 — alias 인덱스
            try:
                c.execute("CREATE INDEX IF NOT EXISTS idx_customers_alias ON customers(alias)")
            except Exception:
                pass
        except Exception:
            pass

        # v5H226z542: teams 에 법인(entity KOR/VN) 컬럼 — 메신저 본사/베트남 부서 분리 표시·매핑.
        #   기존 팀 이름은 그대로 둠(공정 STAGE_OWNERS·권한 매칭 안전). NULL 인 것만 1회 세팅(idempotent).
        try:
            tcols = [r[1] for r in c.execute("PRAGMA table_info(teams)").fetchall()]
            if "entity" not in tcols:
                try:
                    c.execute("ALTER TABLE teams ADD COLUMN entity TEXT")
                except Exception:
                    pass
            # 본사(KOR) 기능팀 — 기존 13팀 + z540 자동생성 '총괄'
            for _nm in ("기술영업팀", "검사기팀", "품질팀", "설계팀", "소프트웨어팀", "전장설계팀",
                        "제조기술1팀", "제조기술2팀", "가공팀", "구매팀", "관리팀", "개발혁신팀",
                        "라이프밸류팀", "총괄"):
                try:
                    c.execute("UPDATE teams SET entity='KOR' WHERE name=? AND (entity IS NULL OR entity='')", (_nm,))
                except Exception:
                    pass
            # 베트남(VN) — z540 자동생성 베트남 전용 이름 + 베트남법인(legacy 버킷)
            for _nm in ("기술팀", "조립팀", "전장팀", "베트남법인"):
                try:
                    c.execute("UPDATE teams SET entity='VN' WHERE name=? AND (entity IS NULL OR entity='')", (_nm,))
                except Exception:
                    pass
        except Exception:
            pass

        # 마이그레이션 (QMS 2차 Pareto·CAPA 심화 · 2026-04-26):
        # corrective_actions / preventive_actions 라이프사이클 컬럼 6종 (CHECK 제약 없이 ALTER ADD)
        # DRAFT → APPROVED → IN_PROGRESS → COMPLETED → VERIFIED
        try:
            cacols = [r[1] for r in c.execute("PRAGMA table_info(corrective_actions)").fetchall()]
            for col, decl in [
                ("lifecycle_status",  "ALTER TABLE corrective_actions ADD COLUMN lifecycle_status TEXT DEFAULT 'DRAFT'"),
                ("approved_by",       "ALTER TABLE corrective_actions ADD COLUMN approved_by INTEGER REFERENCES users(id)"),
                ("approved_at",       "ALTER TABLE corrective_actions ADD COLUMN approved_at TEXT"),
                ("verified_by",       "ALTER TABLE corrective_actions ADD COLUMN verified_by INTEGER REFERENCES users(id)"),
                ("verified_at",       "ALTER TABLE corrective_actions ADD COLUMN verified_at TEXT"),
                ("effectiveness_note","ALTER TABLE corrective_actions ADD COLUMN effectiveness_note TEXT"),
            ]:
                if col not in cacols:
                    try:
                        c.execute(decl)
                    except Exception:
                        pass
            try:
                c.execute("CREATE INDEX IF NOT EXISTS idx_ca_lifecycle ON corrective_actions(lifecycle_status)")
            except Exception:
                pass
        except Exception:
            pass

        try:
            pacols = [r[1] for r in c.execute("PRAGMA table_info(preventive_actions)").fetchall()]
            for col, decl in [
                ("lifecycle_status",  "ALTER TABLE preventive_actions ADD COLUMN lifecycle_status TEXT DEFAULT 'DRAFT'"),
                ("approved_by",       "ALTER TABLE preventive_actions ADD COLUMN approved_by INTEGER REFERENCES users(id)"),
                ("approved_at",       "ALTER TABLE preventive_actions ADD COLUMN approved_at TEXT"),
                ("verified_by",       "ALTER TABLE preventive_actions ADD COLUMN verified_by INTEGER REFERENCES users(id)"),
                ("verified_at",       "ALTER TABLE preventive_actions ADD COLUMN verified_at TEXT"),
                ("effectiveness_note","ALTER TABLE preventive_actions ADD COLUMN effectiveness_note TEXT"),
            ]:
                if col not in pacols:
                    try:
                        c.execute(decl)
                    except Exception:
                        pass
            try:
                c.execute("CREATE INDEX IF NOT EXISTS idx_pa_lifecycle ON preventive_actions(lifecycle_status)")
            except Exception:
                pass
        except Exception:
            pass

        # 사이클 78 (2026-04-27 DEC-1) — 세금계산서 발행여부 체크박스
        # 외부 KNK 회계 시스템에서 발행 후, HAIST WORKS는 발행 여부만 표시
        try:
            ocols = [r[1] for r in c.execute("PRAGMA table_info(orders)").fetchall()]
            for col, decl in [
                ("tax_invoice_issued", "ALTER TABLE orders ADD COLUMN tax_invoice_issued INTEGER DEFAULT 0"),
                ("tax_invoice_no",     "ALTER TABLE orders ADD COLUMN tax_invoice_no TEXT"),
                ("tax_invoice_date",   "ALTER TABLE orders ADD COLUMN tax_invoice_date TEXT"),
                ("tax_invoice_note",   "ALTER TABLE orders ADD COLUMN tax_invoice_note TEXT"),
                # v5H226z367 (대표 지시): 거래명세서 발행일자 — 세금계산서(tax_invoice_date)와 나란히 SO 단위 추적(작업일정표 컬럼)
                ("statement_date",     "ALTER TABLE orders ADD COLUMN statement_date TEXT"),
                # v5H226z467 (대표 지시): 세금계산서 3차 분할(계약금/중도금/잔금). 1차 발행일=기존 tax_invoice_date.
                ("tax_invoice_amt1",   "ALTER TABLE orders ADD COLUMN tax_invoice_amt1 REAL"),
                ("tax_invoice_date2",  "ALTER TABLE orders ADD COLUMN tax_invoice_date2 TEXT"),
                ("tax_invoice_amt2",   "ALTER TABLE orders ADD COLUMN tax_invoice_amt2 REAL"),
                ("tax_invoice_date3",  "ALTER TABLE orders ADD COLUMN tax_invoice_date3 TEXT"),
                ("tax_invoice_amt3",   "ALTER TABLE orders ADD COLUMN tax_invoice_amt3 REAL"),
                # v5H78: 호기별 다중 발주 — 1호기/2호기 ... 호기 라벨 + 비고
                ("unit_label",         "ALTER TABLE orders ADD COLUMN unit_label TEXT"),
                ("unit_note",          "ALTER TABLE orders ADD COLUMN unit_note TEXT"),
                # v5H81: 납품지 (SO 그룹화 키 — 동일 납기 + 동일 납품지 = 1 SO)
                ("ship_to",            "ALTER TABLE orders ADD COLUMN ship_to TEXT"),
                # 호기 수량 (SO 안에 묶인 호기 개수)
                ("unit_qty",           "ALTER TABLE orders ADD COLUMN unit_qty INTEGER DEFAULT 1"),
                # v5H92: 통화 (KRW / USD) — SO 단위
                ("currency",           "ALTER TABLE orders ADD COLUMN currency TEXT DEFAULT 'KRW'"),
                # v5H142 (2026-05-05): SO 종류 — 부모 프로젝트 안에서 SO 단위로 종류 분리
                # EQUIPMENT/CONSUMABLE/SERVICE/OTHER. NULL → EQUIPMENT 폴백.
                ("so_type",            "ALTER TABLE orders ADD COLUMN so_type TEXT DEFAULT 'EQUIPMENT'"),
            ]:
                if col not in ocols:
                    try:
                        c.execute(decl)
                    except Exception:
                        pass
        except Exception:
            pass

        # v5H124 (2026-05-04): receipts_payment.currency — 외화 수금 추적
        # 백워드 호환: 기존 데이터는 NULL → 조회 시 KRW 로 간주
        try:
            rpcols = [r[1] for r in c.execute("PRAGMA table_info(receipts_payment)").fetchall()]
            if "currency" not in rpcols:
                try:
                    c.execute("ALTER TABLE receipts_payment ADD COLUMN currency TEXT DEFAULT 'KRW'")
                except Exception:
                    pass
            if "fx_rate" not in rpcols:
                try:
                    c.execute("ALTER TABLE receipts_payment ADD COLUMN fx_rate REAL")
                except Exception:
                    pass
        except Exception:
            pass

        # v5H81: order_items — 호기 라벨 + 라인 비고
        # v5H177: 호기별 발주일/납기/납품처/통화 override (NULL 이면 SO 부모값 상속)
        try:
            oicols = [r[1] for r in c.execute("PRAGMA table_info(order_items)").fetchall()]
            for col, decl in [
                ("unit_label", "ALTER TABLE order_items ADD COLUMN unit_label TEXT"),
                ("line_note",  "ALTER TABLE order_items ADD COLUMN line_note TEXT"),
                ("order_date", "ALTER TABLE order_items ADD COLUMN order_date TEXT"),
                ("due_date",   "ALTER TABLE order_items ADD COLUMN due_date TEXT"),
                ("ship_to",    "ALTER TABLE order_items ADD COLUMN ship_to TEXT"),
                ("currency",   "ALTER TABLE order_items ADD COLUMN currency TEXT"),
                # v5H186: 호기별 상태 — 진행중/출하/취소/보류 등
                ("unit_status", "ALTER TABLE order_items ADD COLUMN unit_status TEXT DEFAULT '진행중'"),
                # v5H197: 호기별 거래구분 — NULL=프로젝트 상속, 0=내수, 1=수출
                ("is_export",   "ALTER TABLE order_items ADD COLUMN is_export INTEGER"),
                # v5H226c: 소모품 라인 이미지 (엑셀 셀에 박힌 이미지 추출/압축본)
                ("image_path",       "ALTER TABLE order_items ADD COLUMN image_path TEXT"),
                ("image_thumb_path", "ALTER TABLE order_items ADD COLUMN image_thumb_path TEXT"),
                # v5H226z: 부품 형태 출고용 (정식 PACKING LIST) — 메이커/원산지/BOX/규격/입고일정
                ("maker",            "ALTER TABLE order_items ADD COLUMN maker TEXT"),
                ("origin",           "ALTER TABLE order_items ADD COLUMN origin TEXT"),
                ("box_no",           "ALTER TABLE order_items ADD COLUMN box_no TEXT"),
                ("spec",             "ALTER TABLE order_items ADD COLUMN spec TEXT"),
                ("arrival_status",   "ALTER TABLE order_items ADD COLUMN arrival_status TEXT"),
                # v5H226z4: 업체명(supplier) + 단위(unit) 별도 컬럼화
                ("supplier",         "ALTER TABLE order_items ADD COLUMN supplier TEXT"),
                ("unit",             "ALTER TABLE order_items ADD COLUMN unit TEXT"),
                # v5H226z441 (대표 지시): 자재(상품) 입력 항목 — 자재번호 자유입력 신설
                ("material_no",      "ALTER TABLE order_items ADD COLUMN material_no TEXT"),
                # v5H226z460 (대표 지시): 상품(PARTS) 원가관리 — 매입단가 + 마진%(가산율).
                #   판매단가(unit_price) = 매입단가 × (1+마진%/100). 직접수정 시 마진% 역산.
                #   amount = qty × unit_price(판매) 그대로. 원가는 상세화면 can_money 마스킹.
                ("cost_price",       "ALTER TABLE order_items ADD COLUMN cost_price REAL"),
                ("margin_pct",       "ALTER TABLE order_items ADD COLUMN margin_pct REAL"),
                # v5H226z735 (대표 지시): 수출 건 매입/판매단가의 'KRW 원본'(우리 보기용 참고값). 인보이스(USD)는 별도.
                #   환율 역산(×fx) 표시는 1~2원 오차가 나므로, 입력한 KRW를 그대로 저장·표시(정확).
                ("cost_krw",         "ALTER TABLE order_items ADD COLUMN cost_krw REAL"),
                ("sell_krw",         "ALTER TABLE order_items ADD COLUMN sell_krw REAL"),
                # v5H226z5: 정식 PACKING LIST 통관 컬럼 — HS CODE/DUTY/VAT/인보이스/관세/최종/상세/PALLET/중량
                ("hs_code",                "ALTER TABLE order_items ADD COLUMN hs_code TEXT"),
                ("duty_rate",              "ALTER TABLE order_items ADD COLUMN duty_rate REAL"),
                ("vat_rate",               "ALTER TABLE order_items ADD COLUMN vat_rate REAL"),
                ("invoice_unit_price",     "ALTER TABLE order_items ADD COLUMN invoice_unit_price REAL"),
                ("invoice_amount",         "ALTER TABLE order_items ADD COLUMN invoice_amount REAL"),
                ("invoice_unit_price_usd", "ALTER TABLE order_items ADD COLUMN invoice_unit_price_usd REAL"),
                ("invoice_amount_usd",     "ALTER TABLE order_items ADD COLUMN invoice_amount_usd REAL"),
                ("duty_krw",               "ALTER TABLE order_items ADD COLUMN duty_krw REAL"),
                ("duty_usd",               "ALTER TABLE order_items ADD COLUMN duty_usd REAL"),
                ("final_amount_usd",       "ALTER TABLE order_items ADD COLUMN final_amount_usd REAL"),
                ("description",            "ALTER TABLE order_items ADD COLUMN description TEXT"),
                ("pallet_size",            "ALTER TABLE order_items ADD COLUMN pallet_size TEXT"),
                ("weight_kg",              "ALTER TABLE order_items ADD COLUMN weight_kg REAL"),
                # v5H226z484 (대표 지시): 호기별 거래명세서·세금계산서(발행일+금액). 호기마다 납품일 다르면 발행일도 다를 수 있어 호기 단위 저장.
                ("statement_date",     "ALTER TABLE order_items ADD COLUMN statement_date TEXT"),
                ("tax_invoice_date",   "ALTER TABLE order_items ADD COLUMN tax_invoice_date TEXT"),
                ("tax_invoice_amt1",   "ALTER TABLE order_items ADD COLUMN tax_invoice_amt1 REAL"),
                ("tax_invoice_date2",  "ALTER TABLE order_items ADD COLUMN tax_invoice_date2 TEXT"),
                ("tax_invoice_amt2",   "ALTER TABLE order_items ADD COLUMN tax_invoice_amt2 REAL"),
                ("tax_invoice_date3",  "ALTER TABLE order_items ADD COLUMN tax_invoice_date3 TEXT"),
                ("tax_invoice_amt3",   "ALTER TABLE order_items ADD COLUMN tax_invoice_amt3 REAL"),
                # v5H226z711 (대표 지시): 작업일정표 달력에서 상태(출하/취소/보류) 선택 시 '발생일' 기록 — 호기별
                ("status_date",        "ALTER TABLE order_items ADD COLUMN status_date TEXT"),
            ]:
                if col not in oicols:
                    try:
                        c.execute(decl)
                    except Exception:
                        pass
        except Exception:
            pass

        # v5H101: 프로젝트 변경 이력 테이블 (대표 지시: '변경사항 기록 이력서')
        try:
            c.execute("""
                CREATE TABLE IF NOT EXISTS project_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    project_id INTEGER NOT NULL,
                    changed_at TEXT DEFAULT (datetime('now','localtime')),
                    changed_by INTEGER,
                    field TEXT,
                    old_value TEXT,
                    new_value TEXT,
                    note TEXT
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_project_history_pid ON project_history(project_id)")
        except Exception:
            pass

        # v5H112: 고객사 변경 이력 테이블 (자료 보존 + 감사 대응)
        try:
            c.execute("""
                CREATE TABLE IF NOT EXISTS customer_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    customer_id INTEGER NOT NULL,
                    changed_at TEXT DEFAULT (datetime('now','localtime')),
                    changed_by INTEGER,
                    field TEXT,
                    old_value TEXT,
                    new_value TEXT,
                    note TEXT
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_customer_history_cid ON customer_history(customer_id)")
        except Exception:
            pass

        # v5H113: 견적 라인 변경 이력 (M10)
        try:
            c.execute("""
                CREATE TABLE IF NOT EXISTS quotation_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    quotation_id INTEGER NOT NULL,
                    changed_at TEXT DEFAULT (datetime('now','localtime')),
                    changed_by INTEGER,
                    field TEXT,
                    old_value TEXT,
                    new_value TEXT,
                    note TEXT
                )
            """)
            c.execute(
                "CREATE INDEX IF NOT EXISTS idx_quotation_history_qid "
                "ON quotation_history(quotation_id)"
            )
        except Exception:
            pass

        # v5H113: user/team 변경 이력 (LOW#18/#19)
        for _ent in ("user", "team"):
            try:
                c.execute(f"""
                    CREATE TABLE IF NOT EXISTS {_ent}_history (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        {_ent}_id INTEGER NOT NULL,
                        changed_at TEXT DEFAULT (datetime('now','localtime')),
                        changed_by INTEGER,
                        field TEXT,
                        old_value TEXT,
                        new_value TEXT,
                        note TEXT
                    )
                """)
                c.execute(
                    f"CREATE INDEX IF NOT EXISTS idx_{_ent}_history_eid "
                    f"ON {_ent}_history({_ent}_id)"
                )
            except Exception:
                pass

        # v5H89b: 기존 orders.customer_id 가 NULL 인데 project 가 customer_name
        # 또는 customer_id 를 갖고 있으면 backfill (수주관리 리스트 고객사 표시 회복)
        try:
            # 1) projects.customer_id 가 비었지만 customer_name 매칭되는 customers 가 있으면 채움
            c.execute(
                "UPDATE projects SET customer_id = ("
                "  SELECT id FROM customers WHERE customers.name = projects.customer_name LIMIT 1"
                ") "
                "WHERE customer_id IS NULL AND COALESCE(customer_name,'') <> ''"
            )
            # 2) orders.customer_id NULL → project 의 customer_id 로 채움
            c.execute(
                "UPDATE orders SET customer_id = ("
                "  SELECT customer_id FROM projects WHERE projects.id = orders.project_id"
                ") "
                "WHERE customer_id IS NULL AND project_id IS NOT NULL"
            )
        except Exception:
            pass

        # v5H226z474 (대표 지시): 상태 용어 통일 — '납품완료' → '출하' (기존 저장 데이터 일괄 변환·idempotent).
        #   호기(order_items.unit_status)·프로젝트(projects.status/stage)에 '납품완료'로 저장된 기존 행을 '출하'로.
        #   (소모품 consumable_orders·SO orders 는 영어값(SHIPPED 등) 저장이라 대상 아님 / 변경이력 로그는 과거기록이라 보존)
        try:
            _z474n = 0
            for _tbl, _col in (("order_items", "unit_status"),
                               ("projects", "status"), ("projects", "stage")):
                try:
                    _rz = c.execute(f"UPDATE {_tbl} SET {_col}='출하' WHERE {_col}='납품완료'")
                    _z474n += (_rz.rowcount or 0)
                except Exception:
                    pass
            if _z474n:
                print(f"[v5H226z474] 상태 통일: '납품완료'→'출하' {_z474n}행 변환")
        except Exception as _e:
            print(f"[v5H226z474] 납품완료→출하 통일 마이그레이션 실패: {_e}")

        # v5H226t: 상태 cascade 자가치유 백필
        # (1) 모든 호기가 동일 terminal 상태인데 부모 프로젝트가 다르면 → 부모 갱신 (상방)
        # (2) 부모가 terminal 인데 자식 SO/호기가 동기 안 됐으면 → 자식 갱신 (하방)
        try:
            from . import project_workflow as _pwf_csc
            # (1) 상방 backfill — 모든 프로젝트에 대해 호기 종합 검사
            pids = [r[0] for r in c.execute(
                "SELECT DISTINCT id FROM projects WHERE id IN "
                "  (SELECT DISTINCT project_id FROM orders WHERE project_id IS NOT NULL)"
            ).fetchall()]
            up_promoted = 0
            for _pid in pids:
                _r = _pwf_csc.cascade_unit_status_to_project(c, _pid, None)
                if _r.get("changed"):
                    up_promoted += 1
            # (2) 하방 backfill — 프로젝트가 terminal 인데 자식 미반영
            for _pid, _st in c.execute(
                "SELECT id, status FROM projects WHERE status IN ('진행중','출하','취소','보류')"
            ).fetchall():
                _pwf_csc.cascade_project_status_to_so(c, _pid, _st, None)
            if up_promoted:
                print(f"[v5H226t] cascade 백필: 프로젝트 {up_promoted}건 상태 자동 승급")
        except Exception as _e:
            print(f"[v5H226t] cascade 백필 실패: {_e}")


def seed_all():
    """최초 1회: 조직도/사용자/고객사 시드"""
    with db_session() as c:
        # 이미 팀 있으면 skip
        cnt = c.execute("SELECT COUNT(*) FROM teams").fetchone()[0]
        if cnt > 0:
            return

        # 1) Teams
        for code, name, is_lab, sector, order in TEAMS:
            c.execute(
                "INSERT INTO teams(code, name, is_lab, sector, display_order) VALUES(?,?,?,?,?)",
                (code, name, is_lab, sector, order),
            )

        # 2) Admin
        c.execute(
            "INSERT INTO users(name, login_id, password, rank, role) VALUES(?,?,?,?,?)",
            ("시스템관리자", "admin", hash_pw("admin1234"), "관리자", "admin"),
        )

        # 3) Users
        dup: dict = {}
        # 대표님 두 분은 login_id를 명시적으로 부여
        special = {"김동후": "kdh", "김정락": "kjr"}

        for name, team_code, rank, role in USERS:
            if name in special:
                lid = special[name]
            else:
                lid = make_login_id(name, dup)
            team_id = None
            if team_code:
                t = c.execute("SELECT id FROM teams WHERE code=?", (team_code,)).fetchone()
                team_id = t["id"] if t else None
            c.execute(
                "INSERT INTO users(name, login_id, password, team_id, rank, role) VALUES(?,?,?,?,?,?)",
                (name, lid, hash_pw("knk1234"), team_id, rank, role),
            )

        # 4) Team leaders 지정
        # 팀 코드별 "leader" 또는 "executive"인 인원을 leader로 설정
        for code, name, *_ in TEAMS:
            t = c.execute("SELECT id FROM teams WHERE code=?", (code,)).fetchone()
            if not t:
                continue
            # 우선 leader 역할 우선, 없으면 executive 중 해당 팀 배치
            ld = c.execute(
                "SELECT id FROM users WHERE team_id=? AND role='leader' ORDER BY id LIMIT 1",
                (t["id"],),
            ).fetchone()
            if not ld:
                ld = c.execute(
                    "SELECT id FROM users WHERE team_id=? AND role='executive' ORDER BY id LIMIT 1",
                    (t["id"],),
                ).fetchone()
            if ld:
                c.execute("UPDATE teams SET leader_id=? WHERE id=?", (ld["id"], t["id"]))

        # 4-ESC-04) 구매팀장 물류 모듈 초기 권한 (2026-04-22 대표 결재 D01-NEW-PERM — 구매팀장 부여 방식)
        purch_team = c.execute("SELECT id FROM teams WHERE code=?", ("10",)).fetchone()
        if purch_team:
            c.execute(
                "UPDATE users SET can_use_logistics=1 WHERE team_id=? AND role='leader'",
                (purch_team["id"],),
            )

        # (Plan Y S1 권한 시드는 마이그레이션 블록으로 이동 — 매 init idempotent 적용)

        # 5) Customers
        for cname, tier, note in CUSTOMERS:
            c.execute(
                "INSERT INTO customers(name, tier, note) VALUES(?,?,?)",
                (cname, tier, note),
            )

        # 6) Sample Projects
        samples = [
            ("P-2026-001", "제품1 메인보드 ICT 검사기",       "고객사A",  "검사기"),
            ("P-2026-002", "제품1 FCT 라인 검사기",           "고객사A",  "검사기"),
            ("P-2026-003", "전장 제어보드 자동조립 라인",   "고객사B",  "자동화"),
            ("P-2026-004", "카메라 모듈 검사장비",          "드림텍",    "검사기"),
            ("P-2026-005", "EV 배터리 BMS 검사장비",        "한국성전",  "검사기"),
            ("P-2026-006", "사내 자동화 통합 MES",           "내부",      "자동화"),
        ]
        for code, pname, cust, ptype in samples:
            cu = c.execute("SELECT id FROM customers WHERE name=?", (cust,)).fetchone()
            c.execute(
                "INSERT INTO projects(code, name, customer_id, type, status) VALUES(?,?,?,?,?)",
                (code, pname, cu["id"] if cu else None, ptype, "진행중"),
            )

        # P1-1 (2026-04-25 09팀장 결정 §2-2 채택 · 사용자 지시 가상거래처 적용):
        #   시연용 더미 매출 3건 — CEO /home 매출 KPI 활성화
        #   - 외부 브랜드명 0건 (자체 가상 거래처)
        #   - 매출액 ₩50M~₩200M 합리 범위
        #   - 최근 3개월 분산 (당월·당월·전월)
        import datetime as _dt
        _today = _dt.date.today()
        _ym = _today.strftime("%Y-%m")
        _prev_ym = (_today.replace(day=1) - _dt.timedelta(days=1)).strftime("%Y-%m")
        _demo_customers = [
            ("한빛전자",   "B", "[가상·시연용] 더미 거래처"),
            ("미래공업",   "B", "[가상·시연용] 더미 거래처"),
            ("신한정밀",   "C", "[가상·시연용] 더미 거래처"),
        ]
        for _cn, _tier, _note in _demo_customers:
            _exist = c.execute("SELECT id FROM customers WHERE name=?", (_cn,)).fetchone()
            if not _exist:
                c.execute(
                    "INSERT INTO customers(name, tier, note) VALUES(?,?,?)",
                    (_cn, _tier, _note),
                )
        _demo_projects = [
            ("P-DEMO-001", "한빛전자 ICT 검사기 1차",  "한빛전자",  "검사기", 180_000_000, f"{_ym}-05"),
            ("P-DEMO-002", "미래공업 자동화 라인",     "미래공업",  "자동화", 120_000_000, f"{_ym}-12"),
            ("P-DEMO-003", "신한정밀 검사 지그",       "신한정밀",  "검사기",  65_000_000, f"{_prev_ym}-22"),
        ]
        for _code, _pn, _cust, _pt, _amt, _dt_ in _demo_projects:
            _cu = c.execute("SELECT id FROM customers WHERE name=?", (_cust,)).fetchone()
            _exp = c.execute("SELECT id FROM projects WHERE code=?", (_code,)).fetchone()
            if not _exp:
                c.execute(
                    "INSERT INTO projects(code, name, customer_id, type, status, "
                    "order_amount, order_date) VALUES(?,?,?,?,?,?,?)",
                    (_code, _pn, _cu["id"] if _cu else None, _pt, "진행중", _amt, _dt_),
                )


# =====================================================
# SAMPLE DATA GENERATOR (지난 14일치 현실적 데이터)
# =====================================================
import random
from datetime import date as _date, timedelta as _td

# 팀별 현실적인 업무 템플릿 (KNK 실제 업무 반영)
TASK_TEMPLATES = {
    "01": [  # 기술영업팀
        ("{cu} {pj} 견적 제출", "고객대응", "검사기"),
        ("{cu} 신규 BOM 검토 미팅", "고객대응", "공통"),
        ("{cu} 방문 · 사양 협의", "출장", "공통"),
        ("{cu} 발주 진행 체크", "고객대응", "공통"),
        ("{cu} 대응 이슈 회신", "고객대응", "공통"),
        ("{pj} 사양서 전달", "고객대응", "공통"),
        ("주간 수주 현황 정리", "검토", "공통"),
        ("영업 파이프라인 업데이트", "검토", "공통"),
    ],
    "02": [  # 검사기팀
        ("{pj} 1차 셋업 점검", "설계", "검사기"),
        ("{pj} HW 수정 요청 대응", "설계", "검사기"),
        ("{pj} 시제품 전기 검사", "품질", "검사기"),
        ("{pj} 지그 수정 설계", "설계", "검사기"),
        ("{pj} 양산 이관 준비", "제조", "검사기"),
        ("{pj} 회로도 검토", "설계", "검사기"),
        ("검사기 표준 프로세스 개선", "검토", "검사기"),
    ],
    "03": [  # 품질팀
        ("{cu} FCT 라인 품질 이슈 대응", "품질", "검사기"),
        ("{pj} 불량률 집계 · 리포트", "품질", "공통"),
        ("수입검사 샘플링 기준 재검토", "품질", "공통"),
        ("품질 미팅 (일일)", "회의", "공통"),
        ("출하 검사 기록 정리", "품질", "공통"),
        ("{cu} 클레임 대응", "고객대응", "공통"),
    ],
    "04": [  # 설계팀
        ("{pj} 메커니컬 3D 도면", "설계", "공통"),
        ("{pj} 지그 설계 수정", "설계", "공통"),
        ("{pj} 2D 도면 출도", "설계", "공통"),
        ("{pj} 설계 검증 회의", "회의", "공통"),
        ("{pj} 부품 BOM 확정", "설계", "공통"),
        ("{pj} 기존 도면 리비전", "설계", "공통"),
        ("설계 Lesson-Learned 정리", "검토", "공통"),
    ],
    "05": [  # 소프트웨어팀
        ("{pj} 펌웨어 v2.x 개발", "개발", "공통"),
        ("{pj} 시퀀스 스크립트 작성", "개발", "공통"),
        ("{pj} 비전 알고리즘 튜닝", "개발", "공통"),
        ("{pj} 통신 프로토콜 구현", "개발", "공통"),
        ("{pj} UI 화면 개발", "개발", "공통"),
        ("{pj} 현장 디버깅", "개발", "공통"),
        ("{pj} 리포트 기능 추가", "개발", "공통"),
        ("{pj} 통합 테스트", "개발", "공통"),
    ],
    "06": [  # 전장설계팀
        ("{pj} 전장 회로 설계", "설계", "자동화"),
        ("{pj} 배전반 부품 선정", "설계", "자동화"),
        ("{pj} PLC 로직 작성", "설계", "자동화"),
        ("{pj} 전장 도면 출도", "설계", "자동화"),
        ("전장 표준품 정리", "검토", "자동화"),
    ],
    "07": [  # 제조기술1팀
        ("{pj} 검사기 본체 조립", "제조", "검사기"),
        ("{pj} 지그 가공/조립", "제조", "검사기"),
        ("{pj} 양산 이관 대응", "제조", "검사기"),
        ("{pj} 생산 일정 조율", "제조", "검사기"),
        ("일일 제조 미팅", "회의", "공통"),
    ],
    "08": [  # 제조기술2팀
        ("{pj} 자동화 장비 조립", "제조", "자동화"),
        ("{pj} 벨트 얼라인 셋업", "제조", "자동화"),
        ("{pj} 장비 반입 · 현장 셋업", "출장", "자동화"),
        ("{pj} 양산 안정화 작업", "제조", "자동화"),
        ("{pj} 고객 인수 테스트 대응", "제조", "자동화"),
    ],
    "09": [  # 가공팀
        ("{pj} 기구부 가공", "제조", "공통"),
        ("{pj} 지그 부품 가공", "제조", "공통"),
        ("일일 CNC 가동 점검", "제조", "공통"),
        ("가공 스케줄 재편성", "제조", "공통"),
    ],
    "10": [  # 구매팀
        ("{cu} 부품 발주 진행", "구매", "공통"),
        ("협력사 단가 협상", "구매", "공통"),
        ("{pj} 자재 입고 확인", "구매", "공통"),
        ("신규 협력사 등록 심사", "구매", "공통"),
        ("{pj} 긴급 자재 수배", "구매", "공통"),
        ("구매 원가 분석", "검토", "공통"),
    ],
    "11": [  # 관리팀
        ("급여 데이터 정리", "기타", "공통"),
        ("법인 세무 자료 제출", "기타", "공통"),
        ("4대보험 신고", "기타", "공통"),
        ("신입 입사 절차 처리", "기타", "공통"),
        ("경비 정산 승인", "기타", "공통"),
        ("월간 손익 마감 작업", "검토", "공통"),
    ],
    "12": [  # 베트남법인
        ("하노이 현지 {pj} 조립 진행", "제조", "공통"),
        ("본사 자재 입고 확인", "제조", "공통"),
        ("베트남 현지 품질 이슈 대응", "품질", "공통"),
        ("현지 협력사 미팅", "고객대응", "공통"),
        ("본사 기술지원 요청", "기타", "공통"),
    ],
    "13": [  # 개발혁신팀
        ("신규 검사기술 POC 개발", "개발", "검사기"),
        ("차세대 비전 알고리즘 R&D", "개발", "검사기"),
        ("특허 아이디어 검토", "검토", "검사기"),
        ("연구개발 리뷰 미팅", "회의", "검사기"),
    ],
    "14": [  # 라이프밸류팀
        ("LVU 신제품 시장조사", "검토", "공통"),
        ("LVU 샘플 사용 테스트", "검토", "공통"),
        ("LVU 브랜딩 검토", "검토", "공통"),
    ],
}

TEAM_HEADLINES = {
    "01": ["고객사A 제품1 견적 1차 완료, 2차 금주 중 회신 예정", "드림텍 신규 카메라 프로젝트 착수, 사양 협의 중", "금주 견적 5건 완료, 수주 파이프라인 정상"],
    "02": ["P001 ICT 시제품 전기 검사 마무리 단계, 양산 이관 준비", "FCT 라인 지그 수정 완료, 양산 이관 D-3", "BMS 검사기 1차 회로도 완성, 설계 검증 진행"],
    "03": ["고객사A 클레임 1건 대응 완료, 재발 방지 보고서 작성 중", "품질 미팅 정례화, 불량률 1.2% → 0.8% 개선", "FCT 라인 불량률 분석 진행 중, 금주 내 결과 보고"],
    "04": ["P001 메커니컬 도면 80% 완료, 금주 출도 목표", "BMS 검사장비 3D 모델링 진행 중", "자동화 라인 지그 설계 수정 완료"],
    "05": ["ICT 펌웨어 v2.3 릴리즈 완료, 현장 적용 중", "BMS SW 통합 테스트 80%, 금주 완료 예정", "비전 알고리즘 정확도 98.7% 달성"],
    "06": ["자동조립 라인 PLC 로직 작성 완료, 시뮬레이션 단계", "전장 회로 검토 회의 완료, 수정사항 반영 중", "신규 전장 표준화 작업 착수"],
    "07": ["P001 검사기 본체 5대 조립 완료, 금주 출하 예정", "지그 가공 스케줄 정상, 납기 이상 없음", "양산 이관 안정화 진행 중"],
    "08": ["자동조립 라인 셋업 진행률 70%, 현장 파견 중", "드림텍 장비 인수 테스트 대응", "벨트 얼라인 재조정 완료"],
    "09": ["CNC 가동률 92%, 납기 지연 없음", "지그 부품 가공 정상 진행", "설비 정기점검 완료"],
    "10": ["고객사B 관련 부품 긴급 수배 완료", "신규 협력사 2곳 등록 심사 중", "주요 부품 재고 정상"],
    "11": ["월간 급여 마감 진행 중, 4월 10일 완료 예정", "신입 2명 입사 절차 진행", "손익 마감 자료 취합 중"],
    "12": ["하노이 현지 조립 2대 완료, 본사 자재 대기 1건", "현지 품질 이슈 대응 완료", "본사 기술지원 1건 요청"],
    "13": ["차세대 비전 알고리즘 POC 1차 완료", "특허 출원 2건 검토 중", "신규 검사 기법 사내 공유 완료"],
    "14": ["LVU 신규 컨셉 3종 정리 완료, 대표이사 보고 예정", "브랜딩 시안 1차 검토", "샘플 사용자 피드백 수집 중"],
}


def seed_sample_tasks(days_back: int = 14):
    """지난 N일치 현실적인 Task 샘플 시드"""
    random.seed(42)  # 재현 가능
    with db_session() as c:
        # v5H226z135: 운영 모드(데모 초기화 후) — 데모 샘플 task 재생성 차단
        try:
            _r = c.execute("SELECT value FROM app_settings WHERE key='demo_seed_off'").fetchone()
            if _r and str(_r[0]) == '1':
                return 0
        except Exception:
            pass
        # 이미 tasks 있으면 skip
        cnt = c.execute("SELECT COUNT(*) FROM tasks").fetchone()[0]
        if cnt > 5:
            return 0

        # 전 사용자
        users = [dict(r) for r in c.execute(
            """SELECT u.id, u.name, u.rank, u.role, t.code AS team_code, t.id AS team_id
               FROM users u LEFT JOIN teams t ON u.team_id=t.id
               WHERE u.is_active=1 AND u.role!='admin' AND u.team_id IS NOT NULL"""
        ).fetchall()]

        # 프로젝트·고객사 맵
        projects = [dict(r) for r in c.execute("SELECT id, name, customer_id, type FROM projects").fetchall()]
        customers = [dict(r) for r in c.execute("SELECT id, name FROM customers").fetchall()]
        pj_names = {p["id"]: p["name"] for p in projects}
        cu_names = {p["id"]: p.get("customer_id") for p in projects}

        today = _date.today()
        total = 0

        for dback in range(days_back, -1, -1):
            d = today - _td(days=dback)
            wd = d.weekday()  # 0=월
            # 토/일은 카드 거의 없음
            if wd == 5:  # 토
                if random.random() > 0.15:
                    continue
            if wd == 6:  # 일
                continue

            for u in users:
                tc = u["team_code"]
                if tc not in TASK_TEMPLATES:
                    continue
                # 70% 확률로 기록, 관리자급은 덜
                if u["role"] in ("ceo", "executive") and random.random() > 0.55:
                    continue
                if u["role"] == "member" and random.random() > 0.85:
                    continue

                n_tasks = random.choice([2, 3, 3, 3, 4, 4, 5])
                templates = TASK_TEMPLATES[tc]
                for _ in range(n_tasks):
                    title_tpl, cat, sec = random.choice(templates)
                    # 프로젝트 선택 (팀 sector 유사)
                    candidate_projects = [p for p in projects if sec == "공통" or p["type"] == sec]
                    pj = random.choice(candidate_projects) if candidate_projects and random.random() > 0.2 else None
                    cu_id = pj["customer_id"] if pj else (random.choice(customers)["id"] if random.random() > 0.5 else None)
                    cu_name = next((cc["name"] for cc in customers if cc["id"] == cu_id), "") if cu_id else ""
                    pj_name = pj["name"].split()[0] if pj else "사내"
                    title = title_tpl.replace("{cu}", cu_name or "내부").replace("{pj}", pj_name)

                    # 상태: 최근일수록 '완료' 비율 낮음
                    if dback == 0:  # 오늘
                        status = random.choices(["진행중", "완료", "지연", "대기"], weights=[55, 25, 12, 8])[0]
                    elif dback <= 2:
                        status = random.choices(["진행중", "완료", "지연", "대기"], weights=[35, 45, 12, 8])[0]
                    else:
                        status = random.choices(["완료", "진행중", "지연"], weights=[75, 15, 10])[0]

                    hours = round(random.choice([0.5, 1, 1, 1.5, 2, 2, 2.5, 3, 3, 4]), 1)
                    notes = ""
                    if random.random() > 0.8:
                        notes = random.choice([
                            "내부 검토 후 금주 내 완료 예정",
                            "고객사 피드백 대기 중",
                            "협력사 일정 조율 필요",
                            "설계 리비전 반영 완료",
                            "현장 이슈 발생, 대응 중",
                            "다음주 확인 예정",
                        ])

                    c.execute(
                        """INSERT INTO tasks(user_id, work_date, title, category, project_id,
                                              customer_id, status, hours, notes)
                           VALUES(?,?,?,?,?,?,?,?,?)""",
                        (u["id"], d.isoformat(), title, cat,
                         pj["id"] if pj else None, cu_id, status, hours, notes),
                    )
                    total += 1

        # 팀장 한 줄 요약 (최근 5일)
        leaders = [dict(r) for r in c.execute(
            """SELECT t.id AS team_id, t.code, u.id AS leader_id
               FROM teams t JOIN users u ON t.leader_id=u.id"""
        ).fetchall()]
        for dback in range(5):
            d = (today - _td(days=dback))
            if d.weekday() >= 5:
                continue
            for l in leaders:
                code = l["code"]
                hl_list = TEAM_HEADLINES.get(code, ["금일 업무 정상 진행"])
                hl = random.choice(hl_list)
                try:
                    c.execute(
                        """INSERT OR IGNORE INTO team_summaries(team_id, work_date, author_id, headline)
                           VALUES(?,?,?,?)""",
                        (l["team_id"], d.isoformat(), l["leader_id"], hl),
                    )
                except Exception:
                    pass

        return total


# =====================================================
# MGMT CODE IMPORT (관리코드발행목록.xls)
# =====================================================
def _norm_customer(name: str) -> str:
    """고객사명 정규화 — 공백/따옴표 제거"""
    if not name:
        return ""
    s = str(name).strip()
    return s


def _clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in ("nan", "none", "null"):
        return ""
    return s


def parse_mgmt_csv(file_path: str):
    """
    관리코드발행목록 CSV 파싱 (사이클 71: pandas 제거 후 표준 csv 모듈 대체)
    입력 CSV 포맷: 첫 컬럼 sheet_tag(M/T), 이후 시트별 원본 컬럼 순서대로
      M: sheet, seq, div, ym, model, machine, _, customer, author, server_path
      T: sheet, seq, div, ym, model, machine, customer, author
    헤더 행 자유 (seq가 숫자가 아니면 자동 skip)
    Returns: list of dict { mgmt_code, equip_type, year_month, model_name,
                            name, customer_name, author_name, server_path }
    """
    import csv as _csv
    rows = []
    with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = _csv.reader(f)
        for raw in reader:
            if not raw or len(raw) < 5:
                continue
            sheet_name = _clean(raw[0])
            if sheet_name not in ("M", "T"):
                continue
            equip_type = "자동화" if sheet_name == "M" else "검사기"
            seq = _clean(raw[1]) if len(raw) > 1 else ""
            div = _clean(raw[2]) if len(raw) > 2 else ""
            ym = _clean(raw[3]) if len(raw) > 3 else ""
            if not seq or not div or not ym:
                continue
            if div not in ("M", "T"):
                continue
            if ym.endswith(".0"):
                ym = ym[:-2]
            if len(ym) != 4 or not ym.isdigit():
                continue
            try:
                seq_int = int(float(seq))
                seq_str = f"{seq_int:03d}"
            except Exception:
                continue
            mgmt_code = f"{seq_str}{div}{ym}"
            model = _clean(raw[4]) if len(raw) > 4 else ""
            if sheet_name == "M":
                machine = _clean(raw[5]) if len(raw) > 5 else ""
                customer = _norm_customer(_clean(raw[7]) if len(raw) > 7 else "")
                author = _clean(raw[8]) if len(raw) > 8 else ""
                server_path = _clean(raw[9]) if len(raw) > 9 else ""
            else:  # T
                machine = _clean(raw[5]) if len(raw) > 5 else ""
                customer = _norm_customer(_clean(raw[6]) if len(raw) > 6 else "")
                author = _clean(raw[7]) if len(raw) > 7 else ""
                server_path = ""
            if not machine and not model:
                continue
            pj_name = machine or model
            year_month = f"20{ym[:2]}-{ym[2:]}"
            rows.append({
                "mgmt_code": mgmt_code,
                "equip_type": equip_type,
                "div": div,
                "year_month": year_month,
                "model_name": model,
                "name": pj_name,
                "customer_name": customer,
                "author_name": author,
                "server_path": server_path,
            })
    return rows


def parse_mgmt_xls(file_path: str):
    """
    [DEPRECATED 사이클 71 — pandas 제거 결정 2026-04-27]
    구 .xls/.xlsx 파일 직접 파싱은 외부 의존(pandas/openpyxl) 필요로 폐기됨.
    1회성 마이그레이션이므로 사용자가 .xls 를 .csv 로 외부 변환 후
    parse_mgmt_csv(file_path) 사용 또는 import 라우트가 .csv 업로드 안내.
    """
    raise NotImplementedError(
        "관리코드 .xls 직접 파싱은 사이클 71에서 폐기되었습니다 (pandas 제거 결정). "
        ".xls 파일을 Excel 또는 LibreOffice에서 .csv (UTF-8) 로 저장 후 업로드해 주세요. "
        "CSV 컬럼 순서: sheet_tag(M/T), seq, div, ym, model, machine, [customer/_], "
        "[author/customer], [server_path/author], [server_path]"
    )


def import_mgmt_rows(rows: list) -> dict:
    """
    파싱된 rows를 DB에 upsert. 고객사·담당자 매칭 포함.
    Returns: {inserted, updated, customers_created, skipped, errors}
    """
    result = {"inserted": 0, "updated": 0, "customers_created": 0, "skipped": 0, "errors": 0}
    with db_session() as c:
        for r in rows:
            try:
                # 고객사 upsert
                cu_id = None
                if r["customer_name"]:
                    cu = c.execute(
                        "SELECT id FROM customers WHERE name=?", (r["customer_name"],)
                    ).fetchone()
                    if cu:
                        cu_id = cu["id"]
                    else:
                        cur = c.execute(
                            "INSERT INTO customers(name, tier, note) VALUES(?,?,?)",
                            (r["customer_name"], "일반", "자동임포트"),
                        )
                        cu_id = cur.lastrowid
                        result["customers_created"] += 1
                # 담당자 매칭
                lead_id = None
                if r["author_name"]:
                    ur = c.execute(
                        "SELECT id FROM users WHERE name=? AND is_active=1 LIMIT 1",
                        (r["author_name"],),
                    ).fetchone()
                    if ur:
                        lead_id = ur["id"]
                # 프로젝트 upsert (mgmt_code 기준)
                ex = c.execute(
                    "SELECT id FROM projects WHERE mgmt_code=?", (r["mgmt_code"],)
                ).fetchone()
                start_date = f"{r['year_month']}-01"
                if ex:
                    c.execute(
                        """UPDATE projects SET
                               name=?, customer_id=?, type=?, equip_type=?,
                               year_month=?, model_name=?, server_path=?,
                               lead_user_id=?, start_date=?
                           WHERE id=?""",
                        (r["name"], cu_id, r["equip_type"], r["equip_type"],
                         r["year_month"], r["model_name"], r["server_path"],
                         lead_id, start_date, ex["id"]),
                    )
                    result["updated"] += 1
                else:
                    c.execute(
                        """INSERT INTO projects
                           (code, name, customer_id, type, status, start_date,
                            mgmt_code, equip_type, year_month, model_name,
                            server_path, lead_user_id)
                           VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (r["mgmt_code"], r["name"], cu_id, r["equip_type"],
                         "진행중", start_date,
                         r["mgmt_code"], r["equip_type"], r["year_month"],
                         r["model_name"], r["server_path"], lead_id),
                    )
                    result["inserted"] += 1
            except Exception as e:
                result["errors"] += 1
    return result


# =====================================================
# INITIAL PASSWORD DEPLOYMENT
# =====================================================
def regenerate_user_passwords(exclude_logins=("admin",)):
    """
    모든 활성 사용자의 비밀번호를 새로 생성하여 DB에 반영하고 목록을 반환.
    Returns: list of dict {team_name, team_code, name, rank, login_id, password, role}
    """
    import random, string
    rows = []
    with db_session() as c:
        users = c.execute(
            """SELECT u.id, u.name, u.login_id, u.rank, u.role,
                      t.name AS team_name, t.code AS team_code, t.display_order
               FROM users u LEFT JOIN teams t ON u.team_id=t.id
               WHERE u.is_active=1 ORDER BY t.display_order, u.id"""
        ).fetchall()
        for u in users:
            if u["login_id"] in exclude_logins:
                continue
            # knk + 4자리 숫자 (외우기 쉽게)
            pw = "knk" + "".join(random.choices(string.digits, k=4))
            c.execute("UPDATE users SET password=? WHERE id=?",
                      (hash_pw(pw), u["id"]))
            rows.append({
                "team_name": u["team_name"] or "(미배정)",
                "team_code": u["team_code"] or "",
                "name": u["name"],
                "rank": u["rank"] or "",
                "login_id": u["login_id"],
                "password": pw,
                "role": u["role"],
            })
    return rows


def build_password_csv(rows, out_path):
    """
    비밀번호 배포용 CSV 생성 (UTF-8 BOM · Excel 한글 호환)
    사이클 68 (2026-04-27): openpyxl 제거 → csv 표준 모듈 (대표 결정 이행).
    Excel/구글시트에서 .csv 열면 자동 표 형식 — 사용자 영향 거의 없음.
    """
    import csv as _csv
    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        # 첫 행 메타 안내 (Excel에서 1행에 표기되어도 무방)
        f.write("# ㈜케이엔케이 일일업무일지 · 초기 비밀번호 배포용\n")
        f.write("# 첫 로그인 후 반드시 [비밀번호 변경] 메뉴에서 개인 비밀번호로 변경해 주세요. · HAIST Innovation\n")
        w = _csv.writer(f)
        w.writerow(["No", "팀", "이름", "직급", "로그인ID", "초기 비밀번호", "권한"])
        for idx, r in enumerate(rows, 1):
            w.writerow([idx, r["team_name"], r["name"], r["rank"],
                        r["login_id"], r["password"], r["role"]])
    return out_path


# 사이클 69 (2026-04-27): 호환 alias `build_password_xlsx = build_password_csv` 제거.
# 호출자 0건 (grep 검증) — main.py:15·2794는 이미 build_password_csv 직접 호출.

# =====================================================
# COMMENTS & NOTIFICATIONS
# =====================================================
def _parse_mentions(body, c):
    """본문에서 @이름 패턴을 찾아 user id 매핑"""
    import re
    names = set(re.findall(r"@([가-힣A-Za-z0-9_]{2,20})", body or ""))
    if not names:
        return []
    placeholders = ",".join("?" * len(names))
    rows = c.execute(
        f"SELECT id, name FROM users WHERE name IN ({placeholders}) AND is_active=1",
        tuple(names),
    ).fetchall()
    return [(r["id"], r["name"]) for r in rows]


def add_comment(task_id, author_id, body, parent_id=None):
    """댓글 추가 + @멘션 파싱 + 알림 + 활동 로그"""
    with db_session() as c:
        author = c.execute("SELECT role, name, rank, team_id FROM users WHERE id=?", (author_id,)).fetchone()
        is_ceo = 1 if author and author["role"] == "ceo" else 0
        cur = c.execute(
            """INSERT INTO task_comments(task_id, author_id, body, is_ceo_request, parent_id)
               VALUES(?,?,?,?,?)""",
            (task_id, author_id, body.strip(), is_ceo, parent_id),
        )
        comment_id = cur.lastrowid

        task = c.execute(
            """SELECT t.user_id, t.title, t.project_id, u.team_id
               FROM tasks t LEFT JOIN users u ON t.user_id = u.id
               WHERE t.id=?""", (task_id,)
        ).fetchone()
        if not task:
            return comment_id

        recipients = set()
        if task["user_id"] != author_id:
            recipients.add(task["user_id"])
        if parent_id:
            p = c.execute("SELECT author_id FROM task_comments WHERE id=?", (parent_id,)).fetchone()
            if p and p["author_id"] != author_id:
                recipients.add(p["author_id"])

        # @멘션 처리
        mentions = _parse_mentions(body, c)
        mention_uids = set()
        for mid, mname in mentions:
            c.execute("INSERT OR IGNORE INTO comment_mentions(comment_id, user_id) VALUES(?,?)",
                      (comment_id, mid))
            if mid != author_id:
                recipients.add(mid)
                mention_uids.add(mid)

        prefix = "[대표 요청] " if is_ceo else ""
        a_name = author["name"]
        a_rank = author["rank"] or ""
        title = prefix + a_name + " " + a_rank + " 님이 댓글을 남겼습니다"
        snippet = body.strip()[:120]
        link = f"/daily?focus={task_id}#task-{task_id}"
        for uid in recipients:
            kind = "mention" if uid in mention_uids else "comment"
            ttl = ("[멘션] " if kind == "mention" else "") + title
            c.execute(
                """INSERT INTO notifications(user_id, kind, title, body, link, task_id, comment_id)
                   VALUES(?,?,?,?,?,?,?)""",
                (uid, kind, ttl, snippet, link, task_id, comment_id),
            )

        # 활동 로그
        log_activity(c, author_id, "comment",
                     title=f"{a_name} → {task['title'][:50]}",
                     body=snippet,
                     task_id=task_id,
                     project_id=task["project_id"],
                     team_id=task["team_id"])
        return comment_id


# =====================================================
# 활동 로그 / 반응 / 회고
# =====================================================
def log_activity(c, actor_id, kind, title, body=None, task_id=None, project_id=None, team_id=None, meta=None):
    """db_session 이미 열려 있는 상태에서 호출"""
    import json as _json
    c.execute(
        """INSERT INTO activities(actor_id, kind, task_id, project_id, team_id, title, body, meta)
           VALUES(?,?,?,?,?,?,?,?)""",
        (actor_id, kind, task_id, project_id, team_id, title[:200], body, _json.dumps(meta or {}))
    )


def log_activity_standalone(actor_id, kind, title, body=None, task_id=None, project_id=None, team_id=None, meta=None):
    with db_session() as c:
        log_activity(c, actor_id, kind, title, body, task_id, project_id, team_id, meta)


def get_activities(limit=80, team_id=None, actor_id=None, since_id=0):
    with db_session() as c:
        sql = """SELECT a.*, u.name AS actor_name, u.rank AS actor_rank,
                        t.title AS task_title, p.name AS project_name, tm.name AS team_name
                 FROM activities a
                 LEFT JOIN users u ON a.actor_id=u.id
                 LEFT JOIN tasks t ON a.task_id=t.id
                 LEFT JOIN projects p ON a.project_id=p.id
                 LEFT JOIN teams tm ON a.team_id=tm.id
                 WHERE a.id > ?"""
        params = [since_id]
        if team_id:
            sql += " AND a.team_id=?"
            params.append(team_id)
        if actor_id:
            sql += " AND a.actor_id=?"
            params.append(actor_id)
        sql += " ORDER BY a.id DESC LIMIT ?"
        params.append(limit)
        return [dict(r) for r in c.execute(sql, params).fetchall()]


def add_reaction(task_id, user_id, kind):
    """1-click 반응 토글. ack/question/risk/ok"""
    if kind not in ("ack", "question", "risk", "ok"):
        return False
    with db_session() as c:
        existing = c.execute(
            "SELECT id FROM task_reactions WHERE task_id=? AND user_id=? AND kind=?",
            (task_id, user_id, kind)
        ).fetchone()
        if existing:
            c.execute("DELETE FROM task_reactions WHERE id=?", (existing["id"],))
            return "removed"
        c.execute(
            "INSERT INTO task_reactions(task_id, user_id, kind) VALUES(?,?,?)",
            (task_id, user_id, kind)
        )
        # 작성자에게 알림: 대표/임원/팀장 반응은 모든 종류, 직원은 risk/question만
        t = c.execute("SELECT user_id, title FROM tasks WHERE id=?", (task_id,)).fetchone()
        actor = c.execute("SELECT name, rank, role FROM users WHERE id=?", (user_id,)).fetchone()
        if t and t["user_id"] != user_id and actor:
            actor_role = actor["role"] or ""
            is_leader_plus = actor_role in ("ceo", "admin", "executive", "leader")
            if is_leader_plus or kind in ("risk", "question"):
                lab = {"ack":"👍 확인", "question":"❓ 질문", "risk":"⚠️ 리스크", "ok":"✅ OK"}[kind]
                c.execute(
                    """INSERT INTO notifications(user_id, kind, title, body, link, task_id)
                       VALUES(?,?,?,?,?,?)""",
                    (t["user_id"], "reaction",
                     f"{actor['name']} {actor['rank'] or ''} 님이 {lab} 반응",
                     (t["title"] or "")[:120],
                     f"/daily?focus={task_id}#task-{task_id}",
                     task_id)
                )
        return "added"


def get_reactions(task_id):
    with db_session() as c:
        rows = c.execute(
            """SELECT r.kind, u.id AS user_id, u.name, u.rank, u.role
               FROM task_reactions r JOIN users u ON r.user_id=u.id
               WHERE r.task_id=? ORDER BY r.created_at""",
            (task_id,)
        ).fetchall()
        agg = {"ack":[], "question":[], "risk":[], "ok":[]}
        for r in rows:
            agg[r["kind"]].append(dict(r))
        return agg


def get_reactions_bulk(task_ids):
    if not task_ids:
        return {}
    with db_session() as c:
        ph = ",".join("?" * len(task_ids))
        rows = c.execute(
            f"""SELECT task_id, kind, COUNT(*) AS cnt
                FROM task_reactions WHERE task_id IN ({ph})
                GROUP BY task_id, kind""", tuple(task_ids)
        ).fetchall()
        out = {}
        for r in rows:
            out.setdefault(r["task_id"], {})[r["kind"]] = r["cnt"]
        return out


def get_meta_bulk(task_ids):
    """카드 표면에 표시할 메타데이터 (댓글수, 리액션수)를 한 번에 일괄 조회.
       반환: { task_id: {comments:N, ack:N, question:N, risk:N, ok:N, last_comment:'...'} }"""
    if not task_ids:
        return {}
    with db_session() as c:
        ph = ",".join("?" * len(task_ids))
        out = {tid: {"comments": 0, "ack": 0, "question": 0,
                     "risk": 0, "ok": 0, "last_comment": "",
                     "last_comment_at": ""} for tid in task_ids}
        # 댓글 수
        rows = c.execute(
            f"""SELECT task_id, COUNT(*) AS cnt
                FROM task_comments WHERE task_id IN ({ph})
                GROUP BY task_id""", tuple(task_ids)
        ).fetchall()
        for r in rows:
            if r["task_id"] in out:
                out[r["task_id"]]["comments"] = r["cnt"]
        # 마지막 댓글 (미리보기용)
        last_rows = c.execute(
            f"""SELECT tc.task_id, tc.body, tc.created_at, u.name AS user_name
                FROM task_comments tc
                JOIN users u ON tc.author_id=u.id
                WHERE tc.task_id IN ({ph})
                  AND tc.id IN (
                      SELECT MAX(id) FROM task_comments
                      WHERE task_id IN ({ph})
                      GROUP BY task_id
                  )""", tuple(task_ids) + tuple(task_ids)
        ).fetchall()
        for r in last_rows:
            if r["task_id"] in out:
                preview = (r["body"] or "").strip().replace("\n", " ")
                if len(preview) > 40:
                    preview = preview[:40] + "…"
                out[r["task_id"]]["last_comment"] = f"{r['user_name']}: {preview}"
                out[r["task_id"]]["last_comment_at"] = r["created_at"] or ""
        # 리액션 수
        rx_rows = c.execute(
            f"""SELECT task_id, kind, COUNT(*) AS cnt
                FROM task_reactions WHERE task_id IN ({ph})
                GROUP BY task_id, kind""", tuple(task_ids)
        ).fetchall()
        for r in rx_rows:
            if r["task_id"] in out and r["kind"] in out[r["task_id"]]:
                out[r["task_id"]][r["kind"]] = r["cnt"]
        return out


def notify_status_change(task_id, actor_id, old_status, new_status):
    """진행중→지연 등 위험 상태 변경 시 팀장+CEO에게 자동 알림"""
    if new_status != "지연":
        return
    with db_session() as c:
        t = c.execute(
            """SELECT t.title, t.user_id, u.name AS owner_name, u.team_id
               FROM tasks t JOIN users u ON t.user_id=u.id WHERE t.id=?""",
            (task_id,)
        ).fetchone()
        if not t:
            return
        recipients = set()
        # 팀장
        if t["team_id"]:
            leaders = c.execute(
                "SELECT id FROM users WHERE team_id=? AND role IN ('leader','executive') AND is_active=1",
                (t["team_id"],)
            ).fetchall()
            for l in leaders:
                if l["id"] != actor_id:
                    recipients.add(l["id"])
        # CEO
        ceos = c.execute(
            "SELECT id FROM users WHERE role='ceo' AND is_active=1"
        ).fetchall()
        for x in ceos:
            if x["id"] != actor_id:
                recipients.add(x["id"])
        for uid in recipients:
            c.execute(
                """INSERT INTO notifications(user_id, kind, title, body, link, task_id)
                   VALUES(?,?,?,?,?,?)""",
                (uid, "status_delay",
                 f"⚠️ 지연 발생 — {t['owner_name']}",
                 (t["title"] or "")[:120],
                 f"/daily?focus={task_id}#task-{task_id}",
                 task_id)
            )


def get_user_search(q, limit=10):
    """@멘션 자동완성용"""
    with db_session() as c:
        rows = c.execute(
            """SELECT id, name, rank, team_id FROM users
               WHERE is_active=1 AND name LIKE ? ORDER BY name LIMIT ?""",
            (f"{q}%", limit)
        ).fetchall()
        return [dict(r) for r in rows]


def upsert_retro(project_id, author_id, went_well, went_bad, next_action, risk_note):
    with db_session() as c:
        existing = c.execute(
            "SELECT id FROM project_retros WHERE project_id=?", (project_id,)
        ).fetchone()
        if existing:
            c.execute(
                """UPDATE project_retros SET went_well=?, went_bad=?, next_action=?, risk_note=?,
                   author_id=?, updated_at=datetime('now','localtime') WHERE id=?""",
                (went_well, went_bad, next_action, risk_note, author_id, existing["id"])
            )
            return existing["id"]
        cur = c.execute(
            """INSERT INTO project_retros(project_id, author_id, went_well, went_bad, next_action, risk_note)
               VALUES(?,?,?,?,?,?)""",
            (project_id, author_id, went_well, went_bad, next_action, risk_note)
        )
        return cur.lastrowid


def get_retro(project_id):
    with db_session() as c:
        r = c.execute(
            """SELECT pr.*, u.name AS author_name FROM project_retros pr
               LEFT JOIN users u ON pr.author_id=u.id WHERE pr.project_id=?""",
            (project_id,)
        ).fetchone()
        return dict(r) if r else None


def search_all(q, limit=50):
    """카드 + 댓글 + 회고 통합 검색"""
    if not q or len(q.strip()) < 2:
        return {"tasks":[], "comments":[], "retros":[]}
    qq = f"%{q.strip()}%"
    with db_session() as c:
        tasks = c.execute(
            """SELECT t.id, t.title, t.work_date, t.status, t.notes, t.next_plan,
                      u.name AS user_name, p.name AS project_name
               FROM tasks t LEFT JOIN users u ON t.user_id=u.id
               LEFT JOIN projects p ON t.project_id=p.id
               WHERE t.title LIKE ? OR t.notes LIKE ? OR t.next_plan LIKE ?
               ORDER BY t.work_date DESC LIMIT ?""",
            (qq, qq, qq, limit)
        ).fetchall()
        comments = c.execute(
            """SELECT tc.id, tc.body, tc.created_at, tc.task_id,
                      u.name AS author_name, t.title AS task_title
               FROM task_comments tc
               LEFT JOIN users u ON tc.author_id=u.id
               LEFT JOIN tasks t ON tc.task_id=t.id
               WHERE tc.body LIKE ? ORDER BY tc.created_at DESC LIMIT ?""",
            (qq, limit)
        ).fetchall()
        retros = c.execute(
            """SELECT pr.*, p.name AS project_name FROM project_retros pr
               LEFT JOIN projects p ON pr.project_id=p.id
               WHERE pr.went_well LIKE ? OR pr.went_bad LIKE ? OR pr.next_action LIKE ?
               ORDER BY pr.updated_at DESC LIMIT ?""",
            (qq, qq, qq, limit)
        ).fetchall()
        return {
            "tasks":[dict(r) for r in tasks],
            "comments":[dict(r) for r in comments],
            "retros":[dict(r) for r in retros],
        }


def delegate_task(task_id, from_user_id, to_user_id, message=""):
    """업무 위임: from → to, 알림 자동 발송"""
    with db_session() as c:
        c.execute(
            """INSERT INTO task_delegations(task_id, from_user_id, to_user_id, message)
               VALUES(?,?,?,?)""",
            (task_id, from_user_id, to_user_id, message)
        )
        t = c.execute("SELECT title FROM tasks WHERE id=?", (task_id,)).fetchone()
        sender = c.execute("SELECT name, rank FROM users WHERE id=?", (from_user_id,)).fetchone()
        title = (t["title"] if t else "업무")[:80]
        sender_name = f"{sender['name']} {sender['rank'] or ''}".strip() if sender else ""
        # 위임 받는 사람에게 알림
        c.execute(
            """INSERT INTO notifications(user_id, kind, title, body, link, task_id)
               VALUES(?,?,?,?,?,?)""",
            (to_user_id, "delegation",
             f"📌 {sender_name} 님이 업무를 위임했습니다",
             f"[{title}] {message or ''}".strip(),
             f"/daily?focus={task_id}#task-{task_id}",
             task_id)
        )
        # 활동 로그
        log_activity(c, from_user_id, "delegation", f"{sender_name} → 위임", task_id=task_id)
        return True


def get_delegations(task_id):
    with db_session() as c:
        rows = c.execute(
            """SELECT d.*, uf.name AS from_name, uf.rank AS from_rank,
                      ut.name AS to_name, ut.rank AS to_rank
               FROM task_delegations d
               JOIN users uf ON d.from_user_id=uf.id
               JOIN users ut ON d.to_user_id=ut.id
               WHERE d.task_id=? ORDER BY d.created_at DESC""",
            (task_id,)
        ).fetchall()
        return [dict(r) for r in rows]


def resolve_delegation(deleg_id, user_id):
    """위임 완료 처리"""
    with db_session() as c:
        d = c.execute("SELECT * FROM task_delegations WHERE id=?", (deleg_id,)).fetchone()
        if not d or d["to_user_id"] != user_id:
            return False
        c.execute(
            "UPDATE task_delegations SET status='done', resolved_at=datetime('now','localtime') WHERE id=?",
            (deleg_id,)
        )
        # 위임한 사람에게 완료 알림
        t = c.execute("SELECT title FROM tasks WHERE id=?", (d["task_id"],)).fetchone()
        resolver = c.execute("SELECT name, rank FROM users WHERE id=?", (user_id,)).fetchone()
        c.execute(
            """INSERT INTO notifications(user_id, kind, title, body, link, task_id)
               VALUES(?,?,?,?,?,?)""",
            (d["from_user_id"], "delegation",
             f"✅ {resolver['name']} {resolver['rank'] or ''} 님이 위임 업무를 처리했습니다",
             (t["title"] if t else "")[:120],
             f"/daily?focus={d['task_id']}",
             d["task_id"])
        )
        return True


def detect_bottlenecks():
    """병목 자동 탐지: 4가지 규칙"""
    out = []
    with db_session() as c:
        # 1) 지연 3일+
        rows = c.execute(
            """SELECT t.id, t.title, t.work_date, u.name, u.team_id, tm.name AS team_name
               FROM tasks t JOIN users u ON t.user_id=u.id
               LEFT JOIN teams tm ON u.team_id=tm.id
               WHERE t.status='지연' AND date(t.work_date) <= date('now','-3 days')
               ORDER BY t.work_date LIMIT 50"""
        ).fetchall()
        for r in rows:
            out.append({"rule":"지연 3일+", "severity":"high",
                        "title":r["title"], "owner":r["name"],
                        "team":r["team_name"], "task_id":r["id"]})
        # 2) 댓글 5+ 미해결 (미완료 카드)
        rows = c.execute(
            """SELECT t.id, t.title, u.name, tm.name AS team_name,
                      (SELECT COUNT(*) FROM task_comments WHERE task_id=t.id) AS cn
               FROM tasks t JOIN users u ON t.user_id=u.id
               LEFT JOIN teams tm ON u.team_id=tm.id
               WHERE t.status != '완료'
               GROUP BY t.id HAVING cn >= 5 ORDER BY cn DESC LIMIT 30"""
        ).fetchall()
        for r in rows:
            out.append({"rule":f"댓글 {r['cn']}건 미해결", "severity":"mid",
                        "title":r["title"], "owner":r["name"],
                        "team":r["team_name"], "task_id":r["id"]})
        # 3) 한 사람 동시 8개+ 진행중
        rows = c.execute(
            """SELECT u.id, u.name, tm.name AS team_name,
                      COUNT(*) AS cn
               FROM tasks t JOIN users u ON t.user_id=u.id
               LEFT JOIN teams tm ON u.team_id=tm.id
               WHERE t.status='진행중' AND date(t.work_date) >= date('now','-7 days')
               GROUP BY u.id HAVING cn >= 8 ORDER BY cn DESC LIMIT 20"""
        ).fetchall()
        for r in rows:
            out.append({"rule":f"동시 진행 {r['cn']}건 — 과부하", "severity":"mid",
                        "title":f"{r['name']} 업무 과집중", "owner":r["name"],
                        "team":r["team_name"], "task_id":None})
        # 4) 대표 코멘트 24h 무답변
        rows = c.execute(
            """SELECT tc.task_id, tc.body, tc.created_at, t.title, u.name AS owner
               FROM task_comments tc
               JOIN tasks t ON tc.task_id=t.id
               JOIN users u ON t.user_id=u.id
               WHERE tc.is_ceo_request=1
                 AND datetime(tc.created_at) <= datetime('now','-1 day')
                 AND NOT EXISTS (
                    SELECT 1 FROM task_comments tc2
                    WHERE tc2.task_id=tc.task_id AND tc2.id > tc.id
                 )
               LIMIT 30"""
        ).fetchall()
        for r in rows:
            out.append({"rule":"대표 요청 24h 무응답", "severity":"high",
                        "title":r["title"], "owner":r["owner"],
                        "team":None, "task_id":r["task_id"]})
    return out


def get_task_comments(task_id):
    with db_session() as c:
        rows = c.execute(
            """SELECT tc.*, u.name AS author_name, u.rank AS author_rank, u.role AS author_role
               FROM task_comments tc
               JOIN users u ON tc.author_id = u.id
               WHERE tc.task_id=?
               ORDER BY tc.created_at""",
            (task_id,),
        ).fetchall()
        return [dict(r) for r in rows]


def delete_comment(comment_id, user_id):
    """본인 댓글만 삭제 가능 (admin/ceo는 모두)"""
    with db_session() as c:
        u = c.execute("SELECT role FROM users WHERE id=?", (user_id,)).fetchone()
        if not u:
            return False
        cm = c.execute("SELECT author_id FROM task_comments WHERE id=?", (comment_id,)).fetchone()
        if not cm:
            return False
        if cm["author_id"] != user_id and u["role"] not in ("ceo", "admin"):
            return False
        c.execute("DELETE FROM task_comments WHERE id=?", (comment_id,))
        c.execute("DELETE FROM notifications WHERE comment_id=?", (comment_id,))
        return True


def get_notifications(user_id, only_unread=False, limit=30):
    with db_session() as c:
        sql = "SELECT * FROM notifications WHERE user_id=?"
        if only_unread:
            sql += " AND is_read=0"
        sql += " ORDER BY created_at DESC LIMIT ?"
        rows = c.execute(sql, (user_id, limit)).fetchall()
        return [dict(r) for r in rows]


def count_unread(user_id):
    with db_session() as c:
        r = c.execute(
            "SELECT COUNT(*) FROM notifications WHERE user_id=? AND is_read=0",
            (user_id,),
        ).fetchone()
        return r[0] if r else 0


def mark_notification_read(notif_id, user_id):
    with db_session() as c:
        c.execute(
            "UPDATE notifications SET is_read=1 WHERE id=? AND user_id=?",
            (notif_id, user_id),
        )


def mark_all_read(user_id):
    with db_session() as c:
        c.execute(
            "UPDATE notifications SET is_read=1 WHERE user_id=? AND is_read=0",
            (user_id,),
        )


# =====================================================
# 알림시스템 통합 헬퍼 (사이클 2026-04-26)
# 단일 진입점 notify_user — 전사 영역 일관 알림.
# 분류: TASK / SALES / STOCK / QMS / EXPORT / WEEKLY / RATE / PERMISSION / SYSTEM
# 중복 방지: 같은 (user_id, kind, title) 1시간 내 1회.
# v2 본체 무수정 · G1~G5 보존 · 외부 라이브러리 0건.
# =====================================================
NOTIFY_TYPES = {"TASK", "SALES", "STOCK", "QMS", "EXPORT",
                "WEEKLY", "RATE", "PERMISSION", "SYSTEM"}


def notify_user(user_id, type, title, body=None, link=None):
    """전사 통합 알림 헬퍼.
    Args:
        user_id: 수신자 ID
        type: NOTIFY_TYPES 중 하나 (대문자). 미정의 시 SYSTEM.
        title: 알림 제목 (최대 200자 권장)
        body: 본문 (선택)
        link: 클릭 시 이동 경로 (선택)
    Returns:
        True (신규 INSERT) / False (1시간 중복 스킵)
    """
    if not user_id or not title:
        return False
    kind = (type or "SYSTEM").upper()
    if kind not in NOTIFY_TYPES:
        kind = "SYSTEM"
    with db_session() as c:
        # 중복 방지 — 같은 user+kind+title 1시간 내
        dup = c.execute(
            "SELECT 1 FROM notifications "
            "WHERE user_id=? AND kind=? AND title=? "
            "AND created_at >= datetime('now','localtime','-1 hours') "
            "LIMIT 1",
            (user_id, kind, title),
        ).fetchone()
        if dup:
            return False
        c.execute(
            "INSERT INTO notifications(user_id, kind, title, body, link) "
            "VALUES(?,?,?,?,?)",
            (user_id, kind, title, body, link),
        )
        return True


# =====================================================
# HAIST WORKS — 물류 모듈 (parts / 관리코드 발행대장)
# KNK PMS V3 표준 (엑셀정리스킬 §11~12 준수)
# =====================================================
from datetime import datetime as _dt, date as _date

BIZ_CODE = {"검사기": "T", "자동화": "M"}
BIZ_NAME = {v: k for k, v in BIZ_CODE.items()}

STAGES = ["제안작성", "제안제출", "수주확정", "납품"]
NEEDS_CODE_STAGES = ("수주확정", "납품")
# v5H86: 상태가 '이미 수주된' 의미면 stage 와 무관하게 관리코드 발급
WON_STATUSES = ("진행중", "출하")
# v5H97: 거래 구분 (내수/수출)
TRADE_TYPES = ("내수", "수출")
# v5H226z349 (대표 지시): PO유형 'A/S' → '수리'로 이름변경. → 신규/추가/개조/수리/기타.
#   기존 po_type='A/S' 데이터는 초기화 마이그레이션에서 '수리'로 일괄 변환(아래 ALTER 직후).
PO_TYPES = ["신규", "추가", "개조", "수리", "기타"]
# v5H226z349 (대표 지시): 형태 — 제품(완제품·자체 제작)/상품(다품목·부품 상세리스트·PACKING LIST)/기타.
#   기존 '유형(project_type: 신규장비/소모품/수리/기타)'과는 별개 축. 작업일정표 컬럼·일괄등록 양식에서 사용.
# v5H226z455 (대표 지시): 형태 4종 — 완제품/제품/상품/기타. '호기'는 완제품에서만 매김.
#   완제품=장비 전체 완성품(ASSEMBLY·호기 분할) / 제품=반제품·가공부품·2차가공품(SEMI·호기 없음·1줄)
#   / 상품=부품 상세리스트(PARTS·PACKING LIST) / 기타=ETC. (기존 ASSEMBLY 데이터=완제품 유지)
FORM_TYPES = ("완제품", "제품", "상품", "기타")
FORM_TO_SHIP = {"완제품": "ASSEMBLY", "제품": "SEMI", "상품": "PARTS", "기타": "ETC"}
SHIP_TO_FORM = {"ASSEMBLY": "완제품", "SEMI": "제품", "PARTS": "상품", "ETC": "기타"}


def resolve_form_ship(data: dict) -> tuple[str, str]:
    """형태↔출고형태 짝 맞추기. 한쪽만 와도 다른 쪽을 채움.
    우선순위: form_type(폼·일괄등록 신규) → 없으면 shipment_form(레거시)로 역산.
    둘 다 없으면 ('', 'ASSEMBLY') — 형태는 빈값(표시 시 출고형태에서 역산)."""
    ft = (data.get("form_type") or "").strip()
    sf = (data.get("shipment_form") or "").strip().upper()
    if ft in FORM_TYPES:
        return ft, FORM_TO_SHIP[ft]
    if sf in ("ASSEMBLY", "SEMI", "PARTS", "ETC"):
        return SHIP_TO_FORM[sf], sf
    return "", "ASSEMBLY"


LOGI_STATUSES = ["초기협의", "제안서전달", "견적발행", "수주예정", "진행중", "출하", "보류", "취소", "기타"]

# v5H137 (2026-05-05) — 프로젝트 유형 분류 (대표 직접 요청)
# 등록 시점부터 신규장비/소모품/수리/기타 구분 + 소모품·수리는 부모 프로젝트 연결 가능
PROJECT_TYPES = ("NEW_EQUIP", "CONSUMABLE", "SERVICE", "OTHER")
PROJECT_TYPE_LABELS = {
    "NEW_EQUIP":   "🔧 신규 장비",
    "CONSUMABLE":  "📦 소모품·부품",
    "SERVICE":     "🔨 수리·유지보수",
    "OTHER":       "🌐 기타",
}
PROJECT_TYPE_UNIT_LABEL = {
    "NEW_EQUIP":   "{n}호기",
    "CONSUMABLE":  "{n}회차",
    "SERVICE":     "{n}차",
    "OTHER":       "{n}건",
}


def project_unit_label(project_type: str | None, n: int) -> str:
    """v5H137: 프로젝트 유형별 자동 SO 호기 라벨 생성.
    NEW_EQUIP/None → '1호기', CONSUMABLE → '1회차', SERVICE → '1차', OTHER → '1건'.
    백워드 호환: project_type NULL 또는 미지원 값 → '호기' 패턴."""
    pt = (project_type or "NEW_EQUIP").upper()
    if pt not in PROJECT_TYPES:
        pt = "NEW_EQUIP"
    return PROJECT_TYPE_UNIT_LABEL[pt].format(n=n)


# v5H142 (2026-05-05) — 부모 프로젝트 안 SO 단위 종류 분리 (대표 직접 요청)
# 소모품/정비를 매번 새 관리번호 발급하지 않고 부모 SO 로 직접 추가
SO_TYPES = ("EQUIPMENT", "CONSUMABLE", "SERVICE", "OTHER")
SO_TYPE_LABELS = {
    "EQUIPMENT":  ("🔧", "장비"),
    "CONSUMABLE": ("📦", "소모품"),
    "SERVICE":    ("🔨", "정비"),
    "OTHER":      ("🌐", "기타"),
}
SO_TYPE_UNIT_LABEL = {
    "EQUIPMENT":  "{n}호기",
    "CONSUMABLE": "소모품-{n}차",
    "SERVICE":    "정비-{n}차",
    "OTHER":      "{n}건",
}


def so_unit_label(so_type: str | None, project_type: str | None, n: int) -> str:
    """v5H142: SO 단위 라벨. so_type 우선 → 폴백 project_type → 폴백 '{n}호기'."""
    st = (so_type or "").upper()
    if st in SO_TYPE_UNIT_LABEL:
        return SO_TYPE_UNIT_LABEL[st].format(n=n)
    return project_unit_label(project_type, n)


def _logi_now() -> str:
    return _dt.now().strftime("%Y-%m-%d %H:%M:%S")


# ── 부품 마스터 (parts) CRUD ─────────────────────────────
def parts_list(q: str = "", biz_div: str = "", category: str = ""):
    sql = "SELECT * FROM parts WHERE 1=1"
    params: list = []
    if q:
        # v5H226z99 (2026-05-16): purpose(용도) 검색 추가
        # v5H226z110 (2026-05-18): 자재코드 = 모델명 통합 — 별도 모델명 LIKE 제거 (part_no LIKE로 커버)
        sql += " AND (part_no LIKE ? OR part_name LIKE ? OR spec LIKE ? OR maker LIKE ? OR COALESCE(purpose,'') LIKE ?)"
        like = f"%{q}%"
        params += [like, like, like, like, like]
    if biz_div:
        sql += " AND biz_div = ?"
        params.append(biz_div)
    if category:
        sql += " AND category = ?"
        params.append(category)
    sql += " ORDER BY id DESC"
    with db_session() as c:
        return c.execute(sql, params).fetchall()


def parts_get(pid: int):
    with db_session() as c:
        return c.execute("SELECT * FROM parts WHERE id = ?", (pid,)).fetchone()


def purchase_prices_rep(raw) -> dict:
    """v5H226z84: 매입단가 JSON → 목록 표시용 요약.
    반환: {"vendor": 대표 구매사명, "vendor_count": 구매사 수(이름 있는), "price": 대표단가 or None}
    대표 = 가장 최근 날짜의 단가가 속한 구매사 (날짜 없으면 마지막 price>0).
    """
    import json as _json
    out = {"vendor": "", "vendor_count": 0, "price": None}
    if not raw:
        return out
    try:
        vendors = raw if isinstance(raw, list) else _json.loads(raw)
    except (ValueError, TypeError):
        return out
    if not isinstance(vendors, list):
        return out
    named = 0
    best = None       # ((date, seq), price, supplier)
    fallback = None   # (price, supplier)
    for vend in vendors:
        if not isinstance(vend, dict):
            continue
        supplier = str(vend.get("supplier") or "").strip()
        if supplier:
            named += 1
        for ent in (vend.get("entries") or []):
            if not isinstance(ent, dict):
                continue
            try:
                price = float(ent.get("price") or 0)
            except (TypeError, ValueError):
                price = 0.0
            if price <= 0:
                continue
            date = str(ent.get("date") or "").strip()
            try:
                seq = int(ent.get("seq") or 0)
            except (TypeError, ValueError):
                seq = 0
            fallback = (price, supplier)
            if date:
                key = (date, seq)
                if best is None or key > best[0]:
                    best = (key, price, supplier)
    out["vendor_count"] = named
    if best is not None:
        out["vendor"], out["price"] = best[2], best[1]
    elif fallback is not None:
        out["vendor"], out["price"] = fallback[1], fallback[0]
    return out


def parts_delete_impact(pid: int) -> dict:
    """v5H226z81: 자재 삭제 영향도 — 발주/견적 연결 건수 + 재고 + 첨부.
    수정 페이지의 '완전 삭제' 위험 경고에 사용."""
    out = {"stock_qty": 0.0, "po_refs": 0, "quote_refs": 0,
           "movement_refs": 0, "attach_refs": 0}
    with db_session() as c:
        def _count(sql, params=(pid,)):
            try:
                r = c.execute(sql, params).fetchone()
                return int(r[0]) if r and r[0] is not None else 0
            except Exception:
                return 0
        try:
            r = c.execute("SELECT stock_qty FROM parts WHERE id=?", (pid,)).fetchone()
            out["stock_qty"] = float(r[0]) if r and r[0] is not None else 0.0
        except Exception:
            pass
        out["po_refs"]       = _count("SELECT COUNT(*) FROM po_items WHERE part_id=?")
        out["quote_refs"]    = _count("SELECT COUNT(*) FROM quotation_items WHERE part_id=?")
        out["movement_refs"] = _count("SELECT COUNT(*) FROM stock_movements WHERE part_id=?")
        out["attach_refs"]   = _count("SELECT COUNT(*) FROM part_attachments WHERE part_id=?")
    return out


# =====================================================
# v5H226z58 (2026-05-12) — 자재 중복 등록 방지 다층 방어 (A안)
# 1) _normalize_part_key — 공백·대소문자·특수문자 제거 정규화
# 2) _similarity_score — 정규화 일치도 + 부분 포함 + Levenshtein 거리 점수
# 3) parts_find_similar — 유사 자재 검색 (점수 정렬, 상위 limit)
# 사용: parts_create(force=True) 로 우회 가능 / 라우터 /parts/check JSON 응답
# =====================================================
import re as _pq_re

def _normalize_part_key(s: str) -> str:
    """공백·하이픈·언더스코어·특수문자 제거 + 소문자 통일.
    영숫자·한글만 남김. 비교용 정규화 키."""
    return _pq_re.sub(r'[^a-z0-9가-힣]', '', (s or '').lower())


def _levenshtein(a: str, b: str) -> int:
    """편집 거리 (insert/delete/replace) — 짧은 문자열용 자체 구현."""
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    if len(a) < len(b):
        a, b = b, a
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cost = 0 if ca == cb else 1
            cur.append(min(cur[-1] + 1, prev[j] + 1, prev[j - 1] + cost))
        prev = cur
    return prev[-1]


def _similarity_score(a: str, b: str) -> int:
    """0~100 점수. 100=완전동일, ≥85=확정유사, ≥60=주의 권고."""
    if not a or not b:
        return 0
    na, nb = _normalize_part_key(a), _normalize_part_key(b)
    if not na or not nb:
        return 0
    if na == nb:
        return 100
    # 부분 포함 (긴 것이 짧은 것을 완전 포함)
    if na in nb or nb in na:
        ratio = min(len(na), len(nb)) / max(len(na), len(nb))
        return int(60 + 30 * ratio)  # 60~90
    # Levenshtein 거리 기반 (max len 대비)
    dist = _levenshtein(na, nb)
    max_len = max(len(na), len(nb))
    sim = max(0, 100 - int(100 * dist / max_len))
    return sim


def parts_find_similar(name: str, spec: str = "", maker: str = "",
                       exclude_id: int = 0, limit: int = 5,
                       threshold: int = 60) -> list[dict]:
    """유사 자재 검색. 자재명 위주 + spec/maker 보조 점수 가산.
    반환: [{id, part_no, part_name, spec, maker, score}] 점수 내림차순.
    threshold(60) 이상만 반환. 정확히 part_no 일치하는 행은 100점."""
    name = (name or "").strip()
    if not name:
        return []
    spec = (spec or "").strip()
    maker = (maker or "").strip()
    with db_session() as c:
        # 1차 좁히기 — 정규화 키 첫 4자가 같은 행만 후보 (LIKE 인덱스 활용)
        nk = _normalize_part_key(name)
        prefix = nk[:3] if len(nk) >= 3 else nk
        # 광범위 후보 가져오기 — part_name LIKE 또는 part_no LIKE
        like = f"%{name[:6]}%" if len(name) >= 2 else f"%{name}%"
        rows = c.execute(
            """SELECT id, part_no, part_name, spec, maker, biz_div, category
               FROM parts
               WHERE id <> ? AND COALESCE(is_active,1)=1
                 AND (part_name LIKE ? OR part_no LIKE ?)
               LIMIT 200""",
            (int(exclude_id or 0), like, like),
        ).fetchall()
    scored = []
    for r in rows:
        d = dict(r)
        name_score = _similarity_score(name, d.get("part_name") or "")
        # spec/maker 일치 시 가산 (각 +5, 최대 100)
        bonus = 0
        if spec and d.get("spec") and _normalize_part_key(spec) == _normalize_part_key(d.get("spec") or ""):
            bonus += 5
        if maker and d.get("maker") and _normalize_part_key(maker) == _normalize_part_key(d.get("maker") or ""):
            bonus += 5
        score = min(100, name_score + bonus)
        if score >= threshold:
            d["score"] = score
            d["score_name"] = name_score
            d["score_bonus"] = bonus
            scored.append(d)
    scored.sort(key=lambda x: (-x["score"], x["id"]))
    return scored[:max(1, int(limit))]


def _validate_parts_payload(data: dict) -> None:
    """v5H113 LOW#21: parts 도메인 enum/필수 검증."""
    if not (data.get("part_name") or "").strip():
        raise ValueError("부품명(part_name)은 필수입니다.")
    cur = (data.get("currency") or "KRW").strip().upper() or "KRW"
    if cur not in CURRENCY_OPTIONS:
        raise ValueError(f"통화는 {', '.join(CURRENCY_OPTIONS)} 중 하나여야 합니다. (입력: {cur})")
    data["currency"] = cur
    unit = (data.get("unit") or "EA").strip().upper() or "EA"
    # unit 은 자유롭게 허용하되, 잘못된 값 잡기 위한 너그러운 화이트리스트 (경고만)
    if unit not in PART_UNIT_OPTIONS:
        # 사용자 정의 unit 도 허용 (예: '병', '캔') — 단, 길이 제한
        if len(unit) > 8:
            raise ValueError(f"단위(unit)는 8자 이내여야 합니다. (입력: {unit})")
    data["unit"] = unit
    # is_active 화이트리스트
    ia_raw = data.get("is_active", 1)
    try:
        ia = int(ia_raw) if str(ia_raw) not in ("", None) else 1
    except (TypeError, ValueError):
        raise ValueError(f"is_active 는 0 또는 1이어야 합니다. (입력: {ia_raw})")
    if ia not in (0, 1):
        raise ValueError(f"is_active 는 0 또는 1이어야 합니다. (입력: {ia_raw})")
    data["is_active"] = ia
    # std_price(매입단가) 음수 차단
    try:
        if float(data.get("std_price") or 0) < 0:
            raise ValueError("매입단가(std_price)는 0 이상이어야 합니다.")
    except (TypeError, ValueError) as _e:
        if "매입단가" in str(_e):
            raise
        raise ValueError("매입단가(std_price)가 올바른 숫자가 아닙니다.")


def _parse_purchase_prices(raw):
    """v5H226z84/z112: 매입단가 — 협력사 3곳(구매사) × 각 1~5차 JSON 파싱.
    형식: [
      {"supplier":"㈜A", "is_default":true,
       "contact_name":"홍길동", "contact_phone":"010-...", "contact_email":"abc@a.com",
       "entries":[{"seq":1,"date":"2026-05-01","price":52750}, ...]},
      ...
    ]
    반환: (정규화 JSON 문자열 or None, 대표단가 float or None)
    · v5H226z112: 협력사명은 마스터 검증(없으면 NULL로 정규화). 담당자·기본 협력사 필드도 보존.
    """
    import json as _json
    if not raw:
        return None, None
    try:
        vendors = raw if isinstance(raw, list) else _json.loads(raw)
    except (ValueError, TypeError):
        return None, None
    if not isinstance(vendors, list):
        return None, None
    cleaned = []
    best = None
    fallback = None
    default_seen = False  # 첫 is_default=True만 유효
    for vend in vendors[:3]:
        if not isinstance(vend, dict):
            continue
        supplier_raw = str(vend.get("supplier") or "").strip()
        # v5H226z112 — 마스터 검증: 없는 이름은 NULL 로 정규화 (오타 차단)
        supplier = _validate_vendor_name(supplier_raw) or ""
        is_default = bool(vend.get("is_default")) and not default_seen
        if is_default:
            default_seen = True
        contact_name  = str(vend.get("contact_name")  or "").strip()
        contact_phone = str(vend.get("contact_phone") or "").strip()
        contact_email = str(vend.get("contact_email") or "").strip()
        raw_entries = vend.get("entries")
        entries = []
        if isinstance(raw_entries, list):
            for ent in raw_entries[:5]:
                if not isinstance(ent, dict):
                    continue
                try:
                    seq = int(ent.get("seq") or 0)
                except (TypeError, ValueError):
                    seq = 0
                date = str(ent.get("date") or "").strip()
                try:
                    price = float(ent.get("price") or 0)
                except (TypeError, ValueError):
                    price = 0.0
                if price < 0:
                    price = 0.0
                if price <= 0 and not date:
                    continue
                entries.append({"seq": seq, "date": date, "price": price})
                if price > 0:
                    fallback = price
                    if date:
                        key = (date, seq)
                        if best is None or key > best[0]:
                            best = (key, price)
        # 빈 슬롯 건너뛰기 — 협력사·entries·담당자 중 하나라도 있어야 보존
        if supplier or entries or contact_name or contact_phone or contact_email:
            cleaned.append({
                "supplier": supplier,
                "is_default": is_default,
                "contact_name": contact_name,
                "contact_phone": contact_phone,
                "contact_email": contact_email,
                "entries": entries,
            })
    if not cleaned:
        return None, None
    std_price = best[1] if best is not None else fallback
    return _json.dumps(cleaned, ensure_ascii=False), std_price


def _pp_sync_vendor_columns(data: dict, pp_json: str | None) -> None:
    """v5H226z112 (2026-05-19) — purchase_prices JSON 의 supplier 이름을
    parts.vendor1/2/3 컬럼과 default_supplier 컬럼에 자동 미러.
    JSON 내 is_default=true 인 협력사를 default_supplier 로 우선 사용.
    호출 시점은 _parse_purchase_prices 직후 (data 에 in-place 반영)."""
    import json as _json
    if not pp_json:
        return
    try:
        vendors = _json.loads(pp_json)
    except (ValueError, TypeError):
        return
    if not isinstance(vendors, list):
        return
    # 슬롯별 vendor1/2/3 (이미 마스터 검증 완료된 supplier 만 들어옴)
    for idx, vend in enumerate(vendors[:3], start=1):
        if isinstance(vend, dict):
            data[f"vendor{idx}"] = (vend.get("supplier") or "").strip() or None
    # is_default=True 인 협력사 → default_supplier
    default = None
    for vend in vendors:
        if isinstance(vend, dict) and vend.get("is_default"):
            default = (vend.get("supplier") or "").strip()
            if default:
                break
    if not default:
        # 폴백: 첫 번째 협력사 (이름 있는 것)
        for vend in vendors:
            if isinstance(vend, dict):
                _n = (vend.get("supplier") or "").strip()
                if _n:
                    default = _n
                    break
    if default is not None:
        data["default_supplier"] = default


def _validate_vendor_name(name) -> str | None:
    """v5H226z111 (2026-05-19) — 구매사(협력사) 이름 마스터 검증.
    suppliers.name 에 존재하는 이름만 통과. 빈값/매칭실패 → None.
    공백·대소문자 차이는 정규화하여 비교."""
    if not name:
        return None
    s = str(name).strip()
    if not s:
        return None
    try:
        with db_session() as c:
            row = c.execute(
                "SELECT name FROM suppliers WHERE TRIM(name)=? AND COALESCE(is_active,1)=1 LIMIT 1",
                (s,)
            ).fetchone()
            if row:
                return row[0]
            # 대소문자 무시 매칭 한 번 더 시도
            row = c.execute(
                "SELECT name FROM suppliers WHERE LOWER(TRIM(name))=LOWER(?) AND COALESCE(is_active,1)=1 LIMIT 1",
                (s,)
            ).fetchone()
            return row[0] if row else None
    except Exception:
        return None


def parts_create(data: dict, force: bool = False) -> int:
    """자재 신규 등록.
    force=False (기본): 유사 자재(점수 ≥ 85) 발견 시 ValueError + 후보 노출 (3층 방어)
    force=True: 유사 검사 통과 (사용자 confirm 후 재제출용).
    part_no UNIQUE 중복은 force 무관 항상 차단 (DB 제약)."""
    _validate_parts_payload(data)
    # v5H226z83: 매입단가 1~5차 → JSON 정규화 + 마지막 차수를 std_price 에 동기화
    _pp_json, _pp_last = _parse_purchase_prices(data.get("purchase_prices"))
    if _pp_last is not None:
        data["std_price"] = _pp_last
    # v5H226z112: purchase_prices JSON 의 협력사명을 vendor1/2/3에 자동 미러 + 기본 협력사 추출
    _pp_sync_vendor_columns(data, _pp_json)
    # v5H113 LOW#21: part_no 중복 친절 처리
    pno = (data.get("part_no") or "").strip()
    if pno:
        with db_session() as c:
            dup = c.execute("SELECT id, part_name FROM parts WHERE part_no=?", (pno,)).fetchone()
            if dup:
                raise ValueError(
                    f"부품 코드 '{pno}' 는 이미 등록되어 있습니다 "
                    f"(기존: [{dup['id']}] {dup['part_name']}). 다른 코드를 사용하거나 기존 부품을 수정해주세요."
                )
    # v5H226z58 (2026-05-12) — 3층 방어: 자재명 유사도 검사 (force=False 일 때만)
    if not force:
        sims = parts_find_similar(
            name=(data.get("part_name") or "").strip(),
            spec=(data.get("spec") or "").strip(),
            maker=(data.get("maker") or "").strip(),
            exclude_id=0, limit=5, threshold=85,
        )
        if sims:
            # 점수 85 이상 = 확정 유사 → 사용자 confirm 요구
            top = sims[0]
            tail = ", ".join(f"[{s['id']}] {s['part_name']}({s['score']}점)" for s in sims[:3])
            raise ValueError(
                f"유사한 자재 {len(sims)}건 발견 — 정말 신규 등록이 맞나요?\n"
                f"가장 유사: [{top['id']}] {top['part_name']} (점수 {top['score']})\n"
                f"전체: {tail}\n"
                f"기존 자재 수정 권장. 신규가 맞으면 'force=1' 로 재제출하세요."
            )
    # v5H226z58 (2026-05-11): 자재모듈 표준 v2 — 13 컬럼 확장 (item_account/procurement_kind/
    #   category_main/category_series/reorder_point/reorder_qty/conversion_factor/sub_spec1~3/
    #   tax_invoice_name/trade_invoice_name/default_warehouse/hs_code)
    # ※ 컬럼 부재 시 OperationalError → 백필 가드: PRAGMA로 실제 존재 컬럼만 사용
    base_cols = ["part_no", "part_name", "spec", "maker", "origin", "unit",
                 "currency", "std_price", "biz_div", "category", "note",
                 "is_active", "safety_stock", "location", "created_at", "updated_at"]
    now = _logi_now()
    base_values = [
        (data.get("part_no") or "").strip(),
        (data.get("part_name") or "").strip(),
        (data.get("spec") or "").strip(),
        (data.get("maker") or "").strip(),
        (data.get("origin") or "").strip(),
        (data.get("unit") or "EA").strip() or "EA",
        (data.get("currency") or "KRW").strip() or "KRW",
        float(data.get("std_price") or 0),
        (data.get("biz_div") or "").strip(),
        (data.get("category") or "").strip(),
        (data.get("note") or "").strip(),
        1 if data.get("is_active", 1) else 0,
        float(data.get("safety_stock") or 0),
        (data.get("location") or "").strip() or None,
        now, now,
    ]
    # z58 확장 컬럼 (존재 시에만 추가)
    ext_pairs = [
        ("item_account",       (data.get("item_account") or "").strip() or None),
        ("procurement_kind",   (data.get("procurement_kind") or "").strip() or None),
        ("category_main",      (data.get("category_main") or "").strip() or None),
        ("category_series",    (data.get("category_series") or "").strip() or None),
        ("reorder_point",      float(data.get("reorder_point") or 0)),
        ("reorder_qty",        float(data.get("reorder_qty") or 0)),
        ("conversion_factor",  float(data.get("conversion_factor") or 1)),
        ("sub_spec1",          (data.get("sub_spec1") or "").strip() or None),
        ("sub_spec2",          (data.get("sub_spec2") or "").strip() or None),
        ("sub_spec3",          (data.get("sub_spec3") or "").strip() or None),
        ("tax_invoice_name",   (data.get("tax_invoice_name") or "").strip() or None),
        ("trade_invoice_name", (data.get("trade_invoice_name") or "").strip() or None),
        ("default_warehouse",  (data.get("default_warehouse") or "").strip() or None),
        ("hs_code",            (data.get("hs_code") or "").strip() or None),
        ("purchase_prices",    _pp_json),
        # v5H226z99 (2026-05-16): 용도 — 카탈로그 검색 보조 자유 텍스트
        ("purpose",            (data.get("purpose") or "").strip() or None),
        # v5H226z110b (2026-05-19): 기본 협력사명 — deprecated (vendor1로 이전됨)
        ("default_supplier",   (data.get("default_supplier") or "").strip() or None),
        # v5H226z110c (2026-05-19): 제조사 담당자 풀세트
        ("maker_contact_name",  (data.get("maker_contact_name") or "").strip() or None),
        ("maker_contact_phone", (data.get("maker_contact_phone") or "").strip() or None),
        ("maker_contact_email", (data.get("maker_contact_email") or "").strip() or None),
        # v5H226z111 (2026-05-19): 구매사(협력사) 3슬롯 — 마스터 검증 후 저장
        ("vendor1",            _validate_vendor_name(data.get("vendor1"))),
        ("vendor2",            _validate_vendor_name(data.get("vendor2"))),
        ("vendor3",            _validate_vendor_name(data.get("vendor3"))),
    ]
    with db_session() as c:
        # PRAGMA 로 실제 존재 컬럼만 추가 (z58 SQL 미적용 환경 호환)
        existing = {r[1] for r in c.execute("PRAGMA table_info(parts)").fetchall()}
        ext_cols, ext_vals = [], []
        for col, val in ext_pairs:
            if col in existing:
                ext_cols.append(col)
                ext_vals.append(val)
        cols = base_cols + ext_cols
        values = base_values + ext_vals
        placeholders = ",".join(["?"] * len(cols))
        cur = c.execute(
            f"INSERT INTO parts ({','.join(cols)}) VALUES ({placeholders})",
            values,
        )
        return cur.lastrowid


def parts_update(pid: int, data: dict) -> None:
    _validate_parts_payload(data)
    # v5H226z83: 매입단가 1~5차 — "purchase_prices" 키가 전달된 경우에만 처리
    #   (대량등록 등 키 미전달 호출은 기존 컬럼 값 보존)
    _pp_provided = "purchase_prices" in data
    _pp_json = None
    if _pp_provided:
        _pp_json, _pp_last = _parse_purchase_prices(data.get("purchase_prices"))
        if _pp_last is not None:
            data["std_price"] = _pp_last
        # v5H226z112: vendor1/2/3·default_supplier 자동 미러
        _pp_sync_vendor_columns(data, _pp_json)
    # v5H113 LOW#21: 다른 row 와 part_no 중복 차단
    pno = (data.get("part_no") or "").strip()
    if pno:
        with db_session() as c:
            dup = c.execute(
                "SELECT id, part_name FROM parts WHERE part_no=? AND id<>?",
                (pno, pid)
            ).fetchone()
            if dup:
                raise ValueError(
                    f"부품 코드 '{pno}' 는 이미 다른 부품에서 사용 중입니다 "
                    f"(기존: [{dup['id']}] {dup['part_name']})."
                )
    # v5H226z58 (2026-05-11): 자재모듈 표준 v2 — 13 컬럼 확장 (parts_create 와 동일 set)
    base_fields = ["part_no", "part_name", "spec", "maker", "origin", "unit",
                   "currency", "std_price", "biz_div", "category", "note", "is_active",
                   "safety_stock", "location"]
    base_values = [
        (data.get("part_no") or "").strip(),
        (data.get("part_name") or "").strip(),
        (data.get("spec") or "").strip(),
        (data.get("maker") or "").strip(),
        (data.get("origin") or "").strip(),
        (data.get("unit") or "EA").strip() or "EA",
        (data.get("currency") or "KRW").strip() or "KRW",
        float(data.get("std_price") or 0),
        (data.get("biz_div") or "").strip(),
        (data.get("category") or "").strip(),
        (data.get("note") or "").strip(),
        1 if data.get("is_active", 1) else 0,
        float(data.get("safety_stock") or 0),
        (data.get("location") or "").strip() or None,
    ]
    # z58 확장 컬럼 (존재 시에만 SET)
    ext_pairs = [
        ("item_account",       (data.get("item_account") or "").strip() or None),
        ("procurement_kind",   (data.get("procurement_kind") or "").strip() or None),
        ("category_main",      (data.get("category_main") or "").strip() or None),
        ("category_series",    (data.get("category_series") or "").strip() or None),
        ("reorder_point",      float(data.get("reorder_point") or 0)),
        ("reorder_qty",        float(data.get("reorder_qty") or 0)),
        ("conversion_factor",  float(data.get("conversion_factor") or 1)),
        ("sub_spec1",          (data.get("sub_spec1") or "").strip() or None),
        ("sub_spec2",          (data.get("sub_spec2") or "").strip() or None),
        ("sub_spec3",          (data.get("sub_spec3") or "").strip() or None),
        ("tax_invoice_name",   (data.get("tax_invoice_name") or "").strip() or None),
        ("trade_invoice_name", (data.get("trade_invoice_name") or "").strip() or None),
        ("default_warehouse",  (data.get("default_warehouse") or "").strip() or None),
        ("hs_code",            (data.get("hs_code") or "").strip() or None),
    ]
    # v5H226z99 (2026-05-16): 용도(purpose) — 폼에서 전달된 경우에만 SET (미전달 호출은 보존)
    if "purpose" in data:
        ext_pairs.append(("purpose", (data.get("purpose") or "").strip() or None))
    # v5H226z110b (2026-05-19): 기본 공급사명 — 폼 전달 시에만 SET (보존)
    if "default_supplier" in data:
        ext_pairs.append(("default_supplier", (data.get("default_supplier") or "").strip() or None))
    # v5H226z110c (2026-05-19): 제조사 담당자 풀세트 — 폼 전달 시에만 SET (보존)
    for _k in ("maker_contact_name", "maker_contact_phone", "maker_contact_email"):
        if _k in data:
            ext_pairs.append((_k, (data.get(_k) or "").strip() or None))
    # v5H226z111 (2026-05-19): 구매사 3슬롯 — 마스터 검증 후 저장 (없는 이름은 NULL)
    for _k in ("vendor1", "vendor2", "vendor3"):
        if _k in data:
            ext_pairs.append((_k, _validate_vendor_name(data.get(_k))))
    # v5H226z83: purchase_prices 는 폼에서 전달된 경우에만 SET (미전달 호출은 보존)
    if _pp_provided:
        ext_pairs.append(("purchase_prices", _pp_json))
    with db_session() as c:
        existing = {r[1] for r in c.execute("PRAGMA table_info(parts)").fetchall()}
        ext_fields, ext_values = [], []
        for col, val in ext_pairs:
            if col in existing:
                ext_fields.append(col)
                ext_values.append(val)
        fields = base_fields + ext_fields
        values = base_values + ext_values + [_logi_now()]
        sets = ", ".join([f"{f} = ?" for f in fields]) + ", updated_at = ?"
        c.execute(f"UPDATE parts SET {sets} WHERE id = ?", values + [pid])


def parts_delete(pid: int) -> None:
    """v5H119: 공통 헬퍼 _safe_delete_with_cascade 사용. 폴백: v5H112 인라인 로직."""
    with db_session() as c:
        try:
            res = _safe_delete_with_cascade(
                c, "parts", pid, fk_column="part_id",
                explicit_children=[
                    ("UPDATE po_items SET part_id=NULL WHERE part_id=?", (pid,)),
                    ("UPDATE quotation_items SET part_id=NULL WHERE part_id=?", (pid,)),
                    ("DELETE FROM stock_movements WHERE part_id=?", (pid,)),
                    ("DELETE FROM part_prices WHERE part_id=?", (pid,)),
                    ("DELETE FROM parts_safety WHERE part_id=?", (pid,)),
                ],
            )
            if res.get("ok"):
                return
        except Exception:
            pass
        # 폴백 (v5H112 원본 인라인) — 헬퍼 실패/예외 시 안전망
        for sql in [
            "UPDATE po_items SET part_id=NULL WHERE part_id=?",
            "UPDATE quotation_items SET part_id=NULL WHERE part_id=?",
            "DELETE FROM stock_movements WHERE part_id=?",
            "DELETE FROM part_prices WHERE part_id=?",
            "DELETE FROM parts_safety WHERE part_id=?",
        ]:
            try: c.execute(sql, (pid,))
            except Exception: pass
        try:
            all_tables = [r[0] for r in c.execute(
                "SELECT name FROM sqlite_master WHERE type='table' "
                "AND name NOT LIKE 'sqlite_%' AND name <> 'parts'"
            ).fetchall()]
            for tname in all_tables:
                try:
                    cols = [r[1] for r in c.execute(f"PRAGMA table_info({tname})").fetchall()]
                except Exception:
                    continue
                if "part_id" not in cols:
                    continue
                try:
                    c.execute(f"UPDATE {tname} SET part_id=NULL WHERE part_id=?", (pid,))
                except Exception:
                    try: c.execute(f"DELETE FROM {tname} WHERE part_id=?", (pid,))
                    except Exception: pass
        except Exception:
            pass
        c.execute("DELETE FROM parts WHERE id = ?", (pid,))


def parts_count() -> dict:
    with db_session() as c:
        total = c.execute("SELECT COUNT(*) FROM parts").fetchone()[0]
        active = c.execute("SELECT COUNT(*) FROM parts WHERE is_active = 1").fetchone()[0]
        by_div = c.execute(
            "SELECT biz_div, COUNT(*) AS n FROM parts GROUP BY biz_div"
        ).fetchall()
    return {
        "total": total,
        "active": active,
        "by_div": {r["biz_div"] or "(미지정)": r["n"] for r in by_div},
    }


# ── 프로젝트 / 관리코드 발행대장 (projects) CRUD ──────────
def generate_mgmt_code(biz_div: str, today=None) -> str:
    """KNK PMS 8자리 관리코드: [일련3][T/M/K/S][YYMM]. 같은 prefix·연월 내 +1
    v5H150: K(기타) 추가 — project_type=OTHER 일 때 호출측에서 'K' 전달.
    v5H216: S(소모품) 추가.
    v5H222: S 도 projects 테이블에서 검색 (CONSUMABLE 신규 등록은 projects 도메인으로 통일).
            consumable_orders 도 함께 스캔(legacy 데이터와 sequence 충돌 방지)."""
    # v5H226z578 (대표 지시): R(연구과제) 추가 — 부서 자체 연구개발 과제도 관리번호 발번(매출과 별개·rnd_projects 도메인).
    if biz_div not in ("T", "M", "L", "E", "C", "R"):
        raise ValueError(f"biz_div must be T, M, L, E, C, or R (got: {biz_div})")
    today = today or _date.today()
    yymm = today.strftime("%y%m")
    pat = f"%{biz_div}{yymm}"
    # v5H226z386 (대표 지시): 검증(테스트) 모드 — 켜져 있으면 관리번호를 'A' 접두로 발급(A01T2606...).
    #   'A' 시작 = 임시 테스트 데이터 → '검증 데이터 관리' 화면에서 일괄 삭제 대상. (실 관리번호는 항상 숫자 3자리로 시작)
    _test = False
    with db_session() as c:
        try:
            _tr = c.execute("SELECT value FROM app_settings WHERE key='test_mode'").fetchone()
            _test = ((_tr[0] if _tr else "") or "").strip().lower() in ("1", "on", "true", "yes")
        except Exception:
            _test = False
        rows = c.execute(
            "SELECT mgmt_code FROM projects WHERE mgmt_code LIKE ?", (pat,)
        ).fetchall()
        if biz_div == "C":
            # legacy consumable_orders.mgmt_code 도 함께 — sequence 충돌 방지
            try:
                rows = list(rows) + list(c.execute(
                    "SELECT mgmt_code FROM consumable_orders WHERE mgmt_code LIKE ?", (pat,)
                ).fetchall())
            except Exception:
                pass
        if biz_div == "R":
            # v5H226z578: 연구과제는 rnd_projects 도메인 — 거기서 발번 sequence 스캔
            try:
                rows = list(rows) + list(c.execute(
                    "SELECT mgmt_code FROM rnd_projects WHERE mgmt_code LIKE ?", (pat,)
                ).fetchall())
            except Exception:
                pass
    if _test:
        # A 접두 테스트 코드: A + 2자리 일련 + 사업부 + 연월 (A01T2606). 기존 A 코드 중 최대 +1.
        a_seq = 0
        for r in rows:
            code = r["mgmt_code"] or ""
            if len(code) == 8 and code[0] == "A" and code[3] == biz_div and code[4:] == yymm:
                try:
                    a_seq = max(a_seq, int(code[1:3]))
                except ValueError:
                    pass
        return f"A{a_seq + 1:02d}{biz_div}{yymm}"
    max_seq = 0
    for r in rows:
        code = r["mgmt_code"] or ""
        if len(code) == 8 and code[3] == biz_div and code[4:] == yymm:
            try:
                max_seq = max(max_seq, int(code[:3]))
            except ValueError:
                pass
    return f"{max_seq + 1:03d}{biz_div}{yymm}"


def projects_list_logi(q: str = "", biz_div: str = "", stage: str = "",
                       status: str = "", project_type: str = ""):
    """v5H137: project_type 필터 추가 (NEW_EQUIP/CONSUMABLE/SERVICE/OTHER)."""
    sql = ("SELECT p.*, p.name AS project_name "
           "FROM projects p WHERE 1=1")
    params: list = []
    if q:
        # v5H226z217: 장비명(equip_name)도 검색 대상 추가
        sql += (" AND (p.mgmt_code LIKE ? OR p.name LIKE ? OR p.customer_name LIKE ? "
                "OR p.model_name LIKE ? OR p.equip_name LIKE ? OR p.pm_name LIKE ? OR p.sales_name LIKE ?)")
        like = f"%{q}%"
        params += [like] * 7
    if biz_div:
        sql += " AND p.biz_div = ?"
        params.append(biz_div)
    if stage:
        sql += " AND p.stage = ?"
        params.append(stage)
    if status:
        sql += " AND p.status = ?"
        params.append(status)
    if project_type and project_type.upper() in PROJECT_TYPES:
        sql += " AND COALESCE(p.project_type,'NEW_EQUIP') = ?"
        params.append(project_type.upper())
    sql += " ORDER BY p.id DESC"
    with db_session() as c:
        return c.execute(sql, params).fetchall()


def projects_get_logi(pid: int):
    with db_session() as c:
        return c.execute(
            "SELECT *, name AS project_name FROM projects WHERE id = ?",
            (pid,),
        ).fetchone()


def _project_insert_or_update_values(data: dict) -> dict:
    """공통 필드 정리"""
    return {
        "name": (data.get("project_name") or data.get("name") or "").strip(),
        "biz_div": (data.get("biz_div") or "").strip(),
        "customer_name": (data.get("customer_name") or data.get("customer") or "").strip(),
        "model_name": (data.get("model_name") or data.get("model") or "").strip(),
        "stage": (data.get("stage") or "제안작성").strip() or "제안작성",
        "po_type": (data.get("po_type") or "신규").strip() or "신규",
        "status": (data.get("status") or "수주예정").strip() or "수주예정",
        "customer_po": (data.get("customer_po") or "").strip(),
        # v5H226z357 (대표 지시·통화 중대버그): 허용 통화를 단일 진실 소스(CURRENCY_OPTIONS, 6종)로.
        #   기존 (KRW,USD,VND) 3종 제한이 EUR/JPY/CNY 를 조용히 KRW 로 강등시키던 버그 수정.
        "currency": ((data.get("currency") or "KRW").strip().upper()
                     if (data.get("currency") or "KRW").strip().upper() in CURRENCY_OPTIONS
                     else "KRW"),
        "is_export": 1 if str(data.get("is_export") or data.get("trade_type") or "").lower() in ("1","true","수출","export") else 0,
        "order_amount": float(data.get("order_amount") or 0),
        # v5H226z161: 예상 매출(수주 전) — 명시값 우선, 없으면 헤더 order_amount 와 동일. order_amount(확정)와 분리 보존.
        "expected_amount": float(str(data.get("expected_amount")).replace(",", "")) if str(data.get("expected_amount") or "").strip() not in ("",) else float(data.get("order_amount") or 0),
        # v5H226z166: 고객사 담당자
        "cc_name":     (data.get("cc_name") or "").strip() or None,
        "cc_dept":     (data.get("cc_dept") or "").strip() or None,
        "cc_position": (data.get("cc_position") or "").strip() or None,
        "cc_phone":    (data.get("cc_phone") or "").strip() or None,
        "cc_email":    (data.get("cc_email") or "").strip() or None,
        "cc_note":     (data.get("cc_note") or "").strip() or None,
        # v5H226z195/z350: 2단계 발주 — 2차고객사·최종 고객 매출(참고).
        #   z350: '2차거래' 체크박스 폐지 → 고객사2(secondary_customer) 입력 시 2단계 발주로 자동 판정.
        "two_tier_order": 1 if (str(data.get("two_tier_order") or "").lower() in ("1", "true", "on", "yes")
                                or (data.get("secondary_customer") or "").strip()) else 0,
        "secondary_customer": (data.get("secondary_customer") or "").strip() or None,
        "final_amount": (float(str(data.get("final_amount")).replace(",", ""))
                         if str(data.get("final_amount") or "").strip() not in ("",) else None),
        # v5H132: 단가/수량 — 폼에서 받지 않으면 None/1 폴백 (백워드 호환)
        "unit_qty": max(1, min(100, int(float(data.get("unit_qty") or 1)))) if str(data.get("unit_qty") or "").strip() else 1,
        "unit_price": (float(data.get("unit_price")) if str(data.get("unit_price") or "").strip() not in ("", "0") else None),
        "order_date": (data.get("order_date") or "").strip(),
        "due_date": (data.get("due_date") or "").strip(),
        "pm_name": (data.get("pm_name") or data.get("pm") or "").strip(),
        "sales_name": (data.get("sales_name") or data.get("sales") or "").strip(),
        "logi_note": (data.get("logi_note") or data.get("note") or "").strip(),
        # v5H226z375 (대표 지시): 제작요청서 항목 — 자료 경로(제품자료 서버 경로) + 각인(마킹 문구)
        "server_path": (data.get("server_path") or "").strip() or None,
        "engraving":   (data.get("engraving") or "").strip() or None,
        # v5H137: 프로젝트 유형 + 부모 프로젝트 (소모품/수리 연결)
        "project_type": (
            (data.get("project_type") or "NEW_EQUIP").strip().upper()
            if (data.get("project_type") or "NEW_EQUIP").strip().upper() in PROJECT_TYPES
            else "NEW_EQUIP"
        ),
        "parent_project_id": (
            int(data.get("parent_project_id"))
            if str(data.get("parent_project_id") or "").strip().isdigit()
            else None
        ),
        # v5H154: 외화 시 기준환율 + 원화 환산 (KRW 면 None)
        "fx_rate": (
            float(str(data.get("fx_rate") or "").replace(",", ""))
            if str(data.get("fx_rate") or "").strip().replace(",", "").replace(".", "").isdigit()
            and float(str(data.get("fx_rate") or "0").replace(",", "")) > 0
            else None
        ),
        "amount_krw": (
            float(str(data.get("amount_krw") or "").replace(",", ""))
            if str(data.get("amount_krw") or "").strip().replace(",", "").replace(".", "").isdigit()
            and float(str(data.get("amount_krw") or "0").replace(",", "")) > 0
            else None
        ),
        # v5H201: 제안 단계 일정 (수주확정 전 스케줄용). 빈 문자열은 None 으로.
        "proposal_date":  (data.get("proposal_date") or "").strip() or None,
        "quotation_date": (data.get("quotation_date") or "").strip() or None,
        # v5H226z350 (대표 지시): 출고 형태 = 형태(form_type)와 통합 — 한쪽만 와도 짝 맞춤.
        #   제품→ASSEMBLY / 상품→PARTS(부품 상세·PACKING LIST) / 기타→ETC.
        "shipment_form": resolve_form_ship(data)[1],
        # v5H226z211: 장비명 — 관리번호=장비 단위 (프로젝트명 비유일)
        "equip_name": (data.get("equip_name") or "").strip() or None,
        # v5H226z349/z350 (대표 지시): 형태(제품/상품/기타) — 출고형태와 짝(둘 중 온 값으로 통합).
        "form_type": resolve_form_ship(data)[0],
    }


def _sync_project_contact_to_customer(c, customer_id, name, dept, position, phone, email, note):
    """v5H226z166 (대표 지시): 프로젝트의 고객사 담당자를 customer_contacts(고객사 주소록)에
    upsert (이름 기준). 양방향 연동 — 프로젝트에 적으면 고객사 상세에도 자동 등록."""
    if not customer_id or not (name or "").strip():
        return
    name = name.strip()
    try:
        row = c.execute(
            "SELECT id FROM customer_contacts WHERE customer_id=? AND name=? LIMIT 1",
            (customer_id, name)).fetchone()
        if row:
            c.execute(
                "UPDATE customer_contacts SET department=COALESCE(?,department), "
                "position=COALESCE(?,position), phone=COALESCE(?,phone), "
                "email=COALESCE(?,email), note=COALESCE(?,note) WHERE id=?",
                (dept or None, position or None, phone or None, email or None,
                 note or None, row["id"]))
        else:
            c.execute(
                "INSERT INTO customer_contacts(customer_id, role, department, name, "
                "position, phone, mobile, email, is_primary, note) "
                "VALUES(?,?,?,?,?,?,?,?,?,?)",
                (customer_id, "기타", dept or None, name, position or None,
                 phone or None, None, email or None, 1, note or None))
    except Exception as _e:
        print(f"[z166] customer_contact 동기화 실패: {_e}")


def resolve_customer_id(c, name, picked_id=None, prefer_id=None):
    """v5H226z651 (대표 지시): 고객사명 → customer_id 안전 해석 — 데이터 연결성 안전원칙 통일.

    z619 로 같은 상호(name) 여러 종사업장(customers 행 2개 이상)이 가능해진 뒤,
    기존 'WHERE name=? LIMIT 1' 은 종사업장을 임의로 골라 '엉뚱한 곳에 조용히 연결'하거나
    매칭이 빗나가 NULL(=작업일정표 빨강) 이 됐다. 이를 모든 등록/수정 경로에서 안전하게 통일.

    우선순위:
      1) picked_id (콤보박스에서 사용자가 종사업장을 직접 고른 id) — 유효하면 최우선.
      2) 이름 후보가 '정확히 1개' 일 때만 자동연결.
      3) 후보가 2개 이상(같은 상호 여러 종사업장)이면 — prefer_id(기존 연결)가 그 후보 중 하나면 유지,
         아니면 자동연결하지 않음(None=미연결, 수동 종사업장 선택 유도).
      4) 후보 0개면 None(미연결).

    반환: (customer_id 또는 None, reason)
      reason: 'picked' | 'matched' | 'kept' | 'ambiguous' | 'unmatched' | 'empty'
    """
    if picked_id not in (None, ""):
        try:
            _pid = int(picked_id)
            if c.execute("SELECT 1 FROM customers WHERE id=?", (_pid,)).fetchone():
                return (_pid, "picked")
        except (TypeError, ValueError):
            pass
    _name = (name or "").strip()
    if not _name:
        return (None, "empty")
    # v5H226z654 (대표 지시): 등록 상호(name) OR 시스템명(alias) 둘 중 하나라도 정확히 맞으면 매칭
    #   (예: 프로젝트 고객사칸에 '드림텍(본사)'(시스템명)가 저장돼 있어도 그 고객사로 연결). 후보 1개일 때만 자동연결.
    _ids = [r[0] for r in c.execute(
        "SELECT id FROM customers WHERE name=? OR alias=?", (_name, _name)
    ).fetchall()]
    if len(_ids) == 1:
        return (_ids[0], "matched")
    if len(_ids) > 1:
        if prefer_id and prefer_id in _ids:
            return (prefer_id, "kept")
        return (None, "ambiguous")   # 같은 상호 여러 종사업장 등 후보 다수 — 자동연결 금지(수동 유도)
    return (None, "unmatched")


def projects_create_logi(data: dict) -> tuple[int, str | None]:
    """OPS-P1-D3 [D-014]: mgmt_code UNIQUE 제약 race 시 최대 5회 재채번 retry.
    schema: line 78 mgmt_code TEXT UNIQUE — 동시 INSERT 방어."""
    import sqlite3 as _sq
    vals = _project_insert_or_update_values(data)
    now = _logi_now()
    # v5H86: stage 가 NEEDS_CODE_STAGES 이거나 status 가 WON_STATUSES 면 관리코드 발급
    # v5H142 (2026-05-05): CONSUMABLE/SERVICE 는 관리번호 발급 안 함
    # v5H150 (2026-05-05): OTHER 도 관리번호 발급 — prefix 'K' (대표 지시)
    # v5H223 (2026-05-08): CONSUMABLE 도 다른 3종과 동일하게 status 기반 발급 (수주확정 시점)
    _ptype_in = (vals.get("project_type") or "NEW_EQUIP").upper()
    _valid_for_code = (
        (vals["biz_div"] in ("T", "M", "L") and _ptype_in in ("NEW_EQUIP", "CONSUMABLE"))
        or _ptype_in == "OTHER"
    )
    needs_code = (vals["stage"] in NEEDS_CODE_STAGES or vals["status"] in WON_STATUSES) and _valid_for_code
    # v5H226z194 (2026-06-04 대표 지시): 등록 시 관리번호 '항상' 발급. force_code 면 단계 무관 발급
    #   (제안 단계여도 코드 부여 — 단, 상태/단계는 사용자 선택 그대로 유지, 수주확정으로 승격 안 함).
    _force_code = bool(data.get("force_code"))
    gen_code = needs_code or (_force_code and _valid_for_code)
    # status 가 won 인데 stage 가 제안 단계면 stage 도 '수주확정' 으로 승격 (force_code 단독은 단계 안 바꿈)
    if needs_code and vals["stage"] not in NEEDS_CODE_STAGES:
        vals["stage"] = "수주확정"
    # v5H226z148 (2026-06-02 대표 지시): 기존(진행 중) 프로젝트 등록 시 실제 관리코드 직접 지정.
    #   비우면 종전대로 자동 발급. 지정하면 그 코드를 그대로 사용(중복 시 즉시 오류).
    _provided_code = (data.get("mgmt_code") or "").strip().upper()
    last_err = None
    for _attempt in range(5):
        # v5H225: 유형별 prefix — OTHER→E (Etc.), CONSUMABLE→C (Consumable), NEW_EQUIP→biz_div(T/M)
        if _ptype_in == "OTHER":
            _code_prefix = "E"
        elif _ptype_in == "CONSUMABLE":
            _code_prefix = "C"
        else:
            _code_prefix = vals["biz_div"]
        if _provided_code:
            code = _provided_code            # 사용자가 기존 관리코드 지정 → 그대로 사용
        elif gen_code:
            code = generate_mgmt_code(_code_prefix)
        else:
            code = None
        try:
            with db_session() as c:
                # v5H89b: customer_name → customer_id 자동 매핑 (orders FK 연결용)
                # v5H226z651 (대표 지시): 종사업장 안전연결 — 이름 후보 1개일 때만 자동연결, 여러 종사업장이면 미연결(수동). 콤보박스 picked(customer_id) 우선.
                cust_id, _ = resolve_customer_id(c, vals.get("customer_name"),
                                                 picked_id=(data.get("customer_id") if isinstance(data, dict) else None))
                cur = c.execute("""
                    INSERT INTO projects
                    (mgmt_code, name, biz_div, customer_id, customer_name, model_name,
                     stage, po_type, status, customer_po, currency, order_amount,
                     expected_amount,
                     cc_name, cc_dept, cc_position, cc_phone, cc_email, cc_note,
                     two_tier_order, secondary_customer, final_amount,
                     order_date, due_date, pm_name, sales_name, logi_note,
                     is_export, unit_qty, unit_price,
                     project_type, parent_project_id,
                     fx_rate, amount_krw,
                     proposal_date, quotation_date,
                     shipment_form, equip_name, form_type,
                     server_path, engraving,
                     created_at, updated_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (code, vals["name"], vals["biz_div"], cust_id, vals["customer_name"],
                      vals["model_name"], vals["stage"], vals["po_type"], vals["status"],
                      vals["customer_po"], vals["currency"], vals["order_amount"],
                      vals["expected_amount"],
                      vals["cc_name"], vals["cc_dept"], vals["cc_position"],
                      vals["cc_phone"], vals["cc_email"], vals["cc_note"],
                      vals["two_tier_order"], vals["secondary_customer"], vals["final_amount"],
                      vals["order_date"], vals["due_date"], vals["pm_name"],
                      vals["sales_name"], vals["logi_note"], vals["is_export"],
                      vals.get("unit_qty") or 1, vals.get("unit_price"),
                      vals.get("project_type") or "NEW_EQUIP",
                      vals.get("parent_project_id"),
                      vals.get("fx_rate"), vals.get("amount_krw"),
                      vals.get("proposal_date"), vals.get("quotation_date"),
                      vals.get("shipment_form") or "ASSEMBLY", vals.get("equip_name"),
                      vals.get("form_type") or "",
                      vals.get("server_path"), vals.get("engraving"),
                      now, now))
                new_id = cur.lastrowid
                # v5H101: 프로젝트 생성 이벤트 기록
                log_project_change(c, new_id, data.get("_changed_by"),
                                   "프로젝트", "", vals["name"],
                                   note=f"신규 등록 (관리코드 {code or '미발급'})")
                # v5H183: 등록 시 모든 초기값 상세 로깅 (이력이 상세하게 기록되어야 한다는 대표 지시)
                _ptype = (vals.get("project_type") or "NEW_EQUIP").upper()
                _ptype_label = {"NEW_EQUIP":"신규 장비","CONSUMABLE":"소모품",
                                "SERVICE":"수리·정비","OTHER":"기타"}.get(_ptype, _ptype)
                _bd_label = {"T":"T (검사기)","M":"M (자동화)","K":"K (기타)"}.get(
                    vals.get("biz_div") or "", vals.get("biz_div") or "")
                _initial_fields = [
                    ("사업부",       "", _bd_label),
                    ("프로젝트유형", "", _ptype_label),
                    ("PO유형",       "", vals.get("po_type") or ""),
                    ("고객사",       "", vals.get("customer_name") or ""),
                    ("모델",         "", vals.get("model_name") or ""),
                    ("거래구분",     "", "수출" if vals.get("is_export") else "내수"),
                    ("통화",         "", vals.get("currency") or "KRW"),
                    ("기준환율",     "",
                     (f"1 {vals.get('currency')} = {float(vals.get('fx_rate')):,.2f} KRW"
                      if vals.get("fx_rate") and vals.get("currency") and vals.get("currency") != "KRW"
                      else "")),
                    ("수량",         "", str(vals.get("unit_qty") or 1) + "대"),
                    ("단가",
                     "",
                     (f"{float(vals.get('unit_price') or 0):,.0f} {vals.get('currency') or 'KRW'}"
                      if vals.get("unit_price") else "")),
                    ("수주액",
                     "",
                     (f"{float(vals.get('order_amount') or 0):,.0f} {vals.get('currency') or 'KRW'}"
                      + (f" (≈ {float(vals.get('amount_krw') or 0):,.0f} KRW)"
                         if vals.get("amount_krw") else ""))),
                    ("발주일",       "", vals.get("order_date") or ""),
                    ("납기",         "", vals.get("due_date") or ""),
                    ("초기 상태",    "", vals.get("status") or ""),
                    ("PM",           "", vals.get("pm_name") or ""),
                    ("영업담당",     "", vals.get("sales_name") or ""),
                    ("비고",         "", (vals.get("logi_note") or "")[:120]),
                    ("각인",         "", vals.get("engraving") or ""),
                    ("자료경로",     "", vals.get("server_path") or ""),
                ]
                for _label, _ov, _nv in _initial_fields:
                    if _nv:  # 값이 있는 항목만 기록 (빈값은 노이즈)
                        log_project_change(c, new_id, data.get("_changed_by"),
                                           _label, _ov, _nv,
                                           note="등록 시 초기값")
                # v5H226z166: 고객사 담당자 → 주소록 동기화
                _sync_project_contact_to_customer(
                    c, cust_id, vals.get("cc_name"), vals.get("cc_dept"),
                    vals.get("cc_position"), vals.get("cc_phone"),
                    vals.get("cc_email"), vals.get("cc_note"))
                return new_id, code
        except _sq.IntegrityError as e:
            last_err = e
            if "mgmt_code" not in str(e):
                raise  # 다른 컬럼 UNIQUE 위반은 재시도 무의미
            if _provided_code:
                # 사용자가 지정한 코드 충돌 — 재시도해도 같은 코드라 무의미. 즉시 명확한 오류.
                raise ValueError(
                    f"관리코드 '{_provided_code}' 는 이미 등록돼 있습니다. "
                    "다른 코드를 입력하거나, 비워서 자동 발급하세요.")
            continue  # mgmt_code 충돌만 재채번
    raise RuntimeError(f"projects_create_logi: mgmt_code 채번 5회 충돌 — {last_err}")


def log_project_change(c, project_id: int, user_id: int | None,
                        field: str, old_val, new_val, note: str = "") -> None:
    """v5H101: 프로젝트 필드 변경 1건을 project_history 에 기록.
    NULL 안전, 표시용 문자열로 정규화."""
    try:
        ov = "" if old_val is None else str(old_val)
        nv = "" if new_val is None else str(new_val)
        if ov == nv and not note:
            return  # 변동 없음
        c.execute(
            "INSERT INTO project_history(project_id, changed_by, field, "
            "old_value, new_value, note) VALUES(?,?,?,?,?,?)",
            (project_id, user_id, field, ov, nv, note)
        )
    except Exception:
        pass


def get_project_history(project_id: int, limit: int = 50) -> list[dict]:
    """v5H101: 프로젝트 변경 이력 최근 N건 (사용자명 join)."""
    with db_session() as c:
        try:
            rows = c.execute(
                "SELECT ph.*, COALESCE(u.name,'시스템') AS changed_by_name "
                "FROM project_history ph "
                "LEFT JOIN users u ON u.id = ph.changed_by "
                "WHERE ph.project_id=? ORDER BY ph.id DESC LIMIT ?",
                (project_id, int(limit))
            ).fetchall()
            return [dict(r) for r in rows]
        except Exception:
            return []


def get_customer_history(customer_id: int, limit: int = 50) -> list[dict]:
    """v5H113: 고객사 변경 이력 최근 N건 (project_history 동형)."""
    with db_session() as c:
        try:
            rows = c.execute(
                "SELECT ch.*, COALESCE(u.name,'시스템') AS changed_by_name "
                "FROM customer_history ch "
                "LEFT JOIN users u ON u.id = ch.changed_by "
                "WHERE ch.customer_id=? ORDER BY ch.id DESC LIMIT ?",
                (customer_id, int(limit))
            ).fetchall()
            return [dict(r) for r in rows]
        except Exception:
            return []


def get_user_history(user_id: int, limit: int = 50) -> list[dict]:
    """v5H114: 사용자 변경 이력 최근 N건 (project_history 동형)."""
    with db_session() as c:
        try:
            rows = c.execute(
                "SELECT uh.*, COALESCE(u2.name,'시스템') AS changed_by_name "
                "FROM user_history uh "
                "LEFT JOIN users u2 ON u2.id = uh.changed_by "
                "WHERE uh.user_id=? ORDER BY uh.id DESC LIMIT ?",
                (user_id, int(limit))
            ).fetchall()
            return [dict(r) for r in rows]
        except Exception:
            return []


def get_team_history(team_id: int, limit: int = 50) -> list[dict]:
    """v5H114: 팀 변경 이력 최근 N건."""
    with db_session() as c:
        try:
            rows = c.execute(
                "SELECT th.*, COALESCE(u.name,'시스템') AS changed_by_name "
                "FROM team_history th "
                "LEFT JOIN users u ON u.id = th.changed_by "
                "WHERE th.team_id=? ORDER BY th.id DESC LIMIT ?",
                (team_id, int(limit))
            ).fetchall()
            return [dict(r) for r in rows]
        except Exception:
            return []


def get_quotation_history(quotation_id: int, limit: int = 50) -> list[dict]:
    """v5H114: 견적 변경 이력 최근 N건."""
    with db_session() as c:
        try:
            rows = c.execute(
                "SELECT qh.*, COALESCE(u.name,'시스템') AS changed_by_name "
                "FROM quotation_history qh "
                "LEFT JOIN users u ON u.id = qh.changed_by "
                "WHERE qh.quotation_id=? ORDER BY qh.id DESC LIMIT ?",
                (quotation_id, int(limit))
            ).fetchall()
            return [dict(r) for r in rows]
        except Exception:
            return []


def projects_update_logi(pid: int, data: dict) -> str | None:
    current = projects_get_logi(pid)
    if not current:
        return None
    vals = _project_insert_or_update_values(data)
    # v5H101: 변경 전/후 비교 → project_history 자동 기록
    _changed_by = data.get("_changed_by")
    _diff_keys = [
        ("name", "프로젝트명"),
        ("biz_div", "사업부"),
        ("customer_name", "고객사"),
        ("model_name", "모델"),
        ("po_type", "PO유형"),
        ("status", "상태"),
        ("currency", "통화"),
        ("is_export", "거래구분"),
        ("order_amount", "수주액"),
        ("order_date", "발주일"),
        ("due_date", "납기"),
        ("logi_note", "비고"),
        ("project_type", "프로젝트유형"),
        ("parent_project_id", "연관관리번호ID"),
    ]
    _pending_logs = []
    for k, label in _diff_keys:
        old_v = current.get(k) if isinstance(current, dict) else current[k] if k in current.keys() else None
        new_v = vals.get(k)
        if k == "is_export":
            old_v = "수출" if old_v else "내수"
            new_v = "수출" if new_v else "내수"
        elif k == "order_amount":
            try:
                old_v = f"{float(old_v or 0):,.0f}"
                new_v = f"{float(new_v or 0):,.0f}"
            except Exception:
                pass
        if (old_v or "") != (new_v or ""):
            _pending_logs.append((label, old_v, new_v))
    new_code = current["mgmt_code"]
    # v5H86: stage 또는 status 가 won 의미면 관리코드 발급
    # v5H142: CONSUMABLE/SERVICE 는 발급 안 함
    # v5H150: OTHER 는 'K' prefix 로 발급
    # v5H223: CONSUMABLE 도 status 기반 발급 (다른 3종과 동일)
    _ptype_up = (vals.get("project_type") or "NEW_EQUIP").upper()
    needs_code = (
        (vals["stage"] in NEEDS_CODE_STAGES or vals["status"] in WON_STATUSES)
        and (
            (vals["biz_div"] in ("T", "M", "L") and _ptype_up in ("NEW_EQUIP", "CONSUMABLE"))
            or _ptype_up == "OTHER"
        )
    )
    if needs_code and vals["stage"] not in NEEDS_CODE_STAGES:
        vals["stage"] = "수주확정"
    if not new_code and needs_code:
        # v5H225: OTHER → 'E' (Etc.), CONSUMABLE → 'C' (Consumable), NEW_EQUIP → biz_div
        if _ptype_up == "OTHER":
            _code_prefix = "E"
        elif _ptype_up == "CONSUMABLE":
            _code_prefix = "C"
        else:
            _code_prefix = vals["biz_div"]
        new_code = generate_mgmt_code(_code_prefix)
    with db_session() as c:
        # v5H89b: customer_name → customer_id 자동 매핑
        # v5H226z651 (대표 지시): 종사업장 안전연결 — picked(콤보박스) 우선, 후보 1개만 자동연결, 여러 종사업장이면 기존 연결 유지 또는 미연결.
        _ex_cid = (c.execute("SELECT customer_id FROM projects WHERE id=?", (pid,)).fetchone() or [None])[0]
        cust_id, _ = resolve_customer_id(c, vals.get("customer_name"),
                                         picked_id=(data.get("customer_id") if isinstance(data, dict) else None),
                                         prefer_id=_ex_cid)
        c.execute("""
            UPDATE projects
            SET mgmt_code=?, name=?, biz_div=?, customer_id=?, customer_name=?, model_name=?,
                stage=?, po_type=?, status=?, customer_po=?, currency=?,
                order_amount=?, expected_amount=?,
                cc_name=?, cc_dept=?, cc_position=?, cc_phone=?, cc_email=?, cc_note=?,
                two_tier_order=?, secondary_customer=?, final_amount=?,
                order_date=?, due_date=?,
                pm_name=?, sales_name=?, logi_note=?, is_export=?,
                unit_qty=?, unit_price=?,
                project_type=?, parent_project_id=?,
                fx_rate=?, amount_krw=?,
                proposal_date=?, quotation_date=?,
                shipment_form=?, equip_name=?,
                updated_at=?
            WHERE id=?
        """, (new_code, vals["name"], vals["biz_div"], cust_id, vals["customer_name"],
              vals["model_name"], vals["stage"], vals["po_type"], vals["status"],
              vals["customer_po"], vals["currency"], vals["order_amount"],
              vals["expected_amount"],
              vals["cc_name"], vals["cc_dept"], vals["cc_position"],
              vals["cc_phone"], vals["cc_email"], vals["cc_note"],
              vals["two_tier_order"], vals["secondary_customer"], vals["final_amount"],
              vals["order_date"], vals["due_date"], vals["pm_name"],
              vals["sales_name"], vals["logi_note"], vals["is_export"],
              vals.get("unit_qty") or 1, vals.get("unit_price"),
              vals.get("project_type") or "NEW_EQUIP",
              vals.get("parent_project_id"),
              vals.get("fx_rate"), vals.get("amount_krw"),
              vals.get("proposal_date"), vals.get("quotation_date"),
              vals.get("shipment_form") or "ASSEMBLY", vals.get("equip_name"),
              _logi_now(), pid))
        # v5H101: 변경 이력 적재 (UPDATE 성공 후)
        for label, ov, nv in _pending_logs:
            log_project_change(c, pid, _changed_by, label, ov, nv, "")
        # v5H226z166: 고객사 담당자 → 주소록 동기화
        _sync_project_contact_to_customer(
            c, cust_id, vals.get("cc_name"), vals.get("cc_dept"),
            vals.get("cc_position"), vals.get("cc_phone"),
            vals.get("cc_email"), vals.get("cc_note"))
    return new_code


# ==========================================================
# v5H226z208 (2026-06-04 대표 지시): 프로젝트 품목(project_items) CRUD 헬퍼
#   새 규칙 — 관리번호=프로젝트, 그 아래 품목 N개(출고형태별 칸 규칙).
#   ASSEMBLY(ASS'Y)→장비명+모델명 / PARTS(부품)→품명+모델명 / ETC(기타)→품명(모델명 없음)
# ==========================================================
PROJECT_ITEM_SHIPMENT_FORMS = ("ASSEMBLY", "PARTS", "ETC")


def _pi_norm_sf(sf):
    sf = (sf or "").strip().upper()
    if sf in ("PARTS", "부품"):
        return "PARTS"
    if sf in ("ETC", "기타"):
        return "ETC"
    return "ASSEMBLY"


def project_items_list(c, project_id):
    """프로젝트의 품목 목록 (seq, id 순)."""
    rows = c.execute(
        "SELECT * FROM project_items WHERE project_id=? ORDER BY seq, id", (project_id,)
    ).fetchall()
    return [dict(r) for r in rows]


def project_item_create(c, project_id, data: dict) -> int:
    """품목 1건 생성. 반환: 새 품목 id.
    data: shipment_form,item_name,model_name,qty,unit_price,amount,currency,is_export,
          order_customer_id,secondary_customer,final_amount,order_date,due_date,ship_to,note,order_id,seq"""
    sf = _pi_norm_sf(data.get("shipment_form"))
    qty = int(data.get("qty") or 1)
    if qty < 1:
        qty = 1
    up = float(data.get("unit_price") or 0)
    _amt = data.get("amount")
    amt = float(_amt) if (_amt not in (None, "")) else (up * qty)
    model = (data.get("model_name") or "") if sf != "ETC" else ""   # 기타(ETC)는 모델명 없음
    seq = data.get("seq")
    if not seq:
        seq = (c.execute("SELECT COALESCE(MAX(seq),0)+1 FROM project_items WHERE project_id=?",
                         (project_id,)).fetchone()[0]) or 1
    cur = c.execute(
        "INSERT INTO project_items(project_id, seq, shipment_form, item_name, model_name, "
        "qty, unit_price, amount, currency, is_export, order_customer_id, "
        "secondary_customer, final_amount, order_date, due_date, ship_to, note, order_id, updated_at) "
        "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?, datetime('now','localtime'))",
        (project_id, seq, sf, (data.get("item_name") or "").strip(), model,
         qty, up, amt, (data.get("currency") or "KRW").upper(),
         int(data.get("is_export") or 0), data.get("order_customer_id"),
         data.get("secondary_customer"), data.get("final_amount"),
         data.get("order_date"), data.get("due_date"), data.get("ship_to"),
         data.get("note"), data.get("order_id")))
    return cur.lastrowid


def project_item_update(c, item_id, data: dict) -> None:
    """품목 수정 (전달된 필드만)."""
    allow = {"shipment_form", "item_name", "model_name", "qty", "unit_price", "amount",
             "currency", "is_export", "order_customer_id", "secondary_customer",
             "final_amount", "order_date", "due_date", "ship_to", "note", "order_id", "seq"}
    d = dict(data)
    if "shipment_form" in d:
        d["shipment_form"] = _pi_norm_sf(d.get("shipment_form"))
        if d["shipment_form"] == "ETC":
            d["model_name"] = ""   # 기타는 모델명 없음
    # 수량·단가만 바뀌고 금액 미지정 시 금액 자동 재계산 (단가×수량) — 스테일 방지
    if ("qty" in d or "unit_price" in d) and "amount" not in d:
        _cur = c.execute("SELECT qty, unit_price FROM project_items WHERE id=?", (item_id,)).fetchone()
        if _cur:
            _q = int(d.get("qty", _cur["qty"]) or 1)
            _up = float(d.get("unit_price", _cur["unit_price"]) or 0)
            d["amount"] = _up * _q
    sets, vals = [], []
    for k, v in d.items():
        if k in allow:
            sets.append(f"{k}=?")
            vals.append(v)
    if not sets:
        return
    sets.append("updated_at=datetime('now','localtime')")
    vals.append(item_id)
    c.execute(f"UPDATE project_items SET {','.join(sets)} WHERE id=?", vals)


def project_item_delete(c, item_id) -> None:
    c.execute("DELETE FROM project_items WHERE id=?", (item_id,))


# ==========================================================
# v5H226z215 (2026-06-05 대표 지시): A/S 관리(수리 접수·처리) 헬퍼 — 라이프밸류 전용 신설.
#   흐름: 접수 → 수리중 → 발송완료 (+보류/반송). RMA(고객 택배→수리→재발송).
# ==========================================================
AS_STATUSES = ("접수", "수리중", "발송완료", "보류", "반송")


def generate_as_no(c, ref_date=None) -> str:
    """A/S 접수번호 — AS-YYMMDD-NNN (당일 시퀀스)."""
    d = ref_date or _date.today()
    base = f"AS-{d.strftime('%y%m%d')}-"
    rows = c.execute("SELECT as_no FROM as_tickets WHERE as_no LIKE ?", (base + "%",)).fetchall()
    max_seq = 0
    for r in rows:
        an = (r["as_no"] if hasattr(r, "keys") else r[0]) or ""
        tail = an.rsplit("-", 1)
        if len(tail) == 2 and tail[1].isdigit():
            max_seq = max(max_seq, int(tail[1]))
    return f"{base}{max_seq + 1:03d}"


def _as_paid(v) -> int:
    return 1 if str(v or "").strip() in ("1", "on", "true", "yes", "유상") else 0


def _as_fee(v) -> float:
    try:
        return float(str(v or "0").replace(",", "") or 0)
    except ValueError:
        return 0.0


def as_ticket_create(c, data: dict) -> dict:
    """A/S 접수 1건 생성. 반환: {id, as_no}."""
    as_no = generate_as_no(c)
    rd = (data.get("received_date") or "").strip() or _date.today().isoformat()
    status = (data.get("status") or "접수").strip()
    if status not in AS_STATUSES:
        status = "접수"
    cur = c.execute(
        "INSERT INTO as_tickets(as_no, received_date, biz_div, customer_id, customer_name, "
        "customer_phone, customer_address, product_id, product_name, model_name, purchase_date, "
        "symptom, inbound_tracking, outbound_tracking, shipped_date, is_paid, repair_fee, "
        "repair_note, status, assignee_id, note, created_by, updated_at) "
        "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?, datetime('now','localtime'))",
        (as_no, rd, (data.get("biz_div") or "L"),
         data.get("customer_id"), (data.get("customer_name") or "").strip(),
         (data.get("customer_phone") or "").strip(), (data.get("customer_address") or "").strip(),
         data.get("product_id"), (data.get("product_name") or "").strip(),
         (data.get("model_name") or "").strip(), (data.get("purchase_date") or "").strip() or None,
         (data.get("symptom") or "").strip(),
         (data.get("inbound_tracking") or "").strip(), (data.get("outbound_tracking") or "").strip(),
         (data.get("shipped_date") or "").strip() or None,
         _as_paid(data.get("is_paid")), _as_fee(data.get("repair_fee")),
         (data.get("repair_note") or "").strip(),
         status, data.get("assignee_id"), (data.get("note") or "").strip(),
         data.get("created_by")))
    as_id = cur.lastrowid
    try:
        c.execute("INSERT INTO as_status_history(as_id, from_status, to_status, changed_by, note) "
                  "VALUES(?,?,?,?,?)", (as_id, None, status, data.get("created_by"), "접수 등록"))
    except Exception:
        pass
    return {"id": as_id, "as_no": as_no}


def as_ticket_list(c, status: str = "", q: str = ""):
    """A/S 접수 목록 (상태 필터 + 검색)."""
    sql = ("SELECT t.*, (SELECT name FROM customers WHERE customers.id=t.customer_id) AS cust_link, "
           "(SELECT name FROM users WHERE users.id=t.assignee_id) AS assignee_name "
           "FROM as_tickets t WHERE 1=1")
    params: list = []
    if status and status in AS_STATUSES:
        sql += " AND t.status=?"
        params.append(status)
    if q:
        sql += (" AND (t.as_no LIKE ? OR t.customer_name LIKE ? OR t.product_name LIKE ? "
                "OR t.model_name LIKE ? OR t.customer_phone LIKE ?)")
        like = f"%{q}%"
        params += [like] * 5
    sql += " ORDER BY t.id DESC"
    return [dict(r) for r in c.execute(sql, params).fetchall()]


def as_ticket_get(c, as_id):
    """A/S 접수 1건 + 상태 이력."""
    r = c.execute(
        "SELECT t.*, (SELECT name FROM customers WHERE customers.id=t.customer_id) AS cust_link, "
        "(SELECT name FROM users WHERE users.id=t.assignee_id) AS assignee_name "
        "FROM as_tickets t WHERE t.id=?", (as_id,)).fetchone()
    if not r:
        return None
    d = dict(r)
    d["history"] = [dict(h) for h in c.execute(
        "SELECT h.*, (SELECT name FROM users WHERE users.id=h.changed_by) AS changer "
        "FROM as_status_history h WHERE h.as_id=? ORDER BY h.id", (as_id,)).fetchall()]
    return d


def as_ticket_update(c, as_id, data: dict) -> None:
    """A/S 접수 수정 (전달된 필드만). 상태 변경은 as_status_change 로."""
    allow = {"received_date", "biz_div", "customer_id", "customer_name", "customer_phone",
             "customer_address", "product_id", "product_name", "model_name", "purchase_date",
             "symptom", "inbound_tracking", "outbound_tracking", "shipped_date", "is_paid",
             "repair_fee", "repair_note", "assignee_id", "note"}
    sets, vals = [], []
    for k, v in data.items():
        if k not in allow:
            continue
        if k == "is_paid":
            v = _as_paid(v)
        elif k == "repair_fee":
            v = _as_fee(v)
        sets.append(f"{k}=?")
        vals.append(v)
    if not sets:
        return
    sets.append("updated_at=datetime('now','localtime')")
    vals.append(as_id)
    c.execute(f"UPDATE as_tickets SET {','.join(sets)} WHERE id=?", vals)


def as_status_change(c, as_id, to_status: str, changed_by=None, note: str = "") -> dict:
    """A/S 상태 변경 + 이력 기록. '발송완료'로 가면 발송일 자동 기록(비어있으면)."""
    if to_status not in AS_STATUSES:
        return {"ok": False, "message": f"잘못된 상태: {to_status}"}
    cur = c.execute("SELECT status FROM as_tickets WHERE id=?", (as_id,)).fetchone()
    if not cur:
        return {"ok": False, "message": "접수 건을 찾을 수 없음"}
    from_status = cur["status"]
    if from_status == to_status:
        return {"ok": True, "message": "변경 없음", "status": to_status}
    if to_status == "발송완료":
        c.execute("UPDATE as_tickets SET status=?, "
                  "shipped_date=COALESCE(NULLIF(shipped_date,''),?), "
                  "updated_at=datetime('now','localtime') WHERE id=?",
                  (to_status, _date.today().isoformat(), as_id))
    else:
        c.execute("UPDATE as_tickets SET status=?, updated_at=datetime('now','localtime') WHERE id=?",
                  (to_status, as_id))
    c.execute("INSERT INTO as_status_history(as_id, from_status, to_status, changed_by, note) "
              "VALUES(?,?,?,?,?)", (as_id, from_status, to_status, changed_by, note or ""))
    return {"ok": True, "status": to_status, "from": from_status}


def projects_delete_logi(pid: int, force: bool = False) -> None:
    """v5H73b: FK 자식 row 일괄 정리 후 프로젝트 삭제 (FOREIGN KEY 충돌 방지).
    v5H226z193: force=True 면 관리코드(z176) 삭제 잠금을 우회 — '데모 사업부별 초기화' 전용.

    projects 를 참조하는 테이블:
      orders, purchase_orders, tasks, project_phases, project_milestones,
      project_burndown_snapshots, project_forecasts, issues, tickets,
      changes, qc_inspection_reports, work_orders, stock_movements,
      change_impacts (간접)
    각각 SET NULL 또는 DELETE 처리."""
    with db_session() as c:
        # v5H226z176 (2026-06-03 대표 지시): 관리코드(수주확정) 부여 건 삭제 차단.
        #   관리코드는 수주확정 시점에만 발급되는 정식 사업번호 → 매출·납품·수금 이력이
        #   엮인 실거래 기록. 삭제 대신 상태를 '취소'로 변경해야 함. (데이터층 최종 안전망)
        try:
            _mrow = c.execute("SELECT mgmt_code FROM projects WHERE id=?", (pid,)).fetchone()
        except Exception:
            _mrow = None
        _mc = (_mrow[0] if _mrow else None)
        if _mc and str(_mc).strip() and not force:
            raise PermissionError(
                f"수주확정 건(관리코드 {str(_mc).strip()})은 삭제할 수 없습니다. "
                f"매출·납품·수금 이력 보호를 위해, 상세화면에서 상태를 '취소'로 변경하세요.")
        # 1. 자식 row가 자체 자식을 가진 것들 — 먼저 손자부터 정리
        # 1-a) orders 의 자식: order_items, invoices, receipts_payment,
        #      shipments, order_status_history
        order_ids = [r[0] for r in c.execute(
            "SELECT id FROM orders WHERE project_id=?", (pid,)).fetchall()]
        for oid in order_ids:
            for sql in [
                "DELETE FROM order_items WHERE order_id=?",
                "DELETE FROM invoices WHERE order_id=?",
                "DELETE FROM receipts_payment WHERE order_id=?",
                "DELETE FROM shipments WHERE order_id=?",
                "DELETE FROM order_status_history WHERE order_id=?",
                "DELETE FROM production_orders WHERE order_id=?",
            ]:
                try: c.execute(sql, (oid,))
                except Exception: pass
        try: c.execute("DELETE FROM orders WHERE project_id=?", (pid,))
        except Exception: pass

        # v5H226z208: 프로젝트 품목 정리 (orders 삭제 후 — order_id FK 보유)
        try: c.execute("DELETE FROM project_items WHERE project_id=?", (pid,))
        except Exception: pass

        # 1-b) purchase_orders 자식: po_items
        po_ids = [r[0] for r in c.execute(
            "SELECT id FROM purchase_orders WHERE project_id=?", (pid,)).fetchall()]
        for poid in po_ids:
            try: c.execute("DELETE FROM po_items WHERE po_id=?", (poid,))
            except Exception: pass
        try: c.execute("DELETE FROM purchase_orders WHERE project_id=?", (pid,))
        except Exception: pass

        # 1-c) work_orders 자식: work_order_items
        wo_ids = [r[0] for r in c.execute(
            "SELECT id FROM work_orders WHERE project_id=?", (pid,)).fetchall()]
        for woid in wo_ids:
            try: c.execute("DELETE FROM work_order_items WHERE wo_id=?", (woid,))
            except Exception: pass
        try: c.execute("DELETE FROM work_orders WHERE project_id=?", (pid,))
        except Exception: pass

        # 1-d) issues 자식: issue_logs, change_impacts
        issue_ids = [r[0] for r in c.execute(
            "SELECT id FROM issues WHERE project_id=?", (pid,)).fetchall()]
        for iid in issue_ids:
            try: c.execute("DELETE FROM issue_logs WHERE issue_id=?", (iid,))
            except Exception: pass
        try: c.execute("DELETE FROM issues WHERE project_id=?", (pid,))
        except Exception: pass

        # 1-e) changes 자식: change_impacts, change_reads
        change_ids = [r[0] for r in c.execute(
            "SELECT id FROM changes WHERE project_id=?", (pid,)).fetchall()]
        for cid in change_ids:
            for sql in [
                "DELETE FROM change_impacts WHERE change_id=?",
                "DELETE FROM change_reads WHERE change_id=?",
            ]:
                try: c.execute(sql, (cid,))
                except Exception: pass
        try: c.execute("DELETE FROM changes WHERE project_id=?", (pid,))
        except Exception: pass

        # 1-f) project_phases / milestones / burndown / forecasts (CASCADE 일 가능성)
        for sql in [
            "DELETE FROM project_phases WHERE project_id=?",
            "DELETE FROM project_milestones WHERE project_id=?",
            "DELETE FROM project_burndown_snapshots WHERE project_id=?",
            "DELETE FROM project_forecasts WHERE project_id=?",
        ]:
            try: c.execute(sql, (pid,))
            except Exception: pass

        # 1-g) tasks 자식: task_comments, task_reactions, task_meta, etc.
        task_ids = [r[0] for r in c.execute(
            "SELECT id FROM tasks WHERE project_id=?", (pid,)).fetchall()]
        for tid in task_ids:
            for sql in [
                "DELETE FROM task_comments WHERE task_id=?",
                "DELETE FROM task_reactions WHERE task_id=?",
                "DELETE FROM comment_mentions WHERE comment_id IN (SELECT id FROM task_comments WHERE task_id=?)",
                "DELETE FROM task_delegations WHERE task_id=?",
            ]:
                try: c.execute(sql, (tid,))
                except Exception: pass
        # tasks 는 SET NULL 이 안전 — 이력 보존
        try:
            c.execute("UPDATE tasks SET project_id=NULL WHERE project_id=?", (pid,))
        except Exception:
            try: c.execute("DELETE FROM tasks WHERE project_id=?", (pid,))
            except Exception: pass

        # 1-h) 기타 — qc_inspection_reports, tickets, stock_movements
        for sql in [
            "UPDATE qc_inspection_reports SET order_id=NULL WHERE order_id IN (SELECT id FROM orders WHERE project_id=?)",
            "UPDATE tickets SET project_id=NULL WHERE project_id=?",
            "UPDATE stock_movements SET project_id=NULL WHERE project_id=?",
        ]:
            try: c.execute(sql, (pid,))
            except Exception: pass

        # 1-i) v5H226z259 (대표 보고): 안전망 — projects 를 '참조하는 모든 컬럼' 동적 정리.
        #   기존엔 project_id 컬럼만 정리했으나, 이름이 다른 FK 참조
        #   (consumable_order_items.linked_project_id=소모품→장비 연결, projects.parent_project_id 등)는
        #   못 잡아 본체 삭제 시 FK 충돌→삭제 실패. FK 목록 기준으로 모든 참조 컬럼을 SET NULL(또는 DELETE).
        try:
            all_tables = [r[0] for r in c.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
            ).fetchall()]
            for tname in all_tables:
                ref_cols = []
                try:
                    for _fk in c.execute(f"PRAGMA foreign_key_list({tname})").fetchall():
                        # _fk: (id, seq, table, from, to, on_update, on_delete, match)
                        if _fk[2] == "projects" and _fk[3]:
                            ref_cols.append(_fk[3])
                except Exception:
                    pass
                # FK 선언이 없어도 관례적 project_id 컬럼 보강
                try:
                    cols = [r[1] for r in c.execute(f"PRAGMA table_info({tname})").fetchall()]
                except Exception:
                    cols = []
                if tname != "projects" and "project_id" in cols and "project_id" not in ref_cols:
                    ref_cols.append("project_id")
                for _col in set(ref_cols):
                    try:
                        c.execute(f"UPDATE {tname} SET {_col}=NULL WHERE {_col}=?", (pid,))
                    except Exception:
                        try:
                            c.execute(f"DELETE FROM {tname} WHERE {_col}=?", (pid,))
                        except Exception:
                            pass
        except Exception:
            pass

        # 1-j) v5H226z263 (대표 지시·연결성 감사): parent_project_id 명시 정리.
        #   근본원인 = parent_project_id 는 FK 미선언 + projects 자기참조라, 위 z259 동적 패스의
        #   'tname != projects' 가드에서 누락 → 부모 장비 삭제 시 자식(소모품/수리)의 부모참조가
        #   끊긴 채 남아(dangling) 상세의 '연결 관리번호'가 없는 부모를 가리킴. SET NULL 로 끊는다.
        try:
            _pcols = [r[1] for r in c.execute("PRAGMA table_info(projects)").fetchall()]
            if "parent_project_id" in _pcols:
                c.execute("UPDATE projects SET parent_project_id=NULL WHERE parent_project_id=?", (pid,))
        except Exception:
            pass

        # 2. 본체 삭제
        c.execute("DELETE FROM projects WHERE id = ?", (pid,))


def projects_count_logi() -> dict:
    with db_session() as c:
        total = c.execute(
            "SELECT COUNT(*) FROM projects WHERE biz_div IS NOT NULL AND biz_div != ''"
        ).fetchone()[0]
        with_code = c.execute(
            "SELECT COUNT(*) FROM projects WHERE mgmt_code IS NOT NULL AND mgmt_code != ''"
        ).fetchone()[0]
        in_progress = c.execute(
            "SELECT COUNT(*) FROM projects WHERE status = '진행중'"
        ).fetchone()[0]
        by_div = c.execute(
            "SELECT biz_div, COUNT(*) AS n FROM projects WHERE biz_div IS NOT NULL AND biz_div != '' GROUP BY biz_div"
        ).fetchall()
        by_stage = c.execute(
            "SELECT stage, COUNT(*) AS n FROM projects WHERE stage IS NOT NULL AND stage != '' GROUP BY stage"
        ).fetchall()
    return {
        "total": total,
        "with_code": with_code,
        "in_progress": in_progress,
        "by_div": {(BIZ_NAME.get(r["biz_div"]) or r["biz_div"] or "(미지정)"): r["n"]
                   for r in by_div},
        "by_stage": {r["stage"] or "(미지정)": r["n"] for r in by_stage},
    }


# =====================================================
# HAIST WORKS — 발주 모듈 (공급사 + PO 헤더/라인)
# =====================================================
PO_STATUSES = ["작성중", "발주완료", "부분입고", "입고완료", "취소"]
PO_TYPES_KIND = ["일반", "긴급", "정기"]
SHIPPING_TERMS = ["EXW", "FOB", "CIF", "DDP", "국내"]
PAYMENT_TERMS = ["선금", "현금", "30일", "60일", "기타"]


# ── 공급사 (suppliers) CRUD ──────────────────────────────
def suppliers_list(q: str = "", active_only: bool = False):
    sql = "SELECT * FROM suppliers WHERE 1=1"
    params = []
    if q:
        sql += " AND (name LIKE ? OR code LIKE ? OR contact LIKE ? OR country LIKE ?)"
        like = f"%{q}%"
        params += [like] * 4
    if active_only:
        sql += " AND is_active = 1"
    sql += " ORDER BY is_active DESC, name"
    with db_session() as c:
        return c.execute(sql, params).fetchall()


def supplier_get(sid: int):
    with db_session() as c:
        return c.execute("SELECT * FROM suppliers WHERE id = ?", (sid,)).fetchone()


def _validate_currency(data: dict) -> str:
    """v5H113 LOW#20/#21: currency 화이트리스트 정규화."""
    cur = (data.get("currency") or "KRW").strip().upper() or "KRW"
    if cur not in CURRENCY_OPTIONS:
        raise ValueError(f"통화는 {', '.join(CURRENCY_OPTIONS)} 중 하나여야 합니다. (입력: {cur})")
    return cur


def supplier_create(data: dict) -> int:
    # v5H113 LOW#20: currency 화이트리스트 + 이름 필수
    if not (data.get("name") or "").strip():
        raise ValueError("공급사명은 필수입니다.")
    data["currency"] = _validate_currency(data)
    now = _logi_now()
    base_cols = ["name", "code", "contact", "email", "phone", "country", "currency",
                 "payment_terms", "note", "is_active", "created_at", "updated_at"]
    base_vals = [
        (data.get("name") or "").strip(),
        (data.get("code") or "").strip() or None,
        (data.get("contact") or "").strip(),
        (data.get("email") or "").strip(),
        (data.get("phone") or "").strip(),
        (data.get("country") or "").strip(),
        (data.get("currency") or "KRW").strip() or "KRW",
        (data.get("payment_terms") or "").strip(),
        (data.get("note") or "").strip(),
        1 if data.get("is_active", 1) else 0,
        now, now,
    ]
    # v5H226z89/z90: 확장 컬럼 — 존재 시에만 포함 (마이그레이션 호환 가드)
    ext_pairs = [
        ("biz_no",      (str(data.get("biz_no") or "")).strip() or None),
        ("ceo_name",    (str(data.get("ceo_name") or "")).strip() or None),
        ("manage_dept", (str(data.get("manage_dept") or "")).strip() or None),
        ("contact2",    (str(data.get("contact2") or "")).strip() or None),
        ("phone2",      (str(data.get("phone2") or "")).strip() or None),
        ("email2",      (str(data.get("email2") or "")).strip() or None),
        # v5H226z117 (2026-05-19) — 협력사 주소
        ("address",     (str(data.get("address") or "")).strip() or None),
    ]
    with db_session() as c:
        existing = {r[1] for r in c.execute("PRAGMA table_info(suppliers)").fetchall()}
        cols, vals = list(base_cols), list(base_vals)
        for col, val in ext_pairs:
            if col in existing:
                cols.append(col)
                vals.append(val)
        ph = ",".join(["?"] * len(cols))
        cur = c.execute(f"INSERT INTO suppliers ({','.join(cols)}) VALUES ({ph})", vals)
        return cur.lastrowid


def supplier_update(sid: int, data: dict) -> None:
    # v5H113 LOW#20: currency 화이트리스트
    if not (data.get("name") or "").strip():
        raise ValueError("공급사명은 필수입니다.")
    data["currency"] = _validate_currency(data)
    base_fields = ["name", "code", "contact", "email", "phone", "country",
                   "currency", "payment_terms", "note", "is_active"]
    base_vals = [
        (data.get("name") or "").strip(),
        (data.get("code") or "").strip() or None,
        (data.get("contact") or "").strip(),
        (data.get("email") or "").strip(),
        (data.get("phone") or "").strip(),
        (data.get("country") or "").strip(),
        (data.get("currency") or "KRW").strip() or "KRW",
        (data.get("payment_terms") or "").strip(),
        (data.get("note") or "").strip(),
        1 if data.get("is_active", 1) else 0,
    ]
    # v5H226z89/z90: 확장 컬럼 — 존재 시에만 SET
    ext_pairs = [
        ("biz_no",      (str(data.get("biz_no") or "")).strip() or None),
        ("ceo_name",    (str(data.get("ceo_name") or "")).strip() or None),
        ("manage_dept", (str(data.get("manage_dept") or "")).strip() or None),
        ("contact2",    (str(data.get("contact2") or "")).strip() or None),
        ("phone2",      (str(data.get("phone2") or "")).strip() or None),
        ("email2",      (str(data.get("email2") or "")).strip() or None),
        # v5H226z117 (2026-05-19) — 협력사 주소
        ("address",     (str(data.get("address") or "")).strip() or None),
    ]
    with db_session() as c:
        existing = {r[1] for r in c.execute("PRAGMA table_info(suppliers)").fetchall()}
        fields, vals = list(base_fields), list(base_vals)
        for col, val in ext_pairs:
            if col in existing:
                fields.append(col)
                vals.append(val)
        sets = ", ".join(f"{f}=?" for f in fields) + ", updated_at=?"
        c.execute(f"UPDATE suppliers SET {sets} WHERE id=?", vals + [_logi_now(), sid])


def supplier_delete(sid: int) -> None:
    """v5H119: 공통 헬퍼 사용. 폴백: v5H112 인라인."""
    with db_session() as c:
        try:
            res = _safe_delete_with_cascade(
                c, "suppliers", sid, fk_column="supplier_id",
                explicit_children=[
                    ("UPDATE purchase_orders SET supplier_id=NULL WHERE supplier_id=?", (sid,)),
                    ("UPDATE part_prices SET supplier_id=NULL WHERE supplier_id=?", (sid,)),
                ],
            )
            if res.get("ok"):
                return
        except Exception:
            pass
        for sql in [
            "UPDATE purchase_orders SET supplier_id=NULL WHERE supplier_id=?",
            "UPDATE part_prices SET supplier_id=NULL WHERE supplier_id=?",
        ]:
            try: c.execute(sql, (sid,))
            except Exception: pass
        try:
            all_tables = [r[0] for r in c.execute(
                "SELECT name FROM sqlite_master WHERE type='table' "
                "AND name NOT LIKE 'sqlite_%' AND name <> 'suppliers'"
            ).fetchall()]
            for tname in all_tables:
                try:
                    cols = [r[1] for r in c.execute(f"PRAGMA table_info({tname})").fetchall()]
                except Exception:
                    continue
                if "supplier_id" not in cols:
                    continue
                try:
                    c.execute(f"UPDATE {tname} SET supplier_id=NULL WHERE supplier_id=?", (sid,))
                except Exception:
                    try: c.execute(f"DELETE FROM {tname} WHERE supplier_id=?", (sid,))
                    except Exception: pass
        except Exception:
            pass
        c.execute("DELETE FROM suppliers WHERE id = ?", (sid,))


def suppliers_count() -> dict:
    with db_session() as c:
        total = c.execute("SELECT COUNT(*) FROM suppliers").fetchone()[0]
        active = c.execute("SELECT COUNT(*) FROM suppliers WHERE is_active = 1").fetchone()[0]
    return {"total": total, "active": active}


def teams_name_list() -> list[str]:
    """v5H226z90: 부서명 목록 (display_order 순) — 공급사 '관리부서' 드롭다운용."""
    with db_session() as c:
        return [r[0] for r in c.execute(
            "SELECT name FROM teams ORDER BY display_order, id").fetchall()]


# =====================================================
# v5H226z88 (2026-05-14) — 공급사 일괄 등록 (Excel)
# 자재 일괄등록(parts_bulk_import_excel)과 동일 3단계 패턴: 검증·미리보기·확정
# =====================================================
_SUPPLIERS_IMPORT_HEADER_MAP = {
    # name (필수)
    "공급사명": "name", "공급사": "name", "거래처명": "name", "거래처": "name",
    "회사명": "name", "업체명": "name", "업체": "name", "name": "name",
    "supplier": "name", "vendor": "name", "suppliername": "name", "companyname": "name",
    # code
    "코드": "code", "공급사코드": "code", "거래처코드": "code", "업체코드": "code",
    "code": "code", "suppliercode": "code", "vendorcode": "code",
    # biz_no — 사업자등록번호
    "사업자등록번호": "biz_no", "사업자번호": "biz_no", "사업자등록": "biz_no",
    "등록번호": "biz_no", "biz_no": "biz_no", "bizno": "biz_no",
    "businessnumber": "biz_no", "businessno": "biz_no", "사업자": "biz_no",
    # ceo_name — 대표자
    "대표자": "ceo_name", "대표": "ceo_name", "대표자명": "ceo_name", "대표이사": "ceo_name",
    "ceo": "ceo_name", "ceoname": "ceo_name", "representative": "ceo_name",
    # manage_dept — 관리부서
    "관리부서": "manage_dept", "담당부서": "manage_dept", "부서": "manage_dept",
    "managedept": "manage_dept", "dept": "manage_dept", "department": "manage_dept",
    # contact (담당자1)
    "담당자": "contact", "담당": "contact", "담당자명": "contact", "담당자이름": "contact",
    "담당자1": "contact", "contact": "contact", "manager": "contact", "contactperson": "contact",
    # contact2 (담당자2)
    "담당자2": "contact2", "담당자②": "contact2", "보조담당자": "contact2",
    "부담당자": "contact2", "담당자2명": "contact2", "contact2": "contact2",
    # email (담당자1)
    "이메일": "email", "메일": "email", "이메일주소": "email", "이메일1": "email",
    "email": "email", "mail": "email",
    # email2 (담당자2)
    "이메일2": "email2", "메일2": "email2", "email2": "email2",
    # phone (담당자1)
    "전화": "phone", "전화번호": "phone", "연락처": "phone", "휴대폰": "phone",
    "전화1": "phone", "전화번호1": "phone",
    "전화연락처": "phone", "phone": "phone", "tel": "phone", "telephone": "phone",
    # phone2 (담당자2)
    "전화2": "phone2", "전화번호2": "phone2", "연락처2": "phone2",
    "휴대폰2": "phone2", "phone2": "phone2",
    # country
    "국가": "country", "나라": "country", "국적": "country", "country": "country",
    # currency
    "통화": "currency", "화폐": "currency", "currency": "currency",
    # payment_terms
    "결제조건": "payment_terms", "결제": "payment_terms", "지불조건": "payment_terms",
    "결제방식": "payment_terms", "paymentterms": "payment_terms", "terms": "payment_terms",
    # note
    "비고": "note", "메모": "note", "특이사항": "note", "note": "note",
    "remark": "note", "memo": "note",
    # is_active
    "활성": "is_active", "활성여부": "is_active", "사용여부": "is_active",
    "상태": "is_active", "active": "is_active", "isactive": "is_active",
    # v5H226z117 (2026-05-19) — 주소
    "주소": "address", "address": "address", "주소지": "address",
    "회사주소": "address", "본사주소": "address", "소재지": "address",
}


def suppliers_bulk_import_excel(file_path: str, dry_run: bool = True) -> dict:
    """Excel(.xlsx) 공급사 일괄 등록. 1행=헤더, 2행~=데이터.
    dry_run=True: 검증·미리보기만 (DB 변경 0).  dry_run=False: 실제 등록.
    상태: ok / error(필수누락·통화오류) / dup(공급사명 또는 코드 기존 중복)
    반환: {total, parsed, skipped_empty, header_map, unmapped_headers,
           preview[{row_no,data,status,errors}], summary{ok,error,dup,inserted}, inserted_ids}
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "openpyxl 미설치 — pip install openpyxl"}

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return {"error": "엑셀이 비어있습니다"}

    header_map: dict = {}
    unmapped: list = []
    used_fields: set = set()
    for i, h in enumerate(header_row):
        if h is None:
            continue
        key = _norm_header_key(str(h))
        fld = _SUPPLIERS_IMPORT_HEADER_MAP.get(key)
        if fld and fld not in used_fields:
            header_map[i] = fld
            used_fields.add(fld)
        else:
            unmapped.append({"col_idx": i, "header": str(h)})

    if "name" not in header_map.values():
        wb.close()
        return {
            "error": "필수 컬럼 누락 — '공급사명/거래처명/name' 컬럼이 필요합니다",
            "header_map": {str(k): v for k, v in header_map.items()},
            "unmapped_headers": unmapped,
        }

    # 기존 공급사 — 중복 검사용
    with db_session() as c:
        existing = c.execute("SELECT id, name, code FROM suppliers").fetchall()
    name_set: dict = {}
    code_set: dict = {}
    for r in existing:
        nm = (r["name"] or "").strip().lower()
        if nm:
            name_set.setdefault(nm, r["id"])
        cd = (r["code"] or "").strip().lower()
        if cd:
            code_set.setdefault(cd, r["id"])

    preview: list = []
    inserted_ids: list = []
    skipped_empty = total = ok_cnt = err_cnt = dup_cnt = 0
    seen_names: set = set()
    seen_codes: set = set()

    row_no = 1
    for row in rows_iter:
        row_no += 1
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            skipped_empty += 1
            continue
        total += 1
        data: dict = {}
        for ci, fld in header_map.items():
            if ci < len(row):
                val = row[ci]
                if isinstance(val, str):
                    val = val.strip()
                if val is not None and val != "":
                    data[fld] = val
        name = str(data.get("name") or "").strip()
        code = str(data.get("code") or "").strip()
        errors: list = []
        status = "ok"

        if not name:
            errors.append("공급사명 필수")

        # 통화 검증
        cur = str(data.get("currency") or "").strip().upper()
        if cur and cur not in CURRENCY_OPTIONS:
            errors.append(f"통화 값 오류 '{cur}' ({'/'.join(CURRENCY_OPTIONS)})")
            data.pop("currency", None)
        elif cur:
            data["currency"] = cur

        # is_active 정규화
        if "is_active" in data:
            iav = str(data.get("is_active")).strip().lower()
            data["is_active"] = 0 if iav in ("0", "n", "no", "false", "비활성", "중지", "미사용") else 1

        # 중복 검사 (기존 DB + 같은 파일 내)
        if name and not errors:
            nl, cl = name.lower(), code.lower()
            if nl in name_set:
                errors.append(f"공급사명 중복 — 기존 #{name_set[nl]}")
                status = "dup"
            elif nl in seen_names:
                errors.append("공급사명 중복 — 같은 파일 내")
                status = "dup"
            elif cl and cl in code_set:
                errors.append(f"코드 중복 — 기존 #{code_set[cl]}")
                status = "dup"
            elif cl and cl in seen_codes:
                errors.append("코드 중복 — 같은 파일 내")
                status = "dup"
            else:
                seen_names.add(nl)
                if cl:
                    seen_codes.add(cl)

        if errors and status != "dup":
            status = "error"
            err_cnt += 1
        elif status == "dup":
            dup_cnt += 1
        else:
            ok_cnt += 1

        preview.append({"row_no": row_no, "data": data,
                        "status": status, "errors": errors})

    wb.close()

    if not dry_run:
        for p in preview:
            if p["status"] == "ok" and not p["errors"]:
                try:
                    sid = supplier_create(p["data"])
                    inserted_ids.append({"row_no": p["row_no"], "id": sid,
                                          "name": p["data"].get("name"),
                                          "code": p["data"].get("code")})
                    p["status"] = "inserted"
                    p["inserted_id"] = sid
                except Exception as e:
                    p["status"] = "error"
                    p["errors"].append(str(e))
                    err_cnt += 1
                    ok_cnt = max(0, ok_cnt - 1)

    return {
        "total": total, "parsed": total, "skipped_empty": skipped_empty,
        "header_map": {str(k): v for k, v in header_map.items()},
        "unmapped_headers": unmapped,
        "preview": preview[:200],
        "preview_truncated": len(preview) > 200,
        "summary": {"ok": ok_cnt, "error": err_cnt, "dup": dup_cnt,
                    "inserted": len(inserted_ids)},
        "inserted_ids": inserted_ids,
        "dry_run": dry_run,
    }


# ── 발주번호 자동 채번 ─────────────────────────────────
def generate_po_number(today=None) -> str:
    """발주번호 채번: PO-YYMMDD-NNN (같은 날짜 내 순차)"""
    today = today or _date.today()
    yymmdd = today.strftime("%y%m%d")
    prefix = f"PO-{yymmdd}-"
    with db_session() as c:
        rows = c.execute(
            "SELECT po_number FROM purchase_orders WHERE po_number LIKE ?",
            (f"{prefix}%",),
        ).fetchall()
    max_seq = 0
    for r in rows:
        tail = (r["po_number"] or "")[len(prefix):]
        if tail.isdigit():
            max_seq = max(max_seq, int(tail))
    return f"{prefix}{max_seq + 1:03d}"


# ── 발주 (purchase_orders + po_items) CRUD ─────────────

# v5H226z58 (2026-05-11) 자재모듈 표준 v2 — VAT 계산 헬퍼
# 매뉴얼 §3.2 표준 공식:
#   vat_excluded (VAT 미포함): supply = qty×price,        vat = supply × 0.1
#   vat_included (VAT 포함):   supply = (qty×price)/1.1,  vat = base - supply
#   vat_none     (VAT 없음):   supply = qty×price,        vat = 0
def _po_vat_calc(qty: float, price: float, vat_mode: str = "vat_excluded") -> tuple[float, float]:
    base = float(qty or 0) * float(price or 0)
    if vat_mode == "vat_included":
        supply = round(base / 1.1, 2)
        vat = round(base - supply, 2)
    elif vat_mode == "vat_none":
        supply = round(base, 2)
        vat = 0.0
    else:  # vat_excluded (default)
        supply = round(base, 2)
        vat = round(base * 0.1, 2)
    return supply, vat


def _po_backfill_vat(c, po_id: int, vat_mode: str, tax_class: str) -> None:
    """z58 컬럼이 존재하면 헤더·라인에 VAT 백필. 미적용 환경은 skip."""
    try:
        po_cols = {r[1] for r in c.execute("PRAGMA table_info(purchase_orders)").fetchall()}
        line_cols = {r[1] for r in c.execute("PRAGMA table_info(po_items)").fetchall()}
        # 헤더: vat_mode / tax_classification 갱신
        hdr_sets, hdr_vals = [], []
        if "vat_mode" in po_cols:
            hdr_sets.append("vat_mode=?"); hdr_vals.append(vat_mode or "vat_excluded")
        if "tax_classification" in po_cols:
            hdr_sets.append("tax_classification=?"); hdr_vals.append(tax_class or "taxable")
        if hdr_sets:
            c.execute(f"UPDATE purchase_orders SET {','.join(hdr_sets)} WHERE id=?",
                      hdr_vals + [po_id])
        # 라인별 supply_amount / vat_amount 계산
        if "supply_amount" in line_cols and "vat_amount" in line_cols:
            lines = c.execute(
                "SELECT id, quantity, unit_price FROM po_items WHERE po_id=?", (po_id,)
            ).fetchall()
            sup_total, vat_total = 0.0, 0.0
            for ln in lines:
                sup, vat = _po_vat_calc(ln["quantity"], ln["unit_price"], vat_mode)
                sup_total += sup; vat_total += vat
                c.execute(
                    "UPDATE po_items SET supply_amount=?, vat_amount=? WHERE id=?",
                    (sup, vat, ln["id"]),
                )
            # 헤더 합계
            if "supply_total" in po_cols and "vat_total" in po_cols:
                c.execute(
                    "UPDATE purchase_orders SET supply_total=?, vat_total=? WHERE id=?",
                    (round(sup_total, 2), round(vat_total, 2), po_id),
                )
    except Exception:
        # z58 미적용 환경 또는 컬럼 부재 시 silently skip (정상 작동)
        pass


def po_list(q: str = "", status: str = "", supplier_id: int = 0,
            project_id: int = 0):
    sql = """
      SELECT po.*, s.name AS supplier_name, p.name AS project_name,
             p.mgmt_code AS mgmt_code, u.name AS creator_name
      FROM purchase_orders po
      LEFT JOIN suppliers s ON po.supplier_id = s.id
      LEFT JOIN projects p ON po.project_id = p.id
      LEFT JOIN users u ON po.created_by = u.id
      WHERE 1=1
    """
    params = []
    if q:
        sql += """ AND (po.po_number LIKE ? OR s.name LIKE ?
                   OR p.mgmt_code LIKE ? OR p.name LIKE ?)"""
        like = f"%{q}%"
        params += [like] * 4
    if status:
        sql += " AND po.status = ?"
        params.append(status)
    if supplier_id:
        sql += " AND po.supplier_id = ?"
        params.append(supplier_id)
    if project_id:
        sql += " AND po.project_id = ?"
        params.append(project_id)
    sql += " ORDER BY po.id DESC"
    with db_session() as c:
        return c.execute(sql, params).fetchall()


def po_get(po_id: int):
    """발주 헤더 + 라인 + 관련 마스터"""
    with db_session() as c:
        header = c.execute("""
            SELECT po.*, s.name AS supplier_name, s.contact AS supplier_contact,
                   s.email AS supplier_email, s.country AS supplier_country,
                   p.name AS project_name, p.mgmt_code, p.customer_name AS customer,
                   u.name AS creator_name
            FROM purchase_orders po
            LEFT JOIN suppliers s ON po.supplier_id = s.id
            LEFT JOIN projects p ON po.project_id = p.id
            LEFT JOIN users u ON po.created_by = u.id
            WHERE po.id = ?
        """, (po_id,)).fetchone()
        if not header:
            return None, []
        items = c.execute("""
            SELECT * FROM po_items WHERE po_id = ?
            ORDER BY line_no, id
        """, (po_id,)).fetchall()
    return header, items


def po_create(data: dict, items: list[dict], created_by: int = 0) -> tuple[int, str]:
    """발주 헤더 + 라인 일괄 생성. 발주번호는 자동 채번.
    v5H112: 정합성 검증 — 라인 0건/qty<=0/price<0 거부."""
    # v5H112 정합성 검증 (v5H91 패턴)
    valid_items = [it for it in (items or [])
                   if (float(it.get("quantity") or 0) > 0)
                   or (it.get("part_id") and str(it.get("part_id")).strip())]
    if not valid_items:
        raise ValueError("발주 라인이 1개 이상 있어야 합니다. 자재와 수량을 입력해주세요.")
    for idx, it in enumerate(valid_items, start=1):
        q = float(it.get("quantity") or 0)
        p = float(it.get("unit_price") or 0)
        if q <= 0:
            raise ValueError(f"{idx}번째 라인의 수량이 0 이하입니다. 양수로 입력해주세요.")
        if p < 0:
            raise ValueError(f"{idx}번째 라인의 단가가 음수입니다. 0 이상으로 입력해주세요.")
    items = valid_items
    po_number = generate_po_number()
    order_date = (data.get("order_date") or "").strip() or _date.today().isoformat()
    now = _logi_now()
    total = 0.0

    with db_session() as c:
        cur = c.execute("""
            INSERT INTO purchase_orders
            (po_number, project_id, supplier_id, order_date, expected_date,
             currency, exchange_rate, total_amount, status, shipping_terms,
             payment_terms, po_type, created_by, note, created_at, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            po_number,
            int(data.get("project_id") or 0) or None,
            int(data.get("supplier_id") or 0) or None,
            order_date,
            (data.get("expected_date") or "").strip() or None,
            (data.get("currency") or "KRW").strip() or "KRW",
            float(data.get("exchange_rate") or 1),
            0,  # total 임시
            (data.get("status") or "작성중").strip() or "작성중",
            (data.get("shipping_terms") or "").strip(),
            (data.get("payment_terms") or "").strip(),
            (data.get("po_type") or "일반").strip() or "일반",
            int(created_by or 0) or None,
            (data.get("note") or "").strip(),
            now, now,
        ))
        po_id = cur.lastrowid

        # 라인 삽입
        for idx, it in enumerate(items, start=1):
            qty = float(it.get("quantity") or 0)
            price = float(it.get("unit_price") or 0)
            amt = round(qty * price, 2)
            total += amt

            # 부품 스냅샷 채우기
            part_id = int(it.get("part_id") or 0) or None
            part_no = it.get("part_no_snapshot") or ""
            part_name = it.get("part_name_snapshot") or ""
            spec = it.get("spec_snapshot") or ""
            unit = it.get("unit") or "EA"
            if part_id:
                p = c.execute(
                    "SELECT part_no, part_name, spec, unit FROM parts WHERE id=?",
                    (part_id,)
                ).fetchone()
                if p:
                    part_no = part_no or p["part_no"] or ""
                    part_name = part_name or p["part_name"] or ""
                    spec = spec or p["spec"] or ""
                    unit = unit or p["unit"] or "EA"

            c.execute("""
                INSERT INTO po_items
                (po_id, line_no, part_id, part_no_snapshot, part_name_snapshot,
                 spec_snapshot, unit, quantity, unit_price, amount,
                 delivery_date, note)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """, (po_id, idx, part_id, part_no, part_name, spec, unit,
                  qty, price, amt,
                  (it.get("delivery_date") or "").strip() or None,
                  (it.get("note") or "").strip()))

        # total 업데이트
        c.execute("UPDATE purchase_orders SET total_amount=? WHERE id=?",
                  (round(total, 2), po_id))

        # v5H226z58: VAT 백필 (z58 컬럼 존재 시)
        _po_backfill_vat(
            c, po_id,
            (data.get("vat_mode") or "vat_excluded").strip() or "vat_excluded",
            (data.get("tax_classification") or "taxable").strip() or "taxable",
        )

    return po_id, po_number


def po_update(po_id: int, data: dict, items: list[dict]) -> None:
    """v5H112: id 기준 UPSERT — received_qty 보존.
    - id 가 form 에 있고 기존 라인 매칭되면 UPDATE
    - id 없으면 INSERT
    - form 에서 사라진 기존 id 는 DELETE (단 received_qty>0 이면 ValueError)
    - 정합성 검증: qty<=0 / price<0 / 라인 0건 거부
    """
    # v5H112 정합성 검증
    valid_items = [it for it in (items or [])
                   if (float(it.get("quantity") or 0) > 0)
                   or (it.get("part_id") and str(it.get("part_id")).strip())
                   or (it.get("id") and str(it.get("id")).strip())]
    if not valid_items:
        raise ValueError("발주 라인이 1개 이상 있어야 합니다.")
    for idx, it in enumerate(valid_items, start=1):
        q = float(it.get("quantity") or 0)
        p = float(it.get("unit_price") or 0)
        if q <= 0:
            raise ValueError(f"{idx}번째 라인 수량이 0 이하입니다.")
        if p < 0:
            raise ValueError(f"{idx}번째 라인 단가가 음수입니다.")
    items = valid_items
    now = _logi_now()
    total = 0.0
    with db_session() as c:
        # 헤더 업데이트
        c.execute("""
            UPDATE purchase_orders SET
              project_id=?, supplier_id=?, order_date=?, expected_date=?,
              currency=?, exchange_rate=?, status=?, shipping_terms=?,
              payment_terms=?, po_type=?, note=?, updated_at=?
            WHERE id=?
        """, (
            int(data.get("project_id") or 0) or None,
            int(data.get("supplier_id") or 0) or None,
            (data.get("order_date") or "").strip(),
            (data.get("expected_date") or "").strip() or None,
            (data.get("currency") or "KRW").strip() or "KRW",
            float(data.get("exchange_rate") or 1),
            (data.get("status") or "작성중").strip() or "작성중",
            (data.get("shipping_terms") or "").strip(),
            (data.get("payment_terms") or "").strip(),
            (data.get("po_type") or "일반").strip() or "일반",
            (data.get("note") or "").strip(),
            now, po_id,
        ))
        # 기존 라인 조회 (id → row)
        existing = {
            r["id"]: r for r in c.execute(
                "SELECT id, received_qty FROM po_items WHERE po_id=?", (po_id,)
            ).fetchall()
        }
        kept_ids = set()
        for idx, it in enumerate(items, start=1):
            qty = float(it.get("quantity") or 0)
            price = float(it.get("unit_price") or 0)
            amt = round(qty * price, 2)
            total += amt
            part_id = int(it.get("part_id") or 0) or None
            line_id_raw = str(it.get("id") or "").strip()
            line_id = int(line_id_raw) if line_id_raw.isdigit() else 0
            if line_id and line_id in existing:
                # UPDATE — received_qty 보존
                kept_ids.add(line_id)
                c.execute("""
                    UPDATE po_items SET
                      line_no=?, part_id=?, part_no_snapshot=?, part_name_snapshot=?,
                      spec_snapshot=?, unit=?, quantity=?, unit_price=?, amount=?,
                      delivery_date=?, note=?
                    WHERE id=?
                """, (idx, part_id,
                      it.get("part_no_snapshot") or "",
                      it.get("part_name_snapshot") or "",
                      it.get("spec_snapshot") or "",
                      it.get("unit") or "EA",
                      qty, price, amt,
                      (it.get("delivery_date") or "").strip() or None,
                      (it.get("note") or "").strip(),
                      line_id))
            else:
                # INSERT
                c.execute("""
                    INSERT INTO po_items
                    (po_id, line_no, part_id, part_no_snapshot, part_name_snapshot,
                     spec_snapshot, unit, quantity, unit_price, amount,
                     delivery_date, note)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                """, (po_id, idx, part_id,
                      it.get("part_no_snapshot") or "",
                      it.get("part_name_snapshot") or "",
                      it.get("spec_snapshot") or "",
                      it.get("unit") or "EA",
                      qty, price, amt,
                      (it.get("delivery_date") or "").strip() or None,
                      (it.get("note") or "").strip()))
        # 사라진 라인 — received_qty>0 이면 거부
        to_delete = [lid for lid in existing.keys() if lid not in kept_ids]
        for lid in to_delete:
            recv = existing[lid]["received_qty"] or 0
            if recv and recv > 0:
                raise ValueError(
                    f"입고 이력이 있는 라인(id={lid}, 입고수량={recv})은 삭제할 수 없습니다. "
                    "입고 취소 후 다시 시도해주세요."
                )
            c.execute("DELETE FROM po_items WHERE id=?", (lid,))
        c.execute("UPDATE purchase_orders SET total_amount=? WHERE id=?",
                  (round(total, 2), po_id))

        # v5H226z58: VAT 백필 (z58 컬럼 존재 시)
        _po_backfill_vat(
            c, po_id,
            (data.get("vat_mode") or "vat_excluded").strip() or "vat_excluded",
            (data.get("tax_classification") or "taxable").strip() or "taxable",
        )


def po_delete(po_id: int) -> None:
    """v5H119: 공통 헬퍼 사용. 폴백: v5H112 인라인.
    헬퍼 호출 시 본행 'purchase_orders' / 자식 FK 'po_id'."""
    with db_session() as c:
        try:
            res = _safe_delete_with_cascade(
                c, "purchase_orders", po_id, fk_column="po_id",
                explicit_children=[
                    ("DELETE FROM po_items WHERE po_id=?", (po_id,)),
                    ("UPDATE stock_movements SET po_id=NULL, po_item_id=NULL WHERE po_id=?", (po_id,)),
                ],
                keep_tables=("po_items",),
            )
            if res.get("ok"):
                return
        except Exception:
            pass
        try: c.execute("DELETE FROM po_items WHERE po_id=?", (po_id,))
        except Exception: pass
        for sql in [
            "UPDATE stock_movements SET po_id=NULL, po_item_id=NULL WHERE po_id=?",
        ]:
            try: c.execute(sql, (po_id,))
            except Exception: pass
        try:
            all_tables = [r[0] for r in c.execute(
                "SELECT name FROM sqlite_master WHERE type='table' "
                "AND name NOT LIKE 'sqlite_%' AND name <> 'purchase_orders'"
            ).fetchall()]
            for tname in all_tables:
                try:
                    cols = [r[1] for r in c.execute(f"PRAGMA table_info({tname})").fetchall()]
                except Exception:
                    continue
                if "po_id" not in cols:
                    continue
                try:
                    c.execute(f"UPDATE {tname} SET po_id=NULL WHERE po_id=?", (po_id,))
                except Exception:
                    try: c.execute(f"DELETE FROM {tname} WHERE po_id=?", (po_id,))
                    except Exception: pass
        except Exception:
            pass
        c.execute("DELETE FROM purchase_orders WHERE id = ?", (po_id,))


def po_count() -> dict:
    with db_session() as c:
        total = c.execute("SELECT COUNT(*) FROM purchase_orders").fetchone()[0]
        by_status = c.execute(
            "SELECT status, COUNT(*) AS n FROM purchase_orders GROUP BY status"
        ).fetchall()
        sum_amt = c.execute(
            "SELECT currency, SUM(total_amount) AS s FROM purchase_orders "
            "WHERE status != '취소' GROUP BY currency"
        ).fetchall()
    return {
        "total": total,
        "by_status": {r["status"] or "(미지정)": r["n"] for r in by_status},
        "amount_by_currency": {r["currency"] or "KRW": r["s"] or 0 for r in sum_amt},
    }


# =====================================================
# STOCK MOVEMENTS — 입출고 원장 (수불부) 2026-04-20
# 설계: 모든 재고 변동은 stock_movements에 기록, parts.stock_qty는 결과 캐시
# =====================================================
MOVEMENT_KINDS = ["IN", "OUT", "ADJUST", "TRANSFER"]
MOVEMENT_KIND_LABEL = {
    "IN": "입고", "OUT": "출고", "ADJUST": "실사조정", "TRANSFER": "이동",
}


def gen_movement_no(today=None) -> str:
    """수불 번호 자동 채번: SM-YYMMDD-NNN"""
    today = today or _date.today()
    yymmdd = today.strftime("%y%m%d")
    prefix = f"SM-{yymmdd}-"
    try:
        with db_session() as c:
            rows = c.execute(
                "SELECT movement_no FROM stock_movements WHERE movement_no LIKE ?",
                (f"{prefix}%",),
            ).fetchall()
        max_seq = 0
        for r in rows:
            try:
                seq = int(r["movement_no"].rsplit("-", 1)[-1])
                max_seq = max(max_seq, seq)
            except (ValueError, IndexError):
                pass
        return f"{prefix}{max_seq + 1:03d}"
    except sqlite3.OperationalError:
        return f"{prefix}001"


def _recalc_part_stock(part_id: int, c) -> dict:
    """stock_movements 합산 → parts.stock_qty 동기화.
    반환: {prev_qty, new_qty, safety_stock, crossed_low (정상→미달로 전환된 경우 True)}
    """
    # 이전 상태 조회
    prev = c.execute(
        "SELECT stock_qty, safety_stock, part_no, part_name FROM parts WHERE id=?",
        (part_id,),
    ).fetchone()
    prev_qty = (prev["stock_qty"] or 0) if prev else 0
    safety = (prev["safety_stock"] or 0) if prev else 0

    total = c.execute(
        "SELECT COALESCE(SUM(quantity), 0) FROM stock_movements WHERE part_id=?",
        (part_id,),
    ).fetchone()[0] or 0
    c.execute(
        "UPDATE parts SET stock_qty=?, updated_at=? WHERE id=?",
        (total, _logi_now(), part_id),
    )

    # 정상 → 미달로 전환 판별 (안전재고 > 0 설정된 품목만)
    was_ok = safety > 0 and prev_qty >= safety
    is_low = safety > 0 and total < safety
    crossed_low = was_ok and is_low

    return {
        "prev_qty": prev_qty,
        "new_qty": total,
        "safety_stock": safety,
        "crossed_low": crossed_low,
        "part_no": prev["part_no"] if prev else "",
        "part_name": prev["part_name"] if prev else "",
    }


def _auto_ticket_low_stock(part_info: dict, user_id: int):
    """안전재고 미달 전환 시 구매팀에 자동 티켓 생성.
    중복 방지: 동일 부품에 대해 24시간 내 미완료 자재요청 티켓이 있으면 스킵.
    """
    try:
        part_no = part_info.get("part_no", "")
        part_name = part_info.get("part_name", "")
        new_qty = part_info.get("new_qty", 0)
        safety = part_info.get("safety_stock", 0)

        title = f"[자동] 안전재고 미달 — {part_name} ({part_no})"
        with db_session() as c:
            # 중복 방지: 24시간 내 동일 제목 미완료 티켓
            existing = c.execute(
                """SELECT id FROM tickets
                   WHERE title = ? AND status NOT IN ('완료','반려')
                   AND created_at >= datetime('now','-1 day','localtime')""",
                (title,),
            ).fetchone()
            if existing:
                return None
        # 구매팀 id
        recipient_team_id = _team_id_by_name("구매팀") or None
        body = (
            f"안전재고 미달 — 긴급 발주 검토 요청\n\n"
            f"• 부품: [{part_no}] {part_name}\n"
            f"• 현재고: {new_qty}\n"
            f"• 안전재고: {safety}\n"
            f"• 부족 수량: {max(0, safety - new_qty)}\n\n"
            f"※ 이 티켓은 출고/실사 조정으로 재고가 안전재고 아래로 떨어질 때 자동 생성됩니다."
        )
        tid, tno = ticket_create({
            "category": "자재요청",
            "title": title,
            "description": body,
            "urgency": "긴급",
            "source": "auto-stock",
            "recipient_team_id": recipient_team_id,
        }, requester_id=user_id)
        return tid
    except Exception as e:
        print(f"[AUTO-TICKET stock] {e}")
        return None


# =====================================================
# 사이클 51 S2-4차 (2026-04-27) — 안전재고 알림 자동 + 발주 추천
# 헬퍼 +2: check_stock_alerts() / recommend_reorders()
# v2 본체 무수정 · G1~G5 보존 · 외부 라이브러리 0건
# =====================================================
def check_stock_alerts() -> dict:
    """전체 부품 stock_balances vs reorder_point 비교 → 부족 부품 STOCK 알림 발송.
    수신자: 구매팀 멤버 + admin/ceo (멤버십 조회 실패 시 admin/ceo만).
    반환: {"checked": N, "alerts_sent": M, "low_parts": [...]}
    """
    out = {"checked": 0, "alerts_sent": 0, "low_parts": []}
    try:
        with db_session() as c:
            rows = c.execute(
                """SELECT p.id, p.part_no, p.part_name, p.unit,
                          COALESCE(p.reorder_point,0)  AS rop,
                          COALESCE(p.safety_stock,0)   AS safety,
                          COALESCE(p.reorder_qty,0)    AS roq,
                          COALESCE(sb.on_hand,0)       AS on_hand
                   FROM parts p
                   LEFT JOIN stock_balances sb ON sb.part_id = p.id
                   WHERE COALESCE(p.reorder_point,0) > 0
                     AND COALESCE(sb.on_hand,0) < COALESCE(p.reorder_point,0)
                   ORDER BY (COALESCE(p.reorder_point,0) - COALESCE(sb.on_hand,0)) DESC"""
            ).fetchall()
            out["checked"] = len(rows)
            # 수신자: 구매팀 멤버 + admin/ceo (중복 제거)
            recipients = set()
            try:
                team_id = _team_id_by_name("구매팀")
                if team_id:
                    for r in c.execute(
                        "SELECT user_id FROM team_members WHERE team_id=?",
                        (team_id,),
                    ).fetchall():
                        if r["user_id"]:
                            recipients.add(int(r["user_id"]))
            except Exception:
                pass
            try:
                for r in c.execute(
                    "SELECT id FROM users WHERE role IN ('admin','ceo')"
                ).fetchall():
                    if r["id"]:
                        recipients.add(int(r["id"]))
            except Exception:
                pass
        # 알림 발송 (db_session 밖 — notify_user 자체가 별도 세션)
        for row in rows:
            part_no = row["part_no"]
            part_name = row["part_name"] or ""
            on_hand = row["on_hand"] or 0
            rop = row["rop"] or 0
            roq = row["roq"] or 0
            short = max(0, rop - on_hand)
            title = f"[재고알림] 재발주점 미달 — {part_name} ({part_no})"
            body = (
                f"• 부품: [{part_no}] {part_name}\n"
                f"• 현재고: {on_hand}\n"
                f"• 재발주점: {rop}\n"
                f"• 부족: {short}\n"
                f"• 권장 발주량: {roq}"
            )
            for uid in recipients:
                if notify_user(uid, "STOCK", title, body=body,
                               link="/stock/reorder-recommendations"):
                    out["alerts_sent"] += 1
            out["low_parts"].append({
                "part_id": row["id"], "part_no": part_no,
                "part_name": part_name, "on_hand": on_hand,
                "reorder_point": rop, "reorder_qty": roq, "short": short,
            })
    except Exception as e:
        print(f"[check_stock_alerts] {e}")
    return out


def recommend_reorders(limit: int = 200) -> list[dict]:
    """발주 추천 — 부족 부품 + 권장 발주량 + 우선순위.
    우선순위: 부족율(=(rop-on_hand)/rop) 내림차순.
    HIGH ≥0.5 / MID 0.2~0.5 / LOW <0.2
    """
    items: list[dict] = []
    try:
        with db_session() as c:
            rows = c.execute(
                """SELECT p.id              AS part_id,
                          p.part_no         AS part_no,
                          p.part_name       AS part_name,
                          p.unit            AS unit,
                          p.maker           AS maker,
                          COALESCE(p.safety_stock,0)  AS safety,
                          COALESCE(p.reorder_point,0) AS rop,
                          COALESCE(p.reorder_qty,0)   AS roq,
                          COALESCE(sb.on_hand,0)      AS on_hand
                   FROM parts p
                   LEFT JOIN stock_balances sb ON sb.part_id = p.id
                   WHERE COALESCE(p.reorder_point,0) > 0
                     AND COALESCE(sb.on_hand,0) < COALESCE(p.reorder_point,0)
                   ORDER BY (COALESCE(p.reorder_point,0) - COALESCE(sb.on_hand,0)) DESC
                   LIMIT ?""",
                (int(limit),),
            ).fetchall()
        for r in rows:
            rop = r["rop"] or 0
            on_hand = r["on_hand"] or 0
            short = max(0, rop - on_hand)
            ratio = (short / rop) if rop > 0 else 0
            if ratio >= 0.5:
                pri = "HIGH"
            elif ratio >= 0.2:
                pri = "MID"
            else:
                pri = "LOW"
            items.append({
                "part_id":       r["part_id"],
                "part_no":       r["part_no"],
                "part_name":     r["part_name"],
                "unit":          r["unit"] or "EA",
                "maker":         r["maker"] or "",
                "safety_stock":  r["safety"] or 0,
                "reorder_point": rop,
                "reorder_qty":   r["roq"] or 0,
                "on_hand":       on_hand,
                "short":         short,
                "ratio":         round(ratio, 3),
                "priority":      pri,
            })
    except Exception as e:
        print(f"[recommend_reorders] {e}")
    return items


def parts_safety_settings_list(q: str = "") -> list[dict]:
    """안전재고 설정 페이지용 — parts + on_hand 조회 (검색 q 지원)."""
    out: list[dict] = []
    try:
        with db_session() as c:
            sql = (
                "SELECT p.id, p.part_no, p.part_name, p.unit, p.maker, "
                "COALESCE(p.safety_stock,0)  AS safety_stock, "
                "COALESCE(p.reorder_point,0) AS reorder_point, "
                "COALESCE(p.reorder_qty,0)   AS reorder_qty, "
                "COALESCE(sb.on_hand,0)      AS on_hand "
                "FROM parts p LEFT JOIN stock_balances sb ON sb.part_id = p.id "
                "WHERE p.is_active = 1"
            )
            params: list = []
            if q:
                sql += " AND (p.part_no LIKE ? OR p.part_name LIKE ?)"
                like = f"%{q}%"
                params.extend([like, like])
            sql += " ORDER BY p.part_no LIMIT 300"
            for r in c.execute(sql, params).fetchall():
                out.append(dict(r))
    except Exception as e:
        print(f"[parts_safety_settings_list] {e}")
    return out


def parts_safety_update(part_id: int,
                        safety_stock: float,
                        reorder_point: float,
                        reorder_qty: float) -> bool:
    """parts 안전재고 3종 컬럼 일괄 갱신 (음수 → 0 가드)."""
    try:
        ss = max(0.0, float(safety_stock or 0))
        rp = max(0.0, float(reorder_point or 0))
        rq = max(0.0, float(reorder_qty or 0))
        with db_session() as c:
            c.execute(
                "UPDATE parts SET safety_stock=?, reorder_point=?, "
                "reorder_qty=?, updated_at=? WHERE id=?",
                (ss, rp, rq, _logi_now(), int(part_id)),
            )
        return True
    except Exception as e:
        print(f"[parts_safety_update] {e}")
        return False


def _consume_fifo(c, part_id: int, qty: float) -> tuple[float, list[dict]]:
    """FIFO로 IN 레이어 소비.
    반환: (가중평균 단가, 소비한 레이어 리스트)
    레이어가 부족하면 가능한 만큼만 소비 (남은 건 0원으로 처리).
    """
    if qty <= 0:
        return 0.0, []
    layers = c.execute(
        """SELECT id, remaining_qty, unit_price, lot_no, occurred_at
           FROM stock_movements
           WHERE part_id=? AND kind='IN' AND remaining_qty > 0
           ORDER BY occurred_at ASC, id ASC""",
        (part_id,),
    ).fetchall()

    remaining_to_consume = qty
    total_cost = 0.0
    consumed_layers = []
    for r in layers:
        if remaining_to_consume <= 0:
            break
        avail = float(r["remaining_qty"] or 0)
        take = min(avail, remaining_to_consume)
        if take <= 0:
            continue
        price = float(r["unit_price"] or 0)
        total_cost += take * price
        remaining_to_consume -= take
        c.execute(
            "UPDATE stock_movements SET remaining_qty = remaining_qty - ? WHERE id=?",
            (take, r["id"]),
        )
        consumed_layers.append({
            "layer_id": r["id"],
            "lot_no": r["lot_no"],
            "taken": take,
            "unit_price": price,
        })
    # 가중평균 (실 소비량 기준)
    consumed_qty = qty - remaining_to_consume
    avg_price = (total_cost / consumed_qty) if consumed_qty > 0 else 0.0
    return avg_price, consumed_layers


def stock_movement_create(data: dict, user_id: int) -> tuple[int, str]:
    """재고 이동 1건 생성 + parts.stock_qty 자동 갱신 + FIFO 원가 계산

    data 필수:
      - part_id, kind (IN/OUT/ADJUST/TRANSFER), quantity (부호 있는 값)
    선택:
      - unit_price, po_id, po_item_id, project_id, customer_id, reason, location,
        occurred_at, note, unit, lot_no, expiry_date

    FIFO 동작:
      - IN: remaining_qty = quantity 로 레이어 생성
      - OUT: _consume_fifo() 로 IN 레이어 소비, 출고 단가 = 가중평균
      - ADJUST(-): 레이어 소비 (동일)
      - ADJUST(+): 새 IN 레이어로 등록 (조정 사유를 단가 0으로 보존)
    """
    kind = (data.get("kind") or "IN").upper()
    if kind not in MOVEMENT_KINDS:
        raise ValueError(f"Invalid movement kind: {kind}")
    part_id = int(data["part_id"])
    qty = float(data.get("quantity") or 0)
    if kind == "OUT" and qty > 0:
        qty = -qty
    elif kind == "IN" and qty < 0:
        qty = -qty
    unit_price = float(data.get("unit_price") or 0)
    movement_no = gen_movement_no()
    occurred_at = data.get("occurred_at") or _logi_now()

    with db_session() as c:
        # FIFO: 소비가 필요한 경우 먼저 계산
        fifo_avg_price = None
        fifo_layers_info = None
        if qty < 0:  # OUT or ADJUST 음수
            fifo_avg_price, consumed = _consume_fifo(c, part_id, abs(qty))
            if consumed:
                # unit_price가 명시적으로 주어지지 않은 경우 FIFO 단가 사용
                if unit_price == 0 and fifo_avg_price > 0:
                    unit_price = fifo_avg_price
                # 소비 레이어 정보를 note에 간단 기록 (감사 추적)
                lot_summary = ", ".join(
                    f"lot={l['lot_no'] or '#' + str(l['layer_id'])}×{l['taken']:g}@{l['unit_price']:,.0f}"
                    for l in consumed[:5]
                )
                fifo_layers_info = f"FIFO: {lot_summary}" + ("..." if len(consumed) > 5 else "")

        amount = abs(qty) * unit_price

        # IN 이거나 ADJUST 양수면 remaining_qty = |qty|, 아니면 0
        remaining = abs(qty) if qty > 0 else 0

        # note에 FIFO 소비 정보 자동 추가 (기존 note 유지)
        final_note = (data.get("note") or "").strip() or None
        if fifo_layers_info:
            final_note = (final_note + " | " if final_note else "") + fifo_layers_info

        c.execute(
            """INSERT INTO stock_movements
               (movement_no, part_id, kind, quantity, unit, unit_price, amount,
                remaining_qty, lot_no, expiry_date,
                po_id, po_item_id, project_id, customer_id,
                reason, location, occurred_at, note, created_by)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                movement_no, part_id, kind, qty,
                data.get("unit") or "EA",
                unit_price, amount,
                remaining,
                (data.get("lot_no") or "").strip() or None,
                (data.get("expiry_date") or "").strip() or None,
                data.get("po_id"), data.get("po_item_id"),
                data.get("project_id"), data.get("customer_id"),
                (data.get("reason") or "").strip() or None,
                (data.get("location") or "").strip() or None,
                occurred_at,
                final_note,
                user_id,
            ),
        )
        mid = c.execute("SELECT last_insert_rowid()").fetchone()[0]
        stock_info = _recalc_part_stock(part_id, c)

    # 정상 → 미달 전환 시 구매팀에 자동 티켓
    if stock_info.get("crossed_low"):
        _auto_ticket_low_stock(stock_info, user_id)
    return mid, movement_no


def po_receive(po_id: int, receive_lines: list[dict], user_id: int,
               occurred_at: str = "") -> dict:
    """발주서 입고 처리 — 라인별 입고 수량 기록

    receive_lines: [{po_item_id, receive_qty, note?}]
    - stock_movements에 IN 생성
    - po_items.received_qty 누적
    - purchase_orders.status 자동 갱신 (부분입고/입고완료)
    """
    created = []
    with db_session() as c:
        po = c.execute(
            "SELECT * FROM purchase_orders WHERE id=?", (po_id,)
        ).fetchone()
        if not po:
            return {"ok": False, "error": "발주서를 찾을 수 없습니다."}

        for line in receive_lines:
            item_id = int(line.get("po_item_id"))
            qty = float(line.get("receive_qty") or 0)
            if qty <= 0:
                continue
            item = c.execute(
                "SELECT * FROM po_items WHERE id=? AND po_id=?",
                (item_id, po_id),
            ).fetchone()
            if not item or not item["part_id"]:
                continue
            # v5H112: 입고 OVER 차단 — 누적 입고 > 발주 수량 거부
            new_recv = (item["received_qty"] or 0) + qty
            ord_qty = item["quantity"] or 0
            if ord_qty > 0 and new_recv > ord_qty + 0.0001:
                return {
                    "ok": False,
                    "error": (
                        f"입고 수량 초과 — 발주 {ord_qty} / 기존 입고 {item['received_qty'] or 0} / "
                        f"이번 입고 {qty} = 누적 {new_recv}. 발주 수량 내로 입력해주세요."
                    ),
                }
            c.execute(
                "UPDATE po_items SET received_qty=? WHERE id=?",
                (new_recv, item_id),
            )
            # stock_movement 생성 + stock_qty 재계산 (lot_no는 라인에서)
            mid, mno = stock_movement_create({
                "part_id": item["part_id"],
                "kind": "IN",
                "quantity": qty,
                "unit": item["unit"] or "EA",
                "unit_price": item["unit_price"] or 0,
                "po_id": po_id,
                "po_item_id": item_id,
                "project_id": po["project_id"],
                "reason": f"발주 {po['po_number']} 입고",
                "occurred_at": occurred_at or _logi_now(),
                "note": (line.get("note") or "").strip() or None,
                "lot_no": (line.get("lot_no") or "").strip() or None,
                "expiry_date": (line.get("expiry_date") or "").strip() or None,
            }, user_id)
            created.append({"movement_id": mid, "movement_no": mno,
                            "po_item_id": item_id, "qty": qty})

        # PO 상태 자동 갱신
        items_all = c.execute(
            "SELECT quantity, received_qty FROM po_items WHERE po_id=?",
            (po_id,),
        ).fetchall()
        total_ord = sum(r["quantity"] or 0 for r in items_all)
        total_rcv = sum(r["received_qty"] or 0 for r in items_all)
        if total_rcv <= 0:
            new_status = po["status"]
        elif total_rcv >= total_ord:
            new_status = "입고완료"
        else:
            new_status = "부분입고"
        c.execute(
            "UPDATE purchase_orders SET status=?, updated_at=? WHERE id=?",
            (new_status, _logi_now(), po_id),
        )
    return {"ok": True, "created": created, "count": len(created)}


def stock_issue(data: dict, user_id: int) -> tuple[int, str] | None:
    """출고 처리 — 프로젝트/고객사 지정, parts.stock_qty 감소

    data 필수:
      - part_id, quantity (양수로 입력 — 내부에서 음수화)
    """
    qty = float(data.get("quantity") or 0)
    if qty <= 0:
        raise ValueError("출고 수량은 양수여야 합니다.")
    # v5H112: 재고 음수 차단 — 현재고 < 출고 수량 시 거부
    try:
        with db_session() as _c:
            cur = _c.execute(
                "SELECT stock_qty, part_name FROM parts WHERE id=?",
                (int(data["part_id"]),),
            ).fetchone()
            if cur:
                stock = cur["stock_qty"] or 0
                if stock < qty:
                    raise ValueError(
                        f"재고 부족 — '{cur['part_name'] or ''}' 현재고 {stock} < 출고 요청 {qty}. "
                        "재고를 확인하거나 입고 후 다시 시도해주세요."
                    )
    except ValueError:
        raise
    except Exception:
        pass  # 사전조회 실패는 무시 (기존 동작 유지)
    return stock_movement_create({
        "part_id": data["part_id"],
        "kind": "OUT",
        "quantity": qty,  # 내부에서 음수화됨
        "unit_price": data.get("unit_price") or 0,
        "unit": data.get("unit") or "EA",
        "project_id": data.get("project_id"),
        "customer_id": data.get("customer_id"),
        "reason": (data.get("reason") or "현장 출고").strip(),
        "location": data.get("location"),
        "occurred_at": data.get("occurred_at"),
        "note": data.get("note"),
    }, user_id)


def stock_adjust(data: dict, user_id: int) -> tuple[int, str]:
    """실사 조정 — 실사 결과와 시스템 재고 차이를 조정

    data 필수:
      - part_id, quantity (+/- 조정값)
      - reason (실사 사유, 필수)
    v5H113 LOW#17: 사유 빈문자열 차단, 음수 한도(±1,000,000) 검증.
    """
    reason = (data.get("reason") or "").strip()
    if not reason:
        raise ValueError("실사 조정 사유는 필수입니다. (예: 분실/파손/실사 차이 등)")
    try:
        qty = float(data.get("quantity") or 0)
    except (TypeError, ValueError):
        raise ValueError("조정 수량이 올바른 숫자가 아닙니다.")
    if qty == 0:
        raise ValueError("조정 수량이 0입니다. 변경할 차이를 입력해주세요.")
    if abs(qty) > 1_000_000:
        raise ValueError(f"조정 수량({qty}) 절대값이 1,000,000을 초과합니다. 사이트 한도를 확인해주세요.")
    return stock_movement_create({
        "part_id": data["part_id"],
        "kind": "ADJUST",
        "quantity": qty,
        "reason": reason,
        "note": data.get("note"),
        "occurred_at": data.get("occurred_at"),
    }, user_id)


# =====================================================
# 재고 실사·조정 (Top10 #10 — 2026-04-26 P4 자재팀 분기 1회)
# 기존 stock_movements + stock_balances 활용 / 승인 후 ADJUST 기록 생성
# =====================================================
def gen_audit_no(today=None) -> str:
    """실사 번호 자동 채번: AUD-YYYYMM-####"""
    today = today or _date.today()
    yymm = today.strftime("%Y%m")
    prefix = f"AUD-{yymm}-"
    try:
        with db_session() as c:
            rows = c.execute(
                "SELECT audit_no FROM stock_audits WHERE audit_no LIKE ?",
                (f"{prefix}%",),
            ).fetchall()
        max_seq = 0
        for r in rows:
            try:
                seq = int(r["audit_no"].rsplit("-", 1)[-1])
                max_seq = max(max_seq, seq)
            except (ValueError, IndexError):
                pass
        return f"{prefix}{max_seq + 1:04d}"
    except sqlite3.OperationalError:
        return f"{prefix}0001"


def stock_audit_create(led_by: int, start_date: str = "", note: str = "") -> tuple[int, str]:
    """실사 헤더 신규 생성. 시퀀스 idempotent."""
    sd = start_date or _date.today().isoformat()
    audit_no = gen_audit_no()
    with db_session() as c:
        cur = c.execute(
            """INSERT INTO stock_audits (audit_no, start_date, status, led_by, note)
               VALUES (?, ?, 'OPEN', ?, ?)""",
            (audit_no, sd, led_by, note),
        )
        return cur.lastrowid, audit_no


def stock_audits_list(limit: int = 50) -> list[dict]:
    with db_session() as c:
        rows = c.execute(
            """SELECT a.*, u.name AS led_by_name,
                      (SELECT COUNT(*) FROM stock_audit_items ai WHERE ai.audit_id=a.id) AS item_count,
                      (SELECT COUNT(*) FROM stock_audit_items ai
                        WHERE ai.audit_id=a.id AND ai.status<>'PENDING') AS counted_count
               FROM stock_audits a
               LEFT JOIN users u ON a.led_by = u.id
               ORDER BY a.id DESC LIMIT ?""",
            (limit,),
        ).fetchall()
        return [dict(r) for r in rows]


def stock_audit_get(audit_id: int) -> dict | None:
    with db_session() as c:
        a = c.execute(
            """SELECT a.*, u.name AS led_by_name
               FROM stock_audits a LEFT JOIN users u ON a.led_by=u.id
               WHERE a.id=?""",
            (audit_id,),
        ).fetchone()
        if not a:
            return None
        items = c.execute(
            """SELECT ai.*, p.part_no, p.part_name, p.unit AS part_unit,
                      uc.name AS counted_by_name
               FROM stock_audit_items ai
               LEFT JOIN parts p ON ai.part_id=p.id
               LEFT JOIN users uc ON ai.counted_by=uc.id
               WHERE ai.audit_id=? ORDER BY ai.id""",
            (audit_id,),
        ).fetchall()
        return {**dict(a), "items": [dict(r) for r in items]}


def stock_audit_item_upsert(audit_id: int, part_id: int, counted_qty: float,
                            variance_reason: str, user_id: int) -> int:
    """라인 추가/수정 — system_qty는 parts.stock_qty 스냅샷, variance 자동 계산."""
    with db_session() as c:
        sysq_row = c.execute("SELECT stock_qty FROM parts WHERE id=?", (part_id,)).fetchone()
        sysq = float(sysq_row[0]) if sysq_row else 0.0
        variance = float(counted_qty) - sysq
        if abs(variance) < 1e-9:
            st = "MATCHED"
        elif variance < 0:
            st = "SHORT"
        else:
            st = "OVER"
        existing = c.execute(
            "SELECT id FROM stock_audit_items WHERE audit_id=? AND part_id=?",
            (audit_id, part_id),
        ).fetchone()
        now = _dt.now().strftime("%Y-%m-%d %H:%M:%S")
        if existing:
            c.execute(
                """UPDATE stock_audit_items
                   SET system_qty=?, counted_qty=?, variance=?, variance_reason=?,
                       status=?, counted_by=?, counted_at=?
                   WHERE id=?""",
                (sysq, counted_qty, variance, variance_reason, st, user_id, now, existing[0]),
            )
            return existing[0]
        cur = c.execute(
            """INSERT INTO stock_audit_items
               (audit_id, part_id, system_qty, counted_qty, variance,
                variance_reason, status, counted_by, counted_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (audit_id, part_id, sysq, counted_qty, variance,
             variance_reason, st, user_id, now),
        )
        # 자동으로 PENDING 조정 행 생성 (차이 0 아닐 때)
        if st != "MATCHED":
            c.execute(
                """INSERT INTO stock_adjustments
                   (audit_item_id, adjusted_qty, reason, status, created_by)
                   VALUES (?, ?, ?, 'PENDING', ?)""",
                (cur.lastrowid, variance, variance_reason or "실사 차이", user_id),
            )
        return cur.lastrowid


def stock_adjustments_list(status: str = "PENDING", limit: int = 100) -> list[dict]:
    with db_session() as c:
        rows = c.execute(
            """SELECT adj.*, ai.part_id, ai.system_qty, ai.counted_qty, ai.variance,
                      a.audit_no, p.part_no, p.part_name, p.unit AS part_unit,
                      uc.name AS created_by_name, ua.name AS approved_by_name
               FROM stock_adjustments adj
               JOIN stock_audit_items ai ON adj.audit_item_id=ai.id
               JOIN stock_audits a ON ai.audit_id=a.id
               LEFT JOIN parts p ON ai.part_id=p.id
               LEFT JOIN users uc ON adj.created_by=uc.id
               LEFT JOIN users ua ON adj.approved_by=ua.id
               WHERE (?='' OR adj.status=?)
               ORDER BY adj.id DESC LIMIT ?""",
            (status, status, limit),
        ).fetchall()
        return [dict(r) for r in rows]


def stock_adjustment_approve(adj_id: int, approver_id: int) -> tuple[int, str]:
    """조정 승인 — UPDATE adj.status=APPROVED + INSERT stock_movements{kind=ADJUST}.
    반환: (movement_id, movement_no)
    """
    with db_session() as c:
        adj = c.execute(
            """SELECT adj.*, ai.part_id
               FROM stock_adjustments adj
               JOIN stock_audit_items ai ON adj.audit_item_id=ai.id
               WHERE adj.id=?""",
            (adj_id,),
        ).fetchone()
        if not adj:
            raise ValueError("조정 행 없음")
        if adj["status"] != "PENDING":
            raise ValueError("이미 처리된 조정")
    mid, mno = stock_movement_create({
        "part_id": adj["part_id"],
        "kind": "ADJUST",
        "quantity": float(adj["adjusted_qty"]),
        "reason": (adj["reason"] or "실사 조정").strip(),
        "note": f"audit_adjustment_id={adj_id}",
        "occurred_at": None,
    }, approver_id)
    with db_session() as c:
        c.execute(
            """UPDATE stock_adjustments
               SET status='APPROVED', approved_by=?,
                   approved_at=datetime('now','localtime'), movement_id=?
               WHERE id=?""",
            (approver_id, mid, adj_id),
        )
    return mid, mno


def stock_adjustment_reject(adj_id: int, approver_id: int, note: str = "") -> None:
    with db_session() as c:
        c.execute(
            """UPDATE stock_adjustments
               SET status='REJECTED', approved_by=?,
                   approved_at=datetime('now','localtime'), note=?
               WHERE id=? AND status='PENDING'""",
            (approver_id, note, adj_id),
        )


# ---- 재고실사 2차 (2026-04-26): 첨부 + close 워크플로 + 효과 검증 ----
def audit_attachment_create(adjustment_id: int, file_path: str, file_name: str,
                            user_id: int) -> int:
    """첨부 INSERT — 파일 디스크 저장은 라우트에서 (path traversal sanitize 후)."""
    with db_session() as c:
        cur = c.execute(
            """INSERT INTO audit_attachments
               (adjustment_id, file_path, file_name, uploaded_by)
               VALUES (?, ?, ?, ?)""",
            (adjustment_id, file_path, file_name, user_id),
        )
        return cur.lastrowid


def audit_attachments_list(adjustment_id: int) -> list[dict]:
    with db_session() as c:
        rows = c.execute(
            """SELECT att.*, u.name AS uploaded_by_name
               FROM audit_attachments att
               LEFT JOIN users u ON att.uploaded_by=u.id
               WHERE att.adjustment_id=? ORDER BY att.id DESC""",
            (adjustment_id,),
        ).fetchall()
        return [dict(r) for r in rows]


def audit_attachment_get(att_id: int) -> dict | None:
    with db_session() as c:
        r = c.execute(
            "SELECT * FROM audit_attachments WHERE id=?", (att_id,),
        ).fetchone()
        return dict(r) if r else None


# ---- 자재 첨부 (v5H129 — 2026-05-04) ----
def part_attachment_create(part_id: int, file_path: str, file_name: str,
                           file_size: int, mime_type: str, kind: str,
                           user_id: int) -> int:
    """자재 첨부 INSERT — 디스크 저장은 라우트에서 (sanitize 후)."""
    with db_session() as c:
        cur = c.execute(
            """INSERT INTO part_attachments
               (part_id, file_path, file_name, file_size, mime_type, kind, uploaded_by)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (part_id, file_path, file_name, int(file_size or 0),
             mime_type or "", kind or "photo", user_id),
        )
        return cur.lastrowid


def part_attachments_list(part_id: int) -> list[dict]:
    with db_session() as c:
        rows = c.execute(
            """SELECT att.*, u.name AS uploaded_by_name
               FROM part_attachments att
               LEFT JOIN users u ON att.uploaded_by=u.id
               WHERE att.part_id=? ORDER BY att.id DESC""",
            (part_id,),
        ).fetchall()
        return [dict(r) for r in rows]


def part_attachment_get(att_id: int) -> dict | None:
    with db_session() as c:
        r = c.execute(
            "SELECT * FROM part_attachments WHERE id=?", (att_id,),
        ).fetchone()
        return dict(r) if r else None


def part_attachment_delete(att_id: int) -> dict | None:
    """삭제 — DB 레코드 반환(라우트에서 디스크 파일 삭제용)."""
    rec = part_attachment_get(att_id)
    if not rec:
        return None
    with db_session() as c:
        c.execute("DELETE FROM part_attachments WHERE id=?", (att_id,))
    return rec


def stock_audit_close(audit_id: int) -> tuple[bool, str]:
    """실사 CLOSE — 모든 라인 PENDING 0건 + 모든 조정 APPROVED 검증.
    반환: (ok, msg). status='CLOSED', end_date=오늘.
    """
    with db_session() as c:
        a = c.execute("SELECT id, status FROM stock_audits WHERE id=?",
                      (audit_id,)).fetchone()
        if not a:
            return False, "실사 없음"
        if a["status"] == "CLOSED":
            return False, "이미 CLOSED"
        pending = c.execute(
            "SELECT COUNT(*) FROM stock_audit_items WHERE audit_id=? AND status='PENDING'",
            (audit_id,),
        ).fetchone()[0]
        if pending > 0:
            return False, f"라인 PENDING {pending}건 (실측 미완)"
        unresolved = c.execute(
            """SELECT COUNT(*) FROM stock_adjustments adj
               JOIN stock_audit_items ai ON adj.audit_item_id=ai.id
               WHERE ai.audit_id=? AND adj.status='PENDING'""",
            (audit_id,),
        ).fetchone()[0]
        if unresolved > 0:
            return False, f"조정 PENDING {unresolved}건 (승인/반려 필요)"
        c.execute(
            """UPDATE stock_audits
               SET status='CLOSED', end_date=date('now','localtime')
               WHERE id=?""",
            (audit_id,),
        )
        return True, "CLOSED"


def stock_audit_effect_summary(audit_id: int) -> dict:
    """효과 검증 KPI — 본 실사의 ADJUST stock_movements 합계 + 시스템 차이 매칭률.
    반환: {adjust_count, adjust_qty_sum, abs_qty_sum, line_total, line_matched, match_rate}
    """
    with db_session() as c:
        sm = c.execute(
            """SELECT COUNT(sm.id) AS cnt,
                      COALESCE(SUM(sm.quantity), 0) AS qsum,
                      COALESCE(SUM(ABS(sm.quantity)), 0) AS absqsum
               FROM stock_movements sm
               JOIN stock_adjustments adj ON adj.movement_id=sm.id
               JOIN stock_audit_items ai ON adj.audit_item_id=ai.id
               WHERE ai.audit_id=? AND sm.kind='ADJUST'""",
            (audit_id,),
        ).fetchone()
        line = c.execute(
            """SELECT COUNT(*) AS total,
                      SUM(CASE WHEN status='MATCHED' THEN 1 ELSE 0 END) AS matched
               FROM stock_audit_items WHERE audit_id=?""",
            (audit_id,),
        ).fetchone()
        total = int(line["total"] or 0)
        matched = int(line["matched"] or 0)
        rate = (matched / total * 100.0) if total > 0 else 0.0
        return {
            "adjust_count": int(sm["cnt"] or 0),
            "adjust_qty_sum": float(sm["qsum"] or 0),
            "abs_qty_sum": float(sm["absqsum"] or 0),
            "line_total": total,
            "line_matched": matched,
            "match_rate": round(rate, 1),
        }
# ---- /재고실사 2차 ----


def stock_movements_list(part_id: int = 0, kind: str = "", since: str = "",
                         until: str = "", po_id: int = 0, project_id: int = 0,
                         q: str = "", limit: int = 200) -> list[dict]:
    """수불부 조회. 최근 순."""
    sql = """SELECT sm.*,
                    p.part_no, p.part_name, p.unit AS part_unit, p.category,
                    po.po_number,
                    pr.mgmt_code, pr.name AS project_name,
                    cu.name AS customer_name,
                    u.name AS created_by_name
             FROM stock_movements sm
             LEFT JOIN parts p ON sm.part_id = p.id
             LEFT JOIN purchase_orders po ON sm.po_id = po.id
             LEFT JOIN projects pr ON sm.project_id = pr.id
             LEFT JOIN customers cu ON sm.customer_id = cu.id
             LEFT JOIN users u ON sm.created_by = u.id
             WHERE 1=1"""
    params: list = []
    if part_id:
        sql += " AND sm.part_id = ?"
        params.append(part_id)
    if kind and kind.upper() in MOVEMENT_KINDS:
        sql += " AND sm.kind = ?"
        params.append(kind.upper())
    if since:
        sql += " AND sm.occurred_at >= ?"
        params.append(since)
    if until:
        sql += " AND sm.occurred_at <= ?"
        params.append(until + " 23:59:59")
    if po_id:
        sql += " AND sm.po_id = ?"
        params.append(po_id)
    if project_id:
        sql += " AND sm.project_id = ?"
        params.append(project_id)
    if q:
        like = f"%{q}%"
        sql += """ AND (sm.movement_no LIKE ? OR p.part_no LIKE ? OR p.part_name LIKE ?
                        OR sm.reason LIKE ? OR sm.note LIKE ? OR sm.lot_no LIKE ?)"""
        params += [like, like, like, like, like, like]
    sql += " ORDER BY sm.occurred_at DESC, sm.id DESC LIMIT ?"
    params.append(limit)
    with db_session() as c:
        return [dict(r) for r in c.execute(sql, params).fetchall()]


def stock_kpi() -> dict:
    """물류 홈 대시보드용 KPI"""
    today = _date.today().isoformat()
    since_30 = (_date.today() - _td(days=30)).isoformat()
    with db_session() as c:
        parts_total = c.execute("SELECT COUNT(*) FROM parts WHERE is_active=1").fetchone()[0]
        # 재고 금액 = FIFO 남은 레이어의 실단가 합 (정확)
        # Fishbowl 반면교사: 평균 원가 금지 → 실 입고 단가 기반
        row_val = c.execute(
            """SELECT COALESCE(SUM(remaining_qty * unit_price),0)
               FROM stock_movements
               WHERE kind='IN' AND remaining_qty > 0"""
        ).fetchone()
        stock_value = row_val[0] if row_val else 0
        # 폴백: FIFO 데이터 없고 std_price만 있는 경우
        if stock_value == 0:
            row_fb = c.execute(
                "SELECT COALESCE(SUM(stock_qty * std_price),0) FROM parts WHERE is_active=1"
            ).fetchone()
            stock_value = row_fb[0] if row_fb else 0
        # 안전재고 미달
        low_stock = c.execute(
            """SELECT COUNT(*) FROM parts
               WHERE is_active=1 AND safety_stock > 0 AND stock_qty < safety_stock"""
        ).fetchone()[0]
        # 미입고 발주 (작성중/발주완료/부분입고)
        po_pending = c.execute(
            """SELECT COUNT(*) FROM purchase_orders
               WHERE status IN ('작성중','발주완료','부분입고')"""
        ).fetchone()[0]
        # 30일 입/출고 합계
        in_30 = c.execute(
            """SELECT COALESCE(SUM(quantity),0) FROM stock_movements
               WHERE kind='IN' AND occurred_at >= ?""",
            (since_30,),
        ).fetchone()[0] or 0
        out_30 = c.execute(
            """SELECT COALESCE(SUM(quantity),0) FROM stock_movements
               WHERE kind='OUT' AND occurred_at >= ?""",
            (since_30,),
        ).fetchone()[0] or 0
    return {
        "parts_total": parts_total,
        "stock_value": stock_value,
        "low_stock": low_stock,
        "po_pending": po_pending,
        "in_30d": in_30,
        "out_30d": abs(out_30),  # 음수 저장 → 절대값 표시
    }


def parts_usage_distribution(focus_biz_div: str = "공통",
                              days: int = 365, limit: int = 500) -> list[dict]:
    """v5H226z58 (2026-05-12) — 자재의 실제 사업부 사용 분포 계산.
    근거: stock_movements (kind='OUT') → projects.biz_div 조인.
    focus_biz_div: 분석 대상 자재의 현재 사업부 (예: '공통'으로 잘못 분류된 자재 정제용).
    days: 최근 N일 출고 이력만 집계 (None=전체).
    반환: [{part_id, part_no, part_name, biz_div(현재), m_qty, t_qty, unknown_qty, total_qty,
            m_pct, t_pct, unknown_pct, recommended, evidence_count, confidence}]
    confidence: HIGH(분포 90:10+, evidence≥5) / MID(70:30+, evidence≥3) / LOW (그 외)
    recommended:
      - M 90%+ → 'M'
      - T 90%+ → 'T'
      - 양쪽 30%+ → '공통'
      - evidence 부족 → 'review' (수동 검토)
    """
    since_clause = ""
    params: list = []
    if days and days > 0:
        cutoff = (_date.today() - _td(days=int(days))).isoformat()
        since_clause = "AND DATE(sm.occurred_at) >= ?"
        params.append(cutoff)

    with db_session() as c:
        # 대상 자재 후보: 현재 biz_div 가 focus_biz_div 인 활성 자재
        if focus_biz_div:
            base_rows = c.execute(
                """SELECT id, part_no, part_name, biz_div, category
                   FROM parts WHERE COALESCE(is_active,1)=1 AND biz_div = ?
                   ORDER BY id DESC LIMIT ?""",
                (focus_biz_div, int(limit)),
            ).fetchall()
        else:
            base_rows = c.execute(
                """SELECT id, part_no, part_name, biz_div, category
                   FROM parts WHERE COALESCE(is_active,1)=1
                   ORDER BY id DESC LIMIT ?""",
                (int(limit),),
            ).fetchall()

        if not base_rows:
            return []

        pids = [r["id"] for r in base_rows]
        qmarks = ",".join("?" * len(pids))

        # 출고 분포 (stock_movements 출고 → projects.biz_div)
        out_q = c.execute(
            f"""SELECT sm.part_id,
                       COALESCE(p.biz_div,'UNKNOWN') AS bd,
                       SUM(ABS(sm.quantity)) AS qty,
                       COUNT(*) AS cnt
                FROM stock_movements sm
                LEFT JOIN projects p ON sm.project_id = p.id
                WHERE sm.kind='OUT' AND sm.part_id IN ({qmarks}) {since_clause}
                GROUP BY sm.part_id, COALESCE(p.biz_div,'UNKNOWN')""",
            tuple(pids) + tuple(params),
        ).fetchall()

    # 분포 집계
    by_part: dict = {pid: {"M": 0.0, "T": 0.0, "공통": 0.0, "UNKNOWN": 0.0,
                            "evidence": 0} for pid in pids}
    for r in out_q:
        bd = r["bd"] if r["bd"] in ("M", "T", "공통") else "UNKNOWN"
        pid = r["part_id"]
        if pid in by_part:
            by_part[pid][bd] += float(r["qty"] or 0)
            by_part[pid]["evidence"] += int(r["cnt"] or 0)

    out: list[dict] = []
    for base in base_rows:
        pid = base["id"]
        d = by_part[pid]
        total = d["M"] + d["T"] + d["공통"] + d["UNKNOWN"]
        # 알려진(M/T) 분포만으로 비율 계산
        known = d["M"] + d["T"]
        if known > 0:
            m_pct = round(d["M"] * 100 / known, 1)
            t_pct = round(d["T"] * 100 / known, 1)
        else:
            m_pct = 0.0
            t_pct = 0.0
        unknown_pct = round(d["UNKNOWN"] * 100 / total, 1) if total else 0.0
        evidence = int(d["evidence"])

        # 추천 사업부 결정
        if evidence == 0 or known == 0:
            rec = "review"
            conf = "LOW"
        elif m_pct >= 90:
            rec = "M"
            conf = "HIGH" if evidence >= 5 else ("MID" if evidence >= 3 else "LOW")
        elif t_pct >= 90:
            rec = "T"
            conf = "HIGH" if evidence >= 5 else ("MID" if evidence >= 3 else "LOW")
        elif m_pct >= 30 and t_pct >= 30:
            rec = "공통"
            conf = "MID"
        else:
            rec = "review"
            conf = "LOW"

        out.append({
            "part_id": pid,
            "part_no": base["part_no"] or "",
            "part_name": base["part_name"] or "",
            "biz_div": base["biz_div"] or "",
            "category": base["category"] or "",
            "m_qty": round(d["M"], 1),
            "t_qty": round(d["T"], 1),
            "unknown_qty": round(d["UNKNOWN"], 1),
            "total_qty": round(total, 1),
            "m_pct": m_pct,
            "t_pct": t_pct,
            "unknown_pct": unknown_pct,
            "evidence_count": evidence,
            "recommended": rec,
            "confidence": conf,
        })
    # 정렬: 추천 변경(M/T) HIGH → MID → LOW → review 순
    rec_order = {"M": 0, "T": 1, "공통": 2, "review": 3}
    conf_order = {"HIGH": 0, "MID": 1, "LOW": 2}
    out.sort(key=lambda x: (rec_order.get(x["recommended"], 9),
                             conf_order.get(x["confidence"], 9),
                             -x["evidence_count"]))
    return out


# =====================================================
# v5H226z58 (2026-05-12) — 자재 일괄 등록 (Excel)
# 컬럼 매핑 (헤더명 → DB 필드, 대소문자/공백 무시):
#   part_no/자재코드/품번        → part_no (필수)
#   part_name/자재명/품명        → part_name (필수)
#   biz_div/사업부              → biz_div (M/T/공통)
#   spec/규격                   → spec
#   maker/제조사                → maker
#   origin/원산지               → origin
#   unit/단위                   → unit (기본 EA)
#   currency/통화               → currency (기본 KRW)
#   std_price/매입단가/표준단가/단가 → std_price
#   category/분류                → category
#   item_account/품목계정        → item_account (RAW/SUB/SEMI/FIN/CONS)
#   safety_stock/안전재고        → safety_stock
#   location/위치                → location
#   reorder_point/재주문점/ROP   → reorder_point
#   reorder_qty/재주문량/ROQ     → reorder_qty
#   hs_code/HS코드               → hs_code
#   note/비고                    → note
# 옵션:
#   dry_run=True  → 검증만 수행, INSERT 안 함 (미리보기)
#   force=True    → 유사 자재 검사 우회 (중복 방지 3층 방어 끄기)
# =====================================================
_PARTS_IMPORT_HEADER_MAP = {
    # 표준명 (lowercase, 공백/특수문자 제거 키 → DB 필드)
    # part_no — KNK 표준은 제조사 원본 코드 사용 (PRODUCT CODE 우선)
    # ⚠ 모호 키워드 "code" 단독은 제외 — BOM 양식의 "CODE" (KNK 내부 비사용 코드) 와 충돌
    "productcode": "part_no", "제품코드": "part_no", "코드명": "part_no",
    "partno": "part_no", "자재코드": "part_no", "품번": "part_no",
    "자재번호": "part_no", "부품코드": "part_no",
    # part_name
    "productname": "part_name", "제품명": "part_name",
    "partname": "part_name", "자재명": "part_name", "품명": "part_name",
    "name": "part_name", "부품명": "part_name",
    # biz_div
    "bizdiv": "biz_div", "사업부": "biz_div", "사업부서": "biz_div",
    # spec / maker / origin / unit / currency / price
    "spec": "spec", "규격": "spec", "사양": "spec",
    "manufacturer": "maker", "제조사": "maker", "메이커": "maker", "maker": "maker",
    "origin": "origin", "원산지": "origin", "산지": "origin",
    "unit": "unit", "단위": "unit",
    "currency": "currency", "통화": "currency",
    "unitprice": "std_price", "stdprice": "std_price", "표준단가": "std_price",
    "단가": "std_price", "price": "std_price", "기준단가": "std_price",
    "매입단가": "std_price", "purchaseprice": "std_price",  # v5H226z83 — 매입단가 표기 통일
    # 분류
    "category": "category", "분류": "category", "카테고리": "category", "구분": "category",
    "itemaccount": "item_account", "품목계정": "item_account",
    "procurementkind": "procurement_kind", "조달구분": "procurement_kind",
    # 재고
    "safetystock": "safety_stock", "안전재고": "safety_stock",
    "location": "location", "위치": "location", "보관위치": "location",
    "reorderpoint": "reorder_point", "재주문점": "reorder_point", "rop": "reorder_point",
    "reorderqty": "reorder_qty", "재주문량": "reorder_qty", "roq": "reorder_qty",
    "conversionfactor": "conversion_factor", "환산계수": "conversion_factor",
    "hscode": "hs_code", "hs": "hs_code", "관세코드": "hs_code",
    # 비고 / 기타
    "note": "note", "비고": "note", "메모": "note",
    "remarks": "note", "remarksnote": "note",
    "defaultwarehouse": "default_warehouse", "기본창고": "default_warehouse",
    "isactive": "is_active", "활성": "is_active", "사용여부": "is_active",
    # v5H226z109 (2026-05-17) — 누락 컬럼 일괄 추가
    "purpose": "purpose", "용도": "purpose", "사용처": "purpose", "application": "purpose",
    "subspec1": "sub_spec1", "보조규격1": "sub_spec1", "보조사양1": "sub_spec1",
    "subspec2": "sub_spec2", "보조규격2": "sub_spec2", "보조사양2": "sub_spec2",
    "subspec3": "sub_spec3", "보조규격3": "sub_spec3", "보조사양3": "sub_spec3",
    "taxinvoicename": "tax_invoice_name", "세금계산서명": "tax_invoice_name", "세금계산서품명": "tax_invoice_name",
    "tradeinvoicename": "trade_invoice_name", "거래명세서명": "trade_invoice_name", "거래명세서품명": "trade_invoice_name",
    "categorymain": "category_main", "분류대": "category_main", "대분류": "category_main",
    "categoryseries": "category_series", "분류시리즈": "category_series", "시리즈": "category_series",
    # v5H226z110 (2026-05-18) — 자재코드 = 제조사 모델명 (통합). 엑셀에 "모델명" 헤더가 있으면 자재코드로 매핑
    "modelname": "part_no", "model": "part_no", "모델명": "part_no",
    "모델번호": "part_no", "모델": "part_no", "manufacturermodel": "part_no",
    # v5H226z110b (2026-05-19) — 기본 협력사명 (발주 자동 매칭용)
    # v5H226z111 (2026-05-19) — "공급사" UI 표현이 "협력사"로 변경됨에 따라 별칭 추가
    "defaultsupplier": "default_supplier", "기본공급사": "default_supplier",
    "기본공급사명": "default_supplier", "공급사": "default_supplier",
    "공급사명": "default_supplier", "주공급사": "default_supplier",
    "기본협력사": "default_supplier", "기본협력사명": "default_supplier",
    "협력사": "default_supplier", "협력사명": "default_supplier",
    "주협력사": "default_supplier",
    "vendor": "default_supplier", "supplier": "default_supplier",
    "partner": "default_supplier",
    # v5H226z111 (2026-05-19) — 구매사 3슬롯 (협력사 마스터 검증 후 저장)
    "vendor1": "vendor1", "구매사1": "vendor1", "협력사1": "vendor1", "공급사1": "vendor1",
    "vendor2": "vendor2", "구매사2": "vendor2", "협력사2": "vendor2", "공급사2": "vendor2",
    "vendor3": "vendor3", "구매사3": "vendor3", "협력사3": "vendor3", "공급사3": "vendor3",
    # v5H226z110c (2026-05-19) — 제조사 담당자 풀세트
    "makercontactname": "maker_contact_name", "제조사담당자": "maker_contact_name",
    "제조사담당자명": "maker_contact_name", "메이커담당자": "maker_contact_name",
    "manufacturercontact": "maker_contact_name",
    "makercontactphone": "maker_contact_phone", "제조사전화": "maker_contact_phone",
    "제조사담당자전화": "maker_contact_phone", "메이커전화": "maker_contact_phone",
    "makercontactemail": "maker_contact_email", "제조사이메일": "maker_contact_email",
    "제조사담당자이메일": "maker_contact_email", "메이커이메일": "maker_contact_email",
    "manufactureremail": "maker_contact_email",
}

# v5H226z58 (2026-05-12) — BOM 양식 추가 (다국어/복합 헤더 부분 매칭용 우선순위 키워드)
# 헤더 cell 안에 줄바꿈+한국어/베트남어 혼재 시 부분 매칭으로 fallback.
# 매핑 우선순위 (길이 긴 키워드 먼저 매칭 — "productcode" 가 "code" 보다 우선)
_PARTS_IMPORT_PARTIAL_KEYS = [
    # part_no 계열 (PRODUCT CODE 우선)
    ("productcode", "part_no"), ("제품코드", "part_no"), ("코드명", "part_no"),
    ("partno", "part_no"), ("자재코드", "part_no"), ("자재번호", "part_no"),
    ("부품코드", "part_no"), ("품번", "part_no"),
    # part_name 계열 (PRODUCT NAME 우선)
    ("productname", "part_name"), ("제품명", "part_name"),
    ("partname", "part_name"), ("자재명", "part_name"), ("품명", "part_name"),
    ("부품명", "part_name"),
    # maker — MANUFACTURER 우선
    ("manufacturer", "maker"), ("제조사", "maker"), ("메이커", "maker"),
    # std_price — UNIT PRICE 우선
    ("unitprice", "std_price"), ("표준단가", "std_price"), ("매입단가", "std_price"),
    ("기준단가", "std_price"), ("단가", "std_price"),
    # spec
    ("규격", "spec"), ("사양", "spec"), ("spec", "spec"),
    # category — 구분 우선 (BOM 양식)
    ("category", "category"), ("카테고리", "category"), ("분류", "category"), ("구분", "category"),
    # unit
    ("단위", "unit"),  # "unitprice" 보다 뒤에 와야 함 — substring 충돌 방지
    # 재고 / 기타
    ("safetystock", "safety_stock"), ("안전재고", "safety_stock"),
    ("reorderpoint", "reorder_point"), ("재주문점", "reorder_point"),
    ("reorderqty", "reorder_qty"), ("재주문량", "reorder_qty"),
    ("hscode", "hs_code"),
    ("remarks", "note"), ("note", "note"), ("비고", "note"), ("메모", "note"),
    ("bizdiv", "biz_div"), ("사업부", "biz_div"),
    ("origin", "origin"), ("원산지", "origin"),
    # 무시할 BOM 전용 컬럼 (자재마스터와 무관)
    # — TOTAL, AMOUNT, DELIVERY, ORDER STATUS, DELIVERY DATE 등은 매핑 dict 에 없으므로 자동 스킵
]


def _norm_header_key(s: str) -> str:
    """헤더명 정규화 — 영숫자·한글만 남기고 소문자."""
    return _pq_re.sub(r'[^a-z0-9가-힣]', '', (s or '').lower())


def _normalize_bulk_is_active(val) -> int:
    """v5H226z646 (2026-06-24) — 엑셀 일괄등록 '활성' 컬럼을 0/1로 정규화.
    템플릿이 한국어 텍스트('활성'/'비활성')를 쓰는데 _validate_parts_payload는 0/1만 허용해서
    실제 등록 단계에서 대량 실패하던 회귀를 차단.
    한국어/영문/숫자 모두 호환, 알 수 없으면 활성(1) 폴백."""
    if val is None:
        return 1
    if isinstance(val, bool):
        return 1 if val else 0
    if isinstance(val, (int, float)):
        try:
            return 1 if int(val) else 0
        except (TypeError, ValueError):
            return 1
    s = str(val).strip().lower()
    if not s:
        return 1
    if s in ("1", "y", "yes", "true", "t", "활성", "on", "active", "사용", "사용중", "o", "ok"):
        return 1
    if s in ("0", "n", "no", "false", "f", "비활성", "휴면", "off", "inactive", "미사용", "정지", "중지", "x"):
        return 0
    return 1  # 알 수 없는 값 → 기본 활성


def _parts_build_cols_values(data: dict, existing_cols, now: str):
    """v5H226z734: INSERT INTO parts 용 (cols, values) 빌드 — 대량 일괄등록(한 트랜잭션)용.
    data 는 _validate_parts_payload 통과 가정. purchase_prices→std_price 동기화·vendor 미러 포함.
    existing_cols(PRAGMA table_info parts)로 확장 컬럼만 필터(z58 미적용 환경 호환).
    ⚠ 컬럼 정의는 parts_create() 의 base_cols/ext_pairs 와 동일하게 유지할 것(추가 시 둘 다 갱신)."""
    _pp_json, _pp_last = _parse_purchase_prices(data.get("purchase_prices"))
    if _pp_last is not None:
        data["std_price"] = _pp_last
    _pp_sync_vendor_columns(data, _pp_json)
    base_cols = ["part_no", "part_name", "spec", "maker", "origin", "unit",
                 "currency", "std_price", "biz_div", "category", "note",
                 "is_active", "safety_stock", "location", "created_at", "updated_at"]
    base_values = [
        (data.get("part_no") or "").strip(),
        (data.get("part_name") or "").strip(),
        (data.get("spec") or "").strip(),
        (data.get("maker") or "").strip(),
        (data.get("origin") or "").strip(),
        (data.get("unit") or "EA").strip() or "EA",
        (data.get("currency") or "KRW").strip() or "KRW",
        float(data.get("std_price") or 0),
        (data.get("biz_div") or "").strip(),
        (data.get("category") or "").strip(),
        (data.get("note") or "").strip(),
        1 if data.get("is_active", 1) else 0,
        float(data.get("safety_stock") or 0),
        (data.get("location") or "").strip() or None,
        now, now,
    ]
    ext_pairs = [
        ("item_account",       (data.get("item_account") or "").strip() or None),
        ("procurement_kind",   (data.get("procurement_kind") or "").strip() or None),
        ("category_main",      (data.get("category_main") or "").strip() or None),
        ("category_series",    (data.get("category_series") or "").strip() or None),
        ("reorder_point",      float(data.get("reorder_point") or 0)),
        ("reorder_qty",        float(data.get("reorder_qty") or 0)),
        ("conversion_factor",  float(data.get("conversion_factor") or 1)),
        ("sub_spec1",          (data.get("sub_spec1") or "").strip() or None),
        ("sub_spec2",          (data.get("sub_spec2") or "").strip() or None),
        ("sub_spec3",          (data.get("sub_spec3") or "").strip() or None),
        ("tax_invoice_name",   (data.get("tax_invoice_name") or "").strip() or None),
        ("trade_invoice_name", (data.get("trade_invoice_name") or "").strip() or None),
        ("default_warehouse",  (data.get("default_warehouse") or "").strip() or None),
        ("hs_code",            (data.get("hs_code") or "").strip() or None),
        ("purchase_prices",    _pp_json),
        ("purpose",            (data.get("purpose") or "").strip() or None),
        ("default_supplier",   (data.get("default_supplier") or "").strip() or None),
        ("maker_contact_name",  (data.get("maker_contact_name") or "").strip() or None),
        ("maker_contact_phone", (data.get("maker_contact_phone") or "").strip() or None),
        ("maker_contact_email", (data.get("maker_contact_email") or "").strip() or None),
        ("vendor1",            _validate_vendor_name(data.get("vendor1"))),
        ("vendor2",            _validate_vendor_name(data.get("vendor2"))),
        ("vendor3",            _validate_vendor_name(data.get("vendor3"))),
    ]
    ext_cols, ext_vals = [], []
    for _col, _val in ext_pairs:
        if _col in existing_cols:
            ext_cols.append(_col)
            ext_vals.append(_val)
    return base_cols + ext_cols, base_values + ext_vals


def parts_bulk_import_excel(file_path: str, dry_run: bool = True,
                             force_similar: bool = False,
                             full_preview: bool = False) -> dict:
    """Excel(.xlsx) 자재 일괄 등록.
    파싱: 1행 = 헤더, 2행~ = 데이터.
    dry_run=True: 검증·유사 자재 검사만 수행, DB INSERT 안 함 (미리보기 응답).
    dry_run=False: 실제 등록 + 결과 반환.
    force_similar=True: 유사 자재(>=85점) 검사 우회 (개별 행).
    반환: {
      'total': N, 'parsed': N, 'skipped_empty': N,
      'header_map': {col_idx: db_field},
      'unmapped_headers': [...],
      'preview': [{row_no, data, status, errors, similar}],  # 최대 200건
      'summary': {ok, error, similar_warn, unique_dup},
      'inserted_ids': [...]  # dry_run=False 일 때만
    }
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "openpyxl 미설치 — pip install openpyxl"}

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # v5H226z643 (2026-06-24) — 헤더 행 자동 검출 (KNK 발급 양식 호환)
    # 시스템 발급 템플릿은 1행 제목 · 2행 안내 · 3행 공백 · 4행 헤더 구조.
    # 첫 SCAN_MAX 행을 스캔, part_no AND part_name 둘 다 매칭되는 행 중 점수 최고를 헤더로 채택.
    SCAN_MAX = 15
    scan_rows: list = []
    for _ridx, _r in enumerate(rows_iter):
        scan_rows.append(_r)
        if _ridx + 1 >= SCAN_MAX:
            break

    if not scan_rows:
        wb.close()
        return {"error": "엑셀이 비어있습니다"}

    def _try_map_row(row):
        """행을 헤더로 가정해 매핑 시도 → (header_map, unmapped_list)."""
        hm: dict = {}
        unmapped_local: list = []
        used: set = set()
        if row is None:
            return hm, unmapped_local
        for ci, h in enumerate(row):
            if h is None:
                continue
            key = _norm_header_key(str(h))
            # 1) 정확 매칭
            fld = _PARTS_IMPORT_HEADER_MAP.get(key)
            if not fld:
                # 2) 부분 매칭 fallback — BOM 양식 다국어/복합 헤더 지원
                for kw, mapped in _PARTS_IMPORT_PARTIAL_KEYS:
                    if kw in key:
                        if mapped not in used:
                            fld = mapped
                            break
            if fld and fld not in used:
                hm[ci] = fld
                used.add(fld)
            else:
                unmapped_local.append({"col_idx": ci, "header": str(h)})
        return hm, unmapped_local

    header_row_idx = -1
    header_map: dict = {}
    unmapped: list = []
    best_score = 0
    for _ridx, _r in enumerate(scan_rows):
        _hm, _unmapped = _try_map_row(_r)
        # 헤더 후보 — 필수 컬럼 둘 다 매칭 필요. 점수(매핑 수) 최고를 선택.
        _vals = set(_hm.values())
        if "part_no" in _vals and "part_name" in _vals and len(_hm) > best_score:
            best_score = len(_hm)
            header_row_idx = _ridx
            header_map = _hm
            unmapped = _unmapped

    if header_row_idx < 0:
        # 후보 없음 → 1행 기준 매핑 결과로 기존 에러 메시지 반환 (사용자 진단용)
        header_map, unmapped = _try_map_row(scan_rows[0])
        wb.close()
        return {
            "error": "필수 컬럼 누락 — 'part_no/자재코드/품번' 과 'part_name/자재명/품명' 필수",
            "header_map": {k: v for k, v in header_map.items()},
            "unmapped_headers": unmapped,
        }

    preview: list = []
    inserted_ids: list = []
    skipped_empty = 0
    total = 0
    ok_cnt = 0
    err_cnt = 0
    sim_warn = 0
    dup_cnt = 0
    # v5H226z647 (2026-06-25) — 엑셀 내부 part_no 중복 검출용 누적 셋.
    # 같은 part_no가 여러 행에 적혀 있으면 첫 등장만 등록 가능(UNIQUE 제약). 두 번째부터는
    # 미리보기에서 미리 차단 → "정상" 카운트 = 실제 등록 가능 수가 정확히 일치.
    seen_part_nos: dict = {}  # part_no(소문자/공백제거) → 첫 등장 행 번호

    # 헤더 이후 행을 데이터로 처리. 행 번호는 1-indexed (헤더 행 번호 + 1 부터).
    from itertools import chain as _chain
    # v5H226z734 (대량 3000+ 대응): ① 데이터행을 먼저 수집해 건수 파악 ② 대량이면 유사검사 생략
    #   (행마다 자재 테이블 전체 LIKE 스캔이라 O(N×M)·정확 코드중복 검사는 유지) ③ 기존 part_no를
    #   1회 메모리 맵으로(행마다 DB 조회 제거). 등록 INSERT는 아래에서 한 트랜잭션 일괄.
    _data_rows = list(_chain(scan_rows[header_row_idx + 1:], rows_iter))

    def _row_empty(_r):
        return all(v is None or (isinstance(v, str) and not v.strip()) for v in _r)

    _nonempty_cnt = sum(1 for _r in _data_rows if not _row_empty(_r))
    LARGE_IMPORT_SKIP_SIMILAR = 800
    _do_similar = (not force_similar) and (_nonempty_cnt <= LARGE_IMPORT_SKIP_SIMILAR)
    similar_skipped = (not force_similar) and (_nonempty_cnt > LARGE_IMPORT_SKIP_SIMILAR)
    # 기존 자재 part_no → (id, name) 1회 로드 (행마다 SELECT 제거)
    with db_session() as _c0:
        _existing_pno = {}
        for _er in _c0.execute("SELECT part_no, id, part_name FROM parts").fetchall():
            if _er[0] is not None:
                _existing_pno[str(_er[0])] = (_er[1], _er[2])
    row_no = header_row_idx + 1  # 헤더 행 번호. 루프 내 += 1 로 데이터 행 시작.
    for row in _data_rows:
        row_no += 1
        # 빈 행 스킵
        if _row_empty(row):
            skipped_empty += 1
            continue
        total += 1
        # 데이터 추출
        data: dict = {}
        for ci, fld in header_map.items():
            if ci < len(row):
                val = row[ci]
                if isinstance(val, str):
                    val = val.strip()
                if val is not None and val != "":
                    data[fld] = val
        # v5H226z646 (2026-06-24): 한국어/영문 활성 표기 정규화
        # 엑셀 활성 컬럼이 "활성"/"비활성"/"Y"/"N" 등 텍스트일 때 _validate_parts_payload(0/1만 허용)에서
        # 대량 실패가 일어나던 회귀(미리보기 통과 → 실제 등록 0건) 차단.
        if "is_active" in data:
            data["is_active"] = _normalize_bulk_is_active(data["is_active"])
        # 필수 검증
        part_no = (data.get("part_no") or "").strip() if isinstance(data.get("part_no"), str) else str(data.get("part_no") or "").strip()
        part_name = (data.get("part_name") or "").strip() if isinstance(data.get("part_name"), str) else str(data.get("part_name") or "").strip()
        errors: list = []
        sim_hits: list = []
        status = "ok"

        if not part_no:
            errors.append("part_no 필수")
        if not part_name:
            errors.append("part_name 필수")

        # 사업부 검증
        bd = (data.get("biz_div") or "").strip() if isinstance(data.get("biz_div"), str) else str(data.get("biz_div") or "").strip()
        if bd and bd not in ("M", "T", "공통"):
            errors.append(f"사업부 값 오류 '{bd}' (M/T/공통)")
            bd = ""
        data["biz_div"] = bd or "공통"  # 기본 공통

        # part_no UNIQUE 검사
        # v5H226z647 (2026-06-25) — DB 중복 + 엑셀 내부 중복 모두 검사
        # 엑셀 내부 중복(같은 part_no가 위 행에 이미 있음)도 미리보기에서 차단 →
        # "정상" 카운트가 실제 등록 가능 수와 일치 (z647 이전에는 두 번째 행이
        # OK로 표시됐다가 실제 등록 시 silent fail 됨).
        if part_no and not errors:
            # ① 엑셀 내부 중복 (같은 파일 안에서 위 행에 이미 등장)
            #   대소문자·앞뒤 공백 차이 무시 — UNIQUE 제약은 원본 그대로지만 사용자 안내용
            seen_key = part_no.strip().lower()
            if seen_key and seen_key in seen_part_nos:
                first_row = seen_part_nos[seen_key]
                errors.append(f"엑셀 내부 코드 중복 — 같은 코드가 {first_row}행에 이미 있음(첫 등장만 등록 가능)")
                status = "dup"
                dup_cnt += 1
            else:
                if seen_key:
                    seen_part_nos[seen_key] = row_no
                # ② DB 기존 자재와 중복 검사 (z734: 메모리 맵 — 행마다 DB 조회 제거)
                dup = _existing_pno.get(part_no)
                if dup:
                    errors.append(f"코드 중복 — 기존 [{dup[0]}] {dup[1]}")
                    status = "dup"
                    dup_cnt += 1

        # v5H226z644: dry_run에서도 _validate_parts_payload 실행 — 모든 검증 오류를 미리보기에 노출.
        # 통화·단위·매입단가·is_active 등 잠재 이슈를 즉시 표면화. data는 in-place 정규화됨(currency/unit 대문자 등).
        if not errors and status not in ("dup",):
            try:
                _validate_parts_payload(data)
            except (ValueError, TypeError) as _ve:
                errors.append(str(_ve))

        # 유사 자재 검사 (force_similar=False + 대량 아님일 때만 — z734: 대량은 O(N×M)라 생략)
        if part_name and not errors and _do_similar:
            try:
                sims = parts_find_similar(
                    name=part_name,
                    spec=str(data.get("spec") or ""),
                    maker=str(data.get("maker") or ""),
                    limit=3, threshold=85,
                )
                if sims:
                    sim_hits = sims
                    if status == "ok":
                        status = "similar"
                        sim_warn += 1
            except Exception:
                pass

        if errors:
            status = "error" if status == "ok" else status
            err_cnt += 1
        elif status == "ok":
            ok_cnt += 1

        preview.append({
            "row_no": row_no,
            "data": data,
            "status": status,  # ok / error / dup / similar
            "errors": errors,
            "similar": [{
                "id": s["id"], "part_no": s.get("part_no"),
                "part_name": s.get("part_name"), "score": s.get("score"),
            } for s in sim_hits[:3]] if sim_hits else [],
        })

    wb.close()

    # 실제 등록 (dry_run=False 일 때만, status='ok' 또는 'similar' 행만)
    # v5H226z734: 행마다 parts_create(자체 세션·커밋) 대신 '한 트랜잭션 일괄 INSERT'(커밋 1회).
    #   컬럼 빌드는 공용 헬퍼(_parts_build_cols_values) 재사용. dup·유사검사는 위 미리보기 루프에서
    #   이미 끝나 INSERT만 수행(코드중복 행은 status=dup 으로 제외됨). 행별 오류는 격리(나머지 진행).
    if not dry_run:
        _now2 = _logi_now()
        with db_session() as c:
            _existing_cols = {r[1] for r in c.execute("PRAGMA table_info(parts)").fetchall()}
            for p in preview:
                if p["status"] in ("ok", "similar") and not p["errors"]:
                    try:
                        _cols, _vals = _parts_build_cols_values(p["data"], _existing_cols, _now2)
                        _cur = c.execute(
                            f"INSERT INTO parts ({','.join(_cols)}) VALUES ({','.join(['?'] * len(_cols))})",
                            _vals,
                        )
                        inserted_ids.append({"row_no": p["row_no"], "id": _cur.lastrowid,
                                              "part_no": p["data"].get("part_no")})
                        p["status"] = "inserted"
                        p["inserted_id"] = _cur.lastrowid
                    except Exception as e:
                        p["status"] = "error"
                        p["errors"].append(str(e))
                        err_cnt += 1
                        ok_cnt = max(0, ok_cnt - 1)

    return {
        "total": total,
        "parsed": total,
        "skipped_empty": skipped_empty,
        "header_map": {str(k): v for k, v in header_map.items()},
        "unmapped_headers": unmapped,
        # v5H226z648: full_preview=True 면 전체 행 반환(검증결과 엑셀 다운로드용). 기본 False는 200건.
        "preview": preview if full_preview else preview[:200],
        "preview_truncated": len(preview) > 200,
        "summary": {
            "ok": ok_cnt, "error": err_cnt,
            "similar_warn": sim_warn, "dup": dup_cnt,
            "inserted": len(inserted_ids),
        },
        "inserted_ids": inserted_ids,
        "dry_run": dry_run,
        # v5H226z734: 대량 업로드 — 유사 자재 검사 생략 여부(UI 안내용). 정확 코드중복은 항상 검사.
        "similar_skipped": similar_skipped,
        "large_import": _nonempty_cnt > LARGE_IMPORT_SKIP_SIMILAR,
    }


def parts_bulk_import_export_xlsx(file_path: str) -> tuple:
    """v5H226z648 (2026-06-25) — 자재 일괄등록 검증결과를 색칠한 엑셀로 다운로드.
    원본 양식 그대로 유지 + 헤더 끝에 [검증 결과 / 사유] 컬럼 추가 + 상태별 행 색칠.
    구매팀이 받아 어느 행이 중복/오류인지 한눈에 보고 정리할 수 있도록 제공.
    반환: (xlsx_bytes or None, info_dict)"""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from io import BytesIO

    # 1) 전체 행 검증 (full_preview=True 로 모든 행 결과 받음)
    result = parts_bulk_import_excel(file_path, dry_run=True, force_similar=False, full_preview=True)
    if result.get("error"):
        return None, result
    preview_by_row = {p["row_no"]: p for p in result.get("preview", [])}

    # 2) 원본 엑셀 수정 가능 모드로 다시 로드 (read_only=False)
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        return None, {"error": f"엑셀 로드 실패: {e}"}
    ws = wb.active

    # 3) 헤더 행 검출 (z643 자동검출 로직과 동일 — 첫 15행 스캔)
    SCAN_MAX = 15
    header_row_idx = None
    best_len = 0
    max_scan = min(SCAN_MAX, ws.max_row or 0)
    for ri in range(1, max_scan + 1):
        row_vals = []
        for ci in range(1, (ws.max_column or 0) + 1):
            row_vals.append(ws.cell(row=ri, column=ci).value)
        # 인라인 매핑 시도 — 같은 알고리즘
        hm = {}
        used = set()
        for ci, h in enumerate(row_vals):
            if h is None:
                continue
            key = _norm_header_key(str(h))
            fld = _PARTS_IMPORT_HEADER_MAP.get(key)
            if not fld:
                for kw, mapped in _PARTS_IMPORT_PARTIAL_KEYS:
                    if kw in key:
                        if mapped not in used:
                            fld = mapped
                            break
            if fld and fld not in used:
                hm[ci] = fld
                used.add(fld)
        vals = set(hm.values())
        if "part_no" in vals and "part_name" in vals and len(hm) > best_len:
            best_len = len(hm)
            header_row_idx = ri

    if header_row_idx is None:
        wb.close()
        return None, {"error": "헤더 행을 찾을 수 없습니다"}

    # 4) 헤더 행 끝에 새 컬럼 2개 추가 (검증 결과 / 사유)
    last_col = ws.max_column or 1
    col_status = last_col + 1
    col_reason = last_col + 2
    hcell_status = ws.cell(row=header_row_idx, column=col_status)
    hcell_reason = ws.cell(row=header_row_idx, column=col_reason)
    hcell_status.value = "검증 결과"
    hcell_reason.value = "사유"
    bold = Font(bold=True)
    hcell_status.font = bold
    hcell_reason.font = bold
    # 컬럼 폭 (간단 추정)
    try:
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_status)].width = 22
        ws.column_dimensions[get_column_letter(col_reason)].width = 60
    except Exception:
        pass

    # 5) 색 정의 (옅은 톤 — 50대 가독성)
    FILL_DUP_INTERNAL = PatternFill("solid", fgColor="FECACA")  # 빨강 옅음
    FILL_DUP_DB       = PatternFill("solid", fgColor="FEF3C7")  # 노랑 옅음
    FILL_ERROR        = PatternFill("solid", fgColor="FED7AA")  # 주황 옅음
    FILL_SIMILAR      = PatternFill("solid", fgColor="DCFCE7")  # 연두 옅음

    # 6) 각 데이터 행 색칠 + 새 컬럼 채우기
    cnt = {"dup_internal": 0, "dup_db": 0, "error": 0, "similar": 0, "ok": 0}
    for r_no, p in preview_by_row.items():
        if r_no <= header_row_idx:
            continue
        status = p.get("status") or "ok"
        errors = p.get("errors") or []
        fill = None
        label = "✅ 정상"
        reason = ""
        if errors and any("엑셀 내부 코드 중복" in e for e in errors):
            fill = FILL_DUP_INTERNAL
            label = "❌ 엑셀 내부 중복"
            reason = next((e for e in errors if "엑셀 내부 코드 중복" in e), errors[0])
            cnt["dup_internal"] += 1
        elif errors and any("코드 중복 — 기존" in e for e in errors):
            fill = FILL_DUP_DB
            label = "⚠ DB 기존 등록"
            reason = next((e for e in errors if "코드 중복 — 기존" in e), errors[0])
            cnt["dup_db"] += 1
        elif errors:
            fill = FILL_ERROR
            label = "❌ 오류"
            reason = " · ".join(errors[:3])
            cnt["error"] += 1
        elif status == "similar":
            fill = FILL_SIMILAR
            label = "⚠ 유사자재 경고"
            sims = p.get("similar") or []
            reason = " · ".join(
                f"[{s.get('id')}] {s.get('part_name', '')} ({s.get('score', '')}점)"
                for s in sims[:2]
            )
            cnt["similar"] += 1
        else:
            cnt["ok"] += 1
        # 새 컬럼 값
        ws.cell(row=r_no, column=col_status).value = label
        ws.cell(row=r_no, column=col_reason).value = reason
        # 행 색칠 (헤더 제외, 새 컬럼까지)
        if fill is not None:
            for ci in range(1, col_reason + 1):
                ws.cell(row=r_no, column=ci).fill = fill

    # 7) 시트명에 요약 (한글 31자 제한 — 짧게)
    total = result.get("total", 0)
    try:
        ws.title = f"검증결과({total}건)"[:31]
    except Exception:
        pass

    # 8) bytes 반환
    buf = BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue(), {
        "summary": result.get("summary", {}),
        "total": total,
        "header_row_idx": header_row_idx,
        "counts": cnt,
    }


def parts_bulk_reclassify(updates: list[dict], actor_id: int = 0) -> dict:
    """v5H226z58 (2026-05-12) — 자재 사업부 일괄 재분류.
    updates: [{'part_id': int, 'new_biz_div': 'M'|'T'|'공통'}]
    반환: {'ok': N, 'errors': [...]}"""
    ok = 0
    errors = []
    with db_session() as c:
        for u in updates:
            try:
                pid = int(u.get("part_id") or 0)
                bd = (u.get("new_biz_div") or "").strip()
                if pid <= 0 or bd not in ("M", "T", "공통"):
                    errors.append({"part_id": pid, "reason": "invalid input"})
                    continue
                c.execute(
                    "UPDATE parts SET biz_div=?, updated_at=datetime('now','localtime') WHERE id=?",
                    (bd, pid),
                )
                ok += 1
            except Exception as e:
                errors.append({"part_id": u.get("part_id"), "reason": str(e)})
    return {"ok": ok, "errors": errors, "total": len(updates)}


def stock_trend_7d() -> list[dict]:
    """v5H226z58 (2026-05-12) — 자재구매센터 홈 hero 카드용 7일 입출고 추세.
    반환: 오래된 날짜 → 최신 날짜 순서 7개 row.
    각 row: {date: 'MM-DD', wday: '월~일', in_qty, out_qty, in_cnt, out_cnt}
    외부 차트 라이브러리 0건 (호출자가 인라인 SVG 렌더링)."""
    today = _date.today()
    rows = []
    with db_session() as c:
        for d in range(6, -1, -1):  # 6일 전 → 오늘
            target = today - _td(days=d)
            target_s = target.isoformat()
            in_row = c.execute(
                """SELECT COALESCE(SUM(quantity),0) AS q, COUNT(*) AS n
                   FROM stock_movements
                   WHERE kind='IN' AND DATE(occurred_at)=?""",
                (target_s,),
            ).fetchone()
            out_row = c.execute(
                """SELECT COALESCE(SUM(ABS(quantity)),0) AS q, COUNT(*) AS n
                   FROM stock_movements
                   WHERE kind='OUT' AND DATE(occurred_at)=?""",
                (target_s,),
            ).fetchone()
            wday_ko = ['월','화','수','목','금','토','일'][target.weekday()]
            rows.append({
                "date": target.strftime("%m-%d"),
                "iso": target_s,
                "wday": wday_ko,
                "in_qty": float(in_row["q"] or 0),
                "out_qty": float(out_row["q"] or 0),
                "in_cnt": int(in_row["n"] or 0),
                "out_cnt": int(out_row["n"] or 0),
                "is_today": (d == 0),
            })
    return rows


def part_stock_history(part_id: int, limit: int = 50) -> list[dict]:
    """특정 부품의 입출고 이력 (파트 상세에서 사용)"""
    return stock_movements_list(part_id=part_id, limit=limit)


def part_fifo_layers(part_id: int) -> list[dict]:
    """특정 부품의 남아있는 FIFO 레이어 조회 (재고 상세 분석)"""
    with db_session() as c:
        rows = c.execute(
            """SELECT sm.id, sm.movement_no, sm.remaining_qty, sm.unit_price,
                      sm.lot_no, sm.expiry_date, sm.occurred_at,
                      po.po_number, s.name AS supplier_name
               FROM stock_movements sm
               LEFT JOIN purchase_orders po ON sm.po_id = po.id
               LEFT JOIN suppliers s ON po.supplier_id = s.id
               WHERE sm.part_id=? AND sm.kind='IN' AND sm.remaining_qty > 0
               ORDER BY sm.occurred_at ASC, sm.id ASC""",
            (part_id,),
        ).fetchall()
        return [dict(r) for r in rows]


# =====================================================
# Top3 S2 3차 (2026-04-26) — FIFO 레이어 강화 + ABC 분류 + 재고회전율
# 원칙: stock_movements 단일 진실원 활용 · 평균원가 금지 (Fishbowl 반면교사)
# =====================================================
def fifo_layers(part_id: int) -> dict:
    """FIFO 레이어 상세 + 잔량/평균단가 요약. part_fifo_layers 위에 집계 추가."""
    layers = part_fifo_layers(part_id)
    total_qty = sum(float(r.get("remaining_qty") or 0) for r in layers)
    total_val = sum(float(r.get("remaining_qty") or 0) * float(r.get("unit_price") or 0)
                    for r in layers)
    avg_price = (total_val / total_qty) if total_qty > 0 else 0.0
    with db_session() as c:
        p = c.execute("SELECT part_no, part_name, unit FROM parts WHERE id=?",
                      (part_id,)).fetchone()
    return {
        "part_id": part_id,
        "part_no": p["part_no"] if p else "",
        "part_name": p["part_name"] if p else "",
        "unit": (p["unit"] if p else "") or "EA",
        "layers": layers,
        "layers_count": len(layers),
        "total_qty": total_qty,
        "total_value": total_val,
        "avg_unit_price": avg_price,
    }


def abc_classification(days: int = 90) -> list[dict]:
    """최근 N일 (기본 90일) 출고 매출액(qty*unit_price) 기준 ABC 분류.
    A: 누적 상위 80% (기준일 매출 약 20% 부품 — 파레토 원칙).
    B: 80~95%. C: 나머지.
    """
    since = (_date.today() - _td(days=days)).isoformat()
    with db_session() as c:
        rows = c.execute(
            """SELECT sm.part_id, p.part_no, p.part_name, p.unit,
                      SUM(ABS(sm.quantity)) AS out_qty,
                      SUM(ABS(sm.quantity) * COALESCE(sm.unit_price, p.std_price, 0)) AS out_value
               FROM stock_movements sm
               LEFT JOIN parts p ON sm.part_id = p.id
               WHERE sm.kind='OUT' AND sm.occurred_at >= ?
               GROUP BY sm.part_id
               ORDER BY out_value DESC""",
            (since,),
        ).fetchall()
    items = [dict(r) for r in rows]
    total = sum(float(r.get("out_value") or 0) for r in items) or 1.0
    cum = 0.0
    for r in items:
        v = float(r.get("out_value") or 0)
        cum += v
        share = (v / total) * 100.0
        cum_share = (cum / total) * 100.0
        r["share_pct"] = share
        r["cum_share_pct"] = cum_share
        if cum_share <= 80.0:
            r["abc_class"] = "A"
        elif cum_share <= 95.0:
            r["abc_class"] = "B"
        else:
            r["abc_class"] = "C"
    return items


def stock_turnover(days: int = 90) -> list[dict]:
    """부품별 재고 회전율 = 출고량 / 평균재고. 평균재고는 (현재 on_hand + max(0, on_hand+출고량))/2
    근사 (간이 — 일별 스냅샷 미보유 상황). 슬로우/패스트 분류 동반.
    """
    since = (_date.today() - _td(days=days)).isoformat()
    with db_session() as c:
        rows = c.execute(
            """SELECT p.id AS part_id, p.part_no, p.part_name, p.unit,
                      COALESCE(sb.on_hand, 0) AS on_hand,
                      COALESCE((SELECT SUM(ABS(sm.quantity)) FROM stock_movements sm
                                WHERE sm.part_id=p.id AND sm.kind='OUT'
                                  AND sm.occurred_at >= ?), 0) AS out_qty
               FROM parts p
               LEFT JOIN stock_balances sb ON sb.part_id = p.id
               WHERE p.is_active=1
               ORDER BY p.part_no""",
            (since,),
        ).fetchall()
    items = []
    for r in rows:
        d = dict(r)
        on_hand = float(d.get("on_hand") or 0)
        out_qty = float(d.get("out_qty") or 0)
        # 평균재고 근사: 현 잔고 + 시작 시점 추정(현 잔고 + 출고량) → /2
        avg_stock = (on_hand + (on_hand + out_qty)) / 2.0 if (on_hand + out_qty) > 0 else 0.0
        turnover = (out_qty / avg_stock) if avg_stock > 0 else 0.0
        # 패스트(>=2.0회) / 노멀(0.5~2.0) / 슬로우(<0.5, 출고 0 포함)
        if turnover >= 2.0:
            band = "FAST"
        elif turnover >= 0.5:
            band = "NORMAL"
        else:
            band = "SLOW"
        d["avg_stock"] = avg_stock
        d["turnover"] = turnover
        d["band"] = band
        items.append(d)
    items.sort(key=lambda x: x["turnover"], reverse=True)
    return items


# =====================================================
# 환율 관리 (FXLoader 패턴, 수동 입력 시작)
# =====================================================
CURRENCIES = ["KRW", "USD", "VND", "JPY", "CNY", "EUR"]


def get_exchange_rate(date_str: str, from_currency: str, to_currency: str = "KRW") -> float:
    """특정 날짜의 환율 조회. 가장 가까운 과거 레이트 사용 (없으면 1.0 폴백).
    같은 통화면 1.0.
    """
    if from_currency == to_currency:
        return 1.0
    try:
        with db_session() as c:
            r = c.execute(
                """SELECT rate FROM exchange_rates
                   WHERE from_currency=? AND to_currency=? AND rate_date <= ?
                   ORDER BY rate_date DESC LIMIT 1""",
                (from_currency, to_currency, date_str),
            ).fetchone()
            if r:
                return float(r["rate"])
    except Exception:
        pass
    return 1.0  # 환율 없으면 1.0 (운영 시작 전 안전 기본)


def _get_active_fx_rate(currency: str, ref_date: str = None, to_currency: str = "KRW") -> float | None:
    """v5H126: 통화·기준일 기준 가장 가까운 (≤ ref_date) 환율 조회.
    없으면 None (호출측에서 폴백 결정). 같은 통화면 1.0.
    수금/CI 등록 시 fx_rate snapshot 자동 채움 용도."""
    if not currency:
        return None
    currency = currency.upper()
    to_currency = (to_currency or "KRW").upper()
    if currency == to_currency:
        return 1.0
    if not ref_date:
        from datetime import date as _d
        ref_date = _d.today().isoformat()
    try:
        with db_session() as c:
            r = c.execute(
                """SELECT rate FROM exchange_rates
                   WHERE from_currency=? AND to_currency=? AND rate_date <= ?
                   ORDER BY rate_date DESC LIMIT 1""",
                (currency, to_currency, ref_date),
            ).fetchone()
            if r:
                return float(r["rate"])
            # 미래 환율만 있는 케이스 — 가장 빠른 미래 레이트라도 폴백
            r2 = c.execute(
                """SELECT rate FROM exchange_rates
                   WHERE from_currency=? AND to_currency=?
                   ORDER BY rate_date ASC LIMIT 1""",
                (currency, to_currency),
            ).fetchone()
            if r2:
                return float(r2["rate"])
    except Exception:
        pass
    return None


def exchange_rate_create(data: dict, user_id: int) -> int:
    """환율 등록 (수동). 같은 날짜·통화쌍은 UPSERT.
    v5H112: 0/음수 환율 차단."""
    rate = float(data.get("rate") or 0)
    if rate <= 0:
        raise ValueError(f"환율은 0보다 커야 합니다 (입력값: {rate}).")
    if not (data.get("rate_date") and data.get("from_currency")):
        raise ValueError("환율일자와 기준통화는 필수입니다.")
    with db_session() as c:
        c.execute(
            """INSERT INTO exchange_rates(rate_date, from_currency, to_currency, rate, source, note, created_by)
               VALUES (?,?,?,?,?,?,?)
               ON CONFLICT(rate_date, from_currency, to_currency)
               DO UPDATE SET rate=excluded.rate, source=excluded.source,
                             note=excluded.note, created_by=excluded.created_by,
                             created_at=datetime('now','localtime')""",
            (data["rate_date"], data["from_currency"].upper(),
             data.get("to_currency", "KRW").upper(),
             float(data["rate"]),
             data.get("source") or "수동",
             data.get("note"),
             user_id),
        )
        return c.execute("SELECT last_insert_rowid()").fetchone()[0]


def exchange_rates_list(limit: int = 100, currency: str = "") -> list[dict]:
    sql = """SELECT er.*, u.name AS created_by_name
             FROM exchange_rates er LEFT JOIN users u ON er.created_by=u.id
             WHERE 1=1"""
    params = []
    if currency:
        sql += " AND (from_currency=? OR to_currency=?)"
        params += [currency.upper(), currency.upper()]
    sql += " ORDER BY rate_date DESC, from_currency LIMIT ?"
    params.append(limit)
    with db_session() as c:
        return [dict(r) for r in c.execute(sql, params).fetchall()]


def exchange_rates_latest() -> dict:
    """통화별 가장 최근 환율 (to KRW). 대시보드용."""
    with db_session() as c:
        rows = c.execute(
            """SELECT er.* FROM exchange_rates er
               INNER JOIN (
                 SELECT from_currency, to_currency, MAX(rate_date) AS mx
                 FROM exchange_rates WHERE to_currency='KRW'
                 GROUP BY from_currency, to_currency
               ) m ON er.from_currency=m.from_currency
                   AND er.to_currency=m.to_currency
                   AND er.rate_date=m.mx"""
        ).fetchall()
        return {r["from_currency"]: dict(r) for r in rows}


# =====================================================
# 적용일자 단가 관리 (한국 ERP 표준 — emaxit Frame7 참조)
# =====================================================
PRICE_TYPES = ["확정", "가", "견적"]


# v5H113: 도메인 enum 화이트리스트 (LOW #15/#20/#21)
# v5H116: CURRENCIES (line 5591) 와 단일 진실 소스로 동기화 — 두 상수가 갈라지지 않도록 파생.
CURRENCY_OPTIONS = tuple(CURRENCIES)
PART_UNIT_OPTIONS = ("EA", "SET", "BOX", "M", "KG", "L", "PCS", "PR", "PKG")


def part_price_create(data: dict, user_id: int) -> int:
    """적용일자 단가 등록. effective_from 필수.
    v5H113 LOW#15: effective_from <= effective_to 검증, currency 화이트리스트.
    """
    # v5H113 검증
    cur = (data.get("currency") or "KRW").strip().upper()
    if cur not in CURRENCY_OPTIONS:
        raise ValueError(f"통화는 {', '.join(CURRENCY_OPTIONS)} 중 하나여야 합니다.")
    ef = (data.get("effective_from") or "").strip()
    et = (data.get("effective_to") or "").strip()
    if ef and et and ef > et:
        raise ValueError(f"적용 시작일({ef})은 종료일({et})보다 이전이어야 합니다.")
    try:
        if float(data.get("unit_price") or 0) < 0:
            raise ValueError("단가는 0 이상이어야 합니다.")
    except (TypeError, ValueError) as _e:
        if "단가" in str(_e):
            raise
        raise ValueError("단가가 올바른 숫자가 아닙니다.")
    data["currency"] = cur
    with db_session() as c:
        # v5H124 MED: 활성기간 겹침 검증 — 동일 (part_id, supplier_id) 의 기존 row 와 겹치면 경고 (차단 X)
        # 정책: '확정(승인)' price_type 는 동일 키로 겹침 시 거부, '견적' 은 허용 (역사 단가 보존)
        if (data.get("price_type") or "견적") == "확정":
            overlap = c.execute(
                """SELECT id, effective_from, effective_to FROM part_prices
                   WHERE part_id=? AND COALESCE(supplier_id,0)=COALESCE(?,0)
                     AND price_type='확정'
                     AND COALESCE(effective_from,'') <= COALESCE(?,'9999-12-31')
                     AND COALESCE(effective_to,'9999-12-31') >= COALESCE(?,'0000-01-01')""",
                (int(data["part_id"]), data.get("supplier_id"),
                 data.get("effective_to") or "9999-12-31", data["effective_from"]),
            ).fetchone()
            if overlap:
                raise ValueError(
                    f"확정 단가 활성기간 겹침 — 기존 row #{overlap[0]} "
                    f"({overlap[1]}~{overlap[2] or '무기한'}) 와 충돌. "
                    "기존 row 의 effective_to 를 먼저 종료시키거나 견적으로 등록해주세요."
                )
        c.execute(
            """INSERT INTO part_prices
               (part_id, supplier_id, price_type, unit_price, currency,
                effective_from, effective_to, negotiated_at, min_qty, max_qty, note,
                created_by)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                int(data["part_id"]),
                data.get("supplier_id"),
                data.get("price_type") or "견적",
                float(data["unit_price"]),
                data.get("currency") or "KRW",
                data["effective_from"],
                data.get("effective_to"),
                data.get("negotiated_at"),
                float(data.get("min_qty") or 0),
                data.get("max_qty"),
                data.get("note"),
                user_id,
            ),
        )
        return c.execute("SELECT last_insert_rowid()").fetchone()[0]


def part_price_approve(price_id: int, user_id: int):
    """단가 승인 → 확정 전환"""
    with db_session() as c:
        c.execute(
            """UPDATE part_prices
               SET price_type='확정', approved_by=?, approved_at=?
               WHERE id=?""",
            (user_id, _logi_now(), price_id),
        )


def part_prices_list(part_id: int, supplier_id: int = 0,
                     include_inactive: bool = True) -> list[dict]:
    """부품의 단가 이력."""
    sql = """SELECT pp.*, s.name AS supplier_name, u.name AS approved_by_name,
                    cu.name AS created_by_name
             FROM part_prices pp
             LEFT JOIN suppliers s ON pp.supplier_id = s.id
             LEFT JOIN users u ON pp.approved_by = u.id
             LEFT JOIN users cu ON pp.created_by = cu.id
             WHERE pp.part_id = ?"""
    params = [part_id]
    if supplier_id:
        sql += " AND pp.supplier_id = ?"
        params.append(supplier_id)
    if not include_inactive:
        today = _date.today().isoformat()
        sql += " AND pp.effective_from <= ? AND (pp.effective_to IS NULL OR pp.effective_to >= ?)"
        params += [today, today]
    sql += " ORDER BY pp.effective_from DESC, pp.id DESC"
    with db_session() as c:
        return [dict(r) for r in c.execute(sql, params).fetchall()]


def part_active_price(part_id: int, supplier_id: int = 0,
                      date_str: str = "", qty: float = 0) -> dict | None:
    """특정 날짜에 유효한 확정/가 단가 조회. 없으면 None.
    우선순위: 승인된 확정 > 가격 타입 우선순위 > 가장 최신 effective_from
    """
    date_s = date_str or _date.today().isoformat()
    sql = """SELECT pp.*, s.name AS supplier_name
             FROM part_prices pp
             LEFT JOIN suppliers s ON pp.supplier_id = s.id
             WHERE pp.part_id = ?
               AND pp.effective_from <= ?
               AND (pp.effective_to IS NULL OR pp.effective_to >= ?)"""
    params = [part_id, date_s, date_s]
    if supplier_id:
        sql += " AND (pp.supplier_id = ? OR pp.supplier_id IS NULL)"
        params.append(supplier_id)
    if qty > 0:
        sql += " AND (pp.min_qty <= ? AND (pp.max_qty IS NULL OR pp.max_qty >= ?))"
        params += [qty, qty]
    # 확정 > 가 > 견적 순서 우선
    sql += """ ORDER BY
                CASE pp.price_type WHEN '확정' THEN 1 WHEN '가' THEN 2 ELSE 3 END,
                pp.effective_from DESC, pp.id DESC LIMIT 1"""
    with db_session() as c:
        r = c.execute(sql, params).fetchone()
        return dict(r) if r else None


# =====================================================
# 사이클 54 환율·단가 1차 (2026-04-27) — 명시적 헬퍼 +3
# 외부 API 0건. 기존 exchange_rates / part_prices 위에서 동작.
# =====================================================
def get_latest_fx_rate(currency_from: str, currency_to: str = "KRW",
                       as_of: str = "") -> float | None:
    """최신 유효일 환율 조회. 없으면 None.

    - currency_from == currency_to: 1.0 즉시 반환
    - as_of 미지정: 오늘 기준 가장 최근 effective rate
    - SQL 파라미터 바인딩 의무 준수 (?)
    """
    cf = (currency_from or "").upper()
    ct = (currency_to or "KRW").upper()
    if cf == ct:
        return 1.0
    date_s = as_of or _date.today().isoformat()
    try:
        with db_session() as c:
            r = c.execute(
                """SELECT rate FROM exchange_rates
                   WHERE from_currency=? AND to_currency=? AND rate_date<=?
                   ORDER BY rate_date DESC LIMIT 1""",
                (cf, ct, date_s),
            ).fetchone()
            return float(r["rate"]) if r else None
    except Exception:
        return None


def convert_amount(amount: float, currency_from: str,
                   currency_to: str = "KRW", as_of: str = "") -> float:
    """환산값. rate 미존재 시 amount 그대로 반환 (warn은 호출측 책임).

    - 같은 통화: amount 그대로
    - rate 조회 실패: amount 그대로 (안전 폴백)
    """
    try:
        amt = float(amount or 0)
    except Exception:
        return 0.0
    cf = (currency_from or "").upper()
    ct = (currency_to or "KRW").upper()
    if cf == ct:
        return amt
    rate = get_latest_fx_rate(cf, ct, as_of=as_of)
    if rate is None:
        return amt  # rate 미존재 — 원본 통화 금액 그대로 (호출측 warn)
    return amt * rate


def get_latest_part_price(part_id: int, price_type: str = "cost",
                          currency: str = "KRW",
                          as_of: str = "") -> dict | None:
    """최신 단가 조회. 없으면 None.

    price_type 매핑 (외부 API 호환 표준 ↔ 내부 한국 ERP 용어):
      - cost   → '확정' (원가/확정 단가)
      - list   → '가'   (정가/가 단가)
      - export → '견적' (수출가/견적 단가)
    내부 표준 한국어 (확정/가/견적) 도 그대로 허용.
    """
    type_map = {"cost": "확정", "list": "가", "export": "견적"}
    pt = type_map.get((price_type or "").lower(), price_type or "확정")
    cur = (currency or "KRW").upper()
    date_s = as_of or _date.today().isoformat()
    try:
        with db_session() as c:
            r = c.execute(
                """SELECT pp.* FROM part_prices pp
                   WHERE pp.part_id=? AND pp.price_type=? AND pp.currency=?
                     AND pp.effective_from<=?
                     AND (pp.effective_to IS NULL OR pp.effective_to>=?)
                   ORDER BY pp.effective_from DESC, pp.id DESC LIMIT 1""",
                (int(part_id), pt, cur, date_s, date_s),
            ).fetchone()
            return dict(r) if r else None
    except Exception:
        return None


# =====================================================
# 공급사 리드타임 자동 집계 (§6.4 업계 공통)
# =====================================================
def health_check() -> list[dict]:
    """시스템 건전성 점검. 각 기능이 진짜 동작하는지 확인.
    반환: [{name, status, detail, level}] — status: ok/warn/error/info
    """
    checks = []

    # 1. DB 연결
    try:
        with db_session() as c:
            c.execute("SELECT 1").fetchone()
        checks.append({"name": "DB 연결", "status": "ok",
                       "detail": "knk.db 정상 응답", "level": "core"})
    except Exception as e:
        checks.append({"name": "DB 연결", "status": "error",
                       "detail": f"치명: {e}", "level": "core"})
        return checks  # DB 안 되면 다른 검증 의미 없음

    # 2. 테이블 존재 확인
    expected_tables = [
        "users", "teams", "tasks", "projects", "customers",
        "changes", "change_impacts", "change_reads",
        "tickets", "ticket_comments",
        "issues", "issue_logs",
        "parts", "suppliers", "purchase_orders", "po_items",
        "stock_movements", "exchange_rates", "part_prices",
        "boards", "board_posts",
        "app_settings",
    ]
    with db_session() as c:
        existing = {r["name"] for r in c.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        ).fetchall()}
        missing = [t for t in expected_tables if t not in existing]
    if missing:
        checks.append({"name": "DB 스키마",
                       "status": "error",
                       "detail": f"누락 테이블: {missing}",
                       "level": "core"})
    else:
        checks.append({"name": "DB 스키마",
                       "status": "ok",
                       "detail": f"{len(expected_tables)}개 테이블 모두 정상",
                       "level": "core"})

    # 3. 부서명 매칭 (IMPACT_RULES)
    with db_session() as c:
        db_teams = {r["name"] for r in c.execute("SELECT name FROM teams").fetchall()}
    impact_teams = set()
    for ct, biz_map in IMPACT_RULES.items():
        for biz, team_list in biz_map.items():
            if isinstance(team_list, list):
                impact_teams.update(team_list)
    missing_t = impact_teams - db_teams
    if missing_t:
        checks.append({"name": "변경 영향 부서 매핑",
                       "status": "error",
                       "detail": f"DB에 없는 팀명: {sorted(missing_t)} → 영향 부서 0건 발생",
                       "level": "core"})
    else:
        checks.append({"name": "변경 영향 부서 매핑",
                       "status": "ok",
                       "detail": f"{len(impact_teams)}개 팀명 모두 DB에 존재",
                       "level": "core"})

    # 4. 사용자 / 팀 / 프로젝트 데이터
    with db_session() as c:
        n_users = c.execute("SELECT COUNT(*) FROM users WHERE is_active=1").fetchone()[0]
        n_teams = c.execute("SELECT COUNT(*) FROM teams").fetchone()[0]
        n_projects = c.execute("SELECT COUNT(*) FROM projects").fetchone()[0]
        n_parts = c.execute("SELECT COUNT(*) FROM parts WHERE is_active=1").fetchone()[0]
    if n_users < 2:
        checks.append({"name": "사용자 데이터",
                       "status": "warn",
                       "detail": f"활성 사용자 {n_users}명 (운영 시작 전이거나 비밀번호 미배포)",
                       "level": "data"})
    else:
        checks.append({"name": "사용자 데이터",
                       "status": "ok",
                       "detail": f"활성 사용자 {n_users}명, 팀 {n_teams}개, 프로젝트 {n_projects}건, 부품 {n_parts}건",
                       "level": "data"})

    # 5. 알림 채널 마스터 스위치
    # v5H226z114 (2026-05-31): groupware_* 우선, hiworks_* fallback
    notify_ch = get_setting("notify_channel", "off")
    msg_token = (get_setting("groupware_messenger_token", "") or
                 get_setting("hiworks_messenger_token", ""))
    if notify_ch == "off":
        checks.append({"name": "외부 푸시 알림",
                       "status": "info",
                       "detail": "OFF — 사이트 내 알림 / 게시판은 정상. 외부 메신저 푸시 없음.",
                       "level": "external"})
    elif notify_ch in ("groupware", "hiworks"):  # 구 채널명 호환
        if msg_token:
            checks.append({"name": "외부 푸시 알림 (그룹웨어)",
                           "status": "warn",
                           "detail": "그룹웨어 토큰 설정됨. ⚠️ 실제 엔드포인트 동작은 첫 호출 시 확인 필요 (Postman 검증 권장).",
                           "level": "external"})
        else:
            checks.append({"name": "외부 푸시 알림 (그룹웨어)",
                           "status": "error",
                           "detail": "🚨 채널은 groupware지만 토큰 비어있음 → 발송 시 silent skip (실제 발송 0)",
                           "level": "external"})
    elif notify_ch == "smtp":
        checks.append({"name": "외부 푸시 알림 (SMTP)",
                       "status": "warn",
                       "detail": "SMTP 채널은 미구현 (Phase 2). 콘솔 로그만 발생.",
                       "level": "external"})

    # 6. 외부 그룹웨어 링크 — groupware_* 우선, hiworks_* fallback
    apv_url  = (get_setting("groupware_approval_url", "") or get_setting("hiworks_approval_url", ""))
    mail_url = (get_setting("groupware_mail_url",     "") or get_setting("hiworks_mail_url",     ""))
    domain   = (get_setting("groupware_domain",       "") or get_setting("hiworks_domain",       ""))
    if not domain:
        checks.append({"name": "외부 그룹웨어 링크",
                       "status": "warn",
                       "detail": f"기본 URL 사용 중 (회사 도메인 미설정). 결재: {apv_url}",
                       "level": "external"})
    else:
        checks.append({"name": "외부 그룹웨어 링크",
                       "status": "ok",
                       "detail": f"회사 도메인 {domain} 설정됨",
                       "level": "external"})

    # 7. 백업
    import os as _os
    backup_dir = _os.path.join(
        _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))),
        "data", "backups"
    )
    if _os.path.exists(backup_dir):
        backups = [f for f in _os.listdir(backup_dir) if f.endswith(".db")]
        if backups:
            backups.sort(reverse=True)
            last = backups[0]
            try:
                ts = last.replace("knk_", "").replace(".db", "")
                from datetime import datetime as _dtdt
                lt = _dtdt.strptime(ts, "%Y%m%d_%H%M%S")
                age_days = (_dtdt.now() - lt).days
                if age_days > 7:
                    status = "warn"
                    detail = f"마지막 백업 {age_days}일 전 ({last}) — 1주일 초과"
                else:
                    status = "ok"
                    detail = f"{len(backups)}개 백업, 최신 {age_days}일 전 ({last})"
            except Exception:
                status, detail = "ok", f"{len(backups)}개 백업 파일 존재"
            checks.append({"name": "DB 백업", "status": status,
                           "detail": detail, "level": "ops"})
        else:
            checks.append({"name": "DB 백업", "status": "warn",
                           "detail": "백업 디렉토리는 있지만 파일 없음. scripts/backup_db.py 실행 권장",
                           "level": "ops"})
    else:
        checks.append({"name": "DB 백업", "status": "warn",
                       "detail": "data/backups/ 미존재. scripts/backup_db.py 실행 권장",
                       "level": "ops"})

    # 8. 시드/실 데이터 진척도
    with db_session() as c:
        recent_changes = c.execute(
            "SELECT COUNT(*) FROM changes WHERE created_at >= date('now','-7 day')"
        ).fetchone()[0]
        recent_tickets = c.execute(
            "SELECT COUNT(*) FROM tickets WHERE created_at >= date('now','-7 day')"
        ).fetchone()[0]
    if recent_changes == 0 and recent_tickets == 0:
        checks.append({"name": "최근 7일 활동",
                       "status": "info",
                       "detail": "변경/티켓 활동 없음 (운영 시작 전 정상)",
                       "level": "ops"})
    else:
        checks.append({"name": "최근 7일 활동",
                       "status": "ok",
                       "detail": f"변경 {recent_changes}건, 티켓 {recent_tickets}건",
                       "level": "ops"})

    return checks


def supplier_leadtime_stats(supplier_id: int) -> dict:
    """공급사별 PO 리드타임 통계: order_date → 최초 입고일 평균
    IN 수불이 있는 PO만 계산. 없으면 None.
    """
    with db_session() as c:
        rows = c.execute(
            """SELECT po.id, po.order_date,
                      MIN(sm.occurred_at) AS first_receive
               FROM purchase_orders po
               JOIN stock_movements sm ON sm.po_id = po.id AND sm.kind = 'IN'
               WHERE po.supplier_id = ?
                 AND po.order_date IS NOT NULL
               GROUP BY po.id""",
            (supplier_id,),
        ).fetchall()
        diffs = []
        for r in rows:
            try:
                od = _date.fromisoformat(r["order_date"][:10])
                rd = _date.fromisoformat(r["first_receive"][:10])
                days = (rd - od).days
                if 0 <= days <= 365:
                    diffs.append(days)
            except Exception:
                continue
        if not diffs:
            return {"count": 0, "avg_days": None, "min_days": None,
                    "max_days": None, "recent": []}
        diffs.sort()
        return {
            "count": len(diffs),
            "avg_days": round(sum(diffs) / len(diffs), 1),
            "min_days": min(diffs),
            "max_days": max(diffs),
            "median_days": diffs[len(diffs) // 2],
            "recent": diffs[-5:],  # 최근 5건
        }


def part_price_history(part_id: int, limit: int = 30) -> list[dict]:
    """부품별 공급사 단가 이력 (Abram/OpenBOM 모델 §12.4 #9)
    같은 부품에 대한 과거 발주 라인의 단가 추이 조회."""
    with db_session() as c:
        rows = c.execute(
            """SELECT pi.unit_price, pi.quantity, pi.amount, po.order_date,
                      po.po_number, po.currency, s.name AS supplier_name,
                      s.id AS supplier_id
               FROM po_items pi
               JOIN purchase_orders po ON pi.po_id = po.id
               LEFT JOIN suppliers s ON po.supplier_id = s.id
               WHERE pi.part_id=? AND po.status != '취소'
               ORDER BY po.order_date DESC, po.id DESC
               LIMIT ?""",
            (part_id, limit),
        ).fetchall()
        items = [dict(r) for r in rows]

        # 공급사별 최저·최고·최신 단가 집계
        suppliers = {}
        for r in items:
            sid = r.get("supplier_id")
            if not sid:
                continue
            if sid not in suppliers:
                suppliers[sid] = {
                    "supplier_id": sid,
                    "supplier_name": r.get("supplier_name"),
                    "min_price": r["unit_price"],
                    "max_price": r["unit_price"],
                    "latest_price": r["unit_price"],
                    "latest_date": r.get("order_date"),
                    "count": 0,
                }
            s = suppliers[sid]
            s["min_price"] = min(s["min_price"], r["unit_price"] or 0)
            s["max_price"] = max(s["max_price"], r["unit_price"] or 0)
            if (r.get("order_date") or "") > (s.get("latest_date") or ""):
                s["latest_price"] = r["unit_price"]
                s["latest_date"] = r["order_date"]
            s["count"] += 1
        return {"history": items, "by_supplier": list(suppliers.values())}


# =====================================================
# 게시판 (boards / board_posts / board_comments)
# =====================================================

BOARD_CATEGORIES = ["공지", "일반", "자료", "질문"]


def board_get_or_create_company():
    """전사 게시판 ID 반환 (없으면 생성)"""
    with db_session() as c:
        row = c.execute("SELECT id FROM boards WHERE type='company' LIMIT 1").fetchone()
        if row:
            return row["id"]
        c.execute("INSERT INTO boards (name,type) VALUES ('전사 게시판','company')")
        return c.execute("SELECT last_insert_rowid()").fetchone()[0]


def board_get_or_create_team(team_id):
    """부서 게시판 ID 반환 (없으면 생성)"""
    with db_session() as c:
        row = c.execute("SELECT id FROM boards WHERE type='team' AND team_id=?", (team_id,)).fetchone()
        if row:
            return row["id"]
        t = c.execute("SELECT name FROM teams WHERE id=?", (team_id,)).fetchone()
        name = f"{t['name']} 게시판" if t else f"팀{team_id} 게시판"
        c.execute("INSERT INTO boards (name,type,team_id) VALUES (?,'team',?)", (name, team_id))
        return c.execute("SELECT last_insert_rowid()").fetchone()[0]


def board_list_all():
    """전체 게시판 목록"""
    with db_session() as c:
        return c.execute(
            "SELECT b.*, t.name AS team_name FROM boards b LEFT JOIN teams t ON b.team_id=t.id ORDER BY b.type, b.id"
        ).fetchall()


def board_posts_list(board_id, include_pending=False):
    """게시글 목록. include_pending=True면 대기 글도 포함"""
    sql = """SELECT p.*, u.name AS author_name, u.rank AS author_rank,
                    t.name AS author_team,
                    (SELECT COUNT(*) FROM board_comments WHERE post_id=p.id) AS comment_count
             FROM board_posts p
             JOIN users u ON p.author_id=u.id
             LEFT JOIN teams t ON u.team_id=t.id
             WHERE p.board_id=?"""
    if not include_pending:
        sql += " AND p.approval_status='approved'"
    sql += " ORDER BY p.is_pinned DESC, p.created_at DESC"
    with db_session() as c:
        return c.execute(sql, (board_id,)).fetchall()


def board_posts_pending(board_id):
    """승인 대기 글 (팀장용)"""
    sql = """SELECT p.*, u.name AS author_name, u.rank AS author_rank
             FROM board_posts p JOIN users u ON p.author_id=u.id
             WHERE p.board_id=? AND p.approval_status='pending'
             ORDER BY p.created_at DESC"""
    with db_session() as c:
        return c.execute(sql, (board_id,)).fetchall()


def board_post_get(post_id):
    with db_session() as c:
        row = c.execute(
            """SELECT p.*, u.name AS author_name, u.rank AS author_rank,
                      t.name AS author_team, b.type AS board_type, b.team_id AS board_team_id,
                      b.name AS board_name
               FROM board_posts p
               JOIN users u ON p.author_id=u.id
               LEFT JOIN teams t ON u.team_id=t.id
               JOIN boards b ON p.board_id=b.id
               WHERE p.id=?""",
            (post_id,),
        ).fetchone()
        return dict(row) if row else None


def board_post_create(board_id, author_id, title, body, category="일반",
                      approval_status="approved"):
    with db_session() as c:
        c.execute(
            """INSERT INTO board_posts (board_id,author_id,title,body,category,approval_status)
               VALUES (?,?,?,?,?,?)""",
            (board_id, author_id, title.strip(), body.strip(), category, approval_status),
        )
        return c.execute("SELECT last_insert_rowid()").fetchone()[0]


def board_post_update(post_id, title, body, category):
    with db_session() as c:
        c.execute(
            "UPDATE board_posts SET title=?, body=?, category=?, updated_at=datetime('now','localtime') WHERE id=?",
            (title.strip(), body.strip(), category, post_id),
        )


def board_post_delete(post_id):
    """v5H119: 공통 헬퍼 사용. 폴백: v5H115 인라인."""
    with db_session() as c:
        try:
            res = _safe_delete_with_cascade(
                c, "board_posts", post_id, fk_column="post_id",
                explicit_children=[
                    ("DELETE FROM board_comments WHERE post_id=?", (post_id,)),
                ],
                keep_tables=("board_comments",),
            )
            if res.get("ok"):
                return
        except Exception:
            pass
        try:
            c.execute("DELETE FROM board_comments WHERE post_id=?", (post_id,))
        except Exception:
            pass
        try:
            tables = [r[0] for r in c.execute(
                "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
            for tbl in tables:
                if tbl in ("board_posts", "board_comments", "sqlite_sequence"):
                    continue
                try:
                    cols = [r[1] for r in c.execute(f"PRAGMA table_info({tbl})").fetchall()]
                    if "post_id" in cols:
                        try:
                            c.execute(f"UPDATE {tbl} SET post_id=NULL WHERE post_id=?", (post_id,))
                        except Exception:
                            try:
                                c.execute(f"DELETE FROM {tbl} WHERE post_id=?", (post_id,))
                            except Exception:
                                pass
                except Exception:
                    pass
        except Exception:
            pass
        c.execute("DELETE FROM board_posts WHERE id=?", (post_id,))


def board_post_approve(post_id, approver_id):
    with db_session() as c:
        c.execute(
            "UPDATE board_posts SET approval_status='approved', approved_by=?, approved_at=datetime('now','localtime') WHERE id=?",
            (approver_id, post_id),
        )


def board_post_reject(post_id, approver_id, reason=""):
    with db_session() as c:
        c.execute(
            "UPDATE board_posts SET approval_status='rejected', approved_by=?, approved_at=datetime('now','localtime'), reject_reason=? WHERE id=?",
            (approver_id, reason, post_id),
        )


def board_post_toggle_pin(post_id):
    with db_session() as c:
        cur = c.execute("SELECT is_pinned FROM board_posts WHERE id=?", (post_id,)).fetchone()
        if cur:
            c.execute("UPDATE board_posts SET is_pinned=? WHERE id=?",
                      (0 if cur["is_pinned"] else 1, post_id))


def board_post_increment_view(post_id):
    with db_session() as c:
        c.execute("UPDATE board_posts SET view_count=view_count+1 WHERE id=?", (post_id,))


def board_comments_list(post_id):
    with db_session() as c:
        return c.execute(
            """SELECT c.*, u.name AS author_name, u.rank AS author_rank
               FROM board_comments c JOIN users u ON c.author_id=u.id
               WHERE c.post_id=? ORDER BY c.created_at""",
            (post_id,),
        ).fetchall()


def board_comment_create(post_id, author_id, body):
    with db_session() as c:
        c.execute(
            "INSERT INTO board_comments (post_id,author_id,body) VALUES (?,?,?)",
            (post_id, author_id, body.strip()),
        )


def board_comment_delete(comment_id):
    with db_session() as c:
        c.execute("DELETE FROM board_comments WHERE id=?", (comment_id,))


# =====================================================
# 변경 Inform 시스템 — 사전 준비 (2026-04-20)
# 정식 구현은 Research 세션 v2 문서 도착 후
# 설계 청사진: HAIST_WORKS/_DESIGN_변경_Inform.md
# =====================================================

CHANGE_TYPES = ["기구설계", "전장설계", "소프트웨어", "BOM", "도면", "Concept", "사양"]
CHANGE_URGENCIES = ["일반", "긴급", "예약"]
CHANGE_STATUSES = ["작성중", "공지중", "확인완료", "취소"]

# =====================================================
# ISSUES · AS DB — 분류/라우팅 상수 (3순위 ⑦)
# =====================================================
ISSUE_SEVERITIES = ["치명", "심각", "중", "경"]
ISSUE_TYPES      = ["AS", "품질", "설계결함", "SW버그", "기구결함", "전장결함", "기타"]
ISSUE_STATUSES   = ["접수", "원인분석", "조치중", "해결", "재발방지등록", "종결"]

# 이슈 종류 × 사업부 → 기본 책임 부서 (자동 라우팅)
ISSUE_OWNER_RULES = {
    "AS":      {"T": "제조기술1팀", "M": "제조기술2팀"},
    "품질":     {"T": "품질팀",      "M": "품질팀"},
    "설계결함": {"T": "설계팀",      "M": "설계팀"},
    "SW버그":   {"T": "소프트웨어팀", "M": "소프트웨어팀"},
    "기구결함": {"T": "설계팀",      "M": "설계팀"},
    "전장결함": {"T": "전장설계팀",   "M": "전장설계팀"},
    "기타":     {"T": "기술영업팀",   "M": "기술영업팀"},
}

# Abram Scientific 모델 — 외부 도구 출처 (향후 자동 import 대비)
CHANGE_SOURCES = [
    "수동",            # 직원 직접 등록 (현재 기본)
    # KNK 실제 사용 CAD (2026-04 기준): AutoCAD(2D) + SolidWorks(3D) + Inventor(3D), 전장은 AutoCAD
    "AutoCAD",         # 2D 도면 변경 (기구·전장 공용)
    "SolidWorks",      # 3D 모델 변경
    "SolidWorks PDM",  # Check-in 이벤트 (도입 시 자동)
    "Inventor",        # 3D 모델 변경 (AnyCAD로 SW/CATIA/NX/Creo 호환)
    "Altium 365",      # Push to MCAD/ECAD 이벤트 (향후)
    "Git",             # 커밋·태그
    "OpenBOM",         # BOM 변경 (Phase 2)
    "메신저",          # 사내 메신저 메시지 변환
    "메일",
    "기타",
]

# 변경 종류 × 사업부 → 영향 부서 자동 판별 규칙
IMPACT_RULES = {
    "기구설계": {
        "T": ["전장설계팀", "소프트웨어팀", "가공팀", "제조기술1팀", "구매팀"],
        "M": ["전장설계팀", "소프트웨어팀", "가공팀", "제조기술2팀", "구매팀"],
    },
    "전장설계": {
        "T": ["소프트웨어팀", "제조기술1팀", "구매팀"],
        "M": ["소프트웨어팀", "제조기술2팀", "구매팀"],
    },
    "소프트웨어": {
        "T": ["검사기팀", "제조기술1팀"],
        "M": ["제조기술2팀"],
    },
    "BOM": {
        "T": ["구매팀", "제조기술1팀", "가공팀"],
        "M": ["구매팀", "제조기술2팀", "가공팀"],
    },
    "도면": {
        "T": ["가공팀", "제조기술1팀", "품질팀"],
        "M": ["가공팀", "제조기술2팀", "품질팀"],
    },
    "Concept": {"T": "ALL", "M": "ALL"},
    "사양": {
        "T": ["기술영업팀", "검사기팀", "설계팀", "품질팀"],
        "M": ["기술영업팀", "설계팀", "품질팀"],
    },
}


def detect_impact_teams(change_type: str, biz_div: str) -> list[int]:
    """변경 종류 + 사업부 → 영향 받는 팀 ID 리스트.
    biz_div 'T'(검사기) / 'M'(자동화). Concept는 전 팀.
    """
    rule = IMPACT_RULES.get(change_type, {}).get(biz_div, [])
    if not rule:
        return []
    if rule == "ALL":
        with db_session() as c:
            return [r["id"] for r in c.execute("SELECT id FROM teams").fetchall()]
    placeholders = ",".join(["?"] * len(rule))
    with db_session() as c:
        rows = c.execute(
            f"SELECT id FROM teams WHERE name IN ({placeholders})",
            tuple(rule),
        ).fetchall()
    return [r["id"] for r in rows]


def detect_impact_users(team_ids: list[int]) -> list[dict]:
    """영향 팀들의 활성 사용자 목록 (알림 발송 대상)"""
    if not team_ids:
        return []
    placeholders = ",".join(["?"] * len(team_ids))
    with db_session() as c:
        rows = c.execute(
            f"""SELECT u.id, u.name, u.team_id, t.name AS team_name
                FROM users u JOIN teams t ON u.team_id = t.id
                WHERE u.team_id IN ({placeholders}) AND u.is_active = 1""",
            tuple(team_ids),
        ).fetchall()
    return [dict(r) for r in rows]


def gen_change_no(today=None) -> str:
    """변경번호 자동 채번: CHG-YYMMDD-NNN (해당 날짜 +1)"""
    today = today or _date.today()
    yymmdd = today.strftime("%y%m%d")
    prefix = f"CHG-{yymmdd}-"
    # NOTE: changes 테이블이 아직 없으면 001부터 시작
    try:
        with db_session() as c:
            rows = c.execute(
                "SELECT change_no FROM changes WHERE change_no LIKE ?",
                (f"{prefix}%",),
            ).fetchall()
        max_seq = 0
        for r in rows:
            try:
                seq = int(r["change_no"].rsplit("-", 1)[-1])
                max_seq = max(max_seq, seq)
            except (ValueError, IndexError):
                pass
        return f"{prefix}{max_seq + 1:03d}"
    except sqlite3.OperationalError:
        # changes 테이블 미존재 시
        return f"{prefix}001"


def gen_issue_no(today=None) -> str:
    """이슈번호 자동 채번: ISS-YYMM-NNN (월별 누적)"""
    today = today or _date.today()
    yymm = today.strftime("%y%m")
    prefix = f"ISS-{yymm}-"
    try:
        with db_session() as c:
            rows = c.execute(
                "SELECT issue_no FROM issues WHERE issue_no LIKE ?",
                (f"{prefix}%",),
            ).fetchall()
        max_seq = 0
        for r in rows:
            try:
                seq = int(r["issue_no"].rsplit("-", 1)[-1])
                max_seq = max(max_seq, seq)
            except (ValueError, IndexError):
                pass
        return f"{prefix}{max_seq + 1:03d}"
    except sqlite3.OperationalError:
        return f"{prefix}001"


# =====================================================
# APP SETTINGS — 헬퍼 (key-value 저장)
# =====================================================
def get_setting(key: str, default: str = "") -> str:
    """app_settings에서 단일 값 조회. 없으면 default 반환."""
    try:
        with db_session() as c:
            row = c.execute("SELECT value FROM app_settings WHERE key=?", (key,)).fetchone()
            if row and row["value"] is not None:
                return row["value"]
    except sqlite3.OperationalError:
        pass
    return default


def get_settings_all() -> dict:
    """모든 설정을 dict로 반환 (admin 페이지용)."""
    out = {}
    try:
        with db_session() as c:
            for r in c.execute("SELECT key, value, description FROM app_settings ORDER BY key").fetchall():
                out[r["key"]] = {"value": r["value"] or "", "description": r["description"] or ""}
    except sqlite3.OperationalError:
        pass
    return out


def set_setting(key: str, value: str, user_id: int = None, description: str = None):
    """설정 값 갱신. 없으면 생성."""
    with db_session() as c:
        existing = c.execute("SELECT key FROM app_settings WHERE key=?", (key,)).fetchone()
        if existing:
            if description is not None:
                c.execute(
                    "UPDATE app_settings SET value=?, description=?, updated_at=datetime('now','localtime'), updated_by=? WHERE key=?",
                    (value, description, user_id, key),
                )
            else:
                c.execute(
                    "UPDATE app_settings SET value=?, updated_at=datetime('now','localtime'), updated_by=? WHERE key=?",
                    (value, user_id, key),
                )
        else:
            c.execute(
                "INSERT INTO app_settings(key, value, description, updated_by) VALUES(?,?,?,?)",
                (key, value, description or "", user_id),
            )


# =====================================================
# 통합 푸시 알림 — 전자결재·메일 연동 채널 (2026-04-22 대표 결재: 외부 웹훅 폐기)
# =====================================================
def groupware_notify(channel_id: str, text: str) -> bool:
    """사내 메신저 알림 발송 (외부 그룹웨어 API).
    v5H226z110 (2026-05-30): hiworks_notify → groupware_notify 리네임

    notify_channel 설정에 따라:
      - 'hiworks' (기본): 외부 그룹웨어 메신저 API
      - 'off':            개발 모드 (콘솔 로그만)
      - 'smtp':           Phase 2 SMTP 메일 발송 (TBD)
    """
    channel = get_setting("notify_channel", "off").lower()
    if channel == "off":
        print(f"[NOTIFY OFF] channel={channel_id} text={text[:80]}")
        return True
    # v5H226z110: 'hiworks' 채널명 → 'groupware' (구 채널명 호환 fallback)
    if channel in ("groupware", "hiworks"):
        try:
            from .groupware_client import notify as gw_notify
            return gw_notify(text, recipients=channel_id)
        except Exception as e:
            print(f"[그룹웨어 알림 ERR] {e}")
            return False
    print(f"[NOTIFY UNKNOWN channel={channel}] {text[:80]}")
    return False


# v5H226z110: 구 함수명 alias (기존 import 호환, 추후 deprecate)
hiworks_notify = groupware_notify


# ── 변경 Inform CRUD ─────────────────────────────────────────
def changes_list(q="", change_type="", urgency="", status="", scope_user_id=None):
    """변경 목록 조회. scope_user_id 주면 그 사용자가 영향자인 것만"""
    base = """SELECT c.*, u.name AS author_name, u.rank AS author_rank,
                     t.name AS author_team,
                     (SELECT COUNT(*) FROM change_impacts WHERE change_id=c.id) AS impact_count,
                     (SELECT COUNT(*) FROM change_reads WHERE change_id=c.id AND ack_at IS NOT NULL) AS ack_count
              FROM changes c
              JOIN users u ON c.author_id = u.id
              LEFT JOIN teams t ON u.team_id = t.id
              WHERE 1=1"""
    params = []
    if q:
        base += " AND (c.change_no LIKE ? OR c.title LIKE ? OR c.description LIKE ? OR c.target_label LIKE ?)"
        like = f"%{q}%"
        params += [like] * 4
    if change_type:
        base += " AND c.change_type = ?"
        params.append(change_type)
    if urgency:
        base += " AND c.urgency = ?"
        params.append(urgency)
    if status:
        base += " AND c.status = ?"
        params.append(status)
    if scope_user_id:
        base += """ AND c.id IN (
            SELECT change_id FROM change_impacts ci
            LEFT JOIN users u2 ON u2.team_id = ci.impact_team_id
            WHERE u2.id = ? OR ci.impact_user_id = ?
        )"""
        params += [scope_user_id, scope_user_id]
    base += " ORDER BY c.created_at DESC"
    with db_session() as c:
        return c.execute(base, params).fetchall()


def change_get(cid: int) -> dict | None:
    with db_session() as c:
        row = c.execute(
            """SELECT c.*, u.name AS author_name, u.rank AS author_rank,
                      t.name AS author_team
               FROM changes c
               JOIN users u ON c.author_id = u.id
               LEFT JOIN teams t ON u.team_id = t.id
               WHERE c.id = ?""",
            (cid,),
        ).fetchone()
        return dict(row) if row else None


def change_create(data: dict, author_id: int) -> tuple[int, str]:
    """변경 등록 + 영향 자동 판별 + change_reads 미리 생성"""
    change_no = gen_change_no()
    biz_div = (data.get("biz_div") or "").strip()
    change_type = (data.get("change_type") or "").strip()

    with db_session() as c:
        cur = c.execute(
            """INSERT INTO changes
               (change_no, change_type, biz_div, target_kind, target_id, target_label,
                project_id, title, description, before_value, after_value,
                attached_files, urgency, author_id, status, source, source_ref, approval_url)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                change_no, change_type, biz_div,
                data.get("target_kind"), data.get("target_id"),
                data.get("target_label", "").strip(),
                data.get("project_id"),
                data.get("title", "").strip(),
                data.get("description", "").strip(),
                data.get("before_value", "").strip(),
                data.get("after_value", "").strip(),
                data.get("attached_files", ""),
                data.get("urgency", "일반"),
                author_id,
                "공지중",
                data.get("source", "수동"),
                data.get("source_ref", "").strip() if data.get("source_ref") else None,
                data.get("approval_url"),
            ),
        )
        change_id = c.execute("SELECT last_insert_rowid()").fetchone()[0]

        # 영향 부서 자동 판별
        impact_teams = detect_impact_teams(change_type, biz_div) if biz_div in ("T", "M") else []
        for team_id in impact_teams:
            c.execute(
                """INSERT INTO change_impacts
                   (change_id, impact_kind, impact_team_id, auto_detected, impact_reason)
                   VALUES (?, 'team', ?, 1, ?)""",
                (change_id, team_id, f"{change_type} 변경 → 자동 판별"),
            )

        # change_reads 미리 생성 (영향 부서원 전원)
        if impact_teams:
            placeholders = ",".join(["?"] * len(impact_teams))
            users = c.execute(
                f"SELECT id FROM users WHERE team_id IN ({placeholders}) AND is_active = 1",
                tuple(impact_teams),
            ).fetchall()
            for u in users:
                try:
                    c.execute(
                        "INSERT INTO change_reads (change_id, user_id) VALUES (?, ?)",
                        (change_id, u["id"]),
                    )
                except sqlite3.IntegrityError:
                    pass

        # 작성자에게도 read 행 (보낸 사람도 본인이 다 봤다는 의미)
        try:
            c.execute(
                "INSERT INTO change_reads (change_id, user_id, read_at, ack_at) VALUES (?, ?, ?, ?)",
                (change_id, author_id, _logi_now(), _logi_now()),
            )
        except sqlite3.IntegrityError:
            pass

        # notified_at 표시
        c.execute("UPDATE changes SET notified_at=? WHERE id=?", (_logi_now(), change_id))

    return change_id, change_no


def change_get_impacts(cid: int) -> list[dict]:
    """영향 부서/사용자 + 부서별 확인 현황"""
    with db_session() as c:
        impacts = c.execute(
            """SELECT ci.*, t.name AS team_name
               FROM change_impacts ci
               LEFT JOIN teams t ON ci.impact_team_id = t.id
               WHERE ci.change_id = ?""",
            (cid,),
        ).fetchall()
        result = []
        for imp in impacts:
            d = dict(imp)
            if imp["impact_team_id"]:
                # 부서원 확인 현황
                stats = c.execute(
                    """SELECT COUNT(*) AS total,
                              SUM(CASE WHEN cr.ack_at IS NOT NULL THEN 1 ELSE 0 END) AS ack
                       FROM users u
                       LEFT JOIN change_reads cr ON cr.user_id = u.id AND cr.change_id = ?
                       WHERE u.team_id = ? AND u.is_active = 1""",
                    (cid, imp["impact_team_id"]),
                ).fetchone()
                d["total_users"] = stats["total"] or 0
                d["ack_users"] = stats["ack"] or 0
            result.append(d)
    return result


def change_get_reads(cid: int) -> list[dict]:
    with db_session() as c:
        rows = c.execute(
            """SELECT cr.*, u.name AS user_name, u.rank AS user_rank, t.name AS team_name
               FROM change_reads cr
               JOIN users u ON cr.user_id = u.id
               LEFT JOIN teams t ON u.team_id = t.id
               WHERE cr.change_id = ?
               ORDER BY cr.ack_at DESC NULLS LAST, cr.read_at DESC""",
            (cid,),
        ).fetchall()
    return [dict(r) for r in rows]


def change_mark_read(cid: int, user_id: int):
    """페이지 열람 기록"""
    with db_session() as c:
        c.execute(
            """INSERT INTO change_reads (change_id, user_id, read_at)
               VALUES (?, ?, ?)
               ON CONFLICT(change_id, user_id) DO UPDATE SET read_at = COALESCE(read_at, excluded.read_at)""",
            (cid, user_id, _logi_now()),
        )


def change_ack(cid: int, user_id: int, note: str = ""):
    """확인 응답"""
    with db_session() as c:
        c.execute(
            """INSERT INTO change_reads (change_id, user_id, read_at, ack_at, ack_note)
               VALUES (?, ?, ?, ?, ?)
               ON CONFLICT(change_id, user_id) DO UPDATE SET
                   read_at = COALESCE(read_at, excluded.read_at),
                   ack_at = excluded.ack_at,
                   ack_note = excluded.ack_note""",
            (cid, user_id, _logi_now(), _logi_now(), note),
        )
        # 모든 영향자 확인 시 status=확인완료
        remaining = c.execute(
            """SELECT COUNT(*) FROM change_reads
               WHERE change_id = ? AND ack_at IS NULL""",
            (cid,),
        ).fetchone()[0]
        if remaining == 0:
            c.execute(
                "UPDATE changes SET status='확인완료', completed_at=? WHERE id=?",
                (_logi_now(), cid),
            )


def change_delete(cid: int):
    """v5H119: 공통 헬퍼 사용. 폴백: v5H115 인라인."""
    with db_session() as c:
        try:
            res = _safe_delete_with_cascade(
                c, "changes", cid, fk_column="change_id",
                explicit_children=[
                    ("DELETE FROM change_impacts WHERE change_id=?", (cid,)),
                    ("DELETE FROM change_reads WHERE change_id=?", (cid,)),
                ],
                keep_tables=("change_impacts", "change_reads"),
            )
            if res.get("ok"):
                return
        except Exception:
            pass
        for tbl in ("change_impacts", "change_reads"):
            try:
                c.execute(f"DELETE FROM {tbl} WHERE change_id=?", (cid,))
            except Exception:
                pass
        try:
            tables = [r[0] for r in c.execute(
                "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
            for tbl in tables:
                if tbl in ("changes", "change_impacts", "change_reads",
                            "sqlite_sequence"):
                    continue
                try:
                    cols = [r[1] for r in c.execute(f"PRAGMA table_info({tbl})").fetchall()]
                    if "change_id" in cols:
                        try:
                            c.execute(f"UPDATE {tbl} SET change_id=NULL WHERE change_id=?", (cid,))
                        except Exception:
                            try:
                                c.execute(f"DELETE FROM {tbl} WHERE change_id=?", (cid,))
                            except Exception:
                                pass
                except Exception:
                    pass
        except Exception:
            pass
        c.execute("DELETE FROM changes WHERE id=?", (cid,))


def change_unread_count(user_id: int) -> int:
    """내가 영향자인데 아직 ack 안 한 변경 카운트 (사이드바 뱃지)"""
    with db_session() as c:
        row = c.execute(
            """SELECT COUNT(*) FROM change_reads
               WHERE user_id = ? AND ack_at IS NULL""",
            (user_id,),
        ).fetchone()
    return row[0] if row else 0


# === 영향 강도 분류 (알림 피로 방지) ===
# high: 직접 작업 영향 → 즉시 알림 (web + 메신저)
# medium: 일정 협의 필요 → web 알림만
# low: 참고용 → daily digest로 묶기
IMPACT_INTENSITY = {
    # 변경 종류별 직접 영향(high) 부서 — 변경 작업이 직접 부서 일정·산출물에 영향
    ("기구설계", "T"): {"high": ["전장설계팀", "가공팀"], "medium": ["소프트웨어팀", "제조기술1팀"], "low": ["구매팀"]},
    ("기구설계", "M"): {"high": ["전장설계팀", "가공팀"], "medium": ["소프트웨어팀", "제조기술2팀"], "low": ["구매팀"]},
    ("전장설계", "T"): {"high": ["소프트웨어팀"], "medium": ["제조기술1팀"], "low": ["구매팀"]},
    ("전장설계", "M"): {"high": ["소프트웨어팀"], "medium": ["제조기술2팀"], "low": ["구매팀"]},
    ("소프트웨어", "T"): {"high": ["검사기팀"], "medium": ["제조기술1팀"], "low": []},
    ("소프트웨어", "M"): {"high": ["제조기술2팀"], "medium": [], "low": []},
    ("BOM", "T"): {"high": ["구매팀"], "medium": ["제조기술1팀", "가공팀"], "low": []},
    ("BOM", "M"): {"high": ["구매팀"], "medium": ["제조기술2팀", "가공팀"], "low": []},
    ("도면", "T"): {"high": ["가공팀", "제조기술1팀"], "medium": ["품질팀"], "low": []},
    ("도면", "M"): {"high": ["가공팀", "제조기술2팀"], "medium": ["품질팀"], "low": []},
    ("Concept", "T"): {"high": [], "medium": ["기술영업팀", "검사기팀"], "low": "ALL_OTHERS"},
    ("Concept", "M"): {"high": [], "medium": ["기술영업팀"], "low": "ALL_OTHERS"},
    ("사양", "T"): {"high": ["기술영업팀", "검사기팀"], "medium": ["설계팀", "품질팀"], "low": []},
    ("사양", "M"): {"high": ["기술영업팀"], "medium": ["설계팀", "품질팀"], "low": []},
}


def get_impact_intensity(change_type: str, biz_div: str, team_name: str) -> str:
    """변경 종류 + 사업부 + 팀명 → 영향 강도 (high/medium/low)"""
    rule = IMPACT_INTENSITY.get((change_type, biz_div), {})
    if team_name in rule.get("high", []):
        return "high"
    if team_name in rule.get("medium", []):
        return "medium"
    if rule.get("low") == "ALL_OTHERS" and team_name not in rule.get("high", []) + rule.get("medium", []):
        return "low"
    if team_name in rule.get("low", []):
        return "low"
    return "medium"  # default


def changes_user_unread_by_intensity(user_id: int, team_name: str = None) -> dict:
    """미확인 변경을 강도별로 분류 (high/medium/low)"""
    if not team_name:
        with db_session() as c:
            row = c.execute(
                """SELECT t.name FROM users u LEFT JOIN teams t ON u.team_id=t.id
                   WHERE u.id=?""", (user_id,)).fetchone()
            team_name = row["name"] if row and row["name"] else ""

    with db_session() as c:
        rows = c.execute(
            """SELECT c.change_type, c.biz_div FROM changes c
               JOIN change_reads cr ON cr.change_id = c.id
               WHERE cr.user_id = ? AND cr.ack_at IS NULL""",
            (user_id,),
        ).fetchall()
    result = {"high": 0, "medium": 0, "low": 0}
    for r in rows:
        intensity = get_impact_intensity(r["change_type"], r["biz_div"] or "", team_name)
        result[intensity] += 1
    return result


def change_recent_count(user_id: int = None, days: int = 1) -> int:
    """홈 KPI용 — 최근 N일 변경 (전체 또는 내 관련)"""
    sql = """SELECT COUNT(*) FROM changes
             WHERE created_at >= datetime('now', '-' || ? || ' day', 'localtime')"""
    params = [days]
    if user_id:
        sql += """ AND id IN (
            SELECT change_id FROM change_impacts ci
            LEFT JOIN users u ON u.team_id = ci.impact_team_id
            WHERE u.id = ? OR ci.impact_user_id = ?
        )"""
        params += [user_id, user_id]
    with db_session() as c:
        return c.execute(sql, params).fetchone()[0]


# =====================================================
# 요청 티켓 시스템 (HAIST WORKS)
# 설문 1순위 ③ / 10팀 메신저 요청 누락 해결
# =====================================================

TICKET_CATEGORIES = ["자재요청", "긴급가공", "MODIFY", "검수요청", "AS", "기타"]
TICKET_URGENCIES = ["일반", "긴급", "지연"]
TICKET_STATUSES = ["요청", "접수", "처리중", "완료", "반려", "지연"]
TICKET_SOURCES = ["web", "메신저", "메일", "전화"]

# 카테고리별 자동 라우팅 (수신 부서)
TICKET_ROUTING = {
    "자재요청": "구매팀",
    "긴급가공": "가공팀",
    "MODIFY": "설계팀",       # 또는 변경 발생 부서. 일단 기본
    "검수요청": "품질팀",
    "AS": None,               # 사업부에 따라 다름 → 라우팅 함수에서 결정
    "기타": None,
}


def gen_ticket_no(today=None) -> str:
    today = today or _date.today()
    yymmdd = today.strftime("%y%m%d")
    prefix = f"TKT-{yymmdd}-"
    try:
        with db_session() as c:
            rows = c.execute(
                "SELECT ticket_no FROM tickets WHERE ticket_no LIKE ?",
                (f"{prefix}%",),
            ).fetchall()
        max_seq = 0
        for r in rows:
            try:
                seq = int(r["ticket_no"].rsplit("-", 1)[-1])
                max_seq = max(max_seq, seq)
            except (ValueError, IndexError):
                pass
        return f"{prefix}{max_seq + 1:03d}"
    except sqlite3.OperationalError:
        return f"{prefix}001"


def route_ticket_team(category: str, biz_div: str = None) -> int | None:
    """카테고리 → 수신 부서 자동 결정"""
    team_name = TICKET_ROUTING.get(category)
    if category == "AS":
        # AS는 사업부에 따라
        team_name = "검사기팀" if biz_div == "T" else ("자동화팀" if biz_div == "M" else None)
    if not team_name:
        return None
    with db_session() as c:
        row = c.execute("SELECT id FROM teams WHERE name=?", (team_name,)).fetchone()
    return row["id"] if row else None


def tickets_list(scope_user_id=None, scope_team_id=None, status="", category="",
                 urgency="", q=""):
    """목록 조회. scope:
    - scope_user_id: 내가 요청한 + 내가 받은
    - scope_team_id: 특정 팀이 받은
    - 둘 다 없으면 전체"""
    base = """SELECT t.*,
                     u_req.name AS requester_name, u_req.rank AS requester_rank,
                     team_req.name AS requester_team,
                     team_recv.name AS recipient_team_name,
                     u_recv.name AS recipient_user_name,
                     (SELECT COUNT(*) FROM ticket_comments WHERE ticket_id=t.id) AS comment_count
              FROM tickets t
              JOIN users u_req ON t.requester_id = u_req.id
              LEFT JOIN teams team_req ON u_req.team_id = team_req.id
              LEFT JOIN teams team_recv ON t.recipient_team_id = team_recv.id
              LEFT JOIN users u_recv ON t.recipient_user_id = u_recv.id
              WHERE 1=1"""
    params = []
    if scope_user_id:
        base += " AND (t.requester_id = ? OR t.recipient_user_id = ?)"
        params += [scope_user_id, scope_user_id]
    if scope_team_id:
        base += " AND t.recipient_team_id = ?"
        params.append(scope_team_id)
    if status:
        base += " AND t.status = ?"
        params.append(status)
    if category:
        base += " AND t.category = ?"
        params.append(category)
    if urgency:
        base += " AND t.urgency = ?"
        params.append(urgency)
    if q:
        base += " AND (t.ticket_no LIKE ? OR t.title LIKE ? OR t.description LIKE ? OR t.target_label LIKE ?)"
        like = f"%{q}%"
        params += [like] * 4
    base += " ORDER BY CASE t.urgency WHEN '긴급' THEN 1 WHEN '지연' THEN 2 ELSE 3 END, t.created_at DESC"
    with db_session() as c:
        return c.execute(base, params).fetchall()


def ticket_get(tid: int) -> dict | None:
    with db_session() as c:
        row = c.execute(
            """SELECT t.*,
                      u_req.name AS requester_name, u_req.rank AS requester_rank,
                      team_req.name AS requester_team,
                      team_recv.name AS recipient_team_name,
                      u_recv.name AS recipient_user_name,
                      p.mgmt_code AS project_mgmt_code, p.name AS project_name
               FROM tickets t
               JOIN users u_req ON t.requester_id = u_req.id
               LEFT JOIN teams team_req ON u_req.team_id = team_req.id
               LEFT JOIN teams team_recv ON t.recipient_team_id = team_recv.id
               LEFT JOIN users u_recv ON t.recipient_user_id = u_recv.id
               LEFT JOIN projects p ON t.project_id = p.id
               WHERE t.id = ?""",
            (tid,),
        ).fetchone()
        return dict(row) if row else None


def ticket_create(data: dict, requester_id: int) -> tuple[int, str]:
    """티켓 생성 + 자동 라우팅"""
    ticket_no = gen_ticket_no()
    category = (data.get("category") or "기타").strip()
    biz_div = (data.get("biz_div") or "").strip()
    recipient_team_id = data.get("recipient_team_id") or route_ticket_team(category, biz_div)

    with db_session() as c:
        c.execute(
            """INSERT INTO tickets
               (ticket_no, category, title, description, requester_id,
                recipient_team_id, recipient_user_id, project_id, target_label,
                urgency, status, source, due_date, hours_estimated, approval_url)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                ticket_no, category,
                data.get("title", "").strip(),
                data.get("description", "").strip(),
                requester_id,
                recipient_team_id,
                data.get("recipient_user_id"),
                data.get("project_id"),
                data.get("target_label", "").strip(),
                data.get("urgency", "일반"),
                "요청",
                data.get("source", "web"),
                data.get("due_date", ""),
                float(data.get("hours_estimated") or 0) or None,
                (data.get("approval_url") or "").strip() or None,
            ),
        )
        tid = c.execute("SELECT last_insert_rowid()").fetchone()[0]
    return tid, ticket_no


def ticket_edit(tid: int, changes: dict, editor_id: int, editor_name: str = "") -> int:
    """티켓 본문 편집 + 변경 이력 자동 코멘트 (2026-04-28 대표 결재 (나)안).

    changes dict 의 키만 UPDATE. 각 필드 변경마다 시스템 코멘트로 이력 남김:
      "[편집] 납기 04-29 → 04-28 by 이현 매니저"

    허용 필드 (화이트리스트): title, description, due_date, urgency,
                              recipient_team_id, target_label, hours_estimated
    """
    EDITABLE = {"title", "description", "due_date", "urgency",
                "recipient_team_id", "target_label", "hours_estimated"}
    LABELS = {
        "title": "제목", "description": "내용", "due_date": "납기",
        "urgency": "긴급도", "recipient_team_id": "수신팀",
        "target_label": "대상", "hours_estimated": "예상공수",
    }
    applied = 0
    diffs = []
    with db_session() as c:
        cur = c.execute("SELECT * FROM tickets WHERE id=?", (tid,)).fetchone()
        if not cur:
            return 0
        team_name = {}
        # 수신팀 이름 표시용
        if "recipient_team_id" in changes:
            for r in c.execute("SELECT id, name FROM teams").fetchall():
                team_name[r["id"]] = r["name"]
        sets = []
        params = []
        for k, new_v in changes.items():
            if k not in EDITABLE:
                continue
            old_v = cur[k] if k in cur.keys() else None
            # 값 정규화
            if k == "recipient_team_id":
                try:
                    new_v = int(new_v) if new_v not in (None, "") else None
                except (ValueError, TypeError):
                    continue
            elif k == "hours_estimated":
                try:
                    new_v = float(new_v) if new_v not in (None, "") else None
                except (ValueError, TypeError):
                    continue
            else:
                new_v = (new_v or "").strip() if isinstance(new_v, str) else new_v
            if (old_v or "") == (new_v or ""):
                continue
            sets.append(f"{k} = ?")
            params.append(new_v)
            # 표시용
            if k == "recipient_team_id":
                old_label = team_name.get(old_v, str(old_v) if old_v else "(없음)")
                new_label = team_name.get(new_v, str(new_v) if new_v else "(없음)")
            else:
                old_label = str(old_v) if old_v not in (None, "") else "(없음)"
                new_label = str(new_v) if new_v not in (None, "") else "(없음)"
            diffs.append(f"{LABELS.get(k, k)}: {old_label} → {new_label}")
        if not sets:
            return 0
        sets.append("updated_at = ?")
        params.append(_logi_now())
        params.append(tid)
        c.execute(f"UPDATE tickets SET {', '.join(sets)} WHERE id = ?", params)
        applied = len(diffs)
        # 시스템 코멘트 — 이력 (누가/언제/무엇)
        body_lines = ["[편집] " + (editor_name or f"user#{editor_id}") + " 님이 수정"]
        body_lines += [f"  · {d}" for d in diffs]
        c.execute(
            "INSERT INTO ticket_comments (ticket_id, author_id, body, is_status_change) "
            "VALUES (?,?,?,1)",
            (tid, editor_id, "\n".join(body_lines)),
        )
    return applied


def ticket_change_status(tid: int, new_status: str, user_id: int, note: str = ""):
    """상태 변경 + 시스템 코멘트 자동 등록"""
    valid = TICKET_STATUSES
    if new_status not in valid:
        return
    extra_sql = ""
    params = [new_status, _logi_now()]
    if new_status == "완료":
        extra_sql = ", completed_at = ?, complete_note = ?"
        params += [_logi_now(), note]
    elif new_status == "접수":
        extra_sql = ", accept_note = ?"
        params += [note]
    elif new_status == "반려":
        extra_sql = ", reject_reason = ?"
        params += [note]
    sql = f"UPDATE tickets SET status = ?, updated_at = ? {extra_sql} WHERE id = ?"
    params.append(tid)
    with db_session() as c:
        c.execute(sql, params)
        # 시스템 코멘트
        sys_msg = f"[상태 변경] → {new_status}" + (f"\n메모: {note}" if note else "")
        c.execute(
            "INSERT INTO ticket_comments (ticket_id, author_id, body, is_status_change) VALUES (?,?,?,1)",
            (tid, user_id, sys_msg),
        )


def ticket_comments_list(tid: int) -> list[dict]:
    with db_session() as c:
        rows = c.execute(
            """SELECT tc.*, u.name AS author_name, u.rank AS author_rank, t.name AS author_team
               FROM ticket_comments tc
               JOIN users u ON tc.author_id = u.id
               LEFT JOIN teams t ON u.team_id = t.id
               WHERE tc.ticket_id = ?
               ORDER BY tc.created_at""",
            (tid,),
        ).fetchall()
    return [dict(r) for r in rows]


def ticket_add_comment(tid: int, author_id: int, body: str):
    with db_session() as c:
        c.execute(
            "INSERT INTO ticket_comments (ticket_id, author_id, body) VALUES (?,?,?)",
            (tid, author_id, body.strip()),
        )


def ticket_delete(tid: int):
    """v5H119: 공통 헬퍼 사용. 폴백: v5H115 인라인."""
    with db_session() as c:
        try:
            res = _safe_delete_with_cascade(
                c, "tickets", tid, fk_column="ticket_id",
                explicit_children=[
                    ("DELETE FROM ticket_comments WHERE ticket_id=?", (tid,)),
                ],
                keep_tables=("ticket_comments",),
            )
            if res.get("ok"):
                return
        except Exception:
            pass
        try:
            c.execute("DELETE FROM ticket_comments WHERE ticket_id=?", (tid,))
        except Exception:
            pass
        try:
            tables = [r[0] for r in c.execute(
                "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
            for tbl in tables:
                if tbl in ("tickets", "ticket_comments", "sqlite_sequence"):
                    continue
                try:
                    cols = [r[1] for r in c.execute(f"PRAGMA table_info({tbl})").fetchall()]
                    if "ticket_id" in cols:
                        try:
                            c.execute(f"UPDATE {tbl} SET ticket_id=NULL WHERE ticket_id=?", (tid,))
                        except Exception:
                            try:
                                c.execute(f"DELETE FROM {tbl} WHERE ticket_id=?", (tid,))
                            except Exception:
                                pass
                except Exception:
                    pass
        except Exception:
            pass
        c.execute("DELETE FROM tickets WHERE id=?", (tid,))


def tickets_count_for_user(user_id: int, team_id: int = None) -> dict:
    """홈 KPI / 사이드바 뱃지용 — 미처리 카운트"""
    with db_session() as c:
        # 내가 요청한 것 중 미완료
        my_open = c.execute(
            "SELECT COUNT(*) FROM tickets WHERE requester_id = ? AND status NOT IN ('완료','반려')",
            (user_id,),
        ).fetchone()[0]
        # 내가 받은 것 중 미처리 (요청·접수만)
        recv_pending = c.execute(
            """SELECT COUNT(*) FROM tickets
               WHERE (recipient_user_id = ? OR (recipient_team_id = ? AND recipient_user_id IS NULL))
               AND status IN ('요청', '접수')""",
            (user_id, team_id or 0),
        ).fetchone()[0]
    return {"my_open": my_open, "recv_pending": recv_pending}


# =====================================================
# ISSUES · AS DB (3순위 ⑦) — 헬퍼 함수
# =====================================================
def _team_id_by_name(name: str) -> int | None:
    """팀명 → id (없으면 None)"""
    if not name:
        return None
    try:
        with db_session() as c:
            r = c.execute("SELECT id FROM teams WHERE name=?", (name,)).fetchone()
            return r["id"] if r else None
    except Exception:
        return None


def route_issue_team(issue_type: str, biz_div: str) -> int | None:
    """이슈 종류 × 사업부 → 책임 팀 id 자동 판별"""
    rule = ISSUE_OWNER_RULES.get(issue_type, {})
    team_name = rule.get(biz_div) or rule.get("T") or rule.get("M")
    return _team_id_by_name(team_name) if team_name else None


def issue_create(data: dict, created_by: int) -> tuple[int, str]:
    """이슈 등록 + 자동 라우팅 + 접수 로그"""
    issue_no = gen_issue_no()
    issue_type = (data.get("issue_type") or "AS").strip()
    biz_div = (data.get("biz_div") or "").strip()
    owner_team_id = data.get("owner_team_id") or route_issue_team(issue_type, biz_div)

    pid = data.get("project_id")
    mgmt_code = data.get("mgmt_code") or ""
    customer_id = data.get("customer_id")
    customer_name = (data.get("customer_name") or "").strip()

    # 프로젝트가 있으면 mgmt_code/biz_div/customer 백필
    if pid:
        try:
            with db_session() as c:
                p = c.execute(
                    """SELECT p.biz_div, p.mgmt_code, p.name, p.customer_id, cu.name AS cu_name
                       FROM projects p LEFT JOIN customers cu ON p.customer_id=cu.id
                       WHERE p.id=?""",
                    (pid,),
                ).fetchone()
                if p:
                    if not biz_div:
                        biz_div = p["biz_div"] or ""
                    if not mgmt_code:
                        mgmt_code = p["mgmt_code"] or ""
                    if not customer_id and p["customer_id"]:
                        customer_id = p["customer_id"]
                    if not customer_name and p["cu_name"]:
                        customer_name = p["cu_name"]
        except Exception:
            pass

    with db_session() as c:
        c.execute(
            """INSERT INTO issues
               (issue_no, title, severity, issue_type, status,
                customer_id, customer_name, project_id, mgmt_code, biz_div,
                occurred_at, detected_by, description,
                owner_team_id, owner_user_id, created_by)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                issue_no,
                (data.get("title") or "").strip(),
                data.get("severity") or "중",
                issue_type,
                "접수",
                customer_id,
                customer_name,
                pid,
                mgmt_code,
                biz_div,
                (data.get("occurred_at") or "").strip() or _logi_now()[:10],
                (data.get("detected_by") or "").strip(),
                (data.get("description") or "").strip(),
                owner_team_id,
                data.get("owner_user_id"),
                created_by,
            ),
        )
        iid = c.execute("SELECT last_insert_rowid()").fetchone()[0]
        # 접수 로그
        c.execute(
            """INSERT INTO issue_logs (issue_id, user_id, action, content, new_status)
               VALUES (?, ?, '접수', ?, '접수')""",
            (iid, created_by, f"이슈 접수 — {issue_type} / {data.get('severity') or '중'}"),
        )
    return iid, issue_no


def issues_list(scope: str = "all", user_id: int = None, team_id: int = None,
                status: str = "", severity: str = "", issue_type: str = "",
                customer_id: int = None, q: str = "") -> list[dict]:
    """이슈 목록 (필터 지원)"""
    where = []
    params = []
    if scope == "mine_owned" and user_id:
        where.append("(i.owner_user_id = ? OR i.owner_team_id = ?)")
        params += [user_id, team_id or 0]
    elif scope == "team" and team_id:
        where.append("i.owner_team_id = ?")
        params.append(team_id)
    elif scope == "open":
        where.append("i.status NOT IN ('해결','종결')")
    if status:
        where.append("i.status = ?")
        params.append(status)
    if severity:
        where.append("i.severity = ?")
        params.append(severity)
    if issue_type:
        where.append("i.issue_type = ?")
        params.append(issue_type)
    if customer_id:
        where.append("i.customer_id = ?")
        params.append(customer_id)
    if q:
        where.append("(i.title LIKE ? OR i.description LIKE ? OR i.issue_no LIKE ? OR i.mgmt_code LIKE ?)")
        like = f"%{q}%"
        params += [like, like, like, like]
    sql = """SELECT i.*, t.name AS owner_team_name, u.name AS owner_user_name,
                    cu.name AS customer_full_name, cb.name AS created_by_name
             FROM issues i
             LEFT JOIN teams t ON i.owner_team_id = t.id
             LEFT JOIN users u ON i.owner_user_id = u.id
             LEFT JOIN customers cu ON i.customer_id = cu.id
             LEFT JOIN users cb ON i.created_by = cb.id"""
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += """ ORDER BY
                CASE i.severity WHEN '치명' THEN 1 WHEN '심각' THEN 2 WHEN '중' THEN 3 ELSE 4 END,
                CASE WHEN i.status IN ('해결','종결') THEN 9 ELSE 0 END,
                i.created_at DESC"""
    with db_session() as c:
        return [dict(r) for r in c.execute(sql, params).fetchall()]


def issue_get(iid: int) -> dict | None:
    with db_session() as c:
        r = c.execute(
            """SELECT i.*, t.name AS owner_team_name, u.name AS owner_user_name,
                      cu.name AS customer_full_name, cb.name AS created_by_name,
                      ch.change_no AS related_change_no, ch.title AS related_change_title
               FROM issues i
               LEFT JOIN teams t ON i.owner_team_id = t.id
               LEFT JOIN users u ON i.owner_user_id = u.id
               LEFT JOIN customers cu ON i.customer_id = cu.id
               LEFT JOIN users cb ON i.created_by = cb.id
               LEFT JOIN changes ch ON i.related_change_id = ch.id
               WHERE i.id = ?""",
            (iid,),
        ).fetchone()
        return dict(r) if r else None


def issue_logs_get(iid: int) -> list[dict]:
    with db_session() as c:
        return [dict(r) for r in c.execute(
            """SELECT il.*, u.name AS user_name
               FROM issue_logs il LEFT JOIN users u ON il.user_id = u.id
               WHERE il.issue_id = ? ORDER BY il.created_at""",
            (iid,),
        ).fetchall()]


def issue_update(iid: int, data: dict, user_id: int) -> bool:
    """이슈 갱신 (원인/조치/재발방지/상태/책임자) + 로그"""
    with db_session() as c:
        cur = c.execute("SELECT * FROM issues WHERE id=?", (iid,)).fetchone()
        if not cur:
            return False
        old_status = cur["status"]
        sets = []
        params = []
        log_actions = []
        for col in ("title", "severity", "issue_type", "root_cause", "action_taken",
                    "prevention", "owner_team_id", "owner_user_id", "cost_estimate",
                    "related_change_id", "resolved_at"):
            if col in data and data[col] is not None and data[col] != "":
                sets.append(f"{col} = ?")
                params.append(data[col])
                if col in ("root_cause", "action_taken", "prevention"):
                    log_actions.append((col, data[col]))
        new_status = data.get("status")
        if new_status and new_status != old_status:
            sets.append("status = ?")
            params.append(new_status)
            if new_status in ("해결", "종결") and not cur["resolved_at"]:
                sets.append("resolved_at = ?")
                params.append(_logi_now())
        if not sets:
            return False
        sets.append("updated_at = ?")
        params.append(_logi_now())
        params.append(iid)
        c.execute(f"UPDATE issues SET {', '.join(sets)} WHERE id=?", params)
        # 로그
        if new_status and new_status != old_status:
            c.execute(
                """INSERT INTO issue_logs (issue_id, user_id, action, content, old_status, new_status)
                   VALUES (?, ?, '상태변경', ?, ?, ?)""",
                (iid, user_id, f"{old_status} → {new_status}" + (f" / {data.get('note','')}" if data.get('note') else ""),
                 old_status, new_status),
            )
        for action_col, content in log_actions:
            label = {"root_cause": "원인분석", "action_taken": "조치", "prevention": "재발방지"}[action_col]
            c.execute(
                """INSERT INTO issue_logs (issue_id, user_id, action, content)
                   VALUES (?, ?, ?, ?)""",
                (iid, user_id, label, content),
            )
        if data.get("comment"):
            c.execute(
                """INSERT INTO issue_logs (issue_id, user_id, action, content)
                   VALUES (?, ?, '코멘트', ?)""",
                (iid, user_id, data["comment"]),
            )
    return True


def issue_delete(iid: int) -> bool:
    """v5H119: 공통 헬퍼 사용. 폴백: v5H115 인라인."""
    with db_session() as c:
        try:
            res = _safe_delete_with_cascade(
                c, "issues", iid, fk_column="issue_id",
                explicit_children=[
                    ("DELETE FROM issue_logs WHERE issue_id=?", (iid,)),
                ],
                keep_tables=("issue_logs",),
            )
            if res.get("ok"):
                return True
        except Exception:
            pass
        try:
            c.execute("DELETE FROM issue_logs WHERE issue_id=?", (iid,))
        except Exception:
            pass
        try:
            tables = [r[0] for r in c.execute(
                "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
            for tbl in tables:
                if tbl in ("issues", "issue_logs", "sqlite_sequence"):
                    continue
                try:
                    cols = [r[1] for r in c.execute(f"PRAGMA table_info({tbl})").fetchall()]
                    if "issue_id" in cols:
                        try:
                            c.execute(f"UPDATE {tbl} SET issue_id=NULL WHERE issue_id=?", (iid,))
                        except Exception:
                            try:
                                c.execute(f"DELETE FROM {tbl} WHERE issue_id=?", (iid,))
                            except Exception:
                                pass
                except Exception:
                    pass
        except Exception:
            pass
        c.execute("DELETE FROM issues WHERE id=?", (iid,))
    return True


def issues_kpi() -> dict:
    """이슈 KPI: 미해결/심각도/평균 해결시간"""
    with db_session() as c:
        total = c.execute("SELECT COUNT(*) FROM issues").fetchone()[0]
        open_cnt = c.execute("SELECT COUNT(*) FROM issues WHERE status NOT IN ('해결','종결')").fetchone()[0]
        critical = c.execute("SELECT COUNT(*) FROM issues WHERE severity IN ('치명','심각') AND status NOT IN ('해결','종결')").fetchone()[0]
        recent = c.execute(
            """SELECT COUNT(*) FROM issues
               WHERE created_at >= date('now','-30 day','localtime')"""
        ).fetchone()[0]
        # 부서별 미해결
        by_team = [dict(r) for r in c.execute(
            """SELECT t.name AS team_name, COUNT(*) AS cnt
               FROM issues i LEFT JOIN teams t ON i.owner_team_id = t.id
               WHERE i.status NOT IN ('해결','종결')
               GROUP BY i.owner_team_id ORDER BY cnt DESC"""
        ).fetchall()]
    return {"total": total, "open": open_cnt, "critical": critical,
            "recent_30d": recent, "by_team": by_team}


# =====================================================
# 진행률 대시보드 (HAIST WORKS — 1순위 ① / 8팀 공통 요구)
# 12공정 매트릭스 — 베이비 V2 진척과 통합 가능 구조
# =====================================================

# 12공정 정의 (관리코드 단위 진척 추적)
PHASE_DEFS = [
    ("order",     "수주",       1),
    ("concept",   "컨셉",       2),
    ("design",    "기구설계",   3),
    ("elec",      "전장설계",   4),
    ("sw",        "소프트웨어", 5),
    ("machining", "가공",       6),
    ("buying",    "구매",       7),
    ("assembly",  "조립",       8),
    ("qc",        "검수",       9),
    ("ship",      "출하",       10),
    ("setup",    "Set-Up",     11),
    ("knkvn",    "KNKVN 이관", 12),
]
PHASE_CODE_TO_LABEL = {c: l for c, l, _ in PHASE_DEFS}
PHASE_STATUSES = ["예정", "진행", "완료", "지연", "보류"]

# 공정 → 담당 부서 자동 매핑 (사업부별)
PHASE_TEAM_RULES = {
    "order":     {"T": "기술영업팀", "M": "기술영업팀"},
    "concept":   {"T": "기술영업팀", "M": "기술영업팀"},
    "design":    {"T": "설계팀",     "M": "설계팀"},
    "elec":      {"T": "전장설계팀", "M": "전장설계팀"},
    "sw":        {"T": "소프트웨어팀","M": "소프트웨어팀"},
    "machining": {"T": "가공팀",     "M": "가공팀"},
    "buying":    {"T": "구매팀",     "M": "구매팀"},
    "assembly":  {"T": "제조기술1팀","M": "제조기술2팀"},
    "qc":        {"T": "품질팀",     "M": "품질팀"},
    "ship":      {"T": "기술영업팀", "M": "기술영업팀"},
    "setup":     {"T": "검사기팀",   "M": "기술영업팀"},
    "knkvn":     {"T": "기술영업팀", "M": "기술영업팀"},
}


def ensure_phases_for_project(project_id: int):
    """프로젝트에 12공정 자동 시드 (없는 경우만)"""
    with db_session() as c:
        # 사업부 가져오기
        proj = c.execute("SELECT biz_div FROM projects WHERE id=?", (project_id,)).fetchone()
        biz_div = proj["biz_div"] if proj and proj["biz_div"] in ("T", "M", "L") else "T"
        # 기존 phase 확인
        exist = {r["phase_code"] for r in c.execute(
            "SELECT phase_code FROM project_phases WHERE project_id=?", (project_id,)
        ).fetchall()}
        for code, _label, order in PHASE_DEFS:
            if code in exist:
                continue
            # 자동 담당 부서
            team_name = PHASE_TEAM_RULES.get(code, {}).get(biz_div)
            team_id = None
            if team_name:
                row = c.execute("SELECT id FROM teams WHERE name=?", (team_name,)).fetchone()
                team_id = row["id"] if row else None
            c.execute(
                """INSERT INTO project_phases
                   (project_id, phase_code, phase_order, status, progress_pct,
                    assignee_team_id, updated_at)
                   VALUES (?,?,?,?,?,?,?)""",
                (project_id, code, order, "예정", 0, team_id, _logi_now()),
            )


def progress_matrix(biz_div: str = "", customer: str = "", status: str = "",
                    limit: int = 50):
    """진행률 매트릭스 — 프로젝트 × 12공정"""
    sql = """SELECT id, mgmt_code, name, biz_div, customer_name, status, due_date
             FROM projects
             WHERE mgmt_code IS NOT NULL AND mgmt_code != ''"""
    params = []
    if biz_div:
        sql += " AND biz_div = ?"
        params.append(biz_div)
    if customer:
        sql += " AND customer_name LIKE ?"
        params.append(f"%{customer}%")
    if status:
        sql += " AND status = ?"
        params.append(status)
    sql += " ORDER BY id DESC LIMIT ?"
    params.append(limit)

    with db_session() as c:
        projects = [dict(r) for r in c.execute(sql, params).fetchall()]
        # 각 프로젝트의 phase 가져오기
        for p in projects:
            phases = c.execute(
                """SELECT pp.*, t.name AS team_name, u.name AS assignee_name
                   FROM project_phases pp
                   LEFT JOIN teams t ON pp.assignee_team_id = t.id
                   LEFT JOIN users u ON pp.assignee_id = u.id
                   WHERE pp.project_id = ?
                   ORDER BY pp.phase_order""",
                (p["id"],),
            ).fetchall()
            phase_map = {ph["phase_code"]: dict(ph) for ph in phases}
            # 12공정 모두 노출 (없는 건 None)
            p["phases"] = [phase_map.get(code) for code, _, _ in PHASE_DEFS]
            # 전체 진척률 (12공정 평균)
            valid = [ph for ph in phase_map.values() if ph]
            p["overall_pct"] = round(sum(ph["progress_pct"] or 0 for ph in valid) / 12, 1) if valid else 0
            p["delayed_count"] = sum(1 for ph in valid if ph["status"] == "지연")
    return projects


def project_phases_get(project_id: int) -> list[dict]:
    ensure_phases_for_project(project_id)
    with db_session() as c:
        rows = c.execute(
            """SELECT pp.*, t.name AS team_name, u.name AS assignee_name
               FROM project_phases pp
               LEFT JOIN teams t ON pp.assignee_team_id = t.id
               LEFT JOIN users u ON pp.assignee_id = u.id
               WHERE pp.project_id = ?
               ORDER BY pp.phase_order""",
            (project_id,),
        ).fetchall()
    return [dict(r) for r in rows]


def project_phase_update(phase_id: int, data: dict, user_id: int):
    """공정 업데이트 — 1~2 클릭 변경 가능"""
    sets = []
    vals = []
    for f in ["status", "progress_pct", "planned_start", "planned_end",
              "actual_start", "actual_end", "note", "assignee_id"]:
        if f in data:
            sets.append(f"{f} = ?")
            v = data[f]
            if f == "progress_pct" and v not in (None, ""):
                try:
                    v = float(v)
                except (ValueError, TypeError):
                    v = 0
            vals.append(v if v != "" else None)
    if not sets:
        return
    sets.append("updated_by = ?")
    sets.append("updated_at = ?")
    vals += [user_id, _logi_now(), phase_id]
    with db_session() as c:
        c.execute(f"UPDATE project_phases SET {', '.join(sets)} WHERE id = ?", vals)


def progress_summary_for_user(user_id: int, team_id: int = None) -> dict:
    """홈 KPI / 사이드바용 — 내 부서 담당 미완료 공정 카운트"""
    with db_session() as c:
        if team_id:
            my = c.execute(
                """SELECT COUNT(*) FROM project_phases
                   WHERE assignee_team_id = ? AND status NOT IN ('완료', '보류')""",
                (team_id,),
            ).fetchone()[0]
            delayed = c.execute(
                """SELECT COUNT(*) FROM project_phases
                   WHERE assignee_team_id = ? AND status = '지연'""",
                (team_id,),
            ).fetchone()[0]
        else:
            my = 0
            delayed = c.execute(
                "SELECT COUNT(*) FROM project_phases WHERE status = '지연'"
            ).fetchone()[0]
    return {"my_open": my, "delayed": delayed}


# =====================================================
# 환율·단가 강화 (2026-04-26 Top10 #9) — 외부 API 미사용
# =====================================================
def cost_simulation_create(data: dict, user_id: int) -> int:
    """원가 시뮬레이션 INSERT. 외부 환율 API 미사용."""
    base = float(data["unit_price_base"])
    rate = float(data["exchange_rate"])
    target = float(data.get("unit_price_target") or (base * rate))
    margin = float(data.get("margin_pct") or 0)
    with db_session() as c:
        c.execute(
            """INSERT INTO cost_simulations
               (part_id, simulated_at, base_currency, target_currency,
                exchange_rate, unit_price_base, unit_price_target,
                margin_pct, note, created_by)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (int(data["part_id"]),
             data.get("simulated_at") or _logi_now(),
             (data.get("base_currency") or "USD").upper(),
             (data.get("target_currency") or "KRW").upper(),
             rate, base, target, margin, data.get("note"), user_id),
        )
        return c.execute("SELECT last_insert_rowid()").fetchone()[0]


def cost_simulations_list(part_id: int, limit: int = 30) -> list[dict]:
    with db_session() as c:
        return [dict(r) for r in c.execute(
            """SELECT cs.*, u.name AS created_by_name
               FROM cost_simulations cs
               LEFT JOIN users u ON cs.created_by = u.id
               WHERE cs.part_id = ?
               ORDER BY cs.simulated_at DESC, cs.id DESC LIMIT ?""",
            (part_id, limit),
        ).fetchall()]


def price_change_log(part_id: int, supplier_id, old_price, new_price,
                     effective_date: str, user_id: int, note: str = "") -> int:
    """단가 변경 이력 자동 기록. old_price=None 가능 (최초 등록)."""
    op = float(old_price) if old_price else None
    np = float(new_price)
    chg = ((np - op) / op * 100.0) if (op and op > 0) else None
    with db_session() as c:
        c.execute(
            """INSERT INTO price_change_history
               (part_id, supplier_id, old_price, new_price, change_pct,
                effective_date, note, changed_by)
               VALUES (?,?,?,?,?,?,?,?)""",
            (int(part_id),
             int(supplier_id) if supplier_id else None,
             op, np, chg, effective_date, note, user_id),
        )
        return c.execute("SELECT last_insert_rowid()").fetchone()[0]


def price_change_history_list(part_id: int, limit: int = 50) -> list[dict]:
    with db_session() as c:
        return [dict(r) for r in c.execute(
            """SELECT pch.*, s.name AS supplier_name, u.name AS changed_by_name
               FROM price_change_history pch
               LEFT JOIN suppliers s ON pch.supplier_id = s.id
               LEFT JOIN users u ON pch.changed_by = u.id
               WHERE pch.part_id = ?
               ORDER BY pch.effective_date DESC, pch.id DESC LIMIT ?""",
            (part_id, limit),
        ).fetchall()]


def rate_alert_create(target_currency: str, threshold: float,
                      direction: str, user_id: int) -> int:
    with db_session() as c:
        c.execute(
            """INSERT INTO rate_alerts
               (target_currency, threshold, direction, created_by)
               VALUES (?,?,?,?)""",
            (target_currency.upper(), float(threshold),
             direction if direction in ("above", "below") else "above",
             user_id),
        )
        return c.execute("SELECT last_insert_rowid()").fetchone()[0]


def rate_alerts_list(active_only: bool = True) -> list[dict]:
    sql = """SELECT ra.*, u.name AS created_by_name
             FROM rate_alerts ra LEFT JOIN users u ON ra.created_by = u.id"""
    if active_only:
        sql += " WHERE ra.is_active = 1"
    sql += " ORDER BY ra.created_at DESC LIMIT 100"
    with db_session() as c:
        return [dict(r) for r in c.execute(sql).fetchall()]


def exchange_rates_csv_upload(rows: list[dict], user_id: int) -> dict:
    """CSV 일괄 업로드 (rate_date / from_currency / to_currency / rate / source / note).
    외부 API 미호출. UPSERT (날짜·통화쌍)."""
    ok, ng = 0, 0
    errors = []
    for i, r in enumerate(rows, 1):
        try:
            exchange_rate_create({
                "rate_date": r["rate_date"],
                "from_currency": r["from_currency"],
                "to_currency": r.get("to_currency", "KRW"),
                "rate": float(r["rate"]),
                "source": r.get("source") or "CSV",
                "note": r.get("note"),
            }, user_id=user_id)
            ok += 1
        except Exception as e:
            ng += 1
            errors.append(f"행 {i}: {e}")
    return {"ok": ok, "ng": ng, "errors": errors}


# ----- S3-1: 환율 알림 발동 잡 (옵션 A · 자동) -----
def check_rate_alerts(target_currency: str, rate: float) -> int:
    """환율 등록 후 자동 검사. 활성 알림 중 임계 초과 시 triggered_at 갱신.
    반환: 발동된 알림 건수. 외부 발송 없음 (DB 표식만)."""
    cur = (target_currency or "").upper()
    try:
        v = float(rate)
    except Exception:
        return 0
    fired = 0
    with db_session() as c:
        rows = c.execute(
            """SELECT id, threshold, direction FROM rate_alerts
               WHERE is_active=1 AND target_currency=?""",
            (cur,),
        ).fetchall()
        for r in rows:
            th = float(r["threshold"] or 0)
            d = r["direction"] or "above"
            hit = (v >= th) if d == "above" else (v <= th)
            if hit:
                c.execute(
                    "UPDATE rate_alerts SET triggered_at=? WHERE id=?",
                    (_logi_now(), int(r["id"])),
                )
                fired += 1
    return fired


# =====================================================
# 글로벌 검색 — 전사 통합 (2026-04-26 발주, 09 → 01)
# 모든 LIKE 쿼리는 parameter binding (SQL 인젝션 방지).
# 외부 검색엔진 0건 (Elasticsearch 등 절대 금지).
# v2 본체 무수정 · G1~G5 핫패치 보존.
# =====================================================

def _gs_q(q):
    """검색어 정규화 — strip + LIKE 와일드카드. 빈 문자열이면 None."""
    s = (q or "").strip()
    if len(s) < 2:
        return None
    return f"%{s}%"


def search_orders(q, limit=5):
    """수주 검색 — orders.order_no + customers.name + parts.part_no/part_name (order_items 조인)."""
    qq = _gs_q(q)
    if not qq:
        return []
    with db_session() as c:
        rows = c.execute(
            """SELECT DISTINCT o.id, o.order_no, o.order_date, o.due_date, o.total_amount, o.status,
                      cu.name AS customer_name
               FROM orders o
               LEFT JOIN customers cu ON o.customer_id=cu.id
               LEFT JOIN order_items oi ON oi.order_id=o.id
               LEFT JOIN parts p ON oi.part_id=p.id
               WHERE o.order_no LIKE ? OR cu.name LIKE ?
                  OR p.part_no LIKE ? OR p.part_name LIKE ?
               ORDER BY o.order_date DESC LIMIT ?""",
            (qq, qq, qq, qq, int(limit)),
        ).fetchall()
        return [dict(r) for r in rows]


def customers_for_picker():
    """v5H75: 프로젝트 등록 폼 고객사 드롭다운용 — 활성 고객 전부 (id, name)."""
    with db_session() as c:
        try:
            rows = c.execute(
                "SELECT id, name FROM customers WHERE COALESCE(is_active,1)=1 ORDER BY name ASC"
            ).fetchall()
        except Exception:
            rows = c.execute(
                "SELECT id, name FROM customers ORDER BY name ASC"
            ).fetchall()
        return [dict(r) for r in rows]


def search_customers(q, limit=5):
    """고객 검색 — customers.name + tier + note."""
    qq = _gs_q(q)
    if not qq:
        return []
    with db_session() as c:
        rows = c.execute(
            """SELECT id, name, tier, note FROM customers
               WHERE name LIKE ? OR COALESCE(tier,'') LIKE ? OR COALESCE(note,'') LIKE ?
               ORDER BY name ASC LIMIT ?""",
            (qq, qq, qq, int(limit)),
        ).fetchall()
        return [dict(r) for r in rows]


def search_parts(q, limit=5):
    """부품 검색 — parts.part_no + part_name + spec + maker."""
    qq = _gs_q(q)
    if not qq:
        return []
    with db_session() as c:
        rows = c.execute(
            """SELECT id, part_no, part_name, spec, maker, biz_div, category, std_price, currency
               FROM parts
               WHERE is_active=1 AND (
                     part_no LIKE ? OR part_name LIKE ?
                  OR COALESCE(spec,'') LIKE ? OR COALESCE(maker,'') LIKE ?)
               ORDER BY part_no ASC LIMIT ?""",
            (qq, qq, qq, qq, int(limit)),
        ).fetchall()
        return [dict(r) for r in rows]


def search_issues(q, limit=5):
    """이슈 검색 — issues.issue_no + title + description + customer_name + mgmt_code."""
    qq = _gs_q(q)
    if not qq:
        return []
    with db_session() as c:
        rows = c.execute(
            """SELECT i.id, i.issue_no, i.title, i.severity, i.status, i.issue_type,
                      i.customer_name, i.mgmt_code, i.occurred_at
               FROM issues i
               WHERE COALESCE(i.issue_no,'') LIKE ? OR i.title LIKE ?
                  OR COALESCE(i.description,'') LIKE ?
                  OR COALESCE(i.customer_name,'') LIKE ?
                  OR COALESCE(i.mgmt_code,'') LIKE ?
               ORDER BY i.created_at DESC LIMIT ?""",
            (qq, qq, qq, qq, qq, int(limit)),
        ).fetchall()
        return [dict(r) for r in rows]


def search_tickets(q, limit=5):
    """티켓 검색 — tickets.ticket_no + title + description + category + target_label."""
    qq = _gs_q(q)
    if not qq:
        return []
    with db_session() as c:
        rows = c.execute(
            """SELECT id, ticket_no, category, title, status, urgency, target_label,
                      due_date, created_at
               FROM tickets
               WHERE ticket_no LIKE ? OR title LIKE ?
                  OR COALESCE(description,'') LIKE ?
                  OR COALESCE(category,'') LIKE ?
                  OR COALESCE(target_label,'') LIKE ?
               ORDER BY created_at DESC LIMIT ?""",
            (qq, qq, qq, qq, qq, int(limit)),
        ).fetchall()
        return [dict(r) for r in rows]


def search_users(q, limit=5):
    """사용자 검색 — users.name + login_id + email + rank (활성만)."""
    qq = _gs_q(q)
    if not qq:
        return []
    with db_session() as c:
        rows = c.execute(
            """SELECT u.id, u.name, u.login_id, u.email, u.rank, u.role,
                      t.name AS team_name
               FROM users u LEFT JOIN teams t ON u.team_id=t.id
               WHERE u.is_active=1 AND (
                     u.name LIKE ? OR u.login_id LIKE ?
                  OR COALESCE(u.email,'') LIKE ? OR COALESCE(u.rank,'') LIKE ?)
               ORDER BY u.name ASC LIMIT ?""",
            (qq, qq, qq, qq, int(limit)),
        ).fetchall()
        return [dict(r) for r in rows]


def search_boards(q, limit=5):
    """게시판 검색 — board_posts.title + body (승인된 글만)."""
    qq = _gs_q(q)
    if not qq:
        return []
    with db_session() as c:
        rows = c.execute(
            """SELECT bp.id, bp.board_id, bp.title, bp.category, bp.created_at,
                      u.name AS author_name, b.name AS board_name
               FROM board_posts bp
               LEFT JOIN users u ON bp.author_id=u.id
               LEFT JOIN boards b ON bp.board_id=b.id
               WHERE bp.approval_status='approved' AND (
                     bp.title LIKE ? OR COALESCE(bp.body,'') LIKE ?)
               ORDER BY bp.created_at DESC LIMIT ?""",
            (qq, qq, int(limit)),
        ).fetchall()
        return [dict(r) for r in rows]


def search_exports(q, limit=5):
    """수출입 검색 — export_orders.buyer + commercial_invoices.invoice_no."""
    qq = _gs_q(q)
    if not qq:
        return []
    with db_session() as c:
        rows = c.execute(
            """SELECT eo.id, eo.buyer, eo.shipping_terms, eo.payment_terms,
                      eo.port_of_loading, eo.port_of_discharge, eo.status,
                      o.order_no, ci.invoice_no
               FROM export_orders eo
               LEFT JOIN orders o ON eo.order_id=o.id
               LEFT JOIN commercial_invoices ci ON ci.export_order_id=eo.id
               WHERE COALESCE(eo.buyer,'') LIKE ?
                  OR COALESCE(o.order_no,'') LIKE ?
                  OR COALESCE(ci.invoice_no,'') LIKE ?
                  OR COALESCE(eo.port_of_loading,'') LIKE ?
                  OR COALESCE(eo.port_of_discharge,'') LIKE ?
               ORDER BY eo.created_at DESC LIMIT ?""",
            (qq, qq, qq, qq, qq, int(limit)),
        ).fetchall()
        return [dict(r) for r in rows]


def search_audits(q, limit=5):
    """재고실사 검색 — stock_audits.audit_no + note."""
    qq = _gs_q(q)
    if not qq:
        return []
    with db_session() as c:
        rows = c.execute(
            """SELECT sa.id, sa.audit_no, sa.start_date, sa.end_date, sa.status,
                      sa.note, u.name AS led_by_name
               FROM stock_audits sa
               LEFT JOIN users u ON sa.led_by=u.id
               WHERE COALESCE(sa.audit_no,'') LIKE ? OR COALESCE(sa.note,'') LIKE ?
               ORDER BY sa.start_date DESC LIMIT ?""",
            (qq, qq, int(limit)),
        ).fetchall()
        return [dict(r) for r in rows]


# 글로벌 검색 카테고리 → 헬퍼 매핑
GLOBAL_SEARCH_CATEGORIES = {
    "orders":    search_orders,
    "customers": search_customers,
    "parts":     search_parts,
    "issues":    search_issues,
    "tickets":   search_tickets,
    "users":     search_users,
    "boards":    search_boards,
    "exports":   search_exports,
    "audits":    search_audits,
}


def global_search(q, categories=None, limit_per=5):
    """전사 통합 검색. categories=None 이면 전체. limit_per: 카테고리별 상한.
    반환: dict{cat: [rows]}. 빈 카테고리도 키 포함."""
    if categories is None:
        cats = list(GLOBAL_SEARCH_CATEGORIES.keys())
    else:
        cats = [c for c in categories if c in GLOBAL_SEARCH_CATEGORIES]
    out = {}
    for cat in cats:
        try:
            out[cat] = GLOBAL_SEARCH_CATEGORIES[cat](q, limit_per)
        except Exception:
            # 안전 폴백: 어느 한 카테고리 실패가 전체 결과를 막지 않도록.
            out[cat] = []
    return out


# =====================================================
# CEO 통합 대시보드 KPI (2026-04-26 — Top10 9 KPI 통합)
# 단일 헬퍼 ceo_dashboard_kpis() — 매출 / 재고 / 권한 / 수출입 /
# QMS / 주간 / 간트 / 환율 / 알림 9개 영역.
# 모든 쿼리 안전 폴백 (테이블 미존재 / 데이터 0 → 0 반환).
# v2 본체(시안·UI) 미접촉 · 외부 자산 0건 · idempotent.
# =====================================================
def ceo_dashboard_kpis(user_id: int = 0) -> dict:
    """CEO 통합 대시보드용 9 영역 KPI dict 반환.
    - 키: sales / stock / perms / exports / qms / weekly / gantt / fx / notif
    - 각 카테고리 실패 시 0/빈값으로 폴백, 전체 헬퍼는 항상 dict 반환."""
    out = {
        "sales":   {"month": 0.0, "prev_month": 0.0, "growth": 0.0, "unpaid": 0.0},
        "stock":   {"on_hand_kinds": 0, "qc_open": 0},
        "perms":   {"active_tokens": 0, "expiring_7d": 0},
        "exports": {"in_progress": 0, "ship_soon": 0},
        "qms":     {"open": 0, "sla_violation": 0},
        "weekly":  {"completed": 0, "in_progress": 0, "delayed": 0, "rate": 0},
        "gantt":   {"delayed_projects": 0},
        "fx":      {"usd_rate": 0.0, "alerts_active": 0},
        "notif":   {"unread": 0},
    }
    today = date.today() if False else None
    # 표준 datetime import 보장 (모듈 전역 import 없을 시 보강)
    from datetime import date as _date, timedelta as _td
    today = _date.today()
    mon = (today - _td(days=today.weekday())).isoformat()
    sun = (today - _td(days=today.weekday()) + _td(days=6)).isoformat()
    month_first = today.replace(day=1).isoformat()
    prev_month_first = (today.replace(day=1) - _td(days=1)).replace(day=1).isoformat()
    prev_month_last = (today.replace(day=1) - _td(days=1)).isoformat()
    week_later = (today + _td(days=7)).isoformat()
    today_s = today.isoformat()

    with db_session() as c:
        # 1) 매출 (orders 당월·전월 합계 + 미수금)
        try:
            r = c.execute(
                "SELECT COALESCE(SUM(total_amount),0) FROM orders "
                "WHERE order_date>=? AND order_date<=?",
                (month_first, today_s),
            ).fetchone()
            out["sales"]["month"] = float(r[0] or 0)
            r = c.execute(
                "SELECT COALESCE(SUM(total_amount),0) FROM orders "
                "WHERE order_date>=? AND order_date<=?",
                (prev_month_first, prev_month_last),
            ).fetchone()
            out["sales"]["prev_month"] = float(r[0] or 0)
            if out["sales"]["prev_month"] > 0:
                out["sales"]["growth"] = round(
                    (out["sales"]["month"] - out["sales"]["prev_month"])
                    * 100 / out["sales"]["prev_month"], 1)
            # 미수금 = 발행 invoice 합 - 수금 합
            r = c.execute(
                "SELECT COALESCE(SUM(total_amount),0) FROM invoices WHERE status='ISSUED'"
            ).fetchone()
            issued = float(r[0] or 0)
            r = c.execute(
                "SELECT COALESCE(SUM(amount),0) FROM receipts_payment"
            ).fetchone()
            received = float(r[0] or 0)
            out["sales"]["unpaid"] = max(0.0, issued - received)
        except Exception:
            pass

        # 2) 재고 (잔고 품목수 + QC 부적합 진행)
        try:
            r = c.execute(
                "SELECT COUNT(*) FROM stock_balances WHERE on_hand>0"
            ).fetchone()
            out["stock"]["on_hand_kinds"] = int(r[0] or 0)
            r = c.execute(
                "SELECT COUNT(*) FROM qc_inspections WHERE result='FAIL'"
            ).fetchone()
            out["stock"]["qc_open"] = int(r[0] or 0)
        except Exception:
            pass

        # 3) 권한 (활성 위임 토큰 + 만료 임박 7일)
        try:
            r = c.execute(
                "SELECT COUNT(*) FROM delegation_tokens WHERE status='ACTIVE'"
            ).fetchone()
            out["perms"]["active_tokens"] = int(r[0] or 0)
            r = c.execute(
                "SELECT COUNT(*) FROM delegation_tokens "
                "WHERE status='ACTIVE' AND expires_at IS NOT NULL "
                "AND expires_at<=?",
                (week_later,),
            ).fetchone()
            out["perms"]["expiring_7d"] = int(r[0] or 0)
        except Exception:
            pass

        # 4) 수출입 (진행 중 + 7일 내 출하 임박)
        try:
            r = c.execute(
                "SELECT COUNT(*) FROM export_orders "
                "WHERE status NOT IN ('DELIVERED','CANCELLED')"
            ).fetchone()
            out["exports"]["in_progress"] = int(r[0] or 0)
            r = c.execute(
                "SELECT COUNT(*) FROM export_orders "
                "WHERE status NOT IN ('DELIVERED','CANCELLED') "
                "AND ship_date IS NOT NULL AND ship_date<=?",
                (week_later,),
            ).fetchone()
            out["exports"]["ship_soon"] = int(r[0] or 0)
        except Exception:
            pass

        # 5) QMS (오픈 이슈 + SLA 위반 — 14일 미해결 치명/심각)
        try:
            r = c.execute(
                "SELECT COUNT(*) FROM issues WHERE status NOT IN ('해결','종결')"
            ).fetchone()
            out["qms"]["open"] = int(r[0] or 0)
            sla_thr = (today - _td(days=14)).isoformat()
            r = c.execute(
                "SELECT COUNT(*) FROM issues "
                "WHERE status NOT IN ('해결','종결') "
                "AND severity IN ('치명','심각') "
                "AND COALESCE(occurred_at,'')<=?",
                (sla_thr,),
            ).fetchone()
            out["qms"]["sla_violation"] = int(r[0] or 0)
        except Exception:
            pass

        # 6) 주간 (이번 주 진행률)
        try:
            r = c.execute(
                "SELECT "
                "SUM(CASE WHEN status='완료' THEN 1 ELSE 0 END) AS d, "
                "SUM(CASE WHEN status='진행중' THEN 1 ELSE 0 END) AS p, "
                "SUM(CASE WHEN status='지연' THEN 1 ELSE 0 END) AS dl, "
                "COUNT(*) AS tot "
                "FROM tasks WHERE work_date>=? AND work_date<=?",
                (mon, sun),
            ).fetchone()
            d, p, dl, tot = (r["d"] or 0), (r["p"] or 0), (r["dl"] or 0), (r["tot"] or 0)
            out["weekly"]["completed"] = int(d)
            out["weekly"]["in_progress"] = int(p)
            out["weekly"]["delayed"] = int(dl)
            out["weekly"]["rate"] = round(d * 100 / tot) if tot else 0
        except Exception:
            pass

        # 7) 간트 (지연 프로젝트 — end_date 경과 & 진행중)
        try:
            r = c.execute(
                "SELECT COUNT(*) FROM projects "
                "WHERE status='진행중' AND end_date IS NOT NULL AND end_date<?",
                (today_s,),
            ).fetchone()
            out["gantt"]["delayed_projects"] = int(r[0] or 0)
        except Exception:
            pass

        # 8) 환율 (USD 최신 + 활성 알림)
        try:
            r = c.execute(
                "SELECT rate FROM exchange_rates "
                "WHERE from_currency='USD' AND to_currency='KRW' "
                "ORDER BY rate_date DESC LIMIT 1"
            ).fetchone()
            if r:
                out["fx"]["usd_rate"] = float(r[0] or 0)
            r = c.execute(
                "SELECT COUNT(*) FROM rate_alerts WHERE is_active=1"
            ).fetchone()
            out["fx"]["alerts_active"] = int(r[0] or 0)
        except Exception:
            pass

        # 9) 알림 (현재 사용자 UNREAD)
        try:
            if user_id:
                r = c.execute(
                    "SELECT COUNT(*) FROM notifications "
                    "WHERE user_id=? AND is_read=0",
                    (user_id,),
                ).fetchone()
                out["notif"]["unread"] = int(r[0] or 0)
        except Exception:
            pass

    return out


# =====================================================
# Sales 견적 라인 헬퍼 (사이클 58 — 2차 보강)
# =====================================================

def get_quotation_items(quotation_id: int) -> list[dict]:
    """견적 라인 리스트 — quotation_items + parts 조인 (인쇄/수주전환 공통 소스)."""
    if not quotation_id:
        return []
    with db_session() as c:
        rows = c.execute(
            """SELECT qi.id, qi.quotation_id, qi.line_no, qi.part_id,
                      qi.item_name, qi.qty, qi.unit, qi.unit_price,
                      qi.total_price, qi.note, qi.created_at,
                      COALESCE(p.part_no, '') AS part_no,
                      COALESCE(p.part_name, '') AS part_name
               FROM quotation_items qi
               LEFT JOIN parts p ON p.id = qi.part_id
               WHERE qi.quotation_id = ?
               ORDER BY qi.line_no ASC, qi.id ASC""",
            (quotation_id,),
        ).fetchall()
        return [dict(r) for r in rows]


def clone_quotation_to_order(quotation_id: int, due_date: str | None = None,
                              created_by: int = 0) -> tuple[int, str]:
    """견적 + 라인 → 수주(orders) + 라인(order_items) 자동 복제.
    반환: (order_id, order_no). 견적 없으면 (0, '').
    호출 책임: quotations.status='CONFIRMED' 전환은 호출자 별도 처리.
    """
    if not quotation_id:
        return (0, "")
    from datetime import datetime, date
    with db_session() as c:
        q = c.execute(
            """SELECT id, customer_id, total_amount FROM quotations WHERE id=?""",
            (quotation_id,),
        ).fetchone()
        if not q:
            return (0, "")
        # order_no 시퀀스 생성 (SO-YYYYMM-####)
        # v5H226z263 (대표 지시·연결성 감사): 근본원인 = COUNT(*) 기반 채번은 중간 행 삭제 시
        #   번호를 재사용 → 기존 SO 번호와 충돌(orders.order_no UNIQUE 위반)·채번 실패.
        #   조치 = MAX(접미 일련번호)+1 로 바꿔 삭제분 재사용 차단 + UNIQUE 충돌 시 다음 번호로 재시도.
        import re as _re_so
        ym = datetime.now().strftime("%Y%m")
        _rows_so = c.execute(
            "SELECT order_no FROM orders WHERE order_no LIKE ?",
            (f"SO-{ym}-%",),
        ).fetchall()
        _seq = 0
        for _rso in _rows_so:
            _mso = _re_so.match(rf"^SO-{ym}-(\d+)$", str(_rso[0] or ""))
            if _mso:
                _seq = max(_seq, int(_mso.group(1)))
        _seq += 1
        order_id = None
        order_no = ""
        for _attempt in range(8):
            order_no = f"SO-{ym}-{_seq:04d}"
            try:
                cur = c.execute(
                    """INSERT INTO orders
                       (order_no, quote_id, customer_id, order_date, due_date,
                        total_amount, status, created_by)
                       VALUES (?,?,?,?,?,?,'CONFIRMED',?)""",
                    (order_no, q[0], q[1], date.today().isoformat(),
                     due_date, q[2] or 0, created_by or None),
                )
                order_id = cur.lastrowid
                break
            except sqlite3.IntegrityError as _ie_so:
                # v5H226z263: '수주번호(order_no) 중복'일 때만 다음 번호로 재시도.
                #   FK 등 다른 무결성 오류는 즉시 재발생시켜 표면화(엉뚱하게 '발급 충돌'로 가리지 않음).
                _msg = str(_ie_so).lower()
                if "order_no" in _msg or ("unique" in _msg and "order" in _msg):
                    _seq += 1
                    continue
                raise
        if order_id is None:
            raise RuntimeError(f"수주번호(SO) 발급 충돌 — 재시도 초과 (SO-{ym})")
        # 라인 복제 — quotation_items → order_items
        items = c.execute(
            """SELECT part_id, qty, unit_price, total_price
               FROM quotation_items WHERE quotation_id=?
               ORDER BY line_no ASC, id ASC""",
            (quotation_id,),
        ).fetchall()
        for it in items:
            c.execute(
                """INSERT INTO order_items
                   (order_id, part_id, qty, unit_price, amount, allocated_qty)
                   VALUES (?,?,?,?,?,0)""",
                (order_id, it[0], it[1] or 0, it[2] or 0, it[3] or 0),
            )
        return (order_id, order_no)


# =====================================================
# FTA 원산지증명서 헬퍼 (사이클 75 · 2026-04-27)
# 안지연 본업 — KAFTA/KEUFTA/RCEP 등 발급 모듈
# 외부 자산 0건 / SQL 파라미터 바인딩 의무
# =====================================================
def _next_fta_cert_no() -> str:
    """FTA-YYYY-#### 패턴 자동 생성. 현재 연도 내 최대치 +1."""
    from datetime import date as _date
    yr = _date.today().year
    prefix = f"FTA-{yr}-"
    with db_session() as c:
        row = c.execute(
            "SELECT cert_no FROM fta_certificates WHERE cert_no LIKE ? "
            "ORDER BY id DESC LIMIT 1",
            (f"{prefix}%",)
        ).fetchone()
    nxt = 1
    if row and row["cert_no"]:
        try:
            nxt = int(str(row["cert_no"]).split("-")[-1]) + 1
        except Exception:
            nxt = 1
    return f"{prefix}{nxt:04d}"


# =====================================================
# 작업 일정표(운영 보드) 헬퍼 — v5H226z264 (대표 지시)
#   · 진행기간 막대 '보드 전용 메모'(원본 무수정)
#   · 칸 너비·숨김 '내 계정' 저장
#   · 셀(기타사항·부서·담당자·납품위치) 원본 수정 — 화이트리스트만(임의 SQL 차단)
# =====================================================
def schedule_memos_map() -> dict:
    """{(ref_kind, ref_id): memo} — 일정표 진행기간 막대 메모 전체."""
    out = {}
    try:
        with db_session() as c:
            for r in c.execute("SELECT ref_kind, ref_id, memo FROM schedule_board_memos"):
                out[(r[0], int(r[1]))] = r[2] or ""
    except Exception:
        pass
    return out


def schedule_memo_set(ref_kind: str, ref_id: int, memo: str, user_id=None) -> bool:
    """진행기간 막대 보드메모 upsert (원본 데이터 무수정)."""
    if ref_kind not in ("project", "consumable") or not ref_id:
        return False
    from datetime import datetime as _dt
    now = _dt.now().isoformat(timespec="seconds")
    with db_session() as c:
        c.execute(
            """INSERT INTO schedule_board_memos (ref_kind, ref_id, memo, updated_by, updated_at)
               VALUES (?,?,?,?,?)
               ON CONFLICT(ref_kind, ref_id) DO UPDATE SET
                 memo=excluded.memo, updated_by=excluded.updated_by, updated_at=excluded.updated_at""",
            (ref_kind, int(ref_id), (memo or "").strip(), user_id, now))
    return True


def schedule_day_memos_map(ym: str = "") -> dict:
    """v5H226z712 (대표 지시): {(ref_kind, ref_id, ymd): memo} — 날짜별 진행 메모.
    ym('YYYY-MM') 주면 그 달만(빈 메모 제외). 작업일정표 달력 칸별 표시용."""
    out = {}
    try:
        with db_session() as c:
            if ym:
                rows = c.execute("SELECT ref_kind, ref_id, ymd, memo FROM schedule_board_day_memos "
                                 "WHERE ymd LIKE ? AND COALESCE(memo,'')<>''", (ym + "-%",))
            else:
                rows = c.execute("SELECT ref_kind, ref_id, ymd, memo FROM schedule_board_day_memos "
                                 "WHERE COALESCE(memo,'')<>''")
            for r in rows:
                out[(r[0], int(r[1]), r[2])] = r[3] or ""
    except Exception:
        pass
    return out


def schedule_day_memo_set(ref_kind: str, ref_id: int, ymd: str, memo: str, user_id=None) -> bool:
    """날짜별 진행 메모 upsert(빈 메모면 삭제). 원본 데이터 무수정."""
    if ref_kind not in ("project", "consumable") or not ref_id or not ymd:
        return False
    from datetime import datetime as _dt
    now = _dt.now().isoformat(timespec="seconds")
    _m = (memo or "").strip()
    with db_session() as c:
        if not _m:
            c.execute("DELETE FROM schedule_board_day_memos WHERE ref_kind=? AND ref_id=? AND ymd=?",
                      (ref_kind, int(ref_id), ymd))
        else:
            c.execute(
                """INSERT INTO schedule_board_day_memos (ref_kind, ref_id, ymd, memo, updated_by, updated_at)
                   VALUES (?,?,?,?,?,?)
                   ON CONFLICT(ref_kind, ref_id, ymd) DO UPDATE SET
                     memo=excluded.memo, updated_by=excluded.updated_by, updated_at=excluded.updated_at""",
                (ref_kind, int(ref_id), ymd, _m, user_id, now))
    return True


def view_prefs_get(user_id: int, view_key: str) -> str:
    """사용자별 화면 설정(JSON 문자열) 조회. 없으면 ''."""
    if not user_id or not view_key:
        return ""
    try:
        with db_session() as c:
            r = c.execute("SELECT config FROM user_view_prefs WHERE user_id=? AND view_key=?",
                          (int(user_id), view_key)).fetchone()
            return (r[0] or "") if r else ""
    except Exception:
        return ""


def view_prefs_set(user_id: int, view_key: str, config: str) -> bool:
    """사용자별 화면 설정(JSON 문자열) upsert."""
    if not user_id or not view_key:
        return False
    from datetime import datetime as _dt
    now = _dt.now().isoformat(timespec="seconds")
    with db_session() as c:
        c.execute(
            """INSERT INTO user_view_prefs (user_id, view_key, config, updated_at)
               VALUES (?,?,?,?)
               ON CONFLICT(user_id, view_key) DO UPDATE SET
                 config=excluded.config, updated_at=excluded.updated_at""",
            (int(user_id), view_key, config or "", now))
    return True


# 일정표 셀 → 원본 컬럼 매핑 (화이트리스트 — 이 외 필드/테이블은 절대 수정 안 함)
# v5H226z277 (대표 지시): 수량·단가·금액 추가. price/amount 는 라우트에서 can_view_sales 권한 게이트.
_SCHED_CELL_MAP = {
    "project":    {"name": "name",  # v5H226z365 (대표 지시): 프로젝트명 보드 인라인 편집(프로젝트만)
                   # v5H226z366 (대표 지시): 모델명·장비명·PO유형도 보드 인라인 편집(프로젝트만). 거래구분·고객사1은 특수처리(아래).
                   "model": "model_name", "equip": "equip_name", "po_type": "po_type",
                   "note": "logi_note", "dept": "cc_dept", "owner": "cc_name",
                   "qty": "unit_qty", "price": "unit_price",
                   # v5H226z282 (대표 지시): 고객사2·연락처·영업담당자·발주일·납품일 편집
                   "cust2": "secondary_customer", "contact": "cc_phone",
                   "sales_owner": "sales_name",
                   # v5H226z349 (대표 지시): 형태(제품/상품/기타) 보드 인라인 편집
                   "form_type": "form_type",
                   "order_date": "order_date", "due_date": "due_date"},
    "consumable": {"note": "note",      "dept": "cc_dept", "owner": "cc_name", "ship_to": "ship_to",
                   # 소모품은 고객사2·연락처·영업담당자 컬럼이 없어 발주일·납품일만
                   "order_date": "order_date", "due_date": "due_date",
                   # v5H226z367 (대표 지시): 거래명세서·세금계산서 발행일자(소모품도 포함). consumable_orders 컬럼 직접.
                   "statement_date": "statement_date", "tax_invoice_date": "tax_invoice_date"},
}
_SCHED_NUM_FIELDS = {"qty", "price", "amount"}  # 숫자 필드 (콤마 제거 후 저장)


def schedule_cell_update(ref_kind: str, ref_id: int, field: str, value: str) -> tuple[bool, str]:
    """일정표 정보칸(기타사항·부서·담당자·납품위치) 원본 수정.
    화이트리스트(_SCHED_CELL_MAP) 필드만 허용. 반환 (성공, 메시지).
    납품위치(ship_to)는 프로젝트의 경우 '대표 SO(orders 최신)'의 ship_to 를 수정."""
    if ref_kind not in _SCHED_CELL_MAP or not ref_id or not field:
        return (False, "허용되지 않은 대상/필드")
    val = (value or "").strip()
    # v5H226z365 (대표 지시): 프로젝트명은 빈 값으로 지우기 금지 — 클릭 실수로 이름이 사라지는 사고 방지(실패 표면화)
    if field == "name" and not val:
        return (False, "프로젝트명은 비울 수 없습니다")
    # v5H226z366 (대표 지시): 거래구분(수출/내수)=is_export 변환 — 프로젝트만
    if field == "trade":
        if ref_kind != "project":
            return (False, "거래구분은 프로젝트만 편집 가능")
        if val not in ("수출", "내수"):
            return (False, "거래구분은 수출/내수만 가능")
        with db_session() as c:
            c.execute("UPDATE projects SET is_export=? WHERE id=?",
                      (1 if val == "수출" else 0, int(ref_id)))
        return (True, "")
    # v5H226z366 (대표 지시): 고객사1 = customers 안전연결 — '정확히 1곳만 일치'할 때만 자동연결, 애매하면 미매칭(수동확인). 프로젝트만.
    if field == "cust":
        if ref_kind != "project":
            return (False, "고객사1은 프로젝트만 편집 가능")
        if not val:
            return (False, "고객사명은 비울 수 없습니다")
        # v5H226z651 (대표 지시): 안전연결 헬퍼로 통일 — 후보 1개만 자동연결, 여러 종사업장이면 기존 연결 유지(z647) 또는 미매칭.
        with db_session() as c:
            _ex = c.execute("SELECT customer_id FROM projects WHERE id=?", (int(ref_id),)).fetchone()
            _cid, _ = resolve_customer_id(c, val, prefer_id=(_ex[0] if _ex else None))
            c.execute("UPDATE projects SET customer_name=?, customer_id=? WHERE id=?",
                      (val, _cid, int(ref_id)))
        return (True, "matched" if _cid else "unmatched")  # 호출측(엔드포인트)이 연결여부 판단에 사용
    # v5H226z366 (대표 지시): PO유형은 고정 선택지만(빈값=해제 허용). 저장은 아래 일반경로(화이트리스트 컬럼).
    if field == "po_type" and val and val not in PO_TYPES:
        return (False, "PO유형은 신규/추가/개조/수리/기타 중에서만 선택")
    # v5H226z349 (대표 지시): 형태는 고정 선택지(제품/상품/기타)만 — 그 외 값 차단(빈값=해제 허용)
    if field == "form_type" and val and val not in FORM_TYPES:
        return (False, "형태는 제품/상품/기타 중에서만 선택")
    # v5H226z350 (대표 지시): 형태 편집은 출고형태(shipment_form)와 동기화 — 상품→PARTS(부품 상세·PACKING LIST).
    #   프로젝트만 해당(소모품은 form_type 편집 비대상).
    if ref_kind == "project" and field == "form_type":
        _sf = FORM_TO_SHIP.get(val, "ASSEMBLY") if val else "ASSEMBLY"
        with db_session() as c:
            c.execute("UPDATE projects SET form_type=?, shipment_form=? WHERE id=?",
                      (val, _sf, int(ref_id)))
        return (True, "")
    # v5H226z277: 숫자 필드(수량·단가·금액)는 콤마 제거 후 숫자로
    if field in _SCHED_NUM_FIELDS:
        _raw = val.replace(",", "").strip()
        try:
            _num = float(_raw) if _raw else 0
        except ValueError:
            return (False, "숫자만 입력")
        # 금액(project) = order_amount + expected_amount, 단일 SO면 그 SO total_amount 도 일치(상세 자가치유 대비)
        if ref_kind == "project" and field == "amount":
            with db_session() as c:
                c.execute("UPDATE projects SET order_amount=?, expected_amount=? WHERE id=?",
                          (_num, _num, int(ref_id)))
                _so = c.execute("SELECT COUNT(*), MAX(id) FROM orders WHERE project_id=?",
                                (int(ref_id),)).fetchone()
                if _so and _so[0] == 1 and _so[1]:
                    c.execute("UPDATE orders SET total_amount=? WHERE id=?", (_num, int(_so[1])))
            return (True, "")
        # 수량·단가는 화이트리스트 컬럼에 숫자 저장
        _col = _SCHED_CELL_MAP[ref_kind].get(field)
        if not _col:
            return (False, "허용되지 않은 필드")
        _store = int(_num) if field == "qty" else _num
        # v5H226z366 리뷰반영: 대상 테이블을 ref_kind로 동적 선택('projects' 하드코딩 제거) — 향후 소모품 숫자필드 추가 시 엉뚱한 테이블 수정 방지(데이터 안전)
        _ntable = "projects" if ref_kind == "project" else "consumable_orders"
        with db_session() as c:
            _ncols = {r[1] for r in c.execute(f"PRAGMA table_info({_ntable})").fetchall()}
            if _col not in _ncols:
                return (False, f"{_ntable}.{_col} 컬럼 없음")
            c.execute(f"UPDATE {_ntable} SET {_col}=? WHERE id=?", (_store, int(ref_id)))
        return (True, "")
    # 프로젝트 납품위치 = 대표 SO(orders 최신 id)의 ship_to 수정 (별도 처리)
    if ref_kind == "project" and field == "ship_to":
        with db_session() as c:
            _ocols = {r[1] for r in c.execute("PRAGMA table_info(orders)").fetchall()}
            if "ship_to" not in _ocols:
                return (False, "orders.ship_to 컬럼 없음")
            r = c.execute("SELECT id FROM orders WHERE project_id=? ORDER BY id DESC LIMIT 1",
                          (int(ref_id),)).fetchone()
            if not r:
                return (False, "이 프로젝트에 수주(SO)가 없어 납품위치를 저장할 수 없습니다")
            c.execute("UPDATE orders SET ship_to=? WHERE id=?", (val, int(r[0])))
        return (True, "")
    # v5H226z467 (대표 지시): 세금계산서 3차 분할(계약금/중도금/잔금) — 발행일(1/2/3차)+금액(1/2/3차).
    #   단일출처: 프로젝트=대표 SO(orders 최신) / 소모품=consumable_orders. 1차 발행일=기존 tax_invoice_date.
    _TAXINV_FIELDS = {
        "tax_invoice_date":  ("tax_invoice_date",  "date"),
        "tax_invoice_date2": ("tax_invoice_date2", "date"),
        "tax_invoice_date3": ("tax_invoice_date3", "date"),
        "tax_invoice_amt1":  ("tax_invoice_amt1",  "num"),
        "tax_invoice_amt2":  ("tax_invoice_amt2",  "num"),
        "tax_invoice_amt3":  ("tax_invoice_amt3",  "num"),
    }
    if field in _TAXINV_FIELDS:
        _col, _ftype = _TAXINV_FIELDS[field]
        # v5H226z562 (대표 지시): 프로젝트 세금계산서 단일 진실 = 호기(order_items). 이 경로(일괄수정 등)도 호기에 적용.
        #   날짜=호기 전체 동일 / 금액=호기수 균등분배(합계 보존). 소모품은 호기 개념 없어 consumable_orders 그대로.
        if ref_kind == "project":
            with db_session() as c:
                _oic = {r[1] for r in c.execute("PRAGMA table_info(order_items)").fetchall()}
                if _col not in _oic:
                    return (False, f"order_items.{_col} 컬럼 없음")
                _ids = [int(x[0]) for x in c.execute(
                    "SELECT oi.id FROM order_items oi JOIN orders o ON o.id=oi.order_id "
                    "WHERE o.project_id=? AND COALESCE(o.status,'')<>'CANCELLED' ORDER BY oi.id", (int(ref_id),)).fetchall()]
                if not _ids:
                    return (False, "이 프로젝트에 수주(SO) 호기가 없어 세금계산서를 저장할 수 없습니다")
                if _ftype == "date":
                    _store = val or None
                    _ph = ",".join("?" * len(_ids))
                    c.execute(f"UPDATE order_items SET {_col}=? WHERE id IN ({_ph})", [_store] + _ids)
                else:
                    _raw = val.replace(",", "").strip()
                    if _raw == "":
                        _ph = ",".join("?" * len(_ids))
                        c.execute(f"UPDATE order_items SET {_col}=NULL WHERE id IN ({_ph})", _ids)
                    else:
                        try:
                            _tot = float(_raw)
                        except ValueError:
                            return (False, "숫자만 입력")
                        _n = len(_ids); _share = round(_tot / _n, 2); _rem = round(_tot - _share * _n, 2)
                        for _i, _iid in enumerate(_ids):
                            c.execute(f"UPDATE order_items SET {_col}=? WHERE id=?", (_share + (_rem if _i == 0 else 0), _iid))
            return (True, "")
        # 소모품 = consumable_orders (기존 유지)
        if _ftype == "date":
            _store = val or None   # 빈값 = 발행 취소(NULL)
        else:
            _raw = val.replace(",", "").strip()
            try:
                _store = float(_raw) if _raw else None
            except ValueError:
                return (False, "숫자만 입력")
        with db_session() as c:
            _tcols = {r[1] for r in c.execute("PRAGMA table_info(consumable_orders)").fetchall()}
            if _col not in _tcols:
                return (False, f"consumable_orders.{_col} 컬럼 없음")
            _tid = int(ref_id)
            c.execute(f"UPDATE consumable_orders SET {_col}=? WHERE id=?", (_store, _tid))
            if _col == "tax_invoice_date" and "tax_invoice_issued" in _tcols:
                c.execute("UPDATE consumable_orders SET tax_invoice_issued=? WHERE id=?",
                          (1 if _store else 0, _tid))
        return (True, "")
    # v5H226z367/z562 (대표 지시): 거래명세서 발행일자 = 프로젝트 호기(order_items) 전체에 동일 적용(단일 진실).
    if ref_kind == "project" and field == "statement_date":
        _v = val or None   # 빈값 = 발행 취소(NULL)
        with db_session() as c:
            _oic = {r[1] for r in c.execute("PRAGMA table_info(order_items)").fetchall()}
            if "statement_date" not in _oic:
                return (False, "order_items.statement_date 컬럼 없음")
            _ids = [int(x[0]) for x in c.execute(
                "SELECT oi.id FROM order_items oi JOIN orders o ON o.id=oi.order_id "
                "WHERE o.project_id=? AND COALESCE(o.status,'')<>'CANCELLED' ORDER BY oi.id", (int(ref_id),)).fetchall()]
            if not _ids:
                return (False, "이 프로젝트에 수주(SO) 호기가 없어 발행일자를 저장할 수 없습니다")
            _ph = ",".join("?" * len(_ids))
            c.execute(f"UPDATE order_items SET statement_date=? WHERE id IN ({_ph})", [_v] + _ids)
        return (True, "")
    col = _SCHED_CELL_MAP[ref_kind].get(field)
    if not col:
        return (False, "허용되지 않은 필드")
    table = "projects" if ref_kind == "project" else "consumable_orders"
    with db_session() as c:
        # 컬럼 존재 확인(추가형 마이그레이션 전 환경 방어)
        _cols = {r[1] for r in c.execute(f"PRAGMA table_info({table})").fetchall()}
        if col not in _cols:
            return (False, f"{table}.{col} 컬럼 없음")
        # v5H226z367 리뷰반영: 발행일자(거래명세서·세금계산서)는 빈값=미발행이므로 NULL로 저장(프로젝트 경로와 동일·''/NULL 혼재 방지)
        _sv = (val or None) if field in ("statement_date", "tax_invoice_date") else val
        c.execute(f"UPDATE {table} SET {col}=? WHERE id=?", (_sv, int(ref_id)))
    return (True, "")


# =====================================================
# 작업 일정표 = 업무 원장 허브 — 단계 파이프라인 — v5H226z270 (대표 지시)
#   설계: _기획/작업일정표_원장허브_설계_v1.md
#   원칙: 누구나 클릭(담당은 안내) · 누른 사람/시각 기록 · 단계 화이트리스트 · 금액류 권한
# =====================================================
# 단계 정의(순서·라벨·담당안내·그룹). progress 는 동시진행 6 세부.
STAGE_SPEC = [
    {"key": "input",        "label": "입력",          "owner": "기술영업",      "group": "영업"},
    {"key": "quote",        "label": "견적·제안 전달", "owner": "기술영업",      "group": "영업"},
    {"key": "order",        "label": "수주",          "owner": "기술영업",      "group": "영업"},
    {"key": "prod_request", "label": "제작요청",      "owner": "기술영업→전부서", "group": "생산"},
    {"key": "progress",     "label": "진행",          "owner": "생산",          "group": "생산",
     "subs": [("design", "설계"), ("circuit", "회로"), ("program", "프로그램"),
              ("wiring", "전장"), ("assembly", "조립"), ("test", "테스트")]},
    {"key": "verify",       "label": "검증",          "owner": "생산",          "group": "생산"},
    {"key": "label",        "label": "라벨/명판",     "owner": "기술영업",      "group": "생산"},
    {"key": "pack",         "label": "포장",          "owner": "생산",          "group": "생산"},
    {"key": "ship",         "label": "출하",          "owner": "기술영업",      "group": "생산"},
    {"key": "tax_invoice",  "label": "세금계산서",    "owner": "영업·회계",     "group": "매출",
     "sales_only": True},   # 금액·세금계산서 = 권한(영업·관리)만
]
# 화이트리스트 (검증용)
STAGE_KEYS = {s["key"] for s in STAGE_SPEC}
STAGE_SUBS = {s["key"]: {k for k, _ in s.get("subs", [])} for s in STAGE_SPEC}

# =====================================================
# v5H226z369 (대표 지시): 단계 '담당'을 실제 부서명 · 사업부별로 표기.
#   div = 관리번호 4번째 글자(T 검사기 / M 자동화 / L 라이프밸류 / E 기타 / C 소모품).
#   기타(E)=담당 표시 안 함(빈값). 소모품(C)=전 단계 기술영업팀(세금계산서 발행도 기술영업팀).
#   진행(progress) 단계는 헤더 담당 비우고 6세부(설계·회로·프로그램·전장·조립·테스트)에 부서를 실어 표시.
#   ※ 검사기/자동화/라이프밸류 일부는 초안 — 화면 보며 대표 지시로 수정 예정.
# =====================================================
_STG_OWNER_COMMON = {"input": "기술영업팀", "quote": "기술영업팀", "order": "기술영업팀",
                     "prod_request": "기술영업팀→생산", "progress": "",
                     "label": "기술영업팀", "ship": "기술영업팀", "tax_invoice": "관리팀"}
STAGE_OWNERS_BY_DIV = {
    "T": {**_STG_OWNER_COMMON, "verify": "품질팀",       "pack": "제조기술1팀"},
    "M": {**_STG_OWNER_COMMON, "verify": "품질팀",       "pack": "제조기술2팀"},
    "L": {**_STG_OWNER_COMMON, "verify": "라이프밸류팀", "pack": "가공팀"},
    # 소모품: 전 단계 기술영업팀(생산세부 없음·세금계산서도 기술영업팀)
    "C": {k: "기술영업팀" for k in ("input", "quote", "order", "prod_request", "progress",
                                  "verify", "label", "pack", "ship", "tax_invoice")},
    "E": {},   # 기타: 담당 표시 안 함
}
STAGE_SUB_OWNERS_BY_DIV = {
    "T": {"design": "설계팀", "circuit": "설계팀", "program": "개발혁신팀",
          "wiring": "제조기술1팀", "assembly": "제조기술1팀", "test": "품질팀"},
    "M": {"design": "설계팀", "circuit": "전장설계팀", "program": "소프트웨어팀",
          "wiring": "전장설계팀", "assembly": "제조기술2팀", "test": "제조기술2팀"},
    "L": {k: "라이프밸류팀" for k in ("design", "circuit", "program", "wiring", "assembly", "test")},
    "C": {k: "기술영업팀" for k in ("design", "circuit", "program", "wiring", "assembly", "test")},
    "E": {},
}


def stage_spec_for_div(div: str) -> list:
    """사업부(div)에 맞는 담당(owner) + 진행 6세부 담당을 채운 STAGE_SPEC 사본.
    E(기타)=담당 빈값. 미지정 div=기존 일반 담당 유지(안전 폴백). subs는 [key,label,owner] 3요소."""
    div = (div or "").upper()
    omap = STAGE_OWNERS_BY_DIV.get(div)
    smap = STAGE_SUB_OWNERS_BY_DIV.get(div, {})
    out = []
    for s in STAGE_SPEC:
        if div == "E":
            owner = ""
        elif omap is not None and s["key"] in omap:
            owner = omap[s["key"]]
        else:
            owner = s["owner"]   # 미지정 div → 기존 일반 담당 유지
        out.append({
            "key": s["key"], "label": s["label"], "owner": owner,
            "group": s["group"], "sales_only": bool(s.get("sales_only")),
            "subs": [[sk, sl, ("" if div == "E" else smap.get(sk, ""))] for sk, sl in s.get("subs", [])],
        })
    return out


def stage_rows_for(ref_kind: str, ref_id: int) -> dict:
    """한 건의 단계 기록 → {(stage_key, sub_key): dict}."""
    out = {}
    if ref_kind not in ("project", "consumable") or not ref_id:
        return out
    try:
        with db_session() as c:
            # v5H226z477: 착수(시작전→진행중→완료) + 소요시간(h) 포함
            for r in c.execute(
                "SELECT stage_key, sub_key, done, on_date, by_user, by_name, memo, extra, updated_at, "
                "       started_at, started_by_name, hours "
                "FROM project_stage_log WHERE ref_kind=? AND ref_id=?", (ref_kind, int(ref_id))):
                _started = r[9] or ""
                _done = bool(r[2])
                out[(r[0], r[1] or "")] = {
                    "done": _done, "on_date": r[3] or "", "by_user": r[4],
                    "by_name": r[5] or "", "memo": r[6] or "", "extra": r[7] or "",
                    "updated_at": r[8] or "",
                    "started_at": _started, "started_by_name": r[10] or "",
                    "hours": (r[11] if r[11] is not None else ""),
                    # 상태: 완료 / 진행중(착수했고 미완료) / 시작전
                    "state": ("done" if _done else ("doing" if _started else "todo"))}
    except Exception:
        pass
    return out


def stage_set(ref_kind: str, ref_id: int, stage_key: str, sub_key: str = "",
              done=True, on_date: str = "", memo: str = "", extra: str = "",
              user_id=None, user_name: str = "", started: bool = False, hours=None) -> tuple[bool, str]:
    """단계 기록 upsert. 단계/세부는 화이트리스트만. 반환 (성공, 메시지).
    v5H226z477: started=True → '착수'(진행중) 기록(완료정보 미변경) / 그 외 → 완료/해제(+소요시간 hours).
    """
    if ref_kind not in ("project", "consumable") or not ref_id:
        return (False, "대상 오류")
    if stage_key not in STAGE_KEYS:
        return (False, "허용되지 않은 단계")
    sub = (sub_key or "").strip()
    if sub and sub not in STAGE_SUBS.get(stage_key, set()):
        return (False, "허용되지 않은 세부단계")
    from datetime import datetime as _dt
    now = _dt.now().isoformat(timespec="seconds")
    try:
        with db_session() as c:
            if started:
                # v5H226z477: 착수(진행중) — done/on_date/memo 등 완료정보는 손대지 않음. 최초 착수시각만 보존.
                c.execute(
                    """INSERT INTO project_stage_log
                       (ref_kind, ref_id, stage_key, sub_key, done, started_at, started_by, started_by_name, updated_at)
                       VALUES (?,?,?,?,0,?,?,?,?)
                       ON CONFLICT(ref_kind, ref_id, stage_key, sub_key) DO UPDATE SET
                         started_at=COALESCE(NULLIF(started_at,''), excluded.started_at),
                         started_by=excluded.started_by, started_by_name=excluded.started_by_name,
                         updated_at=excluded.updated_at""",
                    (ref_kind, int(ref_id), stage_key, sub, now, user_id, (user_name or "").strip(), now))
            else:
                try:
                    _h = float(hours) if hours not in (None, "") else None
                except Exception:
                    _h = None
                c.execute(
                    """INSERT INTO project_stage_log
                       (ref_kind, ref_id, stage_key, sub_key, done, on_date, by_user, by_name, memo, extra, hours, started_at, updated_at)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                       ON CONFLICT(ref_kind, ref_id, stage_key, sub_key) DO UPDATE SET
                         done=excluded.done, on_date=excluded.on_date, by_user=excluded.by_user,
                         by_name=excluded.by_name, memo=excluded.memo, extra=excluded.extra,
                         hours=excluded.hours,
                         started_at=COALESCE(NULLIF(started_at,''), excluded.started_at),
                         updated_at=excluded.updated_at""",
                    (ref_kind, int(ref_id), stage_key, sub, 1 if done else 0, (on_date or "").strip(),
                     user_id, (user_name or "").strip(), (memo or "").strip(), (extra or "").strip(),
                     _h, now, now))
        return (True, "")
    except Exception as e:
        return (False, str(e)[:160])


def stage_board_map() -> dict:
    """보드 단계 칸용 요약 — {(ref_kind, ref_id): {done:set(stage_key), prog:{sub:bool}, ship_date:str}}.
    완료된 단계 key 집합 + progress 세부 + 출하일(있으면)만 가볍게."""
    out = {}
    try:
        with db_session() as c:
            for r in c.execute(
                "SELECT ref_kind, ref_id, stage_key, sub_key, done, on_date "
                "FROM project_stage_log WHERE done=1"):
                k = (r[0], int(r[1]))
                d = out.setdefault(k, {"done": set(), "prog": {}, "ship_date": ""})
                if r[2] == "progress" and r[3]:
                    d["prog"][r[3]] = True
                else:
                    d["done"].add(r[2])
                if r[2] == "ship" and r[5]:
                    d["ship_date"] = r[5]
    except Exception:
        pass
    return out


def stage_log_all_map() -> dict:
    """v5H226z478 (부서 작업 큐용): 전체 단계기록을 한 번에 →
    {(ref_kind, ref_id, stage_key, sub_key): {state, on_date, by_name, started_at, started_by_name, hours, memo}}.
    state = done / doing(착수·미완) / todo. (행이 없으면 호출측에서 todo로 간주)"""
    out = {}
    try:
        with db_session() as c:
            for r in c.execute(
                "SELECT ref_kind, ref_id, stage_key, sub_key, done, on_date, by_name, "
                "       started_at, started_by_name, hours, memo FROM project_stage_log"):
                _state = "done" if r[4] else ("doing" if r[7] else "todo")
                out[(r[0], int(r[1]), r[2], r[3] or "")] = {
                    "state": _state, "on_date": r[5] or "", "by_name": r[6] or "",
                    "started_at": r[7] or "", "started_by_name": r[8] or "",
                    "hours": (r[9] if r[9] is not None else ""), "memo": r[10] or ""}
    except Exception:
        pass
    return out


def create_fta_certificate(
    fta_type: str,
    customer_id: int = None,
    customer_name: str = None,
    customer_address: str = None,
    customer_country: str = None,
    export_order_id: int = None,
    export_invoice_no: str = None,
    export_date: str = None,
    origin_country: str = "KR",
    total_value: float = 0,
    currency: str = "USD",
    issuer_id: int = None,
    issuer_name: str = None,
    items: list = None,
    remarks: str = None,
    created_by: int = None,
) -> tuple:
    """원산지증명서 신규 발급. (cert_id, cert_no) 반환.
    items: [{part_id, part_name, hs_code, qty, unit, unit_price, origin_country, total}]
    """
    cert_no = _next_fta_cert_no()
    with db_session() as c:
        cur = c.execute(
            """INSERT INTO fta_certificates
               (cert_no, fta_type, customer_id, customer_name, customer_address,
                customer_country, export_order_id, export_invoice_no, export_date,
                origin_country, total_value, currency, issuer_id, issuer_name,
                issued_at, status, remarks, created_by)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,
                       datetime('now','localtime'), 'DRAFT', ?, ?)""",
            (cert_no, fta_type, customer_id, customer_name, customer_address,
             customer_country, export_order_id, export_invoice_no, export_date,
             origin_country, total_value or 0, currency or "USD",
             issuer_id, issuer_name, remarks, created_by),
        )
        cert_id = cur.lastrowid
        if items:
            for idx, it in enumerate(items, start=1):
                c.execute(
                    """INSERT INTO fta_certificate_items
                       (cert_id, line_no, part_id, part_name, hs_code,
                        qty, unit, unit_price, origin_country, total)
                       VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (cert_id, idx,
                     it.get("part_id"), it.get("part_name"),
                     it.get("hs_code"),
                     float(it.get("qty") or 0),
                     it.get("unit"),
                     float(it.get("unit_price") or 0),
                     it.get("origin_country") or origin_country,
                     float(it.get("total") or 0)),
                )
    return (cert_id, cert_no)


def get_fta_certificates(
    status: str = None,
    fta_type: str = None,
    customer_id: int = None,
    limit: int = 200,
) -> list:
    """원산지증명서 목록 (필터 옵션). 최신순."""
    sql = (
        "SELECT fc.*, COALESCE(cu.name, fc.customer_name, '-') AS cust_disp "
        "FROM fta_certificates fc "
        "LEFT JOIN customers cu ON cu.id = fc.customer_id WHERE 1=1"
    )
    args = []
    if status:
        sql += " AND fc.status=?"
        args.append(status)
    if fta_type:
        sql += " AND fc.fta_type=?"
        args.append(fta_type)
    if customer_id:
        sql += " AND fc.customer_id=?"
        args.append(int(customer_id))
    sql += " ORDER BY fc.id DESC LIMIT ?"
    args.append(int(limit))
    with db_session() as c:
        rows = c.execute(sql, tuple(args)).fetchall()
        return [dict(r) for r in rows]


def get_fta_certificate(cert_id: int) -> dict:
    """원산지증명서 상세 + 라인 반환. 없으면 None."""
    with db_session() as c:
        row = c.execute(
            "SELECT fc.*, COALESCE(cu.name, fc.customer_name, '-') AS cust_disp "
            "FROM fta_certificates fc "
            "LEFT JOIN customers cu ON cu.id = fc.customer_id "
            "WHERE fc.id=?",
            (int(cert_id),)
        ).fetchone()
        if not row:
            return None
        cert = dict(row)
        items = c.execute(
            "SELECT fi.*, COALESCE(p.name, fi.part_name, '-') AS part_disp "
            "FROM fta_certificate_items fi "
            "LEFT JOIN parts p ON p.id = fi.part_id "
            "WHERE fi.cert_id=? ORDER BY fi.line_no ASC, fi.id ASC",
            (int(cert_id),)
        ).fetchall()
        cert["items"] = [dict(r) for r in items]
        return cert


# =====================================================
# QC INSPECTION REPORT 헬퍼 (사이클 76 · 2026-04-27)
# 김정록 본업 — 검사기 출하성적서 (QCR-YYYY-####) 발급 모듈
# 외부 자산 0건 / SQL 파라미터 바인딩 의무
# =====================================================
QC_STANDARD_ITEMS = [
    ("반복성",   "≤ 0.5 μm (3σ)"),
    ("정확도",   "100 ± 0.1 mm"),
    ("통신",     "Modbus/TCP RTT < 50 ms"),
    ("외관",     "도장·라벨·결함 없음"),
    ("동작",     "전 사이클 정상 동작"),
    ("안전",     "EMC·접지·인터록 OK"),
]


def _next_qc_report_no() -> str:
    """QCR-YYYY-#### 패턴 자동 생성. 현재 연도 내 최대치 +1."""
    from datetime import date as _date
    yr = _date.today().year
    prefix = f"QCR-{yr}-"
    with db_session() as c:
        row = c.execute(
            "SELECT report_no FROM qc_inspection_reports WHERE report_no LIKE ? "
            "ORDER BY id DESC LIMIT 1",
            (f"{prefix}%",)
        ).fetchone()
    nxt = 1
    if row and row["report_no"]:
        try:
            nxt = int(str(row["report_no"]).split("-")[-1]) + 1
        except Exception:
            nxt = 1
    return f"{prefix}{nxt:04d}"


def create_qc_inspection_report(
    customer_id: int = None,
    customer_name: str = None,
    order_id: int = None,
    order_no: str = None,
    part_id: int = None,
    machine_model: str = None,
    machine_serial: str = None,
    inspection_date: str = None,
    inspector_id: int = None,
    inspector_name: str = None,
    qa_manager_id: int = None,
    qa_manager_name: str = None,
    overall: str = "PASS",
    items: list = None,
    remarks: str = None,
    created_by: int = None,
) -> tuple:
    """검사기 출하성적서 신규 발급. (report_id, report_no) 반환.
    items: [{item_name, spec_value, measured_value, judgment, remarks}]
    """
    report_no = _next_qc_report_no()
    if overall not in ("PASS", "FAIL", "CONDITIONAL_PASS"):
        overall = "PASS"
    with db_session() as c:
        cur = c.execute(
            """INSERT INTO qc_inspection_reports
               (report_no, customer_id, customer_name, order_id, order_no,
                part_id, machine_model, machine_serial, inspection_date,
                inspector_id, inspector_name, qa_manager_id, qa_manager_name,
                overall, status, remarks, created_by)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?, 'DRAFT', ?, ?)""",
            (report_no, customer_id, customer_name, order_id, order_no,
             part_id, machine_model, machine_serial, inspection_date,
             inspector_id, inspector_name, qa_manager_id, qa_manager_name,
             overall, remarks, created_by),
        )
        report_id = cur.lastrowid
        if items:
            for idx, it in enumerate(items, start=1):
                jdg = (it.get("judgment") or "PASS").upper()
                if jdg not in ("PASS", "FAIL", "NA"):
                    jdg = "PASS"
                c.execute(
                    """INSERT INTO qc_inspection_items
                       (report_id, line_no, item_name, spec_value,
                        measured_value, judgment, remarks)
                       VALUES (?,?,?,?,?,?,?)""",
                    (report_id, idx,
                     (it.get("item_name") or "").strip(),
                     it.get("spec_value"),
                     it.get("measured_value"),
                     jdg,
                     it.get("remarks")),
                )
    return (report_id, report_no)


def get_qc_inspection_reports(
    status: str = None,
    overall: str = None,
    customer_id: int = None,
    limit: int = 200,
) -> list:
    """검사기 출하성적서 목록 (필터 옵션). 최신순."""
    sql = (
        "SELECT qr.*, COALESCE(cu.name, qr.customer_name, '-') AS cust_disp, "
        "       COALESCE(us.name, qr.inspector_name, '-') AS inspector_disp "
        "FROM qc_inspection_reports qr "
        "LEFT JOIN customers cu ON cu.id = qr.customer_id "
        "LEFT JOIN users us ON us.id = qr.inspector_id WHERE 1=1"
    )
    args = []
    if status:
        sql += " AND qr.status=?"
        args.append(status)
    if overall:
        sql += " AND qr.overall=?"
        args.append(overall)
    if customer_id:
        sql += " AND qr.customer_id=?"
        args.append(int(customer_id))
    sql += " ORDER BY qr.id DESC LIMIT ?"
    args.append(int(limit))
    with db_session() as c:
        rows = c.execute(sql, tuple(args)).fetchall()
        return [dict(r) for r in rows]


def get_qc_inspection_report(report_id: int) -> dict:
    """검사기 출하성적서 상세 + 라인 반환. 없으면 None."""
    with db_session() as c:
        row = c.execute(
            "SELECT qr.*, COALESCE(cu.name, qr.customer_name, '-') AS cust_disp, "
            "       COALESCE(us.name, qr.inspector_name, '-') AS inspector_disp, "
            "       COALESCE(uq.name, qr.qa_manager_name, '-') AS qa_disp, "
            "       COALESCE(p.part_name, '-') AS part_disp "
            "FROM qc_inspection_reports qr "
            "LEFT JOIN customers cu ON cu.id = qr.customer_id "
            "LEFT JOIN users us ON us.id = qr.inspector_id "
            "LEFT JOIN users uq ON uq.id = qr.qa_manager_id "
            "LEFT JOIN parts p ON p.id = qr.part_id "
            "WHERE qr.id=?",
            (int(report_id),)
        ).fetchone()
        if not row:
            return None
        rep = dict(row)
        items = c.execute(
            "SELECT * FROM qc_inspection_items "
            "WHERE report_id=? ORDER BY line_no ASC, id ASC",
            (int(report_id),)
        ).fetchall()
        rep["items"] = [dict(r) for r in items]
        return rep


# =====================================================
# WORK ORDERS (가공팀 작업지시서) 헬퍼 (2026-04-27 사이클77)
# 04 시뮬 MISSING #3 — 윤영조·이수빈 본업 모듈
# =====================================================

def _next_wo_no() -> str:
    """WO-YYYY-#### 패턴 자동 생성. 현재 연도 내 최대치 +1."""
    from datetime import date as _date
    yr = _date.today().year
    prefix = f"WO-{yr}-"
    with db_session() as c:
        row = c.execute(
            "SELECT wo_no FROM work_orders WHERE wo_no LIKE ? "
            "ORDER BY id DESC LIMIT 1",
            (f"{prefix}%",)
        ).fetchone()
    nxt = 1
    if row and row["wo_no"]:
        try:
            nxt = int(str(row["wo_no"]).split("-")[-1]) + 1
        except Exception:
            nxt = 1
    return f"{prefix}{nxt:04d}"


def create_work_order(
    order_id: int = None,
    project_id: int = None,
    part_id: int = None,
    qty: float = 0,
    assigned_to: int = None,
    assigned_name: str = None,
    created_by: int = None,
    created_by_name: str = None,
    planned_start: str = None,
    planned_end: str = None,
    specifications: str = None,
    items: list = None,
    remarks: str = None,
) -> tuple:
    """가공팀 작업지시서 신규 발급. (wo_id, wo_no) 반환.
    items: [{step_name, duration_min, worker_name, remarks}]
    """
    wo_no = _next_wo_no()
    with db_session() as c:
        cur = c.execute(
            """INSERT INTO work_orders
               (wo_no, order_id, project_id, part_id, qty,
                assigned_to, assigned_name, created_by, created_by_name,
                planned_start, planned_end, specifications,
                status, remarks)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?, 'DRAFT', ?)""",
            (wo_no, order_id, project_id, part_id, qty,
             assigned_to, assigned_name, created_by, created_by_name,
             planned_start, planned_end, specifications, remarks),
        )
        wo_id = cur.lastrowid
        if items:
            for idx, it in enumerate(items, start=1):
                step = (it.get("step_name") or "").strip()
                if not step:
                    continue
                try:
                    dur = int(it.get("duration_min") or 0)
                except Exception:
                    dur = 0
                try:
                    prog = int(it.get("progress") or 0)
                except Exception:
                    prog = 0
                if prog < 0:
                    prog = 0
                if prog > 100:
                    prog = 100
                c.execute(
                    """INSERT INTO work_order_items
                       (wo_id, line_no, step_name, duration_min, progress,
                        worker_id, worker_name, remarks)
                       VALUES (?,?,?,?,?,?,?,?)""",
                    (wo_id, idx, step, dur, prog,
                     it.get("worker_id"),
                     (it.get("worker_name") or "").strip() or None,
                     it.get("remarks")),
                )
    return (wo_id, wo_no)


def get_work_orders(status: str = None, limit: int = 200) -> list:
    """가공팀 작업지시서 목록 (상태 필터). 최신순."""
    sql = (
        "SELECT w.*, COALESCE(o.order_no,'-') AS order_no_disp, "
        "       COALESCE(p.part_name,'-') AS part_disp,"
        "       COALESCE(ua.name, w.assigned_name, '-') AS assigned_disp, "
        "       COALESCE(uc.name, w.created_by_name, '-') AS created_by_disp "
        "FROM work_orders w "
        "LEFT JOIN orders o ON o.id = w.order_id "
        "LEFT JOIN parts p ON p.id = w.part_id "
        "LEFT JOIN users ua ON ua.id = w.assigned_to "
        "LEFT JOIN users uc ON uc.id = w.created_by "
        "WHERE 1=1"
    )
    args = []
    if status:
        sql += " AND w.status=?"
        args.append(status)
    sql += " ORDER BY w.id DESC LIMIT ?"
    args.append(int(limit))
    with db_session() as c:
        rows = c.execute(sql, tuple(args)).fetchall()
        return [dict(r) for r in rows]


def get_work_order(wo_id: int) -> dict:
    """가공팀 작업지시서 상세 + 라인 반환. 없으면 None."""
    with db_session() as c:
        row = c.execute(
            "SELECT w.*, COALESCE(o.order_no,'-') AS order_no_disp, "
            "       COALESCE(p.part_name,'-') AS part_disp,"
            "       COALESCE(p.spec,'-') AS part_spec, COALESCE(p.part_no,'-') AS part_no,"
            "       COALESCE(p.unit,'-') AS part_unit, "
            "       COALESCE(ua.name, w.assigned_name, '-') AS assigned_disp, "
            "       COALESCE(uc.name, w.created_by_name, '-') AS created_by_disp, "
            "       COALESCE(pr.name, '-') AS project_disp "
            "FROM work_orders w "
            "LEFT JOIN orders o ON o.id = w.order_id "
            "LEFT JOIN projects pr ON pr.id = w.project_id "
            "LEFT JOIN parts p ON p.id = w.part_id "
            "LEFT JOIN users ua ON ua.id = w.assigned_to "
            "LEFT JOIN users uc ON uc.id = w.created_by "
            "WHERE w.id=?",
            (int(wo_id),)
        ).fetchone()
        if not row:
            return None
        wo = dict(row)
        items = c.execute(
            "SELECT wi.*, COALESCE(uw.name, wi.worker_name, '-') AS worker_disp "
            "FROM work_order_items wi "
            "LEFT JOIN users uw ON uw.id = wi.worker_id "
            "WHERE wi.wo_id=? ORDER BY wi.line_no ASC, wi.id ASC",
            (int(wo_id),)
        ).fetchall()
        wo["items"] = [dict(r) for r in items]
        # 평균 진행률 계산
        if wo["items"]:
            total = sum(int(it.get("progress") or 0) for it in wo["items"])
            wo["avg_progress"] = total // len(wo["items"])
        else:
            wo["avg_progress"] = 0
        return wo


# =====================================================
# v5H45 (2026-05-03) — 비즈니스 데이터 종합 시드
# 대표 지시: "전 아이콘 항목 내용 자동 보충".
# 대상 테이블 (빈 페이지 원인): parts, suppliers, purchase_orders, po_items,
# stock_movements, receipts, qc_inspections, issues_out, stock_audits,
# quotations, quotation_items, orders, order_items, invoices,
# receipts_payment, issues, corrective_actions, preventive_actions,
# tickets, changes, board_posts, qc_inspection_reports/items,
# work_orders/items, export_orders, commercial_invoices, packing_lists,
# bills_of_lading, project_milestones, project_burndown_snapshots.
# 멱등 가드: parts 테이블에 행이 있으면 전체 skip.
# 외부 자산 0건. 표준 라이브러리만 사용.
# =====================================================
def seed_business_data():
    """매출·자재·품질·진행 모듈 페이지가 빈 채로 노출되지 않도록
    현실적 더미 데이터를 일괄 시드. 멱등 (parts 행 존재 시 skip)."""
    import random as _rand
    _rand.seed(20260503)
    today = _date.today()

    # 멱등 가드: app_settings 마커로 1회만 실행
    with db_session() as c:
        marker = c.execute(
            "SELECT value FROM app_settings WHERE key='seed_business_data_v1'"
        ).fetchone()
        if marker:
            return 0

        # ───────── 0. 참조 마스터 조회 ─────────
        users = [dict(r) for r in c.execute(
            "SELECT id, name, role, team_id FROM users WHERE is_active=1 AND role!='admin'"
        ).fetchall()]
        if not users:
            return 0
        ceo_user = next((u for u in users if u["role"] == "ceo"), users[0])
        purch_users = [u for u in users if u["role"] in ("leader", "executive", "member")][:6] or users[:6]
        qa_users = [u for u in users if u["role"] in ("leader", "member")][:4] or users[:4]
        sales_users = users[:5]

        customers = [dict(r) for r in c.execute(
            "SELECT id, name FROM customers ORDER BY id"
        ).fetchall()]
        projects = [dict(r) for r in c.execute(
            "SELECT id, code, name, customer_id, type FROM projects ORDER BY id"
        ).fetchall()]
        teams = [dict(r) for r in c.execute(
            "SELECT id, code, name FROM teams ORDER BY display_order"
        ).fetchall()]

        # ───────── 1. PARTS 12종 ─────────
        # v5H226z118 (2026-05-19) — 데모 시드: 제조사명은 가상(외부 상표 비거론). 운영 DB 실데이터와 무관
        parts_seed = [
            ("KNK-T-CCD-001",  "CCD 카메라 모듈 5MP",    "1/1.8\" / Global Shutter", "비전테크",   "Japan",   "EA",  450_000, "T", "자재"),
            ("KNK-T-LEN-001",  "M12 광학렌즈 16mm",     "F2.4 / Coated",            "옵티코",     "USA",     "EA",  120_000, "T", "자재"),
            ("KNK-T-LED-001",  "Bar Type LED 조명 200x30", "White 6500K",          "라이트텍",   "Japan",   "EA",  220_000, "T", "자재"),
            ("KNK-T-PCB-001",  "검사기 메인 PCB v3.2",   "8layer FR4",               "KNK설계",    "Korea",   "EA",  380_000, "T", "완성품"),
            ("KNK-M-SVR-001",  "Servo Motor 400W",       "Pulse 2500P/R",            "서보맥스",   "Japan",   "EA",  680_000, "M", "자재"),
            ("KNK-M-CYL-001",  "Pneumatic Cylinder Φ32", "Stroke 100mm",             "에어시스템", "Japan",   "EA",   85_000, "M", "자재"),
            ("KNK-M-PLC-001",  "PLC CPU Module",         "16-Slot",                  "컨트롤텍",   "Japan",   "EA",  720_000, "M", "자재"),
            ("KNK-M-BLT-001",  "Conveyor Belt 1200x100", "PVC Anti-static",          "벨트프로",   "Swiss",   "EA",  140_000, "M", "자재"),
            ("KNK-C-PWR-001",  "DC 24V/5A Power",        "AC100-240V Input",         "파워셀",     "Taiwan",  "EA",   62_000, "공통", "소모품"),
            ("KNK-C-CBL-001",  "Industrial Ethernet Cable 5m", "Cat6A / SHIELD",   "커넥라인",   "Germany", "EA",   28_000, "공통", "소모품"),
            ("KNK-C-FIL-001",  "Air Filter 5μm",          "1/4\" NPT",               "에어시스템", "Japan",   "EA",   18_000, "공통", "소모품"),
            ("KNK-C-LBR-001",  "Industrial Lubricant 1L", "Food grade NSF H1",       "루브리프로", "Germany", "EA",   42_000, "공통", "소모품"),
        ]
        for pn, nm, sp, mk, og, un, pr, bd, cat in parts_seed:
            c.execute(
                "INSERT OR IGNORE INTO parts(part_no, part_name, spec, maker, origin, unit, std_price, biz_div, category) "
                "VALUES(?,?,?,?,?,?,?,?,?)",
                (pn, nm, sp, mk, og, un, pr, bd, cat),
            )
        part_ids = {row["part_no"]: row["id"] for row in c.execute(
            "SELECT id, part_no FROM parts").fetchall()}
        # 시드 part_no가 모두 매핑됐는지 확인 (없으면 해당 시드 스킵)
        parts_seed = [t for t in parts_seed if t[0] in part_ids]
        if not parts_seed:
            return 0

        # ───────── 2. SUPPLIERS 6 ─────────
        # v5H226z118 — 데모 시드: 협력사명은 가상 (외부 상표 비거론)
        sup_seed = [
            ("성진엔지니어링",  "SUP-001", "박상호 차장", "sj@sjeng.kr",     "031-555-1100", "KR",  "KRW", "30일"),
            ("한빛전자부품",    "SUP-002", "김미진 대리", "mj@hb.kr",        "031-777-2200", "KR",  "KRW", "현금"),
            ("동광정밀",        "SUP-003", "이재훈 과장", "lee@dkjm.kr",     "032-310-3300", "KR",  "KRW", "60일"),
            ("비전테크코리아",  "SUP-004", "다나카",      "tn@visiontech.kr","02-555-4400",  "KR",  "USD", "선금"),
            ("에어시스템코리아","SUP-005", "야마다",      "ym@airsys.kr",    "031-460-5500", "KR",  "KRW", "30일"),
            ("파워셀무역",      "SUP-006", "첸",          "chen@powercell.tw","+886-2-66789", "TW",  "USD", "선금"),
        ]
        for nm, cd, ct, em, ph, cnt, cur, pt in sup_seed:
            c.execute(
                "INSERT OR IGNORE INTO suppliers(name, code, contact, email, phone, country, currency, payment_terms) "
                "VALUES(?,?,?,?,?,?,?,?)",
                (nm, cd, ct, em, ph, cnt, cur, pt),
            )
        sup_ids = [r["id"] for r in c.execute("SELECT id FROM suppliers").fetchall()]
        if not sup_ids:
            return 0

        # ───────── 3. PURCHASE ORDERS 8 + LINES + STOCK MOVEMENTS ─────────
        po_status_pool = ["발주완료", "발주완료", "부분입고", "입고완료", "입고완료", "입고완료", "작성중", "발주완료"]
        for i in range(8):
            d = today - _td(days=_rand.randint(2, 35))
            po_no = f"PO-DEMO-{d.strftime('%y%m')}-{i+1:03d}"
            sup = _rand.choice(sup_ids)
            pj  = _rand.choice(projects) if projects and _rand.random() > 0.3 else None
            cur = "USD" if _rand.random() > 0.7 else "KRW"
            xr  = 1380.0 if cur == "USD" else 1.0
            status = po_status_pool[i]
            c.execute(
                "INSERT INTO purchase_orders(po_number, project_id, supplier_id, order_date, expected_date, "
                "currency, exchange_rate, status, shipping_terms, payment_terms, po_type, created_by, note) "
                "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (po_no, pj["id"] if pj else None, sup, d.isoformat(),
                 (d + _td(days=_rand.randint(7, 21))).isoformat(),
                 cur, xr, status,
                 _rand.choice(["국내", "FOB", "CIF", "DDP"]),
                 _rand.choice(["선금", "현금", "30일", "60일"]),
                 _rand.choice(["일반", "긴급", "정기"]),
                 _rand.choice(purch_users)["id"],
                 _rand.choice(["", "", "긴급납기 협의", "QC 보강 요청", "정기 발주"]))
            )
            po_id = c.execute("SELECT last_insert_rowid()").fetchone()[0]
            n_lines = _rand.choice([2, 3, 3, 4])
            line_total = 0
            chosen_parts = _rand.sample(list(parts_seed), k=n_lines)
            for ln, (pn, nm, sp, *rest) in enumerate(chosen_parts, start=1):
                pid = part_ids[pn]
                qty = _rand.choice([5, 10, 20, 30, 50, 100])
                up  = next(p[6] for p in parts_seed if p[0] == pn) * (1 + _rand.uniform(-0.05, 0.05))
                up  = round(up / xr, 2) if cur == "USD" else round(up)
                amt = round(qty * up)
                rcv = qty if status == "입고완료" else (qty // 2 if status == "부분입고" else 0)
                c.execute(
                    "INSERT INTO po_items(po_id, line_no, part_id, part_no_snapshot, part_name_snapshot, "
                    "spec_snapshot, unit, quantity, unit_price, amount, received_qty, delivery_date, note) "
                    "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (po_id, ln, pid, pn, nm, sp, "EA", qty, up, amt, rcv,
                     (d + _td(days=_rand.randint(7, 21))).isoformat(), "")
                )
                po_item_id = c.execute("SELECT last_insert_rowid()").fetchone()[0]
                line_total += amt
                # 입고완료/부분입고 → stock_movements IN 행
                if rcv > 0:
                    occ = (d + _td(days=_rand.randint(2, 14))).strftime("%Y-%m-%d %H:%M")
                    sm_no = f"SM-{occ[2:4]}{occ[5:7]}{occ[8:10]}-{ln:03d}-{po_id}"
                    c.execute(
                        "INSERT INTO stock_movements(movement_no, part_id, kind, quantity, unit, "
                        "unit_price, amount, remaining_qty, lot_no, po_id, po_item_id, occurred_at, "
                        "created_by, reason) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        (sm_no, pid, "IN", rcv, "EA", up * (xr if cur == "USD" else 1),
                         round(rcv * up * (xr if cur == "USD" else 1)),
                         rcv, f"LOT-{po_no}-L{ln}", po_id, po_item_id, occ,
                         _rand.choice(purch_users)["id"], "발주 입고")
                    )
            c.execute("UPDATE purchase_orders SET total_amount=? WHERE id=?", (line_total, po_id))

        # ───────── 4. RECEIPTS (입고 헤더) — 입고완료 PO 기준 ─────────
        finished_pos = [dict(r) for r in c.execute(
            "SELECT id, po_number, total_amount FROM purchase_orders "
            "WHERE status IN ('입고완료','부분입고')").fetchall()]
        for po in finished_pos:
            tot_q = c.execute(
                "SELECT COALESCE(SUM(received_qty),0) FROM po_items WHERE po_id=?", (po["id"],)
            ).fetchone()[0]
            c.execute(
                "INSERT INTO receipts(po_id, received_at, received_by, total_qty, status, note) "
                "VALUES(?,?,?,?,?,?)",
                (po["id"], (today - _td(days=_rand.randint(0, 14))).strftime("%Y-%m-%d %H:%M"),
                 _rand.choice(purch_users)["id"], tot_q,
                 _rand.choice(["PASS", "PASS", "PARTIAL"]), "정상 입고 처리")
            )

        # ───────── 5. QC INSPECTIONS — 입고 일부 (PASS 7 / FAIL 1) ─────────
        po_items_for_qc = [dict(r) for r in c.execute(
            "SELECT pi.id AS poi_id, pi.received_qty AS rq, r.id AS rcp_id "
            "FROM po_items pi JOIN receipts r ON r.po_id=pi.po_id "
            "WHERE pi.received_qty>0 LIMIT 8").fetchall()]
        for i, q in enumerate(po_items_for_qc):
            fail = (i == 3)
            pass_qty = 0 if fail else q["rq"]
            fail_qty = q["rq"] if fail else 0
            c.execute(
                "INSERT INTO qc_inspections(po_item_id, receipt_id, inspector_id, pass_qty, "
                "fail_qty, fail_reason, status) VALUES(?,?,?,?,?,?,?)",
                (q["poi_id"], q["rcp_id"], _rand.choice(qa_users)["id"], pass_qty, fail_qty,
                 "외관 흠집 발견 (cosmetic, 사용가능)" if fail else None,
                 "FAIL" if fail else "PASS")
            )

        # ───────── 6. STOCK ISSUES (출고 요청) 5건 ─────────
        for i in range(5):
            pn = _rand.choice(list(parts_seed))[0]
            pid = part_ids[pn]
            qty = _rand.choice([1, 2, 5])
            stat = _rand.choice(["ISSUED", "ISSUED", "APPROVED", "PENDING", "ISSUED"])
            req_at = (today - _td(days=_rand.randint(0, 10))).strftime("%Y-%m-%d %H:%M")
            iss_at = req_at if stat == "ISSUED" else None
            c.execute(
                "INSERT INTO issues_out(part_id, requester_id, approver_id, requested_at, "
                "issued_at, qty, purpose, status) VALUES(?,?,?,?,?,?,?,?)",
                (pid, _rand.choice(purch_users)["id"], _rand.choice(purch_users)["id"],
                 req_at, iss_at, qty,
                 _rand.choice(["조립용 출고", "AS 부품 교체", "샘플 검토", "긴급 수리"]), stat)
            )
            if stat == "ISSUED":
                up = next(p[6] for p in parts_seed if p[0] == pn)
                c.execute(
                    "INSERT INTO stock_movements(movement_no, part_id, kind, quantity, unit, "
                    "unit_price, amount, occurred_at, created_by, reason) "
                    "VALUES(?,?,?,?,?,?,?,?,?,?)",
                    (f"SM-OUT-{i+1:03d}-{today.strftime('%y%m%d')}", pid, "OUT", -qty, "EA",
                     up, qty * up, iss_at,
                     _rand.choice(purch_users)["id"], "출고 승인")
                )

        # ───────── 7. STOCK AUDITS — 분기 1회 모형 ─────────
        for i, mago in enumerate([90, 60, 30, 0]):
            sd = today - _td(days=mago + _rand.randint(0, 5))
            audit_no = f"AUD-DEMO-{sd.strftime('%Y%m')}-{i+1:04d}"
            stat = ["CLOSED", "CLOSED", "REVIEW", "OPEN"][i]
            c.execute(
                "INSERT INTO stock_audits(audit_no, start_date, end_date, status, led_by, note) "
                "VALUES(?,?,?,?,?,?)",
                (audit_no, sd.isoformat(),
                 (sd + _td(days=2)).isoformat() if stat == "CLOSED" else None,
                 stat, _rand.choice(purch_users)["id"],
                 "분기 정기 실사 / 사이클 카운트")
            )

        # ───────── 8. QUOTATIONS 5 + ITEMS ─────────
        for i in range(5):
            d = today - _td(days=_rand.randint(0, 60))
            qno = f"QT-DEMO-{d.strftime('%Y%m')}-{i+1:04d}"
            cu = _rand.choice(customers) if customers else None
            stat = ["DRAFT", "QUOTED", "QUOTED", "CONFIRMED", "CONFIRMED"][i]
            c.execute(
                "INSERT INTO quotations(quote_no, customer_id, valid_until, version, status, created_by) "
                "VALUES(?,?,?,?,?,?)",
                (qno, cu["id"] if cu else None,
                 (d + _td(days=30)).isoformat(), 1, stat, _rand.choice(sales_users)["id"])
            )
            qid = c.execute("SELECT last_insert_rowid()").fetchone()[0]
            tot = 0
            for ln in range(1, _rand.choice([2, 3, 3]) + 1):
                pn = _rand.choice(list(parts_seed))
                pid = part_ids[pn[0]]
                qty = _rand.choice([1, 2, 3, 5])
                up  = pn[6] * _rand.uniform(1.10, 1.35)
                tp  = round(qty * up)
                c.execute(
                    "INSERT INTO quotation_items(quotation_id, line_no, part_id, item_name, qty, "
                    "unit, unit_price, total_price) VALUES(?,?,?,?,?,?,?,?)",
                    (qid, ln, pid, pn[1], qty, "EA", round(up), tp)
                )
                tot += tp
            c.execute("UPDATE quotations SET total_amount=? WHERE id=?", (tot, qid))

        # ───────── 9. ORDERS 6 + ITEMS + INVOICES + RECEIPTS_PAYMENT ─────────
        order_status_pool = ["CONFIRMED", "IN_PRODUCTION", "READY_TO_SHIP",
                             "SHIPPED", "INVOICED", "PAID"]
        for i in range(6):
            d = today - _td(days=_rand.randint(0, 50))
            sno = f"SO-DEMO-{d.strftime('%Y%m')}-{i+1:04d}"
            cu = _rand.choice(customers) if customers else None
            stat = order_status_pool[i]
            c.execute(
                "INSERT INTO orders(order_no, customer_id, order_date, due_date, status, created_by) "
                "VALUES(?,?,?,?,?,?)",
                (sno, cu["id"] if cu else None, d.isoformat(),
                 (d + _td(days=_rand.randint(20, 60))).isoformat(),
                 stat, _rand.choice(sales_users)["id"])
            )
            oid = c.execute("SELECT last_insert_rowid()").fetchone()[0]
            tot = 0
            for ln in range(1, _rand.choice([2, 3]) + 1):
                pn = _rand.choice(list(parts_seed))
                pid = part_ids[pn[0]]
                qty = _rand.choice([1, 2, 3, 5, 10])
                up  = pn[6] * _rand.uniform(1.15, 1.4)
                amt = round(qty * up)
                c.execute(
                    "INSERT INTO order_items(order_id, part_id, qty, unit_price, amount) "
                    "VALUES(?,?,?,?,?)",
                    (oid, pid, qty, round(up), amt)
                )
                tot += amt
            c.execute("UPDATE orders SET total_amount=? WHERE id=?", (tot, oid))
            # 상태 이력 1행
            c.execute(
                "INSERT INTO order_status_history(order_id, from_status, to_status, changed_by, note) "
                "VALUES(?,?,?,?,?)",
                (oid, "DRAFT", stat, _rand.choice(sales_users)["id"], "정기 진행")
            )
            # 송장 — INVOICED/PAID/SHIPPED 단계
            if stat in ("INVOICED", "PAID", "SHIPPED"):
                vat = round(tot * 0.10)
                ino = f"INV-DEMO-{d.strftime('%Y%m')}-{i+1:04d}"
                c.execute(
                    "INSERT INTO invoices(invoice_no, order_id, customer_id, issue_date, "
                    "amount_excl_vat, vat, total_amount, status, created_by) "
                    "VALUES(?,?,?,?,?,?,?,?,?)",
                    (ino, oid, cu["id"] if cu else None,
                     (d + _td(days=10)).isoformat(),
                     tot, vat, tot + vat, "ISSUED", _rand.choice(sales_users)["id"])
                )
            # 수금 — PAID 일부 / SHIPPED 부분 / INVOICED 미수 split
            if stat == "PAID":
                c.execute(
                    "INSERT INTO receipts_payment(order_id, received_at, amount, method, received_by, note) "
                    "VALUES(?,?,?,?,?,?)",
                    (oid, (d + _td(days=20)).strftime("%Y-%m-%d %H:%M"),
                     round((tot * 1.10)), "이체", _rand.choice(sales_users)["id"], "전액 수금")
                )
            elif stat == "INVOICED":
                # 부분 수금 50%
                c.execute(
                    "INSERT INTO receipts_payment(order_id, received_at, amount, method, received_by, note) "
                    "VALUES(?,?,?,?,?,?)",
                    (oid, (d + _td(days=15)).strftime("%Y-%m-%d %H:%M"),
                     round((tot * 1.10) / 2), "이체", _rand.choice(sales_users)["id"], "부분 수금")
                )

        # ───────── 10. ISSUES (QMS) 5 + CORRECTIVE/PREVENTIVE ─────────
        iss_seed = [
            ("CCD 카메라 노이즈 발생", "심각", "품질", "조치중", "고객사A",   "T", "고객 라인 가동 중 5분에 1회 노이즈 라인", "전원 노이즈 추정", "필터 모듈 추가 적용", ""),
            ("Servo 위치 오차 0.05mm", "중",   "설계결함", "원인분석", "고객사B", "M", "0.05mm 위치 오차 반복 발생", "엔코더 케이블 EMI", "차폐 케이블 교체 검토", ""),
            ("PLC 통신 단선",          "치명", "AS",    "해결",    "드림텍",   "M", "야간 라인 통신 단선", "RJ45 커넥터 산화", "신규 커넥터 교체 / SOP 갱신", "정기 점검 매월 1회"),
            ("LED 조명 휘도 저하",      "경",   "품질",  "접수",    "한국성전", "T", "출하 1주일 만에 휘도 80%로 저하", "", "", ""),
            ("BOM 오기재로 잘못 출고",   "중",   "기타",  "재발방지등록", "기타전장", "T", "BOM Rev2 vs Rev3 혼동", "관리코드 확인 누락", "출고 전 2인 확인 SOP", "체크리스트 도입"),
        ]
        for i, (ttl, sev, ity, st, cust, bd, desc, rc, act, prev) in enumerate(iss_seed):
            cu = next((c2 for c2 in customers if c2["name"] == cust), None)
            d  = today - _td(days=_rand.randint(2, 30))
            iss_no = f"ISS-DEMO-{d.strftime('%y%m')}-{i+1:03d}"
            owner_team = _rand.choice(teams)["id"] if teams else None
            c.execute(
                "INSERT INTO issues(issue_no, title, severity, issue_type, status, customer_id, "
                "customer_name, biz_div, occurred_at, description, root_cause, action_taken, "
                "prevention, owner_team_id, owner_user_id, cost_estimate, created_by) "
                "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (iss_no, ttl, sev, ity, st, cu["id"] if cu else None, cust, bd,
                 d.isoformat(), desc, rc, act, prev,
                 owner_team, _rand.choice(qa_users)["id"],
                 _rand.choice([0, 200_000, 500_000, 1_500_000]),
                 _rand.choice(qa_users)["id"])
            )
            issue_id = c.execute("SELECT last_insert_rowid()").fetchone()[0]
            # 시정조치
            if st in ("조치중", "해결", "재발방지등록"):
                _ca_status = "DONE" if st != "조치중" else "IN_PROGRESS"
                _ca_lifecycle = "COMPLETED" if _ca_status == "DONE" else "IN_PROGRESS"
                c.execute(
                    "INSERT INTO corrective_actions(issue_id, action, responsible, due_date, "
                    "completed_at, status, lifecycle_status, created_by) VALUES(?,?,?,?,?,?,?,?)",
                    (issue_id, act or "임시 조치", _rand.choice(qa_users)["id"],
                     (d + _td(days=14)).isoformat(),
                     (d + _td(days=10)).isoformat() if st != "조치중" else None,
                     _ca_status, _ca_lifecycle,
                     _rand.choice(qa_users)["id"])
                )
                ca_id = c.execute("SELECT last_insert_rowid()").fetchone()[0]
                # 예방조치
                if st == "재발방지등록" and prev:
                    c.execute(
                        "INSERT INTO preventive_actions(corrective_id, action, completed_at, lifecycle_status, created_by) "
                        "VALUES(?,?,?,?,?)",
                        (ca_id, prev, (d + _td(days=12)).isoformat(),
                         "COMPLETED",
                         _rand.choice(qa_users)["id"])
                    )

        # ───────── 11. TICKETS 6 ─────────
        tkt_seed = [
            ("자재요청", "추가 PCB 5장 긴급 요청",      "다음 주 출하분 부족", "긴급", "처리중"),
            ("긴급가공", "Bracket 절삭 가공 의뢰",       "샘플용 3개",        "긴급", "완료"),
            ("MODIFY",   "검사기 SW 임계값 조정",        "0.5 → 0.4mm 변경", "일반", "요청"),
            ("검수요청", "출하 전 외관/동작 검수",       "BOM Rev3 적용분",   "일반", "처리중"),
            ("AS",       "고객사 LED 조명 교체",         "현장 출장 필요",    "긴급", "완료"),
            ("기타",     "사내 정기 안전 점검",          "공장동 1·2층",      "일반", "요청"),
        ]
        for i, (cat, ttl, desc, urg, st) in enumerate(tkt_seed):
            d = today - _td(days=_rand.randint(0, 14))
            tno = f"TKT-DEMO-{d.strftime('%y%m')}-{i+1:03d}"
            req = _rand.choice(users)
            recv = _rand.choice(teams)["id"] if teams else None
            c.execute(
                "INSERT INTO tickets(ticket_no, category, title, description, requester_id, "
                "recipient_team_id, urgency, status, due_date, completed_at, hours_estimated, "
                "hours_actual) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                (tno, cat, ttl, desc, req["id"], recv, urg, st,
                 (d + _td(days=_rand.randint(2, 7))).isoformat(),
                 (d + _td(days=2)).isoformat() if st == "완료" else None,
                 _rand.choice([1, 2, 4, 8]),
                 _rand.choice([1, 2, 3, 5]) if st == "완료" else None)
            )

        # ───────── 12. CHANGES (변경 공지) 4 ─────────
        chg_seed = [
            ("기구설계", "검사기 베이스 두께 +5mm 변경",       "T", "강성 부족 → 처짐 발생",
             "베이스 25mm", "베이스 30mm", "긴급"),
            ("BOM",     "PCB Rev2 → Rev3 적용 (모든 검사기)", "T", "노이즈 필터 회로 추가",
             "Rev2", "Rev3", "일반"),
            ("소프트웨어", "검사 기준값 조정 (0.5→0.4mm)",     "T", "고객 품질 향상 요청",
             "0.5mm", "0.4mm", "예약"),
            ("도면",    "조립도 A-203 개정 (간섭 회피)",       "M", "Cylinder 충돌 방지",
             "A-203 R0", "A-203 R1", "일반"),
        ]
        for i, (cty, ttl, bd, desc, bv, av, urg) in enumerate(chg_seed):
            d = today - _td(days=_rand.randint(1, 20))
            cno = f"CHG-DEMO-{d.strftime('%y%m')}-{i+1:03d}"
            pj  = _rand.choice(projects) if projects else None
            c.execute(
                "INSERT INTO changes(change_no, change_type, biz_div, target_kind, target_id, target_label, "
                "project_id, title, description, before_value, after_value, urgency, author_id, status, "
                "notified_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (cno, cty, bd, "project", pj["id"] if pj else None,
                 pj["name"] if pj else "전체",
                 pj["id"] if pj else None, ttl, desc, bv, av, urg,
                 _rand.choice(users)["id"],
                 _rand.choice(["공지중", "공지중", "확인완료"]),
                 d.strftime("%Y-%m-%d %H:%M"))
            )

        # ───────── 13. BOARD POSTS — 전사·팀 게시판 8 ─────────
        # 전사 게시판 자동 생성
        company_board = c.execute(
            "SELECT id FROM boards WHERE type='company'").fetchone()
        if not company_board:
            c.execute("INSERT INTO boards(name, type) VALUES(?,?)",
                      ("전사 게시판", "company"))
            company_board_id = c.execute("SELECT last_insert_rowid()").fetchone()[0]
        else:
            company_board_id = company_board["id"]

        # 각 팀에도 게시판
        team_board_map = {}
        for t in teams:
            existing = c.execute(
                "SELECT id FROM boards WHERE type='team' AND team_id=?", (t["id"],)
            ).fetchone()
            if not existing:
                c.execute(
                    "INSERT INTO boards(name, type, team_id) VALUES(?,?,?)",
                    (f"{t['name']} 게시판", "team", t["id"])
                )
                bid = c.execute("SELECT last_insert_rowid()").fetchone()[0]
            else:
                bid = existing["id"]
            team_board_map[t["id"]] = bid

        post_seed = [
            (company_board_id, "공지", 1, "5월 전사 안전교육 안내",
             "5월 둘째 주 화요일 14:00, 본사 회의실에서 전사 안전교육이 진행됩니다. 전 직원 필참."),
            (company_board_id, "공지", 1, "여름 휴가 일정 사전 신청 안내",
             "7~8월 여름 휴가 일정을 5월 31일까지 각 팀장에게 신청 부탁드립니다."),
            (company_board_id, "자료", 0, "검사기/자동화 통합 SOP v3.2 배포",
             "표준 작업 절차서 v3.2 가 배포되었습니다. 변경점은 §4 안전 점검 항목입니다."),
            (company_board_id, "일반", 0, "신입사원 환영합니다",
             "이번 주 신규 입사자가 합류했습니다. 따뜻한 환영 부탁드립니다."),
        ]
        # 팀별 1건씩
        for t in teams[:4]:
            post_seed.append(
                (team_board_map[t["id"]], "일반", 0,
                 f"{t['name']} 주간 회의록 ({today.strftime('%m/%d')})",
                 "1) 진행 중 안건 공유\n2) 다음 주 우선순위\n3) 현장 이슈 공유")
            )
        for bid, cat, pin, ttl, body in post_seed:
            c.execute(
                "INSERT INTO board_posts(board_id, author_id, title, body, category, "
                "is_pinned, view_count, approval_status) VALUES(?,?,?,?,?,?,?,?)",
                (bid, _rand.choice(users)["id"], ttl, body, cat, pin,
                 _rand.randint(5, 80), "approved")
            )

        # ───────── 14. QC INSPECTION REPORTS 4 ─────────
        for i in range(4):
            d  = today - _td(days=_rand.randint(0, 25))
            cu = _rand.choice(customers) if customers else None
            qa = _rand.choice(qa_users)
            insp = _rand.choice(qa_users)
            rno = f"QCR-DEMO-{d.strftime('%Y')}-{i+1:04d}"
            ovr = _rand.choice(["PASS", "PASS", "PASS", "CONDITIONAL_PASS"])
            stat = _rand.choice(["ISSUED", "SENT", "DRAFT"])
            c.execute(
                "INSERT INTO qc_inspection_reports(report_no, customer_id, customer_name, "
                "machine_model, machine_serial, inspection_date, inspector_id, inspector_name, "
                "qa_manager_id, qa_manager_name, overall, status, issued_at, created_by, remarks) "
                "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (rno, cu["id"] if cu else None, cu["name"] if cu else "내부",
                 _rand.choice(["HAIST-INS-VX1", "HAIST-INS-VX2", "HAIST-AT-AX1"]),
                 f"SN{_rand.randint(20240, 26050)}",
                 d.isoformat(), insp["id"], insp["name"],
                 qa["id"], qa["name"], ovr, stat,
                 d.strftime("%Y-%m-%d %H:%M") if stat != "DRAFT" else None,
                 insp["id"], "출하 전 표준 검수")
            )
            rid = c.execute("SELECT last_insert_rowid()").fetchone()[0]
            for ln, item in enumerate([
                ("반복성",  "≤0.5μm",   "0.32μm",   "PASS"),
                ("정확도",  "100±0.1mm", "99.97mm",  "PASS"),
                ("통신",    "Latency≤5ms", "3.1ms",  "PASS"),
                ("외관",    "스크래치 無", "양호",     "PASS"),
                ("동작",    "사이클 정상", "정상",     "PASS"),
                ("안전",    "안전 회로 정상", "정상",  "PASS"),
            ], start=1):
                jud = "PASS" if ovr == "PASS" else _rand.choice(["PASS", "PASS", "FAIL"])
                c.execute(
                    "INSERT INTO qc_inspection_items(report_id, line_no, item_name, "
                    "spec_value, measured_value, judgment, remarks) VALUES(?,?,?,?,?,?,?)",
                    (rid, ln, item[0], item[1], item[2], jud,
                     "기준 충족" if jud == "PASS" else "부적합 → 재작업")
                )

        # ───────── 15. WORK ORDERS 3 ─────────
        for i in range(3):
            d  = today - _td(days=_rand.randint(0, 20))
            wno = f"WO-DEMO-{d.strftime('%Y')}-{i+1:04d}"
            pn  = _rand.choice(list(parts_seed))
            pid = part_ids[pn[0]]
            asn = _rand.choice(users)
            cby = _rand.choice(users)
            stat = ["IN_PROGRESS", "COMPLETED", "RELEASED"][i]
            c.execute(
                "INSERT INTO work_orders(wo_no, part_id, qty, assigned_to, assigned_name, "
                "created_by, created_by_name, planned_start, planned_end, "
                "actual_end, specifications, status, remarks) "
                "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (wno, pid, _rand.choice([5, 10, 20]), asn["id"], asn["name"],
                 cby["id"], cby["name"], d.isoformat(),
                 (d + _td(days=5)).isoformat(),
                 (d + _td(days=4)).isoformat() if stat == "COMPLETED" else None,
                 f"{pn[1]} 가공 — 도면 기준 ±0.05mm", stat,
                 "정밀 가공 / 표면 거칠기 Ra1.6")
            )
            wid = c.execute("SELECT last_insert_rowid()").fetchone()[0]
            for ln, step in enumerate(["절삭", "연마", "검수"], start=1):
                prog = 100 if stat == "COMPLETED" else (
                    100 if ln == 1 and stat == "IN_PROGRESS" else
                    50 if ln == 2 and stat == "IN_PROGRESS" else 0)
                w = _rand.choice(users)
                c.execute(
                    "INSERT INTO work_order_items(wo_id, line_no, step_name, duration_min, "
                    "progress, worker_id, worker_name, remarks) VALUES(?,?,?,?,?,?,?,?)",
                    (wid, ln, step, _rand.choice([30, 60, 90, 120]),
                     prog, w["id"], w["name"], "표준 공정")
                )

        # ───────── 16. EXPORT ORDERS 3 + CI/PL/BL ─────────
        # 기존 orders 중 SHIPPED/INVOICED 일부를 수출 오더로 매핑
        ship_orders = [dict(r) for r in c.execute(
            "SELECT id, order_no, total_amount FROM orders "
            "WHERE status IN ('SHIPPED','INVOICED','PAID') LIMIT 3").fetchall()]
        for i, so in enumerate(ship_orders):
            stat = ["SHIPPED", "CI_ISSUED", "BOOKED"][i]
            c.execute(
                "INSERT INTO export_orders(order_id, buyer, shipping_terms, payment_terms, "
                "port_of_loading, port_of_discharge, status, created_by) "
                "VALUES(?,?,?,?,?,?,?,?)",
                (so["id"],
                 _rand.choice(["KNK Vietnam Ltd.", "ASEAN Tech Co.", "Hanoi Auto Inc."]),
                 _rand.choice(["FOB", "CIF", "DDP"]),
                 _rand.choice(["T/T", "L/C", "D/P"]),
                 "BUSAN",
                 _rand.choice(["HAIPHONG", "HOCHIMINH", "BANGKOK"]),
                 stat, _rand.choice(sales_users)["id"])
            )
            eo_id = c.execute("SELECT last_insert_rowid()").fetchone()[0]
            # CI
            usd_amt = round(so["total_amount"] / 1380, 2)
            c.execute(
                "INSERT INTO commercial_invoices(invoice_no, export_order_id, issue_date, "
                "total_amount, currency, signed_by, status) VALUES(?,?,?,?,?,?,?)",
                (f"CI-DEMO-{today.strftime('%Y%m')}-{i+1:04d}", eo_id,
                 today.isoformat(), usd_amt, "USD",
                 _rand.choice(sales_users)["id"], "ISSUED")
            )
            # PL
            c.execute(
                "INSERT INTO packing_lists(pl_no, export_order_id, total_packages, "
                "total_weight, total_volume) VALUES(?,?,?,?,?)",
                (f"PL-DEMO-{today.strftime('%Y%m')}-{i+1:04d}", eo_id,
                 _rand.randint(2, 10), _rand.randint(150, 800), round(_rand.uniform(2.5, 12.0), 2))
            )

        # ───────── 17. PROJECT MILESTONES + BURNDOWN ─────────
        for pj in projects[:6]:
            ms_seed = [
                ("Concept 동결", -45, "DONE"),
                ("설계 완료",     -25, "DONE"),
                ("자재 입고 완료", -10, "IN_PROGRESS"),
                ("조립·시운전",   +5,  "PLANNED"),
                ("출하 검수",     +20, "PLANNED"),
                ("고객 인수",     +35, "PLANNED"),
            ]
            for nm, off, st in ms_seed:
                c.execute(
                    "INSERT INTO project_milestones(project_id, name, target_date, status, "
                    "completed_at) VALUES(?,?,?,?,?)",
                    (pj["id"], nm, (today + _td(days=off)).isoformat(),
                     st, (today + _td(days=off)).isoformat() if st == "DONE" else None)
                )
            # 14일 번다운 — 단조 감소
            base_total = _rand.randint(40, 80)
            done_per_day = base_total / 30
            for dback in range(14, -1, -1):
                snap = today - _td(days=dback)
                done = int(min(base_total, (14 - dback + 5) * done_per_day))
                rem  = max(0, (base_total - done) * 4)
                try:
                    c.execute(
                        "INSERT INTO project_burndown_snapshots(project_id, snap_date, "
                        "total_tasks, completed_tasks, remaining_hours) VALUES(?,?,?,?,?)",
                        (pj["id"], snap.isoformat(), base_total, done, rem)
                    )
                except Exception:
                    pass

        # 마커 기록 — 다음 부팅에서 skip
        c.execute(
            "INSERT OR REPLACE INTO app_settings(key, value, description) "
            "VALUES('seed_business_data_v1', ?, "
            "'v5H45 (2026-05-03) 비즈니스 데이터 시드 완료')",
            (today.isoformat(),)
        )

    return 1


# =====================================================
# v5H50 (2026-05-03) — 최근 7일 task 자동 보충 (대시보드 신선도)
# 시드 데이터가 며칠 묵으면 dashboard·feed·weekly 가 빈 화면으로
# 노출되는 것을 방지. 매 startup 시 실행하되 30건 이상이면 skip.
# 정상 운영 환경에서는 사용자가 직접 작성한 task가 30건 넘으므로 무동작.
# =====================================================
def seed_recent_tasks_topup() -> int:
    """최근 7일에 task가 30건 미만이면 더미 task 자동 보충."""
    import random as _rand
    today = _date.today()
    week_ago = (today - _td(days=6)).isoformat()
    with db_session() as c:
        # v5H226z135: 운영 모드(데모 초기화 후) — 데모 task 재생성 차단
        try:
            _r = c.execute("SELECT value FROM app_settings WHERE key='demo_seed_off'").fetchone()
            if _r and str(_r[0]) == '1':
                return 0
        except Exception:
            pass
        n = c.execute(
            "SELECT COUNT(*) FROM tasks WHERE work_date>=?", (week_ago,)
        ).fetchone()[0]
        if n >= 30:
            return 0
        users = [dict(r) for r in c.execute(
            "SELECT id, name, rank, role, team_id FROM users "
            "WHERE is_active=1 AND role!='admin' AND team_id IS NOT NULL"
        ).fetchall()]
        if not users:
            return 0
        projects = [dict(r) for r in c.execute(
            "SELECT id, name, type, customer_id FROM projects ORDER BY id LIMIT 30"
        ).fetchall()]
        customers = [dict(r) for r in c.execute(
            "SELECT id, name FROM customers"
        ).fetchall()]
        titles = [
            "고객사 견적 검토", "BOM 사양 확인 미팅", "내부 설계 리뷰",
            "협력사 일정 조율", "샘플 검수", "출하 전 점검",
            "테스트 데이터 분석", "공정 개선 검토", "품질 이슈 대응",
            "신규 부품 검증", "월간 리포트 작성", "장비 시운전",
            "검사기 캘리브레이션", "자동화 시퀀스 디버깅",
        ]
        cats = ["고객대응", "내부업무", "협력사", "품질", "설계", "테스트", "보고"]
        st_today = ["진행중", "진행중", "완료", "지연", "대기"]
        st_past = ["완료", "완료", "진행중", "지연"]
        notes_pool = ["", "", "검토 진행 중", "다음 주 마무리",
                      "협의 완료", "샘플 도착 대기"]
        # 시드 시드값: 매일 다르게 (재현성 + 일별 변화)
        _rand.seed(42 + today.toordinal())
        added = 0
        for dback in range(7):
            d = today - _td(days=dback)
            if d.weekday() >= 5:  # 주말 skip
                continue
            sample_users = _rand.sample(users, min(8, len(users)))
            for u in sample_users:
                n_tasks = _rand.choice([1, 2, 2, 3])
                for _ in range(n_tasks):
                    pj = _rand.choice(projects) if projects and _rand.random() > 0.3 else None
                    cu_id = (pj["customer_id"] if pj
                             else (_rand.choice(customers)["id"]
                                   if customers and _rand.random() > 0.5 else None))
                    title = _rand.choice(titles)
                    if pj:
                        title = f'{pj["name"].split()[0]} {title}'
                    status = _rand.choice(st_today if dback == 0 else st_past)
                    hours = round(_rand.choice([0.5, 1, 1.5, 2, 2, 3, 4]), 1)
                    notes = _rand.choice(notes_pool)
                    c.execute(
                        "INSERT INTO tasks(user_id, work_date, title, category, "
                        "project_id, customer_id, status, hours, notes) "
                        "VALUES(?,?,?,?,?,?,?,?,?)",
                        (u["id"], d.isoformat(), title, _rand.choice(cats),
                         pj["id"] if pj else None, cu_id, status, hours, notes)
                    )
                    added += 1
        return added


# =====================================================
# v5H136 (2026-05-05) — PO 라인 ↔ 프로젝트 다대다 연결 헬퍼
# 목적: 장비(관리번호)별 소모품·수리 이력 추적
# =====================================================
def po_item_link_project(po_item_id: int, project_id: int,
                         allocated_qty=None, allocation_pct=None,
                         note: str = None, user_id: int = None) -> int:
    """PO 라인을 프로젝트에 연결. (link.id 반환).
    동일 (po_item_id, project_id) 중복 등록은 거부 (이미 있으면 기존 id 반환)."""
    with db_session() as c:
        existing = c.execute(
            "SELECT id FROM po_item_project_links WHERE po_item_id=? AND project_id=?",
            (int(po_item_id), int(project_id))
        ).fetchone()
        if existing:
            return int(existing[0])
        cur = c.execute(
            """INSERT INTO po_item_project_links
               (po_item_id, project_id, allocated_qty, allocation_pct, note, created_by)
               VALUES (?,?,?,?,?,?)""",
            (int(po_item_id), int(project_id),
             (float(allocated_qty) if allocated_qty not in (None, "", "None") else None),
             (float(allocation_pct) if allocation_pct not in (None, "", "None") else None),
             (note or None),
             (int(user_id) if user_id else None))
        )
        return int(cur.lastrowid)


def po_item_unlink_project(link_id: int) -> bool:
    """링크 1건 삭제."""
    with db_session() as c:
        c.execute("DELETE FROM po_item_project_links WHERE id=?", (int(link_id),))
    return True


def get_po_item_links(po_item_id: int) -> list:
    """PO 라인 1건에 연결된 프로젝트 목록 (project_detail 표시용)."""
    with db_session() as c:
        rows = c.execute(
            """SELECT l.id AS link_id, l.po_item_id, l.project_id,
                      l.allocated_qty, l.allocation_pct, l.note, l.created_at,
                      p.mgmt_code, p.name AS project_name, p.equip_type
               FROM po_item_project_links l
               JOIN projects p ON l.project_id = p.id
               WHERE l.po_item_id = ?
               ORDER BY l.id""",
            (int(po_item_id),)
        ).fetchall()
    return [dict(r) for r in rows]


def get_project_consumables(project_id: int, limit: int = 200) -> dict:
    """프로젝트 1건에 연결된 모든 PO 라인 + 합계.
    반환: {'rows': [...], 'total_amount': float, 'total_qty': float}"""
    with db_session() as c:
        rows = c.execute(
            """SELECT l.id AS link_id,
                      l.allocated_qty, l.allocation_pct, l.note AS link_note,
                      pi.id AS po_item_id, pi.line_no,
                      pi.part_id, pi.part_no_snapshot, pi.part_name_snapshot,
                      pi.spec_snapshot, pi.unit, pi.quantity, pi.unit_price,
                      pi.amount, pi.delivery_date,
                      po.id AS po_id, po.po_number, po.order_date,
                      po.status AS po_status, po.currency,
                      s.name AS supplier_name,
                      pa.code AS part_code, pa.name AS part_name
               FROM po_item_project_links l
               JOIN po_items pi ON l.po_item_id = pi.id
               JOIN purchase_orders po ON pi.po_id = po.id
               LEFT JOIN suppliers s ON po.supplier_id = s.id
               LEFT JOIN parts pa ON pi.part_id = pa.id
               WHERE l.project_id = ?
               ORDER BY po.order_date DESC, po.id DESC, pi.line_no""",
            (int(project_id),)
        ).fetchall()
    out_rows = []
    total_amount = 0.0
    total_qty = 0.0
    for r in rows[:limit]:
        d = dict(r)
        # 분배 수량 우선: allocated_qty > (allocation_pct % of quantity) > quantity
        qty_full = float(d.get("quantity") or 0)
        up = float(d.get("unit_price") or 0)
        if d.get("allocated_qty") is not None:
            eff_qty = float(d["allocated_qty"])
        elif d.get("allocation_pct") is not None:
            eff_qty = qty_full * float(d["allocation_pct"]) / 100.0
        else:
            eff_qty = qty_full
        d["effective_qty"] = eff_qty
        d["effective_amount"] = round(eff_qty * up, 2)
        total_qty += eff_qty
        total_amount += d["effective_amount"]
        out_rows.append(d)
    return {
        "rows": out_rows,
        "total_amount": round(total_amount, 2),
        "total_qty": round(total_qty, 4),
        "count": len(out_rows),
    }


def get_child_projects(parent_project_id: int, limit: int = 200) -> list:
    """v5H141 (2026-05-05): 부모 프로젝트(장비)에 연결된 자식 프로젝트(소모품/수리) 목록 + SO 합계.
    각 row: id, mgmt_code, name, project_type, status, customer_name,
            total_so_amount, total_units, currency, last_so_date, latest_so_no
    """
    with db_session() as c:
        rows = c.execute(
            """SELECT p.id, p.mgmt_code, p.name,
                      COALESCE(p.project_type,'NEW_EQUIP') AS project_type,
                      COALESCE(p.status,'진행중')          AS status,
                      COALESCE(p.customer_name,'')        AS customer_name,
                      COALESCE(p.currency,'KRW')          AS currency,
                      COALESCE(p.unit_qty, 1)             AS total_units,
                      (SELECT COALESCE(SUM(o.total_amount),0)
                         FROM orders o
                        WHERE o.project_id = p.id)        AS total_so_amount,
                      (SELECT MAX(o.order_date)
                         FROM orders o
                        WHERE o.project_id = p.id)        AS last_so_date,
                      (SELECT o.order_no FROM orders o
                        WHERE o.project_id = p.id
                        ORDER BY o.order_date DESC, o.id DESC LIMIT 1) AS latest_so_no
                 FROM projects p
                WHERE p.parent_project_id = ?
                ORDER BY p.created_at DESC, p.id DESC
                LIMIT ?""",
            (int(parent_project_id), int(limit))
        ).fetchall()
    return [dict(r) for r in rows]


def get_part_project_usage(part_id: int, limit: int = 50) -> list:
    """자재 1건이 어떤 프로젝트에서 누적 얼마나 쓰였는지.
    반환: [{project_id, mgmt_code, project_name, total_qty, total_amount, last_order_date, link_count}, ...]
    누적수량 내림차순."""
    with db_session() as c:
        rows = c.execute(
            """SELECT p.id AS project_id, p.mgmt_code, p.name AS project_name,
                      COUNT(l.id) AS link_count,
                      SUM(
                        CASE
                          WHEN l.allocated_qty IS NOT NULL THEN l.allocated_qty
                          WHEN l.allocation_pct IS NOT NULL THEN pi.quantity * l.allocation_pct / 100.0
                          ELSE pi.quantity
                        END
                      ) AS total_qty,
                      SUM(
                        CASE
                          WHEN l.allocated_qty IS NOT NULL THEN l.allocated_qty * pi.unit_price
                          WHEN l.allocation_pct IS NOT NULL THEN pi.quantity * pi.unit_price * l.allocation_pct / 100.0
                          ELSE pi.quantity * pi.unit_price
                        END
                      ) AS total_amount,
                      MAX(po.order_date) AS last_order_date
               FROM po_item_project_links l
               JOIN po_items pi ON l.po_item_id = pi.id
               JOIN purchase_orders po ON pi.po_id = po.id
               JOIN projects p ON l.project_id = p.id
               WHERE pi.part_id = ?
               GROUP BY p.id, p.mgmt_code, p.name
               ORDER BY total_qty DESC
               LIMIT ?""",
            (int(part_id), int(limit))
        ).fetchall()
    return [dict(r) for r in rows]


# =====================================================
# v5H226z136 (2026-05-31): 데모/시드 데이터 초기화 (앱 내 — admin/ceo 전용)
#   미리보기(reset_demo_preview) → 확인 → 실행(reset_demo_apply).
#   보존: 구조·마스터·실직원/대표. 비움: 트랜잭션·데모. 재시드 차단 플래그 설정.
# =====================================================
RESET_KEEP_TABLES = {
    "teams", "app_settings",
    "permissions", "permission_groups", "group_permissions", "user_groups",
    "countries_master", "payment_terms", "exchange_rates",
    "workflow_nodes_master", "workflow_template_nodes", "workflow_templates",
    "boards", "shared_contacts", "user_mail_creds",
    "sqlite_sequence", "users",
}
# 데모/시드 사용자: 사번 없음 + ceo 아님 (실직원·대표 절대 보존)
_RESET_DEMO_USER_WHERE = "(employee_no IS NULL OR TRIM(employee_no)='') AND COALESCE(role,'') <> 'ceo'"

# v5H226z192 (대표 지시): 데모 기간 '항목(그룹)별' 선택 초기화.
#   그룹 단위로 골라서 비움(조건부 삭제 없음). 사용자/설정/마스터/주소록은 그대로 보존.
#   '업체별' = 협력사 그룹. 전체 초기화(reset_demo_*)는 별개로 그대로 유지.
RESET_GROUPS = [
    {"key": "customers", "label": "🤝 고객사", "desc": "고객사·담당자·고객 이력",
     "tables": ["customers", "customer_contacts", "customer_history"]},
    {"key": "sales", "label": "📊 매출·영업", "desc": "프로젝트·수주·견적·납품·수금·송장·세금계산서·생산·수출(CI/PL/BL/통관/FTA)",
     "tables": ["projects", "project_history", "project_milestones", "project_phases", "project_retros",
                "project_forecasts", "project_burndown_snapshots", "project_workflow",
                "project_workflow_edges", "project_workflow_nodes", "project_workflow_node_checkpoints",
                "quotations", "quotation_items", "quotation_history",
                "orders", "order_items", "order_status_history", "production_orders", "ic_invoice_pairs",
                "invoices", "receipts", "receipts_payment", "shipments",
                # v5H226z618 (대표 지시): 묶음 세금계산서 장부 — 매출 초기화 시 함께 비워야
                #   재업로드 후 ref_id 어긋남(세금계산서↔원천 불일치)이 안 남음. 고객사 마스터는 보존(별도 그룹).
                "tax_invoices", "tax_invoice_lines",
                "export_orders", "commercial_invoices", "packing_lists", "packing_items",
                "bills_of_lading", "customs_declarations", "fta_certificates", "fta_certificate_items",
                # v5H226z713 (대표 지시): 작업일정표 보드 메모(행 단위 + 날짜별) — 매출 초기화 시 함께 비워야
                #   재업로드 후 프로젝트 id 재부여로 옛 메모가 엉뚱한 건에 붙는 것 방지(z618 세금계산서와 동일 취지).
                "schedule_board_memos", "schedule_board_day_memos"]},
    {"key": "suppliers", "label": "🏭 협력사(업체)", "desc": "협력사(공급업체)",
     "tables": ["suppliers"]},
    {"key": "materials", "label": "🔩 자재·구매", "desc": "부품·단가·발주·자재요청·재고·실사·원가시뮬",
     "tables": ["parts", "part_prices", "part_attachments", "price_change_history", "catalog_favorites",
                "purchase_orders", "po_items", "po_item_project_links",
                "material_requests", "material_request_items",
                "stock_movements", "stock_adjustments", "stock_audits", "stock_audit_items",
                "cost_simulations", "rate_alerts"]},
    {"key": "consumables", "label": "📦 소모품", "desc": "소모품 발주",
     "tables": ["consumable_orders", "consumable_order_items"]},
    {"key": "quality", "label": "✅ 품질·이슈·변경", "desc": "QC·QMS·시정/예방·작업지시·변경·이슈",
     "tables": ["qc_inspection_reports", "qc_inspection_items", "qc_inspections", "qc_disposition",
                "corrective_actions", "preventive_actions", "qms_audit_log", "audit_attachments",
                "work_orders", "work_order_items",
                "changes", "change_impacts", "change_reads",
                "issues", "issues_out", "issue_logs"]},
    {"key": "tasks", "label": "📝 업무·게시판·티켓", "desc": "일일업무·태스크·게시판·티켓·알림·활동",
     "tables": ["tasks", "task_comments", "task_delegations", "task_reactions", "activities",
                "team_summaries", "team_history",
                "board_posts", "board_comments", "comment_mentions",
                "tickets", "ticket_comments", "notifications"]},
]


def reset_groups_preview() -> list:
    """그룹별로 '지워질 행 수' 미리보기 (변경 없음)."""
    out = []
    with db_session() as c:
        existing = set(r[0] for r in c.execute(
            "SELECT name FROM sqlite_master WHERE type='table'").fetchall())
        for g in RESET_GROUPS:
            tabs = [t for t in g["tables"] if t in existing and t not in RESET_KEEP_TABLES]
            total = 0
            per = []
            for t in tabs:
                try:
                    n = c.execute(f'SELECT COUNT(*) FROM "{t}"').fetchone()[0]
                except Exception:
                    n = 0
                total += n
                if n:
                    per.append({"table": t, "count": n})
            per.sort(key=lambda x: -x["count"])
            out.append({"key": g["key"], "label": g["label"], "desc": g["desc"],
                        "tables": per, "table_count": len(tabs), "total_rows": total})
    return out


def reset_groups_apply(group_keys, actor_name: str = "") -> dict:
    """선택한 그룹만 비움. DB 백업 먼저 → 그룹 테이블 DELETE. 사용자/설정/마스터 보존.
    반환: {ok, backup, cleared:{label:행수}, groups:[label], error?}"""
    import shutil, datetime as _dt
    keys = set(group_keys or [])
    sel = [g for g in RESET_GROUPS if g["key"] in keys]
    if not sel:
        return {"ok": False, "error": "선택된 항목이 없습니다. 하나 이상 선택하세요."}
    backup = ""
    try:
        bdir = os.path.join(os.path.dirname(DB_PATH), "backups")
        os.makedirs(bdir, exist_ok=True)
        ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = os.path.join(bdir, f"knk.db.bak_resetgrp_{ts}")
        shutil.copy2(DB_PATH, backup)
    except Exception as e:
        return {"ok": False, "error": f"백업 실패 — 중단: {e}"}
    try:
        cleared = {}
        with db_session() as c:
            try:
                c.execute("PRAGMA foreign_keys=OFF")
            except Exception:
                pass
            existing = set(r[0] for r in c.execute(
                "SELECT name FROM sqlite_master WHERE type='table'").fetchall())
            for g in sel:
                cnt = 0
                for t in g["tables"]:
                    if t not in existing or t in RESET_KEEP_TABLES:
                        continue
                    try:
                        n = c.execute(f'SELECT COUNT(*) FROM "{t}"').fetchone()[0]
                    except Exception:
                        n = 0
                    try:
                        c.execute(f'DELETE FROM "{t}"')
                        c.execute("DELETE FROM sqlite_sequence WHERE name=?", (t,))
                        cnt += n
                    except Exception:
                        pass
                cleared[g["label"]] = cnt
            try:
                c.execute(
                    "INSERT INTO app_settings(key, value) VALUES('demo_reset_at', ?) "
                    "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                    (f"{_dt.datetime.now().strftime('%Y-%m-%d %H:%M')} by {actor_name} "
                     f"(항목초기화: {', '.join(g['key'] for g in sel)})",),
                )
            except Exception:
                pass
        return {"ok": True, "backup": backup, "cleared": cleared,
                "groups": [g["label"] for g in sel]}
    except Exception as e:
        return {"ok": False, "error": str(e), "backup": backup}


# v5H226z567 (대표 지시): 소모품 전체 초기화 — 소모품 발주/품목/이력 + '소모품 세금계산서'(ref_kind='consumable')만.
#   프로젝트 세금계산서(묶음)는 보존(tax_invoices 통째 삭제 금지). 깨끗이 재등록(z564 발주일코드·z566 고객사분리·z563 형태)용.
def reset_consumables_preview() -> dict:
    """소모품 초기화로 지워질 행 수 미리보기(변경 없음). 소모품 세금계산서는 ref_kind='consumable' 라인 기준."""
    out = {"consumable_orders": 0, "consumable_order_items": 0, "consumable_history": 0,
           "tax_invoices": 0, "tax_invoice_lines": 0}
    with db_session() as c:
        existing = set(r[0] for r in c.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall())
        for t in ("consumable_orders", "consumable_order_items", "consumable_history"):
            if t in existing:
                try:
                    out[t] = c.execute(f'SELECT COUNT(*) FROM "{t}"').fetchone()[0]
                except Exception:
                    out[t] = 0
        if "tax_invoice_lines" in existing:
            try:
                ti_ids = [r[0] for r in c.execute(
                    "SELECT DISTINCT ti_id FROM tax_invoice_lines WHERE ref_kind='consumable'").fetchall()]
                if ti_ids:
                    ph = ",".join("?" * len(ti_ids))
                    out["tax_invoice_lines"] = c.execute(
                        f"SELECT COUNT(*) FROM tax_invoice_lines WHERE ti_id IN ({ph})", ti_ids).fetchone()[0]
                    if "tax_invoices" in existing:
                        out["tax_invoices"] = c.execute(
                            f"SELECT COUNT(*) FROM tax_invoices WHERE id IN ({ph})", ti_ids).fetchone()[0]
            except Exception:
                pass
    out["total"] = sum(v for k, v in out.items() if k != "total")
    return out


def reset_consumables_full(actor_name: str = "") -> dict:
    """소모품 전체 초기화. DB 백업 먼저 → 소모품 세금계산서(라인+장부) 삭제 → 소모품 발주/품목/이력 삭제.
    프로젝트 세금계산서는 손대지 않음(ref_kind='project'/'unit'). 반환 {ok, backup, deleted:{label:n}}."""
    import shutil, datetime as _dt
    backup = ""
    try:
        bdir = os.path.join(os.path.dirname(DB_PATH), "backups")
        os.makedirs(bdir, exist_ok=True)
        ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = os.path.join(bdir, f"knk.db.bak_co_reset_{ts}")
        shutil.copy2(DB_PATH, backup)
    except Exception as e:
        return {"ok": False, "error": f"백업 실패 — 중단: {e}"}
    try:
        deleted = {}
        with db_session() as c:
            try:
                c.execute("PRAGMA foreign_keys=OFF")
            except Exception:
                pass
            existing = set(r[0] for r in c.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall())
            # 1) 소모품 세금계산서(ref_kind='consumable' 라인이 있는 장부)만 삭제 — 프로젝트 묶음 보존
            if "tax_invoice_lines" in existing:
                ti_ids = [r[0] for r in c.execute(
                    "SELECT DISTINCT ti_id FROM tax_invoice_lines WHERE ref_kind='consumable'").fetchall()]
                if ti_ids:
                    ph = ",".join("?" * len(ti_ids))
                    nlines = c.execute(f"SELECT COUNT(*) FROM tax_invoice_lines WHERE ti_id IN ({ph})", ti_ids).fetchone()[0]
                    c.execute(f"DELETE FROM tax_invoice_lines WHERE ti_id IN ({ph})", ti_ids)
                    deleted["소모품 세금계산서 라인"] = nlines
                    if "tax_invoices" in existing:
                        nti = c.execute(f"SELECT COUNT(*) FROM tax_invoices WHERE id IN ({ph})", ti_ids).fetchone()[0]
                        c.execute(f"DELETE FROM tax_invoices WHERE id IN ({ph})", ti_ids)
                        deleted["소모품 세금계산서"] = nti
            # 2) 소모품 발주/품목/이력
            for t in ("consumable_history", "consumable_order_items", "consumable_orders"):
                if t in existing:
                    n = c.execute(f'SELECT COUNT(*) FROM "{t}"').fetchone()[0]
                    c.execute(f'DELETE FROM "{t}"')
                    c.execute("DELETE FROM sqlite_sequence WHERE name=?", (t,))
                    deleted[t] = n
            try:
                c.execute(
                    "INSERT INTO app_settings(key, value) VALUES('demo_reset_at', ?) "
                    "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                    (f"{_dt.datetime.now().strftime('%Y-%m-%d %H:%M')} by {actor_name} (소모품 전체 초기화)",))
            except Exception:
                pass
        return {"ok": True, "backup": backup, "deleted": deleted}
    except Exception as e:
        return {"ok": False, "error": str(e), "backup": backup}


# v5H226z193 (대표 지시): 매출·영업 '사업부별' 초기화 (자동화 M만 등).
#   해당 사업부 프로젝트 + 딸린 수주·납품·수금·송장·생산·수출(CI/PL/BL/통관/FTA) 연쇄삭제.
#   견적(quotations)은 project 직접연결 컬럼이 없어(현 데이터 orders.quote_id 전부 NULL) 사업부 귀속 불가 → 제외.
#   PO·작업지시·이슈·변경·업무·재고 등 타 그룹 데이터는 projects_delete_logi 가 SET NULL(보존·언링크).
SALES_DIVISIONS = [
    {"key": "M", "label": "🤖 자동화 (M)",     "where": "biz_div='M'"},
    {"key": "T", "label": "🔬 검사기 (T)",     "where": "biz_div='T'"},
    {"key": "L", "label": "💄 라이프밸류 (L)", "where": "biz_div='L'"},
    {"key": "E", "label": "🌐 기타 (E)",       "where": "(biz_div='E' OR project_type='OTHER')"},
    {"key": "C", "label": "📦 소모품 (C)",     "where": "(biz_div='C' OR project_type='CONSUMABLE')"},
]


# v5H226z739 (대표 지시): '고아 수주' = orders.project_id 가 NULL 이거나, 가리키는 프로젝트가
#   이미 삭제돼 없는 수주(=관리번호는 사라졌는데 수주번호만 남은 잔재). 사업부별 초기화는 'projects 기준'
#   (projects WHERE biz_div=? → 그 프로젝트의 수주 삭제)이라 이런 고아 수주를 못 잡아 완전 삭제가 안 됐음.
def _count_orphan_orders(c) -> int:
    try:
        return c.execute(
            "SELECT COUNT(*) FROM orders WHERE project_id IS NULL "
            "OR project_id NOT IN (SELECT id FROM projects)").fetchone()[0]
    except Exception:
        return 0


def _sweep_orphan_orders(c) -> int:
    """고아 수주 + 그 자식(order_id 참조 테이블 전체)을 동적 스윕으로 일괄 삭제. 반환=삭제된 수주 수.
    z736 delete_order 와 동일 패턴(테이블이 추가돼도 FK 로 안 죽음). 호출 측에서 PRAGMA foreign_keys=OFF 권장."""
    try:
        oids = [r[0] for r in c.execute(
            "SELECT id FROM orders WHERE project_id IS NULL "
            "OR project_id NOT IN (SELECT id FROM projects)").fetchall()]
    except Exception:
        return 0
    if not oids:
        return 0
    try:
        tbls = [r[0] for r in c.execute(
            "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
    except Exception:
        tbls = []
    for t in tbls:
        if t in ("orders", "sqlite_sequence"):
            continue
        try:
            tc = {r[1] for r in c.execute(f"PRAGMA table_info({t})").fetchall()}
        except Exception:
            continue
        if "order_id" in tc:
            for oid in oids:
                try:
                    c.execute(f"DELETE FROM {t} WHERE order_id=?", (oid,))
                except Exception:
                    pass
    for oid in oids:
        try:
            c.execute("DELETE FROM orders WHERE id=?", (oid,))
        except Exception:
            pass
    return len(oids)


def reset_orphan_orders_count() -> int:
    """고아 수주 수 (미리보기 표시용)."""
    with db_session() as c:
        return _count_orphan_orders(c)


def reset_sales_division_preview() -> list:
    """사업부별 프로젝트/수주 수 미리보기 (변경 없음)."""
    out = []
    with db_session() as c:
        for d in SALES_DIVISIONS:
            try:
                n = c.execute(f"SELECT COUNT(*) FROM projects WHERE {d['where']}").fetchone()[0]
            except Exception:
                n = 0
            try:
                o = c.execute(
                    f"SELECT COUNT(*) FROM orders WHERE project_id IN "
                    f"(SELECT id FROM projects WHERE {d['where']})").fetchone()[0]
            except Exception:
                o = 0
            out.append({"key": d["key"], "label": d["label"], "projects": n, "orders": o})
    return out


def reset_sales_division_apply(divisions, actor_name: str = "") -> dict:
    """선택 사업부의 매출 데이터 초기화. 자동 백업 → 수출체인 삭제 → 프로젝트별 연쇄삭제(force).
    반환: {ok, backup, deleted_projects, export_deleted, failed, divisions, error?}"""
    import shutil, datetime as _dt
    sel = [d for d in SALES_DIVISIONS if d["key"] in set(divisions or [])]
    if not sel:
        return {"ok": False, "error": "선택된 사업부가 없습니다."}
    where = " OR ".join(f"({d['where']})" for d in sel)
    backup = ""
    try:
        bdir = os.path.join(os.path.dirname(DB_PATH), "backups")
        os.makedirs(bdir, exist_ok=True)
        ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = os.path.join(bdir, f"knk.db.bak_resetdiv_{ts}")
        shutil.copy2(DB_PATH, backup)
    except Exception as e:
        return {"ok": False, "error": f"백업 실패 — 중단: {e}"}

    def _inq(n):
        return ",".join("?" * n)
    try:
        # ── phase 1: 대상 pid 수집 + 수출 체인 삭제(수출은 order_id 로 연결 → orders 삭제 전에 처리)
        pids = []
        exp_deleted = 0
        with db_session() as c:
            try:
                c.execute("PRAGMA foreign_keys=OFF")
            except Exception:
                pass
            pids = [r[0] for r in c.execute(f"SELECT id FROM projects WHERE {where}").fetchall()]
            if pids:
                oids = [r[0] for r in c.execute(
                    f"SELECT id FROM orders WHERE project_id IN ({_inq(len(pids))})", pids).fetchall()]
                if oids:
                    eoids = [r[0] for r in c.execute(
                        f"SELECT id FROM export_orders WHERE order_id IN ({_inq(len(oids))})", oids).fetchall()]
                    if eoids:
                        eq = _inq(len(eoids))
                        pl_ids = [r[0] for r in c.execute(
                            f"SELECT id FROM packing_lists WHERE export_order_id IN ({eq})", eoids).fetchall()]
                        cert_ids = [r[0] for r in c.execute(
                            f"SELECT id FROM fta_certificates WHERE export_order_id IN ({eq})", eoids).fetchall()]
                        if pl_ids:
                            try: c.execute(f"DELETE FROM packing_items WHERE pl_id IN ({_inq(len(pl_ids))})", pl_ids)
                            except Exception: pass
                        if cert_ids:
                            try: c.execute(f"DELETE FROM fta_certificate_items WHERE cert_id IN ({_inq(len(cert_ids))})", cert_ids)
                            except Exception: pass
                        for t in ["commercial_invoices", "packing_lists", "bills_of_lading",
                                  "customs_declarations", "fta_certificates"]:
                            try: c.execute(f'DELETE FROM "{t}" WHERE export_order_id IN ({eq})', eoids)
                            except Exception: pass
                        try:
                            c.execute(f"DELETE FROM export_orders WHERE id IN ({eq})", eoids)
                            exp_deleted = len(eoids)
                        except Exception:
                            pass
        if not pids:
            return {"ok": True, "backup": backup, "deleted_projects": 0, "export_deleted": 0,
                    "failed": 0, "divisions": [d["label"] for d in sel], "note": "해당 사업부 프로젝트가 없습니다."}

        # ── phase 2: 프로젝트별 연쇄삭제 (force=관리코드 잠금 우회). 수주·납품·수금·생산 등 함께 삭제.
        deleted = 0
        failed = 0
        for pid in pids:
            try:
                projects_delete_logi(pid, force=True)
                deleted += 1
            except Exception:
                failed += 1

        # ── phase 3 (v5H226z739·대표 지시): 고아 수주 정리 — 사업부 초기화는 'projects 기준'이라
        #    프로젝트가 사라진 뒤 남은 수주(관리번호 없이 수주번호만 존재)를 못 잡음. project_id 가
        #    NULL/끊긴 수주 + 자식을 함께 삭제해야 '완벽 삭제'가 됨. (어느 사업부에도 안 걸리는 잔재)
        orphan_deleted = 0
        try:
            with db_session() as c:
                try:
                    c.execute("PRAGMA foreign_keys=OFF")
                except Exception:
                    pass
                orphan_deleted = _sweep_orphan_orders(c)
        except Exception:
            orphan_deleted = 0

        try:
            with db_session() as c:
                c.execute(
                    "INSERT INTO app_settings(key, value) VALUES('demo_reset_at', ?) "
                    "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                    (f"{_dt.datetime.now().strftime('%Y-%m-%d %H:%M')} by {actor_name} "
                     f"(사업부매출초기화: {','.join(d['key'] for d in sel)}, 고아수주 {orphan_deleted})",),
                )
        except Exception:
            pass
        return {"ok": True, "backup": backup, "deleted_projects": deleted, "export_deleted": exp_deleted,
                "failed": failed, "divisions": [d["label"] for d in sel],
                "orphan_deleted": orphan_deleted}
    except Exception as e:
        return {"ok": False, "error": str(e), "backup": backup}


def _reset_clear_tables(c):
    tabs = [r[0] for r in c.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
    ).fetchall()]
    return sorted(t for t in tabs if t not in RESET_KEEP_TABLES)


def _users_has_empno(c):
    try:
        return any(r[1] == "employee_no" for r in c.execute("PRAGMA table_info(users)").fetchall())
    except Exception:
        return False


def reset_demo_preview() -> dict:
    """무엇이 지워질지 미리보기 (변경 없음)."""
    with db_session() as c:
        clear = _reset_clear_tables(c)
        rows = []
        total = 0
        for t in clear:
            try:
                n = c.execute(f'SELECT COUNT(*) FROM "{t}"').fetchone()[0]
            except Exception:
                n = 0
            total += n
            if n:
                rows.append({"table": t, "count": n})
        rows.sort(key=lambda x: -x["count"])
        has_empno = _users_has_empno(c)
        demo_users = []
        if has_empno:
            demo_users = [dict(r) for r in c.execute(
                f"SELECT id, name, login_id, role, employee_no FROM users "
                f"WHERE {_RESET_DEMO_USER_WHERE} ORDER BY id"
            ).fetchall()]
            keep_users = c.execute(f"SELECT COUNT(*) FROM users WHERE NOT ({_RESET_DEMO_USER_WHERE})").fetchone()[0]
        else:
            keep_users = c.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        return {
            "tables": rows, "total_rows": total, "clear_table_count": len(clear),
            "demo_users": demo_users, "keep_users": keep_users,
            "has_empno": has_empno,
            "keep_tables": sorted(RESET_KEEP_TABLES - {"sqlite_sequence"}),
        }


def reset_demo_apply(actor_name: str = "") -> dict:
    """실제 초기화. DB 파일 백업 먼저 → 비움 → 데모사용자 삭제 → demo_seed_off=1.
    반환: {ok, backup, deleted_users, cleared_tables, error?}"""
    import shutil, datetime as _dt
    backup = ""
    try:
        bdir = os.path.join(os.path.dirname(DB_PATH), "backups")
        os.makedirs(bdir, exist_ok=True)
        ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = os.path.join(bdir, f"knk.db.bak_reset_{ts}")
        shutil.copy2(DB_PATH, backup)
    except Exception as e:
        return {"ok": False, "error": f"백업 실패 — 중단: {e}"}

    try:
        with db_session() as c:
            try:
                c.execute("PRAGMA foreign_keys=OFF")
            except Exception:
                pass
            clear = _reset_clear_tables(c)
            for t in clear:
                try:
                    c.execute(f'DELETE FROM "{t}"')
                    c.execute("DELETE FROM sqlite_sequence WHERE name=?", (t,))
                except Exception:
                    pass
            udel = 0
            if _users_has_empno(c):
                cur = c.execute(f"DELETE FROM users WHERE {_RESET_DEMO_USER_WHERE}")
                udel = cur.rowcount if cur.rowcount and cur.rowcount > 0 else 0
            c.execute(
                "INSERT INTO app_settings(key, value) VALUES('demo_seed_off','1') "
                "ON CONFLICT(key) DO UPDATE SET value='1'"
            )
            try:
                c.execute(
                    "INSERT INTO app_settings(key, value) VALUES('demo_reset_at', ?) "
                    "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                    (f"{_dt.datetime.now().strftime('%Y-%m-%d %H:%M')} by {actor_name}",),
                )
            except Exception:
                pass
        return {"ok": True, "backup": backup, "deleted_users": udel, "cleared_tables": len(clear)}
    except Exception as e:
        return {"ok": False, "error": str(e), "backup": backup}

