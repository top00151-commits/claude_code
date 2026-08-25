#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
후보 ①-a 변환기 시험 — BUDS 표본(005M2606)을 골든으로 전 규칙 검증
실행: python test_bom_inventor_to_partlist.py   → "시험 N건 · 실패 0" 이어야 통과
"""
import os
import random
import sys
import tempfile

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.dirname(HERE))
from app.bom_tools import inventor_to_partlist as B   # noqa: E402
from openpyxl import load_workbook, Workbook  # noqa: E402

KIT = os.path.normpath(os.path.join(HERE, "..", "..", "참고자료", "설계팀",
                                    "BOM 업무 자동화_2026.08.04", "1. BUDS 단차 검사기"))
INV = os.path.join(KIT, "INVENTOR DOWN")
UNITS = [os.path.join(INV, u + ".xlsx") for u in ("AA00", "AB00", "AC00", "AD00")]
# 골든(사람 완성본 60줄)은 키트 실물 — 동봉 양식(DEFAULT_TEMPLATE)은 자료를 비운 판
TPL = os.path.join(KIT, "000M0000 구매품 PART LIST_2026.00.00.xlsx")
TMP = tempfile.mkdtemp(prefix="knk_bomconv_")

FAIL = []
CNT = 0


def check(name, ok, detail=""):
    global CNT
    CNT += 1
    print(("  ✅" if ok else "  ❌") + f" {CNT:02d} {name}" + ("" if ok else f" — {detail}"))
    if not ok:
        FAIL.append(name)


def golden():
    """간이판 견본(사람이 완성한 60줄)을 읽는다."""
    ws = load_workbook(TPL, data_only=True).active
    out = []
    for r in range(8, ws.max_row + 1):
        if ws.cell(row=r, column=5).value in (None, ""):
            continue
        out.append({"구분": str(ws.cell(row=r, column=3).value or "").strip(),
                    "CODE": str(ws.cell(row=r, column=4).value or "").strip(),
                    "품명": str(ws.cell(row=r, column=5).value or "").strip(),
                    "형번": str(ws.cell(row=r, column=6).value or "").strip(),
                    "제조사": str(ws.cell(row=r, column=7).value or "").strip(),
                    "수량": ws.cell(row=r, column=9).value})
    return out


def mutate(src, dst, fn):
    wb = load_workbook(src, data_only=True)
    fn(wb["BOM"] if "BOM" in wb.sheetnames else wb.active)
    wb.save(dst)
    return dst


print("=" * 66)
print("  후보 ①-a 변환기 시험 (골든 = BUDS 005M2606 실물)")
print("=" * 66)

rows, excluded, per_file = B.read_inventor_files(UNITS)
OUT = os.path.join(TMP, "out.xlsx")
res = B.write_partlist(rows, "005M2606", "BUDS 단차 검사기", "한재운", OUT)
G = golden()

# Ⓐ1 줄 수
check("변환 품목 수 = 간이판 견본과 동일(60줄)", len(rows) == len(G) == 60, f"{len(rows)} vs {len(G)}")

# Ⓐ2 형번 집합
si, sg = {x["형번"] for x in rows}, {x["형번"] for x in G}
check("형번 집합 완전 일치(57종)", si == sg and len(si) == 57, f"{len(si)} vs {len(sg)} 차이 {list(si ^ sg)[:3]}")

# Ⓐ3 공유 형번의 구분·CODE·제조사·수량합
def bykey(rs, qk):
    d = {}
    for x in rs:
        e = d.setdefault(x["형번"], {"구분": x["구분"], "CODE": x["CODE"], "제조사": x["제조사"], "합": 0})
        e["합"] += float(x[qk] or 0)
    return d
di, dg = bykey(rows, "수량"), bykey(G, "수량")
bad = [k for k in dg if (di[k]["구분"], di[k]["CODE"], di[k]["제조사"], di[k]["합"])
       != (dg[k]["구분"], dg[k]["CODE"], dg[k]["제조사"], dg[k]["합"])]
check("구분·CODE·제조사·수량합 57/57 일치", not bad, str(bad[:3]))

# Ⓐ4 품명 — 인벤터에 있던 것은 그대로, 빈칸 7건은 사람 몫 목록으로
named_same = all(di and (x["품명"] == "" or any(g["품명"] == x["품명"] and g["형번"] == x["형번"] for g in G))
                 for x in rows)
check("품명: 인벤터에 있던 53건 그대로 보존", named_same)
check("품명 빈칸 7건이 사람 몫 목록에 보고됨", len(res["품명빈칸"]) == 7, str(len(res["품명빈칸"])))

# Ⓐ5 제외 보고
check("제외 보고 = 가공품61·용접1·프로파일1", excluded == {"가공품": 61, "WELDING FRMAE": 1, "AL PROFILE Ass'y": 1}, str(excluded))

# Ⓐ6 번호·자리표시 줄
wo = load_workbook(OUT).active
nums = [wo.cell(row=r, column=2).value for r in range(8, 8 + len(rows))]
check("번호 1..60 연속", nums == list(range(1, 61)))
nr = 8 + len(rows)
check("맨 끝 「전장 구매품 별도」 자리표시 + 파란 배경",
      wo.cell(row=nr, column=3).value == "전장 구매품 별도"
      and str(wo.cell(row=nr, column=3).fill.start_color.rgb) == "FF00B0F0"
      and wo.cell(row=nr, column=2).value == 61)

# Ⓐ7 양식 보존 — 머리글·제목·작성자
wt = load_workbook(TPL).active
hdr_same = all(wo.cell(row=7, column=ci).value == wt.cell(row=7, column=ci).value for ci in range(2, 17))
check("머리글(7행 B~P) 템플릿 그대로", hdr_same)
check("제목·AUTHOR 반영", "005M2606 구매품 BOM" in str(wo["C2"].value)
      and "BUDS 단차 검사기" in str(wo["C2"].value) and wo["C5"].value == "AUTHOR : 한재운")

# Ⓐ8 열 순서를 섞어도 같은 결과 (칸 이름으로 찾으므로)
def shuffle_cols(ws):
    ncol = ws.max_column
    perm = list(range(1, ncol + 1))
    random.Random(42).shuffle(perm)
    data = [[ws.cell(row=r, column=c).value for c in perm] for r in range(1, ws.max_row + 1)]
    for r in range(1, ws.max_row + 1):
        for i in range(ncol):
            ws.cell(row=r, column=i + 1).value = data[r - 1][i]
sh = mutate(UNITS[0], os.path.join(TMP, "AA00_섞음.xlsx"), shuffle_cols)
r1, _, _ = B.read_inventor_files([UNITS[0]])
r2, _, _ = B.read_inventor_files([sh])
strip = lambda rs: [{k: v for k, v in x.items() if k != "원본"} for x in rs]
check("열 순서 섞은 사본 → 동일한 결과", strip(r1) == strip(r2))

# Ⓐ9 필수 칸이 없으면 한국어로 명확히 중단
def kill_spec(ws):
    for ci in range(1, ws.max_column + 1):
        if str(ws.cell(row=1, column=ci).value or "").strip() == "부품 번호":
            ws.cell(row=1, column=ci).value = "없어진칸"
bad_f = mutate(UNITS[0], os.path.join(TMP, "AA00_칸없음.xlsx"), kill_spec)
try:
    B.read_inventor_files([bad_f])
    check("필수 칸 없으면 중단", False, "에러가 나지 않음")
except ValueError as e:
    check("필수 칸 없으면 중단 + 무슨 칸인지 말해줌", "형번" in str(e))

# Ⓐ10 수량은 숫자(int)로
check("수량 전부 숫자형", all(isinstance(x["수량"], int) for x in rows))

# Ⓐ11 처음 보는 구분 값도 조용히 사라지지 않고 보고에 남음
def odd_gubun(ws):
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=4).value or "").strip() == "가공품":
            ws.cell(row=r, column=4).value = "신규구분"
            break
odd = mutate(UNITS[0], os.path.join(TMP, "AA00_신규구분.xlsx"), odd_gubun)
_, ex2, _ = B.read_inventor_files([odd])
check("낯선 구분 값이 제외 보고에 표시됨", ex2.get("신규구분") == 1, str(ex2))

# Ⓐ12 서보모터 파란 표시 (전장 협의 — 실물 간이판의 파란 줄과 동일 관행)
check("서보모터 2줄 파란 표시 + 보고", len(res["서보표시"]) == 2
      and all(str(wo.cell(row=8 + no - 1, column=3).fill.start_color.rgb) == "FF00B0F0"
              for no, _, _ in res["서보표시"]))

# Ⓐ13 산출 파일 재열기
wo2 = load_workbook(OUT, data_only=True).active
n2 = sum(1 for r in range(8, wo2.max_row + 1) if wo2.cell(row=r, column=6).value not in (None, ""))
check("산출 파일 다시 열어 60줄 읽힘", n2 == 60, str(n2))

# Ⓐ14 중복 형번은 줄대로 (합산 금지 — 허동준 매니저 규칙)
check("중복 형번 줄대로 유지(60줄 > 57종)", len(rows) == 60 and len(si) == 57)

# Ⓐ15 폴더 입력이면 '정리용' 파일은 제외 (겹침 방지)
rf, _, pf = B.read_inventor_files([INV])
check("폴더 입력: 정리용 제외 4파일 → 같은 60줄", len(rf) == 60 and len(pf) == 4,
      f"{len(rf)}줄 {len(pf)}파일")

# Ⓐ16 구매품이 0줄이면 안내
def all_off(ws):
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=4).value or "").strip() == "구매품":
            ws.cell(row=r, column=4).value = "가공품"
zero = mutate(UNITS[3], os.path.join(TMP, "AD00_구매품없음.xlsx"), all_off)
rz, _, _ = B.read_inventor_files([zero])
check("구매품 0줄이면 빈 결과(안내는 main에서)", rz == [])

print("-" * 66)
print(f"  시험 {CNT}건 · 실패 {len(FAIL)}건" + ("" if not FAIL else " → " + ", ".join(FAIL)))
sys.exit(0 if not FAIL else 1)
