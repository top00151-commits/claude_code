#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""①-c 마스터 개정 시험 — 골든 = BUDS 실물 사슬 + 008M 실측 관행(B열 낱말·T=0)

시나리오: BUDS 4유닛으로 옛 마스터를 만들고 구매팀 손 기입을 심은 뒤,
AA00 인벤터 BOM 만 실제로 수정(수량변경·삭제·형번변경·추가·실패품 제거)해서 올린다.
AB00·AC00·AD00 은 안 올림 → 한 칸도 안 변해야 한다(전장 무접촉과 같은 원리).

실행: python _검증/test_bom_revise_master.py → "실패 0" 이어야 통과
"""
import os
import shutil
import sys
import tempfile

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.dirname(HERE))
from app.bom_tools import inventor_to_partlist as A   # noqa: E402
from app.bom_tools import merge_to_master as B        # noqa: E402
from app.bom_tools import revise_master as R          # noqa: E402
from openpyxl import load_workbook                    # noqa: E402

KIT = os.path.normpath(os.path.join(HERE, "..", "..", "참고자료", "설계팀",
                                    "BOM 업무 자동화_2026.08.04", "1. BUDS 단차 검사기", "INVENTOR DOWN"))
TMP = tempfile.mkdtemp(prefix="knk_bomrev_")
INV = [os.path.join(KIT, u + ".xlsx") for u in ("AA00", "AB00", "AC00", "AD00")]

FAIL = []
CNT = 0


def check(name, ok, detail=""):
    global CNT
    CNT += 1
    print(("  ✅" if ok else "  ❌") + f" {CNT:02d} {name}" + ("" if ok else f" — {detail}"))
    if not ok:
        FAIL.append(name)


def cells(ws, r):
    return [str(ws.cell(row=r, column=c).value or "") for c in range(2, 28)]


print("=" * 66)
print("  ①-c 마스터 개정 시험 (골든 = BUDS 사슬 + 008M 관행)")
print("=" * 66)

# ── 옛 마스터 만들기 (실물 사슬 그대로) ──
rows, _exc, _pf = A.read_inventor_files(INV)
draft = os.path.join(TMP, "간이판.xlsx")
A.write_partlist(rows, "A005M2606", "BUDS 단차 검사기", "시험", draft)
mrows, _rep = B.read_draft_files([draft])
old = os.path.join(TMP, "옛마스터.xlsx")
B.write_master(mrows, "A005M2606", "BUDS 단차 검사기", "시험", old, sets=1)

wo = load_workbook(old)
wso = wo.active
tot0 = R._find_total_row(wso)
items0 = R._scan_items(wso, tot0)
aa = [x for x in items0 if x["D"] == "AA00"]
ab = [x for x in items0 if x["D"] == "AB00"]
# AA00 안에서 형번이 유일한 줄만 표적으로 (중복 형번 짝지음 꼬임 방지)
specs = {}
for x in aa:
    specs[R._norm_spec(x["F"])] = specs.get(R._norm_spec(x["F"]), 0) + 1
uniq = [x for x in aa if specs[R._norm_spec(x["F"])] == 1 and x["F"]]
check("준비: AA00 유일 형번 줄 5개 이상 · AB00 존재", len(uniq) >= 5 and len(ab) >= 3,
      f"{len(uniq)}/{len(ab)}")
t_keep, t_qty, t_del, t_ren, t_fail = uniq[0], uniq[1], uniq[2], uniq[3], uniq[4]

# 구매팀 손 기입 심기 (보존 검증 대상)
HAND = {t_keep["r"]: {8: "광원전기", 16: 1000, 18: 1100, 20: 1100, 17: "2주"},
        t_qty["r"]: {16: 5000, 8: "파익스"},
        t_del["r"]: {16: 6000, 20: 7777},
        t_ren["r"]: {8: "미스미", 20: 3300},
        ab[0]["r"]: {8: "한국미스미", 16: 2500, 20: 2600}}
for r, kv in HAND.items():
    for ci, v in kv.items():
        wso.cell(row=r, column=ci, value=v)
wso.cell(row=t_fail["r"], column=2, value="실패")      # 실물 관행: 실패 줄 표시(매입액 포함)
wo.save(old)
snap = load_workbook(old).active                        # 개정 전 스냅샷
snap_rows = {}                                          # 안 올린 유닛 줄 = 형번키 → 전 칸
for x in items0:
    if x["D"] in ("AB00", "AC00", "AD00"):
        snap_rows[(x["D"], R._norm_spec(x["F"]), snap_rows.get((x["D"], R._norm_spec(x["F"])), (0,))[0])] = None
snap_map = [(x["D"], x["F"], cells(snap, x["r"])) for x in items0 if x["D"] != "AA00"]

# ── AA00 인벤터 BOM 실제 수정 ──
mod = os.path.join(TMP, "AA00.xlsx")
shutil.copy(INV[0], mod)
wm = load_workbook(mod)
wsm = wm["BOM"] if "BOM" in wm.sheetnames else wm.active
hr, col, _miss = A._find_columns(wsm)


def find_row(spec):
    for r in range(hr + 1, wsm.max_row + 1):
        if str(wsm.cell(row=r, column=col["형번"]).value or "").strip() == spec:
            return r
    return None


qcol = col.get("수량") or col.get("수량예비")
r_qty = find_row(t_qty["F"])
old_q = wsm.cell(row=r_qty, column=qcol).value
wsm.cell(row=r_qty, column=qcol, value=(old_q or 0) + 3)                    # 수량 변경
r_ren = find_row(t_ren["F"])
wsm.cell(row=r_ren, column=col["형번"], value=t_ren["F"] + "N2")            # 형번 변경(품명·제조사 유지)
r_src = find_row(t_keep["F"])
new_r = wsm.max_row + 1                                                     # 부품 추가
wsm.cell(row=new_r, column=col["구분"], value="구매품")
wsm.cell(row=new_r, column=col["주제"], value=wsm.cell(row=r_src, column=col["주제"]).value)
wsm.cell(row=new_r, column=col["code"], value="AA00")
wsm.cell(row=new_r, column=col["품명"], value="NEW PART X")
wsm.cell(row=new_r, column=col["형번"], value="NEW-SPEC-77")
wsm.cell(row=new_r, column=col["제조사"], value="KNK")
wsm.cell(row=new_r, column=qcol, value=2)
for r in sorted([find_row(t_del["F"]), find_row(t_fail["F"])], reverse=True):
    wsm.delete_rows(r)                                                      # 삭제 2건(하나는 실패 기표시)
wm.save(mod)

# ── 개정 실행 ──
out = os.path.join(TMP, "개정본.xlsx")
rep = R.revise(old, [mod], out)
check("보고: 추가1·삭제표시1·수량변경1·형번개정1·이미표시(실패)1",
      len(rep["추가"]) == 1 and len(rep["삭제표시"]) == 1 and len(rep["수량변경"]) == 1
      and len(rep["형번개정"]) == 1 and len(rep["이미표시"]) == 1,
      str({k: len(v) for k, v in rep.items() if isinstance(v, list)}))
check("보고: 안 건드린 유닛 = AB00·AC00·AD00",
      rep["안건드린유닛"] == ["AB00", "AC00", "AD00"], str(rep["안건드린유닛"]))
# 실측 실재 사례: AB00.xlsx 안의 code=AA00 교차 줄(BY3-56-3) → 본 블록 밖 = 무접촉 + 보고
check("본 블록 밖 같은 유닛 줄(BY3-56-3): 삭제표시 대신 무접촉 보고",
      len(rep["블록밖"]) == 1 and rep["블록밖"][0][2] == "BY3-56-3", str(rep["블록밖"]))

wr = load_workbook(out).active
tot1 = R._find_total_row(wr)
items1 = R._scan_items(wr, tot1)
by_spec = {}
for x in items1:
    by_spec.setdefault((x["D"], R._norm_spec(x["F"])), []).append(x)


def one(d, f):
    lst = by_spec.get((d, R._norm_spec(f)), [])
    return lst[0] if lst else None


# 수량 변경 줄 — J만 바뀌고 손 기입 보존
x = one("AA00", t_qty["F"])
check("수량변경 줄: J 갱신 + P·협력사 손기입 보존",
      x and wr.cell(row=x["r"], column=10).value == (old_q or 0) + 3
      and wr.cell(row=x["r"], column=16).value == 5000
      and wr.cell(row=x["r"], column=8).value == "파익스", str(x))

# 삭제 줄 — 남기고 B=삭제·T=0 (008M 실측 관행), P·수량 보존
x = one("AA00", t_del["F"])
check("삭제 줄: 남김 + B=「삭제」", x is not None and x["B"] == "삭제", str(x))
check("삭제 줄: T 7777→0 + 기존단가 6000 보존 + 보고 T내림",
      x and wr.cell(row=x["r"], column=20).value == 0
      and wr.cell(row=x["r"], column=16).value == 6000
      and rep["T내림"] and rep["T내림"][0][2] == 7777)

# 형번 개정 줄 — F=「옛->새」 제자리, 손 기입 보존
x = one("AA00", t_ren["F"] + "N2")
check("형번 개정: F=「옛->새」 제자리 기입", x and x["F"] == t_ren["F"] + "->" + t_ren["F"] + "N2",
      str(x and x["F"]))
check("형번 개정 줄: 협력사·확정단가 손기입 보존",
      x and wr.cell(row=x["r"], column=8).value == "미스미"
      and wr.cell(row=x["r"], column=20).value == 3300)

# 실패 기표시 줄 — 부품이 사라졌어도 B열 안 덮고 보고만
x = one("AA00", t_fail["F"])
check("기표시(실패) 줄: B열 안 덮음 + 이미표시 보고",
      x and x["B"] == "실패" and rep["이미표시"][0][1] == "실패", str(x))

# 추가 줄 — AA00 블록 안 + B=추가 + 수식 자기줄
x = one("AA00", "NEW-SPEC-77")
aa_rows = [y["r"] for y in items1 if y["D"] == "AA00"]
check("추가 줄: AA00 블록 안 + B=「추가」 + 수량·EA",
      x and x["B"] == "추가" and min(aa_rows) <= x["r"] <= max(aa_rows)
      and wr.cell(row=x["r"], column=10).value == 2
      and wr.cell(row=x["r"], column=12).value == "EA", str(x))
check("추가 줄: 실물 수식 5종 자기줄 참조",
      x and all(wr.cell(row=x["r"], column=ci).value == mk(x["r"])
                for ci, mk in R.FORMULA_COLS.items()))

# 밀린 아래 줄 수식 자기줄 교정
below = [y for y in items1 if y["r"] > (x["r"] if x else 0) and y["D"] == "AB00"]
check("삽입 아래(AB00) 줄: K 수식 자기줄 참조로 교정",
      below and wr.cell(row=below[0]["r"], column=11).value == f"=J{below[0]['r']}*$J$5",
      str(below[0] if below else None))

# 안 올린 유닛(AB00·AC00·AD00) 전 칸 무접촉 — 수식 자기줄 재작성만 예외
diffs = []
for d, f, before in snap_map:
    now = one(d, f)
    if not now:
        diffs.append((d, f, "사라짐"))
        continue
    after = cells(wr, now["r"])
    for i, (a, b2) in enumerate(zip(before, after)):
        ci = i + 2
        if ci in R.FORMULA_COLS:          # 줄 밀림에 따른 자기줄 수식 교정은 허용
            want = R.FORMULA_COLS[ci](now["r"])
            if b2 != want:
                diffs.append((d, f, ci, a, b2))
        elif a != b2:
            diffs.append((d, f, ci, a, b2))
check("안 올린 유닛 전 칸 무접촉 (수식 자기줄 제외 차이 0)", not diffs, str(diffs[:4]))
check("AB00 손기입(협력사·단가) 보존", (lambda y: y and wr.cell(row=y["r"], column=8).value == "한국미스미"
      and wr.cell(row=y["r"], column=16).value == 2500)(one("AB00", ab[0]["F"])))

# 합계줄 SUBTOTAL 범위 교정
from openpyxl.utils import get_column_letter as gl
subs_ok = all(wr.cell(row=tot1, column=ci).value == f"=SUBTOTAL(9,{gl(ci)}9:{gl(ci)}{tot1 - 1})"
              for ci in R.SUBTOTAL_COLS)
check("합계줄 SUBTOTAL 범위 = 9~합계줄-1", subs_ok,
      str(wr.cell(row=tot1, column=10).value))

# 품목 수: 60 - 삭제0(남김) + 추가1 = 61
check("품목 줄 수 = 61 (삭제는 남기니 60+추가1)", len(items1) == 61, str(len(items1)))

# ── 무변경(no-op) 실행 — 원본 AA00 그대로 → 아무 것도 안 바뀜 ──
out2 = os.path.join(TMP, "개정_noop.xlsx")
rep2 = R.revise(old, [INV[0]], out2)
check("무변경 실행: 추가·삭제·수량·형번 전부 0 · 수식보정 0",
      not rep2["추가"] and not rep2["삭제표시"] and not rep2["수량변경"]
      and not rep2["형번개정"] and rep2["수식보정"] == 0,
      str({k: len(v) for k, v in rep2.items() if isinstance(v, list)}))
w2 = load_workbook(out2).active
snap2 = load_workbook(old).active
same = all(cells(w2, r) == cells(snap2, r) for r in range(9, tot0 + 1))
check("무변경 실행: 전 칸 원본과 동일", same)

# ── 규칙 밖 입력 ──
try:
    R.revise(INV[0], [mod], os.path.join(TMP, "x.xlsx"))
    check("통일판 아닌 마스터 → 한국어 안내 오류", False, "오류 없음")
except ValueError as e:
    check("통일판 아닌 마스터 → 한국어 안내 오류", "합계줄" in str(e) or "찾지" in str(e), str(e)[:40])

print("-" * 66)
print(f"  시험 {CNT}건 · 실패 {len(FAIL)}건" + ("" if not FAIL else " → " + ", ".join(FAIL)))
sys.exit(0 if not FAIL else 1)
