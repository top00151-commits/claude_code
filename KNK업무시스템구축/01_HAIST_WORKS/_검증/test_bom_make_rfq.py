#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
후보 ② (2/2) 견적요청서 생성기 시험 — 골든 = 허동준 실물 양식(견본 3줄 포함) + 001M 마스터
실행: python test_bom_make_rfq.py → "시험 N건 · 실패 0" 이어야 통과
"""
import os
import sys
import tempfile

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.dirname(HERE))
from app.bom_tools import make_rfq as R      # noqa: E402
from app.bom_tools.make_po import read_master  # noqa: E402
from openpyxl import load_workbook           # noqa: E402

DL = r"C:\Users\top00\Downloads"
GOLD = os.path.join(DL, "매입처 견적 의뢰 요청 자료.xlsx")
MASTER = os.path.join(DL, [x for x in os.listdir(DL)
                           if x.startswith("001M2606") and "PART LIST" in x][0])
TMP = tempfile.mkdtemp(prefix="knk_rfq_")

FAIL = []
CNT = 0


def check(name, ok, detail=""):
    global CNT
    CNT += 1
    print(("  ✅" if ok else "  ❌") + f" {CNT:02d} {name}" + ("" if ok else f" — {detail}"))
    if not ok:
        FAIL.append(name)


print("=" * 66)
print("  후보 ② 견적요청서 생성기 시험 (골든 = 허동준 매니저 실물)")
print("=" * 66)

# Ⓐ1 도출 양식 — 머리글 실물 그대로·값 비움·수식/합계 유지
wt = load_workbook(R.DEFAULT_TEMPLATE)[R.SHEET]
wg = load_workbook(GOLD)[R.SHEET]
hdr = all(str(wt.cell(row=r, column=ci).value or "") == str(wg.cell(row=r, column=ci).value or "")
          for r in (4, 5) for ci in range(2, 21))
check("머리글(4~5행) 실물 그대로", hdr)
check("자료 값 비움 + 번호·'M'·'EA'·금액수식 유지",
      wt.cell(row=6, column=7).value in (None, "") and wt.cell(row=6, column=2).value == 1
      and wt.cell(row=6, column=4).value == "M" and wt.cell(row=6, column=14).value == "EA"
      and str(wt.cell(row=6, column=15).value) == "=K6*M6"
      and str(wt.cell(row=6, column=16).value) == "=L6*M6")
tr = R._find_total_row(wt)
check("합계줄 SUBTOTAL 유지", str(wt.cell(row=tr, column=11).value).startswith("=SUBTOTAL(9,K6:"))

# Ⓐ4 실물 견본 3줄이 우리 수식 규칙과 같은가 (금액 = 단가×요청수량)
wv = load_workbook(GOLD, data_only=True)[R.SHEET]
rule = all((wv.cell(row=r, column=15).value or 0)
           == (wv.cell(row=r, column=11).value or 0) * (wv.cell(row=r, column=13).value or 0)
           for r in (6, 7, 8))
check("실물 견본 3줄: 기존금액 = 기존단가×요청수량 (규칙 일치)", rule)

# Ⓐ5 001M 성화기전 75줄 생성
code, name, sets, rows, vina = read_master(MASTER)
sh = [x for x in rows if x["협력사"] == "성화기전"]
OUT = os.path.join(TMP, "rfq_성화기전.xlsx")
res = R.write_rfq(sh, "성화기전", code, OUT, due="2026-08-29")
w = load_workbook(OUT, data_only=True)[R.SHEET]
check("성화기전 75줄 생성", res["품목"] == 75 == len(sh))
check("관리번호 3분할 = 001·M·2606", (w.cell(row=6, column=3).value, w.cell(row=6, column=4).value,
                                   w.cell(row=6, column=5).value) == ("001", "M", "2606"))
check("순번 1..75 연속", [w.cell(row=6 + i, column=2).value for i in range(75)] == list(range(1, 76)))

# Ⓐ8 기존단가 = 마스터 P(실적가) 그대로 · 변경단가는 비움
ok_price = all((w.cell(row=6 + i, column=11).value or None)
               == (sh[i]["기존단가"] if sh[i]["기존단가"] not in (None, "", 0) else None)
               for i in range(75))
check("기존단가 = 마스터 실적가(P) 그대로 (빈칸은 빈칸)", ok_price)
check("변경단가(L)·소요일(R)·입고가능일(S) 전부 비움 — 매입처 몫",
      all(w.cell(row=6 + i, column=ci).value in (None, "") for i in range(75) for ci in (12, 18, 19)))
# 실적가(P) 빈칸 = 9줄 — 인터뷰 전 실측 "기존단가 빈칸: 성화기전 9" 그대로 (30은 확정단가 T 빈칸 수)
check("신규 단가 요청 9줄 보고 (실적가 빈칸 실측 그대로)", res["신규단가요청"] == 9, str(res["신규단가요청"]))

# Ⓐ11 「옛->새」 유지 (발주서와 반대 규칙)
arrows = sum(1 for i in range(75) if "->" in str(w.cell(row=6 + i, column=8).value or "")
             or "->" in str(w.cell(row=6 + i, column=7).value or "")
             or "->" in str(w.cell(row=6 + i, column=9).value or ""))
check("「옛->새」 표기 그대로 유지 + 보고", arrows > 0 and res["변경표기유지"] == 26, f"{arrows}/{res['변경표기유지']}")

# Ⓐ12 입고요청일 반영
check("입고요청일 = --due 값", str(w.cell(row=6, column=17).value) == "2026-08-29")

# Ⓐ13 파익스 — 어제 실제 메일과 같은 대상: 4줄·신규 2줄
pk = [x for x in rows if x["협력사"] == "파익스"]
OUT2 = os.path.join(TMP, "rfq_파익스.xlsx")
res2 = R.write_rfq(pk, "파익스", code, OUT2, due="2026-08-29")
check("파익스 4줄 · 실적가 없는 신규 2줄 (실측 그대로)", res2["품목"] == 4 and res2["신규단가요청"] == 2,
      str(res2))

# Ⓐ14 --only-missing 흐름
only = [x for x in pk if x["기존단가"] in (None, "", 0)]
OUT3 = os.path.join(TMP, "rfq_파익스_신규만.xlsx")
res3 = R.write_rfq(only, "파익스", code, OUT3)
check("신규 품목만 뽑기 → 2줄", res3["품목"] == 2)

# Ⓐ15 칸 수 초과 → 한국어 안내
try:
    R.write_rfq(sh * 3, "성화기전", code, os.path.join(TMP, "x.xlsx"))
    check("칸 수 초과면 중단", False, "에러 없음")
except ValueError as e:
    check("칸 수 초과면 한국어 안내로 중단", "나눠서" in str(e))

# Ⓐ16 시험용 관리번호(A접두)도 분해
check("A001M2607 → A001·M·2607", R._split_code("A001M2607") == ("A001", "M", "2607"))

# Ⓐ17 재열기
w2 = load_workbook(OUT, data_only=True)[R.SHEET]
n = sum(1 for r in range(6, 200) if w2.cell(row=r, column=8).value not in (None, ""))
check("재열기 75줄", n == 75, str(n))

print("-" * 66)
print(f"  시험 {CNT}건 · 실패 {len(FAIL)}건" + ("" if not FAIL else " → " + ", ".join(FAIL)))
sys.exit(0 if not FAIL else 1)
