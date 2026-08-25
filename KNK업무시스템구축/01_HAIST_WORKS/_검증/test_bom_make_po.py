#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
후보 ② 발주서 생성기 시험 — 골든 = 001M 실물 마스터 + 실제 나간 광원전기 발주서(8/20)
실행: python test_bom_make_po.py → "시험 N건 · 실패 0" 이어야 통과
"""
import datetime
import os
import sys
import tempfile
from collections import Counter

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import bom_make_po as P                       # noqa: E402
from openpyxl import load_workbook            # noqa: E402

DL = r"C:\Users\top00\Downloads"


def dlfind(prefix, needle=""):
    m = [x for x in os.listdir(DL) if x.startswith(prefix) and needle in x and x.endswith(".xlsx")]
    assert m, (prefix, needle)
    return os.path.join(DL, m[0])


MASTER = dlfind("001M2606", "PART LIST")
GOLD_PO = dlfind("001M2606", "발주서")
TMP = tempfile.mkdtemp(prefix="knk_po_")

FAIL = []
CNT = 0


def check(name, ok, detail=""):
    global CNT
    CNT += 1
    print(("  ✅" if ok else "  ❌") + f" {CNT:02d} {name}" + ("" if ok else f" — {detail}"))
    if not ok:
        FAIL.append(name)


print("=" * 66)
print("  후보 ② 발주서 생성기 시험 (골든 = 실제 나간 광원전기 발주서)")
print("=" * 66)

# Ⓐ1 도출 양식 — 값 비움·수식/번호/도장 유지·잡시트 제거
wt = load_workbook(P.DEFAULT_TEMPLATE)
check("도출 양식 시트 = 구매요청서 하나", wt.sheetnames == [P.SHEET])
ws = wt[P.SHEET]
check("품목 값 비움 + 단위 EA 슬롯 + 번호·금액수식 유지",
      all(ws.cell(row=r, column=2).value in (None, "") for r in range(12, 52))
      and ws.cell(row=12, column=6).value == "EA"
      and ws.cell(row=12, column=1).value == 1
      and str(ws.cell(row=12, column=8).value).startswith("=E12"))
check("도장 그림 보존", len(ws._images) == 1)
check("머리 라벨만 남고 값 비움 · 구매팀 연락처는 유지",
      str(ws["A4"].value).strip().endswith(":") and str(ws["A5"].value).strip().endswith(":")
      and "buy1" in str(ws["H7"].value))

# Ⓐ5 마스터 읽기 — 제목에서 관리번호·이름 자동
code, name, sets, rows, vina = P.read_master(MASTER)
check("제목칸에서 관리번호·프로젝트명 자동 인식", code == "001M2606" and "Clip" in name, f"{code}/{name}")
check("VINA 줄 제외 + 수 보고", vina >= 13, str(vina))

# Ⓐ7 골든 — 광원전기 발주서 재현
kw = [x for x in rows if x["협력사"] == "광원전기"]
OUT = os.path.join(TMP, "gen_광원전기.xlsx")
res = P.write_po(kw, "광원전기", code, OUT, date=datetime.date(2026, 8, 20), due="2026-08-25")
check("광원전기 산 줄 32건 → 발주서 32줄", res["품목"] == 32, str(res["품목"]))

def po_items(path):
    w = load_workbook(path, data_only=True)[P.SHEET]
    out = []
    for r in range(12, 200):
        pn = w.cell(row=r, column=2).value
        spec = w.cell(row=r, column=3).value
        if pn in (None, "") and spec in (None, ""):
            continue
        if str(w.cell(row=r, column=1).value or "") == "합 계":
            break
        out.append((str(pn or "").strip(), str(spec or "").strip(),
                    str(w.cell(row=r, column=4).value or "").strip(),
                    float(w.cell(row=r, column=5).value or 0),
                    float(w.cell(row=r, column=7).value or 0)))
    return out

gen = po_items(OUT)
real = po_items(GOLD_PO)
check("품목표가 실물 발주서와 똑같음 (품명·규격·MAKER·수량·단가 32줄 전부)",
      Counter(gen) == Counter(real),
      f"차이 {list((Counter(real) - Counter(gen)).items())[:2]} / {list((Counter(gen) - Counter(real)).items())[:2]}")
check("줄 순서까지 동일", gen == real)
check("합계 금액 = 실물 캐시값 1,819,170원", res["합계"] == 1819170, f"{res['합계']:,}")
check("「옛->새」 변경표기는 새 값만 + 정리 보고", res.get("변경정리", 0) >= 2, str(res.get("변경정리")))

# Ⓐ11 머리
wg = load_workbook(OUT)[P.SHEET]
check("머리: 관리코드·업체명·요청납기·발주일자",
      str(wg["A4"].value).endswith("001M2606") and str(wg["A5"].value).endswith("광원전기")
      and str(wg["A8"].value).endswith("2026-08-25")
      and str(wg["I4"].value)[:10] == "2026-08-20")

# Ⓐ12 성화기전 66+줄 — 슬롯(40) 초과 확장
sh = [x for x in rows if x["협력사"] == "성화기전"]
OUT2 = os.path.join(TMP, "gen_성화기전.xlsx")
res2 = P.write_po(sh, "성화기전", code, OUT2, date=datetime.date(2026, 8, 25))
w2 = load_workbook(OUT2, data_only=False)[P.SHEET]
tr2 = P._find_total_row(w2)
gen2 = po_items(OUT2)
check(f"성화기전 {len(sh)}줄(슬롯 40 초과) 전부 기록 + 줄 확장", len(gen2) == len(sh) and len(sh) > 40,
      f"{len(gen2)} vs {len(sh)}")
check("확장 후 합계 수식이 끝줄까지", str(w2.cell(row=tr2, column=8).value) == f"=SUM(H12:H{tr2 - 1})",
      str(w2.cell(row=tr2, column=8).value))
check("확장 후 번호 연속·금액 수식 전 줄", w2.cell(row=12 + len(sh) - 1, column=1).value == len(sh)
      and str(w2.cell(row=12 + len(sh) - 1, column=8).value).startswith("=E"))
check("확장 후 도장 그림 아래로 밀림", len(w2._images) == 1
      and w2._images[0].anchor._from.row > 54)

# Ⓐ16 태그줄 제외 · 단가 미확정 보고 (마스터 사본 조작)
mut = os.path.join(TMP, "master_mut.xlsx")
wbm = load_workbook(MASTER)
wsm = wbm.active
hit_tag = hit_price = None
for r in range(9, wsm.max_row + 1):
    if str(wsm.cell(row=r, column=8).value or "").strip() == "광원전기":
        if hit_tag is None:
            wsm.cell(row=r, column=2).value = "비교"
            hit_tag = r
        elif hit_price is None:
            wsm.cell(row=r, column=20).value = None
            hit_price = r
wbm.save(mut)
_, _, _, rows_m, _ = P.read_master(mut)
kw_m = [x for x in rows_m if x["협력사"] == "광원전기"]
OUT3 = os.path.join(TMP, "gen_mut.xlsx")
res3 = P.write_po(kw_m, "광원전기", "001M2606", OUT3)
check("「비교」 태그 줄 제외 (32→31)", len(kw_m) == 31, str(len(kw_m)))
check("확정단가 빈 줄 → 단가 빈칸 + 보고 1건", res3["단가미확정"] == 1, str(res3["단가미확정"]))
wv3 = load_workbook(OUT3, data_only=True)[P.SHEET]
prices3 = [wv3.cell(row=r, column=7).value for r in range(12, 12 + 31)]
check("빈 단가는 지어내지 않음(빈칸)", sum(1 for v in prices3 if v in (None, "")) == 1)

# Ⓐ19 수량 = 수량×대수 − 재고 (사본: 사내재고 1 넣기)
wbm2 = load_workbook(MASTER)
wsm2 = wbm2.active
for r in range(9, wsm2.max_row + 1):
    if str(wsm2.cell(row=r, column=8).value or "").strip() == "광원전기":
        wsm2.cell(row=r, column=13).value = 1        # M 사내재고
        base_qty = wsm2.cell(row=r, column=10).value
        break
mut2 = os.path.join(TMP, "master_mut2.xlsx")
wbm2.save(mut2)
_, _, _, rows_m2, _ = P.read_master(mut2)
kw2 = [x for x in rows_m2 if x["협력사"] == "광원전기"]
check("수량 = J×대수 − 사내재고 (재고 1 → 1 감소)", kw2[0]["수량"] == (base_qty or 0) - 1,
      f"{kw2[0]['수량']} vs {base_qty}-1")

# Ⓐ20 --all 흐름: 협력사 수 = 산 줄 있는 협력사 전부
vend = sorted({x["협력사"] for x in rows if x["협력사"]})
outdir = os.path.join(TMP, "all")
os.makedirs(outdir, exist_ok=True)
rc = P.main([MASTER, "--all", "--outdir", outdir, "--date", "2026-08-25"])
made = [f for f in os.listdir(outdir) if f.endswith(".xlsx")]
check(f"--all → 협력사 {len(vend)}곳 발주서 {len(vend)}건", rc == 0 and len(made) == len(vend),
      f"{len(made)} vs {len(vend)}")
check("파일명 규칙 = 관리번호 협력사 발주서_날짜", f"001M2606 광원전기 발주서_20260825.xlsx" in made)

# Ⓐ22 산 줄 없는 협력사 → 안내 에러
try:
    P.write_po([], "없는회사", "001M2606", os.path.join(TMP, "x.xlsx"))
    check("산 줄 없으면 중단", False, "에러 없음")
except ValueError as e:
    check("산 줄 없으면 한국어 안내로 중단", "없는회사" in str(e))

print("-" * 66)
print(f"  시험 {CNT}건 · 실패 {len(FAIL)}건" + ("" if not FAIL else " → " + ", ".join(FAIL)))
sys.exit(0 if not FAIL else 1)
