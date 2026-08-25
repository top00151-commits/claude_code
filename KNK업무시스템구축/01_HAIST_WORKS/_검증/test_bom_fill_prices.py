#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
후보 ③ 단가 사전 시험 — 골든 = 001M 실물(허동준 매니저 손 기입) vs 수불부 실데이터
결정적 시험: P를 지운 001M을 자동으로 채워서 **본인 기입값과 100% 일치**하는지.
실행: python test_bom_fill_prices.py → "시험 N건 · 실패 0" 이어야 통과
"""
import datetime
import os
import sys
import tempfile

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import bom_fill_prices as F                  # noqa: E402
from openpyxl import load_workbook, Workbook  # noqa: E402

DL = r"C:\Users\top00\Downloads"
MASTER = os.path.join(DL, [x for x in os.listdir(DL)
                           if x.startswith("001M2606") and "PART LIST" in x][0])
LEDGER = os.path.join(DL, "2026년 08~12월 수불부_구매 작업_20260818-3.xlsx")
TMP = tempfile.mkdtemp(prefix="knk_price_")

FAIL = []
CNT = 0


def check(name, ok, detail=""):
    global CNT
    CNT += 1
    print(("  ✅" if ok else "  ❌") + f" {CNT:02d} {name}" + ("" if ok else f" — {detail}"))
    if not ok:
        FAIL.append(name)


print("=" * 66)
print("  후보 ③ 단가 사전 시험 (골든 = 001M 손 기입 vs 수불부 실데이터)")
print("=" * 66)

# Ⓐ1 수불부 읽기
by_vk, by_k, n = F.read_ledger(LEDGER)
check("수불부 실적 2,532건 · 열쇠 717/704", n == 2532 and len(by_vk) == 717 and len(by_k) == 704,
      f"{n}/{len(by_vk)}/{len(by_k)}")

# Ⓐ2 결정적 골든 — P를 지우고 자동 기입 → 본인 손 기입과 대조
wm = load_workbook(MASTER)
ws = wm.active
truth = {}
for r in range(9, ws.max_row + 1):
    # 품목 줄만 — 합계줄(P175 SUBTOTAL 수식)은 진리값이 아니다
    if ws.cell(row=r, column=5).value in (None, "") and ws.cell(row=r, column=6).value in (None, ""):
        continue
    v = ws.cell(row=r, column=16).value
    if v not in (None, "", 0):
        truth[r] = v
        ws.cell(row=r, column=16).value = None
BLANK = os.path.join(TMP, "m_blank.xlsx")
wm.save(BLANK)
OUT = os.path.join(TMP, "m_filled.xlsx")
res = F.fill_prices(BLANK, LEDGER, OUT)
# 127 = 원기입 재현 117 + 신규 10 (수불부가 8/25까지라 마스터 저장 뒤 발주된 품목까지 실림 — 도구가 사람보다 최신)
check("빈 마스터에 채움 127줄 (원기입 117 + 이후 발주분 10)", res["채움"] == 127, str(res["채움"]))
check("타매입처 실적 사용 0줄 (전부 그 협력사 실적)", res["타매입처"] == 0, str(res["타매입처"]))

wf = load_workbook(OUT, data_only=True).active
same = diff = 0
for r, tv in truth.items():
    fv = wf.cell(row=r, column=16).value
    if fv in (None, "", 0):
        continue
    if float(fv) == float(tv):
        same += 1
    else:
        diff += 1
check("⭐ 채워진 값이 허동준 매니저 손 기입과 100% 일치 (117/117)", same == 117 and diff == 0,
      f"같음 {same} 다름 {diff}")
check("실적 없는 신규는 빈칸 + 목록 보고", len(res["신규"]) > 0
      and all(wf.cell(row=r, column=16).value in (None, "") for r, _, _ in res["신규"]))

# Ⓐ6 이미 적힌 값은 안 덮는다 + 변동 보고 (원본 그대로 넣으면 채움 0·변동 0 = 100% 재현의 다른 얼굴)
OUT2 = os.path.join(TMP, "m_orig.xlsx")
res2 = F.fill_prices(MASTER, LEDGER, OUT2)
check("이미 적힌 마스터 → 채움은 빈칸(10줄 중 실적 있는 0)·변동 후보 0", res2["변동후보"] == [],
      str(res2["변동후보"][:2]))
w2 = load_workbook(OUT2, data_only=True).active
untouched = all((w2.cell(row=r, column=16).value == truth[r]) for r in truth)
check("적힌 기존단가 전부 그대로 (덮지 않음)", untouched)

# Ⓐ8 변동 감지 — 수불부에 실적이 **있는** 줄 하나를 일부러 틀리게
wm3 = load_workbook(MASTER)
ws3 = wm3.active
target = None
for r in truth:
    spec = F._norm(ws3.cell(row=r, column=6).value)
    ven = str(ws3.cell(row=r, column=8).value or "").strip()
    if (spec, ven) in by_vk:
        target = r
        break
ws3.cell(row=target, column=16).value = 99999999
MUT = os.path.join(TMP, "m_mut.xlsx")
wm3.save(MUT)
res3 = F.fill_prices(MUT, LEDGER, os.path.join(TMP, "m_mut_out.xlsx"))
check("적힌 값 ≠ 최근 실적 → 변동 후보로 보고(덮지 않음)",
      len(res3["변동후보"]) == 1 and res3["변동후보"][0][0] == target, str(res3["변동후보"][:1]))

# Ⓐ9 매입처 이원화 — 같은 형번·다른 협력사·다른 단가 (합성 수불부)
def mini_ledger(path, rows):
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "발주관리대장"
    ws0.append([])
    ws0.append([])
    ws0.append([])
    ws0.cell(row=4, column=1, value="규격\n(형번)")
    ws0.cell(row=4, column=2, value="협력사")
    ws0.cell(row=4, column=3, value="발주\n일자")
    ws0.cell(row=4, column=4, value="단가")
    for i, (s, v, d, p) in enumerate(rows, start=5):
        ws0.cell(row=i, column=1, value=s)
        ws0.cell(row=i, column=2, value=v)
        ws0.cell(row=i, column=3, value=d)
        ws0.cell(row=i, column=4, value=p)
    wb.save(path)
    return path

ML = mini_ledger(os.path.join(TMP, "mini.xlsx"), [
    ("ABC-1", "가업체", datetime.datetime(2026, 8, 1), 1000),
    ("ABC-1", "나업체", datetime.datetime(2026, 8, 20), 2000),   # 다른 매입처가 더 최근
    ("ABC-1", "가업체", datetime.datetime(2026, 8, 10), 1200),   # 가업체의 최근
])
bv, bk, _ = F.read_ledger(ML)
check("매입처 이원화: 그 협력사 것 우선 (가업체→1,200 · 형번만→2,000)",
      bv[("ABC-1", "가업체")][1] == 1200 and bk["ABC-1"][1] == 2000)
check("최근 발주일 우선 (가업체 8/10 값)", bv[("ABC-1", "가업체")][0].date() == datetime.date(2026, 8, 10))

# Ⓐ11 산출 파일 무결성 — P열 밖 칸은 그대로
wa = load_workbook(BLANK, data_only=True).active
wb_ = load_workbook(OUT, data_only=True).active
sample_ok = all(wa.cell(row=r, column=ci).value == wb_.cell(row=r, column=ci).value
                for r in range(9, 60) for ci in (3, 5, 6, 7, 8, 10, 20))
check("기존단가(P) 밖의 칸은 손대지 않음", sample_ok)

# Ⓐ12 협력사 미정(뼈대 단계) — 채우지 않고 과거 실적 참고 목록으로
SC = os.path.join(os.environ.get("TMP", TMP))
skel_src = None
for base in (r"C:\Users\top00\AppData\Local\Temp\claude\C--Users-top00-JR-Claude---\72ba0949-a2b2-4048-9f20-763eff42b897\scratchpad",):
    p = os.path.join(base, "001M2606 통일판 뼈대(전장).xlsx")
    if os.path.exists(p):
        skel_src = p
if skel_src:
    res_sk = F.fill_prices(skel_src, LEDGER, os.path.join(TMP, "skel_out.xlsx"))
    check("전장 뼈대(협력사 미정): 채움 0 + 과거 실적 참고 41건(51줄 중 80%)",
          res_sk["채움"] == 0 and len(res_sk["협력사미정참고"]) == 41,
          f"{res_sk['채움']}/{len(res_sk['협력사미정참고'])}")
    wsk = load_workbook(os.path.join(TMP, "skel_out.xlsx"), data_only=True).active
    check("뼈대의 기존단가는 여전히 전부 빈칸 (판단 대행 안 함)",
          all(wsk.cell(row=r, column=16).value in (None, "") for r in range(9, 60)))
else:
    check("전장 뼈대 표본 없음 — 건너뜀(수동 확인 필요)", False, "뼈대 파일 경로 없음")

# Ⓐ14 잘못된 수불부 → 한국어 에러
bad = os.path.join(TMP, "bad.xlsx")
Workbook().save(bad)
try:
    F.read_ledger(bad)
    check("수불부 아니면 중단", False, "에러 없음")
except ValueError as e:
    check("수불부 아니면 한국어로 중단", "발주관리대장" in str(e))

print("-" * 66)
print(f"  시험 {CNT}건 · 실패 {len(FAIL)}건" + ("" if not FAIL else " → " + ", ".join(FAIL)))
sys.exit(0 if not FAIL else 1)
