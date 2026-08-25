#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
후보 ①-b 취합 시험 — 골든 = 001M2606 실물 마스터(전장 49줄)·008M 머리글·전장 BOM 실물
실행: python test_bom_merge_to_master.py → "시험 N건 · 실패 0" 이어야 통과
"""
import os
import sys
import tempfile

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.dirname(HERE))
from app.bom_tools import merge_to_master as B     # noqa: E402
from app.bom_tools import inventor_to_partlist as A  # noqa: E402
from openpyxl import load_workbook                 # noqa: E402

DL = r"C:\Users\top00\Downloads"


def dlfind(prefix, needle=""):
    m = [x for x in os.listdir(DL) if x.startswith(prefix) and needle in x and x.endswith(".xlsx")]
    assert m, (prefix, needle)
    return os.path.join(DL, m[0])


M001 = dlfind("001M2606", "PART LIST")
M008 = dlfind("008M2603", "PART LIST")
EJ = dlfind("2026.08.13_Crip")                     # 전장 BOM 실물
KIT = os.path.normpath(os.path.join(HERE, "..", "..", "참고자료", "설계팀",
                                    "BOM 업무 자동화_2026.08.04", "1. BUDS 단차 검사기"))
TMP = tempfile.mkdtemp(prefix="knk_bommerge_")
TPL = B.DEFAULT_TEMPLATE

FAIL = []
CNT = 0


def check(name, ok, detail=""):
    global CNT
    CNT += 1
    print(("  ✅" if ok else "  ❌") + f" {CNT:02d} {name}" + ("" if ok else f" — {detail}"))
    if not ok:
        FAIL.append(name)


def norm(s):
    return str(s or "").strip().split("->")[-1].strip().upper().replace(" ", "")


print("=" * 66)
print("  후보 ①-b 취합 시험 (골든 = 001M·008M·전장 BOM 실물)")
print("=" * 66)

# Ⓐ1~2 도출 양식 머리글 = 001M = 008M (칸 단위)
wt = load_workbook(TPL).active
w1 = load_workbook(M001).active
w8 = load_workbook(M008).active
diff1 = [(r, ci) for r in (7, 8) for ci in range(2, 28)
         if str(wt.cell(row=r, column=ci).value or "") != str(w1.cell(row=r, column=ci).value or "")]
diff8 = [(r, ci) for r in (7, 8) for ci in range(2, 28)
         if str(wt.cell(row=r, column=ci).value or "") != str(w8.cell(row=r, column=ci).value or "")]
check("도출 양식 머리글(7~8행 B~AA) == 001M 실물", not diff1, str(diff1[:4]))
check("도출 양식 머리글 == 008M 실물 (통일 보증)", not diff8, str(diff8[:4]))
wtv = load_workbook(TPL, data_only=True).active
empt = all(wtv.cell(row=r, column=ci).value in (None, "")
           for r in range(9, wtv.max_row + 1) for ci in range(2, 28))
check("도출 양식 자료 구역 완전 비움", empt)

# Ⓐ4 전장 BOM 단독 → 뼈대 51줄
rows, rep = B.read_draft_files([EJ])
OUT1 = os.path.join(TMP, "m001_전장뼈대.xlsx")
res = B.write_master(rows, "001M2606", "Clip Attach Machine", "시험", OUT1)
check("전장 BOM 51줄 → 뼈대 51줄", res["품목"] == 51 and len(rows) == 51, str(len(rows)))

wo = load_workbook(OUT1).active
vals = [(str(wo.cell(row=r, column=3).value), str(wo.cell(row=r, column=4).value)) for r in range(9, 9 + 51)]
check("구분·CODE 전 줄 = '전장 부품'·'전장'", all(v == ("전장 부품", "전장") for v in vals))
check("B열(상태낱말 자리)은 비워 둠", all(wo.cell(row=r, column=2).value in (None, "") for r in range(9, 9 + 51)))
check("협력사(H) 비어 있음 — 구매팀 몫", all(wo.cell(row=r, column=8).value in (None, "") for r in range(9, 9 + 51)))

# Ⓐ8 골든 대조 — 001M 실물 마스터의 '전장 부품' 49줄과 (알려진 사람 편집: 표기차·변경·케이블)
mast = {}
for r in range(9, w1.max_row + 1):
    wv = load_workbook(M001, data_only=True).active
    break
wv = load_workbook(M001, data_only=True).active
for r in range(9, wv.max_row + 1):
    if wv.cell(row=r, column=5).value in (None, ""):
        continue
    if str(wv.cell(row=r, column=3).value or "").strip() != "전장 부품":
        continue
    mast[norm(wv.cell(row=r, column=6).value)] = wv.cell(row=r, column=10).value
mine = {norm(x["형번"]): x["수량"] for x in rows}
shared = [k for k in mine if k in mast]
qty_same = sum(1 for k in shared if mine[k] == mast[k])
check("골든 대조: 실물 마스터와 형번 45종 공유(알려진 차이=표기·변경·케이블)", len(shared) == 45, str(len(shared)))
check("골든 대조: 공유 형번 수량 44/45 일치(실측 그대로)", qty_same == 44, str(qty_same))

# Ⓐ10 수식 — 실물과 같은 모양
r0, rl = 9, 9 + 51 - 1
fx = load_workbook(OUT1).active
check("자료줄 수식 5종 실물 동일",
      fx.cell(row=r0, column=11).value == f"=J{r0}*$J$5"
      and fx.cell(row=rl, column=15).value == f"=$K{rl}-($M{rl}+$N{rl})"
      and fx.cell(row=rl, column=21).value == f"=T{rl}*(M{rl}+N{rl})"
      and fx.cell(row=rl, column=22).value == f"=O{rl}*T{rl}"
      and fx.cell(row=rl, column=23).value == f"=U{rl}+V{rl}")
tr = res["합계줄"]
subs = [fx.cell(row=tr, column=ci).value for ci in B.SUBTOTAL_COLS]
check("합계줄 SUBTOTAL 11칸 (J K M N O P R T U V W)",
      all(isinstance(v, str) and v.startswith("=SUBTOTAL(9,") for v in subs)
      and fx.cell(row=tr, column=10).value == f"=SUBTOTAL(9,J9:J{tr - 1})")

# Ⓐ12 취합 — ①-a 산출(BUDS 기구) + 전장 BOM = 111줄 · 자리표시 줄 걷어냄
inv_rows, _, _ = A.read_inventor_files([os.path.join(KIT, "INVENTOR DOWN")])
DRAFT = os.path.join(TMP, "buds_간이판.xlsx")
A.write_partlist(inv_rows, "005M2606", "BUDS 단차 검사기", "한재운", DRAFT)
rows2, rep2 = B.read_draft_files([DRAFT, EJ])
OUT2 = os.path.join(TMP, "취합111.xlsx")
res2 = B.write_master(rows2, "005M2606", "기능시연(기구+전장)", "시험", OUT2)
check("두 파일 취합 60+51 = 111줄 · 순서 = 기구 먼저", res2["품목"] == 111
      and rows2[0]["구분"] == "BOTTOM FRAME" and rows2[-1]["구분"] == "전장 부품")
check("「전장 구매품 별도」 자리표시 줄 걷어내고 보고", rep2["자리표시제거"] == 1, str(rep2["자리표시제거"]))
check("중복 형번 줄대로 유지(합산 안 함)",
      len(rows2) == 111 and len({(x['형번']) for x in rows2}) < 111)

# Ⓐ15 단위 기본값 EA — 간이판 60줄만(전장 BOM 은 단위 EA 가 이미 채워져 옴, 실측 51/51)
check("단위 빈칸(간이판 60줄)만 'EA' 기본 + 건수 보고", res2["단위기본값"] == 60, str(res2["단위기본값"]))

# Ⓐ16 입력에 단가가 있으면 옮기지 않고 보고
def with_price(src, dst):
    wb = load_workbook(src)
    ws = wb.active
    ws.cell(row=9, column=11).value = 12345      # 간이판 K(단가)
    wb.save(dst)
    return dst
pf = with_price(DRAFT, os.path.join(TMP, "단가있음.xlsx"))
rows3, rep3 = B.read_draft_files([pf])
check("단가 값은 옮기지 않고 「옮기지 않음」 보고", any(k == "단가" for _, _, k, _ in rep3["옮기지않음"]),
      str(rep3["옮기지않음"][:2]))
ws3 = load_workbook(B.DEFAULT_TEMPLATE).active   # 뼈대에 단가 안 들어감은 write 결과로
OUT3 = os.path.join(TMP, "단가검증.xlsx")
B.write_master(rows3, "005M2606", "x", "x", OUT3)
w3 = load_workbook(OUT3, data_only=True).active
check("뼈대의 기존단가(P)·확정단가(T) 전부 빈칸", all(
    w3.cell(row=r, column=ci).value in (None, "") for r in range(9, 9 + len(rows3)) for ci in (16, 20)))

# Ⓐ18 엉뚱한 파일 → 한국어 에러
bad = os.path.join(TMP, "엉뚱.xlsx")
from openpyxl import Workbook
wbb = Workbook(); wbb.active["A1"] = "그냥 표"; wbb.save(bad)
try:
    B.read_draft_files([bad])
    check("간이판 아니면 중단", False, "에러 없음")
except ValueError as e:
    check("간이판 아니면 한국어로 중단", "머리글" in str(e))

# Ⓐ19 머리 정보·대수
w2 = load_workbook(OUT2).active
check("제목·AUTHOR·대수(J5) 반영", "005M2606 구매 BOM" in str(w2["C2"].value)
      and str(w2["C5"].value).startswith("AUTHOR :") and w2["J5"].value == 1)
OUT4 = os.path.join(TMP, "2대.xlsx")
B.write_master(rows, "001M2606", "x", "x", OUT4, sets=2)
check("--sets 2 → J5=2 (TOTAL=수량×2)", load_workbook(OUT4).active["J5"].value == 2)

# Ⓐ21 재열기
w2v = load_workbook(OUT2, data_only=True).active
n = sum(1 for r in range(9, w2v.max_row + 1) if w2v.cell(row=r, column=6).value not in (None, ""))
check("취합본 다시 열어 111줄 읽힘", n == 111, str(n))

print("-" * 66)
print(f"  시험 {CNT}건 · 실패 {len(FAIL)}건" + ("" if not FAIL else " → " + ", ".join(FAIL)))
sys.exit(0 if not FAIL else 1)
