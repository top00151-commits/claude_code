#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""WP-04 BOM 도구 공장 — 화면(라우트) 시험 (임시 DB · 운영 무접촉)

⛔ 운영 DB 를 열지 않는다. 임시 폴더 DB + 임시 보관함으로만 시험한다.
   인증(get_user)만 대체하고 인가·저장·다운로드는 실코드 그대로 누른다.
   관리번호는 전부 시험용 A 접두.

사용법:  python _검증/test_bom_tools_routes.py
"""
import io
import json
import os
import sys
import tempfile
import zipfile

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, ROOT)

import app.database as D                                   # noqa: E402

# 🔴 DB 경로는 모듈 상수 — 임시로 갈아끼운다 (환경변수 아님)
TMP = tempfile.mkdtemp(prefix="knk_bomtools_")
_REAL_DB = D.DB_PATH
D.DB_PATH = os.path.join(TMP, "test.db")
assert D.DB_PATH != _REAL_DB
D.init_db()

import app.main as appmain                                 # noqa: E402
from fastapi.testclient import TestClient                  # noqa: E402
from openpyxl import load_workbook                         # noqa: E402

# 보관함도 임시로 (실행 시점에 전역을 읽으므로 갈아끼우면 됨)
appmain._BT_STORE = os.path.join(TMP, "store")
os.makedirs(appmain._BT_STORE, exist_ok=True)

CUR = {"u": None}
appmain.get_user = lambda request: CUR["u"]
cl = TestClient(appmain.app, follow_redirects=False)

RUN = {"id": 701, "name": "구매팀원", "role": "member", "team_id": 10,
       "can_use_logistics": 1, "can_view_logistics": 1}
VIEW = {"id": 702, "name": "조회자", "role": "member", "team_id": 3,
        "can_use_logistics": 0, "can_view_logistics": 1}
# 조회 상시허용 팀(1·2·3·7·8·9·10)과 BOM 작성팀(4·5·6) 밖의 팀이어야 진짜 무권한
NONE = {"id": 703, "name": "무권한", "role": "member", "team_id": 99,
        "can_use_logistics": 0, "can_view_logistics": 0}
with D.db_session() as conn:
    for uu in (RUN, VIEW, NONE):
        conn.execute("INSERT INTO users(id, name, login_id, password, role) VALUES(?,?,?,'x',?)",
                     (uu["id"], uu["name"], f"T{uu['id']}", uu["role"]))

KIT = os.path.normpath(os.path.join(ROOT, "..", "참고자료", "설계팀",
                                    "BOM 업무 자동화_2026.08.04", "1. BUDS 단차 검사기"))
INV = os.path.join(KIT, "INVENTOR DOWN")
DL = r"C:\Users\top00\Downloads"
MASTER = os.path.join(DL, [x for x in os.listdir(DL) if x.startswith("001M2606") and "PART LIST" in x][0])
LEDGER = os.path.join(DL, "2026년 08~12월 수불부_구매 작업_20260818-3.xlsx")
XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def fpart(field, path):
    with open(path, "rb") as fh:
        return (field, (os.path.basename(path), fh.read(), XLSX))


FAIL = []
CNT = 0


def check(name, ok, detail=""):
    global CNT
    CNT += 1
    print(("  ✅" if ok else "  ❌") + f" {CNT:02d} {name}" + ("" if ok else f" — {detail}"))
    if not ok:
        FAIL.append(name)


def loc(r):
    return r.headers.get("location", "")


def last_run():
    with D.db_session() as c:
        r = c.execute("SELECT * FROM bom_tool_runs ORDER BY id DESC LIMIT 1").fetchone()
    return dict(r) if r else None


print("=" * 66)
print("  BOM 도구 공장 — 화면(라우트) 시험 (임시 DB)")
print("=" * 66)

# ── 화면 접근 ──
CUR["u"] = VIEW
r = cl.get("/bom/tools")
check("조회 권한: 화면 200 + 단계 카드 + 실행버튼 잠금 안내",
      r.status_code == 200 and "BOM 도구" in r.text and "인벤터 → 간이판" in r.text
      and "실행은" in r.text)
CUR["u"] = NONE
check("무권한: /home 으로 돌려보냄", cl.get("/bom/tools").status_code == 303)

# ── ①-a 실행 (실제 BUDS 인벤터 4파일 업로드) ──
CUR["u"] = RUN
files = [fpart("files", os.path.join(INV, u + ".xlsx")) for u in ("AA00", "AB00", "AC00", "AD00")]
r = cl.post("/bom/tools/run/draft", data={"code": "A005M2606", "name": "BUDS 시험", "author": "시험"},
            files=files)
check("①-a 실행 → 완료 안내로 돌아감", r.status_code == 303 and "msg=" in loc(r), loc(r)[:80])
run1 = last_run()
check("실행 기록 저장 (단계·관리번호·보고)", run1 and run1["step"] == "draft"
      and run1["mgmt_code"] == "A005M2606" and "품명 빈 줄 7건" in (run1["report"] or ""),
      str(run1 and run1["report"])[:60])
check("입력 4개 + 산출물이 웍스에 저장됨", run1 and len(json.loads(run1["inputs"])) == 4
      and os.path.exists(run1["output_path"]))

# ── 다운로드 (산출물·입력) ──
r = cl.get(f"/bom/tools/download/{run1['id']}?f=out")
check("산출물 다운로드 200", r.status_code == 200 and len(r.content) > 5000)
wo = load_workbook(io.BytesIO(r.content), data_only=True).active
n = sum(1 for rr in range(8, wo.max_row + 1) if wo.cell(row=rr, column=6).value not in (None, ""))
check("내려받은 간이판 = 60줄 (라이브러리 결과와 동일)", n == 60, str(n))
r = cl.get(f"/bom/tools/download/{run1['id']}?f=in0")
check("입력 파일도 언제든 다운로드", r.status_code == 200
      and len(r.content) == os.path.getsize(os.path.join(INV, "AA00.xlsx")))

# ── ③ 실행 (마스터 + 수불부 — 필드 분리) ──
r = cl.post("/bom/tools/run/price", data={},
            files=[fpart("files", MASTER), fpart("ledger", LEDGER)])
run2 = last_run()
check("③ 실행: 수불부 2,532건 대조 보고", r.status_code == 303 and run2["step"] == "price"
      and "2,532건 대조" in (run2["report"] or ""), (run2 or {}).get("report", "")[:60])

# ── ①-b 실행 (간이판 → 통일판 뼈대) — ①-c 의 옛 마스터 재료 ──
with open(run1["output_path"], "rb") as fh:
    _draft_bytes = fh.read()
r = cl.post("/bom/tools/run/master", data={"code": "A005M2606", "name": "BUDS 시험", "sets": "1"},
            files=[("files", ("간이판.xlsx", _draft_bytes, XLSX))])
run_b = last_run()
check("①-b 취합: 60줄 뼈대 기록", r.status_code == 303 and run_b["step"] == "master"
      and "60줄" in (run_b["title"] or ""), (run_b or {}).get("title", ""))

# ── ①-c 실행 (마스터 개정 — AA00 수량 1건 수정해 업로드) ──
import shutil
from app.bom_tools import inventor_to_partlist as _inv
mod = os.path.join(TMP, "AA00_수정.xlsx")
shutil.copy(os.path.join(INV, "AA00.xlsx"), mod)
_wm = load_workbook(mod)
_wsm = _wm["BOM"] if "BOM" in _wm.sheetnames else _wm.active
_hr, _col, _m = _inv._find_columns(_wsm)
_qc = _col.get("수량") or _col.get("수량예비")
for _r in range(_hr + 1, _wsm.max_row + 1):
    if str(_wsm.cell(row=_r, column=_col["구분"]).value or "").strip() == "구매품":
        _wsm.cell(row=_r, column=_qc, value=(_wsm.cell(row=_r, column=_qc).value or 0) + 5)
        break
_wm.save(mod)
with open(run_b["output_path"], "rb") as fh:
    _master_bytes = fh.read()
r = cl.post("/bom/tools/run/revise",
            files=[("master", ("옛마스터.xlsx", _master_bytes, XLSX)), fpart("files", mod)])
run_rev = last_run()
check("①-c 개정: 수량변경 1건 + 기록 저장", r.status_code == 303 and run_rev["step"] == "revise"
      and "수량변경 1건" in (run_rev["report"] or "") and "개정" in (run_rev["title"] or ""),
      loc(r)[:60] + " | " + (run_rev or {}).get("report", "")[:60])
zr0 = cl.get(f"/bom/tools/download/{run_rev['id']}?f=out")
check("①-c 산출물(_개정.xlsx) 다운로드", zr0.status_code == 200
      and run_rev["output_name"].endswith("_개정.xlsx"))
check("①-c 보고: 본 블록 밖 교차 줄(AB00 속 AA00) 무접촉 보고",
      "본 블록 밖" in (run_rev["report"] or ""), (run_rev or {}).get("report", "")[:100])

# ── ② 발주서 (협력사 한 곳) ──
r = cl.post("/bom/tools/run/po", data={"vendor": "광원전기", "due": "2026-09-01"},
            files=[fpart("files", MASTER)])
run3 = last_run()
check("② 발주서(광원전기): 32줄·1,819,170원 보고", r.status_code == 303
      and "광원전기 32줄" in (run3["report"] or "") and "1,819,170" in (run3["report"] or ""),
      (run3 or {}).get("report", "")[:80])

# ── ② 발주서 (전체 → zip 묶음) ──
r = cl.post("/bom/tools/run/po", data={"vendor": ""}, files=[fpart("files", MASTER)])
run4 = last_run()
zr = cl.get(f"/bom/tools/download/{run4['id']}?f=out")
zf = zipfile.ZipFile(io.BytesIO(zr.content))
check("② 발주서(전체): zip 안에 협력사 13곳 발주서", run4["output_name"].endswith(".zip")
      and len(zf.namelist()) == 13, str(len(zf.namelist())))

# ── ② 견적요청서 (파익스 — 어제 실제 메일과 같은 대상) ──
r = cl.post("/bom/tools/run/rfq", data={"vendor": "파익스", "due": "2026-09-01"},
            files=[fpart("files", MASTER)])
run5 = last_run()
check("② 견적요청서(파익스): 신규 2줄 보고", "신규 단가 요청 2줄" in (run5["report"] or ""),
      (run5 or {}).get("report", "")[:60])

# ── 권한·오류 ──
CUR["u"] = VIEW
r = cl.post("/bom/tools/run/draft", data={"code": "A1", "name": "x"},
            files=[fpart("files", os.path.join(INV, "AA00.xlsx"))])
check("조회 권한은 실행 불가(안내)", r.status_code == 303 and "err=" in loc(r))

CUR["u"] = RUN
before = last_run()["id"]
r = cl.post("/bom/tools/run/draft", data={"code": "A2", "name": "x"},
            files=[("files", ("엉뚱.xlsx", b"PK\x03\x04junk", XLSX))])
check("엉뚱한 파일: 한국어 안내 + 기록 안 남김", r.status_code == 303 and "err=" in loc(r)
      and last_run()["id"] == before)

# 저장소 밖 경로 조작 방어
with D.db_session() as c:
    c.execute("INSERT INTO bom_tool_runs(mgmt_code, step, title, inputs, output_name, output_path, report, created_by) "
              "VALUES('A0','draft','조작','[]','x.xlsx',?, '', 701)", (os.path.join(TMP, "밖의파일.xlsx"),))
    bad_id = c.execute("SELECT MAX(id) FROM bom_tool_runs").fetchone()[0]
r = cl.get(f"/bom/tools/download/{bad_id}?f=out")
check("보관함 밖 경로는 주지 않음", r.status_code == 303 and "err=" in loc(r))

# 기록 화면에 지금까지의 실행이 보인다
r = cl.get("/bom/tools")
check("기록 화면: 실행들이 표로 보임 + 다운로드 단추", "실행 기록" in r.text
      and "A005M2606" in r.text and "download" in r.text)

# 메뉴 입구 (대표 화면 지적 2026-08-25 "메뉴 경로가 안 보인다") — 센터 홈 카드 + 사이드바
r = cl.get("/logistics")
check("자재구매센터 홈 카드·사이드바에 「BOM 도구」 입구",
      r.status_code == 200 and r.text.count('href="/bom/tools"') >= 2 and "BOM 도구" in r.text,
      f"{r.status_code}/{r.text.count('href=\"/bom/tools\"')}")

print("-" * 66)
print(f"  시험 {CNT}건 · 실패 {len(FAIL)}건" + ("" if not FAIL else " → " + ", ".join(FAIL)))
# 임시 DB 확인 — 실제 DB 무변경
assert D.DB_PATH.startswith(TMP)
sys.exit(0 if not FAIL else 1)
