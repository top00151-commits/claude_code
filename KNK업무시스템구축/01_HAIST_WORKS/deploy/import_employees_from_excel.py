#!/usr/bin/env python3
"""
import_employees_from_excel.py — HAIST WORKS 130명 직원 일괄 import

발주서: KNK업무시스템구축/_TO_메신저세션/2026-05-29_사번SSO.md
작성: 빅터(01) · 2026-05-30

목적:
    KNK 직원등록 엑셀 2개 → HAIST WORKS users 테이블
    - KNK_직원등록(본사).xlsx 시트 '본사(KOR)' (총 14개 부서)
    - KNK_직원등록(베트남).xlsx 시트 '베트남법인(VN)' (총 10개 부서)

사용법:
    1. 로컬 dry-run (DB 변경 X, 매칭 결과만 확인):
        python deploy/import_employees_from_excel.py \\
            --kor "C:\\path\\to\\KNK_직원등록(본사).xlsx" \\
            --vn  "C:\\path\\to\\KNK_직원등록(베트남).xlsx" \\
            --db  "data/knk.db" \\
            --dry-run

    2. NAS 적용 (실제 import):
        python deploy/import_employees_from_excel.py \\
            --kor "/path/to/KNK_직원등록(본사).xlsx" \\
            --vn  "/path/to/KNK_직원등록(베트남).xlsx" \\
            --db  "/opt/knk_haist/data/main.db" \\
            --apply

매칭 규칙:
    1. 엑셀 직원 = 기존 users 의 (name + email) 매칭 → 갱신
    2. 또는 (name + phone) 매칭 → 갱신
    3. 미매칭 → 신규 INSERT
    4. users 에 있는데 엑셀에 없는 사용자 → 출력만 (수동 판단 위임)

초기 비밀번호 정책:
    - 본사 KOR: 휴대폰 숫자만 (예: '01035401552')
    - VN: '9999' 일괄
    - must_change_pw = 1 (첫 로그인 시 강제 변경)

의존성:
    pip install openpyxl
"""
import argparse
import hashlib
import os
import re
import sqlite3
import sys
from pathlib import Path


# ── 비밀번호 해시 (HAIST WORKS 와 호환) ───────────────────
def hash_pw_pbkdf2(password: str, salt: bytes = None, iterations: int = 200_000) -> str:
    """HAIST WORKS pbkdf2 형식: pbkdf2$200000$<salt_hex>$<hash_hex>"""
    if salt is None:
        salt = os.urandom(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
    return f"pbkdf2${iterations}${salt.hex()}${dk.hex()}"


def initial_pw_kor(phone: str) -> str:
    """본사: 휴대폰 숫자만 (예: '010-3540-1552' → '01035401552')"""
    if not phone:
        return "0000"  # fallback
    return re.sub(r"\D", "", str(phone))


def initial_pw_vn() -> str:
    """VN: 9999 일괄"""
    return "9999"


# ── 엑셀 파싱 ─────────────────────────────────────────────
def parse_kor_sheet(xlsx_path: Path) -> list[dict]:
    """본사(KOR) 시트 파싱 → 직원 dict 리스트"""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)
    if "본사(KOR)" not in wb.sheetnames:
        raise ValueError(f"시트 '본사(KOR)' 없음: {xlsx_path}")
    ws = wb["본사(KOR)"]

    rows = []
    # 헤더: 이름·이메일·휴대폰·사번·직급·부서·이름(영문)
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or not row[0]:  # 빈 행 스킵
            continue
        name_kr, email, phone, emp_no, position, dept_code, name_en = (row + (None,) * 7)[:7]
        if not name_kr or not str(name_kr).strip():
            continue
        rows.append({
            "entity":      "KOR",
            "employee_no": str(emp_no).strip() if emp_no else None,
            "name":        str(name_kr).strip(),
            "name_en":     str(name_en).strip() if name_en else None,
            "name_vi":     None,
            "email":       str(email).strip() if email else None,
            "phone":       str(phone).strip() if phone else None,
            "position":    str(position).strip() if position else None,  # rank 컬럼
            "dept_code":   str(dept_code).strip() if dept_code else None,
        })
    return rows


def parse_vn_sheet(xlsx_path: Path) -> list[dict]:
    """베트남법인(VN) 시트 파싱"""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)
    if "베트남법인(VN)" not in wb.sheetnames:
        raise ValueError(f"시트 '베트남법인(VN)' 없음: {xlsx_path}")
    ws = wb["베트남법인(VN)"]

    rows = []
    # 헤더: 이름(한글)·이메일·휴대폰·사번·직급·부서·이름(베트남어)·이름(영문)
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or not row[0]:
            continue
        name_kr, email, phone, emp_no, position, dept_code, name_vi, name_en = (row + (None,) * 8)[:8]
        if not name_kr or not str(name_kr).strip():
            continue
        rows.append({
            "entity":      "VN",
            "employee_no": str(emp_no).strip() if emp_no else None,
            "name":        str(name_kr).strip(),
            "name_en":     str(name_en).strip() if name_en else None,
            "name_vi":     str(name_vi).strip() if name_vi else None,
            "email":       str(email).strip() if email else None,
            "phone":       str(phone).strip() if phone else None,
            "position":    str(position).strip() if position else None,
            "dept_code":   str(dept_code).strip() if dept_code else None,
        })
    return rows


# ── DB 작업 ───────────────────────────────────────────────
def find_existing_user(c, name: str, email: str = None, phone: str = None):
    """기존 users 매칭: (name+email) 또는 (name+phone)"""
    if email:
        r = c.execute(
            "SELECT * FROM users WHERE name = ? AND email = ?",
            (name, email),
        ).fetchone()
        if r:
            return r
    if phone:
        r = c.execute(
            "SELECT * FROM users WHERE name = ? AND phone = ?",
            (name, phone),
        ).fetchone()
        if r:
            return r
    # name만으로 매칭 (동명이인 가능성 → 1건일 때만)
    rows = c.execute("SELECT * FROM users WHERE name = ?", (name,)).fetchall()
    if len(rows) == 1:
        return rows[0]
    return None


def apply_employee(c, emp: dict, dry_run: bool = True) -> str:
    """
    1명 처리 — 반환: 'updated' / 'inserted' / 'skipped'
    """
    existing = find_existing_user(c, emp["name"], emp.get("email"), emp.get("phone"))

    if existing:
        # 기존 사용자 갱신 (PK·password 보존)
        if dry_run:
            return "updated (dry-run)"
        c.execute("""
            UPDATE users SET
                employee_no = COALESCE(?, employee_no),
                entity      = COALESCE(?, entity),
                login_id    = COALESCE(?, login_id),
                email       = COALESCE(?, email),
                phone       = COALESCE(?, phone),
                name_en     = COALESCE(?, name_en),
                name_vi     = COALESCE(?, name_vi),
                dept_code   = COALESCE(?, dept_code),
                rank        = COALESCE(?, rank)
            WHERE id = ?
        """, (
            emp["employee_no"], emp["entity"],
            emp["employee_no"],  # login_id = 사번
            emp.get("email"), emp.get("phone"),
            emp.get("name_en"), emp.get("name_vi"),
            emp.get("dept_code"), emp.get("position"),
            existing["id"],
        ))
        return "updated"
    else:
        # 신규 INSERT
        if dry_run:
            return "inserted (dry-run)"
        # 초기 비번
        if emp["entity"] == "KOR":
            init_pw = initial_pw_kor(emp.get("phone", ""))
        else:
            init_pw = initial_pw_vn()
        pw_hash = hash_pw_pbkdf2(init_pw)
        c.execute("""
            INSERT INTO users (
                name, login_id, password, email, phone,
                employee_no, entity, name_en, name_vi, dept_code, rank,
                role, is_active, must_change_pw, password_version,
                lang
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, 1, 1, ?)
        """, (
            emp["name"], emp["employee_no"], pw_hash,
            emp.get("email"), emp.get("phone"),
            emp["employee_no"], emp["entity"],
            emp.get("name_en"), emp.get("name_vi"),
            emp.get("dept_code"), emp.get("position"),
            "member",
            "ko" if emp["entity"] == "KOR" else "vi",
        ))
        return "inserted"


# ── 메인 ──────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="HAIST WORKS 직원 일괄 import")
    parser.add_argument("--kor", required=True, help="KNK_직원등록(본사).xlsx 경로")
    parser.add_argument("--vn",  required=True, help="KNK_직원등록(베트남).xlsx 경로")
    parser.add_argument("--db",  required=True, help="SQLite DB 경로 (예: /opt/knk_haist/data/main.db)")
    grp = parser.add_mutually_exclusive_group(required=True)
    grp.add_argument("--dry-run", action="store_true", help="DB 변경 X — 매칭 결과만 출력")
    grp.add_argument("--apply",   action="store_true", help="실제 DB 적용 (백업 권장)")

    args = parser.parse_args()

    kor_path = Path(args.kor)
    vn_path  = Path(args.vn)
    db_path  = Path(args.db)

    if not kor_path.exists():
        sys.exit(f"[ERROR] KOR 엑셀 없음: {kor_path}")
    if not vn_path.exists():
        sys.exit(f"[ERROR] VN 엑셀 없음: {vn_path}")
    if not db_path.exists():
        sys.exit(f"[ERROR] DB 없음: {db_path}")

    print(f"[INFO] KOR 엑셀: {kor_path}")
    print(f"[INFO] VN 엑셀:  {vn_path}")
    print(f"[INFO] DB:       {db_path}")
    print(f"[INFO] 모드:     {'DRY-RUN' if args.dry_run else 'APPLY'}")
    print()

    # 엑셀 파싱
    print("[1/4] 엑셀 파싱 중...")
    kor_emps = parse_kor_sheet(kor_path)
    vn_emps  = parse_vn_sheet(vn_path)
    all_emps = kor_emps + vn_emps
    print(f"  KOR: {len(kor_emps)}명")
    print(f"  VN:  {len(vn_emps)}명")
    print(f"  합계: {len(all_emps)}명")
    print()

    # 사번 중복 검사
    emp_nos = [e["employee_no"] for e in all_emps if e["employee_no"]]
    dup = {n for n in emp_nos if emp_nos.count(n) > 1}
    if dup:
        print(f"[WARN] 사번 중복 발견: {dup}")
        print("  → 진행 전 엑셀 수정 권장. 계속하려면 Enter, 중단은 Ctrl+C")
        input()

    # DB 작업
    print("[2/4] DB 연결 + 트랜잭션 시작...")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    if not args.dry_run:
        # 백업 안내
        print(f"\n[!] DB 변경 직전. 백업 권장:")
        print(f"    cp {db_path} {db_path}.bak_before_emp_import")
        print(f"    계속하려면 Enter, 중단은 Ctrl+C")
        input()

    # 컬럼 존재 검사
    cols = [r[1] for r in c.execute("PRAGMA table_info(users)").fetchall()]
    required = ["employee_no", "entity", "name_en", "name_vi", "phone", "dept_code", "must_change_pw", "password_version"]
    missing = [col for col in required if col not in cols]
    if missing:
        print(f"\n[ERROR] users 테이블에 다음 컬럼 없음: {missing}")
        print(f"  → migrations/v5H226z109_employee_no.sql 먼저 적용하세요.")
        sys.exit(1)

    print("[3/4] 직원 처리 중...")
    stats = {"updated": 0, "inserted": 0, "skipped": 0}
    for emp in all_emps:
        try:
            result = apply_employee(c, emp, dry_run=args.dry_run)
            key = "updated" if "updated" in result else "inserted" if "inserted" in result else "skipped"
            stats[key] += 1
            if args.dry_run:
                emp_label = f"[{emp['employee_no']}] {emp['name']} ({emp['entity']})"
                print(f"  {result:25s} {emp_label}")
        except Exception as e:
            print(f"  [ERROR] {emp['name']}: {e}")
            stats["skipped"] += 1

    # 미매칭 사용자 (DB 에는 있는데 엑셀에 없음)
    print("\n[4/4] DB에만 있는 사용자 (엑셀 미등재) 검색...")
    excel_names = {e["name"] for e in all_emps}
    db_only = c.execute("SELECT id, name, login_id, email FROM users WHERE name NOT IN ({})".format(
        ",".join("?" * len(excel_names))
    ), tuple(excel_names)).fetchall()

    if db_only:
        print(f"  [수동 판단 필요] {len(db_only)}명:")
        for u in db_only:
            print(f"    - id={u['id']:3d}  login={u['login_id']:20s}  name={u['name']:10s}  email={u['email']}")
        print("  → 위 사용자들은 비활성/유지/삭제 중 수동 판단 (대표 결정)")
    else:
        print("  → 미매칭 사용자 없음 (완전 일치)")

    if args.apply:
        conn.commit()
        print("\n[OK] DB 적용 완료")
    else:
        conn.rollback()
        print("\n[OK] dry-run 완료 (DB 변경 안 함)")

    conn.close()

    print(f"\n=== 통계 ===")
    print(f"  갱신: {stats['updated']}")
    print(f"  신규: {stats['inserted']}")
    print(f"  스킵: {stats['skipped']}")
    print(f"  DB 전용(엑셀 미등재): {len(db_only)}")


if __name__ == "__main__":
    main()
