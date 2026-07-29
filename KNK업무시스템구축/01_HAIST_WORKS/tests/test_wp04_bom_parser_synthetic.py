# -*- coding: utf-8 -*-
"""WP-04 파서 — **비식별 자동시험** (실물 파일 없이 항상 실행)

근거: `CHATGPT_WP04_파서보수_완료보고_검토판정_및_V2수정지시_2026-07-28_1435.md`
      §4 P0-1 · §5 P0-2 · §6 P0-3 · §7 P0-4

⛔ **이 시험은 어떤 경우에도 SKIP 하지 않는다.**
   실물 BOM 은 고객 단가가 있어 저장소에 올리지 않으므로, 다른 PC·CI 에서도
   핵심 회귀를 반드시 돌리려면 시험이 **엑셀을 직접 만들어** 써야 한다.
   (실물 대조는 `test_wp04_bom_parser.py` 가 따로 맡고, 파일이 없으면 그쪽만 SKIP 한다.)

⛔ 실제 회사명·품번·단가를 쓰지 않는다. 전부 `TEST-...` 와 가상 금액이다.
"""
import io
import os
import sys
import tempfile

_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.dirname(_HERE))

import openpyxl                                    # noqa: E402
from openpyxl.styles import Font, PatternFill      # noqa: E402

# ⚠ §J 는 실제 라우트를 돌린다. **운영 DB 를 절대 건드리지 않도록** 임시 DB 를
#   먼저 세우고 나서 앱을 들여온다(순서가 뒤바뀌면 기본 DB 경로가 잡힌다).
from app import database as _db                    # noqa: E402
_db.DB_PATH = os.path.join(tempfile.mkdtemp(prefix="wp04syn_"), "test.db")
_db.init_db()

from app import bom                                # noqa: E402

ok = 0
fail = 0


def chk(name, cond, detail=""):
    global ok, fail
    if cond:
        ok += 1
        print(f"  PASS {name}")
    else:
        fail += 1
        print(f"  FAIL {name} {detail}")


_YELLOW = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

# 헤더(구매팀 최종본 양식) — 실물과 같은 낱말을 쓰되 값은 전부 가상
_FULL_HDR = ["NO.", "CATEGORY 구분 / Danh mục", "CODE", "PRODUCT NAME 제품명",
             "PRODUCT CODE 코드명", "MANUFACTURER 제조사", "UNIT COUNT 수량",
             "TOTAL 수량", "UNIT 단위", "사내 재고", "베트남 재고", "발주 수량",
             "UNIT PRICE 단가", "재고 금액", "발주 금액", "AMOINT 합계",
             "DELIVERY 납기", "REMARKS(NOTE) 비고"]
# 합계 열 자체가 없는 양식(설계팀 초기본처럼)
_NOAMT_HDR = ["NO.", "CATEGORY 구분", "CODE", "PRODUCT NAME 제품명",
              "PRODUCT CODE 코드명", "MANUFACTURER 제조사", "UNIT COUNT 수량",
              "TOTAL 수량", "UNIT 단위", "UNIT PRICE 단가", "DELIVERY 납기",
              "REMARKS(NOTE) 비고"]


def _write(ws, header, rows):
    ws.cell(2, 3, "999T9901 구매품 BOM (시험용)")
    for i, h in enumerate(header, start=2):
        ws.cell(7, i, h)
    for r, row in enumerate(rows, start=8):
        for i, v in enumerate(row, start=2):
            ws.cell(r, i, v)


def build_full(path):
    """모든 열이 있는 양식 + 판정서 §4.2 필수 사례 전부."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1. 구매품"
    rows = [
        # NO cat  code  품명            품번               제조사     대분 총량 단위 사내 VN 발주 단가   재고금  발주금  합계    납기      비고
        [1, "UNIT-A", "A1", "TEST NORMAL", "TEST-PART-A", "TESTMAKER", 1, 10, "EA", 2, 3, 5, 100, 200, 500, 1000, "1 WEEK", ""],
        [2, "UNIT-A", "A2", "TEST BLANK AMT", "TEST-PART-B", "TESTMAKER", 1, 4, "EA", 0, 0, 4, 50, None, None, None, "2 WEEK", ""],
        [3, "UNIT-A", "A3", "TEST ZERO AMT", "TEST-PART-C", "TESTMAKER", 1, 6, "EA", 0, 0, 6, 70, 0, 0, 0, "2 WEEK", "무상"],
        [4, "UNIT-A", "A4", "TEST NO PRICE", "TEST-PART-D", "TESTMAKER", 1, 3, "EA", 0, 0, 3, None, None, None, None, "미정", ""],
        [5, "UNIT-B", "B1", "TEST ONE STEP", "TEST-OLD-E->TEST-NEW-E", "TESTMAKER", 1, 2, "EA", 0, 0, 2, 30, 0, 60, 60, "1 WEEK", ""],
        [6, "UNIT-B", "B2", "TEST TWO STEP",
         "TEST-OLD-F->TEST-MID-F->TEST-NEW-F", "TESTOLD->TESTNEW", 1, 2, "EA", 0, 0, 2, 40, 0, 80, 80, "1 WEEK", ""],
        [7, "UNIT-C", "C1", "해외 출장비 (TEST)", "TEST-SVC-1", "TESTMAKER", 1, 1, "EA", 0, 0, 1, 500, 0, 500, 500, "협의", ""],
        [8, "UNIT-C", "C2", "TEST NEW MARK", "TEST-PART-H", "TESTMAKER", 1, 1, "EA", 0, 0, 1, 10, 0, 10, 10, "1 WEEK", ""],
        [9, "UNIT-C", "C3", "TEST DEL MARK", "TEST-PART-I", "TESTMAKER", 1, 1, "EA", 0, 0, 1, 10, 0, 10, 10, "1 WEEK", ""],
    ]
    _write(ws, _FULL_HDR, rows)
    ws.cell(15, 5).fill = _YELLOW          # 8번 줄(15행) 품명 = 노랑 → 신규 표시
    ws.cell(16, 5).font = Font(strike=True)  # 9번 줄(16행) 품명 = 취소선 → 삭제 표시
    ws.cell(16, 6).font = Font(strike=True)
    wb.save(path)


def build_noamt(path):
    """합계 열 **자체가 없는** 양식 — 이때만 계산이 허용된다."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1. 구매품"
    _write(ws, _NOAMT_HDR, [
        [1, "UNIT-A", "A1", "TEST NO AMT COL", "TEST-PART-X", "TESTMAKER", 1, 5, "EA", 100, "1 WEEK", ""],
        [2, "UNIT-A", "A2", "TEST NO AMT NO PRICE", "TEST-PART-Y", "TESTMAKER", 1, 5, "EA", None, "1 WEEK", ""],
    ])
    wb.save(path)


_tmp = tempfile.mkdtemp(prefix="knk_bom_test_")
P_FULL = os.path.join(_tmp, "999T9901 구매품 LIST_TEST.xlsx")
P_NOAMT = os.path.join(_tmp, "999T9902 구매품 LIST_TEST.xlsx")
build_full(P_FULL)
build_noamt(P_NOAMT)

F = bom.parse_bom_file(P_FULL, os.path.basename(P_FULL))["sheets"][0]["items"]
N = bom.parse_bom_file(P_NOAMT, os.path.basename(P_NOAMT))["sheets"][0]["items"]
by = {x["part_no"]: x for x in F}

# ══════════ A. 양식 인식 ══════════
print("\n── A. 양식 인식 (파일 없이도 항상 실행) ──")
chk("A-1 전체 양식 9줄", len(F) == 9, f"{len(F)}줄")
chk("A-2 합계 열 없는 양식 2줄", len(N) == 2, f"{len(N)}줄")
chk("A-3 관리번호 자동추출",
    bom.parse_bom_file(P_FULL, os.path.basename(P_FULL))["mgmt_code"] == "999T9901")

# ══════════ B. P-1/P-4 원본 수량 필드 (이름이 업무 의미와 맞는가) ══════════
print("\n── B. 원본 수량 필드 (P0-4 이름 보정) ──")
chk("B-1 사내 **배분**수량 필드명 source_stock_allocated_kor",
    "source_stock_allocated_kor" in F[0])
chk("B-2 발주 **원본**수량 필드명 source_purchase_qty", "source_purchase_qty" in F[0])
chk("B-3 베트남은 참고수량 stock_ref_vn", "stock_ref_vn" in F[0])
chk("B-4 ⛔운영 오해 이름(stock_qty_kor·order_qty)은 없다",
    "stock_qty_kor" not in F[0] and "order_qty" not in F[0])
chk("B-5 값 읽기: 사내배분 2 · 베트남참고 3 · 발주원본 5",
    (F[0]["source_stock_allocated_kor"], F[0]["stock_ref_vn"], F[0]["source_purchase_qty"]) == (2, 3, 5),
    str((F[0]["source_stock_allocated_kor"], F[0]["stock_ref_vn"], F[0]["source_purchase_qty"])))
chk("B-6 [원본 파일 대조용] 총량 = 사내배분+베트남참고+발주원본 (운영 계산식 아님)",
    F[0]["total_qty"] == F[0]["source_stock_allocated_kor"] + F[0]["stock_ref_vn"] + F[0]["source_purchase_qty"])

# ══════════ C. P0-2 금액 네 상태 ══════════
print("\n── C. P0-2 금액: 원본값·명시적 0·빈칸·열없음 ──")
a, b, c, d = by["TEST-PART-A"], by["TEST-PART-B"], by["TEST-PART-C"], by["TEST-PART-D"]
chk("C-1 ①원본값 있음 → 그대로 보존 (1000)",
    a["amount"] == 1000 and a["amount_source_present"] and not a["amount_is_calculated"],
    f'{a["amount"]}/{a["amount_source_present"]}/{a["amount_is_calculated"]}')
chk("C-2 ①원본값 별도 보관 amount_source_value=1000", a["amount_source_value"] == 1000)
chk("C-3 ③빈칸 + 단가·수량 있음 → **계산**하고 표시 (50×4=200)",
    b["amount"] == 200 and not b["amount_source_present"] and b["amount_is_calculated"],
    f'{b["amount"]}/{b["amount_source_present"]}/{b["amount_is_calculated"]}')
chk("C-4 ⭐②명시적 0 → **계산 금지**, 0 그대로 (70×6=420 이 되면 안 됨)",
    c["amount"] == 0 and c["amount_source_present"] and not c["amount_is_calculated"],
    f'{c["amount"]}/{c["amount_source_present"]}/{c["amount_is_calculated"]}')
chk("C-5 ⭐계산 근거 없음(단가 없음) → **0 이라고 단정하지 않음**",
    not d["amount_source_present"] and not d["amount_is_calculated"],
    f'present={d["amount_source_present"]} calc={d["amount_is_calculated"]}')
chk("C-6 ④합계 열 자체가 없음 + 단가 있음 → 계산 (100×5=500)",
    N[0]["amount"] == 500 and N[0]["amount_is_calculated"] and not N[0]["amount_source_present"],
    f'{N[0]["amount"]}/{N[0]["amount_is_calculated"]}')
chk("C-7 ④합계 열 없음 + 단가도 없음 → 계산 안 함(모름)",
    not N[1]["amount_is_calculated"] and not N[1]["amount_source_present"])
chk("C-8 재고금액·발주금액도 따로 읽힌다 (200 / 500)",
    a["stock_amount"] == 200 and a["order_amount"] == 500,
    f'{a["stock_amount"]}/{a["order_amount"]}')

# ══════════ D. P-4 화살표 ══════════
print("\n── D. 화살표 분해 ──")
e1 = by.get("TEST-NEW-E")
f1 = by.get("TEST-NEW-F")
chk("D-1 A→B 1회 변경 → 새 값이 품번", bool(e1))
chk("D-2 A→B→C 2회 변경 → **마지막 값**이 품번", bool(f1))
chk("D-3 남은 화살표 0줄",
    not [x for x in F if "->" in (x.get("part_no") or "") or "->" in (x.get("maker") or "")])
_fc = [ch for ch in (f1 or {}).get("excel_changes", []) if ch["field"] == "part_no"]
chk("D-4 체인 전체 보존 [A,B,C]",
    bool(_fc) and _fc[0].get("chain") == ["TEST-OLD-F", "TEST-MID-F", "TEST-NEW-F"],
    str(_fc[0].get("chain") if _fc else None))
chk("D-5 제조사 화살표도 분해 (TESTOLD→TESTNEW)", (f1 or {}).get("maker") == "TESTNEW")

# ══════════ E. P0-3 변경 매칭 (§6.3 표 4가지) ══════════
print("\n── E. P0-3 A→B→C 매칭 (기존 저장값이 A든 B든) ──")


def _board(part_nos):
    return [{"category": "UNIT-B", "part_no": p, "part_name": "TEST TWO STEP",
             "id": i} for i, p in enumerate(part_nos, start=1)]


_two = [f1] if f1 else []
p1, a1, _ = bom._pair_items(_board(["TEST-OLD-F"]), _two)
chk("E-1 기존=A → A에서 C로 **변경으로 연결**", len(p1) == 1 and not a1,
    f"pairs={len(p1)} adds={len(a1)}")
p2, a2, _ = bom._pair_items(_board(["TEST-MID-F"]), _two)
chk("E-2 ⭐기존=B(중간값) → B에서 C로 **변경으로 연결**", len(p2) == 1 and not a2,
    f"pairs={len(p2)} adds={len(a2)}")
p3, a3, _ = bom._pair_items(_board(["TEST-OTHER"]), _two)
chk("E-3 기존 없음 → 신규 추가", len(p3) == 0 and len(a3) == 1)
chk("E-3' 신규여도 원문 체인은 보존",
    bool(a3) and any(ch.get("chain") for ch in a3[0].get("excel_changes", [])))
p4, a4, _ = bom._pair_items(_board(["TEST-MID-F", "TEST-MID-F"]), _two)
chk("E-4 ⭐같은 이전값이 **여러 줄** → 자동 확정 금지",
    len(p4) == 0 and len(a4) == 1, f"pairs={len(p4)} adds={len(a4)}")
chk("E-4' 모호 매칭으로 **표시**된다", bool(a4) and a4[0].get("match_ambiguous") is True)
p5, a5, _ = bom._pair_items(_board(["TEST-NEW-F"]), _two)
chk("E-5 기존이 이미 최종값 C → 그대로 같은 줄로 연결", len(p5) == 1 and not a5)

# ══════════ F. P-5 서비스 의심 · 색·취소선 ══════════
print("\n── F. 서비스 의심 · 엑셀 표기 ──")
chk("F-1 출장비 줄에 의심 표시", by["TEST-SVC-1"]["is_service_suspect"] is True)
chk("F-2 일반 자재는 표시 없음", by["TEST-PART-A"]["is_service_suspect"] is False)
chk("F-3 ⭐노랑 배경 = 신규 표시가 읽힌다", by["TEST-PART-H"]["is_new_marked"] is True)
chk("F-4 ⭐취소선 = 삭제 표시가 읽힌다", by["TEST-PART-I"]["is_deleted_marked"] is True)
chk("F-5 표시 없는 줄은 False",
    not by["TEST-PART-A"]["is_new_marked"] and not by["TEST-PART-A"]["is_deleted_marked"])

# ══════════ G. BOM 업로드 검증 ① 원본 ↔ 파서가 읽은 값 ══════════
print("\n── G. 원본↔읽은값 대조 (규칙47 · knkVerifyMsg1024 형식) ──")
V = bom.verify_source_vs_parsed(P_FULL, os.path.basename(P_FULL))
chk("G-1 세션01 공통 형식 그대로 (ok·checked·diff)",
    all(k in V for k in ("ok", "checked", "diff")))
chk("G-2 원본 위치가 시트!행 으로 표시된다", "_src_sheet" in F[0] and "_src_row" in F[0])
chk("G-3 9줄 전부 대조", V["checked"] == 9, str(V["checked"]))
chk("G-4 ⭐불일치 0건 — 파서가 원본을 그대로 읽는다", V["ok"] and not V["diff"],
    str(V["diff"][:2]))
chk("G-5 ⭐화살표 줄이 불일치로 잡히지 않는다(마지막 값이 정답)",
    not [d for d in V["diff"] if d["field"] in ("part_no", "maker")])
chk("G-6 ⭐계산한 금액이 불일치로 잡히지 않는다(규칙대로 검산)",
    not [d for d in V["diff"] if d["field"] == "amount"])
VN = bom.verify_source_vs_parsed(P_NOAMT, os.path.basename(P_NOAMT))
chk("G-7 합계 열 없는 양식도 대조 통과", VN["ok"] and VN["checked"] == 2,
    f'ok={VN["ok"]} checked={VN["checked"]}')
chk("G-8 원본에 없는 값을 지어내지 않는다 — 저장 안 되는 칸은 대조 대상에서 제외",
    "_src_row" not in bom._ITEM_COLS)

# ══════════ H. 중복 시트 감지 ══════════
print("\n── H. 중복 시트 (실물 003M2506 사례) ──")
P_DUP = os.path.join(_tmp, "999T9903 구매품 LIST_TEST.xlsx")
_wb = openpyxl.load_workbook(P_FULL)
_wb.copy_worksheet(_wb["1. 구매품"]).title = "1. 구매품 (2)"
_wb.save(P_DUP)
DUP = bom.scan_duplicate_sheets(P_DUP)
chk("H-1 같은 내용 시트 2개를 찾아낸다", len(DUP) == 1, str(DUP))
chk("H-2 어느 시트인지 알려준다",
    bool(DUP) and set(DUP[0]["sheets"]) == {"1. 구매품", "1. 구매품 (2)"})
VD = bom.verify_source_vs_parsed(P_DUP, os.path.basename(P_DUP))
chk("H-3 ⭐중복이 있으면 ok=False — 그대로 올리면 줄이 두 배", VD["ok"] is False)
chk("H-4 사람 말로 안내한다", "두 배" in (VD.get("skipped") or ""))
chk("H-5 정상 파일은 중복 없음", not bom.scan_duplicate_sheets(P_FULL))
# 빈 시트 오탐 방지 — 실물 002M2505 의 2.가공품·3.공용부 가 줄번호만 있어 100% 같게 나왔다
P_EMPTY = os.path.join(_tmp, "999T9904 구매품 LIST_TEST.xlsx")
_wb2 = openpyxl.Workbook()
for t in ("1. 구매품", "2. 가공품", "3. 공용부"):
    _ws = _wb2.create_sheet(t) if t != "1. 구매품" else _wb2.active
    _ws.title = t
    _write(_ws, _FULL_HDR, [[1, "UNIT-A", "A1", "TEST", "TEST-PART-A", "M", 1, 1, "EA",
                             0, 0, 1, 10, 0, 10, 10, "1W", ""]] if t == "1. 구매품"
           else [[i, None, None, None, None, None, None, None, None,
                  None, None, None, None, None, None, None, None, None] for i in range(1, 6)])
_wb2.save(P_EMPTY)
chk("H-6 ⭐빈 시트끼리는 중복으로 보지 않는다(오탐 방지)",
    not bom.scan_duplicate_sheets(P_EMPTY), str(bom.scan_duplicate_sheets(P_EMPTY)))

# ══════════ I. 화면·라우트에 실제로 붙었는가 ══════════
print("\n── I. 업로드 화면 연결 (한 곳만 고치고 끝내지 않기) ──")
_APP = os.path.dirname(_HERE)
_main = io.open(os.path.join(_APP, "app", "main.py"), encoding="utf-8").read()
_tpl = io.open(os.path.join(_APP, "app", "templates", "bom_upload.html"), encoding="utf-8").read()
chk("I-1 미리보기에서 원본↔읽은값을 대조한다", "verify_source_vs_parsed" in _main)
chk("I-2 대조 결과를 화면으로 넘긴다", "verify_src=verify_src" in _main)
chk("I-3 화면이 결과를 그린다", "verify_src" in _tpl and "vfy" in _tpl)
chk("I-4 일치·불일치·중복·오류 네 갈래를 다 그린다",
    all(k in _tpl for k in ("verify_src.ok", "verify_src.diff",
                            "verify_src.duplicate_sheets", "verify_src.error")))
chk("I-5 적용 직후 저장값을 대조한다", "verify_parsed_vs_saved" in _main)
chk("I-6 대조 결과를 사람에게 알린다", "저장값 대조" in _main)
def _read(*parts):
    """없는 파일은 **크래시가 아니라 빈 문자열** — 뒤 시험까지 못 돌면 변별력을 못 본다."""
    try:
        return io.open(os.path.join(_APP, *parts), encoding="utf-8").read()
    except OSError:
        return ""


_res_tpl = _read("app", "templates", "bom_upload_result.html")
chk("I-7 ⭐대조가 실패해도 저장은 유지한다(결과만 알림)",
    "저장됐습니다" in _res_tpl and "되돌리지 않는다" in _res_tpl)
chk("I-8 CSS 가 있다(글자만 있고 안 보이는 일 없게)", ".vfy-ok" in _tpl and ".vfy-bad" in _tpl)

# ══════════ J. 2026-07-29 판정 보정 5건 ══════════
# 근거: `CHATGPT_WP04_BOM검증_A단계_중간판정_및_B단계_분리요청_2026-07-29_1715.md`
#       + 대표 조건부 승인 2026-07-29 18:29
print("\n── J. 판정 보정 (①상태고지 ②통보잠금 ③중복선택 ④전체상세 ⑤버전문구) ──")
_bom_tpl = _read("app", "templates", "project_bom.html")

_dbsrc = _read("app", "database.py")

# ── ⑤ 업로드 버전 ≠ BOM Revision ──
chk("J-1 완료 안내가 '업로드 버전 vN'이라고 말한다(설계 Revision 과 구분)",
    "업로드 버전 v{res['version_no']}" in _main)
chk("J-2 옛 표현 'v{n} 반영'이 남아 있지 않다",
    "f\"v{res['version_no']} 반영" not in _main)
chk("J-3 보드 화면도 '업로드 버전'이라고 쓴다", "업로드 버전 v<b>" in _bom_tpl)

# ── ① 임시 저장 · Release 전 ──
chk("J-4 반영 버튼이 '임시 저장'이라고 말한다", "보드에 임시 저장" in _tpl)
chk("J-5 버튼 위에 Release 전이라고 적혀 있다",
    "Release 전" in _tpl and "구매근거로 쓸 수 없습니다" in _tpl)
chk("J-6 ⭐BOM 보드에도 임시저장 상태가 표시된다(보는 사람이 오해하지 않게)",
    "임시 저장(업로드 반영) 상태" in _bom_tpl and "공식 구매근거가 아니며" in _bom_tpl)
chk("J-7 상태 고지에 색이 있다(글자만 있고 안 보이는 일 없게)",
    ".banner.stage" in _bom_tpl and ".stage-note" in _tpl)
chk("J-8 완료 안내도 '임시 저장'이라고 말한다", "반영(임시 저장)" in _main)

# ── ② 구매팀 통보 잠금 ──
chk("J-9 잠금 스위치가 WP-01 공용 목록에 있다(배포 전 검사기가 자동으로 본다)",
    "KNK_ENABLE_BOM_UPLOAD_NOTIFY" in _dbsrc
    and "KNK_ENABLE_BOM_UPLOAD_NOTIFY" in _db.WP01_LOCK_SWITCHES)
chk("J-10 ⭐기본값이 잠김이다(환경변수 없으면 통보 안 함)",
    _db.wp01_unlocked("KNK_ENABLE_BOM_UPLOAD_NOTIFY") is False)
chk("J-11 화면이 '구매팀에 통보되지 않는다'고 알려 준다",
    "구매팀에는 통보되지 않습니다" in _tpl)
chk("J-12 통보 코드는 지우지 않고 잠갔다(Release 때 되살릴 수 있게)",
    "_notify_users(c, _uids" in _main)

# ── ③ 중복 후보 시트 ──
chk("J-13 서버에 중복 선택 차단 함수가 있다", hasattr(bom, "duplicate_selection_block"))
chk("J-14 라우트가 그 함수를 부른다", "duplicate_selection_block" in _main)
_dp = [{"sheets": ["1. 구매품", "1. 구매품 (2)"], "rows": 326, "same_ratio": 0.99}]


def _dsb(*a, **k):
    """함수가 아직 없으면 None — 아래 판정이 전부 FAIL 로 떨어진다(크래시로 멈추지 않게)."""
    f = getattr(bom, "duplicate_selection_block", None)
    return f(*a, **k) if f else None


chk("J-15 ⭐둘 다 고르면 막는다", bool(_dsb(_dp, ["1. 구매품", "1. 구매품 (2)"])))
chk("J-16 ⭐아무것도 안 고르면(=전부 적용) 막는다", bool(_dsb(_dp, [])))
chk("J-17 한쪽만 고르면 통과", _dsb(_dp, ["1. 구매품"]) == "")
chk("J-18 '둘 다 사용' 확인하면 통과", _dsb(_dp, [], ack=True) == "")
chk("J-19 중복이 없으면 늘 통과", _dsb([], []) == "")
chk("J-20 화면이 중복 시트를 기본 해제한다",
    "is_dup" in _tpl and "'' if is_dup else 'checked'" in _tpl)
chk("J-21 화면에 '둘 다 사용' 확인칸이 있다", 'name="dup_ack"' in _tpl)
chk("J-22 고르기 전에는 버튼이 잠긴다", "'disabled' if dupset" in _tpl)

# ── ④ 불일치 전체 상세 ──
chk("J-23 결과화면 템플릿이 있다",
    os.path.exists(os.path.join(_APP, "app", "templates", "bom_upload_result.html")))
chk("J-24 라우트가 결과화면을 렌더한다", "bom_upload_result.html" in _main)
chk("J-25 ⭐전체를 싣는다(앞 몇 건만 자르지 않는다)",
    "for d in _diff" in _res_tpl and "_diff[:3]" not in _res_tpl and "diff[:8]" not in _res_tpl)
chk("J-26 ⭐주소창에 상세를 싣지 않는다",
    "x['row']" not in _main and "for x in _d[:3]" not in _main)
chk("J-27 영문 칸이름에 사람 이름표를 붙인다",
    getattr(bom, "field_label", lambda f: f)("part_no") == "모델명(품번)"
    and "field_labels" in _res_tpl)
chk("J-28 목록을 파일로 내려받을 수 있다", "vfyCsvBtn" in _res_tpl)
chk("J-29 ⭐이 화면의 한계를 정직하게 적는다(기록 보존은 다음 단계)",
    "다음 단계에서 대표 승인" in _res_tpl)
chk("J-30 미리보기 불일치도 전부 싣는다(같은 결함을 한쪽만 고치지 않기)",
    "for d in verify_src.diff %}" in _tpl and "verify_src.diff[:8]" not in _tpl)
chk("J-31 '전부 일치'의 한계를 적는다(§5)",
    "인식한 열·행" in _tpl and "설계책임자의 검토를 대신하지 않습니다" in _tpl)

# ── ⑥ 실제 라우트로 '통보 0건' 증명 (대표 지시 6항) ──
print("\n── J' 실제 라우트 실행 — 통보 0건·중복 차단 (임시 DB) ──")
try:
    import base64
    from fastapi.testclient import TestClient
    import app.main as m
    from app import sso_client as _sso

    _sent = {"n": 0}

    def _fake_push(emp_nos, title, body, link):     # ⛔ 실제 메신저로 나가지 않게 가로챈다
        _sent["n"] += len(emp_nos or [])
        return {"ok": True, "sent": len(emp_nos or [])}

    _sso.notify_via_messenger = _fake_push
    _cl = TestClient(m.app)

    with _db.db_session() as _c:
        # 운영 스키마엔 마이그레이션으로 붙는 칸 — 시험 DB 에도 맞춰 준다(있으면 그냥 넘어간다)
        try:
            _c.execute("ALTER TABLE users ADD COLUMN employee_no TEXT")
        except Exception:
            pass
        _c.execute("INSERT INTO projects (mgmt_code, name) VALUES ('999T9999','시험 프로젝트')")
        _PID = _c.execute("SELECT id FROM projects WHERE mgmt_code='999T9999'").fetchone()[0]
        if not _c.execute("SELECT 1 FROM teams WHERE id=10").fetchone():
            _c.execute("INSERT INTO teams (id, code, name) VALUES (10,'T99','구매팀')")
        # 구매팀(team_id=10) 사람 2명 — 잠금이 없었다면 통보가 갔어야 할 대상
        for _i, _n in enumerate(("구매A", "구매B"), start=1):
            _c.execute("INSERT INTO users (name, login_id, password, team_id, is_active, employee_no)"
                       " VALUES (?,?,'x',10,1,?)", (_n, f"T99{_i}", f"T99{_i}"))
        # 올리는 사람 — bom_uploads.uploaded_by 가 users 를 참조하므로 실제 행이 있어야 한다
        _c.execute("INSERT INTO users (name, login_id, password, team_id, is_active)"
                   " VALUES ('시험올린이','T990','x',10,1)")
        _UID = _c.execute("SELECT id FROM users WHERE login_id='T990'").fetchone()[0]
    m.get_user = lambda req: {"id": _UID, "name": "시험올린이", "role": "ceo", "team_id": 10}
    _tgt = 0
    with _db.db_session() as _c:
        _tgt = _c.execute("SELECT COUNT(*) FROM users WHERE team_id=10 AND COALESCE(is_active,1)=1"
                          ).fetchone()[0]
    chk("J'-0 통보 대상이 실제로 존재한다(0명이라 안 간 게 아님)", _tgt >= 2, f"대상 {_tgt}명")

    _b64 = base64.b64encode(io.open(P_FULL, "rb").read()).decode("ascii")
    _r = _cl.post("/bom/upload/apply", follow_redirects=False,
                  data={"file_b64": _b64, "filename": os.path.basename(P_FULL),
                        "project_id": str(_PID), "mode": "merge"})
    chk("J'-1 반영이 정상 처리된다", _r.status_code in (200, 303), str(_r.status_code))
    with _db.db_session() as _c:
        _nrow = _c.execute("SELECT COUNT(*) FROM notifications").fetchone()[0]
        _items = _c.execute("SELECT COUNT(*) FROM bom_items WHERE project_id=?", (_PID,)).fetchone()[0]
    chk("J'-2 ⭐⭐ 인앱 알림 0건 — 업로드 단계에서 구매팀에 통보하지 않았다", _nrow == 0, f"{_nrow}건")
    chk("J'-3 ⭐⭐ 메신저 푸시 0건", _sent["n"] == 0, f"{_sent['n']}건")
    chk("J'-4 통보는 안 했지만 저장은 됐다(기능을 죽인 게 아님)", _items > 0, f"{_items}줄")
    _loc = _r.headers.get("location", "")
    chk("J'-5 완료 안내에 '통보하지 않았습니다'가 있다",
        ("통보하지" in _loc) or ("통보하지" in _r.text))

    # 중복 시트 파일을 화면 없이 그대로 밀어 넣어도 서버가 막는가
    _b64d = base64.b64encode(io.open(P_DUP, "rb").read()).decode("ascii")
    _rd = _cl.post("/bom/upload/apply", follow_redirects=False,
                   data={"file_b64": _b64d, "filename": os.path.basename(P_DUP),
                         "project_id": str(_PID), "mode": "merge"})
    chk("J'-6 ⭐중복 시트를 화면 없이 밀어 넣어도 서버가 막는다",
        _rd.status_code == 303 and "error=" in _rd.headers.get("location", ""),
        _rd.headers.get("location", "")[:80])
    with _db.db_session() as _c:
        _items2 = _c.execute("SELECT COUNT(*) FROM bom_items WHERE project_id=?", (_PID,)).fetchone()[0]
    chk("J'-7 ⭐막혔으니 줄이 두 배가 되지 않았다", _items2 == _items, f"{_items} → {_items2}")

    # 잠금을 열면 원래대로 통보가 나가는가 (기능이 살아 있다는 증명 · 곧바로 되잠근다)
    os.environ["KNK_ENABLE_BOM_UPLOAD_NOTIFY"] = "1"
    _r2 = _cl.post("/bom/upload/apply", follow_redirects=False,
                   data={"file_b64": _b64, "filename": os.path.basename(P_FULL),
                         "project_id": str(_PID), "mode": "merge"})
    with _db.db_session() as _c:
        _nrow2 = _c.execute("SELECT COUNT(*) FROM notifications").fetchone()[0]
    chk("J'-8 잠금을 열면 통보가 나간다(지운 게 아니라 잠근 것)", _nrow2 > 0, f"{_nrow2}건")
    os.environ.pop("KNK_ENABLE_BOM_UPLOAD_NOTIFY", None)
    chk("J'-9 시험 뒤 다시 잠긴다", _db.wp01_unlocked("KNK_ENABLE_BOM_UPLOAD_NOTIFY") is False)
except Exception as _e:
    import traceback
    traceback.print_exc()
    chk("J' 라우트 시험 실행", False, f"{type(_e).__name__}: {_e}")

print(f"\n{'=' * 52}\n결과: PASS {ok} · FAIL {fail}")
sys.exit(0 if fail == 0 else 1)
