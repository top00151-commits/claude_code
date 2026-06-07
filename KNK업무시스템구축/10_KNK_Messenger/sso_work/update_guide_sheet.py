#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
안내 시트 §2 갱신 — 로그인 ID 를 '회사 이메일/영문이름' → '사번' 으로 변경.
발주: _TO_메신저세션/2026-05-29_사번SSO.md §3.3

방식: 키워드 기반 셀 치환(행 번호 고정 아님 → 행이 밀려도 안전).
입력: HR 폴더 원본 2개 (개인정보 포함)
출력: sso_work/ 안에 '_사번SSO' 접미사 사본 (gitignore 처리됨, PII 보호)
      → 대표/담당자 검토 후 HR 원본 교체 (폴더스코프 정책상 빅터가 원본 직접 수정 안 함)

사용법:
  py update_guide_sheet.py            # 기본 HR 경로 → sso_work 출력
  py update_guide_sheet.py --inplace  # (관리자만) HR 원본 직접 수정
"""
import argparse
import os

try:
    import openpyxl
except Exception:
    raise SystemExit("[FATAL] openpyxl 필요")

HR = r"C:\Users\top00\JR\01_업무관련\03_인사업무"
FILES = ["KNK_직원등록(본사).xlsx", "KNK_직원등록(베트남).xlsx"]

# (셀에 포함된 옛 문구 일부) → (새 전체 문구). 부분일치 시 새 문구로 교체.
REPLACE = [
    ("로그인 ID = 회사 이메일",
     "  · 로그인 ID = 사번 (예: 5, 43, 300)   ← 회사 이메일 아님 (2026-05-29 사번 전환)"),
    ("로그인 ID = 영문이름을 공백 제거",
     "  · 로그인 ID = 사번 (예: VN001, VN094)   ← 영문이름 ID 폐지 (2026-05-29 사번 전환)"),
    ("→ 'nguyenvana'",
     "       사번 그대로 입력 — 예: VN001 (영문이름 ID 폐지)"),
    ("로그인 시 'nguyenvana' 만 입력",
     "       로그인 시 사번(예: VN001) 입력"),
    ("사번 표기: 본사 = '001'",
     "사번 표기: 본사 = 순수 숫자(예: 5·43·300, 0 패딩 없음), 베트남 = 'VN001'·'VN002' 식"),
    ("동명이인이면 영문이름에 직접 숫자",
     "  · 사번은 직원마다 고유 — 동명이인도 사번으로 구분 (별도 처리 불필요)"),
    ("같은 영문이름은 같은 ID로 인식",
     "  · 같은 사번은 같은 직원으로 인식 — 재업로드 시 자동 갱신(upsert)"),
]


def patch_guide(ws):
    n = 0
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str):
                continue
            for old, new in REPLACE:
                if old in v:
                    cell.value = new
                    n += 1
                    break
    return n


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--inplace", action="store_true", help="HR 원본 직접 수정 (관리자만)")
    ap.add_argument("--outdir", default=os.path.dirname(os.path.abspath(__file__)))
    args = ap.parse_args()

    for fname in FILES:
        src = os.path.join(HR, fname)
        if not os.path.exists(src):
            print(f"[skip] 없음: {src}")
            continue
        wb = openpyxl.load_workbook(src)
        if "안내" not in wb.sheetnames:
            print(f"[skip] '안내' 시트 없음: {fname}")
            continue
        n = patch_guide(wb["안내"])
        if args.inplace:
            wb.save(src)
            print(f"[inplace] {fname}: {n} 셀 갱신 → 원본 저장")
        else:
            base, ext = os.path.splitext(fname)
            out = os.path.join(args.outdir, f"{base}_사번SSO{ext}")
            wb.save(out)
            print(f"[copy] {fname}: {n} 셀 갱신 → {out}")
    print("\n완료. (기본은 사본 생성 - 검토 후 HR 원본 교체. --inplace 로 원본 직접 수정 가능)")


if __name__ == "__main__":
    main()
