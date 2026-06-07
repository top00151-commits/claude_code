#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
사번 마이그레이션 — 엑셀(본사/베트남) 기준으로 메신저 users 의 login_id(username)
를 사번(employee_no)으로 통일 + employee_no/entity/직급/부서/영문·베트남어 이름 갱신.

발주: _TO_메신저세션/2026-05-29_사번SSO.md §3.2
원칙:
  - users.id(내부 PK) 절대 불변 (모든 외래키 기반)
  - 매칭: display_name == 엑셀 이름  AND  (email 일치 OR 휴대폰 숫자 일치)
  - 매칭됨 → username/employee_no/entity/title/department/영문·VN이름 갱신 (role·비번 보존)
  - 미매칭(엑셀에만) → 신규 INSERT (초기비번 본사=휴대폰숫자 / VN=9999, must_change=1)
  - DB-only(엑셀에 없는 직원) → 목록만 출력. 자동 비활성화 절대 금지 (§8.1, 대표 수동 판단)

사용법:
  # 1) dry-run (기본) — 무엇이 바뀔지만 출력, DB 변경 안 함
  py migrate_employee_no.py --db /opt/knk_messenger/data/messenger.db

  # 2) 실제 적용 (백업 먼저!)
  py migrate_employee_no.py --db /opt/knk_messenger/data/messenger.db --apply

  # 엑셀 경로 지정 (기본 = HR 폴더)
  py migrate_employee_no.py --kor "...본사.xlsx" --vn "...베트남.xlsx" --db ...
"""
import argparse
import os
import sqlite3
import sys
from datetime import datetime, timezone

try:
    import openpyxl
except Exception:
    print("[FATAL] openpyxl 필요: pip install openpyxl")
    sys.exit(2)

try:
    from werkzeug.security import generate_password_hash
except Exception:
    print("[FATAL] werkzeug 필요 (메신저 가상환경에서 실행)")
    sys.exit(2)

DEFAULT_KOR = r"C:\Users\top00\JR\01_업무관련\03_인사업무\KNK_직원등록(본사).xlsx"
DEFAULT_VN = r"C:\Users\top00\JR\01_업무관련\03_인사업무\KNK_직원등록(베트남).xlsx"


def digits(s):
    return "".join(ch for ch in str(s or "") if ch.isdigit())


def norm(s):
    return str(s or "").strip()


def read_sheet(path, sheet_name, entity):
    """엑셀 시트 → 직원 dict 리스트. 컬럼 위치 고정(발주서 매핑)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"[FATAL] 시트 없음: {sheet_name} in {path}")
    ws = wb[sheet_name]
    rows = []
    for ri, r in enumerate(ws.iter_rows(values_only=True), 1):
        if ri == 1:
            continue  # 헤더
        name = norm(r[0])
        emp = norm(r[3])
        if not name or not emp:
            continue  # 빈 행 / 사번 없음 → skip
        if entity == "KOR":
            # 0이름 1이메일 2휴대폰 3사번 4직급 5부서 6영문
            rec = dict(name=name, email=norm(r[1]), phone=norm(r[2]), emp=emp,
                       title=norm(r[4]), dept=norm(r[5]),
                       name_en=norm(r[6]), name_vn="", entity="KOR")
        else:
            # 0이름 1이메일 2휴대폰 3사번 4직급 5부서 6베트남어 7영문
            rec = dict(name=name, email=norm(r[1]), phone=norm(r[2]), emp=emp,
                       title=norm(r[4]), dept=norm(r[5]),
                       name_vn=norm(r[6]), name_en=norm(r[7]), entity="VN")
        rows.append(rec)
    return rows


def initial_password(rec):
    """초기 비번: 본사=휴대폰 숫자 / VN=9999 (발주서 §2.3)."""
    if rec["entity"] == "VN":
        return "9999"
    d = digits(rec["phone"])
    return d or "9999"   # 본사 휴대폰 없으면 9999 폴백 (없을 일 거의 없음)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", required=True, help="메신저 SQLite DB 경로")
    ap.add_argument("--kor", default=DEFAULT_KOR)
    ap.add_argument("--vn", default=DEFAULT_VN)
    ap.add_argument("--apply", action="store_true", help="실제 커밋 (없으면 dry-run)")
    args = ap.parse_args()

    if not os.path.exists(args.db):
        raise SystemExit(f"[FATAL] DB 없음: {args.db}")

    # 1) 엑셀 로드
    emps = []
    emps += read_sheet(args.kor, "본사(KOR)", "KOR")
    emps += read_sheet(args.vn, "베트남법인(VN)", "VN")
    print(f"[load] 엑셀 직원 총 {len(emps)}명 (본사+VN)")

    # 1-1) 엑셀 내 사번 중복 검사
    seen, dups = set(), []
    for e in emps:
        if e["emp"] in seen:
            dups.append(e["emp"])
        seen.add(e["emp"])
    if dups:
        print(f"[WARN] 엑셀 내 사번 중복 {len(dups)}건: {sorted(set(dups))} → 해당 행은 건너뜀")

    # 2) DB 연결
    conn = sqlite3.connect(args.db)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # 내부 직원만(고객사 게스트 제외, 시스템 플레이스홀더 제외)
    db_users = cur.execute(
        "SELECT id, username, display_name, email, phone, employee_no, role, active "
        "FROM users WHERE COALESCE(is_guest,0)=0 AND username != '_deleted_user'"
    ).fetchall()
    print(f"[db] 내부 직원 {len(db_users)}명 로드")

    matched_db_ids = set()
    plan_update, plan_insert, plan_conflict = [], [], []
    used_emp = set()

    for e in emps:
        if e["emp"] in used_emp:
            plan_conflict.append((e["emp"], e["name"], "엑셀 사번 중복"))
            continue
        used_emp.add(e["emp"])
        cand = None
        # 1순위: employee_no 직접 일치 (이미 사번이 반영된 운영 상태 — 가장 신뢰도 높음)
        for u in db_users:
            if u["id"] in matched_db_ids:
                continue
            if norm(u["employee_no"]) and norm(u["employee_no"]) == e["emp"]:
                cand = u
                break
        # 2순위: 이름 동일 + (이메일 동일 OR 휴대폰 숫자 동일) — 미반영(사번 없는) 직원
        if not cand:
            for u in db_users:
                if u["id"] in matched_db_ids:
                    continue
                if norm(u["display_name"]) != e["name"]:
                    continue
                email_ok = bool(e["email"]) and norm(u["email"]).lower() == e["email"].lower()
                phone_ok = bool(digits(e["phone"])) and digits(u["phone"]) == digits(e["phone"])
                if email_ok or phone_ok:
                    cand = u
                    break
        if cand:
            matched_db_ids.add(cand["id"])
            plan_update.append((cand, e))
        else:
            plan_insert.append(e)

    # DB-only (엑셀에 없는 직원) → 수동 판단 목록
    db_only = [u for u in db_users if u["id"] not in matched_db_ids]

    # 3) 실행 — username UNIQUE 전이(transient) 충돌 방지 위해 2-패스.
    #    pass1: 매칭 대상 username 을 충돌 불가능한 임시값('__mig_<id>')으로
    #    pass2: 최종 사번 + 나머지 필드. (최종 사번은 dry-run 에서 유일성 확인됨)
    now = datetime.now(timezone.utc).isoformat()
    applied_u = applied_i = 0
    if args.apply:
        for cand, e in plan_update:
            cur.execute("UPDATE users SET username=? WHERE id=?",
                        (f"__mig_{cand['id']}", cand["id"]))
    for cand, e in plan_update:
        if args.apply:
            cur.execute(
                "UPDATE users SET username=?, employee_no=?, entity=?, title=?, department=?, "
                "display_name_en=COALESCE(NULLIF(?,''), display_name_en), "
                "display_name_vn=COALESCE(NULLIF(?,''), display_name_vn) "
                "WHERE id=?",
                (e["emp"], e["emp"], e["entity"], e["title"], e["dept"],
                 e["name_en"], e["name_vn"], cand["id"]),
            )
        applied_u += 1

    for e in plan_insert:
        pw = generate_password_hash(initial_password(e))
        if args.apply:
            cur.execute(
                "INSERT INTO users (username, password_hash, display_name, role, avatar_color, "
                "created_at, email, phone, title, department, employee_no, entity, "
                "display_name_en, display_name_vn, must_change_password, password_version, active, is_guest) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,1,1,1,0)",
                (e["emp"], pw, e["name"], "staff", "#9CA3AF", now,
                 e["email"] or None, e["phone"] or None, e["title"], e["dept"],
                 e["emp"], e["entity"], e["name_en"] or None, e["name_vn"] or None),
            )
        applied_i += 1

    # 4) 리포트
    print("\n" + "=" * 64)
    print(f"  마이그레이션 {'적용(APPLY)' if args.apply else 'DRY-RUN (변경 없음)'}")
    print("=" * 64)
    print(f"  갱신(매칭)  : {applied_u} 건")
    print(f"  신규(INSERT): {applied_i} 건")
    print(f"  충돌(skip)  : {len(plan_conflict)} 건")
    print(f"  DB-only(수동판단 필요): {len(db_only)} 건")
    if plan_insert:
        print("\n  -- 신규 INSERT 예정 --")
        for e in plan_insert:
            print(f"     + {e['emp']:>6}  {e['name']}  ({e['entity']})  초기비번={initial_password(e)}")
    if db_only:
        print("\n  -- DB-only (엑셀 미등재) → 대표 수동 판단(유지/비활성/삭제) --")
        for u in db_only:
            print(f"     ? id={u['id']} username={u['username']!r} name={u['display_name']!r} "
                  f"role={u['role']} active={u['active']}")
    if plan_conflict:
        print("\n  -- 충돌(건너뜀) --")
        for emp, name, why in plan_conflict:
            print(f"     ! {emp} {name} — {why}")
    print("=" * 64)

    if args.apply:
        conn.commit()
        # 검증
        nulls = cur.execute("SELECT COUNT(*) FROM users WHERE COALESCE(is_guest,0)=0 "
                            "AND username != '_deleted_user' AND (employee_no IS NULL OR employee_no='')").fetchone()[0]
        mism = cur.execute("SELECT COUNT(*) FROM users WHERE COALESCE(is_guest,0)=0 "
                           "AND username != '_deleted_user' AND employee_no IS NOT NULL "
                           "AND username != employee_no").fetchone()[0]
        print(f"[verify] employee_no NULL 직원: {nulls}  ·  username≠employee_no: {mism}")
        print("[APPLY] 커밋 완료.")
    else:
        conn.rollback()
        print("[DRY-RUN] 롤백 — DB 변경 없음. 실제 적용은 --apply 추가.")
    conn.close()


if __name__ == "__main__":
    main()
