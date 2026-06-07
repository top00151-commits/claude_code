"""KNK Messenger — 사내 업무 전용 메신저 (1단계 MVP)

독립 프로젝트. 어느 정도 완성 후 HAIST WORKS와 SSO/사용자 API로 연결 예정.
"""
import os
import re
import sys
import uuid
import json
import base64
import mimetypes
import sqlite3
import secrets
from datetime import datetime, timezone
from functools import wraps

try:
    from pywebpush import webpush, WebPushException
    PYWEBPUSH_OK = True
except ImportError:
    PYWEBPUSH_OK = False

from flask import (
    Flask, request, session, redirect, url_for,
    render_template, jsonify, abort, g, send_from_directory, make_response,
)
from flask_socketio import SocketIO, emit, join_room as sio_join, leave_room as sio_leave
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename


APP_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(APP_DIR, "data", "messenger.db")
UPLOAD_DIR = os.path.join(APP_DIR, "data", "uploads")


# --- .env 자동 로딩 (재시작에도 키 유지) — 대표 지시 2026-05-31 ---
#   런처(run_gunicorn.sh 등)가 .env 를 source 하지 않아도, 앱이 시작 시 직접 읽어
#   OPENAI_API_KEY 등이 재시작 후에도 유지되게 함(번역 멈춤 방지).
#   ★ 이미 설정된 '비어있지 않은' env 값은 덮지 않음(명시적 env 우선).
#     단, 런처가 빈 값(OPENAI_API_KEY=)으로 내보낸 경우엔 .env 의 실제 값으로 채움(번역 멈춤 방지).
#   따옴표·주석(#)·'export ' 접두어 처리. 실패해도 부팅 계속(무해).
def _load_dotenv_file():
    env_path = os.path.join(APP_DIR, ".env")
    if not os.path.exists(env_path):
        return
    try:
        with open(env_path, "r", encoding="utf-8") as f:
            for raw in f:
                line = raw.strip()
                if not line or line.startswith("#"):
                    continue
                if line[:7].lower() == "export ":
                    line = line[7:].strip()
                if "=" not in line:
                    continue
                k, v = line.split("=", 1)
                k = k.strip()
                v = v.strip()
                if len(v) >= 2 and v[0] == v[-1] and v[0] in ("\"", "'"):
                    v = v[1:-1]
                if k and not (os.environ.get(k) or "").strip():
                    os.environ[k] = v
    except Exception as e:
        try:
            print(f"[.env] 로드 실패(무시): {e}", flush=True)
        except Exception:
            pass


_load_dotenv_file()
PORT = int(os.environ.get("KNK_MSG_PORT", "5050"))
# 업로드 한도 — DWG 도면·동영상 대비 500MB 기본. 환경변수로 조정 가능.
# Synology Reverse Proxy / nginx 의 client_max_body_size 도 동시에 키워야 함.
MAX_UPLOAD_MB = int(os.environ.get("KNK_MSG_MAX_UPLOAD_MB", "1000"))   # 요청당 전체 1GB (대표 지시 2026-05-19. nginx client_max_body_size 함께 조정 필요)
MESSAGE_RETENTION_MONTHS = int(os.environ.get("KNK_MSG_RETENTION_MONTHS", "12"))
VAPID_PRIV_PATH = os.path.join(APP_DIR, "data", "vapid_private.pem")
VAPID_CONTACT = os.environ.get("KNK_MSG_CONTACT", "mailto:admin@knknara.co.kr")

# ---------- 운영(인터넷) 배포용 환경변수 ----------
# 운영 모드: KNK_MSG_ENV=production 으로 켜면 보안 헤더·HTTPS 강제·CORS 제한 활성
ENV = os.environ.get("KNK_MSG_ENV", "development").lower()
IS_PRODUCTION = ENV == "production"

# === 최고관리자(소유자) — 대표 지시 2026-05-21 ===
# 이 username 계정은 부팅 시 자동으로 관리자(ceo)·활성 보장되고,
# 강등·비활성·삭제가 차단(보호)되며, 화면에 '👑 최고관리자' 로 표시된다.
# 코드/환경변수로 지정되므로 운영 DB 를 직접 안 만져도 배포(재시작)만으로 적용.
OWNER_USERNAME = os.environ.get("KNK_MSG_OWNER_USERNAME", "5").strip().lower()
# 대표 지시 2026-05-26: 사번 = 로그인 ID 전면 변경 → 김정락(사번 5)이 최고관리자

def _is_owner(username):
    """이 계정이 최고관리자(소유자)인지."""
    return bool(username) and str(username).strip().lower() == OWNER_USERNAME


def _is_team_lead(user):
    """직급(title)에 '팀장' 포함 = 팀장. (대표 지시 2026-05-21 — 채널 생성·관리 권한)"""
    try:
        return bool(user) and ("팀장" in (user["title"] or ""))
    except Exception:
        return False


# 채널 생성·관리 권한 직급 키워드: 대표이사·임원(전무·상무·이사)·팀장·법인장 (대표 지시 2026-05-24)
_CHANNEL_TITLE_KEYWORDS = ("대표", "전무", "상무", "이사", "팀장", "법인장")

def _channel_by_title(user):
    """직급(대표·임원·팀장·법인장) 또는 관리자(ceo) → 채널 생성 자동 허용(잠금, 토글로 해제 불가).
    '사람별 추가 허용'과 구분하기 위한 고정 권한 판정. (대표 지시 2026-05-29)"""
    try:
        if not user:
            return False
        if user["role"] == "ceo":
            return True
        title = user["title"] or ""
        return any(kw in title for kw in _CHANNEL_TITLE_KEYWORDS)
    except Exception:
        return False


def _can_create_channel(user):
    """채널을 만들 수 있는 사람: 직급 자동 허용(_channel_by_title) 또는
    관리자가 직급과 무관하게 추가 허용한 직원(users.channel_create_allowed=1). (대표 지시 2026-05-29)
    채널 생성·삭제·기존 채널 멤버 초대 권한 판정에 공통 사용."""
    try:
        if not user:
            return False
        if _channel_by_title(user):
            return True
        # 관리자가 직급과 무관하게 추가 허용 — 컬럼이 없는 행일 수도 있으니 안전 접근
        try:
            if "channel_create_allowed" in user.keys() and int(user["channel_create_allowed"] or 0) == 1:
                return True
        except Exception:
            pass
        return False
    except Exception:
        return False


def _norm_dept(dept):
    """부서 정규화 — 끝의 (괄호) 제거. 동일 부서 인식용.
    예) '04 설계팀(자동화)'·'04 설계팀(검사기)' → '04 설계팀' (대표 지시 2026-05-21)."""
    return re.sub(r'\s*\([^)]*\)\s*$', '', (dept or '').strip())
# Socket.IO async_mode: 개발=threading(Windows OK), 운영=eventlet(gunicorn worker)
ASYNC_MODE = os.environ.get("KNK_MSG_ASYNC", "threading")
# CORS allowed origins: 콤마 구분. 예) "https://o.knknara.co.kr"
_cors_env = os.environ.get("KNK_MSG_CORS", "*")
CORS_ALLOWED = [o.strip() for o in _cors_env.split(",")] if _cors_env != "*" else "*"
# 하위 경로 배포: KNK_MSG_BASE_PATH=/msg 로 설정하면 앱이 /msg 하위에서 서비스됨.
# 역방향 프록시는 도메인 전체를 그대로 전달하므로 수정 불필요.
# 미설정(빈값) 시 기존처럼 루트(/)에서 동작 — 로컬 개발 하위 호환.
BASE_PATH = os.environ.get("KNK_MSG_BASE_PATH", "").strip().rstrip("/")
# HAIST WORKS 진입 주소 — 메신저 'WORKS 열기' 버튼이 SSO 토큰 발급 후 보낼 곳. (대표 지시 2026-05-31)
#   WORKS 의 토큰 수신 랜딩 경로. 기본 https://works.knknara.co.kr/sso/land (01 세션과 합의 경로).
WORKS_LANDING_URL = os.environ.get("KNK_WORKS_LANDING_URL", "https://works.knknara.co.kr/sso/land").strip().rstrip("/")
# 정적 파일 캐시 (운영은 1일, 개발은 0)
STATIC_CACHE_AGE = int(os.environ.get("KNK_MSG_STATIC_CACHE", "86400" if IS_PRODUCTION else "0"))
# 신뢰할 프록시 수 (nginx 등 리버스 프록시 뒤에서 X-Forwarded-* 신뢰)
TRUSTED_PROXIES = int(os.environ.get("KNK_MSG_PROXIES", "1" if IS_PRODUCTION else "0"))

# --- AI 번역·요약·이력정리 (공급자 선택형) ---
# 공급자 = "openai" (ChatGPT) 또는 "anthropic" (Claude). 기본은 openai.
# 향후 Anthropic 결제 풀리면 TRANSLATE_PROVIDER=anthropic 으로만 바꾸면 됨.
TRANSLATE_PROVIDER = os.environ.get("KNK_MSG_TRANSLATE_PROVIDER", "openai").strip().lower()
if TRANSLATE_PROVIDER not in ("openai", "anthropic"):
    TRANSLATE_PROVIDER = "openai"

# OpenAI (ChatGPT)
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "").strip()
OPENAI_MODEL = os.environ.get("KNK_MSG_OPENAI_MODEL", "gpt-5.4-mini")  # 코드 기본값(폴백). 실제 사용 모델은 _get_openai_model() 가 환경변수>app_settings.ai_model>이 값 순으로 결정. 신형 모델은 openai_create() 가 파라미터 자동 대응.
# 사내 프록시·Azure OpenAI 등 대체 endpoint (선택). 비어있으면 OpenAI 공식.
OPENAI_BASE_URL = os.environ.get("KNK_MSG_OPENAI_BASE_URL", "").strip() or None

# Anthropic (Claude) — 결제 활성 후 사용 예정. 코드는 보존.
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "").strip()

# 모델명 (기존 변수 유지 — Anthropic 전용 의미. 라우터에서 공급자별로 분기).
TRANSLATE_MODEL = os.environ.get("KNK_MSG_TRANSLATE_MODEL", "claude-haiku-4-5")
# 월 비용 한도 (USD). 초과 시 신규 번역 차단 (캐시는 계속 동작).
TRANSLATE_MONTHLY_USD_LIMIT = float(os.environ.get("KNK_MSG_TRANSLATE_USD_LIMIT", "20.0"))

# HAIST WORKS 프로젝트 이력 수신 엔드포인트 (대표 지시 2026-06-01, 01세션 협의 발주).
#  비었으면 자동전송 '미설정'(기본) → 어떤 경우에도 전송 안 함. 01 이 엔드포인트 확정하면 이 env 만 채우면 됨.
#  실제 전송은 이 URL + 관리자 토글(app_settings.works_history_sync=1) '둘 다' 일 때만.
WORKS_HISTORY_URL = os.environ.get("KNK_WORKS_HISTORY_URL", "").strip()

# 현재 공급자의 키가 있는지 (UI 활성 판정용)
def _ai_provider_has_key():
    if TRANSLATE_PROVIDER == "openai":
        return bool(OPENAI_API_KEY)
    if TRANSLATE_PROVIDER == "anthropic":
        return bool(ANTHROPIC_API_KEY)
    return False

def _ai_provider_model_label():
    """시작 배너·상태 응답에서 보여줄 모델 이름."""
    if TRANSLATE_PROVIDER == "openai":
        return _get_openai_model()   # 환경변수>app_settings>기본 (요청 컨텍스트 밖이면 기본값 폴백)
    return TRANSLATE_MODEL

# 데모 모드 — API 키 없이 UI 흐름 테스트용. KNK_MSG_TRANSLATE_MOCK=1 이면 가짜 번역 반환.
_mock_env = os.environ.get("KNK_MSG_TRANSLATE_MOCK", "").strip()
if _mock_env == "0":
    TRANSLATE_MOCK = False  # 명시적 OFF
elif _mock_env == "1":
    TRANSLATE_MOCK = True   # 명시적 ON
else:
    # 자동 분기: 개발 환경에서 키 하나도 없으면 데모 모드로 자동 활성 (편의)
    # 운영 환경(KNK_MSG_ENV=production)에서는 자동 활성 X — 실수로 데모 번역이 운영에 나가는 것 방지
    TRANSLATE_MOCK = (not _ai_provider_has_key()) and (not IS_PRODUCTION)
# 지원 언어 (UI 옵션 + 시스템 프롬프트 분기)
TRANSLATE_LANGS = {
    "ko": "한국어",
    "vi": "Tiếng Việt (베트남어)",
    "en": "English",
    "zh": "中文 (Chinese)",
}


def vapid_private_key():
    """VAPID 개인키 파일 경로 반환 — pywebpush.webpush(vapid_private_key=PATH) 에 그대로 사용.
    이전 코드는 PEM 문자열 자체를 반환했는데, 일부 pywebpush 버전이 이를
    raw 32 bytes base64 로 잘못 해석해 ASN.1 파싱 에러 발생. 파일 경로로 넘기면
    pywebpush 가 내부 from_file 로 PEM 정상 로드."""
    if os.path.exists(VAPID_PRIV_PATH):
        return VAPID_PRIV_PATH
    return None


def vapid_public_key_b64u():
    """VAPID 개인키에서 공개키를 raw 65바이트 → base64url 인코딩으로 추출."""
    if not os.path.exists(VAPID_PRIV_PATH):
        return None
    try:
        from cryptography.hazmat.primitives.asymmetric import ec
        from cryptography.hazmat.primitives import serialization
        with open(VAPID_PRIV_PATH, "rb") as f:
            priv = serialization.load_pem_private_key(f.read(), password=None)
        pub = priv.public_key().public_bytes(
            encoding=serialization.Encoding.X962,
            format=serialization.PublicFormat.UncompressedPoint,
        )
        return base64.urlsafe_b64encode(pub).rstrip(b"=").decode()
    except Exception:
        return None


# ---------- Presence (PC 활성 시 모바일 푸시 억제) ----------
# 메모리 내 SID 매핑: uid -> { sid: {"device": "pc"|"mobile"|"unknown", "active": bool, "ts": float} }
# 단일 worker (eventlet) 환경 가정. 멀티 워커 확장 시 Redis 등으로 교체 필요.
import threading as _pres_threading
import time as _pres_time
_user_connections = {}
_user_conn_lock = _pres_threading.Lock()
# 회사망(사무실) 공인 IP 집합 — 접속 IP 가 여기 있으면 '🏢 회사' 표시. init_db 에서 로딩 + 관리자 변경 시 갱신.
_office_ips = set()

def _presence_register(uid, sid, device="unknown", active=True, idle=False, ip=None):
    """SocketIO 연결 등록·갱신. idle=True 면 앱은 보고 있으나 일정시간 무조작(자동 자리비움).
    ip = 접속자 공인 IP (회사망 여부 판별용 — 메모리에만 보관, 영구 저장 안 함).
    active_room 은 기존 값 보존 (set_active_room 이벤트로만 변경)."""
    if not uid or not sid:
        return
    if device not in ("pc", "mobile", "unknown"):
        device = "unknown"
    with _user_conn_lock:
        existing = _user_connections.setdefault(uid, {}).get(sid, {})
        _user_connections[uid][sid] = {
            "device": device,
            "active": bool(active),
            "idle": bool(idle),
            "ip": ip or "",
            "ts": _pres_time.time(),
            # active_room: 사용자가 현재 보고 있는 방 ID — 푸시 발송 전 비교용 (대표 지시 2026-05-26)
            "active_room": existing.get("active_room"),
        }


def _presence_set_active_room(uid, sid, room_id):
    """사용자가 현재 보고 있는 방 갱신 — push_message_to_room_members 가 같은 방이면 푸시 스킵.
    Chrome PWA 의 SW clients.matchAll() 버그 우회용 (대표 지시 2026-05-26)."""
    if not uid or not sid:
        return
    with _user_conn_lock:
        conn = _user_connections.get(uid, {}).get(sid)
        if conn is not None:
            conn["active_room"] = int(room_id) if room_id else None
            conn["ts"] = _pres_time.time()


def _user_is_viewing_room(uid, room_id):
    """uid 의 어떤 소켓이든 room_id 를 active_room 으로 가지고 있고
    최근(_PC_ACTIVE_STALE_SEC 이내) 활성이면 True → 푸시 스킵.
    소켓 끊김(=백그라운드 추정)이거나 다른 방 보면 False → 푸시 발송."""
    if not uid or not room_id:
        return False
    try:
        rid = int(room_id)
    except Exception:
        return False
    now = _pres_time.time()
    with _user_conn_lock:
        for info in _user_connections.get(uid, {}).values():
            if info.get("active_room") != rid:
                continue
            # 최근 갱신·active 상태만 인정 — 오래된 stale 정보로 푸시 차단 방지
            if now - info.get("ts", 0) > _PC_ACTIVE_STALE_SEC:
                continue
            return True
    return False

def _presence_unregister(uid, sid):
    """SocketIO 연결 해제."""
    if not uid or not sid:
        return
    with _user_conn_lock:
        conns = _user_connections.get(uid)
        if not conns:
            return
        conns.pop(sid, None)
        if not conns:
            _user_connections.pop(uid, None)

# PC presence 가 이 시간(초) 이상 갱신 안 되면 "실제로 보고 있지 않음"(절전·자리비움)으로 간주.
# 클라이언트가 30초마다 heartbeat 를 보내므로, 60초 = heartbeat 2회 누락 시 비활성 판정.
# (대표 지시 2026-05-20: PC 켜둬도 자리 비우면 모바일 알림 와야 함)
_PC_ACTIVE_STALE_SEC = 60   # JS 하트비트 30초 간격 보다 충분히 길게 — 타이밍 어긋남 방지 (2026-05-27 복원)
                            # (5-26 30 으로 줄였더니 PC 더블 알림 발생 — heartbeat 직전 timing 에서 stale 판정)

# 활성 기기가 없을 때(창 비활성·다른 앱·화면 꺼짐·잠금) 마지막 활동 후 이 시간 이상 지나면 '자리비움'.
# 회사 전체 공통값 — 관리자(대표)가 15분/30분 중 선택 (app_settings.away_minutes). 기본 15분.
# 클라가 신호를 못 보내므로 _presence_away_sweep 가 주기 재계산. (대표 지시 2026-05-24)
AWAY_MINUTES_DEFAULT = 15
AWAY_MINUTES_ALLOWED = (15, 30)
_away_minutes_cache = AWAY_MINUTES_DEFAULT   # 메모리 캐시 (init_db·관리자 변경 시 갱신)

def _away_after_sec():
    """자리비움 임계값(초) — 회사 공통 설정값 기반."""
    return _away_minutes_cache * 60

def _set_away_minutes_cache(m):
    """관리자 변경 시 메모리 캐시 갱신 (허용값만)."""
    global _away_minutes_cache
    try:
        m = int(m)
    except Exception:
        return
    if m in AWAY_MINUTES_ALLOWED:
        _away_minutes_cache = m

def _user_has_active_pc(uid):
    """해당 사용자의 PC 연결 중 '실제로 활성'(포커스 + 최근 heartbeat) 게 있으면 True.
    이 때만 모바일 푸시를 발송하지 않음 — PC 활성 감지 + 자리비움 보정.
    PC active 라도 마지막 presence 갱신이 _PC_ACTIVE_STALE_SEC 이상 오래됐으면
    절전·잠금·자리비움으로 보고 False 반환 → 모바일 푸시 정상 발송."""
    if not uid:
        return False
    now = _pres_time.time()
    with _user_conn_lock:
        conns = _user_connections.get(uid, {})
        for sid, info in conns.items():
            if info.get("device") == "pc" and info.get("active"):
                ts = info.get("ts", 0)
                if now - ts <= _PC_ACTIVE_STALE_SEC:
                    return True   # PC 가 진짜 활성 (최근 heartbeat 있음)
    return False


def _user_has_active_session(uid):
    """사용자가 '실제로 앱을 보고 있는' 활성 PC/태블릿 연결이 하나라도 있으면 True → 푸시 skip.
    휴대폰(device='mobile')은 항상 푸시 발송 (대표 지시 2026-05-26 — 휴대폰 알림이 늦는 문제 해결):
      · 휴대폰은 백그라운드 가도 socket 끊김/heartbeat 멈춤 타이밍 차이로 active=True 잔존이 흔함
      · 그래서 stale 60초 안에 메시지 오면 푸시 skip → 알림 안 옴
      · 휴대폰은 인앱 + OS 푸시 동시 OK (사용자가 화면 켜고 있어도 푸시 도착)
    PC만 active 판정 — 보고 있는 동안만 푸시 skip (자리 비우면 즉시 푸시).
    """
    if not uid:
        return False
    now = _pres_time.time()
    with _user_conn_lock:
        for info in _user_connections.get(uid, {}).values():
            # 휴대폰은 active 판정 안 함 → 항상 푸시
            if info.get("device") == "mobile":
                continue
            if info.get("active") and (now - info.get("ts", 0) <= _PC_ACTIVE_STALE_SEC):
                return True
    return False


def _user_is_online(uid):
    """uid 의 활성 SocketIO 연결이 하나라도 있으면 True (= 로그인 상태).
    아무 연결도 없으면 False (= 미접속·오프라인).
    상태 점등용 — 사용자가 별도 상태 설정을 하지 않아도 본 함수가 False 면 'offline' 강제."""
    if not uid:
        return False
    with _user_conn_lock:
        return bool(_user_connections.get(uid))


def _user_has_pc_connection(uid):
    """uid 의 현재 소켓 연결 중 'pc' 기기가 하나라도 있으면 True.
    상태 자동표시용 — PC 접속 있으면 '가능', 휴대폰만이면 '휴대폰'.
    (절전·자리비움 stale 체크는 안 함 — 그건 푸시 억제용 _user_has_active_pc 의 역할)."""
    if not uid:
        return False
    with _user_conn_lock:
        for info in _user_connections.get(uid, {}).values():
            if info.get("device") == "pc":
                return True
    return False


def _active_presence(uid):
    """지금 '실제로 보고 있는(쓰는)' 기기와 그 기기의 idle(자동 자리비움) 여부를 반환: (device, idle).
    조건: active(포커스/화면 ON) + 최근 heartbeat(_PC_ACTIVE_STALE_SEC 이내) 연결만 후보.
    둘 다 활성이면 가장 최근(ts) 기기 = 마지막으로 만진 기기. 아무 활성도 없으면 (None, False).
    상태 자동표시용 — 휴대폰·PC 동시 로그인이어도 '쓰는 기기'로 표기 + 무조작 시 자리비움. (대표 지시 2026-05-23)"""
    if not uid:
        return (None, False)
    now = _pres_time.time()
    best_dev, best_ts, best_idle = None, -1.0, False
    with _user_conn_lock:
        for info in _user_connections.get(uid, {}).values():
            dev = info.get("device")
            if dev not in ("pc", "mobile"):
                continue
            if not info.get("active"):
                continue
            ts = info.get("ts", 0)
            if now - ts > _PC_ACTIVE_STALE_SEC:
                continue  # 활성이라 표시됐지만 heartbeat 끊김(절전·잠금) → 제외
            if ts > best_ts:
                best_ts, best_dev, best_idle = ts, dev, bool(info.get("idle"))
    return (best_dev, best_idle)


def _active_device(uid):
    """가장 최근 활성 기기만 반환 (idle 무관). 기존 호출 호환용."""
    return _active_presence(uid)[0]


def _user_idle_seconds(uid):
    """마지막 presence 갱신(heartbeat·포커스·블러) 이후 경과 초.
    활성(포커스+화면ON) 동안엔 heartbeat 가 30초마다 ts 를 갱신해 작게 유지되고,
    창 비활성·다른 앱·화면 꺼짐이면 ts 가 고정돼 이 값이 '자리 뜬 시간' 근사치가 된다.
    연결이 하나도 없으면 매우 큰 값(=오래 자리 비움) 반환."""
    if not uid:
        return 1e9
    now = _pres_time.time()
    latest = 0.0
    with _user_conn_lock:
        for info in _user_connections.get(uid, {}).values():
            ts = info.get("ts", 0) or 0
            if ts > latest:
                latest = ts
    return (now - latest) if latest > 0 else 1e9


# ── 회사망(사무실) 감지 — 접속 IP 가 등록된 회사 공인 IP 면 '🏢 회사' 표시 ──
def _load_office_ips():
    """office_networks 테이블 → 메모리 캐시 갱신. (요청/앱컨텍스트 내에서 호출)"""
    global _office_ips
    try:
        rows = get_db().execute("SELECT ip FROM office_networks").fetchall()
        _office_ips = set((r["ip"] or "").strip() for r in rows if r["ip"])
    except Exception:
        pass
    return _office_ips

def _real_client_ip():
    """리버스 프록시 뒤 실제 접속자 공인 IP. (프록시가 넣어주는 X-Forwarded-For 우선)"""
    try:
        xff = (request.headers.get("X-Forwarded-For") or "").strip()
        if xff:
            return xff.split(",")[0].strip()
        return request.remote_addr or ""
    except Exception:
        return ""

def _is_office_ip(ip):
    return bool(ip) and ip in _office_ips

def _user_at_office(uid):
    """uid 의 현재 연결 중 하나라도 회사망(등록된 공인 IP)에서 접속했으면 True.
    (연결마다 보관한 IP 를 현재 회사망 목록과 대조 → 목록이 바뀌어도 즉시 재평가됨)"""
    if not uid or not _office_ips:
        return False
    with _user_conn_lock:
        for info in _user_connections.get(uid, {}).values():
            if _is_office_ip(info.get("ip")):
                return True
    return False


def send_push_to_user(user_id, title, body, url=None, tag=None, collect_errors=False, clear=False):
    """특정 사용자의 모든 push 구독에 알림 전송. 410/404는 만료로 간주하고 삭제.
    collect_errors=True 이면 (sent_count, [{id, endpoint, error}], total_subs) 튜플 반환.
    clear=True 이면 알림을 띄우는 대신 '읽음' 신호(type=clear)를 보내 sw.js 가 해당 방(tag)
    알림을 닫고 배지만 갱신하게 한다 — 다른 기기(백그라운드 휴대폰)의 알림 자동 삭제용."""
    errors = []
    if not PYWEBPUSH_OK:
        if collect_errors:
            return 0, [{"error": "PYWEBPUSH_OK=False (pywebpush 모듈 import 실패)"}], 0
        return 0
    priv = vapid_private_key()
    if not priv:
        if collect_errors:
            return 0, [{"error": "VAPID 개인키 파일(data/vapid_private.pem) 없음"}], 0
        return 0
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    try:
        subs = db.execute(
            "SELECT id, endpoint, p256dh, auth FROM push_subscriptions WHERE user_id = ?",
            (user_id,),
        ).fetchall()
        # 사용자별 알림 내용 숨기기 옵션 (대표 지시 2026-05-26):
        #   1이면 title='KNK message', body='' 로 강제 — 잠금화면·메시지창 모두 가림
        try:
            _hide_row = db.execute("SELECT push_hide_preview FROM users WHERE id=?", (user_id,)).fetchone()
            _hide_preview = bool(_hide_row and _hide_row["push_hide_preview"])
        except Exception:
            _hide_preview = False
        if _hide_preview and not clear:
            title = "KNK message"
            body = ""
        # 앱 아이콘 배지용 — 이 사용자의 전체 안 읽은 메시지 총합 (방 목록 unread 와 동일 계산)
        badge_count = 0
        try:
            brow = db.execute("""
                SELECT COALESCE(SUM(cnt), 0) AS total FROM (
                    SELECT (SELECT COUNT(*) FROM messages m
                              WHERE m.room_id = rm.room_id
                                AND m.id > rm.last_read_message_id
                                AND m.user_id != ?
                                AND (m.whisper_to_user_id IS NULL OR m.whisper_to_user_id = ?)
                           ) AS cnt
                      FROM room_members rm
                     WHERE rm.user_id = ?
                )
            """, (user_id, user_id, user_id)).fetchone()
            badge_count = (brow["total"] or 0) if brow else 0
        except Exception:
            badge_count = 0
        total = len(subs)
        sent = 0
        # 구독 병렬 전송 (대표 지시 2026-05-27) — 같은 사용자의 여러 기기 구독에 동시 전송.
        # 옛 sequential 방식: 3 기기 = 3회 HTTP 순차 (150~600ms)
        # 새 parallel 방식  : 3 기기 = 1회분 시간 (50~200ms)
        # FCM·Mozilla AutoPush 등은 동시 호출 OK.
        from concurrent.futures import ThreadPoolExecutor, as_completed
        payload = json.dumps(
            {"type": "clear", "tag": tag, "badge": badge_count}
            if clear else
            {"title": title, "body": body, "url": url or "/chat", "tag": tag, "badge": badge_count}
        )

        def _send_one(s):
            """단일 구독에 webpush 전송. (sent_inc, error_dict_or_none, delete_id_or_none) 반환."""
            ep_prefix = s["endpoint"][:60] + "..." if s["endpoint"] else "?"
            try:
                webpush(
                    subscription_info={
                        "endpoint": s["endpoint"],
                        "keys": {"p256dh": s["p256dh"], "auth": s["auth"]},
                    },
                    data=payload,
                    vapid_private_key=priv,
                    vapid_claims={"sub": VAPID_CONTACT},
                    ttl=43200,  # 12시간 — 휴대폰 일시 끊김에도 도착 유지
                    # Urgency: high — Android Doze 모드/Battery Saver 우회.
                    # Topic 헤더 제거 (2026-05-27): collapse 차단 → 모든 push 도착.
                    headers={"Urgency": "high"},
                )
                return (1, None, None)
            except WebPushException as e:
                code = getattr(e.response, "status_code", None) if e.response is not None else None
                body_resp = ""
                try:
                    if e.response is not None:
                        body_resp = e.response.text[:200]
                except Exception:
                    pass
                err_msg = f"WebPushException code={code} body={body_resp} msg={str(e)[:200]}"
                print(f"[push] FAIL sub_id={s['id']} {ep_prefix} → {err_msg}")
                return (0, {"id": s["id"], "endpoint": ep_prefix, "error": err_msg},
                        s["id"] if code in (404, 410) else None)
            except Exception as e:
                err_msg = f"{type(e).__name__}: {str(e)[:300]}"
                print(f"[push] EXCEPTION sub_id={s['id']} {ep_prefix} → {err_msg}")
                return (0, {"id": s["id"], "endpoint": ep_prefix, "error": err_msg}, None)

        # 사용자당 구독 보통 1~5개 — 작은 풀로 충분
        with ThreadPoolExecutor(max_workers=max(1, min(8, len(subs)))) as ex:
            futures = [ex.submit(_send_one, s) for s in subs]
            for f in as_completed(futures):
                try:
                    inc, err, delete_id = f.result()
                    sent += inc
                    if err:
                        errors.append(err)
                    if delete_id:
                        db.execute("DELETE FROM push_subscriptions WHERE id = ?", (delete_id,))
                        db.commit()
                except Exception as e:
                    print(f"[push] worker error: {e}")

        if collect_errors:
            return sent, errors, total
        return sent
    finally:
        db.close()


def push_message_to_room_members(room_id, sender_user_id, title, body, url=None, tag=None):
    """방의 sender 외 모든 멤버에게 push (백그라운드 알림)."""
    if not PYWEBPUSH_OK:
        return
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    try:
        members = db.execute(
            "SELECT user_id FROM room_members WHERE room_id = ? AND user_id != ?",
            (room_id, sender_user_id),
        ).fetchall()
    finally:
        db.close()
    # tag 에서 room_id 추출 — active_room 비교용 (대표 지시 2026-05-26)
    target_room_id = None
    if tag and isinstance(tag, str) and tag.startswith("room_"):
        try:
            target_room_id = int(tag.replace("room_", ""))
        except Exception:
            target_room_id = None
    for m in members:
        uid = m["user_id"]
        # 1) PC 활성 세션 있음 (자리에 있고 보고 있음) → 푸시 스킵 (인앱 알림으로 충분)
        # 2) 휴대폰이라도 그 방을 active_room 으로 보고 있으면 푸시 스킵 (대표 지시 2026-05-26)
        #     ← Chrome PWA 의 SW clients.matchAll() 버그 우회 — SW 가 못 보면 서버가 본다
        if _user_has_active_session(uid):
            continue
        if target_room_id and _user_is_viewing_room(uid, target_room_id):
            continue
        send_push_to_user(uid, title, body, url=url, tag=tag)
# ============================================================
# 🛡️ 파일 업로드 보안 정책 (대표 지시 2026-05-19 강화)
#   1) 화이트리스트 — 허용된 확장자만 통과
#   2) 블랙리스트 — 위험 실행 파일 명시 차단 (defense-in-depth)
#   3) Magic Number 검증 — 확장자 위조 차단 (예: report.pdf 인데 PE 헤더)
#   4) 더블 확장자 차단 — 예: "report.pdf.exe"
#   5) 단일 파일 크기 — 별도 제한 (PER_FILE_MAX_MB)
# ============================================================
DANGEROUS_EXT = {
    # Windows 실행
    "exe", "scr", "com", "bat", "cmd", "pif", "msi", "msu", "msp", "cpl",
    "vb", "vbs", "vbe", "js", "jse", "wsf", "wsh", "ps1", "ps2", "psc1", "psc2",
    "lnk", "url", "inf", "reg",
    # 매크로 가능 Office
    "docm", "xlsm", "pptm", "dotm", "xltm", "potm",
    # Linux/Mac 실행
    "sh", "bash", "zsh", "csh", "ksh", "fish",
    "app", "dmg", "pkg", "deb", "rpm",
    # 기타 위험
    "jar", "class", "war", "ear",   # Java
    "py", "pl", "rb", "php",         # 스크립트
    "apk", "ipa",                    # 모바일
    # html/htm 은 허용으로 풀림 (대표 지시 2026-05-27 옵션 A — 사내 신뢰 환경)
    # 다운로드 시 Content-Disposition: attachment + application/octet-stream 강제 → 브라우저 인라인 실행 차단
    "xhtml", "shtml",                # HTML 변종은 계속 차단 (드물고 위험)
    "svg",                           # SVG (script 임베드 가능) — 별도 결정 전까지 차단 유지
    "iso", "img",                    # 디스크 이미지
}
# 실행 파일 magic number — 확장자 위조에도 검출
EXECUTABLE_MAGIC = [
    b"MZ",          # Windows PE (exe, dll, scr ...)
    b"\x7fELF",     # Linux ELF
    b"\xca\xfe\xba\xbe",  # Mach-O fat binary
    b"\xcf\xfa\xed\xfe",  # Mach-O 64bit
    b"\xfe\xed\xfa\xce",  # Mach-O 32bit
    b"\xfe\xed\xfa\xcf",  # Mach-O 64bit BE
    b"PK\x03\x04",        # zip (jar/apk 도 zip — 확장자로 추가 판단)
    b"#!",                # Shebang (스크립트)
]
PER_FILE_MAX_MB = int(os.environ.get("KNK_MSG_PER_FILE_MAX_MB", "500"))   # 단일 파일 500MB (대표 지시 2026-05-19)

# ============================================================
# 🛡️ Rate Limiter (In-Memory, 분 단위 슬라이딩 윈도)
#   uid + 액션 별로 분당 호출 횟수 제한. 스팸·DoS 차단.
#   단일 worker 환경 가정. 다중 worker 확장 시 Redis 로 교체.
# ============================================================
import collections as _rl_collections
_rate_buckets = _rl_collections.defaultdict(list)  # (uid, action) -> [timestamps]
_rate_lock = _pres_threading.Lock()

def _check_rate_limit(uid, action, max_per_minute=60):
    """uid 의 액션 분당 횟수 검사. 통과 시 True, 초과 시 False.
    호출 시점을 기록 → 60초 이전 기록은 자동 정리."""
    if not uid:
        return True
    now = _pres_time.time()
    cutoff = now - 60.0
    key = (uid, action)
    with _rate_lock:
        bucket = _rate_buckets[key]
        # 60초 이전 기록 제거
        while bucket and bucket[0] < cutoff:
            bucket.pop(0)
        if len(bucket) >= max_per_minute:
            return False
        bucket.append(now)
    return True

def _is_dangerous_filename(filename):
    """파일명 자체에 위험 패턴이 있는지 확인.
    - 더블 확장자 (report.pdf.exe)
    - 위험 확장자 (exe, bat, ...)
    - 숨김 파일 (.htaccess)"""
    if not filename:
        return True, "파일명이 비어 있습니다"
    fn = filename.lower().strip()
    # 더블 확장자 — 위험 확장자가 마지막이 아닌 중간에 있으면 의심
    parts = fn.split(".")
    if len(parts) >= 3:
        # 마지막에서 두 번째 부분이 위험 확장자면 → 위장 의심
        for p in parts[1:-1]:
            if p in DANGEROUS_EXT:
                return True, f"이중 확장자 의심 (.{p}.~)"
    last_ext = parts[-1] if len(parts) > 1 else ""
    if last_ext in DANGEROUS_EXT:
        return True, f"실행 파일 확장자 차단 (.{last_ext})"
    return False, ""

def _check_executable_magic(file_storage):
    """파일 첫 16바이트 읽어 magic number 검사 — 실행 파일이면 차단.
    호출 후 stream pointer 를 처음으로 되돌려 놓음."""
    try:
        header = file_storage.stream.read(16)
        file_storage.stream.seek(0)
    except Exception:
        return False, ""
    for sig in EXECUTABLE_MAGIC:
        if header.startswith(sig):
            # PK (zip) 은 docx/xlsx 등도 zip 이므로 별도 처리. 여기는 확장자가 이미 화이트리스트 통과한 후라
            # zip 매직이라도 OK. 다만 명시적으로 zip 확장자 아니면서 PK 면 위험.
            if sig == b"PK\x03\x04":
                continue
            # 셔뱅 #! — 텍스트 파일에도 있을 수 있으나 .txt 화이트리스트 안에서 보면 위험 X
            if sig == b"#!":
                continue
            return True, f"실행 파일 헤더 검출 (magic: {sig.hex()})"
    return False, ""


# ============================================================
# 🦠 바이러스 검사 (ClamAV) — 대표 지시 2026-05-23
#   기본 비활성(dormant): 환경변수 KNK_MSG_CLAMAV=1 이고 clamd 에 연결될 때만 동작.
#   설치/켜기 전에는 _scan_file_for_virus() 가 즉시 (False, "disabled") 반환 → 업로드 그대로 통과(무해).
#   서버담당자가 NAS 에 clamd(ClamAV 데몬) 설치 + 환경변수 설정만 하면 자동 활성화.
#
#   활성화 방법(서버담당자):
#     1) ClamAV 설치 + clamd 데몬 실행 (TCP 3310 또는 unix socket), freshclam 으로 DB 갱신
#     2) .env 에 KNK_MSG_CLAMAV=1
#        (TCP)   KNK_MSG_CLAMAV_HOST=127.0.0.1  KNK_MSG_CLAMAV_PORT=3310
#        (socket) KNK_MSG_CLAMAV_SOCKET=/var/run/clamav/clamd.ctl
#     3) (선택) KNK_MSG_CLAMAV_STRICT=1 → clamd 장애 시 업로드 차단(기본은 통과+경고로깅)
#   외부 라이브러리 없이 clamd INSTREAM 프로토콜 직접 구현 (의존성 추가 없음).
# ============================================================
CLAMAV_ENABLED = os.environ.get("KNK_MSG_CLAMAV", "").lower() in ("1", "true", "yes", "on")
CLAMAV_STRICT  = os.environ.get("KNK_MSG_CLAMAV_STRICT", "").lower() in ("1", "true", "yes", "on")
CLAMAV_HOST    = os.environ.get("KNK_MSG_CLAMAV_HOST", "127.0.0.1")
CLAMAV_PORT    = int(os.environ.get("KNK_MSG_CLAMAV_PORT", "3310"))
CLAMAV_SOCKET  = os.environ.get("KNK_MSG_CLAMAV_SOCKET", "")  # 있으면 unix socket 우선
CLAMAV_TIMEOUT = int(os.environ.get("KNK_MSG_CLAMAV_TIMEOUT", "30"))


def _clamd_connect():
    """clamd 소켓 연결 (unix socket 우선, 없으면 TCP). 실패 시 None."""
    import socket as _sock
    try:
        if CLAMAV_SOCKET:
            s = _sock.socket(_sock.AF_UNIX, _sock.SOCK_STREAM)
            s.settimeout(CLAMAV_TIMEOUT)
            s.connect(CLAMAV_SOCKET)
        else:
            s = _sock.create_connection((CLAMAV_HOST, CLAMAV_PORT), timeout=CLAMAV_TIMEOUT)
            s.settimeout(CLAMAV_TIMEOUT)
        return s
    except Exception:
        return None


def _scan_file_for_virus(filepath):
    """파일을 ClamAV(clamd)로 검사.
    반환: (is_infected: bool, detail: str)
      · 비활성/연결불가/검사오류 → (False, '<사유>')  ← 기본은 '통과(무해)'로 처리(가용성 우선)
        단 CLAMAV_STRICT=1 이면 연결불가·오류 시 (True, ...)로 차단.
      · 감염 → (True, '바이러스명')
    호출측: is_infected 면 파일 삭제 + 업로드 거절."""
    if not CLAMAV_ENABLED:
        return (False, "disabled")
    s = _clamd_connect()
    if s is None:
        # clamd 미설치/미기동 — 기본은 통과(업로드 안 끊김), STRICT 면 차단
        print("[clamav] clamd 연결 실패 — 검사 건너뜀", flush=True)
        return ((True, "백신 서버 연결 실패(STRICT)") if CLAMAV_STRICT else (False, "clamd-unreachable"))
    try:
        s.sendall(b"nINSTREAM\n")
        with open(filepath, "rb") as fp:
            while True:
                chunk = fp.read(64 * 1024)
                if not chunk:
                    break
                s.sendall(len(chunk).to_bytes(4, "big") + chunk)
        s.sendall((0).to_bytes(4, "big"))  # 종료 신호 (length 0)
        resp = b""
        while True:
            part = s.recv(4096)
            if not part:
                break
            resp += part
            if b"\n" in part or len(resp) > 8192:
                break
        text = resp.decode("utf-8", "ignore").strip()
        # 정상: "stream: OK"  /  감염: "stream: Eicar-Test-Signature FOUND"
        if "FOUND" in text:
            virus = text.replace("stream:", "").replace("FOUND", "").strip()
            print(f"[clamav] 감염 검출: {virus} ({filepath})", flush=True)
            return (True, virus or "Unknown")
        return (False, "clean")
    except Exception as e:
        print(f"[clamav] 검사 오류: {e}", flush=True)
        return ((True, f"백신 검사 오류(STRICT): {e}") if CLAMAV_STRICT else (False, f"scan-error"))
    finally:
        try: s.close()
        except Exception: pass


ALLOWED_IMAGE_EXT = {"jpg", "jpeg", "png", "gif", "webp", "bmp", "heic"}
ALLOWED_FILE_EXT = ALLOWED_IMAGE_EXT | {
    "pdf", "doc", "docx", "xls", "xlsx", "ppt", "pptx", "hwp", "hwpx",
    "txt", "csv", "zip", "7z", "rar", "dwg", "dxf", "step", "stp", "stl",
    "mp4", "mov", "avi", "mkv", "mp3", "wav",
    # HTML — 사내 트리도·시각화 자료용 (대표 지시 2026-05-27 옵션 A)
    # 다운로드 시 자동 실행 차단 — serve_upload 가 octet-stream 으로 강제
    "html", "htm",
    # 텍스트·문서 보조 형식 (대표 지시 2026-05-28) — 마크다운 우선
    "md", "markdown",     # 마크다운
    "json", "yaml", "yml", # 설정·데이터 교환
    "log",                 # 로그 파일
    # CAD·3D 설계 파일 (대표 지시 2026-06-05 — SolidWorks 등 설계파일 공유). 전부 데이터 파일(실행 아님).
    "x_t", "x_b",                       # Parasolid (중립 교환)
    "sldprt", "sldasm", "slddrw",       # SolidWorks 네이티브
    "igs", "iges",                      # IGES 중립 (step·stp·stl 은 위에 이미 있음)
    "prt", "asm", "ipt", "iam",         # Creo/NX/Pro-E · Inventor
    "catpart", "catproduct",            # CATIA
    "3dm", "sat", "jt", "obj", "3mf",   # Rhino·ACIS·JT·OBJ·3MF
}

# 브라우저가 인라인 실행 가능한 위험 확장자 — 다운로드 강제 + mimetype 변경 필요
UNSAFE_INLINE_EXT = {"html", "htm", "xhtml", "shtml", "svg", "xml", "mhtml", "mht"}

def _load_or_generate_secret():
    """SECRET_KEY 자동 생성·영속화. 환경변수 우선, 없으면 data/secret.key 사용·생성."""
    env_key = os.environ.get("KNK_MSG_SECRET")
    if env_key and len(env_key) >= 16:
        return env_key
    sec_path = os.path.join(APP_DIR, "data", "secret.key")
    os.makedirs(os.path.dirname(sec_path), exist_ok=True)
    if os.path.exists(sec_path):
        with open(sec_path, "r", encoding="utf-8") as f:
            v = f.read().strip()
            if len(v) >= 16:
                return v
    import secrets as _secrets
    new_key = _secrets.token_hex(32)
    with open(sec_path, "w", encoding="utf-8") as f:
        f.write(new_key)
    print(f" * SECRET_KEY 신규 생성·저장: {sec_path}")
    return new_key


app = Flask(__name__)
app.config["SECRET_KEY"] = _load_or_generate_secret()
app.config["JSON_AS_ASCII"] = False
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_MB * 1024 * 1024
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = STATIC_CACHE_AGE
app.config["TEMPLATES_AUTO_RELOAD"] = not IS_PRODUCTION  # 개발만 자동 리로드
app.jinja_env.auto_reload = not IS_PRODUCTION

# 정적 자산 캐시 버스팅 — 서버 시작 시각을 버전으로 사용. 코드 수정 후 재시작 시 자동으로 새 버전.
import time as _time
STATIC_VERSION = str(int(_time.time()))

@app.context_processor
def _inject_asset_version():
    # 상태표시 상세(기기·회사망) on/off — 화면 렌더에 주입. DB 미가용/로그인전 등 예외시 단순(False).
    try:
        _presence_detail = _presence_show_detail(get_db())
    except Exception:
        _presence_detail = False
    return {
        "asset_version": STATIC_VERSION,
        "base_path": BASE_PATH,
        "retention_months": MESSAGE_RETENTION_MONTHS,
        # AI 표식(원형 'AI' 배지) 색 판정 — AI 서비스 가용 여부(전역). True=파랑/False=빨강.
        # (대표 지시 2026-05-31, HAIST WORKS AI 아이콘 복제)
        "ai_on": bool(_ai_provider_has_key() or TRANSLATE_MOCK),
        # 상태표시 상세 — True=기기/회사망 표시, False=🟢 녹색 단순 (대표 지시 2026-06-01)
        "presence_detail": bool(_presence_detail),
        # 시작화면(스플래시/광고) 설정 — 모든 페이지(로그인·대화·게스트)에 주입 (대표 지시 2026-06-06)
        "splash": _splash_config_safe(),
    }


@app.route("/api/version")
def api_version():
    """현재 서버(앱) 버전. 클라이언트가 로드 시 버전과 비교해 새 배포를 감지 → 새로고침 안내.
    (서버 재시작 때마다 STATIC_VERSION 갱신) — 가벼운 공개 엔드포인트."""
    return jsonify({"version": STATIC_VERSION})

# 세션 자동 로그인 유지 — 명시적 로그아웃 전까지 90일 (모바일 사용성)
# session.permanent=True 와 함께 동작. 환경변수로 조정 가능.
from datetime import timedelta as _timedelta
SESSION_DAYS = int(os.environ.get("KNK_MSG_SESSION_DAYS", "90"))
app.config["PERMANENT_SESSION_LIFETIME"] = _timedelta(days=SESSION_DAYS)
app.config["SESSION_REFRESH_EACH_REQUEST"] = True  # 매 요청마다 만료 시간 갱신 (활성 사용자는 영구 유지)

# 운영 환경: 세션 쿠키 보안 강화 + HTTPS 강제
if IS_PRODUCTION:
    app.config["SESSION_COOKIE_SECURE"] = True
    app.config["SESSION_COOKIE_HTTPONLY"] = True
    app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
    app.config["PREFERRED_URL_SCHEME"] = "https"
    # nginx 등 리버스 프록시 뒤에서 X-Forwarded-Proto/Host 신뢰
    if TRUSTED_PROXIES > 0:
        try:
            from werkzeug.middleware.proxy_fix import ProxyFix
            app.wsgi_app = ProxyFix(
                app.wsgi_app,
                x_for=TRUSTED_PROXIES, x_proto=TRUSTED_PROXIES,
                x_host=TRUSTED_PROXIES, x_prefix=TRUSTED_PROXIES,
            )
        except Exception:
            pass

socketio = SocketIO(
    app,
    cors_allowed_origins=CORS_ALLOWED,
    async_mode=ASYNC_MODE,
    # SocketIO 버퍼는 단일 메시지 한도 — HTTP 업로드는 별도 (Flask MAX_CONTENT_LENGTH 가 1GB).
    # 1GB 버퍼는 OOM 위험 (단일 worker, 4-8GB RAM) → 512MB 로 제한 (2026-05-20).
    max_http_buffer_size=512 * 1024 * 1024,
)


# /msg 같은 하위 경로 배포 지원 — PATH_INFO 에서 BASE_PATH 를 떼어 SCRIPT_NAME 으로 옮긴다.
# 결과: url_for() 가 자동으로 접두어를 붙이고, Socket.IO 미들웨어도 정상 동작.
# SocketIO(app) 이후에 감싸야 PrefixMiddleware 가 최외곽이 되어 /msg/socket.io 도 처리됨.
if BASE_PATH:
    class _PrefixMiddleware:
        def __init__(self, wsgi_app, prefix):
            self.wsgi_app = wsgi_app
            self.prefix = prefix

        def __call__(self, environ, start_response):
            path = environ.get("PATH_INFO", "")
            # 헬스체크는 접두어와 무관하게 항상 통과 (docker/healthcheck·배포 스크립트용)
            if path == "/healthz":
                return self.wsgi_app(environ, start_response)
            if path == self.prefix or path.startswith(self.prefix + "/"):
                environ["SCRIPT_NAME"] = environ.get("SCRIPT_NAME", "") + self.prefix
                environ["PATH_INFO"] = path[len(self.prefix):] or "/"
                return self.wsgi_app(environ, start_response)
            # BASE_PATH 외 경로(루트 포함)는 404 — 메신저는 /msg 하위에서만 서비스
            start_response("404 Not Found", [("Content-Type", "text/plain; charset=utf-8")])
            return ["Not Found".encode("utf-8")]

    app.wsgi_app = _PrefixMiddleware(app.wsgi_app, BASE_PATH)


@app.errorhandler(500)
def handle_500(err):
    """500 에러 자동 로깅 + 관리자 socketio 알림 (대표 지시 2026-05-19)."""
    import traceback as _tb
    tb_str = _tb.format_exc()
    try:
        # 파일 로그 — 운영 시점 디버깅 가능하도록 영속화
        log_path = os.path.join(APP_DIR, "data", "error.log")
        os.makedirs(os.path.dirname(log_path), exist_ok=True)
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"\n\n=== [{datetime.now(timezone.utc).isoformat()}] {request.method} {request.path} ===\n")
            try:
                uid = session.get("user_id")
                f.write(f"user_id={uid}, ip={request.remote_addr}, ua={request.user_agent.string[:200]}\n")
            except Exception:
                pass
            f.write(tb_str)
    except Exception:
        pass
    # 관리자(ceo) 들에게 실시간 알림 (socketio) — 최초 100자만
    try:
        short = tb_str.split("\n")[-2] if len(tb_str.split("\n")) >= 2 else str(err)
        socketio.emit("admin_error_alert", {
            "path": request.path,
            "method": request.method,
            "error": short[:200],
            "timestamp": datetime.now(timezone.utc).isoformat(),
        })
    except Exception:
        pass
    return jsonify({"error": "서버 내부 오류가 발생했습니다. 잠시 후 다시 시도해주세요."}), 500


@app.errorhandler(413)
def handle_413(err):
    """파일이 너무 큼 (Flask MAX_CONTENT_LENGTH 초과)."""
    return jsonify({"error": f"파일이 너무 큽니다 (최대 {MAX_UPLOAD_MB}MB)"}), 413


@app.after_request
def no_cache_html_js(resp):
    """동적 응답(HTML/JS/CSS/JSON)에 캐시 방지 헤더. 운영에서도 코드 수정 즉시 반영을 위해 유지.
    + 운영 환경에서는 보안 헤더 추가."""
    if resp.mimetype in ("text/html", "application/javascript", "text/css", "application/json"):
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
    if IS_PRODUCTION:
        resp.headers.setdefault("X-Content-Type-Options", "nosniff")
        resp.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
        resp.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
        resp.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
    return resp


# ---------- DB ----------
def get_db():
    db = getattr(g, "_db", None)
    if db is None:
        os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
        db = g._db = sqlite3.connect(DB_PATH, timeout=20.0)
        db.row_factory = sqlite3.Row
        db.execute("PRAGMA foreign_keys = ON")
        # 동시성 향상 — WAL 모드: 읽기·쓰기 동시 가능, 쓰기 락 범위 축소 (2026-05-20)
        # 75명 동시 사용 시 DB 락 병목 해소.
        db.execute("PRAGMA journal_mode=WAL")
        # 안전성-성능 trade-off: NORMAL — 전원 차단 시 마지막 트랜잭션 손실 가능하나 OS 크래시엔 안전
        db.execute("PRAGMA synchronous=NORMAL")
        # 락 대기 — 동시 INSERT 충돌 시 최대 20초 재시도 (기본 5초 → 20초)
        db.execute("PRAGMA busy_timeout=20000")
        # 캐시 — 페이지 캐시 64MB (기본 2MB → 64MB)
        db.execute("PRAGMA cache_size=-65536")
        # WAL 자동 체크포인트 — 1000 페이지 (기본). 너무 늦추면 .wal 파일 커짐.
    return db


@app.teardown_appcontext
def close_db(_exc):
    db = getattr(g, "_db", None)
    if db is not None:
        db.close()


# ── 사내 메신저 이용·보안 동의서 (대표 지시 2026-06-03) ──
from consent_doc import CONSENT_VERSION, CONSENT_DOC, CONSENT_TEST_MODE, CONSENT_TEST_USERNAMES
CONSENT_INTERVAL_DAYS = 100   # 마지막 동의로부터 N일 경과 시 재동의 강제 / 게스트는 제외


def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH, timeout=20.0)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    # WAL 모드 한 번만 적용 (DB 파일 메타에 영구 저장됨) — 2026-05-20
    try:
        cur.execute("PRAGMA journal_mode=WAL")
        cur.execute("PRAGMA synchronous=NORMAL")
        cur.execute("PRAGMA busy_timeout=20000")
    except Exception as _e:
        print(f"[init_db] PRAGMA 적용 실패 (무시): {_e}", flush=True)
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        display_name TEXT NOT NULL,
        role TEXT DEFAULT 'staff',
        avatar_color TEXT DEFAULT '#3b82f6',
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS rooms (
        id INTEGER PRIMARY KEY,
        name TEXT,
        type TEXT NOT NULL,        -- direct | group | channel
        created_by INTEGER,
        created_at TEXT NOT NULL,
        FOREIGN KEY (created_by) REFERENCES users(id)
    );
    CREATE TABLE IF NOT EXISTS room_members (
        room_id INTEGER NOT NULL,
        user_id INTEGER NOT NULL,
        joined_at TEXT NOT NULL,
        last_read_message_id INTEGER DEFAULT 0,
        PRIMARY KEY (room_id, user_id),
        FOREIGN KEY (room_id) REFERENCES rooms(id) ON DELETE CASCADE,
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS messages (
        id INTEGER PRIMARY KEY,
        room_id INTEGER NOT NULL,
        user_id INTEGER NOT NULL,
        content TEXT NOT NULL,
        kind TEXT DEFAULT 'text',  -- text | image | file | system
        created_at TEXT NOT NULL,
        FOREIGN KEY (room_id) REFERENCES rooms(id) ON DELETE CASCADE,
        FOREIGN KEY (user_id) REFERENCES users(id)
    );
    CREATE INDEX IF NOT EXISTS idx_messages_room_created ON messages(room_id, created_at);

    -- 📢 방별 공지사항: 최신 1건 상단 고정 + 펼치면 과거 공지 전부 + 끝난 공지 삭제.
    -- 삭제는 active=0(데이터 보존, 화면에서만 제거). 등록은 방 참여자 누구나. (대표 지시 2026-05-29)
    CREATE TABLE IF NOT EXISTS room_notices (
        id INTEGER PRIMARY KEY,
        room_id INTEGER NOT NULL,
        content TEXT NOT NULL,
        created_by INTEGER,
        created_at TEXT NOT NULL,
        active INTEGER NOT NULL DEFAULT 1,
        FOREIGN KEY (room_id) REFERENCES rooms(id) ON DELETE CASCADE,
        FOREIGN KEY (created_by) REFERENCES users(id)
    );
    CREATE INDEX IF NOT EXISTS idx_room_notices_room ON room_notices(room_id, active, id);

    -- 멘션함: @이름 으로 호출당한 사람별 기록. read_at IS NULL = 안 읽음. (대표 지시 2026-05-22)
    CREATE TABLE IF NOT EXISTS mentions (
        id INTEGER PRIMARY KEY,
        message_id INTEGER NOT NULL,
        room_id INTEGER NOT NULL,
        mentioned_user_id INTEGER NOT NULL,
        sender_user_id INTEGER NOT NULL,
        created_at TEXT NOT NULL,
        read_at TEXT,                       -- NULL = 안 읽음
        FOREIGN KEY (message_id) REFERENCES messages(id) ON DELETE CASCADE,
        FOREIGN KEY (room_id) REFERENCES rooms(id) ON DELETE CASCADE,
        FOREIGN KEY (mentioned_user_id) REFERENCES users(id) ON DELETE CASCADE,
        FOREIGN KEY (sender_user_id) REFERENCES users(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_mentions_user_unread ON mentions(mentioned_user_id, read_at);
    CREATE UNIQUE INDEX IF NOT EXISTS idx_mentions_msg_user ON mentions(message_id, mentioned_user_id);

    -- 프로젝트(=프로젝트/품목): 일반 메신저의 '방'을 자동 정리 가능한 단위로 승격
    CREATE TABLE IF NOT EXISTS items (
        id INTEGER PRIMARY KEY,
        room_id INTEGER UNIQUE NOT NULL,
        code TEXT,                          -- 관리번호 e.g. 003M2501
        name TEXT NOT NULL,                 -- 프로젝트명
        customer TEXT,                      -- 고객사 e.g. 삼성전자
        status TEXT DEFAULT 'active',       -- active | hold | done | cancelled
        due_date TEXT,                      -- 납기 (ISO date)
        description TEXT,
        created_by INTEGER,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        FOREIGN KEY (room_id) REFERENCES rooms(id) ON DELETE CASCADE,
        FOREIGN KEY (created_by) REFERENCES users(id)
    );
    CREATE INDEX IF NOT EXISTS idx_items_status ON items(status);
    CREATE INDEX IF NOT EXISTS idx_items_customer ON items(customer);

    -- 요청(티켓): 메시지 → 추적 가능한 작업으로 승격
    CREATE TABLE IF NOT EXISTS requests (
        id INTEGER PRIMARY KEY,
        room_id INTEGER NOT NULL,
        message_id INTEGER,
        title TEXT NOT NULL,
        description TEXT,
        requested_by INTEGER NOT NULL,
        assigned_to INTEGER,
        due_date TEXT,
        status TEXT DEFAULT 'open',     -- open | in_progress | done | cancelled
        priority TEXT DEFAULT 'normal',  -- low | normal | high
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        closed_at TEXT,
        FOREIGN KEY (room_id) REFERENCES rooms(id) ON DELETE CASCADE,
        FOREIGN KEY (message_id) REFERENCES messages(id) ON DELETE SET NULL,
        FOREIGN KEY (requested_by) REFERENCES users(id),
        FOREIGN KEY (assigned_to) REFERENCES users(id)
    );
    CREATE INDEX IF NOT EXISTS idx_requests_room_status ON requests(room_id, status);
    CREATE INDEX IF NOT EXISTS idx_requests_assigned_status ON requests(assigned_to, status);
    CREATE INDEX IF NOT EXISTS idx_requests_due ON requests(due_date);

    -- 🐞 버그 신고 — 직원이 메신저 문제를 신고. 전 직원 공유 '버그 신고' 채널에 글로 올라가고,
    --   여기엔 구조화 데이터(상태·기술정보)를 보관 → 유지보수관리자(김정락)만 관리 뷰에서 열람.
    --   (대표 지시 2026-06-03)
    CREATE TABLE IF NOT EXISTS bug_reports (
        id INTEGER PRIMARY KEY,
        reporter_user_id INTEGER,
        title TEXT,
        body TEXT,
        status TEXT NOT NULL DEFAULT 'new',   -- new(접수) | triaged(확인중) | fixed(수정완료) | deployed(배포완료) | wontfix(보류)
        context_json TEXT,                    -- {screen,room,device,browser,os,app_version,ui_lang,url,...}
        room_id INTEGER,                      -- 버그 신고 채널 room_id
        message_id INTEGER,                   -- 채널에 올라간 메시지 id
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        FOREIGN KEY (reporter_user_id) REFERENCES users(id) ON DELETE SET NULL,
        FOREIGN KEY (room_id) REFERENCES rooms(id) ON DELETE SET NULL,
        FOREIGN KEY (message_id) REFERENCES messages(id) ON DELETE SET NULL
    );
    CREATE INDEX IF NOT EXISTS idx_bug_reports_status ON bug_reports(status, id);

    -- 메시지 반응 (👍 ✅ ❤ 등) — "네 알겠습니다" 노이즈 감소용
    CREATE TABLE IF NOT EXISTS message_reactions (
        id INTEGER PRIMARY KEY,
        message_id INTEGER NOT NULL,
        user_id INTEGER NOT NULL,
        emoji TEXT NOT NULL,
        created_at TEXT NOT NULL,
        UNIQUE (message_id, user_id, emoji),
        FOREIGN KEY (message_id) REFERENCES messages(id) ON DELETE CASCADE,
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_reactions_msg ON message_reactions(message_id);

    -- 메시지 전달확인 (acknowledgment) — '읽음'을 넘어선 명시적 확인
    -- "내가 봤고 처리하겠습니다" 의지 표시
    CREATE TABLE IF NOT EXISTS message_acks (
        id INTEGER PRIMARY KEY,
        message_id INTEGER NOT NULL,
        user_id INTEGER NOT NULL,
        ack_type TEXT DEFAULT 'ok',     -- ok | doing | done | reject
        comment TEXT,                    -- 선택: "오늘 5시까지 처리하겠음" 같은 메모
        created_at TEXT NOT NULL,
        UNIQUE (message_id, user_id, ack_type),
        FOREIGN KEY (message_id) REFERENCES messages(id) ON DELETE CASCADE,
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_acks_msg ON message_acks(message_id);
    CREATE INDEX IF NOT EXISTS idx_acks_user ON message_acks(user_id);

    -- 메시지 별표 (중요 결정 마킹) — 사용자별
    CREATE TABLE IF NOT EXISTS message_stars (
        message_id INTEGER NOT NULL,
        user_id INTEGER NOT NULL,
        starred_at TEXT NOT NULL,
        PRIMARY KEY (message_id, user_id),
        FOREIGN KEY (message_id) REFERENCES messages(id) ON DELETE CASCADE,
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_stars_user ON message_stars(user_id);

    -- 첨부 파일 버전 체인 (같은 도면 v1/v2/v3)
    -- parent_message_id 가 가장 최초 버전. version_no 1=v1, 2=v2 ...
    CREATE TABLE IF NOT EXISTS attachment_versions (
        message_id INTEGER PRIMARY KEY,         -- 이 attachment 메시지
        parent_message_id INTEGER NOT NULL,     -- 첫 버전의 message_id (자기 자신이면 자기)
        version_no INTEGER NOT NULL DEFAULT 1,
        room_id INTEGER NOT NULL,
        FOREIGN KEY (message_id) REFERENCES messages(id) ON DELETE CASCADE,
        FOREIGN KEY (parent_message_id) REFERENCES messages(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_attver_parent ON attachment_versions(parent_message_id);
    CREATE INDEX IF NOT EXISTS idx_attver_room ON attachment_versions(room_id);

    -- 메시지 번역 캐시 — 같은 메시지를 두 번 번역하지 않음 (Claude API 비용 절약)
    CREATE TABLE IF NOT EXISTS message_translations (
        id INTEGER PRIMARY KEY,
        message_id INTEGER NOT NULL,
        target_lang TEXT NOT NULL,         -- 'ko' | 'vi' | 'en'
        source_lang TEXT,                  -- 자동 감지 결과
        translated_text TEXT NOT NULL,
        model TEXT,                        -- 'claude-haiku-4-5' 등
        input_tokens INTEGER DEFAULT 0,
        output_tokens INTEGER DEFAULT 0,
        cost_usd REAL DEFAULT 0,
        created_at TEXT NOT NULL,
        UNIQUE (message_id, target_lang),
        FOREIGN KEY (message_id) REFERENCES messages(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_translations_msg ON message_translations(message_id);
    CREATE INDEX IF NOT EXISTS idx_translations_created ON message_translations(created_at);

    -- 사용자 상태 (자리비움·회의중·외근·방해금지·온라인·오프라인)
    -- user_id PRIMARY KEY = 사용자 1명당 1개 행 (UPSERT 패턴).
    CREATE TABLE IF NOT EXISTS user_statuses (
        user_id INTEGER PRIMARY KEY,
        status TEXT NOT NULL DEFAULT 'online',  -- online | away | busy | meeting | external | dnd | offline
        custom_text TEXT,                        -- 사용자 정의 ("HAIST WORKS 운영 중")
        emoji TEXT,                              -- 상태 이모지 (선택)
        until_at TEXT,                           -- 이 시각까지 유지 (NULL=수동 변경까지 영구)
        auto_set INTEGER NOT NULL DEFAULT 0,     -- 1=캘린더 자동 설정, 0=수동
        updated_at TEXT NOT NULL,
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_user_statuses_until ON user_statuses(until_at);

    -- 사용자 캘린더 일정 (간단 — 외부 iCal/Google Calendar 연동은 후속)
    -- 시작 시각에 자동 "회의 중" 전환, 종료 시각에 "온라인" 복귀.
    CREATE TABLE IF NOT EXISTS user_calendar_events (
        id INTEGER PRIMARY KEY,
        user_id INTEGER NOT NULL,
        title TEXT NOT NULL,
        start_at TEXT NOT NULL,    -- ISO datetime
        end_at TEXT NOT NULL,
        kind TEXT DEFAULT 'meeting',  -- meeting | external | busy
        applied INTEGER NOT NULL DEFAULT 0,  -- 0=대기, 1=시작 적용됨, 2=종료 적용됨
        created_at TEXT NOT NULL,
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_cal_user_start ON user_calendar_events(user_id, start_at);
    CREATE INDEX IF NOT EXISTS idx_cal_applied ON user_calendar_events(applied, start_at, end_at);

    -- 프로젝트(프로젝트 방) 이력 스냅샷 — HAIST WORKS 연동 대비.
    -- 하루 1회 자동 + 수동 즉시 갱신. 각 스냅샷은 마지막 스냅샷 이후 새 메시지를 요약.
    CREATE TABLE IF NOT EXISTS project_history (
        id INTEGER PRIMARY KEY,
        room_id INTEGER NOT NULL,
        period_start TEXT,                  -- 이 스냅샷이 다룬 메시지 첫 시각
        period_end TEXT NOT NULL,           -- 마지막 시각
        first_message_id INTEGER,           -- 다룬 범위 (다음 자동 생성의 기준)
        last_message_id INTEGER NOT NULL,
        summary_text TEXT NOT NULL,         -- AI 요약 본문 (한국어)
        message_count INTEGER NOT NULL DEFAULT 0,
        attachment_count INTEGER NOT NULL DEFAULT 0,
        attachments_json TEXT,              -- [{name,size,mime,url,sender,sent_at}, ...]
        model TEXT,
        input_tokens INTEGER DEFAULT 0,
        output_tokens INTEGER DEFAULT 0,
        cost_usd REAL DEFAULT 0,
        created_by INTEGER,                 -- NULL=자동, user_id=수동
        created_at TEXT NOT NULL,
        synced_to_hw INTEGER NOT NULL DEFAULT 0,  -- HAIST WORKS 전송 여부 (이후 사용)
        synced_at TEXT,
        FOREIGN KEY (room_id) REFERENCES rooms(id) ON DELETE CASCADE,
        FOREIGN KEY (created_by) REFERENCES users(id) ON DELETE SET NULL
    );
    CREATE INDEX IF NOT EXISTS idx_ph_room_created ON project_history(room_id, created_at);
    CREATE INDEX IF NOT EXISTS idx_ph_synced ON project_history(synced_to_hw, room_id);

    -- AI 요약 캐시
    -- scope_type 으로 무엇을 요약했는지 구분: 'channel_recent'(방의 최근 N개) | 'thread'(스레드)
    -- scope_key 는 그 식별자: 'room:{id}:last:{N}' 또는 'thread:{parent_msg_id}'
    -- last_message_id 는 캐시 무효화 기준 — 새 메시지가 들어왔으면 다시 생성
    CREATE TABLE IF NOT EXISTS ai_summaries (
        id INTEGER PRIMARY KEY,
        scope_type TEXT NOT NULL,
        scope_key TEXT NOT NULL,
        room_id INTEGER,
        last_message_id INTEGER NOT NULL,
        summary_text TEXT NOT NULL,
        model TEXT,
        input_tokens INTEGER DEFAULT 0,
        output_tokens INTEGER DEFAULT 0,
        cost_usd REAL DEFAULT 0,
        created_by INTEGER,
        created_at TEXT NOT NULL,
        UNIQUE (scope_type, scope_key, last_message_id),
        FOREIGN KEY (room_id) REFERENCES rooms(id) ON DELETE CASCADE,
        FOREIGN KEY (created_by) REFERENCES users(id) ON DELETE SET NULL
    );
    CREATE INDEX IF NOT EXISTS idx_aisum_scope ON ai_summaries(scope_type, scope_key);
    CREATE INDEX IF NOT EXISTS idx_aisum_created ON ai_summaries(created_at);

    -- Web Push 구독 (한 사용자가 여러 디바이스 가능)
    CREATE TABLE IF NOT EXISTS push_subscriptions (
        id INTEGER PRIMARY KEY,
        user_id INTEGER NOT NULL,
        endpoint TEXT UNIQUE NOT NULL,
        p256dh TEXT NOT NULL,
        auth TEXT NOT NULL,
        user_agent TEXT,
        created_at TEXT NOT NULL,
        last_used TEXT,
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_push_user ON push_subscriptions(user_id);

    -- 동시 로그인 제한 (휴대폰 1대 + PC 1대) — (user_id, device_type) 당 활성 세션 1개만.
    -- 같은 종류 기기에서 새로 로그인하면 token 이 덮어써져 옛 기기 세션이 무효화됨.
    CREATE TABLE IF NOT EXISTS active_sessions (
        user_id INTEGER NOT NULL,
        device_type TEXT NOT NULL,   -- 'pc' | 'mobile'
        token TEXT NOT NULL,
        user_agent TEXT,
        ip TEXT,
        created_at TEXT NOT NULL,
        PRIMARY KEY (user_id, device_type),
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
    );
    -- 회사망(사무실) 공인 IP — 접속 IP 가 여기 있으면 '🏢 회사' 표시. (관리자가 등록)
    CREATE TABLE IF NOT EXISTS office_networks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ip TEXT NOT NULL UNIQUE,
        label TEXT,
        created_at TEXT NOT NULL,
        created_by INTEGER
    );
    -- 앱 전역 설정 (key-value) — 예: away_minutes(자리비움 자동전환 분). (대표 지시 2026-05-24)
    CREATE TABLE IF NOT EXISTS app_settings (
        key   TEXT PRIMARY KEY,
        value TEXT
    );
    -- 고객사 게스트 초대 토큰 (대표 지시 2026-05-28)
    --   · 방장/PM/관리자가 발행 → 휴대폰 번호 + 만료 + 토큰
    --   · 고객사 담당자가 QR/링크로 접근 → 휴대폰 매칭 → 게스트 계정 생성·방 입장
    --   · 만료 전 다회 사용 가능 (재로그인)
    CREATE TABLE IF NOT EXISTS guest_invites (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        token TEXT NOT NULL UNIQUE,
        room_id INTEGER NOT NULL,
        invited_by_user_id INTEGER NOT NULL,
        guest_name TEXT NOT NULL,
        guest_company TEXT NOT NULL,
        guest_title TEXT,                -- 직책(선택). NULL/빈문자 허용. (대표 지시 2026-05-28)
        guest_phone TEXT NOT NULL,
        expires_at TEXT,                -- NULL = 무제한
        guest_user_id INTEGER,           -- 게스트가 입장한 후 자동 생성된 user id (NULL=미사용)
        first_used_at TEXT,
        last_used_at TEXT,
        revoked_at TEXT,                 -- NULL = 활성, 값 있음 = 회수됨
        created_at TEXT NOT NULL,
        FOREIGN KEY (room_id) REFERENCES rooms(id) ON DELETE CASCADE,
        FOREIGN KEY (invited_by_user_id) REFERENCES users(id),
        FOREIGN KEY (guest_user_id) REFERENCES users(id)
    );
    CREATE INDEX IF NOT EXISTS idx_guest_invites_token ON guest_invites(token);
    CREATE INDEX IF NOT EXISTS idx_guest_invites_room ON guest_invites(room_id);

    -- 사내 메신저 이용·보안 동의서 동의 이력 (대표 지시 2026-06-03)
    --   · 첫 로그인 + 약 100일마다 재동의 / 게스트(고객사) 제외
    --   · 누가·언제·어떤 버전에 동의했는지 감사 추적 + IP/UA 기록
    CREATE TABLE IF NOT EXISTS consent_agreements (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id    INTEGER NOT NULL,
        version    INTEGER NOT NULL,
        agreed_at  TEXT NOT NULL,
        ip         TEXT,
        user_agent TEXT,
        FOREIGN KEY (user_id) REFERENCES users(id)
    );
    CREATE INDEX IF NOT EXISTS idx_consent_user ON consent_agreements(user_id, version);
    """)
    conn.commit()
    # 회사망 IP 메모리 캐시 초기 로딩 (이후 관리자 변경 시 _load_office_ips 로 갱신)
    global _office_ips
    try:
        _office_ips = set((r["ip"] or "").strip()
                          for r in conn.execute("SELECT ip FROM office_networks").fetchall() if r["ip"])
    except Exception:
        _office_ips = set()

    # 자리비움 자동전환 시간(분) 메모리 캐시 초기 로딩 (관리자 변경 시 _set_away_minutes_cache 로 갱신)
    global _away_minutes_cache
    try:
        _r = conn.execute("SELECT value FROM app_settings WHERE key='away_minutes'").fetchone()
        if _r and str(_r["value"]).isdigit() and int(_r["value"]) in AWAY_MINUTES_ALLOWED:
            _away_minutes_cache = int(_r["value"])
    except Exception:
        pass

    # ---- 컬럼 마이그레이션 ----
    existing_msg_cols = {row["name"] for row in cur.execute("PRAGMA table_info(messages)").fetchall()}
    for col, ddl in [
        ("file_path", "ALTER TABLE messages ADD COLUMN file_path TEXT"),
        ("file_name", "ALTER TABLE messages ADD COLUMN file_name TEXT"),
        ("file_size", "ALTER TABLE messages ADD COLUMN file_size INTEGER"),
        ("file_mime", "ALTER TABLE messages ADD COLUMN file_mime TEXT"),
        # 스레드 — parent_message_id 가 채워지면 스레드 답글.
        # NULL=일반 메시지 (메인 채널에 표시). NOT NULL=답글 (스레드 패널에만 표시).
        ("parent_message_id", "ALTER TABLE messages ADD COLUMN parent_message_id INTEGER"),
        # 인용 답장(Quote Reply) — 본 채널에 답글 + 원본 미니 카드 표시 (스레드와 별개)
        ("quoted_message_id", "ALTER TABLE messages ADD COLUMN quoted_message_id INTEGER"),
        # 전달(Forward) 출처 — 메타데이터 보존
        # 원본 메시지 ID. 원본이 삭제돼도 아래 forwarded_* 캐시로 복원 가능
        ("forwarded_from_message_id", "ALTER TABLE messages ADD COLUMN forwarded_from_message_id INTEGER"),
        ("forwarded_from_user_id", "ALTER TABLE messages ADD COLUMN forwarded_from_user_id INTEGER"),
        # 원본 작성자명·방명·시각 캐시 (원본 삭제 후에도 표시되도록)
        ("forwarded_from_name", "ALTER TABLE messages ADD COLUMN forwarded_from_name TEXT"),
        ("forwarded_from_room_name", "ALTER TABLE messages ADD COLUMN forwarded_from_room_name TEXT"),
        ("forwarded_from_created_at", "ALTER TABLE messages ADD COLUMN forwarded_from_created_at TEXT"),
        # 귓속말 — 특정 한 사용자에게만 보이는 메시지 (sender + recipient 만). NULL=공개.
        ("whisper_to_user_id", "ALTER TABLE messages ADD COLUMN whisper_to_user_id INTEGER"),
        # 앨범 묶음 — 사진 N장을 1개 그리드 메시지로 묶을 때 부여하는 UUID. NULL=단독.
        ("album_id", "ALTER TABLE messages ADD COLUMN album_id TEXT"),
        # 편집 이력 — 마지막 수정 시각 (NULL=한 번도 편집 안 함) (대표 지시 2026-05-19)
        ("edited_at", "ALTER TABLE messages ADD COLUMN edited_at TEXT"),
        # 스레드 보관 연장 — 부모 메시지에만 사용. NULL=기본 30일 규칙, ISO 시각=그 시각까지 🗑 숨김. (대표 지시 2026-05-28)
        ("archive_extended_until", "ALTER TABLE messages ADD COLUMN archive_extended_until TEXT"),
        # 스레드 숨김 — 부모 메시지에만 사용. 1=스레드 목록에서만 숨김(내용·답글은 그대로 보존). (대표 지시 2026-05-29)
        ("thread_hidden", "ALTER TABLE messages ADD COLUMN thread_hidden INTEGER DEFAULT 0"),
        # 여러 방 동시 공유 — 이 메시지를 한 번에 몇 개 방에 같이 보냈는지(2+ 면 '여러 방 공유' 배지). NULL/0/1=일반 (대표 지시 2026-06-04)
        ("share_count", "ALTER TABLE messages ADD COLUMN share_count INTEGER DEFAULT 0"),
        # 여러 방 동시 공유 — 어느 방들에 함께 보냈는지(방 ID JSON 배열). 글쓴이만 '공유한 방 보기'로 확인 (대표 지시 2026-06-04)
        ("share_room_ids", "ALTER TABLE messages ADD COLUMN share_room_ids TEXT"),
        # 여러 방 동시 공유 — 글을 '작성한 방'(현재 방) ID. 공유한 방 목록 팝업에서 '✏️ 작성한 방' 표시용 (대표 지시 2026-06-04)
        ("share_origin_room_id", "ALTER TABLE messages ADD COLUMN share_origin_room_id INTEGER"),
        # 묶음 메시지 — 글+그림을 원본 순서대로 한 말풍선에 담는 파트 배열(JSON). NULL=일반. kind='multipart'. (대표 지시 2026-06-06)
        ("parts", "ALTER TABLE messages ADD COLUMN parts TEXT"),
    ]:
        if col not in existing_msg_cols:
            cur.execute(ddl)
    # 스레드 답글 조회 인덱스
    cur.execute("CREATE INDEX IF NOT EXISTS idx_messages_parent ON messages(parent_message_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_messages_quoted ON messages(quoted_message_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_messages_whisper ON messages(room_id, whisper_to_user_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_messages_album ON messages(album_id)")

    # AI 번역 호출자 추적 — 사용자별 사용량 통계용 (대표 지시 2026-05-27)
    existing_translation_cols = {row["name"] for row in cur.execute("PRAGMA table_info(message_translations)").fetchall()}
    if "requested_by_user_id" not in existing_translation_cols:
        cur.execute("ALTER TABLE message_translations ADD COLUMN requested_by_user_id INTEGER")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_translations_requested_by ON message_translations(requested_by_user_id)")

    existing_item_cols = {row["name"] for row in cur.execute("PRAGMA table_info(items)").fetchall()}
    if "keep_forever" not in existing_item_cols:
        cur.execute("ALTER TABLE items ADD COLUMN keep_forever INTEGER DEFAULT 0")

    # 사용자 활성/비활성 (퇴사자 차단)
    existing_user_cols = {row["name"] for row in cur.execute("PRAGMA table_info(users)").fetchall()}
    if "active" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN active INTEGER NOT NULL DEFAULT 1")
    # 직급·부서 — 사이드바 "사용자" 탭에서 표시
    if "title" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN title TEXT")
    if "department" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN department TEXT")
    # 회사 이메일·전화번호 — 사내 디렉터리
    if "email" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN email TEXT")
    if "phone" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN phone TEXT")
    # 첫 로그인 시 비밀번호 변경 강제 (신규 등록자는 1, 기존 사용자는 0)
    if "must_change_password" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN must_change_password INTEGER NOT NULL DEFAULT 0")
    # 사번 — 회사 내부 직원 식별 번호 (대표 지시 2026-05-19)
    if "employee_no" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN employee_no TEXT")
    # 아바타 사진 URL — 본인 사진 업로드용. NULL 이면 이름 첫 글자 + 배경색 표시 (대표 지시 2026-05-19)
    if "avatar_url" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN avatar_url TEXT")
    if "self_avatar" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN self_avatar TEXT")   # 본인 개인 아바타(직원 셀프). avatar_url=인사카드 사진(관리자) (대표 지시 2026-06-03)
    # 화면 표시 언어 — 'ko'|'vi'|'en'. NULL 이면 법인 기준 자동(베트남=vi, 본사=ko). (대표 지시 2026-05-25, 베트남 법인 사용)
    if "ui_lang" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN ui_lang TEXT")
    # 화면 테마 — 'light'|'dark'|'sage'|'cream'|'sky'. NULL/미설정 = light. (대표 지시 2026-05-28)
    if "ui_theme" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN ui_theme TEXT")
    # AI 요약·프로젝트 이력 사용 권한 — 0/1. 기본 0 (불허).
    # 기존 방장/PM/관리자는 자동 허용 + 이 컬럼이 1 인 직원도 허용. (대표 지시 2026-05-28)
    if "ai_summary_allowed" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN ai_summary_allowed INTEGER DEFAULT 0")
    # 채널 생성 권한 — 직급과 무관하게 관리자가 직원별로 추가 허용. 0/1, 기본 0.
    # (직급=대표·임원·팀장·법인장 또는 ceo 는 이 값과 무관하게 항상 허용) (대표 지시 2026-05-29)
    if "channel_create_allowed" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN channel_create_allowed INTEGER DEFAULT 0")
    # 고객사 게스트 플래그 — 1=외부 사용자(고객사), NULL/0=내부 직원. (대표 지시 2026-05-28)
    if "is_guest" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN is_guest INTEGER DEFAULT 0")
    # 게스트가 속한 방 — 게스트는 이 방 외 다른 방 못 봄. NULL=내부 사용자. (대표 지시 2026-05-28)
    if "guest_room_id" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN guest_room_id INTEGER")
    # 게스트의 회사명 — 표시용. NULL=내부.
    if "guest_company" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN guest_company TEXT")

    # guest_invites 테이블의 추가 컬럼 마이그레이션 (대표 지시 2026-05-28)
    try:
        existing_gi_cols = {row["name"] for row in cur.execute("PRAGMA table_info(guest_invites)").fetchall()}
        if existing_gi_cols and "guest_title" not in existing_gi_cols:
            cur.execute("ALTER TABLE guest_invites ADD COLUMN guest_title TEXT")
        # 고객사 정보 확장 — 부서·담당업무·이메일·참고사항 (대표 지시 2026-05-30)
        if existing_gi_cols:
            # guest_lang: 초대 시 선택한 고객사 언어(ko/vi/en/zh) — 초대메시지·입장페이지·대화방 기본 언어 (대표 지시 2026-05-31)
            for _gc in ("guest_department", "guest_duty", "guest_email", "guest_note", "guest_company_logo", "guest_lang"):
                if _gc not in existing_gi_cols:
                    cur.execute(f"ALTER TABLE guest_invites ADD COLUMN {_gc} TEXT")
    except Exception:
        pass
    # 영문 이름·직급·부서 — 베트남어/영어 화면에서 표시 (한글+영문 엑셀 업로드). NULL 이면 한글로 폴백. (대표 지시 2026-05-25)
    if "display_name_en" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN display_name_en TEXT")
    if "title_en" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN title_en TEXT")
    if "department_en" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN department_en TEXT")
    # 베트남어 원어 이름 — 베트남 직원만. 한국어 모드에서 '{vn} ({한국식발음})' 병기 표시. (대표 지시 2026-05-26)
    if "display_name_vn" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN display_name_vn TEXT")
    # 푸시 알림 내용 숨기기 — 잠금화면·푸시 모두 'KNK message'로만 (대표 지시 2026-05-26)
    if "push_hide_preview" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN push_hide_preview INTEGER NOT NULL DEFAULT 0")
    # 법인 구분 — 'KOR'(본사) | 'VN'(베트남). 사번 SSO 발주 2026-05-29. NULL=미지정(마이그레이션 시 채움).
    if "entity" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN entity TEXT")
    # 비밀번호 버전 — 비번 변경 시 +1. SSO JWT 의 pwv claim 과 비교해 변경 후 기존 토큰 무효화. (방안 A, 발주서 §8.3)
    if "password_version" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN password_version INTEGER NOT NULL DEFAULT 1")
    # HAIST WORKS 사용 권한 — 메신저 관리자가 부여. 1=WORKS 진입 허용. (대표 지시 2026-05-31, 단순화: 메신저 정문 + 'WORKS 열기' 버튼)
    if "works_access" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN works_access INTEGER NOT NULL DEFAULT 0")
    # 운영(유지보수) 권한 — 1이면 대표님이 보는 유지보수 화면(사용현황·버그관리·시스템설정)을 같이 봄.
    #   부여/회수는 기본 유지보수관리자(김정락 사번5)만 가능. (대표 지시 2026-06-03)
    if "ops_allowed" not in existing_user_cols:
        cur.execute("ALTER TABLE users ADD COLUMN ops_allowed INTEGER NOT NULL DEFAULT 0")
    # 사번(employee_no) 고유 인덱스 — 사번=로그인ID 정책. NULL 다수 허용(SQLite UNIQUE 는 NULL 구분).
    #   중복 사번이 있으면 인덱스 생성 실패 → 마이그레이션 스크립트에서 정리. 실패해도 부팅 계속.
    try:
        cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_users_employee_no ON users(employee_no)")
    except Exception as _e:
        print(f"[init_db] employee_no UNIQUE index 보류(중복 존재 가능): {_e}")

    # 사용자 상태 — 'dnd' (방해금지) 옵션 제거됨 → 기존 dnd 사용자는 online(회사)로 환원
    try:
        cur.execute("UPDATE user_statuses SET status='online' WHERE status='dnd'")
    except Exception:
        pass

    # 옛 부서명 '경영지원' (시드값) → KNK 표준 '관리팀' (대표 지시 2026-05-19).
    # 단, 직급이 '대표이사'인 사람은 부서 소속 없음 → 대상 제외.
    try:
        cur.execute(
            "UPDATE users SET department='관리팀' "
            "WHERE department='경영지원' AND (title IS NULL OR title != '대표이사')"
        )
    except Exception:
        pass

    # 마이그레이션 버전 추적 테이블 — 1회성 데이터 변환이 매 재시작마다 반복되지 않도록 (대표 지시 2026-05-19)
    try:
        cur.execute("CREATE TABLE IF NOT EXISTS app_migrations (name TEXT PRIMARY KEY, applied_at TEXT)")
    except Exception:
        pass

    def _migration_applied(name):
        try:
            return cur.execute("SELECT 1 FROM app_migrations WHERE name=?", (name,)).fetchone() is not None
        except Exception:
            return False

    def _mark_migration(name):
        try:
            cur.execute("INSERT OR REPLACE INTO app_migrations (name, applied_at) VALUES (?, ?)",
                        (name, datetime.now(timezone.utc).isoformat()))
        except Exception:
            pass

    # 대표이사 직급 → '총괄' (00) 부서 1회성 배정.
    # 이후 UI 에서 부서 수동 변경 시 덮어쓰지 않음 (마이그레이션 버전 마커).
    if not _migration_applied("v2_ceo_to_chonggwal"):
        try:
            cur.execute("UPDATE users SET department='총괄' WHERE title='대표이사' AND (department IS NULL OR department != '총괄')")
            _mark_migration("v2_ceo_to_chonggwal")
        except Exception:
            pass

    # 베트남법인 부서값 재포맷 — 1회성. 'VN기술팀' → 'VN12-01 기술팀' 등.
    # (이력 보존용 — 이후 v5_vn_format_swap 에서 다시 '12-VN01 기술팀' 형식으로 swap)
    if not _migration_applied("v2_vn_remap"):
        try:
            _vn_remap = {
                "VN기술팀":       "VN12-01 기술팀",
                "VN조립팀":       "VN12-02 조립팀",
                "VN전장팀":       "VN12-03 전장팀",
                "VN설계팀":       "VN12-04 설계팀",
                "VN소프트웨어팀": "VN12-05 소프트웨어팀",
                "VN가공팀":       "VN12-06 가공팀",
                "VN품질팀":       "VN12-07 품질팀",
                "VN구매팀":       "VN12-08 구매팀",
                "VN관리팀":       "VN12-09 관리팀",
            }
            for old_val, new_val in _vn_remap.items():
                cur.execute("UPDATE users SET department=? WHERE department=?", (new_val, old_val))
            _mark_migration("v2_vn_remap")
        except Exception:
            pass

    # 베트남법인 부서값 포맷 변경 (대표 지시 2026-05-20)
    #   'VN12-NN 부서명' → '12-VNNN 부서명'
    #   예: 'VN12-09 관리팀' → '12-VN09 관리팀'
    #   한국법인 코드 12 가 앞으로 나오고 VN 이 베트남 부서번호 앞에 붙음.
    if not _migration_applied("v5_vn_format_swap"):
        try:
            _vn_swap = {
                "VN12-01 기술팀":       "12-VN01 기술팀",
                "VN12-02 조립팀":       "12-VN02 조립팀",
                "VN12-03 전장팀":       "12-VN03 전장팀",
                "VN12-04 설계팀":       "12-VN04 설계팀",
                "VN12-05 소프트웨어팀": "12-VN05 소프트웨어팀",
                "VN12-06 가공팀":       "12-VN06 가공팀",
                "VN12-07 품질팀":       "12-VN07 품질팀",
                "VN12-08 구매팀":       "12-VN08 구매팀",
                "VN12-09 관리팀":       "12-VN09 관리팀",
            }
            _vs_changed = 0
            for old_val, new_val in _vn_swap.items():
                r = cur.execute("UPDATE users SET department=? WHERE department=?", (new_val, old_val))
                _vs_changed += (r.rowcount or 0)
            _mark_migration("v5_vn_format_swap")
            print(f"[migration] v5_vn_format_swap 1회 실행 — {_vs_changed}명 UPDATE", flush=True)
        except Exception as _e:
            print(f"[migration] v5_vn_format_swap 실패: {_e}", flush=True)

    # ── KNK 직원 명단 Excel 기준 부서값 일괄 정정 (대표 지시 2026-05-19) ──
    # 1회성 마이그레이션. 이후 UI 변경은 보존됨.
    # 강제 재실행: 버전 이름 ('v4_excel_dept_remap') 을 새 번호로 바꾸면 다시 1회 실행됨.
    # 모호 케이스: 김동현→설계팀(자동화), 김범수→설계팀(검사기), 윤경호→설계팀(자동화)
    # 임원(김정락·김동후·최홍광)→'총괄'(코드 00). 단, 대화창·요청 표시에서는 deptShortLabel 이 '총괄'을 숨김 (대표 지시 2026-05-22)
    if not _migration_applied("v4_excel_dept_remap"):
        EXCEL_DEPT_REMAP = {
            "anjiyeon@knknara.co.kr": "기술영업팀",
            "b8004@knknara.co.kr": "검사기팀",
            "bumsu.kim@knknara.co.kr": "설계팀(검사기)",
            "buy1@knknara.co.kr": "구매팀",
            "buy2@knknara.co.kr": "구매팀",
            "buy3@knknara.co.kr": "구매팀",
            "buy4@knknara.co.kr": "구매팀",
            "buy5@knknara.co.kr": "구매팀",
            "changho.choi@knknara.co.kr": "소프트웨어팀",
            "cheongmyung.lee@knknara.co.kr": "가공팀",
            "chkcsd@knknara.co.kr": "총괄",  # 최홍광 전무 — 00 총괄
            "chunghee.lee@knknara.co.kr": "소프트웨어팀",
            "chungil71@knknara.co.kr": "제조기술1팀",
            "daeseong.kang@knknara.co.kr": "제조기술2팀",
            "dhkimman@knknara.co.kr": "총괄",  # 김동후 — 00 총괄
            "donghyun.kim@knknara.co.kr": "설계팀(자동화)",
            "dongwook.kim@knknara.co.kr": "소프트웨어팀",
            "giwoon.kim@knknara.co.kr": "소프트웨어팀",
            "hanbin@knknara.co.kr": "검사기팀",
            "hanjung.lee@knknara.co.kr": "소프트웨어팀",
            "hojae.an@knknara.co.kr": "설계팀(검사기)",
            "hyun.lee@knknara.co.kr": "기술영업팀",
            "hyungjin.jeong@knknara.co.kr": "품질팀",
            "hyungryul.kim@knknara.co.kr": "전장설계팀",
            "hyunkyu.choi@knknara.co.kr": "설계팀(자동화)",
            "jaekyeom.na@knknara.co.kr": "라이프밸류팀",
            "jaeun.han@knknara.co.kr": "설계팀(자동화)",
            "jihoon.kim@knknara.co.kr": "검사기팀",
            "jihyeon.park@knknara.co.kr": "제조기술2팀",
            "jinho.joo@knknara.co.kr": "소프트웨어팀",
            "jinho.keum@knknara.co.kr": "제조기술1팀",
            "jks7434@knknara.co.kr": "검사기팀",
            "jongpil.hyeon@knknara.co.kr": "소프트웨어팀",
            "joochang.park@knknara.co.kr": "소프트웨어팀",
            "jun0130@knknara.co.kr": "설계팀(검사기)",
            "jungseok.hwang@knknara.co.kr": "소프트웨어팀",
            "jungwoo.lee@knknara.co.kr": "소프트웨어팀",
            "junyeob.shin@knknara.co.kr": "관리팀",
            "junyoung.ma@knknara.co.kr": "제조기술1팀",
            "khy9631@knknara.co.kr": "검사기팀",
            "kiseon.kim@knknara.co.kr": "라이프밸류팀",
            "kjr7749@knknara.co.kr": "품질팀",
            "knk1@knknara.co.kr": "관리팀",
            "knk2@knknara.co.kr": "관리팀",
            "knk3@knknara.co.kr": "관리팀",
            "knk4@knknara.co.kr": "관리팀",
            "kunghwan.oh@knknara.co.kr": "기술영업팀",
            "kwanghun.yoon@knknara.co.kr": "검사기팀",
            "kwangyoung.shin@knknara.co.kr": "설계팀(자동화)",
            "lhl2425@knknara.co.kr": "기술영업팀",
            "mingyu.jeong@knknara.co.kr": "설계팀(자동화)",
            "ngoclan.le@knknara.co.kr": "구매팀",
            "sales1@knknara.co.kr": "기술영업팀",
            "sangchon.lee@knknara.co.kr": "설계팀(자동화)",
            "sb8664@knknara.co.kr": "가공팀",
            "sejin.kim@knknara.co.kr": "가공팀",
            "seojoon.lee@knknara.co.kr": "검사기팀",
            "seungjin.bae@knknara.co.kr": "기술영업팀",
            "soft@knknara.co.kr": "개발혁신팀",
            "suhyeon.kim@knknara.co.kr": "개발혁신팀",
            "sungjin.jung@knknara.co.kr": "구매팀",
            "sungjin.lee@knknara.co.kr": "검사기팀",
            "sungki.bang@knknara.co.kr": "제조기술2팀",
            "sungsu.park@knknara.co.kr": "라이프밸류팀",
            "taehum.yeon@knknara.co.kr": "제조기술2팀",
            "taehyoung.kim@knknara.co.kr": "검사기팀",
            "taekhun.leem@knknara.co.kr": "제조기술2팀",
            "taewoo.lee@knknara.co.kr": "제조기술1팀",
            "top0015@knknara.co.kr": "총괄",  # 김정락 대표이사 — 00 총괄
            "wangxia1019@knknara.co.kr": "제조기술2팀",
            "yoon5468@knknara.co.kr": "가공팀",
            "yoon5959@knknara.co.kr": "설계팀(자동화)",
            "younghoon.na@knknara.co.kr": "제조기술2팀",
            "youngjun.lee2@knknara.co.kr": "소프트웨어팀",
            "yslee@knknara.co.kr": "12-VN09 관리팀",
        }
        try:
            _ed_changed = 0
            for _email, _dept in EXCEL_DEPT_REMAP.items():
                if _dept is None:
                    _r = cur.execute(
                        "UPDATE users SET department=NULL WHERE LOWER(username)=? AND department IS NOT NULL",
                        (_email.lower(),)
                    )
                else:
                    # 대표이사 제외 가드 제거 — 총괄 배정이 대표이사에게 적용되어야 함
                    _r = cur.execute(
                        "UPDATE users SET department=? "
                        "WHERE LOWER(username)=? "
                        "  AND COALESCE(department,'') != ?",
                        (_dept, _email.lower(), _dept)
                    )
                if _r.rowcount > 0:
                    _ed_changed += 1
            _mark_migration("v4_excel_dept_remap")
            print(f"[migration] v4_excel_dept_remap 1회 실행 — {_ed_changed}명 UPDATE", flush=True)
        except Exception as _e:
            print(f"[migration] Excel 부서 정정 실패: {_e}", flush=True)

    # '총괄' 부서 복원 (대표 지시 2026-05-22) — v5 에서 비웠으나 사용자 목록 정렬·그룹핑이
    # '총괄'(코드 00)을 필요로 함이 확인됨. 임원 3인 부서를 '총괄'로 되돌린다.
    # 표시 정책: 사용자 목록만 '00 총괄' 노출, 대화창·요청은 deptShortLabel 이 '총괄'을 숨김.
    if not _migration_applied("v6_restore_chonggwal"):
        try:
            _rc = 0
            for _email in ("top0015@knknara.co.kr", "dhkimman@knknara.co.kr", "chkcsd@knknara.co.kr"):
                _r = cur.execute(
                    "UPDATE users SET department='총괄' WHERE LOWER(username)=? AND COALESCE(department,'')=''",
                    (_email,)
                )
                _rc += _r.rowcount
            _mark_migration("v6_restore_chonggwal")
            print(f"[migration] v6_restore_chonggwal — {_rc}명 '총괄' 복원", flush=True)
        except Exception as _e:
            print(f"[migration] v6_restore_chonggwal 실패: {_e}", flush=True)

    # 멘션함 과거 소급 — 기존 텍스트 메시지의 '@이름' 을 스캔해 mentions 행 생성(전부 '안 읽음'). (대표 지시 2026-05-22)
    if not _migration_applied("v7_mentions_backfill"):
        try:
            _now_iso = datetime.now(timezone.utc).isoformat()
            _members_by_room = {}
            for _mrow in cur.execute("SELECT room_id, user_id FROM room_members").fetchall():
                _members_by_room.setdefault(_mrow["room_id"], []).append(_mrow["user_id"])
            _uinfo = {}
            for _urow in cur.execute("SELECT id, display_name, username FROM users").fetchall():
                _uinfo[_urow["id"]] = (_urow["display_name"], _urow["username"])
            _mcnt = 0
            for _msg in cur.execute(
                "SELECT id, room_id, user_id, content, created_at FROM messages "
                "WHERE content LIKE '%@%' "
                "  AND COALESCE(whisper_to_user_id,0)=0 "
                "  AND COALESCE(kind,'text')='text'"
            ).fetchall():
                _content = _msg["content"] or ""
                if "@" not in _content:
                    continue
                _sender = _msg["user_id"]
                for _uid_m in _members_by_room.get(_msg["room_id"], []):
                    if _uid_m == _sender:
                        continue
                    _info = _uinfo.get(_uid_m)
                    if not _info:
                        continue
                    # 옛 멘션(@아이디)·새 멘션(@이름) 모두 매칭
                    if _content_mentions_user(_content, _info[0], _info[1]):
                        try:
                            _c = cur.execute(
                                "INSERT OR IGNORE INTO mentions "
                                "(message_id, room_id, mentioned_user_id, sender_user_id, created_at) "
                                "VALUES (?,?,?,?,?)",
                                (_msg["id"], _msg["room_id"], _uid_m, _sender, _msg["created_at"] or _now_iso),
                            )
                            if _c.rowcount:
                                _mcnt += 1
                        except Exception:
                            pass
            _mark_migration("v7_mentions_backfill")
            print(f"[migration] v7_mentions_backfill — {_mcnt}건 과거 멘션 기록", flush=True)
        except Exception as _e:
            print(f"[migration] v7_mentions_backfill 실패: {_e}", flush=True)

    # 첫 사용자(대표) 자동 기본값 — 빈 값일 때만
    # 대표이사(시드 id=1)는 직급만 부여. 부서는 비워둠 — 대표이사는 조직 위에 있어 부서 소속 X (대표 지시 2026-05-19).
    cur.execute("UPDATE users SET title='대표이사' WHERE id=1 AND (title IS NULL OR title='')")

    # 방 이름 고정 플래그 (방장만 이름 변경 가능 vs 멤버 각자 별명 가능)
    existing_room_cols = {row["name"] for row in cur.execute("PRAGMA table_info(rooms)").fetchall()}
    if "name_locked" not in existing_room_cols:
        cur.execute("ALTER TABLE rooms ADD COLUMN name_locked INTEGER NOT NULL DEFAULT 0")

    # 방별 메시지 자동 삭제 일수 (WhatsApp 식: 1=24시간, 7=1주, 30=30일, 90=90일)
    # NULL=영구(글로벌 MESSAGE_RETENTION_MONTHS 만 적용). 방장이 설정 가능.
    if "retention_days" not in existing_room_cols:
        cur.execute("ALTER TABLE rooms ADD COLUMN retention_days INTEGER")

    # 초대 권한 정책 — 'all'(기본: 모든 멤버) | 'host_only'(방장·부방장만)
    if "invite_policy" not in existing_room_cols:
        cur.execute("ALTER TABLE rooms ADD COLUMN invite_policy TEXT NOT NULL DEFAULT 'all'")

    # 채널 소속 범위 (대표 지시 2026-05-20) — NULL(일반/사용자채널) | 'all'(KNK WORLD 전직원)
    # | 'hq'(본사) | 'vn'(베트남). 자동 생성·멤버 자동 동기화·나가기 금지 대상.
    if "channel_scope" not in existing_room_cols:
        cur.execute("ALTER TABLE rooms ADD COLUMN channel_scope TEXT")

    # 방/채널 아바타 이미지 URL — 관리자가 채널 아이콘에 사진 설정 (대표 지시 2026-05-20)
    if "avatar_url" not in existing_room_cols:
        cur.execute("ALTER TABLE rooms ADD COLUMN avatar_url TEXT")

    # 방 이름 다국어 — 베트남어·영어 캐시. 생성/이름변경 시 AI 번역 1회 호출해 저장. (대표 지시 2026-05-28)
    if "name_vi" not in existing_room_cols:
        cur.execute("ALTER TABLE rooms ADD COLUMN name_vi TEXT")
    if "name_en" not in existing_room_cols:
        cur.execute("ALTER TABLE rooms ADD COLUMN name_en TEXT")
    # 방 이름 자동 번역 — 한국어·중국어 칸 추가 (대표 지시 2026-06-05 재가동·다언어)
    if "name_ko" not in existing_room_cols:
        cur.execute("ALTER TABLE rooms ADD COLUMN name_ko TEXT")
    if "name_zh" not in existing_room_cols:
        cur.execute("ALTER TABLE rooms ADD COLUMN name_zh TEXT")

    # 1:1 방 식별 키 — 두 참여자 ID를 정렬해 "min:max" 형식으로 저장 (대표 지시 2026-05-21).
    # 상대가 나가도(room_members 삭제) 기존 방을 다시 찾아 대화 이어가기 위함.
    if "direct_key" not in existing_room_cols:
        cur.execute("ALTER TABLE rooms ADD COLUMN direct_key TEXT")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_rooms_direct_key ON rooms(direct_key)")

    # direct_key 백필 — 기존 1:1 방의 참여자 쌍을 멤버 + 메시지 작성자(나간 사람 포함)에서
    # 재구성해 키를 채운다. 1회성. (대표 지시 2026-05-21)
    if not _migration_applied("v3_direct_key_backfill"):
        try:
            for dr in cur.execute("SELECT id FROM rooms WHERE type='direct'").fetchall():
                rid = dr["id"]
                uids = set()
                for m in cur.execute("SELECT user_id FROM room_members WHERE room_id=?", (rid,)).fetchall():
                    if m["user_id"]:
                        uids.add(m["user_id"])
                # 나간 사람도 메시지(시스템 '나감' 포함) 작성자에서 복원
                for m in cur.execute("SELECT DISTINCT user_id FROM messages WHERE room_id=?", (rid,)).fetchall():
                    if m["user_id"]:
                        uids.add(m["user_id"])
                su = sorted(uids)
                if len(su) >= 2:
                    cur.execute("UPDATE rooms SET direct_key=? WHERE id=?", (f"{su[0]}:{su[1]}", rid))
            _mark_migration("v3_direct_key_backfill")
        except Exception:
            pass

    # created_by → host 자동 백필 — 1회성. 매번 부팅 시 실행되면
    # 사용자가 UI 에서 호스트를 강등한 경우 매번 복귀되는 버그 발생.
    if not _migration_applied("v2_room_host_backfill"):
        try:
            cur.execute("""
                UPDATE room_members
                   SET role = 'host'
                 WHERE role != 'host'
                   AND (room_id, user_id) IN (
                       SELECT id, created_by FROM rooms
                        WHERE created_by IS NOT NULL
                   )
            """)
            _mark_migration("v2_room_host_backfill")
        except Exception:
            pass

    # self 방 이름 '📝 메모' 통일 — 1회성 (옛 '📝 나에게 보내기' 갱신).
    # 매번 실행되면 사용자가 self 방 이름 바꿔도 매번 리셋되는 버그.
    if not _migration_applied("v2_self_room_rename"):
        try:
            cur.execute("UPDATE rooms SET name='📝 메모' WHERE type='self' AND name != '📝 메모'")
            _mark_migration("v2_self_room_rename")
        except Exception:
            pass

    # 방 멤버 역할 (host=방장, sub_host=부방장, member=일반)
    existing_rm_cols = {row["name"] for row in cur.execute("PRAGMA table_info(room_members)").fetchall()}
    if "role" not in existing_rm_cols:
        cur.execute("ALTER TABLE room_members ADD COLUMN role TEXT NOT NULL DEFAULT 'member'")
        # 기존 방의 created_by 사용자를 host 로 백필
        cur.execute("""
            UPDATE room_members SET role='host'
             WHERE (room_id, user_id) IN (
                SELECT id, created_by FROM rooms WHERE created_by IS NOT NULL
             )
        """)

    # 사용자별 방 정렬 — order_value(REAL, 사이 삽입 용이) + pinned(0/1)
    # NULL=자동 정렬, INTEGER/REAL=수동. pinned=1 인 방은 항상 상단 그룹.
    if "order_value" not in existing_rm_cols:
        cur.execute("ALTER TABLE room_members ADD COLUMN order_value REAL")
    if "pinned" not in existing_rm_cols:
        cur.execute("ALTER TABLE room_members ADD COLUMN pinned INTEGER NOT NULL DEFAULT 0")
    # 마지막 읽은 시각 — 읽음 명단의 '언제 읽었는지' 표시용 (대표 지시 2026-05-19)
    if "last_read_at" not in existing_rm_cols:
        cur.execute("ALTER TABLE room_members ADD COLUMN last_read_at TEXT")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_rm_user_order ON room_members(user_id, pinned, order_value)")

    # 방 별명 (멤버 각자 자기 화면에서만 보이는 이름)
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS room_aliases (
        user_id INTEGER NOT NULL,
        room_id INTEGER NOT NULL,
        alias TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        PRIMARY KEY (user_id, room_id),
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
        FOREIGN KEY (room_id) REFERENCES rooms(id) ON DELETE CASCADE
    );
    """)
    conn.commit()

    # 게스트(외부) '대화방 표현' — 보는 사람 각자 자기 화면에서만 보이는 게스트 별칭. (대표 지시 2026-05-30)
    #   직원은 다국어 이름/직급/부서가 등록돼 언어 전환 시 변환되지만, 게스트는 그게 없어
    #   보는 사람이 알아보기 쉽게 직접 이름표를 적게 함. 적은 사람 본인 화면에서만 적용.
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS guest_view_aliases (
        viewer_user_id INTEGER NOT NULL,
        guest_user_id INTEGER NOT NULL,
        room_id INTEGER,
        alias TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        PRIMARY KEY (viewer_user_id, guest_user_id)
    );
    """)
    conn.commit()

    # 공유 묶음 — 여러 방에 한 번에 보낼 때 자주 쓰는 방 조합을 개인이 저장 (대표 지시 2026-06-04)
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS share_bundles (
        id INTEGER PRIMARY KEY,
        user_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        room_ids TEXT NOT NULL,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        UNIQUE (user_id, name),
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
    );
    """)
    conn.commit()

    # ---- FTS5 가상 테이블 (전문 검색) ----
    cur.executescript("""
    CREATE VIRTUAL TABLE IF NOT EXISTS messages_fts USING fts5(
        content, content='messages', content_rowid='id', tokenize='unicode61'
    );
    CREATE TRIGGER IF NOT EXISTS messages_ai AFTER INSERT ON messages BEGIN
        INSERT INTO messages_fts(rowid, content) VALUES (new.id, new.content);
    END;
    CREATE TRIGGER IF NOT EXISTS messages_ad AFTER DELETE ON messages BEGIN
        INSERT INTO messages_fts(messages_fts, rowid, content) VALUES('delete', old.id, old.content);
    END;
    CREATE TRIGGER IF NOT EXISTS messages_au AFTER UPDATE ON messages BEGIN
        INSERT INTO messages_fts(messages_fts, rowid, content) VALUES('delete', old.id, old.content);
        INSERT INTO messages_fts(rowid, content) VALUES (new.id, new.content);
    END;
    """)
    # 기존 데이터 FTS 인덱스 1회 채우기
    cur.execute("SELECT COUNT(*) AS n FROM messages_fts")
    if cur.fetchone()["n"] == 0:
        cur.execute("INSERT INTO messages_fts(rowid, content) SELECT id, content FROM messages")
    conn.commit()

    now = datetime.now(timezone.utc).isoformat()

    # 시드 데이터 제거 (대표 지시 2026-05-26):
    #   옛 코드는 users·전체공지·프로젝트 4종(Watch Molding/WING PLATE/메탈치수/하나머티리얼)을 자동 주입했는데,
    #   cleanup 후에도 부팅마다 재주입되어 잔존이 생김 → 시드 자체를 영구 제거.
    #   신규 직원은 일괄 등록 양식 또는 단건 등록 다이얼로그로 추가.

    # 자동 채널(KNK WORLD/본사/베트남) 생성 + 전 직원 멤버십 동기화 + '전체공지' 삭제 (대표 지시 2026-05-20)
    try:
        _resync_auto_channels(conn)
    except Exception as e:
        print(f"[init_db] 자동채널 동기화 실패(무시): {e}", flush=True)

    # 고아 게스트(회수됐는데 방 멤버만 잔존) 자동 정리 (대표 지시 2026-05-30)
    try:
        _cleanup_revoked_guest_members(conn)
    except Exception as e:
        print(f"[init_db] 게스트 정리 실패(무시): {e}", flush=True)

    # 최고관리자(소유자) 자동 보장 — OWNER_USERNAME 계정이 있으면 ceo·활성 강제 (대표 지시 2026-05-21)
    try:
        conn.execute("UPDATE users SET role='ceo' WHERE LOWER(username)=? AND role!='ceo'", (OWNER_USERNAME,))
        conn.execute("UPDATE users SET active=1 WHERE LOWER(username)=? AND COALESCE(active,1)!=1", (OWNER_USERNAME,))
        conn.commit()
    except Exception as e:
        print(f"[init_db] 최고관리자 보장 실패(무시): {e}", flush=True)

    conn.close()


# ===== 자동 채널 (KNK WORLD / 본사 / 베트남) — 대표 지시 2026-05-20 =====
# 부서값 '02_VN/…' (공식 2026-05-26~) 또는 레거시 '12-VN…' = 베트남, 그 외(또는 미지정) = 본사.
# 3개 채널은 자동 생성·멤버 자동 동기화·나가기 금지(channel_scope 로 식별).
AUTO_CHANNELS = [
    ("all", "🌏 KNK WORLD"),   # 아시아 지구 (대표 지시 2026-05-21)
    ("hq",  "🇰🇷 본사채널"),   # 본사 앞 이모지 태극기로 (대표 지시 2026-05-21)
    ("vn",  "🇻🇳 베트남채널"),
]


def _user_is_vietnam(department):
    """부서값이 '12-VN' (레거시) 또는 '02_VN/' (공식 2026-05-26~) 으로 시작하면 베트남법인 소속.
    공식 부서 마스터 24개 체계 도입 후에는 '02_VN/' 가 정식. '12-VN' 은 이전 등록자 잔존분 호환용."""
    if not department:
        return False
    s = str(department).strip()
    return s.startswith("02_VN/") or s.startswith("12-VN")


# 화면 표시 언어 — 한국어/베트남어/영어 (대표 지시 2026-05-25, 베트남 법인 합류)
UI_LANGS = ("ko", "vi", "en", "zh")

def _user_ui_lang(user):
    """사용자의 화면 표시 언어 결정.
      · 본인이 명시 설정(ui_lang)했으면 그 값.
      · 미설정이면 법인 기준 자동 — 베트남 법인(02_VN/… 또는 레거시 12-VN…)=베트남어(vi), 그 외=한국어(ko).
    user 는 dict/sqlite Row (ui_lang·department 키 사용)."""
    if not user:
        return "ko"
    try:
        lang = (user["ui_lang"] if "ui_lang" in user.keys() else None) if hasattr(user, "keys") else user.get("ui_lang")
    except Exception:
        lang = None
    if lang in UI_LANGS:
        return lang
    try:
        dept = (user["department"] if "department" in user.keys() else None) if hasattr(user, "keys") else user.get("department")
    except Exception:
        dept = None
    return "vi" if _user_is_vietnam(dept) else "ko"


# 화면 테마 — 라이트/다크/세이지/크림/스카이 (대표 지시 2026-05-28)
UI_THEMES = ("light", "dark", "sage", "cream", "sky")

def _user_ui_theme(user):
    """사용자의 화면 테마 결정. 미설정/None/유효하지 않으면 'light'."""
    if not user:
        return "light"
    try:
        t = (user["ui_theme"] if "ui_theme" in user.keys() else None) if hasattr(user, "keys") else user.get("ui_theme")
    except Exception:
        t = None
    return t if t in UI_THEMES else "light"


def _desired_scopes_for(department, active, is_owner=False, is_guest=False):
    """이 사용자가 속해야 할 자동채널 scope 집합.
    최고관리자(소유자)는 본사·베트남 구분 없이 모든 채널 소속 (대표 지시 2026-05-21, 규칙 2).
    게스트(고객사 외부 사용자)는 어떤 자동 채널에도 속하지 않음 — 초대된 방만 (대표 지시 2026-05-29)."""
    if not active or is_guest:
        return set()
    if is_owner:
        return {"all", "hq", "vn"}
    return {"all", ("vn" if _user_is_vietnam(department) else "hq")}


def _auto_channel_ids(db):
    """scope -> room_id 매핑 (없으면 생성). + 레거시 '전체공지' 채널 제거."""
    now = datetime.now(timezone.utc).isoformat()
    # 레거시 '전체공지' 삭제 (대표 지시: 전체공지 삭제) — 메시지·멤버도 명시적 정리(FK 미적용 대비)
    try:
        for o in db.execute(
            "SELECT id FROM rooms WHERE type='channel' AND name='전체공지' AND channel_scope IS NULL"
        ).fetchall():
            rid = o["id"]
            db.execute("DELETE FROM messages WHERE room_id=?", (rid,))
            db.execute("DELETE FROM room_members WHERE room_id=?", (rid,))
            db.execute("DELETE FROM rooms WHERE id=?", (rid,))
    except Exception as e:
        print(f"[auto_channel] 전체공지 삭제 실패(무시): {e}")
    ids = {}
    for scope, cname in AUTO_CHANNELS:
        row = db.execute("SELECT id, name FROM rooms WHERE channel_scope=?", (scope,)).fetchone()
        if row:
            ids[scope] = row["id"]
            # 기존 채널 이름이 바뀐 경우(예: 본사 🏢→🇰🇷) 자동 동기화 (대표 지시 2026-05-21)
            if row["name"] != cname:
                db.execute("UPDATE rooms SET name=? WHERE id=?", (cname, row["id"]))
        else:
            cur = db.execute(
                "INSERT INTO rooms (name, type, created_by, created_at, name_locked, channel_scope) "
                "VALUES (?,?,?,?,?,?)",
                (cname, "channel", 1, now, 1, scope),
            )
            ids[scope] = cur.lastrowid
    return ids


def _sync_user_auto_channels(db, uid):
    """한 사용자의 자동채널 멤버십을 소속에 맞춰 추가/제거 (직원 등록·정보수정 후 호출)."""
    try:
        u = db.execute(
            "SELECT id, username, department, COALESCE(active,1) AS active, COALESCE(is_guest,0) AS is_guest FROM users WHERE id=?", (uid,)
        ).fetchone()
        if not u:
            return
        ids = _auto_channel_ids(db)
        want = _desired_scopes_for(u["department"], u["active"], _is_owner(u["username"]), is_guest=u["is_guest"])
        now = datetime.now(timezone.utc).isoformat()
        for scope, rid in ids.items():
            if scope in want:
                db.execute(
                    "INSERT OR IGNORE INTO room_members (room_id, user_id, joined_at, role) VALUES (?,?,?,?)",
                    (rid, uid, now, "member"),
                )
            else:
                db.execute("DELETE FROM room_members WHERE room_id=? AND user_id=?", (rid, uid))
        db.commit()
    except Exception as e:
        print(f"[auto_channel] sync_user({uid}) 실패: {e}")


def _resync_auto_channels(db):
    """전 직원 자동채널 멤버십 일괄 동기화 (+ 채널 생성·전체공지 삭제). 부팅·일괄등록 시 호출."""
    ids = _auto_channel_ids(db)
    now = datetime.now(timezone.utc).isoformat()
    users = db.execute("SELECT id, username, department, COALESCE(active,1) AS active, COALESCE(is_guest,0) AS is_guest FROM users").fetchall()
    want = {rid: set() for rid in ids.values()}
    for u in users:
        for scope in _desired_scopes_for(u["department"], u["active"], _is_owner(u["username"]), is_guest=u["is_guest"]):
            want[ids[scope]].add(u["id"])
    for scope, rid in ids.items():
        have = set(r["user_id"] for r in db.execute(
            "SELECT user_id FROM room_members WHERE room_id=?", (rid,)
        ).fetchall())
        for uid in (want[rid] - have):
            db.execute(
                "INSERT OR IGNORE INTO room_members (room_id, user_id, joined_at, role) VALUES (?,?,?,?)",
                (rid, uid, now, "member"),
            )
        for uid in (have - want[rid]):
            db.execute("DELETE FROM room_members WHERE room_id=? AND user_id=?", (rid, uid))
    db.commit()


def _cleanup_revoked_guest_members(db):
    """고아 게스트 정리 (대표 지시 2026-05-30) — 부팅 시 자동 실행.
    · 회수(revoked) 만 남고 멤버 행이 안 지워진 옛 게스트(05-29 회수코드 배포 이전 잔존)를 청소.
    · 규칙: 어떤 방의 멤버인 게스트가, 그 방에 '미회수(revoked_at IS NULL)' 초대가 하나도 없으면
            그 방 멤버에서 제거. (정상 게스트는 미회수 초대가 있으므로 유지됨)
    · 이후 어디에도 미회수 초대가 없는 게스트 계정은 active=0 으로 비활성화.
    """
    try:
        # 1) 게스트 방 멤버십 중, 그 방에 미회수 초대가 없는 행 제거
        orphans = db.execute("""
            SELECT rm.room_id, rm.user_id
              FROM room_members rm
              JOIN users u ON u.id = rm.user_id
             WHERE COALESCE(u.is_guest,0) = 1
               AND NOT EXISTS (
                   SELECT 1 FROM guest_invites gi
                    WHERE gi.room_id = rm.room_id
                      AND gi.guest_user_id = rm.user_id
                      AND gi.revoked_at IS NULL
               )
        """).fetchall()
        for r in orphans:
            db.execute("DELETE FROM room_members WHERE room_id=? AND user_id=?",
                       (r["room_id"], r["user_id"]))
        # 2) 어디에도 미회수 초대가 없는 게스트 계정 비활성화
        db.execute("""
            UPDATE users SET active=0
             WHERE COALESCE(is_guest,0) = 1 AND COALESCE(active,1) = 1
               AND id NOT IN (
                   SELECT guest_user_id FROM guest_invites
                    WHERE revoked_at IS NULL AND guest_user_id IS NOT NULL
               )
        """)
        db.commit()
        if orphans:
            print(f"[guest_cleanup] 고아 게스트 멤버 {len(orphans)}건 정리", flush=True)
    except Exception as e:
        print(f"[guest_cleanup] 실패(무시): {e}", flush=True)


# ---------- Auth ----------
# ===== 동시 로그인 제한 (휴대폰 1대 + PC 1대) =====
_MOBILE_UA_RE = re.compile(r"Android|iPhone|iPad|iPod|IEMobile|Windows Phone|BlackBerry|Mobile|Mobi|Tablet", re.I)


def _device_type_from_ua(ua):
    """User-Agent 로 기기 종류 판별 — 휴대폰/태블릿은 'mobile', 그 외(PC)는 'pc'."""
    return "mobile" if (ua and _MOBILE_UA_RE.search(ua)) else "pc"


def _upsert_active_session(uid, dtype, token, ua):
    """(user_id, device_type) 활성 세션 등록/교체 — 같은 종류 옛 토큰을 덮어써 옛 기기를 무효화."""
    db = get_db()
    now = datetime.now(timezone.utc).isoformat()
    db.execute("""
        INSERT INTO active_sessions (user_id, device_type, token, user_agent, ip, created_at)
        VALUES (?,?,?,?,?,?)
        ON CONFLICT(user_id, device_type) DO UPDATE SET
            token=excluded.token, user_agent=excluded.user_agent,
            ip=excluded.ip, created_at=excluded.created_at
    """, (uid, dtype, token, (ua or "")[:200], (request.remote_addr if request else None), now))
    db.commit()


def _session_token_valid(uid):
    """현재 세션이 (user, device_type) 활성 세션과 일치하는지.
    · 토큰 없는 기존 로그인(배포 전)은 자동 인정(grandfather) — 대량 강제 로그아웃 방지.
    · 같은 종류 기기에서 새로 로그인하면 토큰이 안 맞아 False (= 밀려남)."""
    try:
        tok = session.get("sess_token")
        dtype = session.get("device_type")
        ua = request.headers.get("User-Agent", "") if request else ""
        if not tok:
            # 배포 전 로그인 — 토큰 발급 + 등록(adopt)
            dtype = _device_type_from_ua(ua)
            tok = secrets.token_urlsafe(24)
            session["sess_token"] = tok
            session["device_type"] = dtype
            _upsert_active_session(uid, dtype, tok, ua)
            return True
        row = get_db().execute(
            "SELECT token FROM active_sessions WHERE user_id=? AND device_type=?",
            (uid, dtype or "pc"),
        ).fetchone()
        if row is None:
            # 활성행 없음 = 이 세션은 종료됨 (휴대폰 완전 로그아웃으로 삭제됐거나 다른 기기가 차지 후 로그아웃).
            # → 무효 처리해 다음 요청에서 로그인 화면으로. (토큰 없는 '기존 로그인'은 위 grandfather 에서 이미 처리)
            return False
        return row["token"] == tok
    except Exception as e:
        # 검증 중 오류가 나도 로그인은 막지 않음 (안전 측 default)
        print(f"[session] token 검증 오류: {e}")
        return True


def current_user():
    uid = session.get("user_id")
    if not uid:
        return None
    user = get_db().execute("SELECT * FROM users WHERE id = ?", (uid,)).fetchone()
    if not user:
        return None
    # 동시 로그인 제한 — 다른 같은종류 기기에서 로그인했으면 이 세션은 무효
    if not _session_token_valid(uid):
        return None
    return user


def login_required(view):
    @wraps(view)
    def wrapped(*a, **k):
        if not current_user():
            # 외부 브라우저로 '고객사 초대' 재진입 — 미로그인인데 ?g=<토큰> 이 있으면 초대 인증 페이지로. (대표 지시 2026-05-30)
            #   카톡 인앱→기본 브라우저로 열면 게스트 세션이 없어 로그인 화면에 막히던 문제 해결.
            #   토큰 형태(URL-safe)만 허용 + 항상 우리 host_url 에 붙여 리다이렉트(오픈 리다이렉트 불가).
            _g = (request.args.get("g") or "").strip()
            if _g and re.fullmatch(r"[A-Za-z0-9_-]{8,80}", _g):
                return redirect(f"{request.host_url.rstrip('/')}{BASE_PATH}/g/{_g}")
            # user_id 는 있는데 current_user 가 None → 다른 기기 로그인으로 밀려남 (kicked)
            kicked = bool(session.get("user_id"))
            session.clear()
            if kicked:
                return redirect(url_for("login") + "?kicked=1")
            return redirect(url_for("login"))
        return view(*a, **k)
    return wrapped


def _force_logout_same_device_type(uid, dtype):
    """같은 종류 기기로 새 로그인 시, 기존 같은 종류 소켓에 즉시 force_logout 전송.
    밀려난 기기가 실시간으로 로그인 화면으로 빠지게 함 (HTTP 검증과 별개의 즉시 반영)."""
    try:
        sids = []
        with _user_conn_lock:
            for sid, info in _user_connections.get(uid, {}).items():
                if info.get("device") == dtype:
                    sids.append(sid)
        for sid in sids:
            try:
                socketio.emit("force_logout", {"reason": "다른 기기 로그인"}, to=sid)
            except Exception:
                pass
    except Exception as e:
        print(f"[session] force_logout emit 실패: {e}")


def _force_logout_all(uid):
    """이 사용자의 '모든' 소켓에 force_logout 전송 — 휴대폰 로그아웃 시 PC 등 전부 함께 로그아웃."""
    try:
        sids = []
        with _user_conn_lock:
            sids = list(_user_connections.get(uid, {}).keys())
        for sid in sids:
            try:
                socketio.emit("force_logout", {"reason": "휴대폰에서 로그아웃"}, to=sid)
            except Exception:
                pass
    except Exception as e:
        print(f"[session] force_logout_all emit 실패: {e}")


def _presence_offline_status(uid):
    """소켓이 모두 끊겼을 때 남에게 보여줄 상태 —
    '실제 휴대폰' 푸시 구독이 있으면 'mobile'(📱 휴대폰), 없으면(PC 잔재만/로그아웃) 'offline'."""
    try:
        has_mobile = any(
            _device_type_from_ua(r["user_agent"]) == "mobile"
            for r in get_db().execute(
                "SELECT user_agent FROM push_subscriptions WHERE user_id=?", (uid,)
            ).fetchall()
        )
        return "mobile" if has_mobile else "offline"
    except Exception:
        return "offline"


# ---------- Pages ----------
@app.route("/")
def index():
    # 모바일 브라우저 HTML 캐시 완전 우회:
    # 302 redirect 응답 자체도 캐시될 수 있으므로 작은 HTML + JS location.replace 로 매번 unique URL 발생
    # 이렇게 하면 옛 / 응답이 캐시에 박혀 있어도, 그 응답을 받기만 하면 JS 가 새 URL 로 강제 이동
    if current_user():
        body = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><title>KNK</title></head>
<body><script>
location.replace('{BASE_PATH}/chat?v={STATIC_VERSION}&t=' + Date.now() + '&r=' + Math.random());
</script>
<noscript><meta http-equiv="refresh" content="0;url={BASE_PATH}/chat?v={STATIC_VERSION}"></noscript>
</body></html>"""
        resp = make_response(body)
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0, private"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
        return resp
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    # 게스트 기기(쿠키 knk_gt)는 직원 로그인 화면을 절대 안 보이게 — 자기 초대 페이지(/g/토큰)로 보냄.
    #  뒤로가기·PWA 재실행 등 어떤 경로로 /login 에 와도 동일. ?staff=1 이면 우회(직원 로그인 화면 노출). (대표 지시 2026-05-30)
    if request.method == "GET" and not request.args.get("staff") and not session.get("user_id"):
        _cgt = (request.cookies.get("knk_gt") or "").strip()
        if _cgt and re.fullmatch(r"[A-Za-z0-9_-]{8,80}", _cgt):
            return redirect(f"{request.host_url.rstrip('/')}{BASE_PATH}/g/{_cgt}")
    if request.method == "POST":
        u = (request.form.get("username") or "").strip()
        p = request.form.get("password") or ""
        # 로그인 ID 매칭 (대표 지시 2026-05-26):
        #   · 본사  : username = 이메일 (예: top0015@knknara.co.kr)
        #   · 베트남: username = 영문이름 슬러그 (예: nguyenvana — '@knkvn.local' 도메인 없음)
        # 사용자가 입력한 값 그대로 조회. 베트남 직원은 'nguyenvana'만 입력하면 됨.
        row = get_db().execute("SELECT * FROM users WHERE username = ?", (u,)).fetchone()
        # 베트남 직원 대소문자 폴백 — 'NGUYENVANA' 같이 대문자로 입력해도 매칭
        if (not row) and u and ("@" not in u):
            row = get_db().execute("SELECT * FROM users WHERE username = ?", (u.lower(),)).fetchone()
            if row:
                u = u.lower()
        if row and check_password_hash(row["password_hash"], p):
            session.clear()
            session["user_id"] = row["id"]
            # 자동 로그인 유지 — PERMANENT_SESSION_LIFETIME 만큼(기본 90일) 쿠키 유지.
            # 명시적 /logout 호출 전까지 브라우저 닫고 켜도 로그인 상태.
            session.permanent = True
            # 동시 로그인 제한 — 기기 종류(폰/PC)별 토큰 발급 + 등록.
            # 같은 종류로 이미 로그인된 다른 기기는 이 토큰이 덮어써 자동 무효화(밀려남).
            try:
                ua = request.headers.get("User-Agent", "")
                dtype = _device_type_from_ua(ua)
                tok = secrets.token_urlsafe(24)
                session["sess_token"] = tok
                session["device_type"] = dtype
                _upsert_active_session(row["id"], dtype, tok, ua)
                # 같은 종류 기기로 이미 접속 중이던 기기는 즉시 force_logout (실시간 밀어내기)
                _force_logout_same_device_type(row["id"], dtype)
            except Exception as e:
                print(f"[login] active_session 등록 실패: {e}")
            # SSO: 세션에 redirect_uri 가 있으면 JWT 발급 후 Service Provider 로 복귀.
            #   (발주서 §3.5.1·§8.4 — 메신저 로그인 화면 그대로 재사용, 성공 시 토큰 발급)
            _sso_ru = session.pop("sso_redirect_uri", None)
            if _sso_ru and _sso_available():
                _safe_ru = _sso_safe_redirect_uri(_sso_ru)
                if _safe_ru:
                    try:
                        _sso_tok = _sso_issue_token(row)
                        _sep = "&" if ("?" in _safe_ru) else "?"
                        _r2 = redirect(f"{_safe_ru}{_sep}token={_sso_tok}")
                        try:
                            _r2.delete_cookie("knk_gt", path="/")
                        except Exception:
                            pass
                        _sso_log("login_redirect", emp=(row["employee_no"] if "employee_no" in row.keys() else row["username"]))
                        return _r2
                    except Exception as _e:
                        print(f"[SSO] login redirect 실패: {_e}")
            _resp = redirect(url_for("chat") + f"?v={STATIC_VERSION}&t={int(_time.time())}")
            # 직원 로그인 성공 → 이 기기는 더 이상 게스트 기기 아님 → 게스트 표식 쿠키 제거. (대표 지시 2026-05-30)
            try:
                _resp.delete_cookie("knk_gt", path="/")
            except Exception:
                pass
            return _resp
        return render_template("login.html", error="아이디 또는 비밀번호가 올바르지 않습니다.")
    # GET — 강제 로그아웃(밀려남/휴대폰 완전로그아웃) 안내
    kicked_msg = None
    if request.args.get("kicked"):
        r = request.args.get("r") or ""
        if "휴대폰" in r:
            kicked_msg = "휴대폰에서 로그아웃하여 이 PC도 함께 로그아웃되었습니다."
        else:
            kicked_msg = "다른 기기에서 로그인되어 이 기기는 로그아웃되었습니다. (동시 사용은 휴대폰 1대 + PC 1대까지)"
    _gresp = make_response(render_template("login.html", notice=kicked_msg))
    # ?staff 로 직원 로그인에 직접 들어오면 = 이 기기를 '직원 기기'로 선언 → 남아있던 게스트 토큰 쿠키 삭제.
    #   (게스트 테스트했던 기기가 직원 로그인 못 들어가고 고객사 페이지로 빨려가는 문제 영구 해소) (대표 지시 2026-05-31)
    if request.args.get("staff"):
        try:
            _gresp.delete_cookie("knk_gt", path="/")
        except Exception:
            pass
    return _gresp


@app.route("/sso/login")
def sso_login():
    """SSO 진입점 — Service Provider(HAIST WORKS 등) 가 미인증 사용자를 보내는 곳.
    redirect_uri 를 세션에 저장 후 평소 메신저 로그인 화면을 그대로 보여준다(§8.4).
    이미 로그인돼 있으면 재입력 없이 즉시 토큰 발급 후 복귀(진짜 SSO)."""
    ru = (request.args.get("redirect_uri") or "").strip()
    safe = _sso_safe_redirect_uri(ru)
    if safe:
        session["sso_redirect_uri"] = safe
    elif ru:
        print(f"[SSO] sso_login 거부된 redirect_uri: {ru!r}")
    # 이미 메신저에 로그인된 상태면 바로 토큰 발급 후 SP 로 복귀
    if safe and session.get("user_id") and _sso_available():
        try:
            row = get_db().execute("SELECT * FROM users WHERE id=?", (session["user_id"],)).fetchone()
            if row:
                tok = _sso_issue_token(row)
                sep = "&" if ("?" in safe) else "?"
                session.pop("sso_redirect_uri", None)
                _sso_log("sso_login_immediate", uid=session.get("user_id"))
                return redirect(f"{safe}{sep}token={tok}")
        except Exception as e:
            print(f"[SSO] sso_login 즉시발급 실패: {e}")
    return redirect(url_for("login"))


# ============================================================
# 내 사번 찾기 — 로그인 안 한 상태에서 이름으로 사번·부서 조회
# (대표 지시 2026-05-27 — 메신저 테스트 배포 단계 직원 자체 확인용)
# ============================================================
# 보안 고려:
#   · 베트남 9999 공통 비번 + 사번 노출 위험은 first-login 강제 비번 변경으로 커버
#     (must_change_password 가 1 — 누군가 먼저 로그인해도 비번 변경 화면이 강제됨)
#   · 본인 외 정보 노출 최소화: 동명이인 최대 5명, 휴대폰·이메일은 보이지 않음
#   · 남용 방지: 같은 IP 5분 내 5회 실패 시 차단
_lookup_attempts = {}  # ip -> [(ts, success), ...]  최근 5분 보관
_lookup_attempts_lock = _pres_threading.Lock()
_LOOKUP_WINDOW_SEC = 300       # 5분
_LOOKUP_MAX_PER_WINDOW = 20    # 같은 IP, 5분 내 총 시도 20회 (성공 포함, 부르트포스 차단)
_LOOKUP_MAX_FAIL_PER_WINDOW = 5  # 5분 내 실패 5회 시 차단


def _client_ip_for_lookup():
    """nginx 뒤에서도 실제 클라이언트 IP 추출 (남용 방지용)."""
    if TRUSTED_PROXIES:
        xff = request.headers.get("X-Forwarded-For", "")
        if xff:
            return xff.split(",")[0].strip()
    return request.remote_addr or "unknown"


def _lookup_record(ip, success):
    """이력 기록 + 오래된 항목 정리."""
    now = _pres_time.time()
    with _lookup_attempts_lock:
        arr = _lookup_attempts.setdefault(ip, [])
        arr.append((now, bool(success)))
        # 5분 초과 항목 제거
        cutoff = now - _LOOKUP_WINDOW_SEC
        _lookup_attempts[ip] = [(t, s) for (t, s) in arr if t >= cutoff]


def _lookup_is_blocked(ip):
    """차단된 IP 인지. (총 횟수 또는 실패 횟수 초과)"""
    now = _pres_time.time()
    cutoff = now - _LOOKUP_WINDOW_SEC
    with _lookup_attempts_lock:
        arr = [(t, s) for (t, s) in _lookup_attempts.get(ip, []) if t >= cutoff]
        if len(arr) >= _LOOKUP_MAX_PER_WINDOW:
            return True
        fails = sum(1 for (_, s) in arr if not s)
        if fails >= _LOOKUP_MAX_FAIL_PER_WINDOW:
            return True
    return False


def _classify_user_mode(username):
    """username 으로 본사/베트남/불명 구분.
       명확한 시그널만 본사·베트남으로 분류, 애매하면 'unknown'.
       UI 는 unknown 일 때 양쪽 비번 안내를 모두 표시 (오인 방지)."""
    if not username:
        return "unknown"
    u = str(username).strip().lower()
    # 명확한 베트남: 'vn' + 숫자 (예: vn1, vn23), 또는 @knkvn 도메인
    if u.startswith("vn") and len(u) > 2 and u[2:].replace("@knkvn.local", "").isdigit():
        return "vn"
    if "@knkvn" in u:
        return "vn"
    # 명확한 본사: 순수 숫자 사번 (예: "5"), 또는 @knknara 도메인
    if u.isdigit():
        return "kor"
    if "@knknara" in u:
        return "kor"
    # 영문 슬러그(예: 'kjr', 'nguyenvana') — 양쪽 모두 가능. 본인이 확인하도록 unknown.
    return "unknown"


@app.route("/api/lookup_employee", methods=["POST"])
def api_lookup_employee():
    """이름으로 사번·부서 찾기. 로그인 불필요.
    body: {"name": "홍길동"}
    응답: {"results": [{username, display_name, employee_no, department, title, mode, pw_hint}], "total": N}
    실패: 429 (차단) / 400 (이름 짧음)
    """
    ip = _client_ip_for_lookup()
    if _lookup_is_blocked(ip):
        return jsonify({
            "error": "blocked",
            "message": "잠시 후 다시 시도해 주세요. (5분 내 시도가 너무 많습니다)"
        }), 429

    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    if len(name) < 2:
        _lookup_record(ip, success=False)
        return jsonify({
            "error": "too_short",
            "message": "이름을 2자 이상 입력해 주세요."
        }), 400

    # 너무 긴 이름 차단 (DB 부하·SQL 안전 — Like 검색은 % escape)
    if len(name) > 40:
        _lookup_record(ip, success=False)
        return jsonify({"error": "too_long", "message": "이름이 너무 깁니다."}), 400

    # LIKE 와일드카드 안전 처리 (% _ \ 이스케이프)
    safe = name.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
    pattern = f"%{safe}%"

    db = get_db()
    # 한글(display_name) + 베트남어(display_name_vn) + 영문(display_name_en) 3개 이름 모두 검색 (대표 지시 2026-05-29)
    rows = db.execute("""
        SELECT id, username, display_name, display_name_vn, display_name_en,
               employee_no, department, title, role
          FROM users
         WHERE display_name LIKE ? ESCAPE '\\'
            OR COALESCE(display_name_vn,'') LIKE ? ESCAPE '\\'
            OR COALESCE(display_name_en,'') LIKE ? ESCAPE '\\'
         ORDER BY
           CASE WHEN display_name = ?
                  OR COALESCE(display_name_vn,'') = ?
                  OR COALESCE(display_name_en,'') = ? THEN 0 ELSE 1 END,
           length(display_name) ASC,
           id ASC
         LIMIT 5
    """, (pattern, pattern, pattern, name, name, name)).fetchall()

    if not rows:
        _lookup_record(ip, success=False)
        return jsonify({
            "results": [],
            "total": 0,
            "message": "해당 이름의 직원을 찾을 수 없습니다. 관리자에게 문의하세요."
        }), 200

    # 입력한 언어 감지 — 그 언어의 이름을 주(主)로, 나머지는 보조로 표시 (대표 지시 2026-05-29)
    #   한글 글자 → ko / 베트남어 성조부호 → vi / 그 외 라틴문자 → en·vi 모호 (en 우선)
    _vi_diacritics = "ăâđêôơưàáảãạằắẳẵặầấẩẫậèéẻẽẹềếểễệìíỉĩịòóỏõọồốổỗộờớởỡợùúủũụừứửữựỳýỷỹỵ"
    q_low = name.lower()
    if re.search(r"[가-힣]", name):
        _pref = ["ko", "vi", "en"]
    elif any(ch in _vi_diacritics for ch in q_low):
        _pref = ["vi", "en", "ko"]
    else:
        _pref = ["en", "vi", "ko"]

    # 결과 조립
    results = []
    for r in rows:
        username = r["username"] or ""
        mode = _classify_user_mode(username)
        # 사번 표시: employee_no 가 있으면 그것, 없으면 username 그대로
        emp_no_display = r["employee_no"] or username
        if mode == "kor":
            pw_hint = "본인 휴대폰 번호 (숫자만 11자리)"
        elif mode == "vn":
            pw_hint = "9999"
        else:
            pw_hint = "본사: 본인 휴대폰 번호 / 베트남: 9999"
        # 3개 이름 + 입력어에 맞춘 주/보조 이름 계산
        _names = {
            "ko": (r["display_name"] or "").strip(),
            "vi": (r["display_name_vn"] or "").strip(),
            "en": (r["display_name_en"] or "").strip(),
        }
        _matched = [lang for lang in _pref if _names[lang] and q_low in _names[lang].lower()]
        primary_lang = _matched[0] if _matched else ("ko" if _names["ko"] else next((l for l in _pref if _names[l]), "ko"))
        primary_name = _names[primary_lang] or _names["ko"]
        # 보조 이름 — 주 이름과 다른, 값이 있는 나머지 (ko, vi, en 순서 유지)
        other_names = []
        for lang in ("ko", "vi", "en"):
            v = _names[lang]
            if v and v != primary_name and v not in other_names:
                other_names.append(v)
        results.append({
            "username": username,                # 로그인 ID (그대로 입력하면 됨)
            "employee_no": emp_no_display,
            "display_name": r["display_name"] or "",   # (호환 유지)
            "primary_name": primary_name,        # 입력한 언어의 이름 (메인 표시)
            "other_names": other_names,          # 나머지 언어 이름들 (작게 표시)
            "department": r["department"] or "",
            "title": r["title"] or "",
            "mode": mode,                        # "kor" / "vn" / "unknown"
            "pw_hint": pw_hint,
        })

    _lookup_record(ip, success=True)
    return jsonify({"results": results, "total": len(results)})


@app.route("/logout")
def logout():
    # 비대칭 로그아웃 (대표 지시 2026-05-20):
    #  · 휴대폰 로그아웃 = '완전 로그아웃' — 폰+PC 전부 로그아웃 + 모든 푸시 삭제 → 완전 오프라인.
    #  · PC 로그아웃     = '이 PC 만'   — PC 세션·PC 푸시만 정리. 휴대폰은 세션·알림 그대로.
    uid = session.get("user_id")
    dtype = session.get("device_type") or "pc"
    tok = session.get("sess_token")
    ep = session.get("push_endpoint")
    # 게스트(외부 고객) 여부 + 재진입용 초대 토큰 미리 조회 (cleanup·session.clear 전). (대표 지시 2026-05-30)
    _is_guest_logout = False
    _guest_token = None
    if uid:
        try:
            _gdb = get_db()
            _gu = _gdb.execute("SELECT COALESCE(is_guest,0) AS g FROM users WHERE id=?", (uid,)).fetchone()
            if _gu and int(_gu["g"] or 0) == 1:
                _is_guest_logout = True
                _gt = _gdb.execute(
                    "SELECT token FROM guest_invites WHERE guest_user_id=? ORDER BY (revoked_at IS NULL) DESC, id DESC LIMIT 1",
                    (uid,),
                ).fetchone()
                if _gt:
                    _guest_token = _gt["token"]
        except Exception as e:
            print(f"[logout] 게스트 조회 실패(무시): {e}")
    if uid:
        try:
            db = get_db()
            if _is_guest_logout:
                # 게스트(외부 고객, 단일 기기) — 본인 세션·구독만 정리. force_logout 브로드캐스트 안 함.
                #  (force_logout 이 현재 기기를 /logout_local→직원 로그인 으로 밀어내, 게스트가 직원
                #   로그인 화면으로 빠지던 문제 방지. 대표 지시 2026-05-30)
                db.execute("DELETE FROM push_subscriptions WHERE user_id=?", (uid,))
                db.execute("DELETE FROM active_sessions WHERE user_id=?", (uid,))
                db.commit()
            elif dtype == "mobile":
                # 휴대폰 = 모든 기기 완전 로그아웃
                db.execute("DELETE FROM push_subscriptions WHERE user_id=?", (uid,))
                db.execute("DELETE FROM active_sessions WHERE user_id=?", (uid,))
                db.commit()
                _force_logout_all(uid)               # PC 등 다른 기기 즉시 로그아웃
                try:
                    _last_status_bcast[uid] = "offline|0"   # disconnect 중복 emit 방지
                    socketio.emit("user_status_changed", {
                        "user_id": uid, "status": "offline",
                        "custom_text": None, "emoji": None,
                        "label": STATUS_LABEL_KO.get("offline", "오프라인"),
                        "at_office": False,
                    })
                except Exception:
                    pass
            else:
                # PC = 이 기기만 — 휴대폰에 영향 없음
                #  · 이 PC 의 푸시 구독 삭제 (세션 endpoint) + PC(데스크톱) 종류 잔재 정리 (누적 방지)
                #    휴대폰 종류 구독은 보존 → 휴대폰 알림은 그대로 유지. (대표 지시 2026-05-25)
                if ep:
                    db.execute("DELETE FROM push_subscriptions WHERE user_id=? AND endpoint=?", (uid, ep))
                try:
                    for r in db.execute(
                        "SELECT endpoint, user_agent FROM push_subscriptions WHERE user_id=?", (uid,)
                    ).fetchall():
                        if _device_type_from_ua(r["user_agent"]) == "pc":
                            db.execute("DELETE FROM push_subscriptions WHERE user_id=? AND endpoint=?",
                                       (uid, r["endpoint"]))
                except Exception as _e:
                    print(f"[logout] PC 푸시 잔재 정리 실패(무시): {_e}")
                if tok:
                    db.execute("DELETE FROM active_sessions WHERE user_id=? AND token=?", (uid, tok))
                db.commit()
                # 상태 broadcast 는 소켓 disconnect 핸들러가 '휴대폰/오프라인' 정확히 계산해 처리
        except Exception as e:
            print(f"[logout] cleanup 실패: {e}")
    session.clear()
    # 게스트(외부 고객)는 직원용 로그인 화면이 아니라 '자기 초대 페이지(/g/토큰)'로 — QR 재진입/회수·만료 안내. (대표 지시 2026-05-30)
    if _guest_token:
        return redirect(f"{request.host_url.rstrip('/')}{BASE_PATH}/g/{_guest_token}")
    return redirect(url_for("login"))


@app.route("/logout_local")
def logout_local():
    """'이 기기만' 로그아웃 — 다른 기기 로그인으로 밀려난 기기 전용.
    푸시 구독은 건드리지 않음(계정의 다른 기기 알림을 끄면 안 되므로). 세션만 정리."""
    uid = session.get("user_id")
    tok = session.get("sess_token")
    # 게스트면 직원 로그인 대신 자기 초대 페이지로 (밀려난 경우에도 QR 재진입 안내). (대표 지시 2026-05-30)
    _guest_token = None
    if uid:
        try:
            _gdb = get_db()
            _gu = _gdb.execute("SELECT COALESCE(is_guest,0) AS g FROM users WHERE id=?", (uid,)).fetchone()
            if _gu and int(_gu["g"] or 0) == 1:
                _gt = _gdb.execute(
                    "SELECT token FROM guest_invites WHERE guest_user_id=? ORDER BY (revoked_at IS NULL) DESC, id DESC LIMIT 1",
                    (uid,),
                ).fetchone()
                if _gt:
                    _guest_token = _gt["token"]
        except Exception as e:
            print(f"[logout_local] 게스트 조회 실패(무시): {e}")
    if uid and tok:
        try:
            db = get_db()
            # 내 토큰의 활성세션 행만 제거 — 이미 새 기기가 덮어썼으면 불일치로 아무것도 안 지워짐(안전)
            db.execute("DELETE FROM active_sessions WHERE user_id=? AND token=?", (uid, tok))
            db.commit()
        except Exception:
            pass
    session.clear()
    if _guest_token:
        return redirect(f"{request.host_url.rstrip('/')}{BASE_PATH}/g/{_guest_token}")
    return redirect(url_for("login", kicked=1, r=(request.args.get("r") or "")))


@app.route("/capture")
@login_required
def capture_page():
    """별도 캡처 창 — 화면 잡기 + 영역 잘라내기 후 opener(채팅)로 이미지 전달. (대표 지시 2026-06-06)"""
    return render_template("capture.html")


@app.route("/chat")
@login_required
def chat():
    _me = current_user()
    is_guest = _is_guest(_me)
    # 게스트는 self_room 안 만듦 (대표 지시 2026-05-28)
    if not is_guest:
        try:
            _ensure_self_room(session["user_id"])
        except Exception as e:
            print(f"[chat] self_room 보장 실패: {e}")
    # 게스트 — 외부 브라우저 재진입용 초대 토큰(살아있는 것). 방 URL 에 ?g=토큰 으로 달아둠. (대표 지시 2026-05-30)
    me_guest_token = None
    if is_guest:
        try:
            _gt = get_db().execute(
                "SELECT token FROM guest_invites WHERE guest_user_id=? AND revoked_at IS NULL ORDER BY id DESC LIMIT 1",
                (_me["id"],),
            ).fetchone()
            if _gt:
                me_guest_token = _gt["token"]
        except Exception as e:
            print(f"[chat] 게스트 토큰 조회 실패: {e}")
    return render_template("chat.html", me=_me, me_is_owner=_is_owner(_me["username"]),
                           me_is_team_lead=_is_team_lead(_me), me_can_create_channel=_can_create_channel(_me),
                           me_is_maintenance=1 if _is_maintenance_admin(_me) else 0,
                           me_is_maint_owner=1 if _is_maintenance_owner(_me) else 0,
                           bug_room_id=_get_bug_room_id(get_db()),
                           me_ui_lang=_user_ui_lang(_me), me_ui_theme=_user_ui_theme(_me),
                           me_is_guest=1 if is_guest else 0,
                           me_guest_room_id=_me["guest_room_id"] if is_guest else None,
                           me_guest_company=_me["guest_company"] if is_guest else None,
                           me_guest_token=me_guest_token,
                           # HAIST WORKS 진입 — 권한 있는 직원만 'WORKS 열기' 버튼 노출. (대표 지시 2026-05-31)
                           me_works_access=(0 if is_guest else (1 if (("works_access" in _me.keys()) and _me["works_access"]) else 0)),
                           works_landing_url=WORKS_LANDING_URL)


@app.route("/api/me/lang", methods=["GET", "POST"])
@login_required
def api_me_lang():
    """내 화면 표시 언어 조회/변경. POST body: {lang: 'ko'|'vi'|'en'}.
    빈 값/없는 값이면 법인 기준 자동으로 되돌림(컬럼 NULL). (대표 지시 2026-05-25)"""
    me = current_user()
    if request.method == "GET":
        return jsonify({"lang": _user_ui_lang(me), "options": list(UI_LANGS)})
    data = request.get_json(silent=True) or {}
    lang = (data.get("lang") or "").strip().lower()
    db = get_db()
    if lang in UI_LANGS:
        db.execute("UPDATE users SET ui_lang=? WHERE id=?", (lang, me["id"]))
    else:
        db.execute("UPDATE users SET ui_lang=NULL WHERE id=?", (me["id"],))  # 자동(법인 기준)으로 환원
    db.commit()
    return jsonify({"ok": True, "lang": _user_ui_lang(current_user())})


@app.route("/api/me/theme", methods=["GET", "POST"])
@login_required
def api_me_theme():
    """내 화면 테마 조회/변경. POST body: {theme: 'light'|'dark'|'sage'|'cream'|'sky'}.
    빈 값/없는 값이면 light 로 환원(컬럼 NULL). (대표 지시 2026-05-28)"""
    me = current_user()
    if request.method == "GET":
        return jsonify({"theme": _user_ui_theme(me), "options": list(UI_THEMES)})
    data = request.get_json(silent=True) or {}
    theme = (data.get("theme") or "").strip().lower()
    db = get_db()
    if theme in UI_THEMES and theme != "light":
        db.execute("UPDATE users SET ui_theme=? WHERE id=?", (theme, me["id"]))
    else:
        db.execute("UPDATE users SET ui_theme=NULL WHERE id=?", (me["id"],))   # light = 기본 (NULL 로 저장)
    db.commit()
    return jsonify({"ok": True, "theme": _user_ui_theme(current_user())})


@app.route("/dashboard")
@login_required
def dashboard():
    me = current_user()
    # 외부 게스트(고객사)는 전사 프로젝트 대시보드 접근 불가 — 채팅으로 (대표 지시 2026-05-30)
    if _is_guest(me):
        return redirect(url_for("chat"))
    return render_template("dashboard.html", me=me, me_ui_theme=_user_ui_theme(me))




# ---------- API ----------
@app.route("/healthz")
def healthz():
    """헬스체크 — BASE_PATH 와 무관하게 항상 루트(/healthz)에서 응답."""
    return "ok", 200


@app.before_request
def _force_https():
    """http 로 들어온 요청을 https 로 301 영구 전환 (대표 지시 2026-06-04).
    DSM 리버스 프록시가 원래 scheme 을 X-Forwarded-Proto 로 정확히 전달함(실측 확인 2026-06-04).
    - 헬스체크(/healthz)는 제외 — 모니터링이 평문으로도 200 을 받게.
    - X-Forwarded-Proto 헤더가 없으면(내부 직접접속·docker healthcheck) 전환 안 함 → 무한루프·헬스체크 깨짐 방지.
    - https 면 헤더가 'https' 라 전환 안 함 → 루프 없음."""
    if request.headers.get("X-Forwarded-Proto", "").lower() == "http" and request.path != "/healthz":
        return redirect(request.url.replace("http://", "https://", 1), code=301)


@app.route("/sw.js")
def serve_sw():
    """Service Worker — BASE_PATH 하위 경로 배포 시 /msg/sw.js 로 서빙되며 scope 는 /msg/.
    sw.js 내부에서 base 경로는 등록 시 ?base= 쿼리로 전달받는다."""
    return send_from_directory(os.path.join(APP_DIR, "static"), "sw.js", mimetype="application/javascript")


@app.route("/manifest.json")
def serve_manifest():
    """PWA manifest — BASE_PATH 를 반영해 동적 생성 (start_url·scope·아이콘 경로)."""
    bp = BASE_PATH
    data = {
        # id: 앱 고유 식별자. start_url 과 동일하게 두어 기존 설치본과 같은 앱으로 인식.
        "id": f"{bp}/chat",
        "name": "KNK Eum",
        "short_name": "KNK 이음",
        "description": "KNK 사내 업무 전용 메신저 — 프로젝트별 자동 정리·요청 추적·전사 검색",
        "start_url": f"{bp}/chat",
        "scope": f"{bp}/",
        # 설치된 앱 아이콘을 다시 눌러도 새 창을 만들지 않고 기존 창을 화면 앞으로(단일 창).
        "launch_handler": {"client_mode": "focus-existing"},
        "display": "standalone",
        "orientation": "portrait",
        "background_color": "#ffffff",
        "theme_color": "#A5282C",
        "lang": "ko",
        "icons": [
            {"src": f"{bp}/static/icons/icon-192.png", "sizes": "192x192", "type": "image/png", "purpose": "any maskable"},
            {"src": f"{bp}/static/icons/icon-512.png", "sizes": "512x512", "type": "image/png", "purpose": "any maskable"},
            {"src": f"{bp}/static/icons/icon.svg", "sizes": "any", "type": "image/svg+xml", "purpose": "any"},
        ],
        "categories": ["business", "productivity"],
    }
    resp = make_response(jsonify(data))
    resp.headers["Content-Type"] = "application/manifest+json; charset=utf-8"
    return resp


# ─── 고객사 전용 PWA (대표 지시 2026-05-31) ───
#   목적: 고객사가 설치하는 앱은 (1) 홈화면 이름=고객사명·KNK (2) 아이콘=고객사명 자동생성
#         (3) 재실행 시 항상 '고객사 로그인(/g/토큰)'으로 — 직원 로그인으로 빠지던 문제 해결.
_GUEST_ICON_CACHE = {}   # (company, size) -> PNG bytes 메모리 캐시


def _guest_company_for_token(token):
    """초대 토큰 → 고객사명. 토큰 형식 검증 + 조회. 없으면 None."""
    if not token or not re.fullmatch(r"[A-Za-z0-9_-]{8,80}", token):
        return None
    try:
        row = get_db().execute(
            "SELECT guest_company FROM guest_invites WHERE token=?", (token,)
        ).fetchone()
        if row:
            return ((row["guest_company"] or "").strip()) or None
    except Exception as e:
        print(f"[guest_app] 회사명 조회 실패: {e}")
    return None


_KNK_LOGO_SENTINEL = object()
_KNK_LOGO_IMG = _KNK_LOGO_SENTINEL   # 모듈 캐시 (sentinel=미로드 / None=실패 / Image=성공)


def _load_knk_logo():
    """static/icons/icon.svg 에 박힌 KNK 로고 PNG(base64)를 1회 추출해 RGBA(투명) 로 반환(캐시)."""
    global _KNK_LOGO_IMG
    if _KNK_LOGO_IMG is not _KNK_LOGO_SENTINEL:
        return _KNK_LOGO_IMG
    try:
        from PIL import Image
        import io as _io, base64 as _b64
        with open(os.path.join(app.root_path, "static", "icons", "icon.svg"), encoding="utf-8") as fp:
            svg = fp.read()
        mm = re.search(r"base64,([A-Za-z0-9+/=]+)", svg)
        _KNK_LOGO_IMG = Image.open(_io.BytesIO(_b64.b64decode(mm.group(1)))).convert("RGBA") if mm else None
    except Exception as e:
        print(f"[guest_icon] KNK 로고 로드 실패: {e}")
        _KNK_LOGO_IMG = None
    return _KNK_LOGO_IMG


def _make_guest_app_icon(size=512):
    """고객사 홈화면 아이콘 PNG — 차분한 회색(차콜) 배경 + 밝은회색 KNK 로고. (대표 지시 2026-05-31, 회색 시안 C)
    모든 고객사 공통(로고만 / 고객사명은 앱 이름 라벨로). 너무 강렬하지 않게 회색 계열. size 별 메모리 캐시."""
    if size in _GUEST_ICON_CACHE:
        return _GUEST_ICON_CACHE[size]
    from PIL import Image
    import io as _io
    BG = (75, 85, 99)            # 차콜 #4B5563
    LOGO_RGB = (229, 231, 235)   # 밝은 회색 #E5E7EB
    img = Image.new("RGB", (size, size), BG)
    logo = _load_knk_logo()
    if logo is not None:
        try:
            alpha = logo.split()[3]
            tinted = Image.new("RGBA", logo.size, (LOGO_RGB[0], LOGO_RGB[1], LOGO_RGB[2], 0))
            tinted.putalpha(alpha)
            lw, lh = logo.size
            tw = int(size * 0.62)                  # 마스크 안전영역(중앙) 안에 들어오는 크기
            th = max(1, int(tw * lh / lw))
            tinted = tinted.resize((tw, th), Image.LANCZOS)
            img.paste(tinted, ((size - tw) // 2, (size - th) // 2), tinted)
        except Exception as e:
            print(f"[guest_icon] 로고 합성 실패(배경만 출력): {e}")
    buf = _io.BytesIO()
    img.save(buf, format="PNG")
    data = buf.getvalue()
    _GUEST_ICON_CACHE[size] = data
    return data


@app.route("/g/<token>/manifest.json")
def serve_guest_manifest(token):
    """고객사 전용 PWA manifest — 이름=고객사명·KNK, 아이콘=고객사 자동생성, start_url=/g/토큰(재실행 시 고객사 로그인)."""
    bp = BASE_PATH
    company = _guest_company_for_token(token)
    if not company:
        return serve_manifest()          # 토큰 무효/회수 → 일반 KNK manifest 폴백
    short = company if len(company) <= 12 else company[:12]
    data = {
        "id": f"{bp}/g/{token}",
        "name": f"{company} · KNK",
        "short_name": short,
        "description": f"{company} ↔ KNK 이음",
        "start_url": f"{bp}/g/{token}",
        "scope": f"{bp}/",
        "launch_handler": {"client_mode": "focus-existing"},
        "display": "standalone",
        "orientation": "portrait",
        "background_color": "#A5282C",
        "theme_color": "#A5282C",
        "lang": "ko",
        "icons": [
            {"src": f"{bp}/g/{token}/icon-192.png?v={STATIC_VERSION}", "sizes": "192x192", "type": "image/png", "purpose": "any maskable"},
            {"src": f"{bp}/g/{token}/icon-512.png?v={STATIC_VERSION}", "sizes": "512x512", "type": "image/png", "purpose": "any maskable"},
        ],
        "categories": ["business", "productivity"],
    }
    resp = make_response(jsonify(data))
    resp.headers["Content-Type"] = "application/manifest+json; charset=utf-8"
    resp.headers["Cache-Control"] = "no-cache"
    return resp


@app.route("/g/<token>/icon-<int:size>.png")
def serve_guest_icon(token, size):
    """고객사 홈화면 아이콘 — 고객사명으로 동적 생성 PNG (180=apple-touch / 192·512=manifest)."""
    if size not in (180, 192, 512):
        size = 192
    # 아이콘은 모든 고객사 공통(회색 KNK 로고) — 고객사명은 앱 이름 라벨로만. (대표 지시 2026-05-31)
    try:
        png = _make_guest_app_icon(size)
    except Exception as e:
        print(f"[guest_icon] 생성 실패 → 기본 아이콘: {e}")
        return redirect(f"{BASE_PATH}/static/icons/icon-{'512' if size >= 512 else '192'}.png")
    resp = make_response(png)
    resp.headers["Content-Type"] = "image/png"
    resp.headers["Cache-Control"] = "public, max-age=3600"
    return resp


@app.route("/api/verify_password", methods=["POST"])
@login_required
def api_verify_password():
    """화면 잠금 해제용 — 현재 로그인 사용자의 비밀번호 확인 (잠금 해제 외 용도 없음)."""
    me = current_user()
    data = request.get_json(silent=True) or {}
    pw = data.get("password") or ""
    row = get_db().execute("SELECT password_hash FROM users WHERE id = ?", (me["id"],)).fetchone()
    ok = bool(pw) and bool(row) and check_password_hash(row["password_hash"], pw)
    return jsonify({"ok": ok})


@app.route("/api/me")
@login_required
def api_me():
    u = current_user()
    is_guest_flag = 0
    guest_room = None
    guest_company_val = None
    try:
        is_guest_flag = int(u["is_guest"] or 0)
        guest_room = u["guest_room_id"]
        guest_company_val = u["guest_company"]
    except Exception:
        pass
    return jsonify({
        "id": u["id"], "username": u["username"],
        "display_name": u["display_name"], "role": u["role"],
        "avatar_color": u["avatar_color"],
        "is_owner": _is_owner(u["username"]),
        "is_team_lead": _is_team_lead(u),
        # 게스트 정보 (대표 지시 2026-05-28)
        "is_guest": is_guest_flag,
        "guest_room_id": guest_room,
        "guest_company": guest_company_val,
    })


@app.route("/api/users/counts")
@login_required
def api_users_counts():
    """본사·베트남 활성 인원수. 사이드바 '👥 사용자' 탭 라벨 옆에 표시 (대표 지시 2026-05-27).
       시스템 함수 _user_is_vietnam (prefix '02_VN/' 공식 또는 '12-VN' 레거시) 와 동일 패턴 사용.
       베트남이 아니면 모두 본사로 집계 (자동채널 _desired_scopes_for 와 일관).
       비활성(active=0) 제외 + 시스템 플레이스홀더 제외.
       게스트(외부)는 전사 인원 노출 금지 — 자기 방 참여자 수만 (대표 지시 2026-05-31)."""
    me = current_user()
    db = get_db()
    # 게스트(외부) — 전사 인원이 아니라 자기 방 참여자 전원(고객+우리직원) 수만. /api/users 게스트 분기와 동일 기준(_deleted_user 제외).
    if me and _is_guest(me) and me["guest_room_id"]:
        n = db.execute(
            "SELECT COUNT(*) AS c FROM room_members rm "
            "JOIN users u ON u.id = rm.user_id "
            "WHERE rm.room_id = ? AND u.username != '_deleted_user'",
            (me["guest_room_id"],)
        ).fetchone()["c"]
        return jsonify({"total": n, "kor": 0, "vn": 0, "other": 0, "guest": True})
    rows = db.execute("""
        SELECT department FROM users
         WHERE active = 1 AND username != '_deleted_user'
           AND COALESCE(is_guest, 0) = 0
    """).fetchall()
    kor = 0
    vn = 0
    for r in rows:
        if _user_is_vietnam(r["department"]):
            vn += 1
        else:
            kor += 1
    return jsonify({
        "total": kor + vn,
        "kor": kor,
        "vn": vn,
        "other": 0,
    })


@app.route("/api/users")
@login_required
def api_users():
    """전체 사용자 목록 — 사이드바 '👥 사용자' 탭, 멤버 초대, 멘션 자동완성 등에서 공통 사용.
    직급(title)·부서(department) 포함. 비활성(퇴사) 사용자는 active=0 으로 필터링 가능.
    게스트(외부) 사용자는 본인이 속한 방 멤버만 볼 수 있음 (대표 지시 2026-05-28)."""
    me = current_user()
    me_is_ceo = (me["role"] == "ceo") if me else False
    db = get_db()
    if me and _is_guest(me) and me["guest_room_id"]:
        # 게스트 — 자기 방 멤버만 노출
        rows = db.execute(
            "SELECT u.id, u.username, u.display_name, u.display_name_vn, u.display_name_en, u.role, u.avatar_color, u.avatar_url, u.title, u.title_en, u.department, u.department_en, u.email, u.phone, u.employee_no, u.active, "
            "COALESCE(u.ai_summary_allowed, 0) AS ai_summary_allowed, "
            "COALESCE(u.is_guest, 0) AS is_guest, u.guest_company "
            "FROM users u "
            "JOIN room_members rm ON rm.user_id = u.id "
            "WHERE u.username != '_deleted_user' AND rm.room_id = ? "
            "ORDER BY u.display_name ASC",
            (me["guest_room_id"],)
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT id, username, display_name, display_name_vn, display_name_en, role, avatar_color, avatar_url, title, title_en, department, department_en, email, phone, employee_no, active, "
            "COALESCE(ai_summary_allowed, 0) AS ai_summary_allowed, "
            "COALESCE(works_access, 0) AS works_access, "
            "COALESCE(is_guest, 0) AS is_guest, guest_company "
            "FROM users "
            "WHERE username != '_deleted_user' "   # 시스템 플레이스홀더는 디렉터리 응답에서 제외 (대표 지시 2026-05-20)
            "AND COALESCE(is_guest,0)=0 "          # 게스트(외부 고객사)는 전사 직원 디렉터리에서 제외 — 초대된 방 안에서만 표시 (대표 지시 2026-05-29)
            "ORDER BY "
            " CASE WHEN department IS NULL OR department='' THEN 1 ELSE 0 END, "
            " department ASC, "
            " CASE role WHEN 'ceo' THEN 0 ELSE 1 END, "
            " display_name ASC"
        ).fetchall()
    # 보안: role(권한) 은 본인 자신 또는 관리자만 볼 수 있게.
    # 다른 일반 사용자에게는 role 필드를 응답에서 제거 (F12 네트워크 탭으로도 못 보게).
    out = []
    for r in rows:
        d = dict(r)
        d["is_owner"] = _is_owner(d.get("username"))   # 최고관리자(소유자) 여부
        if not me_is_ceo and d["id"] != (me["id"] if me else None):
            d.pop("role", None)
            d.pop("is_owner", None)
        out.append(d)
    return jsonify(out)


@app.route("/api/users/<int:user_id>", methods=["PATCH"])
@login_required
def api_user_patch(user_id):
    """사용자 정보 수정 — 관리자(ceo)만 가능. (대표 지시 2026-05-24: 일반 사용자는 정보 수정 불가)
    body: {title?, department?, display_name?, avatar_color?}"""
    me = current_user()
    if me["role"] != "ceo":
        return jsonify({"error": "정보 수정은 관리자만 가능합니다."}), 403
    data = request.get_json(silent=True) or {}
    fields = {}
    if "title" in data:
        v = (data.get("title") or "").strip()[:40]
        fields["title"] = v or None
    if "department" in data:
        v = (data.get("department") or "").strip()[:40]
        fields["department"] = v or None
    if "email" in data:
        v = (data.get("email") or "").strip()[:100]
        # 비어있어도 OK (지우기). 값 있으면 최소 @ 포함 검증
        if v and "@" not in v:
            return jsonify({"error": "이메일 형식이 올바르지 않습니다 (@ 누락)"}), 400
        fields["email"] = v or None
    if "phone" in data:
        v = (data.get("phone") or "").strip()[:30]
        fields["phone"] = v or None
    if "employee_no" in data:
        v = (data.get("employee_no") or "").strip()[:30]
        fields["employee_no"] = v or None
    # display_name·avatar_color·role·active 는 관리자만 수정 가능
    if me["role"] == "ceo":
        # 최고관리자(소유자) 보호 — 강등·비활성 차단 (대표 지시 2026-05-21)
        _tgt = get_db().execute("SELECT username FROM users WHERE id=?", (user_id,)).fetchone()
        _target_is_owner = _tgt and _is_owner(_tgt["username"])
        if "display_name" in data:
            v = (data.get("display_name") or "").strip()[:40]
            if v:
                fields["display_name"] = v
        if "avatar_color" in data:
            v = (data.get("avatar_color") or "").strip()
            if v and len(v) <= 16:
                fields["avatar_color"] = v
        if "active" in data:
            if _target_is_owner and not data.get("active"):
                return jsonify({"error": "최고관리자 계정은 비활성화할 수 없습니다."}), 400
            fields["active"] = 1 if data.get("active") else 0
        if "works_access" in data:
            # HAIST WORKS 사용 권한 on/off — 관리자만. (대표 지시 2026-05-31)
            fields["works_access"] = 1 if data.get("works_access") else 0
        if "role" in data:
            # 관리자 선정·해지는 최고관리자(소유자)만 가능 (대표 지시 2026-05-21, 규칙 1)
            if not _is_owner(me["username"]):
                return jsonify({"error": "관리자 선정·해지는 최고관리자만 할 수 있습니다."}), 403
            new_role = (data.get("role") or "").strip().lower()
            if new_role not in ("ceo", "staff"):
                return jsonify({"error": "role 은 'ceo' 또는 'staff' 만 가능"}), 400
            if _target_is_owner and new_role != "ceo":
                return jsonify({"error": "최고관리자 계정은 강등할 수 없습니다."}), 400
            # 안전장치: 마지막 관리자가 자기 자신을 강등하려는 경우 차단
            if user_id == me["id"] and new_role == "staff":
                db_pre = get_db()
                ceo_count = db_pre.execute("SELECT COUNT(*) AS n FROM users WHERE role='ceo' AND active=1").fetchone()["n"]
                if ceo_count <= 1:
                    return jsonify({"error": "마지막 관리자는 본인을 강등할 수 없습니다. 다른 사람을 먼저 관리자로 임명한 뒤 강등하세요."}), 400
            fields["role"] = new_role
    if not fields:
        return jsonify({"error": "변경할 필드가 없습니다"}), 400
    db = get_db()
    cols = ", ".join(f"{k} = ?" for k in fields.keys())
    args = list(fields.values()) + [user_id]
    db.execute(f"UPDATE users SET {cols} WHERE id = ?", args)
    db.commit()
    # 부서·활성 변경 시 자동채널(KNK WORLD/본사/베트남) 멤버십 재동기화 (대표 지시 2026-05-20)
    if ("department" in fields) or ("active" in fields):
        try: _sync_user_auto_channels(db, user_id)
        except Exception as e: print(f"[auto_channel] patch sync 실패: {e}")
    row = db.execute(
        "SELECT id, username, display_name, display_name_vn, display_name_en, role, avatar_color, avatar_url, title, title_en, department, department_en, email, phone, employee_no, active FROM users WHERE id=?",
        (user_id,),
    ).fetchone()
    # 실시간 알림 — 다른 클라이언트도 사용자 정보 즉시 갱신
    socketio.emit("user_info_changed", dict(row))
    return jsonify({"ok": True, "user": dict(row)})


# ============================================================
# 📷 아바타 사진 업로드 (대표 지시 2026-05-19)
#   본인 또는 관리자만 업로드 가능. jpg/png/webp/gif, 5MB 이하.
#   저장: data/uploads/avatars/<user_id>.<ext>  (한 사용자 1장)
#   URL: BASE_PATH + /uploads/avatars/<user_id>.<ext>?v=<ts>  (캐시 무력화용 timestamp)
# ============================================================
AVATAR_DIR = os.path.join(UPLOAD_DIR, "avatars")
os.makedirs(AVATAR_DIR, exist_ok=True)
AVATAR_ALLOWED_EXT = {"jpg", "jpeg", "png", "webp", "gif"}
AVATAR_MAX_BYTES = 5 * 1024 * 1024  # 5MB
# 방/채널 아바타 이미지 (관리자가 채널 아이콘에 사진 설정) — data/uploads/room_avatars/<room_id>.<ext>
ROOM_AVATAR_DIR = os.path.join(UPLOAD_DIR, "room_avatars")
os.makedirs(ROOM_AVATAR_DIR, exist_ok=True)

@app.route("/api/users/<int:user_id>/avatar", methods=["POST"])
@login_required
def api_user_avatar_upload(user_id):
    """아바타 사진 업로드 — multipart/form-data 의 'file' 필드.
    권한: 관리자(ceo)만. (사진 보기는 전체 사용자, 등록·변경은 관리자만 — 대표 지시 2026-05-22)
    반환: {ok: True, avatar_url: '...?v=12345'}"""
    me = current_user()
    if me["role"] != "ceo":
        return jsonify({"error": "관리자만 사진을 등록·변경할 수 있습니다"}), 403
    if "file" not in request.files:
        return jsonify({"error": "file 필드에 이미지 첨부 필요"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "파일명 누락"}), 400
    ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
    if ext not in AVATAR_ALLOWED_EXT:
        return jsonify({"error": f"지원 형식: {', '.join(AVATAR_ALLOWED_EXT)}"}), 400
    # 크기 체크 — content_length 신뢰 안 되면 stream 으로 측정
    data = f.read()
    if len(data) > AVATAR_MAX_BYTES:
        return jsonify({"error": f"파일이 너무 큽니다 ({len(data)//1024}KB > {AVATAR_MAX_BYTES//1024//1024}MB)"}), 400
    if len(data) < 100:
        return jsonify({"error": "파일이 너무 작거나 손상되었습니다"}), 400
    # 기존 파일 (다른 확장자) 삭제
    for old_ext in AVATAR_ALLOWED_EXT:
        old_path = os.path.join(AVATAR_DIR, f"{user_id}.{old_ext}")
        if os.path.exists(old_path):
            try: os.remove(old_path)
            except Exception: pass
    # 저장
    save_path = os.path.join(AVATAR_DIR, f"{user_id}.{ext}")
    with open(save_path, "wb") as out:
        out.write(data)
    # DB 갱신 (캐시 무력화용 timestamp 쿼리)
    import time as _t
    ts = int(_t.time())
    rel_url = f"{BASE_PATH}/uploads/avatars/{user_id}.{ext}?v={ts}"
    db = get_db()
    db.execute("UPDATE users SET avatar_url=? WHERE id=?", (rel_url, user_id))
    db.commit()
    # 실시간 broadcast
    row = db.execute(
        "SELECT id, username, display_name, display_name_vn, display_name_en, role, avatar_color, avatar_url, title, title_en, department, department_en, email, phone, employee_no, active FROM users WHERE id=?",
        (user_id,),
    ).fetchone()
    socketio.emit("user_info_changed", dict(row))
    return jsonify({"ok": True, "avatar_url": rel_url})


@app.route("/api/users/<int:user_id>/avatar", methods=["DELETE"])
@login_required
def api_user_avatar_delete(user_id):
    """아바타 사진 제거 — 관리자(ceo)만. (대표 지시 2026-05-22)"""
    me = current_user()
    if me["role"] != "ceo":
        return jsonify({"error": "관리자만 사진을 삭제할 수 있습니다"}), 403
    for ext in AVATAR_ALLOWED_EXT:
        p = os.path.join(AVATAR_DIR, f"{user_id}.{ext}")
        if os.path.exists(p):
            try: os.remove(p)
            except Exception: pass
    db = get_db()
    db.execute("UPDATE users SET avatar_url=NULL WHERE id=?", (user_id,))
    db.commit()
    row = db.execute(
        "SELECT id, username, display_name, display_name_vn, display_name_en, role, avatar_color, avatar_url, title, title_en, department, department_en, email, phone, employee_no, active FROM users WHERE id=?",
        (user_id,),
    ).fetchone()
    socketio.emit("user_info_changed", dict(row))
    return jsonify({"ok": True})


# ── 본인 개인 아바타(self_avatar) — 직원 셀프 등록/삭제 (인사카드 사진 avatar_url 과 별개) (대표 지시 2026-06-03) ──
@app.route("/api/me/avatar", methods=["POST"])
@login_required
def api_me_avatar_upload():
    """본인 개인 아바타 업로드 — multipart 'file'. 인사카드 사진(관리자전용 avatar_url)은 안 건드림. 게스트 제외."""
    me = current_user()
    try:
        if me["is_guest"]:
            return jsonify({"error": "게스트는 사진을 등록할 수 없습니다"}), 403
    except Exception:
        pass
    uid = me["id"]
    if "file" not in request.files:
        return jsonify({"error": "file 필드에 이미지 첨부 필요"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "파일명 누락"}), 400
    ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
    if ext not in AVATAR_ALLOWED_EXT:
        return jsonify({"error": f"지원 형식: {', '.join(AVATAR_ALLOWED_EXT)}"}), 400
    data = f.read()
    if len(data) > AVATAR_MAX_BYTES:
        return jsonify({"error": f"파일이 너무 큽니다 ({len(data)//1024}KB)"}), 400
    if len(data) < 100:
        return jsonify({"error": "파일이 너무 작거나 손상되었습니다"}), 400
    for old_ext in AVATAR_ALLOWED_EXT:
        old_path = os.path.join(AVATAR_DIR, f"self_{uid}.{old_ext}")
        if os.path.exists(old_path):
            try: os.remove(old_path)
            except Exception: pass
    save_path = os.path.join(AVATAR_DIR, f"self_{uid}.{ext}")
    with open(save_path, "wb") as out:
        out.write(data)
    import time as _t
    ts = int(_t.time())
    rel_url = f"{BASE_PATH}/uploads/avatars/self_{uid}.{ext}?v={ts}"
    db = get_db()
    db.execute("UPDATE users SET self_avatar=? WHERE id=?", (rel_url, uid))
    db.commit()
    socketio.emit("self_avatar_changed", {"user_id": uid, "self_avatar": rel_url})
    return jsonify({"ok": True, "self_avatar": rel_url})


@app.route("/api/me/avatar", methods=["DELETE"])
@login_required
def api_me_avatar_delete():
    """본인 개인 아바타 제거 — 인사카드 사진(avatar_url)은 안 건드림."""
    me = current_user()
    uid = me["id"]
    for ext in AVATAR_ALLOWED_EXT:
        p = os.path.join(AVATAR_DIR, f"self_{uid}.{ext}")
        if os.path.exists(p):
            try: os.remove(p)
            except Exception: pass
    db = get_db()
    db.execute("UPDATE users SET self_avatar=NULL WHERE id=?", (uid,))
    db.commit()
    socketio.emit("self_avatar_changed", {"user_id": uid, "self_avatar": None})
    return jsonify({"ok": True})


@app.route("/api/me/push_hide_preview", methods=["GET", "PUT"])
@login_required
def api_me_push_hide_preview():
    """본인의 푸시 알림 내용 숨기기 옵션. 클라이언트 hidePreview 설정과 서버 동기화 (대표 지시 2026-05-26).
    GET → {hide: bool}
    PUT {hide: bool} → 저장
    """
    me = current_user()
    db = get_db()
    if request.method == "PUT":
        body = request.get_json(silent=True) or {}
        v = 1 if body.get("hide") else 0
        db.execute("UPDATE users SET push_hide_preview=? WHERE id=?", (v, me["id"]))
        db.commit()
        return jsonify({"ok": True, "hide": bool(v)})
    row = db.execute("SELECT push_hide_preview FROM users WHERE id=?", (me["id"],)).fetchone()
    return jsonify({"hide": bool(row and row["push_hide_preview"])})


@app.route("/uploads/avatars/<path:filename>")
@login_required
def serve_avatar(filename):
    """업로드된 아바타 이미지 서빙. (대표 지시 2026-05-26: 비로그인 외부 접근 차단,
    사내 직원끼리는 그대로 열람 — `<img src>` 요청에 세션 쿠키 자동 첨부됨)"""
    return send_from_directory(AVATAR_DIR, filename)


@app.route("/uploads/room_avatars/<path:filename>")
@login_required
def serve_room_avatar(filename):
    """업로드된 방/채널 아바타 이미지 서빙. (비로그인 외부 접근 차단 — 2026-05-26)"""
    return send_from_directory(ROOM_AVATAR_DIR, filename)


# ─── 시작화면(스플래시 / 광고) — 관리자(대표)가 메뉴에서 편집 (대표 지시 2026-06-06) ───
SPLASH_DIR = os.path.join(UPLOAD_DIR, "splash")

def _safe_hex_color(v, fallback):
    try:
        s = str(v or "").strip().lower()
        if len(s) == 7 and s[0] == "#" and all(c in "0123456789abcdef" for c in s[1:]):
            return s
    except Exception:
        pass
    return fallback

def _splash_defaults():
    return {
        "enabled": 1,
        "mode": "logo",                # 'logo' = 로고+글자(A) / 'full' = 전체 이미지 광고(B)
        "title": "KNK Eum",
        "tagline": "사람을 잇고, 일을 잇고, 미래를 잇다.",
        "bg_color": "#ffffff",
        "title_color": "#a5282c",
        "tagline_color": "#6b7280",
        "title_size": 26,              # 제목 글자 크기(px)
        "tagline_size": 14,            # 슬로건 글자 크기(px)
        "duration_ms": 1000,
        "image_path": "",              # uploads/splash/ 아래 파일명. 빈값 = 기본 로고(logo.png)
        # 기본 문구의 4개 언어 — 기본값도 다국어로 보이게 미리 채움(저장 안 해도 번역됨) (대표 지시 2026-06-06)
        "title_i18n": {"ko": "KNK Eum", "en": "KNK Eum", "vi": "KNK Eum", "zh": "KNK Eum"},
        "tagline_i18n": {
            "ko": "사람을 잇고, 일을 잇고, 미래를 잇다.",
            "en": "Connecting people, work, and the future.",
            "vi": "Kết nối con người, công việc và tương lai.",
            "zh": "连接人、连接工作、连接未来。",
        },
    }

def _splash_i18n_stale(text, i18n):
    """비영어 원문인데 다른 언어 슬롯이 원문 그대로면 = 아직 번역 안 됨(stale)."""
    t = (text or "").strip()
    if not t:
        return False
    src = _splash_detect_lang(t)
    if src == "en":
        return False   # 영어는 번역 안 하는 게 정상
    d = i18n if isinstance(i18n, dict) else {}
    for lg in ("ko", "en", "vi", "zh"):
        if lg == src:
            continue
        if (d.get(lg) or "") == t:
            return True
    return False

def _splash_detect_lang(text):
    """간단 판별: 한글 있으면 ko / (한글 없이) 한자 있으면 zh / 아니면 en(라틴=영문 취급)."""
    t = text or ""
    if any("가" <= c <= "힣" for c in t):
        return "ko"
    if any("一" <= c <= "鿿" for c in t):
        return "zh"
    return "en"

def _splash_i18n(text):
    """제목/슬로건을 4개 언어(ko/en/vi/zh)로. 영어 입력은 번역 없이 그대로(모두 영문).
    AI 번역 꺼짐·실패 시 원문 그대로 fallback. (대표 지시 2026-06-06)"""
    t = (text or "").strip()
    out = {"ko": t, "en": t, "vi": t, "zh": t}
    if not t:
        return out
    src = _splash_detect_lang(t)
    if src == "en":
        return out   # 영어로 입력 → 번역 없이 모두 영문
    out[src] = t
    for lang in ("ko", "en", "vi", "zh"):
        if lang == src:
            continue
        try:
            result, err = _ai_translate(t, lang)   # result = (번역문, 소스, in토큰, out토큰, 비용)
            tr = (result[0] if (result and not err) else None)
            out[lang] = (tr or t)
        except Exception:
            out[lang] = t
    return out

def _splash_raw(db):
    """저장된 설정(없으면 기본값)을 정규화해 반환 — 검증된 안전한 값만."""
    cfg = _splash_defaults()
    try:
        import json as _json
        r = db.execute("SELECT value FROM app_settings WHERE key='splash_config'").fetchone()
        if r and r["value"]:
            saved = _json.loads(r["value"])
            if isinstance(saved, dict):
                for k in cfg.keys():
                    if k in saved and saved[k] is not None:
                        cfg[k] = saved[k]
    except Exception:
        pass
    try:
        cfg["enabled"] = 1 if int(cfg.get("enabled") or 0) == 1 else 0
    except Exception:
        cfg["enabled"] = 1
    if cfg.get("mode") not in ("logo", "full"):
        cfg["mode"] = "logo"
    cfg["title"] = str(cfg.get("title") or "")[:80]
    cfg["tagline"] = str(cfg.get("tagline") or "")[:160]
    cfg["bg_color"] = _safe_hex_color(cfg.get("bg_color"), "#ffffff")
    cfg["title_color"] = _safe_hex_color(cfg.get("title_color"), "#a5282c")
    cfg["tagline_color"] = _safe_hex_color(cfg.get("tagline_color"), "#6b7280")
    try:
        cfg["title_size"] = max(12, min(64, int(cfg.get("title_size") or 26)))
    except Exception:
        cfg["title_size"] = 26
    try:
        cfg["tagline_size"] = max(9, min(40, int(cfg.get("tagline_size") or 14)))
    except Exception:
        cfg["tagline_size"] = 14
    try:
        cfg["duration_ms"] = max(500, min(10000, int(cfg.get("duration_ms") or 1000)))
    except Exception:
        cfg["duration_ms"] = 1000
    cfg["image_path"] = str(cfg.get("image_path") or "").strip()
    # 다국어(자동번역) 슬롯 정규화 — 4개 언어 dict, 비면 원문으로 fallback (대표 지시 2026-06-06)
    def _ni18n(d, fb):
        o = {}
        if not isinstance(d, dict):
            d = {}
        for lg in ("ko", "en", "vi", "zh"):
            v = d.get(lg)
            # 손상 복구: 번역함수 반환값(번역문,소스,토큰,토큰,비용)이 통째로 저장됐으면 번역문[0]만
            if isinstance(v, (list, tuple)) and v:
                v = v[0]
            elif isinstance(v, str) and v[:1] in ("[", "(") and "None," in v:
                try:
                    import ast as _ast
                    p = _ast.literal_eval(v)
                    if isinstance(p, (list, tuple)) and p:
                        v = p[0]
                except Exception:
                    pass
            o[lg] = (str(v)[:200] if v else (fb or ""))
        return o
    cfg["title_i18n"] = _ni18n(cfg.get("title_i18n"), cfg.get("title"))
    cfg["tagline_i18n"] = _ni18n(cfg.get("tagline_i18n"), cfg.get("tagline"))
    return cfg

def _get_splash_config(db):
    """렌더용 — 정규화된 설정 + 계산된 image_url/has_image."""
    cfg = _splash_raw(db)
    img = cfg.get("image_path") or ""
    if img:
        cfg["image_url"] = f"{BASE_PATH}/uploads/splash/{img}?v={STATIC_VERSION}"
        cfg["has_image"] = 1
    else:
        cfg["image_url"] = f"{BASE_PATH}/static/icons/logo.png?v={STATIC_VERSION}"
        cfg["has_image"] = 0
    return cfg

def _splash_config_safe():
    """모든 템플릿 렌더에 주입 — DB 미가용/예외시에도 기본값 반환(절대 실패 안 함)."""
    try:
        return _get_splash_config(get_db())
    except Exception:
        d = _splash_defaults()
        d["image_url"] = f"{BASE_PATH}/static/icons/logo.png?v={STATIC_VERSION}"
        d["has_image"] = 0
        return d

def _save_splash_raw(db, cfg):
    import json as _json
    raw = {k: cfg[k] for k in _splash_defaults().keys()}
    db.execute(
        "INSERT INTO app_settings (key, value) VALUES ('splash_config', ?) "
        "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (_json.dumps(raw, ensure_ascii=False),),
    )
    db.commit()


@app.route("/uploads/splash/<path:filename>")
def serve_splash_image(filename):
    """시작화면(광고) 이미지 — 로그인 전·외부 게스트 포함 모두에게 보여야 하므로 공개.
    관리자(대표)가 업로드한 이미지만 이 폴더에 들어간다. (send_from_directory 가 경로탈출 차단)"""
    return send_from_directory(SPLASH_DIR, filename)


@app.route("/api/admin/splash", methods=["GET"])
@login_required
def api_admin_splash_get():
    me = current_user()
    if not _is_maintenance_owner(me):   # 시작화면 편집 = 대표(사번5)만 (대표 지시 2026-06-06)
        abort(403)
    db = get_db()
    cur = _splash_raw(db)
    # 예전에 저장돼 아직 번역 안 된 문구가 있으면 편집기 열 때 자동 번역해 저장 (대표 지시 2026-06-06)
    healed = False
    if _splash_i18n_stale(cur.get("title"), cur.get("title_i18n")):
        cur["title_i18n"] = _splash_i18n(cur.get("title"))
        healed = True
    if _splash_i18n_stale(cur.get("tagline"), cur.get("tagline_i18n")):
        cur["tagline_i18n"] = _splash_i18n(cur.get("tagline"))
        healed = True
    if healed:
        try:
            _save_splash_raw(db, cur)
        except Exception:
            pass
    return jsonify(_get_splash_config(db))


@app.route("/api/admin/splash", methods=["POST"])
@login_required
def api_admin_splash_save():
    me = current_user()
    if not _is_maintenance_owner(me):   # 시작화면 편집 = 대표(사번5)만 (대표 지시 2026-06-06)
        abort(403)
    data = request.get_json(silent=True) or {}
    db = get_db()
    cur = _splash_raw(db)
    cfg = dict(cur)
    try:
        cfg["enabled"] = 1 if int(data.get("enabled") or 0) == 1 else 0
    except Exception:
        pass
    if data.get("mode") in ("logo", "full"):
        cfg["mode"] = data.get("mode")
    if "title" in data:
        cfg["title"] = str(data.get("title") or "")[:80]
    if "tagline" in data:
        cfg["tagline"] = str(data.get("tagline") or "")[:160]
    if "bg_color" in data:
        cfg["bg_color"] = _safe_hex_color(data.get("bg_color"), cur["bg_color"])
    if "title_color" in data:
        cfg["title_color"] = _safe_hex_color(data.get("title_color"), cur["title_color"])
    if "tagline_color" in data:
        cfg["tagline_color"] = _safe_hex_color(data.get("tagline_color"), cur["tagline_color"])
    if "title_size" in data:
        try:
            cfg["title_size"] = max(12, min(64, int(data.get("title_size") or 26)))
        except Exception:
            pass
    if "tagline_size" in data:
        try:
            cfg["tagline_size"] = max(9, min(40, int(data.get("tagline_size") or 14)))
        except Exception:
            pass
    if "duration_ms" in data:
        try:
            cfg["duration_ms"] = max(500, min(10000, int(data.get("duration_ms") or 1000)))
        except Exception:
            pass
    if data.get("clear_image"):
        cfg["image_path"] = ""
    # 제목/슬로건이 바뀌었거나 아직 번역 안 됨(stale)이면 4개 언어 재번역 (영어 입력은 그대로). (대표 지시 2026-06-06)
    if cfg.get("title") != cur.get("title") or _splash_i18n_stale(cfg.get("title"), cur.get("title_i18n")):
        cfg["title_i18n"] = _splash_i18n(cfg.get("title"))
    if cfg.get("tagline") != cur.get("tagline") or _splash_i18n_stale(cfg.get("tagline"), cur.get("tagline_i18n")):
        cfg["tagline_i18n"] = _splash_i18n(cfg.get("tagline"))
    _save_splash_raw(db, cfg)
    return jsonify({"ok": True, "splash": _get_splash_config(db)})


@app.route("/api/admin/splash/image", methods=["POST"])
@login_required
def api_admin_splash_image():
    me = current_user()
    if not _is_maintenance_owner(me):   # 시작화면 편집 = 대표(사번5)만 (대표 지시 2026-06-06)
        abort(403)
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error": "파일이 없습니다."}), 400
    ext = ext_of(f.filename)
    if not ext or ext not in ALLOWED_IMAGE_EXT:
        return jsonify({"error": "이미지 파일만 업로드할 수 있습니다 (jpg·png·gif·webp 등)."}), 400
    content_len = request.content_length or 0
    if content_len > 8 * 1024 * 1024:
        return jsonify({"error": "이미지가 너무 큽니다 (8MB 이하)."}), 413
    try:
        is_exec, why = _check_executable_magic(f)
        if is_exec:
            return jsonify({"error": "이미지 파일이 아닙니다."}), 400
    except Exception:
        pass
    os.makedirs(SPLASH_DIR, exist_ok=True)
    safe = secure_filename(f.filename) or "splash"
    if not ext_of(safe):
        safe = f"{safe}.{ext}"
    unique = f"{uuid.uuid4().hex[:12]}_{safe}"
    f.save(os.path.join(SPLASH_DIR, unique))
    db = get_db()
    cfg = _splash_raw(db)
    cfg["image_path"] = unique
    _save_splash_raw(db, cfg)
    return jsonify({"ok": True, "splash": _get_splash_config(db)})


# ──────────────────────────────────────────────────────────────
#  멘션함 API — 내게 온 @멘션 모아보기 / 읽음 처리 (대표 지시 2026-05-22)
# ──────────────────────────────────────────────────────────────
@app.route("/api/mentions")
@login_required
def api_mentions_list():
    """내게 온 멘션 목록 + 안 읽음 개수.
    쿼리: ?limit=50 (기본 50, 최대 200), ?unread=1 (안 읽은 것만)"""
    me = current_user()
    db = get_db()
    try:
        limit = int(request.args.get("limit", 50))
    except (TypeError, ValueError):
        limit = 50
    limit = max(1, min(limit, 200))
    unread_only = request.args.get("unread") in ("1", "true", "yes")
    where = "WHERE mn.mentioned_user_id = ?"
    if unread_only:
        where += " AND mn.read_at IS NULL"
    rows = db.execute(f"""
        SELECT mn.id, mn.message_id, mn.room_id, mn.created_at, mn.read_at,
               m.content, m.kind,
               r.name AS room_name, r.type AS room_type,
               s.id AS sender_id, s.display_name AS sender_name,
               s.title AS sender_title, s.department AS sender_dept
          FROM mentions mn
          JOIN messages m ON m.id = mn.message_id
          JOIN rooms r ON r.id = mn.room_id
          JOIN users s ON s.id = mn.sender_user_id
          {where}
         ORDER BY mn.id DESC
         LIMIT ?
    """, (me["id"], limit)).fetchall()
    items = []
    for r in rows:
        content = r["content"] or ""
        items.append({
            "id": r["id"],
            "message_id": r["message_id"],
            "room_id": r["room_id"],
            "room_name": r["room_name"] or "",
            "room_type": r["room_type"],
            "sender_id": r["sender_id"],
            "sender_name": r["sender_name"],
            "sender_title": r["sender_title"] or "",
            "sender_dept": r["sender_dept"] or "",
            "preview": content[:80],
            "created_at": r["created_at"],
            "read_at": r["read_at"],
        })
    unread = db.execute(
        "SELECT COUNT(*) AS c FROM mentions WHERE mentioned_user_id=? AND read_at IS NULL",
        (me["id"],),
    ).fetchone()["c"]
    return jsonify({"ok": True, "unread_count": unread, "items": items})


@app.route("/api/mentions/unread_count")
@login_required
def api_mentions_unread_count():
    """안 읽은 멘션 개수만 — 로그인 시 배지 표시용."""
    me = current_user()
    db = get_db()
    unread = db.execute(
        "SELECT COUNT(*) AS c FROM mentions WHERE mentioned_user_id=? AND read_at IS NULL",
        (me["id"],),
    ).fetchone()["c"]
    return jsonify({"ok": True, "unread_count": unread})


@app.route("/api/mentions/<int:mention_id>/read", methods=["POST"])
@login_required
def api_mention_read(mention_id):
    """멘션 하나 읽음 처리 (본인 것만)."""
    me = current_user()
    db = get_db()
    now = datetime.now(timezone.utc).isoformat()
    db.execute(
        "UPDATE mentions SET read_at=? WHERE id=? AND mentioned_user_id=? AND read_at IS NULL",
        (now, mention_id, me["id"]),
    )
    db.commit()
    unread = db.execute(
        "SELECT COUNT(*) AS c FROM mentions WHERE mentioned_user_id=? AND read_at IS NULL",
        (me["id"],),
    ).fetchone()["c"]
    return jsonify({"ok": True, "unread_count": unread})


@app.route("/api/mentions/read_all", methods=["POST"])
@login_required
def api_mentions_read_all():
    """내 멘션 모두 읽음 처리."""
    me = current_user()
    db = get_db()
    now = datetime.now(timezone.utc).isoformat()
    db.execute(
        "UPDATE mentions SET read_at=? WHERE mentioned_user_id=? AND read_at IS NULL",
        (now, me["id"]),
    )
    db.commit()
    return jsonify({"ok": True, "unread_count": 0})


@app.route("/api/mentions/<int:mention_id>", methods=["DELETE"])
@login_required
def api_mention_delete(mention_id):
    """멘션 1건을 내 멘션함에서만 제거. (대화방 메시지·멘션은 그대로 — mentions 행만 삭제)"""
    me = current_user()
    db = get_db()
    cur = db.execute(
        "DELETE FROM mentions WHERE id=? AND mentioned_user_id=?",
        (mention_id, me["id"]),
    )
    db.commit()
    unread = db.execute(
        "SELECT COUNT(*) AS c FROM mentions WHERE mentioned_user_id=? AND read_at IS NULL",
        (me["id"],),
    ).fetchone()["c"]
    return jsonify({"ok": True, "deleted": cur.rowcount, "unread_count": unread})


@app.route("/api/mentions/delete_all", methods=["POST"])
@login_required
def api_mentions_delete_all():
    """내 멘션함 전체 비우기. (대화방 메시지·멘션은 그대로 — mentions 행만 삭제)"""
    me = current_user()
    db = get_db()
    cur = db.execute("DELETE FROM mentions WHERE mentioned_user_id=?", (me["id"],))
    db.commit()
    return jsonify({"ok": True, "deleted": cur.rowcount, "unread_count": 0})


# ──────────────────────────────────────────────────────────────
#  스레드함 API — 현재 방 안의 내 스레드 활동 모아보기 (대표 지시 2026-05-27)
#  포함 범위 (옵션 B, 방 단위):
#    · 이 방 안에서 내가 답글 단 스레드
#    · 이 방 안에서 내 부모 메시지에 누군가 답글 단 스레드
# ──────────────────────────────────────────────────────────────
@app.route("/api/rooms/<int:room_id>/threads/my")
@login_required
def api_room_threads_my(room_id):
    """이 방 안의 내 스레드 활동 목록. 답글이 1개 이상 있는 스레드만 표시.
    응답 형식: { items: [{ parent_id, parent_preview, parent_author, parent_created_at,
                           reply_count, last_reply_at, last_reply_author,
                           my_role ('author'|'replier'|'both') }] }"""
    me = current_user()
    db = get_db()
    # 방 멤버여야 조회 가능
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (room_id, me["id"]),
    ).fetchone():
        abort(403)
    try:
        limit = int(request.args.get("limit", 100))
    except (TypeError, ValueError):
        limit = 100
    limit = max(1, min(limit, 300))
    mid = me["id"]
    # 내가 관여한 스레드 부모 ID 집합 (이 방 안에서):
    #   (1) 이 방 안에서 내가 답글 단 스레드 — 답글의 parent_message_id
    #   (2) 이 방 안에서 내 부모 메시지 중 답글이 1개 이상 있는 것
    rows = db.execute("""
        WITH my_thread_ids AS (
            SELECT DISTINCT parent_message_id AS pid
              FROM messages
             WHERE user_id = ? AND parent_message_id IS NOT NULL AND room_id = ?
            UNION
            SELECT id AS pid FROM messages p
             WHERE p.user_id = ? AND p.parent_message_id IS NULL AND p.room_id = ?
               AND EXISTS (SELECT 1 FROM messages c WHERE c.parent_message_id = p.id)
        ),
        thread_stats AS (
            SELECT m.parent_message_id AS pid,
                   COUNT(*) AS reply_count,
                   MAX(m.created_at) AS last_reply_at,
                   MAX(m.id) AS last_reply_id
              FROM messages m
             WHERE m.parent_message_id IN (SELECT pid FROM my_thread_ids)
             GROUP BY m.parent_message_id
        )
        SELECT mp.id AS parent_id,
               mp.content AS parent_content,
               mp.kind AS parent_kind,
               mp.created_at AS parent_created_at,
               mp.user_id AS parent_user_id,
               mp.archive_extended_until AS archive_extended_until,
               up.display_name AS parent_author,
               up.title AS parent_title,
               up.department AS parent_dept,
               ts.reply_count,
               ts.last_reply_at,
               ts.last_reply_id,
               (SELECT u.display_name FROM messages lm
                  JOIN users u ON u.id = lm.user_id
                 WHERE lm.id = ts.last_reply_id) AS last_reply_author,
               CASE
                 WHEN mp.user_id = ? AND EXISTS (
                    SELECT 1 FROM messages c WHERE c.parent_message_id = mp.id AND c.user_id = ?
                 ) THEN 'both'
                 WHEN mp.user_id = ? THEN 'author'
                 ELSE 'replier'
               END AS my_role
          FROM thread_stats ts
          JOIN messages mp ON mp.id = ts.pid
          JOIN users up ON up.id = mp.user_id
         WHERE COALESCE(mp.thread_hidden,0)=0 AND COALESCE(mp.kind,'text')!='deleted'
         ORDER BY ts.last_reply_at DESC
         LIMIT ?
    """, (mid, room_id, mid, room_id, mid, mid, mid, limit)).fetchall()

    can_manage = _can_manage_thread(db, room_id, me)  # 이 방에서의 권한 (방장/PM/관리자)
    is_ceo = _is_ceo(me)

    items = []
    for r in rows:
        # 부모 미리보기 — 제목으로 쓰이므로 200자까지 (CSS clamp 2줄)
        content = r["parent_content"] or ""
        kind = (r["parent_kind"] or "text").lower()
        if kind == "image":
            preview = content[:200] if content else "🖼 사진"
        elif kind == "file":
            preview = content[:200] if content else "📎 파일"
        elif kind == "deleted":
            preview = "(삭제된 메시지)"
        else:
            preview = content[:200]
        # 스레드 삭제/연장 가능 여부 (대표 지시 2026-05-28 옵션 ②)
        archive_ext = r["archive_extended_until"]
        deadline = _thread_archive_deadline(r["parent_created_at"], r["last_reply_at"], archive_ext)
        deletable_now = _thread_is_deletable_now(r["parent_created_at"], r["last_reply_at"], archive_ext)
        can_delete = can_manage and (is_ceo or deletable_now) and kind != "deleted"
        can_extend = can_manage and kind != "deleted"
        items.append({
            "parent_id": r["parent_id"],
            "parent_preview": preview,
            "parent_author": r["parent_author"] or "",
            "parent_title": r["parent_title"] or "",
            "parent_dept": r["parent_dept"] or "",
            "parent_user_id": r["parent_user_id"],
            "parent_created_at": r["parent_created_at"],
            "reply_count": r["reply_count"],
            "last_reply_at": r["last_reply_at"],
            "last_reply_author": r["last_reply_author"] or "",
            "my_role": r["my_role"],
            "archive_extended_until": archive_ext,
            "deletable_at": deadline,
            "deletable_now": deletable_now,
            "can_delete": can_delete,
            "can_extend": can_extend,
        })
    return jsonify({"ok": True, "items": items, "total": len(items)})


# ──────────────────────────────────────────────────────────────
#  스레드 1개 엑셀 다운로드 (대표 지시 2026-05-28)
#  · 부모 메시지 + 모든 답글을 시간순으로 정리
#  · 컬럼: 시각 / 구분(부모/답글) / 작성자 / 직급 / 부서 / 내용
#  · 귓속말은 제외 (개인 메시지)
#  · 방 멤버만 다운로드 가능
# ──────────────────────────────────────────────────────────────
@app.route("/api/threads/<int:parent_id>/export.xlsx")
@login_required
def api_thread_export_xlsx(parent_id):
    me = current_user()
    db = get_db()
    parent = db.execute("""
        SELECT m.*, u.display_name, u.title, u.department
          FROM messages m JOIN users u ON u.id=m.user_id
         WHERE m.id=?
    """, (parent_id,)).fetchone()
    if not parent:
        abort(404)
    # 방 멤버여야 다운로드 가능
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (parent["room_id"], me["id"]),
    ).fetchone():
        abort(403)

    replies = db.execute("""
        SELECT m.*, u.display_name, u.title, u.department
          FROM messages m JOIN users u ON u.id=m.user_id
         WHERE m.parent_message_id=? AND (m.whisper_to_user_id IS NULL)
         ORDER BY m.id ASC
    """, (parent_id,)).fetchall()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        return jsonify({"error": "openpyxl 미설치"}), 500
    from io import BytesIO
    from flask import send_file

    # 사진 임베드용 (Pillow + openpyxl drawing) — 없으면 파일명만 표시 (graceful) (대표 지시 2026-06-04)
    try:
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage
        _img_ok = True
    except Exception:
        _img_ok = False
    _bio_keep = []          # 임베드 BytesIO 참조 유지 (save 전 GC 방지)
    _embed_state = {"n": 0}  # 임베드 장수 (과다 방지 캡)

    from datetime import timedelta as _td_kst
    def _kst_min(s):
        # UTC 저장(공백/ISO T 혼용) → 한국시간 'YYYY-MM-DD HH:MM' (대표 지시 2026-06-04)
        if not s:
            return ""
        try:
            core = str(s).strip().replace("T", " ")[:19]
            return (datetime.strptime(core, "%Y-%m-%d %H:%M:%S") + _td_kst(hours=9)).strftime("%Y-%m-%d %H:%M")
        except Exception:
            return str(s)[:16].replace("T", " ")

    wb = Workbook()
    ws = wb.active
    ws.title = "스레드"

    bold_hdr = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    parent_fill = PatternFill("solid", fgColor="FEF3C7")  # 노랑 (부모 강조)
    wrap = Alignment(wrap_text=True, vertical="top")

    headers = ["시각", "구분", "작성자", "직급", "부서", "내용", "사진"]
    widths  = [20, 8, 18, 14, 24, 80, 24]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = bold_hdr
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    def _content_for(msg):
        kind = (msg["kind"] or "text").lower()
        content = msg["content"] or ""
        if kind == "deleted":
            return "(삭제된 메시지)"
        if kind == "image":
            return f"🖼 사진 — {content}" if content else "🖼 사진"
        if kind == "file":
            fn = msg["file_name"] if "file_name" in msg.keys() else ""
            base = f"📎 파일 — {fn or content}"
            return base
        if kind == "system":
            return f"[시스템] {content}"
        return content

    def _add_row(idx, msg, role_label, fill=None):
        when = _kst_min(msg["created_at"])
        cells = [
            when,
            role_label,
            msg["display_name"] or "",
            msg["title"] or "",
            msg["department"] or "",
            _content_for(msg),
        ]
        for i, v in enumerate(cells, 1):
            c = ws.cell(row=idx, column=i, value=v)
            c.alignment = wrap
            if fill:
                c.fill = fill

    def _embed_photo(rownum, msg):
        # 사진 메시지면 실제 이미지를 '사진'(7번) 열에 축소 임베드 (대표 지시 2026-06-04)
        if not _img_ok or _embed_state["n"] >= 200:
            return
        if (msg["kind"] or "").lower() != "image":
            return
        fp = msg["file_path"] if "file_path" in msg.keys() else None
        if not fp:
            return
        src = os.path.join(UPLOAD_DIR, fp)
        if not os.path.exists(src):
            return
        try:
            pim = PILImage.open(src)
            try:
                from PIL import ImageOps
                pim = ImageOps.exif_transpose(pim)   # 폰 사진 회전 보정
            except Exception:
                pass
            pim = pim.convert("RGB")
            pim.thumbnail((150, 150))
            bio = BytesIO()
            pim.save(bio, format="PNG")
            bio.seek(0)
            _bio_keep.append(bio)
            ws.add_image(XLImage(bio), "%s%d" % (get_column_letter(7), rownum))
            _h = pim.height * 0.75 + 6
            if _h > (ws.row_dimensions[rownum].height or 18):
                ws.row_dimensions[rownum].height = _h
            _embed_state["n"] += 1
        except Exception:
            pass

    _add_row(2, parent, "부모", parent_fill)
    _embed_photo(2, parent)
    for i, r in enumerate(replies, start=3):
        _add_row(i, r, "답글")
        _embed_photo(i, r)

    # 메타 시트
    ws2 = wb.create_sheet("정보")
    info_rows = [
        ("스레드 ID", parent_id),
        ("부모 메시지 작성자", parent["display_name"] or ""),
        ("부모 메시지 시각", _kst_min(parent["created_at"])),
        ("답글 수", len(replies)),
        ("내보낸 사람", me["display_name"] or ""),
        ("내보낸 시각", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for i, (k, v) in enumerate(info_rows, 1):
        ws2.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 40

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"thread_{parent_id}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ──────────────────────────────────────────────────────────────
#  스레드함 API — 이 방 안의 "전체" 스레드 (대표 지시 2026-05-28)
#  · 방 멤버라면 누구나 조회 가능 (특별 권한 제한 없음)
#  · 답글이 1개 이상 있는 모든 스레드를 최신 활동순으로 반환
#  · my_role 에 'none' 케이스 추가 (내가 관여 안 한 스레드)
# ──────────────────────────────────────────────────────────────
@app.route("/api/rooms/<int:room_id>/threads/all")
@login_required
def api_room_threads_all(room_id):
    """이 방 안의 모든 스레드 목록. 답글이 1개 이상 있는 스레드만 표시.
    응답 형식은 /threads/my 와 동일하되 my_role 에 'none' 추가."""
    me = current_user()
    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (room_id, me["id"]),
    ).fetchone():
        abort(403)
    try:
        limit = int(request.args.get("limit", 200))
    except (TypeError, ValueError):
        limit = 200
    limit = max(1, min(limit, 500))
    mid = me["id"]
    rows = db.execute("""
        WITH thread_stats AS (
            SELECT m.parent_message_id AS pid,
                   COUNT(*) AS reply_count,
                   MAX(m.created_at) AS last_reply_at,
                   MAX(m.id) AS last_reply_id
              FROM messages m
              JOIN messages p ON p.id = m.parent_message_id
             WHERE m.parent_message_id IS NOT NULL AND p.room_id = ?
             GROUP BY m.parent_message_id
        )
        SELECT mp.id AS parent_id,
               mp.content AS parent_content,
               mp.kind AS parent_kind,
               mp.created_at AS parent_created_at,
               mp.user_id AS parent_user_id,
               mp.archive_extended_until AS archive_extended_until,
               up.display_name AS parent_author,
               up.title AS parent_title,
               up.department AS parent_dept,
               ts.reply_count,
               ts.last_reply_at,
               ts.last_reply_id,
               (SELECT u.display_name FROM messages lm
                  JOIN users u ON u.id = lm.user_id
                 WHERE lm.id = ts.last_reply_id) AS last_reply_author,
               CASE
                 WHEN mp.user_id = ? AND EXISTS (
                    SELECT 1 FROM messages c WHERE c.parent_message_id = mp.id AND c.user_id = ?
                 ) THEN 'both'
                 WHEN mp.user_id = ? THEN 'author'
                 WHEN EXISTS (
                    SELECT 1 FROM messages c WHERE c.parent_message_id = mp.id AND c.user_id = ?
                 ) THEN 'replier'
                 ELSE 'none'
               END AS my_role
          FROM thread_stats ts
          JOIN messages mp ON mp.id = ts.pid
          JOIN users up ON up.id = mp.user_id
         WHERE COALESCE(mp.thread_hidden,0)=0 AND COALESCE(mp.kind,'text')!='deleted'
         ORDER BY ts.last_reply_at DESC
         LIMIT ?
    """, (room_id, mid, mid, mid, mid, limit)).fetchall()

    can_manage = _can_manage_thread(db, room_id, me)
    is_ceo = _is_ceo(me)

    items = []
    for r in rows:
        content = r["parent_content"] or ""
        kind = (r["parent_kind"] or "text").lower()
        if kind == "image":
            preview = content[:200] if content else "🖼 사진"
        elif kind == "file":
            preview = content[:200] if content else "📎 파일"
        elif kind == "deleted":
            preview = "(삭제된 메시지)"
        else:
            preview = content[:200]
        archive_ext = r["archive_extended_until"]
        deadline = _thread_archive_deadline(r["parent_created_at"], r["last_reply_at"], archive_ext)
        deletable_now = _thread_is_deletable_now(r["parent_created_at"], r["last_reply_at"], archive_ext)
        can_delete = can_manage and (is_ceo or deletable_now) and kind != "deleted"
        can_extend = can_manage and kind != "deleted"
        items.append({
            "parent_id": r["parent_id"],
            "parent_preview": preview,
            "parent_author": r["parent_author"] or "",
            "parent_title": r["parent_title"] or "",
            "parent_dept": r["parent_dept"] or "",
            "parent_user_id": r["parent_user_id"],
            "parent_created_at": r["parent_created_at"],
            "reply_count": r["reply_count"],
            "last_reply_at": r["last_reply_at"],
            "last_reply_author": r["last_reply_author"] or "",
            "my_role": r["my_role"],
            "archive_extended_until": archive_ext,
            "deletable_at": deadline,
            "deletable_now": deletable_now,
            "can_delete": can_delete,
            "can_extend": can_extend,
        })
    return jsonify({"ok": True, "items": items, "total": len(items)})


# ══════════════════════════════════════════════════════════════
#  고객사 게스트 초대 API (대표 지시 2026-05-28)
#    · 방장/PM/관리자 → 외부인을 메신저 방에 초대
#    · 휴대폰 번호 매칭으로 인증
#    · 만료 전 다회 재로그인 가능
#    · 외부 사용자는 그 방 외 다른 기능 차단 (별도 권한 가드)
# ══════════════════════════════════════════════════════════════
import secrets as _secrets_for_token

def _normalize_phone(phone):
    """휴대폰 번호 정규화 — 010-1234-5678 / 01012345678 / +821012345678 → 모두 01012345678 로."""
    if not phone:
        return ""
    s = re.sub(r"[\s\-\(\)]", "", str(phone))
    if s.startswith("+82"):
        s = "0" + s[3:]
    elif s.startswith("82") and len(s) >= 12:
        s = "0" + s[2:]
    return s


@app.route("/api/rooms/<int:room_id>/guest_invites", methods=["POST"])
@login_required
def api_guest_invite_create(room_id):
    """고객사 초대 발행. body: {name, company, phone, expires_days(1|3|7|30|0)} (0=무제한)."""
    me = current_user()
    db = get_db()
    if not _can_invite_guest(db, room_id, me):
        return jsonify({"error": "방장 / PM / 관리자만 고객사를 초대할 수 있습니다."}), 403
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    company = (data.get("company") or "").strip()
    title = (data.get("title") or "").strip()             # 직책(선택) (대표 지시 2026-05-28)
    department = (data.get("department") or "").strip()   # 부서명(선택) (대표 지시 2026-05-30)
    duty = (data.get("duty") or "").strip()               # 담당업무(선택)
    email = (data.get("email") or "").strip()             # 이메일(선택)
    note = (data.get("note") or "").strip()               # 참고사항(선택)
    # 고객사 로고(선택) — data URL(base64). 형식·크기(<=300KB) 안 맞으면 무시하고 로고 없이 진행. (대표 지시 2026-05-31)
    logo = (data.get("logo") or "").strip()
    if logo and not (logo.startswith("data:image/") and len(logo) <= 300000):
        logo = ""
    # 고객사 언어(선택, 기본 ko) — 초대메시지·입장페이지·대화방 기본 언어 (대표 지시 2026-05-31)
    lang = (data.get("lang") or "ko").strip().lower()
    if lang not in UI_LANGS:
        lang = "ko"
    phone = _normalize_phone(data.get("phone") or "")
    expires_days_raw = data.get("expires_days", 7)
    try:
        expires_days = int(expires_days_raw)
    except Exception:
        expires_days = 7
    if expires_days not in (1, 3, 7, 30, 0):
        expires_days = 7
    if not name or not company or not phone:
        return jsonify({"error": "고객명·회사명·전화번호 모두 필수입니다."}), 400
    # 전화번호 — 휴대폰/유선 모두 허용(단, 유선은 게스트 본인인증 매칭 불가). 정규화 후 숫자 8~12자리. (대표 지시 2026-05-30)
    if not re.match(r"^\d{8,12}$", phone):
        return jsonify({"error": "전화번호 형식이 올바르지 않습니다 (숫자만, 예: 010-1234-5678)."}), 400
    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    # 중복 방지 — 같은 방·같은 번호·'같은 이름'의 미사용·미회수·미만료 초대만 재사용.
    #   전화번호가 같아도 이름이 다르면 별개 고객으로 보고 새로 발행 (대표 지시 2026-05-30).
    existing = db.execute("""
        SELECT * FROM guest_invites
         WHERE room_id=? AND guest_phone=? AND guest_name=? AND revoked_at IS NULL AND guest_user_id IS NULL
           AND (expires_at IS NULL OR expires_at > ?)
         ORDER BY id DESC LIMIT 1
    """, (room_id, phone, name, now)).fetchone()
    if existing:
        ex_url = f"{request.host_url.rstrip('/')}{BASE_PATH}/g/{existing['token']}"
        return jsonify({
            "ok": True, "reused": True,
            "id": existing["id"], "token": existing["token"], "url": ex_url,
            "name": existing["guest_name"], "company": existing["guest_company"],
            "title": existing["guest_title"] or "",
            "department": (existing["guest_department"] if "guest_department" in existing.keys() else "") or "",
            "duty": (existing["guest_duty"] if "guest_duty" in existing.keys() else "") or "",
            "email": (existing["guest_email"] if "guest_email" in existing.keys() else "") or "",
            "note": (existing["guest_note"] if "guest_note" in existing.keys() else "") or "",
            "phone": existing["guest_phone"],
            "lang": (existing["guest_lang"] if "guest_lang" in existing.keys() else "") or "ko",
            "expires_at": existing["expires_at"],
        })
    # 토큰 (24 chars URL-safe)
    token = _secrets_for_token.token_urlsafe(18)[:24]
    if expires_days == 0:
        expires_at = None
    else:
        expires_at = (datetime.now(timezone.utc) + _timedelta(days=expires_days)).isoformat(timespec="seconds")
    cur = db.execute("""
        INSERT INTO guest_invites
          (token, room_id, invited_by_user_id, guest_name, guest_company, guest_title,
           guest_department, guest_duty, guest_email, guest_note,
           guest_phone, expires_at, created_at, guest_company_logo, guest_lang)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (token, room_id, me["id"], name, company, (title or None),
          (department or None), (duty or None), (email or None), (note or None),
          phone, expires_at, now, (logo or None), lang))
    db.commit()
    invite_id = cur.lastrowid
    # 중국어 고객을 초대한 방이면 방 이름 zh 번역 생성 (없을 때만) — (대표 지시 2026-06-05)
    if lang == "zh":
        try:
            _rz = db.execute("SELECT name_zh FROM rooms WHERE id=?", (room_id,)).fetchone()
            if _rz and not (_rz["name_zh"] if "name_zh" in _rz.keys() else None):
                _retranslate_room(db, room_id, langs=["zh"])
        except Exception:
            pass
    # 초대 URL — 게스트 로그인 페이지 (Phase 3 에서 구현 예정)
    invite_url = f"{request.host_url.rstrip('/')}{BASE_PATH}/g/{token}"
    return jsonify({
        "ok": True,
        "id": invite_id,
        "token": token,
        "url": invite_url,
        "name": name,
        "company": company,
        "title": title,
        "department": department,
        "duty": duty,
        "email": email,
        "note": note,
        "phone": phone,
        "lang": lang,
        "expires_at": expires_at,
        "expires_days": expires_days,
    })


@app.route("/api/rooms/<int:room_id>/guest_invites", methods=["GET"])
@login_required
def api_guest_invite_list(room_id):
    """그 방의 활성 게스트 초대 목록 — 방장/PM/관리자만."""
    me = current_user()
    db = get_db()
    if not _can_invite_guest(db, room_id, me):
        return jsonify({"error": "권한 없음"}), 403
    rows = db.execute("""
        SELECT gi.id, gi.token, gi.guest_name, gi.guest_company, gi.guest_title,
               gi.guest_department, gi.guest_duty, gi.guest_email, gi.guest_note, gi.guest_phone,
               gi.guest_lang,
               gi.expires_at, gi.first_used_at, gi.last_used_at, gi.created_at,
               gi.revoked_at, gi.guest_user_id,
               u.display_name AS invited_by_name
          FROM guest_invites gi
          LEFT JOIN users u ON u.id = gi.invited_by_user_id
         WHERE gi.room_id = ? AND gi.revoked_at IS NULL
         ORDER BY gi.id DESC
    """, (room_id,)).fetchall()
    items = []
    now_iso = datetime.now(timezone.utc).isoformat(timespec="seconds")
    for r in rows:
        d = dict(r)
        # 만료 여부
        d["expired"] = (d["expires_at"] is not None and d["expires_at"] < now_iso)
        d["url"] = f"{request.host_url.rstrip('/')}{BASE_PATH}/g/{d['token']}"
        items.append(d)
    return jsonify({"ok": True, "items": items})


# ──────────────────────────────────────────────────────────────
#  게스트 로그인 페이지 + 인증 (대표 지시 2026-05-28)
#    · GET /g/<token>                       — 휴대폰 번호 입력 화면
#    · POST /api/guest_invites/<token>/auth — 인증 + 자동 계정 생성 + 방 입장
# ──────────────────────────────────────────────────────────────
@app.route("/g/<token>", methods=["GET"])
def guest_login_page(token):
    # 이미 로그인된 게스트가 앱 재실행 등으로 자기 초대 페이지에 오면 바로 대화방으로. (대표 지시 2026-05-31)
    #   (설치앱 start_url = /g/토큰 이므로 재실행 시 여기로 들어옴 — 세션 살아있으면 로그인 화면 건너뜀)
    if session.get("user_id"):
        try:
            _g = get_db().execute(
                "SELECT COALESCE(is_guest,0) AS g FROM users WHERE id=?", (session["user_id"],)
            ).fetchone()
            if _g and int(_g["g"] or 0) == 1:
                return redirect(url_for("chat") + f"?v={STATIC_VERSION}")
        except Exception:
            pass
    db = get_db()
    inv = db.execute("""
        SELECT gi.*, r.name AS room_name,
               u.display_name AS inviter_name, u.title AS inviter_title
          FROM guest_invites gi
          LEFT JOIN rooms r ON r.id = gi.room_id
          LEFT JOIN users u ON u.id = gi.invited_by_user_id
         WHERE gi.token = ?
    """, (token,)).fetchone()
    revoked = False
    expired = False
    room_name = ""
    if inv:
        revoked = bool(inv["revoked_at"])
        room_name = inv["room_name"] or ""
        if inv["expires_at"]:
            try:
                now_iso = datetime.now(timezone.utc).isoformat(timespec="seconds")
                expired = (inv["expires_at"] < now_iso)
            except Exception:
                expired = False
    return render_template("guest_login.html",
                           invite=dict(inv) if inv else None,
                           revoked=revoked,
                           expired=expired,
                           room_name=room_name,
                           token=token,
                           base_path=BASE_PATH)


@app.route("/api/guest_invites/<token>/auth", methods=["POST"])
def api_guest_invite_auth(token):
    """게스트 인증 — 휴대폰 매칭 → 게스트 user 생성/조회 → 세션 부여 → 방 멤버 추가."""
    db = get_db()
    inv = db.execute("SELECT * FROM guest_invites WHERE token=?", (token,)).fetchone()
    if not inv:
        return jsonify({"error": "유효하지 않은 초대 링크"}), 404
    if inv["revoked_at"]:
        return jsonify({"error": "회수된 초대"}), 403
    if inv["expires_at"]:
        try:
            now_iso = datetime.now(timezone.utc).isoformat(timespec="seconds")
            if inv["expires_at"] < now_iso:
                return jsonify({"error": "만료된 초대"}), 403
        except Exception:
            pass
    data = request.get_json(silent=True) or {}
    phone_in = _normalize_phone(data.get("phone") or "")
    if not phone_in:
        return jsonify({"error": "휴대폰 번호를 입력하세요."}), 400
    if phone_in != _normalize_phone(inv["guest_phone"]):
        return jsonify({"error": "등록된 휴대폰 번호와 일치하지 않습니다."}), 403

    room_id = inv["room_id"]
    now = datetime.now(timezone.utc).isoformat(timespec="seconds")

    # 게스트 user — 이미 있으면 재사용, 없으면 새로 생성
    guest_user_id = inv["guest_user_id"]
    if guest_user_id:
        user_row = db.execute("SELECT * FROM users WHERE id=?", (guest_user_id,)).fetchone()
    else:
        user_row = None

    if not user_row:
        # 새 게스트 계정 생성
        # username 은 토큰 prefix + 짧은 식별자
        gusername = f"guest_{token[:12]}"
        # 충돌 회피 — 같은 username 있으면 suffix
        suffix = 0
        while db.execute("SELECT 1 FROM users WHERE username=?", (gusername,)).fetchone():
            suffix += 1
            gusername = f"guest_{token[:12]}_{suffix}"
        # 비밀번호 — 랜덤 (로그인은 토큰 인증으로만)
        dummy_pw = _secrets_for_token.token_hex(16)
        try:
            pw_hash = generate_password_hash(dummy_pw)
        except Exception:
            pw_hash = ""
        _gdept = (inv["guest_department"] if "guest_department" in inv.keys() else None) or None
        cur = db.execute("""
            INSERT INTO users
              (username, password_hash, display_name, title, department, avatar_color, role,
               is_guest, guest_room_id, guest_company, phone, ai_summary_allowed,
               active, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            gusername, pw_hash, inv["guest_name"], inv["guest_title"] or "", _gdept,
            "#8B5CF6",  # 보라 (게스트 식별)
            "guest", 1, room_id, inv["guest_company"], inv["guest_phone"],
            0, 1, now,
        ))
        db.commit()
        guest_user_id = cur.lastrowid
        # 초대에 guest_user_id 기록
        db.execute("UPDATE guest_invites SET guest_user_id=?, first_used_at=?, last_used_at=? WHERE id=?",
                   (guest_user_id, now, now, inv["id"]))
        db.commit()
        # 방 멤버로 추가
        try:
            db.execute("INSERT OR IGNORE INTO room_members (room_id, user_id, role, joined_at) VALUES (?, ?, ?, ?)",
                       (room_id, guest_user_id, "member", now))
            db.commit()
        except Exception:
            pass
        # 시스템 메시지 — 새 게스트 입장 안내
        try:
            sys_msg = f"🤝 {inv['guest_name']} ({inv['guest_company']}) 님이 입장했습니다."
            cs = db.execute("""
                INSERT INTO messages (room_id, user_id, content, kind, created_at)
                VALUES (?, ?, ?, 'system', ?)
            """, (room_id, guest_user_id, sys_msg, now))
            db.commit()
            socketio.emit("new_message", {
                "id": cs.lastrowid, "room_id": room_id, "user_id": guest_user_id,
                "display_name": inv["guest_name"], "avatar_color": "#8B5CF6",
                "content": sys_msg, "kind": "system", "created_at": now,
            }, to=f"room_{room_id}")
        except Exception:
            pass
    else:
        # 재로그인 — last_used_at 갱신 + 직책·부서·회사 동기화(초대 정보 기준) (대표 지시 2026-05-30)
        db.execute("UPDATE guest_invites SET last_used_at=? WHERE id=?", (now, inv["id"]))
        _gdept = (inv["guest_department"] if "guest_department" in inv.keys() else None) or None
        try:
            db.execute(
                "UPDATE users SET title=?, department=?, guest_company=? WHERE id=? AND COALESCE(is_guest,0)=1",
                (inv["guest_title"] or "", _gdept, inv["guest_company"], guest_user_id),
            )
        except Exception:
            pass
        db.commit()
        # 방 멤버 보장 (만약 빠져있다면 다시 추가)
        if not db.execute("SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
                          (room_id, guest_user_id)).fetchone():
            try:
                db.execute("INSERT INTO room_members (room_id, user_id, role, joined_at) VALUES (?, ?, ?, ?)",
                           (room_id, guest_user_id, "member", now))
                db.commit()
            except Exception:
                pass

    # 입장 페이지에서 선택한 언어를 게스트 화면 언어(ui_lang)로 저장 → 채팅 메뉴도 그 언어로 표시.
    #   (대표 지시 2026-05-31: 고객사가 中文 선택하면 대화방 메뉴도 중국어)
    _glang = (data.get("lang") or "").strip().lower()
    if _glang in UI_LANGS:
        try:
            db.execute("UPDATE users SET ui_lang=? WHERE id=? AND COALESCE(is_guest,0)=1",
                       (_glang, guest_user_id))
            db.commit()
        except Exception:
            pass

    # 세션 부여 — UA 기반 device_type + 토큰 (일반 로그인과 동일 패턴)
    ua = request.headers.get("User-Agent", "") if request else ""
    dtype = _device_type_from_ua(ua)
    tok = _secrets_for_token.token_urlsafe(24)
    session["user_id"] = guest_user_id
    session["device_type"] = dtype
    session["sess_token"] = tok
    session.permanent = True
    try:
        _upsert_active_session(guest_user_id, dtype, tok, ua)
    except Exception:
        pass

    resp = jsonify({"ok": True, "room_id": room_id, "user_id": guest_user_id})
    # 게스트 기기 표식 쿠키 — 이후 /login 으로 와도(뒤로가기·PWA 재실행 등) 직원 로그인 대신
    #  자기 초대 페이지(/g/토큰)로 보내기 위함. 게스트 기기엔 직원 로그인 화면이 절대 안 보이게. (대표 지시 2026-05-30)
    try:
        resp.set_cookie("knk_gt", token, max_age=60 * 24 * 3600, path="/",
                        httponly=True, samesite="Lax", secure=IS_PRODUCTION)
    except Exception:
        pass
    return resp


def _emit_guest_exit_msg(db, room_id, guest_user_id, guest_name, guest_company):
    """게스트 회수/내보내기 시 방에 '나갔습니다' 안내 — 입장 '🤝 … 입장했습니다' 와 대칭. (대표 지시 2026-05-30)"""
    try:
        now = datetime.now(timezone.utc).isoformat(timespec="seconds")
        comp = f" ({guest_company})" if guest_company else ""
        sys_msg = f"🤝 {guest_name or '외부 사용자'}{comp} 님이 방에서 나갔습니다 (초대 회수)."
        cs = db.execute(
            "INSERT INTO messages (room_id, user_id, content, kind, created_at) VALUES (?, ?, ?, 'system', ?)",
            (room_id, guest_user_id, sys_msg, now),
        )
        db.commit()
        socketio.emit("new_message", {
            "id": cs.lastrowid, "room_id": room_id, "user_id": guest_user_id,
            "display_name": guest_name or "외부 사용자", "avatar_color": "#8B5CF6",
            "content": sys_msg, "kind": "system", "created_at": now,
        }, to=f"room_{room_id}")
    except Exception:
        pass


@app.route("/api/guest_invites/<int:invite_id>", methods=["DELETE"])
@login_required
def api_guest_invite_revoke(invite_id):
    """초대 회수 — 방장/PM/관리자만. 회수 후 같은 토큰으로 입장 불가."""
    me = current_user()
    db = get_db()
    inv = db.execute("SELECT * FROM guest_invites WHERE id=?", (invite_id,)).fetchone()
    if not inv:
        return jsonify({"error": "초대를 찾을 수 없습니다."}), 404
    if not _can_invite_guest(db, inv["room_id"], me):
        return jsonify({"error": "권한 없음"}), 403
    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    db.execute("UPDATE guest_invites SET revoked_at=? WHERE id=?", (now, invite_id))
    # 회수 = 이미 입장한 게스트도 방에서 내보내고 계정 비활성 (대표 지시 2026-05-29)
    guid = inv["guest_user_id"]
    was_member = False
    if guid:
        was_member = bool(db.execute(
            "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?", (inv["room_id"], guid)
        ).fetchone())
        db.execute("DELETE FROM room_members WHERE room_id=? AND user_id=?", (inv["room_id"], guid))
        # 이 게스트가 다른 활성(미회수) 초대가 더 없으면 계정 비활성화
        other = db.execute(
            "SELECT 1 FROM guest_invites WHERE guest_user_id=? AND id!=? AND revoked_at IS NULL LIMIT 1",
            (guid, invite_id),
        ).fetchone()
        if not other:
            db.execute("UPDATE users SET active=0 WHERE id=? AND COALESCE(is_guest,0)=1", (guid,))
    db.commit()
    # 실제로 방에 있던 게스트만 '나갔습니다' 안내 (입장 안내와 대칭)
    if was_member:
        _emit_guest_exit_msg(db, inv["room_id"], guid, inv["guest_name"], inv["guest_company"])
    return jsonify({"ok": True})


@app.route("/api/rooms/<int:room_id>/guest_members/<int:user_id>", methods=["DELETE"])
@login_required
def api_room_guest_member_remove(room_id, user_id):
    """방 멤버 목록에서 게스트 직접 내보내기 (대표 지시 2026-05-30) — 방장/PM/관리자만.
    초대 상태가 꼬여(전부 회수 등) 회수 버튼이 없어도 방장이 직접 정리할 수 있는 수동 레버.
    효과: 방에서 제거 + 그 방의 해당 게스트 초대 모두 회수 + (다른 미회수 초대 없으면) 계정 비활성."""
    me = current_user()
    db = get_db()
    if not _can_invite_guest(db, room_id, me):
        return jsonify({"error": "권한 없음"}), 403
    target = db.execute(
        "SELECT id, display_name, guest_company, COALESCE(is_guest,0) AS is_guest FROM users WHERE id=?",
        (user_id,),
    ).fetchone()
    if not target:
        return jsonify({"error": "사용자를 찾을 수 없습니다."}), 404
    if not target["is_guest"]:
        return jsonify({"error": "외부(게스트) 사용자만 내보낼 수 있습니다."}), 400
    was_member = bool(db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?", (room_id, user_id)
    ).fetchone())
    # 1) 방에서 제거
    db.execute("DELETE FROM room_members WHERE room_id=? AND user_id=?", (room_id, user_id))
    # 2) 이 방의 해당 게스트 초대 모두 회수 (재입장 차단)
    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    db.execute(
        "UPDATE guest_invites SET revoked_at=? WHERE room_id=? AND guest_user_id=? AND revoked_at IS NULL",
        (now, room_id, user_id),
    )
    # 3) 다른 방에도 미회수 초대가 없으면 계정 비활성
    other = db.execute(
        "SELECT 1 FROM guest_invites WHERE guest_user_id=? AND revoked_at IS NULL LIMIT 1",
        (user_id,),
    ).fetchone()
    if not other:
        db.execute("UPDATE users SET active=0 WHERE id=? AND COALESCE(is_guest,0)=1", (user_id,))
    db.commit()
    # 방에 '나갔습니다' 안내 (입장 안내와 대칭) — 실제로 방에 있던 게스트만
    if was_member:
        _emit_guest_exit_msg(db, room_id, user_id, target["display_name"], target["guest_company"])
    try:
        _emit_room_event(room_id, "member_removed", {"room_id": room_id, "user_id": user_id})
    except Exception:
        pass
    return jsonify({"ok": True, "removed": target["display_name"]})


# ──────────────────────────────────────────────────────────────
#  스레드 연장 — POST /api/threads/<id>/extend (대표 지시 2026-05-28)
#    · 권한: 방장 / PM / 관리자
#    · 효과: archive_extended_until = now + 30일
#    · 부모 메시지의 archive_extended_until 만 갱신
# ──────────────────────────────────────────────────────────────
@app.route("/api/threads/<int:parent_id>/extend", methods=["POST"])
@login_required
def api_thread_extend(parent_id):
    me = current_user()
    db = get_db()
    parent = db.execute(
        "SELECT id, room_id, kind FROM messages WHERE id=?", (parent_id,)
    ).fetchone()
    if not parent:
        return jsonify({"error": "스레드를 찾을 수 없습니다."}), 404
    if (parent["kind"] or "").lower() == "deleted":
        return jsonify({"error": "이미 삭제된 스레드입니다."}), 400
    if not _can_manage_thread(db, parent["room_id"], me):
        return jsonify({"error": "방장 / PM / 관리자만 보관 연장할 수 있습니다."}), 403
    new_deadline = (datetime.now() + _timedelta(days=THREAD_ARCHIVE_DAYS)).strftime("%Y-%m-%dT%H:%M:%S")
    db.execute("UPDATE messages SET archive_extended_until=? WHERE id=?", (new_deadline, parent_id))
    db.commit()
    return jsonify({"ok": True, "archive_extended_until": new_deadline, "days": THREAD_ARCHIVE_DAYS})


# ──────────────────────────────────────────────────────────────
#  스레드 삭제(숨김) — DELETE /api/threads/<id> (대표 지시 2026-05-29 재정의)
#    · 권한: 방장 / PM / 관리자
#    · 조건: 마지막 답글 후 30일 경과 (관리자는 무시)
#    · 방식: 스레드 목록에서만 숨김(thread_hidden=1). 부모·답글·번역 내용은 그대로 보존.
#    · 채팅 대화에는 메시지가 그대로 남음 ("삭제"는 목록에서 안 보이게 하는 것일 뿐)
# ──────────────────────────────────────────────────────────────
@app.route("/api/threads/<int:parent_id>", methods=["DELETE"])
@login_required
def api_thread_delete(parent_id):
    me = current_user()
    db = get_db()
    parent = db.execute(
        "SELECT id, room_id, kind, created_at, archive_extended_until, COALESCE(thread_hidden,0) AS thread_hidden FROM messages WHERE id=?",
        (parent_id,),
    ).fetchone()
    if not parent:
        return jsonify({"error": "스레드를 찾을 수 없습니다."}), 404
    if int(parent["thread_hidden"] or 0) == 1:
        return jsonify({"error": "이미 목록에서 숨긴 스레드입니다."}), 400
    if not _can_manage_thread(db, parent["room_id"], me):
        return jsonify({"error": "방장 / PM / 관리자만 스레드를 삭제할 수 있습니다."}), 403
    # 마지막 답글 시각 계산
    last_reply_at = db.execute(
        "SELECT MAX(created_at) AS last FROM messages WHERE parent_message_id=?",
        (parent_id,),
    ).fetchone()["last"]
    # 관리자(ceo) 는 기간 무시
    if not _is_ceo(me):
        if not _thread_is_deletable_now(parent["created_at"], last_reply_at, parent["archive_extended_until"]):
            return jsonify({
                "error": f"이 스레드는 아직 보관 기간({THREAD_ARCHIVE_DAYS}일) 이 지나지 않았습니다.",
                "deletable_at": _thread_archive_deadline(parent["created_at"], last_reply_at, parent["archive_extended_until"]),
            }), 403
    # 목록에서만 숨김 — 부모 메시지에 thread_hidden=1 만 설정. 내용·답글·번역 전부 보존. (대표 지시 2026-05-29)
    db.execute("UPDATE messages SET thread_hidden=1 WHERE id=?", (parent_id,))
    affected_replies = db.execute(
        "SELECT COUNT(*) AS c FROM messages WHERE parent_message_id=?", (parent_id,)
    ).fetchone()["c"]
    db.commit()
    # 같은 방 사용자들에게 socket emit (양방향 갱신)
    try:
        socketio.emit("thread_deleted", {"parent_id": parent_id, "room_id": parent["room_id"]}, room=f"room_{parent['room_id']}")
    except Exception:
        pass
    return jsonify({"ok": True, "deleted_replies": affected_replies})


# ─── 📢 방별 공지사항 ───────────────────────────────────────────────
def _notice_to_dict(r, me_id, can_manage):
    return {
        "id": r["id"],
        "content": r["content"],
        "created_by": r["created_by"],
        "created_at": r["created_at"],
        "author": r["author"] if "author" in r.keys() else None,
        "author_title": r["author_title"] if "author_title" in r.keys() else None,
        "can_delete": bool(can_manage or (r["created_by"] == me_id)),
    }


@app.route("/api/rooms/<int:room_id>/notices")
@login_required
def api_room_notices_list(room_id):
    """이 방의 공지 목록 — 최신순(id DESC). 방 멤버면 조회 가능."""
    me = current_user()
    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (room_id, me["id"]),
    ).fetchone():
        abort(403)
    rows = db.execute("""
        SELECT n.id, n.content, n.created_by, n.created_at,
               u.display_name AS author, u.title AS author_title
          FROM room_notices n
          LEFT JOIN users u ON u.id = n.created_by
         WHERE n.room_id=? AND COALESCE(n.active,1)=1
         ORDER BY n.id DESC
    """, (room_id,)).fetchall()
    can_manage = _can_manage_thread(db, room_id, me)
    return jsonify({"items": [_notice_to_dict(r, me["id"], can_manage) for r in rows]})


@app.route("/api/rooms/<int:room_id>/notices", methods=["POST"])
@login_required
def api_room_notice_create(room_id):
    """공지 등록 — 방 참여자 누구나. (대표 지시 2026-05-29)"""
    me = current_user()
    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (room_id, me["id"]),
    ).fetchone():
        abort(403)
    data = request.get_json(silent=True) or {}
    content = (data.get("content") or "").strip()
    if not content:
        return jsonify({"error": "공지 내용을 입력하세요."}), 400
    if len(content) > 2000:
        content = content[:2000]
    now = datetime.now(timezone.utc).isoformat()
    cur = db.execute(
        "INSERT INTO room_notices (room_id, content, created_by, created_at, active) VALUES (?,?,?,?,1)",
        (room_id, content, me["id"], now),
    )
    nid = cur.lastrowid
    db.commit()
    notice = {
        "id": nid, "content": content, "created_by": me["id"], "created_at": now,
        "author": me["display_name"],
        "author_title": (me["title"] if "title" in me.keys() else None),
        "can_delete": True,
    }
    try:
        socketio.emit("notice_added", {"room_id": room_id, "notice": notice}, room=f"room_{room_id}")
    except Exception:
        pass
    return jsonify({"ok": True, "notice": notice})


@app.route("/api/rooms/<int:room_id>/notices/<int:notice_id>", methods=["DELETE"])
@login_required
def api_room_notice_delete(room_id, notice_id):
    """공지 삭제(active=0, 내용 보존) — 등록자 본인 또는 방장/PM/관리자. (대표 지시 2026-05-29)"""
    me = current_user()
    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (room_id, me["id"]),
    ).fetchone():
        abort(403)
    row = db.execute(
        "SELECT id, created_by, COALESCE(active,1) AS active FROM room_notices WHERE id=? AND room_id=?",
        (notice_id, room_id),
    ).fetchone()
    if not row:
        return jsonify({"error": "공지를 찾을 수 없습니다."}), 404
    if int(row["active"] or 0) == 0:
        return jsonify({"ok": True})  # 이미 삭제됨 — 멱등 처리
    if not (row["created_by"] == me["id"] or _can_manage_thread(db, room_id, me)):
        return jsonify({"error": "본인이 등록한 공지 또는 방장 / PM / 관리자만 삭제할 수 있습니다."}), 403
    db.execute("UPDATE room_notices SET active=0 WHERE id=?", (notice_id,))
    db.commit()
    try:
        socketio.emit("notice_deleted", {"room_id": room_id, "notice_id": notice_id}, room=f"room_{room_id}")
    except Exception:
        pass
    return jsonify({"ok": True})


@app.route("/api/rooms/<int:room_id>/avatar", methods=["POST"])
@login_required
def api_room_avatar_upload(room_id):
    """채널/방 아바타 이미지 업로드 — 방장 또는 관리자(ceo). (대표 지시 2026-06-04: 방장도 가능)"""
    me = current_user()
    db = get_db()
    _role = _my_room_role(db, room_id, me["id"])
    if _role != 'host' and me["role"] != "ceo":
        return jsonify({"error": "방장 또는 관리자만 채널 아이콘을 설정할 수 있습니다."}), 403
    room = db.execute("SELECT id FROM rooms WHERE id=?", (room_id,)).fetchone()
    if not room:
        return jsonify({"error": "방이 없습니다."}), 404
    if "file" not in request.files:
        return jsonify({"error": "file 필드에 이미지 첨부 필요"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "파일명 누락"}), 400
    ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
    if ext not in AVATAR_ALLOWED_EXT:
        return jsonify({"error": f"지원 형식: {', '.join(AVATAR_ALLOWED_EXT)}"}), 400
    data = f.read()
    if len(data) > AVATAR_MAX_BYTES:
        return jsonify({"error": f"파일이 너무 큽니다 ({len(data)//1024}KB > {AVATAR_MAX_BYTES//1024//1024}MB)"}), 400
    if len(data) < 100:
        return jsonify({"error": "파일이 너무 작거나 손상되었습니다"}), 400
    for old_ext in AVATAR_ALLOWED_EXT:
        op = os.path.join(ROOM_AVATAR_DIR, f"{room_id}.{old_ext}")
        if os.path.exists(op):
            try: os.remove(op)
            except Exception: pass
    with open(os.path.join(ROOM_AVATAR_DIR, f"{room_id}.{ext}"), "wb") as out:
        out.write(data)
    import time as _t
    rel_url = f"{BASE_PATH}/uploads/room_avatars/{room_id}.{ext}?v={int(_t.time())}"
    db.execute("UPDATE rooms SET avatar_url=? WHERE id=?", (rel_url, room_id))
    db.commit()
    try:
        _emit_room_event(room_id, "room_avatar_changed", {"room_id": room_id, "avatar_url": rel_url})
    except Exception:
        pass
    return jsonify({"ok": True, "avatar_url": rel_url})


@app.route("/api/rooms/<int:room_id>/avatar", methods=["DELETE"])
@login_required
def api_room_avatar_delete(room_id):
    """채널/방 아바타 이미지 제거 — 방장 또는 관리자(ceo). (대표 지시 2026-06-04: 방장도 가능)"""
    me = current_user()
    db = get_db()
    _role = _my_room_role(db, room_id, me["id"])
    if _role != 'host' and me["role"] != "ceo":
        return jsonify({"error": "방장 또는 관리자만 가능"}), 403
    for ext in AVATAR_ALLOWED_EXT:
        p = os.path.join(ROOM_AVATAR_DIR, f"{room_id}.{ext}")
        if os.path.exists(p):
            try: os.remove(p)
            except Exception: pass
    db.execute("UPDATE rooms SET avatar_url=NULL WHERE id=?", (room_id,))
    db.commit()
    try:
        _emit_room_event(room_id, "room_avatar_changed", {"room_id": room_id, "avatar_url": None})
    except Exception:
        pass
    return jsonify({"ok": True})


def _direct_key(uid1, uid2):
    """1:1 방 식별 키 — 두 사용자 ID를 정렬해 'min:max' 로."""
    a, b = sorted((int(uid1), int(uid2)))
    return f"{a}:{b}"


def _socket_rejoin_user(uid, room_id):
    """현재 접속 중인 사용자의 모든 소켓을 방 소켓룸에 합류시킴 (재입장 실시간 반영).
    상대가 나갔다가 1:1 재개될 때, 접속 중이면 즉시 new_message 를 받게 함."""
    try:
        with _user_conn_lock:
            sids = list(_user_connections.get(uid, {}).keys())
        for sid in sids:
            try:
                socketio.server.enter_room(sid, f"room_{room_id}")
            except Exception:
                pass
    except Exception:
        pass


def _find_or_create_direct(db, me_id, other_id):
    """1:1 방 찾기(상대가 나갔어도) 또는 신규 생성. (room_id, existing) 반환.
    기존 방이면 빠진 멤버(나/상대)를 다시 추가해 이전 대화 그대로 이어가기."""
    key = _direct_key(me_id, other_id)
    now = datetime.now(timezone.utc).isoformat()
    # 중복 방(옛 버그로 생성)이 있으면 가장 오래된 방 = 원본 대화 내역이 있는 방 우선
    existing = db.execute(
        "SELECT id FROM rooms WHERE type='direct' AND direct_key=? ORDER BY id ASC LIMIT 1", (key,)
    ).fetchone()
    if not existing:
        # direct_key 없는 옛 방 호환 — 양쪽 모두 현재 멤버인 방 검색 후 키 백필
        existing = db.execute("""
            SELECT r.id FROM rooms r
              JOIN room_members rm1 ON rm1.room_id=r.id AND rm1.user_id=?
              JOIN room_members rm2 ON rm2.room_id=r.id AND rm2.user_id=?
             WHERE r.type='direct' ORDER BY r.id ASC LIMIT 1
        """, (me_id, other_id)).fetchone()
        if existing:
            db.execute("UPDATE rooms SET direct_key=? WHERE id=?", (key, existing["id"]))
    if existing:
        rid = existing["id"]
        # 빠진 멤버 재입장 (이전 대화 유지) — 나가서 사라졌던 쪽 복원
        rejoined = []
        rejoin_msg = None  # 본인(방을 연 사람)이 재입장한 경우의 안내 메시지
        for uid in (me_id, other_id):
            if not db.execute("SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
                              (rid, uid)).fetchone():
                db.execute(
                    "INSERT INTO room_members (room_id, user_id, joined_at, role) VALUES (?,?,?,?)",
                    (rid, uid, now, "member"),
                )
                rejoined.append(uid)
                # '나갔습니다' 와 대칭 — 본인이 다시 들어온 경우만 안내 (대표 지시 2026-05-30).
                #  상대가 나갔던 방을 내가 열어 복원하는 경우는 상대가 능동 행동을 한 게 아니므로 조용히.
                if uid == me_id:
                    urow = db.execute("SELECT display_name, avatar_color FROM users WHERE id=?",
                                      (uid,)).fetchone()
                    dname = urow["display_name"] if urow else "사용자"
                    stext = f"[{dname}] 님이 다시 들어왔습니다."
                    cur2 = db.execute(
                        "INSERT INTO messages (room_id, user_id, content, kind, created_at) VALUES (?,?,?,?,?)",
                        (rid, uid, stext, "system", now),
                    )
                    rejoin_msg = {
                        "id": cur2.lastrowid, "user_id": uid,
                        "display_name": dname,
                        "avatar_color": (urow["avatar_color"] if urow else None),
                        "content": stext, "kind": "system", "created_at": now,
                    }
        db.commit()
        # 재입장한 사용자가 접속 중이면 소켓룸에 즉시 합류 → 실시간 메시지 수신
        for uid in rejoined:
            _socket_rejoin_user(uid, rid)
        # 재입장 안내 emit — 소켓 합류 후에 보내 양쪽(본인·상대) 모두 실시간 수신
        if rejoin_msg:
            socketio.emit("new_message", {**rejoin_msg, "room_id": rid}, to=f"room_{rid}")
        return rid, True
    # 신규 생성
    cur = db.execute(
        "INSERT INTO rooms (name, type, created_by, created_at, name_locked, direct_key) "
        "VALUES (?,?,?,?,1,?)",
        ("", "direct", me_id, now, key),
    )
    rid = cur.lastrowid
    for uid in (me_id, other_id):
        role = "host" if uid == me_id else "member"
        db.execute(
            "INSERT INTO room_members (room_id, user_id, joined_at, role) VALUES (?,?,?,?)",
            (rid, uid, now, role),
        )
    db.commit()
    return rid, False


@app.route("/api/rooms/direct/<int:other_user_id>", methods=["POST"])
@login_required
def api_rooms_direct_open(other_user_id):
    """1:1 채팅방 열기 — 이미 있으면 그 방(상대가 나갔어도 복원), 없으면 새로 생성.
    사이드바 '👥 사용자' 탭에서 사람 클릭 시 호출됨."""
    me = current_user()
    if other_user_id == me["id"]:
        return jsonify({"error": "본인과는 1:1 채팅 불가 — 📝 메모 사용"}), 400
    # 고객사(게스트)는 1:1 대화 신청 불가 — 직원이 고객에게 못 거는 것과 대칭. 초대된 방 안에서만 소통.
    #   (화면 우회로 호출돼도 서버에서 차단하는 안전망. 대표 지시 2026-05-31)
    if _is_guest(me):
        return jsonify({"error": "고객사 사용자는 1:1 대화를 시작할 수 없습니다. 초대된 대화방에서 소통해 주세요."}), 403
    db = get_db()
    other = db.execute(
        "SELECT id, active, COALESCE(is_guest,0) AS is_guest FROM users WHERE id=?", (other_user_id,)
    ).fetchone()
    if not other:
        return jsonify({"error": "사용자 없음"}), 404
    if not other["active"]:
        return jsonify({"error": "비활성 사용자"}), 400
    if other["is_guest"]:
        return jsonify({"error": "고객사 사용자와는 1:1 대화를 만들 수 없습니다. 그룹·프로젝트 방으로 초대해 소통해 주세요."}), 403
    rid, existing = _find_or_create_direct(db, me["id"], other_user_id)
    return jsonify({"room_id": rid, "existing": existing})


@app.route("/api/rooms")
@login_required
def api_rooms():
    me = current_user()
    db = get_db()
    # 🐞 '버그 신고' 채널 — 전 직원 자동 참여(게스트 제외). 방목록 첫 로드 때 합류시킴. (대표 지시 2026-06-03)
    if not _is_guest(me):
        try:
            _brid = _get_bug_room_id(db)
            if _brid:
                db.execute(
                    "INSERT OR IGNORE INTO room_members (room_id, user_id, joined_at, role) VALUES (?,?,?, 'member')",
                    (_brid, me["id"], datetime.now(timezone.utc).isoformat()),
                )
                db.commit()
        except Exception as _be:
            print(f"[bug] 버그채널 자동참여 실패: {_be}", flush=True)
    # 귓속말 필터 — 본인이 송신자/수신자가 아닌 귓속말은 last_message·unread 에서 모두 제외.
    # 다른 사람 사이드바·푸시·미열람 카운트에 안 보이게 (귓속말은 진짜 둘만 보이게).
    rows = db.execute("""
        SELECT r.id, r.name, r.name_ko, r.name_vi, r.name_en, r.name_zh, r.type, r.created_at, r.name_locked, r.created_by,
               r.retention_days, r.invite_policy, r.channel_scope, r.avatar_url,
               rm.role AS my_role,
               (SELECT 1 FROM room_members gm JOIN users gu ON gu.id = gm.user_id
                 WHERE gm.room_id = r.id AND COALESCE(gu.is_guest,0)=1 AND gu.active=1 LIMIT 1) AS has_guest,
               rm.pinned, rm.order_value,
               (SELECT alias FROM room_aliases WHERE room_id=r.id AND user_id=?) AS my_alias,
               it.code AS item_code, it.customer AS item_customer,
               it.status AS item_status, it.due_date AS item_due,
               (SELECT content FROM messages
                  WHERE room_id = r.id
                    AND (whisper_to_user_id IS NULL OR whisper_to_user_id = ? OR user_id = ?)
                    AND (r.type != 'self' OR COALESCE(kind,'text') NOT IN ('system','deleted'))
                  ORDER BY id DESC LIMIT 1) AS last_message,
               (SELECT created_at FROM messages
                  WHERE room_id = r.id
                    AND (whisper_to_user_id IS NULL OR whisper_to_user_id = ? OR user_id = ?)
                    AND (r.type != 'self' OR COALESCE(kind,'text') NOT IN ('system','deleted'))
                  ORDER BY id DESC LIMIT 1) AS last_at,
               (SELECT COUNT(*) FROM messages m
                  WHERE m.room_id = r.id
                    AND m.id > rm.last_read_message_id
                    AND m.user_id != ?
                    AND (m.whisper_to_user_id IS NULL OR m.whisper_to_user_id = ?)
               ) AS unread
          FROM rooms r
          JOIN room_members rm ON rm.room_id = r.id
          LEFT JOIN items it ON it.room_id = r.id
         WHERE rm.user_id = ?
         ORDER BY
            CASE r.type WHEN 'self' THEN 0 ELSE 1 END,
            CASE WHEN rm.pinned = 1 THEN 0 ELSE 1 END,
            CASE WHEN rm.order_value IS NOT NULL THEN 0 ELSE 1 END,
            rm.order_value ASC,
            (last_at IS NULL), last_at DESC, r.id DESC
    """, (
        me["id"],                       # my_alias
        me["id"], me["id"],             # last_message whisper filter (whisper_to_user_id=me OR user_id=me)
        me["id"], me["id"],             # last_at whisper filter
        me["id"], me["id"],             # unread: user_id != me AND (whisper IS NULL OR whisper=me)
        me["id"],                       # rm.user_id = me
    )).fetchall()

    out = []
    for r in rows:
        d = dict(r)
        # 1:1 방은 항상 상대방 이름으로 표시
        if r["type"] == "direct":
            other = db.execute("""
                SELECT u.id, u.display_name, u.avatar_color
                  FROM room_members rm
                  JOIN users u ON u.id = rm.user_id
                 WHERE rm.room_id = ? AND rm.user_id != ?
                 LIMIT 1
            """, (r["id"], me["id"])).fetchone()
            if other:
                # 1:1 상대 고유번호(ID) — 화면이 '이름 글자'가 아니라 ID 로 상대를 찾게 함.
                #   동명이인(예: 표시이름이 둘 다 '짱')일 때 목록·아바타가 엉뚱한 사람으로
                #   표시되던 버그 방지 (대표 지시 2026-06-05). 이름은 표시용으로만 유지.
                d["peer_id"] = other["id"]
                d["name"] = other["display_name"]
                d["avatar_color"] = other["avatar_color"]
        else:
            # 그룹/프로젝트 방: name_locked=0 이고 내 별명 있으면 별명 우선
            if not r["name_locked"] and r["my_alias"]:
                d["display_name_override"] = r["my_alias"]
                d["original_name"] = r["name"]
                d["name"] = r["my_alias"]
        out.append(d)
    return jsonify(out)


@app.route("/api/items", methods=["POST"])
@login_required
def api_items_create():
    me = current_user()
    if _is_guest(me):
        return jsonify({"error": "외부 사용자는 방을 만들 수 없습니다."}), 403
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "프로젝트 이름은 필수입니다."}), 400
    code = (data.get("code") or "").strip() or None
    customer = (data.get("customer") or "").strip() or None
    status = data.get("status") or "active"
    if status not in ("active", "hold", "done", "cancelled"):
        status = "active"
    due_date = data.get("due_date") or None
    description = (data.get("description") or "").strip() or None
    user_ids = list({int(x) for x in (data.get("user_ids") or [])})
    if me["id"] not in user_ids:
        user_ids.append(me["id"])

    db = get_db()
    now = datetime.now(timezone.utc).isoformat()
    # 프로젝트 방은 이름 고정 (방장만 변경 가능)
    cur = db.execute(
        "INSERT INTO rooms (name, type, created_by, created_at, name_locked) VALUES (?,?,?,?,1)",
        (name, "item", me["id"], now),
    )
    rid = cur.lastrowid
    db.execute("""
        INSERT INTO items (room_id, code, name, customer, status, due_date, description,
                           created_by, created_at, updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?)
    """, (rid, code, name, customer, status, due_date, description, me["id"], now, now))
    for uid in user_ids:
        role = 'host' if uid == me["id"] else 'member'
        db.execute(
            "INSERT INTO room_members (room_id, user_id, joined_at, role) VALUES (?,?,?,?)",
            (rid, uid, now, role),
        )
    db.execute(
        "INSERT INTO messages (room_id, user_id, content, kind, created_at) VALUES (?,?,?,?,?)",
        (rid, me["id"], f"프로젝트 [{name}] 생성됨", "system", now),
    )
    db.commit()
    return jsonify({"room_id": rid, "name": name})


@app.route("/api/items/lookup", methods=["GET"])
@login_required
def api_items_lookup():
    """관리번호(관리번호) 자동완성 — 기존 등록된 프로젝트에서 동일·유사 코드 조회.

    Query string:
        code  : 검색 문자열 (전방·중간 부분 일치)
        limit : 최대 반환 건수 (기본 10, 상한 30)

    응답:
        [{code, customer, name, status, due_date, room_id, created_at, source: 'local'}]
        - code 별로 가장 최근(created_at MAX) 1건만 반환 (DISTINCT)
        - 정렬: 정확 일치 > 전방 일치 > 부분 일치, 동률은 최신 우선

    향후: HAIST WORKS 시스템 연동 시 'source' 필드 분기 사용 가능
        (예: source='haist_works' / 'local')
    """
    q = (request.args.get("code") or "").strip()
    try:
        limit = int(request.args.get("limit", 10))
    except (ValueError, TypeError):
        limit = 10
    limit = max(1, min(limit, 30))
    if not q or len(q) < 1:
        return jsonify([])
    db = get_db()
    # SQLite — case-insensitive LIKE (ASCII). 한글은 대소문자 영향 없음.
    like = f"%{q}%"
    # 같은 code 가 여러 번 등록됐을 때 최신 1건만 반환 (서브쿼리 GROUP BY).
    rows = db.execute("""
        SELECT it.id, it.room_id, it.code, it.customer, it.name,
               it.status, it.due_date, it.created_at
          FROM items it
         INNER JOIN (
             SELECT code, MAX(created_at) AS max_ts
               FROM items
              WHERE code IS NOT NULL AND code != '' AND code LIKE ?
              GROUP BY code
         ) latest
            ON it.code = latest.code AND it.created_at = latest.max_ts
         ORDER BY
             CASE WHEN LOWER(it.code) = LOWER(?) THEN 0
                  WHEN LOWER(it.code) LIKE LOWER(?) THEN 1
                  ELSE 2
             END,
             it.created_at DESC
         LIMIT ?
    """, (like, q, f"{q}%", limit)).fetchall()
    out = []
    for r in rows:
        d = dict(r)
        d["source"] = "local"   # 향후 HAIST WORKS 연동 시 'haist_works' 등으로 분기
        out.append(d)
    return jsonify(out)


@app.route("/api/items/<int:room_id>", methods=["GET"])
@login_required
def api_item_get(room_id):
    me = current_user()
    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (room_id, me["id"]),
    ).fetchone():
        abort(403)
    row = db.execute("""
        SELECT it.*, r.name AS room_name
          FROM items it JOIN rooms r ON r.id = it.room_id
         WHERE it.room_id = ?
    """, (room_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    return jsonify(dict(row))


@app.route("/api/items/<int:room_id>", methods=["PATCH"])
@login_required
def api_item_update(room_id):
    """프로젝트 정보 수정. 권한: 방장(host) / PM(sub_host) / 관리자(ceo) (대표 지시 2026-05-27).
       일반 멤버가 '방 설정' 으로 오해해 수정하는 사고 방지."""
    me = current_user()
    data = request.get_json(silent=True) or {}
    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (room_id, me["id"]),
    ).fetchone():
        abort(403)
    # 권한 가드 — 관리자 OR 방장/PM
    is_admin = str((me["role"] if hasattr(me, "keys") else me.get("role")) or "") == "ceo"
    if not is_admin:
        my_role = _my_room_role(db, room_id, me["id"])
        if my_role not in ("host", "sub_host"):
            return jsonify({"error": "프로젝트 정보 수정은 방장·PM 또는 관리자만 가능합니다."}), 403
    fields, args = [], []
    for f in ("code", "customer", "status", "due_date", "description", "name", "keep_forever"):
        if f in data:
            v = data[f]
            if f == "keep_forever":
                v = 1 if v else 0
            elif f == "name":
                v = (v or "").strip() or None
            else:
                v = v or None
            fields.append(f"{f} = ?")
            args.append(v)
    if not fields:
        return jsonify({"ok": True})
    now = datetime.now(timezone.utc).isoformat()
    args.append(now)
    args.append(room_id)
    db.execute(
        f"UPDATE items SET {', '.join(fields)}, updated_at = ? WHERE room_id = ?",
        args,
    )
    if "name" in data:
        db.execute("UPDATE rooms SET name = ? WHERE id = ?", (data["name"], room_id))
    if "status" in data:
        label = {"active": "진행중", "hold": "보류", "done": "완료", "cancelled": "취소"}.get(data["status"], data["status"])
        db.execute(
            "INSERT INTO messages (room_id, user_id, content, kind, created_at) VALUES (?,?,?,?,?)",
            (room_id, me["id"], f"상태 변경 → {label}", "system", now),
        )
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/rooms/<int:room_id>/messages")
@login_required
def api_room_messages(room_id):
    me = current_user()
    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (room_id, me["id"]),
    ).fetchone():
        abort(403)
    # 메인 타임라인은 스레드 부모(parent_message_id IS NULL)만 표시.
    # 답글은 /api/messages/<id>/thread 로 별도 조회.
    # 귓속말 필터: whisper_to_user_id 가 NULL(공개) 이거나 송신자/수신자가 본인일 때만.
    rows = db.execute("""
        SELECT m.id, m.content, m.kind, m.created_at,
               m.file_path, m.file_name, m.file_size, m.file_mime,
               m.parent_message_id AS thread_parent_id,
               m.quoted_message_id,
               m.forwarded_from_message_id, m.forwarded_from_user_id,
               m.forwarded_from_name, m.forwarded_from_room_name, m.forwarded_from_created_at,
               m.whisper_to_user_id,
               m.album_id, m.share_count, m.parts,
               u.id AS user_id, u.display_name, u.avatar_color
          FROM messages m
          JOIN users u ON u.id = m.user_id
         WHERE m.room_id = ?
           AND m.parent_message_id IS NULL
           AND (m.whisper_to_user_id IS NULL
                OR m.whisper_to_user_id = ?
                OR m.user_id = ?)
         ORDER BY m.id ASC
    """, (room_id, me["id"], me["id"])).fetchall()
    out = [dict(r) for r in rows]
    # 묶음 메시지 parts(JSON 문자열) → 배열로 정규화 (kind='multipart') (대표 지시 2026-06-06)
    for _m in out:
        if _m.get("parts"):
            try:
                _m["parts"] = json.loads(_m["parts"])
            except Exception:
                _m["parts"] = None
    if out:
        ids = tuple(r["id"] for r in out)
        placeholders = ",".join("?" for _ in ids)

        # 반응 batch 로드
        rxs = db.execute(f"""
            SELECT mr.message_id, mr.emoji, mr.user_id, u.display_name
              FROM message_reactions mr JOIN users u ON u.id = mr.user_id
             WHERE mr.message_id IN ({placeholders})
        """, ids).fetchall()
        rxmap = {}
        for r in rxs:
            rxmap.setdefault(r["message_id"], []).append({"emoji": r["emoji"], "user_id": r["user_id"], "display_name": r["display_name"]})

        # ack batch 로드
        acks = db.execute(f"""
            SELECT a.message_id, a.user_id, a.ack_type, a.comment, u.display_name
              FROM message_acks a JOIN users u ON u.id = a.user_id
             WHERE a.message_id IN ({placeholders})
        """, ids).fetchall()
        ackmap = {}
        for a in acks:
            ackmap.setdefault(a["message_id"], []).append({
                "user_id": a["user_id"], "ack_type": a["ack_type"],
                "comment": a["comment"], "display_name": a["display_name"],
            })

        # 본인 별표 batch 로드
        my_stars = db.execute(
            f"SELECT message_id FROM message_stars WHERE user_id = ? AND message_id IN ({placeholders})",
            (me["id"],) + ids,
        ).fetchall()
        starred_ids = {s["message_id"] for s in my_stars}

        # 첨부 버전 정보 batch 로드
        avs = db.execute(f"""
            SELECT message_id, version_no, parent_message_id
              FROM attachment_versions
             WHERE message_id IN ({placeholders})
        """, ids).fetchall()
        avmap = {a["message_id"]: dict(a) for a in avs}

        # 같은 parent 의 최신 버전 찾기 (latest 표시용)
        if avmap:
            parent_ids = tuple({a["parent_message_id"] for a in avmap.values()})
            ph2 = ",".join("?" for _ in parent_ids)
            latest_rows = db.execute(f"""
                SELECT parent_message_id, MAX(version_no) AS max_v
                  FROM attachment_versions
                 WHERE parent_message_id IN ({ph2})
                 GROUP BY parent_message_id
            """, parent_ids).fetchall()
            latest_map = {r["parent_message_id"]: r["max_v"] for r in latest_rows}
        else:
            latest_map = {}

        # 캐시된 번역 batch 로드 — 클라이언트가 이미 번역된 메시지는 자동 표시
        translations = db.execute(f"""
            SELECT message_id, target_lang, translated_text
              FROM message_translations
             WHERE message_id IN ({placeholders})
        """, ids).fetchall()
        trmap = {}
        for t in translations:
            trmap.setdefault(t["message_id"], {})[t["target_lang"]] = t["translated_text"]

        # 스레드 답글 카운트 — 각 부모 메시지의 reply_count + 마지막 답글 시각
        thread_rows = db.execute(f"""
            SELECT parent_message_id AS pid,
                   COUNT(*) AS cnt,
                   MAX(created_at) AS last_at,
                   COUNT(DISTINCT user_id) AS participants
              FROM messages
             WHERE parent_message_id IN ({placeholders})
             GROUP BY parent_message_id
        """, ids).fetchall()
        thread_map = {t["pid"]: dict(t) for t in thread_rows}

        # 귓속말 수신자 표시명 batch 로드
        whisper_ids = [r["whisper_to_user_id"] for r in out if r.get("whisper_to_user_id")]
        whisper_name_map = {}
        if whisper_ids:
            wph = ",".join("?" for _ in whisper_ids)
            wnrows = db.execute(
                f"SELECT id, display_name FROM users WHERE id IN ({wph})",
                whisper_ids,
            ).fetchall()
            whisper_name_map = {r["id"]: r["display_name"] for r in wnrows}
        for r in out:
            wid = r.get("whisper_to_user_id")
            if wid:
                r["whisper_to_name"] = whisper_name_map.get(wid)

        # 인용 답장 — quoted_message_id 가 가리키는 원본 메시지 batch 로드 (미니 카드용)
        quoted_ids = [r["quoted_message_id"] for r in out if r.get("quoted_message_id")]
        quoted_map = {}
        if quoted_ids:
            qph = ",".join("?" for _ in quoted_ids)
            qrows = db.execute(f"""
                SELECT m.id, m.content, m.kind, m.created_at, m.file_name,
                       u.display_name, u.avatar_color
                  FROM messages m JOIN users u ON u.id = m.user_id
                 WHERE m.id IN ({qph})
            """, quoted_ids).fetchall()
            quoted_map = {q["id"]: dict(q) for q in qrows}

        for m in out:
            m["reactions"] = rxmap.get(m["id"], [])
            m["acks"] = ackmap.get(m["id"], [])
            m["starred_by_me"] = m["id"] in starred_ids
            m["translations"] = trmap.get(m["id"], {})
            # 스레드 — 답글 정보
            t = thread_map.get(m["id"])
            if t:
                m["thread_reply_count"] = t["cnt"]
                m["thread_last_at"] = t["last_at"]
                m["thread_participants"] = t["participants"]
            else:
                m["thread_reply_count"] = 0
            # 인용 답장 — 원본 메시지 미니 카드용 메타
            if m.get("quoted_message_id"):
                qm = quoted_map.get(m["quoted_message_id"])
                if qm:
                    m["quoted"] = qm
                else:
                    # 원본 삭제된 경우 — 마커만
                    m["quoted"] = {"deleted": True}
            av = avmap.get(m["id"])
            if av:
                m["version_no"] = av["version_no"]
                # 첨부 버전 부모 — 스레드 부모와 다른 개념이라 별도 필드명 사용
                m["attachment_parent_id"] = av["parent_message_id"]
                # 하위호환 — 기존 frontend 가 data-parent-msg-id 로 읽음 (첨부 버전용)
                m["parent_message_id"] = av["parent_message_id"]
                m["is_latest_version"] = (av["version_no"] == latest_map.get(av["parent_message_id"], av["version_no"]))
    return jsonify(out)


@app.route("/api/messages/<int:message_id>/forward", methods=["POST"])
@login_required
def api_message_forward(message_id):
    """전달(Forward) — 출처 보존.
    body: {to_room_ids: [int, ...], add_comment?: str (선택)}
    원본 메시지 1건을 N개 대상 방에 복사. 각 새 메시지에는 forwarded_from_* 메타가 박힘.
    원본이 첨부파일이면 같은 파일 정보도 복사 (실제 파일은 공유)."""
    me = current_user()
    db = get_db()
    src = db.execute("""
        SELECT m.id, m.room_id, m.user_id, m.content, m.kind, m.created_at,
               m.file_path, m.file_name, m.file_size, m.file_mime,
               u.display_name AS author_name,
               r.name AS room_name
          FROM messages m
          JOIN users u ON u.id = m.user_id
          JOIN rooms r ON r.id = m.room_id
         WHERE m.id = ?
    """, (message_id,)).fetchone()
    if not src:
        return jsonify({"error": "원본 메시지 없음"}), 404
    # 출처 방 멤버 권한 (원본 볼 권한 있어야 전달 가능)
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (src["room_id"], me["id"]),
    ).fetchone():
        return jsonify({"error": "원본 방 멤버 아님"}), 403
    data = request.get_json(silent=True) or {}
    to_rooms = data.get("to_room_ids") or []
    add_comment = (data.get("add_comment") or "").strip()
    if not isinstance(to_rooms, list) or not to_rooms:
        return jsonify({"error": "to_room_ids 필요"}), 400
    try:
        to_rooms = [int(x) for x in to_rooms]
    except (ValueError, TypeError):
        return jsonify({"error": "to_room_ids 형식 오류"}), 400
    if len(to_rooms) > 20:
        return jsonify({"error": "한 번에 최대 20개 방"}), 400

    # 사전 검증: 모든 대상 방의 멤버여야 함
    bad = []
    for rid in to_rooms:
        if not db.execute(
            "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
            (rid, me["id"]),
        ).fetchone():
            bad.append(rid)
    if bad:
        return jsonify({"error": f"방 멤버 아님: {bad}"}), 403

    me_row = db.execute("SELECT display_name, avatar_color FROM users WHERE id=?", (me["id"],)).fetchone()
    now = datetime.now(timezone.utc).isoformat()
    new_ids = []
    for rid in to_rooms:
        # 코멘트 + (있으면) 원본 텍스트. 첨부파일은 파일 정보 복사.
        # 텍스트 메시지의 경우 본문은 원본 텍스트를 그대로 옮김.
        new_content = src["content"] if (src["kind"] in ("text", "image", "file") and src["content"]) else ""
        if add_comment:
            # 사용자 코멘트가 있으면 본문 앞에 코멘트, 새 줄 후 원본 (둘 다 표시)
            new_content = add_comment + ("\n\n" + new_content if new_content else "")
        cur = db.execute("""
            INSERT INTO messages
                (room_id, user_id, content, kind, created_at,
                 file_path, file_name, file_size, file_mime,
                 forwarded_from_message_id, forwarded_from_user_id,
                 forwarded_from_name, forwarded_from_room_name, forwarded_from_created_at)
            VALUES (?,?,?,?,?, ?,?,?,?, ?,?,?,?,?)
        """, (
            rid, me["id"], new_content, src["kind"] or "text", now,
            src["file_path"], src["file_name"], src["file_size"], src["file_mime"],
            src["id"], src["user_id"],
            src["author_name"], src["room_name"], src["created_at"],
        ))
        mid = cur.lastrowid
        new_ids.append({"room_id": rid, "message_id": mid})
        # 실시간 broadcast
        payload = {
            "id": mid,
            "room_id": rid,
            "user_id": me["id"],
            "display_name": me_row["display_name"],
            "avatar_color": me_row["avatar_color"],
            "content": new_content,
            "kind": src["kind"] or "text",
            "created_at": now,
            "file_path": src["file_path"],
            "file_name": src["file_name"],
            "file_size": src["file_size"],
            "file_mime": src["file_mime"],
            "forwarded_from_message_id": src["id"],
            "forwarded_from_user_id": src["user_id"],
            "forwarded_from_name": src["author_name"],
            "forwarded_from_room_name": src["room_name"],
            "forwarded_from_created_at": src["created_at"],
        }
        socketio.emit("new_message", payload, to=f"room_{rid}")
    db.commit()

    # 푸시 발송 — 각 방의 송신자 외 멤버
    if PYWEBPUSH_OK:
        body_preview = (add_comment or src["content"] or src["file_name"] or "")[:120]
        for rid in to_rooms:
            row = db.execute("SELECT name FROM rooms WHERE id=?", (rid,)).fetchone()
            room_name = row["name"] if row else "채팅"
            title = f"↗ {me_row['display_name']} 전달 ({room_name})"
            import threading as _t
            _t.Thread(
                target=push_message_to_room_members,
                args=(rid, me["id"], title, body_preview),
                kwargs={"url": f"{BASE_PATH}/chat?room={rid}", "tag": f"room_{rid}"},
                daemon=True,
            ).start()
    return jsonify({"ok": True, "forwarded_to": new_ids, "count": len(new_ids)})


# ───────── 여러 방 동시 공유(broadcast) + 공유 묶음 (대표 지시 2026-06-04) ─────────
@app.route("/api/messages/broadcast", methods=["POST"])
@login_required
def api_messages_broadcast():
    """같은 내용을 여러 방에 동시 전송(공유). 전 직원, 본인이 속한 방만. 게스트 제외.
    body: {content: str, to_room_ids: [int, ...]}
    각 방에 같은 텍스트 게시. share_count=대상 방 수 → 받는 쪽 '여러 방 공유' 배지."""
    me = current_user()
    db = get_db()
    g = db.execute("SELECT is_guest FROM users WHERE id=?", (me["id"],)).fetchone()
    if g and g["is_guest"]:
        return jsonify({"error": "외부 사용자는 사용할 수 없습니다"}), 403
    data = request.get_json(silent=True) or {}
    content = (data.get("content") or "").strip()
    to_rooms = data.get("to_room_ids") or []
    if not content:
        return jsonify({"error": "내용을 입력하세요"}), 400
    if not isinstance(to_rooms, list) or not to_rooms:
        return jsonify({"error": "보낼 방을 선택하세요"}), 400
    try:
        to_rooms = [int(x) for x in to_rooms]
    except (ValueError, TypeError):
        return jsonify({"error": "to_room_ids 형식 오류"}), 400
    to_rooms = list(dict.fromkeys(to_rooms))   # 중복 제거(순서 유지)
    if len(to_rooms) > 20:
        return jsonify({"error": "한 번에 최대 20개 방"}), 400
    # 사전 검증: 모든 대상 방의 멤버여야 함 (본인이 속한 방만)
    bad = []
    for rid in to_rooms:
        if not db.execute("SELECT 1 FROM room_members WHERE room_id=? AND user_id=?", (rid, me["id"])).fetchone():
            bad.append(rid)
    if bad:
        return jsonify({"error": f"본인이 속하지 않은 방이 있습니다: {bad}"}), 403
    me_row = db.execute("SELECT display_name, avatar_color FROM users WHERE id=?", (me["id"],)).fetchone()
    now = datetime.now(timezone.utc).isoformat()
    n = len(to_rooms)
    share_ids_json = json.dumps(to_rooms)   # 각 사본에 '어느 방들에 보냈는지' 동일하게 저장 → 글쓴이가 클릭 시 목록 표시
    # 글을 작성한 '현재 방' — 공유 목록 팝업에서 '✏️ 작성한 방' 표시용. 실제 공유 대상에 포함된 경우만 인정
    _origin = data.get("origin_room_id")
    try:
        origin_room_id = int(_origin) if _origin is not None else None
    except (ValueError, TypeError):
        origin_room_id = None
    if origin_room_id not in to_rooms:
        origin_room_id = None
    new_ids = []
    for rid in to_rooms:
        cur = db.execute(
            "INSERT INTO messages (room_id, user_id, content, kind, created_at, share_count, share_room_ids, share_origin_room_id) VALUES (?,?,?,?,?,?,?,?)",
            (rid, me["id"], content, "text", now, n, share_ids_json, origin_room_id),
        )
        mid = cur.lastrowid
        new_ids.append({"room_id": rid, "message_id": mid})
        socketio.emit("new_message", {
            "id": mid, "room_id": rid, "user_id": me["id"],
            "display_name": me_row["display_name"], "avatar_color": me_row["avatar_color"],
            "content": content, "kind": "text", "created_at": now, "share_count": n,
        }, to=f"room_{rid}")
    db.commit()
    # 푸시 — 각 방 송신자 외 멤버
    if PYWEBPUSH_OK:
        body_preview = content[:120]
        for rid in to_rooms:
            row = db.execute("SELECT name FROM rooms WHERE id=?", (rid,)).fetchone()
            room_name = row["name"] if row else "채팅"
            title = f"👥 {me_row['display_name']} ({room_name})"
            import threading as _t
            _t.Thread(
                target=push_message_to_room_members,
                args=(rid, me["id"], title, body_preview),
                kwargs={"url": f"{BASE_PATH}/chat?room={rid}", "tag": f"room_{rid}"},
                daemon=True,
            ).start()
    return jsonify({"ok": True, "sent_to": new_ids, "count": n})


@app.route("/api/messages/<int:message_id>/share_targets", methods=["GET"])
@login_required
def api_message_share_targets(message_id):
    """여러 방 공유 메시지가 '어느 방들에 함께 보내졌는지' 목록 — 글쓴이 본인만 조회 (대표 지시 2026-06-04).
    각 사본에 share_room_ids(대상 방 ID 배열)가 동일하게 저장돼 있어, 어느 사본을 눌러도 전체 목록을 보여줌."""
    me = current_user()
    db = get_db()
    row = db.execute(
        "SELECT user_id, share_room_ids, share_count, share_origin_room_id FROM messages WHERE id=?",
        (message_id,),
    ).fetchone()
    if not row:
        return jsonify({"error": "메시지를 찾을 수 없습니다"}), 404
    if row["user_id"] != me["id"]:
        return jsonify({"error": "본인이 보낸 글만 확인할 수 있습니다"}), 403
    try:
        ids = json.loads(row["share_room_ids"] or "[]")
    except Exception:
        ids = []
    origin_id = row["share_origin_room_id"]   # 글을 작성한 방(현재 방) — 팝업에서 '✏️ 작성한 방' 표시
    rooms_out = []
    for rid in ids:
        r = db.execute("SELECT id, name, type FROM rooms WHERE id=?", (rid,)).fetchone()
        if not r:
            continue   # 그 사이 삭제된 방은 건너뜀
        nm = (r["name"] or "").strip()
        if r["type"] == "self":
            nm = nm or "📝 메모"
        rooms_out.append({"id": r["id"], "name": nm or f"방 {r['id']}", "is_origin": (r["id"] == origin_id)})
    return jsonify({"rooms": rooms_out, "count": len(rooms_out), "origin_room_id": origin_id})


@app.route("/api/me/share_bundles", methods=["GET", "POST"])
@login_required
def api_share_bundles():
    """공유 묶음(개인용) — GET: 내 묶음 목록 / POST: 저장(같은 이름이면 갱신). body {name, room_ids:[int]}"""
    me = current_user()
    db = get_db()
    if request.method == "GET":
        rows = db.execute(
            "SELECT id, name, room_ids FROM share_bundles WHERE user_id=? ORDER BY updated_at DESC",
            (me["id"],),
        ).fetchall()
        out = []
        for r in rows:
            try:
                ids = json.loads(r["room_ids"])
            except Exception:
                ids = []
            out.append({"id": r["id"], "name": r["name"], "room_ids": ids})
        return jsonify(out)
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    room_ids = data.get("room_ids") or []
    if not name:
        return jsonify({"error": "묶음 이름을 입력하세요"}), 400
    if len(name) > 40:
        return jsonify({"error": "이름은 40자 이내"}), 400
    try:
        room_ids = [int(x) for x in room_ids]
    except (ValueError, TypeError):
        return jsonify({"error": "room_ids 형식 오류"}), 400
    room_ids = list(dict.fromkeys(room_ids))
    if not room_ids:
        return jsonify({"error": "방을 1개 이상 선택하세요"}), 400
    now = datetime.now(timezone.utc).isoformat()
    db.execute("""
        INSERT INTO share_bundles (user_id, name, room_ids, created_at, updated_at)
        VALUES (?,?,?,?,?)
        ON CONFLICT(user_id, name) DO UPDATE SET room_ids=excluded.room_ids, updated_at=excluded.updated_at
    """, (me["id"], name, json.dumps(room_ids), now, now))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/me/share_bundles/<int:bundle_id>", methods=["DELETE"])
@login_required
def api_share_bundle_delete(bundle_id):
    """공유 묶음 삭제 (본인 것만)."""
    me = current_user()
    db = get_db()
    db.execute("DELETE FROM share_bundles WHERE id=? AND user_id=?", (bundle_id, me["id"]))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/messages/<int:message_id>/thread", methods=["GET"])
@login_required
def api_message_thread(message_id):
    """스레드 답글 목록 + 부모 메시지. 사이드 패널 데이터.
    응답: { parent: {...}, replies: [...] }"""
    me = current_user()
    db = get_db()
    parent = db.execute("""
        SELECT m.id, m.room_id, m.content, m.kind, m.created_at,
               m.file_path, m.file_name, m.file_size, m.file_mime,
               u.id AS user_id, u.display_name, u.avatar_color
          FROM messages m
          JOIN users u ON u.id = m.user_id
         WHERE m.id = ?
    """, (message_id,)).fetchone()
    if not parent:
        return jsonify({"error": "not found"}), 404
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (parent["room_id"], me["id"]),
    ).fetchone():
        abort(403)
    replies = db.execute("""
        SELECT m.id, m.content, m.kind, m.created_at,
               m.file_path, m.file_name, m.file_size, m.file_mime,
               u.id AS user_id, u.display_name, u.avatar_color
          FROM messages m
          JOIN users u ON u.id = m.user_id
         WHERE m.parent_message_id = ?
         ORDER BY m.id ASC
    """, (message_id,)).fetchall()
    return jsonify({
        "parent": dict(parent),
        "replies": [dict(r) for r in replies],
    })


@app.route("/api/messages/<int:message_id>/reply", methods=["POST"])
@login_required
def api_message_reply(message_id):
    """스레드 답글 작성 (텍스트 전용).
    body: {content: str}
    답글은 parent_message_id 가 채워진 messages 레코드. 메인 채널 타임라인엔 안 보이고
    스레드 패널에서만 보임. Web Push 는 부모 작성자 + 기존 답글 작성자들에게."""
    me = current_user()
    db = get_db()
    parent = db.execute("SELECT id, room_id, user_id FROM messages WHERE id=?", (message_id,)).fetchone()
    if not parent:
        return jsonify({"error": "not found"}), 404
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (parent["room_id"], me["id"]),
    ).fetchone():
        abort(403)
    data = request.get_json(silent=True) or {}
    content = (data.get("content") or "").strip()
    if not content:
        return jsonify({"error": "content required"}), 400
    if len(content) > 4000:
        content = content[:4000]
    # 부모의 부모가 있으면 (대부분 None 이어야 함) 그것을 사용 → 답글의 답글이 아니라 같은 스레드.
    real_parent_id = message_id
    pp = db.execute("SELECT parent_message_id FROM messages WHERE id=?", (message_id,)).fetchone()
    if pp and pp["parent_message_id"]:
        real_parent_id = pp["parent_message_id"]
    now = datetime.now(timezone.utc).isoformat()
    cur = db.execute(
        "INSERT INTO messages (room_id, user_id, content, kind, created_at, parent_message_id) VALUES (?,?,?,?,?,?)",
        (parent["room_id"], me["id"], content, "text", now, real_parent_id),
    )
    mid = cur.lastrowid
    db.commit()
    u = db.execute("SELECT display_name, avatar_color FROM users WHERE id=?", (me["id"],)).fetchone()
    payload = {
        "id": mid,
        "room_id": parent["room_id"],
        "user_id": me["id"],
        "display_name": u["display_name"],
        "avatar_color": u["avatar_color"],
        "content": content,
        "kind": "text",
        "created_at": now,
        "parent_message_id": real_parent_id,    # 호환용
        "thread_parent_id": real_parent_id,     # 새 키
    }
    # 스레드 패널 실시간 갱신용 이벤트 + 메인 타임라인 reply_count 갱신용
    socketio.emit("thread_reply", payload, to=f"room_{parent['room_id']}")
    socketio.emit("thread_count_changed", {
        "room_id": parent["room_id"],
        "parent_id": real_parent_id,
    }, to=f"room_{parent['room_id']}")
    # 푸시 — 부모 작성자 + 기존 답글 작성자들 (송신자 제외)
    if PYWEBPUSH_OK:
        recipients = set()
        if parent["user_id"] != me["id"]:
            recipients.add(parent["user_id"])
        prev_repliers = db.execute(
            "SELECT DISTINCT user_id FROM messages WHERE parent_message_id=? AND user_id != ?",
            (real_parent_id, me["id"]),
        ).fetchall()
        for r in prev_repliers:
            recipients.add(r["user_id"])
        if recipients:
            room_row = db.execute("SELECT name FROM rooms WHERE id=?", (parent["room_id"],)).fetchone()
            room_name = room_row["name"] if room_row else "채팅"
            title = f"💬 {u['display_name']} (스레드·{room_name})"
            body = content[:120]
            import threading as _t
            def _push_all():
                for rid in recipients:
                    if _user_has_active_session(rid):
                        continue
                    send_push_to_user(
                        rid, title, body,
                        url=f"{BASE_PATH}/chat?room={parent['room_id']}&thread={real_parent_id}",
                        tag=f"thread_{real_parent_id}",
                    )
            _t.Thread(target=_push_all, daemon=True).start()
    return jsonify({"ok": True, "id": mid, "parent_id": real_parent_id})


@app.route("/api/messages/<int:message_id>/react", methods=["POST"])
@login_required
def api_message_react(message_id):
    me = current_user()
    data = request.get_json(silent=True) or {}
    emoji = (data.get("emoji") or "").strip()
    if not emoji or len(emoji) > 16:
        return jsonify({"error": "emoji 필수"}), 400
    db = get_db()
    msg = db.execute("SELECT room_id FROM messages WHERE id=?", (message_id,)).fetchone()
    if not msg:
        return jsonify({"error": "not found"}), 404
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (msg["room_id"], me["id"]),
    ).fetchone():
        abort(403)
    # 토글 — 이미 있으면 삭제, 없으면 추가
    existing = db.execute(
        "SELECT id FROM message_reactions WHERE message_id=? AND user_id=? AND emoji=?",
        (message_id, me["id"], emoji),
    ).fetchone()
    now = datetime.now(timezone.utc).isoformat()
    if existing:
        db.execute("DELETE FROM message_reactions WHERE id=?", (existing["id"],))
        action = "removed"
    else:
        db.execute(
            "INSERT INTO message_reactions (message_id, user_id, emoji, created_at) VALUES (?,?,?,?)",
            (message_id, me["id"], emoji, now),
        )
        action = "added"
    db.commit()
    socketio.emit("reaction_updated", {
        "message_id": message_id,
        "room_id": msg["room_id"],
        "user_id": me["id"],
        "display_name": me["display_name"],
        "emoji": emoji,
        "action": action,
    }, to=f"room_{msg['room_id']}")
    return jsonify({"action": action})


# ---------- 메시지 전달확인 (acknowledgment) ----------
ACK_TYPES = ("ok", "doing", "done", "reject")


@app.route("/api/messages/<int:message_id>/ack", methods=["POST"])
@login_required
def api_message_ack(message_id):
    """전달확인 — '내가 봤고 처리하겠다' 명시. 단순 '읽음' 보다 강한 의지 표시.

    body: { "ack_type": "ok" | "doing" | "done" | "reject", "comment": "..." (선택) }
    토글 동작 — 같은 ack_type 다시 호출하면 해제.
    """
    me = current_user()
    db = get_db()
    msg = db.execute(
        "SELECT m.id, m.room_id, m.user_id, m.content, m.kind FROM messages m WHERE m.id = ?",
        (message_id,),
    ).fetchone()
    if not msg:
        abort(404)
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (msg["room_id"], me["id"]),
    ).fetchone():
        abort(403)

    data = request.get_json(silent=True) or {}
    ack_type = data.get("ack_type", "ok")
    if ack_type not in ACK_TYPES:
        return jsonify({"error": "ack_type 은 ok/doing/done/reject"}), 400
    comment = (data.get("comment") or "").strip() or None

    existing = db.execute(
        "SELECT id FROM message_acks WHERE message_id=? AND user_id=? AND ack_type=?",
        (message_id, me["id"], ack_type),
    ).fetchone()

    now = datetime.now(timezone.utc).isoformat()
    if existing:
        # 토글: 해제
        db.execute("DELETE FROM message_acks WHERE id = ?", (existing["id"],))
        action = "removed"
    else:
        db.execute(
            "INSERT INTO message_acks (message_id, user_id, ack_type, comment, created_at) VALUES (?,?,?,?,?)",
            (message_id, me["id"], ack_type, comment, now),
        )
        action = "added"
    db.commit()

    socketio.emit("ack_updated", {
        "message_id": message_id,
        "user_id": me["id"],
        "display_name": me["display_name"],
        "ack_type": ack_type,
        "action": action,
        "comment": comment,
    }, to=f"room_{msg['room_id']}")
    return jsonify({"action": action, "ack_type": ack_type})


@app.route("/api/messages/<int:message_id>/acks")
@login_required
def api_message_acks_list(message_id):
    """메시지의 ack 누적 + 미확인자 목록."""
    me = current_user()
    db = get_db()
    msg = db.execute("SELECT room_id, user_id FROM messages WHERE id = ?", (message_id,)).fetchone()
    if not msg:
        abort(404)
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?", (msg["room_id"], me["id"])
    ).fetchone():
        abort(403)

    acks = db.execute("""
        SELECT a.user_id, a.ack_type, a.comment, a.created_at,
               u.display_name, u.avatar_color
          FROM message_acks a
          JOIN users u ON u.id = a.user_id
         WHERE a.message_id = ?
         ORDER BY a.created_at
    """, (message_id,)).fetchall()

    # 방 멤버 중 ack 안 한 사람 — 메시지 작성자는 제외 (본인이 본인 글 ack 할 필요 없음)
    members = db.execute("""
        SELECT u.id, u.display_name, u.avatar_color
          FROM room_members rm JOIN users u ON u.id = rm.user_id
         WHERE rm.room_id = ?
    """, (msg["room_id"],)).fetchall()
    acked_ids = {a["user_id"] for a in acks}
    pending = [
        dict(m) for m in members
        if m["id"] not in acked_ids and m["id"] != msg["user_id"]
    ]
    return jsonify({
        "acks": [dict(a) for a in acks],
        "pending": pending,
    })


# ---------- 메시지 별표 (중요 결정 마킹) ----------
@app.route("/api/messages/<int:message_id>/read_status", methods=["GET"])
@login_required
def api_message_read_status(message_id):
    """메시지를 읽은 사람·안 읽은 사람 명단 (대표 지시 2026-05-19).

    동작:
        - 방 멤버 전체 조회
        - 각 멤버의 last_read_message_id 와 비교
        - last_read >= message_id → 읽음
        - 그 미만 → 안 읽음
        - 발신자 본인은 명단에서 제외 (본인은 항상 본 거니까)

    응답:
        {
          message_id,
          read: [{user_id, display_name, avatar_color, avatar_url, last_read_at}],
          unread: [{user_id, display_name, avatar_color, avatar_url}],
          read_count, unread_count, total_members
        }
    """
    me = current_user()
    db = get_db()
    msg = db.execute(
        "SELECT id, room_id, user_id, created_at FROM messages WHERE id=?", (message_id,)
    ).fetchone()
    if not msg:
        return jsonify({"error": "메시지를 찾을 수 없습니다"}), 404
    # 방 멤버인지 확인
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (msg["room_id"], me["id"])
    ).fetchone():
        return jsonify({"error": "이 방 멤버만 조회 가능"}), 403
    # 방 멤버 전원 조회 (발신자 제외)
    rows = db.execute("""
        SELECT u.id, u.display_name, u.avatar_color, u.avatar_url,
               rm.last_read_message_id, rm.last_read_at
          FROM room_members rm
          JOIN users u ON u.id = rm.user_id
         WHERE rm.room_id = ? AND u.id != ? AND u.active = 1
         ORDER BY u.display_name
    """, (msg["room_id"], msg["user_id"])).fetchall()
    read = []
    unread = []
    for r in rows:
        d = {
            "user_id": r["id"],
            "display_name": r["display_name"],
            "avatar_color": r["avatar_color"],
            "avatar_url": r["avatar_url"],
        }
        last_read = r["last_read_message_id"] or 0
        if last_read >= message_id:
            d["last_read_at"] = r["last_read_at"]
            read.append(d)
        else:
            unread.append(d)
    return jsonify({
        "message_id": message_id,
        "sender_user_id": msg["user_id"],
        "read": read,
        "unread": unread,
        "read_count": len(read),
        "unread_count": len(unread),
        "total_members": len(read) + len(unread),
    })


@app.route("/api/messages/<int:message_id>", methods=["PATCH"])
@login_required
def api_message_edit(message_id):
    """메시지 편집 — 본인 텍스트 메시지만 (대표 지시 2026-05-19).

    body: {content: '새 내용'}
    제약:
        - 본인 메시지만 (관리자도 X — 보안)
        - 텍스트 메시지만 (kind='text' 또는 NULL)
        - 삭제된 메시지(kind='deleted')는 편집 불가
        - 시스템 메시지(kind='system')는 편집 불가
        - content 길이 1~4000자
    동작:
        - content 업데이트 + edited_at = 현재 시각
        - 검색 인덱스(FTS) 자동 동기화 (trigger)
        - socketio 'message_edited' broadcast → 다른 사용자 화면 즉시 갱신
    """
    me = current_user()
    db = get_db()
    msg = db.execute(
        "SELECT id, room_id, user_id, kind FROM messages WHERE id = ?",
        (message_id,)
    ).fetchone()
    if not msg:
        return jsonify({"error": "메시지를 찾을 수 없습니다"}), 404
    # 권한 — 본인만
    if msg["user_id"] != me["id"]:
        return jsonify({"error": "본인 메시지만 편집 가능"}), 403
    # 종류 제약
    if msg["kind"] == "deleted":
        return jsonify({"error": "삭제된 메시지는 편집 불가"}), 400
    if msg["kind"] in ("image", "file", "system", "sticker"):
        return jsonify({"error": "사진·파일·스티커·시스템 메시지는 편집 불가"}), 400
    data = request.get_json(silent=True) or {}
    new_content = (data.get("content") or "").strip()
    if not new_content:
        return jsonify({"error": "내용이 비어 있습니다"}), 400
    if len(new_content) > 4000:
        new_content = new_content[:4000]
    now = datetime.now(timezone.utc).isoformat()
    db.execute(
        "UPDATE messages SET content = ?, edited_at = ? WHERE id = ?",
        (new_content, now, message_id)
    )
    db.commit()
    # broadcast
    try:
        socketio.emit("message_edited", {
            "message_id": message_id,
            "room_id": msg["room_id"],
            "content": new_content,
            "edited_at": now,
        }, to=f"room_{msg['room_id']}")
    except Exception as e:
        print(f"[message edit] socketio emit 실패: {e}", flush=True)
    return jsonify({"ok": True, "message_id": message_id, "content": new_content, "edited_at": now})


@app.route("/api/messages/<int:message_id>", methods=["DELETE"])
@login_required
def api_message_delete(message_id):
    """메시지 삭제 (대표 지시 2026-05-19).

    권한:
        - 본인 메시지: 언제든 삭제 가능
        - 관리자(ceo): 누구 메시지든 삭제 가능 (보안·법무 대응)
        - 그 외: 403

    동작 (soft delete):
        - content → '🗑️ 삭제된 메시지'
        - kind → 'deleted'
        - 첨부 파일이 있으면 디스크에서도 제거
        - 인용·답글·전달 메타데이터 NULL 화 (혼란 방지)
        - socketio 'message_deleted' broadcast → 모든 방 멤버 화면 즉시 갱신

    record 자체는 보존 (이력 추적·읽음 카운트·인덱스 일관성)."""
    me = current_user()
    db = get_db()
    msg = db.execute(
        "SELECT id, room_id, user_id, kind, file_path FROM messages WHERE id = ?",
        (message_id,)
    ).fetchone()
    if not msg:
        return jsonify({"error": "메시지를 찾을 수 없습니다"}), 404
    # 권한 체크 — 본인 또는 ceo
    if msg["user_id"] != me["id"] and me["role"] != "ceo":
        return jsonify({"error": "본인이 보낸 메시지만 삭제할 수 있습니다"}), 403
    # 이미 삭제된 메시지면 no-op
    if msg["kind"] == "deleted":
        return jsonify({"ok": True, "message": "이미 삭제됨"})
    # 첨부 파일 디스크에서 제거 — 단, 같은 파일을 참조하는 "다른" 메시지가 없을 때만.
    #   전달(Forward)은 원본과 같은 file_path 를 공유한다. 전달본/원본을 지워도 다른 사본이
    #   남아 있으면 물리 파일은 보존해야 한다 (안 그러면 나머지 사본이 전부 깨짐). (버그수정 2026-06-01)
    if msg["file_path"]:
        other_ref = db.execute(
            "SELECT 1 FROM messages WHERE file_path = ? AND id <> ? LIMIT 1",
            (msg["file_path"], message_id),
        ).fetchone()
        if not other_ref:
            try:
                full_path = os.path.join(UPLOAD_DIR, msg["file_path"])
                if os.path.exists(full_path):
                    os.remove(full_path)
            except Exception as e:
                print(f"[message delete] 파일 제거 실패: {e}", flush=True)
    # soft delete — content/kind 만 교체, record 보존
    now = datetime.now(timezone.utc).isoformat()
    db.execute("""
        UPDATE messages SET
            content = '🗑️ 삭제된 메시지',
            kind = 'deleted',
            file_path = NULL, file_name = NULL, file_size = NULL, file_mime = NULL,
            quoted_message_id = NULL,
            forwarded_from_message_id = NULL, forwarded_from_user_id = NULL,
            forwarded_from_name = NULL, forwarded_from_room_name = NULL, forwarded_from_created_at = NULL,
            album_id = NULL
        WHERE id = ?
    """, (message_id,))
    # AI 번역 캐시도 같이 정리 (대표 지시 2026-05-28).
    #   이전엔 message_translations 행이 남아 메시지 로드 시 inline 으로 번역문이 다시 떴음.
    db.execute("DELETE FROM message_translations WHERE message_id = ?", (message_id,))
    db.commit()
    # 실시간 broadcast — 모든 방 멤버에게 알림
    try:
        socketio.emit("message_deleted", {
            "message_id": message_id,
            "room_id": msg["room_id"],
            "deleted_by": me["id"],
            "deleted_at": now,
        }, to=f"room_{msg['room_id']}")
    except Exception as e:
        print(f"[message delete] socketio emit 실패: {e}", flush=True)
    return jsonify({"ok": True, "message_id": message_id})


@app.route("/api/messages/<int:message_id>/star", methods=["POST"])
@login_required
def api_message_star(message_id):
    """메시지 별표 토글 — 중요 결정·합의·이정표 마킹. 사용자별."""
    me = current_user()
    db = get_db()
    msg = db.execute("SELECT room_id FROM messages WHERE id = ?", (message_id,)).fetchone()
    if not msg:
        abort(404)
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?", (msg["room_id"], me["id"])
    ).fetchone():
        abort(403)
    existing = db.execute(
        "SELECT 1 FROM message_stars WHERE message_id=? AND user_id=?",
        (message_id, me["id"]),
    ).fetchone()
    if existing:
        db.execute("DELETE FROM message_stars WHERE message_id=? AND user_id=?", (message_id, me["id"]))
        action = "removed"
    else:
        db.execute(
            "INSERT INTO message_stars (message_id, user_id, starred_at) VALUES (?,?,?)",
            (message_id, me["id"], datetime.now(timezone.utc).isoformat()),
        )
        action = "added"
    db.commit()
    return jsonify({"action": action})


@app.route("/api/rooms/<int:room_id>/starred")
@login_required
def api_room_starred(room_id):
    """방 안에서 본인이 별표한 메시지 모음 — 인수인계·결정 정리용."""
    me = current_user()
    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?", (room_id, me["id"])
    ).fetchone():
        abort(403)
    rows = db.execute("""
        SELECT m.id, m.content, m.kind, m.created_at, m.file_name, m.file_path,
               u.display_name, u.avatar_color
          FROM message_stars s
          JOIN messages m ON m.id = s.message_id
          JOIN users u ON u.id = m.user_id
         WHERE m.room_id = ? AND s.user_id = ?
         ORDER BY s.starred_at DESC
    """, (room_id, me["id"])).fetchall()
    return jsonify([dict(r) for r in rows])


# ---------- AI 요약·재작성 (Claude Haiku) ----------
# AI 보조 — 채널 일간 요약 / 긴 스레드 요약 / 작성 톤 조정.
# 한국어 1순위 — 한국어 특화로 차별화.
AI_SUMMARY_MODEL = os.environ.get("KNK_MSG_AI_SUMMARY_MODEL", "claude-haiku-4-5")


def _claude_summarize_messages(messages_payload, mode="channel"):
    """메시지 리스트를 요약. mode='channel' (방의 최근 N개) 또는 'thread' (스레드).
    messages_payload: [{display_name, content, created_at, kind}, ...]
    반환: ({summary_text, in_tokens, out_tokens, cost_usd}, None) 또는 (None, err)."""
    try:
        import anthropic
    except ImportError:
        return None, "anthropic SDK 미설치"
    if not ANTHROPIC_API_KEY:
        return None, "ANTHROPIC_API_KEY 환경변수 미설정"
    if not messages_payload:
        return None, "요약할 메시지가 없습니다"

    # 메시지를 자연어 transcript 로 변환 (LLM 가독성)
    lines = []
    for m in messages_payload:
        if m.get("kind") == "system":
            continue
        ts = m.get("created_at", "")
        try:
            from datetime import datetime as _dt
            dt = _dt.fromisoformat(ts.replace("Z", "+00:00")) if ts else None
            ts_short = dt.strftime("%m-%d %H:%M") if dt else ""
        except Exception:
            ts_short = ts[:16]
        name = m.get("display_name", "?")
        content = (m.get("content") or "").strip()
        if not content:
            content = f"[{m.get('kind','파일')}]"
        lines.append(f"[{ts_short}] {name}: {content}")
    transcript = "\n".join(lines)
    if not transcript:
        return None, "요약할 본문이 비어있음"

    mode_label = "스레드 토론" if mode == "thread" else "채팅방 대화"
    system_prompt = f"""You are a Korean business communication assistant for KNK Corporation (industrial machinery / inspection equipment).

Your job: Summarize the following Korean {mode_label} for a busy executive (대표/임원) as a DETAILED, FACTUAL report.

가장 중요한 규칙 (반드시 준수):
- 오직 본문(대화 내용)에 실제로 적힌 사실만 쓴다. 추정·해석·상상·창작·과장 절대 금지.
- 본문에 근거가 없는 내용은 한 글자도 쓰지 않는다. 불확실하면 쓰지 말고, 항목에 내용이 없으면 "없음".
- 최대한 상세하게: 본문에 오간 실질 내용은 빠짐없이 담는다(분량 제한 없음). 단, 없는 내용을 지어내 채우지 않는다.

Output format (Korean — 굵은 제목 그대로 사용):
**핵심 요약**
- 대화 전체의 목적과 결론을 2~4문장으로.
**상세 내용**
- 시간/주제 흐름대로 "누가 무엇을 말했고 어떤 일이 오갔는지" 구체적으로 불릿 정리.
- 사진·파일 공유는 "누가 무엇을 공유함"으로 기록. 숫자·금액·날짜·요청·답변을 구체적으로.
**주요 결정사항**
- 실제로 합의·결정된 것만 불릿. 없으면 "없음".
**미결사항·후속조치**
- 남은 일·확인 필요사항. 담당자·기한이 본문에 있으면 명시. 없으면 "없음".
**관련 인물**
- 대화에 등장한 사람 이름만 나열.

표기 규칙:
- 기술 용어·품번(예: 003M2501, WP-LOA)·고객사명·금액·날짜는 본문 그대로 정확히.
- 정중한 보고서 평어체: "~함", "~확인 필요" 등.
"""
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    try:
        msg = client.messages.create(
            model=AI_SUMMARY_MODEL,
            max_tokens=4000,
            system=[
                {"type": "text", "text": system_prompt, "cache_control": {"type": "ephemeral"}},
            ],
            messages=[
                {"role": "user", "content": f"다음 {mode_label}을 사실 그대로, 최대한 상세히 요약해 주세요:\n\n{transcript}"},
            ],
        )
        summary = "".join(b.text for b in msg.content if hasattr(b, "text")).strip()
        in_t = msg.usage.input_tokens
        out_t = msg.usage.output_tokens
        cost = (in_t / 1_000_000.0) * 1.0 + (out_t / 1_000_000.0) * 5.0
        return {
            "summary_text": summary,
            "in_tokens": in_t,
            "out_tokens": out_t,
            "cost_usd": cost,
            "model": AI_SUMMARY_MODEL,
        }, None
    except Exception as e:
        return None, f"Claude API 오류: {e}"


def _claude_summarize_for_history(messages_payload, item_meta=None):
    """프로젝트(프로젝트 방) 이력용 요약 — HAIST WORKS 프로젝트 이력 포맷.
    messages_payload 는 _claude_summarize_messages 와 같은 dict 리스트.
    item_meta 는 {code, name, customer, status} (있으면 컨텍스트로 전달).
    반환 형식: 기존 함수와 동일 ({summary_text, in_tokens, out_tokens, cost_usd, model}, None)."""
    try:
        import anthropic
    except ImportError:
        return None, "anthropic SDK 미설치"
    if not ANTHROPIC_API_KEY:
        return None, "ANTHROPIC_API_KEY 환경변수 미설정"
    if not messages_payload:
        return None, "요약할 메시지가 없습니다"

    # transcript 빌드
    lines = []
    for m in messages_payload:
        if m.get("kind") == "system":
            continue
        ts = m.get("created_at", "")
        try:
            from datetime import datetime as _dt
            dt = _dt.fromisoformat(ts.replace("Z", "+00:00")) if ts else None
            ts_short = dt.strftime("%m-%d %H:%M") if dt else ""
        except Exception:
            ts_short = ts[:16]
        name = m.get("display_name", "?")
        content = (m.get("content") or "").strip()
        kind = m.get("kind", "text")
        file_name = m.get("file_name")
        if kind == "image" and file_name:
            content = f"[사진: {file_name}] " + content
        elif kind == "file" and file_name:
            content = f"[파일: {file_name}] " + content
        if not content:
            content = f"[{kind}]"
        lines.append(f"[{ts_short}] {name}: {content}")
    transcript = "\n".join(lines)
    if not transcript:
        return None, "요약할 본문이 비어있음"

    item_ctx = ""
    if item_meta:
        bits = []
        if item_meta.get("code"): bits.append(f"품번 {item_meta['code']}")
        if item_meta.get("name"): bits.append(f"프로젝트명 {item_meta['name']}")
        if item_meta.get("customer"): bits.append(f"고객사 {item_meta['customer']}")
        if item_meta.get("status"): bits.append(f"상태 {item_meta['status']}")
        if bits:
            item_ctx = "\n[프로젝트 컨텍스트] " + " · ".join(bits) + "\n"

    system_prompt = f"""You are a Korean business communication assistant for KNK Corporation (industrial machinery / inspection equipment).

Your job: Summarize the following Korean project conversation as PROJECT HISTORY for HAIST WORKS (ERP system).
This summary will be saved as the official project log.

Output structure (Korean, in this exact order):

**기간 요약** (1~2 문장 — 이번 기간에 무엇이 진행되었는지 핵심)

**주요 결정사항**
- (담당자) 결정 내용 (날짜)
- ...

**진척 상황**
- 이번 기간 완료된 작업
- 진행 중인 작업

**미결 사항 / 후속조치**
- (담당자) 해야 할 것 (기한)
- ...

**관련 인물**
- 이름 나열 (쉼표 구분)

**언급된 외부 거래처·품번**
- 거래처명, 품번 등 (있으면)

Rules:
- 최대 600 한글 글자.
- 기술 용어·품번(예: 003M2501, WP-LOA)·고객사명은 원문 그대로.
- 추측 금지. 본문에 없으면 "없음".
- 정중한 평어체 (보고서 톤). "~함", "~확인 필요", "~예정".
- 시간은 가능하면 명시 (예: 5월 17일 14:00).
{item_ctx}"""

    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    try:
        msg = client.messages.create(
            model=AI_SUMMARY_MODEL,
            max_tokens=2048,
            system=[
                {"type": "text", "text": system_prompt, "cache_control": {"type": "ephemeral"}},
            ],
            messages=[
                {"role": "user", "content": f"다음 프로젝트 대화를 이력 보고서로 요약해 주세요:\n\n{transcript}"},
            ],
        )
        summary = "".join(b.text for b in msg.content if hasattr(b, "text")).strip()
        in_t = msg.usage.input_tokens
        out_t = msg.usage.output_tokens
        cost = (in_t / 1_000_000.0) * 1.0 + (out_t / 1_000_000.0) * 5.0
        return {
            "summary_text": summary,
            "in_tokens": in_t,
            "out_tokens": out_t,
            "cost_usd": cost,
            "model": AI_SUMMARY_MODEL,
        }, None
    except Exception as e:
        return None, f"Claude API 오류: {e}"


def _generate_project_history(room_id, created_by_uid=None):
    """방의 마지막 history 이후 새 메시지를 AI 요약 + 첨부 정리해서 1개 스냅샷 생성.
    created_by_uid 가 None 이면 자동(워커), 값이 있으면 수동(사용자).
    반환: (history_dict, None) 또는 (None, err_str). 새 메시지 0개면 (None, 'no_new')."""
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    try:
        # 마지막 스냅샷 last_message_id
        last_snap = db.execute(
            "SELECT last_message_id FROM project_history WHERE room_id=? ORDER BY id DESC LIMIT 1",
            (room_id,),
        ).fetchone()
        last_mid_cursor = last_snap["last_message_id"] if last_snap else 0

        # 그 이후 새 메시지 (시스템 제외, 스레드 답글 제외 — 메인 타임라인만)
        rows = db.execute("""
            SELECT m.id, m.content, m.kind, m.created_at,
                   m.file_path, m.file_name, m.file_size, m.file_mime,
                   u.display_name
              FROM messages m JOIN users u ON u.id = m.user_id
             WHERE m.room_id = ? AND m.id > ?
               AND m.kind != 'system'
               AND m.parent_message_id IS NULL
             ORDER BY m.id ASC
        """, (room_id, last_mid_cursor)).fetchall()

        if not rows:
            return None, "no_new"
        if len(rows) < 2 and created_by_uid is None:
            # 자동 워커: 메시지 1개만으론 의미있는 요약 어려움. 다음 사이클에 시도.
            return None, "too_few"

        # 프로젝트 메타
        item_meta = None
        item_row = db.execute("""
            SELECT it.code, it.name, it.customer, it.status
              FROM items it WHERE it.room_id=?
        """, (room_id,)).fetchone()
        if item_row:
            item_meta = dict(item_row)

        # AI 요약 (공급자 라우터)
        result, err = _ai_summarize_for_history([dict(r) for r in rows], item_meta=item_meta)
        if err:
            return None, err

        # 첨부 정리
        attachments = []
        for r in rows:
            if r["file_path"]:
                attachments.append({
                    "message_id": r["id"],
                    "name": r["file_name"] or r["file_path"],
                    "size": r["file_size"],
                    "mime": r["file_mime"],
                    "url": f"{BASE_PATH}/uploads/{r['file_path']}",
                    "sender": r["display_name"],
                    "sent_at": r["created_at"],
                })

        now = datetime.now(timezone.utc).isoformat()
        first_mid = rows[0]["id"]
        last_mid = rows[-1]["id"]
        period_start = rows[0]["created_at"]
        period_end = rows[-1]["created_at"]

        cur = db.execute("""
            INSERT INTO project_history
                (room_id, period_start, period_end, first_message_id, last_message_id,
                 summary_text, message_count, attachment_count, attachments_json,
                 model, input_tokens, output_tokens, cost_usd,
                 created_by, created_at)
            VALUES (?,?,?,?,?, ?,?,?,?, ?,?,?,?, ?,?)
        """, (
            room_id, period_start, period_end, first_mid, last_mid,
            result["summary_text"], len(rows), len(attachments),
            json.dumps(attachments, ensure_ascii=False),
            result["model"], result["in_tokens"], result["out_tokens"], result["cost_usd"],
            created_by_uid, now,
        ))
        hid = cur.lastrowid
        db.commit()
        # WORKS 자동전송 (게이팅: env URL + 관리자 토글 둘 다 ON 일 때만).
        #  best-effort — 전송 실패/미설정이어도 스냅샷 생성에는 영향 없음. (대표 지시 2026-06-01)
        synced_now = False
        try:
            if _works_history_sync_enabled():
                synced_now = _send_history_to_works(hid)
        except Exception as _we:
            print(f"[works_sync] history {hid} send exception: {_we}")
        return {
            "id": hid,
            "room_id": room_id,
            "period_start": period_start,
            "period_end": period_end,
            "first_message_id": first_mid,
            "last_message_id": last_mid,
            "summary_text": result["summary_text"],
            "message_count": len(rows),
            "attachment_count": len(attachments),
            "attachments": attachments,
            "model": result["model"],
            "cost_usd": result["cost_usd"],
            "created_by": created_by_uid,
            "created_at": now,
            "synced_to_hw": 1 if synced_now else 0,
        }, None
    finally:
        db.close()


# ──────────────────────────────────────────────────────────────
#  HAIST WORKS 프로젝트 이력 자동전송 (대표 지시 2026-06-01 · 01세션 협의 발주)
#   · 게이팅: env KNK_WORKS_HISTORY_URL + 관리자 토글(app_settings.works_history_sync=1) '둘 다' ON
#   · 인증: 기존 SSO RS256 개인키로 서명한 단기 서비스 JWT (aud=haist-works)
#   · best-effort: 실패해도 스냅샷·요약 기능엔 영향 없음. 미전송 건은 하루 1회 재시도 sweep.
#   · WORKS 엔드포인트 확정 전(URL 미설정)엔 전 구간 no-op → 지금은 영향 없음(미리 구현).
# ──────────────────────────────────────────────────────────────
def _works_history_sync_enabled():
    if not WORKS_HISTORY_URL:
        return False
    try:
        db = sqlite3.connect(DB_PATH); db.row_factory = sqlite3.Row
        try:
            r = db.execute("SELECT value FROM app_settings WHERE key='works_history_sync'").fetchone()
            return bool(r and str(r["value"]) == "1")
        finally:
            db.close()
    except Exception:
        return False


def _works_service_token():
    """WORKS 수신 API 호출용 단기 서비스 JWT (RS256, 기존 SSO 개인키 재사용)."""
    import time as _t
    now = int(_t.time())
    payload = {
        "iss": "knk-messenger",
        "aud": "haist-works",
        "purpose": "project_history_sync",
        "iat": now,
        "exp": now + 120,
    }
    return _pyjwt.encode(payload, _sso_load_private_key(), algorithm="RS256")


def _send_history_to_works(history_id):
    """프로젝트 이력 1건을 HAIST WORKS 수신 API로 전송. 성공 시 synced_to_hw=1·synced_at 기록.
    반환 True/False. 예외는 던지지 않음(호출부 보호)."""
    if not WORKS_HISTORY_URL:
        return False
    import urllib.request
    db = sqlite3.connect(DB_PATH); db.row_factory = sqlite3.Row
    try:
        h = db.execute("SELECT * FROM project_history WHERE id=?", (history_id,)).fetchone()
        if not h:
            return False
        item = db.execute("SELECT code, name, customer, status FROM items WHERE room_id=?", (h["room_id"],)).fetchone()
        try:
            atts = json.loads(h["attachments_json"] or "[]")
        except Exception:
            atts = []
        payload = {
            "source": "knk-messenger",
            "history_id": h["id"],
            "room_id": h["room_id"],
            "project": (dict(item) if item else None),   # {code,name,customer,status}
            "period": {"start": h["period_start"], "end": h["period_end"]},
            "message_count": h["message_count"],
            "attachment_count": h["attachment_count"],
            "summary_text": h["summary_text"],
            "attachments": atts,
            "model": h["model"],
            "created_mode": ("auto" if h["created_by"] is None else "manual"),
            "created_at": h["created_at"],
        }
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(
            WORKS_HISTORY_URL, data=body, method="POST",
            headers={
                "Content-Type": "application/json; charset=utf-8",
                "Authorization": "Bearer " + _works_service_token(),
            },
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            ok = (200 <= getattr(resp, "status", 200) < 300)
        if ok:
            synced_at = datetime.now(timezone.utc).isoformat()
            db.execute("UPDATE project_history SET synced_to_hw=1, synced_at=? WHERE id=?", (synced_at, history_id))
            db.commit()
            return True
        return False
    except Exception as e:
        print(f"[works_sync] send error (history {history_id}): {e}")
        return False
    finally:
        db.close()


def _works_retry_unsynced(limit=50):
    """미전송(synced_to_hw=0) 이력을 재시도 (하루 1회 worker 에서 호출). 게이팅 OFF 면 no-op."""
    if not _works_history_sync_enabled():
        return 0
    db = sqlite3.connect(DB_PATH); db.row_factory = sqlite3.Row
    try:
        rows = db.execute(
            "SELECT id FROM project_history WHERE synced_to_hw=0 ORDER BY id ASC LIMIT ?",
            (limit,),
        ).fetchall()
    finally:
        db.close()
    sent = 0
    for r in rows:
        try:
            if _send_history_to_works(r["id"]):
                sent += 1
        except Exception:
            pass
    if sent:
        print(f"[works_sync] retry sweep — {sent} sent")
    return sent


def _auto_generate_project_histories():
    """모든 프로젝트 방에 대해 자동 이력 생성 (하루 1회 호출)."""
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    try:
        rooms = db.execute("""
            SELECT id, name FROM rooms WHERE type='item'
        """).fetchall()
    finally:
        db.close()
    generated = 0
    for r in rooms:
        try:
            hist, err = _generate_project_history(r["id"], created_by_uid=None)
            if hist:
                generated += 1
                print(f"[project_history] auto-generated for room {r['id']} ({r['name']})")
            elif err and err not in ("no_new", "too_few"):
                print(f"[project_history] room {r['id']} failed: {err}")
        except Exception as e:
            print(f"[project_history] room {r['id']} exception: {e}")
    print(f"[project_history] auto run complete — generated {generated} snapshots")
    return generated


# 자동 이력 워커 — 하루 1회 (24시간 = 86400 초)
_phist_thread_started = False
def _start_project_history_worker():
    global _phist_thread_started
    if _phist_thread_started:
        return
    _phist_thread_started = True
    def _loop():
        import time as _t
        # 서버 시작 직후 30분 대기 (안정화 후 첫 실행)
        _t.sleep(30 * 60)
        while True:
            try:
                _auto_generate_project_histories()
            except Exception as e:
                print(f"[project_history worker] error: {e}")
            # 미전송 이력 재시도 (WORKS 자동전송 게이팅 OFF 면 no-op)
            try:
                _works_retry_unsynced()
            except Exception as e:
                print(f"[works_sync] retry sweep error: {e}")
            _t.sleep(24 * 60 * 60)   # 24시간
    import threading as _th
    _th.Thread(target=_loop, daemon=True).start()


REWRITE_TONES = {
    "formal":       "정중한 공식 비즈니스 한국어 (존댓말, ~합니다체)",
    "short":        "동일한 의미를 유지하면서 가능한 한 짧게 (1~2 문장)",
    "professional": "기술·업무 보고서 톤. 간결하고 명확하게",
    "casual":       "동료 간 가벼운 말투 (~네요, ~할게요)",
    "polite":       "외부 거래처에 보내는 정중한 한국어",
}


def _claude_rewrite(text, tone="formal"):
    """작성 톤 조정 (AI 작성 도움)."""
    try:
        import anthropic
    except ImportError:
        return None, "anthropic SDK 미설치"
    if not ANTHROPIC_API_KEY:
        return None, "ANTHROPIC_API_KEY 환경변수 미설정"
    text = (text or "").strip()
    if not text:
        return None, "재작성할 텍스트가 비어있음"
    tone_desc = REWRITE_TONES.get(tone, REWRITE_TONES["formal"])
    system_prompt = f"""You are a Korean business writing assistant for KNK Corporation.

Rewrite the user's draft message in the following tone:
{tone_desc}

Rules:
- 원문 의미·정보는 100% 보존.
- 기술 용어·품번(003M2501, WP-LOA 등)·고객사명·@멘션·이모지는 그대로.
- 한국어로만 출력. 설명·메모 없이 다듬은 메시지 본문만 출력.
- 1줄 입력은 1줄 출력. 여러 줄은 줄바꿈 유지.
"""
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    try:
        msg = client.messages.create(
            model=AI_SUMMARY_MODEL,
            max_tokens=1024,
            system=[
                {"type": "text", "text": system_prompt, "cache_control": {"type": "ephemeral"}},
            ],
            messages=[
                {"role": "user", "content": text},
            ],
        )
        rewritten = "".join(b.text for b in msg.content if hasattr(b, "text")).strip()
        in_t = msg.usage.input_tokens
        out_t = msg.usage.output_tokens
        cost = (in_t / 1_000_000.0) * 1.0 + (out_t / 1_000_000.0) * 5.0
        return {
            "text": rewritten,
            "in_tokens": in_t,
            "out_tokens": out_t,
            "cost_usd": cost,
            "model": AI_SUMMARY_MODEL,
        }, None
    except Exception as e:
        return None, f"Claude API 오류: {e}"


# ---------- AI 번역 (Claude Haiku) ----------
# KNK 업계(사출·금형·메탈·도면) 용어 사전. 번역 일관성 + 정확도 향상.
KNK_GLOSSARY = """
[KNK 업무 용어 — 번역 시 일관성 유지]  (zh = 중국어 통용 표기, 의역 금지)
- 사출 = injection molding (vi: ép phun nhựa, zh: 注塑)
- 금형 = mold (vi: khuôn mẫu, zh: 模具)
- 메탈 = metal (vi: kim loại, zh: 金属)
- 도면 = drawing / blueprint (vi: bản vẽ, zh: 图纸)
- 납기 = delivery date / due date (vi: ngày giao hàng, zh: 交货期)
- 검사기 = inspection machine / inspector (vi: máy kiểm tra, zh: 检测机/检测设备)
- 검사장비 = inspection equipment (zh: 检测设备)
- 자동화 = automation (zh: 自动化)
- 비전검사 = vision inspection (zh: 视觉检测)
- 센서 = sensor (zh: 传感器)
- 지그 / 治具 = jig / fixture (zh: 治具/夹具)
- 치수 = dimension (vi: kích thước, zh: 尺寸)
- 공차 = tolerance (zh: 公差)
- LOA = Letter of Agreement
- 검수 = inspection / acceptance (vi: nghiệm thu, zh: 验收)
- 품번 = part number (vi: mã sản phẩm, zh: 零件号 — 단, 실제 품번 코드는 원문 유지)
- 발주 = purchase order (vi: đặt hàng, zh: 采购订单)
- 견적 = quotation (vi: báo giá, zh: 报价)
- 협력사 = supplier / partner (vi: nhà cung cấp, zh: 供应商)
- 고객사 = customer (vi: khách hàng, zh: 客户)
- 케이엔케이 / KNK = ㈜케이엔케이 (zh: KNK — 영문 유지)
- 하이스트 / HAIST = HAIST Innovation (zh: HAIST — 영문 유지)
- 베트남법인 = Vietnam branch (vi: chi nhánh Việt Nam, zh: 越南分公司)
[직급 — 베트남법인 확정 번역. 메시지에 직급이 나오면 반드시 이 용어로 일관 번역]
- 대표이사 = CEO (vi: Tổng Giám đốc)
- 전무이사/전무 = Senior Managing Director (vi: Phó Tổng Giám đốc)
- 상무이사/상무 = Managing Director (vi: Giám đốc điều hành)
- 이사 = Director (vi: Giám đốc; 팀 담당이면 Giám đốc phụ trách phòng ban)
- 법인장 = General Director (vi: Giám đốc pháp nhân / Giám đốc chi nhánh)
- 부장 = General Manager (vi: Trưởng phòng / Trưởng bộ phận)
- 차장 = Deputy General Manager (vi: Phó Trưởng Bộ Phận)
- 과장 = Manager (vi: Quản đốc)
- 대리 = Assistant Manager (vi: Phó quản đốc)
- 주임 = Supervisor (vi: Chủ nhiệm)
- 프로 = Pro (vi: Chuyên viên)
- 매니저 = Manager (vi: Quản lý)
- 팀장 = Team Lead (vi: Trưởng nhóm)
- 사원 = Staff (vi: Nhân viên)
- 반장 = Line Supervisor (vi: Trưởng chuyền) — 라인 전체 총괄
- 조장 = Section Leader (vi: Tổ trưởng) — 라인 내 특정 파트(조) 책임
"""


def _zh_translate_rules(zh_variant=None):
    """중국어 번역 추가 규칙. zh_variant=='traditional' 이면 번체, 그 외 간체(기본).
    (대표 지시 2026-05-31: 간체 기본 / 직전 상대 메시지 번체면 번체 / 您 격식 / 전문용어 통용표기 / 이름·부서 영문)"""
    if zh_variant == "traditional":
        script = "Traditional Chinese (繁體中文) — use Traditional characters."
    else:
        script = "Simplified Chinese (简体中文) — use Simplified characters."
    return f"""[Chinese (中文) output rules — STRICT]
- Script: {script}
  (Default is Simplified; use Traditional ONLY when told the recipient's recent message was Traditional.)
- Register: formal business Chinese. Address the reader with the honorific 您 (NOT 你).
- Technical terms (inspection equipment / automation): use the STANDARD industry term, never paraphrase.
  검사기→检测机/检测设备 · 검사장비→检测设备 · 금형→模具 · 사출→注塑 · 치수→尺寸 · 공차→公差 ·
  도면→图纸 · 자동화→自动化 · 비전검사→视觉检测 · 지그/치구→治具 · 납기→交货期 · 발주→采购订单 · 견적→报价 · 검수→验收
- Keep personal NAMES, DEPARTMENT names, part-number codes, brand names (KNK/HAIST) in original/English — do NOT translate them.
"""


# 번체(繁體) 전용 고빈도 한자 — 간체엔 없는 형태. 직전 상대 메시지 번체 판정용. (대표 지시 2026-05-31)
_TRAD_ONLY_CHARS = frozenset(
    "們國這個時後來對開關實點為應員機檢測設備數體認確請與區號樣義產處發變還進過動務圖"
    "預訊結約紙線經親觀語讀際報價標樓萬學寫顯餘麗齊膽臺灣顧電話頭題種樂業東車専專門"
    "閉雙構樣樹橋歡權歲歷濟營績織總統綜緊縣繼績罷聯聲職聽肅膚與興舉舊艱華蓋藥處號補"
)


def _detect_han_variant(text):
    """텍스트가 '번체 중국어'로 보이면 'traditional', 아니면 None(간체/기타).
    ★ 한글이 섞여 있으면 한국어로 보고 None — 한국어 한자(漢字)는 번체형이라 오판 방지."""
    if not text:
        return None
    for ch in text:
        if "가" <= ch <= "힣":   # 한글 음절 → 한국어 메시지 → 판정 안 함
            return None
    trad = sum(1 for ch in text if ch in _TRAD_ONLY_CHARS)
    return "traditional" if trad >= 1 else None


# ---------- OpenAI (ChatGPT) 공급자 ----------
# Anthropic 결제 활성 전까지 기본 공급자. 동일 시그니처·반환값으로 _claude_* 와 호환.
# 모델 기본 = gpt-4o-mini (저렴·빠름·번역 품질 충분). 환경변수 KNK_MSG_OPENAI_MODEL 로 교체 가능.
# (2026-06-01 gpt-5.4-mini 로 교체 시도했으나 OpenAI 가 해당 ID 를 인식하지 못해 번역 실패 → 롤백.
#  새 모델로 바꾸려면 OpenAI 카탈로그의 정확한 모델 ID 확인 후 KNK_MSG_OPENAI_MODEL 또는 이 기본값 교체.)

# gpt-4o-mini 가격 (2026-05 기준): $0.15/MTok input, $0.60/MTok output
# gpt-4o 가격: $2.50/MTok input, $10/MTok output
_OPENAI_PRICING = {
    "gpt-4o-mini":     (0.15, 0.60),
    "gpt-4o":          (2.50, 10.00),
    "gpt-4o-2024-08-06": (2.50, 10.00),
    "gpt-4-turbo":     (10.0, 30.0),
    "gpt-3.5-turbo":   (0.50, 1.50),
}


def _openai_client():
    """OpenAI 클라이언트 생성. 미설치·미설정이면 (None, err)."""
    try:
        from openai import OpenAI
    except ImportError:
        return None, "openai SDK 미설치 (requirements.txt 의 openai 설치 필요)"
    if not OPENAI_API_KEY:
        return None, "OPENAI_API_KEY 환경변수 미설정 (OpenAI키설정.bat 실행 후 재시작)"
    kwargs = {"api_key": OPENAI_API_KEY}
    if OPENAI_BASE_URL:
        kwargs["base_url"] = OPENAI_BASE_URL
    return OpenAI(**kwargs), None


# ── 신형 OpenAI 모델 파라미터 자동 대응 (가이드 §1, WORKS 동일 구현, 대표 지시 2026-06-01) ──
#  GPT-5 / o 계열 신형 모델은 max_tokens 대신 max_completion_tokens 를 요구하고,
#  temperature 미지원일 수 있음. 구·신형 양쪽을 자동 처리.
#  ※ '파라미터 오류'일 때만 변형 재시도 — 진짜 '모델 없음/키 오류'는 즉시 그대로 전달(조용히 죽지 않게).
def openai_create(client, model, messages, max_tokens=1024, temperature=0.2):
    variants = [
        {"max_tokens": max_tokens, "temperature": temperature},             # 구형 표준
        {"max_completion_tokens": max_tokens, "temperature": temperature},  # 신형
        {"max_completion_tokens": max_tokens},                              # 신형 + temperature 미지원
        {"max_tokens": max_tokens},                                         # 구형 + temperature 미지원
    ]
    last = None
    for v in variants:
        try:
            return client.chat.completions.create(model=model, messages=messages, **v)
        except Exception as e:
            last = e
            m = str(e).lower()
            is_param = ("unsupported_parameter" in m or "not supported" in m
                        or "max_completion_tokens" in m or "max_tokens" in m
                        or "temperature" in m or "unsupported value" in m)
            if not is_param:
                raise   # 파라미터 문제 아님(모델없음·키오류 등) → 그대로 올림
    raise last


# ── 현재 사용할 OpenAI 모델 결정 — 우선순위: 환경변수 > app_settings.ai_model > 코드 기본값 ──
#  관리자 화면에서 app_settings.ai_model 변경 시 재시작 없이 즉시 반영. (대표 지시 2026-06-01)
_OPENAI_MODEL_ENV = os.environ.get("KNK_MSG_OPENAI_MODEL", "").strip()
def _get_openai_model(db=None):
    if _OPENAI_MODEL_ENV:
        return _OPENAI_MODEL_ENV   # 환경변수 명시 시 잠금(최우선)
    try:
        if db is None:
            db = get_db()
        r = db.execute("SELECT value FROM app_settings WHERE key='ai_model'").fetchone()
        if r and str(r["value"]).strip():
            return str(r["value"]).strip()
    except Exception:
        pass
    return OPENAI_MODEL


def _openai_calc_cost(model, in_tokens, out_tokens):
    """OpenAI 사용량을 USD 로 환산. 가격표에 없는 모델은 0 으로 간주(추후 보정)."""
    rates = _OPENAI_PRICING.get(model)
    if not rates:
        # 알 수 없는 모델 → 가장 가까운 4o-mini 가격으로 fallback (관리자에게 알리기 위해 약간 비싸게)
        rates = _OPENAI_PRICING.get("gpt-4o-mini", (0.15, 0.60))
    in_rate, out_rate = rates
    return (in_tokens / 1_000_000.0) * in_rate + (out_tokens / 1_000_000.0) * out_rate


def _openai_translate(text, target_lang_code, zh_variant=None):
    """ChatGPT 호출. 시그니처·반환값은 _claude_translate 와 동일."""
    target_name = TRANSLATE_LANGS.get(target_lang_code)
    if not target_name:
        return None, f"지원 안 되는 언어: {target_lang_code}"

    client, err = _openai_client()
    if err:
        return None, err

    zh_block = _zh_translate_rules(zh_variant) if target_lang_code == "zh" else ""
    system_prompt = f"""You are a professional Korean / Vietnamese / English / Chinese business translator for KNK Corporation, a Korean industrial machinery and inspection equipment company with a Vietnam branch.

{KNK_GLOSSARY}
{zh_block}
Rules:
1. Translate ONLY into {target_name}. Do NOT add explanations or notes.
2. Preserve technical part numbers (e.g. 003M2501, WP-LOA), brand names, file paths, URLs, @mentions.
3. Use formal business tone (한국어: 존댓말, 베트남어: anh/chị + ạ, English: professional, 中文: 您 honorific).
4. If the source is already in {target_name}, return the source as-is (do not translate to itself).
5. Keep emoji and formatting (newlines, bullets) as-is.
6. Output ONLY the translated text, nothing else."""

    try:
        model = _get_openai_model()
        resp = openai_create(
            client, model,
            [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"Translate the following to {target_name}:\n\n{text}"},
            ],
            max_tokens=1024, temperature=0.2,
        )
        translated = (resp.choices[0].message.content or "").strip()
        in_t = resp.usage.prompt_tokens
        out_t = resp.usage.completion_tokens
        cost = _openai_calc_cost(model, in_t, out_t)
        return (translated, None, in_t, out_t, cost), None
    except Exception as e:
        return None, f"OpenAI API 오류: {e}"


def _openai_summarize_messages(messages_payload, mode="channel"):
    """ChatGPT 호출 — 채널·스레드 요약. 시그니처·반환값은 _claude_summarize_messages 와 동일."""
    client, err = _openai_client()
    if err:
        return None, err
    if not messages_payload:
        return None, "요약할 메시지가 없습니다"

    # transcript 빌드 + 사진(문서 캡처 포함) 비전 첨부 수집 — gpt-4o-mini 비전으로 이미지 속 글자까지 읽음 (대표 지시 2026-05-31)
    import base64 as _b64
    from io import BytesIO as _BIO
    lines = []
    image_blocks = []   # OpenAI 비전 content 블록
    IMG_CAP = 3         # 요약당 분석 사진 최대 장수 — gpt-4o-mini 비전은 장당 토큰이 커서(고해상도 ~2.5만) 컨텍스트(128k) 보호
    img_n = 0
    img_total = 0
    for m in messages_payload:
        if m.get("kind") == "system":
            continue
        ts = m.get("created_at", "")
        try:
            from datetime import datetime as _dt
            dt = _dt.fromisoformat(ts.replace("Z", "+00:00")) if ts else None
            ts_short = dt.strftime("%m-%d %H:%M") if dt else ""
        except Exception:
            ts_short = ts[:16]
        name = m.get("display_name", "?")
        content = (m.get("content") or "").strip()
        # 사진 메시지 → 실제 이미지를 비전으로 첨부해 글자(문서 캡처 등)까지 읽게 함
        if m.get("kind") == "image" and m.get("file_path"):
            img_total += 1
            if img_n < IMG_CAP:
                try:
                    _src = os.path.join(UPLOAD_DIR, m["file_path"])
                    if os.path.exists(_src):
                        from PIL import Image as _PImg
                        _im = _PImg.open(_src).convert("RGB")
                        _im.thumbnail((1024, 1024))   # 문서 글자 가독성 + 토큰 절약(mini 비전 토큰 큼)
                        _bio = _BIO(); _im.save(_bio, format="JPEG", quality=80); _bio.seek(0)
                        _b = _b64.b64encode(_bio.read()).decode("ascii")
                        img_n += 1
                        lines.append(f"[{ts_short}] {name}: [사진 {img_n}]{(' ' + content) if content else ''}")
                        image_blocks.append({
                            "type": "image_url",
                            "image_url": {"url": f"data:image/jpeg;base64,{_b}", "detail": "high"},
                        })
                        continue
                except Exception:
                    pass
        if not content:
            content = f"[{m.get('kind','파일')}]"
        lines.append(f"[{ts_short}] {name}: {content}")
    if img_total > img_n:
        lines.append(f"(참고: 사진 {img_total}장 중 {img_n}장만 분석에 포함됨 — 비용 보호)")
    transcript = "\n".join(lines)
    if not transcript:
        return None, "요약할 본문이 비어있음"

    mode_label = "스레드 토론" if mode == "thread" else "채팅방 대화"
    system_prompt = f"""You are a Korean business communication assistant for KNK Corporation (industrial machinery / inspection equipment).

Your job: Summarize the following Korean {mode_label} for a busy executive (대표/임원) as a DETAILED, FACTUAL report.

가장 중요한 규칙 (반드시 준수):
- 오직 본문(대화 내용)에 실제로 적힌 사실만 쓴다. 추정·해석·상상·창작·과장 절대 금지.
- 본문에 근거가 없는 내용은 한 글자도 쓰지 않는다. 불확실하면 쓰지 말고, 항목에 내용이 없으면 "없음".
- 최대한 상세하게: 본문에 오간 실질 내용은 빠짐없이 담는다(분량 제한 없음). 단, 없는 내용을 지어내 채우지 않는다.

Output format (Korean — 굵은 제목 그대로 사용):
**핵심 요약**
- 대화 전체의 목적과 결론을 2~4문장으로.
**상세 내용**
- 시간/주제 흐름대로 "누가 무엇을 말했고 어떤 일이 오갔는지" 구체적으로 불릿 정리.
- 사진·파일 공유는 "누가 무엇을 공유함"으로 기록. 숫자·금액·날짜·요청·답변을 구체적으로.
**주요 결정사항**
- 실제로 합의·결정된 것만 불릿. 없으면 "없음".
**미결사항·후속조치**
- 남은 일·확인 필요사항. 담당자·기한이 본문에 있으면 명시. 없으면 "없음".
**관련 인물**
- 대화에 등장한 사람 이름만 나열.

표기 규칙:
- 기술 용어·품번(예: 003M2501, WP-LOA)·고객사명·금액·날짜는 본문 그대로 정확히.
- 함께 첨부된 '[사진 N]' 이미지(문서 캡처·스크린샷 포함)의 글자·표·내용도 읽어 사실대로 반영한다. 사진에서 읽은 것도 추측 없이 보이는 그대로만.
- 정중한 보고서 평어체: "~함", "~확인 필요" 등."""

    _utext = (f"다음 {mode_label}을 사실 그대로, 최대한 상세히 요약해 주세요. "
              f"함께 첨부된 '[사진 N]' 이미지(문서 캡처 포함)의 글자/내용도 읽어 반영해 주세요:\n\n{transcript}")
    # 사진이 있으면 비전 content 리스트, 없으면 기존처럼 문자열
    _user_content = ([{"type": "text", "text": _utext}] + image_blocks) if image_blocks else _utext

    _model = _get_openai_model()
    def _call_openai(_uc):
        return openai_create(
            client, _model,
            [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": _uc},
            ],
            max_tokens=4000, temperature=0.2,
        )

    try:
        try:
            resp = _call_openai(_user_content)
        except Exception:
            # 사진(비전) 첨부로 실패(컨텍스트 초과·모델 미지원 등) 시 → 텍스트만으로 자동 재시도. 요약은 무조건 되게.
            if image_blocks:
                resp = _call_openai(_utext)
            else:
                raise
        summary = (resp.choices[0].message.content or "").strip()
        in_t = resp.usage.prompt_tokens
        out_t = resp.usage.completion_tokens
        cost = _openai_calc_cost(_model, in_t, out_t)
        return {
            "summary_text": summary,
            "in_tokens": in_t,
            "out_tokens": out_t,
            "cost_usd": cost,
            "model": _model,
        }, None
    except Exception as e:
        return None, f"OpenAI API 오류: {e}"


def _openai_summarize_for_history(messages_payload, item_meta=None):
    """ChatGPT 호출 — 프로젝트 이력 요약. 시그니처는 _claude_summarize_for_history 와 동일."""
    client, err = _openai_client()
    if err:
        return None, err
    if not messages_payload:
        return None, "요약할 메시지가 없습니다"

    # transcript 빌드
    lines = []
    for m in messages_payload:
        if m.get("kind") == "system":
            continue
        ts = m.get("created_at", "")
        try:
            from datetime import datetime as _dt
            dt = _dt.fromisoformat(ts.replace("Z", "+00:00")) if ts else None
            ts_short = dt.strftime("%m-%d %H:%M") if dt else ""
        except Exception:
            ts_short = ts[:16]
        name = m.get("display_name", "?")
        content = (m.get("content") or "").strip()
        kind = m.get("kind", "text")
        file_name = m.get("file_name")
        if kind == "image" and file_name:
            content = f"[사진: {file_name}] " + content
        elif kind == "file" and file_name:
            content = f"[파일: {file_name}] " + content
        if not content:
            content = f"[{kind}]"
        lines.append(f"[{ts_short}] {name}: {content}")
    transcript = "\n".join(lines)
    if not transcript:
        return None, "요약할 본문이 비어있음"

    item_ctx = ""
    if item_meta:
        bits = []
        if item_meta.get("code"): bits.append(f"품번 {item_meta['code']}")
        if item_meta.get("name"): bits.append(f"프로젝트명 {item_meta['name']}")
        if item_meta.get("customer"): bits.append(f"고객사 {item_meta['customer']}")
        if item_meta.get("status"): bits.append(f"상태 {item_meta['status']}")
        if bits:
            item_ctx = "\n[프로젝트 컨텍스트] " + " · ".join(bits) + "\n"

    system_prompt = f"""You are a Korean business communication assistant for KNK Corporation (industrial machinery / inspection equipment).

Your job: Summarize the following Korean project conversation as PROJECT HISTORY for HAIST WORKS (ERP system).
This summary will be saved as the official project log.

Output structure (Korean, in this exact order):

**기간 요약** (1~2 문장 — 이번 기간에 무엇이 진행되었는지 핵심)

**주요 결정사항**
- (담당자) 결정 내용 (날짜)
- ...

**진척 상황**
- 이번 기간 완료된 작업
- 진행 중인 작업

**미결 사항 / 후속조치**
- (담당자) 해야 할 것 (기한)
- ...

**관련 인물**
- 이름 나열 (쉼표 구분)

**언급된 외부 거래처·품번**
- 거래처명, 품번 등 (있으면)

Rules:
- 최대 600 한글 글자.
- 기술 용어·품번(예: 003M2501, WP-LOA)·고객사명은 원문 그대로.
- 추측 금지. 본문에 없으면 "없음".
- 정중한 평어체 (보고서 톤). "~함", "~확인 필요", "~예정".
- 시간은 가능하면 명시 (예: 5월 17일 14:00).
{item_ctx}"""

    try:
        _model = _get_openai_model()
        resp = openai_create(
            client, _model,
            [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"다음 프로젝트 대화를 이력 보고서로 요약해 주세요:\n\n{transcript}"},
            ],
            max_tokens=2048, temperature=0.3,
        )
        summary = (resp.choices[0].message.content or "").strip()
        in_t = resp.usage.prompt_tokens
        out_t = resp.usage.completion_tokens
        cost = _openai_calc_cost(_model, in_t, out_t)
        return {
            "summary_text": summary,
            "in_tokens": in_t,
            "out_tokens": out_t,
            "cost_usd": cost,
            "model": _model,
        }, None
    except Exception as e:
        return None, f"OpenAI API 오류: {e}"


# ---------- 공급자 라우터 ----------
# 모든 호출부는 _ai_* 를 사용. 내부적으로 TRANSLATE_PROVIDER 에 따라 분기.

def _ai_translate(text, target_lang_code, zh_variant=None):
    """공급자 자동 선택. 데모 모드면 _claude_translate 의 mock 분기를 그대로 사용.
       전사 토글(app_settings.ai_translate_enabled = 0) 시 차단. (대표 지시 2026-05-28)
       zh_variant: 중국어 대상일 때 'traditional'(번체)/None(간체 기본). (대표 지시 2026-05-31)"""
    # 전사 번역 토글 확인
    try:
        if not _ai_translate_enabled(get_db()):
            return None, "AI 번역이 관리자에 의해 비활성화되어 있습니다."
    except Exception:
        pass
    if TRANSLATE_MOCK:
        return _claude_translate(text, target_lang_code, zh_variant)  # _claude_translate 안의 mock 사용
    if TRANSLATE_PROVIDER == "openai":
        return _openai_translate(text, target_lang_code, zh_variant)
    return _claude_translate(text, target_lang_code, zh_variant)


def _ai_summarize_messages(messages_payload, mode="channel"):
    if TRANSLATE_PROVIDER == "openai":
        return _openai_summarize_messages(messages_payload, mode=mode)
    return _claude_summarize_messages(messages_payload, mode=mode)


def _ai_summarize_for_history(messages_payload, item_meta=None):
    if TRANSLATE_PROVIDER == "openai":
        return _openai_summarize_for_history(messages_payload, item_meta=item_meta)
    return _claude_summarize_for_history(messages_payload, item_meta=item_meta)


def _claude_translate(text, target_lang_code, zh_variant=None):
    """Claude Haiku 호출. 성공 시 (translated_text, source_lang, in_tokens, out_tokens, cost_usd)."""
    target_name = TRANSLATE_LANGS.get(target_lang_code)
    if not target_name:
        return None, f"지원 안 되는 언어: {target_lang_code}"

    # === 데모 모드: API 키 없이 UI 흐름 테스트 (무료) ===
    if TRANSLATE_MOCK:
        # 간단한 KNK 핵심 용어 미니 사전 — 실제 번역처럼 보이는 데모용 변환
        mini_dict = {
            "vi": {
                "안녕하세요": "Xin chào",
                "감사합니다": "Cảm ơn",
                "도면": "bản vẽ",
                "사출": "ép phun nhựa",
                "금형": "khuôn mẫu",
                "납기": "ngày giao hàng",
                "오늘": "hôm nay",
                "내일": "ngày mai",
                "부탁드립니다": "vui lòng",
                "확인": "xác nhận",
                "메탈": "kim loại",
                "검사기": "máy kiểm tra",
                "치수": "kích thước",
                "수정": "sửa đổi",
            },
            "en": {
                "안녕하세요": "Hello",
                "감사합니다": "Thank you",
                "도면": "drawing",
                "사출": "injection molding",
                "금형": "mold",
                "납기": "due date",
                "오늘": "today",
                "내일": "tomorrow",
                "부탁드립니다": "please",
                "확인": "confirm",
                "메탈": "metal",
                "검사기": "inspection machine",
                "치수": "dimension",
                "수정": "modify",
            },
            "ko": {
                "Xin chào": "안녕하세요",
                "Cảm ơn": "감사합니다",
                "Hello": "안녕하세요",
                "Thank you": "감사합니다",
            },
            "zh": {
                "안녕하세요": "您好",
                "감사합니다": "谢谢",
                "도면": "图纸",
                "사출": "注塑",
                "금형": "模具",
                "납기": "交货期",
                "검사기": "检测机",
                "치수": "尺寸",
                "확인": "确认",
            },
        }
        translated = text
        for k, v in mini_dict.get(target_lang_code, {}).items():
            translated = translated.replace(k, v)
        # 실제 번역과 구분되도록 명시
        translated = f"[데모 {target_lang_code.upper()}] {translated}"
        return (translated, None, 0, 0, 0.0), None

    # === 실제 Claude API 호출 ===
    try:
        import anthropic
    except ImportError:
        return None, "anthropic SDK 미설치 (requirements.txt 의 anthropic 설치 필요)"

    if not ANTHROPIC_API_KEY:
        return None, "ANTHROPIC_API_KEY 환경변수 미설정"

    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    zh_block = _zh_translate_rules(zh_variant) if target_lang_code == "zh" else ""
    system_prompt = f"""You are a professional Korean / Vietnamese / English / Chinese business translator for KNK Corporation, a Korean industrial machinery and inspection equipment company with a Vietnam branch.

{KNK_GLOSSARY}
{zh_block}
Rules:
1. Translate ONLY into {target_name}. Do NOT add explanations or notes.
2. Preserve technical part numbers (e.g. 003M2501, WP-LOA), brand names, file paths, URLs, @mentions.
3. Use formal business tone (한국어: 존댓말, 베트남어: anh/chị + ạ, English: professional, 中文: 您 honorific).
4. If the source is already in {target_name}, return the source as-is (do not translate to itself).
5. Keep emoji and formatting (newlines, bullets) as-is.
6. Output ONLY the translated text, nothing else.
"""

    try:
        msg = client.messages.create(
            model=TRANSLATE_MODEL,
            max_tokens=1024,
            system=[
                {"type": "text", "text": system_prompt, "cache_control": {"type": "ephemeral"}},
            ],
            messages=[
                {"role": "user", "content": f"Translate the following to {target_name}:\n\n{text}"},
            ],
        )
        translated = "".join(b.text for b in msg.content if hasattr(b, "text")).strip()
        in_t = msg.usage.input_tokens
        out_t = msg.usage.output_tokens
        # Haiku 4.5 가격: $1/MTok input, $5/MTok output (대략)
        cost = (in_t / 1_000_000.0) * 1.0 + (out_t / 1_000_000.0) * 5.0
        return (translated, None, in_t, out_t, cost), None
    except Exception as e:
        return None, f"Claude API 오류: {e}"


def _translate_monthly_cost():
    """이번 달 누적 번역 비용 (USD)."""
    db = get_db()
    row = db.execute(
        "SELECT COALESCE(SUM(cost_usd), 0) AS total FROM message_translations WHERE date(created_at) >= date('now', 'start of month')"
    ).fetchone()
    return float(row["total"] or 0)


def _can_use_ai_summary(db, room_id, user):
    """AI 요약·프로젝트 이력 사용 권한.
       (대표 지시 2026-05-27 → 2026-05-28 직원별 권한 / v2 ceo 자동통과 제거 / v3 게스트 차단)
       허용 조건 OR:
         1) 해당 방의 방장(host) / PM(sub_host) — 방 단위 자동 허용
         2) users.ai_summary_allowed = 1 인 직원 (관리자가 명시적 부여, ceo 포함)
       ※ 게스트(외부)는 무조건 차단."""
    if not user:
        return False
    # 게스트 차단 (대표 지시 2026-05-28)
    if _is_guest(user):
        return False
    # 1) 방장 / PM
    role = _my_room_role(db, room_id, user["id"])
    if role in ("host", "sub_host"):
        return True
    # 2) 명시적 권한 부여 (ceo 도 동일하게 이 컬럼 사용)
    try:
        if int(user["ai_summary_allowed"] or 0) == 1:
            return True
    except Exception:
        pass
    return False


# ──────────────────────────────────────────────────────────────
#  AI 번역 전사 토글 — app_settings.ai_translate_enabled (대표 지시 2026-05-28)
#    · 기본값 '1' (활성)
#    · 0 으로 설정 시 모든 번역 호출 차단
# ──────────────────────────────────────────────────────────────
def _ai_translate_enabled(db):
    try:
        r = db.execute("SELECT value FROM app_settings WHERE key='ai_translate_enabled'").fetchone()
        if r is None:
            return True  # 미설정 = 활성
        return str(r["value"]) == "1"
    except Exception:
        return True


# ──────────────────────────────────────────────────────────────
#  상태 표시 상세 토글 — app_settings.presence_show_detail (대표 지시 2026-06-01)
#    · 기본값 미설정 = '0'(단순): 접속하면 '🟢 접속' 하나로 — 기기(컴퓨터/휴대폰)·회사망(회사) 숨김
#    · '1'(상세): 기존처럼 💻 컴퓨터/📱 휴대폰 + (회사) 표시
#    · 화면 '표시'만 제어 — 푸시 라우팅(PC활성→폰푸시 억제)·퇴근 자동복귀 등 내부 로직은 영향 없음
#    · 직원이 "회사망 접속이 노출돼 불안" 하다는 피드백 → 단순 표시를 기본값으로
# ──────────────────────────────────────────────────────────────
def _presence_show_detail(db):
    try:
        r = db.execute("SELECT value FROM app_settings WHERE key='presence_show_detail'").fetchone()
        if r is None:
            return False  # 미설정 = 단순(녹색)
        return str(r["value"]) == "1"
    except Exception:
        return False


def _translate_room_name(name, langs=("ko", "vi", "en"), db=None):
    """방 이름을 주어진 언어들로 자동 번역. (대표 지시 2026-05-28 → 2026-06-05 재가동·다언어)
       반환: {lang: 번역문 or None}. AI OFF/실패면 해당 값 None.
       비용: 언어당 약 0.0005달러 (짧은 텍스트). 생성·이름변경·중국고객초대 시 1회만."""
    out = {l: None for l in langs}
    if not name or not name.strip():
        return out
    if db is None:
        try:
            db = get_db()
        except Exception:
            return out
    if not _ai_translate_enabled(db):
        return out  # 전사 OFF
    for l in langs:
        if l not in ("ko", "vi", "en", "zh"):
            continue
        try:
            res, err = _ai_translate(name, l)
            if not err and res:
                txt = (res[0] or "").strip()
                if txt:
                    out[l] = txt
        except Exception:
            pass
    return out


def _room_has_zh_guest(db, room_id):
    """그 방에 중국어(zh) 고객 초대가 있는지 — 회수 안 된 것 기준. (대표 지시 2026-06-05)"""
    try:
        r = db.execute(
            "SELECT 1 FROM guest_invites WHERE room_id=? AND guest_lang='zh' AND revoked_at IS NULL LIMIT 1",
            (room_id,),
        ).fetchone()
        return bool(r)
    except Exception:
        return False


def _retranslate_room(db, room_id, langs=None):
    """방 이름을 번역해 rooms.name_ko/vi/en/zh 에 저장 (동기). 실패해도 조용히 통과(원본 유지).
       langs=None 이면 ko/vi/en (+ 중국 고객 있으면 zh). direct/self 방은 건너뜀."""
    try:
        row = db.execute("SELECT name, type FROM rooms WHERE id=?", (room_id,)).fetchone()
        if not row:
            return
        if (row["type"] if "type" in row.keys() else None) in ("direct", "self"):
            return
        name = (row["name"] or "").strip()
        if not name:
            return
        if langs is None:
            langs = ["ko", "vi", "en"]
            if _room_has_zh_guest(db, room_id):
                langs.append("zh")
        cols = [l for l in langs if l in ("ko", "vi", "en", "zh")]
        if not cols:
            return
        # 한국어 원본(한글 포함)이면 ko 칸은 원본 그대로 — AI 변형·비용 방지
        is_ko_src = bool(re.search(r"[가-힣]", name))
        translate_cols = [l for l in cols if not (l == "ko" and is_ko_src)]
        tr = _translate_room_name(name, langs=translate_cols, db=db) if translate_cols else {}
        if "ko" in cols and is_ko_src:
            tr["ko"] = name
        sets = ", ".join("name_{}=?".format(l) for l in cols)
        vals = [tr.get(l) for l in cols] + [room_id]
        db.execute("UPDATE rooms SET {} WHERE id=?".format(sets), vals)
        db.commit()
    except Exception:
        pass


# ──────────────────────────────────────────────────────────────
#  스레드 삭제/연장 권한 helper (대표 지시 2026-05-28, 옵션 ②)
#    · 권한: 방장 / PM / 관리자 (AI 요약과 동일 정책)
#    · 기간: 마지막 답글 후 30일 경과 OR archive_extended_until 지난 경우
#    · 관리자(ceo) 는 기간 무시하고 즉시 가능
#    · 답글이 0개인 스레드는 부모 메시지 생성일 기준
# ──────────────────────────────────────────────────────────────
THREAD_ARCHIVE_DAYS = 30  # 마지막 답글 후 N일이 지나면 삭제 가능

def _can_manage_thread(db, room_id, user):
    """방장 / PM / 관리자 여부 — extend 와 delete 둘 다 동일."""
    if not user:
        return False
    try:
        if str((user["role"] if hasattr(user, "keys") else user.get("role")) or "") == "ceo":
            return True
    except Exception:
        pass
    role = _my_room_role(db, room_id, user["id"])
    return role in ("host", "sub_host")

def _is_ceo(user):
    if not user:
        return False
    try:
        return str((user["role"] if hasattr(user, "keys") else user.get("role")) or "") == "ceo"
    except Exception:
        return False


def _is_guest(user):
    """외부 게스트(고객사) 여부 — users.is_guest = 1 (대표 지시 2026-05-28)."""
    if not user:
        return False
    try:
        return int(user["is_guest"] or 0) == 1
    except Exception:
        return False


def _can_invite_guest(db, room_id, user):
    """고객사 초대 발행 권한 — 그룹·프로젝트 방에서만 (1:1·채널·메모는 불가). (대표 지시 2026-05-30)
       그룹·프로젝트: 방장/PM/관리자.  게스트는 초대 불가.
       1:1 제외 이유: 사적 1:1에 늦게 초대된 외부인이 과거 대화를 모두 보게 되는 문제 방지."""
    if not user or _is_guest(user):
        return False
    room = db.execute("SELECT type FROM rooms WHERE id=?", (room_id,)).fetchone()
    if not room or room["type"] not in ("group", "item"):
        return False
    if _is_ceo(user):
        return True
    role = _my_room_role(db, room_id, user["id"])
    return role in ("host", "sub_host")

def _thread_archive_deadline(parent_created_at, last_reply_at, archive_extended_until):
    """이 시각 이후에 🗑 표시 (deletable_at). ISO 8601 문자열 반환."""
    # archive_extended_until 이 있으면 그게 우선
    if archive_extended_until:
        return archive_extended_until
    # 답글 있으면 last_reply_at + 30일, 없으면 parent_created_at + 30일
    base = last_reply_at or parent_created_at
    if not base:
        return None
    try:
        # SQLite 는 보통 'YYYY-MM-DD HH:MM:SS' 또는 ISO 형식. 두 케이스 모두 대응.
        s = base.replace("T", " ").split(".")[0]
        dt = datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S")
        deadline = dt + _timedelta(days=THREAD_ARCHIVE_DAYS)
        return deadline.strftime("%Y-%m-%dT%H:%M:%S")
    except Exception:
        return None

def _thread_is_deletable_now(parent_created_at, last_reply_at, archive_extended_until):
    """현재 시각이 archive deadline 을 지났는가."""
    deadline = _thread_archive_deadline(parent_created_at, last_reply_at, archive_extended_until)
    if not deadline:
        return False
    try:
        dl = datetime.strptime(deadline[:19].replace("T", " "), "%Y-%m-%d %H:%M:%S")
        return datetime.now() >= dl
    except Exception:
        return False


@app.route("/api/rooms/<int:room_id>/summarize", methods=["POST"])
@login_required
def api_room_summarize(room_id):
    """방의 최근 N개 메시지(또는 since 이후) AI 요약.
    body: {limit?: int=80, since?: 'YYYY-MM-DD', force?: bool}
    캐시 동작: 마지막 메시지 ID 기준 캐싱. 같은 ID 면 재생성 안 함 (비용 절감).
    권한: 테스트 기간 방장·PM·관리자만 (대표 지시 2026-05-27)."""
    me = current_user()
    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (room_id, me["id"]),
    ).fetchone():
        abort(403)
    # 게스트(고객사)는 내부 AI 요약 진입 불가 (보기·생성 모두) (대표 지시 2026-06-01)
    if _is_guest(me):
        return jsonify({"error": "외부 사용자는 AI 요약을 사용할 수 없습니다."}), 403
    data = request.get_json(silent=True) or {}
    peek = bool(data.get("peek"))            # True=저장본만 조회(AI 호출 X, 무료, 멤버면 OK)
    limit = int(data.get("limit") or 80)
    if limit < 5: limit = 5
    if limit > 300: limit = 300
    since = data.get("since")  # 'YYYY-MM-DD'
    force = bool(data.get("force"))
    scope_key0 = (f"room:{room_id}:since:{since}" if since else f"room:{room_id}:last:{limit}")

    # ── 조회 전용(peek): 저장된 요약만 반환. AI 호출 없음 → 비용 0. 같은 방 멤버는 권한 없어도 열람 가능. (대표 지시 2026-06-01) ──
    if peek:
        saved = db.execute("""
            SELECT summary_text, created_at, model, last_message_id
              FROM ai_summaries
             WHERE scope_type='channel' AND scope_key=?
             ORDER BY id DESC LIMIT 1
        """, (scope_key0,)).fetchone()
        if not saved:
            return jsonify({"has_saved": False})
        new_cnt = db.execute("""
            SELECT COUNT(*) AS c FROM messages
             WHERE room_id=? AND parent_message_id IS NULL AND kind != 'system' AND id > ?
        """, (room_id, saved["last_message_id"])).fetchone()["c"]
        return jsonify({
            "has_saved": True,
            "summary": saved["summary_text"],
            "cached": True,
            "cached_at": saved["created_at"],
            "model": saved["model"],
            "new_messages_since": new_cnt,
        })

    # ── 생성(AI 실행, 비용 발생): 권한 필요. 권한 없으면 저장본 보기만 가능. (대표 지시 2026-06-01) ──
    if not _can_use_ai_summary(db, room_id, me):
        return jsonify({"error": "AI 요약 생성 권한이 없습니다. (방장·PM·관리자만 가능 · 저장된 요약은 보실 수 있어요)"}), 403

    # 메시지 수집 — 스레드 부모만 (메인 타임라인 기준). 시스템 메시지 제외.
    if since:
        rows = db.execute("""
            SELECT m.id, m.content, m.kind, m.created_at,
                   m.file_path, m.file_mime,
                   u.display_name
              FROM messages m
              JOIN users u ON u.id = m.user_id
             WHERE m.room_id = ? AND m.parent_message_id IS NULL
               AND m.kind != 'system'
               AND date(m.created_at) >= date(?)
             ORDER BY m.id ASC
        """, (room_id, since)).fetchall()
        scope_key = f"room:{room_id}:since:{since}"
    else:
        rows = db.execute("""
            SELECT m.id, m.content, m.kind, m.created_at,
                   m.file_path, m.file_mime,
                   u.display_name
              FROM messages m
              JOIN users u ON u.id = m.user_id
             WHERE m.room_id = ? AND m.parent_message_id IS NULL
               AND m.kind != 'system'
             ORDER BY m.id DESC
             LIMIT ?
        """, (room_id, limit)).fetchall()
        rows = list(reversed(rows))   # 시간 순으로
        scope_key = f"room:{room_id}:last:{limit}"

    if not rows:
        return jsonify({"summary": "요약할 메시지가 없습니다.", "cached": False, "message_count": 0})

    last_msg_id = rows[-1]["id"]
    # 캐시 조회
    if not force:
        cached = db.execute("""
            SELECT summary_text, created_at, model, input_tokens, output_tokens
              FROM ai_summaries
             WHERE scope_type='channel' AND scope_key=? AND last_message_id=?
             ORDER BY id DESC LIMIT 1
        """, (scope_key, last_msg_id)).fetchone()
        if cached:
            return jsonify({
                "summary": cached["summary_text"],
                "cached": True,
                "cached_at": cached["created_at"],
                "model": cached["model"],
                "message_count": len(rows),
            })

    # AI 호출 (공급자 라우터: openai 또는 anthropic)
    payload = [dict(r) for r in rows]
    result, err = _ai_summarize_messages(payload, mode="channel")
    if err:
        return jsonify({"error": err}), 500
    # 캐시 저장
    now = datetime.now(timezone.utc).isoformat()
    db.execute("""
        INSERT INTO ai_summaries
            (scope_type, scope_key, room_id, last_message_id, summary_text,
             model, input_tokens, output_tokens, cost_usd, created_by, created_at)
        VALUES ('channel', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (scope_key, room_id, last_msg_id, result["summary_text"],
          result["model"], result["in_tokens"], result["out_tokens"],
          result["cost_usd"], me["id"], now))
    db.commit()
    return jsonify({
        "summary": result["summary_text"],
        "cached": False,
        "model": result["model"],
        "input_tokens": result["in_tokens"],
        "output_tokens": result["out_tokens"],
        "cost_usd": result["cost_usd"],
        "message_count": len(rows),
    })


@app.route("/api/messages/<int:message_id>/summarize_thread", methods=["POST"])
@login_required
def api_thread_summarize(message_id):
    """스레드(부모 + 답글 전체) AI 요약. 캐시는 마지막 답글 ID 기준.
    권한: 테스트 기간 방장·PM·관리자만 (대표 지시 2026-05-27)."""
    me = current_user()
    db = get_db()
    parent = db.execute("""
        SELECT m.id, m.room_id, m.content, m.kind, m.created_at,
               u.display_name
          FROM messages m JOIN users u ON u.id = m.user_id
         WHERE m.id = ?
    """, (message_id,)).fetchone()
    if not parent:
        return jsonify({"error": "not found"}), 404
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (parent["room_id"], me["id"]),
    ).fetchone():
        abort(403)
    if not _can_use_ai_summary(db, parent["room_id"], me):
        return jsonify({"error": "AI 요약은 테스트 기간 동안 방장·PM 또는 관리자만 사용할 수 있습니다."}), 403
    replies = db.execute("""
        SELECT m.id, m.content, m.kind, m.created_at,
               u.display_name
          FROM messages m JOIN users u ON u.id = m.user_id
         WHERE m.parent_message_id = ?
         ORDER BY m.id ASC
    """, (message_id,)).fetchall()
    all_rows = [dict(parent)] + [dict(r) for r in replies]
    if len(all_rows) < 3:
        return jsonify({"summary": "요약할 만큼 답글이 충분하지 않습니다 (3개 이상 필요).", "cached": False})
    data = request.get_json(silent=True) or {}
    force = bool(data.get("force"))
    last_id = replies[-1]["id"] if replies else parent["id"]
    scope_key = f"thread:{message_id}"
    if not force:
        cached = db.execute("""
            SELECT summary_text, created_at, model
              FROM ai_summaries
             WHERE scope_type='thread' AND scope_key=? AND last_message_id=?
             ORDER BY id DESC LIMIT 1
        """, (scope_key, last_id)).fetchone()
        if cached:
            return jsonify({
                "summary": cached["summary_text"],
                "cached": True,
                "cached_at": cached["created_at"],
                "model": cached["model"],
                "message_count": len(all_rows),
            })
    result, err = _ai_summarize_messages(all_rows, mode="thread")
    if err:
        return jsonify({"error": err}), 500
    now = datetime.now(timezone.utc).isoformat()
    db.execute("""
        INSERT INTO ai_summaries
            (scope_type, scope_key, room_id, last_message_id, summary_text,
             model, input_tokens, output_tokens, cost_usd, created_by, created_at)
        VALUES ('thread', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (scope_key, parent["room_id"], last_id, result["summary_text"],
          result["model"], result["in_tokens"], result["out_tokens"],
          result["cost_usd"], me["id"], now))
    db.commit()
    return jsonify({
        "summary": result["summary_text"],
        "cached": False,
        "model": result["model"],
        "message_count": len(all_rows),
        "input_tokens": result["in_tokens"],
        "output_tokens": result["out_tokens"],
        "cost_usd": result["cost_usd"],
    })


@app.route("/api/ai/rewrite", methods=["POST"])
@login_required
def api_ai_rewrite():
    """작성 톤 조정 (AI 작성 도움).
    body: {text: str, tone: 'formal'|'short'|'professional'|'casual'|'polite'}"""
    data = request.get_json(silent=True) or {}
    text = (data.get("text") or "").strip()
    tone = data.get("tone") or "formal"
    if tone not in REWRITE_TONES:
        return jsonify({"error": f"tone 은 {list(REWRITE_TONES.keys())} 중 하나"}), 400
    if not text:
        return jsonify({"error": "text 필수"}), 400
    if len(text) > 4000:
        return jsonify({"error": "text 4000 자 제한"}), 400
    result, err = _claude_rewrite(text, tone=tone)
    if err:
        return jsonify({"error": err}), 500
    return jsonify({
        "text": result["text"],
        "model": result["model"],
        "input_tokens": result["in_tokens"],
        "output_tokens": result["out_tokens"],
        "cost_usd": result["cost_usd"],
    })


@app.route("/api/rooms/<int:room_id>/history", methods=["GET"])
@login_required
def api_room_history_list(room_id):
    """프로젝트 이력 목록 — 시간 역순.
    권한: 테스트 기간 방장·PM·관리자만 (대표 지시 2026-05-27)."""
    me = current_user()
    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (room_id, me["id"]),
    ).fetchone():
        abort(403)
    if not _can_use_ai_summary(db, room_id, me):  # 게스트(외부)는 프로젝트 이력 사용 불가 (대표 지시 2026-05-30)
        return jsonify({"error": "프로젝트 이력은 테스트 기간 동안 방장·PM 또는 관리자만 사용할 수 있습니다."}), 403
    rows = db.execute("""
        SELECT id, period_start, period_end, first_message_id, last_message_id,
               summary_text, message_count, attachment_count, attachments_json,
               model, cost_usd, created_by, created_at, synced_to_hw, synced_at
          FROM project_history
         WHERE room_id = ?
         ORDER BY id DESC
         LIMIT 200
    """, (room_id,)).fetchall()
    out = []
    for r in rows:
        d = dict(r)
        try:
            d["attachments"] = json.loads(d.pop("attachments_json") or "[]")
        except Exception:
            d["attachments"] = []
        # created_by display_name 채우기
        if d.get("created_by"):
            u = db.execute("SELECT display_name FROM users WHERE id=?", (d["created_by"],)).fetchone()
            d["created_by_name"] = u["display_name"] if u else None
            d["created_mode"] = "manual"
        else:
            d["created_by_name"] = None
            d["created_mode"] = "auto"
        out.append(d)
    return jsonify(out)


@app.route("/api/rooms/<int:room_id>/history/generate", methods=["POST"])
@login_required
def api_room_history_generate(room_id):
    """수동 즉시 이력 생성.
    권한: 테스트 기간 방장·PM·관리자만 (대표 지시 2026-05-27, 이전 TODO 해소)."""
    me = current_user()
    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (room_id, me["id"]),
    ).fetchone():
        abort(403)
    if not _can_use_ai_summary(db, room_id, me):
        return jsonify({"error": "프로젝트 이력은 테스트 기간 동안 방장·PM 또는 관리자만 사용할 수 있습니다."}), 403
    # 프로젝트 방만 — 일반 방은 이력 비활성
    rtype = db.execute("SELECT type FROM rooms WHERE id=?", (room_id,)).fetchone()
    if not rtype or rtype["type"] != "item":
        return jsonify({"error": "이력은 프로젝트 방에서만 생성 가능합니다"}), 400
    hist, err = _generate_project_history(room_id, created_by_uid=me["id"])
    if err == "no_new":
        return jsonify({"error": "마지막 이력 이후 새 메시지가 없습니다", "no_new": True}), 200
    if err == "too_few":
        return jsonify({"error": "새 메시지가 적어 의미 있는 요약이 어렵습니다 (수동은 1개부터 가능)", "too_few": True}), 200
    if err:
        return jsonify({"error": err}), 500
    return jsonify({"ok": True, "history": hist})


@app.route("/api/rooms/<int:room_id>/history/<int:history_id>", methods=["GET"])
@login_required
def api_room_history_get(room_id, history_id):
    """단일 이력 상세.
    권한: 테스트 기간 방장·PM·관리자만 (대표 지시 2026-05-27)."""
    me = current_user()
    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (room_id, me["id"]),
    ).fetchone():
        abort(403)
    if not _can_use_ai_summary(db, room_id, me):  # 게스트(외부)는 프로젝트 이력 사용 불가 (대표 지시 2026-05-30)
        return jsonify({"error": "프로젝트 이력은 테스트 기간 동안 방장·PM 또는 관리자만 사용할 수 있습니다."}), 403
    r = db.execute("""
        SELECT * FROM project_history WHERE id=? AND room_id=?
    """, (history_id, room_id)).fetchone()
    if not r:
        return jsonify({"error": "not found"}), 404
    d = dict(r)
    try:
        d["attachments"] = json.loads(d.pop("attachments_json") or "[]")
    except Exception:
        d["attachments"] = []
    return jsonify(d)


@app.route("/api/messages/<int:message_id>/translate", methods=["POST"])
@login_required
def api_message_translate(message_id):
    """메시지 번역 — 캐시 우선, 없으면 Claude Haiku 호출.

    body: { "target_lang": "ko" | "vi" | "en" }
    """
    me = current_user()
    db = get_db()
    msg = db.execute(
        "SELECT id, room_id, content, kind FROM messages WHERE id = ?", (message_id,)
    ).fetchone()
    if not msg:
        abort(404)
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?", (msg["room_id"], me["id"])
    ).fetchone():
        abort(403)

    if msg["kind"] not in ("text", None) or not (msg["content"] or "").strip():
        return jsonify({"error": "번역 가능한 텍스트 메시지가 아닙니다."}), 400

    data = request.get_json(silent=True) or {}
    target_lang = data.get("target_lang", "vi")
    if target_lang not in TRANSLATE_LANGS:
        return jsonify({"error": f"지원 언어: {', '.join(TRANSLATE_LANGS.keys())}"}), 400

    # 1) 캐시 조회
    cached = db.execute(
        "SELECT translated_text, source_lang, model, created_at FROM message_translations WHERE message_id=? AND target_lang=?",
        (message_id, target_lang),
    ).fetchone()
    if cached:
        return jsonify({
            "message_id": message_id,
            "target_lang": target_lang,
            "translated_text": cached["translated_text"],
            "source_lang": cached["source_lang"],
            "model": cached["model"],
            "from_cache": True,
        })

    # 2) API 키 또는 데모 모드 확인 (공급자 무관 — OpenAI / Anthropic 둘 다 검사)
    if not _ai_provider_has_key() and not TRANSLATE_MOCK:
        return jsonify({
            "error": "번역 서비스가 설정되지 않았습니다.",
            "hint": f"현재 공급자={TRANSLATE_PROVIDER}. OpenAI 면 OPENAI_API_KEY, Anthropic 이면 ANTHROPIC_API_KEY 환경변수 + 서버 재시작.",
        }), 503

    # 3) 월 비용 한도 가드
    monthly = _translate_monthly_cost()
    if monthly >= TRANSLATE_MONTHLY_USD_LIMIT:
        return jsonify({
            "error": f"이번 달 번역 비용 한도(${TRANSLATE_MONTHLY_USD_LIMIT}) 초과",
            "monthly_cost_usd": monthly,
            "hint": "캐시된 번역은 계속 사용 가능. 한도를 늘리려면 KNK_MSG_TRANSLATE_USD_LIMIT 환경변수 조정.",
        }), 429

    # 4) AI 호출 (공급자 라우터)
    result, err = _ai_translate(msg["content"], target_lang)
    if err:
        return jsonify({"error": err}), 500

    translated, source_lang, in_t, out_t, cost = result

    # 5) 캐시 저장 — requested_by_user_id 기록 (사용자별 사용량 통계용)
    now = datetime.now(timezone.utc).isoformat()
    db.execute(
        """INSERT INTO message_translations
           (message_id, target_lang, source_lang, translated_text, model,
            input_tokens, output_tokens, cost_usd, created_at, requested_by_user_id)
           VALUES (?,?,?,?,?,?,?,?,?,?)""",
        (message_id, target_lang, source_lang, translated, TRANSLATE_MODEL,
         in_t, out_t, cost, now, me["id"]),
    )
    db.commit()

    return jsonify({
        "message_id": message_id,
        "target_lang": target_lang,
        "translated_text": translated,
        "source_lang": source_lang,
        "model": TRANSLATE_MODEL,
        "from_cache": False,
        "cost_usd": round(cost, 6),
        "monthly_cost_usd": round(monthly + cost, 4),
    })


@app.route("/api/translate/status")
@login_required
def api_translate_status():
    """번역 기능 상태 — UI 가 메뉴 표시 여부 결정용."""
    enabled = _ai_provider_has_key() or TRANSLATE_MOCK
    if TRANSLATE_MOCK:
        model_label = "DEMO (mock)"
    else:
        model_label = _ai_provider_model_label()
    return jsonify({
        "enabled": enabled,
        "provider": TRANSLATE_PROVIDER,  # "openai" 또는 "anthropic" — UI 가 라벨 분기 가능
        "mock_mode": TRANSLATE_MOCK,
        "model": model_label,
        "languages": TRANSLATE_LANGS,
        "monthly_cost_usd": round(_translate_monthly_cost(), 4),
        "monthly_limit_usd": TRANSLATE_MONTHLY_USD_LIMIT,
    })


# ============================================================
# 관리자: AI 번역 사용량 통계 (대표 지시 2026-05-27)
# ============================================================
# 추적 정책:
#   1) requested_by_user_id 가 있으면 → 그 사용자가 호출자 (정확)
#   2) NULL 이면 (기존 데이터) → 메시지 작성자(messages.user_id) 로 추정
# 권한: role='ceo' (관리자) 만 — _is_owner 또는 user.role 검사

def _is_admin_user(user):
    """관리자 여부 — role=ceo 또는 OWNER_USERNAME 본인.
    user 는 dict 또는 sqlite3.Row 일 수 있음 → dict() 로 정규화 (Row 는 .get() 미지원이라 500 나던 버그 수정 2026-06-01)."""
    if not user:
        return False
    try:
        u = dict(user)
    except Exception:
        return False
    if str(u.get("role") or "") == "ceo":
        return True
    if _is_owner(u.get("username")):
        return True
    return False


# ── 🐞 버그 신고 + 유지보수관리자 ─────────────────────────────────────────
# 유지보수관리자 = 사용현황·버그관리 등 '운영 데이터'를 볼 수 있는 특정 계정.
#   일반 '관리자'(ceo/owner)보다 더 좁게 못박음(기본=김정락 사번 5). 환경변수로 추가 가능.
#   (대표 지시 2026-06-03 — "유지보수관리자만 / 김정락만 확인")
MAINTENANCE_ADMIN_USERNAMES = set(
    u.strip() for u in os.environ.get("KNK_MSG_MAINTENANCE_ADMINS", "5").split(",") if u.strip()
)
def _is_maintenance_owner(user):
    """운영 담당자를 '지정'(부여/회수)할 수 있는 최상위 = 기본 유지보수관리자(김정락 사번5).
    운영 권한자라도 또 다른 운영 담당자를 지정하지는 못함. (대표 지시 2026-06-03)"""
    if not user:
        return False
    try:
        u = dict(user)
    except Exception:
        return False
    return str(u.get("username") or "") in MAINTENANCE_ADMIN_USERNAMES
def _is_maintenance_admin(user):
    """유지보수 화면(사용현황·버그관리·시스템설정)을 보고 다룰 수 있는 사람.
    = 기본 유지보수관리자(김정락) 또는 '운영' 권한 부여자(ops_allowed=1). (대표 지시 2026-06-03)"""
    if not user:
        return False
    try:
        u = dict(user)
    except Exception:
        return False
    if str(u.get("username") or "") in MAINTENANCE_ADMIN_USERNAMES:
        return True
    try:
        return int(u.get("ops_allowed") or 0) == 1
    except Exception:
        return False


BUG_ROOM_NAME = "메신저 불편 신고 여기에"
BUG_STATUSES = ("new", "triaged", "fixed", "deployed", "wontfix")
def _get_bug_room_id(db, create=True):
    """버그 신고 전용 채널의 room_id. app_settings('bug_room_id')에 캐시. 없으면 생성.
    전 직원 공유 채널(게스트 제외) — 멤버 합류는 api_rooms 에서 지연 처리. (대표 지시 2026-06-03)"""
    try:
        row = db.execute("SELECT value FROM app_settings WHERE key='bug_room_id'").fetchone()
    except Exception:
        row = None
    if row and str(row["value"]).isdigit():
        rid = int(row["value"])
        _ex = db.execute("SELECT name FROM rooms WHERE id=?", (rid,)).fetchone()
        if _ex:
            # 방 이름이 바뀌었으면 기존 방 제목도 자동 갱신 (대표 지시 2026-06-03)
            if _ex["name"] != BUG_ROOM_NAME:
                db.execute("UPDATE rooms SET name=? WHERE id=?", (BUG_ROOM_NAME, rid))
                db.commit()
            return rid
    if not create:
        return None
    now = datetime.now(timezone.utc).isoformat()
    cur = db.execute(
        "INSERT INTO rooms (name, type, created_by, created_at, name_locked) VALUES (?,?,?,?,1)",
        (BUG_ROOM_NAME, "channel", None, now),
    )
    rid = cur.lastrowid
    db.execute(
        "INSERT INTO app_settings (key, value) VALUES ('bug_room_id', ?) "
        "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (str(rid),),
    )
    db.commit()
    return rid


# ──────────────────────────────────────────────────────────────
#  AI 사용 권한 관리 — 관리자 전용 (대표 지시 2026-05-28)
#    · GET  /api/admin/ai_permissions       — 사용자 목록 + 권한 + 번역 토글
#    · POST /api/admin/ai_permissions/<uid> — { allowed: 0|1 } 직원별 허용
#    · POST /api/admin/ai_translate_toggle  — { enabled: 0|1 } 전사 번역
# ──────────────────────────────────────────────────────────────
@app.route("/api/admin/ai_permissions")
@login_required
def api_admin_ai_permissions():
    me = current_user()
    if me["role"] != "ceo":
        abort(403)
    db = get_db()
    users_rows = db.execute("""
        SELECT id, username, display_name, title, department, role,
               COALESCE(ai_summary_allowed, 0) AS ai_summary_allowed
          FROM users
         ORDER BY department NULLS LAST, display_name ASC
    """).fetchall()
    translate_enabled = _ai_translate_enabled(db)
    presence_detail = _presence_show_detail(db)
    return jsonify({
        "ok": True,
        "translate_enabled": 1 if translate_enabled else 0,
        "presence_detail": 1 if presence_detail else 0,
        "ai_model": _get_openai_model(db),                 # 현재 사용 OpenAI 모델 (대표 지시 2026-06-01)
        "ai_model_locked": bool(_OPENAI_MODEL_ENV),        # 환경변수 잠금 시 화면 변경 불가
        "ai_provider": TRANSLATE_PROVIDER,
        "works_history_sync": 1 if _works_history_sync_enabled() else 0,   # WORKS 이력 자동전송 ON 여부
        "works_history_url_set": 1 if WORKS_HISTORY_URL else 0,            # WORKS 엔드포인트 설정 여부(미설정이면 토글 무의미)
        "users": [dict(r) for r in users_rows],
    })


@app.route("/api/admin/ai_permissions/<int:user_id>", methods=["POST"])
@login_required
def api_admin_ai_permission_set(user_id):
    me = current_user()
    if me["role"] != "ceo":
        abort(403)
    data = request.get_json(silent=True) or {}
    allowed = 1 if int(data.get("allowed") or 0) == 1 else 0
    db = get_db()
    db.execute("UPDATE users SET ai_summary_allowed=? WHERE id=?", (allowed, user_id))
    db.commit()
    return jsonify({"ok": True, "user_id": user_id, "allowed": allowed})


# ──────────────────────────────────────────────────────────────
#  채널 생성 권한 관리 — 관리자 전용 (대표 지시 2026-05-29)
#    · GET  /api/admin/channel_permissions       — 직원 목록 + 권한(직급자동 by_title / 추가허용)
#    · POST /api/admin/channel_permissions/<uid> — { allowed: 0|1 } 직급 무관 추가 허용
#  직급(대표·임원·팀장·법인장)·관리자(ceo)는 by_title=1 로 항상 허용(UI 잠금).
# ──────────────────────────────────────────────────────────────
@app.route("/api/admin/channel_permissions")
@login_required
def api_admin_channel_permissions():
    me = current_user()
    if me["role"] != "ceo":
        abort(403)
    db = get_db()
    rows = db.execute("""
        SELECT id, username, display_name, title, department, role,
               COALESCE(channel_create_allowed, 0) AS channel_create_allowed
          FROM users
         WHERE username != '_deleted_user' AND COALESCE(is_guest, 0) = 0
         ORDER BY department NULLS LAST, display_name ASC
    """).fetchall()
    out = []
    for r in rows:
        d = dict(r)
        d["by_title"] = 1 if _channel_by_title(r) else 0   # 직급/관리자 자동 허용(잠금)
        out.append(d)
    return jsonify({"ok": True, "users": out})


@app.route("/api/admin/channel_permissions/<int:user_id>", methods=["POST"])
@login_required
def api_admin_channel_permission_set(user_id):
    me = current_user()
    if me["role"] != "ceo":
        abort(403)
    data = request.get_json(silent=True) or {}
    allowed = 1 if int(data.get("allowed") or 0) == 1 else 0
    db = get_db()
    db.execute("UPDATE users SET channel_create_allowed=? WHERE id=?", (allowed, user_id))
    db.commit()
    return jsonify({"ok": True, "user_id": user_id, "allowed": allowed})


@app.route("/api/admin/ai_translate_toggle", methods=["POST"])
@login_required
def api_admin_ai_translate_toggle():
    me = current_user()
    if not _is_maintenance_admin(me):   # 🔧시스템설정 — 유지보수(대표님·운영자) (대표 지시 2026-06-03)
        abort(403)
    data = request.get_json(silent=True) or {}
    enabled = "1" if int(data.get("enabled") or 0) == 1 else "0"
    db = get_db()
    db.execute("""
        INSERT INTO app_settings (key, value) VALUES ('ai_translate_enabled', ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
    """, (enabled,))
    db.commit()
    return jsonify({"ok": True, "enabled": int(enabled)})


@app.route("/api/admin/presence_detail_toggle", methods=["POST"])
@login_required
def api_admin_presence_detail_toggle():
    """상태표시 상세(기기·회사망) on/off — 관리자 전용. (대표 지시 2026-06-01)
    enabled=1: 💻 컴퓨터/📱 휴대폰 + (회사) 표시 / 0: '🟢 접속' 하나로 단순.
    화면 표시만 바뀌며, 변경 즉시 열린 모든 클라이언트에 broadcast(새로고침 불필요)."""
    me = current_user()
    if not _is_maintenance_admin(me):   # 🔧시스템설정 — 유지보수(대표님·운영자) (대표 지시 2026-06-03)
        abort(403)
    data = request.get_json(silent=True) or {}
    enabled = "1" if int(data.get("enabled") or 0) == 1 else "0"
    db = get_db()
    db.execute("""
        INSERT INTO app_settings (key, value) VALUES ('presence_show_detail', ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
    """, (enabled,))
    db.commit()
    try:
        socketio.emit("presence_detail_changed", {"detail": int(enabled)})
    except Exception:
        pass
    return jsonify({"ok": True, "enabled": int(enabled)})


@app.route("/api/admin/ai_model", methods=["POST"])
@login_required
def api_admin_ai_model_set():
    """OpenAI 모델 변경 — 관리자 전용 (app_settings.ai_model). 재시작 없이 다음 호출부터 적용.
    (대표 지시 2026-06-01, 가이드 §3. 환경변수 KNK_MSG_OPENAI_MODEL 설정 시엔 그게 우선이라 화면 변경 무시)"""
    me = current_user()
    if not _is_maintenance_admin(me):   # 🔧시스템설정 — 유지보수(대표님·운영자) (대표 지시 2026-06-03)
        abort(403)
    if _OPENAI_MODEL_ENV:
        return jsonify({"ok": False, "error": "서버 환경변수(KNK_MSG_OPENAI_MODEL)로 모델이 고정돼 있어 화면에서 바꿀 수 없습니다."}), 400
    data = request.get_json(silent=True) or {}
    model = (data.get("model") or "").strip()[:80]
    db = get_db()
    db.execute("""
        INSERT INTO app_settings (key, value) VALUES ('ai_model', ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
    """, (model,))
    db.commit()
    return jsonify({"ok": True, "model": _get_openai_model(db)})


@app.route("/api/admin/ai_model/test", methods=["POST"])
@login_required
def api_admin_ai_model_test():
    """입력한(또는 현재) 모델로 짧게 1회 실제 호출 — 연결 즉시 검증(조용히 죽는 것 방지).
    (대표 지시 2026-06-01, 가이드 §3)"""
    me = current_user()
    if not _is_maintenance_admin(me):   # 🔧시스템설정 — 유지보수(대표님·운영자) (대표 지시 2026-06-03)
        abort(403)
    data = request.get_json(silent=True) or {}
    model = (data.get("model") or "").strip() or _get_openai_model(get_db())
    if TRANSLATE_PROVIDER != "openai":
        return jsonify({"ok": False, "model": model, "error": "현재 AI 공급자가 OpenAI 가 아닙니다."})
    client, err = _openai_client()
    if err:
        return jsonify({"ok": False, "model": model, "error": err})
    try:
        resp = openai_create(
            client, model,
            [{"role": "user", "content": "Reply with the single word: OK"}],
            max_tokens=5, temperature=0.0,
        )
        ans = (resp.choices[0].message.content or "").strip()
        return jsonify({"ok": True, "model": model, "answer": ans[:80]})
    except Exception as e:
        return jsonify({"ok": False, "model": model, "error": str(e)[:300]})


@app.route("/api/admin/works_history_toggle", methods=["POST"])
@login_required
def api_admin_works_history_toggle():
    """HAIST WORKS 프로젝트 이력 자동전송 ON/OFF — 관리자 전용. (대표 지시 2026-06-01)
    ON 이라도 서버에 KNK_WORKS_HISTORY_URL 이 설정돼 있어야 실제 전송됨(미설정이면 no-op)."""
    me = current_user()
    if not _is_maintenance_admin(me):   # 🔧시스템설정 — 유지보수(대표님·운영자) (대표 지시 2026-06-03)
        abort(403)
    data = request.get_json(silent=True) or {}
    enabled = "1" if int(data.get("enabled") or 0) == 1 else "0"
    db = get_db()
    db.execute("""
        INSERT INTO app_settings (key, value) VALUES ('works_history_sync', ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
    """, (enabled,))
    db.commit()
    return jsonify({
        "ok": True,
        "enabled": int(enabled),
        "url_set": 1 if WORKS_HISTORY_URL else 0,
        "effective": 1 if (enabled == "1" and WORKS_HISTORY_URL) else 0,
    })


@app.route("/api/admin/translate_usage")
@login_required
def api_admin_translate_usage():
    """사용자별 번역 사용량 통계.
    Query params:
      period: 'this_month' (기본) / 'last_month' / 'last_7days' / 'all'
      dept:   'all' (기본) / '본사' / '베트남' / 특정 부서명
    응답: { period, total_cost_usd, total_calls, total_in_tokens, total_out_tokens,
            monthly_limit_usd, users: [{user_id, display_name, department, title, calls, in_tokens, out_tokens, cost_usd}] }
    """
    me = current_user()
    if not _is_admin_user(me):
        abort(403)

    period = (request.args.get("period") or "this_month").lower()
    dept = (request.args.get("dept") or "all").strip()

    # 기간 SQL 조건
    if period == "this_month":
        date_cond = "AND date(t.created_at) >= date('now', 'start of month')"
        period_label = "이번 달"
    elif period == "last_month":
        date_cond = "AND date(t.created_at) >= date('now', 'start of month', '-1 month') AND date(t.created_at) < date('now', 'start of month')"
        period_label = "지난 달"
    elif period == "last_7days":
        date_cond = "AND date(t.created_at) >= date('now', '-7 days')"
        period_label = "최근 7일"
    elif period == "all":
        date_cond = ""
        period_label = "전체"
    else:
        date_cond = "AND date(t.created_at) >= date('now', 'start of month')"
        period_label = "이번 달"
        period = "this_month"

    # 사용자 기준: requested_by_user_id 우선, NULL 이면 메시지 작성자 fallback
    # COALESCE 로 통합
    db = get_db()
    dept_filter = ""
    dept_params = []
    if dept and dept != "all":
        # LIKE 검색 (부서명 일부 매칭) — 본사·베트남 같이 큰 범위 모두 포함
        dept_filter = " AND COALESCE(u.department, '') LIKE ?"
        dept_params = [f"%{dept}%"]

    sql = f"""
        SELECT
            COALESCE(t.requested_by_user_id, m.user_id) AS uid,
            u.display_name AS display_name,
            u.display_name_vn AS display_name_vn,
            u.display_name_en AS display_name_en,
            u.department AS department,
            u.title AS title,
            u.username AS username,
            COUNT(*) AS calls,
            COALESCE(SUM(t.input_tokens), 0) AS in_tokens,
            COALESCE(SUM(t.output_tokens), 0) AS out_tokens,
            COALESCE(SUM(t.cost_usd), 0) AS cost_usd
          FROM message_translations t
          JOIN messages m ON m.id = t.message_id
          LEFT JOIN users u ON u.id = COALESCE(t.requested_by_user_id, m.user_id)
         WHERE 1=1 {date_cond} {dept_filter}
         GROUP BY uid
         ORDER BY cost_usd DESC, calls DESC
    """
    rows = db.execute(sql, dept_params).fetchall()

    users = []
    total_cost = 0.0
    total_calls = 0
    total_in = 0
    total_out = 0
    for r in rows:
        users.append({
            "user_id": r["uid"],
            "display_name": r["display_name"] or f"(삭제된 사용자 #{r['uid']})",
            "display_name_vn": r["display_name_vn"] or "",
            "display_name_en": r["display_name_en"] or "",
            "department": r["department"] or "",
            "title": r["title"] or "",
            "is_vn": bool(_user_is_vietnam(r["department"] or "")),
            "username": r["username"] or "",
            "calls": int(r["calls"] or 0),
            "in_tokens": int(r["in_tokens"] or 0),
            "out_tokens": int(r["out_tokens"] or 0),
            "cost_usd": round(float(r["cost_usd"] or 0), 6),
        })
        total_cost += float(r["cost_usd"] or 0)
        total_calls += int(r["calls"] or 0)
        total_in += int(r["in_tokens"] or 0)
        total_out += int(r["out_tokens"] or 0)

    # 부서 목록 (필터 드롭다운용) — 사용량 있는 부서만
    dept_rows = db.execute("""
        SELECT DISTINCT COALESCE(u.department, '') AS dept
          FROM users u
         WHERE COALESCE(u.department, '') != ''
         ORDER BY dept
    """).fetchall()
    departments = [r["dept"] for r in dept_rows]

    return jsonify({
        "period": period,
        "period_label": period_label,
        "dept": dept,
        "departments": departments,
        "total_cost_usd": round(total_cost, 6),
        "total_calls": total_calls,
        "total_in_tokens": total_in,
        "total_out_tokens": total_out,
        "monthly_limit_usd": TRANSLATE_MONTHLY_USD_LIMIT,
        # 이번 달 한도 진행률 (period 무관 — 한도 자체는 월 단위)
        "monthly_cost_usd": round(_translate_monthly_cost(), 6),
        "provider": TRANSLATE_PROVIDER,
        "users": users,
    })


@app.route("/api/admin/translate_usage/export")
@login_required
def api_admin_translate_usage_export():
    """사용량 엑셀 다운로드. Query params 는 api_admin_translate_usage 와 동일."""
    me = current_user()
    if not _is_admin_user(me):
        abort(403)

    # 동일 데이터 조회 — 위 endpoint 로직 재사용 (Flask test client 안 쓰고 직접)
    period = (request.args.get("period") or "this_month").lower()
    dept = (request.args.get("dept") or "all").strip()

    if period == "this_month":
        date_cond = "AND date(t.created_at) >= date('now', 'start of month')"
        period_label = "이번 달"
    elif period == "last_month":
        date_cond = "AND date(t.created_at) >= date('now', 'start of month', '-1 month') AND date(t.created_at) < date('now', 'start of month')"
        period_label = "지난 달"
    elif period == "last_7days":
        date_cond = "AND date(t.created_at) >= date('now', '-7 days')"
        period_label = "최근 7일"
    elif period == "all":
        date_cond = ""
        period_label = "전체"
    else:
        date_cond = "AND date(t.created_at) >= date('now', 'start of month')"
        period_label = "이번 달"

    db = get_db()
    dept_filter = ""
    dept_params = []
    if dept and dept != "all":
        dept_filter = " AND COALESCE(u.department, '') LIKE ?"
        dept_params = [f"%{dept}%"]

    sql = f"""
        SELECT
            COALESCE(t.requested_by_user_id, m.user_id) AS uid,
            u.display_name AS display_name,
            u.display_name_vn AS display_name_vn,
            u.display_name_en AS display_name_en,
            u.department AS department,
            u.title AS title,
            u.username AS username,
            COUNT(*) AS calls,
            COALESCE(SUM(t.input_tokens), 0) AS in_tokens,
            COALESCE(SUM(t.output_tokens), 0) AS out_tokens,
            COALESCE(SUM(t.cost_usd), 0) AS cost_usd
          FROM message_translations t
          JOIN messages m ON m.id = t.message_id
          LEFT JOIN users u ON u.id = COALESCE(t.requested_by_user_id, m.user_id)
         WHERE 1=1 {date_cond} {dept_filter}
         GROUP BY uid
         ORDER BY cost_usd DESC, calls DESC
    """
    rows = db.execute(sql, dept_params).fetchall()

    # openpyxl 로 xlsx 생성
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({"error": "openpyxl 미설치 — 서버 관리자에게 문의"}), 500

    wb = Workbook()

    # 본사/베트남법인 분리 (대표 지시 2026-06-05) — 행을 법인별로 나눠 시트 2개로 출력
    hq_rows = [r for r in rows if not _user_is_vietnam(r["department"] or "")]
    vn_rows = [r for r in rows if _user_is_vietnam(r["department"] or "")]

    _thin = Side(border_style="thin", color="D1D5DB")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _headers = ["순위", "이름", "직급", "부서", "로그인ID", "호출수", "토큰(in/out)", "비용 (USD)"]
    _widths = [6, 14, 10, 18, 14, 9, 16, 12]

    def _fill_usage_sheet(ws, corp_label, data_rows):
        # 상단 제목
        ws["A1"] = f"AI 번역 사용량 — {corp_label}"
        ws["A1"].font = Font(bold=True, size=14, color="A5282C")
        ws.merge_cells("A1:H1")
        ws["A2"] = f"기간: {period_label} / 소속: {corp_label} / 부서필터: {dept if dept != 'all' else '전체'} / 공급자: {TRANSLATE_PROVIDER}"
        ws["A2"].font = Font(size=10, color="6B7280")
        ws.merge_cells("A2:H2")
        # 헤더
        header_fill = PatternFill(start_color="A5282C", end_color="A5282C", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, h in enumerate(_headers, start=1):
            c = ws.cell(row=4, column=col_idx, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _border
        # 데이터 행
        s_cost = 0.0
        s_calls = 0
        s_in = 0
        s_out = 0
        for idx, r in enumerate(data_rows, start=1):
            row_num = 4 + idx
            calls = int(r["calls"] or 0)
            in_t = int(r["in_tokens"] or 0)
            out_t = int(r["out_tokens"] or 0)
            cost = float(r["cost_usd"] or 0)
            ws.cell(row=row_num, column=1, value=idx).alignment = Alignment(horizontal="center")
            _vn_nm = (r["display_name_vn"] or "").strip()
            _disp_nm = (f"{_vn_nm} ({r['display_name']})" if (_vn_nm and r["display_name"]) else (r["display_name"] or f"(삭제된 사용자 #{r['uid']})"))
            ws.cell(row=row_num, column=2, value=_disp_nm)
            ws.cell(row=row_num, column=3, value=r["title"] or "")
            ws.cell(row=row_num, column=4, value=r["department"] or "")
            ws.cell(row=row_num, column=5, value=r["username"] or "")
            ws.cell(row=row_num, column=6, value=calls).alignment = Alignment(horizontal="right")
            ws.cell(row=row_num, column=7, value=f"{in_t:,} / {out_t:,}").alignment = Alignment(horizontal="right")
            ws.cell(row=row_num, column=8, value=round(cost, 6)).alignment = Alignment(horizontal="right")
            for col_idx in range(1, 9):
                ws.cell(row=row_num, column=col_idx).border = _border
            s_cost += cost
            s_calls += calls
            s_in += in_t
            s_out += out_t
        # 합계 행
        total_row = 4 + len(data_rows) + 1
        ws.cell(row=total_row, column=1, value="합계").font = Font(bold=True)
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right")
        ws.cell(row=total_row, column=6, value=s_calls).font = Font(bold=True)
        ws.cell(row=total_row, column=6).alignment = Alignment(horizontal="right")
        ws.cell(row=total_row, column=7, value=f"{s_in:,} / {s_out:,}").font = Font(bold=True)
        ws.cell(row=total_row, column=7).alignment = Alignment(horizontal="right")
        ws.cell(row=total_row, column=8, value=round(s_cost, 6)).font = Font(bold=True, color="A5282C")
        ws.cell(row=total_row, column=8).alignment = Alignment(horizontal="right")
        for col_idx in range(1, 9):
            ws.cell(row=total_row, column=col_idx).border = _border
            ws.cell(row=total_row, column=col_idx).fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        if not data_rows:
            ws.cell(row=5, column=1, value="해당 기간 사용 내역 없음").font = Font(color="6B7280")
        # 컬럼 너비
        for col_idx, w in enumerate(_widths, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = w

    ws1 = wb.active
    ws1.title = "본사"
    _fill_usage_sheet(ws1, "🇰🇷 본사", hq_rows)
    ws2 = wb.create_sheet("베트남법인")
    _fill_usage_sheet(ws2, "🇻🇳 베트남법인", vn_rows)

    # 파일로 send
    import io
    from flask import send_file as _send_file
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"AI번역사용량_{period}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return _send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


@app.route("/api/messages/send", methods=["POST"])
@login_required
def api_messages_send():
    """전송 시점 번역 지원 메시지 송신.

    body:
      room_id (int)        — 방 ID
      content (str)        — 원문
      translate_to (list[str], 선택) — 함께 첨부할 번역 언어들. 예: ["vi"], ["vi","en"]

    동작:
      1. 멤버십 확인
      2. translate_to 있으면 Claude 호출해서 번역 (캐시 가능 — 메시지 저장 후 message_translations 에 기록)
      3. 메시지 저장 (원문은 그대로)
      4. 번역들을 message_translations 에 저장 (메시지 로드 시 자동 inline 표시됨)
      5. socket new_message 로 broadcast (translations dict 포함 → 양국어 즉시 표시)

    이게 Socket 'send' 이벤트와 다른 점:
      - 번역 비용·시간 때문에 sync REST 로 분리
      - 보내는 사람이 양국어 동시에 만들어 발송 (일반 메신저 번역의 한계 극복: 받은 사람만 번역되는 문제 X)
    """
    me = current_user()
    data = request.get_json(silent=True) or {}
    room_id = data.get("room_id")
    content = (data.get("content") or "").strip()
    translate_to = data.get("translate_to") or []

    if not isinstance(translate_to, list):
        translate_to = [translate_to] if translate_to else []
    translate_to = [t for t in translate_to if t in TRANSLATE_LANGS]

    if not room_id or not content:
        return jsonify({"error": "room_id 와 content 필수"}), 400
    try:
        room_id = int(room_id)
    except (TypeError, ValueError):
        return jsonify({"error": "room_id 가 숫자가 아닙니다"}), 400
    if len(content) > 4000:
        content = content[:4000]

    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (room_id, me["id"]),
    ).fetchone():
        abort(403)

    # 1) 번역 수행 (있으면)
    translations = {}  # lang -> translated_text
    translation_meta = {}  # lang -> (in_t, out_t, cost_usd, source_lang)

    if translate_to:
        if not _ai_provider_has_key() and not TRANSLATE_MOCK:
            return jsonify({
                "error": "번역 서비스가 설정되지 않았습니다.",
                "hint": f"현재 공급자={TRANSLATE_PROVIDER}. OpenAI 면 OPENAI_API_KEY, Anthropic 이면 ANTHROPIC_API_KEY 환경변수 + 서버 재시작.",
            }), 503

        # 월 비용 가드
        monthly = _translate_monthly_cost()
        if monthly >= TRANSLATE_MONTHLY_USD_LIMIT:
            return jsonify({
                "error": f"이번 달 번역 비용 한도(${TRANSLATE_MONTHLY_USD_LIMIT}) 초과",
                "hint": "한도를 늘리려면 KNK_MSG_TRANSLATE_USD_LIMIT 환경변수 조정",
            }), 429

        # 중국어 대상이면 간체/번체 결정 — 방의 직전 '상대(나 외)' 텍스트 메시지가 번체면 번체.
        # (대표 지시 2026-05-31: 기본 간체, 직전 상대 메시지 번체면 번체)
        zh_variant = None
        if "zh" in translate_to:
            prev = db.execute(
                "SELECT content FROM messages WHERE room_id=? AND user_id!=? AND kind='text' "
                "ORDER BY id DESC LIMIT 1",
                (room_id, me["id"]),
            ).fetchone()
            if prev:
                zh_variant = _detect_han_variant(prev["content"])

        for lang in translate_to:
            result, err = _ai_translate(content, lang, zh_variant if lang == "zh" else None)
            if err:
                return jsonify({"error": f"{lang} 번역 실패: {err}"}), 500
            translated, source_lang, in_t, out_t, cost = result
            translations[lang] = translated
            translation_meta[lang] = (in_t, out_t, cost, source_lang)

    # 2) 메시지 저장
    now = datetime.now(timezone.utc).isoformat()
    cur = db.execute(
        "INSERT INTO messages (room_id, user_id, content, kind, created_at) VALUES (?,?,?,?,?)",
        (room_id, me["id"], content, "text", now),
    )
    mid = cur.lastrowid

    # 3) 번역 캐시 저장 — requested_by_user_id 기록 (메시지 작성자가 곧 요청자)
    for lang, text in translations.items():
        in_t, out_t, cost, source_lang = translation_meta[lang]
        db.execute(
            """INSERT INTO message_translations
               (message_id, target_lang, source_lang, translated_text, model,
                input_tokens, output_tokens, cost_usd, created_at, requested_by_user_id)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (mid, lang, source_lang, text, TRANSLATE_MODEL,
             in_t, out_t, cost, now, me["id"]),
        )
    db.commit()

    # 4) 사용자 정보 + broadcast
    u = db.execute(
        "SELECT display_name, avatar_color FROM users WHERE id=?", (me["id"],)
    ).fetchone()
    _sender_name = u["display_name"]
    payload = {
        "id": mid,
        "room_id": room_id,
        "user_id": me["id"],
        "display_name": u["display_name"],
        "avatar_color": u["avatar_color"],
        "content": content,
        "kind": "text",
        "created_at": now,
        "translations": translations,  # {lang: text} - 클라이언트가 즉시 inline 렌더
    }
    socketio.emit("new_message", payload, to=f"room_{room_id}")
    # @멘션 처리는 전달 '이후' 백그라운드로 (전송 지연 방지)
    if content and "@" in content:
        try:
            socketio.start_background_task(
                _process_mentions_bg, mid, room_id, me["id"], content, None, now, _sender_name
            )
        except Exception as _me:
            print(f"[mention] api_messages_send 백그라운드 시작 실패: {_me}", flush=True)

    # Web Push — 백그라운드 알림 (휴대폰 PWA 가 닫혀있어도 OS 알림 도착).
    # socket 'send' 이벤트는 이미 push 호출함. REST 라우트도 동일하게 처리.
    if PYWEBPUSH_OK:
        try:
            r = db.execute("SELECT name FROM rooms WHERE id=?", (room_id,)).fetchone()
            room_name = r["name"] if r and r["name"] else "채팅"
        except Exception:
            room_name = "채팅"
        push_title = f"💬 {u['display_name']} ({room_name})"
        push_body = content[:120]
        import threading
        threading.Thread(
            target=push_message_to_room_members,
            args=(room_id, me["id"], push_title, push_body),
            kwargs={"url": f"{BASE_PATH}/chat?room={room_id}", "tag": f"room_{room_id}"},
            daemon=True,
        ).start()

    return jsonify(payload)


# ---------- 파일 버전 체인 ----------
@app.route("/api/files/<int:message_id>/versions")
@login_required
def api_file_versions(message_id):
    """이 첨부 메시지가 속한 파일 버전 체인 전체."""
    me = current_user()
    db = get_db()
    msg = db.execute("SELECT room_id FROM messages WHERE id = ?", (message_id,)).fetchone()
    if not msg:
        abort(404)
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?", (msg["room_id"], me["id"])
    ).fetchone():
        abort(403)
    av = db.execute(
        "SELECT parent_message_id FROM attachment_versions WHERE message_id = ?", (message_id,)
    ).fetchone()
    if not av:
        return jsonify([])
    rows = db.execute("""
        SELECT av.message_id, av.version_no, av.parent_message_id,
               m.file_name, m.file_path, m.file_size, m.file_mime,
               m.user_id, m.created_at,
               u.display_name, u.avatar_color
          FROM attachment_versions av
          JOIN messages m ON m.id = av.message_id
          JOIN users u ON u.id = m.user_id
         WHERE av.parent_message_id = ?
         ORDER BY av.version_no DESC
    """, (av["parent_message_id"],)).fetchall()
    return jsonify([dict(r) for r in rows])


# ---------- 파일 업로드 / 다운로드 ----------
def ext_of(filename):
    return (filename.rsplit(".", 1)[-1] or "").lower() if "." in filename else ""


def is_image_ext(ext):
    return ext in ALLOWED_IMAGE_EXT


@app.route("/api/upload", methods=["POST"])
@login_required
def api_upload():
    me = current_user()
    room_id = request.form.get("room_id")
    if not room_id:
        return jsonify({"error": "room_id가 필요합니다."}), 400
    try:
        room_id = int(room_id)
    except (TypeError, ValueError):
        return jsonify({"error": "room_id가 숫자가 아닙니다."}), 400

    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?", (room_id, me["id"])
    ).fetchone():
        abort(403)

    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error": "파일이 없습니다."}), 400

    original = f.filename
    ext = ext_of(original)

    # 🛡️ 보안 검사 1: 위험 확장자 + 이중 확장자 차단
    dangerous, why = _is_dangerous_filename(original)
    if dangerous:
        return jsonify({"error": f"보안 정책에 따라 차단됨: {why}"}), 400

    # 🛡️ 보안 검사 2: 화이트리스트 통과
    if ext and ext not in ALLOWED_FILE_EXT:
        return jsonify({"error": f"허용되지 않는 확장자(.{ext})"}), 400

    # 🛡️ 보안 검사 3: 실행 파일 magic number — 확장자 위조 차단
    is_exec, why = _check_executable_magic(f)
    if is_exec:
        return jsonify({"error": f"실행 파일은 업로드할 수 없습니다 — {why}"}), 400

    # 🛡️ 보안 검사 4: 단일 파일 크기 — Content-Length 사전 검사
    content_len = request.content_length or 0
    if content_len > PER_FILE_MAX_MB * 1024 * 1024:
        return jsonify({"error": f"파일이 너무 큽니다 (단일 {PER_FILE_MAX_MB}MB 초과)"}), 413

    # 🛡️ 보안 검사 5: 사용자별 분당 업로드 횟수 제한 (Rate Limit)
    # 사진 일괄 업로드 (30장) + 약간의 여유 = 60개/분 (대표 지시 2026-05-19)
    if not _check_rate_limit(me["id"], "upload", max_per_minute=60):
        return jsonify({"error": "업로드 속도 제한 — 잠시 후 다시 시도하세요 (1분에 60개)"}), 429

    safe_base = secure_filename(original) or "file"
    if not ext_of(safe_base):
        safe_base = f"{safe_base}.{ext}" if ext else safe_base
    unique = f"{uuid.uuid4().hex[:12]}_{safe_base}"
    room_dir = os.path.join(UPLOAD_DIR, str(room_id))
    os.makedirs(room_dir, exist_ok=True)
    fpath = os.path.join(room_dir, unique)
    f.save(fpath)
    size = os.path.getsize(fpath)
    # 🦠 보안 검사 6: 바이러스 검사 (ClamAV) — 비활성 시 통과(무해). 감염이면 파일 삭제 + 거절.
    try:
        _infected, _vdetail = _scan_file_for_virus(fpath)
    except Exception as _ve:
        _infected, _vdetail = (False, f"scan-exc:{_ve}")
    if _infected:
        try: os.remove(fpath)
        except Exception: pass
        return jsonify({"error": f"바이러스가 검출되어 업로드가 차단되었습니다 ({_vdetail})"}), 400
    mime = f.mimetype or mimetypes.guess_type(unique)[0] or "application/octet-stream"

    kind = "image" if is_image_ext(ext) else "file"
    rel_path = f"{room_id}/{unique}"
    now = datetime.now(timezone.utc).isoformat()

    # 앨범 묶음 ID (선택) — 클라이언트가 같은 album_id 로 N번 업로드 호출 → 그리드 1개 메시지로 렌더.
    # 사진(kind='image') 일 때만 의미가 있어 file 타입은 강제로 무시.
    album_id = (request.form.get("album_id") or "").strip() or None
    if album_id and kind != "image":
        album_id = None

    # 스레드 답글 첨부(선택) — thread_parent_id 가 오면 메인 타임라인이 아니라
    #   해당 부모 메시지의 스레드(답글)로 저장한다. (대표 지시 2026-06-04)
    thread_parent_id = None
    _tp_raw = (request.form.get("thread_parent_id") or "").strip()
    if _tp_raw:
        try:
            _tp = int(_tp_raw)
        except (TypeError, ValueError):
            _tp = None
        if _tp:
            _prow = db.execute(
                "SELECT id, room_id, parent_message_id FROM messages WHERE id=?", (_tp,)
            ).fetchone()
            # 같은 방의 메시지여야 하고, 답글의 답글이면 최상위 부모로 모은다.
            if _prow and _prow["room_id"] == room_id:
                thread_parent_id = _prow["parent_message_id"] or _prow["id"]
        # 스레드 첨부는 앨범(메인 그리드) 개념을 쓰지 않는다.
        if thread_parent_id:
            album_id = None

    if thread_parent_id:
        cur = db.execute("""
            INSERT INTO messages (room_id, user_id, content, kind, file_path, file_name, file_size, file_mime, album_id, created_at, parent_message_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, (room_id, me["id"], original, kind, rel_path, original, size, mime, album_id, now, thread_parent_id))
    else:
        cur = db.execute("""
            INSERT INTO messages (room_id, user_id, content, kind, file_path, file_name, file_size, file_mime, album_id, created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (room_id, me["id"], original, kind, rel_path, original, size, mime, album_id, now))
    mid = cur.lastrowid

    # ── 스레드 답글 첨부: 버전체인·메인 타임라인 대신 스레드에만 표시하고 종료 ──
    if thread_parent_id:
        db.commit()
        u = db.execute("SELECT display_name, avatar_color FROM users WHERE id=?", (me["id"],)).fetchone()
        payload = {
            "id": mid,
            "room_id": room_id,
            "user_id": me["id"],
            "display_name": u["display_name"],
            "avatar_color": u["avatar_color"],
            "content": original,
            "kind": kind,
            "file_path": rel_path,
            "file_name": original,
            "file_size": size,
            "file_mime": mime,
            "created_at": now,
            "parent_message_id": thread_parent_id,
            "thread_parent_id": thread_parent_id,
        }
        socketio.emit("thread_reply", payload, to=f"room_{room_id}")
        socketio.emit("thread_count_changed", {
            "room_id": room_id,
            "parent_id": thread_parent_id,
        }, to=f"room_{room_id}")
        # 푸시 — 부모 작성자 + 기존 답글 작성자들 (송신자 제외)
        if PYWEBPUSH_OK:
            recipients = set()
            _par = db.execute("SELECT user_id FROM messages WHERE id=?", (thread_parent_id,)).fetchone()
            if _par and _par["user_id"] != me["id"]:
                recipients.add(_par["user_id"])
            _prev = db.execute(
                "SELECT DISTINCT user_id FROM messages WHERE parent_message_id=? AND user_id != ?",
                (thread_parent_id, me["id"]),
            ).fetchall()
            for _r in _prev:
                recipients.add(_r["user_id"])
            if recipients:
                _rr = db.execute("SELECT name FROM rooms WHERE id=?", (room_id,)).fetchone()
                _rname = _rr["name"] if _rr else "채팅"
                _title = f"💬 {u['display_name']} (스레드·{_rname})"
                _body = ("[사진] " if kind == "image" else "[파일] ") + original
                import threading as _t
                def _push_all():
                    for _rid in recipients:
                        if _user_has_active_session(_rid):
                            continue
                        send_push_to_user(
                            _rid, _title, _body,
                            url=f"{BASE_PATH}/chat?room={room_id}&thread={thread_parent_id}",
                            tag=f"thread_{thread_parent_id}",
                        )
                _t.Thread(target=_push_all, daemon=True).start()
        return jsonify(payload)

    # 동일 방·동일 원본 파일명으로 이전 업로드 있으면 버전 체인 연결
    prev = db.execute("""
        SELECT av.parent_message_id, av.version_no
          FROM messages m
          JOIN attachment_versions av ON av.message_id = m.id
         WHERE m.room_id = ? AND m.file_name = ? AND m.id != ?
         ORDER BY av.version_no DESC LIMIT 1
    """, (room_id, original, mid)).fetchone()

    if prev:
        parent_id = prev["parent_message_id"]
        version_no = (prev["version_no"] or 1) + 1
    else:
        # 첫 버전 — 자기 자신이 부모
        # 같은 파일명의 가장 오래된 메시지를 부모로 (그동안 attachment_versions 없던 케이스)
        oldest = db.execute(
            "SELECT id FROM messages WHERE room_id=? AND file_name=? AND kind IN ('image','file') ORDER BY id ASC LIMIT 1",
            (room_id, original),
        ).fetchone()
        parent_id = oldest["id"] if oldest else mid
        # 부모도 attachment_versions 가 없으면 v1 으로 backfill
        if oldest and oldest["id"] != mid:
            existing_parent_av = db.execute(
                "SELECT 1 FROM attachment_versions WHERE message_id=?", (oldest["id"],)
            ).fetchone()
            if not existing_parent_av:
                db.execute(
                    "INSERT INTO attachment_versions (message_id, parent_message_id, version_no, room_id) VALUES (?,?,?,?)",
                    (oldest["id"], oldest["id"], 1, room_id),
                )
            # 그 사이 다른 버전 카운트 — 최대 version_no 기반
            max_v = db.execute(
                "SELECT MAX(version_no) AS m FROM attachment_versions WHERE parent_message_id = ?",
                (parent_id,),
            ).fetchone()
            version_no = (max_v["m"] or 1) + 1
        else:
            version_no = 1

    db.execute(
        "INSERT INTO attachment_versions (message_id, parent_message_id, version_no, room_id) VALUES (?,?,?,?)",
        (mid, parent_id, version_no, room_id),
    )
    db.commit()

    u = db.execute("SELECT display_name, avatar_color FROM users WHERE id=?", (me["id"],)).fetchone()
    payload = {
        "id": mid,
        "room_id": room_id,
        "user_id": me["id"],
        "display_name": u["display_name"],
        "avatar_color": u["avatar_color"],
        "content": original,
        "kind": kind,
        "file_path": rel_path,
        "file_name": original,
        "file_size": size,
        "file_mime": mime,
        "created_at": now,
        "version_no": version_no,
        "parent_message_id": parent_id,
        "album_id": album_id,
    }
    socketio.emit("new_message", payload, to=f"room_{room_id}")
    return jsonify(payload)


@app.route("/api/upload_part", methods=["POST"])
@login_required
def api_upload_part():
    """묶음 메시지용 파일 1개 저장 — 메시지는 만들지 않고 파일 메타만 반환.
       (글+그림 한 말풍선: 클라이언트가 이미지 파트를 먼저 올려 file_path 를 받고,
        /api/rooms/<id>/messages/multipart 로 순서대로 묶어 1개 메시지 전송) (대표 지시 2026-06-06)"""
    me = current_user()
    room_id = request.form.get("room_id")
    if not room_id:
        return jsonify({"error": "room_id가 필요합니다."}), 400
    try:
        room_id = int(room_id)
    except (TypeError, ValueError):
        return jsonify({"error": "room_id가 숫자가 아닙니다."}), 400
    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?", (room_id, me["id"])
    ).fetchone():
        abort(403)
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error": "파일이 없습니다."}), 400
    original = f.filename
    ext = ext_of(original)
    dangerous, why = _is_dangerous_filename(original)
    if dangerous:
        return jsonify({"error": f"보안 정책에 따라 차단됨: {why}"}), 400
    if ext and ext not in ALLOWED_FILE_EXT:
        return jsonify({"error": f"허용되지 않는 확장자(.{ext})"}), 400
    is_exec, why = _check_executable_magic(f)
    if is_exec:
        return jsonify({"error": f"실행 파일은 업로드할 수 없습니다 — {why}"}), 400
    content_len = request.content_length or 0
    if content_len > PER_FILE_MAX_MB * 1024 * 1024:
        return jsonify({"error": f"파일이 너무 큽니다 (단일 {PER_FILE_MAX_MB}MB 초과)"}), 413
    if not _check_rate_limit(me["id"], "upload", max_per_minute=60):
        return jsonify({"error": "업로드 속도 제한 — 잠시 후 다시 시도하세요 (1분에 60개)"}), 429
    safe_base = secure_filename(original) or "file"
    if not ext_of(safe_base):
        safe_base = f"{safe_base}.{ext}" if ext else safe_base
    unique = f"{uuid.uuid4().hex[:12]}_{safe_base}"
    room_dir = os.path.join(UPLOAD_DIR, str(room_id))
    os.makedirs(room_dir, exist_ok=True)
    fpath = os.path.join(room_dir, unique)
    f.save(fpath)
    size = os.path.getsize(fpath)
    try:
        _infected, _vdetail = _scan_file_for_virus(fpath)
    except Exception as _ve:
        _infected, _vdetail = (False, f"scan-exc:{_ve}")
    if _infected:
        try:
            os.remove(fpath)
        except Exception:
            pass
        return jsonify({"error": f"바이러스가 검출되어 업로드가 차단되었습니다 ({_vdetail})"}), 400
    mime = f.mimetype or mimetypes.guess_type(unique)[0] or "application/octet-stream"
    kind = "image" if is_image_ext(ext) else "file"
    rel_path = f"{room_id}/{unique}"
    return jsonify({
        "file_path": rel_path, "file_name": original,
        "file_size": size, "file_mime": mime, "kind": kind,
    })


@app.route("/api/rooms/<int:room_id>/messages/multipart", methods=["POST"])
@login_required
def api_messages_multipart(room_id):
    """묶음 메시지 전송 — 글/그림 파트 배열을 순서대로 1개 말풍선 메시지로 저장·전송. (대표 지시 2026-06-06)"""
    me = current_user()
    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?", (room_id, me["id"])
    ).fetchone():
        abort(403)
    data = request.get_json(silent=True) or {}
    raw_parts = data.get("parts") or []
    if not isinstance(raw_parts, list) or not raw_parts:
        return jsonify({"error": "parts 가 필요합니다."}), 400
    if len(raw_parts) > 40:
        return jsonify({"error": "묶음은 최대 40조각까지 가능합니다."}), 400
    parts = []
    text_acc = []
    for p in raw_parts:
        if not isinstance(p, dict):
            continue
        ptype = p.get("type")
        if ptype == "text":
            t = p.get("text")
            if not isinstance(t, str):
                continue
            t = t[:4000]
            if t.strip() == "":
                continue
            parts.append({"type": "text", "text": t})
            text_acc.append(t)
        elif ptype == "image":
            fp = p.get("file_path")
            if not isinstance(fp, str) or not fp:
                continue
            # 보안 — 이 방에 업로드된 파일만 허용 (경로 주입 차단)
            if ".." in fp or not fp.startswith(f"{room_id}/"):
                return jsonify({"error": "허용되지 않는 파일 경로"}), 400
            if not os.path.exists(os.path.join(UPLOAD_DIR, fp)):
                return jsonify({"error": "업로드된 파일을 찾을 수 없습니다."}), 400
            parts.append({
                "type": "image",
                "file_path": fp,
                "file_name": str(p.get("file_name") or "image")[:255],
                "file_size": int(p.get("file_size") or 0),
                "file_mime": str(p.get("file_mime") or "image/png")[:120],
            })
    if not parts:
        return jsonify({"error": "보낼 조각이 없습니다."}), 400
    # content — 미리보기·검색·알림용. 글 합치되 없으면 [사진].
    content = "\n".join(text_acc).strip()
    if not content:
        content = "[사진]"
    if len(content) > 8000:
        content = content[:8000]
    now = datetime.now(timezone.utc).isoformat()
    parts_json = json.dumps(parts, ensure_ascii=False)
    cur = db.execute(
        "INSERT INTO messages (room_id, user_id, content, kind, created_at, parts) VALUES (?,?,?,?,?,?)",
        (room_id, me["id"], content, "multipart", now, parts_json),
    )
    mid = cur.lastrowid
    db.commit()
    u = db.execute(
        "SELECT display_name, avatar_color FROM users WHERE id=?", (me["id"],)
    ).fetchone()
    payload = {
        "id": mid, "room_id": room_id, "user_id": me["id"],
        "display_name": u["display_name"], "avatar_color": u["avatar_color"],
        "content": content, "kind": "multipart", "created_at": now,
        "parts": parts,
    }
    socketio.emit("new_message", payload, to=f"room_{room_id}")
    # 푸시 — 텍스트 전송과 동일 패턴
    if PYWEBPUSH_OK:
        try:
            r = db.execute("SELECT name FROM rooms WHERE id=?", (room_id,)).fetchone()
            room_name = r["name"] if r and r["name"] else "채팅"
        except Exception:
            room_name = "채팅"
        push_title = f"💬 {u['display_name']} ({room_name})"
        push_body = content[:120]
        import threading
        threading.Thread(
            target=push_message_to_room_members,
            args=(room_id, me["id"], push_title, push_body),
            kwargs={"url": f"{BASE_PATH}/chat?room={room_id}", "tag": f"room_{room_id}"},
            daemon=True,
        ).start()
    return jsonify(payload)


def _is_safe_remote_url(url):
    """SSRF 방지 — http(s) 이미지 주소만, 사설/내부 IP·비표준 포트 차단. (대표 지시 2026-06-06)"""
    import socket, ipaddress
    from urllib.parse import urlparse
    try:
        p = urlparse(url)
    except Exception:
        return False, "잘못된 주소"
    if p.scheme not in ("http", "https"):
        return False, "http/https 만 허용"
    host = p.hostname
    if not host:
        return False, "호스트 없음"
    port = p.port or (443 if p.scheme == "https" else 80)
    if port not in (80, 443):
        return False, "허용되지 않는 포트"
    try:
        infos = socket.getaddrinfo(host, port, proto=socket.IPPROTO_TCP)
    except Exception:
        return False, "주소 해석 실패"
    for info in infos:
        ip = info[4][0]
        try:
            ipobj = ipaddress.ip_address(ip)
        except Exception:
            return False, "IP 확인 실패"
        if (ipobj.is_private or ipobj.is_loopback or ipobj.is_link_local
                or ipobj.is_reserved or ipobj.is_multicast or ipobj.is_unspecified):
            return False, "내부망 주소는 가져올 수 없습니다"
    return True, ""


@app.route("/api/fetch_remote_image")
@login_required
def api_fetch_remote_image():
    """외부(타 출처) 이미지를 서버가 대신 받아 전달 — 묶음 메시지용. SSRF 가드 포함. (대표 지시 2026-06-06)"""
    import urllib.request, urllib.error
    me = current_user()
    url = (request.args.get("url") or "").strip()
    if not url or len(url) > 2000:
        return jsonify({"error": "url 이 비었거나 너무 깁니다."}), 400
    if not _check_rate_limit(me["id"], "img_proxy", max_per_minute=120):
        return jsonify({"error": "요청이 많습니다. 잠시 후 다시."}), 429
    ok, why = _is_safe_remote_url(url)
    if not ok:
        return jsonify({"error": why}), 400

    class _SafeRedirect(urllib.request.HTTPRedirectHandler):
        def redirect_request(self, req, fp, code, msg, headers, newurl):
            _ok, _ = _is_safe_remote_url(newurl)
            if not _ok:
                raise urllib.error.HTTPError(newurl, code, "redirect blocked", headers, fp)
            return urllib.request.HTTPRedirectHandler.redirect_request(self, req, fp, code, msg, headers, newurl)

    MAXB = 15 * 1024 * 1024
    try:
        opener = urllib.request.build_opener(_SafeRedirect())
        req_obj = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 KNK-Messenger"})
        with opener.open(req_obj, timeout=8) as r:
            ctype = (r.headers.get("Content-Type") or "").split(";")[0].strip().lower()
            if not ctype.startswith("image/"):
                return jsonify({"error": "이미지가 아닙니다."}), 400
            data = r.read(MAXB + 1)
    except Exception as e:
        return jsonify({"error": f"가져오기 실패: {e}"}), 502
    if not data:
        return jsonify({"error": "빈 응답"}), 502
    if len(data) > MAXB:
        return jsonify({"error": "이미지가 너무 큽니다(15MB)."}), 413
    resp = make_response(data)
    resp.headers["Content-Type"] = ctype
    resp.headers["Cache-Control"] = "private, max-age=300"
    return resp


@app.route("/uploads/<int:room_id>/<path:filename>")
@login_required
def serve_upload(room_id, filename):
    me = current_user()
    db = get_db()
    # 권한 — 둘 중 하나면 통과:
    #   (1) 주소에 박힌 방(파일이 물리적으로 저장된 방)의 멤버
    #   (2) 이 파일(file_path)을 참조하는 메시지가 있는 방의 멤버
    #       → 전달(Forward)본 포함. 전달본은 원본 방의 file_path 를 그대로 공유하므로,
    #         전달받은 방의 멤버도 같은 파일을 볼 수 있어야 한다 (미리보기·다운로드).
    #         전달은 원본 방 멤버만 가능하므로 권한 누수 없음. (버그수정 2026-06-01)
    ok = db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?", (room_id, me["id"])
    ).fetchone()
    if not ok:
        file_ref = f"{room_id}/{filename}"
        ok = db.execute(
            """SELECT 1
                 FROM messages m
                 JOIN room_members rm ON rm.room_id = m.room_id
                WHERE m.file_path = ? AND rm.user_id = ?
                LIMIT 1""",
            (file_ref, me["id"]),
        ).fetchone()
    if not ok:
        abort(403)

    # 다운로드(?dl=1) — '보낸 그대로의 원본 파일명'으로 강제 저장.
    #   휴대폰 브라우저는 <a download> 속성을 무시하므로, 서버가 Content-Disposition 으로
    #   DB 의 원본 file_name 을 지정한다(UTF-8 → 한글 보존). 같은 이름이 여러 개여도
    #   브라우저가 알아서 '이름 (1)' 로 구분 저장. (대표 지시 2026-06-02)
    if request.args.get("dl"):
        row = db.execute(
            "SELECT file_name FROM messages WHERE file_path = ? AND file_name IS NOT NULL AND file_name <> '' ORDER BY id ASC LIMIT 1",
            (f"{room_id}/{filename}",),
        ).fetchone()
        dl_name = row["file_name"] if row else filename
        return send_from_directory(
            os.path.join(UPLOAD_DIR, str(room_id)),
            filename,
            as_attachment=True,
            download_name=dl_name,
        )

    # XSS 방어 (대표 지시 2026-05-27 옵션 A):
    # HTML·SVG·XML 등 브라우저가 자동 실행할 수 있는 파일은
    #   1) Content-Disposition: attachment 로 강제 다운로드 (인라인 렌더 차단)
    #   2) mimetype application/octet-stream 으로 브라우저 자동 실행 방지
    # → 메신저 화면에서 절대 자동 실행 안 되고, 다운로드 후 본인 PC 에서만 실행.
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in UNSAFE_INLINE_EXT:
        return send_from_directory(
            os.path.join(UPLOAD_DIR, str(room_id)),
            filename,
            as_attachment=True,
            mimetype="application/octet-stream",
        )
    return send_from_directory(os.path.join(UPLOAD_DIR, str(room_id)), filename)


@app.route("/api/albums/<album_id>/zip")
@login_required
def api_album_zip(album_id):
    """앨범(같은 album_id 의 image 메시지들)을 ZIP 으로 묶어 다운로드.
    권한: 해당 방의 멤버여야 한다.
    """
    import zipfile
    import io
    from flask import send_file as _send_file

    if not album_id or len(album_id) > 80:
        abort(400)
    me = current_user()
    db = get_db()
    rows = db.execute("""
        SELECT m.file_path, m.file_name, m.room_id, m.created_at
          FROM messages m
         WHERE m.album_id = ?
           AND m.kind = 'image'
           AND m.file_path IS NOT NULL
         ORDER BY m.id ASC
    """, (album_id,)).fetchall()
    if not rows:
        abort(404)
    room_id = rows[0]["room_id"]
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?", (room_id, me["id"])
    ).fetchone():
        abort(403)

    buf = io.BytesIO()
    used = set()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for r in rows:
            src = os.path.join(UPLOAD_DIR, r["file_path"])
            if not os.path.exists(src):
                continue
            name = r["file_name"] or os.path.basename(r["file_path"])
            # 동명 충돌 방지 — 사진명_1.jpg, _2.jpg 식
            arc = name
            n = 1
            while arc in used:
                stem, ext = os.path.splitext(name)
                arc = f"{stem}_{n}{ext}"
                n += 1
            used.add(arc)
            zf.write(src, arcname=arc)
    if not used:
        abort(404)
    buf.seek(0)

    short = (album_id or "album")[:8]
    zip_name = f"album_{short}.zip"
    return _send_file(
        buf,
        mimetype="application/zip",
        as_attachment=True,
        download_name=zip_name,
    )


@app.route("/api/rooms/<int:room_id>/attachments")
@login_required
def api_room_attachments(room_id):
    me = current_user()
    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?", (room_id, me["id"])
    ).fetchone():
        abort(403)
    kind = request.args.get("kind", "all")  # all | image | file
    sql = """
        SELECT m.id, m.kind, m.file_path, m.file_name, m.file_size, m.file_mime, m.created_at,
               u.id AS user_id, u.display_name, u.avatar_color
          FROM messages m JOIN users u ON u.id = m.user_id
         WHERE m.room_id = ? AND m.file_path IS NOT NULL
    """
    params = [room_id]
    if kind == "image":
        sql += " AND m.kind = 'image'"
    elif kind == "file":
        sql += " AND m.kind = 'file'"
    sql += " ORDER BY m.id DESC"
    rows = db.execute(sql, params).fetchall()
    return jsonify([dict(r) for r in rows])


# ---------- 요청(티켓) ----------
@app.route("/api/rooms/<int:room_id>/requests")
@login_required
def api_requests_list(room_id):
    me = current_user()
    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?", (room_id, me["id"])
    ).fetchone():
        abort(403)
    status = request.args.get("status")
    sql = """
        SELECT q.*, ub.display_name AS requested_by_name, ub.avatar_color AS requested_by_color,
               ua.display_name AS assigned_to_name, ua.avatar_color AS assigned_to_color,
               m.content AS source_message
          FROM requests q
          JOIN users ub ON ub.id = q.requested_by
          LEFT JOIN users ua ON ua.id = q.assigned_to
          LEFT JOIN messages m ON m.id = q.message_id
         WHERE q.room_id = ?
    """
    params = [room_id]
    if status:
        sql += " AND q.status = ?"
        params.append(status)
    sql += " ORDER BY (q.status='open') DESC, (q.due_date IS NULL), q.due_date ASC, q.id DESC"
    rows = db.execute(sql, params).fetchall()
    my_role = _my_room_role(db, room_id, me["id"])   # 한 방이라 1회 계산
    out = []
    for r in rows:
        d = dict(r)
        d.update(_request_perms(my_role, me["id"], r))   # 역할별 동작 권한 플래그
        out.append(d)
    return jsonify(out)


def _notify_requests_updated_users(room_id, *user_ids):
    """요청 변경을 대상 사용자(담당자·요청자) 본인 SID로 직접 전송.
    방을 안 열어둔 상태(메인창만 보는 중)에서도 '내 요청' 배지가 실시간으로 갱신되게 한다.
    (방 단위 emit 은 그 방 소켓에 join 한 클라이언트만 받으므로 한계가 있음 — 멘션과 동일 방식 적용)"""
    targets = set()
    for uid in user_ids:
        if not uid:
            continue
        try:
            targets.add(int(uid))
        except (TypeError, ValueError):
            pass
    for uid in targets:
        for sid in list(_user_connections.get(uid, {}).keys()):
            try:
                socketio.emit("requests_updated", {"room_id": room_id}, to=sid)
            except Exception:
                pass


@app.route("/api/requests", methods=["POST"])
@login_required
def api_request_create():
    me = current_user()
    data = request.get_json(silent=True) or {}
    room_id = data.get("room_id")
    title = (data.get("title") or "").strip()
    if not room_id or not title:
        return jsonify({"error": "방과 제목은 필수입니다."}), 400
    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?", (room_id, me["id"])
    ).fetchone():
        abort(403)
    now = datetime.now(timezone.utc).isoformat()
    cur = db.execute("""
        INSERT INTO requests (room_id, message_id, title, description, requested_by, assigned_to,
                              due_date, status, priority, created_at, updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
    """, (
        room_id,
        data.get("message_id"),
        title,
        (data.get("description") or "").strip() or None,
        me["id"],
        data.get("assigned_to"),
        data.get("due_date") or None,
        "open",
        data.get("priority") or "normal",
        now, now,
    ))
    qid = cur.lastrowid

    # 시스템 메시지로 알림
    assignee_name = ""
    if data.get("assigned_to"):
        a = db.execute("SELECT display_name FROM users WHERE id=?", (data.get("assigned_to"),)).fetchone()
        if a:
            assignee_name = f" → {a['display_name']}"
    due_part = f" (납기 {data.get('due_date')})" if data.get("due_date") else ""
    sys_msg = f"📌 요청 등록{assignee_name}{due_part}: {title}"
    cur = db.execute(
        "INSERT INTO messages (room_id, user_id, content, kind, created_at) VALUES (?,?,?,?,?)",
        (room_id, me["id"], sys_msg, "system", now),
    )
    sys_mid = cur.lastrowid
    db.commit()

    # 채팅방에 시스템 메시지 emit
    socketio.emit("new_message", {
        "id": sys_mid, "room_id": room_id, "user_id": me["id"],
        "display_name": me["display_name"], "avatar_color": me["avatar_color"],
        "content": sys_msg, "kind": "system", "created_at": now,
    }, to=f"room_{room_id}")
    socketio.emit("requests_updated", {"room_id": room_id}, to=f"room_{room_id}")
    # 담당자·요청자 본인에게 직접 — 그 방을 안 열어둬도 '내 요청' 배지 즉시 갱신
    _notify_requests_updated_users(room_id, data.get("assigned_to"), me["id"])

    # Web Push: 담당자 직접 통지
    if PYWEBPUSH_OK and data.get("assigned_to") and int(data["assigned_to"]) != me["id"]:
        import threading
        room = db.execute("SELECT name FROM rooms WHERE id=?", (room_id,)).fetchone()
        room_name = room["name"] if room else ""
        push_title = f"📌 새 요청 — {me['display_name']}"
        push_body = f"[{room_name}] {title}" + (f" (납기 {data.get('due_date')})" if data.get("due_date") else "")
        threading.Thread(
            target=send_push_to_user,
            args=(int(data["assigned_to"]), push_title, push_body),
            kwargs={"url": f"{BASE_PATH}/chat?room={room_id}", "tag": f"req_{qid}"},
            daemon=True,
        ).start()

    return jsonify({"id": qid})


# 요청 상태 라벨 (보류=on_hold 포함) — 대표 지시 2026-05-21
REQ_STATUS_LABELS = {"open": "열림", "in_progress": "진행중", "done": "완료",
                     "cancelled": "취소", "on_hold": "보류"}
# 상태 전환 → 필요한 권한 플래그 (open 은 보류해제/재오픈 둘 중 하나로 별도 처리)
_REQ_STATUS_PERM = {"in_progress": "can_start", "done": "can_complete",
                    "cancelled": "can_cancel", "on_hold": "can_hold"}


def _request_perms(my_role, me_id, row):
    """요청에 대해 현재 사용자가 할 수 있는 동작 플래그.
    담당자 = 시작·완료 / 요청자 = 취소·보류·보류해제·재오픈 / 방장·PM = 전체(override).
    담당자 미지정이면 요청자가 시작·완료까지 가능. (대표 지시 2026-05-21)"""
    is_assignee = (row["assigned_to"] == me_id)
    is_requester = (row["requested_by"] == me_id)
    unassigned = (row["assigned_to"] is None)
    is_manager = my_role in ("host", "sub_host")   # 방장 또는 PM
    can_assignee = is_assignee or (unassigned and is_requester) or is_manager
    can_requester = is_requester or is_manager
    st = row["status"]
    return {
        "can_start":    bool(can_assignee and st == "open"),
        "can_complete": bool(can_assignee and st in ("open", "in_progress")),
        "can_cancel":   bool(can_requester and st in ("open", "in_progress", "on_hold")),
        "can_hold":     bool(can_requester and st in ("open", "in_progress")),
        "can_unhold":   bool(can_requester and st == "on_hold"),
        "can_reopen":   bool(can_requester and st in ("done", "cancelled")),
        "is_assignee":  bool(is_assignee),
        "is_requester": bool(is_requester),
        "is_manager":   bool(is_manager),
    }


@app.route("/api/requests/<int:req_id>", methods=["PATCH"])
@login_required
def api_request_update(req_id):
    me = current_user()
    data = request.get_json(silent=True) or {}
    db = get_db()
    row = db.execute("SELECT * FROM requests WHERE id=?", (req_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    my_role = _my_room_role(db, row["room_id"], me["id"])
    perms = _request_perms(my_role, me["id"], row)
    is_manager = perms["is_manager"]
    is_requester = perms["is_requester"]

    new_status = data.get("status")
    status_changing = (new_status is not None and new_status != row["status"])
    if status_changing:
        # 역할별 상태 전환 권한 검사
        if new_status == "open":
            allowed = perms["can_unhold"] or perms["can_reopen"]
        else:
            pk = _REQ_STATUS_PERM.get(new_status)
            allowed = bool(pk and perms.get(pk))
        if not allowed:
            return jsonify({"error": "이 동작을 할 권한이 없습니다."}), 403
        # 완료는 메시지 필수
        if new_status == "done" and not (data.get("message") or "").strip():
            return jsonify({"error": "완료 메시지를 입력해주세요."}), 400
    else:
        # 상태 외 필드(제목·담당자·납기 등) 수정 — 요청자 또는 방장·PM 만
        edits = [f for f in ("title", "description", "assigned_to", "due_date", "priority") if f in data]
        if edits and not (is_requester or is_manager):
            return jsonify({"error": "이 요청을 수정할 권한이 없습니다."}), 403

    fields, args = [], []
    for f in ("title", "description", "assigned_to", "due_date", "status", "priority"):
        if f in data:
            v = data[f]
            if isinstance(v, str):
                v = v.strip() or None
            fields.append(f"{f} = ?")
            args.append(v)
    if not fields:
        return jsonify({"ok": True})
    now = datetime.now(timezone.utc).isoformat()
    # ★ 순서 주의: SQL 은 "{fields}, updated_at=? WHERE id=?" 라서 args 도
    #   [필드값들..., (closed_at 값 있으면), updated_at(now), req_id] 순이어야 함.
    #   closed_at 을 fields 에 추가하면 그 값이 updated_at 보다 앞에 와야 한다.
    if new_status in ("done", "cancelled"):
        fields.append("closed_at = ?")
        args.append(now)
    elif new_status in ("open", "in_progress", "on_hold"):
        fields.append("closed_at = ?")   # 재개·보류 → 종료시각 해제
        args.append(None)
    args.append(now)        # updated_at (NOT NULL)
    args.append(req_id)     # WHERE id
    db.execute(f"UPDATE requests SET {', '.join(fields)}, updated_at = ? WHERE id = ?", args)
    db.commit()

    emitted = []
    # 완료 메시지 — 일반 메시지로 방에 게시 (요청자가 보고 알림 받음)
    done_msg = (data.get("message") or "").strip() if new_status == "done" else ""
    if done_msg:
        body = f"✅ [요청 완료: {row['title']}] {done_msg}"
        cm = db.execute(
            "INSERT INTO messages (room_id, user_id, content, kind, created_at) VALUES (?,?,?,?,?)",
            (row["room_id"], me["id"], body, "text", now),
        )
        db.commit()
        emitted.append({"id": cm.lastrowid, "room_id": row["room_id"], "user_id": me["id"],
                        "display_name": me["display_name"], "avatar_color": me["avatar_color"],
                        "content": body, "kind": "text", "created_at": now})

    # 상태 변경 시스템 메시지
    if status_changing:
        sys_msg = f"📌 요청 [{row['title']}] → {REQ_STATUS_LABELS.get(new_status, new_status)}"
        cs = db.execute(
            "INSERT INTO messages (room_id, user_id, content, kind, created_at) VALUES (?,?,?,?,?)",
            (row["room_id"], me["id"], sys_msg, "system", now),
        )
        db.commit()
        emitted.append({"id": cs.lastrowid, "room_id": row["room_id"], "user_id": me["id"],
                        "display_name": me["display_name"], "avatar_color": me["avatar_color"],
                        "content": sys_msg, "kind": "system", "created_at": now})

    for payload in emitted:
        socketio.emit("new_message", payload, to=f"room_{row['room_id']}")
    socketio.emit("requests_updated", {"room_id": row["room_id"]}, to=f"room_{row['room_id']}")
    # 담당자·요청자 본인에게 직접 — 방을 안 열어둬도 양쪽 '내 요청' 배지 즉시 갱신
    _notify_requests_updated_users(row["room_id"], row["assigned_to"], row["requested_by"], me["id"])

    # 완료 시 요청자에게 푸시 (요청자가 앱을 안 보고 있으면)
    if done_msg and PYWEBPUSH_OK:
        try:
            requester = row["requested_by"]
            if requester and requester != me["id"] and not _user_has_active_session(requester):
                rn = db.execute("SELECT name FROM rooms WHERE id=?", (row["room_id"],)).fetchone()
                room_name = rn["name"] if rn else "채팅"
                import threading as _t
                _t.Thread(target=send_push_to_user,
                          args=(requester, f"✅ 요청 완료 ({room_name})", f"{row['title']} — {done_msg[:80]}"),
                          kwargs={"url": f"{BASE_PATH}/chat?room={row['room_id']}", "tag": f"room_{row['room_id']}"},
                          daemon=True).start()
        except Exception:
            pass

    return jsonify({"ok": True})


@app.route("/api/my/requests")
@login_required
def api_my_requests():
    """내 요청 종합 — 내가 '받은 요청'(assigned_to=나) + 내가 '보낸 요청'(requested_by=나) 모두.
    요청자(예: 대표)가 위임한 요청도 한곳에서 추적·상태변경 가능하게. (대표 지시 2026-05-21)"""
    me = current_user()
    db = get_db()
    rows = db.execute("""
        SELECT q.*, r.name AS room_name, it.customer AS item_customer, it.code AS item_code,
               ub.display_name AS requested_by_name, ub.avatar_color AS requested_by_color,
               ua.display_name AS assigned_to_name, ua.avatar_color AS assigned_to_color
          FROM requests q
          JOIN rooms r ON r.id = q.room_id
          LEFT JOIN items it ON it.room_id = q.room_id
          JOIN users ub ON ub.id = q.requested_by
          LEFT JOIN users ua ON ua.id = q.assigned_to
         WHERE (q.assigned_to = ? OR q.requested_by = ?) AND q.status IN ('open','in_progress','on_hold')
         ORDER BY (q.due_date IS NULL), q.due_date ASC, q.id DESC
    """, (me["id"], me["id"])).fetchall()
    out = []
    _role_cache = {}
    for r in rows:
        d = dict(r)
        rid = r["room_id"]
        if rid not in _role_cache:
            _role_cache[rid] = _my_room_role(db, rid, me["id"])   # 방마다 1회만 계산
        d.update(_request_perms(_role_cache[rid], me["id"], r))
        out.append(d)
    return jsonify(out)


# ---------- 검색 ----------
def fts_query_safe(q):
    # FTS5 query: 단어 단위로 분리 후 prefix 매칭. 특수문자 제거.
    tokens = re.findall(r"[\w가-힣]+", q)
    if not tokens:
        return None
    return " ".join(t + "*" for t in tokens)


@app.route("/api/rooms/<int:room_id>/summary")
@login_required
def api_room_summary(room_id):
    """방 요약 — 프로젝트 카드 헤더 / 다이제스트용 카운트"""
    me = current_user()
    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?", (room_id, me["id"])
    ).fetchone():
        abort(403)
    counts = db.execute("""
        SELECT
            (SELECT COUNT(*) FROM messages WHERE room_id=? AND kind='text') AS text_count,
            (SELECT COUNT(*) FROM messages WHERE room_id=? AND kind='image') AS image_count,
            (SELECT COUNT(*) FROM messages WHERE room_id=? AND kind='file') AS file_count,
            (SELECT COUNT(*) FROM requests WHERE room_id=? AND status='open') AS open_requests,
            (SELECT COUNT(*) FROM requests WHERE room_id=? AND status='in_progress') AS active_requests,
            (SELECT COUNT(*) FROM requests WHERE room_id=? AND status='done') AS done_requests,
            (SELECT MAX(created_at) FROM messages WHERE room_id=?) AS last_activity,
            (SELECT COUNT(*) FROM room_members WHERE room_id=?) AS members
    """, (room_id, room_id, room_id, room_id, room_id, room_id, room_id, room_id)).fetchone()
    return jsonify(dict(counts))


@app.route("/api/items/dashboard")
@login_required
def api_items_dashboard():
    """전체 프로젝트 대시보드 — 카운트·최근활동 한눈"""
    me = current_user()
    if _is_guest(me):
        return jsonify({"error": "외부 사용자는 사용할 수 없습니다."}), 403
    db = get_db()
    rows = db.execute("""
        SELECT r.id AS room_id, r.name, it.code, it.customer, it.status, it.due_date,
               (SELECT COUNT(*) FROM messages WHERE room_id=r.id AND kind='image') AS image_count,
               (SELECT COUNT(*) FROM messages WHERE room_id=r.id AND kind='file') AS file_count,
               (SELECT COUNT(*) FROM requests WHERE room_id=r.id AND status='open') AS open_requests,
               (SELECT COUNT(*) FROM requests WHERE room_id=r.id AND status='in_progress') AS active_requests,
               (SELECT MAX(created_at) FROM messages WHERE room_id=r.id) AS last_activity
          FROM rooms r
          JOIN items it ON it.room_id = r.id
          JOIN room_members rm ON rm.room_id = r.id AND rm.user_id = ?
         ORDER BY (it.status = 'active') DESC, last_activity DESC NULLS LAST
    """, (me["id"],)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/digest")
@login_required
def api_digest():
    """오늘 / 이번주 — 로그인 직후 보여주는 자동 다이제스트.

    일반 메신저에서 묻혀서 못 보고 지나가는 일을 막기 위한 핵심 기능.
    """
    me = current_user()
    if _is_guest(me):
        return jsonify({"error": "외부 사용자는 사용할 수 없습니다."}), 403
    db = get_db()
    today_iso = datetime.now(timezone.utc).date().isoformat()

    overdue = db.execute("""
        SELECT q.id, q.title, q.due_date, r.name AS room_name, r.id AS room_id,
               it.customer, it.code
          FROM requests q
          JOIN rooms r ON r.id = q.room_id
          LEFT JOIN items it ON it.room_id = q.room_id
         WHERE q.assigned_to = ? AND q.status IN ('open','in_progress')
           AND q.due_date IS NOT NULL AND q.due_date < ?
         ORDER BY q.due_date ASC
    """, (me["id"], today_iso)).fetchall()

    today_due = db.execute("""
        SELECT q.id, q.title, q.due_date, r.name AS room_name, r.id AS room_id,
               it.customer, it.code
          FROM requests q
          JOIN rooms r ON r.id = q.room_id
          LEFT JOIN items it ON it.room_id = q.room_id
         WHERE q.assigned_to = ? AND q.status IN ('open','in_progress')
           AND q.due_date = ?
    """, (me["id"], today_iso)).fetchall()

    upcoming = db.execute("""
        SELECT q.id, q.title, q.due_date, r.name AS room_name, r.id AS room_id,
               it.customer, it.code
          FROM requests q
          JOIN rooms r ON r.id = q.room_id
          LEFT JOIN items it ON it.room_id = q.room_id
         WHERE q.assigned_to = ? AND q.status IN ('open','in_progress')
           AND q.due_date IS NOT NULL AND q.due_date > ?
           AND date(q.due_date) <= date(?, '+7 days')
         ORDER BY q.due_date ASC
    """, (me["id"], today_iso, today_iso)).fetchall()

    no_due = db.execute("""
        SELECT COUNT(*) AS n FROM requests
         WHERE assigned_to = ? AND status IN ('open','in_progress') AND due_date IS NULL
    """, (me["id"],)).fetchone()["n"]

    requested_open = db.execute("""
        SELECT q.id, q.title, q.due_date, q.status,
               r.name AS room_name, r.id AS room_id,
               ua.display_name AS assigned_to_name,
               it.customer, it.code
          FROM requests q
          JOIN rooms r ON r.id = q.room_id
          LEFT JOIN items it ON it.room_id = q.room_id
          LEFT JOIN users ua ON ua.id = q.assigned_to
         WHERE q.requested_by = ? AND q.status IN ('open','in_progress')
         ORDER BY (q.due_date IS NULL), q.due_date ASC
    """, (me["id"],)).fetchall()

    stale_items = db.execute("""
        SELECT r.id AS room_id, r.name, it.customer, it.code, it.status,
               (SELECT MAX(created_at) FROM messages WHERE room_id=r.id) AS last_activity
          FROM rooms r
          JOIN items it ON it.room_id = r.id
          JOIN room_members rm ON rm.room_id = r.id AND rm.user_id = ?
         WHERE it.status = 'active'
           AND (SELECT MAX(created_at) FROM messages WHERE room_id=r.id) IS NOT NULL
           AND date((SELECT MAX(created_at) FROM messages WHERE room_id=r.id)) < date(?, '-7 days')
         ORDER BY last_activity ASC
         LIMIT 10
    """, (me["id"], today_iso)).fetchall()

    return jsonify({
        "overdue": [dict(r) for r in overdue],
        "today_due": [dict(r) for r in today_due],
        "upcoming": [dict(r) for r in upcoming],
        "no_due_count": no_due,
        "requested_open": [dict(r) for r in requested_open],
        "stale_items": [dict(r) for r in stale_items],
    })


@app.route("/api/rooms/<int:room_id>/export.xlsx")
@login_required
def api_room_export_xlsx(room_id):
    """대화방 기록 Excel 내보내기 — 6시트(개요/메시지/요청/첨부/스레드/멤버).

    일반 메신저로 못 했던 기능: 프로젝트 단위로 모든 이력을 한 파일로.
    감사·법무 보고·인수인계용. KNK 브랜드 서식(레드 헤더·맑은 고딕) + 한국시간(KST) 표기.
    메시지 시트에 사진 썸네일 인라인(작게) + 모든 첨부 '열기'/동영상 '▶재생' 하이퍼링크,
    '스레드' 시트에 원글+답글 정리 (대표 지시 2026-05-31).

    기간 필터 (대표 지시 2026-05-28):
      ?from=YYYY-MM-DD&to=YYYY-MM-DD — 메시지·요청·첨부 시트만 적용
      (개요는 전체 통계 + 적용된 기간 표시)
      파라미터 없거나 잘못된 경우 전체 반환.
    """
    me = current_user()
    db = get_db()
    if _is_guest(me):
        return jsonify({"error": "외부 사용자는 대화 내용을 다운로드할 수 없습니다."}), 403
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?", (room_id, me["id"])
    ).fetchone():
        abort(403)

    # ── 기간 필터 파라미터 파싱 (YYYY-MM-DD)
    from_raw = (request.args.get("from") or "").strip()
    to_raw = (request.args.get("to") or "").strip()
    def _valid_date(s):
        try:
            datetime.strptime(s, "%Y-%m-%d"); return s
        except Exception:
            return None
    from_date = _valid_date(from_raw) if from_raw else None
    to_date   = _valid_date(to_raw) if to_raw else None
    # SQL WHERE 절: created_at >= from 00:00 AND created_at < to+1 00:00 (to 포함)
    period_label = "전체"
    if from_date and to_date:
        period_label = f"{from_date} ~ {to_date}"
    elif from_date:
        period_label = f"{from_date} ~"
    elif to_date:
        period_label = f"~ {to_date}"

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from flask import send_file

    # 기간 조건 빌드 (created_at 컬럼명을 받아서 ? 바인딩 생성)
    def _date_where(col):
        wh = []
        params = []
        if from_date:
            wh.append(f"{col} >= ?")
            params.append(from_date + " 00:00:00")
        if to_date:
            wh.append(f"{col} < date(?, '+1 day')")
            params.append(to_date)
        return ((" AND " + " AND ".join(wh)) if wh else ""), params

    room = db.execute("SELECT * FROM rooms WHERE id=?", (room_id,)).fetchone()
    item = db.execute("SELECT * FROM items WHERE room_id=?", (room_id,)).fetchone()

    m_where, m_params = _date_where("m.created_at")
    msgs = db.execute(f"""
        SELECT m.*, u.display_name, u.title AS author_title, u.department AS author_dept
          FROM messages m JOIN users u ON u.id=m.user_id
         WHERE m.room_id=? {m_where}
         ORDER BY m.id ASC
    """, (room_id, *m_params)).fetchall()

    q_where, q_params = _date_where("q.created_at")
    reqs = db.execute(f"""
        SELECT q.*, ub.display_name AS requested_by_name, ua.display_name AS assigned_to_name,
               ub.title AS req_title, ub.department AS req_dept,
               ua.title AS asg_title, ua.department AS asg_dept
          FROM requests q
          JOIN users ub ON ub.id=q.requested_by
          LEFT JOIN users ua ON ua.id=q.assigned_to
         WHERE q.room_id=? {q_where}
         ORDER BY q.id ASC
    """, (room_id, *q_params)).fetchall()

    a_where, a_params = _date_where("m.created_at")
    attachments = db.execute(f"""
        SELECT m.id, m.kind, m.file_name, m.file_size, m.created_at, m.file_path, u.display_name,
               u.title AS author_title, u.department AS author_dept
          FROM messages m JOIN users u ON u.id=m.user_id
         WHERE m.room_id=? AND m.file_path IS NOT NULL {a_where}
         ORDER BY m.id ASC
    """, (room_id, *a_params)).fetchall()
    members = db.execute("""
        SELECT u.username, u.display_name, u.role, rm.joined_at,
               u.title AS author_title, u.department AS author_dept
          FROM room_members rm JOIN users u ON u.id=rm.user_id
         WHERE rm.room_id=? ORDER BY rm.joined_at ASC
    """, (room_id,)).fetchall()

    # ── 스타일 (KNK 브랜드: 레드 헤더 · 맑은 고딕) — 대표 지시 2026-05-31 업무용 정돈 ──
    from datetime import timedelta as _td
    FONT = "맑은 고딕"
    RED = "A5282C"; ZEBRA = "F7F7F8"; GRAYTX = "4A4A4A"
    f_base = Font(name=FONT, size=10)
    f_head = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    f_title = Font(name=FONT, bold=True, size=15, color="FFFFFF")
    f_label = Font(name=FONT, bold=True, size=10, color=GRAYTX)
    fill_head = PatternFill("solid", fgColor=RED)
    fill_zebra = PatternFill("solid", fgColor=ZEBRA)
    fill_img = PatternFill("solid", fgColor="FBEAEA")
    fill_file = PatternFill("solid", fgColor="EFF6FF")
    fill_sys = PatternFill("solid", fgColor="F3F4F6")
    _thin = Side(style="thin", color="E5E7EB")
    border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center")
    vtop = Alignment(vertical="top")

    def _kst(s):
        # 저장형식 혼용 대응 — 옛 'YYYY-MM-DD HH:MM:SS'(공백) + 새 ISO 'YYYY-MM-DDTHH:MM:SS.ffffff+00:00'
        # 둘 다 UTC 저장 → +9h 한국시간 'YYYY-MM-DD HH:MM' (초·소수점·시간대 제거) (대표 지시 2026-06-04)
        if not s:
            return ""
        try:
            core = str(s).strip().replace("T", " ")[:19]
            dt = datetime.strptime(core, "%Y-%m-%d %H:%M:%S") + _td(hours=9)
            return dt.strftime("%Y-%m-%d %H:%M")
        except Exception:
            return str(s)[:16].replace("T", " ")

    def _txt(v, limit=32000):
        # openpyxl 셀 길이 한계(32767) 보호
        s = "" if v is None else str(v)
        return s[:limit]

    def _person(name, title, dept):
        # 사람 칸 — '이름 직급 부서' (없는 항목은 생략) (대표 지시 2026-06-04)
        parts = []
        for v in (name, title, dept):
            v = str(v or "").strip()
            if v:
                parts.append(v)
        return " ".join(parts)

    KIND_LABEL = {"text": "메시지", "image": "사진", "file": "파일", "system": "시스템"}
    REQ_STATUS = {"open": "열림", "in_progress": "진행중", "done": "완료", "cancelled": "취소"}
    ITEM_STATUS = {"active": "진행중", "hold": "보류", "done": "완료", "cancelled": "취소"}

    def setup_header(ws, headers, widths=None):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = f_head; c.fill = fill_head; c.alignment = center; c.border = border
        if widths:
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

    def _rowstyle(ws, i, ncol, zebra=True, kind_fill=None):
        for col in range(1, ncol + 1):
            c = ws.cell(row=i, column=col)
            c.font = f_base
            c.border = border
            # 이미 명시 설정된 정렬(center/wrap 등)은 보존, 미설정 셀만 상단정렬
            _al = c.alignment
            if not (_al and (_al.horizontal or _al.vertical or _al.wrap_text)):
                c.alignment = vtop
            if kind_fill is not None:
                c.fill = kind_fill
            elif zebra and (i % 2 == 1):
                c.fill = fill_zebra

    wb = Workbook()

    # ── Sheet 1: 개요 (표지) ──
    ws1 = wb.active
    ws1.title = "개요"
    ws1.merge_cells("A1:B1")
    t = ws1["A1"]; t.value = "KNK 이음 — 대화방 기록 보고서"
    t.font = f_title; t.fill = fill_head
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[1].height = 34
    img_count = sum(1 for m in msgs if m["kind"] == "image")
    rows1 = [
        ["내보내기 일시", _kst(datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")) + " (KST)"],
        ["내보낸 사람", me["display_name"]],
        ["기간", period_label],
        ["", ""],
        ["방 이름", room["name"] if room else ""],
        ["타입", {"item": "프로젝트", "channel": "채널", "group": "그룹채팅", "direct": "1:1"}.get(room["type"] if room else "", "")],
    ]
    if item:
        rows1.extend([
            ["고객사", item["customer"] or ""],
            ["관리번호", item["code"] or ""],
            ["상태", ITEM_STATUS.get(item["status"], item["status"] or "")],
            ["납기", item["due_date"] or ""],
            ["영구보존", "예" if item["keep_forever"] else "아니오"],
            ["설명", item["description"] or ""],
        ])
    rows1.extend([
        ["", ""],
        ["통계", ""],
        ["  · 메시지 수", sum(1 for m in msgs if m["kind"] == "text")],
        ["  · 사진", img_count],
        ["  · 파일", sum(1 for m in msgs if m["kind"] == "file")],
        ["  · 시스템 메시지", sum(1 for m in msgs if m["kind"] == "system")],
        ["  · 요청 (전체)", len(reqs)],
        ["  · 요청 (열림+진행중)", sum(1 for r in reqs if r["status"] in ("open", "in_progress"))],
        ["  · 멤버 수", len(members)],
    ])
    for i, (k, v) in enumerate(rows1, 3):
        ck = ws1.cell(row=i, column=1, value=k); ck.font = f_label
        cv = ws1.cell(row=i, column=2, value=v); cv.font = f_base; cv.alignment = wrap
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 64
    _fr = 3 + len(rows1) + 1
    ws1.merge_cells(start_row=_fr, start_column=1, end_row=_fr, end_column=2)
    _fc = ws1.cell(row=_fr, column=1, value="㈜케이엔케이 | HAIST Innovation  ·  Human & AI create the Best")
    _fc.font = Font(name=FONT, size=9, italic=True, color="9CA3AF")

    # ── Sheet 2: 메시지 타임라인 (사진 인라인 + 모든 첨부 열기/재생 링크 — 대표 지시 2026-05-31) ──
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage
    link_font = Font(name=FONT, size=10, color="2563EB", underline="single")
    _bio_keep = []  # 임베드 이미지 BytesIO 참조 유지 (save 전 GC 방지)

    def _file_url(mm):
        try:
            return url_for("serve_upload", room_id=room_id,
                           filename=os.path.basename(mm["file_path"]), _external=True)
        except Exception:
            return None

    def _is_video(mm):
        mime = (mm["file_mime"] or "").lower()
        if mime.startswith("video/"):
            return True
        fn = (mm["file_name"] or "").lower()
        ext = fn.rsplit(".", 1)[-1] if "." in fn else ""
        return ext in ("mp4", "mov", "avi", "mkv", "webm", "m4v", "3gp", "wmv", "flv")

    ws2 = wb.create_sheet("메시지")
    setup_header(ws2, ["#", "일시(KST)", "보낸이", "구분", "내용", "미리보기", "파일/링크", "크기(B)"],
                 [6, 18, 22, 8, 55, 14, 34, 12])
    _EMBED_CAP = 400
    _embedded = 0
    for idx, m in enumerate(msgs):
        i = idx + 2
        ws2.cell(row=i, column=1, value=m["id"])
        ws2.cell(row=i, column=2, value=_kst(m["created_at"]))
        ws2.cell(row=i, column=3, value=_person(m["display_name"], m["author_title"], m["author_dept"]))
        ws2.cell(row=i, column=4, value=KIND_LABEL.get(m["kind"], m["kind"])).alignment = center
        ws2.cell(row=i, column=5, value=_txt(m["content"])).alignment = wrap
        ws2.cell(row=i, column=8, value=m["file_size"] or "")
        # 미리보기 — 사진 작은 썸네일(최대 80px) 인라인 (확대는 엑셀에서)
        if m["kind"] == "image" and m["file_path"]:
            _src = os.path.join(UPLOAD_DIR, m["file_path"])
            if os.path.exists(_src) and _embedded < _EMBED_CAP:
                try:
                    _pim = PILImage.open(_src).convert("RGB"); _pim.thumbnail((80, 80))
                    _bio = BytesIO(); _pim.save(_bio, format="PNG"); _bio.seek(0)
                    _bio_keep.append(_bio)
                    ws2.add_image(XLImage(_bio), "F%d" % i)
                    ws2.row_dimensions[i].height = max(_pim.height * 0.75 + 6, 18)
                    _embedded += 1
                except Exception:
                    pass
        _kf = {"image": fill_img, "file": fill_file, "system": fill_sys}.get(m["kind"])
        _rowstyle(ws2, i, 8, kind_fill=_kf)
        # 파일/링크 — 모든 첨부에 하이퍼링크 (_rowstyle 뒤에 링크 폰트 적용)
        if m["file_path"]:
            _url = _file_url(m)
            _lbl = ("▶ 재생 — " if _is_video(m) else "열기 — ") + (m["file_name"] or os.path.basename(m["file_path"]))
            _lc = ws2.cell(row=i, column=7, value=_lbl)
            if _url:
                _lc.hyperlink = _url
            _lc.font = link_font
            _lc.alignment = wrap
    if msgs:
        ws2.auto_filter.ref = ws2.dimensions

    # ── Sheet 3: 요청 ──
    ws3 = wb.create_sheet("요청")
    setup_header(ws3, ["#", "상태", "우선순위", "제목", "상세", "요청자", "담당자", "납기", "등록일", "마감일"], [6, 10, 10, 35, 50, 20, 20, 12, 18, 18])
    for idx, r in enumerate(reqs):
        i = idx + 2
        ws3.cell(row=i, column=1, value=r["id"])
        ws3.cell(row=i, column=2, value=REQ_STATUS.get(r["status"], r["status"])).alignment = center
        ws3.cell(row=i, column=3, value=r["priority"]).alignment = center
        ws3.cell(row=i, column=4, value=_txt(r["title"])).alignment = wrap
        ws3.cell(row=i, column=5, value=_txt(r["description"])).alignment = wrap
        ws3.cell(row=i, column=6, value=_person(r["requested_by_name"], r["req_title"], r["req_dept"]))
        ws3.cell(row=i, column=7, value=_person(r["assigned_to_name"], r["asg_title"], r["asg_dept"]))
        ws3.cell(row=i, column=8, value=r["due_date"] or "")
        ws3.cell(row=i, column=9, value=_kst(r["created_at"]))
        ws3.cell(row=i, column=10, value=_kst(r["closed_at"]))
        _rowstyle(ws3, i, 10)
    if reqs:
        ws3.auto_filter.ref = ws3.dimensions

    # ── Sheet 4: 첨부 ──
    ws4 = wb.create_sheet("첨부")
    setup_header(ws4, ["#", "구분", "파일명", "크기(B)", "올린이", "일시(KST)", "사진"], [6, 8, 50, 12, 22, 18, 14])
    for idx, a in enumerate(attachments):
        i = idx + 2
        ws4.cell(row=i, column=1, value=a["id"])
        ws4.cell(row=i, column=2, value=KIND_LABEL.get(a["kind"], a["kind"])).alignment = center
        ws4.cell(row=i, column=3, value=a["file_name"] or "")
        ws4.cell(row=i, column=4, value=a["file_size"] or "")
        ws4.cell(row=i, column=5, value=_person(a["display_name"], a["author_title"], a["author_dept"]))
        ws4.cell(row=i, column=6, value=_kst(a["created_at"]))
        # 사진 첨부면 실제 이미지를 '사진'(7번/G) 열에 축소 임베드 — 메시지·스레드 시트와 동일 (대표 지시 2026-06-05)
        if a["kind"] == "image" and a["file_path"] and _embedded < _EMBED_CAP:
            _asrc = os.path.join(UPLOAD_DIR, a["file_path"])
            if os.path.exists(_asrc):
                try:
                    _apim = PILImage.open(_asrc).convert("RGB"); _apim.thumbnail((80, 80))
                    _abio = BytesIO(); _apim.save(_abio, format="PNG"); _abio.seek(0)
                    _bio_keep.append(_abio)
                    ws4.add_image(XLImage(_abio), "G%d" % i)
                    ws4.row_dimensions[i].height = max(_apim.height * 0.75 + 6, 18)
                    _embedded += 1
                except Exception:
                    pass
        _rowstyle(ws4, i, 7)
    if attachments:
        ws4.auto_filter.ref = ws4.dimensions

    # ── Sheet 5: 스레드 (원글 + 답글 묶음 — 대표 지시 2026-05-31) ──
    #   parent_message_id 가 NULL 이면서 답글이 있는 메시지 = 스레드 원글. 그 답글들을 묶어 정리.
    try:
        from collections import defaultdict as _dd
        _tw, _tp = _date_where("mp.created_at")
        _parents = db.execute(f"""
            SELECT mp.id, mp.created_at, mp.content, mp.kind, mp.file_path, up.display_name AS author,
                   up.title AS author_title, up.department AS author_dept
              FROM messages mp JOIN users up ON up.id=mp.user_id
             WHERE mp.room_id=? AND mp.parent_message_id IS NULL
               AND COALESCE(mp.thread_hidden,0)=0
               AND EXISTS (SELECT 1 FROM messages c WHERE c.parent_message_id=mp.id)
               {_tw}
             ORDER BY mp.id ASC
        """, (room_id, *_tp)).fetchall()
        _reps = db.execute("""
            SELECT mc.parent_message_id AS pid, mc.created_at, mc.content, mc.kind, mc.file_path, uc.display_name AS author,
                   uc.title AS author_title, uc.department AS author_dept
              FROM messages mc JOIN users uc ON uc.id=mc.user_id
             WHERE mc.room_id=? AND mc.parent_message_id IS NOT NULL
             ORDER BY mc.id ASC
        """, (room_id,)).fetchall()
        _rep_by = _dd(list)
        for _r in _reps:
            _rep_by[_r["pid"]].append(_r)
        wst = wb.create_sheet("스레드")
        setup_header(wst, ["스레드", "구분", "일시(KST)", "보낸이", "내용", "사진"], [8, 8, 18, 22, 70, 14])
        _fill_parent = PatternFill("solid", fgColor="FBEAEA")
        # 사진 메시지면 실제 이미지를 '사진'(6번/F) 열에 축소 임베드 — 메시지 시트와 동일 (대표 지시 2026-06-04)
        def _t_embed(rownum, mm):
            nonlocal _embedded
            if mm["kind"] != "image" or not mm["file_path"] or _embedded >= _EMBED_CAP:
                return
            _ts = os.path.join(UPLOAD_DIR, mm["file_path"])
            if not os.path.exists(_ts):
                return
            try:
                _tpim = PILImage.open(_ts).convert("RGB"); _tpim.thumbnail((80, 80))
                _tbio = BytesIO(); _tpim.save(_tbio, format="PNG"); _tbio.seek(0)
                _bio_keep.append(_tbio)
                wst.add_image(XLImage(_tbio), "F%d" % rownum)
                wst.row_dimensions[rownum].height = max(_tpim.height * 0.75 + 6, 18)
                _embedded += 1
            except Exception:
                pass
        _ti = 2
        _tn = 0
        for _p in _parents:
            _tn += 1
            wst.cell(row=_ti, column=1, value=_tn).alignment = center
            wst.cell(row=_ti, column=2, value="원글").alignment = center
            wst.cell(row=_ti, column=3, value=_kst(_p["created_at"]))
            wst.cell(row=_ti, column=4, value=_person(_p["author"], _p["author_title"], _p["author_dept"]))
            wst.cell(row=_ti, column=5, value=_txt(_p["content"])).alignment = wrap
            _rowstyle(wst, _ti, 6, zebra=False, kind_fill=_fill_parent)
            wst.cell(row=_ti, column=4).font = Font(name=FONT, size=10, bold=True)
            _t_embed(_ti, _p)
            _ti += 1
            for _c in _rep_by.get(_p["id"], []):
                wst.cell(row=_ti, column=2, value="↳ 답글").alignment = center
                wst.cell(row=_ti, column=3, value=_kst(_c["created_at"]))
                wst.cell(row=_ti, column=4, value=_person(_c["author"], _c["author_title"], _c["author_dept"]))
                wst.cell(row=_ti, column=5, value=_txt(_c["content"])).alignment = wrap
                _rowstyle(wst, _ti, 6, zebra=False)
                _t_embed(_ti, _c)
                _ti += 1
        if not _parents:
            wst.cell(row=2, column=1, value="스레드(답글)가 없습니다.").font = f_base
        wst.freeze_panes = "A2"
    except Exception:
        pass

    # ── Sheet 6: 멤버 ──
    ws5 = wb.create_sheet("멤버")
    setup_header(ws5, ["아이디", "이름", "역할", "참여일(KST)"], [14, 24, 10, 18])
    for idx, m in enumerate(members):
        i = idx + 2
        ws5.cell(row=i, column=1, value=m["username"])
        ws5.cell(row=i, column=2, value=_person(m["display_name"], m["author_title"], m["author_dept"]))
        ws5.cell(row=i, column=3, value=m["role"])
        ws5.cell(row=i, column=4, value=_kst(m["joined_at"]))
        _rowstyle(ws5, i, 4)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    safe_name = re.sub(r'[\\/:"*?<>|]+', "_", room["name"] or f"room{room_id}")
    # 파일명에도 기간 표시 (기간 선택 시) — 대표 지시 2026-05-28
    if from_date and to_date:
        fname_period = f"{from_date.replace('-','')}-{to_date.replace('-','')}"
    elif from_date:
        fname_period = f"{from_date.replace('-','')}-now"
    elif to_date:
        fname_period = f"~{to_date.replace('-','')}"
    else:
        fname_period = "전체"
    fname = f"KNK이음_{safe_name}_{fname_period}_{today}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/rooms/<int:room_id>/timeline")
@login_required
def api_room_timeline(room_id):
    """프로젝트 타임라인 — 날짜별로 사진·파일·요청·결정 그룹.

    신규 담당자가 인수받을 때 처음부터 끝까지 한 페이지로 보기 위한 용도.
    """
    me = current_user()
    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?", (room_id, me["id"])
    ).fetchone():
        abort(403)
    msgs = db.execute("""
        SELECT m.id, m.kind, m.content, m.file_path, m.file_name, m.file_size, m.created_at,
               u.display_name, u.avatar_color
          FROM messages m JOIN users u ON u.id = m.user_id
         WHERE m.room_id = ?
         ORDER BY m.id ASC
    """, (room_id,)).fetchall()
    reqs = db.execute("""
        SELECT q.*, ub.display_name AS requested_by_name, ua.display_name AS assigned_to_name
          FROM requests q
          JOIN users ub ON ub.id = q.requested_by
          LEFT JOIN users ua ON ua.id = q.assigned_to
         WHERE q.room_id = ?
         ORDER BY q.id ASC
    """, (room_id,)).fetchall()
    return jsonify({
        "messages": [dict(r) for r in msgs],
        "requests": [dict(r) for r in reqs],
    })


@app.route("/api/push/vapid_public")
@login_required
def api_push_vapid_public():
    pk = vapid_public_key_b64u()
    if not pk:
        return jsonify({"error": "VAPID 키 없음 — generate_vapid.py 실행 필요"}), 503
    return jsonify({"public_key": pk, "enabled": PYWEBPUSH_OK})


@app.route("/api/push/subscribe", methods=["POST"])
@login_required
def api_push_subscribe():
    me = current_user()
    data = request.get_json(silent=True) or {}
    sub = data.get("subscription") or {}
    endpoint = sub.get("endpoint")
    keys = sub.get("keys") or {}
    p256dh = keys.get("p256dh")
    auth = keys.get("auth")
    if not (endpoint and p256dh and auth):
        return jsonify({"error": "subscription 형식 오류",
                        "got": {"endpoint": bool(endpoint), "p256dh": bool(p256dh), "auth": bool(auth)}}), 400
    db = get_db()
    now = datetime.now(timezone.utc).isoformat()
    try:
        db.execute("""
            INSERT INTO push_subscriptions (user_id, endpoint, p256dh, auth, user_agent, created_at, last_used)
            VALUES (?,?,?,?,?,?,?)
            ON CONFLICT(endpoint) DO UPDATE SET
                user_id = excluded.user_id,
                p256dh = excluded.p256dh,
                auth = excluded.auth,
                last_used = excluded.last_used
        """, (me["id"], endpoint, p256dh, auth, request.headers.get("User-Agent", "")[:200], now, now))
        db.commit()
    except Exception as e:
        print(f"[push/subscribe] INSERT 실패 user={me['id']}: {e}")
        return jsonify({"error": f"DB INSERT 실패: {e}"}), 500
    # 이 기기(세션)의 push endpoint 기록 — PC 로그아웃 시 '이 PC 의 푸시만' 정확히 삭제하기 위함.
    try:
        session["push_endpoint"] = endpoint
    except Exception:
        pass
    # 저장 확인 — INSERT 직후 다시 조회해서 row 수 반환
    count = db.execute(
        "SELECT COUNT(*) FROM push_subscriptions WHERE user_id=?", (me["id"],)
    ).fetchone()[0]
    print(f"[push/subscribe] OK user={me['id']} count={count} ua={request.headers.get('User-Agent','')[:80]}")
    return jsonify({"ok": True, "subscription_count": count})


@app.route("/api/push/diag")
@login_required
def api_push_diag():
    """현재 사용자의 push 구독 상태 진단 — 다이얼로그에 표시."""
    me = current_user()
    db = get_db()
    rows = db.execute(
        """SELECT id, substr(endpoint, 1, 60) AS ep_prefix, user_agent,
                  datetime(created_at) AS created, datetime(last_used) AS last_used
             FROM push_subscriptions WHERE user_id=? ORDER BY id""",
        (me["id"],),
    ).fetchall()
    return jsonify({
        "user_id": me["id"],
        "display_name": me["display_name"],
        "pywebpush_ok": PYWEBPUSH_OK,
        "vapid_key_present": bool(vapid_public_key_b64u()),
        "subscription_count": len(rows),
        "subscriptions": [{
            "id": r["id"],
            "endpoint": r["ep_prefix"] + "...",
            "user_agent": (r["user_agent"] or "")[:80],
            "created": r["created"],
            "last_used": r["last_used"],
        } for r in rows],
    })


@app.route("/api/push/unsubscribe", methods=["POST"])
@login_required
def api_push_unsubscribe():
    me = current_user()
    data = request.get_json(silent=True) or {}
    endpoint = data.get("endpoint")
    if not endpoint:
        return jsonify({"error": "endpoint 필요"}), 400
    db = get_db()
    db.execute("DELETE FROM push_subscriptions WHERE user_id=? AND endpoint=?", (me["id"], endpoint))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/push/unsubscribe_all", methods=["POST"])
@login_required
def api_push_unsubscribe_all():
    """내 계정의 모든 기기 푸시 구독 삭제 — '완전 오프라인' 확실한 탈출구.
    로그아웃 정리가 실패해 '휴대폰' 상태로 끼어있을 때 즉시 정리용."""
    me = current_user()
    db = get_db()
    n = db.execute(
        "SELECT COUNT(*) FROM push_subscriptions WHERE user_id=?", (me["id"],)
    ).fetchone()[0]
    db.execute("DELETE FROM push_subscriptions WHERE user_id=?", (me["id"],))
    db.commit()
    return jsonify({"ok": True, "deleted": n})


@app.route("/api/push/test", methods=["POST"])
@login_required
def api_push_test():
    """테스트 푸시 발송 + 실패 시 상세 에러 반환 (진단용).
    본인→본인 직접 발송 (억제 로직 우회) — 푸시 인프라 자체 동작 확인용."""
    me = current_user()
    sent, errors, total = send_push_to_user(
        me["id"],
        "🔔 KNK 이음 테스트",
        "푸시 알림이 정상 작동합니다.",
        url="/chat", tag="test",
        collect_errors=True,
    )
    return jsonify({"sent": sent, "total_subscriptions": total, "errors": errors})


@app.route("/api/push/test_simulate", methods=["POST"])
@login_required
def api_push_test_simulate():
    """실제 메시지 수신 시나리오 시뮬레이션 — 똑똑한 억제 로직(_user_has_active_pc) 적용.
    '다른 사람이 보낸 메시지' 처럼 동작:
      · PC 가 실제 활성(focus+최근 heartbeat) → 모바일 푸시 스킵 (정상 동작 확인)
      · PC 비활성/자리비움/끔 → 모바일 푸시 발송 (백그라운드 알림 확인)
    """
    me = current_user()
    pc_active = _user_has_active_pc(me["id"])
    if pc_active:
        return jsonify({
            "pc_active": True,
            "skipped": True,
            "message": "PC가 활성 상태라 모바일 푸시를 스킵했습니다 (똑똑한 억제 정상 동작). "
                       "휴대폰만 두고 PC를 끄거나 1분 이상 자리를 비운 뒤 다시 시도하세요.",
        })
    sent, errors, total = send_push_to_user(
        me["id"],
        "💬 테스트발송 — 새 메시지",
        "실제 수신 시나리오 시뮬레이션입니다. 이 알림이 보이면 백그라운드 푸시 정상!",
        url="/chat", tag="test_sim",
        collect_errors=True,
    )
    return jsonify({"pc_active": False, "skipped": False, "sent": sent, "total_subscriptions": total, "errors": errors})


@app.route("/api/admin/cleanup", methods=["POST"])
@login_required
def api_admin_cleanup():
    """메시지 자동삭제 — N개월 이전 메시지(첨부 포함) 제거. 단 keep_forever=1 프로젝트·system 메시지 보존.

    수동 실행 또는 외부 스케줄러(Windows 작업스케줄러·cron)에서 호출.
    """
    me = current_user()
    if me["role"] != "ceo":
        abort(403)
    db = get_db()
    cutoff = (datetime.now(timezone.utc).date().replace(day=1)).isoformat()
    # cutoff = "오늘 - N개월" 의 첫날
    months = MESSAGE_RETENTION_MONTHS
    # 1단계: 전역 보존 정책 (N개월). keep_forever=1 프로젝트은 제외.
    # SQLite date() 산술
    rows = db.execute("""
        SELECT m.id, m.file_path
          FROM messages m
          JOIN rooms r ON r.id = m.room_id
          LEFT JOIN items it ON it.room_id = m.room_id
         WHERE m.kind != 'system'
           AND date(m.created_at) < date('now', ?)
           AND COALESCE(it.keep_forever, 0) = 0
    """, (f"-{months} months",)).fetchall()
    # 2단계: 방별 retention_days 가 설정되면 그 일수 이전 메시지도 추가 삭제.
    # WhatsApp 식 자동 삭제 — 방장이 보안·정리 목적으로 설정.
    per_room_rows = db.execute("""
        SELECT m.id, m.file_path
          FROM messages m
          JOIN rooms r ON r.id = m.room_id
         WHERE r.retention_days IS NOT NULL
           AND r.retention_days > 0
           AND m.kind != 'system'
           AND julianday('now') - julianday(m.created_at) > r.retention_days
    """).fetchall()
    # 중복 제거 (전역에서 이미 삭제될 ID 는 2단계에서 빼기)
    seen_ids = {r["id"] for r in rows}
    for pr in per_room_rows:
        if pr["id"] not in seen_ids:
            rows.append(pr)
            seen_ids.add(pr["id"])
    deleted = 0
    deleted_files = 0
    for r in rows:
        if r["file_path"]:
            try:
                fp = os.path.join(UPLOAD_DIR, r["file_path"])
                if os.path.exists(fp):
                    os.remove(fp)
                    deleted_files += 1
            except OSError:
                pass
        db.execute("DELETE FROM messages WHERE id = ?", (r["id"],))
        deleted += 1
    db.commit()
    return jsonify({"deleted_messages": deleted, "deleted_files": deleted_files, "retention_months": months})


@app.route("/api/admin/cleanup/preview")
@login_required
def api_admin_cleanup_preview():
    """삭제 미리보기 — 실제 삭제는 안 하고 카운트만 반환."""
    me = current_user()
    if me["role"] != "ceo":
        abort(403)
    db = get_db()
    months = MESSAGE_RETENTION_MONTHS
    row = db.execute("""
        SELECT COUNT(*) AS n,
               SUM(CASE WHEN m.file_path IS NOT NULL THEN 1 ELSE 0 END) AS files
          FROM messages m
          JOIN rooms r ON r.id = m.room_id
          LEFT JOIN items it ON it.room_id = m.room_id
         WHERE m.kind != 'system'
           AND date(m.created_at) < date('now', ?)
           AND COALESCE(it.keep_forever, 0) = 0
    """, (f"-{months} months",)).fetchone()
    return jsonify({
        "would_delete_messages": row["n"] or 0,
        "would_delete_files": row["files"] or 0,
        "retention_months": months,
        "ceo_only": True,
    })


# ============================================================
# 🧪 부하 테스트 전용 API (대표 지시 2026-05-20)
#   격리된 테스트 계정·방·메시지 일괄 생성/삭제. admin 전용.
#   계정 username 패턴: loadtest_001@knktest.local ~ loadtest_NNN@knktest.local
#   방 이름 prefix: [LOADTEST]
#   confirm_token 필수 (URL 또는 body): "I_UNDERSTAND_LOAD_TEST"
# ============================================================
LOAD_TEST_USERNAME_PREFIX = "loadtest_"
LOAD_TEST_USERNAME_DOMAIN = "@knktest.local"
LOAD_TEST_ROOM_NAME_PREFIX = "[LOADTEST]"
LOAD_TEST_CONFIRM_TOKEN = "I_UNDERSTAND_LOAD_TEST"


@app.route("/api/admin/load_test/setup", methods=["POST"])
@login_required
def api_load_test_setup():
    """N개 격리 테스트 계정 + 1개 격리 방 자동 생성.
    body: {count: 75, confirm_token: 'I_UNDERSTAND_LOAD_TEST'}
    응답: {users: [{username, password, id}, ...], room_id, room_name}"""
    me = current_user()
    if me["role"] != "ceo":
        abort(403)
    data = request.get_json(silent=True) or {}
    if data.get("confirm_token") != LOAD_TEST_CONFIRM_TOKEN:
        return jsonify({"error": f"confirm_token 필수: '{LOAD_TEST_CONFIRM_TOKEN}'"}), 400
    count = int(data.get("count", 75))
    if count < 1 or count > 200:
        return jsonify({"error": "count 는 1~200 사이"}), 400
    db = get_db()
    now = datetime.now(timezone.utc).isoformat()
    created_users = []
    for i in range(1, count + 1):
        username = f"{LOAD_TEST_USERNAME_PREFIX}{i:03d}{LOAD_TEST_USERNAME_DOMAIN}"
        password = f"{LOAD_TEST_USERNAME_PREFIX}{i:03d}_pwd"
        display_name = f"부하테스트{i:03d}"
        # 이미 존재하면 그 id 사용
        existing = db.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone()
        if existing:
            uid = existing["id"]
            # 비밀번호 갱신 (이전 테스트 잔존 대비)
            db.execute("UPDATE users SET password_hash=?, must_change_password=0, active=1 WHERE id=?",
                       (generate_password_hash(password), uid))
        else:
            cur = db.execute(
                "INSERT INTO users (username, password_hash, display_name, role, avatar_color, created_at, "
                " email, title, department, must_change_password, active) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (username, generate_password_hash(password), display_name, "staff", "#9CA3AF",
                 now, username, "부하테스트", "테스트", 0, 1)
            )
            uid = cur.lastrowid
        created_users.append({"id": uid, "username": username, "password": password})
    # 격리 방 생성 (이미 존재하면 재활용)
    room_name = f"{LOAD_TEST_ROOM_NAME_PREFIX} 부하 테스트 방"
    existing_room = db.execute("SELECT id FROM rooms WHERE name=?", (room_name,)).fetchone()
    if existing_room:
        room_id = existing_room["id"]
        # 기존 멤버 다 비우고 재구성
        db.execute("DELETE FROM room_members WHERE room_id=?", (room_id,))
    else:
        cur = db.execute(
            "INSERT INTO rooms (name, type, created_by, created_at, invite_policy) "
            "VALUES (?, 'group', ?, ?, 'all')",
            (room_name, me["id"], now)
        )
        room_id = cur.lastrowid
    # admin + 75개 테스트 계정 모두 멤버로
    db.execute("INSERT INTO room_members (room_id, user_id, role, joined_at) VALUES (?, ?, 'host', ?)",
               (room_id, me["id"], now))
    for u in created_users:
        db.execute("INSERT INTO room_members (room_id, user_id, role, joined_at) VALUES (?, ?, 'member', ?)",
                   (room_id, u["id"], now))
    db.commit()
    return jsonify({
        "users": created_users,
        "room_id": room_id,
        "room_name": room_name,
        "count": len(created_users),
    })


@app.route("/api/admin/load_test/cleanup", methods=["POST"])
@login_required
def api_load_test_cleanup():
    """loadtest_* 계정 + [LOADTEST] 방 + 관련 메시지·파일 일괄 삭제.
    body: {confirm_token: 'I_UNDERSTAND_LOAD_TEST'}"""
    me = current_user()
    if me["role"] != "ceo":
        abort(403)
    data = request.get_json(silent=True) or {}
    if data.get("confirm_token") != LOAD_TEST_CONFIRM_TOKEN:
        return jsonify({"error": f"confirm_token 필수: '{LOAD_TEST_CONFIRM_TOKEN}'"}), 400
    db = get_db()
    # 1) [LOADTEST] 방의 모든 메시지·파일 삭제
    rooms = db.execute("SELECT id FROM rooms WHERE name LIKE ?", (f"{LOAD_TEST_ROOM_NAME_PREFIX}%",)).fetchall()
    deleted_files = 0
    deleted_msgs = 0
    deleted_rooms = 0
    for r in rooms:
        rid = r["id"]
        # 파일 디스크 삭제
        files = db.execute("SELECT file_path FROM messages WHERE room_id=? AND file_path IS NOT NULL", (rid,)).fetchall()
        for f in files:
            try:
                fpath = os.path.join(UPLOAD_DIR, f["file_path"])
                if os.path.exists(fpath):
                    os.remove(fpath)
                    deleted_files += 1
            except Exception:
                pass
        # 방 디렉토리 자체 삭제 시도
        try:
            room_dir = os.path.join(UPLOAD_DIR, str(rid))
            if os.path.isdir(room_dir):
                import shutil as _sh
                _sh.rmtree(room_dir, ignore_errors=True)
        except Exception:
            pass
        # 메시지 + 멤버 + 방 삭제 (FK CASCADE 의존)
        c = db.execute("DELETE FROM messages WHERE room_id=?", (rid,))
        deleted_msgs += c.rowcount or 0
        db.execute("DELETE FROM room_members WHERE room_id=?", (rid,))
        db.execute("DELETE FROM requests WHERE room_id=?", (rid,))
        db.execute("DELETE FROM rooms WHERE id=?", (rid,))
        deleted_rooms += 1
    # 2) loadtest_* 계정 삭제 — 다른 곳에 있는 멤버십·메시지도 정리
    test_users = db.execute(
        "SELECT id FROM users WHERE username LIKE ?",
        (f"{LOAD_TEST_USERNAME_PREFIX}%{LOAD_TEST_USERNAME_DOMAIN}",)
    ).fetchall()
    deleted_users = 0
    for u in test_users:
        uid = u["id"]
        db.execute("DELETE FROM room_members WHERE user_id=?", (uid,))
        db.execute("DELETE FROM messages WHERE user_id=?", (uid,))
        db.execute("DELETE FROM users WHERE id=?", (uid,))
        deleted_users += 1
    db.commit()
    return jsonify({
        "deleted_users": deleted_users,
        "deleted_rooms": deleted_rooms,
        "deleted_messages": deleted_msgs,
        "deleted_files": deleted_files,
    })


@app.route("/api/admin/load_test/status")
@login_required
def api_load_test_status():
    me = current_user()
    if me["role"] != "ceo":
        abort(403)
    db = get_db()
    test_users = db.execute(
        "SELECT COUNT(*) AS n FROM users WHERE username LIKE ?",
        (f"{LOAD_TEST_USERNAME_PREFIX}%{LOAD_TEST_USERNAME_DOMAIN}",)
    ).fetchone()
    test_rooms = db.execute("SELECT id, name FROM rooms WHERE name LIKE ?",
                            (f"{LOAD_TEST_ROOM_NAME_PREFIX}%",)).fetchall()
    test_msgs = 0
    for r in test_rooms:
        c = db.execute("SELECT COUNT(*) AS n FROM messages WHERE room_id=?", (r["id"],)).fetchone()
        test_msgs += c["n"]
    return jsonify({
        "test_users_count": test_users["n"],
        "test_rooms": [{"id": r["id"], "name": r["name"]} for r in test_rooms],
        "test_messages_count": test_msgs,
    })


@app.route("/api/admin/backup_now", methods=["POST"])
@login_required
def api_admin_backup_now():
    """수동 백업 즉시 실행 — deploy/backup.sh 호출. admin 전용."""
    me = current_user()
    if me["role"] != "ceo":
        abort(403)
    try:
        import subprocess as _sp
        backup_script = os.path.join(APP_DIR, "deploy", "backup.sh")
        if not os.path.exists(backup_script):
            return jsonify({"error": f"backup.sh not found at {backup_script}"}), 500
        result = _sp.run(["bash", backup_script], capture_output=True, text=True, timeout=600)
        return jsonify({
            "exit_code": result.returncode,
            "stdout": result.stdout[-2000:],
            "stderr": result.stderr[-2000:],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/users", methods=["POST"])
@login_required
def api_users_create():
    """관리자만 — 직원 등록.
    표준 정책 (대표 지시 2026-05-26 전면 변경): 사번 = 로그인 ID.
      • mode='kor' (본사): username = employee_no (소문자), password = 휴대폰 숫자만
      • mode='vn'  (베트남): username = 'vn' + employee_no, password = '9999'
      • mode 미지정 시 옛 호환 (username·password 직접 지정) 허용
    body: {mode, display_name, employee_no, phone, email?, title?, department?,
           display_name_vn?, display_name_en?, role?, avatar_color?}"""
    me = current_user()
    if me["role"] != "ceo":
        abort(403)
    data = request.get_json(silent=True) or {}
    mode = (data.get("mode") or "").strip().lower()
    display_name = (data.get("display_name") or "").strip()
    display_name_vn = (data.get("display_name_vn") or "").strip()
    display_name_en = (data.get("display_name_en") or "").strip()
    email = (data.get("email") or "").strip().lower()
    phone = (data.get("phone") or "").strip()
    title = (data.get("title") or "").strip()[:40] or None
    department = (data.get("department") or "").strip()[:40] or None
    employee_no_raw = (data.get("employee_no") or "").strip()[:30]
    role = data.get("role") or "staff"
    avatar_color = data.get("avatar_color") or "#3b82f6"
    # 호환: username·password 직접 지정 가능 (특수 케이스)
    explicit_username = (data.get("username") or "").strip().lower()
    explicit_password = data.get("password")

    if not display_name:
        return jsonify({"error": "이름(display_name) 필수"}), 400

    # 본사 모드 — 사번=ID, 휴대폰=초기 비번
    if mode == "kor":
        if not employee_no_raw:
            return jsonify({"error": "사번 필수 (사번이 로그인 ID 입니다)"}), 400
        if not phone:
            return jsonify({"error": "휴대폰 번호 필수 (초기 비밀번호로 사용)"}), 400
        digits = "".join(ch for ch in phone if ch.isdigit())
        if len(digits) < 9:
            return jsonify({"error": "전화번호 자릿수 부족 (숫자 9자리 이상)"}), 400
        employee_no = employee_no_raw
        username = employee_no_raw.strip().lower()
        password = digits
        must_change = 1
    # 베트남 모드 — 'VN'+사번=ID, 초기 비번 9999
    elif mode == "vn":
        if not employee_no_raw:
            return jsonify({"error": "사번 필수 (사번이 로그인 ID 입니다)"}), 400
        if not display_name_en:
            return jsonify({"error": "이름(영문) 필수"}), 400
        # 사번에 이미 'VN' 접두가 붙어 있으면 제거 후 다시 부착 (이중 방지)
        raw = employee_no_raw.upper().strip()
        if raw.upper().startswith("VN"):
            raw = raw[2:].strip()
        if not raw:
            return jsonify({"error": "사번 숫자 부분 필수"}), 400
        employee_no = "VN" + raw
        username = ("vn" + raw).lower()
        password = "9999"
        must_change = 1
    # 옛 호환 모드 — username/password 직접 지정 (관리 스크립트 등)
    elif explicit_username:
        username = explicit_username
        password = explicit_password or "knk1234"
        must_change = 1 if data.get("must_change_password", True) else 0
        employee_no = employee_no_raw or None
    else:
        return jsonify({"error": "mode(kor/vn) 또는 username 지정 필요"}), 400

    # 영문 직급·부서 자동 매핑
    title_en = TITLE_TO_EN.get(title or "", "") or None
    department_en = DEPT_TO_EN.get(department or "", "") or None

    db = get_db()
    if db.execute("SELECT 1 FROM users WHERE username=?", (username,)).fetchone():
        return jsonify({"error": f"이미 존재하는 ID: {username}"}), 400
    if employee_no and db.execute("SELECT 1 FROM users WHERE employee_no=?", (employee_no,)).fetchone():
        return jsonify({"error": f"이미 존재하는 사번: {employee_no}"}), 400
    now = datetime.now(timezone.utc).isoformat()
    cur = db.execute(
        "INSERT INTO users (username, password_hash, display_name, role, avatar_color, "
        " created_at, email, phone, title, department, employee_no, must_change_password, "
        " display_name_vn, display_name_en, title_en, department_en) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (username, generate_password_hash(password), display_name, role, avatar_color,
         now, email or None, phone or None, title, department, employee_no, must_change,
         display_name_vn or None, display_name_en or None, title_en, department_en),
    )
    db.commit()
    # 신규 직원 → 자동채널(KNK WORLD + 본사/베트남) 자동 가입 (대표 지시 2026-05-20)
    try: _sync_user_auto_channels(db, cur.lastrowid)
    except Exception as e: print(f"[auto_channel] create sync 실패: {e}")
    # broadcast — 사용자 목록 즉시 갱신
    new_row = db.execute(
        "SELECT id, username, display_name, display_name_vn, display_name_en, role, avatar_color, title, title_en, department, department_en, email, phone, employee_no, active "
        "FROM users WHERE id=?", (cur.lastrowid,)
    ).fetchone()
    socketio.emit("user_info_changed", dict(new_row))
    if mode == "vn":
        pw_hint = "초기 비밀번호 9999 (베트남 공통)"
    elif mode == "kor":
        pw_hint = "전화번호 숫자만 (대시·공백 제외)"
    else:
        pw_hint = "관리자 지정 비밀번호"
    return jsonify({
        "id": cur.lastrowid,
        "username": username,
        "employee_no": employee_no,
        "display_name": display_name,
        "initial_password": password,
        "initial_password_hint": pw_hint,
    })


def _ensure_deleted_placeholder_user(db):
    """삭제된 사용자의 메시지·요청을 이전받을 플레이스홀더 사용자.
    username='_deleted_user', active=0 (로그인 불가). 1회만 생성."""
    row = db.execute("SELECT id FROM users WHERE username = '_deleted_user'").fetchone()
    if row:
        return row["id"]
    now = datetime.now(timezone.utc).isoformat()
    cur = db.execute(
        "INSERT INTO users (username, password_hash, display_name, role, avatar_color, "
        " created_at, active) VALUES (?,?,?,?,?,?,?)",
        ("_deleted_user", "", "(삭제된 사용자)", "staff", "#9CA3AF", now, 0),
    )
    return cur.lastrowid


@app.route("/api/users/<int:user_id>", methods=["DELETE"])
@login_required
def api_user_delete(user_id):
    """사용자 완전 삭제 (퇴사 등). 관리자(ceo) 만 가능.
    동작:
      1. 메시지·요청·방생성 기록을 '(삭제된 사용자)' 플레이스홀더로 이전 (이력 보존)
      2. 사용자 행 DELETE → CASCADE 로 room_members·push_subscriptions·reactions 자동 정리
    안전장치:
      - 본인 삭제 불가
      - 마지막 관리자 삭제 불가
      - 초기 대표 계정(id=1) 삭제 불가 (비활성화는 가능)"""
    me = current_user()
    if me["role"] != "ceo":
        return jsonify({"error": "관리자만 삭제 가능"}), 403
    if user_id == me["id"]:
        return jsonify({"error": "본인은 삭제할 수 없습니다. 다른 관리자에게 요청하세요."}), 400
    db = get_db()
    row = db.execute("SELECT id, display_name, role FROM users WHERE id=?", (user_id,)).fetchone()
    if not row:
        return jsonify({"error": "사용자 없음"}), 404
    # id=1 = 초기 대표 계정 — 절대 삭제 금지 (시드 보호)
    if user_id == 1:
        return jsonify({"error": "초기 대표 계정(id=1)은 삭제할 수 없습니다. 비활성화만 가능합니다."}), 400
    # 최고관리자(소유자) — 삭제 금지 (대표 지시 2026-05-21)
    _own = db.execute("SELECT username FROM users WHERE id=?", (user_id,)).fetchone()
    if _own and _is_owner(_own["username"]):
        return jsonify({"error": "최고관리자 계정은 삭제할 수 없습니다."}), 400
    # 플레이스홀더 사용자에게 본인을 삭제하려는 경우 — 차단
    if row["display_name"] == "(삭제된 사용자)":
        return jsonify({"error": "시스템 사용자는 삭제할 수 없습니다."}), 400
    # 마지막 관리자 삭제 차단
    if row["role"] == "ceo":
        ceo_count = db.execute("SELECT COUNT(*) AS n FROM users WHERE role='ceo' AND active=1").fetchone()["n"]
        if ceo_count <= 1:
            return jsonify({"error": "마지막 관리자는 삭제할 수 없습니다. 다른 관리자를 먼저 임명하세요."}), 400
    deleted_name = row["display_name"]
    # 플레이스홀더 사용자 확보
    placeholder_id = _ensure_deleted_placeholder_user(db)
    # 메시지·요청·방생성 등 NOT NULL FK 가 있는 것들 → 플레이스홀더로 이전
    db.execute("UPDATE messages SET user_id = ? WHERE user_id = ?", (placeholder_id, user_id))
    db.execute("UPDATE messages SET forwarded_from_user_id = ? WHERE forwarded_from_user_id = ?", (placeholder_id, user_id))
    db.execute("UPDATE rooms SET created_by = ? WHERE created_by = ?", (placeholder_id, user_id))
    try:
        db.execute("UPDATE requests SET requested_by = ? WHERE requested_by = ?", (placeholder_id, user_id))
        db.execute("UPDATE requests SET assigned_to = NULL WHERE assigned_to = ?", (user_id,))
    except Exception:
        pass
    try:
        db.execute("UPDATE items SET created_by = ? WHERE created_by = ?", (placeholder_id, user_id))
    except Exception:
        pass
    # 사용자 삭제 — CASCADE 로 room_members/push_subscriptions/reactions/acks/stars 등 자동 정리
    db.execute("DELETE FROM users WHERE id = ?", (user_id,))
    db.commit()
    # 실시간 broadcast — 사용자 목록 즉시 갱신
    socketio.emit("user_deleted", {"user_id": user_id, "display_name": deleted_name})
    return jsonify({"ok": True, "deleted_user_id": user_id, "display_name": deleted_name})


@app.route("/api/users/<int:user_id>/reset_password", methods=["POST"])
@login_required
def api_user_reset_password(user_id):
    """관리자가 사용자 비밀번호를 전화번호(숫자만)로 초기화 (대표 지시 2026-05-21, 규칙 5).
    초기화 후 must_change_password=1 → 사용자 첫 로그인 시 새 비밀번호 설정 강제."""
    me = current_user()
    if me["role"] != "ceo":
        return jsonify({"error": "관리자만 비밀번호를 초기화할 수 있습니다."}), 403
    db = get_db()
    row = db.execute("SELECT id, username, display_name, phone FROM users WHERE id=?", (user_id,)).fetchone()
    if not row:
        return jsonify({"error": "사용자 없음"}), 404
    # 최고관리자 비밀번호는 탈취 방지 — 본인(최고관리자)만 변경 가능
    if _is_owner(row["username"]) and not _is_owner(me["username"]):
        return jsonify({"error": "최고관리자 비밀번호는 본인만 변경할 수 있습니다."}), 403
    digits = "".join(ch for ch in (row["phone"] or "") if ch.isdigit())
    if not digits:
        return jsonify({"error": "이 사용자의 전화번호가 없어 초기화할 수 없습니다. 먼저 전화번호를 등록하세요."}), 400
    db.execute("UPDATE users SET password_hash=?, must_change_password=1, "
               "password_version=COALESCE(password_version,1)+1 WHERE id=?",
               (generate_password_hash(digits), user_id))
    db.commit()
    return jsonify({"ok": True, "temp_password": digits, "display_name": row["display_name"]})


@app.route("/api/admin/user_diag/<int:user_id>")
@login_required
def api_admin_user_diag(user_id):
    """관리자 전용 — 사용자의 실시간 접속·세션·알림 등록 현황 (읽기 전용 진단).
    상태표시('📱 휴대폰/⚫ 오프라인' 등)가 왜 그렇게 나오는지 확인용.
    민감정보(세션 토큰·푸시 endpoint·암호키)는 포함하지 않음. (대표 지시 2026-05-25)"""
    me = current_user()
    if me["role"] != "ceo":
        return jsonify({"error": "관리자만 사용할 수 있습니다."}), 403
    db = get_db()
    u = db.execute("SELECT id, username, display_name FROM users WHERE id=?",
                   (user_id,)).fetchone()
    if not u:
        return jsonify({"error": "사용자 없음"}), 404
    saved = _get_user_status(user_id)   # status·custom_text 는 user_statuses 테이블에 있음
    online = _user_is_online(user_id)
    has_pc = _user_has_pc_connection(user_id)
    active_dev, idle = _active_presence(user_id)
    computed = _computed_user_status(user_id)
    sessions = [
        {"device_type": r["device_type"], "ip": r["ip"],
         "user_agent": (r["user_agent"] or "")[:70], "created_at": r["created_at"]}
        for r in db.execute(
            "SELECT device_type, ip, user_agent, created_at FROM active_sessions "
            "WHERE user_id=? ORDER BY created_at DESC", (user_id,)).fetchall()
    ]
    pushes = [
        {"user_agent": (r["user_agent"] or "")[:70], "created_at": r["created_at"], "last_used": r["last_used"]}
        for r in db.execute(
            "SELECT user_agent, created_at, last_used FROM push_subscriptions "
            "WHERE user_id=? ORDER BY created_at DESC", (user_id,)).fetchall()
    ]
    return jsonify({
        "user": {"id": u["id"], "username": u["username"], "display_name": u["display_name"],
                 "saved_status": saved.get("status"), "custom_text": saved.get("custom_text")},
        "live": {
            "online": bool(online),
            "has_pc_connection": bool(has_pc),
            "active_device": active_dev,
            "idle": bool(idle),
            "computed_status": computed.get("status"),
            "computed_label": computed.get("label"),
            "at_office": bool(computed.get("at_office")),
        },
        "active_sessions": sessions,
        "push_subscriptions": pushes,
    })


@app.route("/api/me/password", methods=["PUT"])
@login_required
def api_me_password():
    """본인 비밀번호 변경. body: {current_password, new_password}
    must_change_password=1 인 첫 로그인 사용자도 이 엔드포인트로 변경."""
    me = current_user()
    data = request.get_json(silent=True) or {}
    cur_pw = data.get("current_password") or ""
    new_pw = data.get("new_password") or ""
    if not new_pw or len(new_pw) < 6:
        return jsonify({"error": "새 비밀번호는 6자 이상"}), 400
    if new_pw == cur_pw:
        return jsonify({"error": "현재 비밀번호와 동일 — 변경 의미 없음"}), 400
    db = get_db()
    row = db.execute("SELECT password_hash FROM users WHERE id=?", (me["id"],)).fetchone()
    if not row or not check_password_hash(row["password_hash"], cur_pw):
        return jsonify({"error": "현재 비밀번호가 올바르지 않습니다"}), 400
    db.execute(
        "UPDATE users SET password_hash=?, must_change_password=0, "
        "password_version=COALESCE(password_version,1)+1 WHERE id=?",
        (generate_password_hash(new_pw), me["id"]),
    )
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/me/must_change_password", methods=["GET"])
@login_required
def api_me_must_change():
    """현재 사용자가 비밀번호 변경 강제 상태인지 조회 (UI 다이얼로그 트리거용)."""
    me = current_user()
    db = get_db()
    row = db.execute("SELECT must_change_password FROM users WHERE id=?", (me["id"],)).fetchone()
    return jsonify({"must_change": bool(row and row["must_change_password"])})


# ═══════════════════════════════════════════════════════════════════════════
# 사번 SSO — 메신저 = 사내 단일 인증 서버(Identity Provider). JWT RS256.
#   발주: _TO_메신저세션/2026-05-29_사번SSO.md  ·  착수: 2026-05-31
#   엔드포인트는 BASE_PATH 자동 적용 → 실제 경로 /msg/api/sso/*
# ═══════════════════════════════════════════════════════════════════════════
import threading as _threading

try:
    import jwt as _pyjwt                     # PyJWT[crypto] — RS256 발급·검증
except Exception:                            # 미설치 시 SSO 엔드포인트만 503, 앱은 정상 부팅
    _pyjwt = None

SSO_KEYS_DIR = os.environ.get("KNK_MSG_KEYS_DIR", os.path.join(APP_DIR, "keys"))
SSO_PRIVATE_KEY_PATH = os.path.join(SSO_KEYS_DIR, "jwt_rs256_private.pem")
SSO_PUBLIC_KEY_PATH = os.path.join(SSO_KEYS_DIR, "jwt_rs256_public.pem")
SSO_ISSUER = os.environ.get("KNK_MSG_SSO_ISSUER", "https://haist.knknara.co.kr/msg/")
SSO_AUDIENCES = [a.strip() for a in os.environ.get(
    "KNK_MSG_SSO_AUDIENCES", "haist-works,knk-internal").split(",") if a.strip()]
SSO_TOKEN_TTL = int(os.environ.get("KNK_MSG_SSO_TTL", "3600"))   # access_token 1시간
SSO_RL_LIMIT = int(os.environ.get("KNK_MSG_SSO_RL_LIMIT", "10")) # /token 분당 허용 횟수
SSO_RL_WINDOW = 60
# 직원 명부(Directory) API 서버-서버 공유키 — WORKS(NAS)→메신저 내부 호출용. (발주 2026-05-31)
#   양쪽 OS 환경변수 KNK_SSO_SERVICE_KEY 에 동일 비밀값을 넣어야 작동. 코드·git 에 절대 넣지 않음.
#   미설정(빈값)이면 /api/sso/directory 엔드포인트는 503 으로 닫혀 절대 열리지 않음.
SSO_SERVICE_KEY = os.environ.get("KNK_SSO_SERVICE_KEY", "").strip()

_sso_key_cache = {"priv": None, "pub": None}
_sso_rl_lock = _threading.Lock()
_sso_rl = {}   # ip -> [최근 요청 timestamp ...]


def _sso_available():
    return (_pyjwt is not None
            and os.path.exists(SSO_PRIVATE_KEY_PATH)
            and os.path.exists(SSO_PUBLIC_KEY_PATH))


def _sso_load_private_key():
    if _sso_key_cache["priv"] is None:
        with open(SSO_PRIVATE_KEY_PATH, "r", encoding="utf-8") as f:
            _sso_key_cache["priv"] = f.read()
    return _sso_key_cache["priv"]


def _sso_load_public_key():
    if _sso_key_cache["pub"] is None:
        with open(SSO_PUBLIC_KEY_PATH, "r", encoding="utf-8") as f:
            _sso_key_cache["pub"] = f.read()
    return _sso_key_cache["pub"]


def _rg(row, k, default=None):
    """sqlite3.Row 안전 getter — 컬럼 없으면 default."""
    try:
        return row[k] if k in row.keys() else default
    except Exception:
        return default


def _sso_user_claims(row):
    """users row → JWT 사용자 정보 claim (실시간 DB 값)."""
    return {
        "name_kr": _rg(row, "display_name"),
        "name_en": _rg(row, "display_name_en"),
        "name_vi": _rg(row, "display_name_vn"),
        "dept": _rg(row, "department"),
        "position": _rg(row, "title"),
        "entity": _rg(row, "entity"),
        "email": _rg(row, "email"),
        "is_admin": (str(_rg(row, "role") or "") == "ceo"),
        # HAIST WORKS 진입 허용 여부 — WORKS 는 이 값이 true 일 때만 입장 허용. (대표 지시 2026-05-31)
        "works_access": bool(_rg(row, "works_access", 0)),
    }


def _sso_issue_token(row):
    """users row → 서명된 RS256 JWT 문자열."""
    import time as _t
    now = int(_t.time())
    emp = _rg(row, "employee_no") or _rg(row, "username")
    payload = {
        "iss": SSO_ISSUER,
        "sub": str(emp),
        "aud": SSO_AUDIENCES,
        "iat": now,
        "exp": now + SSO_TOKEN_TTL,
        "jti": str(uuid.uuid4()),
        "uid": _rg(row, "id"),
        "pwv": int(_rg(row, "password_version", 1) or 1),  # 비번 변경 추적
    }
    payload.update(_sso_user_claims(row))
    return _pyjwt.encode(payload, _sso_load_private_key(), algorithm="RS256")


def _sso_verify_token(token, verify_aud=True):
    """JWT 검증 + pwv(비번버전) 일치 확인. (payload, error) 반환."""
    if _pyjwt is None:
        return None, "pyjwt_missing"
    try:
        decoded = _pyjwt.decode(
            token, _sso_load_public_key(), algorithms=["RS256"],
            audience=(SSO_AUDIENCES if verify_aud else None),
            issuer=SSO_ISSUER,
            options={"verify_aud": verify_aud},
        )
    except Exception as e:
        return None, f"invalid_token:{type(e).__name__}"
    sub = decoded.get("sub")
    db = get_db()
    row = db.execute(
        "SELECT password_version, active FROM users "
        "WHERE employee_no=? OR LOWER(username)=LOWER(?)", (sub, sub)
    ).fetchone()
    if not row:
        return None, "user_not_found"
    if not _rg(row, "active", 1):
        return None, "user_inactive"
    cur_pwv = int(_rg(row, "password_version", 1) or 1)
    if int(decoded.get("pwv", -1)) != cur_pwv:
        return None, "password_changed"   # 비번 변경 후 발급된 토큰 → SP 가 재로그인 유도
    return decoded, None


def _sso_err(code, desc, status):
    resp = jsonify({"error": code, "error_description": desc})
    resp.status_code = status
    return resp


def _sso_cors(resp):
    """사내 도메인(*.knknara.co.kr)만 CORS 허용."""
    origin = request.headers.get("Origin", "")
    host = ""
    if origin:
        try:
            from urllib.parse import urlparse
            host = (urlparse(origin).hostname or "")
        except Exception:
            host = ""
    if host == "knknara.co.kr" or host.endswith(".knknara.co.kr"):
        resp.headers["Access-Control-Allow-Origin"] = origin
        resp.headers["Vary"] = "Origin"
        resp.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
        resp.headers["Access-Control-Allow-Headers"] = "Authorization, Content-Type"
        resp.headers["Access-Control-Max-Age"] = "600"
    return resp


def _sso_safe_redirect_uri(uri):
    """오픈 리다이렉트 방지 — 사내 도메인(*.knknara.co.kr) https 만 허용.
    dev 환경은 http + localhost 허용. 통과 시 uri 그대로, 아니면 None."""
    if not uri:
        return None
    try:
        from urllib.parse import urlparse
        p = urlparse(uri)
    except Exception:
        return None
    host = (p.hostname or "")
    scheme_ok = (p.scheme == "https") or (not IS_PRODUCTION and p.scheme == "http")
    if not scheme_ok:
        return None
    if host == "knknara.co.kr" or host.endswith(".knknara.co.kr"):
        return uri
    if (not IS_PRODUCTION) and host in ("localhost", "127.0.0.1"):
        return uri
    return None


def _sso_require_https():
    """운영 환경에서 HTTPS 외 거부. dev 는 통과. (거부 응답 or None)"""
    if not IS_PRODUCTION:
        return None
    if request.is_secure or request.headers.get("X-Forwarded-Proto", "") == "https":
        return None
    return _sso_cors(_sso_err("https_required", "HTTPS 필수", 403))


def _sso_rate_limited(ip):
    import time as _t
    now = _t.time()
    with _sso_rl_lock:
        lst = [t for t in _sso_rl.get(ip, []) if now - t < SSO_RL_WINDOW]
        if len(lst) >= SSO_RL_LIMIT:
            _sso_rl[ip] = lst
            return True
        lst.append(now)
        _sso_rl[ip] = lst
        return False


def _sso_log(event, **kw):
    parts = " ".join(f"{k}={v}" for k, v in kw.items())
    try:
        print(f"[SSO] {event} ip={request.remote_addr} {parts}", flush=True)
    except Exception:
        pass


@app.route("/api/sso/token", methods=["POST", "OPTIONS"])
def api_sso_token():
    """사번/비번 → JWT 발급. 인증 없음(자격증명 자체가 인증)."""
    if request.method == "OPTIONS":
        return _sso_cors(jsonify({"ok": True}))
    deny = _sso_require_https()
    if deny is not None:
        return deny
    if not _sso_available():
        return _sso_cors(_sso_err("sso_unavailable", "SSO 키 미설치 — 관리자 문의", 503))
    ip = request.remote_addr or "?"
    if _sso_rate_limited(ip):
        _sso_log("token_ratelimited")
        return _sso_cors(_sso_err("rate_limited", "요청이 많습니다. 잠시 후 다시 시도", 429))
    data = request.get_json(silent=True) or {}
    emp = str(data.get("employee_no") or data.get("username") or "").strip()
    pw = data.get("password") or ""
    if not emp or not pw:
        return _sso_cors(_sso_err("invalid_request", "employee_no 와 password 가 필요합니다", 400))
    db = get_db()
    row = db.execute(
        "SELECT * FROM users WHERE (employee_no=? OR LOWER(username)=LOWER(?)) "
        "AND COALESCE(active,1)=1 AND COALESCE(is_guest,0)=0",
        (emp, emp)
    ).fetchone()
    if not row or not check_password_hash(row["password_hash"], pw):
        _sso_log("token_fail", emp=emp)
        return _sso_cors(_sso_err("invalid_credentials", "사번 또는 비밀번호 불일치", 401))
    token = _sso_issue_token(row)
    _sso_log("token_ok", emp=emp, uid=_rg(row, "id"))
    return _sso_cors(jsonify({
        "access_token": token,
        "token_type": "Bearer",
        "expires_in": SSO_TOKEN_TTL,
        "scope": "openid profile employee",
        "must_change_password": bool(_rg(row, "must_change_password", 0)),
    }))


@app.route("/api/sso/userinfo", methods=["GET", "OPTIONS"])
def api_sso_userinfo():
    """JWT 검증 + 사용자 정보(실시간 DB 최신값) 반환. 헤더: Authorization: Bearer <JWT>."""
    if request.method == "OPTIONS":
        return _sso_cors(jsonify({"ok": True}))
    deny = _sso_require_https()
    if deny is not None:
        return deny
    if not _sso_available():
        return _sso_cors(_sso_err("sso_unavailable", "SSO 키 미설치", 503))
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        return _sso_cors(_sso_err("invalid_request", "Authorization: Bearer 토큰 필요", 401))
    payload, err = _sso_verify_token(auth[7:].strip())
    if err:
        _sso_log("userinfo_fail", reason=err)
        return _sso_cors(_sso_err("invalid_token", err, 401))
    db = get_db()
    row = db.execute(
        "SELECT * FROM users WHERE employee_no=? OR LOWER(username)=LOWER(?)",
        (payload.get("sub"), payload.get("sub"))
    ).fetchone()
    if not row:
        return _sso_cors(_sso_err("user_not_found", "사용자 없음", 404))
    info = {
        "sub": str(_rg(row, "employee_no") or _rg(row, "username")),
        "employee_no": _rg(row, "employee_no"),
        "uid": _rg(row, "id"),
    }
    info.update(_sso_user_claims(row))
    return _sso_cors(jsonify(info))


@app.route("/api/sso/public-key", methods=["GET", "OPTIONS"])
def api_sso_public_key():
    """RS256 public key — Service Provider 가 토큰 검증에 사용. PEM 기본, ?format=jwk 지원."""
    if request.method == "OPTIONS":
        return _sso_cors(jsonify({"ok": True}))
    if not os.path.exists(SSO_PUBLIC_KEY_PATH):
        return _sso_cors(_sso_err("sso_unavailable", "public key 미설치", 503))
    pem = _sso_load_public_key()
    if request.args.get("format") == "jwk" and _pyjwt is not None:
        try:
            from cryptography.hazmat.primitives.serialization import load_pem_public_key
            pub = load_pem_public_key(pem.encode("utf-8"))
            jwk = _pyjwt.algorithms.RSAAlgorithm.to_jwk(pub)   # JSON 문자열
            return _sso_cors(app.response_class(jwk, mimetype="application/json"))
        except Exception as e:
            return _sso_cors(_sso_err("jwk_error", str(e), 500))
    return _sso_cors(app.response_class(pem, mimetype="application/x-pem-file"))


@app.route("/api/sso/directory", methods=["GET"])
def api_sso_directory():
    """직원 명부(Directory) — WORKS 전 직원 동기화용. **서버-서버 전용**(브라우저용 아님).
    인증: 헤더 X-SSO-Service-Key == env KNK_SSO_SERVICE_KEY (상수시간 비교).
      · env 미설정 → 503(절대 안 열림), 키 누락/불일치 → 403.
    응답(200, JSON):
      {"users":[{employee_no,name_kr,name_en,name_vi,dept,position,entity,
                 email,phone,is_admin,works_access,active}, ...],
       "count":N, "generated_at":ISO8601, "source":"knk-messenger"}
    범위: 게스트(고객사) 제외한 전 직원(재직·퇴직 모두, active 플래그로 구분).
    개인정보(메일·연락처) 포함 → 내부망 + 서비스키로만 접근. (발주 2026-05-31)
    """
    import hmac as _hmac
    # 1) 서비스키 미설정이면 기능 자체가 닫힘 — 빈 키로 우회 불가
    if not SSO_SERVICE_KEY:
        _sso_log("directory_unconfigured")
        return _sso_err("service_unconfigured", "directory 서비스키(KNK_SSO_SERVICE_KEY) 미설정", 503)
    # 2) 헤더 키 상수시간 비교
    provided = (request.headers.get("X-SSO-Service-Key", "") or "").strip()
    if not provided or not _hmac.compare_digest(provided, SSO_SERVICE_KEY):
        _sso_log("directory_forbidden")
        return _sso_err("forbidden", "유효한 X-SSO-Service-Key 필요", 403)
    db = get_db()
    rows = db.execute(
        """SELECT id, username, employee_no, display_name, display_name_en, display_name_vn,
                  department, title, entity, email, phone, role,
                  COALESCE(works_access, 0) AS works_access,
                  COALESCE(active, 1) AS active
             FROM users
            WHERE COALESCE(is_guest, 0) = 0
            ORDER BY (employee_no IS NULL) ASC, id ASC"""
    ).fetchall()
    users = []
    for r in rows:
        users.append({
            # 사번(필수) — 없으면 username 으로 대체(사번=로그인ID 정책상 대부분 존재)
            "employee_no": str(_rg(r, "employee_no") or _rg(r, "username") or ""),
            "name_kr": _rg(r, "display_name"),
            "name_en": _rg(r, "display_name_en"),
            "name_vi": _rg(r, "display_name_vn"),
            "dept": _rg(r, "department"),
            "position": _rg(r, "title"),
            "entity": _rg(r, "entity"),
            "email": _rg(r, "email"),
            "phone": _rg(r, "phone"),
            "is_admin": (str(_rg(r, "role") or "") == "ceo"),
            "works_access": bool(_rg(r, "works_access", 0)),
            "active": bool(_rg(r, "active", 1)),
        })
    _sso_log("directory_ok", count=len(users))
    return jsonify({
        "users": users,
        "count": len(users),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "knk-messenger",
    })


@app.route("/api/sso/revoke", methods=["POST", "OPTIONS"])
def api_sso_revoke():
    """토큰 무효화 — 해당 사용자 password_version +1 → 발급된 모든 토큰 거부(전체 로그아웃)."""
    if request.method == "OPTIONS":
        return _sso_cors(jsonify({"ok": True}))
    deny = _sso_require_https()
    if deny is not None:
        return deny
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        return _sso_cors(_sso_err("invalid_request", "Authorization: Bearer 토큰 필요", 401))
    payload, err = _sso_verify_token(auth[7:].strip())
    if err:
        return _sso_cors(_sso_err("invalid_token", err, 401))
    db = get_db()
    db.execute(
        "UPDATE users SET password_version=COALESCE(password_version,1)+1 "
        "WHERE employee_no=? OR LOWER(username)=LOWER(?)",
        (payload.get("sub"), payload.get("sub"))
    )
    db.commit()
    _sso_log("revoke_ok", sub=payload.get("sub"))
    return _sso_cors(jsonify({"ok": True, "revoked": True}))


# ═══ 비번 위임 (HAIST WORKS 연동) — 서버↔서버 전용 API 3종 (대표 지시 2026-05-31) ═══
#   WORKS 는 자체 로그인창(사번+비번)을 유지하되 비번 검증을 메신저에 위임. 메신저=비밀번호 단일 저장소.
#   공유키 헤더 X-SSO-Service-Key 로만 인증(브라우저 노출 금지) · 내부망(localhost) 호출 · HTTPS/CORS 불요.
SSO_SERVICE_KEY = (os.environ.get("KNK_SSO_SERVICE_KEY") or "").strip()
SSO_PW_RL_LIMIT = int(os.environ.get("KNK_MSG_SSO_PW_RL_LIMIT", "5"))   # 사번당 분당 실패 허용
SSO_PW_RL_WINDOW = 60
_sso_pw_rl_lock = _threading.Lock()
_sso_pw_fail = {}   # employee_no -> [실패 timestamp ...]


def _sso_service_key_ok():
    """서버↔서버 공유키 검증 — 상수시간 비교. 키 미설정(서버)·헤더 누락/불일치 → False (fail-closed)."""
    if not SSO_SERVICE_KEY:
        return False
    import hmac as _hmac
    provided = (request.headers.get("X-SSO-Service-Key") or "").strip()
    return bool(provided) and _hmac.compare_digest(provided, SSO_SERVICE_KEY)


def _sso_pw_fail_count(emp):
    import time as _t
    now = _t.time()
    with _sso_pw_rl_lock:
        lst = [t for t in _sso_pw_fail.get(emp, []) if now - t < SSO_PW_RL_WINDOW]
        _sso_pw_fail[emp] = lst
        return len(lst)


def _sso_pw_fail_record(emp):
    import time as _t
    with _sso_pw_rl_lock:
        _sso_pw_fail.setdefault(emp, []).append(_t.time())


def _sso_pw_fail_clear(emp):
    with _sso_pw_rl_lock:
        _sso_pw_fail.pop(emp, None)


def _sso_is_vn_user(row):
    """VN(베트남 법인) 직원 판정 — entity > 사번/아이디(_classify_user_mode) > 부서 순. 초기비번 규칙 분기용."""
    ent = (_rg(row, "entity") or "").strip().upper()
    if ent == "VN":
        return True
    if ent == "KOR":
        return False
    mode = _classify_user_mode(_rg(row, "username") or "")
    if mode == "vn":
        return True
    if mode == "kor":
        return False
    return _user_is_vietnam(_rg(row, "department") or "")


@app.route("/api/sso/verify-password", methods=["POST"])
def api_sso_verify_password():
    """[서버 전용] 사번+비번 검증. 성공 {valid:true, pwv, user{...}} / 실패 {valid:false} (정보유출 방지 위해 항상 200).
    공유키(X-SSO-Service-Key) 필수 · 사번당 무차별 대입 방지(분당 N회 실패)."""
    if not _sso_service_key_ok():
        _sso_log("verify_pw_forbidden")
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json(silent=True) or {}
    emp = str(data.get("employee_no") or "").strip()
    pw = data.get("password") or ""
    if not emp or not pw:
        return jsonify({"valid": False}), 200
    if _sso_pw_fail_count(emp) >= SSO_PW_RL_LIMIT:
        _sso_log("verify_pw_ratelimited", emp=emp)
        return jsonify({"valid": False, "error": "rate_limited"}), 429
    db = get_db()
    row = db.execute(
        "SELECT * FROM users WHERE (employee_no=? OR LOWER(username)=LOWER(?)) "
        "AND COALESCE(active,1)=1 AND COALESCE(is_guest,0)=0",
        (emp, emp)
    ).fetchone()
    if not row or not check_password_hash(row["password_hash"], pw):
        _sso_pw_fail_record(emp)
        _sso_log("verify_pw_fail", emp=emp)
        return jsonify({"valid": False}), 200
    _sso_pw_fail_clear(emp)
    user = {"employee_no": _rg(row, "employee_no") or _rg(row, "username")}
    user.update(_sso_user_claims(row))
    _sso_log("verify_pw_ok", emp=emp, uid=_rg(row, "id"))
    return jsonify({
        "valid": True,
        "pwv": int(_rg(row, "password_version", 1) or 1),
        "must_change_password": bool(_rg(row, "must_change_password", 0)),
        "user": user,
    }), 200


@app.route("/api/sso/admin-reset-password", methods=["POST"])
def api_sso_admin_reset_password():
    """[서버 전용] 관리자 초기화 — 본사=휴대폰 숫자만 / VN='9999' 로 리셋 + must_change + pwv+1(→ 양쪽 즉시 로그아웃).
    공유키 필수. (요청자 admin/ceo 여부는 WORKS 가 사전 확인)"""
    if not _sso_service_key_ok():
        _sso_log("admin_reset_forbidden")
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json(silent=True) or {}
    emp = str(data.get("employee_no") or "").strip()
    reset_by = str(data.get("reset_by") or "").strip()
    if not emp:
        return jsonify({"ok": False, "reason": "invalid_request"}), 200
    db = get_db()
    row = db.execute(
        "SELECT id, username, employee_no, phone, entity, department "
        "FROM users WHERE (employee_no=? OR LOWER(username)=LOWER(?)) AND COALESCE(is_guest,0)=0",
        (emp, emp)
    ).fetchone()
    if not row:
        return jsonify({"ok": False, "reason": "not_found"}), 200
    if _is_owner(row["username"]):
        # 최고관리자(소유자) 비번은 원격 초기화 금지 — 본인만 변경
        _sso_log("admin_reset_blocked_owner", emp=emp, by=reset_by)
        return jsonify({"ok": False, "reason": "owner_protected"}), 200
    if _sso_is_vn_user(row):
        temp_pw, hint = "9999", "9999 (베트남 법인 공통 초기비번)"
    else:
        temp_pw = "".join(ch for ch in (row["phone"] or "") if ch.isdigit())
        if not temp_pw:
            _sso_log("admin_reset_no_phone", emp=emp, by=reset_by)
            return jsonify({"ok": False, "reason": "no_phone"}), 200
        hint = "휴대폰 번호 (숫자만)"
    db.execute(
        "UPDATE users SET password_hash=?, must_change_password=1, "
        "password_version=COALESCE(password_version,1)+1 WHERE id=?",
        (generate_password_hash(temp_pw), row["id"])
    )
    db.commit()
    new_pwv = int((db.execute(
        "SELECT password_version FROM users WHERE id=?", (row["id"],)
    ).fetchone()["password_version"]) or 1)
    _sso_log("admin_reset_ok", emp=emp, by=reset_by, new_pwv=new_pwv)
    return jsonify({"ok": True, "new_pwv": new_pwv, "temp_pw_hint": hint}), 200


@app.route("/api/sso/pwv", methods=["GET"])
def api_sso_pwv():
    """[서버 전용] 사번으로 현재 password_version 조회(토큰 없이). WORKS 세션이 5분 주기로 비번 변경 감지용."""
    if not _sso_service_key_ok():
        return jsonify({"error": "forbidden"}), 403
    emp = str(request.args.get("employee_no") or "").strip()
    if not emp:
        return jsonify({"employee_no": emp, "pwv": None}), 200
    db = get_db()
    row = db.execute(
        "SELECT password_version FROM users "
        "WHERE (employee_no=? OR LOWER(username)=LOWER(?)) AND COALESCE(is_guest,0)=0",
        (emp, emp)
    ).fetchone()
    if not row:
        return jsonify({"employee_no": emp, "pwv": None}), 200
    return jsonify({"employee_no": emp, "pwv": int(_rg(row, "password_version", 1) or 1)}), 200


# ─── 한국어 직급·부서 → 영문 자동 매핑 (대표 지시 2026-05-26) ───
# 일괄 등록 시 사용자가 영문을 직접 입력하지 않아도 서버가 자동으로 채워줌
TITLE_TO_EN = {
    # 본사 (KNK 자체 직급)
    "대표이사":      "CEO",
    "전무이사":      "Senior Executive",
    "상무이사":      "Managing Director",
    "이사":          "Director",
    "이사(팀장)":    "Director / Team Lead",
    "매니저":        "Manager",
    "매니저(팀장)":  "Manager / Team Lead",
    "프로":          "Professional",
    "프로(팀장)":    "Professional / Team Lead",
    # 베트남법인
    "법인장":        "General Director",
    "부장":          "General Manager",
    "차장":          "Deputy General Manager",
    "과장":          "Manager",
    "대리":          "Assistant Manager",
    "주임":          "Senior Staff",
    # 공통
    "사원":          "Staff",
}

DEPT_TO_EN = {
    # 본사
    "01_KOR/00_총괄":        "General Management",
    "01_KOR/01_기술영업팀":  "Technical Sales Team",
    "01_KOR/02_검사기팀":    "Inspection Equipment Team",
    "01_KOR/03_품질팀":      "Quality Team",
    "01_KOR/04_설계팀":      "Design Team",
    "01_KOR/05_소프트웨어팀":"Software Team",
    "01_KOR/06_전장설계팀":  "Electrical Design Team",
    "01_KOR/07_제조기술1팀": "Manufacturing Engineering Team 1",
    "01_KOR/08_제조기술2팀": "Manufacturing Engineering Team 2",
    "01_KOR/09_가공팀":      "Machining Team",
    "01_KOR/10_구매팀":      "Purchasing Team",
    "01_KOR/11_관리팀":      "Management Team",
    "01_KOR/12_개발혁신팀":  "Development & Innovation Team",
    "01_KOR/21_라이프밸류팀":"Life Value Team",
    # 베트남법인
    "02_VN/00_총괄":         "VN General Management",
    "02_VN/01_기술팀":       "VN Technical Team",
    "02_VN/02_설계팀":       "VN Design Team",
    "02_VN/03_소프트웨어팀": "VN Software Team",
    "02_VN/04_가공팀":       "VN Machining Team",
    "02_VN/05_조립팀":       "VN Assembly Team",
    "02_VN/06_전장팀":       "VN Electrical Team",
    "02_VN/07_품질팀":       "VN Quality Team",
    "02_VN/08_관리팀":       "VN Management Team",
    "02_VN/09_구매팀":       "VN Purchasing Team",
}


def _make_sample_avatar_image(size=512, label_main="512×512", label_sub="권장 크기 견본 (Sample)"):
    """양식에 박아둘 견본 사진을 동적 생성 (대표 지시 2026-05-26).
    Pillow로 회색 배경 + 점선 원 + 가운데 텍스트를 그려 JPEG BytesIO 로 반환.

    한글 폰트가 시스템에 없을 수 있어 NotoSans/맑은고딕 등을 순서대로 시도,
    실패하면 영문 기본 폰트로 폴백.
    """
    from PIL import Image, ImageDraw, ImageFont
    import io as _io
    img = Image.new("RGB", (size, size), (243, 244, 246))   # 옅은 회색 #F3F4F6
    d = ImageDraw.Draw(img)
    # 테두리 (두꺼운 회색)
    d.rectangle([(0, 0), (size - 1, size - 1)], outline=(156, 163, 175), width=6)
    # 가운데 원 (얼굴 위치 안내)
    cx, cy, cr = size // 2, size // 2, int(size * 0.36)
    for i, w in enumerate(range(0, 360, 12)):
        a0 = w * 3.14159 / 180
        # 점선 원 — 작은 호 12개씩 끊어 그리기
        d.arc([cx - cr, cy - cr, cx + cr, cy + cr], start=w, end=w + 7, fill=(156, 163, 175), width=3)

    # 폰트 — 한글 지원 폰트 시도, 실패 시 영문 기본
    font_big = None
    font_small = None
    for fp in [
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "C:/Windows/Fonts/malgun.ttf",
        "C:/Windows/Fonts/malgunbd.ttf",
        "/System/Library/Fonts/AppleSDGothicNeo.ttc",
    ]:
        try:
            font_big = ImageFont.truetype(fp, size=size // 8)
            font_small = ImageFont.truetype(fp, size=size // 18)
            break
        except Exception:
            continue
    if font_big is None:
        # 영문 기본 — 한글 깨질 수 있어 label_sub 는 영문으로 교체
        font_big = ImageFont.load_default()
        font_small = ImageFont.load_default()
        label_sub = "Sample / Recommended size"

    # 텍스트 가운데 정렬
    def _text_size(text, font):
        try:
            bx = d.textbbox((0, 0), text, font=font)
            return bx[2] - bx[0], bx[3] - bx[1]
        except Exception:
            return font.getsize(text) if hasattr(font, "getsize") else (len(text) * 8, 12)
    w1, h1 = _text_size(label_main, font_big)
    w2, h2 = _text_size(label_sub, font_small)
    total_h = h1 + 10 + h2
    y_start = (size - total_h) // 2
    d.text(((size - w1) // 2, y_start), label_main, fill=(31, 41, 55), font=font_big)
    d.text(((size - w2) // 2, y_start + h1 + 10), label_sub, fill=(75, 85, 99), font=font_small)

    buf = _io.BytesIO()
    img.save(buf, "JPEG", quality=88, optimize=True)
    buf.seek(0)
    return buf


@app.route("/api/users/bulk/template")
@login_required
def api_users_bulk_template():
    """직원 일괄 등록용 엑셀 양식 다운로드. 관리자만."""
    me = current_user()
    if me["role"] != "ceo":
        abort(403)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.drawing.image import Image as XLImage
    except ImportError:
        return jsonify({"error": "openpyxl 미설치"}), 500
    import io

    wb = Workbook()
    # ─── 헤더·스타일 공통 정의 ───
    # 본사: 7컬럼 (베트남어 이름·영문 직급·영문 부서 제거 — 대표 지시 2026-05-26)
    # 베트남: 8컬럼 (영문 직급·영문 부서만 제거)
    # 영문 직급·부서는 서버가 매핑 테이블로 자동 채움
    # 대표 지시 2026-05-26: 사번 = 로그인 ID (전면 변경)
    headers_kor = [
        "이름 * (한글)",
        "회사 이메일",
        "휴대폰 (= 초기 비밀번호)",
        "사번 * (= 로그인 ID)",
        "직급", "부서 *",
        "이름(영문)",
    ]
    headers_vn = [
        "이름 * (한국식 발음 한글)",
        "회사 이메일 (선택 — 사용하는 직원만)",
        "휴대폰 (선택)",
        "사번 * (= 로그인 ID, 'VN' 자동 부착)",
        "직급", "부서 *",
        "이름(베트남어) — 필수",
        "이름(영문) — 필수",
    ]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill_kor = PatternFill(start_color="A5282C", end_color="A5282C", fill_type="solid")   # 본사 — 빨강
    header_fill_vn  = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")   # 베트남 — 초록(테마 구분)
    border_thin = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    stripe_fill = PatternFill(start_color="FBFBFB", end_color="FBFBFB", fill_type="solid")
    # 컬럼 너비 — 본사는 앞 7개, 베트남은 앞 8개 사용
    widths_full = [14, 36, 22, 12, 16, 22, 22, 22]   # 이름·이메일·휴대폰·사번·직급·부서·베트남어·영문

    # ─── 부서 목록 (시트별 분리) ───
    _dept_list_kor = [
        "01_KOR/00_총괄","01_KOR/01_기술영업팀","01_KOR/02_검사기팀","01_KOR/03_품질팀",
        "01_KOR/04_설계팀","01_KOR/05_소프트웨어팀","01_KOR/06_전장설계팀",
        "01_KOR/07_제조기술1팀","01_KOR/08_제조기술2팀","01_KOR/09_가공팀",
        "01_KOR/10_구매팀","01_KOR/11_관리팀","01_KOR/12_개발혁신팀","01_KOR/21_라이프밸류팀",
    ]
    _dept_list_vn = [
        "02_VN/00_총괄","02_VN/01_기술팀","02_VN/02_설계팀","02_VN/03_소프트웨어팀",
        "02_VN/04_가공팀","02_VN/05_조립팀","02_VN/06_전장팀","02_VN/07_품질팀",
        "02_VN/08_관리팀","02_VN/09_구매팀",
    ]
    ws_dept_kor = wb.create_sheet("_부서_KOR")
    for i, d in enumerate(_dept_list_kor, start=1):
        ws_dept_kor.cell(row=i, column=1, value=d)
    ws_dept_kor.sheet_state = "hidden"
    ws_dept_vn = wb.create_sheet("_부서_VN")
    for i, d in enumerate(_dept_list_vn, start=1):
        ws_dept_vn.cell(row=i, column=1, value=d)
    ws_dept_vn.sheet_state = "hidden"

    # ─── 직급 목록 (시트별 분리) ───
    _title_list_kor = [
        "대표이사", "전무이사", "상무이사",
        "이사", "이사(팀장)",
        "매니저", "매니저(팀장)",
        "프로", "프로(팀장)",
        "사원",
    ]
    _title_list_vn = [
        "법인장", "부장", "차장", "과장", "대리", "주임", "사원",
    ]
    ws_title_kor = wb.create_sheet("_직급_KOR")
    for i, t in enumerate(_title_list_kor, start=1):
        ws_title_kor.cell(row=i, column=1, value=t)
    ws_title_kor.sheet_state = "hidden"
    ws_title_vn = wb.create_sheet("_직급_VN")
    for i, t in enumerate(_title_list_vn, start=1):
        ws_title_vn.cell(row=i, column=1, value=t)
    ws_title_vn.sheet_state = "hidden"

    # ─── 헬퍼: 시트 1개 셋업 (헤더 + 줄무늬 + 드롭다운) ───
    def _setup_data_sheet(ws_target, headers_list, hdr_fill, dept_sheet_name, dept_count, title_sheet_name, title_count):
        ncols = len(headers_list)   # 본사 7 / 베트남 8
        ws_target.append(headers_list)
        for col_idx in range(1, ncols + 1):
            cell = ws_target.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_thin
        ws_target.row_dimensions[1].height = 42
        for i in range(1, ncols + 1):
            ws_target.column_dimensions[chr(64 + i)].width = widths_full[i - 1]
        for _ in range(100):
            ws_target.append([""] * ncols)
        for r in range(2, 102):
            is_stripe = (r % 2 == 0)
            for c in range(1, ncols + 1):
                cell = ws_target.cell(row=r, column=c)
                cell.border = border_thin
                if is_stripe:
                    cell.fill = stripe_fill
        # 부서 드롭다운 (F열)
        dv_d = DataValidation(type="list", formula1=f"={dept_sheet_name}!$A$1:$A${dept_count}", allow_blank=True)
        dv_d.error = "드롭다운(▼)에서 선택하세요"
        dv_d.errorTitle = "부서 형식 오류"
        ws_target.add_data_validation(dv_d)
        dv_d.add("F2:F101")
        # 직급 드롭다운 (E열)
        dv_t = DataValidation(type="list", formula1=f"={title_sheet_name}!$A$1:$A${title_count}", allow_blank=True)
        dv_t.error = "드롭다운(▼)에서 선택하세요"
        dv_t.errorTitle = "직급 형식 오류"
        ws_target.add_data_validation(dv_t)
        dv_t.add("E2:E101")

    # ─── 시트 1: 본사 (KOR) ───
    ws = wb.active
    ws.title = "본사(KOR)"
    _setup_data_sheet(ws, headers_kor, header_fill_kor, "_부서_KOR", len(_dept_list_kor), "_직급_KOR", len(_title_list_kor))

    # ─── 시트 2: 베트남법인 (VN) ───
    ws_vn = wb.create_sheet("베트남법인(VN)")
    _setup_data_sheet(ws_vn, headers_vn, header_fill_vn, "_부서_VN", len(_dept_list_vn), "_직급_VN", len(_title_list_vn))

    # ─── 시트 2: 안내 ───
    ws2 = wb.create_sheet("안내")
    notes = [
        ["KNK 이음 — 직원 일괄 등록 안내", ""],
        ["", ""],
        ["1. 필수 필드 (별표 * 표시)", ""],
        ["", "본사(KOR): 이름 · 사번 · 부서 — 3개 필수 (사번 = 로그인 ID)"],
        ["", "베트남(VN): 이름 · 사번 · 부서 · 이름(영문) — 4개 필수 (사번 = 로그인 ID, 'VN' 자동 부착)"],
        ["", "직급은 비워둬도 됨 (나중에 본인 또는 관리자가 수정 가능)"],
        ["", "권한도 별도 — 등록 후 관리자가 메신저에서 부여"],
        ["", ""],
        ["2. 로그인 ID 와 초기 비밀번호 ★ 사번 = 로그인 ID (대표 지시 2026-05-26 전면 변경)", ""],
        ["", "[본사 KOR]"],
        ["", "  · 로그인 ID = 사번 (예: '5', '67', '123')"],
        ["", "  · 초기 비밀번호 = 휴대폰 숫자만 (있을 때) / 없으면 '9999'"],
        ["", "  · 회사 이메일 = 별개 컬럼 — 기록·표시용 (로그인엔 사용 안 함)"],
        ["", "[베트남 VN]"],
        ["", "  · 로그인 ID = 'VN' + 사번 (자동 부착, 예: 'VN1', 'VN23')"],
        ["", "       양식의 사번 칸에 '1' 입력 → DB에 'VN1' 저장 / 로그인 시 'VN1' 입력"],
        ["", "  · 초기 비밀번호 = 9999"],
        ["", "  · 회사 이메일 = 선택 (사용하는 직원만 입력) / 안 쓰면 공란"],
        ["", "[공통]"],
        ["", "  · 사번 = 로그인 ID 라 사번 중복 불가 (본사 5번 + 베트남 VN5번은 다른 ID이므로 충돌 X)"],
        ["", "  · 같은 사번 재업로드 시 정보만 갱신(upsert) — 새 직원 안 만들어짐"],
        ["", ""],
        ["2-1. 베트남 직원 이름 3종 (대표 지시 2026-05-26) ★", ""],
        ["", "베트남 직원은 이름을 3가지 형식으로 모두 등록해 주세요:"],
        ["", "  ① A열  '이름'         = 한국식 발음 한글 표기   예: 응우옌 반 아"],
        ["", "  ② H열  '이름(베트남어)' = 베트남어 원어            예: Nguyễn Văn A"],
        ["", "  ③ I열  '이름(영문)'    = 영문                    예: Nguyen Van A"],
        ["", ""],
        ["", "화면 표시 (한국어 모드):"],
        ["", "  '{베트남어} ({한국식 발음 한글})' 형태로 병기 표시"],
        ["", "  예: 'Nguyễn Văn A (응우옌 반 아)'  — 한국 직원이 발음·기억 쉽게"],
        ["", ""],
        ["", "본사 직원은 H열(베트남어)는 비워두면 됩니다."],
        ["", ""],
        ["3. 첫 로그인 동작 (본사·베트남 공통)", ""],
        ["", "직원이 처음 로그인하면 비밀번호 변경 다이얼로그가 강제로 노출됩니다."],
        ["", "본인만 아는 비밀번호로 변경한 후에야 메신저 사용 가능합니다."],
        ["", ""],
        ["4. 권한 옵션", ""],
        ["", "'관리자' — 직원 등록·다른 사람 정보 수정·계정 비활성화·삭제 가능"],
        ["", "'일반' (또는 빈칸) — 본인 정보만 수정. 평소 사용자"],
        ["", ""],
        ["5. 이메일/ID 중복 시 동작 (자동 갱신 = upsert)", ""],
        ["", "이미 등록된 ID는 정보가 자동 갱신(이름·직급·부서·영문 등)됩니다."],
        ["", "  · 비밀번호·권한은 보존 (실수로 변경 방지)"],
        ["", "  · 빈 칸은 기존값 유지 (부분 시트로 인한 실수 삭제 방지)"],
        ["", "새 ID는 신규 등록 — 결과 화면에 '신규 N건 · 수정 M건' 표시"],
        ["", ""],
        ["6. 양식 작성 팁", ""],
        ["", "데이터 시트가 둘로 나뉘어 있습니다:"],
        ["", "  · 「본사(KOR)」 시트 — 7컬럼 (빨강 헤더): 이름·이메일·휴대폰·사번·직급·부서·이름영문"],
        ["", "  · 「베트남법인(VN)」 시트 — 8컬럼 (초록 헤더): 위 + 이름(베트남어)"],
        ["", "각 시트의 부서·직급 드롭다운은 해당 법인 것만 표시됩니다."],
        ["", "★ 영문 직급·영문 부서는 서버가 자동으로 채워줍니다 — 엑셀에 직접 입력 불필요"],
        ["", "   (안내 시트 7·8번의 매핑 표를 기준으로 자동 변환)"],
        ["", "예시는 별도 「예시」 시트에 있습니다 — 한 번 보고 본사·베트남 시트에 직접 입력하세요."],
        ["", "100명 이상 등록 필요 시 여러 번 나눠 업로드 — 두 시트에 채운 다음 한 번에 업로드 OK"],
        ["", ""],
        ["7. 정식 부서 목록 (한국어 → 영문)", "부서 칸에 아래 형식 그대로 입력 (법인 접두어 포함)"],
        ["[🇰🇷 본사 — 01_KOR (14개)]", ""],
        ["  01_KOR/00_총괄",                 "General Management"],
        ["  01_KOR/01_기술영업팀",           "Technical Sales Team"],
        ["  01_KOR/02_검사기팀",             "Inspection Equipment Team"],
        ["  01_KOR/03_품질팀",               "Quality Team"],
        ["  01_KOR/04_설계팀",               "Design Team"],
        ["  01_KOR/05_소프트웨어팀",         "Software Team"],
        ["  01_KOR/06_전장설계팀",           "Electrical Design Team"],
        ["  01_KOR/07_제조기술1팀",          "Manufacturing Engineering Team 1"],
        ["  01_KOR/08_제조기술2팀",          "Manufacturing Engineering Team 2"],
        ["  01_KOR/09_가공팀",               "Machining Team"],
        ["  01_KOR/10_구매팀",               "Purchasing Team"],
        ["  01_KOR/11_관리팀",               "Management Team"],
        ["  01_KOR/12_개발혁신팀",           "Development & Innovation Team"],
        ["  01_KOR/21_라이프밸류팀",         "Life Value Team"],
        ["[🇻🇳 베트남법인 — 02_VN (10개)]", ""],
        ["  02_VN/00_총괄",                  "VN General Management"],
        ["  02_VN/01_기술팀",                "VN Technical Team"],
        ["  02_VN/02_설계팀",                "VN Design Team"],
        ["  02_VN/03_소프트웨어팀",          "VN Software Team"],
        ["  02_VN/04_가공팀",                "VN Machining Team"],
        ["  02_VN/05_조립팀",                "VN Assembly Team"],
        ["  02_VN/06_전장팀",                "VN Electrical Team"],
        ["  02_VN/07_품질팀",                "VN Quality Team"],
        ["  02_VN/08_관리팀",                "VN Management Team"],
        ["  02_VN/09_구매팀",                "VN Purchasing Team"],
        ["", ""],
        ["※ 부서(영문) 칸에는 위 영문명을 그대로 입력하세요.", ""],
        ["", ""],
        ["8. 직급 한↔영 표기 (KNK 표준)", "직급 칸은 드롭다운(▼)에서 선택, 영문 직급은 아래 표 참고"],
        ["[🇰🇷 본사 — KNK 자체 직급]", ""],
        ["  대표이사",          "CEO"],
        ["  전무이사",          "Senior Executive"],
        ["  상무이사",          "Managing Director"],
        ["  이사",              "Director"],
        ["  이사(팀장)",        "Director / Team Lead"],
        ["  매니저",            "Manager"],
        ["  매니저(팀장)",      "Manager / Team Lead"],
        ["  프로",              "Professional"],
        ["  프로(팀장)",        "Professional / Team Lead"],
        ["[🇻🇳 베트남법인 — 일반 직급]", ""],
        ["  법인장",            "General Director"],
        ["  부장",              "General Manager"],
        ["  차장",              "Deputy General Manager"],
        ["  과장",              "Manager"],
        ["  대리",              "Assistant Manager"],
        ["  주임",              "Senior Staff"],
        ["[공통]", ""],
        ["  사원",              "Staff"],
        ["", ""],
        ["※ 팀장은 별도 직급이 아니라 부가 역할 — 이사·매니저·프로 어느 직급에서도 팀장이 될 수 있습니다.", ""],
        ["", ""],
        ["9. 사진 등록 안내", ""],
        ["", "사진은 일괄 등록 양식에 포함되지 않습니다 (대표 지시 2026-05-26)."],
        ["", "사진은 메신저에서 직원 카드 → '사진 등록' 버튼으로 1명씩 등록합니다."],
        ["", "드래그·확대/축소로 위치를 직접 조정한 뒤 저장."],
    ]
    for r, (a, b) in enumerate(notes, start=1):
        ws2.cell(row=r, column=1, value=a)
        ws2.cell(row=r, column=2, value=b)
    ws2.column_dimensions['A'].width = 32
    ws2.column_dimensions['B'].width = 84

    # 견본 사진 섹션은 제거됨 (대표 지시 2026-05-26) — 사진은 메신저에서 1명씩 등록

    # ─── 시트 3: 예시 (참고용, 등록 안 됨) — 대표 지시 2026-05-26 ───
    # 데이터 시트(직원등록)에는 예시를 두지 않고 별도 시트로 분리 → 사용자가 예시 삭제할 필요 없음
    ws3 = wb.create_sheet("예시")
    ws3.append(["이름", "회사 이메일", "휴대폰", "사번", "직급", "부서",
                "이름(베트남어)", "이름(영문)"])
    examples = [
        ["홍길동",        "hong@knknara.co.kr",     "010-1234-5678", "1",   "사원",          "01_KOR/01_기술영업팀",  "",              "Hong Gil-dong"],
        ["이순신",        "lee@knknara.co.kr",      "010-2345-6789", "2",   "매니저",        "01_KOR/05_소프트웨어팀", "",              "Lee Sun-sin"],
        ["김영업",        "kim.sales@knknara.co.kr","010-3456-7890", "3",   "이사(팀장)",    "01_KOR/01_기술영업팀",  "",              "Kim Sales"],
        ["응우옌 반 아",  "",                        "",               "1",   "과장",          "02_VN/01_기술팀",      "Nguyễn Văn A",  "Nguyen Van A"],
    ]
    for r in examples:
        ws3.append(r)
    # 헤더 스타일 (예시 시트도 동일 톤 — 본사 색상 사용)
    for col_idx in range(1, 9):
        cell = ws3.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill_kor
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin
    ws3.row_dimensions[1].height = 36
    # 예시 데이터 행 — 회색·기울임으로 '예시' 강조
    example_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    for r in range(2, 6):
        for c in range(1, 9):
            cell = ws3.cell(row=r, column=c)
            cell.fill = example_fill
            cell.font = Font(italic=True, color="6B7280")
            cell.border = border_thin
    # 너비
    widths3 = [14, 28, 18, 10, 14, 22, 18, 20]
    for i, w in enumerate(widths3, start=1):
        ws3.column_dimensions[chr(64 + i)].width = w
    # 상단 안내 (행 7~10)
    ws3.cell(row=7, column=1, value="◆ 이 시트는 참고용입니다 — 데이터는 「본사(KOR)」 또는 「베트남법인(VN)」 시트에 입력하세요.").font = Font(bold=True, color="A5282C", size=11)
    ws3.cell(row=8, column=1, value="  · 영문 직급·영문 부서는 자동으로 채워집니다 — 직접 입력 불필요 (안내 시트의 매핑 표 참고)").font = Font(color="334155", size=10)
    ws3.cell(row=9, column=1, value="  · 본사: 이름·이메일·휴대폰 모두 필수 / 베트남: 이메일·휴대폰은 선택(비워두면 공란), 이름 3종(한국식·베트남어·영문)은 필수").font = Font(color="334155", size=10)
    ws3.cell(row=10, column=1, value="  · 권한·사진은 별도 — 등록 후 관리자가 메신저에서 부여(권한) / 카드 → '사진 등록'(사진)").font = Font(color="334155", size=10)
    ws3.merge_cells(start_row=7, start_column=1, end_row=7, end_column=8)
    ws3.merge_cells(start_row=8, start_column=1, end_row=8, end_column=8)
    ws3.merge_cells(start_row=9, start_column=1, end_row=9, end_column=8)
    ws3.merge_cells(start_row=10, start_column=1, end_row=10, end_column=8)

    # ─── 안내 시트 KNK 톤 마감 ───
    # 1행 제목: 큰 글씨 + KNK 빨강 배경 + 흰 글씨
    title_cell = ws2.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="A5282C", end_color="A5282C", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.cell(row=1, column=2).fill = PatternFill(start_color="A5282C", end_color="A5282C", fill_type="solid")
    ws2.row_dimensions[1].height = 28

    # 섹션 제목(번호로 시작하는 행) 강조
    import re as _re_h
    section_fill = PatternFill(start_color="FDECEC", end_color="FDECEC", fill_type="solid")
    section_font = Font(bold=True, color="A5282C", size=11)
    bracket_font = Font(bold=True, color="334155", size=11)
    for r in range(2, len(notes) + 1):
        a_val = ws2.cell(row=r, column=1).value or ""
        if _re_h.match(r"^\d+\.\s", a_val):
            ws2.cell(row=r, column=1).font = section_font
            ws2.cell(row=r, column=1).fill = section_fill
            ws2.cell(row=r, column=2).fill = section_fill
            ws2.cell(row=r, column=2).font = section_font
        elif a_val.startswith("[") and a_val.endswith("]"):
            ws2.cell(row=r, column=1).font = bracket_font
        # 본문 들여쓰기 행은 회색 톤
        b_val = ws2.cell(row=r, column=2).value or ""
        if not a_val and b_val:
            ws2.cell(row=r, column=2).font = Font(color="334155", size=10)
        # 부서 매핑 행은 등폭 느낌 (왼쪽 정렬)
        if a_val.startswith("  01_KOR/") or a_val.startswith("  02_VN/"):
            ws2.cell(row=r, column=1).font = Font(name="Consolas", size=10, color="0F172A")
            ws2.cell(row=r, column=2).font = Font(size=10, color="334155")

    # 다운로드
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file
    return send_file(
        buf,
        as_attachment=True,
        download_name="KNK_직원등록_양식.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _extract_images_by_row(ws):
    """엑셀 워크시트에서 셀에 삽입된 사진을 추출 → {row_idx(1-based): (bytes, ext)} 반환.
    대표 지시 2026-05-26 — 일괄 등록 시 사진까지 함께 등록하기 위한 헬퍼.

    openpyxl의 _images 비공개 속성을 사용한다 (안정적). 각 이미지의 anchor 위치(행)로 매칭.
    사용자가 어떤 셀에 두든 그 셀의 행을 기준으로 사용자에게 매칭됨 (보통 K열에 둘 것을 권장).
    한 행에 사진이 여러 개면 마지막 것을 사용.
    """
    result = {}
    try:
        imgs = getattr(ws, "_images", []) or []
    except Exception:
        imgs = []
    if not imgs:
        return result
    for img in imgs:
        try:
            anchor = getattr(img, "anchor", None)
            row_idx = None
            if anchor is not None:
                _from = getattr(anchor, "_from", None) or getattr(anchor, "from_", None)
                if _from is not None and hasattr(_from, "row"):
                    row_idx = int(_from.row) + 1   # 0-based → 1-based
            if not row_idx or row_idx < 2:
                continue   # 헤더 행 또는 좌상단 외부
            # 이미지 바이너리 추출 — openpyxl 버전에 따라 _data() 메서드 또는 ref 사용
            data = None
            try:
                d = img._data() if callable(getattr(img, "_data", None)) else None
                if d: data = d
            except Exception: pass
            if not data:
                # fallback: img.ref 가 BytesIO 일 수도 있음
                ref = getattr(img, "ref", None)
                if ref is not None:
                    try:
                        data = ref.getvalue() if hasattr(ref, "getvalue") else (ref.read() if hasattr(ref, "read") else None)
                    except Exception: pass
            if not data or not isinstance(data, (bytes, bytearray)):
                continue
            # 포맷 식별 (PNG/JPEG/GIF/WEBP 시그니처)
            if data[:8] == b"\x89PNG\r\n\x1a\n":
                ext = "png"
            elif data[:3] == b"\xff\xd8\xff":
                ext = "jpg"
            elif data[:6] in (b"GIF87a", b"GIF89a"):
                ext = "gif"
            elif data[:4] == b"RIFF" and data[8:12] == b"WEBP":
                ext = "webp"
            else:
                # 기타 — 일단 PNG 로 저장 시도
                ext = "png"
            if ext not in AVATAR_ALLOWED_EXT:
                continue   # 허용되지 않은 형식은 스킵
            result[row_idx] = (bytes(data), ext)
        except Exception as e:
            print(f"[bulk] 이미지 추출 실패 (row={row_idx}): {e}")
            continue
    return result


def _save_avatar_for_user(uid, data, ext):
    """사용자 한 명의 사진을 저장. Pillow 로 512×512 center-crop JPEG 로 변환.
    실패 시 원본 그대로 저장 (확장자 유지). 대표 지시 2026-05-26.
    """
    try:
        from PIL import Image
        import io as _io
        img = Image.open(_io.BytesIO(data))
        img = img.convert("RGB")   # RGBA·팔레트 등 → RGB (JPEG 호환)
        # cover-crop (center) — 짧은 변을 기준으로 정사각형으로 자름
        w, h = img.size
        side = min(w, h)
        left = (w - side) // 2
        top = (h - side) // 2
        img = img.crop((left, top, left + side, top + side))
        if side > 512:
            img = img.resize((512, 512), Image.LANCZOS)
        out_path = os.path.join(AVATAR_DIR, f"{uid}.jpg")
        # 기존 다른 확장자 파일은 정리
        for old_ext in AVATAR_ALLOWED_EXT:
            old = os.path.join(AVATAR_DIR, f"{uid}.{old_ext}")
            if os.path.exists(old) and old != out_path:
                try: os.remove(old)
                except Exception: pass
        img.save(out_path, "JPEG", quality=85, optimize=True)
        return "jpg"
    except Exception as e:
        # Pillow 실패 — 원본 바이너리를 그대로 저장 (확장자 유지)
        print(f"[bulk] Pillow 변환 실패 uid={uid}: {e} — 원본 저장")
        try:
            out_path = os.path.join(AVATAR_DIR, f"{uid}.{ext}")
            for old_ext in AVATAR_ALLOWED_EXT:
                old = os.path.join(AVATAR_DIR, f"{uid}.{old_ext}")
                if os.path.exists(old) and old != out_path:
                    try: os.remove(old)
                    except Exception: pass
            with open(out_path, "wb") as fp:
                fp.write(data)
            return ext
        except Exception as e2:
            print(f"[bulk] 사진 저장 실패 uid={uid}: {e2}")
            return None


@app.route("/api/users/bulk", methods=["POST"])
@login_required
def api_users_bulk():
    """엑셀 일괄 등록. 관리자만. multipart/form-data 의 'file' 필드에 .xlsx.
    결과: {created: [...], skipped: [{row, name, reason}], errors: [{row, name, error}]}"""
    me = current_user()
    if me["role"] != "ceo":
        return jsonify({"error": "관리자만 일괄 등록 가능"}), 403
    if "file" not in request.files:
        return jsonify({"error": "file 필드에 엑셀 첨부 필요"}), 400
    f = request.files["file"]
    if not f.filename or not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "엑셀 파일(.xlsx) 만 업로드 가능"}), 400
    try:
        from openpyxl import load_workbook
    except ImportError:
        return jsonify({"error": "openpyxl 미설치"}), 500
    try:
        wb = load_workbook(f, data_only=True)
    except Exception as e:
        return jsonify({"error": f"엑셀 파일 열기 실패: {e}"}), 400

    # 데이터 시트 자동 식별 (대표 지시 2026-05-26 시트 분리):
    #   '본사' 또는 '베트남'/'VN'/'KOR' 이 시트 이름에 포함되면 데이터 시트
    #   '_'로 시작하면 hidden 참조 시트 (스킵)
    #   '안내', '예시'는 스킵
    data_sheets = []
    for sheet_name in wb.sheetnames:
        sn_lower = sheet_name.lower()
        if sheet_name.startswith("_"):
            continue
        if sheet_name in ("안내", "예시"):
            continue
        # 시트 이름에 본사/베트남/VN/KOR 포함 → 데이터 시트로 인식
        if ("본사" in sheet_name or "베트남" in sheet_name or
            "kor" in sn_lower or "vn" in sn_lower or
            sheet_name == "직원등록"):   # 옛 단일 시트 호환
            data_sheets.append(wb[sheet_name])
    if not data_sheets:
        # 안전망 — 인식 못하면 첫 시트라도 처리
        data_sheets = [wb.active]

    images_by_row = {}   # 사진 일괄 등록은 제거됨 (대표 지시 2026-05-26)

    # 예시 행 자동 스킵
    EXAMPLE_EMAILS = {"hong@knknara.co.kr", "lee@knknara.co.kr", "kim.sales@knknara.co.kr"}
    EXAMPLE_NAMES = {"홍길동", "이순신", "김영업", "응우옌 반 아", "Nguyễn Văn A"}

    # 베트남법인 자동 ID 도메인 + 초기 비번 (대표 지시 2026-05-26)
    #   본사(KOR)  : 회사 이메일 그대로 사용 → 기존 흐름
    #   베트남(VN) : 이메일 칸 비어 있으면 영문이름 슬러그로 자동 생성 → '{slug}@knkvn.local'
    #              초기 비밀번호 = '9999' (첫 로그인 시 변경 강제)
    VN_DOMAIN = "knkvn.local"
    VN_DEFAULT_PW = "9999"

    def _slug_for_vn(en_name: str) -> str:
        """영문이름 → ID 슬러그. 영숫자만 남기고 소문자화. 빈 결과는 ''."""
        if not en_name:
            return ""
        return "".join(ch for ch in en_name if ch.isalnum()).lower()

    def _unique_vn_username(slug: str) -> str:
        """베트남 자동 ID — 영문이름 슬러그를 그대로 username 으로 사용 (대표 지시 2026-05-26).
        옛 정책: '{slug}@knkvn.local' → 새 정책: '{slug}' (도메인 제거, 더 단순)
        같은 슬러그면 기존 username 재사용 (upsert). 동명이인은 영문이름에 직접 숫자 추가."""
        return slug

    created = []
    updated = []
    skipped = []
    errors = []
    db = get_db()
    now = datetime.now(timezone.utc).isoformat()

    # 데이터 시트 순회 (본사·베트남 분리, 또는 옛 단일 '직원등록' 시트 호환)
    for ws in data_sheets:
     # 시트 종류 판단 — 본사(7컬럼) / 베트남(8컬럼) / 옛 단일 시트(10~11컬럼)
     sheet_name_upper = (ws.title or "").upper()
     is_vn_sheet = ("베트남" in (ws.title or "")) or ("VN" in sheet_name_upper)
     # 헤더 한 줄 스킵 (row=1). row=2 부터 데이터.
     for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue  # 빈 행 스킵
        # 7컬럼: 이름·이메일·전화·사번·직급·부서·권한
        name = (str(row[0]) if row[0] is not None else "").strip()
        email = (str(row[1]) if row[1] is not None else "").strip().lower()
        phone = (str(row[2]) if row[2] is not None else "").strip()
        employee_no = (str(row[3]) if len(row) > 3 and row[3] is not None else "").strip()
        title = (str(row[4]) if len(row) > 4 and row[4] is not None else "").strip()
        department = (str(row[5]) if len(row) > 5 and row[5] is not None else "").strip()
        # 부서 코드 정규화 (대표 지시 2026-05-20 갱신)
        #  · '01 기술영업팀'    → '기술영업팀'           (본사 — code prefix 제거)
        #  · '12-VN01 기술팀'  → '12-VN01 기술팀'      (베트남법인 — DB에도 코드 포함하여 통째로 저장)
        #  · 'VN12-01 기술팀'  → '12-VN01 기술팀'      (legacy 포맷 자동 변환)
        import re as _re_dept
        # 새 체계 (2026-05-26 대표 지시): '01_KOR/NN_부서' / '02_VN/NN_부서' — 접두어 그대로 저장
        m_new_corp = _re_dept.match(r"^\s*(0[12]_(?:KOR|VN)/\d{2}(?:\.\d)?_.+?)\s*$", department)
        m_vn_new = _re_dept.match(r"^\s*12-VN(\d{2})\s+(.+)$", department) if not m_new_corp else None
        m_vn_legacy = _re_dept.match(r"^\s*VN12-(\d{2})\s+(.+)$", department) if (not m_new_corp and not m_vn_new) else None
        m_kr = _re_dept.match(r"^\s*(\d{2})\s+(.+)$", department) if (not m_new_corp and not m_vn_new and not m_vn_legacy) else None
        if m_new_corp:
            department = m_new_corp.group(1).strip()
        elif m_vn_new:
            department = f"12-VN{m_vn_new.group(1)} {m_vn_new.group(2).strip()}"
        elif m_vn_legacy:
            department = f"12-VN{m_vn_legacy.group(1)} {m_vn_legacy.group(2).strip()}"
        elif m_kr:
            department = m_kr.group(2).strip()
        # 권한 컬럼 제거됨 (대표 지시 2026-05-26) — 모든 신규 사용자는 'staff'로 등록, 관리자가 메신저에서 별도 부여
        role_raw = ""
        # 시트별 컬럼 인덱스 분기 (대표 지시 2026-05-26 슬림화):
        #   · 본사 시트(7컬럼) :     A이름 / B이메일 / C휴대폰 / D사번 / E직급 / F부서 / G이름영문
        #     → display_name_vn 없음, display_name_en = row[6]
        #   · 베트남 시트(8컬럼) :   A이름 / B이메일 / C휴대폰 / D사번 / E직급 / F부서 / G이름베트남어 / H이름영문
        #     → display_name_vn = row[6], display_name_en = row[7]
        #   · 옛 단일 시트 호환 :    row[6]=베트남어, row[7]=영문 (이전 정책 그대로)
        if is_vn_sheet:
            display_name_vn = (str(row[6]) if len(row) > 6 and row[6] is not None else "").strip()
            display_name_en = (str(row[7]) if len(row) > 7 and row[7] is not None else "").strip()
        else:
            # 본사: 베트남어 이름 컬럼 없음 (G가 바로 영문이름)
            display_name_vn = ""
            display_name_en = (str(row[6]) if len(row) > 6 and row[6] is not None else "").strip()
        # 영문 직급·부서 — 매핑 테이블에서 자동 채움 (대표 지시 2026-05-26)
        #   사용자가 엑셀에서 따로 입력 안 함. 매핑에 없는 직급/부서면 빈 값 (한글로 폴백).
        title_en = TITLE_TO_EN.get(title, "")
        department_en = DEPT_TO_EN.get(department, "")
        # 예시 행 자동 스킵
        if email in EXAMPLE_EMAILS or name in EXAMPLE_NAMES:
            skipped.append({"row": row_idx, "name": name, "reason": "예시 행 (자동 스킵)"})
            continue
        # 권한 변환
        role = "ceo" if role_raw == "관리자" else "staff"
        # 검증
        if not name:
            errors.append({"row": row_idx, "name": "(이름 없음)", "error": "이름 누락"})
            continue

        # 로그인 ID = 사번 (대표 지시 2026-05-26 전면 변경):
        #   본사 (KOR)    : username = 사번 (예: '5', '67', '123')
        #   베트남 (VN)   : username = 'VN' + 사번 (예: 'VN1', 'VN23')
        #   email         = 사용자가 적은 값 (없으면 None)
        is_vn = bool(department) and department.startswith("02_VN/")
        auto_generated_vn_id = False   # 옛 자동 ID 정책 폐기 (호환 용도로만 변수 유지)
        if not employee_no or not employee_no.strip():
            errors.append({"row": row_idx, "name": name,
                           "error": "사번(employee_no)은 로그인 ID가 됩니다 — 필수 입력"})
            continue
        emp = employee_no.strip()
        if is_vn:
            # 베트남: 사번에 'VN' 자동 부착 (이미 VN 접두어 있으면 중복 부착 X)
            if not emp.upper().startswith("VN"):
                emp = f"VN{emp}"
            else:
                emp = "VN" + emp[2:]   # 대소문자 정규화
            user_login_id = emp
            email_for_db = email if (email and "@" in email) else None
        else:
            # 본사: 사번 그대로
            user_login_id = emp
            email_for_db = email if (email and "@" in email) else None

        # 비밀번호 결정
        #   본사: 휴대폰 디지트 (입력 있으면) / 없으면 '9999'
        #   베트남: '9999' 기본 (휴대폰 입력해도 9999 — 일관성)
        if is_vn:
            digits = VN_DEFAULT_PW
        else:
            if not phone:
                # 본사도 휴대폰 없으면 임시 비번 '9999' — 첫 로그인 시 변경 강제
                digits = VN_DEFAULT_PW
            else:
                digits = "".join(ch for ch in phone if ch.isdigit())
                if len(digits) < 4:
                    digits = VN_DEFAULT_PW   # 너무 짧으면 9999 폴백
        # 기존 직원이면 정보만 갱신(upsert), 신규면 추가. (대표 지시 2026-05-25)
        existing = db.execute("SELECT id FROM users WHERE username=?", (user_login_id,)).fetchone()
        if existing:
            # 정보 갱신 — 비밀번호·권한(role)·활성·아바타는 절대 건드리지 않음(보호).
            #   빈 칸은 기존값 유지(부분 시트로 인한 실수 삭제 방지). 전화번호는 갱신하되 비번은 불변.
            try:
                upd_pairs = [
                    ("display_name", name), ("title", title), ("department", department),
                    ("employee_no", employee_no), ("phone", phone),
                    ("display_name_vn", display_name_vn),   # 베트남어 이름 (대표 지시 2026-05-26)
                    ("display_name_en", display_name_en), ("title_en", title_en),
                    ("department_en", department_en),
                    # 이메일은 베트남이면 사용자 입력값(또는 None), 본사면 항상 그대로
                    # email_for_db 가 None 이면 upd_pairs 에서 자동 스킵됨(빈 값 스킵 로직)
                    # → 베트남 직원이 이메일을 비웠으면 기존 이메일 보존. 명시적으로 적었으면 갱신.
                    ("email", email_for_db),
                ]
                sets, args = [], []
                for col, val in upd_pairs:
                    if val:
                        sets.append(f"{col}=?")
                        args.append(val)
                if sets:
                    args.append(existing["id"])
                    db.execute(f"UPDATE users SET {', '.join(sets)} WHERE id=?", args)
                # 사진이 해당 행에 있으면 아바타 갱신 (대표 지시 2026-05-26 — 일괄 등록 시 사진 등록)
                _photo_applied = False
                if row_idx in images_by_row:
                    _data, _ext = images_by_row[row_idx]
                    saved_ext = _save_avatar_for_user(existing["id"], _data, _ext)
                    if saved_ext:
                        import time as _t
                        rel_url = f"{BASE_PATH}/uploads/avatars/{existing['id']}.{saved_ext}?v={int(_t.time())}"
                        db.execute("UPDATE users SET avatar_url=? WHERE id=?", (rel_url, existing["id"]))
                        _photo_applied = True
                updated.append({"row": row_idx, "name": name, "email": user_login_id, "photo": _photo_applied})
            except Exception as e:
                errors.append({"row": row_idx, "name": name, "error": f"DB 수정 오류: {e}"})
            continue
        # 신규 등록
        try:
            cur = db.execute(
                "INSERT INTO users (username, password_hash, display_name, role, avatar_color, "
                " created_at, email, phone, title, department, employee_no, must_change_password, "
                " display_name_vn, display_name_en, title_en, department_en) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (
                    user_login_id,   # username = 본사:이메일 / 베트남:자동 ID (대표 지시 2026-05-26 ID·이메일 분리)
                    generate_password_hash(digits),
                    name, role, "#3b82f6",
                    now,
                    email_for_db,    # email = 본사:이메일 / 베트남:사용자 입력값 또는 NULL
                    phone,
                    title or None, department or None, employee_no or None, 1,
                    display_name_vn or None,
                    display_name_en or None, title_en or None, department_en or None,
                ),
            )
            new_uid = cur.lastrowid
            # 사진이 해당 행에 있으면 아바타 등록 (대표 지시 2026-05-26)
            _photo_applied = False
            if row_idx in images_by_row and new_uid:
                _data, _ext = images_by_row[row_idx]
                saved_ext = _save_avatar_for_user(new_uid, _data, _ext)
                if saved_ext:
                    import time as _t
                    rel_url = f"{BASE_PATH}/uploads/avatars/{new_uid}.{saved_ext}?v={int(_t.time())}"
                    db.execute("UPDATE users SET avatar_url=? WHERE id=?", (rel_url, new_uid))
                    _photo_applied = True
            created.append({
                "row": row_idx, "name": name,
                "email": user_login_id,     # 결과 화면에는 username(=로그인 ID) 표시 (이메일 칸과 혼동 방지)
                "user_email": email_for_db,  # 실제 이메일 칸 값 (베트남은 NULL일 수 있음)
                "phone_initial_pw": digits,
                "auto_id_vn": auto_generated_vn_id,
                "photo": _photo_applied,
            })
        except Exception as e:
            errors.append({"row": row_idx, "name": name, "error": f"DB 오류: {e}"})
    db.commit()
    # 등록·수정 후 자동채널(KNK WORLD/본사/베트남) 멤버십 재동기화 — 부서 변경이 본사/베트남 배정에 영향 (대표 지시 2026-05-20)
    if created or updated:
        try: _resync_auto_channels(db)
        except Exception as e: print(f"[auto_channel] bulk resync 실패: {e}")
    # broadcast — 추가·수정된 사용자 목록 즉시 갱신
    _changed_emails = [c["email"] for c in created] + [u["email"] for u in updated]
    if _changed_emails:
        rows = db.execute(
            "SELECT id, username, display_name, display_name_vn, display_name_en, role, avatar_color, title, title_en, department, department_en, email, phone, employee_no, active "
            "FROM users WHERE username IN ({})".format(",".join("?" for _ in _changed_emails)),
            _changed_emails,
        ).fetchall()
        for r in rows:
            socketio.emit("user_info_changed", dict(r))
    return jsonify({
        "created_count": len(created),
        "updated_count": len(updated),
        "skipped_count": len(skipped),
        "error_count": len(errors),
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    })


def _parse_search_filters(q):
    """고급 검색 필터 파싱.
       지원 키: from:이름, in:방이름, has:file|image|link, before:YYYY-MM-DD, after:YYYY-MM-DD
       이름·방에는 # 또는 @ 접두 가능 (예: in:#영업, from:@김정락).
       반환: (plain_text, filters)"""
    filters = {
        'from':  [],    # 발신자 이름·아이디 (OR)
        'in':    [],    # 방 이름 (OR)
        'has':   [],    # file, image, link 중 (AND)
        'before': None, # YYYY-MM-DD
        'after':  None, # YYYY-MM-DD
    }
    parts = []
    # 인용("") 안의 다중 토큰도 한 단어로 묶음 → 일단 단순 split.
    for tok in (q or "").split():
        m = re.match(r'^(from|in|has|before|after):(.+)$', tok, re.IGNORECASE)
        if not m:
            parts.append(tok)
            continue
        key = m.group(1).lower()
        val = m.group(2).strip().lstrip('#@').strip()
        if not val:
            continue
        if key in ('from', 'in'):
            filters[key].append(val)
        elif key == 'has':
            v = val.lower()
            if v in ('file', 'image', 'link', 'photo', 'pic'):
                # photo/pic 도 image 동의어
                if v in ('photo', 'pic'):
                    v = 'image'
                if v not in filters['has']:
                    filters['has'].append(v)
        elif key in ('before', 'after'):
            # YYYY-MM-DD 형식 검증
            if re.match(r'^\d{4}-\d{2}-\d{2}$', val):
                filters[key] = val
    return ' '.join(parts).strip(), filters


@app.route("/api/search")
@login_required
def api_search():
    """전문 검색 + 고급 필터.
       사용 예: "발주 from:김정락 has:file before:2026-05-01 in:#영업\""""
    me = current_user()
    raw_q = (request.args.get("q") or "").strip()
    if not raw_q:
        return jsonify([])
    plain_q, filters = _parse_search_filters(raw_q)
    db = get_db()

    # ----- 동적 WHERE 구성 -----
    where_clauses = []
    args = [me["id"]]   # 첫 인자는 room_members.user_id (방 멤버 권한)

    # from: — 발신자 (display_name 또는 username LIKE) — OR
    if filters['from']:
        sub_or = []
        for name in filters['from']:
            sub_or.append("(u.display_name LIKE ? OR u.username LIKE ?)")
            args.extend([f"%{name}%", f"%{name}%"])
        where_clauses.append("(" + " OR ".join(sub_or) + ")")

    # in: — 방 이름 OR (별명·실제 이름 둘 다 매칭)
    if filters['in']:
        sub_or = []
        for rn in filters['in']:
            sub_or.append("r.name LIKE ?")
            args.append(f"%{rn}%")
        where_clauses.append("(" + " OR ".join(sub_or) + ")")

    # has: — file/image/link AND (다중이면 모두 충족)
    for h in filters['has']:
        if h == 'file':
            where_clauses.append("m.file_path IS NOT NULL")
        elif h == 'image':
            where_clauses.append("(m.kind='image' OR (m.file_mime IS NOT NULL AND m.file_mime LIKE 'image/%'))")
        elif h == 'link':
            where_clauses.append("(m.content LIKE '%http://%' OR m.content LIKE '%https://%')")

    # before/after: — date 비교
    if filters['before']:
        where_clauses.append("date(m.created_at) < date(?)")
        args.append(filters['before'])
    if filters['after']:
        where_clauses.append("date(m.created_at) > date(?)")
        args.append(filters['after'])

    # FTS 또는 LIKE 본문 검색 (plain_q 가 있을 때만)
    use_fts = False
    fts_q = None
    if plain_q:
        fts_q = fts_query_safe(plain_q)
        if fts_q:
            use_fts = True

    if use_fts:
        # FTS MATCH 경로 — content 검색
        sql = f"""
            SELECT m.id, m.content, m.kind, m.created_at, m.room_id, m.file_path, m.file_name, m.file_mime,
                   u.display_name, u.avatar_color, u.username,
                   r.name AS room_name, r.type AS room_type,
                   it.customer AS item_customer, it.code AS item_code,
                   CASE WHEN m.parent_message_id IS NULL THEN 'message' ELSE 'thread' END AS result_type
              FROM messages_fts fts
              JOIN messages m ON m.id = fts.rowid
              JOIN users u ON u.id = m.user_id
              JOIN rooms r ON r.id = m.room_id
              JOIN room_members rm ON rm.room_id = r.id AND rm.user_id = ?
              LEFT JOIN items it ON it.room_id = r.id
             WHERE messages_fts MATCH ?
               {('AND ' + ' AND '.join(where_clauses)) if where_clauses else ''}
             ORDER BY m.id DESC
             LIMIT 50
        """
        # args 순서: me["id"], fts_q, [filter args ...]
        full_args = [me["id"], fts_q] + args[1:]
        msg_rows = db.execute(sql, full_args).fetchall()
    elif where_clauses:
        # 필터만 있고 본문 검색 없음 — 직접 messages 스캔 (FTS 없이)
        sql = f"""
            SELECT m.id, m.content, m.kind, m.created_at, m.room_id, m.file_path, m.file_name, m.file_mime,
                   u.display_name, u.avatar_color, u.username,
                   r.name AS room_name, r.type AS room_type,
                   it.customer AS item_customer, it.code AS item_code,
                   CASE WHEN m.parent_message_id IS NULL THEN 'message' ELSE 'thread' END AS result_type
              FROM messages m
              JOIN users u ON u.id = m.user_id
              JOIN rooms r ON r.id = m.room_id
              JOIN room_members rm ON rm.room_id = r.id AND rm.user_id = ?
              LEFT JOIN items it ON it.room_id = r.id
             WHERE {' AND '.join(where_clauses)}
             ORDER BY m.id DESC
             LIMIT 50
        """
        msg_rows = db.execute(sql, args).fetchall()
    else:
        # 검색어가 전부 stop-word 등으로 fts_query_safe 가 비어졌고 필터도 없음 → 빈 결과
        msg_rows = []

    # 프로젝트 메타 검색 (LIKE) — plain_q 기준. 필터링 모드(예: from:만)면 스킵.
    item_results = []
    if plain_q:
        tokens = re.findall(r"[\w가-힣]+", plain_q)
        if tokens:
            like_clauses = []
            like_args = [me["id"]]
            for t in tokens:
                like_clauses.append("(it.name LIKE ? OR it.customer LIKE ? OR it.code LIKE ? OR it.description LIKE ?)")
                for _ in range(4):
                    like_args.append(f"%{t}%")
            item_rows = db.execute(f"""
                SELECT r.id AS room_id, r.name AS room_name, r.type AS room_type,
                       it.customer AS item_customer, it.code AS item_code,
                       it.status AS item_status, it.description AS item_desc,
                       'item' AS result_type
                  FROM items it
                  JOIN rooms r ON r.id = it.room_id
                  JOIN room_members rm ON rm.room_id = r.id AND rm.user_id = ?
                 WHERE {' AND '.join(like_clauses)}
                 ORDER BY r.id DESC
                 LIMIT 20
            """, like_args).fetchall()
            item_results = [dict(r) for r in item_rows]

    # ----- 사용자 이름 검색 (우선순위 1) — 게스트 제외 / 게스트 사용자는 검색 불가 (대표 지시 2026-06-03) -----
    #   사용자(직원 디렉터리)는 회사 전체 공개 정보 → 방 참여 여부와 무관하게 검색 허용.
    #   (방 안의 정보 = 방·대화내용·스레드·요청 은 아래에서 전부 '내 멤버십'으로 제한 — 대표 지시 2026-06-06)
    user_results = []
    matched_user_ids = []
    if plain_q and not _is_guest(me):
        utokens = re.findall(r"[\w가-힣]+", plain_q)
        if utokens:
            ulike = []
            uargs = []
            for t in utokens:
                ulike.append("(u.display_name LIKE ? OR u.username LIKE ?)")
                uargs.extend([f"%{t}%", f"%{t}%"])
            user_rows = db.execute(f"""
                SELECT u.id AS user_id, u.display_name, u.username, u.avatar_color,
                       u.title AS user_title, u.department AS user_department,
                       'user' AS result_type
                  FROM users u
                 WHERE COALESCE(u.is_guest,0)=0 AND ({' AND '.join(ulike)})
                 ORDER BY u.display_name
                 LIMIT 20
            """, uargs).fetchall()
            user_results = [dict(r) for r in user_rows]
            matched_user_ids = [r["user_id"] for r in user_results]

    # ----- 방 검색 (우선순위 2) — 반드시 '내가 멤버인 방'만 (대표 지시 2026-06-06: 내가 없는 방 정보는 검색 금지) -----
    #   (a) 검색된 사람이 참여한 방 ∩ 내가 참여한 방  (예: '정민규' → 정민규와 내가 함께 있는 방)
    #   (b) 방 이름이 검색어와 일치하는 방 ∩ 내가 참여한 방  (예: '베트남' → 베트남 방)
    #   두 결과를 room_id 로 합치고 중복 제거. 둘 다 room_members 로 내 멤버십 강제.
    room_results = []
    _seen_rooms = set()
    if plain_q:
        # (a) 사람 공유 방 — 검색된 사용자가 멤버이고 나도 멤버인 방 (자기 자신 검색 시 방 폭주 방지 위해 본인 제외)
        _other_ids = [uid for uid in matched_user_ids if uid != me["id"]]
        if _other_ids:
            ph = ",".join("?" * len(_other_ids))
            shared_rows = db.execute(f"""
                SELECT r.id AS room_id, r.name AS room_name, r.type AS room_type,
                       'room' AS result_type
                  FROM rooms r
                  JOIN room_members rm_me ON rm_me.room_id = r.id AND rm_me.user_id = ?
                  JOIN room_members rm_u  ON rm_u.room_id  = r.id AND rm_u.user_id IN ({ph})
                 GROUP BY r.id
                 ORDER BY r.id DESC
                 LIMIT 20
            """, [me["id"]] + _other_ids).fetchall()
            for r in shared_rows:
                if r["room_id"] not in _seen_rooms:
                    _seen_rooms.add(r["room_id"]); room_results.append(dict(r))
        # (b) 방 이름 매칭 — 내가 멤버인 방
        rtokens = re.findall(r"[\w가-힣]+", plain_q)
        if rtokens:
            rlike = []
            rargs = [me["id"]]
            for t in rtokens:
                rlike.append("r.name LIKE ?")
                rargs.append(f"%{t}%")
            room_rows = db.execute(f"""
                SELECT r.id AS room_id, r.name AS room_name, r.type AS room_type,
                       'room' AS result_type
                  FROM rooms r
                  JOIN room_members rm ON rm.room_id = r.id AND rm.user_id = ?
                 WHERE r.name IS NOT NULL AND ({' AND '.join(rlike)})
                 ORDER BY r.id DESC
                 LIMIT 20
            """, rargs).fetchall()
            for r in room_rows:
                if r["room_id"] not in _seen_rooms:
                    _seen_rooms.add(r["room_id"]); room_results.append(dict(r))

    # ----- 대화내용 / 스레드 분리 (대표 지시 2026-06-06) — 둘 다 위 msg_rows 에서 이미 내 멤버십 강제됨 -----
    _msgs = [dict(r) for r in msg_rows]
    conv_results   = [m for m in _msgs if m.get("result_type") == "message"]   # 대화내용(일반 메시지)
    thread_results = [m for m in _msgs if m.get("result_type") == "thread"]    # 스레드(답글)

    # 우선순위: 1.사용자 → 2.방(내 멤버) → 3.대화내용 → 스레드 → 요청 (대표 지시 2026-06-06)
    return jsonify(user_results + room_results + conv_results + thread_results + item_results)


# ── 사내 메신저 이용·보안 동의서 API (대표 지시 2026-06-03) ──
# 동의서 한글 강제 대상 — 베트남법인(02_VN/) 소속이지만 한국인 (대표 지시 2026-06-05).
#   현재 해당자: 이용식(VN001)·박지만(VN003)·권혁인(VN269). 셋 다 display_name_vn(베트남어 이름) 없음.
#   → 아래 _consent_lang_for 의 '한글 이름만' 규칙으로도 자동 ko 지만, 사번으로도 못박아 이중 안전.
CONSENT_KO_FORCE_USERNAMES = {"VN001", "VN003", "VN269"}


def _consent_lang_for(user):
    """동의서 표시 언어를 소속·이름 기준으로 자동 선택 (대표 지시 2026-06-05).
    · 베트남법인(부서 02_VN/…) 소속 + 베트남어 이름(display_name_vn) 있음 → 'vi'
    · 그 외(본사 / 베트남법인이지만 한글 이름만 / 한국인 강제명단) → 'ko'
    UI 화면 언어와 무관하게 '누구인지'로 동의서 언어를 정한다."""
    try:
        uname = str(user["username"]).strip()
    except Exception:
        uname = ""
    if uname in CONSENT_KO_FORCE_USERNAMES:
        return "ko"
    try:
        dept = user["department"]
    except Exception:
        dept = None
    try:
        vn = (user["display_name_vn"] or "").strip()
    except Exception:
        vn = ""
    if _user_is_vietnam(dept) and vn:
        return "vi"
    return "ko"


def _consent_needed(db, user_id):
    """현재 버전 동의가 없거나, 마지막 동의로부터 CONSENT_INTERVAL_DAYS 일이 지났으면 True."""
    row = db.execute(
        "SELECT MAX(agreed_at) AS last_at FROM consent_agreements WHERE user_id=? AND version=?",
        (user_id, CONSENT_VERSION),
    ).fetchone()
    last_at = row["last_at"] if row else None
    if not last_at:
        return True
    try:
        last_dt = datetime.fromisoformat(last_at)
        if last_dt.tzinfo is None:
            last_dt = last_dt.replace(tzinfo=timezone.utc)
        return (datetime.now(timezone.utc) - last_dt).days >= CONSENT_INTERVAL_DAYS
    except Exception:
        return True


@app.route("/api/me/consent")
@login_required
def api_me_consent():
    """동의 필요 여부 + 동의서 본문(사용자 언어). 게스트는 항상 needed=False."""
    me = current_user()
    if _is_guest(me):
        return jsonify({"needed": False})
    db = get_db()
    # 테스트 모드: 지정 계정(사번)만 + 매번 항상 표시 / 정식: 전 직원 + 100일 주기 (대표 지시 2026-06-03)
    if CONSENT_TEST_MODE:
        needed = (str(me["username"]) in CONSENT_TEST_USERNAMES)
    else:
        needed = _consent_needed(db, me["id"])
    # 동의서 언어 — UI 화면 언어가 아니라 '소속·이름' 기준 자동 (대표 지시 2026-06-05):
    #   베트남법인(02_VN/) + 베트남어 이름 있음 → vi / 그 외(본사·한국인 명단·한글 이름만) → ko
    lang = _consent_lang_for(me)
    doc = CONSENT_DOC["vi"] if lang == "vi" else CONSENT_DOC["ko"]
    return jsonify({
        "needed": needed,
        "version": CONSENT_VERSION,
        "title": doc["title"],
        "html": doc["html"],
        "agree_label": doc["agree"],
        "decline_label": doc["decline"],
        "scroll_hint": doc["scroll_hint"],
        "declined_msg": doc["declined_msg"],
    })


@app.route("/api/me/consent", methods=["POST"])
@login_required
def api_me_consent_agree():
    """현재 버전 동의 기록(시각·IP·UA). 게스트는 차단."""
    me = current_user()
    if _is_guest(me):
        return jsonify({"error": "외부 사용자는 대상이 아닙니다."}), 403
    db = get_db()
    now = datetime.now(timezone.utc).isoformat()
    ip = (request.headers.get("X-Forwarded-For", "").split(",")[0].strip()
          or request.remote_addr or "")
    ua = (request.headers.get("User-Agent", "") or "")[:300]
    db.execute(
        "INSERT INTO consent_agreements (user_id, version, agreed_at, ip, user_agent) VALUES (?,?,?,?,?)",
        (me["id"], CONSENT_VERSION, now, ip, ua),
    )
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/admin/consent_status")
@login_required
def api_admin_consent_status():
    """전 직원 사용동의 현황 — 현재 버전(CONSENT_VERSION) 기준 동의/미동의 명단. 관리자 전용. (대표 지시 2026-06-05)
    게스트·비활성 제외. 미동의자 먼저 정렬(독려 목적)."""
    me = current_user()
    if not _is_admin_user(me):
        abort(403)
    db = get_db()
    rows = db.execute(
        """
        SELECT u.id, u.username, u.display_name, u.display_name_vn, u.department, u.title,
               (SELECT MAX(agreed_at) FROM consent_agreements
                 WHERE user_id=u.id AND version=?) AS agreed_at
          FROM users u
         WHERE COALESCE(u.is_guest,0)=0 AND u.active=1
        """,
        (CONSENT_VERSION,),
    ).fetchall()
    members = []
    agreed = 0
    hq_total = hq_agreed = vn_total = vn_agreed = 0
    for r in rows:
        is_vn = _user_is_vietnam(r["department"])
        has = bool(r["agreed_at"])
        if has:
            agreed += 1
        if is_vn:
            vn_total += 1
            vn_agreed += 1 if has else 0
        else:
            hq_total += 1
            hq_agreed += 1 if has else 0
        members.append({
            "user_id": r["id"], "username": r["username"],
            "display_name": r["display_name"], "display_name_vn": r["display_name_vn"],
            "department": r["department"], "title": r["title"],
            "is_vn": is_vn, "agreed": has, "agreed_at": r["agreed_at"],
        })
    members.sort(key=lambda m: (m["agreed"], m["department"] or "", m["display_name"] or ""))
    return jsonify({
        "version": CONSENT_VERSION,
        "total": len(rows), "agreed": agreed, "pending": len(rows) - agreed,
        "hq": {"total": hq_total, "agreed": hq_agreed, "pending": hq_total - hq_agreed},
        "vn": {"total": vn_total, "agreed": vn_agreed, "pending": vn_total - vn_agreed},
        "members": members,
    })


@app.route("/api/admin/consent_status/export.xlsx")
@login_required
def api_admin_consent_status_export():
    """사용동의 현황 엑셀 (본사/베트남 시트 분리, 미동의 강조). 관리자 전용. (대표 지시 2026-06-05)"""
    me = current_user()
    if not _is_admin_user(me):
        abort(403)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl 미설치 — 서버 관리자에게 문의"}), 500
    db = get_db()
    rows = db.execute(
        """
        SELECT u.username, u.display_name, u.display_name_vn, u.department, u.title,
               (SELECT MAX(agreed_at) FROM consent_agreements
                 WHERE user_id=u.id AND version=?) AS agreed_at
          FROM users u
         WHERE COALESCE(u.is_guest,0)=0 AND u.active=1
        """,
        (CONSENT_VERSION,),
    ).fetchall()
    hq_rows = [r for r in rows if not _user_is_vietnam(r["department"] or "")]
    vn_rows = [r for r in rows if _user_is_vietnam(r["department"] or "")]

    _thin = Side(border_style="thin", color="D1D5DB")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _headers = ["이름", "베트남어 이름", "직급", "부서", "로그인ID", "동의여부", "동의일시"]
    _widths = [14, 18, 10, 18, 12, 9, 22]

    def _fill_sheet(ws, corp_label, data_rows):
        ws["A1"] = f"사용동의 현황 (v{CONSENT_VERSION}) — {corp_label}"
        ws["A1"].font = Font(bold=True, size=14, color="A5282C")
        ws.merge_cells("A1:G1")
        _ag = sum(1 for r in data_rows if r["agreed_at"])
        ws["A2"] = f"총 {len(data_rows)}명 / 동의 {_ag}명 / 미동의 {len(data_rows) - _ag}명"
        ws["A2"].font = Font(size=10, color="6B7280")
        ws.merge_cells("A2:G2")
        hf = PatternFill(start_color="A5282C", end_color="A5282C", fill_type="solid")
        hfont = Font(bold=True, color="FFFFFF")
        for ci, h in enumerate(_headers, start=1):
            c = ws.cell(row=4, column=ci, value=h)
            c.fill = hf
            c.font = hfont
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _border
        sorted_rows = sorted(data_rows, key=lambda r: (bool(r["agreed_at"]), r["display_name"] or ""))
        for idx, r in enumerate(sorted_rows, start=1):
            rn = 4 + idx
            has = bool(r["agreed_at"])
            at = (r["agreed_at"] or "")
            at = at[:19].replace("T", " ") if has else ""
            vals = [r["display_name"] or "", r["display_name_vn"] or "", r["title"] or "",
                    r["department"] or "", r["username"] or "", "동의" if has else "미동의", at]
            for ci, v in enumerate(vals, start=1):
                c = ws.cell(row=rn, column=ci, value=v)
                c.border = _border
                if ci == 6 and not has:
                    c.font = Font(bold=True, color="DC2626")
        for ci, w in enumerate(_widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "본사"
    _fill_sheet(ws1, "본사", hq_rows)
    ws2 = wb.create_sheet("베트남법인")
    _fill_sheet(ws2, "베트남법인", vn_rows)

    import io
    from urllib.parse import quote
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"사용동의현황_v{CONSENT_VERSION}.xlsx"
    resp = make_response(bio.read())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(fname)}"
    return resp


@app.route("/api/rooms", methods=["POST"])
@login_required
def api_rooms_create():
    me = current_user()
    if _is_guest(me):
        return jsonify({"error": "외부 사용자는 새 방을 만들 수 없습니다."}), 403
    data = request.get_json(silent=True) or {}
    user_ids = list({int(x) for x in (data.get("user_ids") or [])})
    if me["id"] not in user_ids:
        user_ids.append(me["id"])
    if len(user_ids) < 2:
        return jsonify({"error": "최소 2명 이상이어야 합니다."}), 400
    type_ = data.get("type") or ("direct" if len(user_ids) == 2 else "group")
    # 채널 생성 — 관리자(ceo)·대표이사·임원(전무·상무·이사)·팀장·법인장 (대표 지시 2026-05-24)
    #   초대 범위: 부서 제한 없이 등록된 전체 직원 누구나 가능.
    if type_ == "channel":
        if not _can_create_channel(me):
            return jsonify({"error": "채널은 관리자·대표이사·임원·팀장·법인장만 만들 수 있습니다."}), 403
    name = (data.get("name") or "").strip() or None
    # 이름 고정 — 방장이 지정한 이름을 다른 멤버가 변경/별명 불가
    name_locked = 1 if data.get("name_locked") else 0
    # direct 방은 항상 name_locked=0 (이름 없음, 각자 상대방 이름으로 표시)
    if type_ == "direct":
        name_locked = 0
    db = get_db()
    now = datetime.now(timezone.utc).isoformat()

    if type_ == "direct" and len(user_ids) == 2:
        other = [u for u in user_ids if u != me["id"]][0]
        # 상대가 나갔어도 기존 방 재사용 + 빠진 멤버 복원 (이전 대화 유지)
        rid, existing = _find_or_create_direct(db, me["id"], other)
        return jsonify({"id": rid, "existing": existing, "name_locked": 0})

    # 방 이름 AI 자동 번역 취소 (대표 지시 2026-05-28 번복)
    #   이유: 방 멤버만 보이는 구조라 다른 언어 사용자가 그 방을 볼 일 없음 → AI 비용 낭비.
    #   helper(_translate_room_name) 와 DB 컬럼(name_vi/en) 은 보존 (필요 시 재활성화).
    cur = db.execute(
        "INSERT INTO rooms (name, type, created_by, created_at, name_locked) VALUES (?,?,?,?,?)",
        (name, type_, me["id"], now, name_locked),
    )
    rid = cur.lastrowid
    # 생성자는 host, 나머지는 member
    for uid in user_ids:
        role = 'host' if uid == me["id"] else 'member'
        db.execute(
            "INSERT INTO room_members (room_id, user_id, joined_at, role) VALUES (?,?,?,?)",
            (rid, uid, now, role),
        )
    db.commit()
    # 방 이름 자동 번역 재가동 (대표 지시 2026-06-05) — ko/vi/en (생성 직후엔 게스트 없음). 동기·실패무해.
    _retranslate_room(db, rid, langs=["ko", "vi", "en"])
    return jsonify({"id": rid, "existing": False, "name_locked": name_locked})


# ── 🐞 버그 신고 API ──────────────────────────────────────────────────────
@app.route("/api/bug_reports", methods=["POST"])
@login_required
def api_bug_report_create():
    """직원이 버그 신고 → 전 직원 '버그 신고' 채널에 글로 올리고 + 구조화 보관.
    스크린샷은 신고 후 클라이언트가 기존 업로드 API 로 같은 방에 붙임. (대표 지시 2026-06-03)"""
    me = current_user()
    if _is_guest(me):
        return jsonify({"error": "외부 사용자는 버그 신고를 사용할 수 없습니다."}), 403
    data = request.get_json(silent=True) or {}
    title = (data.get("title") or "").strip()[:200]
    body = (data.get("body") or "").strip()[:4000]
    if not title and not body:
        return jsonify({"error": "내용을 입력해 주세요."}), 400
    ctx = data.get("context") or {}
    keys = ("screen", "room", "device", "browser", "os", "app_version", "ui_lang", "url", "screen_size")
    safe_ctx = {}
    for k in keys:
        v = ctx.get(k)
        if v is not None and str(v).strip():
            safe_ctx[k] = str(v).strip()[:300]
    import json as _json
    ctx_json = _json.dumps(safe_ctx, ensure_ascii=False)
    db = get_db()
    now = datetime.now(timezone.utc).isoformat()
    rid = _get_bug_room_id(db)
    db.execute(
        "INSERT OR IGNORE INTO room_members (room_id, user_id, joined_at, role) VALUES (?,?,?, 'member')",
        (rid, me["id"], now),
    )
    # 채널에 올라갈, 사람이 읽을 메시지 (제목 + 내용 + 기기/버전 한 줄)
    msg = "🐞 [버그 신고]"
    if title:
        msg += " " + title
    if body:
        msg += "\n" + body
    tail = []
    if safe_ctx.get("device"):
        tail.append("기기 " + safe_ctx["device"])
    if safe_ctx.get("app_version"):
        tail.append("버전 " + safe_ctx["app_version"])
    if tail:
        msg += "\n— " + " · ".join(tail)
    cur = db.execute(
        "INSERT INTO messages (room_id, user_id, content, kind, created_at) VALUES (?,?,?,?,?)",
        (rid, me["id"], msg, "text", now),
    )
    mid = cur.lastrowid
    db.execute(
        "INSERT INTO bug_reports (reporter_user_id, title, body, status, context_json, room_id, message_id, created_at, updated_at) "
        "VALUES (?,?,?,?,?,?,?,?,?)",
        (me["id"], title, body, "new", ctx_json, rid, mid, now, now),
    )
    db.commit()
    u = db.execute("SELECT display_name, avatar_color FROM users WHERE id=?", (me["id"],)).fetchone()
    payload = {
        "id": mid, "room_id": rid, "user_id": me["id"],
        "display_name": u["display_name"], "avatar_color": u["avatar_color"],
        "content": msg, "kind": "text", "created_at": now,
    }
    try:
        socketio.emit("new_message", payload, to=f"room_{rid}")
    except Exception:
        pass
    return jsonify({"ok": True, "room_id": rid, "message_id": mid})


@app.route("/api/admin/bug_reports")
@login_required
def api_admin_bug_reports():
    """유지보수관리자(김정락) 전용 — 버그 신고 전체 목록 + 상태별 집계."""
    me = current_user()
    if not _is_maintenance_admin(me):
        return jsonify({"error": "권한이 없습니다."}), 403
    db = get_db()
    status = (request.args.get("status") or "").strip()
    q = ("SELECT b.id, b.reporter_user_id, b.title, b.body, b.status, b.context_json, "
         "       b.room_id, b.message_id, b.created_at, b.updated_at, "
         "       u.display_name AS reporter_name, u.username AS reporter_username "
         "  FROM bug_reports b LEFT JOIN users u ON u.id = b.reporter_user_id")
    args = []
    if status in BUG_STATUSES:
        q += " WHERE b.status = ?"
        args.append(status)
    q += " ORDER BY b.id DESC LIMIT 500"
    reports = [dict(r) for r in db.execute(q, args).fetchall()]
    counts = {s: 0 for s in BUG_STATUSES}
    for r in db.execute("SELECT status, COUNT(*) AS c FROM bug_reports GROUP BY status").fetchall():
        if r["status"] in counts:
            counts[r["status"]] = r["c"]
    return jsonify({"reports": reports, "counts": counts, "total": sum(counts.values())})


@app.route("/api/admin/bug_reports/<int:bid>/status", methods=["POST"])
@login_required
def api_admin_bug_report_status(bid):
    """유지보수관리자 전용 — 버그 신고 처리상태 변경."""
    me = current_user()
    if not _is_maintenance_admin(me):
        return jsonify({"error": "권한이 없습니다."}), 403
    data = request.get_json(silent=True) or {}
    st = (data.get("status") or "").strip()
    if st not in BUG_STATUSES:
        return jsonify({"error": "잘못된 상태입니다."}), 400
    db = get_db()
    now = datetime.now(timezone.utc).isoformat()
    db.execute("UPDATE bug_reports SET status=?, updated_at=? WHERE id=?", (st, now, bid))
    db.commit()
    return jsonify({"ok": True, "status": st})


# ── 📊 사용 현황 (채택률·기능별·부서별) — 유지보수관리자(김정락) 전용 ──────────
#   대표 지시 2026-06-03: "사용현황·기록은 나(유지보수담당자)만 본다."
#   대부분 '있는 데이터' 집계 — 활성 사용자 = 기간 내 실제 메시지 발신자(=진짜 업무 사용).
@app.route("/api/admin/usage_overview")
@login_required
def api_admin_usage_overview():
    me = current_user()
    if not _is_maintenance_admin(me):
        return jsonify({"error": "유지보수관리자 전용입니다."}), 403

    period = (request.args.get("period") or "this_month").lower()
    if period == "this_month":
        since, until, label = "date('now','start of month')", None, "이번 달"
    elif period == "last_month":
        since, until, label = "date('now','start of month','-1 month')", "date('now','start of month')", "지난 달"
    elif period == "last_7days":
        since, until, label = "date('now','-7 days')", None, "최근 7일"
    elif period == "last_30days":
        since, until, label = "date('now','-30 days')", None, "최근 30일"
    elif period == "all":
        since, until, label = None, None, "전체"
    else:
        since, until, label, period = "date('now','start of month')", None, "이번 달", "this_month"

    def dcond(col):
        # period 는 위에서 고정 집합으로 검증됨 → since/until 은 신뢰 가능한 상수식(주입 위험 없음)
        s = ""
        if since:
            s += f" AND date({col}) >= {since}"
        if until:
            s += f" AND date({col}) < {until}"
        return s

    db = get_db()
    NONGUEST = "COALESCE(u.is_guest,0)=0"
    REALMSG = "COALESCE(m.kind,'text') NOT IN ('system','deleted')"

    def scalar(sql):
        r = db.execute(sql).fetchone()
        return int((r["c"] if r else 0) or 0)

    # 전체 직원(비게스트·활성)
    total_users = scalar("SELECT COUNT(*) c FROM users u WHERE COALESCE(u.is_guest,0)=0 AND COALESCE(u.active,1)=1")

    # 기간 내 활성 사용자(실제 발신자) uid 집합
    active_uids = set()
    for r in db.execute(f"""SELECT DISTINCT m.user_id AS uid
          FROM messages m JOIN users u ON u.id=m.user_id
         WHERE {NONGUEST} AND {REALMSG} {dcond('m.created_at')}""").fetchall():
        if r["uid"] is not None:
            active_uids.add(r["uid"])
    active_users = len(active_uids)

    # DAU/WAU/MAU (채택률 핵심 — period 무관 고정 계산)
    def senders_since(expr):
        return scalar(f"""SELECT COUNT(DISTINCT m.user_id) c
              FROM messages m JOIN users u ON u.id=m.user_id
             WHERE {NONGUEST} AND {REALMSG} AND date(m.created_at) >= {expr}""")
    dau = senders_since("date('now')")
    wau = senders_since("date('now','-7 days')")
    mau = senders_since("date('now','-30 days')")

    # 부서별 / 본사·베트남 — 전 직원을 파이썬에서 활성 매칭
    hq_total = hq_active = vn_total = vn_active = 0
    dept_map = {}
    for u in db.execute("SELECT id, department FROM users u WHERE COALESCE(u.is_guest,0)=0 AND COALESCE(u.active,1)=1").fetchall():
        is_vn = _user_is_vietnam(u["department"])
        act = u["id"] in active_uids
        if is_vn:
            vn_total += 1; vn_active += 1 if act else 0
        else:
            hq_total += 1; hq_active += 1 if act else 0
        dept = (u["department"] or "(미지정)")
        d = dept_map.setdefault(dept, {"total": 0, "active": 0})
        d["total"] += 1
        if act:
            d["active"] += 1
    dept_list = [{"department": k, "total": v["total"], "active": v["active"]} for k, v in dept_map.items()]
    dept_list.sort(key=lambda x: (-x["active"], -x["total"], x["department"]))

    # 기능별 사용량(기간 내)
    msg_count = scalar(f"SELECT COUNT(*) c FROM messages m JOIN users u ON u.id=m.user_id WHERE {NONGUEST} AND COALESCE(m.kind,'text') IN ('text','sticker') {dcond('m.created_at')}")
    file_count = scalar(f"SELECT COUNT(*) c FROM messages m JOIN users u ON u.id=m.user_id WHERE {NONGUEST} AND COALESCE(m.kind,'text') IN ('image','file') {dcond('m.created_at')}")
    tr_count = scalar(f"SELECT COUNT(*) c FROM message_translations t WHERE 1=1 {dcond('t.created_at')}")
    sum_count = scalar(f"SELECT COUNT(*) c FROM ai_summaries s WHERE s.created_by IS NOT NULL {dcond('s.created_at')}")
    req_count = scalar(f"SELECT COUNT(*) c FROM requests r WHERE 1=1 {dcond('r.created_at')}")
    mention_count = scalar(f"SELECT COUNT(*) c FROM mentions mn WHERE 1=1 {dcond('mn.created_at')}")
    reaction_count = scalar(f"SELECT COUNT(*) c FROM message_reactions mr WHERE 1=1 {dcond('mr.created_at')}")
    features = [
        {"key": "message", "label": "메시지", "emoji": "💬", "count": msg_count},
        {"key": "file", "label": "사진·파일", "emoji": "📎", "count": file_count},
        {"key": "translate", "label": "AI 번역", "emoji": "🌐", "count": tr_count},
        {"key": "summary", "label": "AI 요약", "emoji": "📝", "count": sum_count},
        {"key": "request", "label": "업무 요청", "emoji": "📌", "count": req_count},
        {"key": "mention", "label": "멘션(@)", "emoji": "🔔", "count": mention_count},
        {"key": "reaction", "label": "반응", "emoji": "👍", "count": reaction_count},
    ]
    features.sort(key=lambda x: -x["count"])

    # 개인별 상세(유지보수관리자만) — 기간 내 활동량 상위
    per_user = []
    for r in db.execute(f"""SELECT u.id, u.display_name, u.display_name_vn, u.department, u.title,
               SUM(CASE WHEN COALESCE(m.kind,'text') IN ('text','sticker') THEN 1 ELSE 0 END) AS msgs,
               SUM(CASE WHEN COALESCE(m.kind,'text') IN ('image','file') THEN 1 ELSE 0 END) AS files,
               COUNT(*) AS total, MAX(m.created_at) AS last_at
          FROM messages m JOIN users u ON u.id=m.user_id
         WHERE {NONGUEST} AND {REALMSG} {dcond('m.created_at')}
         GROUP BY u.id ORDER BY total DESC LIMIT 200""").fetchall():
        per_user.append({
            "user_id": r["id"], "display_name": r["display_name"] or "?",
            "display_name_vn": r["display_name_vn"] or "", "is_vn": bool(_user_is_vietnam(r["department"] or "")),
            "department": r["department"] or "", "title": r["title"] or "",
            "msgs": int(r["msgs"] or 0), "files": int(r["files"] or 0), "total": int(r["total"] or 0),
            "last_at": str(r["last_at"] or "")[:16].replace("T", " "),
        })

    return jsonify({
        "period": period, "period_label": label,
        "adoption": {
            "total_users": total_users, "active_users": active_users,
            "active_pct": round(active_users * 100.0 / total_users, 1) if total_users else 0,
            "dau": dau, "wau": wau, "mau": mau,
            "hq_total": hq_total, "hq_active": hq_active,
            "vn_total": vn_total, "vn_active": vn_active,
        },
        "features": features,
        "departments": dept_list,
        "per_user": per_user,
    })


# ── 👤 운영(유지보수) 담당자 지정 — 기본 유지보수관리자(김정락)만 ──────────────
@app.route("/api/admin/ops_admins")
@login_required
def api_admin_ops_admins():
    """운영 담당자 목록 — 지정 권한자(대표님=유지보수 owner)만 조회."""
    me = current_user()
    if not _is_maintenance_owner(me):
        return jsonify({"error": "권한이 없습니다."}), 403
    db = get_db()
    rows = db.execute("""
        SELECT id, username, display_name, display_name_vn, department, title, COALESCE(ops_allowed,0) AS ops_allowed
          FROM users
         WHERE COALESCE(is_guest,0)=0 AND COALESCE(active,1)=1
         ORDER BY (COALESCE(ops_allowed,0)=1) DESC, display_name
    """).fetchall()
    owners = set(MAINTENANCE_ADMIN_USERNAMES)
    out = []
    for r in rows:
        out.append({
            "user_id": r["id"], "display_name": r["display_name"] or "?",
            "display_name_vn": r["display_name_vn"] or "", "is_vn": bool(_user_is_vietnam(r["department"] or "")),
            "department": r["department"] or "", "title": r["title"] or "",
            "ops_allowed": int(r["ops_allowed"] or 0),
            "is_owner": str(r["username"] or "") in owners,
        })
    return jsonify({"users": out})


@app.route("/api/admin/ops_admins/<int:uid>", methods=["POST"])
@login_required
def api_admin_ops_admins_set(uid):
    """운영 권한 부여/회수 — 대표님(유지보수 owner)만. 기본 유지보수관리자 본인은 변경 불가(항상 활성)."""
    me = current_user()
    if not _is_maintenance_owner(me):
        return jsonify({"error": "권한이 없습니다."}), 403
    data = request.get_json(silent=True) or {}
    allowed = 1 if int(data.get("allowed") or 0) == 1 else 0
    db = get_db()
    target = db.execute("SELECT username FROM users WHERE id=?", (uid,)).fetchone()
    if not target:
        return jsonify({"error": "사용자를 찾을 수 없습니다."}), 404
    if str(target["username"] or "") in MAINTENANCE_ADMIN_USERNAMES:
        return jsonify({"error": "기본 유지보수관리자는 항상 활성입니다."}), 400
    db.execute("UPDATE users SET ops_allowed=? WHERE id=?", (allowed, uid))
    db.commit()
    return jsonify({"ok": True, "ops_allowed": allowed})


@app.route("/api/admin/system_config")
@login_required
def api_admin_system_config():
    """🔧 시스템 설정 현재값 — 유지보수(대표님·운영자) 전용.
    설정 토글(번역ON/OFF·상태표시·AI모델·WORKS)이 🔐권한→🔧유지보수로 이동하며,
    상태 로딩을 ceo 전용 권한표에서 분리. (대표 지시 2026-06-03)"""
    me = current_user()
    if not _is_maintenance_admin(me):
        return jsonify({"error": "유지보수 전용입니다."}), 403
    db = get_db()
    return jsonify({
        "translate_enabled": 1 if _ai_translate_enabled(db) else 0,
        "presence_detail": 1 if _presence_show_detail(db) else 0,
        "ai_model": _get_openai_model(db),
        "ai_model_locked": 1 if _OPENAI_MODEL_ENV else 0,
        "works_history_sync": 1 if _works_history_sync_enabled() else 0,
        "works_history_url_set": 1 if WORKS_HISTORY_URL else 0,
    })


@app.route("/api/rooms/<int:room_id>/membership", methods=["DELETE"])
@login_required
def api_leave_room(room_id):
    """방 나가기 — 본인을 room_members에서 제거 + 시스템 메시지로 다른 멤버에 알림."""
    me = current_user()
    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?", (room_id, me["id"])
    ).fetchone():
        return jsonify({"error": "이 방의 멤버가 아닙니다."}), 404

    room = db.execute("SELECT type, name, channel_scope FROM rooms WHERE id=?", (room_id,)).fetchone()
    if not room:
        return jsonify({"error": "방이 없습니다."}), 404
    # 자동 채널(KNK WORLD/본사/베트남)은 전사·소속 채널이라 나갈 수 없음 (대표 지시 2026-05-20)
    if room["channel_scope"]:
        return jsonify({"error": "전사·소속 채널은 나갈 수 없습니다."}), 400

    # 진행 중인 '받은 요청'이 있으면 나가기 차단 — 완료 후 나갈 수 있게 (대표 지시 2026-05-21)
    pending = db.execute(
        "SELECT COUNT(*) AS n FROM requests WHERE room_id=? AND assigned_to=? AND status IN ('open','in_progress','on_hold')",
        (room_id, me["id"]),
    ).fetchone()
    if pending and pending["n"] > 0:
        return jsonify({
            "error": f"이 방에 처리할 요청이 {pending['n']}건 있습니다. 완료(또는 취소·보류 해제)한 뒤 나갈 수 있습니다."
        }), 400

    now = datetime.now(timezone.utc).isoformat()
    sys_text = f"[{me['display_name']}] 님이 나갔습니다."
    cur = db.execute(
        "INSERT INTO messages (room_id, user_id, content, kind, created_at) VALUES (?,?,?,?,?)",
        (room_id, me["id"], sys_text, "system", now),
    )
    sys_mid = cur.lastrowid
    db.execute(
        "DELETE FROM room_members WHERE room_id=? AND user_id=?", (room_id, me["id"])
    )
    db.commit()

    # 남은 멤버에게 시스템 메시지 emit
    socketio.emit("new_message", {
        "id": sys_mid, "room_id": room_id, "user_id": me["id"],
        "display_name": me["display_name"], "avatar_color": me["avatar_color"],
        "content": sys_text, "kind": "system", "created_at": now,
    }, to=f"room_{room_id}")

    return jsonify({"ok": True, "room_type": room["type"], "room_name": room["name"]})


# ============================================================
# 방 권한 관리: host(방장)·sub_host(부방장)·member(일반)
# ============================================================
def _my_room_role(db, room_id, user_id):
    r = db.execute("SELECT role FROM room_members WHERE room_id=? AND user_id=?",
                   (room_id, user_id)).fetchone()
    if not r:
        return None
    role = r["role"]
    # 채널에서는 등록된 관리자(ceo)·최고관리자 모두가 방장 기능 수행 가능 (대표 지시 2026-05-21)
    # → 멤버이기만 하면 ceo 는 effective 'host' 로 취급 (rename·멤버관리·초대정책·삭제 등).
    if role != 'host':
        room = db.execute("SELECT type FROM rooms WHERE id=?", (room_id,)).fetchone()
        if room and room["type"] == 'channel':
            u = db.execute("SELECT role FROM users WHERE id=?", (user_id,)).fetchone()
            if u and u["role"] == 'ceo':
                return 'host'
    return role


def _room_members_full(db, room_id, viewer_id=None):
    """방 멤버 + 역할 + 표시정보 일괄. UI 멤버 패널용.
    viewer_id 주면 그 사람이 각 게스트에게 지정한 '대화방 표현'(개인 별칭) view_alias 동봉. (대표 지시 2026-05-30)"""
    return [dict(r) for r in db.execute("""
        SELECT u.id, u.username, u.display_name, u.avatar_color,
               COALESCE(u.is_guest,0) AS is_guest,
               rm.role, rm.joined_at,
               (SELECT alias FROM guest_view_aliases
                 WHERE viewer_user_id=? AND guest_user_id=u.id) AS view_alias
          FROM room_members rm JOIN users u ON u.id = rm.user_id
         WHERE rm.room_id = ?
         ORDER BY CASE rm.role WHEN 'host' THEN 0 WHEN 'sub_host' THEN 1 ELSE 2 END, u.display_name
    """, (viewer_id, room_id)).fetchall()]


def _emit_room_event(room_id, event, payload):
    """방 멤버 전체에 이벤트 브로드캐스트."""
    socketio.emit(event, payload, to=f"room_{room_id}")


@app.route("/api/rooms/<int:room_id>", methods=["DELETE"])
@login_required
def api_room_delete(room_id):
    """채널(대화방) 삭제 (대표 지시 2026-05-21).
    관리자(ceo)는 모든 채널, 팀장은 본인이 만든 채널만. 자동 채널은 삭제 불가."""
    me = current_user()
    db = get_db()
    room = db.execute("SELECT id, type, channel_scope, name, created_by FROM rooms WHERE id=?", (room_id,)).fetchone()
    if not room:
        return jsonify({"error": "방을 찾을 수 없습니다."}), 404
    if room["channel_scope"]:
        return jsonify({"error": "자동 채널(KNK WORLD/본사/베트남)은 삭제할 수 없습니다."}), 400
    if room["type"] != "channel":
        return jsonify({"error": "채널만 삭제할 수 있습니다."}), 400
    # 권한: 관리자(ceo) 또는 본인이 만든 채널의 생성권한자(임원·팀장·법인장 등) (대표 지시 2026-05-24)
    if me["role"] != "ceo":
        if not (_can_create_channel(me) and room["created_by"] == me["id"]):
            return jsonify({"error": "채널은 관리자 또는 채널을 만든 사람만 삭제할 수 있습니다."}), 403
    name = room["name"]
    _emit_room_event(room_id, "room_deleted", {"room_id": room_id, "name": name})
    # 의존 데이터 명시 정리 (FK CASCADE 미적용 대비) — 전체공지 삭제와 동일 패턴
    db.execute("DELETE FROM messages WHERE room_id=?", (room_id,))
    db.execute("DELETE FROM room_members WHERE room_id=?", (room_id,))
    db.execute("DELETE FROM rooms WHERE id=?", (room_id,))
    db.commit()
    return jsonify({"ok": True, "deleted": name})


@app.route("/api/rooms/<int:room_id>/purge", methods=["POST"])
@login_required
def api_room_purge_direct(room_id):
    """1:1 대화방 '완전 삭제' — 소유자(대표 본인) 전용. (대표 지시 2026-06-05)

    테스트로 만든 1:1 방의 메시지·첨부·방 자체를 영구 삭제한다. 방을 지우면 direct_key 도
    사라져, 같은 사람과 다시 1:1 을 열면 _find_or_create_direct 가 새 방을 만든다(잔여 대화 0).
    양쪽 사용자 화면에서 사라진다(room_deleted emit).

    ※ 권한은 _is_owner(소유자=OWNER_USERNAME) 단 한 명. role='ceo' 인 다른 관리자도 사용 불가.
    ※ 1:1(direct) 방에만 적용 — 그룹·프로젝트·채널은 거부(기존 채널 삭제 endpoint 사용).
    """
    me = current_user()
    # 권한: 소유자(대표 본인)만 — 다른 ceo(관리자)도 불가
    # ※ current_user() 는 sqlite3.Row 라 .get() 없음 → 반드시 대괄호 접근 (me["username"])
    if not _is_owner(me["username"]):
        return jsonify({"error": "이 기능은 소유자(대표) 본인만 사용할 수 있습니다."}), 403
    db = get_db()
    room = db.execute("SELECT id, type, name FROM rooms WHERE id=?", (room_id,)).fetchone()
    if not room:
        return jsonify({"error": "방을 찾을 수 없습니다."}), 404
    if room["type"] != "direct":
        return jsonify({"error": "1:1 대화방만 완전 삭제할 수 있습니다."}), 400
    # 본인이 참여한 방만 (남의 1:1 방 임의 삭제 방지)
    if not db.execute("SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
                      (room_id, me["id"])).fetchone():
        return jsonify({"error": "본인이 참여한 1:1 방만 삭제할 수 있습니다."}), 403

    # 1) 양쪽 화면에서 즉시 사라지게 — 실제 삭제 전에 알림 (소켓룸 멤버 모두 수신)
    _emit_room_event(room_id, "room_deleted", {"room_id": room_id, "name": room["name"]})

    # 2) 첨부 파일 디스크 정리 (uploads/<room_id>/ 통째)
    try:
        import shutil
        _updir = os.path.join(UPLOAD_DIR, str(room_id))
        if os.path.isdir(_updir):
            shutil.rmtree(_updir, ignore_errors=True)
    except Exception as _e:
        print(f"[purge] 첨부 삭제 실패 room {room_id}: {_e}", flush=True)

    # 3) 메시지 딸림 데이터(message_id FK) — 메시지 삭제 전에 먼저 (best-effort)
    for _tbl in ("message_reactions", "message_acks", "message_stars", "message_translations"):
        try:
            db.execute(
                f"DELETE FROM {_tbl} WHERE message_id IN (SELECT id FROM messages WHERE room_id=?)",
                (room_id,),
            )
        except Exception:
            pass
    # 4) room_id FK 데이터 — best-effort (해당 테이블/행 없어도 안전하게 통과)
    for _tbl in ("messages", "room_members", "room_notices", "mentions", "requests",
                 "attachment_versions", "ai_summaries", "room_aliases",
                 "guest_view_aliases", "guest_invites"):
        try:
            db.execute(f"DELETE FROM {_tbl} WHERE room_id=?", (room_id,))
        except Exception:
            pass
    # 5) 방 자체 삭제 → direct_key 소멸 → 다시 1:1 열면 깨끗한 새 방
    db.execute("DELETE FROM rooms WHERE id=?", (room_id,))
    db.commit()
    print(f"[purge] 1:1 방 완전삭제 room={room_id} by owner uid={me['id']}", flush=True)
    return jsonify({"ok": True, "deleted": room["name"]})


@app.route("/api/rooms/<int:room_id>/members", methods=["GET"])
@login_required
def api_room_members(room_id):
    """방 멤버 목록 + 각 역할 + 내 별명 + name_locked."""
    me = current_user()
    db = get_db()
    if not _my_room_role(db, room_id, me["id"]):
        abort(403)
    room = db.execute(
        "SELECT id, name, type, created_by, name_locked, retention_days, invite_policy, avatar_url FROM rooms WHERE id=?",
        (room_id,),
    ).fetchone()
    if not room:
        abort(404)
    alias = db.execute("SELECT alias FROM room_aliases WHERE room_id=? AND user_id=?",
                       (room_id, me["id"])).fetchone()
    return jsonify({
        "room": {
            "id": room["id"], "name": room["name"], "type": room["type"],
            "created_by": room["created_by"], "name_locked": bool(room["name_locked"]),
            "retention_days": room["retention_days"],
            "invite_policy": room["invite_policy"] or "all",
            # 방 설정 화면의 '아이콘 사진 변경/제거' 버튼 표시에 필요 (대표 지시 2026-05-20)
            "avatar_url": room["avatar_url"],
        },
        "members": _room_members_full(db, room_id, me["id"]),
        "my_alias": alias["alias"] if alias else None,
        "my_role": _my_room_role(db, room_id, me["id"]),
    })


@app.route("/api/rooms/<int:room_id>/alias", methods=["POST", "DELETE"])
@login_required
def api_room_alias(room_id):
    """내 화면에서만 보이는 방 별명. name_locked=1 이면 거부."""
    me = current_user()
    db = get_db()
    room = db.execute("SELECT name_locked, type FROM rooms WHERE id=?", (room_id,)).fetchone()
    if not room:
        abort(404)
    if not _my_room_role(db, room_id, me["id"]):
        abort(403)
    if room["type"] == "direct":
        return jsonify({"error": "1:1 방은 별명 설정 불가"}), 400
    if room["name_locked"]:
        return jsonify({"error": "이 방은 이름 고정. 방장만 변경 가능"}), 400
    now = datetime.now(timezone.utc).isoformat()
    if request.method == "DELETE":
        db.execute("DELETE FROM room_aliases WHERE room_id=? AND user_id=?", (room_id, me["id"]))
        db.commit()
        return jsonify({"ok": True, "alias": None})
    data = request.get_json(silent=True) or {}
    alias = (data.get("alias") or "").strip()
    if not alias:
        db.execute("DELETE FROM room_aliases WHERE room_id=? AND user_id=?", (room_id, me["id"]))
        db.commit()
        return jsonify({"ok": True, "alias": None})
    if len(alias) > 50:
        return jsonify({"error": "별명은 50자 이내"}), 400
    db.execute("""
        INSERT INTO room_aliases (user_id, room_id, alias, updated_at) VALUES (?,?,?,?)
        ON CONFLICT(user_id, room_id) DO UPDATE SET alias=excluded.alias, updated_at=excluded.updated_at
    """, (me["id"], room_id, alias, now))
    db.commit()
    return jsonify({"ok": True, "alias": alias})


@app.route("/api/rooms/<int:room_id>/guest_alias", methods=["POST", "DELETE"])
@login_required
def api_guest_view_alias(room_id):
    """게스트(외부) '대화방 표현' — 보는 사람 본인 화면에서만 보이는 개인 별칭. (대표 지시 2026-05-30)
       body: {guest_user_id, alias}. 빈 alias 또는 DELETE → 초기화.
       직원은 다국어 변환이 되지만 게스트는 안 돼서, 보는 사람이 알아보기 쉽게 직접 지정."""
    me = current_user()
    db = get_db()
    # 보는 사람은 이 방 멤버여야 함 (게스트 본인은 방설정 접근 불가 → 사실상 직원만)
    if not _my_room_role(db, room_id, me["id"]):
        abort(403)
    data = request.get_json(silent=True) or {}
    try:
        guest_id = int(data.get("guest_user_id") or 0)
    except Exception:
        guest_id = 0
    if not guest_id:
        return jsonify({"error": "대상 사용자가 지정되지 않았습니다."}), 400
    # 대상이 이 방의 '게스트' 멤버인지 확인 (외부 사용자에게만 허용)
    tgt = db.execute("""
        SELECT COALESCE(u.is_guest,0) AS is_guest
          FROM room_members rm JOIN users u ON u.id = rm.user_id
         WHERE rm.room_id=? AND rm.user_id=?
    """, (room_id, guest_id)).fetchone()
    if not tgt:
        return jsonify({"error": "이 방의 멤버가 아닙니다."}), 404
    if not int(tgt["is_guest"] or 0):
        return jsonify({"error": "외부(고객) 사용자에게만 설정할 수 있습니다."}), 400
    now = datetime.now(timezone.utc).isoformat()
    if request.method == "DELETE":
        db.execute("DELETE FROM guest_view_aliases WHERE viewer_user_id=? AND guest_user_id=?", (me["id"], guest_id))
        db.commit()
        return jsonify({"ok": True, "alias": None})
    alias = (data.get("alias") or "").strip()
    if not alias:
        db.execute("DELETE FROM guest_view_aliases WHERE viewer_user_id=? AND guest_user_id=?", (me["id"], guest_id))
        db.commit()
        return jsonify({"ok": True, "alias": None})
    if len(alias) > 40:
        return jsonify({"error": "표현은 40자 이내로 입력하세요."}), 400
    db.execute("""
        INSERT INTO guest_view_aliases (viewer_user_id, guest_user_id, room_id, alias, updated_at)
        VALUES (?,?,?,?,?)
        ON CONFLICT(viewer_user_id, guest_user_id)
        DO UPDATE SET alias=excluded.alias, room_id=excluded.room_id, updated_at=excluded.updated_at
    """, (me["id"], guest_id, room_id, alias, now))
    db.commit()
    return jsonify({"ok": True, "alias": alias})


@app.route("/api/rooms/<int:room_id>/name", methods=["PATCH"])
@login_required
def api_room_rename(room_id):
    """방 이름 변경 — 방장 또는 관리자(ceo). body: {name, name_locked?}"""
    me = current_user()
    db = get_db()
    role = _my_room_role(db, room_id, me["id"])
    if role != 'host' and me["role"] != 'ceo':
        return jsonify({"error": "방장 또는 관리자만 가능"}), 403
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "이름은 비어있을 수 없음"}), 400
    if len(name) > 100:
        return jsonify({"error": "이름은 100자 이내"}), 400
    name_locked = 1 if data.get("name_locked") else 0
    # 방 이름 AI 자동 번역 취소 (대표 지시 2026-05-28 번복) — 비용 우려
    db.execute("UPDATE rooms SET name=?, name_locked=? WHERE id=?",
               (name, name_locked, room_id))
    # name_locked=1 이면 기존 alias 모두 삭제 (방장이 강제)
    if name_locked:
        db.execute("DELETE FROM room_aliases WHERE room_id=?", (room_id,))
    db.commit()
    # 방 이름 바뀌면 번역 갱신 (대표 지시 2026-06-05) — ko/vi/en + 중국 고객 있으면 zh. 동기·실패무해.
    _retranslate_room(db, room_id)
    now = datetime.now(timezone.utc).isoformat()
    # 관리자(실제 방장 아님)의 강제 변경이면 누가 했는지 남겨 방장에게 통보 (대표 지시 2026-06-04)
    # role 은 effective(채널에선 ceo도 host) 라, 실제 방장 여부는 room_members.role 직접 확인
    _rm = db.execute("SELECT role FROM room_members WHERE room_id=? AND user_id=?", (room_id, me["id"])).fetchone()
    if _rm and _rm["role"] == 'host':
        sys_text = f"방 이름이 [{name}] 으로 변경됨"
    else:
        sys_text = f"⚠️ 관리자({me['display_name']})님이 방 이름을 [{name}] (으)로 강제 변경했습니다 (방장 통보)"
    cur = db.execute(
        "INSERT INTO messages (room_id, user_id, content, kind, created_at) VALUES (?,?,?,?,?)",
        (room_id, me["id"], sys_text, "system", now),
    )
    db.commit()
    _emit_room_event(room_id, "room_renamed", {
        "room_id": room_id, "name": name, "name_locked": bool(name_locked),
        "by": me["display_name"],
    })
    _emit_room_event(room_id, "new_message", {
        "id": cur.lastrowid, "room_id": room_id, "user_id": me["id"],
        "display_name": me["display_name"], "avatar_color": me["avatar_color"],
        "content": sys_text, "kind": "system", "created_at": now,
    })
    return jsonify({"ok": True, "name": name, "name_locked": bool(name_locked)})


@app.route("/api/admin/retranslate_rooms", methods=["POST"])
@login_required
def api_admin_retranslate_rooms():
    """기존 방 이름 일괄 자동 번역 (관리자 전용). 청크 단위로 처리해 타임아웃 방지.
       body: {after_id?: int, limit?: int}. 응답: {processed, last_id, remaining, total}.
       프론트는 remaining==0 까지 after_id=last_id 로 반복 호출. (대표 지시 2026-06-05)"""
    me = current_user()
    if not _is_admin_user(me):
        return jsonify({"error": "권한 없음"}), 403
    db = get_db()
    if not _ai_translate_enabled(db):
        return jsonify({"error": "AI 번역이 꺼져 있습니다 (관리자 탭에서 켜세요)"}), 400
    data = request.get_json(silent=True) or {}
    try:
        after_id = int(data.get("after_id") or 0)
    except Exception:
        after_id = 0
    try:
        limit = int(data.get("limit") or 4)
    except Exception:
        limit = 4
    limit = max(1, min(limit, 12))
    rows = db.execute(
        """SELECT id FROM rooms
            WHERE id > ? AND type NOT IN ('direct','self')
              AND name IS NOT NULL AND TRIM(name) <> ''
            ORDER BY id ASC LIMIT ?""",
        (after_id, limit),
    ).fetchall()
    last_id = after_id
    for r in rows:
        _retranslate_room(db, r["id"])
        last_id = r["id"]
    remaining = db.execute(
        """SELECT COUNT(*) AS c FROM rooms
            WHERE id > ? AND type NOT IN ('direct','self')
              AND name IS NOT NULL AND TRIM(name) <> ''""",
        (last_id,),
    ).fetchone()["c"]
    total = db.execute(
        """SELECT COUNT(*) AS c FROM rooms
            WHERE type NOT IN ('direct','self') AND name IS NOT NULL AND TRIM(name) <> ''"""
    ).fetchone()["c"]
    return jsonify({"ok": True, "processed": len(rows), "last_id": last_id,
                    "remaining": remaining, "total": total})


@app.route("/api/rooms/<int:room_id>/retention", methods=["GET", "PUT"])
@login_required
def api_room_retention(room_id):
    """방별 메시지 자동 삭제 일수 설정 (WhatsApp 식).
    GET: 현재 설정 반환. PUT: body {retention_days: int|null}.
    null=영구 보존(글로벌 정책만 적용), 1=24시간, 7=1주, 30=30일, 90=90일.
    방장만 변경 가능 — 멤버는 GET 만."""
    me = current_user()
    db = get_db()
    # 멤버 확인
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (room_id, me["id"]),
    ).fetchone():
        return jsonify({"error": "방 멤버 아님"}), 403
    if request.method == "GET":
        row = db.execute("SELECT retention_days FROM rooms WHERE id=?", (room_id,)).fetchone()
        return jsonify({"retention_days": row["retention_days"] if row else None})
    # PUT
    if _my_room_role(db, room_id, me["id"]) != 'host':
        return jsonify({"error": "방장만 변경 가능"}), 403
    data = request.get_json(silent=True) or {}
    raw = data.get("retention_days")
    if raw is None:
        rd = None
    else:
        try:
            rd = int(raw)
            if rd not in (1, 7, 30, 90):
                return jsonify({"error": "retention_days 는 1/7/30/90 또는 null"}), 400
        except (ValueError, TypeError):
            return jsonify({"error": "retention_days 형식 오류"}), 400
    db.execute("UPDATE rooms SET retention_days=? WHERE id=?", (rd, room_id))
    db.commit()
    # 시스템 메시지로 변경 사실 기록
    now = datetime.now(timezone.utc).isoformat()
    label = "영구 보존" if rd is None else (
        "24시간 후" if rd == 1 else f"{rd}일 후"
    )
    sys_text = f"메시지 자동 삭제 정책: {label}"
    cur = db.execute(
        "INSERT INTO messages (room_id, user_id, content, kind, created_at) VALUES (?,?,?,?,?)",
        (room_id, me["id"], sys_text, "system", now),
    )
    db.commit()
    _emit_room_event(room_id, "new_message", {
        "id": cur.lastrowid, "room_id": room_id, "user_id": me["id"],
        "display_name": me["display_name"], "avatar_color": me["avatar_color"],
        "content": sys_text, "kind": "system", "created_at": now,
    })
    _emit_room_event(room_id, "room_retention_changed", {
        "room_id": room_id, "retention_days": rd,
    })
    return jsonify({"ok": True, "retention_days": rd})


# ---------- 사용자 상태 + 캘린더 동기화 ----------
# 대표 지시 2026-05-19: '온라인'→'회사' 라벨 변경, 방해금지 제거,
# 해외출장·국내출장·휴가 신규 추가.
# (status 키 'online' 은 코드 호환 위해 유지하고 라벨만 변경)
VALID_STATUSES = ("online", "away", "busy", "meeting", "external", "overseas", "domestic", "vacation", "offwork", "offline")
STATUS_LABEL_KO = {
    "online":   "💻 컴퓨터",
    "mobile":   "📱 휴대폰",
    "away":     "🪑 자리비움",
    "busy":     "🔴 바쁨",
    "meeting":  "🤝 회의 중",
    "external": "🚗 외근",
    "overseas": "✈️ 해외출장",
    "domestic": "🚆 국내출장",
    "vacation": "🌴 휴가",
    "offwork":  "🏠 퇴근",
    "offline":  "⚫ 오프라인",
}


def _get_user_status(uid):
    """사용자의 현재 상태 dict 반환. 없으면 online 디폴트."""
    db = get_db()
    row = db.execute("SELECT * FROM user_statuses WHERE user_id=?", (uid,)).fetchone()
    if not row:
        return {
            "user_id": uid, "status": "online", "custom_text": None,
            "emoji": None, "until_at": None, "auto_set": 0,
            "label": STATUS_LABEL_KO["online"],
        }
    d = dict(row)
    # until_at 지났으면 online 으로 간주 (영구 적용은 별도 cron 에서)
    if d.get("until_at"):
        try:
            from datetime import datetime as _dt
            ua = _dt.fromisoformat(d["until_at"].replace("Z", "+00:00"))
            if ua < datetime.now(timezone.utc):
                d["status"] = "online"
                d["until_at"] = None
        except Exception:
            pass
    d["label"] = STATUS_LABEL_KO.get(d["status"], d["status"])
    return d


def _computed_user_status(uid):
    """uid 의 '표시용' 상태 dict — 자동표시 규칙 적용 (다른 사람에게 보이는 값).
      · 접속 중 + 기본('online') → PC 연결 있으면 'online'(가능), 휴대폰만이면 'mobile'(📱 휴대폰)
      · 접속 중 + 특수상태(회의중·외근 등) → 그대로 (사용자 선택)
      · 미접속 → 푸시 있으면 'mobile', 없으면 'offline'
    호출 측에서 app context 보장 필요 (get_db 사용)."""
    base = _get_user_status(uid)
    status = base["status"]
    # 자리비움(away) 폐지 — 옛 데이터·잔재는 자동(online)으로 강등 (대표 지시 2026-05-25)
    if status == "away":
        status = "online"
    # 🏠 퇴근(offwork) — 수동 상태. 접속/미접속 무관하게 그대로 유지 (대표 지시 2026-05-25).
    #  PC 로 다시 메신저에 접속하면 on_presence 에서 자동(online) 복귀시킴.
    if status == "offwork":
        base["status"] = "offwork"
        base["label"] = STATUS_LABEL_KO["offwork"]
        base["at_office"] = False   # 퇴근이면 회사망 표기 의미 없음
        return base
    # 🤝회의·🚗외근·✈️해외·🚆국내출장·🌴휴가 — 기한(until_at) 전까지는 접속/미접속 무관하게 유지
    #   (만료된 until_at 은 _get_user_status 에서 이미 online 으로 강등됨). 퇴근(offwork)과 같은 '유지' 처리. (대표 지시 2026-06-06)
    if status in ("meeting", "external", "overseas", "domestic", "vacation") and base.get("until_at"):
        base["status"] = status
        base["label"] = STATUS_LABEL_KO.get(status, status)
        base["at_office"] = bool(_user_is_online(uid) and _user_at_office(uid))
        return base
    if _user_is_online(uid):
        if status == "online":
            # 자동표시 — 컴퓨터로 접속해 있으면 '💻 컴퓨터'(창 안 봐도), 휴대폰만이면 '📱 휴대폰'.
            status = "online" if _user_has_pc_connection(uid) else "mobile"
        # else: 특수상태(회의중·외근·휴가·바쁨·출장 등)는 사용자 선택이라 그대로 — 우선순위
    else:
        # 미접속 — '실제 휴대폰' 알림 등록이 있을 때만 '📱 휴대폰'. PC 등록만 남았거나 없으면 '⚫ 오프라인'.
        #   (PC 알림 등록 잔재가 휴대폰으로 오표시되던 문제 수정 — 대표 지시 2026-05-25)
        has_mobile_push = False
        try:
            for r in get_db().execute(
                "SELECT user_agent FROM push_subscriptions WHERE user_id=?", (uid,)
            ).fetchall():
                if _device_type_from_ua(r["user_agent"]) == "mobile":
                    has_mobile_push = True
                    break
        except Exception:
            has_mobile_push = False
        status = "mobile" if has_mobile_push else "offline"
        base["custom_text"] = None
        base["emoji"] = None
        base["until_at"] = None
    base["status"] = status
    base["label"] = STATUS_LABEL_KO.get(status, status)
    base["at_office"] = bool(_user_is_online(uid) and _user_at_office(uid))
    return base


# 직전에 broadcast 한 표시상태 캐시 — 동일 상태 중복 emit 방지 (uid -> status). eventlet 단일스레드 가정.
_last_status_bcast = {}

# 🏠 퇴근(offwork) 자동복귀 대기 집합 — '완전 미접속 → 재접속' 한 사용자가 퇴근 상태면 여기 등록.
#  이후 presence 가 ① PC(데스크탑) 이거나 ② 회사 인터넷망(등록된 회사 IP) 접속으로 확인되면
#  자동(online: 컴퓨터/휴대폰)으로 복귀시킴.
#  · 휴대폰으로만 잠깐 접속(저녁에 집에서 확인 등)하면 복귀 안 함 → 퇴근 유지.
#  · 단, 휴대폰이라도 '회사망'에 접속했으면 출근으로 자동 전환 (사무실 출근 신호).
#  · 책상에서 퇴근 누르고 그대로 켜둔 경우는 '재접속'이 아니라 등록 안 됨 → 퇴근 유지(즉시 표기).
#  (대표 지시 2026-05-25: 컴퓨터로 다시 접속 OR 회사망 접속 시 자동 출근)
_offwork_clear_pending = set()

def _broadcast_status_if_changed(uid):
    """uid 의 표시 상태를 계산해, 직전 broadcast 와 다를 때만 user_status_changed emit.
    호출 측에서 app context 보장 필요."""
    try:
        s = _computed_user_status(uid)
        key = f"{s['status']}|{1 if s.get('at_office') else 0}"   # 상태 OR 회사망 여부가 바뀌면 broadcast
        if _last_status_bcast.get(uid) == key:
            return
        _last_status_bcast[uid] = key
        socketio.emit("user_status_changed", {
            "user_id": uid, "status": s["status"],
            "custom_text": s.get("custom_text"), "emoji": s.get("emoji"),
            "label": s["label"], "at_office": bool(s.get("at_office")),
            "until_at": s.get("until_at"),   # 기한(회의·외근 시각 / 출장·휴가 날짜) 실시간 반영 (대표 지시 2026-06-06)
        })
    except Exception as e:
        print(f"[status] broadcast 실패: {e}")


@app.route("/api/me/status", methods=["GET", "PUT"])
@login_required
def api_me_status():
    """내 상태 조회 / 변경.
    GET: 현재 상태 + 모든 가능한 상태 enum.
    PUT body: {status, custom_text?, emoji?, until_at?}"""
    me = current_user()
    if request.method == "GET":
        return jsonify({
            "current": _get_user_status(me["id"]),
            "options": [{"value": k, "label": v} for k, v in STATUS_LABEL_KO.items()],
        })
    db = get_db()
    data = request.get_json(silent=True) or {}
    status = data.get("status") or "online"
    if status not in VALID_STATUSES:
        return jsonify({"error": f"status 는 {VALID_STATUSES} 중 하나"}), 400
    custom_text = (data.get("custom_text") or "").strip()[:80] or None
    emoji = (data.get("emoji") or "").strip()[:8] or None
    until_at = data.get("until_at") or None
    now = datetime.now(timezone.utc).isoformat()
    db.execute("""
        INSERT INTO user_statuses (user_id, status, custom_text, emoji, until_at, auto_set, updated_at)
        VALUES (?,?,?,?,?,0,?)
        ON CONFLICT(user_id) DO UPDATE SET
            status=excluded.status, custom_text=excluded.custom_text,
            emoji=excluded.emoji, until_at=excluded.until_at,
            auto_set=0, updated_at=excluded.updated_at
    """, (me["id"], status, custom_text, emoji, until_at, now))
    db.commit()
    cur = _get_user_status(me["id"])  # 본인에게 돌려줄 값 = 본인이 고른 상태
    # 다른 사람에게는 자동표시 규칙(가능/휴대폰) 적용한 값으로 broadcast
    disp = _computed_user_status(me["id"])
    _last_status_bcast[me["id"]] = f"{disp['status']}|{1 if disp.get('at_office') else 0}"
    socketio.emit("user_status_changed", {
        "user_id": me["id"], "status": disp["status"],
        "custom_text": disp.get("custom_text"), "emoji": disp.get("emoji"),
        "label": disp["label"], "at_office": bool(disp.get("at_office")),
        "until_at": disp.get("until_at"),   # 기한(회의·외근 시각 / 출장·휴가 날짜) 실시간 반영 (대표 지시 2026-06-06)
    })
    return jsonify({"ok": True, "current": cur})


@app.route("/api/users/statuses", methods=["GET"])
@login_required
def api_users_statuses():
    """전체 사용자 현재 상태 일괄 조회 — 사이드바·메시지 아바타 색점용."""
    db = get_db()
    me = current_user()
    my_uid = me["id"] if me else None
    rows = db.execute("""
        SELECT u.id AS user_id, u.display_name, u.self_avatar,
               COALESCE(us.status, 'online') AS status,
               us.custom_text, us.emoji, us.until_at
          FROM users u
          LEFT JOIN user_statuses us ON us.user_id = u.id
         WHERE u.active = 1
    """).fetchall()
    # '실제 휴대폰' 푸시 구독이 있는 사용자 집합 = '📱 휴대폰'. PC(데스크톱) 구독만 있으면 제외(→오프라인).
    # (대표 지시 2026-05-20: 오프라인 vs 휴대폰 구분 / 2026-05-25: PC 푸시 잔재 휴대폰 오표시 수정)
    try:
        push_uids = set()
        for r in db.execute("SELECT user_id, user_agent FROM push_subscriptions").fetchall():
            if _device_type_from_ua(r["user_agent"]) == "mobile":
                push_uids.add(r["user_id"])
    except Exception:
        push_uids = set()
    out = []
    now_iso = datetime.now(timezone.utc).isoformat()
    for r in rows:
        d = dict(r)
        # ★ 본인은 항상 online 으로 간주 (API 호출 자체가 활성 세션 증거)
        #   SocketIO connect 전 API 호출 race condition 회피 (대표 지시 2026-05-19).
        is_self = (my_uid is not None and d["user_id"] == my_uid)
        uid = d["user_id"]
        online = _user_is_online(uid)
        # 만료된 until_at 은 기본(online)으로 강등 (DB 미반영, 표시상만)
        if d.get("until_at") and d["until_at"] < now_iso:
            d["status"] = "online"
            d["until_at"] = None
        set_status = d["status"]  # 사용자가 설정한 상태 (없으면 'online')
        if set_status == "away":
            set_status = "online"   # 자리비움 폐지 — 자동(컴퓨터/휴대폰)으로 강등 (대표 지시 2026-05-25)
        # 🏠 퇴근(offwork) — 접속/미접속 무관하게 그대로 유지. PC 재접속 시 on_presence 에서 자동 복귀. (대표 지시 2026-05-25)
        if set_status == "offwork":
            d["status"] = "offwork"
            d["label"] = STATUS_LABEL_KO["offwork"]
            d["at_office"] = False
            out.append(d)
            continue
        # 🤝회의·🚗외근·✈️해외·🚆국내출장·🌴휴가 — 기한(until_at) 전까지는 접속/미접속 무관하게 유지
        #   (만료된 until_at 은 위에서 이미 online 으로 강등). 휴가·출장은 보통 미접속이라 이 유지가 핵심. (대표 지시 2026-06-06)
        if set_status in ("meeting", "external", "overseas", "domestic", "vacation") and d.get("until_at"):
            d["status"] = set_status
            d["label"] = STATUS_LABEL_KO.get(set_status, set_status)
            d["at_office"] = bool((online or is_self) and _user_at_office(uid))
            out.append(d)
            continue
        # 상태 자동표시: 접속+기본 → 컴퓨터(PC 연결)/휴대폰(휴대폰만), 미접속 → 휴대폰(푸시)/오프라인
        #   · 특수상태(회의중·외근·휴가·바쁨·출장)는 사용자 선택이라 그대로
        if online or is_self:
            if set_status == "online":
                if is_self and not online:
                    d["status"] = "online"   # 본인 로드 직후 소켓 미연결 race → 기본 컴퓨터
                else:
                    d["status"] = "online" if _user_has_pc_connection(uid) else "mobile"
            else:
                d["status"] = set_status   # 수동 특수상태 우선
        else:
            # 미접속 — 알림 가능(푸시 있음)이면 휴대폰, 로그아웃이면 오프라인
            d["status"] = "mobile" if uid in push_uids else "offline"
            d["custom_text"] = None
            d["emoji"] = None
            d["until_at"] = None
        d["label"] = STATUS_LABEL_KO.get(d["status"], d["status"])
        d["at_office"] = bool((online or is_self) and _user_at_office(uid))   # 🏢 회사망 접속 여부
        out.append(d)
    return jsonify(out)


def _rebroadcast_all_status():
    """현재 접속 중인 모든 사용자 상태 재계산·broadcast (회사망 목록 변경 후 즉시 반영용)."""
    try:
        for _uid in list(_user_connections.keys()):
            _broadcast_status_if_changed(_uid)
    except Exception as e:
        print(f"[office] rebroadcast 실패: {e}")


@app.route("/api/office_networks", methods=["GET", "POST"])
@login_required
def api_office_networks():
    """회사망(사무실) 공인 IP 관리 — 관리자 전용.
    GET: 현재 접속 IP + 등록 목록. POST: 현재 접속 IP(또는 body.ip)를 회사망으로 등록."""
    me = current_user()
    if me["role"] != "ceo":
        return jsonify({"error": "관리자 전용"}), 403
    db = get_db()
    if request.method == "POST":
        data = request.get_json(silent=True) or {}
        ip = (data.get("ip") or _real_client_ip() or "").strip()
        label = (data.get("label") or "").strip()[:60] or None
        if not ip:
            return jsonify({"error": "접속 IP 를 확인할 수 없습니다."}), 400
        now = datetime.now(timezone.utc).isoformat()
        try:
            db.execute("INSERT OR IGNORE INTO office_networks (ip, label, created_at, created_by) VALUES (?,?,?,?)",
                       (ip, label, now, me["id"]))
            db.commit()
        except Exception as e:
            return jsonify({"error": str(e)}), 500
        _load_office_ips()
        _rebroadcast_all_status()   # 등록 즉시 회사망 접속자들 🏢 표시
    rows = db.execute("SELECT id, ip, label, created_at FROM office_networks ORDER BY id").fetchall()
    return jsonify({"current_ip": _real_client_ip(), "networks": [dict(r) for r in rows]})


@app.route("/api/office_networks/<int:nid>", methods=["DELETE"])
@login_required
def api_office_network_delete(nid):
    me = current_user()
    if me["role"] != "ceo":
        return jsonify({"error": "관리자 전용"}), 403
    db = get_db()
    db.execute("DELETE FROM office_networks WHERE id=?", (nid,))
    db.commit()
    _load_office_ips()
    _rebroadcast_all_status()
    return jsonify({"ok": True})


@app.route("/api/admin/notif_off", methods=["GET"])
@login_required
def api_admin_notif_off():
    """알림(푸시) 미설정 직원 목록 — 관리자 전용. (회사 '알림 필수' 정책 집행용)
    푸시 구독이 하나도 없는 활성 사용자 = 실시간 알림 못 받음 = '알림 OFF'."""
    me = current_user()
    if me["role"] != "ceo":
        return jsonify({"error": "관리자 전용"}), 403
    db = get_db()
    rows = db.execute("""
        SELECT u.id, u.display_name, u.department, u.title,
               CASE WHEN ps.user_id IS NULL THEN 0 ELSE 1 END AS has_push
          FROM users u
          LEFT JOIN (SELECT DISTINCT user_id FROM push_subscriptions) ps ON ps.user_id = u.id
         WHERE u.active = 1 AND u.username != '_deleted_user'
         ORDER BY u.department, u.display_name
    """).fetchall()
    off = [dict(r) for r in rows if not r["has_push"]]
    return jsonify({"off": off, "total_active": len(rows), "off_count": len(off)})


@app.route("/api/me/calendar", methods=["GET", "POST"])
@login_required
def api_me_calendar():
    """내 캘린더 일정 — GET: 목록 / POST: 추가.
    POST body: {title, start_at, end_at, kind?}"""
    me = current_user()
    db = get_db()
    if request.method == "GET":
        # 향후 7일치만 (UI 부하 차단)
        rows = db.execute("""
            SELECT id, title, start_at, end_at, kind, applied
              FROM user_calendar_events
             WHERE user_id=? AND date(end_at) >= date('now', '-1 day')
             ORDER BY start_at ASC
             LIMIT 100
        """, (me["id"],)).fetchall()
        return jsonify([dict(r) for r in rows])
    data = request.get_json(silent=True) or {}
    title = (data.get("title") or "").strip()
    start_at = (data.get("start_at") or "").strip()
    end_at = (data.get("end_at") or "").strip()
    kind = data.get("kind") or "meeting"
    if kind not in ("meeting", "external", "busy"):
        kind = "meeting"
    if not title or not start_at or not end_at:
        return jsonify({"error": "title/start_at/end_at 필수"}), 400
    now = datetime.now(timezone.utc).isoformat()
    cur = db.execute("""
        INSERT INTO user_calendar_events (user_id, title, start_at, end_at, kind, applied, created_at)
        VALUES (?,?,?,?,?,0,?)
    """, (me["id"], title, start_at, end_at, kind, now))
    db.commit()
    return jsonify({"ok": True, "id": cur.lastrowid})


@app.route("/api/me/calendar/<int:event_id>", methods=["DELETE"])
@login_required
def api_me_calendar_delete(event_id):
    me = current_user()
    db = get_db()
    db.execute("DELETE FROM user_calendar_events WHERE id=? AND user_id=?", (event_id, me["id"]))
    db.commit()
    return jsonify({"ok": True})


def _apply_calendar_status_transitions():
    """캘린더 일정 시작·종료에 따라 사용자 상태 자동 전환.
    /api/me/status PUT 으로 수동 변경되면 auto_set=0 으로 갱신돼 자동전환 차단.
    cron 또는 요청 hook 에서 주기 호출."""
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    try:
        now_iso = datetime.now(timezone.utc).isoformat()
        # 시작됐고 아직 적용 안 된 일정 → 회의 중 자동 설정
        starting = db.execute("""
            SELECT id, user_id, title, end_at, kind
              FROM user_calendar_events
             WHERE applied = 0 AND start_at <= ? AND end_at > ?
        """, (now_iso, now_iso)).fetchall()
        for ev in starting:
            # 현재 사용자 상태 — 수동 설정이면 건드리지 않음
            us = db.execute("SELECT status, auto_set FROM user_statuses WHERE user_id=?", (ev["user_id"],)).fetchone()
            if us and us["auto_set"] == 0 and us["status"] in ("dnd", "external"):
                # 수동으로 더 강한 상태 설정돼 있으면 그것을 존중
                pass
            else:
                new_status = "meeting" if ev["kind"] == "meeting" else (
                    "external" if ev["kind"] == "external" else "busy"
                )
                db.execute("""
                    INSERT INTO user_statuses (user_id, status, custom_text, until_at, auto_set, updated_at)
                    VALUES (?,?,?,?,1,?)
                    ON CONFLICT(user_id) DO UPDATE SET
                        status=excluded.status, custom_text=excluded.custom_text,
                        until_at=excluded.until_at, auto_set=1, updated_at=excluded.updated_at
                """, (ev["user_id"], new_status, ev["title"], ev["end_at"], now_iso))
            db.execute("UPDATE user_calendar_events SET applied=1 WHERE id=?", (ev["id"],))
        # 종료된 일정 (applied=1) → online 복귀 (auto_set=1 인 경우만)
        ending = db.execute("""
            SELECT id, user_id FROM user_calendar_events
             WHERE applied = 1 AND end_at <= ?
        """, (now_iso,)).fetchall()
        for ev in ending:
            us = db.execute("SELECT status, auto_set FROM user_statuses WHERE user_id=?", (ev["user_id"],)).fetchone()
            if us and us["auto_set"] == 1:
                db.execute("""
                    UPDATE user_statuses
                       SET status='online', custom_text=NULL, until_at=NULL, auto_set=0, updated_at=?
                     WHERE user_id=?
                """, (now_iso, ev["user_id"]))
            db.execute("UPDATE user_calendar_events SET applied=2 WHERE id=?", (ev["id"],))
        db.commit()
    finally:
        db.close()


# 캘린더 자동 전환을 60초마다 백그라운드에서 실행
_cal_thread_started = False
def _start_calendar_worker():
    global _cal_thread_started
    if _cal_thread_started:
        return
    _cal_thread_started = True
    def _loop():
        import time as _t
        while True:
            try:
                _apply_calendar_status_transitions()
            except Exception as e:
                print(f"[calendar worker] error: {e}")
            _t.sleep(60)
    import threading as _th
    _th.Thread(target=_loop, daemon=True).start()


# ============================================================
# 🛡️ 자동 백업 워커 (대표 지시 2026-05-19)
#   매일 새벽 3시 (KST) deploy/backup.sh 실행. cron 별도 설정 불필요.
#   threading.Thread daemon — eventlet 환경에서도 안전 (subprocess 사용).
# ============================================================
_backup_thread_started = False

def _start_backup_worker():
    global _backup_thread_started
    if _backup_thread_started:
        return
    _backup_thread_started = True
    import threading as _th, subprocess as _sp, time as _t
    from datetime import datetime as _dt, timezone as _tz, timedelta as _td
    def _loop():
        # KST = UTC+9
        kst = _tz(_td(hours=9))
        while True:
            try:
                now_kst = _dt.now(kst)
                # 다음 새벽 3시 (KST) 계산
                target = now_kst.replace(hour=3, minute=0, second=0, microsecond=0)
                if now_kst >= target:
                    target = target + _td(days=1)
                wait_seconds = (target - now_kst).total_seconds()
                # 한 번에 너무 오래 자지 않도록 — 6시간씩 끊어서 sleep
                while wait_seconds > 0:
                    chunk = min(wait_seconds, 6 * 3600)
                    _t.sleep(chunk)
                    wait_seconds -= chunk
                # 백업 실행
                backup_script = os.path.join(APP_DIR, "deploy", "backup.sh")
                if os.path.exists(backup_script):
                    try:
                        # 백업 로그는 data/backup.log 에 누적
                        log_path = os.path.join(APP_DIR, "data", "backup.log")
                        with open(log_path, "a", encoding="utf-8") as logf:
                            _sp.run(
                                ["bash", backup_script],
                                stdout=logf, stderr=_sp.STDOUT,
                                timeout=1800,  # 30분 타임아웃
                            )
                        print(f"[backup worker] 백업 완료 ({_dt.now(kst).isoformat()})", flush=True)
                    except _sp.TimeoutExpired:
                        print("[backup worker] 백업 타임아웃 (30분 초과)", flush=True)
                    except Exception as e:
                        print(f"[backup worker] 백업 실행 에러: {e}", flush=True)
                else:
                    print(f"[backup worker] 백업 스크립트 없음: {backup_script}", flush=True)
            except Exception as e:
                print(f"[backup worker] loop error: {e}", flush=True)
                _t.sleep(3600)  # 에러 시 1시간 후 재시도
    _th.Thread(target=_loop, daemon=True).start()


@app.route("/api/rooms/<int:room_id>/invite_policy", methods=["PUT"])
@login_required
def api_room_invite_policy(room_id):
    """방의 초대 권한 정책 변경 — 방장·PM 가능(권한 동일).
    body: {invite_policy: 'all' | 'host_only'}
    'all' = 모든 멤버 초대 가능 (기본). 'host_only' = 방장·PM만 초대 가능."""
    me = current_user()
    db = get_db()
    role = _my_room_role(db, room_id, me["id"])
    if role not in ('host', 'sub_host'):
        return jsonify({"error": "방장·PM만 변경 가능"}), 403
    data = request.get_json(silent=True) or {}
    policy = data.get("invite_policy")
    if policy not in ("all", "host_only"):
        return jsonify({"error": "invite_policy 는 'all' 또는 'host_only'"}), 400
    db.execute("UPDATE rooms SET invite_policy=? WHERE id=?", (policy, room_id))
    db.commit()
    now = datetime.now(timezone.utc).isoformat()
    label = "모든 멤버 가능" if policy == "all" else "방장·PM만 가능"
    sys_text = f"초대 권한: {label}"
    cur = db.execute(
        "INSERT INTO messages (room_id, user_id, content, kind, created_at) VALUES (?,?,?,?,?)",
        (room_id, me["id"], sys_text, "system", now),
    )
    db.commit()
    _emit_room_event(room_id, "room_invite_policy_changed", {
        "room_id": room_id, "invite_policy": policy,
    })
    _emit_room_event(room_id, "new_message", {
        "id": cur.lastrowid, "room_id": room_id, "user_id": me["id"],
        "display_name": me["display_name"], "avatar_color": me["avatar_color"],
        "content": sys_text, "kind": "system", "created_at": now,
    })
    return jsonify({"ok": True, "invite_policy": policy})


@app.route("/api/rooms/<int:room_id>/order", methods=["PUT"])
@login_required
def api_room_order(room_id):
    """사용자별 방 순서 조정.
    body: {action: 'pin' | 'unpin' | 'top' | 'bottom' | 'up' | 'down' | 'reset'}
    pin/unpin: rm.pinned 토글. top/bottom/up/down: order_value 조정. reset: 둘 다 NULL.
    각 사용자별 독립적으로 동작 (server 저장)."""
    me = current_user()
    db = get_db()
    rm = db.execute(
        "SELECT pinned, order_value FROM room_members WHERE room_id=? AND user_id=?",
        (room_id, me["id"]),
    ).fetchone()
    if not rm:
        return jsonify({"error": "방 멤버 아님"}), 403
    data = request.get_json(silent=True) or {}
    action = data.get("action")
    if action not in ("pin", "unpin", "top", "bottom", "up", "down", "reset"):
        return jsonify({"error": "action 은 pin/unpin/top/bottom/up/down/reset"}), 400

    # 본인이 속한 모든 방의 정렬 정보 (self 제외, pinned/non-pinned 그룹별)
    # pin 그룹 안에서 order_value 정렬, 일반 그룹 안에서 order_value 정렬
    is_pinned = bool(rm["pinned"])
    my_ov = rm["order_value"]

    def _peers_in_group(pinned_flag):
        """같은 그룹(pin/일반) 의 다른 방들 — order_value 가 설정된 것만."""
        return db.execute("""
            SELECT room_id, order_value FROM room_members
             WHERE user_id=? AND pinned=?
               AND room_id != ?
               AND order_value IS NOT NULL
             ORDER BY order_value ASC
        """, (me["id"], pinned_flag, room_id)).fetchall()

    def _all_in_group_sorted(pinned_flag):
        """그룹 안 모든 방 + 자동정렬 기준(last_at) 까지 합쳐 위→아래 순으로."""
        # api_rooms 와 동일한 정렬 로직: order_value 있는 것 먼저, 없는 것은 last_at desc
        return db.execute("""
            SELECT rm.room_id, rm.order_value,
                   (SELECT created_at FROM messages
                     WHERE room_id = rm.room_id ORDER BY id DESC LIMIT 1) AS last_at
              FROM room_members rm
              JOIN rooms r ON r.id = rm.room_id
             WHERE rm.user_id=? AND rm.pinned=? AND r.type != 'self'
             ORDER BY
                CASE WHEN rm.order_value IS NOT NULL THEN 0 ELSE 1 END,
                rm.order_value ASC,
                (last_at IS NULL), last_at DESC, rm.room_id DESC
        """, (me["id"], pinned_flag)).fetchall()

    now = datetime.now(timezone.utc).isoformat()

    if action == "reset":
        db.execute(
            "UPDATE room_members SET pinned=0, order_value=NULL WHERE room_id=? AND user_id=?",
            (room_id, me["id"]),
        )
    elif action == "pin":
        # 핀 그룹 최상단으로 (가장 작은 order_value 보다 더 작게)
        peers = _peers_in_group(1)
        new_ov = (peers[0]["order_value"] - 1.0) if peers else 0.0
        db.execute(
            "UPDATE room_members SET pinned=1, order_value=? WHERE room_id=? AND user_id=?",
            (new_ov, room_id, me["id"]),
        )
    elif action == "unpin":
        db.execute(
            "UPDATE room_members SET pinned=0 WHERE room_id=? AND user_id=?",
            (room_id, me["id"]),
        )
    elif action == "top":
        peers = _peers_in_group(1 if is_pinned else 0)
        new_ov = (peers[0]["order_value"] - 1.0) if peers else 0.0
        db.execute(
            "UPDATE room_members SET order_value=? WHERE room_id=? AND user_id=?",
            (new_ov, room_id, me["id"]),
        )
    elif action == "bottom":
        # 같은 그룹의 모든 방(자동정렬 포함) 중 가장 큰 order_value 보다 1 크게
        # order_value NULL 인 방들도 고려: 그것들은 last_at 기준 자동정렬이라
        # bottom 으로 보내려면 가장 큰 order_value + 1 보다 더 큰 값으로 설정
        peers = _peers_in_group(1 if is_pinned else 0)
        max_ov = peers[-1]["order_value"] if peers else 0.0
        # 자동정렬 방까지 아래로 보내려면 매우 큰 값 — 그러나 자동정렬은 항상 order_value 있는 것 다음에 표시되므로
        # max_ov + 1 이면 그룹의 명시 정렬된 방들 중 맨 아래. 자동정렬 방들은 더 아래에 표시됨.
        # → 사용자 의도(맨 아래)는 자동정렬 위치 너머. 따라서 9999 같이 매우 큰 값 사용.
        new_ov = 9999.0 + (max_ov or 0)
        db.execute(
            "UPDATE room_members SET order_value=? WHERE room_id=? AND user_id=?",
            (new_ov, room_id, me["id"]),
        )
    elif action in ("up", "down"):
        ordered = _all_in_group_sorted(1 if is_pinned else 0)
        # 현재 인덱스
        idx = next((i for i, x in enumerate(ordered) if x["room_id"] == room_id), -1)
        if idx < 0:
            return jsonify({"error": "정렬 위치 찾기 실패"}), 500
        if action == "up" and idx == 0:
            return jsonify({"ok": True, "no_change": True, "reason": "이미 최상단"})
        if action == "down" and idx == len(ordered) - 1:
            return jsonify({"ok": True, "no_change": True, "reason": "이미 최하단"})
        # 인접한 두 방 사이로 삽입 (REAL 사이값 = 평균)
        if action == "up":
            above = ordered[idx - 1]
            above_ov = above["order_value"]
            if idx >= 2:
                above_above = ordered[idx - 2]
                aa_ov = above_above["order_value"]
                if aa_ov is not None and above_ov is not None:
                    new_ov = (aa_ov + above_ov) / 2
                else:
                    new_ov = (above_ov if above_ov is not None else 0.0) - 1.0
            else:
                new_ov = (above_ov if above_ov is not None else 0.0) - 1.0
        else:  # down
            below = ordered[idx + 1]
            below_ov = below["order_value"]
            if idx + 2 < len(ordered):
                below_below = ordered[idx + 2]
                bb_ov = below_below["order_value"]
                if bb_ov is not None and below_ov is not None:
                    new_ov = (below_ov + bb_ov) / 2
                else:
                    new_ov = (below_ov if below_ov is not None else 0.0) + 1.0
            else:
                new_ov = (below_ov if below_ov is not None else 0.0) + 1.0
        db.execute(
            "UPDATE room_members SET order_value=? WHERE room_id=? AND user_id=?",
            (new_ov, room_id, me["id"]),
        )
    db.commit()
    return jsonify({"ok": True, "action": action})


def _ensure_self_room(uid):
    """사용자의 '나에게 보내기' 1인방 보장. 없으면 생성.
    '나와의 채팅' 모델.
    type='self' 이며 room_members 엔트리 1개(본인)만 가짐."""
    db = get_db()
    row = db.execute("""
        SELECT r.id FROM rooms r
         JOIN room_members rm ON rm.room_id = r.id
         WHERE r.type='self' AND rm.user_id=?
         LIMIT 1
    """, (uid,)).fetchone()
    if row:
        return row["id"]
    now = datetime.now(timezone.utc).isoformat()
    cur = db.execute(
        "INSERT INTO rooms (name, type, created_by, created_at, name_locked) VALUES (?,?,?,?,?)",
        ("📝 메모", "self", uid, now, 1),  # name_locked=1 → 별명 비활성
    )
    rid = cur.lastrowid
    db.execute(
        "INSERT INTO room_members (room_id, user_id, joined_at, role) VALUES (?,?,?,?)",
        (rid, uid, now, "host"),
    )
    db.execute(
        "INSERT INTO messages (room_id, user_id, content, kind, created_at) VALUES (?,?,?,?,?)",
        (rid, uid, "📝 메모·임시 파일·즉시 기록용 1인방입니다. 다른 사람은 볼 수 없습니다.", "system", now),
    )
    db.commit()
    return rid


@app.route("/api/me/self_room", methods=["GET"])
@login_required
def api_me_self_room():
    """'나에게 보내기' 1인방 ID 반환. 없으면 자동 생성."""
    me = current_user()
    rid = _ensure_self_room(me["id"])
    return jsonify({"room_id": rid})


@app.route("/api/rooms/<int:room_id>/members/<int:user_id>/role", methods=["POST"])
@login_required
def api_room_set_role(room_id, user_id):
    """멤버 역할 변경 — 방장·PM 가능(권한 동일). body: {role: 'sub_host'|'member'} (host 는 transfer 로만)"""
    me = current_user()
    db = get_db()
    if _my_room_role(db, room_id, me["id"]) not in ('host', 'sub_host'):
        return jsonify({"error": "방장·PM만 가능"}), 403
    if user_id == me["id"]:
        return jsonify({"error": "본인 역할은 변경할 수 없습니다"}), 400
    data = request.get_json(silent=True) or {}
    new_role = data.get("role")
    if new_role not in ('sub_host', 'member'):
        return jsonify({"error": "role 은 sub_host(PM) 또는 member"}), 400
    target = db.execute("SELECT role FROM room_members WHERE room_id=? AND user_id=?",
                        (room_id, user_id)).fetchone()
    if not target:
        return jsonify({"error": "멤버가 아님"}), 404
    # 방장(host) 의 역할은 이 경로로 변경 불가 (위임 transfer 로만) — PM 이 방장을 강등 못 하게 보호
    if target["role"] == 'host':
        return jsonify({"error": "방장 역할은 위임으로만 변경됩니다"}), 403
    db.execute("UPDATE room_members SET role=? WHERE room_id=? AND user_id=?",
               (new_role, room_id, user_id))
    db.commit()
    target_user = db.execute("SELECT display_name FROM users WHERE id=?", (user_id,)).fetchone()
    label = 'PM' if new_role == 'sub_host' else '일반 멤버'
    now = datetime.now(timezone.utc).isoformat()
    sys_text = f"[{target_user['display_name']}] 님이 {label} 으로 지정됨"
    cur = db.execute(
        "INSERT INTO messages (room_id, user_id, content, kind, created_at) VALUES (?,?,?,?,?)",
        (room_id, me["id"], sys_text, "system", now),
    )
    db.commit()
    _emit_room_event(room_id, "member_role_changed", {
        "room_id": room_id, "user_id": user_id, "role": new_role,
    })
    _emit_room_event(room_id, "new_message", {
        "id": cur.lastrowid, "room_id": room_id, "user_id": me["id"],
        "display_name": me["display_name"], "avatar_color": me["avatar_color"],
        "content": sys_text, "kind": "system", "created_at": now,
    })
    return jsonify({"ok": True})


@app.route("/api/rooms/<int:room_id>/transfer-host", methods=["POST"])
@login_required
def api_room_transfer_host(room_id):
    """방장 위임 — 현재 방장만. body: {to_user_id}"""
    me = current_user()
    db = get_db()
    if _my_room_role(db, room_id, me["id"]) != 'host':
        return jsonify({"error": "방장만 가능"}), 403
    data = request.get_json(silent=True) or {}
    to_uid = int(data.get("to_user_id") or 0)
    if not to_uid or to_uid == me["id"]:
        return jsonify({"error": "대상이 자기자신일 수 없음"}), 400
    target = db.execute("SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
                        (room_id, to_uid)).fetchone()
    if not target:
        return jsonify({"error": "대상이 방 멤버가 아님"}), 404
    db.execute("UPDATE room_members SET role='member' WHERE room_id=? AND user_id=?",
               (room_id, me["id"]))
    db.execute("UPDATE room_members SET role='host' WHERE room_id=? AND user_id=?",
               (room_id, to_uid))
    db.execute("UPDATE rooms SET created_by=? WHERE id=?", (to_uid, room_id))
    db.commit()
    to_user = db.execute("SELECT display_name FROM users WHERE id=?", (to_uid,)).fetchone()
    now = datetime.now(timezone.utc).isoformat()
    sys_text = f"방장이 [{me['display_name']}] → [{to_user['display_name']}] 로 변경됨"
    cur = db.execute(
        "INSERT INTO messages (room_id, user_id, content, kind, created_at) VALUES (?,?,?,?,?)",
        (room_id, me["id"], sys_text, "system", now),
    )
    db.commit()
    _emit_room_event(room_id, "host_transferred", {
        "room_id": room_id, "new_host_id": to_uid, "old_host_id": me["id"],
    })
    _emit_room_event(room_id, "new_message", {
        "id": cur.lastrowid, "room_id": room_id, "user_id": me["id"],
        "display_name": me["display_name"], "avatar_color": me["avatar_color"],
        "content": sys_text, "kind": "system", "created_at": now,
    })
    return jsonify({"ok": True})


@app.route("/api/rooms/<int:room_id>/invite", methods=["POST"])
@login_required
def api_room_invite(room_id):
    """방 멤버 초대 — invite_policy 에 따라 권한 분기.
    'all'(기본): 모든 방 멤버가 초대 가능. 'host_only': 방장·PM만.
    body: {user_ids: [int]}"""
    me = current_user()
    db = get_db()
    # 게스트(외부)는 회원 초대 권한 없음 (대표 지시 2026-05-30)
    if _is_guest(me):
        return jsonify({"error": "외부 사용자는 멤버를 초대할 수 없습니다."}), 403
    # 본인이 멤버인지 확인
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
        (room_id, me["id"]),
    ).fetchone():
        return jsonify({"error": "방 멤버 아님"}), 403
    # 방의 초대 정책 조회
    room_row = db.execute("SELECT invite_policy, type, created_by FROM rooms WHERE id=?", (room_id,)).fetchone()
    if not room_row:
        return jsonify({"error": "방을 찾을 수 없음"}), 404
    # 1:1·self 방은 멤버 추가 불가
    if room_row["type"] in ("direct", "self"):
        return jsonify({"error": "1:1 또는 1인방은 멤버를 추가할 수 없습니다"}), 400
    invite_policy = room_row["invite_policy"] or "all"
    my_role = _my_room_role(db, room_id, me["id"])
    if invite_policy == "host_only" and my_role not in ('host', 'sub_host'):
        return jsonify({"error": "이 방은 방장·PM만 초대할 수 있습니다"}), 403
    data = request.get_json(silent=True) or {}
    user_ids = list({int(x) for x in (data.get("user_ids") or [])})
    if not user_ids:
        return jsonify({"error": "초대할 사용자 ID 필요"}), 400
    # 채널 생성권한자(비ceo)는 '본인이 만든 채널에만' 초대 가능 (대표 지시 2026-05-24).
    #   단, 그 채널의 초대 정책이 'all'(모든 멤버 초대 가능)이면 방장이 명시적으로 개방한 것이므로
    #   이 제한을 적용하지 않는다 → 설정대로 모든 멤버 초대 허용. (대표 지시 2026-06-02)
    if invite_policy != "all" and me["role"] != "ceo" and _can_create_channel(me) and room_row["type"] == "channel":
        if room_row["created_by"] != me["id"]:
            return jsonify({"error": "본인이 만든 채널에만 초대할 수 있습니다."}), 403
        # 등록된 전체 직원 누구나 초대 가능 (부서 제한 제거)
    now = datetime.now(timezone.utc).isoformat()
    added = []
    for uid in user_ids:
        # 이미 멤버면 skip
        if db.execute("SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
                      (room_id, uid)).fetchone():
            continue
        u = db.execute("SELECT id, display_name, active FROM users WHERE id=?",
                       (uid,)).fetchone()
        if not u or not u["active"]:
            continue
        db.execute("INSERT INTO room_members (room_id, user_id, joined_at, role) VALUES (?,?,?,'member')",
                   (room_id, uid, now))
        added.append(u["display_name"])
    if not added:
        return jsonify({"error": "추가된 사용자 없음 (이미 멤버이거나 비활성)"}), 400
    db.commit()
    sys_text = f"[{', '.join(added)}] 님이 초대됨 (by {me['display_name']})"
    cur = db.execute(
        "INSERT INTO messages (room_id, user_id, content, kind, created_at) VALUES (?,?,?,?,?)",
        (room_id, me["id"], sys_text, "system", now),
    )
    db.commit()
    _emit_room_event(room_id, "members_added", {"room_id": room_id, "added": added})
    _emit_room_event(room_id, "new_message", {
        "id": cur.lastrowid, "room_id": room_id, "user_id": me["id"],
        "display_name": me["display_name"], "avatar_color": me["avatar_color"],
        "content": sys_text, "kind": "system", "created_at": now,
    })
    return jsonify({"ok": True, "added": added})


@app.route("/api/rooms/<int:room_id>/members/<int:user_id>/kick", methods=["POST"])
@login_required
def api_room_kick(room_id, user_id):
    """멤버 내보내기. 방장·PM 가능(권한 동일). 단, 방장(host)은 내보낼 수 없음(위임/나가기로만)."""
    me = current_user()
    db = get_db()
    my_role = _my_room_role(db, room_id, me["id"])
    if my_role not in ('host', 'sub_host'):
        return jsonify({"error": "방장·PM만 내보낼 수 있습니다"}), 403
    if user_id == me["id"]:
        return jsonify({"error": "본인은 /leave 로 나가기"}), 400
    target = db.execute("SELECT u.display_name, rm.role FROM room_members rm "
                        "JOIN users u ON u.id=rm.user_id "
                        "WHERE rm.room_id=? AND rm.user_id=?",
                        (room_id, user_id)).fetchone()
    if not target:
        return jsonify({"error": "멤버가 아님"}), 404
    # 방장(host)은 보호 — PM·방장도 방장은 내보낼 수 없음 (위임 후/스스로 나가기만)
    if target["role"] == 'host':
        return jsonify({"error": "방장은 내보낼 수 없습니다 (위임 후 진행하세요)"}), 403
    db.execute("DELETE FROM room_members WHERE room_id=? AND user_id=?", (room_id, user_id))
    db.commit()
    now = datetime.now(timezone.utc).isoformat()
    sys_text = f"[{target['display_name']}] 님이 방에서 나갔습니다 (by {me['display_name']})"
    cur = db.execute(
        "INSERT INTO messages (room_id, user_id, content, kind, created_at) VALUES (?,?,?,?,?)",
        (room_id, me["id"], sys_text, "system", now),
    )
    db.commit()
    _emit_room_event(room_id, "member_kicked", {"room_id": room_id, "user_id": user_id})
    _emit_room_event(room_id, "new_message", {
        "id": cur.lastrowid, "room_id": room_id, "user_id": me["id"],
        "display_name": me["display_name"], "avatar_color": me["avatar_color"],
        "content": sys_text, "kind": "system", "created_at": now,
    })
    return jsonify({"ok": True})


@app.route("/api/rooms/<int:room_id>/read", methods=["POST"])
@login_required
def api_mark_read(room_id):
    me = current_user()
    db = get_db()
    # 읽기 전 — 이 방에 내가 안 읽은(남이 보낸) 메시지가 있었는지 확인.
    #  있었다면 읽음 처리 후, 다른 기기(특히 백그라운드 휴대폰)의 이 방 알림을 clear 푸시로 닫는다.
    #  (안 읽은 게 없었으면 알림도 없으니 불필요한 푸시를 보내지 않음 — 푸시 남용/배터리 방지)
    prevrow = db.execute(
        "SELECT last_read_message_id FROM room_members WHERE room_id=? AND user_id=?",
        (room_id, me["id"]),
    ).fetchone()
    prev_read = (prevrow["last_read_message_id"] or 0) if prevrow else 0
    had_unread = db.execute(
        "SELECT 1 FROM messages WHERE room_id=? AND id>? AND user_id!=? "
        "AND (whisper_to_user_id IS NULL OR whisper_to_user_id=?) LIMIT 1",
        (room_id, prev_read, me["id"], me["id"]),
    ).fetchone() is not None
    last = db.execute("SELECT MAX(id) AS m FROM messages WHERE room_id=?", (room_id,)).fetchone()
    if last and last["m"]:
        now_iso = datetime.now(timezone.utc).isoformat()
        db.execute(
            "UPDATE room_members SET last_read_message_id=?, last_read_at=? WHERE room_id=? AND user_id=?",
            (last["m"], now_iso, room_id, me["id"]),
        )
        db.commit()
        # 같은 방의 다른 클라이언트에 읽음 알림 → 그쪽 UI에서 "안 읽음 N" 숫자 갱신
        socketio.emit("read_status", {
            "room_id": room_id,
            "user_id": me["id"],
            "last_read": last["m"],
            "last_read_at": now_iso,
        }, to=f"room_{room_id}")
    # 다른 기기의 이 방 OS 알림 자동 닫기 — 백그라운드 휴대폰은 소켓이 끊겨 푸시로만 닿으므로
    # clear 푸시를 보내 sw.js 가 tag=room_<id> 알림을 닫고 배지를 갱신하게 한다. (대표 지시 2026-05-20)
    #  ?clear=0 (옛 호환) 인 경우만 생략 — 본인 기기 포함 모든 구독에 보내 모바일 알림 회수 (대표 지시 2026-05-26)
    suppress_clear = (request.args.get("clear") == "0")
    if had_unread and PYWEBPUSH_OK and not suppress_clear:
        try:
            # 500ms 지연 — 메시지 푸시(미리 발사된 새 메시지 알림)가 모바일에 먼저 도달한 뒤
            # clear 푸시가 따라가서 닫게 함. 순서 보장 안 하면 clear 가 먼저 도착해 빈 곳에서 닫고
            # 그 다음 메시지 푸시가 알림을 새로 띄워 잔존하는 race condition 발생.
            def _delayed_clear():
                try:
                    socketio.sleep(0.5)   # eventlet 호환 sleep
                except Exception:
                    try:
                        import time as _t; _t.sleep(0.5)
                    except Exception:
                        pass
                send_push_to_user(me["id"], "", "", url=None,
                                   tag=f"room_{room_id}", clear=True)
            socketio.start_background_task(_delayed_clear)
        except Exception:
            pass
    return jsonify({"ok": True, "last_read": last["m"] if last else 0})


@app.route("/api/rooms/<int:room_id>/read_status")
@login_required
def api_read_status(room_id):
    """방 멤버별 마지막으로 읽은 메시지 ID 반환 (읽음/안읽음 표시용)."""
    me = current_user()
    db = get_db()
    if not db.execute(
        "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?", (room_id, me["id"])
    ).fetchone():
        abort(403)
    rows = db.execute("""
        SELECT rm.user_id, rm.last_read_message_id, rm.last_read_at,
               u.display_name, u.avatar_color, u.avatar_url, u.title, u.department,
               COALESCE(u.is_guest,0) AS is_guest, COALESCE(u.active,1) AS active, u.guest_company,
               (SELECT alias FROM guest_view_aliases
                 WHERE viewer_user_id=? AND guest_user_id=u.id) AS view_alias
          FROM room_members rm JOIN users u ON u.id = rm.user_id
         WHERE rm.room_id = ?
    """, (me["id"], room_id)).fetchall()
    return jsonify({
        "members": [dict(r) for r in rows],
        "total": len(rows),
    })


# ---------- SocketIO ----------
@socketio.on("connect")
def on_connect():
    """클라이언트 SocketIO 연결 시점에 사용자가 속한 모든 방에 자동 join.
    일반 메신저식 동작 — 활성 방이 아니어도 모든 방의 new_message broadcast 를 받아
    소리·토스트·사이드바 깜빡임 등 알림이 동작.
    + Presence 등록 — 클라이언트가 곧이어 "presence" 이벤트로 device/active 갱신."""
    uid = session.get("user_id")
    if not uid:
        return
    try:
        with app.app_context():
            db = get_db()
            rows = db.execute(
                "SELECT room_id FROM room_members WHERE user_id=?", (uid,)
            ).fetchall()
            for r in rows:
                sio_join(f"room_{r['room_id']}")
    except Exception as e:
        print(f"[socket connect] auto-join 실패: {e}")
    # 🏠 퇴근(offwork) 중 '완전 미접속 → 재접속' 이면 PC 복귀 대기 등록 (대표 지시 2026-05-25).
    #  이 connect 등록 직전까지 연결이 0 개였으면(=완전 미접속) + 퇴근 상태 → 대기 집합에 추가.
    #  이후 on_presence 에서 device=='pc' 확인되면 자동(online) 복귀.
    try:
        if not _user_is_online(uid):   # 이번 sid 등록 전 = 직전까지 완전 미접속
            with app.app_context():
                if _get_user_status(uid).get("status") == "offwork":
                    _offwork_clear_pending.add(uid)
    except Exception as e:
        print(f"[offwork] 재접속 감지 실패: {e}")
    # Presence — 일단 device 미상·active=True 로 등록.
    # 온라인 진입 broadcast 는 곧이어 오는 on_presence(기기 종류 확정 후)에서 처리 →
    # PC 면 '가능', 휴대폰만이면 '휴대폰' 으로 정확히 표시 (connect 시점엔 기기 미상이라 깜빡임 방지).
    try:
        _presence_register(uid, request.sid, device="unknown", active=True, ip=_real_client_ip())
    except Exception as e:
        print(f"[presence] register 실패: {e}")


@socketio.on("disconnect")
def on_disconnect():
    """SocketIO 연결 해제 시 presence 에서 제거.
    마지막 연결이 끊겼으면 다른 사용자에게 offline 알림."""
    uid = session.get("user_id")
    if not uid:
        return
    try:
        _presence_unregister(uid, request.sid)
        # 완전 미접속이 되면 퇴근 복귀 대기 정리 (다음 재접속 때 새로 판정)
        if not _user_is_online(uid):
            _offwork_clear_pending.discard(uid)
        # 연결 해제 후 표시 상태 재계산 broadcast:
        #  · PC 끊기고 휴대폰 남음 → '가능'→'휴대폰' 으로 전환
        #  · 마지막 연결까지 끊김 → 푸시 있으면 '휴대폰', 없으면 '오프라인'
        with app.app_context():
            _broadcast_status_if_changed(uid)
    except Exception as e:
        print(f"[presence] unregister 실패: {e}")


@socketio.on("presence")
def on_presence(data):
    """클라이언트의 device_type + active 상태 보고.
    페이지 visibility 변경, focus/blur 마다 호출됨.
    이 정보로 _user_has_active_pc(uid) 가 푸시 스킵 여부를 판단."""
    uid = session.get("user_id")
    if not uid or not isinstance(data, dict):
        return
    device = data.get("device")
    if device not in ("pc", "mobile"):
        device = "unknown"
    active = bool(data.get("active", True))
    idle = bool(data.get("idle", False))
    try:
        _presence_register(uid, request.sid, device=device, active=active, idle=idle, ip=_real_client_ip())
        # 기기 종류(pc/mobile) 확정 → 표시 상태(가능/휴대폰) 변동 시 broadcast
        with app.app_context():
            # 🏠 퇴근 중 재접속 확인 → 자동(online: 컴퓨터/휴대폰) 복귀 (대표 지시 2026-05-25)
            #  복귀 조건: ① PC(데스크탑)로 접속  OR  ② 회사 인터넷망(회사 IP)에 접속(휴대폰 포함)
            if uid in _offwork_clear_pending and (device == "pc" or _user_at_office(uid)):
                _offwork_clear_pending.discard(uid)
                try:
                    _db = get_db()
                    _now = datetime.now(timezone.utc).isoformat()
                    _db.execute(
                        "UPDATE user_statuses SET status='online', custom_text=NULL, emoji=NULL, "
                        "until_at=NULL, auto_set=1, updated_at=? WHERE user_id=? AND status='offwork'",
                        (_now, uid),
                    )
                    _db.commit()
                except Exception as e2:
                    print(f"[offwork] 재접속 자동복귀 실패: {e2}")
            _broadcast_status_if_changed(uid)
    except Exception as e:
        print(f"[presence] update 실패: {e}")


@socketio.on("join")
def on_join(data):
    rid = data.get("room_id") if isinstance(data, dict) else None
    if rid:
        sio_join(f"room_{rid}")


@socketio.on("leave")
def on_leave(data):
    rid = data.get("room_id") if isinstance(data, dict) else None
    if rid:
        sio_leave(f"room_{rid}")


@socketio.on("set_active_room")
def on_set_active_room(data):
    """클라이언트가 현재 보고 있는 방 알림 — 푸시 발송 시 비교용 (대표 지시 2026-05-26).
    Chrome PWA 의 SW clients.matchAll() 버그 우회 — 서버가 직접 추적해 푸시 차단.
    body: {room_id: N | null}. null/0 = 방 안 들어간 상태(목록 화면 등)."""
    uid = session.get("user_id")
    if not uid:
        return
    try:
        rid = None
        if isinstance(data, dict):
            try:
                rid = int(data.get("room_id") or 0) or None
            except Exception:
                rid = None
        _presence_set_active_room(uid, request.sid, rid)
    except Exception:
        pass


# ============================================================
# 스티커 — static/stickers/manifest.json 의 허용 파일 + 라벨
# ============================================================
STICKER_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "stickers")
_STICKER_LABELS = None
def _sticker_labels():
    """{파일명: 라벨} dict. manifest.json 1회 로드 후 캐시."""
    global _STICKER_LABELS
    if _STICKER_LABELS is None:
        labels = {}
        try:
            with open(os.path.join(STICKER_DIR, "manifest.json"), encoding="utf-8") as fp:
                for item in json.load(fp):
                    f = item.get("file")
                    if f:
                        labels[f] = item.get("label", "")
        except Exception:
            labels = {}
        _STICKER_LABELS = labels
    return _STICKER_LABELS


# ──────────────────────────────────────────────────────────────
#  @멘션 처리 — 멘션함 (대표 지시 2026-05-22)
#  메시지 본문의 '@이름' 을 같은 방 멤버 display_name 과 정확 매칭해
#  mentions 행을 만들고, 멘션된 사용자에게 실시간 'mention_added' 알림.
#  귓속말(whisper_to)·스티커는 제외(수신자만 보이거나 본문에 이름 없음).
# ──────────────────────────────────────────────────────────────
def _mention_name_candidates(display_name, username):
    """한 사용자를 '@'로 부를 때 쓰일 수 있는 후보 이름들.
    - display_name (현재 멘션 형식, 2026-05-22~)
    - username (옛 멘션 형식 — @아이디)
    - username 이 이메일이면 @ 앞부분(local part) 도 (예: buy1@... → buy1)"""
    cands = set()
    dn = (display_name or "").strip()
    if dn:
        cands.add(dn)
    un = (username or "").strip()
    if un:
        cands.add(un)
        if "@" in un:
            local = un.split("@", 1)[0].strip()
            if local:
                cands.add(local)
    return cands


def _content_mentions_user(content, display_name, username):
    """content 안에 이 사용자를 가리키는 '@이름' 토큰이 있으면 True."""
    for cand in _mention_name_candidates(display_name, username):
        if re.search(r"@" + re.escape(cand) + r"(?![\w가-힣])", content):
            return True
    return False


def _process_mentions(db, message_id, room_id, sender_id, content, whisper_to, created_at, sender_name=""):
    """반환: 멘션된 user_id 리스트. (DB INSERT + socket emit 까지 수행)"""
    if not content or "@" not in content or whisper_to:
        return []
    members = db.execute(
        "SELECT u.id, u.display_name, u.username FROM room_members rm "
        " JOIN users u ON u.id = rm.user_id WHERE rm.room_id=?",
        (room_id,),
    ).fetchall()
    mentioned = []
    for mem in members:
        if mem["id"] == sender_id:
            continue  # 본인 멘션 제외
        # '@이름' 토큰(이름/아이디) — 바로 뒤에 한글/영숫자/_ 가 붙지 않을 때만 일치
        if _content_mentions_user(content, mem["display_name"], mem["username"]):
            try:
                c = db.execute(
                    "INSERT OR IGNORE INTO mentions "
                    "(message_id, room_id, mentioned_user_id, sender_user_id, created_at) "
                    "VALUES (?,?,?,?,?)",
                    (message_id, room_id, mem["id"], sender_id, created_at),
                )
                if c.rowcount:
                    mentioned.append(mem["id"])
            except Exception:
                pass
    if mentioned:
        db.commit()
        for uid_m in mentioned:
            sids = []
            try:
                with _user_conn_lock:
                    sids = list((_user_connections.get(uid_m, {})).keys())
            except Exception:
                sids = []
            for sid in sids:
                try:
                    socketio.emit("mention_added", {
                        "room_id": room_id,
                        "message_id": message_id,
                        "sender_name": sender_name,
                    }, to=sid)
                except Exception:
                    pass
    return mentioned


def _process_mentions_bg(message_id, room_id, sender_id, content, whisper_to, created_at, sender_name=""):
    """멘션 처리를 '메시지 전달 이후' 백그라운드에서 수행 — 전송 지연 방지. (대표 지시 2026-05-23)
    on_send 는 new_message 를 먼저 emit 하고, 이 작업을 start_background_task 로 분리해
    멘션 DB 조회·기록·배지 알림이 메시지 도착을 막지 않게 한다. (멘션은 배지용이라 약간 늦어도 무방)"""
    try:
        with app.app_context():
            db = get_db()
            _process_mentions(db, message_id, room_id, sender_id, content, whisper_to, created_at, sender_name)
    except Exception as e:
        print(f"[mention] 백그라운드 처리 실패: {e}", flush=True)


@socketio.on("send")
def on_send(data):
    uid = session.get("user_id")
    if not uid or not isinstance(data, dict):
        return
    # 🛡️ Rate Limit — 분당 60개 메시지 (1초당 1개 평균. 스팸 차단)
    if not _check_rate_limit(uid, "send_message", max_per_minute=60):
        try:
            socketio.emit("rate_limited", {"action": "send_message", "message": "메시지 전송 속도 제한 — 1분에 60개"}, to=request.sid)
        except Exception:
            pass
        return
    room_id = data.get("room_id")
    content = (data.get("content") or "").strip()
    # 스티커 전송 — sticker=파일명. static/stickers 의 허용된 파일만 처리. content 는 라벨로 대체.
    sticker_file = (data.get("sticker") or "").strip()
    is_sticker = False
    if sticker_file:
        _labels = _sticker_labels()
        if sticker_file in _labels:
            is_sticker = True
            content = _labels.get(sticker_file) or "스티커"
        else:
            return  # 허용되지 않은 스티커 파일명 — 무시
    if not room_id:
        return
    if not is_sticker and not content:
        return
    if len(content) > 4000:
        content = content[:4000]
    # 인용 답장 — quoted_message_id (선택). 본 채널에 답글 + 원본 미니 카드.
    quoted_id = data.get("quoted_message_id")
    try:
        quoted_id = int(quoted_id) if quoted_id else None
    except (ValueError, TypeError):
        quoted_id = None
    # 귓속말 — whisper_to_user_id (선택). 송신자·수신자만 보이는 메시지.
    whisper_to = data.get("whisper_to_user_id")
    try:
        whisper_to = int(whisper_to) if whisper_to else None
    except (ValueError, TypeError):
        whisper_to = None
    if whisper_to == uid:
        whisper_to = None   # 자기 자신에게 귓속말은 의미 없음

    with app.app_context():
        db = get_db()
        if not db.execute(
            "SELECT 1 FROM room_members WHERE room_id=? AND user_id=?",
            (room_id, uid),
        ).fetchone():
            return
        # quoted_id 가 있으면 같은 방의 메시지인지 검증 (다른 방 메시지를 인용은 안 됨 — 전달 기능 사용)
        if quoted_id:
            qrow = db.execute(
                "SELECT room_id FROM messages WHERE id=?", (quoted_id,)
            ).fetchone()
            if not qrow or qrow["room_id"] != room_id:
                quoted_id = None
        # 귓속말 대상이 같은 방 멤버인지 확인
        whisper_to_name = None
        if whisper_to:
            wrow = db.execute(
                "SELECT u.display_name FROM users u "
                " JOIN room_members rm ON rm.user_id=u.id "
                " WHERE u.id=? AND rm.room_id=?",
                (whisper_to, room_id),
            ).fetchone()
            if not wrow:
                whisper_to = None  # 같은 방이 아니면 귓속말 불가
            else:
                whisper_to_name = wrow["display_name"]
        now = datetime.now(timezone.utc).isoformat()
        kind = "sticker" if is_sticker else "text"
        cur = db.execute(
            "INSERT INTO messages (room_id, user_id, content, kind, created_at, quoted_message_id, whisper_to_user_id, file_name) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (room_id, uid, content, kind, now, quoted_id, whisper_to, sticker_file if is_sticker else None),
        )
        mid = cur.lastrowid
        db.commit()
        u = db.execute(
            "SELECT display_name, avatar_color FROM users WHERE id=?", (uid,)
        ).fetchone()
        _sender_name = u["display_name"]   # 멘션 백그라운드 처리에 사용 (아래 emit 후)
        # quoted 메타데이터 — 클라이언트가 즉시 카드로 렌더할 수 있도록 함께 보냄
        quoted_meta = None
        if quoted_id:
            q = db.execute("""
                SELECT m.id, m.content, m.kind, m.created_at, m.file_name,
                       u.display_name, u.avatar_color
                  FROM messages m JOIN users u ON u.id = m.user_id
                 WHERE m.id = ?
            """, (quoted_id,)).fetchone()
            if q:
                quoted_meta = dict(q)

    payload = {
        "id": mid,
        "room_id": room_id,
        "user_id": uid,
        "display_name": u["display_name"],
        "avatar_color": u["avatar_color"],
        "content": content,
        "kind": kind,
        "created_at": now,
    }
    if is_sticker:
        payload["file_name"] = sticker_file
    if quoted_id:
        payload["quoted_message_id"] = quoted_id
        payload["quoted"] = quoted_meta
    if whisper_to:
        payload["whisper_to_user_id"] = whisper_to
        payload["whisper_to_name"] = whisper_to_name
    # 귓속말은 방 전체 broadcast 안 함 — 송신자·수신자 SID 에만 개별 emit
    if whisper_to:
        target_sids = set()
        with _user_conn_lock:
            for u_to_check in (uid, whisper_to):
                for sid, info in (_user_connections.get(u_to_check, {})).items():
                    target_sids.add(sid)
        for sid in target_sids:
            try:
                socketio.emit("new_message", payload, to=sid)
            except Exception as _e:
                pass
    else:
        socketio.emit("new_message", payload, to=f"room_{room_id}")

    # @멘션 처리는 메시지 전달 '이후' 백그라운드로 — 전송 지연 방지. (스티커 제외, 귓속말은 함수 내부에서 제외)
    if not is_sticker and content and "@" in content:
        try:
            socketio.start_background_task(
                _process_mentions_bg, mid, room_id, uid, content, whisper_to, now, _sender_name
            )
        except Exception as _me:
            print(f"[mention] 백그라운드 태스크 시작 실패: {_me}", flush=True)

    # Web Push — 백그라운드 알림 (송신자 제외 모든 방 멤버)
    # 귓속말이면 푸시도 수신자 1명에게만 (송신자·다른 멤버 제외).
    if PYWEBPUSH_OK and whisper_to:
        # 귓속말 푸시
        with app.app_context():
            db3 = get_db()
            r3 = db3.execute("SELECT name FROM rooms WHERE id=?", (room_id,)).fetchone()
            room_name = r3["name"] if r3 else "채팅"
        if not _user_has_active_session(whisper_to):
            import threading as _t
            _t.Thread(
                target=send_push_to_user,
                args=(whisper_to, f"🤫 {u['display_name']} 귓속말 ({room_name})", content[:120]),
                # tag = 방+발신자 단위 — 같은 발신자의 귓속말은 1장 카드에 누적 (대표 지시 2026-05-26 팝업 통합)
                kwargs={"url": f"{BASE_PATH}/chat?room={room_id}", "tag": f"whisper_{room_id}_{uid}"},
                daemon=True,
            ).start()
        return  # 일반 푸시 흐름 스킵
    if PYWEBPUSH_OK:
        # 방 이름 조회 + 메시지에 멘션 있는지 검사
        with app.app_context():
            db2 = get_db()
            r = db2.execute("SELECT name, type FROM rooms WHERE id=?", (room_id,)).fetchone()
            room_name = (r["name"] if r else "") or ""
        # 빈 방 이름(1:1 등)이면 괄호 생략 — '김정락 ()' 같은 빈 괄호 방지 (대표 지시 2026-05-26)
        title = f"💬 {u['display_name']} ({room_name})" if room_name else f"💬 {u['display_name']}"
        body = content[:120]
        # 비동기 스레드로 발송 (pywebpush는 HTTP 호출이라 블로킹)
        # tag 를 '방 단위'(room_N)로 — 같은 방 알림은 1개 카드로 합쳐짐. 새 메시지는
        # renotify 로 그 카드를 갱신·재알림. 알림창이 메시지 수만큼 쌓이는 문제 해결. (2026-05-20)
        # 앱 아이콘 배지 숫자는 sw.js 의 setAppBadge(payload.badge=안읽음 총합) 가 담당 → tag 와 무관.
        import threading
        threading.Thread(
            target=push_message_to_room_members,
            args=(room_id, uid, title, body),
            # tag = 방 단위 — 같은 방 메시지는 1장 카드에 누적 (대표 지시 2026-05-26 팝업 통합)
            kwargs={"url": f"{BASE_PATH}/chat?room={room_id}", "tag": f"room_{room_id}"},
            daemon=True,
        ).start()


def _local_ips():
    """이 PC의 LAN/VPN IP들을 모두 반환."""
    ips = set()
    try:
        import socket as _sock
        host = _sock.gethostname()
        for info in _sock.getaddrinfo(host, None):
            ip = info[4][0]
            if ip and ":" not in ip and not ip.startswith("127."):
                ips.add(ip)
    except Exception:
        pass
    return sorted(ips)


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    init_db()
    _start_calendar_worker()  # 캘린더 자동 상태 전환 백그라운드 워커
    _start_project_history_worker()  # 프로젝트 이력 하루 1회 자동 생성
    _start_backup_worker()  # 매일 새벽 3시(KST) 자동 백업 (대표 지시 2026-05-19)
    print()
    print(" ============================================")
    print("  KNK Messenger - server start")
    print(" ============================================")
    print(f"   PC (this server):  http://localhost:{PORT}")
    for ip in _local_ips():
        print(f"   employee URL:      http://{ip}:{PORT}")
    print(f"   port:              {PORT}  (open in firewall)")
    print(f"   upload dir:        {UPLOAD_DIR}")
    print(f"   DB:                {DB_PATH}")
    print(f"   retention:         {MESSAGE_RETENTION_MONTHS} months")
    print(f"   env:               {ENV}  async_mode={ASYNC_MODE}")
    # AI 공급자 상태 표시
    if TRANSLATE_MOCK:
        print(f"   AI translation:    DEMO MODE (auto-enabled, no API key)")
        print(f"                      -> 가짜 번역 + KNK 핵심 용어만 진짜 변환")
        print(f"                      -> 진짜 번역 원하시면: OpenAI키설정.bat (또는 번역키설정.bat)")
    elif TRANSLATE_PROVIDER == "openai" and OPENAI_API_KEY:
        print(f"   AI translation:    REAL [OpenAI] {OPENAI_MODEL}  (${TRANSLATE_MONTHLY_USD_LIMIT}/month limit)")
    elif TRANSLATE_PROVIDER == "anthropic" and ANTHROPIC_API_KEY:
        print(f"   AI translation:    REAL [Anthropic] {TRANSLATE_MODEL}  (${TRANSLATE_MONTHLY_USD_LIMIT}/month limit)")
    else:
        provider_name = "OpenAI" if TRANSLATE_PROVIDER == "openai" else "Anthropic"
        print(f"   AI translation:    DISABLED — {provider_name} key 미설정")
    print(" ============================================")
    print()
    socketio.run(app, host="0.0.0.0", port=PORT, debug=False, allow_unsafe_werkzeug=True)
