"""전직원 Excel → 기술영업팀 추출하여 KNK 메신저에 시드.

기존 시드 계정(kjr/hong/lee)은 유지. 신규로 기술영업팀 멤버를 추가.
이미 같은 username 존재하면 스킵.
"""
import os
import sys
import sqlite3
from datetime import datetime, timezone
from openpyxl import load_workbook

# 윈도우 콘솔 한글 출력
sys.stdout.reconfigure(encoding="utf-8")

# Pillow 폰트 처럼 werkzeug 직접 사용
from werkzeug.security import generate_password_hash

XLSX = r"C:\Users\top00\JR\Claude 코드\KNK업무시스템구축\참고자료\전직원_로그인계정_2026-04-28.xlsx"
DB = r"C:\Users\top00\JR\Claude 코드\KNK업무시스템구축\10_KNK_Messenger\data\messenger.db"

# 부서명 키워드 — '기술영업' / '영업' 포함
TARGET_KEYWORDS = ["기술영업", "기영", "영업"]

# 부서별 색상 팔레트
DEPT_COLORS = {
    "기술영업": "#06b6d4",   # 청록
    "영업": "#0ea5e9",        # 파랑
    "default": "#3b82f6",
}

wb = load_workbook(XLSX, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
print(f"[엑셀] 헤더: {header}")
print(f"[엑셀] 총 {len(rows) - 1}명")

# 컬럼 인덱스 추정
def find_col(keys):
    for i, h in enumerate(header):
        if h and any(k in str(h) for k in keys):
            return i
    return None

c_dept = find_col(["부서"])
c_rank = find_col(["직급"])
c_name = find_col(["이름"])
c_id = find_col(["아이디", "ID"])
c_pw = find_col(["비밀", "PW"])

print(f"[컬럼] 부서={c_dept} 직급={c_rank} 이름={c_name} 아이디={c_id} 비번={c_pw}")

# 부서명 후보 모두 수집
all_depts = sorted({(r[c_dept] or "").strip() for r in rows[1:] if r[c_dept]})
print(f"[부서 목록 — 검색 참조용]")
for d in all_depts:
    print(f"  · {d}")
print()

# 기술영업 매칭
matches = []
for r in rows[1:]:
    if not r[c_name] or not r[c_id]:
        continue
    dept = (r[c_dept] or "").strip()
    if any(k in dept for k in TARGET_KEYWORDS):
        matches.append({
            "dept": dept,
            "rank": (r[c_rank] or "").strip() if c_rank is not None else "",
            "name": str(r[c_name]).strip(),
            "username": str(r[c_id]).strip().lower(),
            "password": str(r[c_pw] or "knk1234").strip(),
        })

print(f"[매칭] 기술영업 키워드({TARGET_KEYWORDS}) 일치 {len(matches)}명:")
for m in matches:
    print(f"  · {m['dept']} / {m['rank']} / {m['name']} ({m['username']})")
print()

# DB 시드
conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

now = datetime.now(timezone.utc).isoformat()
added = 0
skipped = 0
for m in matches:
    if cur.execute("SELECT 1 FROM users WHERE username = ?", (m["username"],)).fetchone():
        skipped += 1
        continue
    color = DEPT_COLORS.get("기술영업" if "기술영업" in m["dept"] else "영업", DEPT_COLORS["default"])
    role = "ceo" if m["rank"] in ("대표이사", "회장") else "manager" if m["rank"] in ("이사", "상무", "전무", "부장", "팀장") else "staff"
    display = f"{m['name']} {m['rank']}".strip() if m["rank"] else m["name"]
    cur.execute(
        "INSERT INTO users (username, password_hash, display_name, role, avatar_color, created_at) VALUES (?,?,?,?,?,?)",
        (m["username"], generate_password_hash(m["password"]), display, role, color, now),
    )
    added += 1
    print(f"  ✓ 추가: {display} ({m['username']})")

conn.commit()

# 부서별 채널 자동 생성 (기술영업팀)
if matches:
    # 이미 기술영업팀 채널 있는지 확인
    existing_room = cur.execute(
        "SELECT id FROM rooms WHERE name = ? AND type = 'channel'", ("기술영업팀",)
    ).fetchone()
    if not existing_room:
        cur.execute(
            "INSERT INTO rooms (name, type, created_by, created_at) VALUES (?,?,?,?)",
            ("기술영업팀", "channel", 1, now),
        )
        rid = cur.lastrowid
        # 김정락 대표(uid=1) + 매칭된 모든 신규 직원 추가
        cur.execute(
            "INSERT INTO room_members (room_id, user_id, joined_at) VALUES (?,?,?)",
            (rid, 1, now),
        )
        for m in matches:
            u = cur.execute("SELECT id FROM users WHERE username = ?", (m["username"],)).fetchone()
            if u:
                cur.execute(
                    "INSERT OR IGNORE INTO room_members (room_id, user_id, joined_at) VALUES (?,?,?)",
                    (rid, u["id"], now),
                )
        cur.execute(
            "INSERT INTO messages (room_id, user_id, content, kind, created_at) VALUES (?,?,?,?,?)",
            (rid, 1, "기술영업팀 채널 — 베타 시범 운영 시작합니다. 카톡과 병행하면서 묻혀있던 요청·자료를 여기로 모아보시죠.", "system", now),
        )
        print(f"\n[채널] 기술영업팀 채널 생성 (room_id={rid}), 멤버 {len(matches) + 1}명")
    else:
        print(f"\n[채널] 기술영업팀 채널 이미 존재 (room_id={existing_room['id']})")

conn.commit()
conn.close()

print(f"\n[완료] 추가 {added}명 / 스킵(이미 있음) {skipped}명")
