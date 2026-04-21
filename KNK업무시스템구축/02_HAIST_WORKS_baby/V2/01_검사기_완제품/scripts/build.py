"""
╔══════════════════════════════════════════════════════════════╗
║  KNK PMS V4 — 01_검사기_완제품 build.py                       ║
║  엑셀 파일 구조 생성 + 서식 + 보호 + Comment                  ║
║  실행: python build.py                                        ║
╚══════════════════════════════════════════════════════════════╝
"""

import os, sys, glob, shutil, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# config import
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import *


# ═══════════════════════════════════════════════════════════════
# 빌드 유틸리티
# ═══════════════════════════════════════════════════════════════

def _add_dropdown(ws, col, options, row_start=5, row_end=2000, allow_blank=True):
    """드롭다운 유효성 검사 추가"""
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    dv.error = "목록에서 선택하세요"
    dv.errorTitle = "입력 오류"
    col_letter = get_column_letter(col)
    dv.add(f"{col_letter}{row_start}:{col_letter}{row_end}")
    ws.add_data_validation(dv)


def _apply_dropdowns_pms(ws, row_end=2000):
    """PMS 1_프로젝트등록 시트에 드롭다운 적용

    스펙 §18 (v2026.04): 부서담당자는 드롭다운 제거 — 담당자 이름 직접 입력.
    공란이면 sync 시점에 자동 제외된다.
    """
    for col, (options, _) in DROPDOWN_MAP.items():
        _add_dropdown(ws, col, options, row_end=row_end)


# ═══════════════════════════════════════════════════════════════
# PMS 통합 파일 빌드 (7시트)
# ═══════════════════════════════════════════════════════════════

def build_pms():
    """PMS 통합 파일 생성"""
    fp = pms_path()
    print(f"  [BUILD] PMS: {os.path.basename(fp)}")
    wb = Workbook()

    # ── 1_프로젝트등록 ──
    ws1 = wb.active
    ws1.title = "1_프로젝트등록"
    ws1.sheet_properties.tabColor = TAB_INPUT

    # v2026.04e: PMS 33열 — 부서(8) + 매출·수금(2) 포함
    total_col = MAX_COL
    setup_r1(ws1, f"㈜케이엔케이 │ {TYPE_NAME} │ 프로젝트 등록 │ {YEAR}", total_col)
    r3_full = dict(R3_MAP_PROJ)
    apply_r3_guide(ws1, total_col, r3_full)
    labels = list(PMS_HEADER_LABELS) + list(DEPTS) + ["계산서\n발행일", "입금일"]
    r4_full = dict(R4_MAP_PROJ)
    apply_r4_header(ws1, total_col, labels, r4_full)
    # 서식
    money_cols = [C_UPRICE, C_TOTAL]
    format_data_rows(ws1, total_col, money_cols=money_cols)
    # 보호
    apply_protection(ws1, total_col, r3_full)
    # 드롭다운
    _apply_dropdowns_pms(ws1)
    # 컬럼 너비
    auto_fit_columns(ws1)

    # ── 2_진행현황 ──
    ws2 = wb.create_sheet("2_진행현황")
    ws2.sheet_properties.tabColor = TAB_VIEW
    _build_progress_sheet(ws2)

    # ── 3_출하현황 ──
    ws3 = wb.create_sheet("3_출하현황")
    ws3.sheet_properties.tabColor = TAB_VIEW
    _build_ship_summary_sheet(ws3)

    # ── 4_완료이력 ──
    ws4 = wb.create_sheet("4_완료이력")
    ws4.sheet_properties.tabColor = TAB_VIEW2
    _build_archive_sheet(ws4)

    # ── 5_매핑조회(수주번호) ──
    ws5 = wb.create_sheet("5_매핑조회(수주번호)")
    ws5.sheet_properties.tabColor = TAB_VIEW2
    _build_mapping_sheet(ws5)

    # ── 6_관리코드발행대장 ──
    ws6 = wb.create_sheet("6_관리코드발행대장")
    ws6.sheet_properties.tabColor = TAB_VIEW2
    _build_code_ledger_sheet(ws6)

    # ── 7_수주번호생성대장 ──
    ws7 = wb.create_sheet("7_수주번호생성대장")
    ws7.sheet_properties.tabColor = TAB_VIEW2
    _build_sjnum_ledger_sheet(ws7)

    # ── 8_매출마감 (v2026.04e — 계산서·입금 완료 건 아카이브) ──
    ws8 = wb.create_sheet("8_매출마감")
    ws8.sheet_properties.tabColor = TAB_VIEW2
    _build_settled_sheet(ws8)

    wb.save(fp)
    print(f"    → 저장 완료: {fp}")
    return fp


def _build_settled_sheet(ws):
    """8_매출마감 시트 — 계산서 발행 + 입금 완료된 건 최종 아카이브 (1_프로젝트등록과 동일 구조)"""
    setup_r1(ws, f"㈜케이엔케이 │ {TYPE_NAME} │ 매출마감 │ {YEAR}", MAX_COL)
    r3 = {c: "auto" for c in range(1, MAX_COL + 1)}
    apply_r3_guide(ws, MAX_COL, r3)
    labels = list(PMS_HEADER_LABELS) + list(DEPTS) + ["계산서\n발행일", "입금일"]
    apply_r4_header(ws, MAX_COL, labels, dict(R4_MAP_PROJ))
    format_data_rows(ws, MAX_COL, money_cols=[C_UPRICE, C_TOTAL])
    apply_protection(ws, MAX_COL, r3)
    auto_fit_columns(ws)


def _build_progress_sheet(ws):
    """2_진행현황 시트 (부서별 진척률 자동 수집)"""
    # 컬럼: NO, 관리코드, 수주번호, 고객사, 품명, 모델, 진행상태, 전체진척률,
    #        + 각 부서(세부항목별 진척률 + 소계)
    # 동적이므로 기본 헤더만 설정
    base_cols = ["NO","관리코드","수주번호","고객사","모델","품명","진행상태","전체\n진척률(%)"]
    col = len(base_cols)
    all_labels = list(base_cols)

    for dept in DEPTS:
        subs = DEPT_SUB_ITEMS.get(dept, [])
        for sub in subs:
            all_labels.append(f"{dept}\n{sub}")
            col += 1
        all_labels.append(f"{dept}\n소계")
        col += 1

    total_col = len(all_labels)
    setup_r1(ws, f"㈜케이엔케이 │ {TYPE_NAME} │ 진행현황 │ {YEAR}", total_col)

    # R3 전체 auto
    r3 = {c: "auto" for c in range(1, total_col + 1)}
    apply_r3_guide(ws, total_col, r3)

    # R4 헤더 (부서별 색상 구분)
    for c in range(1, total_col + 1):
        cell = ws.cell(row=4, column=c)
        cell.value = all_labels[c - 1]
        cell.font = FT_HEAD
        cell.alignment = AL_C
        cell.border = THIN
        if c <= len(base_cols):
            cell.fill = FILL_KNK_RED
        else:
            # 부서별 색상 계산
            dept_idx = _get_dept_index_for_col(c, base_cols, DEPTS)
            color = DEPT_HEADER_COLORS[dept_idx % len(DEPT_HEADER_COLORS)]
            cell.fill = PatternFill("solid", fgColor=color)

    # 서식
    pct_cols = list(range(len(base_cols), total_col + 1))
    format_data_rows(ws, total_col, pct_cols=pct_cols)
    apply_protection(ws, total_col, r3)
    auto_fit_columns(ws)


def _get_dept_index_for_col(col, base_cols, depts):
    """진행현황에서 컬럼번호 → 부서 인덱스 계산"""
    offset = col - len(base_cols) - 1
    idx = 0
    for i, dept in enumerate(depts):
        n = len(DEPT_SUB_ITEMS.get(dept, [])) + 1  # 세부항목 + 소계
        if offset < n:
            return i
        offset -= n
    return len(depts) - 1


def _build_ship_summary_sheet(ws):
    """3_출하현황 시트"""
    mc = SHIP_SUMMARY_MAX_COL
    setup_r1(ws, f"㈜케이엔케이 │ {TYPE_NAME} │ 출하현황 │ {YEAR}", mc)
    r3 = {c: "auto" for c in range(1, mc + 1)}
    apply_r3_guide(ws, mc, r3)
    apply_r4_header(ws, mc, SHIP_SUMMARY_LABELS, R4_MAP_SHIP)
    money_cols = [16]
    money_usd = [17]
    format_data_rows(ws, mc, money_cols=money_cols, money_usd_cols=money_usd)
    apply_protection(ws, mc, r3)
    auto_fit_columns(ws)


def _build_archive_sheet(ws):
    """4_완료이력 시트 (1_프로젝트등록과 동일 구조, 매출·수금 컬럼 포함)"""
    setup_r1(ws, f"㈜케이엔케이 │ {TYPE_NAME} │ 완료이력 │ {YEAR}", MAX_COL)
    r3 = {c: "auto" for c in range(1, MAX_COL + 1)}
    apply_r3_guide(ws, MAX_COL, r3)
    labels = list(PMS_HEADER_LABELS) + list(DEPTS) + ["계산서\n발행일", "입금일"]
    apply_r4_header(ws, MAX_COL, labels, R4_MAP_PROJ)
    format_data_rows(ws, MAX_COL, money_cols=[C_UPRICE, C_TOTAL])
    apply_protection(ws, MAX_COL, r3)
    auto_fit_columns(ws)


def _build_mapping_sheet(ws):
    """5_매핑조회(수주번호) 시트"""
    cols = ["NO","관리코드","수주번호","고객사","모델","품명","진행상태","수주일"]
    mc = len(cols)
    setup_r1(ws, f"㈜케이엔케이 │ {TYPE_NAME} │ 매핑조회(수주번호) │ {YEAR}", mc)
    r3 = {c: "auto" for c in range(1, mc + 1)}
    apply_r3_guide(ws, mc, r3)
    r4_map = {c: "id" for c in range(1, mc + 1)}
    apply_r4_header(ws, mc, cols, r4_map)
    format_data_rows(ws, mc, row_end=1000)
    apply_protection(ws, mc, r3, row_end=1000)
    auto_fit_columns(ws)


def _build_code_ledger_sheet(ws):
    """6_관리코드발행대장 시트"""
    cols = ["NO","관리코드","수주번호","고객사","품명","발행일","발행사유"]
    mc = len(cols)
    setup_r1(ws, f"㈜케이엔케이 │ {TYPE_NAME} │ 관리코드발행대장 │ {YEAR}", mc)
    r3 = {c: "auto" for c in range(1, mc + 1)}
    apply_r3_guide(ws, mc, r3)
    r4_map = {c: "id" for c in range(1, mc + 1)}
    apply_r4_header(ws, mc, cols, r4_map)
    format_data_rows(ws, mc, row_end=5000)
    apply_protection(ws, mc, r3, row_end=5000)
    auto_fit_columns(ws)


def _build_sjnum_ledger_sheet(ws):
    """7_수주번호생성대장 시트"""
    cols = ["NO","수주번호","관리코드","고객사","품명","생성일"]
    mc = len(cols)
    setup_r1(ws, f"㈜케이엔케이 │ {TYPE_NAME} │ 수주번호생성대장 │ {YEAR}", mc)
    r3 = {c: "auto" for c in range(1, mc + 1)}
    apply_r3_guide(ws, mc, r3)
    r4_map = {c: "id" for c in range(1, mc + 1)}
    apply_r4_header(ws, mc, cols, r4_map)
    format_data_rows(ws, mc, row_end=5000)
    apply_protection(ws, mc, r3, row_end=5000)
    auto_fit_columns(ws)


# ═══════════════════════════════════════════════════════════════
# 부서입력 파일 빌드
# ═══════════════════════════════════════════════════════════════

def build_dept_file(dept):
    """부서입력 파일 생성"""
    fp = dept_path(dept)
    print(f"  [BUILD] 부서입력: {dept} → {os.path.basename(fp)}")
    wb = Workbook()
    ws = wb.active
    ws.title = "부서진척입력"
    ws.sheet_properties.tabColor = TAB_INPUT

    subs = DEPT_SUB_ITEMS.get(dept, [])
    n_sub = len(subs)
    # v2026.04b: 자동연동 11열 (C10=담당자는 사용자 입력) + 세부항목(n_sub) + 상태 + 메모
    # 순서: NO, 관리코드, 수주번호, 고객사, 모델, 품명, PO유형, 영업단계, 진행상태, 담당자, 전체진척률(%)
    auto_cols = ["NO","관리코드","수주번호","고객사","모델","품명",
                 "PO유형","영업단계","진행상태","담당자","전체진척률(%)"]
    n_auto = len(auto_cols)                          # 11
    C_DEPT_PIC = 10                                   # 담당자 (사용자 직접 입력)
    C_PROG     = 11                                   # 전체진척률(%) (auto)
    c_status = n_auto + n_sub + 1
    c_memo   = n_auto + n_sub + 2
    total_col = c_memo

    # R1
    setup_r1(ws, f"㈜케이엔케이 │ {TYPE_NAME} │ {dept} │ {YEAR}", total_col)

    # R3 가이드 — C10(담당자)만 input, 나머지 auto_cols는 auto
    r3 = {}
    for c in range(1, n_auto + 1):
        r3[c] = "auto"
    r3[C_DEPT_PIC] = "input"                          # 담당자 — 부서원 직접 입력
    for c in range(n_auto + 1, n_auto + n_sub + 1):
        r3[c] = "input"                                # 세부항목 진척률
    r3[c_status] = "select"
    r3[c_memo] = "memo"
    apply_r3_guide(ws, total_col, r3)

    # R4 헤더
    labels = list(auto_cols)
    for sub in subs:
        labels.append(sub)
    labels.append("상태")
    labels.append("메모")
    r4_map = {}
    for c in range(1, n_auto + 1):
        r4_map[c] = "id"
    r4_map[C_DEPT_PIC] = "dept"                       # 담당자는 부서색
    for c in range(n_auto + 1, total_col + 1):
        r4_map[c] = "status"
    apply_r4_header(ws, total_col, labels, r4_map)

    # 서식 — 세부항목 + 전체진척률 퍼센트
    pct_cols = list(range(n_auto + 1, n_auto + n_sub + 1)) + [C_PROG]
    format_data_rows(ws, total_col, pct_cols=pct_cols)

    # 드롭다운: 상태
    dept_statuses = ["미착수","진행중","완료","N/A"]
    _add_dropdown(ws, c_status, dept_statuses)
    # 세부항목 진척률: 0~100 (제한 없이 숫자 입력)

    # 보호
    apply_protection(ws, total_col, r3)
    auto_fit_columns(ws)

    wb.save(fp)
    print(f"    → 저장 완료")
    return fp



def clean_existing():
    """기존 파일 삭제"""
    patterns = [
        os.path.join(DIR_DATA, "*.xlsx"),
    ]
    count = 0
    for pattern in patterns:
        for f in glob.glob(pattern):
            os.remove(f)
            count += 1
    if count:
        print(f"  [CLEAN] {count}개 기존 파일 삭제")


def build_all():
    """전체 빌드 총괄"""
    print("=" * 60)
    print(f"  KNK PMS V4 — {TYPE_NAME} 빌드 시작")
    print(f"  시각: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # 1. 기존 파일 삭제
    clean_existing()

    # 2. PMS 통합 파일
    build_pms()

    # 3. 부서입력 파일 (8개 부서)
    for dept in DEPTS:
        build_dept_file(dept)


    print("=" * 60)
    print(f"  빌드 완료! 총 {1 + len(DEPTS)}개 파일 생성")
    print("=" * 60)


if __name__ == "__main__":
    build_all()
