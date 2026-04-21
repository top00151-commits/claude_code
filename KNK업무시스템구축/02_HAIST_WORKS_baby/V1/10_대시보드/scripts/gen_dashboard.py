"""
╔══════════════════════════════════════════════════════════════╗
║  KNK PMS V4 — 10_대시보드 gen_dashboard.py                    ║
║  5개 소스 통합 집계 → 경영현황 대시보드 생성                   ║
║  실행: python gen_dashboard.py                                ║
╚══════════════════════════════════════════════════════════════╝
"""

import os, sys, glob, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import *


# ═══════════════════════════════════════════════════════════════
# 데이터 수집
# ═══════════════════════════════════════════════════════════════

def _val(ws, r, c):
    v = ws.cell(row=r, column=c).value
    return v if v is not None else ""

def _num(ws, r, c):
    v = ws.cell(row=r, column=c).value
    try: return float(v)
    except (TypeError, ValueError): return 0


# R4 헤더 텍스트 기반 컬럼 인덱스 룩업 (캐시 포함)
_header_cache = {}
def _header_index(ws, header_text):
    """ws의 R4에서 header_text와 부분 일치하는 첫 컬럼 번호 반환 (없으면 None)."""
    key = (id(ws), header_text)
    if key in _header_cache:
        return _header_cache[key]
    norm_target = str(header_text).replace("\n", "").replace(" ", "").strip()
    found = None
    for c in range(1, ws.max_column + 1):
        h = ws.cell(4, c).value
        if not h:
            continue
        norm = str(h).replace("\n", "").replace(" ", "").strip()
        if norm_target in norm:
            found = c
            break
    _header_cache[key] = found
    return found

def _val_by_header(ws, r, header_text):
    c = _header_index(ws, header_text)
    return _val(ws, r, c) if c else ""

def _num_by_header(ws, r, header_text):
    c = _header_index(ws, header_text)
    return _num(ws, r, c) if c else 0


def _collect_product_data(source_dir, type_name):
    """완제품(01, 02) PMS에서 프로젝트 데이터 수집"""
    projects = []
    pms_files = glob.glob(os.path.join(source_dir, "KNK_*_PMS_*.xlsx"))
    if not pms_files:
        return projects

    fp = pms_files[0]
    try:
        wb = load_workbook(fp, data_only=True)
        ws = wb["1_프로젝트등록"]
        for r in range(5, ws.max_row + 1):
            code = _val(ws, r, 2)
            if not code:
                continue
            # v2026.04c: 01 검사기(23열, C6=제품구분)와 02 자동화(22열)의 컬럼 위치가 다름.
            # R4 헤더 텍스트 기반 룩업으로 구조 차이 흡수.
            projects.append({
                "type": type_name,
                "code": str(code),
                "sj":       _val_by_header(ws, r, "수주번호"),
                "cust":     _val_by_header(ws, r, "고객사"),
                "cust_pic": _val_by_header(ws, r, "고객사담당자"),
                "prodtype": _val_by_header(ws, r, "제품구분"),    # 01만 존재 (02는 None)
                "model":    _val_by_header(ws, r, "모델"),
                "prod":     _val_by_header(ws, r, "품명"),
                "amount":   _num_by_header(ws, r, "금액"),
                "currency": _val_by_header(ws, r, "통화"),
                "status":   _val_by_header(ws, r, "진행상태"),
                "prog":     _num_by_header(ws, r, "진척률"),
                "due":      _val_by_header(ws, r, "납기일"),
                "dday":     _num_by_header(ws, r, "D-day"),
                "update": "",
            })
        wb.close()
    except Exception as e:
        print(f"  [WARN] {type_name} 읽기 실패: {e}")

    return projects


def _collect_parts_data(source_dir, type_name):
    """부품출하(03, 04) 마스터에서 데이터 수집"""
    projects = []
    master_files = glob.glob(os.path.join(source_dir, "KNK_*_마스터_*.xlsx"))
    if not master_files:
        return projects

    fp = master_files[0]
    try:
        wb = load_workbook(fp, data_only=True)
        ws = wb.active
        for r in range(5, ws.max_row + 1):
            code = _val(ws, r, 2)
            if not code:
                continue
            projects.append({
                "type": type_name,
                "code": str(code),
                "sj": _val(ws, r, 3),
                "cust": _val(ws, r, 4),
                "prod": _val(ws, r, 5),
                "amount": _num(ws, r, 16),
                "currency": "KRW",
                "status": _val(ws, r, 19) if ws.max_column >= 19 else "",
                "prog": _num(ws, r, 11),
                "due": "",
                "dday": 0,
                "update": "",
                "inv_usd": _num(ws, r, 17),
            })
        wb.close()
    except Exception as e:
        print(f"  [WARN] {type_name} 읽기 실패: {e}")

    return projects


def _collect_consumables_data(source_dir):
    """소모품(05) 출하관리에서 데이터 수집"""
    projects = []
    files = glob.glob(os.path.join(source_dir, "KNK_*_출하대장_*.xlsx"))
    if not files:
        return projects

    fp = files[0]
    try:
        wb = load_workbook(fp, data_only=True)
        ws = wb["출하관리"]
        for r in range(5, ws.max_row + 1):
            if not _val(ws, r, 3):  # 품명 없으면 스킵
                continue
            projects.append({
                "type": "소모품",
                "code": str(_val(ws, r, 1)),
                "sj": _val(ws, r, 2),
                "cust": "",
                "prod": _val(ws, r, 3),
                "amount": _num(ws, r, 12),
                "currency": _val(ws, r, 10),
                "status": _val(ws, r, 25),
                "prog": 0,
                "due": "",
                "dday": 0,
                "update": "",
            })
        wb.close()
    except Exception as e:
        print(f"  [WARN] 소모품 읽기 실패: {e}")

    return projects


# ═══════════════════════════════════════════════════════════════
# 대시보드 생성
# ═══════════════════════════════════════════════════════════════

def generate():
    """경영현황 대시보드 생성"""
    print("=" * 60)
    print(f"  KNK PMS V4 — 경영현황 대시보드 생성")
    print(f"  시각: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # 데이터 수집
    all_projects = []
    type_summaries = {}

    # 2026-04-15: 03/04 부품출하 보류 — 01/02/05 3-모듈로 축소
    source_configs = [
        ("01_검사기_완제품", "검사기_완제품", "product"),
        ("02_자동화_완제품", "자동화_완제품", "product"),
        ("05_소모품", "소모품", "consumables"),
    ]

    for key, type_name, category in source_configs:
        src = SOURCES.get(key, "")
        if not os.path.exists(src):
            print(f"  [SKIP] {key} 폴더 없음")
            continue

        if category == "product":
            data = _collect_product_data(src, type_name)
        elif category == "parts":
            data = _collect_parts_data(src, type_name)
        else:
            data = _collect_consumables_data(src)

        all_projects.extend(data)

        # 유형별 집계
        active = [p for p in data if p.get("status") not in ("납품완료", "완료", "취소")]
        done = [p for p in data if p.get("status") in ("납품완료", "완료")]
        total_krw = sum(p["amount"] for p in data if p.get("currency", "KRW") != "USD")
        total_usd = sum(p["amount"] for p in data if p.get("currency") == "USD")
        inv_usd = sum(p.get("inv_usd", 0) for p in data)
        progs = [p["prog"] for p in active if p["prog"] and p["prog"] > 0]
        avg_prog = sum(progs) / len(progs) if progs else 0

        type_summaries[type_name] = {
            "active": len(active), "done": len(done), "total": len(data),
            "krw": total_krw, "usd": total_usd, "inv_usd": inv_usd,
            "avg_prog": avg_prog,
            "delayed": sum(1 for p in active if p.get("dday", 0) and p["dday"] < 0),
        }
        print(f"  {type_name}: {len(data)}건 수집")

    # 대시보드 생성
    fp = dashboard_path()
    wb = Workbook()

    # ── 시트1: 매출총괄 ──
    ws1 = wb.active
    ws1.title = "매출총괄"
    ws1.sheet_properties.tabColor = TAB_VIEW

    mc1 = SUMMARY_MAX_COL
    setup_r1(ws1, f"㈜케이엔케이 │ 경영현황 │ 매출총괄 │ {YEAR}", mc1)
    r3 = {c: "auto" for c in range(1, mc1 + 1)}
    apply_r3_guide(ws1, mc1, r3)
    r4 = {c: "id" for c in range(1, mc1 + 1)}
    apply_r4_header(ws1, mc1, SUMMARY_LABELS, r4)

    row = 5
    for type_name, s in type_summaries.items():
        ws1.cell(row=row, column=1).value = type_name
        ws1.cell(row=row, column=2).value = s["active"]
        ws1.cell(row=row, column=3).value = s["done"]
        ws1.cell(row=row, column=4).value = s["total"]
        ws1.cell(row=row, column=5).value = s["krw"]
        ws1.cell(row=row, column=5).number_format = ACCT
        ws1.cell(row=row, column=5).alignment = AL_R
        ws1.cell(row=row, column=6).value = s["usd"]
        ws1.cell(row=row, column=6).number_format = ACCT_USD
        ws1.cell(row=row, column=6).alignment = AL_R
        ws1.cell(row=row, column=7).value = s["inv_usd"]
        ws1.cell(row=row, column=7).number_format = ACCT_USD
        ws1.cell(row=row, column=7).alignment = AL_R
        ws1.cell(row=row, column=8).value = s["avg_prog"] / 100 if s["avg_prog"] else None
        ws1.cell(row=row, column=8).number_format = PCT
        ws1.cell(row=row, column=9).value = s["delayed"]
        row += 1

    # 합계행
    ws1.cell(row=row, column=1).value = "합계"
    ws1.cell(row=row, column=1).font = FT_BOLD
    for c in range(2, 10):
        total = sum(ws1.cell(row=r, column=c).value or 0 for r in range(5, row))
        if c == 8:  # 평균 진척률은 평균
            vals = [ws1.cell(row=r, column=c).value or 0 for r in range(5, row) if ws1.cell(row=r, column=c).value]
            ws1.cell(row=row, column=c).value = sum(vals) / len(vals) if vals else None
            ws1.cell(row=row, column=c).number_format = PCT
        else:
            ws1.cell(row=row, column=c).value = total
        if c in (5,):
            ws1.cell(row=row, column=c).number_format = ACCT
            ws1.cell(row=row, column=c).alignment = AL_R
        if c in (6, 7):
            ws1.cell(row=row, column=c).number_format = ACCT_USD
            ws1.cell(row=row, column=c).alignment = AL_R

    format_data_rows(ws1, mc1, money_cols=[5], money_usd_cols=[6, 7], pct_cols=[8], row_start=5, row_end=row)
    apply_protection(ws1, mc1, r3, row_end=row + 10)
    auto_fit_columns(ws1)

    # ── 시트2: 전체 프로젝트 목록 ──
    ws2 = wb.create_sheet("프로젝트목록")
    ws2.sheet_properties.tabColor = TAB_VIEW

    mc2 = PROJECT_MAX_COL
    setup_r1(ws2, f"㈜케이엔케이 │ 경영현황 │ 전체 프로젝트 │ {YEAR}", mc2)
    r3_2 = {c: "auto" for c in range(1, mc2 + 1)}
    apply_r3_guide(ws2, mc2, r3_2)
    r4_2 = {c: "id" for c in range(1, mc2 + 1)}
    apply_r4_header(ws2, mc2, PROJECT_LABELS, r4_2)

    row = 5
    for idx, p in enumerate(all_projects, 1):
        ws2.cell(row=row, column=1).value = idx
        ws2.cell(row=row, column=2).value = p["type"]
        ws2.cell(row=row, column=3).value = p["code"]
        ws2.cell(row=row, column=4).value = p["sj"]
        ws2.cell(row=row, column=5).value = p["cust"]
        ws2.cell(row=row, column=6).value = p["prod"]
        ws2.cell(row=row, column=7).value = p["amount"]
        ws2.cell(row=row, column=7).number_format = ACCT
        ws2.cell(row=row, column=7).alignment = AL_R
        ws2.cell(row=row, column=8).value = p["status"]
        ws2.cell(row=row, column=9).value = p["prog"] / 100 if p["prog"] else None
        ws2.cell(row=row, column=9).number_format = PCT
        ws2.cell(row=row, column=10).value = p["due"]
        ws2.cell(row=row, column=11).value = p["dday"] if p["dday"] else None
        ws2.cell(row=row, column=12).value = p["update"]
        row += 1

    format_data_rows(ws2, mc2, money_cols=[7], pct_cols=[9], row_start=5, row_end=max(row, 2000))
    apply_protection(ws2, mc2, r3_2)
    auto_fit_columns(ws2)

    wb.save(fp)
    print(f"\n  대시보드 저장: {fp}")
    print(f"  총 {len(all_projects)}건 프로젝트 집계")
    print("=" * 60)
    return fp

if __name__ == "__main__":
    generate()
