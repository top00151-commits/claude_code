"""
KNK 일일업무일지 v2 - Phase 1 MVP
Task Card 기반 일일업무 + 팀장 뷰 + 경영진 대시보드 + 개인 히스토리
"""
from fastapi import FastAPI, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
import os, io, calendar, tempfile
from datetime import datetime, timedelta, date
from .i18n import LANGS, t as i18n_t, get_all_translations
from .database import (db_session, init_db, seed_all, seed_sample_tasks, hash_pw,
                        parse_mgmt_xls, import_mgmt_rows,
                        regenerate_user_passwords, build_password_xlsx,
                        add_comment, get_task_comments, delete_comment,
                        get_notifications, count_unread, mark_notification_read,
                        mark_all_read,
                        log_activity, log_activity_standalone, get_activities,
                        add_reaction, get_reactions, get_reactions_bulk, get_meta_bulk,
                        notify_status_change, get_user_search,
                        upsert_retro, get_retro, search_all, detect_bottlenecks,
                        delegate_task, get_delegations, resolve_delegation)

BASE = os.path.dirname(os.path.dirname(__file__))
app = FastAPI(title="KNK 일일업무일지 v2")
app.add_middleware(SessionMiddleware, secret_key="knk-haist-2026-phase1")
app.mount("/static", StaticFiles(directory=os.path.join(BASE, "static")), name="static")
tpl = Jinja2Templates(directory=os.path.join(BASE, "app", "templates"))


@app.on_event("startup")
def startup():
    init_db()
    seed_all()
    seed_sample_tasks(14)


CATEGORIES = ["설계", "제조", "고객대응", "회의", "출장", "검토", "개발", "구매", "품질", "기타"]
STATUSES = ["진행중", "완료", "지연", "대기", "보류"]


# =====================================================
# HELPERS
# =====================================================
def ctx(request, name, **kwargs):
    # 사용자 언어 결정
    user = kwargs.get("user")
    lang = "ko"
    if user and isinstance(user, dict):
        lang = user.get("lang") or "ko"
    elif hasattr(request, "session"):
        lang = request.session.get("lang", "ko")

    # 번역 사전 생성
    i = get_all_translations(lang)

    base = {
        "categories": CATEGORIES,
        "statuses": STATUSES,
        "today": date.today().isoformat(),
        "now": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "lang": lang,
        "i": i,
        "LANGS": LANGS,
        # HAIST WORKS 브랜드 (통합 후)
        "app_name": "HAIST WORKS",
        "app_subtitle": "KNK 통합 업무 플랫폼",
        "brand_slogan": "Human & AI create the Best",
    }
    # 글로벌 알림 카운트 (로그인 상태일 때만)
    uid = request.session.get("user_id") if hasattr(request, "session") else None
    if uid:
        try:
            base["unread_notif"] = count_unread(uid)
        except Exception:
            base["unread_notif"] = 0
    else:
        base["unread_notif"] = 0
    base.update(kwargs)
    return tpl.TemplateResponse(request=request, name=name, context=base)


def get_user(req: Request):
    uid = req.session.get("user_id")
    if not uid:
        return None
    with db_session() as c:
        row = c.execute(
            """SELECT u.*, t.name AS team_name, t.code AS team_code, t.is_lab AS team_is_lab,
                      t.sector AS team_sector
               FROM users u LEFT JOIN teams t ON u.team_id=t.id
               WHERE u.id=? AND u.is_active=1""",
            (uid,),
        ).fetchone()
        return dict(row) if row else None


def require(req: Request, roles=None):
    u = get_user(req)
    if not u:
        return None
    if roles and u["role"] not in roles:
        return None
    return u


def can_use_logistics(user) -> bool:
    """HAIST WORKS 물류 모듈 접근 권한.
    - admin / ceo / executive: 항상 허용
    - 그 외: users.can_use_logistics 플래그가 1일 때만
    """
    if not user:
        return False
    role = user.get("role") if isinstance(user, dict) else user["role"]
    if role in ("admin", "ceo", "executive"):
        return True
    flag = user.get("can_use_logistics") if isinstance(user, dict) else user["can_use_logistics"]
    return bool(flag)


def status_color(s):
    return {
        "진행중": ("#FFF8E1", "#F57F17"),
        "완료":   ("#E8F5E9", "#2E7D32"),
        "지연":   ("#FFEBEE", "#A5282C"),
        "대기":   ("#F5F5F5", "#4A4A4A"),
        "보류":   ("#F5F5F5", "#4A4A4A"),
    }.get(s, ("#F5F5F5", "#4A4A4A"))


def fetch_projects(c):
    return [dict(r) for r in c.execute(
        """SELECT p.id, p.code, p.name, p.type, c.name AS customer_name
           FROM projects p LEFT JOIN customers c ON p.customer_id=c.id
           WHERE p.status='진행중' ORDER BY p.id"""
    ).fetchall()]


def fetch_customers(c):
    return [dict(r) for r in c.execute(
        "SELECT id, name, tier FROM customers ORDER BY tier DESC, name"
    ).fetchall()]


# =====================================================
# AUTH
# =====================================================
@app.get("/login", response_class=HTMLResponse)
async def login_page(req: Request):
    return ctx(req, "login.html", error=None)


@app.post("/login")
async def login_post(req: Request, login_id: str = Form(...), password: str = Form(...)):
    with db_session() as c:
        u = c.execute(
            "SELECT * FROM users WHERE login_id=? AND password=? AND is_active=1",
            (login_id.strip(), hash_pw(password)),
        ).fetchone()
        if u:
            req.session["user_id"] = u["id"]
            r = u["role"]
            if r == "admin":
                return RedirectResponse("/dashboard", 303)
            if r == "ceo":
                return RedirectResponse("/dashboard", 303)
            if r in ("leader", "executive"):
                return RedirectResponse("/team", 303)
            return RedirectResponse("/daily", 303)
    return ctx(req, "login.html", error="아이디 또는 비밀번호가 올바르지 않습니다.")


@app.get("/logout")
async def logout(req: Request):
    req.session.clear()
    return RedirectResponse("/login", 303)


# =====================================================
# ROOT
# =====================================================
@app.get("/")
async def root(req: Request):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    return RedirectResponse("/home", 303)


# =====================================================
# HOME — 직관적 단일 페이지 (역할별 자동 분기)
# =====================================================
@app.get("/home", response_class=HTMLResponse)
@app.get("/home/{sel_date}", response_class=HTMLResponse)
async def home_page(req: Request, sel_date: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if not sel_date:
        sel_date = date.today().isoformat()

    prev_d = (datetime.strptime(sel_date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
    next_d = (datetime.strptime(sel_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")

    with db_session() as c:
        teams = [dict(r) for r in c.execute(
            """SELECT t.*, u.name AS leader_name, u.rank AS leader_rank
               FROM teams t LEFT JOIN users u ON t.leader_id=u.id
               ORDER BY t.display_order"""
        ).fetchall()]

        projects = fetch_projects(c)
        customers = fetch_customers(c)

        # 전체 사용자 수
        total_users = c.execute(
            "SELECT COUNT(*) FROM users WHERE is_active=1 AND role!='admin'"
        ).fetchone()[0]
        today_reporters = c.execute(
            "SELECT COUNT(DISTINCT user_id) FROM tasks WHERE work_date=?",
            (sel_date,),
        ).fetchone()[0]
        participation_rate = round(today_reporters * 100 / total_users) if total_users else 0

        # 팀별 데이터 구축
        team_data = []
        for tm in teams:
            members = [dict(r) for r in c.execute(
                """SELECT id, name, rank, role FROM users
                   WHERE team_id=? AND is_active=1
                   ORDER BY CASE role WHEN 'ceo' THEN 0 WHEN 'executive' THEN 1
                            WHEN 'leader' THEN 2 ELSE 3 END, id""",
                (tm["id"],),
            ).fetchall()]
            mids = [m["id"] for m in members]
            if not mids:
                continue
            ph = ",".join("?" * len(mids))
            tasks = [dict(r) for r in c.execute(
                f"""SELECT t.*, u.name AS user_name, u.rank AS user_rank,
                           p.name AS project_name, p.code AS project_code,
                           cu.name AS customer_name,
                           (SELECT COUNT(*) FROM task_comments WHERE task_id=t.id) AS comment_count
                    FROM tasks t JOIN users u ON t.user_id=u.id
                    LEFT JOIN projects p ON t.project_id=p.id
                    LEFT JOIN customers cu ON t.customer_id=cu.id
                    WHERE t.user_id IN ({ph}) AND t.work_date=?
                    ORDER BY u.id, CASE t.status WHEN '지연' THEN 0 WHEN '진행중' THEN 1
                             WHEN '대기' THEN 2 WHEN '보류' THEN 3 ELSE 4 END, t.id""",
                mids + [sel_date],
            ).fetchall()]

            reported = len(set(t["user_id"] for t in tasks))
            delay_count = len([t for t in tasks if t["status"] == "지연"])
            progress_count = len([t for t in tasks if t["status"] == "진행중"])
            done_count = len([t for t in tasks if t["status"] == "완료"])

            # 신호등: 지연 있으면 빨강, 참여 50% 미만 노랑, 나머지 초록
            if delay_count > 0:
                signal = "red"
            elif reported < len(members) * 0.5:
                signal = "yellow"
            else:
                signal = "green"

            # 멤버별 그룹
            member_tasks = {}
            for m in members:
                mt = [t for t in tasks if t["user_id"] == m["id"]]
                member_tasks[m["id"]] = mt

            team_data.append({
                "team": tm,
                "members": members,
                "tasks": tasks,
                "member_tasks": member_tasks,
                "reported": reported,
                "total_members": len(members),
                "delay": delay_count,
                "progress": progress_count,
                "done": done_count,
                "total": len(tasks),
                "signal": signal,
            })

        # 내 업무 (직원/팀장 공통)
        my_tasks = [dict(r) for r in c.execute(
            """SELECT t.*, p.name AS project_name, p.code AS project_code,
                      cu.name AS customer_name,
                      (SELECT COUNT(*) FROM task_comments WHERE task_id=t.id) AS comment_count
               FROM tasks t
               LEFT JOIN projects p ON t.project_id=p.id
               LEFT JOIN customers cu ON t.customer_id=cu.id
               WHERE t.user_id=? AND t.work_date=?
               ORDER BY CASE t.status WHEN '지연' THEN 0 WHEN '진행중' THEN 1
                        WHEN '대기' THEN 2 WHEN '보류' THEN 3 ELSE 4 END, t.id""",
            (u["id"], sel_date),
        ).fetchall()]

        # 어제 미완료 (이월 후보)
        yday = (datetime.strptime(sel_date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
        pending_yday = [dict(r) for r in c.execute(
            """SELECT t.*, p.name AS project_name, cu.name AS customer_name
               FROM tasks t LEFT JOIN projects p ON t.project_id=p.id
               LEFT JOIN customers cu ON t.customer_id=cu.id
               WHERE t.user_id=? AND t.work_date=? AND t.status IN ('진행중','지연','대기')
               AND NOT EXISTS (SELECT 1 FROM tasks t2 WHERE t2.carry_from_id=t.id AND t2.work_date=?)
               ORDER BY t.id""",
            (u["id"], yday, sel_date),
        ).fetchall()]

        # 전사 요약
        all_delay = sum(td["delay"] for td in team_data)
        all_tasks = sum(td["total"] for td in team_data)

    # HAIST WORKS — 물류 KPI (간단 집계)
    try:
        from . import database as _logi_db
        logi_parts_stats = _logi_db.parts_count()
        logi_proj_stats = _logi_db.projects_count_logi()
    except Exception:
        logi_parts_stats = {"total": 0, "active": 0, "by_div": {}}
        logi_proj_stats = {"total": 0, "with_code": 0, "in_progress": 0,
                           "by_div": {}, "by_stage": {}}

    return ctx(
        req, "home.html",
        user=u, sel_date=sel_date, prev_date=prev_d, next_date=next_d,
        team_data=team_data, my_tasks=my_tasks, pending_yday=pending_yday,
        projects=projects, customers=customers,
        participation_rate=participation_rate,
        today_reporters=today_reporters, total_users=total_users,
        all_delay=all_delay, all_tasks=all_tasks,
        logi_parts_stats=logi_parts_stats, logi_proj_stats=logi_proj_stats,
    )


# =====================================================
# DAILY — 개인 일일 업무카드
# =====================================================
@app.get("/daily", response_class=HTMLResponse)
@app.get("/daily/{sel_date}", response_class=HTMLResponse)
async def daily_page(req: Request, sel_date: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if not sel_date:
        sel_date = date.today().isoformat()

    prev_d = (datetime.strptime(sel_date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
    next_d = (datetime.strptime(sel_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")

    with db_session() as c:
        tasks = [dict(r) for r in c.execute(
            """SELECT t.*, p.name AS project_name, p.code AS project_code,
                      c.name AS customer_name,
                      (SELECT COUNT(*) FROM task_comments WHERE task_id=t.id) AS comment_count,
                      (SELECT MAX(is_ceo_request) FROM task_comments WHERE task_id=t.id) AS has_ceo_req
               FROM tasks t
               LEFT JOIN projects p ON t.project_id=p.id
               LEFT JOIN customers c ON t.customer_id=c.id
               WHERE t.user_id=? AND t.work_date=?
               ORDER BY t.status, t.id""",
            (u["id"], sel_date),
        ).fetchall()]
        # 각 카드에 댓글 첨부
        for t in tasks:
            t["comments"] = [dict(r) for r in c.execute(
                """SELECT tc.*, u.name AS author_name, u.rank AS author_rank, u.role AS author_role
                   FROM task_comments tc JOIN users u ON tc.author_id=u.id
                   WHERE tc.task_id=? ORDER BY tc.created_at""",
                (t["id"],),
            ).fetchall()]

        # 어제 미완료 (carry-forward 후보)
        yday = (datetime.strptime(sel_date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
        pending_yday = [dict(r) for r in c.execute(
            """SELECT t.*, p.name AS project_name, c.name AS customer_name
               FROM tasks t
               LEFT JOIN projects p ON t.project_id=p.id
               LEFT JOIN customers c ON t.customer_id=c.id
               WHERE t.user_id=? AND t.work_date=? AND t.status IN ('진행중','지연','대기')
               AND NOT EXISTS (
                   SELECT 1 FROM tasks t2 WHERE t2.carry_from_id=t.id AND t2.work_date=?
               )
               ORDER BY t.id""",
            (u["id"], yday, sel_date),
        ).fetchall()]

        projects = fetch_projects(c)
        customers = fetch_customers(c)

        # 통계 - 이번 주
        wk_mon = (datetime.strptime(sel_date, "%Y-%m-%d") -
                  timedelta(days=datetime.strptime(sel_date, "%Y-%m-%d").weekday())).strftime("%Y-%m-%d")
        wk_sun = (datetime.strptime(wk_mon, "%Y-%m-%d") + timedelta(days=6)).strftime("%Y-%m-%d")
        week_stats = c.execute(
            """SELECT COUNT(*) AS total,
                      SUM(CASE WHEN status='완료' THEN 1 ELSE 0 END) AS done,
                      SUM(CASE WHEN status='진행중' THEN 1 ELSE 0 END) AS progress,
                      SUM(CASE WHEN status='지연' THEN 1 ELSE 0 END) AS delay,
                      COALESCE(SUM(hours),0) AS hours
               FROM tasks WHERE user_id=? AND work_date>=? AND work_date<=?""",
            (u["id"], wk_mon, wk_sun),
        ).fetchone()
        week_stats = dict(week_stats)

    return ctx(
        req, "daily.html",
        user=u, tasks=tasks, sel_date=sel_date,
        prev_date=prev_d, next_date=next_d,
        pending_yday=pending_yday,
        projects=projects, customers=customers,
        week_stats=week_stats, week_range=f"{wk_mon} ~ {wk_sun}",
    )


@app.post("/api/task")
async def api_create_task(req: Request):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    d = await req.json()
    with db_session() as c:
        cur = c.execute(
            """INSERT INTO tasks(user_id, work_date, title, category, project_id, customer_id,
                                  status, hours, notes, next_plan, due_date)
               VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
            (
                u["id"],
                d.get("work_date") or date.today().isoformat(),
                (d.get("title") or "").strip(),
                d.get("category") or "기타",
                d.get("project_id") or None,
                d.get("customer_id") or None,
                d.get("status") or "진행중",
                float(d.get("hours") or 0),
                d.get("notes") or "",
                (d.get("next_plan") or "").strip(),
                d.get("due_date") or None,
            ),
        )
        new_id = cur.lastrowid
        log_activity(c, u["id"], "task_create",
                     title=f"{u['name']} 신규 카드: {(d.get('title') or '')[:60]}",
                     body=(d.get("notes") or "")[:200],
                     task_id=new_id,
                     project_id=d.get("project_id") or None,
                     team_id=u.get("team_id"))
    # 지연으로 생성된 경우도 알림
    if (d.get("status") or "") == "지연":
        notify_status_change(new_id, u["id"], "", "지연")
    return JSONResponse({"ok": True, "id": new_id})


@app.put("/api/task/{tid}")
async def api_update_task(req: Request, tid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    d = await req.json()
    with db_session() as c:
        prev = c.execute("SELECT user_id, status, title, project_id FROM tasks WHERE id=?", (tid,)).fetchone()
        if not prev or prev["user_id"] != u["id"]:
            return JSONResponse({"error": "권한 없음"}, 403)
        new_status = d.get("status") or "진행중"
        c.execute(
            """UPDATE tasks SET title=?, category=?, project_id=?, customer_id=?,
                               status=?, hours=?, notes=?, next_plan=?, due_date=?,
                               updated_at=datetime('now','localtime')
               WHERE id=?""",
            (
                (d.get("title") or "").strip(),
                d.get("category") or "기타",
                d.get("project_id") or None,
                d.get("customer_id") or None,
                new_status,
                float(d.get("hours") or 0),
                d.get("notes") or "",
                (d.get("next_plan") or "").strip(),
                d.get("due_date") or None,
                tid,
            ),
        )
        if prev["status"] != new_status:
            log_activity(c, u["id"], "task_status",
                         title=f"{u['name']}: {prev['title'][:50]} — {prev['status']} → {new_status}",
                         task_id=tid, project_id=prev["project_id"], team_id=u.get("team_id"))
    if prev["status"] != new_status and new_status == "지연":
        notify_status_change(tid, u["id"], prev["status"], new_status)
    return JSONResponse({"ok": True})


@app.delete("/api/task/{tid}")
async def api_delete_task(req: Request, tid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    with db_session() as c:
        c.execute("DELETE FROM tasks WHERE id=? AND user_id=?", (tid, u["id"]))
    return JSONResponse({"ok": True})


@app.post("/api/task/{tid}/status")
async def api_quick_status(req: Request, tid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    d = await req.json()
    new_status = d.get("status") or "진행중"
    with db_session() as c:
        prev = c.execute("SELECT status, title, user_id, project_id FROM tasks WHERE id=?", (tid,)).fetchone()
        if not prev or prev["user_id"] != u["id"]:
            return JSONResponse({"error":"권한 없음"}, 403)
        c.execute(
            "UPDATE tasks SET status=?, updated_at=datetime('now','localtime') WHERE id=?",
            (new_status, tid),
        )
        if prev["status"] != new_status:
            log_activity(c, u["id"], "task_status",
                         title=f"{u['name']}: {prev['title'][:50]} — {prev['status']} → {new_status}",
                         task_id=tid, project_id=prev["project_id"], team_id=u.get("team_id"))
    if prev["status"] != new_status and new_status == "지연":
        notify_status_change(tid, u["id"], prev["status"], new_status)
    return JSONResponse({"ok": True})


@app.post("/api/carry-forward")
async def api_carry_forward(req: Request):
    """어제 미완료 카드 → 오늘로 이월 (전부 또는 선택)"""
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    d = await req.json()
    target = d.get("date") or date.today().isoformat()
    ids = d.get("ids") or []
    with db_session() as c:
        if not ids:
            yday = (datetime.strptime(target, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
            rows = c.execute(
                """SELECT id FROM tasks WHERE user_id=? AND work_date=?
                   AND status IN ('진행중','지연','대기')""",
                (u["id"], yday),
            ).fetchall()
            ids = [r["id"] for r in rows]
        cnt = 0
        for src_id in ids:
            src = c.execute("SELECT * FROM tasks WHERE id=? AND user_id=?",
                            (src_id, u["id"])).fetchone()
            if not src:
                continue
            exists = c.execute(
                "SELECT id FROM tasks WHERE carry_from_id=? AND work_date=?",
                (src_id, target),
            ).fetchone()
            if exists:
                continue
            c.execute(
                """INSERT INTO tasks(user_id, work_date, title, category, project_id,
                                      customer_id, status, hours, notes, due_date, carry_from_id)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    u["id"], target, src["title"], src["category"], src["project_id"],
                    src["customer_id"], "진행중", 0, src["notes"], src["due_date"], src_id,
                ),
            )
            cnt += 1
    return JSONResponse({"ok": True, "count": cnt})


# =====================================================
# SUMMARY — 통합 요약 (일/주/월 × 개인/부서/전사)
# =====================================================
@app.get("/summary", response_class=HTMLResponse)
async def summary_page(req: Request, period: str = "weekly", scope: str = "me", ref: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if period not in ("daily", "weekly", "monthly"):
        period = "weekly"
    if scope not in ("me", "team", "all"):
        scope = "me"
    # 권한: 일반 직원은 me/team만, all은 ceo/admin/executive
    if scope == "all" and u["role"] not in ("ceo", "admin", "executive"):
        scope = "team" if u.get("team_id") else "me"
    if scope == "team" and not u.get("team_id"):
        scope = "me"

    today = date.today()
    if not ref:
        ref = today.isoformat()
    rd = datetime.strptime(ref, "%Y-%m-%d").date()

    # 기간 범위 계산
    if period == "daily":
        frm = to = ref
        prev_ref = (rd - timedelta(days=1)).isoformat()
        next_ref = (rd + timedelta(days=1)).isoformat()
        period_label = ref
    elif period == "weekly":
        mon = rd - timedelta(days=rd.weekday())
        sun = mon + timedelta(days=6)
        frm, to = mon.isoformat(), sun.isoformat()
        prev_ref = (mon - timedelta(days=7)).isoformat()
        next_ref = (mon + timedelta(days=7)).isoformat()
        period_label = f"{frm} ~ {to}"
    else:  # monthly
        first = rd.replace(day=1)
        next_month_first = (first + timedelta(days=32)).replace(day=1)
        last = next_month_first - timedelta(days=1)
        frm, to = first.isoformat(), last.isoformat()
        prev_first = (first - timedelta(days=1)).replace(day=1)
        prev_ref = prev_first.isoformat()
        next_ref = next_month_first.isoformat()
        period_label = first.strftime("%Y년 %m월")

    # 스코프 필터 SQL 조각
    if scope == "me":
        scope_filter = "AND t.user_id = ?"
        scope_args = (u["id"],)
        scope_label = f"{u['name']} {u['rank']}"
    elif scope == "team":
        scope_filter = "AND uu.team_id = ?"
        scope_args = (u["team_id"],)
        scope_label = u["team_name"] or "(부서 미배정)"
    else:
        scope_filter = ""
        scope_args = ()
        scope_label = "전사"

    with db_session() as c:
        # 통계
        sql = f"""SELECT COUNT(*) AS total,
                         SUM(CASE WHEN t.status='완료' THEN 1 ELSE 0 END) AS done,
                         SUM(CASE WHEN t.status='진행중' THEN 1 ELSE 0 END) AS progress,
                         SUM(CASE WHEN t.status='지연' THEN 1 ELSE 0 END) AS delay,
                         SUM(CASE WHEN t.status='대기' THEN 1 ELSE 0 END) AS waiting,
                         COALESCE(SUM(t.hours),0) AS hours
                  FROM tasks t JOIN users uu ON t.user_id = uu.id
                  WHERE t.work_date BETWEEN ? AND ? {scope_filter}"""
        stats = dict(c.execute(sql, (frm, to) + scope_args).fetchone())
        for k in stats:
            stats[k] = stats[k] or 0
        completion_rate = round(stats["done"] * 100 / stats["total"]) if stats["total"] else 0

        # 카드 목록 (내러티브용)
        cards_sql = f"""SELECT t.*, uu.name AS user_name, uu.rank, tm.name AS team_name,
                              p.name AS project_name, cu.name AS customer_name
                       FROM tasks t JOIN users uu ON t.user_id = uu.id
                       LEFT JOIN teams tm ON uu.team_id = tm.id
                       LEFT JOIN projects p ON t.project_id=p.id
                       LEFT JOIN customers cu ON t.customer_id=cu.id
                       WHERE t.work_date BETWEEN ? AND ? {scope_filter}
                       ORDER BY CASE t.status WHEN '지연' THEN 0 WHEN '진행중' THEN 1
                                              WHEN '대기' THEN 2 ELSE 3 END, t.work_date DESC, t.id"""
        cards = [dict(r) for r in c.execute(cards_sql, (frm, to) + scope_args).fetchall()]

        # 팀별/사용자별/프로젝트별 집계
        team_agg = [dict(r) for r in c.execute(
            f"""SELECT tm.name AS team_name, COUNT(*) AS cnt,
                       SUM(CASE WHEN t.status='완료' THEN 1 ELSE 0 END) AS done,
                       SUM(CASE WHEN t.status='지연' THEN 1 ELSE 0 END) AS delay,
                       COALESCE(SUM(t.hours),0) AS hours
                FROM tasks t JOIN users uu ON t.user_id = uu.id
                LEFT JOIN teams tm ON uu.team_id = tm.id
                WHERE t.work_date BETWEEN ? AND ? {scope_filter}
                GROUP BY tm.name ORDER BY cnt DESC""",
            (frm, to) + scope_args,
        ).fetchall()]

        user_agg = [dict(r) for r in c.execute(
            f"""SELECT uu.name AS user_name, uu.rank, COUNT(*) AS cnt,
                       SUM(CASE WHEN t.status='완료' THEN 1 ELSE 0 END) AS done,
                       COALESCE(SUM(t.hours),0) AS hours
                FROM tasks t JOIN users uu ON t.user_id = uu.id
                WHERE t.work_date BETWEEN ? AND ? {scope_filter}
                GROUP BY uu.id ORDER BY cnt DESC LIMIT 20""",
            (frm, to) + scope_args,
        ).fetchall()]

        project_agg = [dict(r) for r in c.execute(
            f"""SELECT p.name AS project_name, p.code AS project_code, COUNT(*) AS cnt,
                       COALESCE(SUM(t.hours),0) AS hours
                FROM tasks t JOIN users uu ON t.user_id = uu.id
                JOIN projects p ON t.project_id = p.id
                WHERE t.work_date BETWEEN ? AND ? {scope_filter}
                GROUP BY p.id ORDER BY cnt DESC LIMIT 15""",
            (frm, to) + scope_args,
        ).fetchall()]

        # 월간일 때만 전월 대비 비교
        prev_stats = None
        delta = None
        if period == "monthly":
            prev_first = (rd.replace(day=1) - timedelta(days=1)).replace(day=1)
            prev_last = rd.replace(day=1) - timedelta(days=1)
            prev_stats = dict(c.execute(sql, (prev_first.isoformat(), prev_last.isoformat()) + scope_args).fetchone())
            for k in prev_stats:
                prev_stats[k] = prev_stats[k] or 0
            def pct(cur, prev):
                if not prev:
                    return None
                return round((cur - prev) * 100 / prev, 1)
            delta = {
                "total": pct(stats["total"], prev_stats["total"]),
                "done": pct(stats["done"], prev_stats["done"]),
                "delay": pct(stats["delay"], prev_stats["delay"]),
                "hours": pct(stats["hours"], prev_stats["hours"]),
            }

        # 일별 추이 (월간일 때)
        daily_trend = []
        if period == "monthly":
            daily_trend = [dict(r) for r in c.execute(
                f"""SELECT t.work_date AS d, COUNT(*) AS cnt,
                           SUM(CASE WHEN t.status='완료' THEN 1 ELSE 0 END) AS done,
                           COALESCE(SUM(t.hours),0) AS hours
                    FROM tasks t JOIN users uu ON t.user_id = uu.id
                    WHERE t.work_date BETWEEN ? AND ? {scope_filter}
                    GROUP BY t.work_date ORDER BY t.work_date""",
                (frm, to) + scope_args,
            ).fetchall()]

    return ctx(req, "summary.html", user=u,
               period=period, scope=scope, ref=ref,
               period_label=period_label, scope_label=scope_label,
               frm=frm, to=to, prev_ref=prev_ref, next_ref=next_ref,
               stats=stats, completion_rate=completion_rate,
               cards=cards[:80], total_cards=len(cards),
               team_agg=team_agg, user_agg=user_agg, project_agg=project_agg,
               prev_stats=prev_stats, delta=delta, daily_trend=daily_trend,
               can_all=u["role"] in ("ceo","admin","executive"),
               can_team=u.get("team_id") is not None,
               active="summary")


# =====================================================
# COMMENTS — 업무카드 댓글/요청사항
# =====================================================
@app.get("/api/task/{tid}/comments")
async def api_list_comments(req: Request, tid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    return JSONResponse({"ok": True, "comments": get_task_comments(tid)})


@app.post("/api/task/{tid}/comment")
async def api_add_comment(req: Request, tid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    d = await req.json()
    body = (d.get("body") or "").strip()
    if not body:
        return JSONResponse({"error": "내용을 입력하세요"}, 400)
    parent_id = d.get("parent_id") or None
    cid = add_comment(tid, u["id"], body, parent_id)
    return JSONResponse({"ok": True, "id": cid})


@app.delete("/api/comment/{cid}")
async def api_delete_comment(req: Request, cid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    ok = delete_comment(cid, u["id"])
    return JSONResponse({"ok": ok})


# =====================================================
# TASK DETAIL — 모달용 단일 카드 정보 (어디서든 호출)
# =====================================================
@app.get("/api/task/{tid}/detail")
async def api_task_detail(req: Request, tid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error":"로그인 필요"}, 401)
    with db_session() as c:
        t = c.execute(
            """SELECT t.*, u.name AS user_name, u.rank AS user_rank,
                      p.name AS project_name, cs.name AS customer_name,
                      tm.name AS team_name
               FROM tasks t LEFT JOIN users u ON t.user_id=u.id
               LEFT JOIN projects p ON t.project_id=p.id
               LEFT JOIN customers cs ON t.customer_id=cs.id
               LEFT JOIN teams tm ON u.team_id=tm.id
               WHERE t.id=?""", (tid,)
        ).fetchone()
        if not t:
            return JSONResponse({"error":"카드 없음"}, 404)
    comments = get_task_comments(tid)
    reactions = get_reactions(tid)
    delegations = get_delegations(tid)
    return JSONResponse({"ok":True, "task":dict(t), "comments":comments,
                         "reactions":reactions, "delegations":delegations})


# =====================================================
# REACTIONS — 1-click 빠른 피드백
# =====================================================
@app.post("/api/task/{tid}/reaction")
async def api_reaction(req: Request, tid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error":"로그인 필요"}, 401)
    d = await req.json()
    res = add_reaction(tid, u["id"], d.get("kind") or "")
    if not res:
        return JSONResponse({"error":"잘못된 반응"}, 400)
    return JSONResponse({"ok":True, "result":res, "reactions":get_reactions(tid)})


# =====================================================
# 번역 API
# =====================================================
@app.post("/api/set-lang")
async def api_set_lang(req: Request):
    """사용자 UI 언어 변경"""
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    data = await req.json()
    lang = data.get("lang", "ko")
    if lang not in LANGS:
        lang = "ko"
    with db_session() as c:
        c.execute("UPDATE users SET lang=? WHERE id=?", (lang, u["id"]))
    req.session["lang"] = lang
    return JSONResponse({"ok": True, "lang": lang})


@app.post("/api/translate")
async def api_translate(req: Request):
    u = get_user(req)
    if not u:
        return JSONResponse({"error":"로그인 필요"}, 401)
    d = await req.json()
    text = (d.get("text") or "").strip()
    target = d.get("target") or "vi"  # 기본: 베트남어
    if not text:
        return JSONResponse({"error":"텍스트 없음"}, 400)
    # 방법1: deep_translator
    try:
        from deep_translator import GoogleTranslator
        result = GoogleTranslator(source='auto', target=target).translate(text)
        if result:
            return JSONResponse({"ok": True, "translated": result, "target": target})
    except Exception as e1:
        print(f"[번역] GoogleTranslator 실패: {e1}")
    # 방법2: MyMemoryTranslator (fallback)
    try:
        from deep_translator import MyMemoryTranslator
        src = 'ko-KR'
        tgt_map = {'vi':'vi-VN','en':'en-GB','ko':'ko-KR','ja':'ja-JP','zh-CN':'zh-CN'}
        result = MyMemoryTranslator(source=src, target=tgt_map.get(target,'en-GB')).translate(text)
        if result:
            return JSONResponse({"ok": True, "translated": result, "target": target})
    except Exception as e2:
        print(f"[번역] MyMemoryTranslator 실패: {e2}")
    # 방법3: urllib로 직접 Google Translate 호출
    try:
        import urllib.request, urllib.parse, json as _json
        encoded = urllib.parse.quote(text)
        url = f"https://translate.googleapis.com/translate_a/single?client=gtx&sl=auto&tl={target}&dt=t&q={encoded}"
        req2 = urllib.request.Request(url, headers={"User-Agent":"Mozilla/5.0"})
        with urllib.request.urlopen(req2, timeout=10) as resp:
            data = _json.loads(resp.read().decode())
            translated = "".join(seg[0] for seg in data[0] if seg[0])
            return JSONResponse({"ok": True, "translated": translated, "target": target})
    except Exception as e3:
        print(f"[번역] urllib 직접호출 실패: {e3}")
    return JSONResponse({"ok": False, "error": "번역 서비스에 연결할 수 없습니다. 인터넷 연결을 확인해주세요."}, 500)


# =====================================================
# 업무 위임 (Delegation)
# =====================================================
@app.post("/api/task/{tid}/delegate")
async def api_delegate(req: Request, tid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error":"로그인 필요"}, 401)
    d = await req.json()
    to_id = d.get("to_user_id")
    msg = (d.get("message") or "").strip()
    if not to_id:
        return JSONResponse({"error":"위임 대상을 선택하세요"}, 400)
    delegate_task(tid, u["id"], int(to_id), msg)
    return JSONResponse({"ok": True})


@app.post("/api/delegation/{did}/resolve")
async def api_resolve_delegation(req: Request, did: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error":"로그인 필요"}, 401)
    ok = resolve_delegation(did, u["id"])
    if not ok:
        return JSONResponse({"error":"본인만 완료 처리 가능"}, 403)
    return JSONResponse({"ok": True})


# =====================================================
# @멘션 자동완성
# =====================================================
@app.get("/api/users/search")
async def api_user_search(req: Request, q: str = ""):
    u = get_user(req)
    if not u:
        return JSONResponse({"error":"로그인 필요"}, 401)
    return JSONResponse({"ok":True, "users":get_user_search(q.strip(), 8)})


# =====================================================
# SIDEBAR TREE — Notion 스타일 좌측 트리 데이터
# =====================================================
@app.get("/api/sidebar-tree")
async def api_sidebar_tree(req: Request):
    """좌측 사이드바에 표시할 트리 구조:
       팀 목록 > 하위 프로젝트 > (선택시 해당 페이지 이동)
       + 오늘 카드 수, 지연 카운트"""
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    today_s = date.today().isoformat()
    with db_session() as c:
        teams = [dict(r) for r in c.execute(
            """SELECT t.id, t.name, t.code, t.is_lab, t.display_order,
                      u.name AS leader_name
               FROM teams t LEFT JOIN users u ON t.leader_id=u.id
               ORDER BY t.display_order"""
        ).fetchall()]
        for t in teams:
            stats = c.execute(
                """SELECT COUNT(*) AS total,
                          SUM(CASE WHEN tk.status='지연' THEN 1 ELSE 0 END) AS delay
                   FROM tasks tk JOIN users u ON tk.user_id=u.id
                   WHERE u.team_id=? AND tk.work_date=?""",
                (t["id"], today_s)
            ).fetchone()
            t["today_count"] = stats["total"] or 0
            t["delay_count"] = stats["delay"] or 0
            # 팀 하위 활성 프로젝트
            projects = [dict(r) for r in c.execute(
                """SELECT DISTINCT p.id, p.name, p.status
                   FROM projects p
                   JOIN tasks tk ON tk.project_id=p.id
                   JOIN users u ON tk.user_id=u.id
                   WHERE u.team_id=? AND p.status IN ('active','진행중','planning')
                   ORDER BY p.name""",
                (t["id"],)
            ).fetchall()]
            t["projects"] = projects
        # 즐겨찾기 / 내 카드 요약
        my_today = c.execute(
            "SELECT COUNT(*) FROM tasks WHERE user_id=? AND work_date=?",
            (u["id"], today_s)
        ).fetchone()[0]
    return JSONResponse({"ok": True, "teams": teams, "my_today": my_today,
                         "user_role": u["role"], "user_team_id": u.get("team_id")})


# =====================================================
# ACTIVITIES — 실시간 활동 피드
# =====================================================
@app.get("/api/activities")
async def api_activities(req: Request, scope: str = "all", since_id: int = 0):
    u = get_user(req)
    if not u:
        return JSONResponse({"error":"로그인 필요"}, 401)
    team_id = None
    actor_id = None
    if scope == "team" and u.get("team_id"):
        team_id = u["team_id"]
    elif scope == "me":
        actor_id = u["id"]
    items = get_activities(limit=80, team_id=team_id, actor_id=actor_id, since_id=since_id)
    return JSONResponse({"ok":True, "items":items})


@app.get("/now", response_class=HTMLResponse)
async def now_feed(req: Request, scope: str = "all"):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    team_id = None; actor_id = None
    if scope == "team" and u.get("team_id"):
        team_id = u["team_id"]
    elif scope == "me":
        actor_id = u["id"]
    items = get_activities(limit=80, team_id=team_id, actor_id=actor_id)
    return ctx(req, "now.html", user=u, items=items, scope=scope,
               can_team=u.get("team_id") is not None, active="now")


# =====================================================
# 통합 검색
# =====================================================
@app.get("/search", response_class=HTMLResponse)
async def search_page(req: Request, q: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    res = search_all(q, 50) if q else {"tasks":[], "comments":[], "retros":[]}
    return ctx(req, "search.html", user=u, q=q, res=res, active="search")


# =====================================================
# 프로젝트 회고 (Retro)
# =====================================================
@app.post("/api/project/{pid}/retro")
async def api_retro_save(req: Request, pid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error":"로그인 필요"}, 401)
    d = await req.json()
    rid = upsert_retro(pid, u["id"],
                       (d.get("went_well") or "").strip(),
                       (d.get("went_bad") or "").strip(),
                       (d.get("next_action") or "").strip(),
                       (d.get("risk_note") or "").strip())
    log_activity_standalone(u["id"], "retro",
                            title=f"{u['name']} 프로젝트 회고 작성",
                            project_id=pid, team_id=u.get("team_id"))
    return JSONResponse({"ok":True, "id":rid})


# =====================================================
# 코크핏 (팀장/CEO 라이브 시야)
# =====================================================
@app.get("/cockpit", response_class=HTMLResponse)
async def cockpit_page(req: Request):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if u["role"] not in ("leader","executive","ceo","admin"):
        return RedirectResponse("/dashboard", 303)
    today_iso = date.today().isoformat()
    week_ago = (date.today() - timedelta(days=7)).isoformat()
    is_global = u["role"] in ("ceo","admin","executive")
    team_filter = ""
    params = []
    if not is_global and u.get("team_id"):
        team_filter = "AND u.team_id = ?"
        params = [u["team_id"]]

    with db_session() as c:
        # 우리팀(또는 전사) 멤버
        members = c.execute(
            f"""SELECT u.id, u.name, u.rank, tm.name AS team_name,
                       (SELECT COUNT(*) FROM tasks WHERE user_id=u.id AND work_date=?) AS today_cn,
                       (SELECT COUNT(*) FROM tasks WHERE user_id=u.id AND status='지연') AS delay_cn,
                       (SELECT COUNT(*) FROM tasks WHERE user_id=u.id AND status='진행중') AS prog_cn,
                       (SELECT COUNT(*) FROM tasks t JOIN task_comments tc ON tc.task_id=t.id
                        WHERE t.user_id=u.id AND tc.is_ceo_request=1
                        AND NOT EXISTS (SELECT 1 FROM task_comments tc2
                                        WHERE tc2.task_id=tc.task_id AND tc2.id>tc.id)) AS unanswered_ceo
                FROM users u LEFT JOIN teams tm ON u.team_id=tm.id
                WHERE u.is_active=1 AND u.role NOT IN ('admin') {team_filter}
                ORDER BY tm.display_order, u.id""",
            tuple([today_iso] + params)
        ).fetchall()

        # 막힌 카드 (지연 + 코멘트 미해결)
        stuck = c.execute(
            f"""SELECT t.id, t.title, t.work_date, u.name AS owner_name, tm.name AS team_name,
                       (SELECT COUNT(*) FROM task_comments WHERE task_id=t.id) AS cn
                FROM tasks t JOIN users u ON t.user_id=u.id
                LEFT JOIN teams tm ON u.team_id=tm.id
                WHERE t.status='지연' {team_filter}
                ORDER BY t.work_date LIMIT 30""",
            tuple(params)
        ).fetchall()

        # 미작성자 (오늘)
        missing = c.execute(
            f"""SELECT u.id, u.name, u.rank, tm.name AS team_name
                FROM users u LEFT JOIN teams tm ON u.team_id=tm.id
                WHERE u.is_active=1 AND u.role NOT IN ('admin','ceo')
                AND NOT EXISTS (SELECT 1 FROM tasks WHERE user_id=u.id AND work_date=?)
                {team_filter} ORDER BY tm.display_order, u.name""",
            tuple([today_iso] + params)
        ).fetchall()

    bn = detect_bottlenecks() if is_global else []
    return ctx(req, "cockpit.html", user=u,
               members=[dict(r) for r in members],
               stuck=[dict(r) for r in stuck],
               missing=[dict(r) for r in missing],
               bottlenecks=bn,
               is_global=is_global,
               active="cockpit")


# =====================================================
# 병목 자동 감지 페이지
# =====================================================
@app.get("/bottlenecks", response_class=HTMLResponse)
async def bottlenecks_page(req: Request):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if u["role"] not in ("leader","executive","ceo","admin"):
        return RedirectResponse("/dashboard", 303)
    items = detect_bottlenecks()
    return ctx(req, "bottlenecks.html", user=u, items=items, active="bottlenecks")


# =====================================================
# NOTIFICATIONS — 알림 (벨 아이콘 + 드롭다운)
# =====================================================
@app.get("/api/notifications")
async def api_notifications(req: Request, only_unread: int = 0):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    items = get_notifications(u["id"], only_unread=bool(only_unread), limit=30)
    return JSONResponse({"ok": True, "items": items, "unread": count_unread(u["id"])})


@app.post("/api/notification/{nid}/read")
async def api_notif_read(req: Request, nid: int):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    mark_notification_read(nid, u["id"])
    return JSONResponse({"ok": True})


@app.post("/api/notifications/read-all")
async def api_notif_read_all(req: Request):
    u = get_user(req)
    if not u:
        return JSONResponse({"error": "로그인 필요"}, 401)
    mark_all_read(u["id"])
    return JSONResponse({"ok": True})


@app.get("/notifications", response_class=HTMLResponse)
async def notifications_page(req: Request):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    items = get_notifications(u["id"], limit=100)
    return ctx(req, "notifications.html", user=u, items=items, active="notifications")


# =====================================================
# HISTORY — 개인 히스토리 (내가 한 일 검색/조회)
# =====================================================
@app.get("/history", response_class=HTMLResponse)
async def history_page(req: Request, q: str = "", frm: str = "", to: str = "", status: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if not frm:
        frm = (date.today() - timedelta(days=30)).isoformat()
    if not to:
        to = date.today().isoformat()
    # 유효한 status만 필터 허용
    valid_statuses = {"완료", "진행중", "지연", "대기", "보류"}
    if status and status not in valid_statuses:
        status = ""
    with db_session() as c:
        sql = """SELECT t.*, p.name AS project_name, c.name AS customer_name
                 FROM tasks t
                 LEFT JOIN projects p ON t.project_id=p.id
                 LEFT JOIN customers c ON t.customer_id=c.id
                 WHERE t.user_id=? AND t.work_date>=? AND t.work_date<=?"""
        params = [u["id"], frm, to]
        if q:
            sql += " AND (t.title LIKE ? OR t.notes LIKE ?)"
            params += [f"%{q}%", f"%{q}%"]
        if status:
            sql += " AND t.status=?"
            params.append(status)
        sql += " ORDER BY t.work_date DESC, t.id DESC"
        tasks = [dict(r) for r in c.execute(sql, params).fetchall()]

        summary = c.execute(
            """SELECT COUNT(*) AS total,
                      SUM(CASE WHEN status='완료' THEN 1 ELSE 0 END) AS done,
                      COALESCE(SUM(hours),0) AS hours
               FROM tasks WHERE user_id=? AND work_date>=? AND work_date<=?""",
            (u["id"], frm, to),
        ).fetchone()
        summary = dict(summary)

        by_category = [dict(r) for r in c.execute(
            """SELECT category, COUNT(*) AS cnt, COALESCE(SUM(hours),0) AS hrs
               FROM tasks WHERE user_id=? AND work_date>=? AND work_date<=?
               GROUP BY category ORDER BY cnt DESC""",
            (u["id"], frm, to),
        ).fetchall()]

    return ctx(req, "history.html",
               user=u, tasks=tasks, q=q, frm=frm, to=to, status=status,
               summary=summary, by_category=by_category)


# =====================================================
# TEAM — 팀장 뷰
# =====================================================
@app.get("/team", response_class=HTMLResponse)
@app.get("/team/{sel_date}", response_class=HTMLResponse)
async def team_page(req: Request, sel_date: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if u["role"] not in ("leader", "executive", "ceo", "admin"):
        return RedirectResponse("/daily", 303)
    if not sel_date:
        sel_date = date.today().isoformat()

    # CEO/admin은 쿼리로 team 선택 가능, 기본 전체
    with db_session() as c:
        teams = [dict(r) for r in c.execute(
            "SELECT * FROM teams ORDER BY display_order"
        ).fetchall()]

        target_team_id = req.query_params.get("team_id")
        if u["role"] in ("leader", "executive"):
            target_team_id = u["team_id"]
        elif target_team_id:
            target_team_id = int(target_team_id)
        else:
            target_team_id = teams[0]["id"] if teams else None

        members = []
        if target_team_id:
            members = [dict(r) for r in c.execute(
                """SELECT id, name, rank, role FROM users
                   WHERE team_id=? AND is_active=1
                   ORDER BY CASE role
                        WHEN 'ceo' THEN 0 WHEN 'executive' THEN 1
                        WHEN 'leader' THEN 2 ELSE 3 END, id""",
                (target_team_id,),
            ).fetchall()]

        mids = [m["id"] for m in members]
        day_tasks = []
        week_tasks = []
        if mids:
            ph = ",".join("?" * len(mids))
            day_tasks = [dict(r) for r in c.execute(
                f"""SELECT t.*, u.name AS user_name, u.rank AS user_rank,
                           p.name AS project_name, cu.name AS customer_name
                    FROM tasks t JOIN users u ON t.user_id=u.id
                    LEFT JOIN projects p ON t.project_id=p.id
                    LEFT JOIN customers cu ON t.customer_id=cu.id
                    WHERE t.user_id IN ({ph}) AND t.work_date=?
                    ORDER BY u.id, t.status, t.id""",
                mids + [sel_date],
            ).fetchall()]

            # 이번 주
            d0 = datetime.strptime(sel_date, "%Y-%m-%d")
            mon = (d0 - timedelta(days=d0.weekday())).strftime("%Y-%m-%d")
            sun = (d0 - timedelta(days=d0.weekday()) + timedelta(days=6)).strftime("%Y-%m-%d")
            week_tasks = [dict(r) for r in c.execute(
                f"""SELECT t.*, u.name AS user_name FROM tasks t JOIN users u ON t.user_id=u.id
                    WHERE t.user_id IN ({ph}) AND t.work_date>=? AND t.work_date<=?""",
                mids + [mon, sun],
            ).fetchall()]

            # 카드 표면 메타(댓글/리액션 카운트) 일괄 조회
            meta = get_meta_bulk([t["id"] for t in day_tasks])
            for t in day_tasks:
                m = meta.get(t["id"], {})
                t["meta_comments"] = m.get("comments", 0)
                t["meta_ack"] = m.get("ack", 0)
                t["meta_question"] = m.get("question", 0)
                t["meta_risk"] = m.get("risk", 0)
                t["meta_ok"] = m.get("ok", 0)
                t["meta_last_comment"] = m.get("last_comment", "")
                t["meta_has_activity"] = bool(
                    t["meta_comments"] or t["meta_ack"] or t["meta_question"]
                    or t["meta_risk"] or t["meta_ok"]
                )

        # 팀 전체 통계 (오늘)
        stats_by_user = {}
        for m in members:
            ut = [t for t in day_tasks if t["user_id"] == m["id"]]
            stats_by_user[m["id"]] = {
                "total": len(ut),
                "done": len([t for t in ut if t["status"] == "완료"]),
                "progress": len([t for t in ut if t["status"] == "진행중"]),
                "delay": len([t for t in ut if t["status"] == "지연"]),
                "hours": sum(t["hours"] or 0 for t in ut),
                "reported": 1 if ut else 0,
            }

        team_summary = {
            "members": len(members),
            "reported": sum(1 for s in stats_by_user.values() if s["reported"]),
            "total": len(day_tasks),
            "done": len([t for t in day_tasks if t["status"] == "완료"]),
            "delay": len([t for t in day_tasks if t["status"] == "지연"]),
            "progress": len([t for t in day_tasks if t["status"] == "진행중"]),
            "week_total": len(week_tasks),
            "week_done": len([t for t in week_tasks if t["status"] == "완료"]),
        }

    prev_d = (datetime.strptime(sel_date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
    next_d = (datetime.strptime(sel_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")

    return ctx(req, "team.html",
               user=u, teams=teams, target_team_id=target_team_id,
               members=members, day_tasks=day_tasks, stats_by_user=stats_by_user,
               team_summary=team_summary, sel_date=sel_date,
               prev_date=prev_d, next_date=next_d)


# =====================================================
# DASHBOARD — 대표이사 전사 뷰
# =====================================================
@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard_page(req: Request):
    u = require(req, ["ceo", "admin", "executive"])
    if not u:
        return RedirectResponse("/login", 303)
    today = date.today()
    mon = (today - timedelta(days=today.weekday())).isoformat()
    sun = (today - timedelta(days=today.weekday()) + timedelta(days=6)).isoformat()
    today_s = today.isoformat()

    with db_session() as c:
        teams = [dict(r) for r in c.execute(
            """SELECT t.*, u.name AS leader_name, u.rank AS leader_rank
               FROM teams t LEFT JOIN users u ON t.leader_id=u.id
               ORDER BY t.display_order"""
        ).fetchall()]

        for t in teams:
            mc = c.execute(
                "SELECT COUNT(*) FROM users WHERE team_id=? AND is_active=1",
                (t["id"],),
            ).fetchone()[0]
            t["member_count"] = mc
            s = c.execute(
                """SELECT COUNT(*) AS total,
                          SUM(CASE WHEN status='완료' THEN 1 ELSE 0 END) AS done,
                          SUM(CASE WHEN status='진행중' THEN 1 ELSE 0 END) AS progress,
                          SUM(CASE WHEN status='지연' THEN 1 ELSE 0 END) AS delay,
                          COALESCE(SUM(hours),0) AS hours
                   FROM tasks tk JOIN users u ON tk.user_id=u.id
                   WHERE u.team_id=? AND tk.work_date>=? AND tk.work_date<=?""",
                (t["id"], mon, sun),
            ).fetchone()
            t["week_stats"] = {k: (s[k] or 0) for k in s.keys()}
            td = c.execute(
                """SELECT COUNT(*) AS total FROM tasks tk JOIN users u ON tk.user_id=u.id
                   WHERE u.team_id=? AND tk.work_date=?""",
                (t["id"], today_s),
            ).fetchone()
            t["today_count"] = td["total"] or 0
            # 참여율: 오늘 카드 작성자 수 / 팀원 수
            rp = c.execute(
                """SELECT COUNT(DISTINCT tk.user_id) FROM tasks tk JOIN users u ON tk.user_id=u.id
                   WHERE u.team_id=? AND tk.work_date=?""",
                (t["id"], today_s),
            ).fetchone()[0]
            t["reported"] = rp
            t["participation"] = round(rp * 100 / mc) if mc else 0

        total_stats = c.execute(
            """SELECT COUNT(*) AS total,
                      SUM(CASE WHEN status='완료' THEN 1 ELSE 0 END) AS done,
                      SUM(CASE WHEN status='지연' THEN 1 ELSE 0 END) AS delay,
                      SUM(CASE WHEN status='진행중' THEN 1 ELSE 0 END) AS progress,
                      COALESCE(SUM(hours),0) AS hours
               FROM tasks WHERE work_date>=? AND work_date<=?""",
            (mon, sun),
        ).fetchone()
        total_stats = {k: (total_stats[k] or 0) for k in total_stats.keys()}

        total_users = c.execute(
            "SELECT COUNT(*) FROM users WHERE is_active=1 AND role!='admin'"
        ).fetchone()[0]
        today_reporters = c.execute(
            "SELECT COUNT(DISTINCT user_id) FROM tasks WHERE work_date=?",
            (today_s,),
        ).fetchone()[0]
        participation_rate = round(today_reporters * 100 / total_users) if total_users else 0

        delays = [dict(r) for r in c.execute(
            """SELECT t.*, u.name AS user_name, tm.name AS team_name,
                      p.name AS project_name, cu.name AS customer_name
               FROM tasks t JOIN users u ON t.user_id=u.id
               LEFT JOIN teams tm ON u.team_id=tm.id
               LEFT JOIN projects p ON t.project_id=p.id
               LEFT JOIN customers cu ON t.customer_id=cu.id
               WHERE t.status='지연' AND t.work_date>=? AND t.work_date<=?
               ORDER BY t.work_date DESC LIMIT 10""",
            (mon, sun),
        ).fetchall()]

        customers = [dict(r) for r in c.execute(
            """SELECT cu.name AS customer_name, COUNT(*) AS cnt,
                      COALESCE(SUM(t.hours),0) AS hours
               FROM tasks t JOIN customers cu ON t.customer_id=cu.id
               WHERE t.work_date>=? AND t.work_date<=?
               GROUP BY cu.name ORDER BY cnt DESC LIMIT 10""",
            (mon, sun),
        ).fetchall()]

        # 내러티브: 팀별 진행중 핵심 카드 3건씩 + 다음 계획
        narratives = []
        for t in teams:
            cards = [dict(r) for r in c.execute(
                """SELECT t.title, t.status, t.next_plan, u.name AS user_name, u.rank,
                          p.name AS project_name
                   FROM tasks t JOIN users u ON t.user_id=u.id
                   LEFT JOIN projects p ON t.project_id=p.id
                   WHERE u.team_id=? AND t.work_date=?
                     AND t.status IN ('진행중','지연')
                   ORDER BY CASE t.status WHEN '지연' THEN 0 ELSE 1 END, t.id
                   LIMIT 3""",
                (t["id"], today_s),
            ).fetchall()]
            if cards:
                narratives.append({"team": t, "cards": cards})

    return ctx(req, "dashboard.html",
               user=u, teams=teams, total_stats=total_stats,
               mon=mon, sun=sun, today_s=today_s,
               participation_rate=participation_rate,
               today_reporters=today_reporters, total_users=total_users,
               delays=delays, customers=customers, narratives=narratives)


# =====================================================
# TEAM DAILY SUMMARY — 팀장 "오늘의 한 줄"
# =====================================================
@app.post("/api/team-summary")
async def api_team_summary(req: Request):
    u = require(req, ["leader", "executive", "ceo", "admin"])
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    d = await req.json()
    tid = d.get("team_id") or u.get("team_id")
    wdate = d.get("work_date") or date.today().isoformat()
    headline = (d.get("headline") or "").strip()
    notes = d.get("notes") or ""
    if not headline or not tid:
        return JSONResponse({"error": "내용/팀 필수"}, 400)
    with db_session() as c:
        ex = c.execute(
            "SELECT id FROM team_summaries WHERE team_id=? AND work_date=?",
            (tid, wdate),
        ).fetchone()
        if ex:
            c.execute(
                """UPDATE team_summaries SET headline=?, notes=?, author_id=?,
                       updated_at=datetime('now','localtime') WHERE id=?""",
                (headline, notes, u["id"], ex["id"]),
            )
        else:
            c.execute(
                """INSERT INTO team_summaries(team_id, work_date, author_id, headline, notes)
                   VALUES(?,?,?,?,?)""",
                (tid, wdate, u["id"], headline, notes),
            )
    return JSONResponse({"ok": True})


# =====================================================
# WEEKLY — 주간 자동 요약
# =====================================================
@app.get("/weekly", response_class=HTMLResponse)
@app.get("/weekly/{wk_mon}", response_class=HTMLResponse)
async def weekly_page(req: Request, wk_mon: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if not wk_mon:
        td = date.today()
        wk_mon = (td - timedelta(days=td.weekday())).isoformat()
    mon = datetime.strptime(wk_mon, "%Y-%m-%d").date()
    sun = mon + timedelta(days=6)
    prev_mon = (mon - timedelta(days=7)).isoformat()
    next_mon = (mon + timedelta(days=7)).isoformat()

    with db_session() as c:
        # 개인 요약
        my_tasks = [dict(r) for r in c.execute(
            """SELECT t.*, p.name AS project_name, cu.name AS customer_name
               FROM tasks t LEFT JOIN projects p ON t.project_id=p.id
               LEFT JOIN customers cu ON t.customer_id=cu.id
               WHERE t.user_id=? AND t.work_date>=? AND t.work_date<=?
               ORDER BY t.work_date, t.id""",
            (u["id"], mon.isoformat(), sun.isoformat()),
        ).fetchall()]
        my_stats = {
            "total": len(my_tasks),
            "done": sum(1 for t in my_tasks if t["status"] == "완료"),
            "progress": sum(1 for t in my_tasks if t["status"] == "진행중"),
            "delay": sum(1 for t in my_tasks if t["status"] == "지연"),
            "hours": round(sum(t["hours"] or 0 for t in my_tasks), 1),
        }
        # 분류별
        my_by_cat = {}
        for t in my_tasks:
            k = t["category"] or "기타"
            my_by_cat.setdefault(k, {"cnt": 0, "hours": 0})
            my_by_cat[k]["cnt"] += 1
            my_by_cat[k]["hours"] += t["hours"] or 0

        # 팀 요약 (팀장/임원/CEO/admin만)
        team_data = None
        if u["role"] in ("leader", "executive", "ceo", "admin"):
            tid = req.query_params.get("team_id")
            if u["role"] in ("leader", "executive"):
                tid = u["team_id"]
            elif tid:
                tid = int(tid)
            if tid:
                t_row = c.execute("SELECT * FROM teams WHERE id=?", (tid,)).fetchone()
                members = [dict(r) for r in c.execute(
                    "SELECT id, name, rank FROM users WHERE team_id=? AND is_active=1", (tid,)
                ).fetchall()]
                mids = [m["id"] for m in members]
                t_tasks = []
                if mids:
                    ph = ",".join("?" * len(mids))
                    t_tasks = [dict(r) for r in c.execute(
                        f"""SELECT tk.*, u.name AS user_name, p.name AS project_name,
                                  cu.name AS customer_name
                           FROM tasks tk JOIN users u ON tk.user_id=u.id
                           LEFT JOIN projects p ON tk.project_id=p.id
                           LEFT JOIN customers cu ON tk.customer_id=cu.id
                           WHERE tk.user_id IN ({ph}) AND tk.work_date>=? AND tk.work_date<=?
                           ORDER BY u.id, tk.work_date""",
                        mids + [mon.isoformat(), sun.isoformat()],
                    ).fetchall()]
                # 팀원별 집계
                per_user = {}
                for m in members:
                    per_user[m["id"]] = {"name": m["name"], "rank": m["rank"],
                                          "total": 0, "done": 0, "hours": 0, "tasks": []}
                for t in t_tasks:
                    per_user.setdefault(t["user_id"], {"name": t["user_name"], "rank": "",
                                                        "total": 0, "done": 0, "hours": 0, "tasks": []})
                    per_user[t["user_id"]]["total"] += 1
                    if t["status"] == "완료":
                        per_user[t["user_id"]]["done"] += 1
                    per_user[t["user_id"]]["hours"] += t["hours"] or 0
                    per_user[t["user_id"]]["tasks"].append(t)
                # 프로젝트별
                pj_agg = {}
                for t in t_tasks:
                    k = t["project_name"] or "(기타)"
                    pj_agg.setdefault(k, {"cnt": 0, "hours": 0, "done": 0})
                    pj_agg[k]["cnt"] += 1
                    pj_agg[k]["hours"] += t["hours"] or 0
                    if t["status"] == "완료":
                        pj_agg[k]["done"] += 1
                # 고객사별
                cu_agg = {}
                for t in t_tasks:
                    k = t["customer_name"] or "(기타)"
                    cu_agg.setdefault(k, {"cnt": 0, "hours": 0})
                    cu_agg[k]["cnt"] += 1
                    cu_agg[k]["hours"] += t["hours"] or 0
                team_data = {
                    "team": dict(t_row),
                    "members": members,
                    "per_user": per_user,
                    "pj_agg": sorted(pj_agg.items(), key=lambda x: -x[1]["cnt"])[:15],
                    "cu_agg": sorted(cu_agg.items(), key=lambda x: -x[1]["cnt"])[:15],
                    "total": len(t_tasks),
                    "done": sum(1 for t in t_tasks if t["status"] == "완료"),
                    "delay": sum(1 for t in t_tasks if t["status"] == "지연"),
                    "hours": round(sum(t["hours"] or 0 for t in t_tasks), 1),
                }

        # 전사 요약 (CEO/admin/executive)
        all_data = None
        if u["role"] in ("ceo", "admin", "executive"):
            all_tasks = [dict(r) for r in c.execute(
                """SELECT tk.*, u.name AS user_name, tm.name AS team_name, tm.code AS team_code,
                          p.name AS project_name, cu.name AS customer_name
                   FROM tasks tk JOIN users u ON tk.user_id=u.id
                   LEFT JOIN teams tm ON u.team_id=tm.id
                   LEFT JOIN projects p ON tk.project_id=p.id
                   LEFT JOIN customers cu ON tk.customer_id=cu.id
                   WHERE tk.work_date>=? AND tk.work_date<=?""",
                (mon.isoformat(), sun.isoformat()),
            ).fetchall()]
            by_team = {}
            for t in all_tasks:
                k = t["team_name"] or "(미배정)"
                by_team.setdefault(k, {"code": t["team_code"], "cnt": 0, "done": 0, "hours": 0})
                by_team[k]["cnt"] += 1
                if t["status"] == "완료":
                    by_team[k]["done"] += 1
                by_team[k]["hours"] += t["hours"] or 0
            by_cust = {}
            for t in all_tasks:
                if not t["customer_name"]:
                    continue
                by_cust.setdefault(t["customer_name"], {"cnt": 0, "hours": 0})
                by_cust[t["customer_name"]]["cnt"] += 1
                by_cust[t["customer_name"]]["hours"] += t["hours"] or 0
            all_data = {
                "total": len(all_tasks),
                "done": sum(1 for t in all_tasks if t["status"] == "완료"),
                "delay": sum(1 for t in all_tasks if t["status"] == "지연"),
                "hours": round(sum(t["hours"] or 0 for t in all_tasks), 1),
                "by_team": sorted(by_team.items(), key=lambda x: -x[1]["cnt"]),
                "by_cust": sorted(by_cust.items(), key=lambda x: -x[1]["cnt"])[:10],
            }

        teams_all = [dict(r) for r in c.execute(
            "SELECT * FROM teams ORDER BY display_order"
        ).fetchall()]

    return ctx(req, "weekly.html",
               user=u, my_tasks=my_tasks, my_stats=my_stats, my_by_cat=my_by_cat,
               team_data=team_data, all_data=all_data,
               wk_mon=mon.isoformat(), wk_sun=sun.isoformat(),
               prev_mon=prev_mon, next_mon=next_mon, teams_all=teams_all,
               active="weekly")


# =====================================================
# PROJECT DETAIL
# =====================================================
@app.get("/project/{pid}", response_class=HTMLResponse)
async def project_detail(req: Request, pid: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        p = c.execute(
            """SELECT p.*, cu.name AS customer_name FROM projects p
               LEFT JOIN customers cu ON p.customer_id=cu.id WHERE p.id=?""",
            (pid,),
        ).fetchone()
        if not p:
            return RedirectResponse("/", 303)
        p = dict(p)
        # 참여 전체 카드
        tasks = [dict(r) for r in c.execute(
            """SELECT tk.*, u.name AS user_name, u.rank AS user_rank,
                      tm.name AS team_name, tm.code AS team_code
               FROM tasks tk JOIN users u ON tk.user_id=u.id
               LEFT JOIN teams tm ON u.team_id=tm.id
               WHERE tk.project_id=? ORDER BY tk.work_date DESC, tk.id DESC""",
            (pid,),
        ).fetchall()]
        stats = {
            "total": len(tasks),
            "done": sum(1 for t in tasks if t["status"] == "완료"),
            "progress": sum(1 for t in tasks if t["status"] == "진행중"),
            "delay": sum(1 for t in tasks if t["status"] == "지연"),
            "hours": round(sum(t["hours"] or 0 for t in tasks), 1),
        }
        # 팀별 집계
        by_team = {}
        for t in tasks:
            k = t["team_name"] or "(미배정)"
            by_team.setdefault(k, {"cnt": 0, "hours": 0, "members": set()})
            by_team[k]["cnt"] += 1
            by_team[k]["hours"] += t["hours"] or 0
            by_team[k]["members"].add(t["user_name"])
        by_team_list = [(k, {"cnt": v["cnt"], "hours": v["hours"], "members": sorted(v["members"])})
                        for k, v in sorted(by_team.items(), key=lambda x: -x[1]["cnt"])]
        # 참여자 개별
        by_user = {}
        for t in tasks:
            k = f"{t['user_name']} ({t['team_name'] or '-'})"
            by_user.setdefault(k, {"cnt": 0, "hours": 0, "last": None})
            by_user[k]["cnt"] += 1
            by_user[k]["hours"] += t["hours"] or 0
            if not by_user[k]["last"] or t["work_date"] > by_user[k]["last"]:
                by_user[k]["last"] = t["work_date"]
        by_user_list = sorted(by_user.items(), key=lambda x: -x[1]["cnt"])[:20]

        # 타임라인용 — 일자별 그룹화
        timeline = {}
        for t in tasks:
            timeline.setdefault(t["work_date"], []).append(t)
        timeline_list = sorted(timeline.items(), reverse=True)
        # 댓글 통합
        all_comments = [dict(r) for r in c.execute(
            """SELECT tc.*, u.name AS author_name, u.rank AS author_rank,
                      tk.title AS task_title, tk.id AS tid
               FROM task_comments tc
               JOIN tasks tk ON tc.task_id=tk.id
               JOIN users u ON tc.author_id=u.id
               WHERE tk.project_id=? ORDER BY tc.created_at DESC LIMIT 30""",
            (pid,)
        ).fetchall()]
    retro = get_retro(pid)
    return ctx(req, "project_detail.html",
               user=u, p=p, tasks=tasks[:50], stats=stats,
               by_team=by_team_list, by_user=by_user_list, total_tasks=len(tasks),
               timeline=timeline_list[:30], all_comments=all_comments, retro=retro)


# =====================================================
# CUSTOMER DETAIL
# =====================================================
@app.get("/customer/{cid}", response_class=HTMLResponse)
async def customer_detail(req: Request, cid: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        cu = c.execute("SELECT * FROM customers WHERE id=?", (cid,)).fetchone()
        if not cu:
            return RedirectResponse("/", 303)
        cu = dict(cu)
        # 고객사 프로젝트
        pjts = [dict(r) for r in c.execute(
            "SELECT * FROM projects WHERE customer_id=? ORDER BY id DESC", (cid,),
        ).fetchall()]
        # 최근 2주 카드
        since = (date.today() - timedelta(days=30)).isoformat()
        tasks = [dict(r) for r in c.execute(
            """SELECT tk.*, u.name AS user_name, tm.name AS team_name, p.name AS project_name
               FROM tasks tk JOIN users u ON tk.user_id=u.id
               LEFT JOIN teams tm ON u.team_id=tm.id
               LEFT JOIN projects p ON tk.project_id=p.id
               WHERE tk.customer_id=? AND tk.work_date>=? ORDER BY tk.work_date DESC""",
            (cid, since),
        ).fetchall()]
        stats = {
            "total": len(tasks),
            "done": sum(1 for t in tasks if t["status"] == "완료"),
            "delay": sum(1 for t in tasks if t["status"] == "지연"),
            "hours": round(sum(t["hours"] or 0 for t in tasks), 1),
        }
        by_team = {}
        for t in tasks:
            k = t["team_name"] or "(미배정)"
            by_team.setdefault(k, {"cnt": 0, "hours": 0})
            by_team[k]["cnt"] += 1
            by_team[k]["hours"] += t["hours"] or 0
        by_team_list = sorted(by_team.items(), key=lambda x: -x[1]["cnt"])

    return ctx(req, "customer_detail.html",
               user=u, cu=cu, pjts=pjts, tasks=tasks[:80],
               stats=stats, by_team=by_team_list, total_tasks=len(tasks))


# =====================================================
# CALENDAR (개인 월간 뷰)
# =====================================================
@app.get("/calendar", response_class=HTMLResponse)
@app.get("/calendar/{month}", response_class=HTMLResponse)
async def calendar_page(req: Request, month: str = "", scope: str = "me"):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if not month:
        month = date.today().strftime("%Y-%m")
    # 리더 이상이 아니면 me로 강제
    if scope == "team" and u["role"] not in ("leader", "executive", "ceo", "admin"):
        scope = "me"
    y, m = int(month[:4]), int(month[5:7])
    with db_session() as c:
        if scope == "team" and u.get("team_id"):
            rows = [dict(r) for r in c.execute(
                """SELECT work_date, status, COUNT(*) AS cnt, COALESCE(SUM(hours),0) AS hours
                   FROM tasks t JOIN users u ON t.user_id=u.id
                   WHERE u.team_id=? AND work_date LIKE ?
                   GROUP BY work_date, status""",
                (u["team_id"], f"{month}%"),
            ).fetchall()]
        else:
            rows = [dict(r) for r in c.execute(
                """SELECT work_date, status, COUNT(*) AS cnt, COALESCE(SUM(hours),0) AS hours
                   FROM tasks WHERE user_id=? AND work_date LIKE ?
                   GROUP BY work_date, status""",
                (u["id"], f"{month}%"),
            ).fetchall()]
    by_date = {}
    for r in rows:
        by_date.setdefault(r["work_date"], {"total": 0, "done": 0, "progress": 0,
                                              "delay": 0, "hours": 0})
        by_date[r["work_date"]]["total"] += r["cnt"]
        by_date[r["work_date"]]["hours"] += r["hours"]
        if r["status"] == "완료":
            by_date[r["work_date"]]["done"] += r["cnt"]
        elif r["status"] == "진행중":
            by_date[r["work_date"]]["progress"] += r["cnt"]
        elif r["status"] == "지연":
            by_date[r["work_date"]]["delay"] += r["cnt"]

    first_wd, ndays = calendar.monthrange(y, m)
    weeks = []
    day = 1 - first_wd
    for _ in range(6):
        week = []
        for wd in range(7):
            if day < 1 or day > ndays:
                week.append({"day": 0, "date": "", "d": None, "is_today": False, "wd": wd})
            else:
                ds = f"{y}-{m:02d}-{day:02d}"
                week.append({
                    "day": day, "date": ds, "d": by_date.get(ds),
                    "is_today": ds == date.today().isoformat(), "wd": wd,
                })
            day += 1
        if any(c["day"] > 0 for c in week):
            weeks.append(week)

    prev_m = f"{y-1 if m==1 else y}-{12 if m==1 else m-1:02d}"
    next_m = f"{y+1 if m==12 else y}-{1 if m==12 else m+1:02d}"
    month_total = sum(d["total"] for d in by_date.values())
    month_done = sum(d["done"] for d in by_date.values())
    month_hours = round(sum(d["hours"] for d in by_date.values()), 1)

    return ctx(req, "calendar.html",
               user=u, weeks=weeks, month=month, prev_m=prev_m, next_m=next_m,
               month_total=month_total, month_done=month_done, month_hours=month_hours,
               scope=scope, can_team=u["role"] in ("leader","executive","ceo","admin") and u.get("team_id") is not None,
               active="calendar")


# =====================================================
# FEED (부서간 오늘 피드)
# =====================================================
@app.get("/feed", response_class=HTMLResponse)
@app.get("/feed/{sel_date}", response_class=HTMLResponse)
async def feed_page(req: Request, sel_date: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if not sel_date:
        sel_date = date.today().isoformat()
    with db_session() as c:
        # 팀별 오늘 한 줄 요약
        summaries = [dict(r) for r in c.execute(
            """SELECT ts.*, t.name AS team_name, t.code AS team_code, t.is_lab,
                      u.name AS author_name
               FROM team_summaries ts JOIN teams t ON ts.team_id=t.id
               LEFT JOIN users u ON ts.author_id=u.id
               WHERE ts.work_date=? ORDER BY t.display_order""",
            (sel_date,),
        ).fetchall()]
        # 전 팀 오늘 카드 (요약)
        tasks = [dict(r) for r in c.execute(
            """SELECT tk.*, u.name AS user_name, u.rank AS user_rank, u.team_id AS team_id,
                      t.name AS team_name, t.code AS team_code,
                      p.name AS project_name, cu.name AS customer_name
               FROM tasks tk JOIN users u ON tk.user_id=u.id
               LEFT JOIN teams t ON u.team_id=t.id
               LEFT JOIN projects p ON tk.project_id=p.id
               LEFT JOIN customers cu ON tk.customer_id=cu.id
               WHERE tk.work_date=? ORDER BY t.display_order, u.id, tk.id""",
            (sel_date,),
        ).fetchall()]
        # 메타데이터(댓글수+리액션수) 일괄 조회 → 카드 표면 배지로 표시
        meta = get_meta_bulk([t["id"] for t in tasks])
        # 현재 사용자 멘션/응답대기 카운트 (나를 @멘션한 댓글이 달린 카드)
        mentioned_ids = set()
        if tasks:
            ph = ",".join("?" * len(tasks))
            mrows = c.execute(
                f"""SELECT DISTINCT tc.task_id
                    FROM comment_mentions cm
                    JOIN task_comments tc ON cm.comment_id=tc.id
                    WHERE cm.user_id=? AND tc.task_id IN ({ph})""",
                (u["id"],) + tuple(t["id"] for t in tasks)
            ).fetchall()
            mentioned_ids = {r["task_id"] for r in mrows}
        for t in tasks:
            m = meta.get(t["id"], {})
            t["meta_comments"] = m.get("comments", 0)
            t["meta_ack"] = m.get("ack", 0)
            t["meta_question"] = m.get("question", 0)
            t["meta_risk"] = m.get("risk", 0)
            t["meta_ok"] = m.get("ok", 0)
            t["meta_last_comment"] = m.get("last_comment", "")
            t["meta_mentioned_me"] = 1 if t["id"] in mentioned_ids else 0
            t["meta_has_activity"] = bool(
                t["meta_comments"] or t["meta_ack"] or t["meta_question"]
                or t["meta_risk"] or t["meta_ok"]
            )
            # 정렬 키: 지연(0) → 리스크(1) → 멘션(2) → 활동(3) → 진행중(4) → 대기(5) → 완료(6)
            if t["status"] == "지연":
                sk = 0
            elif t["meta_risk"]:
                sk = 1
            elif t["meta_mentioned_me"]:
                sk = 2
            elif t["meta_has_activity"]:
                sk = 3
            elif t["status"] == "진행중":
                sk = 4
            elif t["status"] in ("대기", "보류"):
                sk = 5
            else:
                sk = 6
            t["_sort"] = sk

        by_team = {}
        for t in tasks:
            k = t["team_name"] or "(미배정)"
            by_team.setdefault(k, {"code": t["team_code"], "tasks": [],
                                    "has_urgent": False})
            by_team[k]["tasks"].append(t)
            if t["_sort"] <= 2:  # 지연/리스크/멘션 중 하나라도 있으면 긴급팀
                by_team[k]["has_urgent"] = True
        # 각 팀 내부 정렬: 우선순위 → 담당자 → id
        for k, d in by_team.items():
            d["tasks"].sort(key=lambda x: (x["_sort"], x.get("user_name") or "", x["id"]))
        by_team_list = sorted(by_team.items(), key=lambda x: x[1]["code"] or "99")

        # 전체 카운트 (필터칩용)
        all_total = len(tasks)
        all_done = sum(1 for t in tasks if t["status"] == "완료")
        all_progress = sum(1 for t in tasks if t["status"] == "진행중")
        all_delay = sum(1 for t in tasks if t["status"] == "지연")
        all_wait = sum(1 for t in tasks if t["status"] in ("대기", "보류"))
        all_risk = sum(1 for t in tasks if t["meta_risk"])
        all_mentioned = sum(1 for t in tasks if t["meta_mentioned_me"])
    prev_d = (datetime.strptime(sel_date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
    next_d = (datetime.strptime(sel_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
    # 멀티뷰용 flat JSON (칸반/캘린더)
    import json as _json
    flat_tasks = _json.dumps([{
        "id": t["id"], "title": t["title"], "status": t["status"],
        "user_name": t["user_name"], "user_rank": t.get("user_rank") or "",
        "team_name": t.get("team_name") or "", "project_name": t.get("project_name") or "",
        "customer_name": t.get("customer_name") or "",
        "hours": t.get("hours") or 0,
        "cm": t.get("meta_comments", 0), "risk": t.get("meta_risk", 0),
        "question": t.get("meta_question", 0), "ack": t.get("meta_ack", 0),
    } for t in tasks], ensure_ascii=False)
    # 역할 기반 기본 펼침 로직용 — 내 팀 id
    my_team_id = u.get("team_id")
    return ctx(req, "feed.html",
               user=u, summaries=summaries, by_team=by_team_list,
               sel_date=sel_date, prev_date=prev_d, next_date=next_d,
               all_total=all_total, all_done=all_done, all_progress=all_progress,
               all_delay=all_delay, all_wait=all_wait,
               all_risk=all_risk, all_mentioned=all_mentioned,
               my_team_id=my_team_id, flat_tasks_json=flat_tasks,
               sel_date_obj=sel_date, active="feed")


# =====================================================
# ADMIN — 관리자 페이지
# =====================================================
@app.get("/admin", response_class=HTMLResponse)
async def admin_page(req: Request):
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        teams = [dict(r) for r in c.execute(
            """SELECT t.*, u.name AS leader_name,
                      (SELECT COUNT(*) FROM users WHERE team_id=t.id AND is_active=1) AS member_count
               FROM teams t LEFT JOIN users u ON t.leader_id=u.id
               ORDER BY t.display_order"""
        ).fetchall()]
        users = [dict(r) for r in c.execute(
            """SELECT u.*, t.name AS team_name FROM users u
               LEFT JOIN teams t ON u.team_id=t.id
               WHERE u.role!='admin' ORDER BY t.display_order, u.role DESC, u.id"""
        ).fetchall()]
        projects = [dict(r) for r in c.execute(
            """SELECT p.*, cu.name AS customer_name FROM projects p
               LEFT JOIN customers cu ON p.customer_id=cu.id ORDER BY p.id DESC"""
        ).fetchall()]
        customers = [dict(r) for r in c.execute("SELECT * FROM customers ORDER BY tier DESC, id").fetchall()]
    return ctx(req, "admin.html",
               user=u, teams=teams, users=users, projects=projects, customers=customers,
               active="admin")


@app.post("/api/admin/user")
async def api_admin_user(req: Request):
    u = require(req, ["admin"])
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    d = await req.json()
    with db_session() as c:
        if d.get("id"):
            c.execute(
                "UPDATE users SET name=?, team_id=?, rank=?, role=?, is_active=? WHERE id=?",
                (d["name"], d.get("team_id") or None, d.get("rank", ""),
                 d.get("role", "member"), int(d.get("is_active", 1)), d["id"]),
            )
            if d.get("password"):
                c.execute("UPDATE users SET password=? WHERE id=?", (hash_pw(d["password"]), d["id"]))
        else:
            c.execute(
                """INSERT INTO users(name, login_id, password, team_id, rank, role)
                   VALUES(?,?,?,?,?,?)""",
                (d["name"], d["login_id"], hash_pw(d.get("password", "knk1234")),
                 d.get("team_id") or None, d.get("rank", ""), d.get("role", "member")),
            )
    return JSONResponse({"ok": True})


@app.post("/api/admin/project")
async def api_admin_project(req: Request):
    u = require(req, ["admin", "ceo"])
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    d = await req.json()
    with db_session() as c:
        if d.get("id"):
            c.execute(
                "UPDATE projects SET code=?, name=?, customer_id=?, type=?, status=? WHERE id=?",
                (d.get("code", ""), d["name"], d.get("customer_id") or None,
                 d.get("type", ""), d.get("status", "진행중"), d["id"]),
            )
        else:
            c.execute(
                "INSERT INTO projects(code, name, customer_id, type, status) VALUES(?,?,?,?,?)",
                (d.get("code", ""), d["name"], d.get("customer_id") or None,
                 d.get("type", ""), d.get("status", "진행중")),
            )
    return JSONResponse({"ok": True})


@app.post("/api/admin/customer")
async def api_admin_customer(req: Request):
    u = require(req, ["admin", "ceo"])
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    d = await req.json()
    with db_session() as c:
        if d.get("id"):
            c.execute(
                "UPDATE customers SET name=?, tier=?, note=? WHERE id=?",
                (d["name"], d.get("tier", "일반"), d.get("note", ""), d["id"]),
            )
        else:
            c.execute(
                "INSERT INTO customers(name, tier, note) VALUES(?,?,?)",
                (d["name"], d.get("tier", "일반"), d.get("note", "")),
            )
    return JSONResponse({"ok": True})


# =====================================================
# MGMT CODE IMPORT — 관리코드발행목록.xls 업로드
# =====================================================
@app.post("/api/admin/import-mgmt")
async def api_admin_import_mgmt(req: Request, file: UploadFile = File(...)):
    u = require(req, ["admin", "ceo"])
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    fn = (file.filename or "").lower()
    if not (fn.endswith(".xls") or fn.endswith(".xlsx")):
        return JSONResponse({"error": "xls/xlsx 파일만 업로드 가능"}, 400)
    try:
        data = await file.read()
        suffix = ".xls" if fn.endswith(".xls") else ".xlsx"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tf:
            tf.write(data)
            tmp_path = tf.name
        try:
            rows = parse_mgmt_xls(tmp_path)
            result = import_mgmt_rows(rows)
            result["parsed"] = len(rows)
            result["filename"] = file.filename
            return JSONResponse({"ok": True, "result": result})
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass
    except Exception as e:
        return JSONResponse({"error": f"파싱 실패: {type(e).__name__}: {e}"}, 500)


# =====================================================
# INITIAL PASSWORD REGENERATION (A안 킥오프)
# =====================================================
@app.post("/api/admin/regenerate-passwords")
async def api_regen_pw(req: Request):
    u = require(req, ["admin", "ceo"])
    if not u:
        return JSONResponse({"error": "권한 없음"}, 401)
    rows = regenerate_user_passwords()
    out_dir = os.path.join(BASE, "data")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "초기비밀번호_배포용.xlsx")
    build_password_xlsx(rows, out_path)
    return JSONResponse({"ok": True, "count": len(rows),
                         "download": "/admin/download-passwords"})


@app.get("/admin/download-passwords")
async def dl_passwords(req: Request):
    u = require(req, ["admin", "ceo"])
    if not u:
        return RedirectResponse("/login", 303)
    out_path = os.path.join(BASE, "data", "초기비밀번호_배포용.xlsx")
    if not os.path.exists(out_path):
        return JSONResponse({"error": "먼저 재생성을 수행하세요"}, 404)
    with open(out_path, "rb") as f:
        data = f.read()
    from urllib.parse import quote
    fn = "KNK_초기비밀번호_배포용.xlsx"
    headers = {
        "Content-Disposition": f"attachment; filename=KNK_passwords.xlsx; filename*=UTF-8''{quote(fn)}"
    }
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


# =====================================================
# PROFILE — 비밀번호 변경
# =====================================================
@app.get("/profile", response_class=HTMLResponse)
async def profile_page(req: Request, msg: str = "", err: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    return ctx(req, "profile.html", user=u, msg=msg, err=err, active="profile")


@app.post("/profile/change-password")
async def change_password(req: Request,
                          current_pw: str = Form(...),
                          new_pw: str = Form(...),
                          confirm_pw: str = Form(...)):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    if new_pw != confirm_pw:
        return RedirectResponse("/profile?err=새 비밀번호가 일치하지 않습니다", 303)
    if len(new_pw) < 6:
        return RedirectResponse("/profile?err=비밀번호는 6자 이상", 303)
    with db_session() as c:
        row = c.execute("SELECT password FROM users WHERE id=?", (u["id"],)).fetchone()
        if not row or row["password"] != hash_pw(current_pw):
            return RedirectResponse("/profile?err=현재 비밀번호가 맞지 않습니다", 303)
        c.execute("UPDATE users SET password=? WHERE id=?",
                  (hash_pw(new_pw), u["id"]))
    return RedirectResponse("/profile?msg=비밀번호가 변경되었습니다", 303)


# =====================================================
# REMINDERS — 팀장용 오늘 미작성자 리스트
# =====================================================
@app.get("/admin/reminders", response_class=HTMLResponse)
async def reminders_page(req: Request, sel_date: str = ""):
    u = require(req, ["leader", "executive", "ceo", "admin"])
    if not u:
        return RedirectResponse("/login", 303)
    if not sel_date:
        sel_date = date.today().isoformat()
    with db_session() as c:
        teams = [dict(r) for r in c.execute(
            """SELECT t.*, ul.name AS leader_name, ul.email AS leader_email
               FROM teams t LEFT JOIN users ul ON t.leader_id=ul.id
               ORDER BY t.display_order"""
        ).fetchall()]
        # 팀장이면 자기 팀만 표시
        if u["role"] == "leader":
            teams = [t for t in teams if t["id"] == u["team_id"]]
        for t in teams:
            members = [dict(r) for r in c.execute(
                """SELECT u.id, u.name, u.rank, u.email
                   FROM users u WHERE u.team_id=? AND u.is_active=1
                   AND u.role!='admin' ORDER BY u.id""",
                (t["id"],),
            ).fetchall()]
            reported_ids = {r["user_id"] for r in c.execute(
                """SELECT DISTINCT user_id FROM tasks
                   WHERE work_date=? AND user_id IN
                   (SELECT id FROM users WHERE team_id=?)""",
                (sel_date, t["id"]),
            ).fetchall()}
            t["members"] = members
            t["missing"] = [m for m in members if m["id"] not in reported_ids]
            t["reported_count"] = len(members) - len(t["missing"])
            t["total"] = len(members)
            t["participation"] = round(t["reported_count"] * 100 / max(t["total"], 1))
    return ctx(req, "reminders.html", user=u, teams=teams,
               sel_date=sel_date, active="reminders")


# =====================================================
# EXCEL EXPORT — 주간 요약 엑셀
# =====================================================
@app.get("/export/weekly")
async def export_weekly(req: Request, wk_mon: str = ""):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return JSONResponse({"error": "openpyxl 미설치"}, 500)
    if not wk_mon:
        td = date.today()
        wk_mon = (td - timedelta(days=td.weekday())).isoformat()
    mon = datetime.strptime(wk_mon, "%Y-%m-%d").date()
    sun = mon + timedelta(days=6)

    wb = Workbook()
    red_fill = PatternFill(start_color="A5282C", end_color="A5282C", fill_type="solid")
    head_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="맑은 고딕", size=10)
    thin = Side(border_style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    with db_session() as c:
        # Sheet 1: 전사 요약
        ws = wb.active
        ws.title = "전사 요약"
        ws["A1"] = f"㈜케이엔케이 주간 업무 요약"
        ws["A1"].font = Font(name="맑은 고딕", size=14, bold=True, color="A5282C")
        ws["A2"] = f"{mon} ~ {sun}"
        ws["A2"].font = Font(name="맑은 고딕", size=10, color="4A4A4A")
        headers = ["팀코드", "팀명", "팀장", "인원", "카드수", "완료", "지연", "공수(h)"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = head_font
            cell.fill = red_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        teams = [dict(r) for r in c.execute(
            """SELECT t.*, u.name AS leader_name,
                      (SELECT COUNT(*) FROM users WHERE team_id=t.id AND is_active=1) AS mc
               FROM teams t LEFT JOIN users u ON t.leader_id=u.id
               ORDER BY t.display_order"""
        ).fetchall()]
        row = 5
        for t in teams:
            s = c.execute(
                """SELECT COUNT(*) AS total,
                          SUM(CASE WHEN status='완료' THEN 1 ELSE 0 END) AS done,
                          SUM(CASE WHEN status='지연' THEN 1 ELSE 0 END) AS delay,
                          COALESCE(SUM(hours),0) AS hours
                   FROM tasks tk JOIN users u ON tk.user_id=u.id
                   WHERE u.team_id=? AND tk.work_date>=? AND tk.work_date<=?""",
                (t["id"], mon.isoformat(), sun.isoformat()),
            ).fetchone()
            vals = [t["code"], t["name"], t["leader_name"] or "-", t["mc"],
                    s["total"] or 0, s["done"] or 0, s["delay"] or 0, round(s["hours"] or 0, 1)]
            for col, v in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.font = body_font
                cell.border = border
                if col >= 4:
                    cell.alignment = Alignment(horizontal="right")
            row += 1
        for col, w in enumerate([8, 18, 14, 8, 10, 10, 10, 12], start=1):
            ws.column_dimensions[chr(64 + col)].width = w

        # Sheet 2: 카드 상세
        ws2 = wb.create_sheet("카드 상세")
        headers2 = ["날짜", "팀", "이름", "직급", "제목", "분류", "프로젝트", "고객사", "상태", "공수(h)"]
        for col, h in enumerate(headers2, start=1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = head_font
            cell.fill = red_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        rows = c.execute(
            """SELECT tk.work_date, t.name AS team_name, u.name AS user_name, u.rank,
                      tk.title, tk.category, p.name AS pj, cu.name AS cu,
                      tk.status, tk.hours
               FROM tasks tk JOIN users u ON tk.user_id=u.id
               LEFT JOIN teams t ON u.team_id=t.id
               LEFT JOIN projects p ON tk.project_id=p.id
               LEFT JOIN customers cu ON tk.customer_id=cu.id
               WHERE tk.work_date>=? AND tk.work_date<=?
               ORDER BY tk.work_date, t.display_order, u.id""",
            (mon.isoformat(), sun.isoformat()),
        ).fetchall()
        for ri, r in enumerate(rows, start=2):
            for col, v in enumerate(r, start=1):
                cell = ws2.cell(row=ri, column=col, value=v)
                cell.font = body_font
                cell.border = border
        for col, w in enumerate([12, 14, 10, 10, 40, 10, 24, 14, 10, 10], start=1):
            ws2.column_dimensions[chr(64 + col)].width = w
        ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"KNK_주간요약_{mon}_{sun}.xlsx"
    from urllib.parse import quote
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )


# =====================================================
# HAIST WORKS — 게시판 라우트
# =====================================================
from .database import (board_get_or_create_company, board_get_or_create_team,
                        board_posts_list, board_posts_pending, board_post_get,
                        board_post_create, board_post_update, board_post_delete,
                        board_post_approve, board_post_reject, board_post_toggle_pin,
                        board_post_increment_view,
                        board_comments_list, board_comment_create, board_comment_delete,
                        BOARD_CATEGORIES)


def _is_team_leader(user, team_id):
    """사용자가 해당 팀의 팀장인지 확인"""
    if not user:
        return False
    if user.get("role") in ("admin", "ceo", "executive"):
        return True
    if user.get("role") == "leader" and user.get("team_id") == team_id:
        return True
    return False


@app.get("/board/company", response_class=HTMLResponse)
async def board_company(req: Request):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    bid = board_get_or_create_company()
    posts = board_posts_list(bid)
    can_write = u["role"] in ("admin", "ceo", "executive")
    return ctx(req, "board_list.html", user=u, active="board_company",
               board_id=bid, board_name="전사 게시판", board_type="company",
               posts=posts, can_write=can_write, is_leader=False,
               pending_count=0, CATEGORIES=BOARD_CATEGORIES)


@app.get("/board/team", response_class=HTMLResponse)
async def board_team_my(req: Request):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    tid = u.get("team_id")
    # 팀 미소속 (admin/ceo 등) → 전체 부서 게시판 목록
    if not tid:
        with db_session() as c:
            teams = c.execute(
                """SELECT t.id, t.name, t.sector,
                          (SELECT COUNT(*) FROM board_posts bp
                           JOIN boards b ON bp.board_id=b.id
                           WHERE b.type='team' AND b.team_id=t.id
                           AND bp.approval_status='approved') AS post_count,
                          (SELECT COUNT(*) FROM board_posts bp
                           JOIN boards b ON bp.board_id=b.id
                           WHERE b.type='team' AND b.team_id=t.id
                           AND bp.approval_status='pending') AS pending_count
                   FROM teams t ORDER BY t.display_order"""
            ).fetchall()
        return ctx(req, "board_teams.html", user=u, active="board_team",
                   teams=teams)
    bid = board_get_or_create_team(tid)
    is_leader = _is_team_leader(u, tid)
    posts = board_posts_list(bid, include_pending=is_leader)
    pending = board_posts_pending(bid) if is_leader else []
    return ctx(req, "board_list.html", user=u, active="board_team",
               board_id=bid, board_name=f"{u.get('team_name','')} 게시판",
               board_type="team", team_id=tid,
               posts=posts, can_write=True, is_leader=is_leader,
               pending_count=len(pending), pending_posts=pending,
               CATEGORIES=BOARD_CATEGORIES)


@app.get("/board/team/{team_id}", response_class=HTMLResponse)
async def board_team_specific(req: Request, team_id: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    bid = board_get_or_create_team(team_id)
    is_leader = _is_team_leader(u, team_id)
    with db_session() as c:
        t = c.execute("SELECT name FROM teams WHERE id=?", (team_id,)).fetchone()
    team_name = t["name"] if t else f"팀{team_id}"
    posts = board_posts_list(bid, include_pending=is_leader)
    pending = board_posts_pending(bid) if is_leader else []
    can_write = (u.get("team_id") == team_id) or u["role"] in ("admin", "ceo", "executive")
    return ctx(req, "board_list.html", user=u, active="board_team",
               board_id=bid, board_name=f"{team_name} 게시판",
               board_type="team", team_id=team_id,
               posts=posts, can_write=can_write, is_leader=is_leader,
               pending_count=len(pending), pending_posts=pending,
               CATEGORIES=BOARD_CATEGORIES)


@app.get("/board/new", response_class=HTMLResponse)
async def board_new_form(req: Request, board_id: int = 0):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    return ctx(req, "board_form.html", user=u, active="board",
               board_id=board_id, post=None, CATEGORIES=BOARD_CATEGORIES)


@app.post("/board/new")
async def board_new_submit(req: Request,
                           board_id: int = Form(...),
                           title: str = Form(...),
                           body: str = Form(""),
                           category: str = Form("일반")):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    # 전사 게시판: admin/ceo/executive만 → 바로 approved
    # 부서 게시판: 팀장/경영진 → approved, 일반 부서원 → pending
    with db_session() as c:
        board = c.execute("SELECT type, team_id FROM boards WHERE id=?", (board_id,)).fetchone()
    if not board:
        return RedirectResponse("/home", 303)

    if board["type"] == "company":
        status = "approved"
        redirect_url = "/board/company"
    else:
        if _is_team_leader(u, board["team_id"]):
            status = "approved"
        else:
            status = "pending"
        redirect_url = f"/board/team/{board['team_id']}"

    board_post_create(board_id, u["id"], title, body, category, status)
    return RedirectResponse(redirect_url, 303)


@app.get("/board/post/{post_id}", response_class=HTMLResponse)
async def board_post_detail(req: Request, post_id: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    post = board_post_get(post_id)
    if not post:
        return RedirectResponse("/home", 303)
    board_post_increment_view(post_id)
    comments = board_comments_list(post_id)
    is_leader = _is_team_leader(u, post.get("board_team_id"))
    is_author = post["author_id"] == u["id"]
    return ctx(req, "board_detail.html", user=u, active="board",
               post=post, comments=comments,
               is_leader=is_leader, is_author=is_author)


@app.post("/board/post/{post_id}/comment")
async def board_add_comment(req: Request, post_id: int, body: str = Form(...)):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    board_comment_create(post_id, u["id"], body)
    return RedirectResponse(f"/board/post/{post_id}", 303)


@app.post("/board/post/{post_id}/approve")
async def board_approve(req: Request, post_id: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    post = board_post_get(post_id)
    if post and _is_team_leader(u, post.get("board_team_id")):
        board_post_approve(post_id, u["id"])
    return RedirectResponse(f"/board/team/{post['board_team_id']}" if post else "/home", 303)


@app.post("/board/post/{post_id}/reject")
async def board_reject(req: Request, post_id: int, reason: str = Form("")):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    post = board_post_get(post_id)
    if post and _is_team_leader(u, post.get("board_team_id")):
        board_post_reject(post_id, u["id"], reason)
    return RedirectResponse(f"/board/team/{post['board_team_id']}" if post else "/home", 303)


@app.post("/board/post/{post_id}/pin")
async def board_pin(req: Request, post_id: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    post = board_post_get(post_id)
    if post and _is_team_leader(u, post.get("board_team_id")):
        board_post_toggle_pin(post_id)
    redir = f"/board/team/{post['board_team_id']}" if post and post["board_type"] == "team" else "/board/company"
    return RedirectResponse(redir, 303)


@app.post("/board/post/{post_id}/delete")
async def board_delete(req: Request, post_id: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    post = board_post_get(post_id)
    if post and (post["author_id"] == u["id"] or _is_team_leader(u, post.get("board_team_id"))):
        board_post_delete(post_id)
    redir = f"/board/team/{post['board_team_id']}" if post and post["board_type"] == "team" else "/board/company"
    return RedirectResponse(redir, 303)


@app.post("/board/comment/{cid}/delete")
async def board_del_comment(req: Request, cid: int):
    u = get_user(req)
    if not u:
        return RedirectResponse("/login", 303)
    with db_session() as c:
        cm = c.execute("SELECT bc.post_id, bc.author_id FROM board_comments bc WHERE bc.id=?", (cid,)).fetchone()
    if cm and (cm["author_id"] == u["id"] or u["role"] in ("admin", "ceo")):
        board_comment_delete(cid)
        return RedirectResponse(f"/board/post/{cm['post_id']}", 303)
    return RedirectResponse("/home", 303)


# =====================================================
# HAIST WORKS — 물류 모듈 라우트
# =====================================================
from . import database as _logi


# ── 관리자: 물류 권한 토글 ─────────────────────────────
@app.post("/api/admin/user-logistics")
async def admin_toggle_logistics(request: Request,
                                 user_id: int = Form(...),
                                 can_use_logistics: str = Form("0")):
    me = require(request, roles=["admin", "ceo"])
    if not me:
        return JSONResponse({"ok": False, "error": "관리자 권한 필요"}, status_code=403)
    flag = 1 if can_use_logistics == "1" else 0
    with db_session() as c:
        c.execute(
            "UPDATE users SET can_use_logistics = ? WHERE id = ?",
            (flag, user_id),
        )
    return JSONResponse({"ok": True, "user_id": user_id, "can_use_logistics": flag})


# ── 물류 대시보드 ─────────────────────────────────────────
@app.get("/logistics", response_class=HTMLResponse)
async def logi_dashboard(request: Request):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    parts_stats = _logi.parts_count()
    proj_stats = _logi.projects_count_logi()
    return ctx(request, "logistics_home.html",
               user=u, active="logistics",
               parts_stats=parts_stats, proj_stats=proj_stats)


# ── 부품 마스터 (parts) ────────────────────────────────
@app.get("/parts", response_class=HTMLResponse)
async def parts_list_page(request: Request, q: str = "", biz_div: str = "",
                          category: str = ""):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    rows = _logi.parts_list(q=q, biz_div=biz_div, category=category)
    return ctx(request, "parts.html",
               user=u, active="parts",
               parts=rows, q=q, biz_div=biz_div, category=category)


@app.get("/parts/new", response_class=HTMLResponse)
async def parts_new_form(request: Request):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    return ctx(request, "part_form.html", user=u, active="parts", part=None)


@app.post("/parts/new")
async def parts_new_submit(
    request: Request,
    part_no: str = Form(...), part_name: str = Form(...),
    spec: str = Form(""), maker: str = Form(""), origin: str = Form(""),
    unit: str = Form("EA"), currency: str = Form("KRW"),
    std_price: str = Form("0"), biz_div: str = Form(""),
    category: str = Form(""), note: str = Form(""), is_active: str = Form("1"),
):
    _u = get_user(request)
    if not _u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(_u):
        return RedirectResponse("/home", 303)
    _logi.parts_create({
        "part_no": part_no, "part_name": part_name, "spec": spec,
        "maker": maker, "origin": origin, "unit": unit,
        "currency": currency, "std_price": std_price,
        "biz_div": biz_div, "category": category, "note": note,
        "is_active": is_active,
    })
    return RedirectResponse("/parts", status_code=303)


@app.get("/parts/{pid}/edit", response_class=HTMLResponse)
async def parts_edit_form(request: Request, pid: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    p = _logi.parts_get(pid)
    if not p:
        return RedirectResponse("/parts", status_code=303)
    return ctx(request, "part_form.html", user=u, active="parts", part=p)


@app.post("/parts/{pid}/edit")
async def parts_edit_submit(
    request: Request,
    pid: int, part_no: str = Form(...), part_name: str = Form(...),
    spec: str = Form(""), maker: str = Form(""), origin: str = Form(""),
    unit: str = Form("EA"), currency: str = Form("KRW"),
    std_price: str = Form("0"), biz_div: str = Form(""),
    category: str = Form(""), note: str = Form(""), is_active: str = Form("1"),
):
    _u = get_user(request)
    if not _u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(_u):
        return RedirectResponse("/home", 303)
    _logi.parts_update(pid, {
        "part_no": part_no, "part_name": part_name, "spec": spec,
        "maker": maker, "origin": origin, "unit": unit,
        "currency": currency, "std_price": std_price,
        "biz_div": biz_div, "category": category, "note": note,
        "is_active": is_active,
    })
    return RedirectResponse("/parts", status_code=303)


@app.post("/parts/{pid}/delete")
async def parts_delete_submit(request: Request, pid: int):
    _u = get_user(request)
    if not _u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(_u):
        return RedirectResponse("/home", 303)
    _logi.parts_delete(pid)
    return RedirectResponse("/parts", status_code=303)


# ── 프로젝트 / 관리코드 발행대장 ─────────────────────────
@app.get("/projects", response_class=HTMLResponse)
async def projects_list_page(request: Request, q: str = "", biz_div: str = "",
                             stage: str = "", status: str = ""):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    rows = _logi.projects_list_logi(q=q, biz_div=biz_div, stage=stage, status=status)
    return ctx(request, "projects.html",
               user=u, active="logi_projects",
               projects=rows, q=q, biz_div=biz_div, stage=stage, status=status,
               STAGES=_logi.STAGES, STATUSES=_logi.LOGI_STATUSES)


@app.get("/projects/new", response_class=HTMLResponse)
async def projects_new_form(request: Request):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    return ctx(request, "project_form.html",
               user=u, active="logi_projects",
               project=None,
               STAGES=_logi.STAGES, STATUSES=_logi.LOGI_STATUSES,
               PO_TYPES=_logi.PO_TYPES)


@app.post("/projects/new")
async def projects_new_submit(
    request: Request,
    biz_div: str = Form(...), project_name: str = Form(...),
    customer: str = Form(""), model: str = Form(""),
    stage: str = Form("제안작성"), po_type: str = Form("신규"),
    status: str = Form("수주예정"), customer_po: str = Form(""),
    currency: str = Form("KRW"), order_amount: str = Form("0"),
    order_date: str = Form(""), due_date: str = Form(""),
    pm: str = Form(""), sales: str = Form(""), note: str = Form(""),
):
    _u = get_user(request)
    if not _u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(_u):
        return RedirectResponse("/home", 303)
    _logi.projects_create_logi({
        "biz_div": biz_div, "project_name": project_name, "customer": customer,
        "model": model, "stage": stage, "po_type": po_type, "status": status,
        "customer_po": customer_po, "currency": currency,
        "order_amount": order_amount, "order_date": order_date, "due_date": due_date,
        "pm": pm, "sales": sales, "note": note,
    })
    return RedirectResponse("/projects", status_code=303)


@app.get("/projects/{pid}/edit", response_class=HTMLResponse)
async def projects_edit_form(request: Request, pid: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    p = _logi.projects_get_logi(pid)
    if not p:
        return RedirectResponse("/projects", status_code=303)
    return ctx(request, "project_form.html",
               user=u, active="logi_projects",
               project=p,
               STAGES=_logi.STAGES, STATUSES=_logi.LOGI_STATUSES,
               PO_TYPES=_logi.PO_TYPES)


@app.post("/projects/{pid}/edit")
async def projects_edit_submit(
    request: Request,
    pid: int, biz_div: str = Form(...), project_name: str = Form(...),
    customer: str = Form(""), model: str = Form(""),
    stage: str = Form("제안작성"), po_type: str = Form("신규"),
    status: str = Form("수주예정"), customer_po: str = Form(""),
    currency: str = Form("KRW"), order_amount: str = Form("0"),
    order_date: str = Form(""), due_date: str = Form(""),
    pm: str = Form(""), sales: str = Form(""), note: str = Form(""),
):
    _u = get_user(request)
    if not _u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(_u):
        return RedirectResponse("/home", 303)
    _logi.projects_update_logi(pid, {
        "biz_div": biz_div, "project_name": project_name, "customer": customer,
        "model": model, "stage": stage, "po_type": po_type, "status": status,
        "customer_po": customer_po, "currency": currency,
        "order_amount": order_amount, "order_date": order_date, "due_date": due_date,
        "pm": pm, "sales": sales, "note": note,
    })
    return RedirectResponse("/projects", status_code=303)


@app.post("/projects/{pid}/delete")
async def projects_delete_submit(request: Request, pid: int):
    _u = get_user(request)
    if not _u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(_u):
        return RedirectResponse("/home", 303)
    _logi.projects_delete_logi(pid)
    return RedirectResponse("/projects", status_code=303)


# =====================================================
# HAIST WORKS — 공급사 (suppliers) 라우트
# =====================================================
@app.get("/suppliers", response_class=HTMLResponse)
async def suppliers_list_page(request: Request, q: str = ""):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    rows = _logi.suppliers_list(q=q)
    return ctx(request, "suppliers.html",
               user=u, active="suppliers",
               suppliers=rows, q=q)


@app.get("/suppliers/new", response_class=HTMLResponse)
async def suppliers_new_form(request: Request):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    return ctx(request, "supplier_form.html",
               user=u, active="suppliers", supplier=None,
               PAYMENT_TERMS=_logi.PAYMENT_TERMS)


@app.post("/suppliers/new")
async def suppliers_new_submit(
    request: Request,
    name: str = Form(...), code: str = Form(""), contact: str = Form(""),
    email: str = Form(""), phone: str = Form(""), country: str = Form(""),
    currency: str = Form("KRW"), payment_terms: str = Form(""),
    note: str = Form(""), is_active: str = Form("1"),
):
    _u = get_user(request)
    if not _u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(_u):
        return RedirectResponse("/home", 303)
    _logi.supplier_create({
        "name": name, "code": code, "contact": contact, "email": email,
        "phone": phone, "country": country, "currency": currency,
        "payment_terms": payment_terms, "note": note, "is_active": is_active,
    })
    return RedirectResponse("/suppliers", status_code=303)


@app.get("/suppliers/{sid}/edit", response_class=HTMLResponse)
async def suppliers_edit_form(request: Request, sid: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    s = _logi.supplier_get(sid)
    if not s:
        return RedirectResponse("/suppliers", 303)
    return ctx(request, "supplier_form.html",
               user=u, active="suppliers", supplier=s,
               PAYMENT_TERMS=_logi.PAYMENT_TERMS)


@app.post("/suppliers/{sid}/edit")
async def suppliers_edit_submit(
    request: Request, sid: int,
    name: str = Form(...), code: str = Form(""), contact: str = Form(""),
    email: str = Form(""), phone: str = Form(""), country: str = Form(""),
    currency: str = Form("KRW"), payment_terms: str = Form(""),
    note: str = Form(""), is_active: str = Form("1"),
):
    _u = get_user(request)
    if not _u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(_u):
        return RedirectResponse("/home", 303)
    _logi.supplier_update(sid, {
        "name": name, "code": code, "contact": contact, "email": email,
        "phone": phone, "country": country, "currency": currency,
        "payment_terms": payment_terms, "note": note, "is_active": is_active,
    })
    return RedirectResponse("/suppliers", status_code=303)


@app.post("/suppliers/{sid}/delete")
async def suppliers_delete_submit(request: Request, sid: int):
    _u = get_user(request)
    if not _u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(_u):
        return RedirectResponse("/home", 303)
    _logi.supplier_delete(sid)
    return RedirectResponse("/suppliers", status_code=303)


# =====================================================
# HAIST WORKS — 발주 (purchase_orders) 라우트
# =====================================================
@app.get("/po", response_class=HTMLResponse)
async def po_list_page(request: Request, q: str = "", status: str = ""):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    rows = _logi.po_list(q=q, status=status)
    return ctx(request, "po_list.html",
               user=u, active="po",
               orders=rows, q=q, status=status,
               PO_STATUSES=_logi.PO_STATUSES)


@app.get("/po/new", response_class=HTMLResponse)
async def po_new_form(request: Request):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    # 드롭다운용 데이터
    suppliers = _logi.suppliers_list(active_only=True)
    # 관리코드 발급된 프로젝트만 선택 가능
    projects = _logi.projects_list_logi()
    projects_with_code = [p for p in projects if p["mgmt_code"]]
    parts_all = _logi.parts_list()
    return ctx(request, "po_form.html",
               user=u, active="po", po=None, items=[],
               suppliers=suppliers, projects=projects_with_code,
               parts_all=parts_all,
               PO_STATUSES=_logi.PO_STATUSES,
               SHIPPING_TERMS=_logi.SHIPPING_TERMS,
               PAYMENT_TERMS=_logi.PAYMENT_TERMS,
               PO_TYPES_KIND=_logi.PO_TYPES_KIND)


@app.post("/po/new")
async def po_new_submit(request: Request):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    form = await request.form()
    header = {
        "project_id": form.get("project_id", ""),
        "supplier_id": form.get("supplier_id", ""),
        "order_date": form.get("order_date", ""),
        "expected_date": form.get("expected_date", ""),
        "currency": form.get("currency", "KRW"),
        "exchange_rate": form.get("exchange_rate", "1"),
        "status": form.get("status", "작성중"),
        "shipping_terms": form.get("shipping_terms", ""),
        "payment_terms": form.get("payment_terms", ""),
        "po_type": form.get("po_type", "일반"),
        "note": form.get("note", ""),
    }
    # 라인 파싱: item_part_id[], item_qty[], item_price[], item_delivery[], item_note[]
    part_ids = form.getlist("item_part_id")
    qtys = form.getlist("item_qty")
    prices = form.getlist("item_price")
    delivs = form.getlist("item_delivery")
    notes = form.getlist("item_note")
    items = []
    for i, pid in enumerate(part_ids):
        if not pid and not (qtys[i] if i < len(qtys) else ""):
            continue
        items.append({
            "part_id": pid,
            "quantity": qtys[i] if i < len(qtys) else "0",
            "unit_price": prices[i] if i < len(prices) else "0",
            "delivery_date": delivs[i] if i < len(delivs) else "",
            "note": notes[i] if i < len(notes) else "",
        })
    po_id, po_num = _logi.po_create(header, items, created_by=u["id"])
    return RedirectResponse(f"/po/{po_id}", status_code=303)


@app.get("/po/{po_id}", response_class=HTMLResponse)
async def po_detail(request: Request, po_id: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    header, items = _logi.po_get(po_id)
    if not header:
        return RedirectResponse("/po", 303)
    return ctx(request, "po_detail.html",
               user=u, active="po", po=header, items=items)


@app.get("/po/{po_id}/edit", response_class=HTMLResponse)
async def po_edit_form(request: Request, po_id: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    header, items = _logi.po_get(po_id)
    if not header:
        return RedirectResponse("/po", 303)
    suppliers = _logi.suppliers_list(active_only=True)
    projects = [p for p in _logi.projects_list_logi() if p["mgmt_code"]]
    parts_all = _logi.parts_list()
    return ctx(request, "po_form.html",
               user=u, active="po", po=header, items=items,
               suppliers=suppliers, projects=projects,
               parts_all=parts_all,
               PO_STATUSES=_logi.PO_STATUSES,
               SHIPPING_TERMS=_logi.SHIPPING_TERMS,
               PAYMENT_TERMS=_logi.PAYMENT_TERMS,
               PO_TYPES_KIND=_logi.PO_TYPES_KIND)


@app.post("/po/{po_id}/edit")
async def po_edit_submit(request: Request, po_id: int):
    u = get_user(request)
    if not u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(u):
        return RedirectResponse("/home", 303)
    form = await request.form()
    header = {
        "project_id": form.get("project_id", ""),
        "supplier_id": form.get("supplier_id", ""),
        "order_date": form.get("order_date", ""),
        "expected_date": form.get("expected_date", ""),
        "currency": form.get("currency", "KRW"),
        "exchange_rate": form.get("exchange_rate", "1"),
        "status": form.get("status", "작성중"),
        "shipping_terms": form.get("shipping_terms", ""),
        "payment_terms": form.get("payment_terms", ""),
        "po_type": form.get("po_type", "일반"),
        "note": form.get("note", ""),
    }
    part_ids = form.getlist("item_part_id")
    qtys = form.getlist("item_qty")
    prices = form.getlist("item_price")
    delivs = form.getlist("item_delivery")
    notes = form.getlist("item_note")
    items = []
    for i, pid in enumerate(part_ids):
        if not pid and not (qtys[i] if i < len(qtys) else ""):
            continue
        items.append({
            "part_id": pid,
            "quantity": qtys[i] if i < len(qtys) else "0",
            "unit_price": prices[i] if i < len(prices) else "0",
            "delivery_date": delivs[i] if i < len(delivs) else "",
            "note": notes[i] if i < len(notes) else "",
        })
    _logi.po_update(po_id, header, items)
    return RedirectResponse(f"/po/{po_id}", status_code=303)


@app.post("/po/{po_id}/delete")
async def po_delete_submit(request: Request, po_id: int):
    _u = get_user(request)
    if not _u:
        return RedirectResponse("/login", 303)
    if not can_use_logistics(_u):
        return RedirectResponse("/home", 303)
    _logi.po_delete(po_id)
    return RedirectResponse("/po", status_code=303)
