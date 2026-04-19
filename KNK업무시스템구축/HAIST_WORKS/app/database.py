"""
KNK 일일업무일지 v2 - Database Layer
Phase 1 MVP - Task Card 기반 구조
"""
import sqlite3, os, hashlib
from contextlib import contextmanager

DB_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "knk.db")

def hash_pw(pw: str) -> str:
    return hashlib.sha256(("knk-haist-" + pw).encode()).hexdigest()

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

@contextmanager
def db_session():
    conn = get_db()
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# =====================================================
# SCHEMA
# =====================================================
SCHEMA = """
CREATE TABLE IF NOT EXISTS teams (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    code TEXT UNIQUE NOT NULL,
    name TEXT NOT NULL,
    leader_id INTEGER,
    is_lab INTEGER DEFAULT 0,
    sector TEXT,
    parent_team_id INTEGER REFERENCES teams(id),
    display_order INTEGER DEFAULT 0
);

CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    login_id TEXT UNIQUE NOT NULL,
    password TEXT NOT NULL,
    email TEXT,
    team_id INTEGER REFERENCES teams(id),
    rank TEXT,
    role TEXT DEFAULT 'member',
    is_active INTEGER DEFAULT 1,
    lang TEXT DEFAULT 'ko',
    created_at TEXT DEFAULT (datetime('now','localtime'))
);

CREATE TABLE IF NOT EXISTS customers (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT UNIQUE NOT NULL,
    tier TEXT DEFAULT '일반',
    note TEXT
);

CREATE TABLE IF NOT EXISTS projects (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    code TEXT,
    name TEXT NOT NULL,
    customer_id INTEGER REFERENCES customers(id),
    type TEXT,
    status TEXT DEFAULT '진행중',
    pm_id INTEGER REFERENCES users(id),
    start_date TEXT,
    end_date TEXT,
    mgmt_code TEXT UNIQUE,
    equip_type TEXT,
    year_month TEXT,
    model_name TEXT,
    server_path TEXT,
    lead_user_id INTEGER REFERENCES users(id),
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_projects_mgmt ON projects(mgmt_code);
CREATE INDEX IF NOT EXISTS idx_projects_equip ON projects(equip_type);

CREATE TABLE IF NOT EXISTS tasks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER REFERENCES users(id),
    work_date TEXT NOT NULL,
    title TEXT NOT NULL,
    category TEXT,
    project_id INTEGER REFERENCES projects(id),
    customer_id INTEGER REFERENCES customers(id),
    status TEXT DEFAULT '진행중',
    hours REAL DEFAULT 0,
    notes TEXT,
    next_plan TEXT,
    due_date TEXT,
    carry_from_id INTEGER REFERENCES tasks(id),
    created_at TEXT DEFAULT (datetime('now','localtime')),
    updated_at TEXT DEFAULT (datetime('now','localtime'))
);

CREATE TABLE IF NOT EXISTS task_comments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    task_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
    author_id INTEGER NOT NULL REFERENCES users(id),
    body TEXT NOT NULL,
    is_ceo_request INTEGER DEFAULT 0,
    parent_id INTEGER REFERENCES task_comments(id),
    created_at TEXT DEFAULT (datetime('now','localtime'))
);

CREATE TABLE IF NOT EXISTS notifications (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL REFERENCES users(id),
    kind TEXT NOT NULL,
    title TEXT NOT NULL,
    body TEXT,
    link TEXT,
    task_id INTEGER REFERENCES tasks(id) ON DELETE CASCADE,
    comment_id INTEGER REFERENCES task_comments(id) ON DELETE CASCADE,
    is_read INTEGER DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);

CREATE TABLE IF NOT EXISTS team_summaries (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    team_id INTEGER REFERENCES teams(id),
    work_date TEXT NOT NULL,
    author_id INTEGER REFERENCES users(id),
    headline TEXT NOT NULL,
    notes TEXT,
    created_at TEXT DEFAULT (datetime('now','localtime')),
    updated_at TEXT DEFAULT (datetime('now','localtime')),
    UNIQUE(team_id, work_date)
);

CREATE INDEX IF NOT EXISTS idx_tasks_user_date ON tasks(user_id, work_date);
CREATE INDEX IF NOT EXISTS idx_tasks_date ON tasks(work_date);
CREATE INDEX IF NOT EXISTS idx_tasks_status ON tasks(status);
CREATE INDEX IF NOT EXISTS idx_tasks_project ON tasks(project_id);
CREATE INDEX IF NOT EXISTS idx_tasks_customer ON tasks(customer_id);
CREATE INDEX IF NOT EXISTS idx_users_team ON users(team_id);
CREATE INDEX IF NOT EXISTS idx_team_summaries_team_date ON team_summaries(team_id, work_date);
CREATE INDEX IF NOT EXISTS idx_comments_task ON task_comments(task_id);
CREATE INDEX IF NOT EXISTS idx_comments_author ON task_comments(author_id);
CREATE INDEX IF NOT EXISTS idx_notif_user_read ON notifications(user_id, is_read);
CREATE INDEX IF NOT EXISTS idx_notif_created ON notifications(created_at);

-- ============ Phase 1+ : 협업 신경망 ============
CREATE TABLE IF NOT EXISTS activities (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    actor_id INTEGER NOT NULL REFERENCES users(id),
    kind TEXT NOT NULL,            -- task_create / task_update / task_status / comment / reaction / retro
    task_id INTEGER REFERENCES tasks(id) ON DELETE CASCADE,
    project_id INTEGER REFERENCES projects(id) ON DELETE SET NULL,
    team_id INTEGER REFERENCES teams(id) ON DELETE SET NULL,
    title TEXT NOT NULL,
    body TEXT,
    meta TEXT,                     -- JSON
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_act_created ON activities(created_at DESC);
CREATE INDEX IF NOT EXISTS idx_act_team ON activities(team_id);
CREATE INDEX IF NOT EXISTS idx_act_actor ON activities(actor_id);
CREATE INDEX IF NOT EXISTS idx_act_task ON activities(task_id);

CREATE TABLE IF NOT EXISTS task_reactions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    task_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
    user_id INTEGER NOT NULL REFERENCES users(id),
    kind TEXT NOT NULL,            -- ack / question / risk / ok
    created_at TEXT DEFAULT (datetime('now','localtime')),
    UNIQUE(task_id, user_id, kind)
);
CREATE INDEX IF NOT EXISTS idx_react_task ON task_reactions(task_id);

CREATE TABLE IF NOT EXISTS project_retros (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
    author_id INTEGER NOT NULL REFERENCES users(id),
    went_well TEXT,
    went_bad TEXT,
    next_action TEXT,
    risk_note TEXT,
    created_at TEXT DEFAULT (datetime('now','localtime')),
    updated_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_retro_project ON project_retros(project_id);

-- 댓글 멘션
CREATE TABLE IF NOT EXISTS comment_mentions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    comment_id INTEGER NOT NULL REFERENCES task_comments(id) ON DELETE CASCADE,
    user_id INTEGER NOT NULL REFERENCES users(id),
    UNIQUE(comment_id, user_id)
);

CREATE TABLE IF NOT EXISTS task_delegations (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    task_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
    from_user_id INTEGER NOT NULL REFERENCES users(id),
    to_user_id INTEGER NOT NULL REFERENCES users(id),
    message TEXT,
    status TEXT DEFAULT 'pending',
    created_at TEXT DEFAULT (datetime('now','localtime')),
    resolved_at TEXT
);
CREATE INDEX IF NOT EXISTS idx_deleg_task ON task_delegations(task_id);
CREATE INDEX IF NOT EXISTS idx_deleg_to ON task_delegations(to_user_id, status);

-- ============ 물류 모듈 (HAIST WORKS) ============
-- 부품 마스터
CREATE TABLE IF NOT EXISTS parts (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    part_no     TEXT NOT NULL UNIQUE,
    part_name   TEXT NOT NULL,
    spec        TEXT,
    maker       TEXT,
    origin      TEXT,
    unit        TEXT DEFAULT 'EA',
    currency    TEXT DEFAULT 'KRW',
    std_price   REAL DEFAULT 0,
    biz_div     TEXT,                -- T 검사기 / M 자동화 / 공통
    category    TEXT,                -- 완성품/자재/소모품
    note        TEXT,
    is_active   INTEGER DEFAULT 1,
    created_at  TEXT DEFAULT (datetime('now','localtime')),
    updated_at  TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_parts_part_no ON parts(part_no);
CREATE INDEX IF NOT EXISTS idx_parts_biz_div ON parts(biz_div);
CREATE INDEX IF NOT EXISTS idx_parts_category ON parts(category);

-- 공급사 마스터 (발주 대상 거래처)
CREATE TABLE IF NOT EXISTS suppliers (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    name            TEXT NOT NULL,
    code            TEXT UNIQUE,
    contact         TEXT,
    email           TEXT,
    phone           TEXT,
    country         TEXT,
    currency        TEXT DEFAULT 'KRW',
    payment_terms   TEXT,             -- 선금/현금/30일/60일
    note            TEXT,
    is_active       INTEGER DEFAULT 1,
    created_at      TEXT DEFAULT (datetime('now','localtime')),
    updated_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_sup_name ON suppliers(name);

-- 발주 헤더
CREATE TABLE IF NOT EXISTS purchase_orders (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    po_number       TEXT NOT NULL UNIQUE,           -- PO-YYMMDD-NNN
    project_id      INTEGER REFERENCES projects(id),-- 관리코드 연결 (선택)
    supplier_id     INTEGER REFERENCES suppliers(id),
    order_date      TEXT NOT NULL,                   -- YYYY-MM-DD
    expected_date   TEXT,                            -- 예상 입고일
    currency        TEXT DEFAULT 'KRW',
    exchange_rate   REAL DEFAULT 1,
    total_amount    REAL DEFAULT 0,                  -- 라인 합계 (자동)
    status          TEXT DEFAULT '작성중',            -- 작성중/발주완료/부분입고/입고완료/취소
    shipping_terms  TEXT,                            -- EXW/FOB/CIF/DDP/국내
    payment_terms   TEXT,                            -- 선금/현금/30일/60일
    po_type         TEXT DEFAULT '일반',              -- 일반/긴급/정기
    created_by      INTEGER REFERENCES users(id),
    note            TEXT,
    created_at      TEXT DEFAULT (datetime('now','localtime')),
    updated_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_po_number ON purchase_orders(po_number);
CREATE INDEX IF NOT EXISTS idx_po_project ON purchase_orders(project_id);
CREATE INDEX IF NOT EXISTS idx_po_supplier ON purchase_orders(supplier_id);
CREATE INDEX IF NOT EXISTS idx_po_status ON purchase_orders(status);

-- 발주 라인
CREATE TABLE IF NOT EXISTS po_items (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    po_id               INTEGER NOT NULL REFERENCES purchase_orders(id) ON DELETE CASCADE,
    line_no             INTEGER DEFAULT 0,
    part_id             INTEGER REFERENCES parts(id),
    part_no_snapshot    TEXT,       -- 발주 시점 부품번호
    part_name_snapshot  TEXT,
    spec_snapshot       TEXT,
    unit                TEXT DEFAULT 'EA',
    quantity            REAL NOT NULL DEFAULT 0,
    unit_price          REAL NOT NULL DEFAULT 0,
    amount              REAL DEFAULT 0,    -- quantity * unit_price
    received_qty        REAL DEFAULT 0,    -- 누적 입고 수량 (4단계)
    delivery_date       TEXT,
    note                TEXT
);
CREATE INDEX IF NOT EXISTS idx_poitem_po ON po_items(po_id);
CREATE INDEX IF NOT EXISTS idx_poitem_part ON po_items(part_id);

-- ============ 게시판 (HAIST WORKS) ============
-- 게시판 마스터: 전사 / 부서별
CREATE TABLE IF NOT EXISTS boards (
    id        INTEGER PRIMARY KEY AUTOINCREMENT,
    name      TEXT NOT NULL,          -- '전사 게시판' / '검사기팀 게시판' 등
    type      TEXT NOT NULL,          -- 'company' 전사 / 'team' 부서
    team_id   INTEGER REFERENCES teams(id),  -- 부서 게시판일 때
    created_at TEXT DEFAULT (datetime('now','localtime')),
    UNIQUE(type, team_id)
);

-- 게시글
CREATE TABLE IF NOT EXISTS board_posts (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    board_id        INTEGER NOT NULL REFERENCES boards(id),
    author_id       INTEGER NOT NULL REFERENCES users(id),
    title           TEXT NOT NULL,
    body            TEXT,
    category        TEXT DEFAULT '일반',       -- 공지/일반/자료/질문
    is_pinned       INTEGER DEFAULT 0,         -- 고정 (관리자/팀장만)
    view_count      INTEGER DEFAULT 0,
    approval_status TEXT DEFAULT 'approved',    -- approved/pending/rejected
    approved_by     INTEGER REFERENCES users(id),
    approved_at     TEXT,
    reject_reason   TEXT,
    created_at      TEXT DEFAULT (datetime('now','localtime')),
    updated_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_bp_board ON board_posts(board_id);
CREATE INDEX IF NOT EXISTS idx_bp_author ON board_posts(author_id);
CREATE INDEX IF NOT EXISTS idx_bp_approval ON board_posts(approval_status);

-- 댓글
CREATE TABLE IF NOT EXISTS board_comments (
    id        INTEGER PRIMARY KEY AUTOINCREMENT,
    post_id   INTEGER NOT NULL REFERENCES board_posts(id) ON DELETE CASCADE,
    author_id INTEGER NOT NULL REFERENCES users(id),
    body      TEXT NOT NULL,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_bc_post ON board_comments(post_id);

-- ============ 변경 Inform 시스템 (HAIST WORKS) ============
-- 변경 사건 본체
CREATE TABLE IF NOT EXISTS changes (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    change_no       TEXT NOT NULL UNIQUE,         -- CHG-YYMMDD-NNN
    change_type     TEXT NOT NULL,                 -- 기구설계/전장설계/소프트웨어/BOM/도면/Concept/사양
    biz_div         TEXT,                          -- T 검사기 / M 자동화 (영향 부서 자동 판별용)
    target_kind     TEXT,                          -- project / part / document
    target_id       INTEGER,
    target_label    TEXT,                          -- 사람이 읽는 라벨 (예: "001T2604 검사기")
    project_id      INTEGER REFERENCES projects(id),
    title           TEXT NOT NULL,
    description     TEXT,
    before_value    TEXT,
    after_value     TEXT,
    attached_files  TEXT,                          -- JSON 배열
    urgency         TEXT DEFAULT '일반',            -- 일반/긴급/예약
    author_id       INTEGER NOT NULL REFERENCES users(id),
    status          TEXT DEFAULT '공지중',          -- 작성중/공지중/확인완료/취소
    notified_at     TEXT,
    completed_at    TEXT,
    created_at      TEXT DEFAULT (datetime('now','localtime')),
    updated_at      TEXT DEFAULT (datetime('now','localtime'))
);
CREATE INDEX IF NOT EXISTS idx_chg_no ON changes(change_no);
CREATE INDEX IF NOT EXISTS idx_chg_project ON changes(project_id);
CREATE INDEX IF NOT EXISTS idx_chg_status ON changes(status);
CREATE INDEX IF NOT EXISTS idx_chg_author ON changes(author_id);
CREATE INDEX IF NOT EXISTS idx_chg_created ON changes(created_at DESC);

-- 영향 부서/사용자
CREATE TABLE IF NOT EXISTS change_impacts (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    change_id       INTEGER NOT NULL REFERENCES changes(id) ON DELETE CASCADE,
    impact_kind     TEXT NOT NULL,                 -- team / user
    impact_team_id  INTEGER REFERENCES teams(id),
    impact_user_id  INTEGER REFERENCES users(id),
    auto_detected   INTEGER DEFAULT 1,              -- 1=자동, 0=수동 추가
    impact_reason   TEXT
);
CREATE INDEX IF NOT EXISTS idx_cimp_change ON change_impacts(change_id);
CREATE INDEX IF NOT EXISTS idx_cimp_team ON change_impacts(impact_team_id);

-- 확인 추적
CREATE TABLE IF NOT EXISTS change_reads (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    change_id       INTEGER NOT NULL REFERENCES changes(id) ON DELETE CASCADE,
    user_id         INTEGER NOT NULL REFERENCES users(id),
    read_at         TEXT,
    ack_at          TEXT,
    ack_note        TEXT,
    UNIQUE(change_id, user_id)
);
CREATE INDEX IF NOT EXISTS idx_cread_change ON change_reads(change_id);
CREATE INDEX IF NOT EXISTS idx_cread_user ON change_reads(user_id);
"""

# =====================================================
# ORG CHART SEED DATA (2026.03.11 기준)
# =====================================================
TEAMS = [
    # (code, name, is_lab, sector, order)
    ("01", "기술영업팀", 0, "공통",  1),
    ("02", "검사기팀",   1, "검사기", 2),
    ("03", "품질팀",     0, "검사기", 3),
    ("04", "설계팀",     1, "공통",  4),
    ("05", "소프트웨어팀", 1, "자동화", 5),
    ("06", "전장설계팀", 1, "자동화", 6),
    ("07", "제조기술1팀", 0, "검사기", 7),
    ("08", "제조기술2팀", 0, "자동화", 8),
    ("09", "가공팀",     0, "공통",  9),
    ("10", "구매팀",     0, "공통", 10),
    ("11", "관리팀",     0, "공통", 11),
    ("12", "베트남법인", 0, "공통", 12),
    ("13", "개발혁신팀", 1, "검사기", 13),
    ("14", "라이프밸류팀", 1, "공통", 14),
]

# 사용자 시드: (이름, 팀코드, 직급, 역할)
# role: ceo / executive / leader / member
USERS = [
    # 경영진
    ("김동후", None, "대표이사", "ceo"),
    ("김정락", None, "대표이사", "ceo"),
    ("이한빈", "02", "상무",     "executive"),  # 연구소장 + 검사기팀장
    ("윤경호", "04", "상무",     "executive"),  # 설계팀장
    ("최홍광", None, "전무",     "executive"),  # 인사총괄
    ("최보현", "13", "상무",     "executive"),  # 개발혁신팀장

    # 01 기술영업팀
    ("이해림", "01", "이사",     "leader"),
    ("이현",   "01", "매니저", "member"),
    ("오경환", "01", "프로",   "member"),
    ("배승진", "01", "프로",   "member"),
    ("안지연", "01", "프로",   "member"),
    ("이새롬", "01", "프로",   "member"),

    # 02 검사기팀 (팀장 이한빈 상무 - 이미 등록)
    ("이치권", "02", "이사",   "member"),
    ("이성진", "02", "매니저", "member"),
    ("길희용", "02", "매니저", "member"),
    ("윤광훈", "02", "프로",   "member"),
    ("김태형", "02", "프로",   "member"),
    ("이서준", "02", "사원",   "member"),
    ("김지훈", "02", "사원",   "member"),
    ("지경숙", "02", "반장",   "member"),

    # 03 품질팀
    ("김정록", "03", "매니저", "leader"),
    ("정형진", "03", "사원",   "member"),

    # 04 설계팀 (팀장 윤경호 상무 - 이미 등록)
    ("이영준",  "04", "매니저", "member"),  # 검사기
    ("김범수",  "04", "프로",   "member"),
    ("안호재",  "04", "사원",   "member"),
    ("정민규",  "04", "매니저", "member"),  # 자동화
    ("이상천",  "04", "매니저", "member"),
    ("한재운",  "04", "매니저", "member"),
    ("신광용",  "04", "매니저", "member"),
    ("김선주",  "04", "매니저", "member"),
    ("김동현",  "04", "프로",   "member"),
    ("최현규",  "04", "프로",   "member"),
    ("김원",    "04", "사원",   "member"),

    # 05 소프트웨어팀
    ("이한중", "05", "매니저", "leader"),
    ("최창호", "05", "매니저", "member"),
    ("황정석", "05", "매니저", "member"),
    ("현종필", "05", "매니저", "member"),
    ("이정우", "05", "매니저", "member"),
    ("주진호", "05", "매니저", "member"),
    ("김동욱", "05", "매니저", "member"),
    ("차상권", "05", "매니저", "member"),
    ("이영준2","05", "매니저", "member"),
    ("이충희", "05", "프로",   "member"),
    ("박주창", "05", "프로",   "member"),
    ("김기운", "05", "프로",   "member"),

    # 06 전장설계팀
    ("김형렬", "06", "매니저", "leader"),
    ("박재석", "06", "사원",   "member"),

    # 07 제조기술1팀
    ("노충일", "07", "매니저", "leader"),
    ("마준영", "07", "사원",   "member"),
    ("이태우", "07", "사원",   "member"),
    ("금진호", "07", "사원",   "member"),

    # 08 제조기술2팀
    ("임택훈", "08", "매니저", "leader"),
    ("서재희", "08", "매니저", "member"),
    ("연태흠", "08", "매니저", "member"),
    ("방성기", "08", "프로",   "member"),
    ("김한성", "08", "프로",   "member"),
    ("박지현", "08", "사원",   "member"),
    ("나영훈", "08", "사원",   "member"),
    ("강대성", "08", "사원",   "member"),

    # 09 가공팀
    ("윤영조", "09", "매니저", "leader"),
    ("이수빈", "09", "매니저", "member"),
    ("이청명", "09", "사원",   "member"),

    # 10 구매팀
    ("정성진", "10", "매니저", "leader"),
    ("허동준", "10", "매니저", "member"),
    ("오용균", "10", "프로",   "member"),
    ("김선미", "10", "프로",   "member"),
    ("이홍규", "10", "프로",   "member"),
    ("박성준", "10", "사원",   "member"),
    ("란",     "10", "사원",   "member"),

    # 11 관리팀
    ("박지은", "11", "매니저", "leader"),
    ("엄혜린", "11", "매니저", "member"),
    ("엄주영", "11", "프로",   "member"),
    ("최혜연", "11", "프로",   "member"),

    # 12 베트남법인
    ("이용식", "12", "법인장", "leader"),
    ("땀",     "12", "부장",   "member"),
    ("박지만", "12", "부장",   "member"),
    ("탕",     "12", "차장",   "member"),
    ("쑤아잉", "12", "차장",   "member"),

    # 13 개발혁신팀 (팀장 최보현 상무 - 이미 등록)
    ("박승환", "13", "매니저", "member"),
    ("김수현", "13", "사원",   "member"),

    # 14 라이프밸류팀
    ("나재겸", "14", "매니저", "leader"),
    ("김기선", "14", "매니저", "member"),
    ("박성수", "14", "사원",   "member"),
]

# 주요 고객사
CUSTOMERS = [
    ("삼성전자",   "주요", "1차 협력사"),
    ("삼성전기",   "주요", "1차 협력사"),
    ("드림텍",     "주요", ""),
    ("한국성전",   "주요", ""),
    ("기타전장",   "일반", "전장사업 관련"),
    ("내부",       "일반", "사내 프로젝트"),
]


# =====================================================
# LOGIN ID 생성 규칙
# =====================================================
def make_login_id(name: str, dup_map: dict) -> str:
    base = name.lower().replace(" ", "")
    # 한글은 그대로 초성처리 대신 name 그대로 사용하되, 영문/숫자가 아니면 중복 방지만
    if base in dup_map:
        dup_map[base] += 1
        return f"{base}{dup_map[base]}"
    dup_map[base] = 1
    return base


# =====================================================
# INIT + SEED
# =====================================================
def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    with db_session() as c:
        c.executescript(SCHEMA)
        # 마이그레이션: 기존 DB에 next_plan 컬럼이 없으면 추가
        cols = [r[1] for r in c.execute("PRAGMA table_info(tasks)").fetchall()]
        if "next_plan" not in cols:
            try:
                c.execute("ALTER TABLE tasks ADD COLUMN next_plan TEXT")
            except Exception:
                pass
        # 마이그레이션: users에 lang 컬럼 추가
        ucols = [r[1] for r in c.execute("PRAGMA table_info(users)").fetchall()]
        if "lang" not in ucols:
            try:
                c.execute("ALTER TABLE users ADD COLUMN lang TEXT DEFAULT 'ko'")
            except Exception:
                pass
        # 마이그레이션: 물류 모듈 접근 권한 (HAIST WORKS)
        if "can_use_logistics" not in ucols:
            try:
                c.execute("ALTER TABLE users ADD COLUMN can_use_logistics INTEGER DEFAULT 0")
            except Exception:
                pass
        # 게시판 시드: 전사 게시판 + 팀별 게시판 자동 생성
        board_cnt = c.execute("SELECT COUNT(*) FROM boards").fetchone()[0]
        if board_cnt == 0:
            c.execute("INSERT OR IGNORE INTO boards (name, type, team_id) VALUES ('전사 게시판', 'company', NULL)")
            teams_rows = c.execute("SELECT id, name FROM teams ORDER BY display_order").fetchall()
            for t in teams_rows:
                c.execute("INSERT OR IGNORE INTO boards (name, type, team_id) VALUES (?, 'team', ?)",
                          (f"{t['name']} 게시판", t["id"]))
        # 마이그레이션: projects에 물류 모듈 컬럼 추가 (HAIST WORKS)
        pcols = [r[1] for r in c.execute("PRAGMA table_info(projects)").fetchall()]
        _logi_adds = [
            ("biz_div",       "TEXT"),                  # T 검사기 / M 자동화
            ("stage",         "TEXT DEFAULT '제안작성'"),
            ("po_type",       "TEXT DEFAULT '신규'"),
            ("customer_name", "TEXT"),                   # 텍스트 고객사 (customer_id FK와 병행)
            ("customer_po",   "TEXT"),
            ("currency",      "TEXT DEFAULT 'KRW'"),
            ("order_amount",  "REAL DEFAULT 0"),
            ("order_date",    "TEXT"),
            ("due_date",      "TEXT"),
            ("pm_name",       "TEXT"),                   # 텍스트 PM (pm_id FK와 병행)
            ("sales_name",    "TEXT"),
            ("logi_note",     "TEXT"),                   # 물류 비고
            ("updated_at",    "TEXT"),
        ]
        for col, decl in _logi_adds:
            if col not in pcols:
                try:
                    c.execute(f"ALTER TABLE projects ADD COLUMN {col} {decl}")
                except Exception:
                    pass


def seed_all():
    """최초 1회: 조직도/사용자/고객사 시드"""
    with db_session() as c:
        # 이미 팀 있으면 skip
        cnt = c.execute("SELECT COUNT(*) FROM teams").fetchone()[0]
        if cnt > 0:
            return

        # 1) Teams
        for code, name, is_lab, sector, order in TEAMS:
            c.execute(
                "INSERT INTO teams(code, name, is_lab, sector, display_order) VALUES(?,?,?,?,?)",
                (code, name, is_lab, sector, order),
            )

        # 2) Admin
        c.execute(
            "INSERT INTO users(name, login_id, password, rank, role) VALUES(?,?,?,?,?)",
            ("시스템관리자", "admin", hash_pw("admin1234"), "관리자", "admin"),
        )

        # 3) Users
        dup: dict = {}
        # 대표님 두 분은 login_id를 명시적으로 부여
        special = {"김동후": "kdh", "김정락": "kjr"}

        for name, team_code, rank, role in USERS:
            if name in special:
                lid = special[name]
            else:
                lid = make_login_id(name, dup)
            team_id = None
            if team_code:
                t = c.execute("SELECT id FROM teams WHERE code=?", (team_code,)).fetchone()
                team_id = t["id"] if t else None
            c.execute(
                "INSERT INTO users(name, login_id, password, team_id, rank, role) VALUES(?,?,?,?,?,?)",
                (name, lid, hash_pw("knk1234"), team_id, rank, role),
            )

        # 4) Team leaders 지정
        # 팀 코드별 "leader" 또는 "executive"인 인원을 leader로 설정
        for code, name, *_ in TEAMS:
            t = c.execute("SELECT id FROM teams WHERE code=?", (code,)).fetchone()
            if not t:
                continue
            # 우선 leader 역할 우선, 없으면 executive 중 해당 팀 배치
            ld = c.execute(
                "SELECT id FROM users WHERE team_id=? AND role='leader' ORDER BY id LIMIT 1",
                (t["id"],),
            ).fetchone()
            if not ld:
                ld = c.execute(
                    "SELECT id FROM users WHERE team_id=? AND role='executive' ORDER BY id LIMIT 1",
                    (t["id"],),
                ).fetchone()
            if ld:
                c.execute("UPDATE teams SET leader_id=? WHERE id=?", (ld["id"], t["id"]))

        # 5) Customers
        for cname, tier, note in CUSTOMERS:
            c.execute(
                "INSERT INTO customers(name, tier, note) VALUES(?,?,?)",
                (cname, tier, note),
            )

        # 6) Sample Projects
        samples = [
            ("P-2026-001", "갤럭시 S27 메인보드 ICT",       "삼성전자",  "검사기"),
            ("P-2026-002", "갤럭시 S27 FCT 라인",           "삼성전자",  "검사기"),
            ("P-2026-003", "전장 제어보드 자동조립 라인",   "삼성전기",  "자동화"),
            ("P-2026-004", "카메라 모듈 검사장비",          "드림텍",    "검사기"),
            ("P-2026-005", "EV 배터리 BMS 검사장비",        "한국성전",  "검사기"),
            ("P-2026-006", "사내 자동화 통합 MES",           "내부",      "자동화"),
        ]
        for code, pname, cust, ptype in samples:
            cu = c.execute("SELECT id FROM customers WHERE name=?", (cust,)).fetchone()
            c.execute(
                "INSERT INTO projects(code, name, customer_id, type, status) VALUES(?,?,?,?,?)",
                (code, pname, cu["id"] if cu else None, ptype, "진행중"),
            )


# =====================================================
# SAMPLE DATA GENERATOR (지난 14일치 현실적 데이터)
# =====================================================
import random
from datetime import date as _date, timedelta as _td

# 팀별 현실적인 업무 템플릿 (KNK 실제 업무 반영)
TASK_TEMPLATES = {
    "01": [  # 기술영업팀
        ("{cu} {pj} 견적 제출", "고객대응", "검사기"),
        ("{cu} 신규 BOM 검토 미팅", "고객대응", "공통"),
        ("{cu} 방문 · 사양 협의", "출장", "공통"),
        ("{cu} 발주 진행 체크", "고객대응", "공통"),
        ("{cu} 대응 이슈 회신", "고객대응", "공통"),
        ("{pj} 사양서 전달", "고객대응", "공통"),
        ("주간 수주 현황 정리", "검토", "공통"),
        ("영업 파이프라인 업데이트", "검토", "공통"),
    ],
    "02": [  # 검사기팀
        ("{pj} 1차 셋업 점검", "설계", "검사기"),
        ("{pj} HW 수정 요청 대응", "설계", "검사기"),
        ("{pj} 시제품 전기 검사", "품질", "검사기"),
        ("{pj} 지그 수정 설계", "설계", "검사기"),
        ("{pj} 양산 이관 준비", "제조", "검사기"),
        ("{pj} 회로도 검토", "설계", "검사기"),
        ("검사기 표준 프로세스 개선", "검토", "검사기"),
    ],
    "03": [  # 품질팀
        ("{cu} FCT 라인 품질 이슈 대응", "품질", "검사기"),
        ("{pj} 불량률 집계 · 리포트", "품질", "공통"),
        ("수입검사 샘플링 기준 재검토", "품질", "공통"),
        ("품질 미팅 (일일)", "회의", "공통"),
        ("출하 검사 기록 정리", "품질", "공통"),
        ("{cu} 클레임 대응", "고객대응", "공통"),
    ],
    "04": [  # 설계팀
        ("{pj} 메커니컬 3D 도면", "설계", "공통"),
        ("{pj} 지그 설계 수정", "설계", "공통"),
        ("{pj} 2D 도면 출도", "설계", "공통"),
        ("{pj} 설계 검증 회의", "회의", "공통"),
        ("{pj} 부품 BOM 확정", "설계", "공통"),
        ("{pj} 기존 도면 리비전", "설계", "공통"),
        ("설계 Lesson-Learned 정리", "검토", "공통"),
    ],
    "05": [  # 소프트웨어팀
        ("{pj} 펌웨어 v2.x 개발", "개발", "공통"),
        ("{pj} 시퀀스 스크립트 작성", "개발", "공통"),
        ("{pj} 비전 알고리즘 튜닝", "개발", "공통"),
        ("{pj} 통신 프로토콜 구현", "개발", "공통"),
        ("{pj} UI 화면 개발", "개발", "공통"),
        ("{pj} 현장 디버깅", "개발", "공통"),
        ("{pj} 리포트 기능 추가", "개발", "공통"),
        ("{pj} 통합 테스트", "개발", "공통"),
    ],
    "06": [  # 전장설계팀
        ("{pj} 전장 회로 설계", "설계", "자동화"),
        ("{pj} 배전반 부품 선정", "설계", "자동화"),
        ("{pj} PLC 로직 작성", "설계", "자동화"),
        ("{pj} 전장 도면 출도", "설계", "자동화"),
        ("전장 표준품 정리", "검토", "자동화"),
    ],
    "07": [  # 제조기술1팀
        ("{pj} 검사기 본체 조립", "제조", "검사기"),
        ("{pj} 지그 가공/조립", "제조", "검사기"),
        ("{pj} 양산 이관 대응", "제조", "검사기"),
        ("{pj} 생산 일정 조율", "제조", "검사기"),
        ("일일 제조 미팅", "회의", "공통"),
    ],
    "08": [  # 제조기술2팀
        ("{pj} 자동화 장비 조립", "제조", "자동화"),
        ("{pj} 벨트 얼라인 셋업", "제조", "자동화"),
        ("{pj} 장비 반입 · 현장 셋업", "출장", "자동화"),
        ("{pj} 양산 안정화 작업", "제조", "자동화"),
        ("{pj} 고객 인수 테스트 대응", "제조", "자동화"),
    ],
    "09": [  # 가공팀
        ("{pj} 기구부 가공", "제조", "공통"),
        ("{pj} 지그 부품 가공", "제조", "공통"),
        ("일일 CNC 가동 점검", "제조", "공통"),
        ("가공 스케줄 재편성", "제조", "공통"),
    ],
    "10": [  # 구매팀
        ("{cu} 부품 발주 진행", "구매", "공통"),
        ("협력사 단가 협상", "구매", "공통"),
        ("{pj} 자재 입고 확인", "구매", "공통"),
        ("신규 협력사 등록 심사", "구매", "공통"),
        ("{pj} 긴급 자재 수배", "구매", "공통"),
        ("구매 원가 분석", "검토", "공통"),
    ],
    "11": [  # 관리팀
        ("급여 데이터 정리", "기타", "공통"),
        ("법인 세무 자료 제출", "기타", "공통"),
        ("4대보험 신고", "기타", "공통"),
        ("신입 입사 절차 처리", "기타", "공통"),
        ("경비 정산 승인", "기타", "공통"),
        ("월간 손익 마감 작업", "검토", "공통"),
    ],
    "12": [  # 베트남법인
        ("하노이 현지 {pj} 조립 진행", "제조", "공통"),
        ("본사 자재 입고 확인", "제조", "공통"),
        ("베트남 현지 품질 이슈 대응", "품질", "공통"),
        ("현지 협력사 미팅", "고객대응", "공통"),
        ("본사 기술지원 요청", "기타", "공통"),
    ],
    "13": [  # 개발혁신팀
        ("신규 검사기술 POC 개발", "개발", "검사기"),
        ("차세대 비전 알고리즘 R&D", "개발", "검사기"),
        ("특허 아이디어 검토", "검토", "검사기"),
        ("연구개발 리뷰 미팅", "회의", "검사기"),
    ],
    "14": [  # 라이프밸류팀
        ("LVU 신제품 시장조사", "검토", "공통"),
        ("LVU 샘플 사용 테스트", "검토", "공통"),
        ("LVU 브랜딩 검토", "검토", "공통"),
    ],
}

TEAM_HEADLINES = {
    "01": ["삼성전자 S27 견적 1차 완료, 2차 금주 중 회신 예정", "드림텍 신규 카메라 프로젝트 착수, 사양 협의 중", "금주 견적 5건 완료, 수주 파이프라인 정상"],
    "02": ["S27 ICT 시제품 전기 검사 마무리 단계, 양산 이관 준비", "FCT 라인 지그 수정 완료, 양산 이관 D-3", "BMS 검사기 1차 회로도 완성, 설계 검증 진행"],
    "03": ["삼성전자 클레임 1건 대응 완료, 재발 방지 보고서 작성 중", "품질 미팅 정례화, 불량률 1.2% → 0.8% 개선", "FCT 라인 불량률 분석 진행 중, 금주 내 결과 보고"],
    "04": ["S27 메커니컬 도면 80% 완료, 금주 출도 목표", "BMS 검사장비 3D 모델링 진행 중", "자동화 라인 지그 설계 수정 완료"],
    "05": ["ICT 펌웨어 v2.3 릴리즈 완료, 현장 적용 중", "BMS SW 통합 테스트 80%, 금주 완료 예정", "비전 알고리즘 정확도 98.7% 달성"],
    "06": ["자동조립 라인 PLC 로직 작성 완료, 시뮬레이션 단계", "전장 회로 검토 회의 완료, 수정사항 반영 중", "신규 전장 표준화 작업 착수"],
    "07": ["S27 검사기 본체 5대 조립 완료, 금주 출하 예정", "지그 가공 스케줄 정상, 납기 이상 없음", "양산 이관 안정화 진행 중"],
    "08": ["자동조립 라인 셋업 진행률 70%, 현장 파견 중", "드림텍 장비 인수 테스트 대응", "벨트 얼라인 재조정 완료"],
    "09": ["CNC 가동률 92%, 납기 지연 없음", "지그 부품 가공 정상 진행", "설비 정기점검 완료"],
    "10": ["삼성전기 부품 긴급 수배 완료", "신규 협력사 2곳 등록 심사 중", "주요 부품 재고 정상"],
    "11": ["월간 급여 마감 진행 중, 4월 10일 완료 예정", "신입 2명 입사 절차 진행", "손익 마감 자료 취합 중"],
    "12": ["하노이 현지 조립 2대 완료, 본사 자재 대기 1건", "현지 품질 이슈 대응 완료", "본사 기술지원 1건 요청"],
    "13": ["차세대 비전 알고리즘 POC 1차 완료", "특허 출원 2건 검토 중", "신규 검사 기법 사내 공유 완료"],
    "14": ["LVU 신규 컨셉 3종 정리 완료, 대표이사 보고 예정", "브랜딩 시안 1차 검토", "샘플 사용자 피드백 수집 중"],
}


def seed_sample_tasks(days_back: int = 14):
    """지난 N일치 현실적인 Task 샘플 시드"""
    random.seed(42)  # 재현 가능
    with db_session() as c:
        # 이미 tasks 있으면 skip
        cnt = c.execute("SELECT COUNT(*) FROM tasks").fetchone()[0]
        if cnt > 5:
            return 0

        # 전 사용자
        users = [dict(r) for r in c.execute(
            """SELECT u.id, u.name, u.rank, u.role, t.code AS team_code, t.id AS team_id
               FROM users u LEFT JOIN teams t ON u.team_id=t.id
               WHERE u.is_active=1 AND u.role!='admin' AND u.team_id IS NOT NULL"""
        ).fetchall()]

        # 프로젝트·고객사 맵
        projects = [dict(r) for r in c.execute("SELECT id, name, customer_id, type FROM projects").fetchall()]
        customers = [dict(r) for r in c.execute("SELECT id, name FROM customers").fetchall()]
        pj_names = {p["id"]: p["name"] for p in projects}
        cu_names = {p["id"]: p.get("customer_id") for p in projects}

        today = _date.today()
        total = 0

        for dback in range(days_back, -1, -1):
            d = today - _td(days=dback)
            wd = d.weekday()  # 0=월
            # 토/일은 카드 거의 없음
            if wd == 5:  # 토
                if random.random() > 0.15:
                    continue
            if wd == 6:  # 일
                continue

            for u in users:
                tc = u["team_code"]
                if tc not in TASK_TEMPLATES:
                    continue
                # 70% 확률로 기록, 관리자급은 덜
                if u["role"] in ("ceo", "executive") and random.random() > 0.55:
                    continue
                if u["role"] == "member" and random.random() > 0.85:
                    continue

                n_tasks = random.choice([2, 3, 3, 3, 4, 4, 5])
                templates = TASK_TEMPLATES[tc]
                for _ in range(n_tasks):
                    title_tpl, cat, sec = random.choice(templates)
                    # 프로젝트 선택 (팀 sector 유사)
                    candidate_projects = [p for p in projects if sec == "공통" or p["type"] == sec]
                    pj = random.choice(candidate_projects) if candidate_projects and random.random() > 0.2 else None
                    cu_id = pj["customer_id"] if pj else (random.choice(customers)["id"] if random.random() > 0.5 else None)
                    cu_name = next((cc["name"] for cc in customers if cc["id"] == cu_id), "") if cu_id else ""
                    pj_name = pj["name"].split()[0] if pj else "사내"
                    title = title_tpl.replace("{cu}", cu_name or "내부").replace("{pj}", pj_name)

                    # 상태: 최근일수록 '완료' 비율 낮음
                    if dback == 0:  # 오늘
                        status = random.choices(["진행중", "완료", "지연", "대기"], weights=[55, 25, 12, 8])[0]
                    elif dback <= 2:
                        status = random.choices(["진행중", "완료", "지연", "대기"], weights=[35, 45, 12, 8])[0]
                    else:
                        status = random.choices(["완료", "진행중", "지연"], weights=[75, 15, 10])[0]

                    hours = round(random.choice([0.5, 1, 1, 1.5, 2, 2, 2.5, 3, 3, 4]), 1)
                    notes = ""
                    if random.random() > 0.8:
                        notes = random.choice([
                            "내부 검토 후 금주 내 완료 예정",
                            "고객사 피드백 대기 중",
                            "협력사 일정 조율 필요",
                            "설계 리비전 반영 완료",
                            "현장 이슈 발생, 대응 중",
                            "다음주 확인 예정",
                        ])

                    c.execute(
                        """INSERT INTO tasks(user_id, work_date, title, category, project_id,
                                              customer_id, status, hours, notes)
                           VALUES(?,?,?,?,?,?,?,?,?)""",
                        (u["id"], d.isoformat(), title, cat,
                         pj["id"] if pj else None, cu_id, status, hours, notes),
                    )
                    total += 1

        # 팀장 한 줄 요약 (최근 5일)
        leaders = [dict(r) for r in c.execute(
            """SELECT t.id AS team_id, t.code, u.id AS leader_id
               FROM teams t JOIN users u ON t.leader_id=u.id"""
        ).fetchall()]
        for dback in range(5):
            d = (today - _td(days=dback))
            if d.weekday() >= 5:
                continue
            for l in leaders:
                code = l["code"]
                hl_list = TEAM_HEADLINES.get(code, ["금일 업무 정상 진행"])
                hl = random.choice(hl_list)
                try:
                    c.execute(
                        """INSERT OR IGNORE INTO team_summaries(team_id, work_date, author_id, headline)
                           VALUES(?,?,?,?)""",
                        (l["team_id"], d.isoformat(), l["leader_id"], hl),
                    )
                except Exception:
                    pass

        return total


# =====================================================
# MGMT CODE IMPORT (관리코드발행목록.xls)
# =====================================================
def _norm_customer(name: str) -> str:
    """고객사명 정규화 — 공백/따옴표 제거"""
    if not name:
        return ""
    s = str(name).strip()
    return s


def _clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in ("nan", "none", "null"):
        return ""
    return s


def parse_mgmt_xls(file_path: str):
    """
    관리코드발행목록.xls 파싱
    Returns: list of dict { mgmt_code, equip_type, year_month, model_name,
                            name, customer_name, author_name, server_path }
    """
    import pandas as pd
    rows = []
    sheets = pd.read_excel(file_path, sheet_name=None, header=None)
    for sheet_name, df in sheets.items():
        if sheet_name not in ("M", "T"):
            continue
        equip_type = "자동화" if sheet_name == "M" else "검사기"
        # 2행(idx=1)이 헤더, 3행(idx=2)부터 데이터
        for i in range(2, len(df)):
            row = df.iloc[i]
            seq = _clean(row[0])
            div = _clean(row[1])
            ym = _clean(row[2])
            if not seq or not div or not ym:
                continue
            if div not in ("M", "T"):
                continue
            # ym: '2409.0' -> '2409'
            if ym.endswith(".0"):
                ym = ym[:-2]
            if len(ym) != 4 or not ym.isdigit():
                continue
            # 관리코드 형식: {seq}{div}{ym}  예) 012M2409
            try:
                seq_int = int(float(seq))
                seq_str = f"{seq_int:03d}"
            except Exception:
                continue
            mgmt_code = f"{seq_str}{div}{ym}"
            model = _clean(row[3])
            if sheet_name == "M":
                machine = _clean(row[4])  # 명판 설비명
                customer = _norm_customer(_clean(row[6]))
                author = _clean(row[7])
                server_path = _clean(row[8])
            else:  # T
                machine = _clean(row[4])  # 품명
                customer = _norm_customer(_clean(row[5]))
                author = _clean(row[6])
                server_path = ""
            if not machine and not model:
                continue
            # 프로젝트명 = 품명/설비명 우선, 없으면 모델명
            pj_name = machine or model
            year_month = f"20{ym[:2]}-{ym[2:]}"  # 2409 -> 2024-09
            rows.append({
                "mgmt_code": mgmt_code,
                "equip_type": equip_type,
                "div": div,
                "year_month": year_month,
                "model_name": model,
                "name": pj_name,
                "customer_name": customer,
                "author_name": author,
                "server_path": server_path,
            })
    return rows


def import_mgmt_rows(rows: list) -> dict:
    """
    파싱된 rows를 DB에 upsert. 고객사·담당자 매칭 포함.
    Returns: {inserted, updated, customers_created, skipped, errors}
    """
    result = {"inserted": 0, "updated": 0, "customers_created": 0, "skipped": 0, "errors": 0}
    with db_session() as c:
        for r in rows:
            try:
                # 고객사 upsert
                cu_id = None
                if r["customer_name"]:
                    cu = c.execute(
                        "SELECT id FROM customers WHERE name=?", (r["customer_name"],)
                    ).fetchone()
                    if cu:
                        cu_id = cu["id"]
                    else:
                        cur = c.execute(
                            "INSERT INTO customers(name, tier, note) VALUES(?,?,?)",
                            (r["customer_name"], "일반", "자동임포트"),
                        )
                        cu_id = cur.lastrowid
                        result["customers_created"] += 1
                # 담당자 매칭
                lead_id = None
                if r["author_name"]:
                    ur = c.execute(
                        "SELECT id FROM users WHERE name=? AND is_active=1 LIMIT 1",
                        (r["author_name"],),
                    ).fetchone()
                    if ur:
                        lead_id = ur["id"]
                # 프로젝트 upsert (mgmt_code 기준)
                ex = c.execute(
                    "SELECT id FROM projects WHERE mgmt_code=?", (r["mgmt_code"],)
                ).fetchone()
                start_date = f"{r['year_month']}-01"
                if ex:
                    c.execute(
                        """UPDATE projects SET
                               name=?, customer_id=?, type=?, equip_type=?,
                               year_month=?, model_name=?, server_path=?,
                               lead_user_id=?, start_date=?
                           WHERE id=?""",
                        (r["name"], cu_id, r["equip_type"], r["equip_type"],
                         r["year_month"], r["model_name"], r["server_path"],
                         lead_id, start_date, ex["id"]),
                    )
                    result["updated"] += 1
                else:
                    c.execute(
                        """INSERT INTO projects
                           (code, name, customer_id, type, status, start_date,
                            mgmt_code, equip_type, year_month, model_name,
                            server_path, lead_user_id)
                           VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (r["mgmt_code"], r["name"], cu_id, r["equip_type"],
                         "진행중", start_date,
                         r["mgmt_code"], r["equip_type"], r["year_month"],
                         r["model_name"], r["server_path"], lead_id),
                    )
                    result["inserted"] += 1
            except Exception as e:
                result["errors"] += 1
    return result


# =====================================================
# INITIAL PASSWORD DEPLOYMENT
# =====================================================
def regenerate_user_passwords(exclude_logins=("admin",)):
    """
    모든 활성 사용자의 비밀번호를 새로 생성하여 DB에 반영하고 목록을 반환.
    Returns: list of dict {team_name, team_code, name, rank, login_id, password, role}
    """
    import random, string
    rows = []
    with db_session() as c:
        users = c.execute(
            """SELECT u.id, u.name, u.login_id, u.rank, u.role,
                      t.name AS team_name, t.code AS team_code, t.display_order
               FROM users u LEFT JOIN teams t ON u.team_id=t.id
               WHERE u.is_active=1 ORDER BY t.display_order, u.id"""
        ).fetchall()
        for u in users:
            if u["login_id"] in exclude_logins:
                continue
            # knk + 4자리 숫자 (외우기 쉽게)
            pw = "knk" + "".join(random.choices(string.digits, k=4))
            c.execute("UPDATE users SET password=? WHERE id=?",
                      (hash_pw(pw), u["id"]))
            rows.append({
                "team_name": u["team_name"] or "(미배정)",
                "team_code": u["team_code"] or "",
                "name": u["name"],
                "rank": u["rank"] or "",
                "login_id": u["login_id"],
                "password": pw,
                "role": u["role"],
            })
    return rows


def build_password_xlsx(rows, out_path):
    """
    비밀번호 배포용 Excel 생성 (KNK CI 적용)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "초기비밀번호_배포용"

    R = "A5282C"  # KNK RED
    RD = "8B1E22"
    LG = "F5F5F5"
    W = "FFFFFF"

    thin = Side(border_style="thin", color="E0E0E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="맑은 고딕", bold=True, color=W, size=11)
    title_font = Font(name="맑은 고딕", bold=True, color=W, size=14)
    cell_font = Font(name="맑은 고딕", size=10)
    pw_font = Font(name="Consolas", bold=True, size=11, color=RD)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", indent=1)
    header_fill = PatternFill("solid", start_color=R, end_color=R)
    title_fill = PatternFill("solid", start_color=RD, end_color=RD)
    alt_fill = PatternFill("solid", start_color=LG, end_color=LG)

    # 타이틀
    ws.merge_cells("A1:G2")
    ws["A1"] = "㈜케이엔케이 일일업무일지 · 초기 비밀번호 배포용"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = title_fill
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22

    # 안내
    ws.merge_cells("A3:G3")
    ws["A3"] = "첫 로그인 후 반드시 [비밀번호 변경] 메뉴에서 개인 비밀번호로 변경해 주세요. · HAIST Innovation"
    ws["A3"].font = Font(name="맑은 고딕", italic=True, size=9, color="4A4A4A")
    ws["A3"].alignment = center
    ws.row_dimensions[3].height = 18

    # 헤더
    headers = ["No", "팀", "이름", "직급", "로그인ID", "초기 비밀번호", "권한"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[4].height = 24

    # 데이터
    for idx, r in enumerate(rows, 1):
        row = 4 + idx
        vals = [idx, r["team_name"], r["name"], r["rank"],
                r["login_id"], r["password"], r["role"]]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = pw_font if ci == 6 else cell_font
            cell.alignment = center if ci in (1, 4, 5, 6, 7) else left
            cell.border = border
            if idx % 2 == 0:
                cell.fill = alt_fill

    widths = [5, 22, 12, 10, 14, 14, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    ws.freeze_panes = "A5"
    wb.save(out_path)
    return out_path

# =====================================================
# COMMENTS & NOTIFICATIONS
# =====================================================
def _parse_mentions(body, c):
    """본문에서 @이름 패턴을 찾아 user id 매핑"""
    import re
    names = set(re.findall(r"@([가-힣A-Za-z0-9_]{2,20})", body or ""))
    if not names:
        return []
    placeholders = ",".join("?" * len(names))
    rows = c.execute(
        f"SELECT id, name FROM users WHERE name IN ({placeholders}) AND is_active=1",
        tuple(names),
    ).fetchall()
    return [(r["id"], r["name"]) for r in rows]


def add_comment(task_id, author_id, body, parent_id=None):
    """댓글 추가 + @멘션 파싱 + 알림 + 활동 로그"""
    with db_session() as c:
        author = c.execute("SELECT role, name, rank, team_id FROM users WHERE id=?", (author_id,)).fetchone()
        is_ceo = 1 if author and author["role"] == "ceo" else 0
        cur = c.execute(
            """INSERT INTO task_comments(task_id, author_id, body, is_ceo_request, parent_id)
               VALUES(?,?,?,?,?)""",
            (task_id, author_id, body.strip(), is_ceo, parent_id),
        )
        comment_id = cur.lastrowid

        task = c.execute(
            """SELECT t.user_id, t.title, t.project_id, u.team_id
               FROM tasks t LEFT JOIN users u ON t.user_id = u.id
               WHERE t.id=?""", (task_id,)
        ).fetchone()
        if not task:
            return comment_id

        recipients = set()
        if task["user_id"] != author_id:
            recipients.add(task["user_id"])
        if parent_id:
            p = c.execute("SELECT author_id FROM task_comments WHERE id=?", (parent_id,)).fetchone()
            if p and p["author_id"] != author_id:
                recipients.add(p["author_id"])

        # @멘션 처리
        mentions = _parse_mentions(body, c)
        mention_uids = set()
        for mid, mname in mentions:
            c.execute("INSERT OR IGNORE INTO comment_mentions(comment_id, user_id) VALUES(?,?)",
                      (comment_id, mid))
            if mid != author_id:
                recipients.add(mid)
                mention_uids.add(mid)

        prefix = "[대표 요청] " if is_ceo else ""
        a_name = author["name"]
        a_rank = author["rank"] or ""
        title = prefix + a_name + " " + a_rank + " 님이 댓글을 남겼습니다"
        snippet = body.strip()[:120]
        link = f"/daily?focus={task_id}#task-{task_id}"
        for uid in recipients:
            kind = "mention" if uid in mention_uids else "comment"
            ttl = ("[멘션] " if kind == "mention" else "") + title
            c.execute(
                """INSERT INTO notifications(user_id, kind, title, body, link, task_id, comment_id)
                   VALUES(?,?,?,?,?,?,?)""",
                (uid, kind, ttl, snippet, link, task_id, comment_id),
            )

        # 활동 로그
        log_activity(c, author_id, "comment",
                     title=f"{a_name} → {task['title'][:50]}",
                     body=snippet,
                     task_id=task_id,
                     project_id=task["project_id"],
                     team_id=task["team_id"])
        return comment_id


# =====================================================
# 활동 로그 / 반응 / 회고
# =====================================================
def log_activity(c, actor_id, kind, title, body=None, task_id=None, project_id=None, team_id=None, meta=None):
    """db_session 이미 열려 있는 상태에서 호출"""
    import json as _json
    c.execute(
        """INSERT INTO activities(actor_id, kind, task_id, project_id, team_id, title, body, meta)
           VALUES(?,?,?,?,?,?,?,?)""",
        (actor_id, kind, task_id, project_id, team_id, title[:200], body, _json.dumps(meta or {}))
    )


def log_activity_standalone(actor_id, kind, title, body=None, task_id=None, project_id=None, team_id=None, meta=None):
    with db_session() as c:
        log_activity(c, actor_id, kind, title, body, task_id, project_id, team_id, meta)


def get_activities(limit=80, team_id=None, actor_id=None, since_id=0):
    with db_session() as c:
        sql = """SELECT a.*, u.name AS actor_name, u.rank AS actor_rank,
                        t.title AS task_title, p.name AS project_name, tm.name AS team_name
                 FROM activities a
                 LEFT JOIN users u ON a.actor_id=u.id
                 LEFT JOIN tasks t ON a.task_id=t.id
                 LEFT JOIN projects p ON a.project_id=p.id
                 LEFT JOIN teams tm ON a.team_id=tm.id
                 WHERE a.id > ?"""
        params = [since_id]
        if team_id:
            sql += " AND a.team_id=?"
            params.append(team_id)
        if actor_id:
            sql += " AND a.actor_id=?"
            params.append(actor_id)
        sql += " ORDER BY a.id DESC LIMIT ?"
        params.append(limit)
        return [dict(r) for r in c.execute(sql, params).fetchall()]


def add_reaction(task_id, user_id, kind):
    """1-click 반응 토글. ack/question/risk/ok"""
    if kind not in ("ack", "question", "risk", "ok"):
        return False
    with db_session() as c:
        existing = c.execute(
            "SELECT id FROM task_reactions WHERE task_id=? AND user_id=? AND kind=?",
            (task_id, user_id, kind)
        ).fetchone()
        if existing:
            c.execute("DELETE FROM task_reactions WHERE id=?", (existing["id"],))
            return "removed"
        c.execute(
            "INSERT INTO task_reactions(task_id, user_id, kind) VALUES(?,?,?)",
            (task_id, user_id, kind)
        )
        # 작성자에게 알림: 대표/임원/팀장 반응은 모든 종류, 직원은 risk/question만
        t = c.execute("SELECT user_id, title FROM tasks WHERE id=?", (task_id,)).fetchone()
        actor = c.execute("SELECT name, rank, role FROM users WHERE id=?", (user_id,)).fetchone()
        if t and t["user_id"] != user_id and actor:
            actor_role = actor["role"] or ""
            is_leader_plus = actor_role in ("ceo", "admin", "executive", "leader")
            if is_leader_plus or kind in ("risk", "question"):
                lab = {"ack":"👍 확인", "question":"❓ 질문", "risk":"⚠️ 리스크", "ok":"✅ OK"}[kind]
                c.execute(
                    """INSERT INTO notifications(user_id, kind, title, body, link, task_id)
                       VALUES(?,?,?,?,?,?)""",
                    (t["user_id"], "reaction",
                     f"{actor['name']} {actor['rank'] or ''} 님이 {lab} 반응",
                     (t["title"] or "")[:120],
                     f"/daily?focus={task_id}#task-{task_id}",
                     task_id)
                )
        return "added"


def get_reactions(task_id):
    with db_session() as c:
        rows = c.execute(
            """SELECT r.kind, u.id AS user_id, u.name, u.rank, u.role
               FROM task_reactions r JOIN users u ON r.user_id=u.id
               WHERE r.task_id=? ORDER BY r.created_at""",
            (task_id,)
        ).fetchall()
        agg = {"ack":[], "question":[], "risk":[], "ok":[]}
        for r in rows:
            agg[r["kind"]].append(dict(r))
        return agg


def get_reactions_bulk(task_ids):
    if not task_ids:
        return {}
    with db_session() as c:
        ph = ",".join("?" * len(task_ids))
        rows = c.execute(
            f"""SELECT task_id, kind, COUNT(*) AS cnt
                FROM task_reactions WHERE task_id IN ({ph})
                GROUP BY task_id, kind""", tuple(task_ids)
        ).fetchall()
        out = {}
        for r in rows:
            out.setdefault(r["task_id"], {})[r["kind"]] = r["cnt"]
        return out


def get_meta_bulk(task_ids):
    """카드 표면에 표시할 메타데이터 (댓글수, 리액션수)를 한 번에 일괄 조회.
       반환: { task_id: {comments:N, ack:N, question:N, risk:N, ok:N, last_comment:'...'} }"""
    if not task_ids:
        return {}
    with db_session() as c:
        ph = ",".join("?" * len(task_ids))
        out = {tid: {"comments": 0, "ack": 0, "question": 0,
                     "risk": 0, "ok": 0, "last_comment": "",
                     "last_comment_at": ""} for tid in task_ids}
        # 댓글 수
        rows = c.execute(
            f"""SELECT task_id, COUNT(*) AS cnt
                FROM task_comments WHERE task_id IN ({ph})
                GROUP BY task_id""", tuple(task_ids)
        ).fetchall()
        for r in rows:
            if r["task_id"] in out:
                out[r["task_id"]]["comments"] = r["cnt"]
        # 마지막 댓글 (미리보기용)
        last_rows = c.execute(
            f"""SELECT tc.task_id, tc.body, tc.created_at, u.name AS user_name
                FROM task_comments tc
                JOIN users u ON tc.author_id=u.id
                WHERE tc.task_id IN ({ph})
                  AND tc.id IN (
                      SELECT MAX(id) FROM task_comments
                      WHERE task_id IN ({ph})
                      GROUP BY task_id
                  )""", tuple(task_ids) + tuple(task_ids)
        ).fetchall()
        for r in last_rows:
            if r["task_id"] in out:
                preview = (r["body"] or "").strip().replace("\n", " ")
                if len(preview) > 40:
                    preview = preview[:40] + "…"
                out[r["task_id"]]["last_comment"] = f"{r['user_name']}: {preview}"
                out[r["task_id"]]["last_comment_at"] = r["created_at"] or ""
        # 리액션 수
        rx_rows = c.execute(
            f"""SELECT task_id, kind, COUNT(*) AS cnt
                FROM task_reactions WHERE task_id IN ({ph})
                GROUP BY task_id, kind""", tuple(task_ids)
        ).fetchall()
        for r in rx_rows:
            if r["task_id"] in out and r["kind"] in out[r["task_id"]]:
                out[r["task_id"]][r["kind"]] = r["cnt"]
        return out


def notify_status_change(task_id, actor_id, old_status, new_status):
    """진행중→지연 등 위험 상태 변경 시 팀장+CEO에게 자동 알림"""
    if new_status != "지연":
        return
    with db_session() as c:
        t = c.execute(
            """SELECT t.title, t.user_id, u.name AS owner_name, u.team_id
               FROM tasks t JOIN users u ON t.user_id=u.id WHERE t.id=?""",
            (task_id,)
        ).fetchone()
        if not t:
            return
        recipients = set()
        # 팀장
        if t["team_id"]:
            leaders = c.execute(
                "SELECT id FROM users WHERE team_id=? AND role IN ('leader','executive') AND is_active=1",
                (t["team_id"],)
            ).fetchall()
            for l in leaders:
                if l["id"] != actor_id:
                    recipients.add(l["id"])
        # CEO
        ceos = c.execute(
            "SELECT id FROM users WHERE role='ceo' AND is_active=1"
        ).fetchall()
        for x in ceos:
            if x["id"] != actor_id:
                recipients.add(x["id"])
        for uid in recipients:
            c.execute(
                """INSERT INTO notifications(user_id, kind, title, body, link, task_id)
                   VALUES(?,?,?,?,?,?)""",
                (uid, "status_delay",
                 f"⚠️ 지연 발생 — {t['owner_name']}",
                 (t["title"] or "")[:120],
                 f"/daily?focus={task_id}#task-{task_id}",
                 task_id)
            )


def get_user_search(q, limit=10):
    """@멘션 자동완성용"""
    with db_session() as c:
        rows = c.execute(
            """SELECT id, name, rank, team_id FROM users
               WHERE is_active=1 AND name LIKE ? ORDER BY name LIMIT ?""",
            (f"{q}%", limit)
        ).fetchall()
        return [dict(r) for r in rows]


def upsert_retro(project_id, author_id, went_well, went_bad, next_action, risk_note):
    with db_session() as c:
        existing = c.execute(
            "SELECT id FROM project_retros WHERE project_id=?", (project_id,)
        ).fetchone()
        if existing:
            c.execute(
                """UPDATE project_retros SET went_well=?, went_bad=?, next_action=?, risk_note=?,
                   author_id=?, updated_at=datetime('now','localtime') WHERE id=?""",
                (went_well, went_bad, next_action, risk_note, author_id, existing["id"])
            )
            return existing["id"]
        cur = c.execute(
            """INSERT INTO project_retros(project_id, author_id, went_well, went_bad, next_action, risk_note)
               VALUES(?,?,?,?,?,?)""",
            (project_id, author_id, went_well, went_bad, next_action, risk_note)
        )
        return cur.lastrowid


def get_retro(project_id):
    with db_session() as c:
        r = c.execute(
            """SELECT pr.*, u.name AS author_name FROM project_retros pr
               LEFT JOIN users u ON pr.author_id=u.id WHERE pr.project_id=?""",
            (project_id,)
        ).fetchone()
        return dict(r) if r else None


def search_all(q, limit=50):
    """카드 + 댓글 + 회고 통합 검색"""
    if not q or len(q.strip()) < 2:
        return {"tasks":[], "comments":[], "retros":[]}
    qq = f"%{q.strip()}%"
    with db_session() as c:
        tasks = c.execute(
            """SELECT t.id, t.title, t.work_date, t.status, t.notes, t.next_plan,
                      u.name AS user_name, p.name AS project_name
               FROM tasks t LEFT JOIN users u ON t.user_id=u.id
               LEFT JOIN projects p ON t.project_id=p.id
               WHERE t.title LIKE ? OR t.notes LIKE ? OR t.next_plan LIKE ?
               ORDER BY t.work_date DESC LIMIT ?""",
            (qq, qq, qq, limit)
        ).fetchall()
        comments = c.execute(
            """SELECT tc.id, tc.body, tc.created_at, tc.task_id,
                      u.name AS author_name, t.title AS task_title
               FROM task_comments tc
               LEFT JOIN users u ON tc.author_id=u.id
               LEFT JOIN tasks t ON tc.task_id=t.id
               WHERE tc.body LIKE ? ORDER BY tc.created_at DESC LIMIT ?""",
            (qq, limit)
        ).fetchall()
        retros = c.execute(
            """SELECT pr.*, p.name AS project_name FROM project_retros pr
               LEFT JOIN projects p ON pr.project_id=p.id
               WHERE pr.went_well LIKE ? OR pr.went_bad LIKE ? OR pr.next_action LIKE ?
               ORDER BY pr.updated_at DESC LIMIT ?""",
            (qq, qq, qq, limit)
        ).fetchall()
        return {
            "tasks":[dict(r) for r in tasks],
            "comments":[dict(r) for r in comments],
            "retros":[dict(r) for r in retros],
        }


def delegate_task(task_id, from_user_id, to_user_id, message=""):
    """업무 위임: from → to, 알림 자동 발송"""
    with db_session() as c:
        c.execute(
            """INSERT INTO task_delegations(task_id, from_user_id, to_user_id, message)
               VALUES(?,?,?,?)""",
            (task_id, from_user_id, to_user_id, message)
        )
        t = c.execute("SELECT title FROM tasks WHERE id=?", (task_id,)).fetchone()
        sender = c.execute("SELECT name, rank FROM users WHERE id=?", (from_user_id,)).fetchone()
        title = (t["title"] if t else "업무")[:80]
        sender_name = f"{sender['name']} {sender['rank'] or ''}".strip() if sender else ""
        # 위임 받는 사람에게 알림
        c.execute(
            """INSERT INTO notifications(user_id, kind, title, body, link, task_id)
               VALUES(?,?,?,?,?,?)""",
            (to_user_id, "delegation",
             f"📌 {sender_name} 님이 업무를 위임했습니다",
             f"[{title}] {message or ''}".strip(),
             f"/daily?focus={task_id}#task-{task_id}",
             task_id)
        )
        # 활동 로그
        log_activity(c, from_user_id, "delegation", f"{sender_name} → 위임", task_id=task_id)
        return True


def get_delegations(task_id):
    with db_session() as c:
        rows = c.execute(
            """SELECT d.*, uf.name AS from_name, uf.rank AS from_rank,
                      ut.name AS to_name, ut.rank AS to_rank
               FROM task_delegations d
               JOIN users uf ON d.from_user_id=uf.id
               JOIN users ut ON d.to_user_id=ut.id
               WHERE d.task_id=? ORDER BY d.created_at DESC""",
            (task_id,)
        ).fetchall()
        return [dict(r) for r in rows]


def resolve_delegation(deleg_id, user_id):
    """위임 완료 처리"""
    with db_session() as c:
        d = c.execute("SELECT * FROM task_delegations WHERE id=?", (deleg_id,)).fetchone()
        if not d or d["to_user_id"] != user_id:
            return False
        c.execute(
            "UPDATE task_delegations SET status='done', resolved_at=datetime('now','localtime') WHERE id=?",
            (deleg_id,)
        )
        # 위임한 사람에게 완료 알림
        t = c.execute("SELECT title FROM tasks WHERE id=?", (d["task_id"],)).fetchone()
        resolver = c.execute("SELECT name, rank FROM users WHERE id=?", (user_id,)).fetchone()
        c.execute(
            """INSERT INTO notifications(user_id, kind, title, body, link, task_id)
               VALUES(?,?,?,?,?,?)""",
            (d["from_user_id"], "delegation",
             f"✅ {resolver['name']} {resolver['rank'] or ''} 님이 위임 업무를 처리했습니다",
             (t["title"] if t else "")[:120],
             f"/daily?focus={d['task_id']}",
             d["task_id"])
        )
        return True


def detect_bottlenecks():
    """병목 자동 탐지: 4가지 규칙"""
    out = []
    with db_session() as c:
        # 1) 지연 3일+
        rows = c.execute(
            """SELECT t.id, t.title, t.work_date, u.name, u.team_id, tm.name AS team_name
               FROM tasks t JOIN users u ON t.user_id=u.id
               LEFT JOIN teams tm ON u.team_id=tm.id
               WHERE t.status='지연' AND date(t.work_date) <= date('now','-3 days')
               ORDER BY t.work_date LIMIT 50"""
        ).fetchall()
        for r in rows:
            out.append({"rule":"지연 3일+", "severity":"high",
                        "title":r["title"], "owner":r["name"],
                        "team":r["team_name"], "task_id":r["id"]})
        # 2) 댓글 5+ 미해결 (미완료 카드)
        rows = c.execute(
            """SELECT t.id, t.title, u.name, tm.name AS team_name,
                      (SELECT COUNT(*) FROM task_comments WHERE task_id=t.id) AS cn
               FROM tasks t JOIN users u ON t.user_id=u.id
               LEFT JOIN teams tm ON u.team_id=tm.id
               WHERE t.status != '완료'
               GROUP BY t.id HAVING cn >= 5 ORDER BY cn DESC LIMIT 30"""
        ).fetchall()
        for r in rows:
            out.append({"rule":f"댓글 {r['cn']}건 미해결", "severity":"mid",
                        "title":r["title"], "owner":r["name"],
                        "team":r["team_name"], "task_id":r["id"]})
        # 3) 한 사람 동시 8개+ 진행중
        rows = c.execute(
            """SELECT u.id, u.name, tm.name AS team_name,
                      COUNT(*) AS cn
               FROM tasks t JOIN users u ON t.user_id=u.id
               LEFT JOIN teams tm ON u.team_id=tm.id
               WHERE t.status='진행중' AND date(t.work_date) >= date('now','-7 days')
               GROUP BY u.id HAVING cn >= 8 ORDER BY cn DESC LIMIT 20"""
        ).fetchall()
        for r in rows:
            out.append({"rule":f"동시 진행 {r['cn']}건 — 과부하", "severity":"mid",
                        "title":f"{r['name']} 업무 과집중", "owner":r["name"],
                        "team":r["team_name"], "task_id":None})
        # 4) 대표 코멘트 24h 무답변
        rows = c.execute(
            """SELECT tc.task_id, tc.body, tc.created_at, t.title, u.name AS owner
               FROM task_comments tc
               JOIN tasks t ON tc.task_id=t.id
               JOIN users u ON t.user_id=u.id
               WHERE tc.is_ceo_request=1
                 AND datetime(tc.created_at) <= datetime('now','-1 day')
                 AND NOT EXISTS (
                    SELECT 1 FROM task_comments tc2
                    WHERE tc2.task_id=tc.task_id AND tc2.id > tc.id
                 )
               LIMIT 30"""
        ).fetchall()
        for r in rows:
            out.append({"rule":"대표 요청 24h 무응답", "severity":"high",
                        "title":r["title"], "owner":r["owner"],
                        "team":None, "task_id":r["task_id"]})
    return out


def get_task_comments(task_id):
    with db_session() as c:
        rows = c.execute(
            """SELECT tc.*, u.name AS author_name, u.rank AS author_rank, u.role AS author_role
               FROM task_comments tc
               JOIN users u ON tc.author_id = u.id
               WHERE tc.task_id=?
               ORDER BY tc.created_at""",
            (task_id,),
        ).fetchall()
        return [dict(r) for r in rows]


def delete_comment(comment_id, user_id):
    """본인 댓글만 삭제 가능 (admin/ceo는 모두)"""
    with db_session() as c:
        u = c.execute("SELECT role FROM users WHERE id=?", (user_id,)).fetchone()
        if not u:
            return False
        cm = c.execute("SELECT author_id FROM task_comments WHERE id=?", (comment_id,)).fetchone()
        if not cm:
            return False
        if cm["author_id"] != user_id and u["role"] not in ("ceo", "admin"):
            return False
        c.execute("DELETE FROM task_comments WHERE id=?", (comment_id,))
        c.execute("DELETE FROM notifications WHERE comment_id=?", (comment_id,))
        return True


def get_notifications(user_id, only_unread=False, limit=30):
    with db_session() as c:
        sql = "SELECT * FROM notifications WHERE user_id=?"
        if only_unread:
            sql += " AND is_read=0"
        sql += " ORDER BY created_at DESC LIMIT ?"
        rows = c.execute(sql, (user_id, limit)).fetchall()
        return [dict(r) for r in rows]


def count_unread(user_id):
    with db_session() as c:
        r = c.execute(
            "SELECT COUNT(*) FROM notifications WHERE user_id=? AND is_read=0",
            (user_id,),
        ).fetchone()
        return r[0] if r else 0


def mark_notification_read(notif_id, user_id):
    with db_session() as c:
        c.execute(
            "UPDATE notifications SET is_read=1 WHERE id=? AND user_id=?",
            (notif_id, user_id),
        )


def mark_all_read(user_id):
    with db_session() as c:
        c.execute(
            "UPDATE notifications SET is_read=1 WHERE user_id=? AND is_read=0",
            (user_id,),
        )


# =====================================================
# HAIST WORKS — 물류 모듈 (parts / 관리코드 발행대장)
# KNK PMS V3 표준 (엑셀정리스킬 §11~12 준수)
# =====================================================
from datetime import datetime as _dt, date as _date

BIZ_CODE = {"검사기": "T", "자동화": "M"}
BIZ_NAME = {v: k for k, v in BIZ_CODE.items()}

STAGES = ["제안작성", "제안제출", "수주확정", "납품", "개조", "A/S"]
NEEDS_CODE_STAGES = ("수주확정", "납품", "개조", "A/S")
PO_TYPES = ["신규", "추가"]
LOGI_STATUSES = ["수주예정", "진행중", "납품완료", "취소", "보류"]


def _logi_now() -> str:
    return _dt.now().strftime("%Y-%m-%d %H:%M:%S")


# ── 부품 마스터 (parts) CRUD ─────────────────────────────
def parts_list(q: str = "", biz_div: str = "", category: str = ""):
    sql = "SELECT * FROM parts WHERE 1=1"
    params: list = []
    if q:
        sql += " AND (part_no LIKE ? OR part_name LIKE ? OR spec LIKE ? OR maker LIKE ?)"
        like = f"%{q}%"
        params += [like, like, like, like]
    if biz_div:
        sql += " AND biz_div = ?"
        params.append(biz_div)
    if category:
        sql += " AND category = ?"
        params.append(category)
    sql += " ORDER BY id DESC"
    with db_session() as c:
        return c.execute(sql, params).fetchall()


def parts_get(pid: int):
    with db_session() as c:
        return c.execute("SELECT * FROM parts WHERE id = ?", (pid,)).fetchone()


def parts_create(data: dict) -> int:
    cols = ["part_no", "part_name", "spec", "maker", "origin", "unit",
            "currency", "std_price", "biz_div", "category", "note",
            "is_active", "created_at", "updated_at"]
    now = _logi_now()
    values = [
        (data.get("part_no") or "").strip(),
        (data.get("part_name") or "").strip(),
        (data.get("spec") or "").strip(),
        (data.get("maker") or "").strip(),
        (data.get("origin") or "").strip(),
        (data.get("unit") or "EA").strip() or "EA",
        (data.get("currency") or "KRW").strip() or "KRW",
        float(data.get("std_price") or 0),
        (data.get("biz_div") or "").strip(),
        (data.get("category") or "").strip(),
        (data.get("note") or "").strip(),
        1 if data.get("is_active", 1) else 0,
        now, now,
    ]
    placeholders = ",".join(["?"] * len(cols))
    with db_session() as c:
        cur = c.execute(
            f"INSERT INTO parts ({','.join(cols)}) VALUES ({placeholders})",
            values,
        )
        return cur.lastrowid


def parts_update(pid: int, data: dict) -> None:
    fields = ["part_no", "part_name", "spec", "maker", "origin", "unit",
              "currency", "std_price", "biz_div", "category", "note", "is_active"]
    sets = ", ".join([f"{f} = ?" for f in fields]) + ", updated_at = ?"
    values = [
        (data.get("part_no") or "").strip(),
        (data.get("part_name") or "").strip(),
        (data.get("spec") or "").strip(),
        (data.get("maker") or "").strip(),
        (data.get("origin") or "").strip(),
        (data.get("unit") or "EA").strip() or "EA",
        (data.get("currency") or "KRW").strip() or "KRW",
        float(data.get("std_price") or 0),
        (data.get("biz_div") or "").strip(),
        (data.get("category") or "").strip(),
        (data.get("note") or "").strip(),
        1 if data.get("is_active", 1) else 0,
        _logi_now(),
    ]
    with db_session() as c:
        c.execute(f"UPDATE parts SET {sets} WHERE id = ?", values + [pid])


def parts_delete(pid: int) -> None:
    with db_session() as c:
        c.execute("DELETE FROM parts WHERE id = ?", (pid,))


def parts_count() -> dict:
    with db_session() as c:
        total = c.execute("SELECT COUNT(*) FROM parts").fetchone()[0]
        active = c.execute("SELECT COUNT(*) FROM parts WHERE is_active = 1").fetchone()[0]
        by_div = c.execute(
            "SELECT biz_div, COUNT(*) AS n FROM parts GROUP BY biz_div"
        ).fetchall()
    return {
        "total": total,
        "active": active,
        "by_div": {r["biz_div"] or "(미지정)": r["n"] for r in by_div},
    }


# ── 프로젝트 / 관리코드 발행대장 (projects) CRUD ──────────
def generate_mgmt_code(biz_div: str, today=None) -> str:
    """KNK PMS 8자리 관리코드: [일련3][T/M][YYMM]. 같은 사업부·연월 내 +1"""
    if biz_div not in ("T", "M"):
        raise ValueError(f"biz_div must be T or M (got: {biz_div})")
    today = today or _date.today()
    yymm = today.strftime("%y%m")
    pat = f"%{biz_div}{yymm}"
    with db_session() as c:
        rows = c.execute(
            "SELECT mgmt_code FROM projects WHERE mgmt_code LIKE ?", (pat,)
        ).fetchall()
    max_seq = 0
    for r in rows:
        code = r["mgmt_code"] or ""
        if len(code) == 8 and code[3] == biz_div and code[4:] == yymm:
            try:
                max_seq = max(max_seq, int(code[:3]))
            except ValueError:
                pass
    return f"{max_seq + 1:03d}{biz_div}{yymm}"


def projects_list_logi(q: str = "", biz_div: str = "", stage: str = "",
                       status: str = ""):
    sql = ("SELECT p.*, p.name AS project_name "
           "FROM projects p WHERE 1=1")
    params: list = []
    if q:
        sql += (" AND (p.mgmt_code LIKE ? OR p.name LIKE ? OR p.customer_name LIKE ? "
                "OR p.model_name LIKE ? OR p.pm_name LIKE ? OR p.sales_name LIKE ?)")
        like = f"%{q}%"
        params += [like] * 6
    if biz_div:
        sql += " AND p.biz_div = ?"
        params.append(biz_div)
    if stage:
        sql += " AND p.stage = ?"
        params.append(stage)
    if status:
        sql += " AND p.status = ?"
        params.append(status)
    sql += " ORDER BY p.id DESC"
    with db_session() as c:
        return c.execute(sql, params).fetchall()


def projects_get_logi(pid: int):
    with db_session() as c:
        return c.execute(
            "SELECT *, name AS project_name FROM projects WHERE id = ?",
            (pid,),
        ).fetchone()


def _project_insert_or_update_values(data: dict) -> dict:
    """공통 필드 정리"""
    return {
        "name": (data.get("project_name") or data.get("name") or "").strip(),
        "biz_div": (data.get("biz_div") or "").strip(),
        "customer_name": (data.get("customer_name") or data.get("customer") or "").strip(),
        "model_name": (data.get("model_name") or data.get("model") or "").strip(),
        "stage": (data.get("stage") or "제안작성").strip() or "제안작성",
        "po_type": (data.get("po_type") or "신규").strip() or "신규",
        "status": (data.get("status") or "수주예정").strip() or "수주예정",
        "customer_po": (data.get("customer_po") or "").strip(),
        "currency": (data.get("currency") or "KRW").strip() or "KRW",
        "order_amount": float(data.get("order_amount") or 0),
        "order_date": (data.get("order_date") or "").strip(),
        "due_date": (data.get("due_date") or "").strip(),
        "pm_name": (data.get("pm_name") or data.get("pm") or "").strip(),
        "sales_name": (data.get("sales_name") or data.get("sales") or "").strip(),
        "logi_note": (data.get("logi_note") or data.get("note") or "").strip(),
    }


def projects_create_logi(data: dict) -> tuple[int, str | None]:
    vals = _project_insert_or_update_values(data)
    # 자동 채번
    code = None
    if vals["stage"] in NEEDS_CODE_STAGES and vals["biz_div"] in ("T", "M"):
        code = generate_mgmt_code(vals["biz_div"])
    now = _logi_now()
    with db_session() as c:
        cur = c.execute("""
            INSERT INTO projects
            (mgmt_code, name, biz_div, customer_name, model_name, stage, po_type,
             status, customer_po, currency, order_amount, order_date, due_date,
             pm_name, sales_name, logi_note, created_at, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (code, vals["name"], vals["biz_div"], vals["customer_name"],
              vals["model_name"], vals["stage"], vals["po_type"], vals["status"],
              vals["customer_po"], vals["currency"], vals["order_amount"],
              vals["order_date"], vals["due_date"], vals["pm_name"],
              vals["sales_name"], vals["logi_note"], now, now))
        return cur.lastrowid, code


def projects_update_logi(pid: int, data: dict) -> str | None:
    current = projects_get_logi(pid)
    if not current:
        return None
    vals = _project_insert_or_update_values(data)
    new_code = current["mgmt_code"]
    if not new_code and vals["stage"] in NEEDS_CODE_STAGES and vals["biz_div"] in ("T", "M"):
        new_code = generate_mgmt_code(vals["biz_div"])
    with db_session() as c:
        c.execute("""
            UPDATE projects
            SET mgmt_code=?, name=?, biz_div=?, customer_name=?, model_name=?,
                stage=?, po_type=?, status=?, customer_po=?, currency=?,
                order_amount=?, order_date=?, due_date=?,
                pm_name=?, sales_name=?, logi_note=?, updated_at=?
            WHERE id=?
        """, (new_code, vals["name"], vals["biz_div"], vals["customer_name"],
              vals["model_name"], vals["stage"], vals["po_type"], vals["status"],
              vals["customer_po"], vals["currency"], vals["order_amount"],
              vals["order_date"], vals["due_date"], vals["pm_name"],
              vals["sales_name"], vals["logi_note"], _logi_now(), pid))
    return new_code


def projects_delete_logi(pid: int) -> None:
    with db_session() as c:
        c.execute("DELETE FROM projects WHERE id = ?", (pid,))


def projects_count_logi() -> dict:
    with db_session() as c:
        total = c.execute(
            "SELECT COUNT(*) FROM projects WHERE biz_div IS NOT NULL AND biz_div != ''"
        ).fetchone()[0]
        with_code = c.execute(
            "SELECT COUNT(*) FROM projects WHERE mgmt_code IS NOT NULL AND mgmt_code != ''"
        ).fetchone()[0]
        in_progress = c.execute(
            "SELECT COUNT(*) FROM projects WHERE status = '진행중'"
        ).fetchone()[0]
        by_div = c.execute(
            "SELECT biz_div, COUNT(*) AS n FROM projects WHERE biz_div IS NOT NULL AND biz_div != '' GROUP BY biz_div"
        ).fetchall()
        by_stage = c.execute(
            "SELECT stage, COUNT(*) AS n FROM projects WHERE stage IS NOT NULL AND stage != '' GROUP BY stage"
        ).fetchall()
    return {
        "total": total,
        "with_code": with_code,
        "in_progress": in_progress,
        "by_div": {(BIZ_NAME.get(r["biz_div"]) or r["biz_div"] or "(미지정)"): r["n"]
                   for r in by_div},
        "by_stage": {r["stage"] or "(미지정)": r["n"] for r in by_stage},
    }


# =====================================================
# HAIST WORKS — 발주 모듈 (공급사 + PO 헤더/라인)
# =====================================================
PO_STATUSES = ["작성중", "발주완료", "부분입고", "입고완료", "취소"]
PO_TYPES_KIND = ["일반", "긴급", "정기"]
SHIPPING_TERMS = ["EXW", "FOB", "CIF", "DDP", "국내"]
PAYMENT_TERMS = ["선금", "현금", "30일", "60일", "기타"]


# ── 공급사 (suppliers) CRUD ──────────────────────────────
def suppliers_list(q: str = "", active_only: bool = False):
    sql = "SELECT * FROM suppliers WHERE 1=1"
    params = []
    if q:
        sql += " AND (name LIKE ? OR code LIKE ? OR contact LIKE ? OR country LIKE ?)"
        like = f"%{q}%"
        params += [like] * 4
    if active_only:
        sql += " AND is_active = 1"
    sql += " ORDER BY is_active DESC, name"
    with db_session() as c:
        return c.execute(sql, params).fetchall()


def supplier_get(sid: int):
    with db_session() as c:
        return c.execute("SELECT * FROM suppliers WHERE id = ?", (sid,)).fetchone()


def supplier_create(data: dict) -> int:
    now = _logi_now()
    with db_session() as c:
        cur = c.execute("""
            INSERT INTO suppliers
            (name, code, contact, email, phone, country, currency,
             payment_terms, note, is_active, created_at, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            (data.get("name") or "").strip(),
            (data.get("code") or "").strip() or None,
            (data.get("contact") or "").strip(),
            (data.get("email") or "").strip(),
            (data.get("phone") or "").strip(),
            (data.get("country") or "").strip(),
            (data.get("currency") or "KRW").strip() or "KRW",
            (data.get("payment_terms") or "").strip(),
            (data.get("note") or "").strip(),
            1 if data.get("is_active", 1) else 0,
            now, now,
        ))
        return cur.lastrowid


def supplier_update(sid: int, data: dict) -> None:
    with db_session() as c:
        c.execute("""
            UPDATE suppliers SET
              name=?, code=?, contact=?, email=?, phone=?, country=?,
              currency=?, payment_terms=?, note=?, is_active=?, updated_at=?
            WHERE id=?
        """, (
            (data.get("name") or "").strip(),
            (data.get("code") or "").strip() or None,
            (data.get("contact") or "").strip(),
            (data.get("email") or "").strip(),
            (data.get("phone") or "").strip(),
            (data.get("country") or "").strip(),
            (data.get("currency") or "KRW").strip() or "KRW",
            (data.get("payment_terms") or "").strip(),
            (data.get("note") or "").strip(),
            1 if data.get("is_active", 1) else 0,
            _logi_now(), sid,
        ))


def supplier_delete(sid: int) -> None:
    with db_session() as c:
        c.execute("DELETE FROM suppliers WHERE id = ?", (sid,))


def suppliers_count() -> dict:
    with db_session() as c:
        total = c.execute("SELECT COUNT(*) FROM suppliers").fetchone()[0]
        active = c.execute("SELECT COUNT(*) FROM suppliers WHERE is_active = 1").fetchone()[0]
    return {"total": total, "active": active}


# ── 발주번호 자동 채번 ─────────────────────────────────
def generate_po_number(today=None) -> str:
    """발주번호 채번: PO-YYMMDD-NNN (같은 날짜 내 순차)"""
    today = today or _date.today()
    yymmdd = today.strftime("%y%m%d")
    prefix = f"PO-{yymmdd}-"
    with db_session() as c:
        rows = c.execute(
            "SELECT po_number FROM purchase_orders WHERE po_number LIKE ?",
            (f"{prefix}%",),
        ).fetchall()
    max_seq = 0
    for r in rows:
        tail = (r["po_number"] or "")[len(prefix):]
        if tail.isdigit():
            max_seq = max(max_seq, int(tail))
    return f"{prefix}{max_seq + 1:03d}"


# ── 발주 (purchase_orders + po_items) CRUD ─────────────
def po_list(q: str = "", status: str = "", supplier_id: int = 0,
            project_id: int = 0):
    sql = """
      SELECT po.*, s.name AS supplier_name, p.name AS project_name,
             p.mgmt_code AS mgmt_code, u.name AS creator_name
      FROM purchase_orders po
      LEFT JOIN suppliers s ON po.supplier_id = s.id
      LEFT JOIN projects p ON po.project_id = p.id
      LEFT JOIN users u ON po.created_by = u.id
      WHERE 1=1
    """
    params = []
    if q:
        sql += """ AND (po.po_number LIKE ? OR s.name LIKE ?
                   OR p.mgmt_code LIKE ? OR p.name LIKE ?)"""
        like = f"%{q}%"
        params += [like] * 4
    if status:
        sql += " AND po.status = ?"
        params.append(status)
    if supplier_id:
        sql += " AND po.supplier_id = ?"
        params.append(supplier_id)
    if project_id:
        sql += " AND po.project_id = ?"
        params.append(project_id)
    sql += " ORDER BY po.id DESC"
    with db_session() as c:
        return c.execute(sql, params).fetchall()


def po_get(po_id: int):
    """발주 헤더 + 라인 + 관련 마스터"""
    with db_session() as c:
        header = c.execute("""
            SELECT po.*, s.name AS supplier_name, s.contact AS supplier_contact,
                   s.email AS supplier_email, s.country AS supplier_country,
                   p.name AS project_name, p.mgmt_code, p.customer_name AS customer,
                   u.name AS creator_name
            FROM purchase_orders po
            LEFT JOIN suppliers s ON po.supplier_id = s.id
            LEFT JOIN projects p ON po.project_id = p.id
            LEFT JOIN users u ON po.created_by = u.id
            WHERE po.id = ?
        """, (po_id,)).fetchone()
        if not header:
            return None, []
        items = c.execute("""
            SELECT * FROM po_items WHERE po_id = ?
            ORDER BY line_no, id
        """, (po_id,)).fetchall()
    return header, items


def po_create(data: dict, items: list[dict], created_by: int = 0) -> tuple[int, str]:
    """발주 헤더 + 라인 일괄 생성. 발주번호는 자동 채번."""
    po_number = generate_po_number()
    order_date = (data.get("order_date") or "").strip() or _date.today().isoformat()
    now = _logi_now()
    total = 0.0

    with db_session() as c:
        cur = c.execute("""
            INSERT INTO purchase_orders
            (po_number, project_id, supplier_id, order_date, expected_date,
             currency, exchange_rate, total_amount, status, shipping_terms,
             payment_terms, po_type, created_by, note, created_at, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            po_number,
            int(data.get("project_id") or 0) or None,
            int(data.get("supplier_id") or 0) or None,
            order_date,
            (data.get("expected_date") or "").strip() or None,
            (data.get("currency") or "KRW").strip() or "KRW",
            float(data.get("exchange_rate") or 1),
            0,  # total 임시
            (data.get("status") or "작성중").strip() or "작성중",
            (data.get("shipping_terms") or "").strip(),
            (data.get("payment_terms") or "").strip(),
            (data.get("po_type") or "일반").strip() or "일반",
            int(created_by or 0) or None,
            (data.get("note") or "").strip(),
            now, now,
        ))
        po_id = cur.lastrowid

        # 라인 삽입
        for idx, it in enumerate(items, start=1):
            qty = float(it.get("quantity") or 0)
            price = float(it.get("unit_price") or 0)
            amt = round(qty * price, 2)
            total += amt

            # 부품 스냅샷 채우기
            part_id = int(it.get("part_id") or 0) or None
            part_no = it.get("part_no_snapshot") or ""
            part_name = it.get("part_name_snapshot") or ""
            spec = it.get("spec_snapshot") or ""
            unit = it.get("unit") or "EA"
            if part_id:
                p = c.execute(
                    "SELECT part_no, part_name, spec, unit FROM parts WHERE id=?",
                    (part_id,)
                ).fetchone()
                if p:
                    part_no = part_no or p["part_no"] or ""
                    part_name = part_name or p["part_name"] or ""
                    spec = spec or p["spec"] or ""
                    unit = unit or p["unit"] or "EA"

            c.execute("""
                INSERT INTO po_items
                (po_id, line_no, part_id, part_no_snapshot, part_name_snapshot,
                 spec_snapshot, unit, quantity, unit_price, amount,
                 delivery_date, note)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """, (po_id, idx, part_id, part_no, part_name, spec, unit,
                  qty, price, amt,
                  (it.get("delivery_date") or "").strip() or None,
                  (it.get("note") or "").strip()))

        # total 업데이트
        c.execute("UPDATE purchase_orders SET total_amount=? WHERE id=?",
                  (round(total, 2), po_id))

    return po_id, po_number


def po_update(po_id: int, data: dict, items: list[dict]) -> None:
    """발주 헤더 + 라인 전체 교체 방식 (단순화)"""
    now = _logi_now()
    total = 0.0
    with db_session() as c:
        # 헤더 업데이트
        c.execute("""
            UPDATE purchase_orders SET
              project_id=?, supplier_id=?, order_date=?, expected_date=?,
              currency=?, exchange_rate=?, status=?, shipping_terms=?,
              payment_terms=?, po_type=?, note=?, updated_at=?
            WHERE id=?
        """, (
            int(data.get("project_id") or 0) or None,
            int(data.get("supplier_id") or 0) or None,
            (data.get("order_date") or "").strip(),
            (data.get("expected_date") or "").strip() or None,
            (data.get("currency") or "KRW").strip() or "KRW",
            float(data.get("exchange_rate") or 1),
            (data.get("status") or "작성중").strip() or "작성중",
            (data.get("shipping_terms") or "").strip(),
            (data.get("payment_terms") or "").strip(),
            (data.get("po_type") or "일반").strip() or "일반",
            (data.get("note") or "").strip(),
            now, po_id,
        ))
        # 라인 전체 삭제 후 재삽입 (received_qty 보존은 4단계에서 처리)
        c.execute("DELETE FROM po_items WHERE po_id = ?", (po_id,))
        for idx, it in enumerate(items, start=1):
            qty = float(it.get("quantity") or 0)
            price = float(it.get("unit_price") or 0)
            amt = round(qty * price, 2)
            total += amt
            part_id = int(it.get("part_id") or 0) or None
            c.execute("""
                INSERT INTO po_items
                (po_id, line_no, part_id, part_no_snapshot, part_name_snapshot,
                 spec_snapshot, unit, quantity, unit_price, amount,
                 delivery_date, note)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """, (po_id, idx, part_id,
                  it.get("part_no_snapshot") or "",
                  it.get("part_name_snapshot") or "",
                  it.get("spec_snapshot") or "",
                  it.get("unit") or "EA",
                  qty, price, amt,
                  (it.get("delivery_date") or "").strip() or None,
                  (it.get("note") or "").strip()))
        c.execute("UPDATE purchase_orders SET total_amount=? WHERE id=?",
                  (round(total, 2), po_id))


def po_delete(po_id: int) -> None:
    with db_session() as c:
        c.execute("DELETE FROM purchase_orders WHERE id = ?", (po_id,))


def po_count() -> dict:
    with db_session() as c:
        total = c.execute("SELECT COUNT(*) FROM purchase_orders").fetchone()[0]
        by_status = c.execute(
            "SELECT status, COUNT(*) AS n FROM purchase_orders GROUP BY status"
        ).fetchall()
        sum_amt = c.execute(
            "SELECT currency, SUM(total_amount) AS s FROM purchase_orders "
            "WHERE status != '취소' GROUP BY currency"
        ).fetchall()
    return {
        "total": total,
        "by_status": {r["status"] or "(미지정)": r["n"] for r in by_status},
        "amount_by_currency": {r["currency"] or "KRW": r["s"] or 0 for r in sum_amt},
    }


# =====================================================
# 게시판 (boards / board_posts / board_comments)
# =====================================================

BOARD_CATEGORIES = ["공지", "일반", "자료", "질문"]


def board_get_or_create_company():
    """전사 게시판 ID 반환 (없으면 생성)"""
    with db_session() as c:
        row = c.execute("SELECT id FROM boards WHERE type='company' LIMIT 1").fetchone()
        if row:
            return row["id"]
        c.execute("INSERT INTO boards (name,type) VALUES ('전사 게시판','company')")
        return c.execute("SELECT last_insert_rowid()").fetchone()[0]


def board_get_or_create_team(team_id):
    """부서 게시판 ID 반환 (없으면 생성)"""
    with db_session() as c:
        row = c.execute("SELECT id FROM boards WHERE type='team' AND team_id=?", (team_id,)).fetchone()
        if row:
            return row["id"]
        t = c.execute("SELECT name FROM teams WHERE id=?", (team_id,)).fetchone()
        name = f"{t['name']} 게시판" if t else f"팀{team_id} 게시판"
        c.execute("INSERT INTO boards (name,type,team_id) VALUES (?,'team',?)", (name, team_id))
        return c.execute("SELECT last_insert_rowid()").fetchone()[0]


def board_list_all():
    """전체 게시판 목록"""
    with db_session() as c:
        return c.execute(
            "SELECT b.*, t.name AS team_name FROM boards b LEFT JOIN teams t ON b.team_id=t.id ORDER BY b.type, b.id"
        ).fetchall()


def board_posts_list(board_id, include_pending=False):
    """게시글 목록. include_pending=True면 대기 글도 포함"""
    sql = """SELECT p.*, u.name AS author_name, u.rank AS author_rank,
                    t.name AS author_team,
                    (SELECT COUNT(*) FROM board_comments WHERE post_id=p.id) AS comment_count
             FROM board_posts p
             JOIN users u ON p.author_id=u.id
             LEFT JOIN teams t ON u.team_id=t.id
             WHERE p.board_id=?"""
    if not include_pending:
        sql += " AND p.approval_status='approved'"
    sql += " ORDER BY p.is_pinned DESC, p.created_at DESC"
    with db_session() as c:
        return c.execute(sql, (board_id,)).fetchall()


def board_posts_pending(board_id):
    """승인 대기 글 (팀장용)"""
    sql = """SELECT p.*, u.name AS author_name, u.rank AS author_rank
             FROM board_posts p JOIN users u ON p.author_id=u.id
             WHERE p.board_id=? AND p.approval_status='pending'
             ORDER BY p.created_at DESC"""
    with db_session() as c:
        return c.execute(sql, (board_id,)).fetchall()


def board_post_get(post_id):
    with db_session() as c:
        row = c.execute(
            """SELECT p.*, u.name AS author_name, u.rank AS author_rank,
                      t.name AS author_team, b.type AS board_type, b.team_id AS board_team_id,
                      b.name AS board_name
               FROM board_posts p
               JOIN users u ON p.author_id=u.id
               LEFT JOIN teams t ON u.team_id=t.id
               JOIN boards b ON p.board_id=b.id
               WHERE p.id=?""",
            (post_id,),
        ).fetchone()
        return dict(row) if row else None


def board_post_create(board_id, author_id, title, body, category="일반",
                      approval_status="approved"):
    with db_session() as c:
        c.execute(
            """INSERT INTO board_posts (board_id,author_id,title,body,category,approval_status)
               VALUES (?,?,?,?,?,?)""",
            (board_id, author_id, title.strip(), body.strip(), category, approval_status),
        )
        return c.execute("SELECT last_insert_rowid()").fetchone()[0]


def board_post_update(post_id, title, body, category):
    with db_session() as c:
        c.execute(
            "UPDATE board_posts SET title=?, body=?, category=?, updated_at=datetime('now','localtime') WHERE id=?",
            (title.strip(), body.strip(), category, post_id),
        )


def board_post_delete(post_id):
    with db_session() as c:
        c.execute("DELETE FROM board_posts WHERE id=?", (post_id,))


def board_post_approve(post_id, approver_id):
    with db_session() as c:
        c.execute(
            "UPDATE board_posts SET approval_status='approved', approved_by=?, approved_at=datetime('now','localtime') WHERE id=?",
            (approver_id, post_id),
        )


def board_post_reject(post_id, approver_id, reason=""):
    with db_session() as c:
        c.execute(
            "UPDATE board_posts SET approval_status='rejected', approved_by=?, approved_at=datetime('now','localtime'), reject_reason=? WHERE id=?",
            (approver_id, reason, post_id),
        )


def board_post_toggle_pin(post_id):
    with db_session() as c:
        cur = c.execute("SELECT is_pinned FROM board_posts WHERE id=?", (post_id,)).fetchone()
        if cur:
            c.execute("UPDATE board_posts SET is_pinned=? WHERE id=?",
                      (0 if cur["is_pinned"] else 1, post_id))


def board_post_increment_view(post_id):
    with db_session() as c:
        c.execute("UPDATE board_posts SET view_count=view_count+1 WHERE id=?", (post_id,))


def board_comments_list(post_id):
    with db_session() as c:
        return c.execute(
            """SELECT c.*, u.name AS author_name, u.rank AS author_rank
               FROM board_comments c JOIN users u ON c.author_id=u.id
               WHERE c.post_id=? ORDER BY c.created_at""",
            (post_id,),
        ).fetchall()


def board_comment_create(post_id, author_id, body):
    with db_session() as c:
        c.execute(
            "INSERT INTO board_comments (post_id,author_id,body) VALUES (?,?,?)",
            (post_id, author_id, body.strip()),
        )


def board_comment_delete(comment_id):
    with db_session() as c:
        c.execute("DELETE FROM board_comments WHERE id=?", (comment_id,))


# =====================================================
# 변경 Inform 시스템 — 사전 준비 (2026-04-20)
# 정식 구현은 Research 세션 v2 문서 도착 후
# 설계 청사진: HAIST_WORKS/_DESIGN_변경_Inform.md
# =====================================================

CHANGE_TYPES = ["기구설계", "전장설계", "소프트웨어", "BOM", "도면", "Concept", "사양"]
CHANGE_URGENCIES = ["일반", "긴급", "예약"]
CHANGE_STATUSES = ["작성중", "공지중", "확인완료", "취소"]

# 변경 종류 × 사업부 → 영향 부서 자동 판별 규칙
IMPACT_RULES = {
    "기구설계": {
        "T": ["전장설계팀", "소프트웨어팀", "가공팀", "제조기술1팀", "구매팀"],
        "M": ["전장설계팀", "소프트웨어팀", "가공팀", "제조기술2팀", "구매팀"],
    },
    "전장설계": {
        "T": ["소프트웨어팀", "제조기술1팀", "구매팀"],
        "M": ["소프트웨어팀", "제조기술2팀", "구매팀"],
    },
    "소프트웨어": {
        "T": ["검사기팀", "제조기술1팀"],
        "M": ["제조기술2팀"],
    },
    "BOM": {
        "T": ["구매팀", "제조기술1팀", "가공팀"],
        "M": ["구매팀", "제조기술2팀", "가공팀"],
    },
    "도면": {
        "T": ["가공팀", "제조기술1팀", "품질팀"],
        "M": ["가공팀", "제조기술2팀", "품질팀"],
    },
    "Concept": {"T": "ALL", "M": "ALL"},
    "사양": {
        "T": ["기술영업팀", "검사기팀", "설계팀", "품질팀"],
        "M": ["기술영업팀", "설계팀", "품질팀"],
    },
}


def detect_impact_teams(change_type: str, biz_div: str) -> list[int]:
    """변경 종류 + 사업부 → 영향 받는 팀 ID 리스트.
    biz_div 'T'(검사기) / 'M'(자동화). Concept는 전 팀.
    """
    rule = IMPACT_RULES.get(change_type, {}).get(biz_div, [])
    if not rule:
        return []
    if rule == "ALL":
        with db_session() as c:
            return [r["id"] for r in c.execute("SELECT id FROM teams").fetchall()]
    placeholders = ",".join(["?"] * len(rule))
    with db_session() as c:
        rows = c.execute(
            f"SELECT id FROM teams WHERE name IN ({placeholders})",
            tuple(rule),
        ).fetchall()
    return [r["id"] for r in rows]


def detect_impact_users(team_ids: list[int]) -> list[dict]:
    """영향 팀들의 활성 사용자 목록 (알림 발송 대상)"""
    if not team_ids:
        return []
    placeholders = ",".join(["?"] * len(team_ids))
    with db_session() as c:
        rows = c.execute(
            f"""SELECT u.id, u.name, u.team_id, t.name AS team_name
                FROM users u JOIN teams t ON u.team_id = t.id
                WHERE u.team_id IN ({placeholders}) AND u.is_active = 1""",
            tuple(team_ids),
        ).fetchall()
    return [dict(r) for r in rows]


def gen_change_no(today=None) -> str:
    """변경번호 자동 채번: CHG-YYMMDD-NNN (해당 날짜 +1)"""
    today = today or _date.today()
    yymmdd = today.strftime("%y%m%d")
    prefix = f"CHG-{yymmdd}-"
    # NOTE: changes 테이블이 아직 없으면 001부터 시작
    try:
        with db_session() as c:
            rows = c.execute(
                "SELECT change_no FROM changes WHERE change_no LIKE ?",
                (f"{prefix}%",),
            ).fetchall()
        max_seq = 0
        for r in rows:
            try:
                seq = int(r["change_no"].rsplit("-", 1)[-1])
                max_seq = max(max_seq, seq)
            except (ValueError, IndexError):
                pass
        return f"{prefix}{max_seq + 1:03d}"
    except sqlite3.OperationalError:
        # changes 테이블 미존재 시
        return f"{prefix}001"


# 카카오워크 Webhook 더미 (실제 URL은 사용자 발급 후 config로)
def kakao_webhook_send(channel_id: str, text: str, blocks: list = None) -> bool:
    """카카오워크 Webhook 메시지 발송. 정식 구현 전 stub.
    Phase 2에서 사용자가 webhook URL 발급한 후 실제 호출 코드로 교체."""
    # TODO: requests.post(KAKAO_WEBHOOK_URL[channel_id], json={...})
    print(f"[KAKAO STUB] channel={channel_id} text={text[:80]}...")
    return True


# ── 변경 Inform CRUD ─────────────────────────────────────────
def changes_list(q="", change_type="", urgency="", status="", scope_user_id=None):
    """변경 목록 조회. scope_user_id 주면 그 사용자가 영향자인 것만"""
    base = """SELECT c.*, u.name AS author_name, u.rank AS author_rank,
                     t.name AS author_team,
                     (SELECT COUNT(*) FROM change_impacts WHERE change_id=c.id) AS impact_count,
                     (SELECT COUNT(*) FROM change_reads WHERE change_id=c.id AND ack_at IS NOT NULL) AS ack_count
              FROM changes c
              JOIN users u ON c.author_id = u.id
              LEFT JOIN teams t ON u.team_id = t.id
              WHERE 1=1"""
    params = []
    if q:
        base += " AND (c.change_no LIKE ? OR c.title LIKE ? OR c.description LIKE ? OR c.target_label LIKE ?)"
        like = f"%{q}%"
        params += [like] * 4
    if change_type:
        base += " AND c.change_type = ?"
        params.append(change_type)
    if urgency:
        base += " AND c.urgency = ?"
        params.append(urgency)
    if status:
        base += " AND c.status = ?"
        params.append(status)
    if scope_user_id:
        base += """ AND c.id IN (
            SELECT change_id FROM change_impacts ci
            LEFT JOIN users u2 ON u2.team_id = ci.impact_team_id
            WHERE u2.id = ? OR ci.impact_user_id = ?
        )"""
        params += [scope_user_id, scope_user_id]
    base += " ORDER BY c.created_at DESC"
    with db_session() as c:
        return c.execute(base, params).fetchall()


def change_get(cid: int) -> dict | None:
    with db_session() as c:
        row = c.execute(
            """SELECT c.*, u.name AS author_name, u.rank AS author_rank,
                      t.name AS author_team
               FROM changes c
               JOIN users u ON c.author_id = u.id
               LEFT JOIN teams t ON u.team_id = t.id
               WHERE c.id = ?""",
            (cid,),
        ).fetchone()
        return dict(row) if row else None


def change_create(data: dict, author_id: int) -> tuple[int, str]:
    """변경 등록 + 영향 자동 판별 + change_reads 미리 생성"""
    change_no = gen_change_no()
    biz_div = (data.get("biz_div") or "").strip()
    change_type = (data.get("change_type") or "").strip()

    with db_session() as c:
        cur = c.execute(
            """INSERT INTO changes
               (change_no, change_type, biz_div, target_kind, target_id, target_label,
                project_id, title, description, before_value, after_value,
                attached_files, urgency, author_id, status)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                change_no, change_type, biz_div,
                data.get("target_kind"), data.get("target_id"),
                data.get("target_label", "").strip(),
                data.get("project_id"),
                data.get("title", "").strip(),
                data.get("description", "").strip(),
                data.get("before_value", "").strip(),
                data.get("after_value", "").strip(),
                data.get("attached_files", ""),
                data.get("urgency", "일반"),
                author_id,
                "공지중",
            ),
        )
        change_id = c.execute("SELECT last_insert_rowid()").fetchone()[0]

        # 영향 부서 자동 판별
        impact_teams = detect_impact_teams(change_type, biz_div) if biz_div in ("T", "M") else []
        for team_id in impact_teams:
            c.execute(
                """INSERT INTO change_impacts
                   (change_id, impact_kind, impact_team_id, auto_detected, impact_reason)
                   VALUES (?, 'team', ?, 1, ?)""",
                (change_id, team_id, f"{change_type} 변경 → 자동 판별"),
            )

        # change_reads 미리 생성 (영향 부서원 전원)
        if impact_teams:
            placeholders = ",".join(["?"] * len(impact_teams))
            users = c.execute(
                f"SELECT id FROM users WHERE team_id IN ({placeholders}) AND is_active = 1",
                tuple(impact_teams),
            ).fetchall()
            for u in users:
                try:
                    c.execute(
                        "INSERT INTO change_reads (change_id, user_id) VALUES (?, ?)",
                        (change_id, u["id"]),
                    )
                except sqlite3.IntegrityError:
                    pass

        # 작성자에게도 read 행 (보낸 사람도 본인이 다 봤다는 의미)
        try:
            c.execute(
                "INSERT INTO change_reads (change_id, user_id, read_at, ack_at) VALUES (?, ?, ?, ?)",
                (change_id, author_id, _logi_now(), _logi_now()),
            )
        except sqlite3.IntegrityError:
            pass

        # notified_at 표시
        c.execute("UPDATE changes SET notified_at=? WHERE id=?", (_logi_now(), change_id))

    return change_id, change_no


def change_get_impacts(cid: int) -> list[dict]:
    """영향 부서/사용자 + 부서별 확인 현황"""
    with db_session() as c:
        impacts = c.execute(
            """SELECT ci.*, t.name AS team_name
               FROM change_impacts ci
               LEFT JOIN teams t ON ci.impact_team_id = t.id
               WHERE ci.change_id = ?""",
            (cid,),
        ).fetchall()
        result = []
        for imp in impacts:
            d = dict(imp)
            if imp["impact_team_id"]:
                # 부서원 확인 현황
                stats = c.execute(
                    """SELECT COUNT(*) AS total,
                              SUM(CASE WHEN cr.ack_at IS NOT NULL THEN 1 ELSE 0 END) AS ack
                       FROM users u
                       LEFT JOIN change_reads cr ON cr.user_id = u.id AND cr.change_id = ?
                       WHERE u.team_id = ? AND u.is_active = 1""",
                    (cid, imp["impact_team_id"]),
                ).fetchone()
                d["total_users"] = stats["total"] or 0
                d["ack_users"] = stats["ack"] or 0
            result.append(d)
    return result


def change_get_reads(cid: int) -> list[dict]:
    with db_session() as c:
        rows = c.execute(
            """SELECT cr.*, u.name AS user_name, u.rank AS user_rank, t.name AS team_name
               FROM change_reads cr
               JOIN users u ON cr.user_id = u.id
               LEFT JOIN teams t ON u.team_id = t.id
               WHERE cr.change_id = ?
               ORDER BY cr.ack_at DESC NULLS LAST, cr.read_at DESC""",
            (cid,),
        ).fetchall()
    return [dict(r) for r in rows]


def change_mark_read(cid: int, user_id: int):
    """페이지 열람 기록"""
    with db_session() as c:
        c.execute(
            """INSERT INTO change_reads (change_id, user_id, read_at)
               VALUES (?, ?, ?)
               ON CONFLICT(change_id, user_id) DO UPDATE SET read_at = COALESCE(read_at, excluded.read_at)""",
            (cid, user_id, _logi_now()),
        )


def change_ack(cid: int, user_id: int, note: str = ""):
    """확인 응답"""
    with db_session() as c:
        c.execute(
            """INSERT INTO change_reads (change_id, user_id, read_at, ack_at, ack_note)
               VALUES (?, ?, ?, ?, ?)
               ON CONFLICT(change_id, user_id) DO UPDATE SET
                   read_at = COALESCE(read_at, excluded.read_at),
                   ack_at = excluded.ack_at,
                   ack_note = excluded.ack_note""",
            (cid, user_id, _logi_now(), _logi_now(), note),
        )
        # 모든 영향자 확인 시 status=확인완료
        remaining = c.execute(
            """SELECT COUNT(*) FROM change_reads
               WHERE change_id = ? AND ack_at IS NULL""",
            (cid,),
        ).fetchone()[0]
        if remaining == 0:
            c.execute(
                "UPDATE changes SET status='확인완료', completed_at=? WHERE id=?",
                (_logi_now(), cid),
            )


def change_delete(cid: int):
    with db_session() as c:
        c.execute("DELETE FROM changes WHERE id=?", (cid,))


def change_unread_count(user_id: int) -> int:
    """내가 영향자인데 아직 ack 안 한 변경 카운트 (사이드바 뱃지)"""
    with db_session() as c:
        row = c.execute(
            """SELECT COUNT(*) FROM change_reads
               WHERE user_id = ? AND ack_at IS NULL""",
            (user_id,),
        ).fetchone()
    return row[0] if row else 0


def change_recent_count(user_id: int = None, days: int = 1) -> int:
    """홈 KPI용 — 최근 N일 변경 (전체 또는 내 관련)"""
    sql = """SELECT COUNT(*) FROM changes
             WHERE created_at >= datetime('now', '-' || ? || ' day', 'localtime')"""
    params = [days]
    if user_id:
        sql += """ AND id IN (
            SELECT change_id FROM change_impacts ci
            LEFT JOIN users u ON u.team_id = ci.impact_team_id
            WHERE u.id = ? OR ci.impact_user_id = ?
        )"""
        params += [user_id, user_id]
    with db_session() as c:
        return c.execute(sql, params).fetchone()[0]


