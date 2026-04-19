"""
╔══════════════════════════════════════════════════════════════╗
║  KNK PMS V4 — 03_검사기_부품출하 sync.py                      ║
║  물류상세 INV 자동계산 → 마스터 집계                           ║
║  실행: python sync.py                                         ║
╚══════════════════════════════════════════════════════════════╝

처리 흐름:
  ① 물류상세 INV 자동계산 (단가×상승률, ÷환율, ×관세율)
  ② 물류상세 → 마스터(조달진행현황) 집계 write-back
"""

import os, sys, glob, datetime, logging
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import *
from build import build_detail


def _setup_logger():
    os.makedirs(log_dir(), exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = os.path.join(log_dir(), f"sync_{TYPE_CODE}_{ts}.log")
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)


def _val(ws, r, c):
    v = ws.cell(row=r, column=c).value
    return v if v is not None else ""

def _num(ws, r, c):
    v = ws.cell(row=r, column=c).value
    try: return float(v)
    except (TypeError, ValueError): return 0

def _date_val(ws, r, c):
    v = ws.cell(row=r, column=c).value
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    return None


def _calc_inv_details(log):
    """① 물류상세 INV 자동계산"""
    if not os.path.exists(DIR_LOGISTICS):
        return
    files = glob.glob(os.path.join(DIR_LOGISTICS, "*_물류상세.xlsx"))
    for fp in files:
        wb = load_workbook(fp)
        ws = wb.active
        changed = False
        for r in range(5, ws.max_row + 1):
            qty = _num(ws, r, LI["수량"])
            uprice = _num(ws, r, LI["단가"])
            rate = _num(ws, r, LI["상승률"])
            exrate = _num(ws, r, LI["환율"])
            duty_pct = _num(ws, r, LI["DUTY"])

            if not qty and not uprice:
                continue

            amount = qty * uprice
            fmt_cell(ws, r, LI["금액"], amount, is_money=True)

            if rate:
                inv_up = uprice * rate
                inv_amt = amount * rate
                fmt_cell(ws, r, LI["INV단가"], inv_up, is_money=True)
                fmt_cell(ws, r, LI["INV금액"], inv_amt, is_money=True)

                if exrate and exrate > 0:
                    inv_up_usd = inv_up / exrate
                    inv_amt_usd = inv_amt / exrate
                    fmt_cell(ws, r, LI["INV단가USD"], inv_up_usd, is_money=True, is_usd=True)
                    fmt_cell(ws, r, LI["INV금액USD"], inv_amt_usd, is_money=True, is_usd=True)

                    if duty_pct:
                        fmt_cell(ws, r, LI["관세KRW"], inv_amt * duty_pct, is_money=True)
                        fmt_cell(ws, r, LI["관세USD"], inv_amt_usd * duty_pct, is_money=True, is_usd=True)

                changed = True

        if changed:
            wb.save(fp)
            log.info(f"  INV 계산: {os.path.basename(fp)}")
        wb.close()


def _sync_master(log):
    """② 물류상세 → 마스터 집계"""
    fp = master_path()
    if not os.path.exists(fp):
        log.error("마스터 없음 — build.py 먼저 실행")
        return

    # 물류상세 집계
    summaries = {}
    if os.path.exists(DIR_LOGISTICS):
        for dfp in glob.glob(os.path.join(DIR_LOGISTICS, "*_물류상세.xlsx")):
            wb = load_workbook(dfp, data_only=True)
            ws = wb.active
            code = None; sj = ""; cust = ""; prod = ""
            total = 0; ordered = 0; received = 0; shipped = 0
            total_amt = 0; total_inv_usd = 0
            ship_dates = []

            for r in range(5, ws.max_row + 1):
                if not _val(ws, r, LI["품명"]):
                    continue
                if not code:
                    code = str(_val(ws, r, LI["관리코드"]))
                    sj = str(_val(ws, r, LI["수주번호"]))

                total += 1
                if _val(ws, r, LI["발주일"]): ordered += 1
                if str(_val(ws, r, LI["입고상태"])) == "입고완료": received += 1
                if str(_val(ws, r, LI["출하상태"])) == "출하완료": shipped += 1
                total_amt += _num(ws, r, LI["금액"])
                total_inv_usd += _num(ws, r, LI["INV금액USD"])
                sd = _date_val(ws, r, LI["출하일"])
                if sd: ship_dates.append(sd)

            if code:
                ship_dates.sort()
                summaries[code] = {
                    "sj": sj, "total": total, "ordered": ordered,
                    "received": received, "shipped": shipped,
                    "unprocessed": total - shipped,
                    "proc_rate": received / total if total else 0,
                    "ship_rate": shipped / total if total else 0,
                    "ship_count": len(set(ship_dates)),
                    "last_ship": ship_dates[-1] if ship_dates else None,
                    "total_amt": total_amt, "total_inv_usd": total_inv_usd,
                }
            wb.close()

    # 마스터에 write-back
    wb = load_workbook(fp)
    ws = wb.active

    # 기존 마스터 행의 수동 입력 데이터 보존
    existing = {}
    for r in range(5, ws.max_row + 1):
        code = _val(ws, r, 2)
        if code:
            existing[str(code)] = {
                "cust": _val(ws, r, 4), "prod": _val(ws, r, 5),
                "next_ship": _val(ws, r, 15), "status": _val(ws, r, 19),
                "reg_date": _val(ws, r, 18),
            }

    # 모든 코드 수집 (기존 + 신규)
    all_codes = list(dict.fromkeys(list(existing.keys()) + list(summaries.keys())))

    row = 5
    for idx, code in enumerate(all_codes, 1):
        s = summaries.get(code, {})
        e = existing.get(code, {})

        ws.cell(row=row, column=1).value = idx
        ws.cell(row=row, column=2).value = code
        ws.cell(row=row, column=3).value = s.get("sj", "")
        ws.cell(row=row, column=4).value = e.get("cust", "")
        ws.cell(row=row, column=5).value = e.get("prod", "")
        ws.cell(row=row, column=6).value = s.get("total", 0)
        ws.cell(row=row, column=7).value = s.get("ordered", 0)
        ws.cell(row=row, column=8).value = s.get("received", 0)
        ws.cell(row=row, column=9).value = s.get("shipped", 0)
        ws.cell(row=row, column=10).value = s.get("unprocessed", 0)

        pr = s.get("proc_rate", 0)
        sr = s.get("ship_rate", 0)
        ws.cell(row=row, column=11).value = pr if pr else None
        ws.cell(row=row, column=11).number_format = PCT
        ws.cell(row=row, column=12).value = sr if sr else None
        ws.cell(row=row, column=12).number_format = PCT

        ws.cell(row=row, column=13).value = s.get("ship_count", 0)
        ws.cell(row=row, column=14).value = s.get("last_ship")
        ws.cell(row=row, column=15).value = e.get("next_ship", "")  # 수동입력 보존
        fmt_cell(ws, row, 16, s.get("total_amt", 0), is_money=True)
        fmt_cell(ws, row, 17, s.get("total_inv_usd", 0), is_money=True, is_usd=True)
        ws.cell(row=row, column=18).value = e.get("reg_date") or TODAY.isoformat()
        ws.cell(row=row, column=19).value = e.get("status", "진행중")

        dfp = detail_path(code)
        ws.cell(row=row, column=20).value = f"{code}_물류상세.xlsx" if os.path.exists(dfp) else ""
        row += 1

    wb.save(fp)
    log.info(f"  마스터 집계: {len(all_codes)}건")


def _auto_create_details(log):
    """① 마스터에 등록된 관리코드별 물류상세 파일 자동생성"""
    os.makedirs(DIR_LOGISTICS, exist_ok=True)
    fp = master_path()
    if not os.path.exists(fp):
        log.info("  마스터 파일 없음 — 스킵")
        return
    wb = load_workbook(fp, data_only=False)
    ws = wb.active
    created = 0
    for r in range(5, ws.max_row + 1):
        code = _val(ws, r, 2)  # C_CODE=2
        if not code:
            break
        dfp = detail_path(code)
        if os.path.exists(dfp):
            continue
        sj   = str(_val(ws, r, 3))
        cust = str(_val(ws, r, 4))
        prod = str(_val(ws, r, 5))
        build_detail(str(code), sj_num=sj, cust=cust, prod=prod)
        created += 1
    wb.close()
    log.info(f"  물류상세 자동생성: {created}건")


def sync_all():
    log = _setup_logger()
    log.info("=" * 60)
    log.info(f"  KNK PMS V4 — {TYPE_NAME} 동기화 시작")
    log.info("=" * 60)

    log.info("① 물류상세 자동생성")
    _auto_create_details(log)

    log.info("② INV 자동계산")
    _calc_inv_details(log)

    log.info("③ 마스터 집계")
    _sync_master(log)

    log.info("=" * 60)
    log.info("  동기화 완료!")
    log.info("=" * 60)


if __name__ == "__main__":
    sync_all()
