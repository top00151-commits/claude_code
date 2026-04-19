"""
╔══════════════════════════════════════════════════════════════╗
║  KNK PMS V4 — 05_소모품 build.py                              ║
║  출하대장 (소모품목록 + 출하관리) 2시트 구조                    ║
║  실행: python build.py                                        ║
╚══════════════════════════════════════════════════════════════╝
"""

import os, sys, glob, datetime
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import *


def _add_dropdown(ws, col, options, row_start=5, row_end=2000, allow_blank=True):
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    dv.error = "목록에서 선택하세요"
    col_letter = get_column_letter(col)
    dv.add(f"{col_letter}{row_start}:{col_letter}{row_end}")
    ws.add_data_validation(dv)


def build_consumables():
    """소모품 출하대장 생성 (2시트)"""
    fp = consumables_path()
    print(f"  [BUILD] 소모품: {os.path.basename(fp)}")
    wb = Workbook()

    # ── 시트1: 소모품목록 ──
    ws1 = wb.active
    ws1.title = "소모품목록"
    ws1.sheet_properties.tabColor = TAB_INPUT

    mc1 = CONS_LIST_MAX_COL
    setup_r1(ws1, f"㈜케이엔케이 │ {TYPE_NAME} │ 소모품목록 │ {YEAR}", mc1)
    apply_r3_guide(ws1, mc1, R3_MAP_CONS_LIST)
    apply_r4_header(ws1, mc1, CONS_LIST_LABELS, R4_MAP_CONS_LIST)
    money_cols = [7]  # 기본단가
    format_data_rows(ws1, mc1, money_cols=money_cols)
    _add_dropdown(ws1, 5, UNITS)  # 단위
    _add_dropdown(ws1, 8, CONS_CATEGORIES)  # 카테고리
    apply_protection(ws1, mc1, R3_MAP_CONS_LIST)
    auto_fit_columns(ws1)

    # ── 시트2: 출하관리 ──
    ws2 = wb.create_sheet("출하관리")
    ws2.sheet_properties.tabColor = TAB_INPUT

    mc2 = LI_MAX_COL
    setup_r1(ws2, f"㈜케이엔케이 │ {TYPE_NAME} │ 출하관리 │ {YEAR}", mc2)
    apply_r3_guide(ws2, mc2, R3_MAP_CONSUMABLES)
    apply_r4_header(ws2, mc2, LI_HEADER_LABELS, LI_COL_AREA, LI_AREA_COLORS)

    money_cols = [LI["단가"], LI["금액"], LI["INV단가"], LI["INV금액"], LI["관세KRW"]]
    money_usd = [LI["INV단가USD"], LI["INV금액USD"], LI["관세USD"]]
    format_data_rows(ws2, mc2, money_cols=money_cols, money_usd_cols=money_usd, row_end=5000)

    _add_dropdown(ws2, LI["통화"], CURRENCIES, row_end=5000)
    _add_dropdown(ws2, LI["입고상태"], RECEIPT_STATUSES, row_end=5000)
    _add_dropdown(ws2, LI["출하상태"], SHIP_ITEM_STATUSES, row_end=5000)
    _add_dropdown(ws2, LI["거래유형"], TRADE_TYPES, row_end=5000)
    _add_dropdown(ws2, LI["통관구분"], CUSTOMS_TYPES, row_end=5000)
    _add_dropdown(ws2, LI["통관분류"], SHIP_TYPES, row_end=5000)
    _add_dropdown(ws2, LI["통관상태"], SHIP_STATUSES, row_end=5000)
    _add_dropdown(ws2, LI["단위"], UNITS, row_end=5000)
    _add_dropdown(ws2, LI["원산지"], ORIGINS, row_end=5000)

    apply_protection(ws2, mc2, R3_MAP_CONSUMABLES, row_end=5000)
    auto_fit_columns(ws2)

    wb.save(fp)
    print(f"    → 저장 완료")
    return fp


def clean_existing():
    count = 0
    for f in glob.glob(os.path.join(DIR_DATA, "*.xlsx")):
        os.remove(f)
        count += 1
    if count:
        print(f"  [CLEAN] {count}개 기존 파일 삭제")


def build_all():
    print("=" * 60)
    print(f"  KNK PMS V4 — {TYPE_NAME} 빌드 시작")
    print(f"  시각: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    clean_existing()
    build_consumables()
    print("=" * 60)
    print(f"  빌드 완료! 출하대장 1개 (2시트) 생성")
    print("=" * 60)


if __name__ == "__main__":
    build_all()
