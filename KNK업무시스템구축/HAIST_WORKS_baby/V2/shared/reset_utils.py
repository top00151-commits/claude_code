"""
╔══════════════════════════════════════════════════════════════╗
║  shared/reset_utils.py                                       ║
║  xlsx 데이터 비움 유틸 — 현재 데이터를 백업 후 R5+ 내용 제거  ║
╚══════════════════════════════════════════════════════════════╝

사용:
  from shared.reset_utils import reset_module
  reset_module(MODULE_DIR)

동작:
  - 모듈 폴더의 모든 xlsx(백업·참고용 제외)를 _참고용_원본데이터/ 하위에 타임스탬프 접미사로 복사
  - 각 xlsx의 모든 시트 R5+ 데이터(value + comment) 제거
  - R1~R4(타이틀·슬로건·가이드·헤더) 및 서식·병합·보호·드롭다운은 그대로 유지
  - 이후 sync → apply_standard 실행으로 스펙 준수 상태의 빈 양식 완성
"""
import os
import shutil
import datetime
import glob
from openpyxl import load_workbook

BACKUP_DIR_NAME = "_참고용_원본데이터"
EXCLUDE_KEYWORDS = ("_backup_", "_데이터참고용_", BACKUP_DIR_NAME)


def _is_backup_or_ref(path):
    base = os.path.basename(path)
    return any(k in base for k in EXCLUDE_KEYWORDS)


def _list_target_xlsx(module_dir):
    """모듈 폴더 직속 xlsx만 대상 (하위 폴더 xlsx 제외)."""
    pat = os.path.join(module_dir, "*.xlsx")
    return [p for p in glob.glob(pat) if not _is_backup_or_ref(p)]


def _backup(xlsx_path, backup_dir, ts):
    os.makedirs(backup_dir, exist_ok=True)
    base = os.path.basename(xlsx_path)
    name, ext = os.path.splitext(base)
    target = os.path.join(backup_dir, f"{name}_{ts}{ext}")
    shutil.copy2(xlsx_path, target)
    return target


def _clear_sheet_data(ws, row_start=5):
    """시트의 R{row_start}+ 영역에서 값 + 주석 제거."""
    if ws.max_row < row_start:
        return 0
    cleared = 0
    for r in range(row_start, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            if cell.value is not None:
                cell.value = None
                cleared += 1
            if cell.comment is not None:
                cell.comment = None
    return cleared


def reset_module(module_dir, log=print):
    """모듈 폴더의 모든 xlsx를 백업 후 R5+ 비움."""
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = os.path.join(module_dir, BACKUP_DIR_NAME, ts)

    targets = _list_target_xlsx(module_dir)
    if not targets:
        log(f"  ⚠ 대상 xlsx 없음: {module_dir}")
        return

    log(f"  백업 폴더: {backup_dir}")
    log(f"  대상 파일: {len(targets)}개")
    log("")

    total_cleared = 0
    for fp in targets:
        base = os.path.basename(fp)
        log(f"▶ {base}")
        # 1. 백업
        bak = _backup(fp, backup_dir, ts)
        log(f"  ✓ 백업: {os.path.basename(bak)}")

        # 2. R5+ 비움
        wb = load_workbook(fp)
        sheet_cleared = {}
        for sh_name in wb.sheetnames:
            ws = wb[sh_name]
            n = _clear_sheet_data(ws, row_start=5)
            sheet_cleared[sh_name] = n
            total_cleared += n
        for sh, n in sheet_cleared.items():
            log(f"  ✓ {sh}: {n}개 셀 비움")

        wb.save(fp)
        log(f"  ✓ 저장 완료")
        log("")

    log(f"완료 — 총 {total_cleared}개 셀 비움")
    log(f"원본 보존: {backup_dir}")
