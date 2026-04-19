"""
╔══════════════════════════════════════════════════════════════╗
║  KNK PMS V4 — 03_검사기_부품출하 build.py                     ║
║  마스터(조달진행현황) + 물류상세(프로젝트별 BOM 40열)          ║
║  실행: python build.py                                        ║
╚══════════════════════════════════════════════════════════════╝
"""

import os, sys, glob, datetime
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import *


def _add_dropdown(ws, col, options, row_start=5, row_end=2000, allow_blank=True):
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    dv.error = "목록에서 선택하세요"
    col_letter = get_column_letter(col)
    dv.add(f"{col_letter}{row_start}:{col_letter}{row_end}")
    ws.add_data_validation(dv)


def build_master():
    """마스터(조달진행현황) 파일 생성"""
    fp = master_path()
    print(f"  [BUILD] 마스터: {os.path.basename(fp)}")
    wb = Workbook()
    ws = wb.active
    ws.title = "조달진행현황"
    ws.sheet_properties.tabColor = TAB_VIEW

    mc = MASTER_MAX_COL
    setup_r1(ws, f"㈜케이엔케이 │ {TYPE_NAME} │ 조달진행현황 │ {YEAR}", mc)
    apply_r3_guide(ws, mc, R3_MAP_MASTER)
    apply_r4_header(ws, mc, MASTER_LABELS, R4_MAP_MASTER, MASTER_AREA_COLORS)

    money_cols = [16]; money_usd = [17]; pct_cols = [11, 12]
    format_data_rows(ws, mc, money_cols=money_cols, money_usd_cols=money_usd, pct_cols=pct_cols)
    _add_dropdown(ws, 19, MASTER_STATUSES)
    apply_protection(ws, mc, R3_MAP_MASTER)
    auto_fit_columns(ws)

    wb.save(fp)
    print(f"    → 저장 완료")
    return fp


def build_detail(mgmt_code, sj_num="", cust="", prod=""):
    """물류상세 상세파일 생성 (40열)"""
    fp = detail_path(mgmt_code)
    os.makedirs(os.path.dirname(fp), exist_ok=True)
    print(f"  [BUILD] 물류상세: {mgmt_code}")
    wb = Workbook()
    ws = wb.active
    ws.title = "물류입력"
    ws.sheet_properties.tabColor = TAB_INPUT

    mc = LI_MAX_COL
    title = f"㈜케이엔케이 │ {TYPE_NAME} │ {mgmt_code} │ {sj_num} │ {cust} │ {prod}"
    setup_r1(ws, title, mc)
    apply_r3_guide(ws, mc, R3_MAP_LI_DETAIL)
    apply_r4_header(ws, mc, LI_HEADER_LABELS, LI_COL_AREA, LI_AREA_COLORS)

    money_cols = [LI["단가"], LI["금액"], LI["INV단가"], LI["INV금액"], LI["관세KRW"]]
    money_usd = [LI["INV단가USD"], LI["INV금액USD"], LI["관세USD"]]
    format_data_rows(ws, mc, money_cols=money_cols, money_usd_cols=money_usd, row_end=5000)

    _add_dropdown(ws, LI["통화"], CURRENCIES, row_end=5000)
    _add_dropdown(ws, LI["입고상태"], RECEIPT_STATUSES, row_end=5000)
    _add_dropdown(ws, LI["출하상태"], SHIP_ITEM_STATUSES, row_end=5000)
    _add_dropdown(ws, LI["거래유형"], TRADE_TYPES, row_end=5000)
    _add_dropdown(ws, LI["통관구분"], CUSTOMS_TYPES, row_end=5000)
    _add_dropdown(ws, LI["통관분류"], SHIP_TYPES, row_end=5000)
    _add_dropdown(ws, LI["통관상태"], SHIP_STATUSES, row_end=5000)
    _add_dropdown(ws, LI["단위"], UNITS, row_end=5000)
    _add_dropdown(ws, LI["원산지"], ORIGINS, row_end=5000)

    apply_protection(ws, mc, R3_MAP_LI_DETAIL, row_end=5000)
    auto_fit_columns(ws)

    wb.save(fp)
    print(f"    → 저장 완료")
    return fp


def clean_existing():
    patterns = [
        os.path.join(DIR_DATA, "*.xlsx"),
        os.path.join(DIR_LOGISTICS, "*.xlsx"),
    ]
    count = 0
    for pattern in patterns:
        for f in glob.glob(pattern):
            os.remove(f)
            count += 1
    if count:
        print(f"  [CLEAN] {count}개 기존 파일 삭제")


def build_all():
    print("=" * 60)
    print(f"  KNK PMS V4 — {TYPE_NAME} 빌드 시작")
    print(f"  시각: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    clean_existing()
    build_master()
    print("=" * 60)
    print(f"  빌드 완료! 마스터 1개 생성 (물류상세는 sync 시 자동생성)")
    print("=" * 60)


if __name__ == "__main__":
    build_all()
